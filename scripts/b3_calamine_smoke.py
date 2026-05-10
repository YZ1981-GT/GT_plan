"""B3 calamine 解析速度基准（smoke 级）。

对比 openpyxl(read_only) vs python-calamine 在真实样本上的解析速度。
只测"把所有数据行读出来"的时间，不做 identify/convert/write。
"""
from __future__ import annotations

import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

SAMPLES = [
    ("YG4001-30", ROOT / "数据/YG4001-30重庆医药集团宜宾医药有限公司新健康大药房临港店-余额表+序时账.xlsx"),
    ("YG36", ROOT / "数据/YG36-重庆医药集团四川物流有限公司2025.xlsx"),
    ("YG2101", ROOT / "数据/YG2101-重庆医药集团四川医药有限公司2025年-科目余额表+序时账.xlsx"),
]


def _bench_openpyxl(path: Path) -> tuple[float, int, list[str]]:
    import openpyxl

    t0 = time.time()
    total_rows = 0
    sheets = []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
        for sn in sheets:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
    finally:
        wb.close()
    return time.time() - t0, total_rows, sheets


def _bench_calamine(path: Path) -> tuple[float, int, list[str]]:
    from python_calamine import CalamineWorkbook

    t0 = time.time()
    total_rows = 0
    sheets = []
    wb = CalamineWorkbook.from_path(str(path))
    sheets = list(wb.sheet_names)
    for sn in sheets:
        sheet = wb.get_sheet_by_name(sn).to_python()
        total_rows += len(sheet)
    return time.time() - t0, total_rows, sheets


def main() -> int:
    print(f"{'样本':<15} {'大小 MB':<10} {'engine':<12} {'耗时 s':<10} {'行数':<10} {'sheets'}")
    print("-" * 100)
    for name, path in SAMPLES:
        if not path.exists():
            print(f"{name:<15} -- 文件不存在")
            continue
        size_mb = path.stat().st_size / 1e6

        try:
            elapsed_o, rows_o, sheets_o = _bench_openpyxl(path)
            print(f"{name:<15} {size_mb:<10.2f} {'openpyxl':<12} {elapsed_o:<10.2f} {rows_o:<10} {sheets_o}")
        except Exception as exc:
            print(f"{name:<15} {size_mb:<10.2f} openpyxl    FAIL: {exc}")

        try:
            elapsed_c, rows_c, sheets_c = _bench_calamine(path)
            print(f"{name:<15} {size_mb:<10.2f} {'calamine':<12} {elapsed_c:<10.2f} {rows_c:<10} {sheets_c}")
        except Exception as exc:
            print(f"{name:<15} {size_mb:<10.2f} calamine    FAIL: {exc}")

        try:
            speedup = elapsed_o / elapsed_c if elapsed_c > 0 else 0
            print(f"{'  → 加速':<15} {'':<10} {'':<12} {speedup:.1f}x")
        except Exception:
            pass
        print()

    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.exit(main())
