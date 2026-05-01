#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
从审计报告模板Excel中提取报表行次，生成 report_config_seed.json
"""
import json, os, re, sys
import openpyxl
try:
    import xlrd
except ImportError:
    print("pip install xlrd"); sys.exit(1)

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOE_CON = os.path.join(BASE, "审计报告模板", "国企版", "合并", "1.1-2025国企财务报表.xlsx")
SOE_STD = os.path.join(BASE, "审计报告模板", "国企版", "单体", "1.1-2025国企财务报表.xlsx")
LST_CON = os.path.join(BASE, "审计报告模板", "上市版", "合并_上市",
                        "2.股份年审－经审计的财务报表-202601.xlsx")
LST_STD = os.path.join(BASE, "审计报告模板", "上市版", "单体_上市",
                        "2.股份年审－经审计的财务报表-202601.xlsx")
OUTPUT  = os.path.join(BASE, "backend", "data", "report_config_seed.json")

# ── helpers ──
SKIP_NAMES = {
    "项目", "项            目", "项       目", "项   目", "项  目",
    "编制单位", "金额单位", "栏       次", "栏            次", "栏次",
    "——", "资  产", "资产", "负债和所有者权益", "负债和所有者权益（或股东权益）",
    "负债和股东权益", "合并", "公司", "本期金额", "上期金额",
    "期末余额", "期初余额", "上年年末余额", "附注",
}

def _skip(name):
    if not name: return True
    s = name.strip()
    if not s or len(s) < 2: return True
    if s in SKIP_NAMES: return True
    if re.match(r'^20\d{2}年', s): return True
    if re.match(r'^\d+(\.\d+)?$', s): return True  # pure number / row number
    if s.startswith("编制单位"): return True
    if s.startswith("金额单位"): return True
    if s.startswith("单位："): return True
    if "资产负债表" in s and ("合并" in s or "公司" in s or "续" in s): return True
    if s in ("合 并 及 公 司 资 产 负 债 表", "合 并 及 公 司 资 产 负 债 表 （续）",
             "公 司 资 产 负 债 表", "公 司 资 产 负 债 表 （续）"): return True
    if "利 润 表" in s or "现 金 流 量 表" in s: return True
    if "权益变动表" in s and len(s) > 8: return True
    if "减值准备情况表" in s and len(s) > 8: return True
    # Footer / note lines
    if s.startswith("企业负责人"): return True
    if s.startswith("公司法定代表人"): return True
    if s.startswith("注：") or s.startswith("注:"): return True
    if s.startswith("提示：") or s.startswith("提示:"): return True
    if s.startswith("【提示"): return True
    if s.startswith("财会["): return True
    if s.startswith("（1）本表") or s.startswith("（2）其他:"): return True
    if "合并及公司利润表" in s: return True
    if "合并及公司现金流量表" in s: return True
    if "公司利润表" in s and len(s) < 10: return True
    if "公司现金流量表" in s and len(s) < 12: return True
    if "负债合计" in s and "负" in s and "债" in s and "合" in s and "计" in s and len(s) > 6:
        # "负 债 合 计" with spaces - keep it but clean
        pass
    return False

def _clean(s):
    if not s: return ""
    s = str(s).strip()
    # collapse multiple spaces between Chinese characters
    s = re.sub(r'(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])', '', s)
    # but keep single space after special prefixes
    s = re.sub(r'\s{2,}', ' ', s)
    return s.strip()

def _is_total(name):
    for kw in ("合计", "小计", "总计", "总额"):
        if kw in name: return True
    return False

def _indent(name):
    s = name.strip()
    # category headers
    if s.endswith("：") or s.endswith(":"): return 0
    if _is_total(s): return 0
    if re.match(r'^[一二三四五六七八九十]+、', s): return 0
    if re.match(r'^(减：|加：|其中：)', s): return 1
    # "其中：" inside name
    if "其中：" in s: return 2
    return 1

# ── SOE xlsx extraction ──
def _cv(ws, r, c):
    try:
        v = ws.cell(row=r, column=c).value
        return _clean(str(v)) if v is not None else ""
    except: return ""

def extract_soe_bs(wb):
    """Extract balance sheet from SOE xlsx (two sheets merged)"""
    rows = []
    for sn in wb.sheetnames:
        if "资产负债表" in sn and "企财01" in sn:
            ws = wb[sn]
            # Find data start (after header rows)
            start = 5
            for r in range(1, 10):
                v = _cv(ws, r, 1)
                if "流动资产" in v or "流动负债" in v:
                    start = r
                    break
            for r in range(start, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
    return rows

def extract_soe_is(wb):
    rows = []
    for sn in wb.sheetnames:
        if "利润表" in sn and "企财02" in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

def extract_soe_cfs(wb):
    rows = []
    for sn in wb.sheetnames:
        if "现金流量表" in sn and "企财03" in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

def extract_soe_eq(wb, sheet_keyword):
    """Extract equity statement rows"""
    rows = []
    for sn in wb.sheetnames:
        if "权益变动表" in sn and sheet_keyword in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

def extract_soe_impairment(wb, sheet_keyword):
    rows = []
    for sn in wb.sheetnames:
        if "资产减值准备" in sn and sheet_keyword in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

# ── Listed xlsx extraction (also openpyxl) ──
def extract_listed_bs(wb):
    rows = []
    for sn in wb.sheetnames:
        if "资产负债表" in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
    return rows

def extract_listed_is(wb):
    rows = []
    for sn in wb.sheetnames:
        if "利润表" in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

def extract_listed_cfs(wb):
    rows = []
    for sn in wb.sheetnames:
        if "现金流量表" in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

def extract_listed_eq(wb, sheet_keyword):
    rows = []
    for sn in wb.sheetnames:
        if "权益变动表" in sn and sheet_keyword in sn:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                name = _cv(ws, r, 1)
                if name and not _skip(name):
                    rows.append(name)
            break
    return rows

# ── Build seed JSON ──
def build_rows(names, prefix):
    """Convert list of row names to seed row objects"""
    result = []
    for i, name in enumerate(names, 1):
        code = f"{prefix}-{i:03d}"
        result.append({
            "row_code": code,
            "row_number": i,
            "row_name": name,
            "indent_level": _indent(name),
            "is_total_row": _is_total(name),
            "formula": None,
            "formula_category": None,
            "formula_description": None,
            "formula_source": None,
            "parent_row_code": None,
        })
    return result

def build_config(report_type, applicable_standard, template_variant, scope, desc, names, prefix):
    return {
        "report_type": report_type,
        "applicable_standard": applicable_standard,
        "template_variant": template_variant,
        "scope": scope,
        "description": desc,
        "rows": build_rows(names, prefix),
    }

# ── Main ──
def main():
    all_configs = []

    # ── 国企合并 ──
    print("读取国企合并...")
    wb = openpyxl.load_workbook(SOE_CON, data_only=True, read_only=True)
    bs = extract_soe_bs(wb)
    is_ = extract_soe_is(wb)
    cfs = extract_soe_cfs(wb)
    eq = extract_soe_eq(wb, "合并")
    imp = extract_soe_impairment(wb, "合并")
    wb.close()
    print(f"  BS={len(bs)} IS={len(is_)} CFS={len(cfs)} EQ={len(eq)} IMP={len(imp)}")
    all_configs.append(build_config("balance_sheet", "soe_consolidated", "soe", "consolidated",
                                     "balance_sheet — 致同国企版合并", bs, "BS"))
    all_configs.append(build_config("income_statement", "soe_consolidated", "soe", "consolidated",
                                     "income_statement — 致同国企版合并", is_, "IS"))
    all_configs.append(build_config("cash_flow_statement", "soe_consolidated", "soe", "consolidated",
                                     "cash_flow_statement — 致同国企版合并", cfs, "CFS"))
    all_configs.append(build_config("equity_statement", "soe_consolidated", "soe", "consolidated",
                                     "equity_statement — 致同国企版合并", eq, "EQ"))

    # ── 国企单体 ──
    print("读取国企单体...")
    wb = openpyxl.load_workbook(SOE_STD, data_only=True, read_only=True)
    bs = extract_soe_bs(wb)
    is_ = extract_soe_is(wb)
    cfs = extract_soe_cfs(wb)
    # 单体版只有一个权益变动表sheet
    eq = extract_soe_eq(wb, "企财04")
    imp = extract_soe_impairment(wb, "企财06")
    wb.close()
    print(f"  BS={len(bs)} IS={len(is_)} CFS={len(cfs)} EQ={len(eq)} IMP={len(imp)}")
    all_configs.append(build_config("balance_sheet", "soe_standalone", "soe", "standalone",
                                     "balance_sheet — 致同国企版单体", bs, "BS"))
    all_configs.append(build_config("income_statement", "soe_standalone", "soe", "standalone",
                                     "income_statement — 致同国企版单体", is_, "IS"))
    all_configs.append(build_config("cash_flow_statement", "soe_standalone", "soe", "standalone",
                                     "cash_flow_statement — 致同国企版单体", cfs, "CFS"))
    all_configs.append(build_config("equity_statement", "soe_standalone", "soe", "standalone",
                                     "equity_statement — 致同国企版单体", eq, "EQ"))

    # ── 上市合并 ──
    print("读取上市合并...")
    wb = openpyxl.load_workbook(LST_CON, data_only=True, read_only=True)
    bs = extract_listed_bs(wb)
    is_ = extract_listed_is(wb)
    cfs = extract_listed_cfs(wb)
    eq = extract_listed_eq(wb, "合并")
    wb.close()
    print(f"  BS={len(bs)} IS={len(is_)} CFS={len(cfs)} EQ={len(eq)}")
    all_configs.append(build_config("balance_sheet", "listed_consolidated", "listed", "consolidated",
                                     "balance_sheet — 致同上市版合并", bs, "BS"))
    all_configs.append(build_config("income_statement", "listed_consolidated", "listed", "consolidated",
                                     "income_statement — 致同上市版合并", is_, "IS"))
    all_configs.append(build_config("cash_flow_statement", "listed_consolidated", "listed", "consolidated",
                                     "cash_flow_statement — 致同上市版合并", cfs, "CFS"))
    all_configs.append(build_config("equity_statement", "listed_consolidated", "listed", "consolidated",
                                     "equity_statement — 致同上市版合并", eq, "EQ"))

    # ── 上市单体 ──
    print("读取上市单体...")
    wb = openpyxl.load_workbook(LST_STD, data_only=True, read_only=True)
    bs = extract_listed_bs(wb)
    is_ = extract_listed_is(wb)
    cfs = extract_listed_cfs(wb)
    eq = extract_listed_eq(wb, "公司")
    wb.close()
    print(f"  BS={len(bs)} IS={len(is_)} CFS={len(cfs)} EQ={len(eq)}")
    all_configs.append(build_config("balance_sheet", "listed_standalone", "listed", "standalone",
                                     "balance_sheet — 致同上市版单体", bs, "BS"))
    all_configs.append(build_config("income_statement", "listed_standalone", "listed", "standalone",
                                     "income_statement — 致同上市版单体", is_, "IS"))
    all_configs.append(build_config("cash_flow_statement", "listed_standalone", "listed", "standalone",
                                     "cash_flow_statement — 致同上市版单体", cfs, "CFS"))
    all_configs.append(build_config("equity_statement", "listed_standalone", "listed", "standalone",
                                     "equity_statement — 致同上市版单体", eq, "EQ"))

    # ── Cash Flow Supplement (from note templates) ──
    # These rows are the same for all 4 standards
    cfs_supp_rows = [
        "1.将净利润调节为经营活动现金流量：",
        "净利润",
        "加：资产减值损失",
        "信用减值损失",
        "固定资产折旧、油气资产折耗、生产性生物资产折旧",
        "使用权资产折旧",
        "无形资产摊销",
        "长期待摊费用摊销",
    ]
    cfs_supp_rows += [
        '处置固定资产、无形资产和其他长期资产的损失（收益以"\uff0d"号填列）',
        '固定资产报废损失（收益以"\uff0d"号填列）',
        '公允价值变动损失（收益以"\uff0d"号填列）',
        '财务费用（收益以"\uff0d"号填列）',
        '投资损失（收益以"\uff0d"号填列）',
        '递延所得税资产减少（增加以"\uff0d"号填列）',
        '递延所得税负债增加（减少以"\uff0d"号填列）',
        '存货的减少（增加以"\uff0d"号填列）',
        '经营性应收项目的减少（增加以"\uff0d"号填列）',
        '经营性应付项目的增加（减少以"\uff0d"号填列）',
        "其他",
        "经营活动产生的现金流量净额",
        "2.不涉及现金收支的重大投资和筹资活动：",
        "债务转为资本",
        "一年内到期的可转换公司债券",
        "新增使用权资产",
        "3.现金及现金等价物净变动情况：",
        "现金的期末余额",
        "减：现金的期初余额",
        "加：现金等价物的期末余额",
        "减：现金等价物的期初余额",
        "现金及现金等价物净增加额",
    ]

    for std_key, variant, scope, desc_suffix in [
        ("soe_consolidated", "soe", "consolidated", "致同国企版合并"),
        ("soe_standalone", "soe", "standalone", "致同国企版单体"),
        ("listed_consolidated", "listed", "consolidated", "致同上市版合并"),
        ("listed_standalone", "listed", "standalone", "致同上市版单体"),
    ]:
        all_configs.append(build_config(
            "cash_flow_supplement", std_key, variant, scope,
            f"cash_flow_supplement — {desc_suffix}",
            cfs_supp_rows, "CFSS"
        ))

    # Print summary
    print(f"\n总计 {len(all_configs)} 个配置块")
    for cfg in all_configs:
        print(f"  {cfg['applicable_standard']}/{cfg['report_type']}: {len(cfg['rows'])} 行")

    # Print detailed rows for verification
    for cfg in all_configs:
        print(f"\n{'='*60}")
        print(f"{cfg['applicable_standard']} / {cfg['report_type']}")
        print(f"{'='*60}")
        for row in cfg["rows"]:
            indent = "  " * row["indent_level"]
            total = " [合计]" if row["is_total_row"] else ""
            print(f"  {row['row_code']:>8} {indent}{row['row_name']}{total}")

    # Write output
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(all_configs, f, ensure_ascii=False, indent=2)
    print(f"\n已写入 {OUTPUT}")
    print(f"文件大小: {os.path.getsize(OUTPUT)} bytes")

if __name__ == "__main__":
    main()
