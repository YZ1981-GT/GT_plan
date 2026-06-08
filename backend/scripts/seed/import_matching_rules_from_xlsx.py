"""从「年度审计报告模板使用对照表.xlsx」导入企业子类型推荐规则。

产出 ``backend/data/audit_report_templates/matching_rules.json``，供
``MatchingRulesService.recommend_company_subtype`` 加载。

对照表本身是一张人工阅读的二维矩阵（公司类型 × 意见类型 → 模板A/B/C/D），
难以直接、稳定地解析为可执行的匹配条件，因此本脚本采用「读 xlsx 校验类别
存在 + 手工编码匹配条件」的策略：从 xlsx 中提取四类模板（A/B/C/D）的人工
说明文本作为规则描述/关键词来源，再补充可匹配的项目属性条件
（entity_type / scenario / industry 关键词）。

需求映射（requirements §1.5 / §7）：
  - A (type_a): 上市公司、三板创新层及公开发债
  - B (type_b): 三板基础层、银行、保险、期货、证券
  - C (type_c): 其他公众利益实体（含中央企业集团公司）
  - D (type_d): 非公众利益实体（其他企业）

用法（cwd=backend）：
    ..\\.venv\\Scripts\\python.exe scripts/seed/import_matching_rules_from_xlsx.py

Validates: Requirements 7.1, 7.3, 7.4
"""

from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - openpyxl 必装
    load_workbook = None  # type: ignore


# ── 路径常量 ───────────────────────────────────────────────────────────────
_TEMPLATE_DIR = Path(__file__).resolve().parents[2] / "data" / "audit_report_templates"
_XLSX_PATH = _TEMPLATE_DIR / "年度审计报告模板使用对照表.xlsx"
_OUTPUT_PATH = _TEMPLATE_DIR / "matching_rules.json"


def _inspect_xlsx() -> dict[str, list[str]]:
    """读取对照表，提取四类模板的人工说明文本（用于规则描述/关键词补充）。

    返回 {"A": [...文本...], "B": [...], "C": [...], "D": [...]}；
    若 openpyxl 不可用或文件缺失则返回空 dict（不阻断手工规则生成）。
    """
    found: dict[str, list[str]] = {"A": [], "B": [], "C": [], "D": []}
    if load_workbook is None or not _XLSX_PATH.exists():
        return found

    wb = load_workbook(str(_XLSX_PATH), data_only=True)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            text = val.strip()
            if not text:
                continue
            # 模板说明文本形如「模板A【简、详…】」
            for letter in ("A", "B", "C", "D"):
                if text.startswith(f"模板{letter}"):
                    found[letter].append(text)
    return found


def build_rules() -> dict:
    """构造 matching_rules.json 的内容（手工编码 + xlsx 说明补充）。"""
    descriptions = _inspect_xlsx()

    def _first_desc(letter: str, default: str) -> str:
        items = descriptions.get(letter) or []
        return items[0] if items else default

    rules = [
        {
            "id": "A_listed_public_debt",
            "subtype": "type_a",
            "label": "模板A",
            "description": "上市公司、三板创新层及公开发债",
            "priority": 40,
            "match": {
                "scenario": ["listed", "ipo"],
                "entity_type": ["listed"],
                "template_type": ["listed"],
                "keywords": [
                    "上市", "创新层", "三板创新", "公开发行债券",
                    "公开发债", "发债", "公开发行公司债",
                ],
            },
            "source_note": _first_desc("A", "模板A 上市/创新层/公开发债"),
        },
        {
            "id": "B_basic_layer_financial",
            "subtype": "type_b",
            "label": "模板B",
            "description": "三板基础层、银行、保险、期货、证券",
            "priority": 40,
            "match": {
                "scenario": [],
                "entity_type": [],
                "template_type": [],
                "keywords": [
                    "基础层", "三板基础", "银行", "保险", "期货",
                    "证券", "农村信用", "村镇银行", "合作银行",
                ],
            },
            "source_note": _first_desc("B", "模板B 三板基础层/银行/保险/期货/证券"),
        },
        {
            "id": "C_other_pie",
            "subtype": "type_c",
            "label": "模板C",
            "description": "其他公众利益实体（含中央企业集团公司）",
            "priority": 30,
            "match": {
                "scenario": [],
                "entity_type": [],
                "template_type": [],
                "keywords": [
                    "中央企业", "央企", "中央企业集团", "公众利益实体",
                    "国有资产监督管理", "吸收公众存款", "公众",
                ],
            },
            "source_note": _first_desc("C", "模板C 其他公众利益实体"),
        },
        {
            # D = 非公众利益实体，是「兜底类别」：不以宽泛的 entity_type=private /
            # scenario=normal 作为正向规则（否则几乎所有项目都会与更具体的 A/B/C
            # 关键词规则产生歧义）。D 主要通过 requirements §1.4 fallback
            # （non_listed → type_d）到达；仅在出现明确「非公众」字样时作为规则命中。
            "id": "D_non_pie",
            "subtype": "type_d",
            "label": "模板D",
            "description": "非公众利益实体（其他企业）",
            "priority": 10,
            "match": {
                "scenario": [],
                "entity_type": [],
                "template_type": [],
                "keywords": ["非公众利益实体", "非公众"],
            },
            "source_note": _first_desc("D", "模板D 非公众利益实体"),
        },
    ]

    return {
        "version": "2025-v1",
        "source": _XLSX_PATH.name,
        "description": (
            "企业子类型（type_a/b/c/d）推荐规则。规则按 priority 从高到低评估，"
            "匹配项目属性（entity_type / scenario / template_type / industry / "
            "company_name 关键词）。无匹配时按 fallback 兜底（requirements §1.4）。"
        ),
        "fallback": {"listed": "type_a", "non_listed": "type_d"},
        "rules": rules,
    }


def main() -> None:
    payload = build_rules()
    _OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    n = len(payload["rules"])
    print(f"[import_matching_rules] 已写入 {_OUTPUT_PATH} （{n} 条规则）")
    # 回显 xlsx 提取到的说明文本数量，确认确实读到了对照表
    desc = _inspect_xlsx()
    summary = {k: len(v) for k, v in desc.items()}
    print(f"[import_matching_rules] xlsx 模板说明提取计数: {summary}")


if __name__ == "__main__":
    main()
