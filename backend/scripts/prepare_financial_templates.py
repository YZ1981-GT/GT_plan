#!/usr/bin/env python3
"""泛化报表模板占位整理（4 变体 × 主表 sheet），数据驱动自 report_config_seed.json.

Task 12.1a（泛化自 prepare_financial_template_poc.py）。

为 4 个财报模板（soe_standalone / soe_consolidated / listed_standalone /
listed_consolidated）注入内联占位符：

- **双列（standalone 物理 2 列）**：``{{row:CODE:current}}`` → 期末/本期，
  ``{{row:CODE:prior}}`` → 期初/上期。
- **四列（consolidated / 物理含「合并/公司」子表头）**：合并列写
  ``{{row:CODE:current}}`` / ``{{row:CODE:prior}}``；**仅 consolidated 变体**额外
  在公司（母公司个别）列写 ``{{row:CODE:current:parent}}`` /
  ``{{row:CODE:prior:parent}}``。standalone 变体即使模板物理有公司列也不写
  ``:parent``（个别报表无母公司个别数概念，公司列留空）。

域规则（与 design §6 / §9 一致）：
- **按表头文本检测列**（期末余额/期初余额；本期金额/上期金额；本年金额/上年金额），
  不硬编码列序号；consolidated 读子表头「合并/公司」拆列。
- **绝不写入 附注列（col 2）或指引提示列**——只写已检测到的数值数据列。
- 跳过公式格（``data_type=='f'`` 或值以 ``=`` 开头）。
- 仅写入空格或已是同 code ``{{row:}}`` 占位的格（幂等，绝不覆盖正文/公式/其他占位）。
- 按规范化 row_name 顺序匹配模板 A 列（``_norm_row_name``：去空白 + 去前导 △▲#）；
  跳过段标题行（流动资产：等，名称以 ：/: 结尾）。
- balance_sheet 的主表 + 续表共用 ``balance_sheet`` seed（跨两 sheet 各匹配子集）。
- equity_statement / impairment_provision 为矩阵表（本年金额 × 权益构成列 /
  年初·本期增加·本期减少 列），**不适用 current/prior 方案**，本脚本跳过并在
  汇总中报告为「未映射（矩阵结构）」。

Usage:
    python backend/scripts/prepare_financial_templates.py --dry-run
    python backend/scripts/prepare_financial_templates.py --write
    python backend/scripts/prepare_financial_templates.py --variant soe_consolidated --write
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
TPL = _BACKEND / "data" / "audit_report_templates" / "financial_statements"
SEED = _BACKEND / "data" / "report_config_seed.json"

VARIANTS = [
    "soe_standalone",
    "soe_consolidated",
    "listed_standalone",
    "listed_consolidated",
]

# 注入的主表（矩阵表 equity/impairment 不在内）
_INJECT_REPORT_TYPES = {
    "balance_sheet",
    "income_statement",
    "cash_flow_statement",
}

# sheet 名关键字 → report_type
_SHEET_KEYWORDS = [
    ("资产负债表", "balance_sheet"),
    ("利润表", "income_statement"),
    ("现金流量表", "cash_flow_statement"),
    ("权益变动表", "equity_statement"),
    ("减值准备", "impairment_provision"),
]

_CURRENT_KW = ("期末余额", "本期金额", "本年金额")
_PRIOR_KW = ("期初余额", "上年年末余额", "上期金额", "上年金额")

_ROW_PH_RE = re.compile(r"^\{\{row:([^:}]+):(current|prior)(?::parent)?\}\}$")


def _norm_row_name(text: str) -> str:
    s = re.sub(r"\s+", "", (text or ""))
    return re.sub(r"^[△▲#*]+", "", s)


def _report_type_for_sheet(sheet_name: str) -> str | None:
    for kw, rt in _SHEET_KEYWORDS:
        if kw in sheet_name:
            return rt
    return None


def _load_seed_rows(variant: str, report_type: str) -> list[dict]:
    configs = json.loads(SEED.read_text(encoding="utf-8"))
    for block in configs:
        if (
            block.get("report_type") == report_type
            and block.get("applicable_standard") == variant
        ):
            return block.get("rows", [])
    return []


def _find_period_header_row(ws, max_scan: int = 8) -> int | None:
    """定位含「期末/本期/本年」与「期初/上期/上年」关键字的表头行。"""
    best_row = None
    best_score = 0
    for r in range(1, min(max_scan, ws.max_row) + 1):
        score = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            if any(k in v for k in _CURRENT_KW) or any(k in v for k in _PRIOR_KW):
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score >= 1 else None


def _classify_period_col(text: str) -> str | None:
    if any(k in text for k in _CURRENT_KW):
        return "current"
    if any(k in text for k in _PRIOR_KW):
        return "prior"
    return None


def _detect_columns(ws, header_row: int) -> dict | None:
    """检测 current/prior 数据列；如有「合并/公司」子表头则拆 main/parent.

    返回::
        {
          "current": col, "current_parent": col|None,
          "prior": col, "prior_parent": col|None,
          "data_cols": set(all numeric data columns to allow writing),
        }
    """
    # 1) 标记 current/prior 区块边界
    blocks: list[tuple[int, str]] = []  # (start_col, period)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str):
            period = _classify_period_col(v)
            if period:
                blocks.append((c, period))
    if not blocks:
        return None

    # 区块范围：每个区块从其起始列延伸到下一个区块起始列前 / max_column
    ranges: list[tuple[str, int, int]] = []  # (period, start, end_inclusive)
    for i, (start, period) in enumerate(blocks):
        end = (blocks[i + 1][0] - 1) if i + 1 < len(blocks) else ws.max_column
        ranges.append((period, start, end))

    # 2) 子表头行（header_row+1）是否含 合并/公司
    sub_row = header_row + 1
    has_sub = False
    sub_vals: dict[int, str] = {}
    if sub_row <= ws.max_row:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(sub_row, c).value
            if isinstance(v, str) and ("合并" in v or "公司" in v):
                sub_vals[c] = v
                has_sub = True

    result = {
        "current": None,
        "current_parent": None,
        "prior": None,
        "prior_parent": None,
        "data_cols": set(),
    }

    for period, start, end in ranges:
        if has_sub:
            main_col = None
            parent_col = None
            for c in range(start, end + 1):
                sv = sub_vals.get(c)
                if sv is None:
                    continue
                if "合并" in sv and main_col is None:
                    main_col = c
                elif "公司" in sv and parent_col is None:
                    parent_col = c
            # 回退：子表头缺失时主列取区块起始列
            if main_col is None:
                main_col = start
            result[period] = main_col
            result[f"{period}_parent"] = parent_col
            for c in (main_col, parent_col):
                if c:
                    result["data_cols"].add(c)
        else:
            result[period] = start
            result["data_cols"].add(start)

    if result["current"] is None and result["prior"] is None:
        return None
    return result


def _build_name_to_row(ws) -> dict[str, list[int]]:
    """A 列规范名 → 行号列表（同名多行保序）。"""
    mapping: dict[str, list[int]] = defaultdict(list)
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        key = _norm_row_name(str(val))
        if key:
            mapping[key].append(r)
    return mapping


def _is_section_header(row_name: str) -> bool:
    s = (row_name or "").strip()
    return s.endswith("：") or s.endswith(":")


def _is_formula_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row, col)
    if getattr(cell, "data_type", None) == "f":
        return True
    val = cell.value
    return isinstance(val, str) and val.startswith("=")


def _can_write(ws, row: int, col: int, code: str) -> bool:
    """仅当格为空 / 已是本 code 占位 / 已是任意 {{row:}} 占位时允许写（幂等）。"""
    if _is_formula_cell(ws, row, col):
        return False
    val = ws.cell(row, col).value
    if val is None:
        return True
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return True
        m = _ROW_PH_RE.match(s)
        if m and m.group(1) == code:
            return True
    return False


def _inject_header(ws, write: bool, stats: dict) -> None:
    """注入表头占位：{{period_end_date}}（标题日期区）+ 编制单位：{{company_full_name}}。"""
    # 编制单位行：current text 含「编制单位」
    for r in range(1, min(6, ws.max_row) + 1):
        for c in range(1, min(3, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "编制单位" in v and "{{company_full_name}}" not in v:
                target = "编制单位：{{company_full_name}}"
                if v != target:
                    stats["headers"] += 1
                    if write:
                        ws.cell(r, c).value = target
    # 期间日期行：row2/row3 含纯日期（数字序列号或 年..月..日）或已是占位
    for r in (2, 3):
        if r > ws.max_row:
            continue
        for c in range(1, min(3, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if _is_formula_cell(ws, r, c):
                continue
            if v is None:
                continue
            s = str(v).strip()
            is_date_like = bool(re.search(r"\d{4}年.*月.*日", s)) or s.isdigit()
            if is_date_like and "{{" not in s:
                stats["headers"] += 1
                if write:
                    ws.cell(r, c).value = "{{period_end_date}}"


def _inject_sheet(
    ws,
    report_type: str,
    seed_rows: list[dict],
    *,
    allow_parent: bool,
    write: bool,
    stats: dict,
) -> None:
    header_row = _find_period_header_row(ws)
    if header_row is None:
        stats["skipped_sheets"].append((ws.title, "no period header row"))
        return
    cols = _detect_columns(ws, header_row)
    if cols is None:
        stats["skipped_sheets"].append((ws.title, "no current/prior columns"))
        return

    name_to_rows = _build_name_to_row(ws)
    used_rows: set[int] = set()

    for cfg in seed_rows:
        code = cfg.get("row_code")
        raw_name = cfg.get("row_name", "")
        if _is_section_header(raw_name):
            continue
        key = _norm_row_name(raw_name)
        if not key:
            continue
        candidates = [r for r in name_to_rows.get(key, []) if r not in used_rows]
        if not candidates:
            continue
        excel_row = candidates[0]
        used_rows.add(excel_row)

        for period in ("current", "prior"):
            main_col = cols.get(period)
            if main_col:
                token = f"{{{{row:{code}:{period}}}}}"
                if _can_write(ws, excel_row, main_col, code):
                    if ws.cell(excel_row, main_col).value != token:
                        stats["row_ph"] += 1
                        if write:
                            ws.cell(excel_row, main_col).value = token
                else:
                    stats["skipped_cells"] += 1
            # 公司列：仅 consolidated 变体
            parent_col = cols.get(f"{period}_parent")
            if allow_parent and parent_col:
                ptoken = f"{{{{row:{code}:{period}:parent}}}}"
                if _can_write(ws, excel_row, parent_col, code):
                    if ws.cell(excel_row, parent_col).value != ptoken:
                        stats["row_ph_parent"] += 1
                        if write:
                            ws.cell(excel_row, parent_col).value = ptoken
                else:
                    stats["skipped_cells"] += 1


def prepare_variant(variant: str, *, write: bool) -> dict:
    path = TPL / f"{variant}.xlsx"
    if not path.is_file():
        raise FileNotFoundError(path)
    allow_parent = variant.endswith("_consolidated")
    wb = openpyxl.load_workbook(path)
    stats: dict = {
        "variant": variant,
        "headers": 0,
        "row_ph": 0,
        "row_ph_parent": 0,
        "skipped_cells": 0,
        "skipped_sheets": [],
        "per_sheet": {},
        "unmapped_matrix": [],
    }

    for name in wb.sheetnames:
        if name == "GT_Custom":
            continue
        ws = wb[name]
        rt = _report_type_for_sheet(name)
        if rt in ("equity_statement", "impairment_provision"):
            stats["unmapped_matrix"].append((name, rt))
            continue
        if rt not in _INJECT_REPORT_TYPES:
            continue
        before = stats["row_ph"] + stats["row_ph_parent"]
        _inject_header(ws, write, stats)
        seed_rows = _load_seed_rows(variant, rt)
        if not seed_rows:
            stats["skipped_sheets"].append((name, f"no seed for {rt}"))
            continue
        _inject_sheet(
            ws, rt, seed_rows, allow_parent=allow_parent, write=write, stats=stats
        )
        stats["per_sheet"][name] = (stats["row_ph"] + stats["row_ph_parent"]) - before

    if write:
        wb.save(path)
    return stats


def _print_stats(stats: dict) -> None:
    print(f"\n[{stats['variant']}]")
    print(
        f"  headers={stats['headers']} row_ph(main)={stats['row_ph']} "
        f"row_ph(parent)={stats['row_ph_parent']} skipped_cells={stats['skipped_cells']}"
    )
    for name, n in stats["per_sheet"].items():
        print(f"    + {name!r}: {n} placeholders injected")
    for name, rt in stats["unmapped_matrix"]:
        print(f"    ~ unmapped (matrix structure): {name!r} [{rt}]")
    for name, reason in stats["skipped_sheets"]:
        print(f"    ! skipped sheet {name!r}: {reason}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare financial templates (all variants)")
    parser.add_argument("--variant", choices=VARIANTS)
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if not args.write:
        args.dry_run = True

    targets = [args.variant] if args.variant else VARIANTS
    print(f"prepare_financial_templates @ {_dt.datetime.now():%Y-%m-%d %H:%M:%S} "
          f"(write={args.write})")
    grand = {"row_ph": 0, "row_ph_parent": 0, "headers": 0}
    for variant in targets:
        try:
            stats = prepare_variant(variant, write=args.write)
        except FileNotFoundError as e:
            print(f"  MISSING: {e}", file=sys.stderr)
            continue
        _print_stats(stats)
        grand["row_ph"] += stats["row_ph"]
        grand["row_ph_parent"] += stats["row_ph_parent"]
        grand["headers"] += stats["headers"]

    print(
        f"\nGRAND TOTAL: headers={grand['headers']} "
        f"row_ph(main)={grand['row_ph']} row_ph(parent)={grand['row_ph_parent']}"
    )
    if not args.write:
        print("(dry-run; pass --write to apply)")


if __name__ == "__main__":
    main()
