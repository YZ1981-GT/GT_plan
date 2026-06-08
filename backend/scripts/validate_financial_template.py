#!/usr/bin/env python3
"""校验财务报表模板：公式格无占位符、无 GT_Custom 数据污染.

Task 0.5.2

Usage:
    python backend/scripts/validate_financial_template.py --variant soe_standalone
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
TPL = _BACKEND / "data" / "audit_report_templates" / "financial_statements"

PLACEHOLDER_RE = re.compile(r"\{\{[^}]+\}\}")


def validate(path: Path) -> list[str]:
    issues: list[str] = []
    wb = openpyxl.load_workbook(path, data_only=False)
    placeholder_count = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == "GT_Custom":
            continue
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if not isinstance(val, str):
                    continue
                if not PLACEHOLDER_RE.search(val):
                    continue
                placeholder_count += 1
                if val.strip().startswith("=") or (
                    isinstance(ws.cell(r, c).value, str)
                    and str(ws.cell(r, c).value).startswith("=")
                ):
                    issues.append(
                        f"{sheet_name}!{r},{c} placeholder in formula cell"
                    )
                if val.startswith("=SUM(") and "{{" in val:
                    issues.append(f"{sheet_name}!{r},{c} placeholder inside SUM")

    if placeholder_count < 20:
        issues.append(f"expected >=20 placeholders, found {placeholder_count}")

    return issues


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--variant", default="soe_standalone")
    args = parser.parse_args()
    path = TPL / f"{args.variant}.xlsx"
    issues = validate(path)
    if issues:
        print(f"FAIL {path.name}:")
        for i in issues:
            print(f"  - {i}")
        raise SystemExit(1)
    print(f"OK {path.name}")


if __name__ == "__main__":
    main()
