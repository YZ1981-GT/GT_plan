# -*- coding: utf-8 -*-
"""重庆和平药房四表数据导入脚本

处理特殊格式：
1. 科目余额表：双行合并表头 + 核算维度列 → tb_balance + tb_aux_balance
2. 序时账：两个文件合并 + 核算维度列 → tb_ledger + tb_aux_ledger
3. 核算维度解析：多维度拆分（金融机构:YG0001,工商银行; 银行账户:xxx; 成本中心:xxx）

用法：
  python scripts/import_heping_data.py --project-id <UUID> [--year 2025] [--dry-run]
"""

import asyncio
import argparse
import hashlib
import re
import sys
import uuid as _uuid
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

# 添加 backend 到 sys.path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# 核算维度解析
# ─────────────────────────────────────────────────────────────────────────────

# 用于给没有编码的辅助核算名称生成唯一占位编码
_AUX_CODE_CACHE: dict[str, str] = {}  # (aux_type, aux_name) -> generated_code


def _generate_aux_code(aux_type: str, aux_name: str) -> str:
    """为没有编码的辅助核算名称生成唯一占位编码（不重复且名称唯一）。"""
    key = f"{aux_type}:{aux_name}"
    if key in _AUX_CODE_CACHE:
        return _AUX_CODE_CACHE[key]
    # 用 hash 前8位 + 类型前缀生成
    prefix = aux_type[:2].upper() if aux_type else "XX"
    h = hashlib.md5(key.encode()).hexdigest()[:6].upper()
    code = f"AUTO_{prefix}_{h}"
    _AUX_CODE_CACHE[key] = code
    return code


def parse_aux_dimensions(dim_str: str) -> list[dict]:
    """解析核算维度字符串为多条辅助核算记录。

    输入格式示例：
      "金融机构:YG0001,工商银行; 银行账户:3100021509024876090; 成本中心:YG130108,财务部"
      "客户:111111,个人客户"
      "成本中心:459125,兴塘路店"

    返回：
      [
        {"aux_type": "金融机构", "aux_code": "YG0001", "aux_name": "工商银行"},
        {"aux_type": "银行账户", "aux_code": "AUTO_YI_xxx", "aux_name": "3100021509024876090"},
        {"aux_type": "成本中心", "aux_code": "YG130108", "aux_name": "财务部"},
      ]
    """
    if not dim_str or not dim_str.strip():
        return []

    results = []
    # 按 ; 或 ；分割多个维度
    parts = re.split(r'[;；]', dim_str.strip())

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # 格式：类型:编码,名称  或  类型:名称（无编码）
        colon_idx = part.find(':')
        if colon_idx < 0:
            colon_idx = part.find('：')
        if colon_idx < 0:
            continue

        aux_type = part[:colon_idx].strip()
        value_part = part[colon_idx + 1:].strip()

        if not aux_type or not value_part:
            continue

        # 尝试拆分 编码,名称
        comma_idx = value_part.find(',')
        if comma_idx < 0:
            comma_idx = value_part.find('，')

        if comma_idx > 0:
            aux_code = value_part[:comma_idx].strip()
            aux_name = value_part[comma_idx + 1:].strip()
        else:
            # 没有逗号 → 整个值可能是编码也可能是名称
            # 如果全是数字/字母，当作编码（名称=编码）
            # 否则当作名称，自动生成编码
            if re.match(r'^[A-Za-z0-9\-_.]+$', value_part):
                aux_code = value_part
                aux_name = value_part
            else:
                aux_name = value_part
                aux_code = _generate_aux_code(aux_type, aux_name)

        results.append({
            "aux_type": aux_type,
            "aux_code": aux_code,
            "aux_name": aux_name,
        })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# 科目余额表解析（双行合并表头）
# ─────────────────────────────────────────────────────────────────────────────

# 余额表列映射（基于实际表头）
# Row 3: 科目编码(1) 科目名称(2) 核算维度(3) 组织编码(4) 年初余额(5) 期初余额(7) 本期发生额(9) 本年累计(11) 期末余额(13)
# Row 4:                                          借方金额(5) 贷方金额(6) 借方金额(7) 贷方金额(8) 借方金额(9) 贷方金额(10) 借方金额(11) 贷方金额(12) 借方金额(13) 贷方金额(14)

_BALANCE_COL_MAP = {
    1: "account_code",
    2: "account_name",
    3: "aux_dimensions",  # 核算维度 → 需要特殊处理
    4: "company_code",
    # 年初余额
    5: "opening_debit",
    6: "opening_credit",
    # 期初余额
    7: "period_opening_debit",
    8: "period_opening_credit",
    # 本期发生额
    9: "debit_amount",
    10: "credit_amount",
    # 本年累计
    11: "year_debit",
    12: "year_credit",
    # 期末余额
    13: "closing_debit",
    14: "closing_credit",
}


def _safe_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        d = Decimal(str(val))
        return d if d != 0 else None
    except (InvalidOperation, ValueError):
        return None


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _infer_level(code: str) -> int:
    """从科目编码推断级次。"""
    code = code.replace(".", "").replace("-", "")
    if len(code) <= 4:
        return 1
    elif len(code) <= 6:
        return 2
    elif len(code) <= 8:
        return 3
    elif len(code) <= 10:
        return 4
    return 5


def parse_balance_file(filepath: str) -> tuple[list[dict], list[dict]]:
    """解析科目余额表 → (tb_balance_rows, tb_aux_balance_rows)"""
    print(f"[BALANCE] 解析: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    balance_rows = []
    aux_balance_rows = []
    data_start_row = 5  # 数据从第5行开始（前4行是标题+表头）

    for row_idx in range(data_start_row, ws.max_row + 1):
        account_code = _safe_str(ws.cell(row_idx, 1).value)
        if not account_code:
            continue

        account_name = _safe_str(ws.cell(row_idx, 2).value)
        aux_dim_str = _safe_str(ws.cell(row_idx, 3).value)
        company_code = _safe_str(ws.cell(row_idx, 4).value) or "default"

        opening_debit = _safe_decimal(ws.cell(row_idx, 5).value)
        opening_credit = _safe_decimal(ws.cell(row_idx, 6).value)
        debit_amount = _safe_decimal(ws.cell(row_idx, 9).value)
        credit_amount = _safe_decimal(ws.cell(row_idx, 10).value)
        closing_debit = _safe_decimal(ws.cell(row_idx, 13).value)
        closing_credit = _safe_decimal(ws.cell(row_idx, 14).value)

        # 计算净额（借方为正，贷方为负）
        opening_balance = (opening_debit or Decimal(0)) - (opening_credit or Decimal(0))
        closing_balance = (closing_debit or Decimal(0)) - (closing_credit or Decimal(0))

        if aux_dim_str:
            # 有核算维度 → 拆分为辅助余额表记录
            dims = parse_aux_dimensions(aux_dim_str)
            for dim in dims:
                aux_balance_rows.append({
                    "account_code": account_code,
                    "account_name": account_name,
                    "aux_type": dim["aux_type"],
                    "aux_code": dim["aux_code"],
                    "aux_name": dim["aux_name"],
                    "company_code": company_code,
                    "opening_balance": opening_balance,
                    "debit_amount": debit_amount,
                    "credit_amount": credit_amount,
                    "closing_balance": closing_balance,
                    "currency_code": "CNY",
                })
        else:
            # 无核算维度 → 余额表记录
            balance_rows.append({
                "account_code": account_code,
                "account_name": account_name,
                "level": _infer_level(account_code),
                "company_code": company_code,
                "opening_balance": opening_balance,
                "debit_amount": debit_amount,
                "credit_amount": credit_amount,
                "closing_balance": closing_balance,
                "currency_code": "CNY",
            })

    wb.close()
    print(f"[BALANCE] 余额表: {len(balance_rows)} 行, 辅助余额表: {len(aux_balance_rows)} 行")
    return balance_rows, aux_balance_rows


# ─────────────────────────────────────────────────────────────────────────────
# 序时账解析（支持多文件合并 + 核算维度拆分）
# ─────────────────────────────────────────────────────────────────────────────

def _parse_period(period_str: str) -> Optional[int]:
    """解析期间字符串为会计月份。如 '2025年1期' → 1"""
    m = re.search(r'(\d+)\s*期', str(period_str))
    return int(m.group(1)) if m else None


def _parse_date_val(val) -> Optional[date]:
    """解析日期值。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_ledger_files(filepaths: list[str]) -> tuple[list[dict], list[dict]]:
    """解析序时账文件（支持多文件合并） → (tb_ledger_rows, tb_aux_ledger_rows)"""
    ledger_rows = []
    aux_ledger_rows = []

    for filepath in filepaths:
        print(f"[LEDGER] 解析: {filepath}")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        row_count = 0
        is_header = True
        for row in ws.iter_rows(values_only=True):
            if is_header:
                is_header = False
                continue  # 跳过表头行

            # 列映射（基于实际表头）
            # 1:序号 2:核算组织 3:账簿类型 4:期间 5:记账日期 6:凭证类型 7:凭证号
            # 8:状态 9:过账 10:摘要 11:科目编码 12:科目名称 13:借方 14:贷方
            # 15:核算维度 16:主表项目 17:补充资料 18:最后操作人 ...
            if len(row) < 15:
                continue

            account_code = _safe_str(row[10])  # col 11 (0-indexed: 10)
            if not account_code:
                continue

            voucher_date = _parse_date_val(row[4])  # col 5
            voucher_no = _safe_str(row[6])  # col 7
            if not voucher_date or not voucher_no:
                continue

            account_name = _safe_str(row[11])  # col 12
            period_str = _safe_str(row[3])  # col 4
            accounting_period = _parse_period(period_str)
            voucher_type = _safe_str(row[5])  # col 6
            debit_amount = _safe_decimal(row[12])  # col 13
            credit_amount = _safe_decimal(row[13])  # col 14
            aux_dim_str = _safe_str(row[14]) if len(row) > 14 else ""  # col 15
            summary = _safe_str(row[9])  # col 10
            preparer = _safe_str(row[22]) if len(row) > 22 else ""  # col 23 制单人

            # 序时账主记录（始终写入）
            ledger_row = {
                "account_code": account_code,
                "account_name": account_name,
                "voucher_date": voucher_date,
                "voucher_no": voucher_no,
                "voucher_type": voucher_type,
                "accounting_period": accounting_period,
                "debit_amount": debit_amount,
                "credit_amount": credit_amount,
                "summary": summary,
                "preparer": preparer,
                "company_code": "default",
                "currency_code": "CNY",
            }
            ledger_rows.append(ledger_row)

            # 有核算维度 → 同时生成辅助明细账记录
            if aux_dim_str:
                dims = parse_aux_dimensions(aux_dim_str)
                for dim in dims:
                    aux_ledger_rows.append({
                        "account_code": account_code,
                        "account_name": account_name,
                        "aux_type": dim["aux_type"],
                        "aux_code": dim["aux_code"],
                        "aux_name": dim["aux_name"],
                        "voucher_date": voucher_date,
                        "voucher_no": voucher_no,
                        "voucher_type": voucher_type,
                        "accounting_period": accounting_period,
                        "debit_amount": debit_amount,
                        "credit_amount": credit_amount,
                        "summary": summary,
                        "preparer": preparer,
                        "company_code": "default",
                        "currency_code": "CNY",
                    })

            row_count += 1
            if row_count % 100000 == 0:
                print(f"  已处理 {row_count} 行...")

        wb.close()
        print(f"[LEDGER] {filepath}: {row_count} 行")

    print(f"[LEDGER] 合计: 序时账 {len(ledger_rows)} 行, 辅助明细账 {len(aux_ledger_rows)} 行")
    return ledger_rows, aux_ledger_rows



# ─────────────────────────────────────────────────────────────────────────────
# 数据库写入
# ─────────────────────────────────────────────────────────────────────────────

async def write_to_db(
    project_id: _uuid.UUID,
    year: int,
    balance_rows: list[dict],
    aux_balance_rows: list[dict],
    ledger_rows: list[dict],
    aux_ledger_rows: list[dict],
    dry_run: bool = False,
):
    """将解析好的四表数据写入数据库。"""
    from app.core.database import async_session
    from app.models.audit_platform_models import (
        TbBalance, TbLedger, TbAuxBalance, TbAuxLedger,
        ImportBatch, ImportStatus,
    )
    import sqlalchemy as sa

    print(f"\n[统计]")
    print(f"  tb_balance:     {len(balance_rows)} 行")
    print(f"  tb_aux_balance: {len(aux_balance_rows)} 行")
    print(f"  tb_ledger:      {len(ledger_rows)} 行")
    print(f"  tb_aux_ledger:  {len(aux_ledger_rows)} 行")
    # 展示辅助核算维度统计
    aux_types = {}
    for r in aux_balance_rows + aux_ledger_rows:
        t = r.get("aux_type", "?")
        aux_types[t] = aux_types.get(t, 0) + 1
    print(f"\n  辅助核算维度分布:")
    for t, c in sorted(aux_types.items(), key=lambda x: -x[1]):
        print(f"    {t}: {c} 条")

    if dry_run:
        print("\n[DRY RUN] 仅解析不写入数据库。")
        return

    CHUNK_SIZE = 10000

    async with async_session() as db:
        try:
            # 写入四表
            table_data = [
                ("tb_balance", TbBalance, balance_rows),
                ("tb_aux_balance", TbAuxBalance, aux_balance_rows),
                ("tb_ledger", TbLedger, ledger_rows),
                ("tb_aux_ledger", TbAuxLedger, aux_ledger_rows),
            ]

            for data_type, model, rows in table_data:
                if not rows:
                    print(f"[DB] {data_type}: 无数据，跳过")
                    continue

                # 创建导入批次
                batch = ImportBatch(
                    project_id=project_id,
                    year=year,
                    source_type="script",
                    file_name=f"heping_{data_type}",
                    data_type=data_type,
                    status=ImportStatus.processing,
                    started_at=datetime.utcnow(),
                )
                db.add(batch)
                await db.flush()

                # 软删除旧数据
                tbl = model.__table__
                await db.execute(
                    sa.update(tbl)
                    .where(
                        tbl.c.project_id == project_id,
                        tbl.c.year == year,
                        tbl.c.is_deleted == sa.false(),
                    )
                    .values(is_deleted=True)
                )

                # 批量写入
                record_count = 0
                for i in range(0, len(rows), CHUNK_SIZE):
                    chunk = rows[i:i + CHUNK_SIZE]
                    recs = []
                    for row in chunk:
                        rec = {
                            "id": _uuid.uuid4(),
                            "project_id": project_id,
                            "year": year,
                            "import_batch_id": batch.id,
                            "is_deleted": False,
                            **row,
                        }
                        recs.append(rec)
                    if recs:
                        await db.execute(tbl.insert(), recs)
                        record_count += len(recs)

                    if (i + CHUNK_SIZE) % 50000 == 0:
                        print(f"  {data_type}: 已写入 {record_count} 行...")

                batch.record_count = record_count
                batch.status = ImportStatus.completed
                batch.completed_at = datetime.utcnow()
                print(f"[DB] {data_type}: 写入 {record_count} 行")

            await db.commit()
            print("\n[DB] 全部写入完成！")

        except Exception as e:
            await db.rollback()
            print(f"\n[DB] 写入失败: {e}")
            import traceback
            traceback.print_exc()
            raise


# ─────────────────────────────────────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description="重庆和平药房四表数据导入")
    parser.add_argument("--project-id", required=True, help="项目UUID")
    parser.add_argument("--year", type=int, default=2025, help="审计年度（默认2025）")
    parser.add_argument("--dry-run", action="store_true", help="仅解析不写入数据库")
    parser.add_argument("--balance-file", default=None, help="科目余额表文件路径")
    parser.add_argument("--ledger-files", nargs="+", default=None, help="序时账文件路径（支持多个）")
    args = parser.parse_args()

    project_id = _uuid.UUID(args.project_id)
    year = args.year

    # 默认文件路径
    base_dir = Path(__file__).resolve().parent.parent.parent  # 项目根目录
    balance_file = args.balance_file or str(
        base_dir / "数据" / "科目余额表-重庆和平药房连锁有限责任公司2025.xlsx"
    )
    ledger_files = args.ledger_files or [
        str(base_dir / "数据" / "25年序时账" / "序时账-重庆和平药房连锁有限责任公司 20250101-1011.xlsx"),
        str(base_dir / "数据" / "25年序时账" / "序时账-重庆和平药房连锁有限责任公司 20251012-1031.xlsx"),
    ]

    print("=" * 60)
    print("重庆和平药房四表数据导入")
    print(f"  项目ID: {project_id}")
    print(f"  年度: {year}")
    print(f"  余额表: {balance_file}")
    print(f"  序时账: {ledger_files}")
    print(f"  模式: {'DRY RUN' if args.dry_run else '写入数据库'}")
    print("=" * 60)

    # 1. 解析科目余额表
    balance_rows, aux_balance_rows = parse_balance_file(balance_file)

    # 2. 解析序时账（多文件合并）
    ledger_rows, aux_ledger_rows = parse_ledger_files(ledger_files)

    # 3. 写入数据库
    await write_to_db(
        project_id, year,
        balance_rows, aux_balance_rows,
        ledger_rows, aux_ledger_rows,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    asyncio.run(main())
