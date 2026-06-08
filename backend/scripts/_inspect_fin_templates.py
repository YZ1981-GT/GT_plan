#!/usr/bin/env python3
"""临时检查：打印 4 个财务报表模板的 sheet 名 + 前几行各列文本，验证列布局。用完即删。"""
from __future__ import annotations

from pathlib import Path
import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
TPL = _BACKEND / "data" / "audit_report_templates" / "financial_statements"

VARIANTS = ["soe_standalone", "soe_consolidated", "listed_standalone", "listed_consolidated"]


def main() -> None:
    for variant in VARIANTS:
        path = TPL / f"{variant}.xlsx"
        wb = openpyxl.load_workbook(path, data_only=False)
        print(f"\n{'='*70}\n{variant}  sheets={wb.sheetnames}\n{'='*70}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n--- sheet [{sn}]  max_row={ws.max_row} max_col={ws.max_column}")
            for r in range(1, min(8, ws.max_row + 1)):
                cells = []
                for c in range(1, min(9, ws.max_column + 1)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        s = str(v).replace("\n", " ")[:22]
                        cells.append(f"{openpyxl.utils.get_column_letter(c)}{r}={s}")
                if cells:
                    print("   " + " | ".join(cells))


if __name__ == "__main__":
    main()
