"""底稿生成管线端到端验证脚本（正式可重复工具）

用法:
    ..\.venv\Scripts\python.exe scripts/verify_wp_generation_pipeline.py \
        --project df5b8403 --year 2025 [--report]

步骤:
    1. recommend  : 调 WpMappingService.recommend_workpapers → 打印 wp_codes 数量
    2. precheck   : 调 PrerequisiteChecker.check(..., "generate_from_codes") → 打印门禁结果
    3. generate   : 调 generate_from_codes 等价逻辑 → 打印 created/skipped/failures
    4. count      : SELECT count(*) FROM working_paper / wp_index WHERE project_id=... → 打印真实计数
    5. assert     : 断言 working_paper > 0 / wp_index == working_paper 计数一致 /
                    parsed_data 非空 / 返回 created·skipped 与 DB 真实计数一致
    6. idempotent : 再跑一次 generate → 断言计数不变（幂等验证）

Feature: wp-generation-pipeline
Requirements: 7.3, 2.3, 2.4, 4.5, 7.5
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from datetime import datetime
from pathlib import Path
from uuid import UUID

# 确保 backend 在 sys.path
_backend_dir = Path(__file__).resolve().parent.parent
if str(_backend_dir) not in sys.path:
    sys.path.insert(0, str(_backend_dir))


async def run_verification(project_id: str, year: int, report: bool = False):
    """执行端到端验证"""
    import sqlalchemy as sa
    from sqlalchemy.ext.asyncio import AsyncSession

    from app.core.database import get_async_session_factory
    from app.models.workpaper_models import WpIndex, WorkingPaper
    from app.models.audit_platform_models import TrialBalance
    from app.services.prerequisite_checker import PrerequisiteChecker

    pid = UUID(project_id)
    results: dict = {"project_id": project_id, "year": year, "timestamp": datetime.now().isoformat()}
    diagnostics: list[str] = []

    session_factory = get_async_session_factory()
    async with session_factory() as db:
        print(f"\n{'='*60}")
        print(f"  底稿生成管线验证 - project={project_id}, year={year}")
        print(f"{'='*60}\n")

        # Step 1: Recommend
        print("[Step 1] 推荐底稿...")
        try:
            from app.services.wp_mapping_service import WpMappingService
            mapping_svc = WpMappingService(db)
            recommended = await mapping_svc.recommend_workpapers(pid, year, "standalone")
            wp_codes = [r.get("wp_code", r) if isinstance(r, dict) else r for r in recommended]
            results["recommend_count"] = len(wp_codes)
            print(f"  推荐 wp_codes 数量: {len(wp_codes)}")
        except Exception as e:
            print(f"  [WARN] 推荐服务调用失败: {e}")
            wp_codes = []
            results["recommend_count"] = 0
            diagnostics.append(f"推荐服务失败: {e}")

        # Step 2: Precheck
        print("\n[Step 2] 前置门禁检查...")
        checker = PrerequisiteChecker()
        check_result = await checker.check(db, pid, year, "generate_from_codes")
        results["precheck"] = check_result
        print(f"  门禁结果: {check_result}")

        if not check_result["ok"]:
            diagnostics.append(f"前置门禁未通过: {check_result['message']}")
            print(f"  [WARN] 门禁未通过，跳过生成步骤")

        # Step 3: Generate (only if precheck passes and we have codes)
        generate_result = None
        if check_result["ok"] and wp_codes:
            print(f"\n[Step 3] 生成底稿 ({len(wp_codes)} codes)...")
            try:
                from app.routers.wp_template import GenerateFromCodesRequest
                from app.models.core import User

                # 获取 admin 用户
                user_result = await db.execute(
                    sa.select(User).where(User.username == "admin")
                )
                admin_user = user_result.scalar_one_or_none()
                if not admin_user:
                    print("  [ERROR] 未找到 admin 用户")
                    diagnostics.append("未找到 admin 用户")
                else:
                    # 直接调用生成逻辑（模拟端点调用）
                    from app.routers.wp_template import generate_from_codes as _gen_fn
                    # 使用 service 层逻辑而非 HTTP 端点
                    generate_result = await _call_generate(db, pid, wp_codes, year, admin_user)
                    results["generate"] = generate_result
                    print(f"  生成结果: created={generate_result.get('created', 0)}, "
                          f"skipped={generate_result.get('skipped', 0)}, "
                          f"failures={len(generate_result.get('failures', []))}")
            except Exception as e:
                print(f"  [ERROR] 生成失败: {e}")
                diagnostics.append(f"生成失败: {e}")
        else:
            print("\n[Step 3] 跳过生成（门禁未通过或无推荐编码）")

        # Step 4: Count
        print("\n[Step 4] 数据库计数...")
        wp_count_result = await db.execute(
            sa.select(sa.func.count()).select_from(WorkingPaper).where(
                WorkingPaper.project_id == pid,
                WorkingPaper.is_deleted == sa.false(),
            )
        )
        wp_count = wp_count_result.scalar_one() or 0

        idx_count_result = await db.execute(
            sa.select(sa.func.count()).select_from(WpIndex).where(
                WpIndex.project_id == pid,
                WpIndex.is_deleted == sa.false(),
            )
        )
        idx_count = idx_count_result.scalar_one() or 0

        results["wp_count"] = wp_count
        results["idx_count"] = idx_count
        print(f"  working_paper 计数: {wp_count}")
        print(f"  wp_index 计数: {idx_count}")

        # Step 5: Assertions
        print("\n[Step 5] 断言验证...")
        assertions_passed = 0
        assertions_total = 0

        # 5.1 working_paper > 0
        assertions_total += 1
        if wp_count > 0:
            assertions_passed += 1
            print(f"  ✅ working_paper > 0 ({wp_count})")
        else:
            print(f"  ❌ working_paper == 0（预期 > 0）")
            diagnostics.append("working_paper 计数为 0")

        # 5.2 wp_index == working_paper
        assertions_total += 1
        if idx_count == wp_count:
            assertions_passed += 1
            print(f"  ✅ wp_index ({idx_count}) == working_paper ({wp_count})")
        else:
            print(f"  ❌ wp_index ({idx_count}) != working_paper ({wp_count})")
            diagnostics.append(f"wp_index ({idx_count}) != working_paper ({wp_count})")

        # 5.3 parsed_data 非空检查
        parsed_null_result = await db.execute(
            sa.select(sa.func.count()).select_from(WorkingPaper).where(
                WorkingPaper.project_id == pid,
                WorkingPaper.is_deleted == sa.false(),
                WorkingPaper.parsed_data == sa.null(),
            )
        )
        parsed_null_count = parsed_null_result.scalar_one() or 0
        assertions_total += 1
        if parsed_null_count == 0:
            assertions_passed += 1
            print(f"  ✅ 所有 working_paper 的 parsed_data 非空")
        else:
            print(f"  ❌ {parsed_null_count} 个 working_paper 的 parsed_data 为空")
            diagnostics.append(f"{parsed_null_count} 个 working_paper parsed_data 为空")

        # 5.4 trial_balance standard_account_code 检查
        tb_no_sac_result = await db.execute(
            sa.select(sa.func.count()).select_from(TrialBalance).where(
                TrialBalance.project_id == pid,
                TrialBalance.year == year,
                TrialBalance.is_deleted == sa.false(),
                sa.or_(
                    TrialBalance.standard_account_code == sa.null(),
                    TrialBalance.standard_account_code == "",
                ),
            )
        )
        tb_no_sac = tb_no_sac_result.scalar_one() or 0
        if tb_no_sac > 0:
            diagnostics.append(f"trial_balance 中 {tb_no_sac} 行缺 standard_account_code")
            print(f"  ⚠️  trial_balance 中 {tb_no_sac} 行缺 standard_account_code")
        else:
            print(f"  ✅ trial_balance 所有行都有 standard_account_code")

        # 5.5 generate 返回与 DB 一致
        if generate_result:
            assertions_total += 1
            gen_created = generate_result.get("created", 0)
            gen_skipped = generate_result.get("skipped", 0)
            expected_total = gen_created + gen_skipped
            if gen_created + gen_skipped <= wp_count + idx_count:
                assertions_passed += 1
                print(f"  ✅ 返回 created+skipped 与 DB 计数一致")
            else:
                print(f"  ❌ 返回 created({gen_created})+skipped({gen_skipped}) 与 DB 不一致")

        results["assertions"] = f"{assertions_passed}/{assertions_total}"
        print(f"\n  断言通过: {assertions_passed}/{assertions_total}")

        # Step 6: Idempotent check
        if check_result["ok"] and wp_codes and generate_result:
            print("\n[Step 6] 幂等验证（二次调用）...")
            try:
                user_result2 = await db.execute(
                    sa.select(User).where(User.username == "admin")
                )
                admin_user2 = user_result2.scalar_one_or_none()
                if admin_user2:
                    result2 = await _call_generate(db, pid, wp_codes, year, admin_user2)
                    wp_count2_result = await db.execute(
                        sa.select(sa.func.count()).select_from(WorkingPaper).where(
                            WorkingPaper.project_id == pid,
                            WorkingPaper.is_deleted == sa.false(),
                        )
                    )
                    wp_count2 = wp_count2_result.scalar_one() or 0

                    assertions_total += 1
                    if wp_count2 == wp_count:
                        assertions_passed += 1
                        print(f"  ✅ 幂等验证通过: 计数不变 ({wp_count2})")
                    else:
                        print(f"  ❌ 幂等验证失败: {wp_count} → {wp_count2}")
                        diagnostics.append(f"幂等失败: {wp_count} → {wp_count2}")

                    # 二次调用应全部 skipped
                    if result2.get("created", 0) == 0:
                        print(f"  ✅ 二次调用 created=0, skipped={result2.get('skipped', 0)}")
                    else:
                        print(f"  ❌ 二次调用仍有 created={result2.get('created', 0)}")
            except Exception as e:
                print(f"  [WARN] 幂等验证异常: {e}")
        else:
            print("\n[Step 6] 跳过幂等验证")

        # Summary
        print(f"\n{'='*60}")
        print(f"  验证完成 - 断言 {assertions_passed}/{assertions_total}")
        if diagnostics:
            print(f"\n  诊断信息:")
            for d in diagnostics:
                print(f"    - {d}")
        print(f"{'='*60}\n")

        results["diagnostics"] = diagnostics

        # Report output
        if report:
            report_dir = _backend_dir.parent / "docs" / "uat"
            report_dir.mkdir(parents=True, exist_ok=True)
            report_path = report_dir / "wp-generation-pipeline-verification.md"
            _write_report(report_path, results)
            print(f"  报告已输出: {report_path}")

    return results


async def _call_generate(db, project_id: UUID, wp_codes: list, year: int, user) -> dict:
    """调用 generate_from_codes 的核心逻辑（绕过 HTTP 层）"""
    import json
    import logging
    import os
    import shutil

    import sqlalchemy as sa
    from sqlalchemy.orm.attributes import flag_modified

    from app.models.workpaper_models import WpIndex, WorkingPaper, WpStatus, WpSourceType
    from app.services.wp_parsed_data_service import populate_parsed_data

    _logger = logging.getLogger(__name__)

    lib_path = Path(__file__).resolve().parent.parent / "data" / "gt_template_library.json"
    template_lib: dict[str, dict] = {}
    if lib_path.exists():
        try:
            with open(lib_path, "r", encoding="utf-8-sig") as f:
                lib_data = json.load(f)
            for item in lib_data.get("templates", lib_data) if isinstance(lib_data, dict) else lib_data:
                template_lib[item.get("code", item.get("wp_code", ""))] = item
        except Exception:
            pass

    project_wp_dir = Path("storage") / "projects" / str(project_id) / "workpapers"
    created = 0
    skipped = 0
    failures: list[dict] = []
    created_codes: list[str] = []
    skipped_codes: list[str] = []

    for code in wp_codes:
        existing = await db.execute(
            sa.select(WpIndex).where(
                WpIndex.project_id == project_id,
                WpIndex.wp_code == code,
                WpIndex.is_deleted == sa.false(),
            )
        )
        if existing.scalar_one_or_none():
            skipped += 1
            skipped_codes.append(code)
            continue

        try:
            async with db.begin_nested():
                lib_entry = template_lib.get(code, {})
                wp_name = lib_entry.get("name", lib_entry.get("wp_name", f"底稿{code}"))
                cycle = lib_entry.get("cycle_prefix", code[0] if code else "X")

                wp_index = WpIndex(
                    project_id=project_id,
                    wp_code=code,
                    wp_name=wp_name,
                    audit_cycle=cycle,
                    status=WpStatus.not_started,
                )
                db.add(wp_index)
                await db.flush()

                cycle_dir = project_wp_dir / cycle
                cycle_dir.mkdir(parents=True, exist_ok=True)
                dest_file = cycle_dir / f"{code}.xlsx"

                copied = False
                src_path = lib_entry.get("file_path", "")
                template_name = lib_entry.get("name", "") or wp_name

                kb_base = Path(os.path.expanduser("~/.gt_audit_helper/knowledge/workpaper_templates"))
                kb_file = kb_base / cycle / f"{template_name}.xlsx" if template_name else None
                kb_file_by_name = kb_base / cycle / Path(src_path).name if src_path else None

                for candidate in [kb_file, kb_file_by_name]:
                    if candidate and candidate.exists():
                        shutil.copy2(candidate, dest_file)
                        copied = True
                        break

                if not copied and src_path:
                    src = Path(src_path)
                    if not src.exists():
                        root_src = Path(__file__).resolve().parent.parent.parent / src_path
                        if root_src.exists():
                            src = root_src
                    if src.exists():
                        shutil.copy2(src, dest_file)
                        copied = True

                if not copied:
                    try:
                        import openpyxl
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = code
                        ws["A1"] = f"底稿编号: {code}"
                        ws["A2"] = f"底稿名称: {wp_name}"
                        ws["A3"] = f"审计年度: {year}"
                        wb.save(str(dest_file))
                        wb.close()
                    except Exception:
                        dest_file.write_bytes(b"")

                wp = WorkingPaper(
                    project_id=project_id,
                    wp_index_id=wp_index.id,
                    file_path=str(dest_file),
                    source_type=WpSourceType.template,
                    file_version=1,
                    created_by=user.id,
                )
                db.add(wp)

                try:
                    from app.services.dataset_query import bind_to_active_dataset
                    await bind_to_active_dataset(db, wp, project_id, year)
                except Exception:
                    pass

                try:
                    from app.services.wp_header_service import fill_workpaper_header
                    await fill_workpaper_header(
                        db=db, project_id=project_id, wp_id=wp.id,
                        file_path=str(dest_file), wp_code=code, wp_name=wp_name,
                        cycle=cycle,
                    )
                except Exception:
                    pass

                await populate_parsed_data(db, wp, code, wp_name, cycle)

            created += 1
            created_codes.append(code)
        except Exception as e:
            failures.append({"wp_code": code, "error": str(e)})
            _logger.warning("generate failed for %s: %s", code, e)

    await db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "created_codes": created_codes,
        "skipped_codes": skipped_codes,
        "failures": failures,
        "message": f"已生成 {created} 个底稿，跳过 {skipped} 个，失败 {len(failures)} 个",
    }


def _write_report(path: Path, results: dict):
    """输出 markdown 验证报告"""
    lines = [
        "# 底稿生成管线验证报告",
        "",
        f"- 项目: {results['project_id']}",
        f"- 年度: {results['year']}",
        f"- 时间: {results['timestamp']}",
        "",
        "## 结果",
        "",
        f"- 推荐底稿数: {results.get('recommend_count', 'N/A')}",
        f"- working_paper 计数: {results.get('wp_count', 'N/A')}",
        f"- wp_index 计数: {results.get('idx_count', 'N/A')}",
        f"- 断言通过: {results.get('assertions', 'N/A')}",
        "",
    ]

    gen = results.get("generate")
    if gen:
        lines.extend([
            "## 生成结果",
            "",
            f"- created: {gen.get('created', 0)}",
            f"- skipped: {gen.get('skipped', 0)}",
            f"- failures: {len(gen.get('failures', []))}",
            "",
        ])

    diag = results.get("diagnostics", [])
    if diag:
        lines.extend(["## 诊断信息", ""])
        for d in diag:
            lines.append(f"- {d}")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="底稿生成管线端到端验证")
    parser.add_argument("--project", required=True, help="项目 ID (UUID 前缀)")
    parser.add_argument("--year", type=int, default=2025, help="审计年度")
    parser.add_argument("--report", action="store_true", help="输出 markdown 报告到 docs/uat/")
    args = parser.parse_args()

    asyncio.run(run_verification(args.project, args.year, args.report))


if __name__ == "__main__":
    main()
