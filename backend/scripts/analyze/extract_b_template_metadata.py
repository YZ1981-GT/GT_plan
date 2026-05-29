"""从致同 2025 修订版 B 类底稿模板中提取元数据，生成 wp_template_metadata seed 数据。

B 类覆盖：初步业务活动（B1-B5）+ 风险评估（B10-B60）
用法: python backend/scripts/extract_b_template_metadata.py
"""
import json, re, sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl required. pip install openpyxl")
try:
    from docx import Document as DocxDocument
except ImportError:
    sys.exit("ERROR: python-docx required. pip install python-docx")

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATE_BASE = REPO_ROOT / "致同通用审计程序及底稿模板（2025年修订）" / "1.致同审计程序及底稿模板（2025年）"
B1_DIR = TEMPLATE_BASE / "1.初步业务活动（B1-B5）"
B10_DIR = TEMPLATE_BASE / "2.风险评估（B11-B60）"
OUTPUT_FILE = REPO_ROOT / "backend" / "data" / "wp_template_metadata_b_seed.json"

UNIVER_CODES = {"B1-1","B1-2","B13","B15","B50-1","B50-2","B50-3","B50-4",
                "B51","B51-3","B51-5","B52","B60-1"}
FORM_CODES = {"B1A","B1B","B1-5","B2","B2-5","B3","B10","B11","B12",
              "B18","B19","B20","B22","B23","B40","B50"}
FORMULA_RE = re.compile(r"(SUM|IF|VLOOKUP|INDEX|MATCH|SUMIF|SUMPRODUCT|AVERAGE|COUNT)\(", re.I)
RE_WP_CODE = re.compile(r"^(B\d+[A-Z]?(?:-\d+)?)")
RE_FIELD = re.compile(r"[【《]([^】》]+)[】》]|(\{\{[^}]+\}\})")

OBJECTIVES = {
    "B1": "完成业务承接/保持评估，确认是否接受或保持委托",
    "B2": "与前任注册会计师沟通，了解变更原因及重要事项",
    "B3": "确认项目组成员独立性，识别并评价独立性威胁",
    "B5": "签署业务约定书，明确审计范围和双方责任",
    "B10": "了解被审计单位及其环境，识别重大错报风险",
    "B11": "通过检查相关信息或文件了解被审计单位",
    "B12": "通过访谈了解被审计单位及其环境",
    "B13": "执行初步分析程序，识别异常变动和潜在风险",
    "B15": "确定重要性水平，为审计计划提供基础",
    "B18": "了解内部审计职能及其工作",
    "B19": "识别关联方及关联方交易",
    "B20": "了解被审计单位内部控制",
    "B22": "了解并评价企业层面控制环境",
    "B23": "了解、评价并测试业务层面控制",
    "B30": "制定集团审计策略和计划",
    "B40": "组织项目组讨论，交流审计风险信息",
    "B50": "汇总风险评估结果，确定重大错报风险",
    "B51": "评估舞弊风险因素", "B52": "评估管理层凌驾于控制之上的风险",
    "B60": "制定总体审计策略和具体审计计划",
}


def parse_wp_code(filename: str) -> str:
    stem = Path(filename).stem
    if stem.startswith("B5附件"):
        return "B5"
    m = RE_WP_CODE.match(stem)
    return m.group(1) if m else ""


def parse_wp_name(filename: str) -> str:
    stem = Path(filename).stem
    cleaned = re.sub(r"^B\d+[A-Z]?(?:-\d+(?:-\d+)?)?\s*", "", stem)
    return re.sub(r"[（(][^)）]*[)）]$", "", cleaned).strip() or stem


def primary_code(wp_code: str) -> str:
    m = re.match(r"^(B\d+)[A-D]$", wp_code)
    if m:
        return m.group(1)
    m = re.match(r"^(B\d+)", wp_code)
    return m.group(1) if m else wp_code


def determine_type(wp_code: str, ext: str, has_formulas: bool) -> str:
    if wp_code in UNIVER_CODES:
        return "univer"
    if wp_code in FORM_CODES:
        return "form"
    if ext in (".docx", ".doc"):
        return "word"
    return "univer" if has_formulas else "form"


def extract_xlsx(filepath: Path) -> dict:
    try:
        wb = load_workbook(str(filepath), read_only=True, data_only=False)
    except Exception as e:
        print(f"  ⚠ xlsx 打开失败: {filepath.name} ({e})")
        return {}
    sheets = wb.sheetnames
    formula_cells, form_schema, has_formulas = [], [], False
    for sn in sheets:
        try:
            ws = wb[sn]
        except Exception:
            continue
        labels = []
        for row in ws.iter_rows(max_row=30):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val.startswith("=") and FORMULA_RE.search(val):
                    has_formulas = True
                    if len(formula_cells) < 10:
                        ref = f"{cell.column_letter}{cell.row}" if hasattr(cell, "column_letter") else f"R{cell.row}"
                        formula_cells.append({"sheet": sn, "cell": ref, "formula": val[:120]})
            if row and row[0].value and isinstance(row[0].value, str):
                lbl = row[0].value.strip()
                if lbl and len(lbl) < 50 and not lbl.startswith("="):
                    labels.append(lbl)
        if not form_schema and labels:
            for lbl in labels[:15]:
                key = re.sub(r"[^\w]", "_", lbl).strip("_").lower()[:30]
                if not key:
                    continue
                ft = "date" if ("日期" in lbl or "时间" in lbl) else \
                     "checkbox" if ("是否" in lbl or "确认" in lbl) else \
                     "select" if ("等级" in lbl or "类型" in lbl or "结论" in lbl) else "input"
                form_schema.append({"key": key, "label": lbl, "type": ft})
    try:
        wb.close()
    except Exception:
        pass
    return {"sheets": sheets, "formula_cells": formula_cells, "has_formulas": has_formulas, "form_schema": form_schema}


def extract_docx(filepath: Path) -> dict:
    try:
        doc = DocxDocument(str(filepath))
    except Exception as e:
        print(f"  ⚠ docx 打开失败: {filepath.name} ({e})")
        return {}
    headings, fields, para_count = [], [], 0
    for para in doc.paragraphs:
        para_count += 1
        if para.style and para.style.name and "Heading" in para.style.name:
            headings.append(para.text.strip())
        for m in RE_FIELD.finditer(para.text):
            field = m.group(1) or m.group(2)
            if field and field not in fields:
                fields.append(field)
    name = filepath.stem
    subtype = "agreement" if "约定书" in name else "letter" if ("沟通函" in name or "通知" in name) else \
              "declaration" if "声明" in name else "memo" if ("备忘录" in name or "记录" in name) else \
              "report" if ("评价" in name or "报告" in name) else "document"
    return {"headings": headings[:10], "paragraph_count": para_count,
            "template_fields": fields[:20], "doc_subtype": subtype}


def main():
    import io, sys as _sys
    _sys.stdout = io.TextIOWrapper(_sys.stdout.buffer, encoding="utf-8", errors="replace")
    dirs = [d for d in (B1_DIR, B10_DIR) if d.exists()]
    if not dirs:
        sys.exit(f"ERROR: 模板目录不存在: {B1_DIR} / {B10_DIR}")
    print("扫描 B 类模板目录:")
    for d in dirs:
        print(f"   {d.relative_to(REPO_ROOT)}")

    all_files: list[Path] = []
    for d in dirs:
        for ext in ("*.xlsx", "*.xlsm", "*.docx", "*.doc"):
            all_files.extend(d.rglob(ext))
    all_files = [f for f in all_files if not f.name.startswith("~$") and not f.name.startswith("~WRL")]
    all_files.sort(key=lambda p: p.name)
    print(f"\n🔍 找到 {len(all_files)} 个模板文件\n")

    grouped: dict[str, list[dict]] = {}
    for fp in all_files:
        wc = parse_wp_code(fp.name)
        if not wc:
            continue
        pc = primary_code(wc)
        grouped.setdefault(pc, []).append({"filepath": fp, "wp_code": wc})

    entries, stats = [], {"xlsx": 0, "docx": 0, "form": 0, "univer": 0, "word": 0}
    for pc in sorted(grouped, key=lambda c: (int(re.search(r"\d+", c).group()), c)):
        files_info = grouped[pc]
        tpl_files, all_sheets, all_formulas, form_schema = [], [], [], []
        has_formulas, headings, fields, doc_subtype, fmts = False, [], [], None, set()

        for fi in files_info:
            fp, ext = fi["filepath"], fi["filepath"].suffix.lower()
            tpl_files.append(fp.name)
            fmts.add(ext.lstrip("."))
            if ext in (".xlsx", ".xlsm"):
                stats["xlsx"] += 1
                meta = extract_xlsx(fp)
                if meta:
                    all_sheets.extend(meta.get("sheets", []))
                    all_formulas.extend(meta.get("formula_cells", []))
                    has_formulas = has_formulas or meta.get("has_formulas", False)
                    if not form_schema:
                        form_schema = meta.get("form_schema", [])
            elif ext in (".docx", ".doc"):
                stats["docx"] += 1
                if ext == ".docx":
                    meta = extract_docx(fp)
                    if meta:
                        headings.extend(meta.get("headings", []))
                        fields.extend(meta.get("template_fields", []))
                        doc_subtype = doc_subtype or meta.get("doc_subtype")

        comp = determine_type(files_info[0]["wp_code"], files_info[0]["filepath"].suffix.lower(), has_formulas)
        stats[comp] += 1
        entry = {
            "wp_code": pc,
            "wp_name": parse_wp_name(files_info[0]["filepath"].name),
            "component_type": comp,
            "audit_stage": "planning",
            "cycle": None,
            "file_format": "docx" if "docx" in fmts else ("xlsm" if "xlsm" in fmts else "xlsx"),
            "template_files": tpl_files,
            "sheets": list(dict.fromkeys(all_sheets)) or None,
            "form_schema": form_schema if comp == "form" else None,
            "procedure_steps": None,
            "formula_cells": all_formulas[:10] if all_formulas else [],
            "audit_objective": OBJECTIVES.get(pc, f"完成{parse_wp_name(files_info[0]['filepath'].name)}相关审计程序"),
        }
        if comp == "word":
            entry["doc_subtype"] = doc_subtype
            entry["template_fields"] = fields[:15] or None
            entry["headings"] = headings[:10] or None
        entries.append(entry)

    output = {
        "description": "B 类底稿模板元数据（初步业务活动+风险评估）",
        "version": "2025-R1-auto",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "entries": entries,
        "stats": {"total_files_scanned": len(all_files), "total_entries": len(entries),
                  "xlsx_files": stats["xlsx"], "docx_files": stats["docx"],
                  "by_component_type": {"form": stats["form"], "univer": stats["univer"], "word": stats["word"]}},
    }
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print("=" * 60)
    print("📊 B 类模板元数据提取完成")
    print("=" * 60)
    print(f"  总文件数: {len(all_files)} (xlsx: {stats['xlsx']}, docx: {stats['docx']})")
    print(f"  总条目数: {len(entries)}")
    print(f"  组件类型: form={stats['form']}, univer={stats['univer']}, word={stats['word']}")
    print(f"\n  条目列表:")
    for e in entries:
        print(f"    {e['wp_code']:8s} [{e['component_type']:6s}] {e['wp_name']}")
    print(f"\n✅ 输出: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
