"""从致同 2025 修订版 D-N 循环底稿模板中提取元数据，生成 wp_template_metadata seed 数据。
用法: python backend/scripts/extract_dn_template_metadata.py
"""
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATE_BASE = (REPO_ROOT / "致同通用审计程序及底稿模板（2025年修订）"
                 / "1.致同审计程序及底稿模板（2025年）" / "4.风险应对-实质性程序（D-N）")
MAPPING_FILE = REPO_ROOT / "backend" / "data" / "wp_account_mapping.json"
OUTPUT_FILE = REPO_ROOT / "backend" / "data" / "wp_template_metadata_dn_seed.json"

CYCLE_NAMES = {
    "D": "收入循环", "E": "货币资金循环", "F": "存货循环", "G": "投资循环",
    "H": "固定资产循环", "I": "无形资产循环", "J": "职工薪酬循环",
    "K": "管理循环", "L": "债务循环", "M": "权益循环", "N": "税金循环",
}
CYCLE_ASSERTIONS = {
    "D": ["existence", "completeness", "accuracy", "cutoff"],
    "E": ["existence", "completeness", "rights_and_obligations"],
    "F": ["existence", "valuation", "completeness"],
    "G": ["existence", "valuation", "rights_and_obligations"],
    "H": ["existence", "valuation", "completeness"],
    "I": ["existence", "valuation", "rights_and_obligations"],
    "J": ["completeness", "accuracy", "obligations"],
    "K": ["completeness", "accuracy", "classification"],
    "L": ["completeness", "accuracy", "obligations"],
    "M": ["existence", "rights_and_obligations", "presentation"],
    "N": ["completeness", "accuracy", "valuation"],
}
AUDIT_OBJECTIVES = {
    "应收账款": "验证应收账款期末余额的存在性、计价和分摊",
    "应收票据": "验证应收票据期末余额的存在性和权利",
    "预收账款": "验证预收账款期末余额的完整性和准确性",
    "营业收入": "验证营业收入的发生、完整性和截止",
    "货币资金": "验证货币资金期末余额的存在性和完整性",
    "存货": "验证存货期末余额的存在性和计价",
    "固定资产": "验证固定资产期末余额的存在性和计价",
    "在建工程": "验证在建工程期末余额的存在性和计价",
    "无形资产": "验证无形资产期末余额的存在性和计价",
    "长期股权投资": "验证长期股权投资期末余额的存在性和计价",
    "应付职工薪酬": "验证应付职工薪酬的完整性和准确性",
    "短期借款": "验证短期借款期末余额的完整性和义务",
    "长期借款": "验证长期借款期末余额的完整性和义务",
    "实收资本": "验证实收资本的存在性和权利",
    "所得税": "验证所得税费用的完整性和准确性",
    "应交税费": "验证应交税费期末余额的完整性和准确性",
    "函证": "通过函证程序验证相关余额的存在性",
}
RE_WP_CODE = re.compile(r"^([A-N]\d+)")
RE_WP_CODE_RANGE = re.compile(r"^([A-N]\d+)-\d+至[A-N]\d+-\d+")
RE_CROSS_SHEET = re.compile(r"'([^']+)'![A-Z]+\d+")
RE_FORMULA_FUNC = re.compile(r"(SUMIF|VLOOKUP|INDEX|MATCH|SUMPRODUCT|IF)\(", re.IGNORECASE)


def load_account_mapping() -> dict:
    if not MAPPING_FILE.exists():
        return {}
    with open(MAPPING_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {m["wp_code"]: m for m in data.get("mappings", []) if m.get("wp_code")}


def parse_wp_code(filename: str) -> str:
    stem = Path(filename).stem
    m = RE_WP_CODE_RANGE.match(stem)
    if m:
        return m.group(1)
    m = RE_WP_CODE.match(stem)
    return m.group(1) if m else ""


def parse_wp_name(filename: str) -> str:
    stem = Path(filename).stem
    cleaned = re.sub(r"^[A-N]\d+[-\d至]*\s*", "", stem)
    cleaned = re.sub(r"[（(]Leap[^)）]*[)）]", "", cleaned).strip()
    return cleaned or stem


def _classify_formula(formula: str, sheet_name: str, cell) -> dict | None:
    cell_ref = f"{cell.column_letter}{cell.row}" if hasattr(cell, "column_letter") else f"R{cell.row}"
    cross_refs = RE_CROSS_SHEET.findall(formula)
    if cross_refs:
        return {"sheet": sheet_name, "cell_ref": cell_ref,
                "formula_type": "cross_sheet_reference",
                "raw_formula": formula[:200], "referenced_sheets": list(set(cross_refs))}
    func_match = RE_FORMULA_FUNC.search(formula)
    if func_match:
        return {"sheet": sheet_name, "cell_ref": cell_ref,
                "formula_type": func_match.group(1).upper(), "raw_formula": formula[:200]}
    return None


def extract_file_metadata(filepath: Path) -> dict | None:
    try:
        wb = load_workbook(str(filepath), read_only=True, data_only=False)
    except Exception as e:
        print(f"  ⚠ 无法打开: {filepath.name} ({e})")
        return None
    sheets = wb.sheetnames
    formula_cells, procedure_steps = [], []
    # conclusion_cell 分优先级收集候选
    conclusion_candidates: list[tuple[int, dict]] = []  # (priority, {sheet, cell})
    for sheet_name in sheets:
        try:
            ws = wb[sheet_name]
        except Exception:
            continue
        is_procedure_sheet = "程序" in sheet_name or "检查" in sheet_name or "分析" in sheet_name
        is_audit_table = "审定表" in sheet_name
        if is_procedure_sheet:
            procedure_steps.append(sheet_name)
        for row in ws.iter_rows(max_row=50):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val.startswith("="):
                    info = _classify_formula(val, sheet_name, cell)
                    if info:
                        formula_cells.append(info)
                # conclusion_cell 检测：按优先级收集
                if "审计结论" in val or "总体结论" in val or ("结论" in val and cell.row >= 20):
                    ref = f"{cell.column_letter}{cell.row}" if hasattr(cell, "column_letter") else f"R{cell.row}"
                    # 优先级：程序表(1) > 审定表(2) > 其他(3)；行号越大越优先（结论通常在底部）
                    if is_procedure_sheet:
                        priority = 100 + cell.row  # 程序表最高优先
                    elif is_audit_table:
                        priority = 50 + cell.row
                    else:
                        priority = cell.row
                    conclusion_candidates.append((priority, {"sheet": sheet_name, "cell": ref}))
    try:
        wb.close()
    except Exception:
        pass
    # 选择最高优先级的 conclusion_cell
    conclusion_cell = None
    if conclusion_candidates:
        conclusion_candidates.sort(key=lambda x: x[0], reverse=True)
        conclusion_cell = conclusion_candidates[0][1]
    return {"sheets": sheets, "formula_cells": formula_cells[:50],
            "procedure_steps": procedure_steps, "conclusion_cell": conclusion_cell}


def infer_audit_objective(wp_name: str) -> str:
    for key, obj in AUDIT_OBJECTIVES.items():
        if key in wp_name:
            return obj
    return f"验证{wp_name}相关余额的真实性和准确性"


def find_cross_wp_references(formula_cells: list, wp_code: str) -> list[str]:
    refs = set()
    for fc in formula_cells:
        for ref_sheet in fc.get("referenced_sheets", []):
            m = RE_WP_CODE.match(ref_sheet)
            if m and m.group(1) != wp_code:
                refs.add(m.group(1))
    return sorted(refs)


def main():
    if not TEMPLATE_BASE.exists():
        print(f"ERROR: 模板目录不存在: {TEMPLATE_BASE}")
        sys.exit(1)
    print(f"📂 扫描目录: {TEMPLATE_BASE}")
    print(f"📄 映射文件: {MAPPING_FILE}\n")
    account_mapping = load_account_mapping()

    all_files: list[Path] = []
    for ext in ("*.xlsx", "*.xlsm"):
        all_files.extend(TEMPLATE_BASE.rglob(ext))
    all_files = [f for f in all_files if not f.name.startswith("~$")]
    all_files.sort(key=lambda p: p.name)
    print(f"🔍 找到 {len(all_files)} 个模板文件\n")

    grouped: dict[str, list[dict]] = {}
    stats_by_cycle: dict[str, int] = {}
    for filepath in all_files:
        wp_code = parse_wp_code(filepath.name)
        if not wp_code:
            continue
        cycle = wp_code[0]
        stats_by_cycle[cycle] = stats_by_cycle.get(cycle, 0) + 1
        if wp_code not in grouped:
            grouped[wp_code] = []
        meta = extract_file_metadata(filepath)
        grouped[wp_code].append({
            "filename": filepath.name,
            "file_format": filepath.suffix.lstrip(".").lower(),
            "meta": meta,
        })

    entries = []
    total_formula_cells = 0
    total_cross_refs = 0
    for wp_code in sorted(grouped.keys()):
        files = grouped[wp_code]
        cycle = wp_code[0]
        all_sheets, all_formulas, all_procedures = [], [], []
        conclusion, file_formats, template_files = None, set(), []
        for f in files:
            template_files.append(f["filename"])
            file_formats.add(f["file_format"])
            if f["meta"]:
                all_sheets.extend(f["meta"]["sheets"])
                all_formulas.extend(f["meta"]["formula_cells"])
                all_procedures.extend(f["meta"]["procedure_steps"])
                if not conclusion and f["meta"]["conclusion_cell"]:
                    conclusion = f["meta"]["conclusion_cell"]
        all_sheets = list(dict.fromkeys(all_sheets))
        all_procedures = list(dict.fromkeys(all_procedures))
        total_formula_cells += len(all_formulas)
        cross_refs = find_cross_wp_references(all_formulas, wp_code)
        total_cross_refs += len(cross_refs)
        wp_name = parse_wp_name(files[0]["filename"])
        mapping_info = account_mapping.get(wp_code, {})
        if mapping_info.get("wp_name"):
            wp_name = mapping_info["wp_name"]
        entries.append({
            "wp_code": wp_code,
            "wp_name": wp_name,
            "component_type": "univer",
            "audit_stage": "substantive",
            "cycle": cycle,
            "cycle_name": CYCLE_NAMES.get(cycle, ""),
            "file_format": "xlsm" if "xlsm" in file_formats else "xlsx",
            "template_files": template_files,
            "sheets": all_sheets,
            "procedure_steps": all_procedures or None,
            "formula_cells": all_formulas[:20] if all_formulas else [],
            "conclusion_cell": conclusion,
            "audit_objective": infer_audit_objective(wp_name),
            "related_assertions": CYCLE_ASSERTIONS.get(cycle, []),
            "linked_accounts": mapping_info.get("account_codes", []),
            "note_section": mapping_info.get("note_section"),
            "cross_wp_references": cross_refs or [],
        })

    output = {
        "description": "D-N 循环底稿模板元数据（从致同 2025 模板自动提取）",
        "version": "2025-R1-auto",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "entries": entries,
        "stats": {
            "total_files_scanned": len(all_files),
            "total_entries": len(entries),
            "by_cycle": {c: stats_by_cycle.get(c, 0) for c in "DEFGHIJKLMN"},
        },
    }
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print("=" * 60)
    print("📊 提取完成摘要")
    print("=" * 60)
    print(f"  总文件数: {len(all_files)}")
    print(f"  总条目数: {len(entries)}")
    print(f"  公式单元格: {total_formula_cells}")
    print(f"  跨底稿引用: {total_cross_refs}\n")
    print("  按循环分布:")
    for c in "DEFGHIJKLMN":
        print(f"    {c} {CYCLE_NAMES.get(c, '')}: {stats_by_cycle.get(c, 0)} 文件")
    print(f"\n✅ 输出: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
