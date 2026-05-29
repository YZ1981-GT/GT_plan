"""从致同 2025 修订版 C/A/S 类底稿模板中提取元数据，生成 wp_template_metadata seed 数据。

C 类：风险应对—一般性程序与控制测试（C1-C26）
A 类：完成阶段（A1-A30）
S 类：特定项目程序（S1-S35）

用法: python backend/scripts/extract_cas_template_metadata.py
"""
import io, json, re, sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl required. pip install openpyxl")
try:
    from docx import Document as DocxDocument
except ImportError:
    sys.exit("ERROR: python-docx required. pip install python-docx")

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATE_BASE = REPO_ROOT / "致同通用审计程序及底稿模板（2025年修订）" / "1.致同审计程序及底稿模板（2025年）"
C_DIR = TEMPLATE_BASE / "3.风险应对-一般性程序与控制测试（C1-C26）"
A_DIR = TEMPLATE_BASE / "5.完成阶段（A1-A30）"
S_DIR = TEMPLATE_BASE / "6.特定项目程序（S）"
OUTPUT_FILE = REPO_ROOT / "backend" / "data" / "wp_template_metadata_cas_seed.json"

RE_WP_CODE = re.compile(r"^([CAS]\d+)(?:-\d+)?")
FORMULA_RE = re.compile(r"(SUM|IF|VLOOKUP|INDEX|MATCH|SUMIF|SUMPRODUCT|AVERAGE|COUNT)\(", re.I)
RE_FIELD = re.compile(r"[【《]([^】》]+)[】》]|(\{\{[^}]+\}\})")


def parse_wp_code(filename: str) -> str:
    """Extract wp_code like C1, C21-1, A9-2, S12A from filename."""
    stem = Path(filename).stem
    # Handle codes like C21-1, A9-1, S12A
    m = re.match(r"^([CAS]\d+[A-Z]?(?:-\d+)?)", stem)
    return m.group(1) if m else ""


def primary_code(wp_code: str) -> str:
    """Get primary grouping code: C21-1 -> C21, A9-2 -> A9, S12A -> S12."""
    m = re.match(r"^([CAS]\d+)", wp_code)
    return m.group(1) if m else wp_code


def parse_wp_name(filename: str) -> str:
    stem = Path(filename).stem
    cleaned = re.sub(r"^[CAS]\d+[A-Z]?(?:-\d+(?:-\d+)?)?\s*", "", stem)
    cleaned = re.sub(r"\d{4,6}$", "", cleaned)  # trailing date like 202504
    return cleaned.strip() or stem


def determine_audit_stage(wp_code: str) -> str:
    if wp_code.startswith("C"):
        return "risk_response"
    if wp_code.startswith("A"):
        return "completion"
    return "special"


def determine_component_type(wp_code: str, ext: str) -> str:
    """Determine component_type based on class rules."""
    if ext in (".docx", ".doc"):
        return "word"
    pc = primary_code(wp_code)
    num_match = re.match(r"[CAS](\d+)", pc)
    num = int(num_match.group(1)) if num_match else 0

    if pc.startswith("C"):
        if 1 <= num <= 15:
            return "hybrid"  # 穿行测试+控制矩阵+偏差评价
        if 21 <= num <= 26:
            return "univer"  # IT/会计分录, has formulas
        return "hybrid" if ext == ".xlsm" else "form"
    if pc.startswith("A"):
        if 1 <= num <= 10:
            return "form"  # checklist or word (word handled above)
        return "form"  # A11-A30 checklists
    # S class
    return "form"  # most are structured checklists


def extract_xlsx(filepath: Path) -> dict:
    try:
        wb = load_workbook(str(filepath), read_only=True, data_only=False)
    except Exception as e:
        print(f"  ⚠ xlsx 打开失败: {filepath.name} ({e})")
        return {}
    sheets = wb.sheetnames
    formula_cells, form_schema, has_formulas = [], [], False
    for sn in sheets:
        try:
            ws = wb[sn]
        except Exception:
            continue
        labels = []
        for row in ws.iter_rows(max_row=30):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val.startswith("=") and FORMULA_RE.search(val):
                    has_formulas = True
                    if len(formula_cells) < 10:
                        ref = f"{cell.column_letter}{cell.row}" if hasattr(cell, "column_letter") else f"R{cell.row}"
                        formula_cells.append({"sheet": sn, "cell": ref, "formula": val[:120]})
            if row and row[0].value and isinstance(row[0].value, str):
                lbl = row[0].value.strip()
                if lbl and len(lbl) < 50 and not lbl.startswith("="):
                    labels.append(lbl)
        if not form_schema and labels:
            for lbl in labels[:15]:
                key = re.sub(r"[^\w]", "_", lbl).strip("_").lower()[:30]
                if not key:
                    continue
                ft = "date" if ("日期" in lbl or "时间" in lbl) else \
                     "checkbox" if ("是否" in lbl or "确认" in lbl) else \
                     "select" if ("等级" in lbl or "类型" in lbl or "结论" in lbl) else "input"
                form_schema.append({"key": key, "label": lbl, "type": ft})
    try:
        wb.close()
    except Exception:
        pass
    return {"sheets": sheets, "formula_cells": formula_cells, "has_formulas": has_formulas, "form_schema": form_schema}


def extract_docx(filepath: Path) -> dict:
    try:
        doc = DocxDocument(str(filepath))
    except Exception as e:
        print(f"  ⚠ docx 打开失败: {filepath.name} ({e})")
        return {}
    headings, fields, para_count = [], [], 0
    for para in doc.paragraphs:
        para_count += 1
        if para.style and para.style.name and "Heading" in para.style.name:
            headings.append(para.text.strip())
        for m in RE_FIELD.finditer(para.text):
            field = m.group(1) or m.group(2)
            if field and field not in fields:
                fields.append(field)
    name = filepath.stem
    subtype = "letter" if ("沟通函" in name or "通知" in name or "函" in name) else \
              "declaration" if "声明" in name else \
              "report" if ("评价" in name or "报告" in name or "总结" in name) else \
              "memo" if ("备忘录" in name or "记录" in name) else "document"
    return {"headings": headings[:10], "paragraph_count": para_count,
            "template_fields": fields[:20], "doc_subtype": subtype}


def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    dirs = [d for d in (C_DIR, A_DIR, S_DIR) if d.exists()]
    if not dirs:
        sys.exit(f"ERROR: 模板目录不存在")
    print("扫描 C/A/S 类模板目录:")
    for d in dirs:
        print(f"   {d.relative_to(REPO_ROOT)}")

    all_files: list[Path] = []
    for d in dirs:
        for ext in ("*.xlsx", "*.xlsm", "*.xls", "*.docx", "*.doc"):
            all_files.extend(d.rglob(ext))
    all_files = [f for f in all_files if not f.name.startswith("~$") and not f.name.startswith("~WRL")]
    all_files.sort(key=lambda p: p.name)
    print(f"\n🔍 找到 {len(all_files)} 个模板文件\n")

    grouped: dict[str, list[dict]] = {}
    for fp in all_files:
        wc = parse_wp_code(fp.name)
        if not wc:
            continue
        pc = primary_code(wc)
        grouped.setdefault(pc, []).append({"filepath": fp, "wp_code": wc})

    entries, stats = [], {"xlsx": 0, "xlsm": 0, "docx": 0, "hybrid": 0, "form": 0, "univer": 0, "word": 0}
    for pc in sorted(grouped, key=lambda c: (c[0], int(re.search(r"\d+", c).group()))):
        files_info = grouped[pc]
        tpl_files, all_sheets, all_formulas, form_schema = [], [], [], []
        has_formulas, headings, fields, doc_subtype, fmts = False, [], [], None, set()

        for fi in files_info:
            fp, ext = fi["filepath"], fi["filepath"].suffix.lower()
            tpl_files.append(fp.name)
            fmts.add(ext.lstrip("."))
            if ext in (".xlsx", ".xlsm"):
                stats["xlsm" if ext == ".xlsm" else "xlsx"] += 1
                meta = extract_xlsx(fp)
                if meta:
                    all_sheets.extend(meta.get("sheets", []))
                    all_formulas.extend(meta.get("formula_cells", []))
                    has_formulas = has_formulas or meta.get("has_formulas", False)
                    if not form_schema:
                        form_schema = meta.get("form_schema", [])
            elif ext in (".docx", ".doc"):
                stats["docx"] += 1
                if ext == ".docx":
                    meta = extract_docx(fp)
                    if meta:
                        headings.extend(meta.get("headings", []))
                        fields.extend(meta.get("template_fields", []))
                        doc_subtype = doc_subtype or meta.get("doc_subtype")

        comp = determine_component_type(files_info[0]["wp_code"], files_info[0]["filepath"].suffix.lower())
        stats[comp] += 1
        entry = {
            "wp_code": pc,
            "wp_name": parse_wp_name(files_info[0]["filepath"].name),
            "component_type": comp,
            "audit_stage": determine_audit_stage(pc),
            "cycle": None,
            "file_format": "docx" if "docx" in fmts else ("xlsm" if "xlsm" in fmts else "xlsx"),
            "template_files": tpl_files,
            "sheets": list(dict.fromkeys(all_sheets)) or None,
            "form_schema": form_schema if comp == "form" else None,
            "procedure_steps": None,
            "formula_cells": all_formulas[:10] if all_formulas else [],
            "audit_objective": f"完成{parse_wp_name(files_info[0]['filepath'].name)}相关审计程序",
        }
        if comp == "word":
            entry["doc_subtype"] = doc_subtype
            entry["template_fields"] = fields[:15] or None
            entry["headings"] = headings[:10] or None
        entries.append(entry)

    output = {
        "description": "C/A/S 类底稿模板元数据",
        "version": "2025-R1-auto",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "entries": entries,
        "stats": {
            "total_files_scanned": len(all_files),
            "total_entries": len(entries),
            "xlsx_files": stats["xlsx"],
            "xlsm_files": stats["xlsm"],
            "docx_files": stats["docx"],
            "by_component_type": {
                "hybrid": stats["hybrid"], "form": stats["form"],
                "univer": stats["univer"], "word": stats["word"],
            },
        },
    }
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print("=" * 60)
    print("📊 C/A/S 类模板元数据提取完成")
    print("=" * 60)
    print(f"  总文件数: {len(all_files)} (xlsx: {stats['xlsx']}, xlsm: {stats['xlsm']}, docx: {stats['docx']})")
    print(f"  总条目数: {len(entries)}")
    print(f"  组件类型: hybrid={stats['hybrid']}, form={stats['form']}, univer={stats['univer']}, word={stats['word']}")
    print(f"  审计阶段: C(risk_response)={sum(1 for e in entries if e['audit_stage']=='risk_response')}, "
          f"A(completion)={sum(1 for e in entries if e['audit_stage']=='completion')}, "
          f"S(special)={sum(1 for e in entries if e['audit_stage']=='special')}")
    print(f"\n  条目列表:")
    for e in entries:
        print(f"    {e['wp_code']:8s} [{e['component_type']:6s}] [{e['audit_stage']:14s}] {e['wp_name']}")
    print(f"\n✅ 输出: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
