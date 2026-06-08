#!/usr/bin/env python3
"""POC：整理 soe_standalone.xlsx 资产负债表 — 表头占位 + BS 行 {{row:}}.

Task 0.0.3

Usage:
    python backend/scripts/prepare_financial_template_poc.py --dry-run
    python backend/scripts/prepare_financial_template_poc.py --write
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
XLSX = _BACKEND / "data/audit_report_templates/financial_statements/soe_standalone.xlsx"
SEED = _BACKEND / "data/report_config_seed.json"

BS_SHEET_PREFIX = "1,2-资产负债表"
ROW_PLACEHOLDER_RE = re.compile(r"^\{\{row:([^:]+):(current|prior)\}\}$")
MIN_DATA_ROWS = 20  # BS-002 … BS-021


def _norm_row_name(text: str) -> str:
    s = re.sub(r"\s+", "", (text or ""))
    return re.sub(r"^[△▲#]+", "", s)


def _load_bs_rows(standard: str = "soe_standalone") -> list[dict]:
    configs = json.loads(SEED.read_text(encoding="utf-8"))
    for block in configs:
        if (
            block.get("report_type") == "balance_sheet"
            and block.get("applicable_standard") == standard
        ):
            return block.get("rows", [])
    raise KeyError(f"balance_sheet config not found for {standard}")


def _find_balance_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.startswith(BS_SHEET_PREFIX) and "续" not in name:
            return wb[name]
    raise KeyError("balance sheet sheet not found")


def _build_name_to_row(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        key = _norm_row_name(str(val))
        if key:
            mapping[key] = r
    return mapping


def _is_formula_cell(ws, row: int, col: int) -> bool:
    val = ws.cell(row, col).value
    return isinstance(val, str) and val.startswith("=")


def prepare_workbook(wb: openpyxl.Workbook, *, write: bool) -> dict[str, int]:
    stats = {"headers": 0, "row_placeholders": 0, "skipped_formula": 0}
    ws = _find_balance_sheet(wb)
    rows_cfg = _load_bs_rows()
    name_to_row = _build_name_to_row(ws)

    # 表头
    header_ops = [
        (2, 1, "{{period_end_date}}"),
        (3, 1, "编制单位：{{company_full_name}}"),
    ]
    for r, c, text in header_ops:
        cur = ws.cell(r, c).value
        if cur != text:
            stats["headers"] += 1
            if write:
                ws.cell(r, c).value = text

    # 数据行：跳过 BS-001（大标题行），取后续 MIN_DATA_ROWS 条
    data_rows = [row for row in rows_cfg if row.get("row_code", "").startswith("BS-")]
    data_rows = [r for r in data_rows if r["row_code"] != "BS-001"][:MIN_DATA_ROWS]

    for cfg in data_rows:
        code = cfg["row_code"]
        name = _norm_row_name(cfg.get("row_name", ""))
        excel_row = name_to_row.get(name)
        if excel_row is None:
            continue
        for col, suffix in ((3, "current"), (4, "prior")):
            if _is_formula_cell(ws, excel_row, col):
                stats["skipped_formula"] += 1
                continue
            token = f"{{{{row:{code}:{suffix}}}}}"
            if ws.cell(excel_row, col).value != token:
                stats["row_placeholders"] += 1
                if write:
                    ws.cell(excel_row, col).value = token

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare financial template POC")
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if not args.write:
        args.dry_run = True

    if not XLSX.is_file():
        print(f"Missing: {XLSX}", file=sys.stderr)
        raise SystemExit(1)

    wb = openpyxl.load_workbook(XLSX)
    stats = prepare_workbook(wb, write=args.write)
    print(f"headers={stats['headers']} row_placeholders={stats['row_placeholders']} "
          f"skipped_formula={stats['skipped_formula']}")
    if args.write:
        wb.save(XLSX)
        print(f"Wrote {XLSX}")
    else:
        print("(dry-run; pass --write to apply)")


if __name__ == "__main__":
    main()
