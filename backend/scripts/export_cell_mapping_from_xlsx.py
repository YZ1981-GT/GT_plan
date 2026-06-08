#!/usr/bin/env python3
"""从 xlsx 内联 ``{{row:…}}`` 扫描生成 cell_mapping.json（全量 4 变体）.

Task 0.5.1 / 0.0.4b（POC）→ Task 12.1（全量泛化）。

- 按 sheet 名关键字判定 report_type（资产负债表/利润表/现金流量表/权益变动表/减值准备），
  自适应 soe_* 与 listed_* 不同 sheet 命名，不再硬编码 soe_standalone 别名。
- 扫描所有非 GT_Custom sheet 的内联 ``{{row:CODE:current|prior}}`` 与表头 ``{{key}}``。
- balance_sheet 主表 + 续表共用同一 report_type（按 sheet 关键字归并）。
- 输出 per-variant：sheet_aliases / headers / rows。

Usage:
    python backend/scripts/export_cell_mapping_from_xlsx.py --variant soe_standalone --write
    python backend/scripts/export_cell_mapping_from_xlsx.py --all --write
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

_BACKEND = Path(__file__).resolve().parent.parent
TPL_ROOT = _BACKEND / "data" / "audit_report_templates"
OUT = TPL_ROOT / "cell_mapping.json"

VARIANTS = [
    "soe_standalone",
    "soe_consolidated",
    "listed_standalone",
    "listed_consolidated",
]

# {{row:CODE:current|prior}} 或 {{row:CODE:current|prior:parent}}（公司/母公司个别列）
ROW_RE = re.compile(r"^\{\{row:([^:}]+):(current|prior)(?::(parent))?\}\}$")
HEADER_RE = re.compile(r"\{\{([a-z_]+)\}\}")

# sheet 名关键字 → report_type（与 prepare_financial_templates 一致）
_SHEET_KEYWORDS = [
    ("资产负债表", "balance_sheet"),
    ("利润表", "income_statement"),
    ("现金流量表", "cash_flow_statement"),
    ("权益变动表", "equity_statement"),
    ("减值准备", "asset_impairment"),
]


def _report_type_for_sheet(sheet_name: str) -> str | None:
    for kw, rt in _SHEET_KEYWORDS:
        if kw in sheet_name:
            return rt
    return None


def _build_sheet_aliases(wb: openpyxl.Workbook) -> dict[str, object]:
    """report_type → sheet 名（多张归并为 list）+ hidden_in_export。"""
    aliases: dict[str, list[str]] = {}
    hidden: list[str] = []
    for name in wb.sheetnames:
        if name == "GT_Custom":
            hidden.append(name)
            continue
        rt = _report_type_for_sheet(name)
        if rt:
            aliases.setdefault(rt, []).append(name)
    out: dict[str, object] = {}
    for rt, names in aliases.items():
        out[rt] = names[0] if len(names) == 1 else names
    if hidden:
        out["hidden_in_export"] = hidden
    return out


def scan_workbook(path: Path, variant_key: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    rows_map: dict[str, dict] = {}
    headers: dict[str, dict[str, str]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "GT_Custom":
            continue
        ws = wb[sheet_name]
        rt = _report_type_for_sheet(sheet_name)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if not isinstance(val, str):
                    continue
                text = val.strip()
                m = ROW_RE.match(text)
                if m:
                    code, period, is_parent = m.group(1), m.group(2), m.group(3)
                    entry = rows_map.setdefault(
                        code,
                        {
                            "row_code": code,
                            "sheet": rt or sheet_name,
                            "row_name": None,
                            "fill_empty_as": "blank",
                        },
                    )
                    coord = f"{get_column_letter(c)}{r}"
                    if is_parent:
                        entry[f"{period}_parent"] = coord
                    else:
                        entry[period] = coord
                    name_val = ws.cell(r, 1).value
                    if name_val and not entry.get("row_name"):
                        entry["row_name"] = str(name_val).strip()
                    note_val = ws.cell(r, 2).value
                    if isinstance(note_val, str) and "{{" in note_val:
                        entry["note_ref"] = f"B{r}"
                else:
                    for hm in HEADER_RE.finditer(text):
                        key = hm.group(1)
                        # row: 占位由上面处理；此处只收表头键
                        if key.startswith("row"):
                            continue
                        scope = rt or sheet_name
                        headers.setdefault(scope, {})[key] = (
                            f"{get_column_letter(c)}{r}"
                        )

    return {
        "template_key": variant_key,
        "xlsx_file": f"financial_statements/{path.name}",
        "sheet_aliases": _build_sheet_aliases(wb),
        "headers": headers,
        "rows": rows_map,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export cell_mapping.json from xlsx")
    parser.add_argument("--variant", default="soe_standalone", choices=VARIANTS)
    parser.add_argument("--all", action="store_true", help="处理全部 4 个变体")
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args()

    targets = VARIANTS if args.all else [args.variant]

    payloads: dict[str, dict] = {}
    for variant in targets:
        xlsx = TPL_ROOT / "financial_statements" / f"{variant}.xlsx"
        if not xlsx.is_file():
            print(f"Missing: {xlsx}", file=sys.stderr)
            continue
        payload = scan_workbook(xlsx, variant)
        payloads[variant] = payload
        rt_counts: dict[str, int] = {}
        parent_count = 0
        for entry in payload["rows"].values():
            rt_counts[entry["sheet"]] = rt_counts.get(entry["sheet"], 0) + 1
            if entry.get("current_parent") or entry.get("prior_parent"):
                parent_count += 1
        print(
            f"{variant}: {len(payload['rows'])} row codes "
            f"({parent_count} with :parent coords), "
            f"{len(payload.get('headers', {}))} header scopes; per-type={rt_counts}"
        )

    if args.write:
        if OUT.exists():
            root = json.loads(OUT.read_text(encoding="utf-8"))
        else:
            root = {"version": "poc-v1", "variants": {}}
        root["version"] = "v1"
        root.setdefault("variants", {})
        for variant, payload in payloads.items():
            root["variants"][variant] = payload
        OUT.write_text(
            json.dumps(root, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"Wrote {OUT} ({len(payloads)} variant(s))")
    else:
        for variant, payload in payloads.items():
            print(json.dumps(payload, ensure_ascii=False, indent=2)[:1500])


if __name__ == "__main__":
    main()
