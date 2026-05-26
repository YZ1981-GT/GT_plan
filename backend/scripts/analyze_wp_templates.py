"""
扫描 backend/wp_templates/ 下全部 xlsx 模板（476 个），按 sheet 提取特征 +
自动归类到 9 类（A/B/C/D/E/F/G/H/I），输出：
- .kiro/specs/workpaper-html-renderer/workpaper_template_analysis.json（机器读）
- .kiro/specs/workpaper-html-renderer/workpaper_template_analysis.md（人工读，按循环分章节）

特征提取（每 sheet）：
- max_row × max_col
- 合并单元格数 + 合并单元格密度（合并数 / 行数）
- 长文本 cell 数（>100 字符）
- 公式 cell 数（=开头）
- 是否含步骤词（步骤一/二/三）
- 是否含三段式索引（序号 / 内容 / 索引号）
- 首列前 10 行内容样本

归类规则（按优先级有序）：
1. I 占位/数据：含 GT_Custom/Data/List/Lists/Instructions/sheet1/temp_/参数 等
2. H 辅助说明：含"修订说明/编制说明/填表说明/文号规则/参考-XX"
3. A 程序表：含"实质性程序表"/末尾 [A-S]\d+A/含"程序表"
4. B 底稿目录：含"底稿目录"或末尾"目录"
5. C 附注披露：含"附注披露"/"披露信息"
6. E 控制测试：编码 C\d+ 起始 + 含"控制测试/评价偏差"
7. D 检查/政策/业务模式/函证/盘点/访谈/复核：覆盖 10 类子模式
8. G 测算表：含"测算/测试/计量/调节" + 公式 ≥ 5
9. F 数据表：含"审定/明细/分析/调整分录/汇总" 或 公式 ≥ 5
10. 兜底：未分类（输出 _pending_classification 标记）

用法：
    python backend/scripts/analyze_wp_templates.py
"""
import json
import re
import os
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# 路径
TEMPLATE_ROOT = Path(r"D:\GT_plan\backend\wp_templates")
OUTPUT_DIR = Path(r"D:\GT_plan\.kiro\specs\workpaper-html-renderer")
JSON_OUTPUT = OUTPUT_DIR / "workpaper_template_analysis.json"
MD_OUTPUT = OUTPUT_DIR / "workpaper_template_analysis.md"

CYCLE_NAMES = {
    'A': '报表/调整',
    'B': '控制了解',
    'C': '控制测试',
    'D': '销售收入',
    'E': '货币资金',
    'F': '采购存货',
    'G': '投资',
    'H': '固定资产+在建+使用权+租赁',
    'I': '无形资产+商誉+开发支出',
    'J': '职工薪酬+股份支付',
    'K': '管理',
    'L': '筹资',
    'M': '股东权益',
    'N': '税费',
    'S': '专项程序',
    'T': '工具',
    '_': '参考',
}


def extract_sheet_features(ws):
    """提取单 sheet 特征"""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    merged_count = len(ws.merged_cells.ranges) if hasattr(ws, 'merged_cells') else 0

    long_text = 0
    formula_count = 0
    has_step = False
    has_index_triplet = False
    first_col_samples = []
    has_assertion_5 = False  # 5 项认定

    # 限制扫描范围避免超大 sheet 拖慢
    scan_max_row = min(max_row, 200)
    scan_max_col = min(max_col, 30)

    for row_idx, row in enumerate(ws.iter_rows(
        min_row=1, max_row=scan_max_row,
        min_col=1, max_col=scan_max_col,
        values_only=False), 1):
        for c_idx, cell in enumerate(row, 1):
            v = cell.value
            if v is None:
                continue
            s = str(v)
            # 公式判定（openpyxl data_only=False 时公式以 = 开头）
            if isinstance(v, str) and v.startswith('='):
                formula_count += 1
            if len(s) > 100:
                long_text += 1
            if '步骤一' in s or '步骤二' in s:
                has_step = True
            if '存在' in s and '完整性' in s:
                has_assertion_5 = True
            # 收集首列样本
            if c_idx == 1 and row_idx <= 12 and s.strip():
                first_col_samples.append(s[:60])
        if row_idx >= 12:
            break

    # 三段式索引判定（含 序号+内容+索引号）
    full_text = ' '.join(first_col_samples)
    if '序号' in full_text and ('索引号' in full_text or '索引' in full_text):
        has_index_triplet = True

    return {
        'max_row': max_row,
        'max_col': max_col,
        'merged_count': merged_count,
        'merged_density': round(merged_count / max(max_row, 1), 2),
        'long_text_cells': long_text,
        'formula_cells': formula_count,
        'has_step_words': has_step,
        'has_index_triplet': has_index_triplet,
        'has_assertion_5': has_assertion_5,
        'first_col_samples': first_col_samples[:10],
    }


def classify_sheet(sheet_name, features):
    """按规则归类（返回 class_code, reason）"""
    s = sheet_name.strip()

    # 1. I 占位/数据
    if any(k in s for k in ['GT_Custom', 'Data', 'Lists', 'Instructions', 'temp_', '参数']):
        return 'I-占位', '占位关键词'
    if re.match(r'^(Sheet|sheet)\d*$', s) or s in ('0000', '0001'):
        return 'I-占位', 'Sheet 默认名'
    if 'List' in s and len(s) < 8:
        return 'I-占位', '简短 List'
    if '不归档' in s or s.endswith('-删除') or s.endswith(' (2)'):
        return 'I-占位', '不归档/已删除/重复副本'

    # 2. H 辅助说明
    if any(k in s for k in ['修订说明', '编制说明', '填表说明', '文号规则']):
        return 'H-辅助说明', '说明类'
    if s.startswith('参考-') or s.startswith('参考中'):
        return 'H-辅助说明', '参考引用'
    if re.match(r'^\d+\.\s', s) and len(s) < 30:
        return 'H-辅助说明', '编号说明段'
    if s == '说明' or s.endswith('说明') and len(s) < 15:
        return 'H-辅助说明', '说明'

    # 3. A 程序表（最优先于 D，因为程序表常含"检查"等关键词）
    if '实质性程序表' in s:
        return 'A-实质性程序', '实质性程序表'
    if re.search(r'[A-S]\d+A\s*$', s):
        return 'A-程序表', '末尾 *A'
    if '替代程序' in s:
        return 'A-替代程序', '替代程序'
    if '程序表' in s and '汇总' not in s:
        return 'A-一般程序表', '程序表'

    # 4. B 底稿目录
    if '底稿目录' in s or (s.endswith('目录') and len(s) < 20):
        return 'B-底稿目录', '底稿目录'

    # 5. C 附注披露
    if '附注披露' in s or '披露信息' in s:
        return 'C-附注披露', '附注披露'

    # 6. E 控制测试
    if '控制测试' in s and '汇总' in s:
        return 'E-控制测试汇总', 'C 控制测试汇总'
    if '控制测试' in s:
        return 'E-控制测试单条', 'C 控制测试单项'
    if '评价控制偏差' in s or '评价偏差' in s or features['has_step_words']:
        return 'E-评价控制偏差', '控制偏差/含步骤词'

    # 7. D 检查/政策/业务模式/函证/盘点/访谈/复核（多子类）
    if '业务模式' in s:
        return 'D-业务模式', '业务模式'
    if '会计政策' in s or '政策检查' in s:
        return 'D-政策检查', '政策检查'
    if '函证' in s or '询证' in s or '回函' in s or '发函记录' in s or '银行账户情况承诺' in s:
        return 'D-函证', '函证'
    if '盘点' in s or '监盘' in s:
        return 'D-盘点', '盘点'
    if '访谈' in s or '调查问卷' in s:
        return 'D-访谈', '访谈'
    if '复核' in s and ('记录' in s or '签发' in s or '复核表' in s or '项目经理' in s or '合伙人' in s):
        return 'D-复核记录', '复核记录'
    if '风险评估' in s or ('特别风险' in s) or ('舞弊风险' in s):
        return 'D-风险评估', '风险评估'
    if '互转审核' in s:
        return 'D-检查表', '互转审核'
    if '识别未披露' in s or '核查清单' in s:
        return 'D-检查表', '核查清单/识别未披露'
    if '系统核对' in s or 'ERP' in s.upper() or ('流水' in s and '核对' in s) or '对账' in s:
        return 'D-系统核对', '系统/流水核对'
    if '检查表' in s:
        return 'D-检查表', '检查表'
    if re.search(r'[检核查][\u4e00-\u9fff]+表', s):
        return 'D-检查表', '检查类'
    # B 循环子区：B10-B30 业务流程了解 / B11-X 信息摘录
    if re.match(r'^B(10|11|12|13|18|20|21|22|23|24|25|30|40|50|51|52)', s) and ('了解' in s or '摘录' in s or '记录' in s or '评价' in s):
        return 'D-业务了解', 'B 循环业务了解'

    # 7.1 补丁：S 循环主程序表（如"违反法规行为S1"/"对环境事项的考虑S10"）
    if re.search(r'S\d+\s*$', s) and (('行为' in s or '事项' in s or '考虑' in s or '审计' in s)):
        return 'A-S 议题程序', 'S 议题主程序'
    # 7.2 补丁：C22 IT 控制测试（SA-X / PE-X / PM-X / RP 编号）
    if re.match(r'^(SA|PE|PM|RP|NS|IC|CIA)[\s\-]', s) or re.match(r'^(SA|PE|PM|RP|NS|IC|CIA)\d', s):
        return 'E-IT 控制测试', 'IT 控制测试编码'
    # 7.3 补丁：分析性复核（横向/纵向/比率/财务比率分析）
    if any(k in s for k in ['横向分析', '纵向分析', '比率分析', '同行业对比']):
        return 'F-分析表', '分析性复核'
    if '分析' in s and features['formula_cells'] >= 5:
        return 'F-分析表', '含公式的分析'
    # 7.4 补丁：实质性分析 / 业务分析（K8-4/K9-4/F1-4/F4-4 等）
    if '实质性分析' in s:
        return 'F-分析表', '实质性分析'
    # 7.5 补丁：合同检查/逾期检查（L1-6/L3-6/F3-5 等）
    if '合同检查' in s or '逾期' in s and '检查' in s:
        return 'D-检查表', '合同/逾期检查'
    # 7.6 补丁：贷款合同/账面核对（L 循环）
    if '贷款合同' in s or '账面核对' in s:
        return 'D-检查表', '账面核对类'

    # 7.7 补丁：S 议题子程序（编号 S\d+-\d+）
    if re.match(r'^S\d+\-\d+', s) and '记录' in s:
        return 'D-访谈', 'S 议题沟通记录'
    if re.match(r'^S\d+\-\d+', s):
        return 'A-一般程序表', 'S 议题子程序'

    # 7.8 补丁：表头（请先填写）/相关资源 / 准则及应用指南 → 占位/辅助
    if '请先填写' in s or '表头' in s and len(s) < 20:
        return 'I-占位', '表头占位'
    if s == '相关资源' or '相关资源' in s and len(s) < 10:
        return 'H-辅助说明', '相关资源'
    if '准则' in s and ('指南' in s or '规定' in s):
        return 'H-辅助说明', '准则参考'
    if '指引' in s or '会计提示' in s:
        return 'H-辅助说明', '指引提示'
    if 'IPO' in s and ('提示' in s or '解答' in s or '审计重点' in s):
        return 'H-辅助说明', 'IPO 提示'
    if '环境法规' in s or '市场平均收益率' in s or '示例' in s and len(s) < 30:
        return 'H-辅助说明', '参考资料'
    if s == '说明' or s.endswith('规定') and len(s) < 15:
        return 'H-辅助说明', '说明/规定'
    # 7.9 补丁：参考-XXXX 准则/指引（无连字符版本）
    if s.startswith('参考') or '参考示例' in s:
        return 'H-辅助说明', '参考'

    # 8. G 测算表（含"测算/测试/计量/调节" + 公式 ≥ 5）
    if '测算' in s or '测试' in s:
        return 'G-测算', '测算/测试'
    if ('初始计量' in s or '后续计量' in s or '计量' in s) and features['formula_cells'] >= 3:
        return 'G-测算', '计量类含公式'
    if '简化的追溯调整法' in s or '追溯调整' in s:
        return 'G-测算', '追溯调整'
    if '资本化' in s and '判断' in s:
        return 'G-测算', '资本化判断'
    if '三阶段划分' in s:
        return 'G-测算', 'ECL 三阶段'
    if '公允价值' in s and '调节' in s:
        return 'G-测算', '公允价值调节'

    # 9. F 数据表
    if '审定表' in s or s == '审定':
        return 'F-审定表', '审定表'
    if '明细表' in s:
        return 'F-明细表', '明细表'
    if '分析表' in s or ('分析' in s and '比较' in s):
        return 'F-分析表', '分析表'
    if '调整分录' in s:
        return 'F-调整分录', '调整分录'
    if '汇总' in s and '表' in s:
        return 'F-汇总表', '汇总表'
    # 编号章节式（"七、发出商品"等）
    if re.match(r'^[一二三四五六七八九十]+、', s):
        return 'F-明细表', '编号章节明细'
    # CFS 编制类
    if 'CF' in s or '现金流量' in s:
        return 'F-CFS 编制', 'CFS 编制'

    # 9.1 补丁：A 程序表中的多公式 sheet（如"完成阶段的分析性复核A1-13"）
    if features['formula_cells'] >= 10:
        return 'F-数据表', '高公式密度数据表'
    # 9.2 补丁：含"检查"且公式 >= 5 的多归 F 数据表
    if '检查' in s and features['formula_cells'] >= 5:
        return 'F-检查数据表', '检查类含公式'
    # 9.3 补丁：A 程序表完成阶段（含审计程序文字但无 *A 后缀）
    if s.startswith('审计程序') or s.startswith('完成阶段'):
        return 'A-一般程序表', 'A 完成阶段程序'
    # 9.4 补丁：业务约定书 / 报告文号
    if '业务约定' in s or '报告文号' in s:
        return 'A-一般程序表', '业务约定/报告文号'

    # 9.5 补丁：A 类沟通/错报/缺陷评价子区
    if '沟通记录' in s or '沟通模板' in s:
        return 'D-访谈', '沟通记录/模板'
    if '错报' in s or '披露错报' in s:
        return 'D-检查表', '错报评价'
    if '缺陷评价' in s or '缺陷汇总' in s:
        return 'D-检查表', '内控缺陷评价'
    if '管理层凌驾' in s:
        return 'D-检查表', '管理层凌驾'
    if '识别' in s and '关联方' in s:
        return 'D-检查表', '识别关联方'
    if '业务承接' in s or '业务保持' in s:
        return 'D-检查表', '业务承接/保持'
    if '前任注册会计师' in s:
        return 'D-访谈', '与前任沟通'
    if '组成部分' in s and '注册会计师' in s:
        return 'D-访谈', '组成部分会计师'
    if '项目现场负责人' in s and '复核' in s:
        return 'D-复核记录', '项目现场负责人复核'
    if '了解企业层面控制' in s or '控制环境' in s or '风险评估流程' in s or '信息与沟通' in s and '详细' in s:
        return 'D-业务了解', '了解企业层面控制（详细法）'
    if '详细法' in s:
        return 'D-业务了解', '详细法了解'
    if '前瞻性调整' in s or '打分卡' in s or '量化分析' in s and features['formula_cells'] >= 5:
        return 'G-测算', '前瞻性 ECL 调整'
    # 9.6 补丁：B 循环 B22A-X / B19-X 子区（业务了解模板）
    if re.match(r'^B(19|20|22|23|24|25|26|27|28|29|30)[A-Z]?[\-\s]', s):
        return 'D-业务了解', 'B 循环了解子区'
    if re.match(r'^B[12]\w?\s', s) and ('承接' in s or '保持' in s or '沟通' in s):
        return 'D-检查表', 'B1/B2 业务表'
    # 9.7 补丁：A1-13 / A14 / A13 等子区（多公式归 F 数据表已被前规则覆盖，余下归对应类）
    if re.match(r'^A1[3-9]', s) or re.match(r'^A2[0-3]', s):
        return 'D-检查表', 'A1X/A2X 评价复核子区'
    # 9.8 补丁：B15 重要性计算表（保留 Univer，含公式）
    if 'B15' in s or '重要性' in s and '计算' in s:
        return 'F-数据表', 'B15 重要性计算'
    # 9.9 补丁：政府补助/特殊议题
    if '政府补助' in s:
        return 'D-检查表', '政府补助检查'
    if '资产负债表日后事' in s or '日后事项' in s:
        return 'D-检查表', '期后事项'
    # 9.10 补丁：知识产权许可 / 主要责任人代理人 等专项
    if '知识产权' in s or '主要责任人' in s or '代理人' in s:
        return 'A-S 议题程序', '专项议题分析'
    # 9.11 补丁：成本倒轧/银行余额调节/账面核对（多公式数据表）
    if '余额调节' in s or '倒轧' in s:
        return 'F-数据表', '调节/倒轧表'
    # 9.12 补丁：管理层提供的清单 / 自我评价报告参考格式
    if '清单' in s and '管理层' in s:
        return 'D-检查表', '管理层清单'
    if '参考格式' in s:
        return 'H-辅助说明', '参考格式'
    # 9.13 补丁：假期清单 / 政府文件等（参考资料）
    if re.search(r'\d{4}假期', s) or '假期清单' in s:
        return 'I-占位', '假期清单'

    # ==================== 第二批补丁（覆盖 88 个 single-cycle pending） ====================

    # 9.14 S 议题专项主程序（编号 S\d+ 末尾或 S\d+ 单独成 sheet）
    if re.match(r'^[一二三四五六七八九十]?[\u4e00-\u9fff、（）()0-9 ]*S\d+\s*$', s) or re.search(r'S\d+\s*$', s):
        return 'A-S 议题程序', 'S 议题主程序'
    # 9.15 S 议题"提示"页（短/中等长度，纯说明）
    if s.strip() == '提示':
        return 'H-辅助说明', 'S 议题提示'
    # 9.16 S 议题"披露格式参考 IC*-X" / "核查事项清单" / "【附】XXX"
    if re.search(r'披露格式参考', s) or '核查事项清单' in s or s.startswith('【附】'):
        return 'H-辅助说明', 'S 议题参考'
    # 9.17 S3 简化追溯调整 / 首次执行新准则的调整
    if '首次执行' in s and ('准则' in s or '调整' in s):
        return 'G-测算', 'S3 准则首次执行调整'
    if '会计估计变更' in s and '审计程序' in s:
        return 'A-S 议题程序', 'S3 会计估计变更程序'
    # 9.18 S 数据资产 / 套期活动 / 营业收入扣除核查 等
    if '数据资产' in s and '基本情况' in s:
        return 'D-检查表', 'S 数据资产基本情况'
    if '套期活动' in s:
        return 'A-S 议题程序', '套期活动'
    if '营业收入扣除' in s and '核查' in s:
        return 'D-检查表', 'S 营业收入扣除核查'
    if 'S' in s[:3] and '内控调查' in s:
        return 'D-检查表', 'S 内控调查表'
    # 9.19 S 议题环境法规 / 电子商务法律法规 等长篇法规列表
    if '法规' in s and ('环境' in s or '电子商务' in s):
        return 'H-辅助说明', '法规列表'

    # 9.20 IT 一般控制（B22A-4-4 系列 / IT 缺陷会议纪要等）
    if 'IT' in s and ('一般控制' in s or '缺陷' in s or '控制环境' in s):
        return 'D-业务了解', 'IT 控制了解'
    if 'ITGC' in s.upper() or 'SoD analysis' in s:
        return 'D-业务了解', 'IT 职责分离'
    if s in ('安全和维护', '作业调度和接口', '清单'):
        return 'D-业务了解', 'IT 一般控制子模块'
    if '财务报告过程' in s and '模板' in s:
        return 'D-业务了解', '财务报告过程'
    if 'T3' in s and '缺陷' in s and '评价' in s:
        return 'D-检查表', '缺陷一同评价'

    # 9.21 B 循环了解类（B10-X / B11-X / B19 / B22 / B50 / B51）
    if re.match(r'^B(10|11|19|22)', s) and ('结构' in s or '法律监管' in s or '信息' in s or '记录' in s or '清单' in s):
        return 'D-业务了解', 'B 循环了解'
    if 'B19' in s and '关联方清单' in s:
        return 'D-检查表', 'B19 关联方清单'
    if 'B50' in s and ('风险因素' in s or '财务报表层次' in s):
        return 'D-风险评估', 'B50 风险汇总'
    if 'B51' in s and '警觉' in s:
        return 'D-风险评估', 'B51 警觉情形'
    if 'B30' in s and ('特别风险' in s or 'Significant risk' in s):
        return 'D-风险评估', 'B30 特别风险记录'
    # B 业务承接保持/独立性（B1A/B1B/B3）
    if re.match(r'^B[12][AB]?', s) and ('承接' in s or '保持' in s or '独立性' in s):
        return 'D-检查表', 'B 业务承接保持'
    if 'B3' in s and '独立性' in s:
        return 'D-检查表', 'B3 独立性确认'
    if '管理层凌驾' in s:
        return 'D-业务了解', '管理层凌驾'

    # 9.22 B60 工时预算
    if 'B6' in s and ('工时' in s or '预算' in s):
        return 'F-数据表', 'B60 工时预算'

    # 9.23 A 循环错报评价 / 缺陷评价（A13 / A14）
    if re.match(r'^A1[34]', s):
        return 'D-检查表', 'A13/A14 错报缺陷评价'
    if 'IT控制缺陷沟通会议纪要' in s:
        return 'D-访谈', 'IT 缺陷沟通纪要'

    # 9.24 A1-11 报告签发 / A1-18 准则衔接
    if 'A1-11' in s or '报告签发' in s and len(s) < 20:
        return 'F-数据表', 'A1-11 报告签发流转'
    if 'A1-18' in s or '准则衔接影响' in s:
        return 'F-数据表', 'A1-18 准则衔接核对'

    # 9.25 A15 持续经营决策图
    if 'A15' in s or ('持续经营' in s and ('决策' in s or '影响' in s)):
        return 'D-检查表', 'A15 持续经营'

    # 9.26 A31 审计标识 / A30 归档检查
    if 'A31' in s or '审计标识' in s:
        return 'B-底稿目录', 'A31 审计标识'
    if 'A30' in s or '电子底稿归档' in s:
        return 'D-复核记录', 'A30 归档检查'

    # 9.27 A5 现金流量子区 / 财务承诺 / 或有事项
    if 'A5-1' in s and '勾稽' in s:
        return 'F-数据表', 'A5-1 现金流勾稽'
    if 'A5-2' in s or '财务承诺' in s:
        return 'D-检查表', 'A5-2 财务承诺'
    if 'A5-3' in s or '或有事项' in s and '程序' in s:
        return 'A-一般程序表', 'A5-3 或有事项程序'

    # 9.28 A7 关联交易汇总
    if 'A7' in s and ('关联交易' in s or '关联往来' in s):
        return 'F-审定表', 'A7 关联交易汇总'

    # 9.29 A10 治理层沟通模板
    if '沟通记录模板' in s:
        return 'D-访谈', 'A10 沟通记录模板'

    # 9.30 C 循环额外补丁（C21/C24/C25）
    if 'C21' in s and 'IT' in s and '专业技能' in s:
        return 'D-业务了解', 'C21 IT 专业技能项目组'
    if 'C24' in s and ('完整性' in s or '会计分录' in s):
        return 'F-检查数据表', 'C24 会计分录细测'
    if 'C25' in s or '利用内部审计' in s:
        return 'D-业务了解', 'C25 利用内审'

    # 9.31 D4 重要客户销售价格分析（不含"价格分析"通用规则覆盖）
    if 'D4-10' in s or ('重要客户' in s and '销售价格' in s):
        return 'F-分析表', 'D4 客户销售价格分析'

    # 9.32 E1 银行余额调节 / 银行流水
    if 'E1' in s and ('银行' in s and ('调节' in s or '查询' in s or '流水' in s)):
        return 'F-数据表', 'E1 银行核查'

    # 9.33 F2 关联方采购定价核查
    if 'F2' in s and '关联方采购定价' in s:
        return 'D-检查表', 'F2 关联方采购定价'

    # 9.34 F5 成本倒轧
    if 'F5' in s and '成本倒轧' in s:
        return 'F-数据表', 'F5 成本倒轧'

    # 9.35 G1 结存表 / G7 长投基础信息
    if 'G1' in s and '结存' in s:
        return 'F-明细表', 'G1 结存表'
    if 'G7' in s and ('被投资单位' in s or '初始判断' in s):
        return 'D-检查表', 'G7 长投基础'

    # 9.36 H2 在建工程审核 / H8 租赁底稿
    if 'H2' in s and '在建工程' in s and '审核' in s:
        return 'D-复核记录', 'H2 在建工程审核'
    if 'H8' in s and ('租赁的识别' in s or '租赁期' in s or '租赁变更' in s):
        return 'D-检查表', 'H8 租赁判断'

    # 9.37 J3 股份支付情况表
    if 'J3' in s and '股份支付情况' in s:
        return 'F-明细表', 'J3 股份支付情况'

    # 9.38 N2 应交税金认定 / N5 当期所得税
    if 'N2' in s and '应交税金' in s and '认定' in s:
        return 'D-检查表', 'N2 应交税金认定'
    if 'N5' in s and '当期所得税' in s:
        return 'G-测算', 'N5 当期所得税计算'

    # 9.39 S15 每股收益（基本+稀释）
    if 'S15' in s and ('每股收益' in s or 'EPS' in s.upper()):
        return 'G-测算', 'S15 每股收益计算'

    # ==================== 第三批补丁（覆盖剩余 11 个 single-cycle pending） ====================

    # 9.40 B22B 控制矩阵选项清单
    if '选项清单列表' == s.strip() or 'B22B' in s:
        return 'I-占位', 'B22B 控制矩阵选项'

    # 9.41 B50 相关资源（固有风险因素参考）
    if '相关资源' in s and ('固有风险' in s or '风险因素' in s):
        return 'H-辅助说明', 'B50 风险因素参考'

    # 9.42 C21 IT 专业技能（C21 主程序）
    if 'C21' in s:
        return 'A-一般程序表', 'C21 IT 专业技能项目组主表'

    # 9.43 E1 信用报告查询 / 关键岗位流水
    if 'E1' in s and ('信用报告' in s or '董监高' in s or '关键岗位' in s):
        return 'D-检查表', 'E1 信用/流水核查'

    # 9.44 IPO 首发业务解答（参考）
    if '首发业务解答' in s:
        return 'H-辅助说明', 'IPO 首发业务解答参考'

    # 9.45 S 议题"附、内控调查表" / "内部控制调查表"
    if '内控调查表' in s or '内部控制调查表' in s:
        return 'D-检查表', 'S 内控调查表'

    # 9.46 S4/S5 商业实质判断 / 债务重组损益时点
    if '商业实质' in s and '判断' in s:
        return 'D-业务模式', 'S4 商业实质判断'
    if '损益确认时点' in s or '损益确认' in s:
        return 'D-业务模式', 'S5 损益确认时点'

    # 9.47 会计监管风险提示（参考资料类）
    if '会计监管风险提示' in s or '风险提示第' in s:
        return 'H-辅助说明', '监管风险提示'

    # 10. 兜底标记（待人工归类）
    return '_pending', '未匹配规则'


def recommend_render(class_code, features):
    """根据归类 + 特征推荐渲染策略"""
    if class_code.startswith('A-'):
        return 'HTML 中控台'
    if class_code.startswith('B-'):
        return 'HTML 表单（编制信息+索引导航）'
    if class_code.startswith('C-'):
        return 'HTML 嵌套表（多级子表）'
    if class_code.startswith('D-'):
        if class_code == 'D-业务模式':
            return 'HTML 是否问答型'
        if class_code == 'D-政策检查':
            return 'HTML 段落型'
        if class_code in ('D-函证', 'D-盘点', 'D-访谈'):
            return 'HTML 表单（专属子组件）'
        if class_code == 'D-复核记录':
            return 'HTML 表单（电子签）'
        return 'HTML 表单（表格型检查）'
    if class_code.startswith('E-'):
        if class_code == 'E-评价控制偏差':
            return 'HTML stepper'
        return 'HTML 表单'
    if class_code.startswith('F-'):
        return '保留 Univer'
    if class_code.startswith('G-'):
        return '保留 Univer（测算）'
    if class_code.startswith('H-'):
        return '静态展示'
    if class_code.startswith('I-'):
        return '跳过渲染'
    return 'PENDING-待人工归类'


def main():
    if not TEMPLATE_ROOT.exists():
        print(f"ERROR: 模板目录不存在: {TEMPLATE_ROOT}")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    result = {
        'meta': {
            'generated_at': datetime.now().isoformat(),
            'template_root': str(TEMPLATE_ROOT),
            'total_xlsx': 0,
            'total_sheets': 0,
        },
        'cycles': {},
        'class_summary': Counter(),
        'render_summary': Counter(),
        'pending_sheets': [],  # 兜底标记的 sheet
    }

    cycle_dirs = sorted([d for d in TEMPLATE_ROOT.iterdir() if d.is_dir()])

    for cycle_dir in cycle_dirs:
        cycle = cycle_dir.name
        if cycle.startswith('_'):
            cycle_key = cycle  # _reference 等保留
        else:
            cycle_key = cycle

        cycle_data = {
            'cycle': cycle_key,
            'cycle_name': CYCLE_NAMES.get(cycle_key, cycle_key),
            'templates': [],
        }

        xlsx_files = sorted(cycle_dir.glob('*.xlsx'))
        for xlsx_path in xlsx_files:
            result['meta']['total_xlsx'] += 1
            template_data = {
                'filename': xlsx_path.name,
                'sheets': [],
            }
            try:
                wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)
                for sheet_name in wb.sheetnames:
                    result['meta']['total_sheets'] += 1
                    ws = wb[sheet_name]
                    features = extract_sheet_features(ws)
                    class_code, reason = classify_sheet(sheet_name, features)
                    render = recommend_render(class_code, features)

                    sheet_record = {
                        'name': sheet_name,
                        'class': class_code,
                        'reason': reason,
                        'render': render,
                        'features': features,
                    }
                    template_data['sheets'].append(sheet_record)
                    result['class_summary'][class_code] += 1
                    result['render_summary'][render] += 1

                    if class_code == '_pending':
                        result['pending_sheets'].append({
                            'cycle': cycle_key,
                            'template': xlsx_path.name,
                            'sheet': sheet_name,
                            'features': features,
                        })
                wb.close()
            except Exception as e:
                template_data['error'] = str(e)
                print(f"  WARN 解析失败 {xlsx_path.name}: {e}")

            cycle_data['templates'].append(template_data)

        result['cycles'][cycle_key] = cycle_data
        print(f"  扫描 {cycle_key}: {len(xlsx_files)} 个 xlsx")

    # 转 Counter 为 dict
    result['class_summary'] = dict(result['class_summary'])
    result['render_summary'] = dict(result['render_summary'])

    # 写 JSON
    with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\n[OK] JSON: {JSON_OUTPUT}")

    # 总览输出
    print(f"\n=== 总览 ===")
    print(f"模板 xlsx: {result['meta']['total_xlsx']}")
    print(f"sheets: {result['meta']['total_sheets']}")
    print(f"\n=== 归类分布 ===")
    for cls, cnt in sorted(result['class_summary'].items(), key=lambda x: -x[1]):
        pct = cnt / result['meta']['total_sheets'] * 100
        print(f"  {cls}: {cnt} ({pct:.1f}%)")
    print(f"\n=== 渲染策略分布 ===")
    for r, cnt in sorted(result['render_summary'].items(), key=lambda x: -x[1]):
        pct = cnt / result['meta']['total_sheets'] * 100
        print(f"  {r}: {cnt} ({pct:.1f}%)")
    print(f"\n待人工归类 sheet: {len(result['pending_sheets'])}")


if __name__ == '__main__':
    main()
