"""调试：对比 read_only 和完整模式的解析差异"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from app.services.smart_import_engine import detect_header_rows, merge_header_rows

filepath = r"基础数据\和平药房2024\科目余额表-重庆和平药房连锁有限责任公司2024.xlsx"

print("=== read_only=True ===")
wb1 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
ws1 = wb1.worksheets[0]
# 打印前5行原始数据
for i, row in enumerate(ws1.iter_rows(max_row=5, values_only=True)):
    cells = [str(c)[:30] if c else "" for c in row]
    print(f"  Row {i}: {cells[:8]}... (len={len(row)})")
# 表头检测
ws1_b = wb1.worksheets[0]  # 需要重新获取（iter_rows 消耗了迭代器）
wb1.close()

# 重新打开
wb1 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
ws1 = wb1.worksheets[0]
hs, hc = detect_header_rows(ws1)
print(f"  header_start={hs}, header_count={hc}")
headers = merge_header_rows(ws1, hs, hc)
print(f"  headers: {headers[:8]}")
wb1.close()

print("\n=== read_only=False (完整模式) ===")
wb2 = openpyxl.load_workbook(filepath, data_only=True)
ws2 = wb2.worksheets[0]
for i in range(1, 6):
    cells = [str(ws2.cell(i, c).value)[:30] if ws2.cell(i, c).value else "" for c in range(1, 15)]
    print(f"  Row {i}: {cells[:8]}...")
hs2, hc2 = detect_header_rows(ws2)
print(f"  header_start={hs2}, header_count={hc2}")
headers2 = merge_header_rows(ws2, hs2, hc2)
print(f"  headers: {headers2[:8]}")
wb2.close()
