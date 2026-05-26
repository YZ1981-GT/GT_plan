"""
generate_wp_render_schema.py — 从模板 xlsx 反向生成 wp_render_schema YAML 草稿

设计目标（design §10.1 H1 风险缓解）：
- 减少手工维护成本：1346 sheet × 50-80 wp_code 全人工编写 schema 不现实
- 复用 analyze_wp_templates 既有 60+ 归类规则（避免重复实现）
- 输出 YAML 草稿 + `# TODO: 人工审核` 标注，提示需人工修正的关键字段
- 不覆盖手工编写的 YAML（默认 skip-existing），生成结果存放 generated/ 子目录

字段提取策略（每 sheet）：
- fixed_cells: A1-A4 / 索引号 / 页码 等模板头部（启发式扫描）
- dynamic_table: 检测表头行 + 数据起始行 + 列字段（基于合并单元格 + 列标题）
- formulas: 收集所有公式 cell（preserve）
- merged_cells: 计数 + 主要范围
- static_text: 检测纯文本行（无公式无输入）
- cross_refs: 留空（# TODO 人工补充）

用法（多场景示例）：

    # 1. 帮助
    python backend/scripts/generate_wp_render_schema.py --help

    # 2. 干跑（只统计不写文件）
    python backend/scripts/generate_wp_render_schema.py --dry-run

    # 3. 单 wp_code 过滤（生成 D2 系列）
    python backend/scripts/generate_wp_render_schema.py --wp-code-filter D2 --dry-run

    # 4. 实际生成（默认 skip-existing，跳过 backend/data/wp_render_schema/*.yaml 已存在的 wp_code）
    python backend/scripts/generate_wp_render_schema.py

    # 5. 强制覆盖（注意：不覆盖 generated/ 之外的手工 YAML）
    python backend/scripts/generate_wp_render_schema.py --overwrite

    # 6. 自定义输入输出目录
    python backend/scripts/generate_wp_render_schema.py \
        --template-dir backend/wp_templates/D \
        --output-dir backend/data/wp_render_schema/generated

输出：
- 默认输出至 backend/data/wp_render_schema/generated/{wp_code}.yaml（与手工 YAML 隔离）
- 生成 _generation_report.json 包含统计信息（成功/跳过/失败）

Requirements: 2.2 原则 2（配置驱动）+ design §10.1 H1 风险缓解
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl pyyaml", file=sys.stderr)
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("ERROR: pyyaml not installed. Run: pip install pyyaml", file=sys.stderr)
    sys.exit(1)

# 复用 analyze_wp_templates 的归类逻辑
SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from analyze_wp_templates import (  # noqa: E402
    classify_sheet,
    extract_sheet_features,
    recommend_render,
)

logger = logging.getLogger(__name__)

# ─── 默认路径 ─────────────────────────────────────────────────────
WORKSPACE_ROOT = SCRIPT_DIR.parent.parent  # d:/GT_plan
DEFAULT_TEMPLATE_ROOT = WORKSPACE_ROOT / "backend" / "wp_templates"
DEFAULT_OUTPUT_DIR = WORKSPACE_ROOT / "backend" / "data" / "wp_render_schema" / "generated"
HANDCRAFTED_DIR = WORKSPACE_ROOT / "backend" / "data" / "wp_render_schema"

# ─── 9 类 → componentType 映射（与 wp_classification_service 一致） ───
_CLASS_TO_COMPONENT: dict[str, str] = {
    "A-": "a-program-console",
    "B-": "b-index",
    "C-": "c-note-table",
    "E-": "e-control-test",
    "F-": "univer",
    "G-": "univer",
    "H-": "h-static-doc",
    "I-": "skip",
}

_D_SUB_ROUTING: dict[str, str] = {
    "D-函证": "d-form-confirmation",
    "D-盘点": "d-form-confirmation",
    "D-访谈": "d-form-confirmation",
    "D-政策检查": "d-form-paragraph",
    "D-业务模式": "d-form-qa",
    "D-复核记录": "d-form-review",
}

_D_DEFAULT = "d-form-table"

# ─── wp_code 提取（从文件名前缀） ──────────────────────────────
_WP_CODE_PATTERN = re.compile(r"^([A-Z]\d+(?:[A-Z]?(?:-\d+)?)?)")


def derive_component_type(class_code: str | None) -> str:
    """归类 → componentType 映射（D 类有子路由）"""
    if not class_code:
        return "univer"
    if class_code.startswith("D-"):
        return _D_SUB_ROUTING.get(class_code, _D_DEFAULT)
    for prefix, ctype in _CLASS_TO_COMPONENT.items():
        if class_code.startswith(prefix):
            return ctype
    return "univer"


def extract_wp_code_from_filename(filename: str) -> str | None:
    """从模板文件名提取 wp_code（如 'D2-1至D2-4 应收账款...xlsx' → 'D2'）"""
    m = _WP_CODE_PATTERN.match(filename)
    if m:
        return m.group(1).split("-")[0]  # 去掉子序号，保留主 wp_code
    return None


def detect_fixed_cells(ws) -> dict[str, str]:
    """启发式检测固定头部 cell

    常见模式：
    - A1: 事务所名称
    - A2: 表名
    - A3: 被审计单位（公式 =底稿目录!A2）
    - A4: 截止日（公式 =底稿目录!A3）
    - 右上角 (H/I/J列 row 3-4): 索引号 / 页码
    """
    fixed: dict[str, str] = {}

    # A1-A4 标准头部
    for row in (1, 2, 3, 4):
        cell = ws.cell(row=row, column=1)
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        addr = f"A{row}"
        if row == 1 and "致同" in s:
            fixed[addr] = "致同会计师事务所"
        elif row == 2 and len(s) < 50 and not s.startswith("="):
            fixed[addr] = s  # 表名（保留模板默认值）
        elif row == 3:
            # =底稿目录!A2 → ${entity_name}
            if isinstance(v, str) and v.startswith("="):
                fixed[addr] = "${entity_name}"
            elif s:
                fixed[addr] = "${entity_name}"
        elif row == 4:
            if isinstance(v, str) and v.startswith("="):
                fixed[addr] = "${period_end}"
            elif s:
                fixed[addr] = "${period_end}"

    # 索引号 / 页码（扫 row 3-4 右半部分 col 5-15）
    max_col_scan = min(ws.max_column or 0, 15)
    for row in (3, 4):
        for col in range(5, max_col_scan + 1):
            cell = ws.cell(row=row, column=col)
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            addr = f"{get_column_letter(col)}{row}"
            if row == 3 and ("索引" in s or re.match(r"^[A-Z]\d+", s)):
                fixed[addr] = "${index_no}"
            elif row == 4 and ("页" in s or re.match(r"^\d+/\d+", s)):
                fixed[addr] = "${page_no}"

    return fixed


def detect_dynamic_table(ws, max_scan_rows: int = 200) -> dict[str, Any] | None:
    """检测动态表格区（表头行 + 数据起始行 + 列字段）

    启发式策略：
    1. 扫描前 30 行找含多个非空列标题的行（候选表头）
    2. 表头行下一行作为数据起始行
    3. 数据行扫描到底（end_row: dynamic）
    4. 列字段名取表头文本（人工后续修正）
    """
    max_row = min(ws.max_row or 0, max_scan_rows)
    max_col = min(ws.max_column or 0, 30)
    if max_row < 3 or max_col < 2:
        return None

    # 一次性扫表前 30 行（避免 read_only 随机访问陷阱）
    rows_data: dict[int, dict[int, Any]] = {}
    for row_cells in ws.iter_rows(min_row=1, max_row=min(30, max_row), values_only=False):
        for cell in row_cells:
            if cell.value is not None:
                rows_data.setdefault(cell.row, {})[cell.column] = cell.value

    # 找候选表头行：>= 3 个非空 cell 且前一行较空（结构边界）
    header_row = None
    for row_idx in sorted(rows_data.keys()):
        if row_idx < 5 or row_idx > 25:  # 表头通常在 5-25 行
            continue
        non_empty = sum(
            1
            for v in rows_data[row_idx].values()
            if v is not None and str(v).strip() and not (isinstance(v, str) and v.startswith("="))
        )
        if non_empty >= 3:
            # 列标题特征：短文本（< 30 字符）且不是公式
            short_count = sum(
                1
                for v in rows_data[row_idx].values()
                if isinstance(v, (str, int, float))
                and len(str(v)) < 30
                and not str(v).startswith("=")
                and str(v).strip()
            )
            if short_count >= 3:
                header_row = row_idx
                break

    if header_row is None:
        return None

    # 提取列字段（启发式 field 名：从中文标题派生 snake_case）
    columns: dict[str, dict[str, Any]] = {}
    header_cells = rows_data.get(header_row, {})
    for col_idx in sorted(header_cells.keys())[:max_col]:
        v = header_cells[col_idx]
        if v is None or not str(v).strip():
            continue
        label = str(v).strip()[:30]
        col_letter = get_column_letter(col_idx)
        columns[col_letter] = {
            "field": f"col_{col_letter.lower()}",  # TODO: 人工改 snake_case 字段名
            "type": "text",
            "label": label,
        }

    if not columns:
        return None

    return {
        "start_row": header_row + 1,
        "end_row": "dynamic",  # 数据行数由用户决定
        "header_row": header_row,
        "columns": columns,
    }


def detect_formulas(ws, max_scan_rows: int = 300) -> list[str]:
    """收集所有公式 cell 地址"""
    formulas: list[str] = []
    max_row = min(ws.max_row or 0, max_scan_rows)
    max_col = min(ws.max_column or 0, 50)
    for row_cells in ws.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False
    ):
        for cell in row_cells:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                formulas.append(cell.coordinate)
    return formulas


def detect_merged_info(ws) -> dict[str, Any]:
    """合并单元格信息"""
    merged_ranges = list(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else []
    return {
        "count": len(merged_ranges),
        "ranges_sample": [str(r) for r in merged_ranges[:10]],  # 前 10 个用于人工核对
    }


def detect_static_text_rows(ws, header_row: int | None, max_scan_rows: int = 200) -> list[Any]:
    """检测静态文本行（无公式无输入位的纯文本说明行）

    策略：
    - 表头之前 + 长段落 row（> 50 字符 + 1-2 列含值 + 无公式）
    - 输出 row 范围（list of int / str like '5-13'）
    """
    static_rows: list[int] = []
    max_row = min(ws.max_row or 0, max_scan_rows)
    if header_row is None:
        boundary = 16  # 默认假设
    else:
        boundary = header_row

    for row_cells in ws.iter_rows(
        min_row=1, max_row=min(boundary, max_row), min_col=1, max_col=15, values_only=False
    ):
        if not row_cells:
            continue
        row_idx = row_cells[0].row
        non_empty = [c for c in row_cells if c.value is not None and str(c.value).strip()]
        if not non_empty:
            continue
        # 含公式 = 不算静态文本
        has_formula = any(isinstance(c.value, str) and c.value.startswith("=") for c in non_empty)
        if has_formula:
            continue
        # 至少有一个长文本 cell（说明性段落）
        has_long = any(isinstance(c.value, str) and len(c.value) > 50 for c in non_empty)
        # 或仅 1-2 列填充（章节标题）
        few_cells = len(non_empty) <= 2
        if has_long or few_cells:
            static_rows.append(row_idx)

    # 折叠连续 row 为范围字符串
    if not static_rows:
        return []
    static_rows.sort()
    ranges: list[Any] = []
    start = prev = static_rows[0]
    for r in static_rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        ranges.append(f"{start}-{prev}" if start != prev else start)
        start = prev = r
    ranges.append(f"{start}-{prev}" if start != prev else start)
    return ranges


def derive_applicable_standards(class_code: str | None, wp_code: str) -> list[str]:
    """启发式推导 applicable_standards

    保守策略：返回 4 标准全集，由人工根据底稿性质收窄。
    """
    return [
        "soe_standalone",
        "listed_standalone",
        "soe_consolidated",
        "listed_consolidated",
    ]


def build_sheet_schema(
    ws,
    sheet_name: str,
    class_code: str | None,
    component_type: str,
    cycle_letter: str,
) -> dict[str, Any]:
    """构造单 sheet 的 schema dict"""
    fixed_cells = detect_fixed_cells(ws)
    dynamic_table = detect_dynamic_table(ws)
    formulas = detect_formulas(ws)
    merged_info = detect_merged_info(ws)
    static_rows = detect_static_text_rows(
        ws, dynamic_table["header_row"] if dynamic_table else None
    )

    sheet_schema: dict[str, Any] = OrderedDict()
    sheet_schema["component_type"] = component_type

    if class_code:
        sheet_schema["class_code"] = class_code

    if fixed_cells:
        sheet_schema["fixed_cells"] = fixed_cells

    if dynamic_table:
        sheet_schema["dynamic_table"] = dynamic_table
    else:
        sheet_schema["dynamic_table"] = (
            "# TODO: 人工审核 — 未自动检测到表格结构，请手工定义 start_row/header_row/columns"
        )

    if static_rows:
        sheet_schema["static_text"] = {
            "rows": static_rows,
            "description": "# TODO: 人工审核 — 自动检测的静态文本行范围",
        }

    if formulas:
        sheet_schema["formulas"] = {
            "preserve": True,
            "cells": formulas[:30],  # 限制前 30 个，避免 YAML 过长
            "_total_count": len(formulas),
        }

    if merged_info["count"] > 0:
        sheet_schema["merged_cells"] = {
            "preserve": True,
            "count": merged_info["count"],
            "ranges_sample": merged_info["ranges_sample"],
        }

    # cross_refs 留 TODO（自动检测精度不够，需人工根据 cross_wp_references.json 补）
    sheet_schema["cross_refs"] = (
        "# TODO: 人工审核 — 跨底稿引用关系需查 cross_wp_references.json 后手工补充"
    )

    return sheet_schema


def build_workbook_schema(
    xlsx_path: Path,
    wp_code: str,
    template_version: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """生成单 xlsx 的整本 schema

    Returns:
        (schema_dict, stats_dict)
    """
    rel_path = xlsx_path.relative_to(WORKSPACE_ROOT).as_posix()
    cycle_letter = wp_code[0] if wp_code else "?"

    schema: dict[str, Any] = OrderedDict()
    schema["wp_code"] = wp_code
    schema["template_path"] = rel_path
    schema["template_version"] = template_version
    schema["applicable_standards"] = derive_applicable_standards(None, wp_code)
    schema["_generated"] = {
        "tool": "generate_wp_render_schema.py",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_xlsx": xlsx_path.name,
        "_note": "此 YAML 由工具自动生成，关键字段（dynamic_table 边界 / cross_refs / "
        "applicable_standards 收窄）需人工审核修正",
    }

    sheets_section: dict[str, Any] = OrderedDict()
    stats = {"sheets_total": 0, "sheets_classified": 0, "sheets_pending": 0, "errors": []}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=False)
    except Exception as e:
        stats["errors"].append(f"load_workbook 失败: {e}")
        return schema, stats

    for sheet_name in wb.sheetnames:
        stats["sheets_total"] += 1
        try:
            ws = wb[sheet_name]
            features = extract_sheet_features(ws)
            class_code, _reason = classify_sheet(sheet_name, features)
            component_type = derive_component_type(class_code)
            if class_code and not class_code.startswith("_"):
                stats["sheets_classified"] += 1
            else:
                stats["sheets_pending"] += 1

            sheet_schema = build_sheet_schema(
                ws, sheet_name, class_code, component_type, cycle_letter
            )
            sheets_section[sheet_name] = sheet_schema
        except Exception as e:
            stats["errors"].append(f"sheet={sheet_name}: {e}")
            sheets_section[sheet_name] = {
                "component_type": "univer",
                "_error": f"# TODO: 人工审核 — 自动生成失败: {e}",
            }

    wb.close()
    schema["sheets"] = sheets_section
    return schema, stats


# 注册 OrderedDict 的 YAML 表示器（保持字段顺序）
def _represent_ordered_dict(dumper, data):
    return dumper.represent_dict(data.items())


yaml.SafeDumper.add_representer(OrderedDict, _represent_ordered_dict)


def write_yaml(schema: dict[str, Any], output_path: Path) -> None:
    """写 YAML 文件，附带头部注释"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    header = (
        f"# {output_path.name}\n"
        f"# 由 backend/scripts/generate_wp_render_schema.py 自动生成\n"
        f"# 生成时间: {datetime.now().isoformat(timespec='seconds')}\n"
        "# Requirements: 2.2 原则 2（配置驱动）\n"
        "# ⚠️  此文件为草稿，关键字段需人工审核：\n"
        "#   - dynamic_table 边界（start_row/end_row/header_row）\n"
        "#   - cross_refs 跨底稿引用（查 cross_wp_references.json）\n"
        "#   - applicable_standards 适用准则收窄（4 标准全集 → 实际适用）\n"
        "#   - columns 字段名（col_a/col_b/... → 业务字段 snake_case）\n"
        "#   - fixed_cells 占位符（${entity_name} 等是否符合实际模板）\n\n"
    )

    body = yaml.safe_dump(
        schema,
        allow_unicode=True,
        sort_keys=False,
        default_flow_style=False,
        width=120,
    )

    output_path.write_text(header + body, encoding="utf-8")


def collect_handcrafted_wp_codes() -> set[str]:
    """收集手工 YAML 已覆盖的 wp_code（避免覆盖手工成果）"""
    if not HANDCRAFTED_DIR.exists():
        return set()
    codes: set[str] = set()
    for yaml_path in HANDCRAFTED_DIR.glob("*.yaml"):
        try:
            data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
            if isinstance(data, dict) and "wp_code" in data:
                code = str(data["wp_code"])
                codes.add(code)
                # 也加入主 wp_code（如 D-D2-13.yaml 的 wp_code='D-D2-13'，主 wp_code='D2'）
                m = _WP_CODE_PATTERN.match(code)
                if m:
                    codes.add(m.group(1).split("-")[0])
        except Exception as e:
            logger.warning("解析手工 YAML 失败 %s: %s", yaml_path, e)
    return codes


def iter_template_files(
    template_dir: Path, wp_code_filter: str | None
) -> list[tuple[str, Path]]:
    """枚举 (wp_code, xlsx_path) 列表（按 wp_code 去重，每个主 wp_code 仅取首个文件）"""
    if not template_dir.exists():
        raise FileNotFoundError(f"模板目录不存在: {template_dir}")

    seen_codes: dict[str, Path] = {}
    for xlsx_path in sorted(template_dir.rglob("*.xlsx")):
        # 跳过临时文件
        if xlsx_path.name.startswith("~$") or xlsx_path.name.startswith("."):
            continue
        wp_code = extract_wp_code_from_filename(xlsx_path.name)
        if not wp_code:
            continue
        if wp_code_filter and not wp_code.startswith(wp_code_filter):
            continue
        # 去重：同 wp_code 多个文件（如 D2-1至D2-4 / D2-5 / D2-6至D2-13）只保留第一个
        # （后续 schema 中 sheets section 是按 sheet_name 区分，不影响）
        seen_codes.setdefault(wp_code, xlsx_path)
    return list(seen_codes.items())


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="从 xlsx 模板反向生成 wp_render_schema YAML 草稿",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--template-dir",
        type=Path,
        default=DEFAULT_TEMPLATE_ROOT,
        help=f"模板根目录（默认 {DEFAULT_TEMPLATE_ROOT}）",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"输出目录（默认 {DEFAULT_OUTPUT_DIR}）",
    )
    parser.add_argument(
        "--wp-code-filter",
        type=str,
        default=None,
        help="按 wp_code 前缀过滤（如 'D2' / 'A' / 'E1'）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只统计不写文件",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="强制覆盖已存在的 YAML（默认 skip-existing）",
    )
    parser.add_argument(
        "--include-handcrafted",
        action="store_true",
        help="包含手工 YAML 已覆盖的 wp_code（默认排除避免覆盖手工成果）",
    )
    parser.add_argument(
        "--template-version",
        type=str,
        default="v2025-R5",
        help="模板版本号（默认 v2025-R5）",
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="详细日志",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    # 收集手工 YAML 覆盖的 wp_code（避免重复生成）
    handcrafted = set() if args.include_handcrafted else collect_handcrafted_wp_codes()
    if handcrafted:
        logger.info("手工 YAML 已覆盖 %d 个 wp_code: %s", len(handcrafted), sorted(handcrafted))

    # 枚举模板文件
    try:
        targets = iter_template_files(args.template_dir, args.wp_code_filter)
    except FileNotFoundError as e:
        logger.error(str(e))
        return 2

    logger.info("发现 %d 个 wp_code 候选（去重后）", len(targets))

    # 全局统计
    summary: dict[str, Any] = {
        "tool": "generate_wp_render_schema.py",
        "args": {
            "template_dir": str(args.template_dir),
            "output_dir": str(args.output_dir),
            "wp_code_filter": args.wp_code_filter,
            "dry_run": args.dry_run,
            "overwrite": args.overwrite,
            "include_handcrafted": args.include_handcrafted,
        },
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stats": {
            "candidates": len(targets),
            "skipped_handcrafted": 0,
            "skipped_existing": 0,
            "generated": 0,
            "failed": 0,
        },
        "class_summary": Counter(),
        "files": [],
    }

    for wp_code, xlsx_path in targets:
        if wp_code in handcrafted:
            summary["stats"]["skipped_handcrafted"] += 1
            logger.debug("[skip 手工] %s", wp_code)
            continue

        output_path = args.output_dir / f"{wp_code}.yaml"
        if output_path.exists() and not args.overwrite:
            summary["stats"]["skipped_existing"] += 1
            logger.debug("[skip 已存在] %s → %s", wp_code, output_path)
            continue

        try:
            logger.info("生成 %s ← %s", wp_code, xlsx_path.name)
            schema, stats = build_workbook_schema(
                xlsx_path, wp_code, args.template_version
            )

            # 统计 class_code
            for sheet_data in schema.get("sheets", {}).values():
                if isinstance(sheet_data, dict) and "class_code" in sheet_data:
                    summary["class_summary"][sheet_data["class_code"]] += 1

            file_record = {
                "wp_code": wp_code,
                "source_xlsx": xlsx_path.name,
                "output_path": str(output_path.relative_to(WORKSPACE_ROOT)),
                "sheets_total": stats["sheets_total"],
                "sheets_classified": stats["sheets_classified"],
                "sheets_pending": stats["sheets_pending"],
                "errors": stats["errors"],
            }

            if args.dry_run:
                logger.info(
                    "  [dry-run] %d sheet（已分类 %d / pending %d）",
                    stats["sheets_total"],
                    stats["sheets_classified"],
                    stats["sheets_pending"],
                )
            else:
                write_yaml(schema, output_path)
                logger.info(
                    "  ✓ 写入 %s（%d sheet）",
                    output_path.relative_to(WORKSPACE_ROOT),
                    stats["sheets_total"],
                )

            summary["stats"]["generated"] += 1
            summary["files"].append(file_record)

        except Exception as e:
            logger.exception("生成失败 %s: %s", wp_code, e)
            summary["stats"]["failed"] += 1
            summary["files"].append({"wp_code": wp_code, "error": str(e)})

    # 写入总报告（即使 dry-run 也写，便于审阅）
    if not args.dry_run:
        report_path = args.output_dir / "_generation_report.json"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        # Counter → dict for JSON serialization
        summary["class_summary"] = dict(summary["class_summary"])
        report_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        logger.info("生成报告: %s", report_path.relative_to(WORKSPACE_ROOT))

    # 控制台总结
    s = summary["stats"]
    print()
    print("=" * 60)
    print("生成总结")
    print("=" * 60)
    print(f"候选 wp_code:        {s['candidates']}")
    print(f"跳过（已有手工 YAML）: {s['skipped_handcrafted']}")
    print(f"跳过（output 已存在）: {s['skipped_existing']}")
    print(f"成功生成:            {s['generated']}")
    print(f"失败:                {s['failed']}")
    if summary["class_summary"]:
        print()
        print("归类分布（按 sheet 计）:")
        for cls, count in sorted(
            dict(summary["class_summary"]).items(), key=lambda x: -x[1]
        )[:20]:
            print(f"  {cls:30s} {count:5d}")
    print()

    return 0 if s["failed"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
