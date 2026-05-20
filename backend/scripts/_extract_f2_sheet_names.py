"""一次性脚本：提取 F2 模板真实 sheet 名清单（用完即删）。

用于 F spec followup P0-2 修复 — 替换 prefill_formula_mapping.json 中
F-cycle 部分臆造的 sheet 名（如「明细汇总表F2-2」），改用真名。
"""

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("backend/wp_templates/F")

target_files = [
    "F2-1至F2-14 存货及跌价准备-审定明细表类（Leap-常规程序）.xlsx",
    "F2-18至F2-20 存货及跌价准备-分析类（Leap应对措施-分析程序）.xlsx",
    "F2-21至F2-26 存货及跌价准备 - 盘点类（Leap应对措施- 存货监盘）.xlsx",
    "F2-38至F2-44  存货及跌价准备 -计价测试（Leap-存货程序）.xlsx",
    "F2-47至F2-49 存货及跌价准备 -跌价准备测试（Leap应对措施-会计估计）.xlsx",
]

for fname in target_files:
    fp = ROOT / fname
    if not fp.exists():
        print(f"MISSING: {fname}")
        continue
    print(f"\n=== {fname} ===")
    wb = load_workbook(fp, read_only=True, data_only=True)
    for s in wb.sheetnames:
        print(f"  {s}")
    wb.close()
