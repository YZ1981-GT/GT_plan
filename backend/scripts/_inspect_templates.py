#!/usr/bin/env python3
"""一次性：检查 4 个 xlsx 真实 header 结构。用完即删。"""
from __future__ import annotations

from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
FS = _BACKEND / "data/audit_report_templates/financial_statements"


def dump_sheet_head(variant: str, sheet_substr: str, nrows: int = 8, ncols: int = 10) -> None:
    path = FS / f"{variant}.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    target = None
    for sn in wb.sheetnames:
        if sheet_substr in sn:
            target = sn
            break
    if not target:
        print(f"  no sheet matching '{sheet_substr}' in {variant}")
        return
    ws = wb[target]
    print(f"\n--- {variant} :: {target} (first {nrows} rows) ---")
    for r in range(1, min(nrows, ws.max_row) + 1):
        cells = []
        for c in range(1, min(ncols, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                s = str(v)
                if len(s) > 22:
                    s = s[:22] + "…"
                cells.append(f"{chr(64+c)}{r}={s}")
        if cells:
            print("   " + " | ".join(cells))


PLANS = {
    "soe_standalone": ["资产负债表(企财01表）续", "利润表", "现金流量表", "所有者权益变动表", "资产减值准备"],
    "soe_consolidated": ["资产负债表(企财01表）", "权益变动表（企财04表-合并）", "权益变动表（企财04表-母公司）"],
    "listed_standalone": ["资产负债表", "资产负债表续", "利润表", "现金流量表", "合并股东", "公司股东"],
    "listed_consolidated": ["资产负债表"],
}

if __name__ == "__main__":
    for variant, subs in PLANS.items():
        for sub in subs:
            dump_sheet_head(variant, sub)
