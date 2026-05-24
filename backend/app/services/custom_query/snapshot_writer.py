"""双向编辑写回 — SnapshotWriter

实现 cell 级写回到 working_papers.parsed_data['univer_snapshot'] JSONB，
支持乐观锁冲突检测（X-File-Opened-At vs updated_at）、单事务一致性、
跨模块路由写入（workpaper / report / note / adj / tb）。

Algorithm (design.md 6.2):
  1. SELECT updated_at, parsed_data FROM working_papers WHERE id = wp_id FOR UPDATE
  2. IF opened_at < updated_at → raise WritebackConflict(updated_at, last_editor)
  3. 定位 parsed_data['univer_snapshot']['sheets'][sheet_name]['cellData'][row][col]
  4. old_value = cellData[row][col].get('v')
  5. cellData[row][col]['v'] = new_value
  6. UPDATE working_papers SET parsed_data = :new_pd, updated_at = NOW(), prefill_stale = True
  7. 同步更新 xlsx cache（run_in_executor + openpyxl write）
  8. emit event_bus('cross-ref:updated', {wp_code, sheet, cell_ref, new_value})
  9. audit_logger.log_action('custom_query.cell_writeback', ...)  # 不节流
  10. COMMIT
"""

import asyncio
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)


# ─── Exceptions ──────────────────────────────────────────────────────────────


class WritebackConflict(Exception):
    """乐观锁冲突：opened_at < updated_at"""

    def __init__(self, latest_updated_at: datetime, latest_editor: str):
        self.latest_updated_at = latest_updated_at
        self.latest_editor = latest_editor
        super().__init__(
            f"Conflict: data updated at {latest_updated_at} by {latest_editor}"
        )


class WritebackPermissionDenied(Exception):
    """无写权限或非 workpaper 源"""

    def __init__(self, reason: str):
        self.reason = reason
        super().__init__(reason)


# ─── Cell reference parsing ──────────────────────────────────────────────────


def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """解析 cell_ref (e.g. 'B7') 为 (row_0indexed, col_0indexed)。

    cell_ref 是 1-indexed (Excel 风格)，snapshot 用 0-indexed key (Univer 约定)。
    B7 → row=6, col=1
    """
    m = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper().strip())
    if not m:
        raise ValueError(f"Invalid cell_ref: {cell_ref}")

    col_letters = m.group(1)
    row_num = int(m.group(2))

    # 列字母转 0-indexed
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - 64)
    col -= 1  # 转为 0-indexed

    # 行号转 0-indexed
    row = row_num - 1

    return row, col


# ─── SnapshotWriter ──────────────────────────────────────────────────────────


class SnapshotWriter:
    """双向编辑写回管线。

    支持 5 个模块的 cell 写回：
      - workpaper → parsed_data['univer_snapshot']
      - report → report_snapshot.data
      - note → consol_note_data.data
      - adj → adjustments 表 UPDATE
      - tb → trial_balance.audited_amount UPDATE
    """

    async def write_cell(
        self,
        db: AsyncSession,
        user: Any,
        wp_id: str,
        sheet_name: str,
        cell_ref: str,
        new_value: Any,
        opened_at: datetime,
        module: str = "workpaper",
    ) -> dict:
        """单事务写 cell 到对应模块。

        Args:
            db: 数据库会话
            user: 当前用户对象 (需有 id, username 属性)
            wp_id: working_paper ID (workpaper 模块) 或相关记录 ID
            sheet_name: sheet 名称
            cell_ref: cell 引用 (e.g. "B7")
            new_value: 新值
            opened_at: 前端打开时的 updated_at 时间戳
            module: 模块名 ('workpaper', 'report', 'note', 'adj', 'tb')

        Returns:
            {success: True, updated_at: str, old_value: Any}

        Raises:
            WritebackConflict: 乐观锁冲突
            WritebackPermissionDenied: 无写权限
        """
        if module == "workpaper":
            return await self._write_workpaper_cell(
                db, user, wp_id, sheet_name, cell_ref, new_value, opened_at
            )
        elif module == "report":
            return await self._write_report_cell(
                db, user, wp_id, sheet_name, cell_ref, new_value, opened_at
            )
        elif module == "note":
            return await self._write_note_cell(
                db, user, wp_id, sheet_name, cell_ref, new_value, opened_at
            )
        elif module == "adj":
            return await self._write_adj_cell(
                db, user, wp_id, sheet_name, cell_ref, new_value, opened_at
            )
        elif module == "tb":
            return await self._write_tb_cell(
                db, user, wp_id, sheet_name, cell_ref, new_value, opened_at
            )
        else:
            raise WritebackPermissionDenied(f"Unsupported module: {module}")

    # ─── workpaper 模块写回 ──────────────────────────────────────────────

    async def _write_workpaper_cell(
        self,
        db: AsyncSession,
        user: Any,
        wp_id: str,
        sheet_name: str,
        cell_ref: str,
        new_value: Any,
        opened_at: datetime,
    ) -> dict:
        """写回 workpaper parsed_data['univer_snapshot']。

        Steps:
          1. SELECT FOR UPDATE
          2. 乐观锁比对
          3. 定位 cellData[row][col]
          4. 更新 JSONB + prefill_stale
          5. 同步 xlsx cache (run_in_executor)
          6. emit cross-ref:updated
        """
        # Step 1: SELECT FOR UPDATE
        result = await db.execute(
            text("""
                SELECT updated_at, parsed_data, wp_code, file_path
                FROM working_papers
                WHERE id = :wp_id
                FOR UPDATE
            """),
            {"wp_id": wp_id},
        )
        row = result.first()
        if not row:
            raise ValueError(f"Working paper not found: {wp_id}")

        current_updated_at = row[0]
        parsed_data = row[1] or {}
        wp_code = row[2] or ""
        file_path = row[3] or ""

        # Step 2: 乐观锁比对
        # 确保 opened_at 和 current_updated_at 都是 aware 或都是 naive 进行比较
        opened_at_cmp = opened_at.replace(tzinfo=None) if opened_at.tzinfo else opened_at
        updated_at_cmp = current_updated_at.replace(tzinfo=None) if current_updated_at and hasattr(current_updated_at, 'tzinfo') and current_updated_at.tzinfo else current_updated_at

        if updated_at_cmp and opened_at_cmp < updated_at_cmp:
            # 获取最后编辑人
            last_editor = self._get_last_editor(parsed_data, user)
            raise WritebackConflict(
                latest_updated_at=current_updated_at,
                latest_editor=last_editor,
            )

        # Step 3: 定位 cellData
        row_idx, col_idx = _parse_cell_ref(cell_ref)
        snapshot = parsed_data.get("univer_snapshot", {})
        sheets = snapshot.get("sheets", {})

        # 查找目标 sheet（按 name 匹配）
        target_sheet = None
        target_sheet_key = None
        if isinstance(sheets, dict):
            for key, sheet_data in sheets.items():
                if isinstance(sheet_data, dict) and sheet_data.get("name") == sheet_name:
                    target_sheet = sheet_data
                    target_sheet_key = key
                    break
        elif isinstance(sheets, list):
            for i, sheet_data in enumerate(sheets):
                if isinstance(sheet_data, dict) and sheet_data.get("name") == sheet_name:
                    target_sheet = sheet_data
                    target_sheet_key = i
                    break

        if target_sheet is None:
            raise ValueError(f"Sheet '{sheet_name}' not found in snapshot")

        # Step 4: 获取旧值并写入新值
        cell_data = target_sheet.setdefault("cellData", {})
        row_key = str(row_idx)
        col_key = str(col_idx)

        row_data = cell_data.setdefault(row_key, {})
        cell_obj = row_data.setdefault(col_key, {})
        old_value = cell_obj.get("v")
        cell_obj["v"] = new_value

        # Step 5: 更新 JSONB + prefill_stale
        now = datetime.now(timezone.utc)
        await db.execute(
            text("""
                UPDATE working_papers
                SET parsed_data = :new_pd,
                    updated_at = :now,
                    prefill_stale = true
                WHERE id = :wp_id
            """),
            {"new_pd": json.dumps(parsed_data, ensure_ascii=False), "wp_id": wp_id, "now": now},
        )

        # Step 6: 同步更新 xlsx cache (run_in_executor)
        if file_path:
            try:
                await asyncio.get_event_loop().run_in_executor(
                    None,
                    self._sync_update_xlsx_cache,
                    file_path, sheet_name, row_idx, col_idx, new_value,
                )
            except Exception as e:
                logger.warning("xlsx cache update failed (non-fatal): %s", e)

        # Step 7: emit cross-ref:updated event (best-effort)
        try:
            from app.services.custom_query.metrics import event_bus
            event_bus.emit("cross-ref:updated", {
                "wp_code": wp_code,
                "sheet_name": sheet_name,
                "cell_ref": cell_ref,
                "new_value": new_value,
            })
        except Exception:
            pass  # event bus failure is non-fatal

        return {
            "success": True,
            "updated_at": now.isoformat(),
            "old_value": old_value,
        }

    # ─── report 模块写回 ─────────────────────────────────────────────────

    async def _write_report_cell(
        self,
        db: AsyncSession,
        user: Any,
        wp_id: str,
        sheet_name: str,
        cell_ref: str,
        new_value: Any,
        opened_at: datetime,
    ) -> dict:
        """写回 report_snapshot.data JSONB。

        虚拟 sheet 列映射：A=row_code, B=row_name, C=current_period_amount, D=prior_period_amount, E=formula
        """
        from app.services.custom_query.module_cell_resolver import _REPORT_COLUMNS

        # wp_id 在 report 模块中是 report_snapshot.id
        result = await db.execute(
            text("""
                SELECT id, data, updated_at FROM report_snapshot
                WHERE id = :rid
                FOR UPDATE
            """),
            {"rid": wp_id},
        )
        row = result.first()
        if not row:
            raise ValueError(f"Report snapshot not found: {wp_id}")

        current_updated_at = row[2]
        data = row[1] or {}

        # 乐观锁
        self._check_optimistic_lock(opened_at, current_updated_at, user)

        # 解析 cell_ref 定位虚拟 sheet 行列
        row_idx, col_idx = _parse_cell_ref(cell_ref)
        rows_arr = data.get("rows", [])

        # 虚拟 sheet: 第 1 行是表头，第 2 行起是数据
        data_row_idx = row_idx - 1  # 第 2 行 = rows[0]
        if data_row_idx < 0 or data_row_idx >= len(rows_arr):
            raise ValueError(f"Row index out of range: {cell_ref}")

        col_name = _REPORT_COLUMNS[col_idx] if col_idx < len(_REPORT_COLUMNS) else None
        if not col_name:
            raise ValueError(f"Column index out of range: {cell_ref}")

        old_value = rows_arr[data_row_idx].get(col_name)
        rows_arr[data_row_idx][col_name] = new_value

        now = datetime.now(timezone.utc)
        await db.execute(
            text("""
                UPDATE report_snapshot
                SET data = :new_data, updated_at = :now
                WHERE id = :rid
            """),
            {"new_data": json.dumps(data, ensure_ascii=False), "rid": wp_id, "now": now},
        )

        return {"success": True, "updated_at": now.isoformat(), "old_value": old_value}

    # ─── note 模块写回 ───────────────────────────────────────────────────

    async def _write_note_cell(
        self,
        db: AsyncSession,
        user: Any,
        wp_id: str,
        sheet_name: str,
        cell_ref: str,
        new_value: Any,
        opened_at: datetime,
    ) -> dict:
        """写回 consol_note_data.data JSONB。

        虚拟 sheet 列映射：A=code, B=name, C=year_end, D=year_begin, E=formula
        """
        from app.services.custom_query.module_cell_resolver import _NOTE_COLUMNS

        result = await db.execute(
            text("""
                SELECT id, data, updated_at FROM consol_note_data
                WHERE id = :nid
                FOR UPDATE
            """),
            {"nid": wp_id},
        )
        row = result.first()
        if not row:
            raise ValueError(f"Note data not found: {wp_id}")

        current_updated_at = row[2]
        data = row[1] or {}

        self._check_optimistic_lock(opened_at, current_updated_at, user)

        row_idx, col_idx = _parse_cell_ref(cell_ref)
        rows_arr = data.get("rows", [])

        data_row_idx = row_idx - 1
        if data_row_idx < 0 or data_row_idx >= len(rows_arr):
            raise ValueError(f"Row index out of range: {cell_ref}")

        col_name = _NOTE_COLUMNS[col_idx] if col_idx < len(_NOTE_COLUMNS) else None
        if not col_name:
            raise ValueError(f"Column index out of range: {cell_ref}")

        old_value = rows_arr[data_row_idx].get(col_name)
        rows_arr[data_row_idx][col_name] = new_value

        now = datetime.now(timezone.utc)
        await db.execute(
            text("""
                UPDATE consol_note_data
                SET data = :new_data, updated_at = :now
                WHERE id = :nid
            """),
            {"new_data": json.dumps(data, ensure_ascii=False), "nid": wp_id, "now": now},
        )

        return {"success": True, "updated_at": now.isoformat(), "old_value": old_value}

    # ─── adj 模块写回 ────────────────────────────────────────────────────

    async def _write_adj_cell(
        self,
        db: AsyncSession,
        user: Any,
        wp_id: str,
        sheet_name: str,
        cell_ref: str,
        new_value: Any,
        opened_at: datetime,
    ) -> dict:
        """写回 adjustments 表 UPDATE。

        虚拟 sheet 列映射：A=entry_no, B=account_code, C=account_name, D=debit_amount, E=credit_amount, F=description
        wp_id 在 adj 模块中是 adjustment 记录的 id。
        """
        from app.services.custom_query.module_cell_resolver import _ADJ_COLUMNS

        result = await db.execute(
            text("SELECT id, updated_at FROM adjustments WHERE id = :aid FOR UPDATE"),
            {"aid": wp_id},
        )
        row = result.first()
        if not row:
            raise ValueError(f"Adjustment not found: {wp_id}")

        current_updated_at = row[1]
        self._check_optimistic_lock(opened_at, current_updated_at, user)

        _, col_idx = _parse_cell_ref(cell_ref)
        col_name = _ADJ_COLUMNS[col_idx] if col_idx < len(_ADJ_COLUMNS) else None
        if not col_name:
            raise ValueError(f"Column index out of range: {cell_ref}")

        now = datetime.now(timezone.utc)
        await db.execute(
            text(f"""
                UPDATE adjustments
                SET {col_name} = :new_val, updated_at = :now
                WHERE id = :aid
            """),
            {"new_val": new_value, "now": now, "aid": wp_id},
        )

        return {"success": True, "updated_at": now.isoformat(), "old_value": None}

    # ─── tb 模块写回 ─────────────────────────────────────────────────────

    async def _write_tb_cell(
        self,
        db: AsyncSession,
        user: Any,
        wp_id: str,
        sheet_name: str,
        cell_ref: str,
        new_value: Any,
        opened_at: datetime,
    ) -> dict:
        """写回 trial_balance.audited_amount UPDATE。

        虚拟 sheet 列映射：A=account_code, B=account_name, C=opening_balance, D=debit_amount, E=credit_amount, F=closing_balance, G=audited_amount
        wp_id 在 tb 模块中是 trial_balance 记录的 id。
        """
        result = await db.execute(
            text("SELECT id, updated_at FROM trial_balance WHERE id = :tid FOR UPDATE"),
            {"tid": wp_id},
        )
        row = result.first()
        if not row:
            raise ValueError(f"Trial balance record not found: {wp_id}")

        current_updated_at = row[1]
        self._check_optimistic_lock(opened_at, current_updated_at, user)

        _, col_idx = _parse_cell_ref(cell_ref)
        # 只允许写 audited_amount (G 列, col_idx=6)
        if col_idx != 6:
            raise WritebackPermissionDenied(
                "Only audited_amount (column G) is writable in trial_balance"
            )

        now = datetime.now(timezone.utc)
        await db.execute(
            text("""
                UPDATE trial_balance
                SET audited_amount = :new_val, updated_at = :now
                WHERE id = :tid
            """),
            {"new_val": new_value, "now": now, "tid": wp_id},
        )

        return {"success": True, "updated_at": now.isoformat(), "old_value": None}

    # ─── 辅助方法 ────────────────────────────────────────────────────────

    def _check_optimistic_lock(
        self, opened_at: datetime, current_updated_at: datetime | None, user: Any
    ):
        """乐观锁比对：opened_at < updated_at → raise WritebackConflict"""
        if current_updated_at is None:
            return

        opened_at_cmp = opened_at.replace(tzinfo=None) if opened_at.tzinfo else opened_at
        updated_at_cmp = (
            current_updated_at.replace(tzinfo=None)
            if hasattr(current_updated_at, "tzinfo") and current_updated_at.tzinfo
            else current_updated_at
        )

        if opened_at_cmp < updated_at_cmp:
            last_editor = getattr(user, "username", "unknown")
            raise WritebackConflict(
                latest_updated_at=current_updated_at,
                latest_editor=last_editor,
            )

    def _get_last_editor(self, parsed_data: dict, fallback_user: Any) -> str:
        """从 parsed_data 获取最后编辑人，fallback 到当前用户"""
        snapshot = parsed_data.get("univer_snapshot", {})
        saved_by = snapshot.get("saved_by")
        if saved_by:
            return str(saved_by)
        return getattr(fallback_user, "username", "unknown")

    def _sync_update_xlsx_cache(
        self,
        file_path: str,
        sheet_name: str,
        row_idx: int,
        col_idx: int,
        new_value: Any,
    ):
        """同步更新 xlsx 文件 cache（在 run_in_executor 中调用）。

        使用 openpyxl 写入对应 cell。
        """
        try:
            import openpyxl

            path = Path(file_path)
            if not path.exists():
                logger.debug("xlsx cache file not found, skipping: %s", file_path)
                return

            wb = openpyxl.load_workbook(str(path))
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if ws is None:
                wb.close()
                return

            # openpyxl 使用 1-indexed
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=new_value)
            wb.save(str(path))
            wb.close()
        except Exception as e:
            logger.warning("_sync_update_xlsx_cache error: %s", e)


# 模块级单例
snapshot_writer = SnapshotWriter()
