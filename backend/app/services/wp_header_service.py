"""底稿表头填充服务 — 生成底稿时自动填写表头信息和交叉索引

致同标准底稿表头包含：
- 事务所名称（致同会计师事务所）
- 编制单位（client_name）
- 审计期间（audit_period_start ~ audit_period_end）
- 底稿名称
- 索引号（wp_code）
- 编制人 / 编制日期
- 复核人 / 复核日期
- 交叉索引（关联底稿编号和名称）
"""

from __future__ import annotations

import json
import logging
from datetime import date
from pathlib import Path
from typing import Any
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.core import Project
from app.models.workpaper_models import WorkingPaper, WpIndex

_logger = logging.getLogger(__name__)

# 致同标准底稿表头区域定义（行号从1开始）
# 实际模板中表头位置可能不同，这里定义兜底的空白底稿表头布局
_HEADER_LAYOUT = {
    "firm_name": {"row": 1, "col": 1, "value": "致同会计师事务所（特殊普通合伙）"},
    "client_label": {"row": 2, "col": 1, "value": "编制单位："},
    "client_name": {"row": 2, "col": 2},
    "period_label": {"row": 2, "col": 4, "value": "审计期间："},
    "period_value": {"row": 2, "col": 5},
    "wp_name_label": {"row": 3, "col": 1, "value": "底稿名称："},
    "wp_name_value": {"row": 3, "col": 2},
    "index_label": {"row": 3, "col": 4, "value": "索引号："},
    "index_value": {"row": 3, "col": 5},
    "preparer_label": {"row": 4, "col": 1, "value": "编制人："},
    "preparer_value": {"row": 4, "col": 2},
    "prep_date_label": {"row": 4, "col": 3, "value": "日期："},
    "prep_date_value": {"row": 4, "col": 4},
    "reviewer_label": {"row": 4, "col": 5, "value": "复核人："},
    "reviewer_value": {"row": 4, "col": 6},
    "rev_date_label": {"row": 4, "col": 7, "value": "日期："},
    "rev_date_value": {"row": 4, "col": 8},
    "xref_label": {"row": 5, "col": 1, "value": "交叉索引："},
    "xref_value": {"row": 5, "col": 2},
}

# 审计阶段中文名
_CYCLE_NAMES = {
    "B": "准备阶段（初步业务活动/风险评估）",
    "C": "准备阶段（风险应对）",
    "D": "实施阶段（销售与收款循环）",
    "E": "实施阶段（货币资金循环）",
    "F": "实施阶段（采购与存货循环）",
    "G": "实施阶段（投资循环）",
    "H": "实施阶段（固定资产循环）",
    "I": "实施阶段（无形资产循环）",
    "J": "实施阶段（薪酬循环）",
    "K": "实施阶段（管理费用与其他循环）",
    "L": "实施阶段（债务循环）",
    "M": "实施阶段（权益循环）",
    "N": "实施阶段（税金循环）",
    "A": "完成阶段",
    "S": "特殊事项",
    "Q": "关联方",
}


def get_cross_ref_text(wp_code: str, mapping_data: list[dict]) -> str:
    """根据底稿编码生成交叉索引文本

    规则：
    1. 同循环的其他底稿（如 E1-1 引用 E1-2、E1-3）
    2. 通过 wp_account_mapping 关联的底稿
    3. 通用引用（审定表引用程序表、程序表引用审定表）
    """
    cycle = wp_code[0] if wp_code else ""
    refs = []

    # 同循环底稿
    for m in mapping_data:
        if m.get("cycle") == cycle and m.get("wp_code") != wp_code:
            refs.append(f"{m['wp_code']}({m.get('wp_name', '')})")

    # 通用交叉引用
    if wp_code.endswith("-1") or "审定表" in str(mapping_data):
        # 审定表通常引用程序表
        prog_code = wp_code.split("-")[0] if "-" in wp_code else wp_code
        refs.append(f"{prog_code}(程序表)")

    # 限制长度
    if len(refs) > 5:
        refs = refs[:5] + [f"等共{len(refs)}项"]

    return "、".join(refs) if refs else ""


async def fill_workpaper_header(
    db: AsyncSession,
    project_id: UUID,
    wp_id: UUID,
    file_path: str,
    wp_code: str,
    wp_name: str,
    cycle: str | None = None,
    assigned_to_name: str | None = None,
    reviewer_name: str | None = None,
    is_custom: bool = False,
) -> dict[str, Any]:
    """填充底稿表头信息

    对于从模板复制的底稿：搜索已有表头区域并填充
    对于空白生成的底稿或自定义底稿(is_custom=True)：按标准布局写入表头

    通用规则（适用于所有底稿）：
    1. 编制单位 = Project.client_name
    2. 审计期间 = audit_period_start ~ audit_period_end
    3. 索引号 = wp_code
    4. 底稿名称 = wp_name
    5. 交叉索引 = 同循环关联底稿自动生成
    6. 审计阶段 = 根据循环前缀映射中文名
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return {"status": "error", "message": "openpyxl 未安装"}

    fp = Path(file_path)
    if not fp.exists():
        return {"status": "error", "message": f"文件不存在: {fp}"}

    # 获取项目信息
    proj = (await db.execute(
        sa.select(Project).where(Project.id == project_id)
    )).scalar_one_or_none()

    client_name = proj.client_name if proj else "未知单位"
    ws_data = proj.wizard_state or {} if proj else {}
    basic_info = ws_data.get("steps", {}).get("basic_info", {}).get("data", {})
    audit_year = basic_info.get("audit_year", date.today().year)
    period_start = proj.audit_period_start if proj else None
    period_end = proj.audit_period_end if proj else None

    if not period_start:
        period_start = date(audit_year, 1, 1)
    if not period_end:
        period_end = date(audit_year, 12, 31)

    period_text = f"{period_start.strftime('%Y年%m月%d日')}至{period_end.strftime('%Y年%m月%d日')}"

    # 加载交叉索引数据
    mapping_path = Path(__file__).parent.parent.parent / "data" / "wp_account_mapping.json"
    mapping_data = []
    if mapping_path.exists():
        try:
            with open(mapping_path, "r", encoding="utf-8-sig") as f:
                mapping_data = json.load(f).get("mappings", [])
        except Exception:
            pass

    xref_text = get_cross_ref_text(wp_code, mapping_data)

    # 打开 Excel
    try:
        wb = openpyxl.load_workbook(str(fp))
    except Exception as e:
        return {"status": "error", "message": f"打开文件失败: {e}"}

    ws = wb.active

    # 策略：先尝试搜索已有表头关键词并填充，找不到则按标准布局写入
    # 自定义底稿(is_custom)强制写入标准表头
    filled_by_search = False
    if not is_custom:
        filled_by_search = _try_fill_existing_header(
            ws, client_name, period_text, wp_name, wp_code,
            assigned_to_name, reviewer_name, xref_text,
        )

    if not filled_by_search:
        # 空白底稿或自定义底稿：按标准布局写入表头
        _write_standard_header(
            ws, client_name, period_text, wp_name, wp_code,
            cycle, assigned_to_name, reviewer_name, xref_text,
        )

    try:
        wb.save(str(fp))
    except Exception as e:
        return {"status": "error", "message": f"保存失败: {e}"}
    finally:
        wb.close()

    _logger.info("fill_header: wp=%s code=%s client=%s", wp_id, wp_code, client_name)
    return {"status": "ok", "filled_by_search": filled_by_search}


def _try_fill_existing_header(
    ws, client_name: str, period_text: str, wp_name: str, wp_code: str,
    preparer: str | None, reviewer: str | None, xref_text: str,
) -> bool:
    """尝试在已有模板中搜索表头关键词并填充右侧单元格"""
    filled = False
    _KW_MAP = {
        "编制单位": client_name,
        "被审计单位": client_name,
        "审计期间": period_text,
        "会计期间": period_text,
        "底稿名称": wp_name,
        "索引号": wp_code,
        "索引": wp_code,
        "编制人": preparer or "",
        "编制": preparer or "",
        "复核人": reviewer or "",
        "复核": reviewer or "",
        "交叉索引": xref_text,
    }

    for row in ws.iter_rows(min_row=1, max_row=10, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                s = cell.value.strip().rstrip("：:").strip()
                for kw, fill_value in _KW_MAP.items():
                    if s == kw or s.endswith(kw):
                        # 填充右侧单元格
                        right = ws.cell(row=cell.row, column=cell.column + 1)
                        if right.value is None or str(right.value).strip() == "":
                            right.value = fill_value
                            filled = True
                        break
    return filled


def _write_standard_header(
    ws, client_name: str, period_text: str, wp_name: str, wp_code: str,
    cycle: str | None, preparer: str | None, reviewer: str | None, xref_text: str,
):
    """空白底稿写入标准表头"""
    from openpyxl.styles import Font, Alignment, PatternFill

    header_font = Font(name="仿宋_GB2312", size=10)
    title_font = Font(name="仿宋_GB2312", size=12, bold=True)
    header_fill = PatternFill(start_color="F4F0FA", end_color="F4F0FA", fill_type="solid")

    # 第1行：事务所名称
    ws.cell(row=1, column=1, value="致同会计师事务所（特殊普通合伙）").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # 第2行：编制单位 + 审计期间
    ws.cell(row=2, column=1, value="编制单位：").font = header_font
    ws.cell(row=2, column=2, value=client_name).font = header_font
    ws.cell(row=2, column=5, value="审计期间：").font = header_font
    ws.cell(row=2, column=6, value=period_text).font = header_font

    # 第3行：底稿名称 + 索引号
    ws.cell(row=3, column=1, value="底稿名称：").font = header_font
    ws.cell(row=3, column=2, value=wp_name).font = header_font
    ws.cell(row=3, column=5, value="索引号：").font = header_font
    ws.cell(row=3, column=6, value=wp_code).font = Font(name="Arial Narrow", size=11, bold=True)

    # 第4行：编制人 + 复核人
    ws.cell(row=4, column=1, value="编制人：").font = header_font
    ws.cell(row=4, column=2, value=preparer or "").font = header_font
    ws.cell(row=4, column=3, value="日期：").font = header_font
    ws.cell(row=4, column=4, value="").font = header_font
    ws.cell(row=4, column=5, value="复核人：").font = header_font
    ws.cell(row=4, column=6, value=reviewer or "").font = header_font
    ws.cell(row=4, column=7, value="日期：").font = header_font

    # 第5行：交叉索引
    ws.cell(row=5, column=1, value="交叉索引：").font = header_font
    ws.cell(row=5, column=2, value=xref_text).font = header_font

    # 第6行：审计阶段
    stage = _CYCLE_NAMES.get(cycle, "") if cycle else ""
    if stage:
        ws.cell(row=5, column=5, value="审计阶段：").font = header_font
        ws.cell(row=5, column=6, value=stage).font = header_font

    # 第7行空行分隔
    # 第8行开始是数据区域（审定表的列标题）

    # 设置列宽
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 14

    # 表头区域浅紫色背景
    for r in range(1, 6):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
