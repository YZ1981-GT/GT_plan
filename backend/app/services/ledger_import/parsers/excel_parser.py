"""Excel 流式解析器（openpyxl）。

职责（见 design.md §11 / Sprint 2 Task 21 / Sprint 6 Task S6-11）：

- 按 chunk 生成行数据（50,000 行/chunk，经验值兼顾内存与 COPY 效率）。
- 复用 `detector.detect_header_row` 的表头结果，从 `data_start_row` 开始流式读。
- openpyxl read_only=True + iter_rows 避免将整个文件加载到内存。
- read_only 对部分 xlsx（合并表头/特殊样式）返回 0 行，自动回退 read_only=False。
- bytes 入口和 path 入口共用底层 `_iter_from_workbook` 函数（S6-11 合并双实现）。

每个 chunk 是一个 list[list[Any]]，每行是一个 list of cell values
（str/int/float/None — 不在此处强制转 str，由 writer 负责类型转换）。
"""

from __future__ import annotations

import io
import logging
from contextlib import contextmanager
from typing import Any, Generator

import openpyxl

__all__ = ["iter_excel_rows", "iter_excel_rows_from_path", "CHUNK_SIZE"]

logger = logging.getLogger(__name__)

CHUNK_SIZE = 50_000  # rows per chunk (design §11)


@contextmanager
def _open_workbook(source, *, read_only: bool):
    """统一的 workbook 打开/关闭上下文（bytes 或 path）。"""
    wb = None
    try:
        wb = openpyxl.load_workbook(source, read_only=read_only, data_only=True)
        yield wb
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass


def _iter_from_workbook(
    wb,
    sheet_name: str,
    *,
    data_start_row: int,
    chunk_size: int,
    forward_fill_cols: list[int] | None = None,
) -> Generator[list[list[Any]], None, None]:
    """已打开的 workbook → 流式 yield chunk。

    forward_fill_cols：合并单元格向下填充（None → 上一行非空值）。
    """
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    # openpyxl 1-based；data_start_row 0-based
    min_row = data_start_row + 1
    chunk: list[list[Any]] = []

    # forward-fill 上一行记忆值（跨 chunk 保持）
    prev_values: dict[int, Any] = {}
    ff_cols = set(forward_fill_cols or [])

    for row_tuple in ws.iter_rows(min_row=min_row, values_only=True):
        row_list = list(row_tuple)
        # 执行 forward-fill
        if ff_cols:
            for col_idx in ff_cols:
                if col_idx < len(row_list):
                    val = row_list[col_idx]
                    if val is None or (isinstance(val, str) and not val.strip()):
                        # 空值 → 用上一行值填
                        if col_idx in prev_values:
                            row_list[col_idx] = prev_values[col_idx]
                    else:
                        # 非空值 → 更新记忆
                        prev_values[col_idx] = val

        chunk.append(row_list)
        if len(chunk) >= chunk_size:
            yield chunk
            chunk = []

    if chunk:
        yield chunk


def _iter_with_fallback(
    source,
    source_desc: str,
    sheet_name: str,
    data_start_row: int,
    chunk_size: int,
    *,
    forward_fill_cols: list[int] | None = None,
) -> Generator[list[list[Any]], None, None]:
    """尝试 read_only=True 流式；0 行时回退 read_only=False。"""
    yielded = 0
    try:
        with _open_workbook(source, read_only=True) as wb:
            for chunk in _iter_from_workbook(
                wb, sheet_name,
                data_start_row=data_start_row,
                chunk_size=chunk_size,
                forward_fill_cols=forward_fill_cols,
            ):
                yielded += len(chunk)
                yield chunk
    except RuntimeError:
        raise
    except Exception as exc:
        logger.exception("Excel parsing (read_only) failed for '%s'", sheet_name)
        raise RuntimeError(
            f"Excel parsing failed for sheet '{sheet_name}': {exc}"
        ) from exc

    if yielded > 0:
        return

    # Fallback: read_only=False（某些 xlsx read_only 会返回 0 行）
    logger.warning(
        "Excel read_only returned 0 rows for sheet '%s' (%s), "
        "falling back to full-load mode",
        sheet_name, source_desc,
    )
    try:
        # bytes 需要重置游标；path 重新打开
        if isinstance(source, io.BytesIO):
            source.seek(0)
        with _open_workbook(source, read_only=False) as wb:
            yield from _iter_from_workbook(
                wb, sheet_name,
                data_start_row=data_start_row,
                chunk_size=chunk_size,
                forward_fill_cols=forward_fill_cols,
            )
    except RuntimeError:
        raise
    except Exception as exc:
        logger.exception("Excel parsing (full-load) failed for '%s'", sheet_name)
        raise RuntimeError(
            f"Excel parsing failed for sheet '{sheet_name}': {exc}"
        ) from exc


def iter_excel_rows(
    content: bytes,
    sheet_name: str,
    *,
    data_start_row: int = 1,
    chunk_size: int = CHUNK_SIZE,
    forward_fill_cols: list[int] | None = None,
) -> Generator[list[list[Any]], None, None]:
    """Yield chunks from xlsx bytes. read_only + 自动回退。

    Parameters
    ----------
    forward_fill_cols : list[int] | None
        0-based 列索引，这些列的 None 值会被上一行的非空值填充。
        典型场景：科目编码列（银行存款 4 行只在第 1 行显示）。默认 None = 不填充。
    """
    source = io.BytesIO(content)
    yield from _iter_with_fallback(
        source, "bytes", sheet_name, data_start_row, chunk_size,
        forward_fill_cols=forward_fill_cols,
    )


def iter_excel_rows_from_path(
    path: str,
    sheet_name: str,
    *,
    data_start_row: int = 1,
    chunk_size: int = CHUNK_SIZE,
    forward_fill_cols: list[int] | None = None,
) -> Generator[list[list[Any]], None, None]:
    """Yield chunks from xlsx path. 不全量读入内存（read_only）+ 自动回退。

    forward_fill_cols：S6-12 合并单元格向下填充（如合并的科目编码列）。
    """
    yield from _iter_with_fallback(
        path, f"path={path}", sheet_name, data_start_row, chunk_size,
        forward_fill_cols=forward_fill_cols,
    )
