"""探测层 — 文件类型 / 前 20 行 / 表头行 / 合并单元格 识别。

职责（见 design.md §4 / §18 / Sprint 1 Task 4-6）：

- ``detect_file(content, filename)``   : 返回 ``FileDetection``，支持
  xlsx / xlsm / csv / tsv / zip，只读前 20 行（openpyxl read_only +
  iter_rows(max_row=20)）。
- ``.xls`` → ``XLS_NOT_SUPPORTED`` 致命错误（不引入 xlrd 依赖）。
- 未知扩展名 → ``UNSUPPORTED_FILE_TYPE`` 致命错误。
- 文件损坏 → ``CORRUPTED_FILE`` 致命错误。

表头识别（Task 5，design §18 + requirements 需求 15）：
``_detect_header_row`` 支持
- 标题行跳过（前 5 行中 ``len(non_empty) <= 2`` 且含 "公司 / 年度 / 报表 /
  科目余额表 / 序时账 / 凭证明细" 等关键词的单行横幅）；
- 合并表头识别（连续 2-3 行合并 → 单行表头），向左 forward-fill 模拟合并
  单元格语义后与下行按 "top.bottom" 拼接；
- 兜底：首个"非空单元格数 ≥ 3"的行。

ZIP 处理：CP437 → gbk 自动重解码中文条目名（需求 14），递归 detect_file
内部文件；内部 sheet 的 ``file_name`` 前缀 ``archive.zip!inner.xlsx``。

编码探测：委托给 ``encoding_detector.decode_content``（Task 7 已实装）——
BOM → 候选列表试解码 → chardet 兜底 → latin1 终极兜底。
"""

import csv
import io
import logging
import os
import re
import zipfile
from datetime import date, datetime, time
from typing import Any, Optional

from .detection_types import FileDetection, SheetDetection
from .encoding_detector import decode_content
from .errors import ErrorCode, make_error

logger = logging.getLogger(__name__)

# 只读前 20 行（对齐 requirements 需求 1 / design §11.1）
PREVIEW_ROW_LIMIT = 20

# CSV 探测只需读取的字节数（64KB 足够编码探测 + 前 20 行）
_CSV_PROBE_BYTES = 65536

# 支持的扩展名分发
_XLSX_EXTS = {".xlsx", ".xlsm"}
_CSV_EXTS = {".csv", ".tsv"}
_XLS_EXTS = {".xls"}
_ZIP_EXTS = {".zip"}


# ---------------------------------------------------------------------------
# 文件名元信息提取（F6 / Sprint 10 Task 10.12-10.14）
# ---------------------------------------------------------------------------

# 表类型关键字（匹配文件名基名，不区分大小写）
_FILENAME_TABLE_KEYWORDS: dict[str, list[str]] = {
    "balance": ["科目余额", "余额表", "试算", "试算平衡", "总账", "trial balance", "tb"],
    "ledger": ["序时账", "序时", "凭证明细", "凭证", "总账明细", "明细账", "general ledger", "journal", "gl"],
    "aux_balance": ["辅助余额", "核算项目余额", "辅助核算余额"],
    "aux_ledger": ["辅助明细", "核算项目明细", "辅助序时"],
}

# 期间信息（年月）提取模式
# 优先级顺序（关键：更具体的先匹配）：
# 1. 25年10月 / 2025年10月（年月都带中文后缀）
# 2. 24.10 / 2024.10 / 2024-10 / 2024/10（点/斜杠/短横分隔）
# 3. 202410（6 位连续）
#
# 为保证 4 位年份不被吃掉为 "2 位+2 位"，采用多条正则按顺序匹配。
_FILENAME_YEAR_MONTH_PATTERNS = [
    # (年)年(月)月：明确的中文分隔
    re.compile(r"(?P<year>20\d{2}|\d{2})年(?P<month>0?[1-9]|1[0-2])月"),
    # 2 位年 + 月格式（如 24.10 / 24-10 / 24_10 / 24/10）- year 严格 2 位
    re.compile(r"(?:^|[^0-9])(?P<year>\d{2})[./\-_](?P<month>0?[1-9]|1[0-2])(?:[^0-9]|$)"),
    # 4 位年 + 月格式（如 2024.10 / 2024-10 / 2024_10 / 2024/10）
    re.compile(r"(?P<year>20\d{2})[./\-_](?P<month>0?[1-9]|1[0-2])(?:[^0-9]|$)"),
    # 仅年年月（不带分隔符）：202410
    re.compile(r"(?:^|[^0-9])(?P<year>20\d{2})(?P<month>0?[1-9]|1[0-2])(?:[^0-9]|$)"),
]
# 仅年份
_FILENAME_YEAR_ONLY_RE = re.compile(r"(?:^|[^0-9])(?P<year>20\d{2}|19\d{2})(?:[^0-9]|$)")


def _extract_filename_hints(filename: str) -> dict:
    """从文件名提取表类型 / 年月信息（F6 文件名元信息利用）。

    Args:
        filename: 含路径或不含路径的文件名（含扩展名）。

    Returns:
        dict 含以下可选键：
        - ``table_type``: ``"balance"``/``"ledger"``/``"aux_balance"``/``"aux_ledger"``（不含 unknown）
        - ``table_confidence``: 65 或 70（仅命中时返回）
        - ``matched_keyword``: 触发命中的关键字
        - ``year``: int（4 位年份）
        - ``month``: int（1-12）
        - ``file_stem``: 无扩展名的文件名基名

    如果未能提取任何信息，返回空 dict。
    """
    hints: dict = {}
    if not filename:
        return hints

    # 去路径 + 扩展名，统一小写做关键字匹配
    stem = os.path.splitext(os.path.basename(filename))[0]
    hints["file_stem"] = stem
    stem_lower = stem.lower()

    # 1) 表类型匹配：优先匹配更长/更具体的关键字
    best_match: Optional[tuple[str, str, int]] = None  # (table_type, keyword, length)
    for tt, keywords in _FILENAME_TABLE_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in stem_lower:
                # 更长的 keyword 更具体
                if best_match is None or len(kw) > best_match[2]:
                    best_match = (tt, kw, len(kw))

    if best_match is not None:
        tt, kw, _ = best_match
        hints["table_type"] = tt
        hints["matched_keyword"] = kw
        # 较长关键字（>=4 char）给 75，较短给 65
        hints["table_confidence"] = 75 if len(kw) >= 4 else 65

    # 2) 期间信息提取：按优先级尝试多条正则
    for pat in _FILENAME_YEAR_MONTH_PATTERNS:
        ym_match = pat.search(stem)
        if ym_match:
            year_raw = ym_match.group("year")
            month_raw = ym_match.group("month")
            # 2 位年份 → 20xx
            if len(year_raw) == 2:
                year = 2000 + int(year_raw)
            else:
                year = int(year_raw)
            month = int(month_raw)
            if 1 <= month <= 12 and 2000 <= year <= 2100:
                hints["year"] = year
                hints["month"] = month
                break
    else:
        # 兜底只查年份
        y_only = _FILENAME_YEAR_ONLY_RE.search(stem)
        if y_only:
            year = int(y_only.group("year"))
            if 2000 <= year <= 2100:
                hints["year"] = year

    return hints


# ---------------------------------------------------------------------------
# 表头规范化（F7 方括号 + 组合表头，Sprint 10 Task 10.16-10.18）
# ---------------------------------------------------------------------------

# 方括号剥壳：[XX] → XX；全角【XX】→ XX
_BRACKET_STRIP_RE = re.compile(r"[\[【](.+?)[\]】]")

# 组合分隔符：# / | / @ 等
_COMPOUND_SEP_RE = re.compile(r"\s*[#|@]\s*")


def _normalize_header(cell: str) -> tuple[str, list[str]]:
    """规范化单个表头单元格（F7）。

    处理：
    1. 剥离方括号：``[凭证号码]`` → ``凭证号码``
    2. 组合表头拆分：``凭证号码#日期`` / ``[凭证号码]#[日期]`` → 主名 + 子字段列表

    返回 ``(primary, sub_fields)``：
    - ``primary``：主标识（拆分后的第一个字段，始终非空）
    - ``sub_fields``：拆出的子字段（含 primary 自身，空时为 [primary]）

    示例：
      - ``"[凭证号码]"`` → ``("凭证号码", ["凭证号码"])``
      - ``"[凭证号码]#[日期]"`` → ``("凭证号码#日期", ["凭证号码", "日期"])``
      - ``"凭证号码#日期"`` → ``("凭证号码#日期", ["凭证号码", "日期"])``
      - ``"科目编码"`` → ``("科目编码", ["科目编码"])``
    """
    if cell is None:
        return "", []
    s = str(cell).strip()
    if not s:
        return "", []

    # 1) 剥方括号：反复替换直到没有方括号对（但保留未配对的）
    def _strip_once(m: re.Match) -> str:
        return m.group(1)

    s_stripped = _BRACKET_STRIP_RE.sub(_strip_once, s)

    # 2) 检测组合分隔符
    if _COMPOUND_SEP_RE.search(s_stripped):
        sub_fields = [p.strip() for p in _COMPOUND_SEP_RE.split(s_stripped) if p.strip()]
        if sub_fields:
            # primary 用 # 连接所有子字段（作为唯一 header 标识）
            primary = "#".join(sub_fields)
            return primary, sub_fields

    # 3) 无组合：单字段
    return s_stripped, [s_stripped]


def _normalize_header_row(cells: list[str]) -> tuple[list[str], dict[int, list[str]]]:
    """对整行表头做规范化。

    Returns:
        ``(normalized_cells, compound_headers)``
        - ``normalized_cells``: 与输入等长，每个 cell 都已剥方括号 + 主标识
        - ``compound_headers``: ``{col_index: [子字段, ...]}``，仅含拆出多个子字段的列
    """
    normalized: list[str] = []
    compound: dict[int, list[str]] = {}
    for idx, cell in enumerate(cells):
        primary, subs = _normalize_header(cell)
        normalized.append(primary)
        if len(subs) > 1:
            compound[idx] = subs
    return normalized, compound


def detect_file_from_path(path: str, filename: Optional[str] = None) -> FileDetection:
    """从文件路径探测，支持 600MB+ 大文件而不全量读入内存。

    - CSV/TSV：只读前 64KB 做编码探测 + 前 20 行解析，用 ``os.path.getsize``
      获取文件大小（不读完整文件）。
    - XLSX/XLSM：openpyxl read_only=True 流式读前 20 行（内存占用 ~10MB 级别，
      与文件大小无关）。
    - ZIP：需要完整读入（ZIP 目录在文件末尾），大 ZIP 仍走全量读取。
    - XLS / 未知扩展名：同 ``detect_file``。

    Args:
        path: 文件绝对或相对路径。
        filename: 显示用文件名（默认取 basename）。

    Returns:
        FileDetection（与 detect_file 返回值完全相同）。
    """
    if filename is None:
        filename = os.path.basename(path)

    ext = os.path.splitext(filename)[1].lower()
    size = os.path.getsize(path)
    logger.debug("detect_file_from_path: path=%s size=%d ext=%s", path, size, ext)

    if ext in _CSV_EXTS:
        hint = "\t" if ext == ".tsv" else None
        return _detect_csv_from_path(path, filename, file_size=size, delimiter_hint=hint)

    if ext in _XLSX_EXTS:
        return _detect_xlsx_from_path(path, filename, file_size=size)

    # XLS / ZIP / 未知扩展名：读完整文件走原有逻辑
    with open(path, "rb") as f:
        content = f.read()
    return detect_file(content, filename)


def detect_file(content: bytes, filename: str) -> FileDetection:
    """探测单个文件，返回 ``FileDetection``。

    永远不抛异常：任何解析错误都收敛到 ``FileDetection.errors``。

    未知扩展名：``file_type`` 字段受 ``Literal["xlsx","xls","csv","zip"]``
    限制，无"unknown"值；此时回退为 ``"csv"`` 并附 ``UNSUPPORTED_FILE_TYPE``
    fatal 错误，调用方通过 ``errors`` 判断实际可用性（对齐 design §10.1）。
    """

    ext = os.path.splitext(filename)[1].lower()
    size = len(content)
    logger.debug("detect_file: filename=%s size=%d ext=%s", filename, size, ext)

    if ext in _XLSX_EXTS:
        return _detect_xlsx(content, filename)
    if ext in _CSV_EXTS:
        hint = "\t" if ext == ".tsv" else None
        return _detect_csv(content, filename, delimiter_hint=hint)
    if ext in _XLS_EXTS:
        return FileDetection(
            file_name=filename,
            file_size_bytes=size,
            file_type="xls",
            errors=[
                make_error(
                    ErrorCode.XLS_NOT_SUPPORTED,
                    message=(
                        f".xls 格式暂不支持：{filename}；"
                        "请将文件另存为 .xlsx 后重新上传"
                    ),
                    file=filename,
                    suggestion="打开文件 → 另存为 → 选择 Excel 工作簿 (*.xlsx)",
                )
            ],
        )
    if ext in _ZIP_EXTS:
        return _detect_zip(content, filename)

    logger.debug("detect_file: unsupported extension %s for %s", ext, filename)
    return FileDetection(
        file_name=filename,
        file_size_bytes=size,
        file_type="csv",
        errors=[
            make_error(
                ErrorCode.UNSUPPORTED_FILE_TYPE,
                message=(
                    f"不支持的文件类型：{filename}"
                    f"（扩展名 {ext or '(无)'}）"
                ),
                file=filename,
                suggestion="请上传 .xlsx / .xlsm / .csv / .tsv / .zip 文件",
            )
        ],
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _detect_xlsx(content: bytes, filename: str) -> FileDetection:
    """探测 xlsx/xlsm：逐 sheet 读前 20 行，异常收敛到 errors。

    read_only 回退策略：openpyxl read_only=True 对含大量合并单元格的文件
    可能只返回实际有值的单元格（不填充到 max_column），导致每行只有 1-2 列。
    检测到此情况后自动回退到 read_only=False 重新解析。
    """

    size = len(content)
    fd = FileDetection(file_name=filename, file_size_bytes=size, file_type="xlsx")

    try:
        import openpyxl  # 局部 import 避免模块加载开销
    except ImportError as exc:  # pragma: no cover - 环境缺包才会命中
        logger.exception("openpyxl import failed")
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"缺少 openpyxl 依赖，无法解析 xlsx 文件：{exc}",
                file=filename,
            )
        )
        return fd

    # 尝试 read_only=True 先（性能优先）
    fd = _detect_xlsx_with_mode(openpyxl, content, filename, fd, read_only=True)

    # 检测行宽异常：如果任何 sheet 的 preview_rows 非空但最大行宽 ≤ 2，
    # 说明 read_only 模式未正确展开合并单元格，需回退
    needs_fallback = False
    for sheet_det in fd.sheets:
        if sheet_det.preview_rows and len(sheet_det.preview_rows) >= 3:
            max_width = max(len(row) for row in sheet_det.preview_rows)
            if max_width <= 2:
                needs_fallback = True
                logger.info(
                    "read_only mode returned narrow rows (max_width=%d) for "
                    "sheet %r in %s, falling back to read_only=False",
                    max_width,
                    sheet_det.sheet_name,
                    filename,
                )
                break

    if needs_fallback:
        # 重置 sheets 和 errors，用 read_only=False 重新解析
        fd.sheets = []
        fd.errors = []
        fd = _detect_xlsx_with_mode(openpyxl, content, filename, fd, read_only=False)

    return fd


def _detect_xlsx_with_mode(
    openpyxl: Any,
    content: bytes,
    filename: str,
    fd: FileDetection,
    *,
    read_only: bool,
) -> FileDetection:
    """用指定的 read_only 模式解析 xlsx，结果写入 fd。"""

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=read_only,
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001 - 任何 openpyxl 异常都视为损坏
        logger.exception(
            "openpyxl.load_workbook(read_only=%s) failed for %s",
            read_only,
            filename,
        )
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"无法打开 xlsx 文件 {filename}：{exc}",
                file=filename,
                suggestion="请检查文件是否损坏或被其他程序占用",
            )
        )
        return fd

    try:
        for sheet_name in wb.sheetnames:
            try:
                sheet_det = _detect_xlsx_sheet(
                    wb, sheet_name, filename, read_only=read_only
                )
                fd.sheets.append(sheet_det)
            except Exception as exc:  # noqa: BLE001 - 单 sheet 失败不影响其他 sheet
                logger.exception(
                    "failed to parse sheet %s in %s", sheet_name, filename
                )
                fd.errors.append(
                    make_error(
                        ErrorCode.CORRUPTED_FILE,
                        message=(
                            f"解析 sheet {sheet_name!r} 失败"
                            f"（文件 {filename}）：{exc}"
                        ),
                        file=filename,
                        sheet=sheet_name,
                    )
                )
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass

    return fd


def _detect_xlsx_sheet(
    wb: Any, sheet_name: str, filename: str, *, read_only: bool = True
) -> SheetDetection:
    """解析单个 sheet 的前 20 行，返回 ``SheetDetection``。

    read_only=False 时，openpyxl 会正确展开合并单元格区域，每行宽度
    等于 max_column（而非只返回有值的单元格）。
    """

    ws = wb[sheet_name]
    preview_rows: list[list[str]] = []

    if read_only:
        # read_only 模式：iter_rows 返回 tuple of values
        for row in ws.iter_rows(max_row=PREVIEW_ROW_LIMIT, values_only=True):
            preview_rows.append([_coerce_cell(c) for c in row])
            if len(preview_rows) >= PREVIEW_ROW_LIMIT:
                break
    else:
        # 非 read_only 模式：iter_rows 返回 Cell 对象，需取 .value
        # 确保每行宽度一致（填充到 max_column）
        max_col = ws.max_column or 1
        for row in ws.iter_rows(
            min_row=1, max_row=PREVIEW_ROW_LIMIT, max_col=max_col, values_only=True
        ):
            preview_rows.append([_coerce_cell(c) for c in row])
            if len(preview_rows) >= PREVIEW_ROW_LIMIT:
                break

    if ws.max_row is not None:
        row_count_estimate = ws.max_row
    else:
        row_count_estimate = len(preview_rows)

    data_start_row, merged_headers = _detect_header_row(preview_rows)
    # header_row_index 是"表头所在行"的 0-based 索引（合并表头时取下行）
    header_row_index = max(data_start_row - 1, 0)

    # F7: 规范化表头（剥方括号 + 拆组合表头）
    normalized_headers, compound_headers = _normalize_header_row(merged_headers)

    # F6: 提取文件名元信息作为识别降级信号
    filename_hint = _extract_filename_hints(filename)

    detection_evidence: dict = {
        "header_cells": normalized_headers,
        "header_cells_raw": merged_headers,
        "merged_header": _is_merged_header(preview_rows, data_start_row, merged_headers),
        "compound_headers": compound_headers,
        "filename_hint": filename_hint,
        "amount_unit": _extract_amount_unit(preview_rows, data_start_row),
    }

    return SheetDetection(
        file_name=filename,
        sheet_name=sheet_name,
        row_count_estimate=row_count_estimate,
        header_row_index=header_row_index,
        data_start_row=data_start_row,
        table_type="unknown",
        table_type_confidence=0,
        confidence_level="manual_required",
        adapter_id=None,
        column_mappings=[],
        has_aux_dimension=False,
        aux_dimension_columns=[],
        preview_rows=preview_rows,
        detection_evidence=detection_evidence,
        warnings=[],
    )


# ---------------------------------------------------------------------------
# CSV / TSV
# ---------------------------------------------------------------------------


def _detect_csv(
    content: bytes,
    filename: str,
    *,
    delimiter_hint: Optional[str] = None,
) -> FileDetection:
    """探测 csv/tsv：编码自适应，只读前 20 行。"""

    size = len(content)
    fd = FileDetection(file_name=filename, file_size_bytes=size, file_type="csv")

    text, encoding = _decode_csv_content(content)
    if text is None:
        fd.errors.append(
            make_error(
                ErrorCode.ENCODING_DETECTION_FAILED,
                message=f"无法识别 CSV 编码：{filename}",
                file=filename,
                suggestion="请将文件另存为 UTF-8 或 GBK 编码后重新上传",
            )
        )
        return fd

    fd.encoding = encoding

    # 单 sheet 名：取文件基名（无扩展名）；空字符串时保留 ""
    sheet_name = os.path.splitext(os.path.basename(filename))[0]

    try:
        delimiter = _pick_delimiter(text, delimiter_hint)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        preview_rows: list[list[str]] = []
        for row in reader:
            preview_rows.append([_coerce_cell(c) for c in row])
            if len(preview_rows) >= PREVIEW_ROW_LIMIT:
                break
    except Exception as exc:  # noqa: BLE001 - csv 解析失败一律视为损坏
        logger.exception("csv parse failed for %s", filename)
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"解析 CSV 失败 {filename}：{exc}",
                file=filename,
            )
        )
        return fd

    data_start_row, merged_headers = _detect_header_row(preview_rows)
    header_row_index = max(data_start_row - 1, 0)
    row_count_estimate = len(preview_rows)

    # F7: 规范化表头（剥方括号 + 拆组合表头）
    normalized_headers, compound_headers = _normalize_header_row(merged_headers)

    # F6: 提取文件名元信息作为识别降级信号
    filename_hint = _extract_filename_hints(filename)

    detection_evidence: dict = {
        "header_cells": normalized_headers,
        "header_cells_raw": merged_headers,
        "merged_header": _is_merged_header(preview_rows, data_start_row, merged_headers),
        "compound_headers": compound_headers,
        "filename_hint": filename_hint,
        "amount_unit": _extract_amount_unit(preview_rows, data_start_row),
    }

    fd.sheets.append(
        SheetDetection(
            file_name=filename,
            sheet_name=sheet_name,
            row_count_estimate=row_count_estimate,
            header_row_index=header_row_index,
            data_start_row=data_start_row,
            table_type="unknown",
            table_type_confidence=0,
            confidence_level="manual_required",
            adapter_id=None,
            column_mappings=[],
            has_aux_dimension=False,
            aux_dimension_columns=[],
            preview_rows=preview_rows,
            detection_evidence=detection_evidence,
            warnings=[],
        )
    )
    return fd


def _decode_csv_content(content: bytes) -> tuple[Optional[str], Optional[str]]:
    """委托给 ``encoding_detector.decode_content``，返回 ``(text, encoding)``。

    Task 7 已把编码探测抽离为独立模块（BOM → 候选列表 → chardet → latin1）。
    这里保留兼容签名，丢弃 confidence（detector.py 现有逻辑只需要 text/encoding，
    失败时返回 ``(None, None)``）。
    """

    text, encoding, _confidence = decode_content(content)
    if text is None:
        return None, None
    return text, encoding


def _pick_delimiter(text: str, hint: Optional[str]) -> str:
    """选择 CSV 分隔符：hint → Sniffer → 逗号兜底。"""

    if hint is not None:
        return hint
    probe = text[:4096]
    if not probe:
        return ","
    try:
        dialect = csv.Sniffer().sniff(probe, delimiters=",\t;|")
        return dialect.delimiter
    except csv.Error:
        return ","


# ---------------------------------------------------------------------------
# ZIP
# ---------------------------------------------------------------------------


def _detect_zip(content: bytes, filename: str) -> FileDetection:
    """探测 zip：递归 detect 内部文件；内部 sheet 前缀 ``archive.zip!inner``。"""

    size = len(content)
    fd = FileDetection(file_name=filename, file_size_bytes=size, file_type="zip")

    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except Exception as exc:  # noqa: BLE001 - 含 BadZipFile 等各类打开错误
        logger.exception("zipfile open failed for %s", filename)
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"无法打开 ZIP 文件 {filename}:{exc}",
                file=filename,
            )
        )
        return fd

    try:
        supported_exts = _XLSX_EXTS | _CSV_EXTS | _XLS_EXTS | _ZIP_EXTS
        for info in zf.infolist():
            if info.is_dir():
                continue
            inner_name = _decode_zip_entry_name(info)
            inner_ext = os.path.splitext(inner_name)[1].lower()

            if inner_ext not in supported_exts:
                logger.debug(
                    "zip: skip unsupported entry %s in %s", inner_name, filename
                )
                continue

            try:
                inner_content = zf.read(info)
            except Exception as exc:  # noqa: BLE001
                logger.exception(
                    "zip: read entry %s failed in %s", inner_name, filename
                )
                fd.errors.append(
                    make_error(
                        ErrorCode.CORRUPTED_FILE,
                        message=(
                            f"读取 ZIP 条目失败 {inner_name}"
                            f"（位于 {filename}）：{exc}"
                        ),
                        file=filename,
                    )
                )
                continue

            inner_fd = detect_file(inner_content, inner_name)

            # 内部文件的 errors 上浮（保留原 file 定位）
            fd.errors.extend(inner_fd.errors)

            prefix = f"{filename}!{inner_name}"
            for sheet in inner_fd.sheets:
                sheet.file_name = prefix
                fd.sheets.append(sheet)
    finally:
        try:
            zf.close()
        except Exception:  # noqa: BLE001
            pass

    return fd


def _decode_zip_entry_name(info: zipfile.ZipInfo) -> str:
    """解码 ZIP 条目名：UTF-8 flag 优先；否则 CP437 → gbk 重解码（需求 14）。"""

    # flag_bits 第 11 位（0x800）为 1 表示 UTF-8 编码的文件名
    if info.flag_bits & 0x800:
        return info.filename

    # zipfile 默认按 CP437 解码，中文易乱码；尝试重编码为 gbk
    try:
        raw = info.filename.encode("cp437")
        return raw.decode("gbk")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return info.filename


# ---------------------------------------------------------------------------
# 单元格强制类型 + 表头行 stub
# ---------------------------------------------------------------------------


def _coerce_cell(value: Any) -> str:
    """把 openpyxl / csv 读出的任意单元格值转成 str。

    规则：
    - ``None`` → ``""``
    - ``datetime`` / ``date`` / ``time`` → ISO 字符串
    - ``bytes`` → utf-8 解码（errors=replace）
    - 其余类型 → ``str(value)``
    """

    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        try:
            return value.isoformat()
        except Exception:  # noqa: BLE001 - 损坏的 datetime 兜底
            return str(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8", errors="replace")
        except Exception:  # noqa: BLE001
            return str(value)
    return str(value)


def _detect_header_row(rows: list[list[str]]) -> tuple[int, list[str]]:
    """识别表头行并返回 ``(data_start_row_index, merged_headers)``。

    通用算法（v2.1，对齐 design §28.2）：基于"行间值多样性"判断表头边界。

    算法：
    1. 对前 10 行计算每行的：
       - ``unique_value_count``：去重后非空单元格值的数量
       - ``fill_ratio``：非空单元格数 / 总列数
    2. **标题/横幅行判定**：``unique_value_count <= 2``
       （所有单元格值相同 = 合并单元格横幅，或只有 1-2 个不同值 = 元信息行）
       → 跳过这些行
    3. **表头行判定**：跳过横幅后，第一行满足
       ``fill_ratio >= 0.3 AND unique_value_count >= 3``
       且值主要为文本（非纯数字）→ 表头起始行
    4. **合并表头判定**：如果表头起始行的下一行也满足表头条件
       （fill_ratio >= 0.3, unique >= 3, 文本值），且两行呈互补模式
       （上行稀疏+下行密集，或上行有分组名下行有子列名）
       → 合并：forward-fill 上行 + "top.bottom" 拼接
    5. **数据起始行**：表头结束后的第一行

    返回
    ----
    ``(data_start_row_index, merged_headers)``

    - ``data_start_row_index``：数据起始行的 0-based 索引
    - ``merged_headers``：合成后的表头列表
    """

    if not rows:
        return 0, []

    # 全部行均无非空单元格 → 直接返回 (0, [])
    if not any(any(str(c or "").strip() for c in row) for row in rows):
        return 0, []

    # --- 计算前 10 行的行特征 ---
    scan_limit = min(10, len(rows))
    row_features: list[tuple[int, float, list[str], bool]] = []
    # (unique_count, fill_ratio, non_empty_values, is_text_row)

    for i in range(scan_limit):
        row = rows[i]
        cells = [str(c or "").strip() for c in row]
        width = len(cells) if cells else 1
        non_empty = [c for c in cells if c]
        unique_values = set(non_empty)
        unique_count = len(unique_values)
        fill_ratio = len(non_empty) / width if width else 0.0

        # 判断是否为文本行（非纯数字）
        numeric_count = 0
        for v in non_empty:
            candidate = v.replace(",", "").replace(" ", "")
            try:
                float(candidate)
                numeric_count += 1
            except (ValueError, TypeError):
                pass
        is_text = (len(non_empty) - numeric_count) >= max(1, len(non_empty) * 0.3) if non_empty else False

        row_features.append((unique_count, fill_ratio, non_empty, is_text))

    # --- 1. 跳过标题/横幅行 ---
    # 标题行特征：unique_value_count <= 2 AND looks like banner
    skip = 0
    for i in range(scan_limit):
        unique_count, fill_ratio, non_empty, is_text = row_features[i]

        if not non_empty:
            # 全空行：如果在连续 skip 序列中，继续跳过
            if skip == i:
                skip = i + 1
            continue

        # 核心判据：unique_value_count <= 2 → 可能是横幅/元信息行
        if unique_count <= 2:
            # 额外验证：确实像横幅（含关键词或所有值相同）
            if _looks_like_banner_v2(non_empty, unique_count):
                skip = i + 1
            else:
                # unique <= 2 但不像横幅 → 可能是稀疏表头的上行
                break
        else:
            break

    # 已全部跳过 → 兜底
    if skip >= len(rows):
        return skip, []

    # --- 2. 寻找表头行 ---
    # 从 skip 位置开始，找第一行满足表头条件的行
    header_start = skip
    found_header = False

    for i in range(skip, scan_limit):
        unique_count, fill_ratio, non_empty, is_text = row_features[i]

        if not non_empty:
            continue

        # 表头条件：fill_ratio >= 0.3 AND unique >= 3 AND 主要是文本
        if fill_ratio >= 0.3 and unique_count >= 3 and is_text:
            header_start = i
            found_header = True
            break
        # 稀疏表头候选：unique >= 2 AND is_text（可能是合并表头的上行）
        elif unique_count >= 2 and is_text and fill_ratio > 0:
            # 检查下一行是否是密集表头行（互补模式）
            next_i = i + 1
            if next_i < scan_limit:
                next_unique, next_fill, next_non_empty, next_is_text = row_features[next_i]
                if next_fill >= 0.3 and next_unique >= 3 and next_is_text:
                    # 这是合并表头：当前行是稀疏上行，下一行是密集下行
                    header_start = i
                    found_header = True
                    break
            # 如果下一行不满足，继续寻找
            continue

    if not found_header:
        # 兜底：取 skip 后第一个非空行
        for i in range(skip, len(rows)):
            cells = [str(c or "").strip() for c in rows[i]]
            if any(cells):
                header = cells
                return i + 1, header
        return skip, []

    # --- 3. 合并表头判定 ---
    # 检查 header_start 的下一行是否也是表头行（互补模式）
    next_idx = header_start + 1
    if next_idx < scan_limit:
        cur_unique = row_features[header_start][0]
        cur_fill = row_features[header_start][1]
        next_unique, next_fill, next_non_empty, next_is_text = row_features[next_idx]

        # 合并表头条件：
        # A) 当前行稀疏（unique <= 少数分组名）+ 下行密集（unique >= 3）
        # B) 或者两行都满足表头条件且有互补位置/重复值
        # C) 下行 unique 少但 fill_ratio 高且有重复值（典型"借方/贷方"子列模式）
        # CRITICAL: 下行必须看起来像表头（短标签），不能是数据行
        is_merged_candidate = False

        # 先检查下行是否像表头（而非数据行）
        # 数据行特征（任一命中即判定为数据行）：
        # - 平均值长度 > 10 字符
        # - 长值（>15字符）≥ 3 个
        # - 第一个非空值是纯数字（序号列）
        # - 含日期格式值（YYYY-MM-DD / YYYY/MM/DD）
        # - 含金额格式值（带逗号的数字如 "39,414.01"）
        # - unique 值数量 > 上行 unique 的 2 倍（数据行值通常都不同）
        bottom_row_candidate = rows[next_idx]
        bottom_cells = [str(c or "").strip() for c in bottom_row_candidate]
        bottom_non_empty_cells = [c for c in bottom_cells if c]
        avg_len = (sum(len(c) for c in bottom_non_empty_cells) / len(bottom_non_empty_cells)
                   if bottom_non_empty_cells else 0)
        long_values = sum(1 for c in bottom_non_empty_cells if len(c) > 15)
        looks_like_data_row = avg_len > 10 or long_values >= 3

        if not looks_like_data_row and bottom_non_empty_cells:
            # 额外数据行检测：第一个非空值是纯数字（序号）
            first_val = bottom_non_empty_cells[0]
            try:
                int(first_val)
                # 第一列是整数 → 很可能是序号列 → 数据行
                looks_like_data_row = True
            except (ValueError, TypeError):
                pass

        if not looks_like_data_row and bottom_non_empty_cells:
            # 额外数据行检测：含日期格式值
            import re as _re
            _date_pattern = _re.compile(r"^\d{4}[-/]\d{2}[-/]\d{2}")
            _amount_pattern = _re.compile(r"^[\d,]+\.\d{2}$")
            date_count = sum(1 for v in bottom_non_empty_cells if _date_pattern.match(v))
            amount_count = sum(1 for v in bottom_non_empty_cells if _amount_pattern.match(v.replace(",", "")))
            if date_count >= 1 or amount_count >= 1:
                looks_like_data_row = True

        if not looks_like_data_row and next_is_text and next_fill >= 0.3:
            if cur_unique <= 4 and next_unique >= 3:
                # 经典合并表头：上行是分组名（2-4个），下行是子列名
                is_merged_candidate = True
            elif cur_fill >= 0.3 and cur_unique >= 3 and next_unique >= 3:
                # 两行都像表头 — 检查是否为"分组+子列"模式
                top_row = rows[header_start]
                top_cells_check = [str(c or "").strip() for c in top_row]
                top_non_empty_vals = [c for c in top_cells_check if c]
                has_repeats = len(top_non_empty_vals) > len(set(top_non_empty_vals))

                # 上下行的值集合不完全相同
                top_set = set(top_non_empty_vals)
                bottom_set = set(bottom_non_empty_cells)
                different_values = top_set != bottom_set

                if has_repeats and different_values:
                    is_merged_candidate = True
                elif _has_complementary_positions(
                    top_cells_check + [""] * (max(len(top_cells_check), len(bottom_cells)) - len(top_cells_check)),
                    bottom_cells + [""] * (max(len(top_cells_check), len(bottom_cells)) - len(bottom_cells))
                ):
                    is_merged_candidate = True
            elif (
                cur_fill >= 0.3
                and cur_unique >= 3
                and next_unique >= 2
                and next_fill >= 0.5
                and len(bottom_non_empty_cells) > next_unique * 2
            ):
                # 子列重复模式：下行 unique 少（如只有"借方金额/贷方金额"）
                # 但 fill_ratio 高且有大量重复值 → 典型的分组子列表头
                # 条件：non_empty 数量 > unique * 2（说明值在重复）
                is_merged_candidate = True

        if is_merged_candidate:
            top_row = rows[header_start]
            bottom_row = rows[next_idx]
            width = max(len(top_row), len(bottom_row))
            top_padded = [str(c or "").strip() for c in top_row] + [""] * (width - len(top_row))
            bottom_padded = [str(c or "").strip() for c in bottom_row] + [""] * (width - len(bottom_row))

            # forward-fill top：空格继承左邻非空值
            filled_top: list[str] = []
            last = ""
            for c in top_padded:
                if c:
                    last = c
                filled_top.append(last)

            merged: list[str] = []
            for t, b in zip(filled_top, bottom_padded):
                if t and b and t != b:
                    merged.append(f"{t}.{b}")
                elif b:
                    merged.append(b)
                elif t:
                    merged.append(t)
                else:
                    merged.append("")
            return next_idx + 1, merged

    # --- 4. 单行表头 ---
    header = [str(c or "").strip() for c in rows[header_start]]
    return header_start + 1, header


def _looks_like_banner(non_empty_cells: list[str]) -> bool:
    """判断非空单元格拼起来是否像一张"XX 公司 YYYY 年度 科目余额表"横幅。

    真正的空行（``non_empty_cells`` 为空）不视为横幅，避免把"全空输入"误判
    为大段标题跳过后返回一个非零的 data_start_row。
    """

    if not non_empty_cells:
        return False
    joined = "".join(non_empty_cells)
    banner_keywords = (
        "公司",
        "年度",
        "报表",
        "科目余额表",
        "序时账",
        "凭证明细",
    )
    return any(kw in joined for kw in banner_keywords)


def _extract_amount_unit(preview_rows: list[list[str]], data_start_row: int) -> str | None:
    """从 data_start_row 之前的横幅行中提取金额单位。

    常见模式：
    - "金额单位：万元"
    - "单位：万元"
    - "金额单位:元"
    - "单位（万元）"
    - 单元格值直接是 "万元" / "元" / "千元"

    返回 "万元" / "元" / "千元" / None（未识别）
    """
    import re

    unit_pattern = re.compile(
        r'(?:金额)?单位[：:\s]*([万千百]?元)|'
        r'(?:金额)?单位[（(]([万千百]?元)[)）]|'
        r'^([万千百]元)$'
    )

    for i in range(min(data_start_row, len(preview_rows))):
        row = preview_rows[i]
        for cell in row:
            if not cell:
                continue
            cell_str = str(cell).strip()
            m = unit_pattern.search(cell_str)
            if m:
                unit = m.group(1) or m.group(2) or m.group(3)
                if unit:
                    return unit
    return None


def _looks_like_banner_v2(non_empty_values: list[str], unique_count: int) -> bool:
    """v2.1 通用横幅判定：unique_count <= 2 时进一步确认是否为横幅。

    判定逻辑：
    - 所有值相同（合并单元格横幅）→ True
    - 含横幅关键词（公司/年度/报表/余额表/序时账/凭证/核算组织/期间）→ True
    - unique_count == 1 → True（单一值重复 = 合并单元格）
    - unique_count == 2 且含元信息关键词 → True
    """
    if not non_empty_values:
        return False

    # 所有值相同 = 合并单元格横幅
    if unique_count <= 1:
        return True

    joined = "".join(non_empty_values)
    banner_keywords = (
        "公司",
        "年度",
        "报表",
        "科目余额表",
        "序时账",
        "凭证明细",
        "核算组织",
        "会计期间",
        "开始期间",
        "截止期间",
        "编制单位",
        "金额单位",
    )
    return any(kw in joined for kw in banner_keywords)


def _has_complementary_positions(top: list[str], bottom: list[str]) -> bool:
    """检测两行是否在列位置上互补（上行空的位置下行有值，反之亦然）。

    用于合并表头判定：经典模式是上行只在分组起始位置有值（如"年初余额"），
    下行在每个子列都有值（如"借方金额"、"贷方金额"）。
    """
    if not top or not bottom:
        return False

    width = min(len(top), len(bottom))
    # 统计上行空但下行非空的位置数
    top_empty_bottom_filled = sum(
        1 for i in range(width) if not top[i] and bottom[i]
    )
    # 如果有较多这样的互补位置，说明是合并表头
    return top_empty_bottom_filled >= 2


def _is_merged_header(
    rows: list[list[str]],
    data_start_row: int,
    merged_headers: list[str],
) -> bool:
    """在 detection_evidence 里标记是否为合并表头。

    合并表头的判据：
    - ``data_start_row >= 2``（即至少跨了 2 行表头）
    - 合成的表头里至少有一个 "top.bottom" 风格的拼接项（含 ``.`` 且两侧皆非空）
    """

    if data_start_row < 2:
        return False
    for cell in merged_headers:
        if cell and "." in cell:
            left, _, right = cell.partition(".")
            if left and right:
                return True
    return False


__all__ = [
    "detect_file",
    "detect_file_from_path",
    "_extract_filename_hints",
    "_normalize_header",
    "_normalize_header_row",
]


# ---------------------------------------------------------------------------
# Path-based detection (大文件支持，不全量读入内存)
# ---------------------------------------------------------------------------


def _detect_csv_from_path(
    path: str,
    filename: str,
    *,
    file_size: int,
    delimiter_hint: Optional[str] = None,
) -> FileDetection:
    """从文件路径探测 CSV，只读前 64KB + 前 20 行。

    对于 600MB 的 CSV，内存占用 < 1MB（仅 64KB probe + 20 行文本）。
    """
    from .encoding_detector import detect_encoding

    fd = FileDetection(file_name=filename, file_size_bytes=file_size, file_type="csv")

    # 只读前 64KB 做编码探测
    try:
        with open(path, "rb") as f:
            probe_bytes = f.read(_CSV_PROBE_BYTES)
    except OSError as exc:
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"无法读取文件 {filename}：{exc}",
                file=filename,
            )
        )
        return fd

    # 编码探测
    encoding, confidence = detect_encoding(probe_bytes)
    fd.encoding = encoding

    # 用探测到的编码打开文件，流式读前 20 行
    sheet_name = os.path.splitext(os.path.basename(filename))[0]
    try:
        with open(path, "r", encoding=encoding, errors="replace", newline="") as f:
            # 先读前 4KB 做分隔符探测
            probe_text = f.read(4096)
            f.seek(0)

            delimiter = _pick_delimiter(probe_text, delimiter_hint)
            reader = csv.reader(f, delimiter=delimiter)
            preview_rows: list[list[str]] = []
            for row in reader:
                preview_rows.append([_coerce_cell(c) for c in row])
                if len(preview_rows) >= PREVIEW_ROW_LIMIT:
                    break
    except Exception as exc:  # noqa: BLE001
        logger.exception("csv parse from path failed for %s", filename)
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"解析 CSV 失败 {filename}：{exc}",
                file=filename,
            )
        )
        return fd

    data_start_row, merged_headers = _detect_header_row(preview_rows)
    header_row_index = max(data_start_row - 1, 0)

    # 估算行数：file_size / avg_row_bytes（基于前 20 行）
    if preview_rows:
        total_chars = sum(sum(len(c) for c in row) for row in preview_rows)
        avg_row_bytes = max(total_chars // len(preview_rows), 1)
        row_count_estimate = file_size // avg_row_bytes
    else:
        row_count_estimate = 0

    # F7: 规范化表头（剥方括号 + 拆组合表头）
    normalized_headers, compound_headers = _normalize_header_row(merged_headers)

    # F6: 提取文件名元信息作为识别降级信号
    filename_hint = _extract_filename_hints(filename)

    detection_evidence: dict = {
        "header_cells": normalized_headers,
        "header_cells_raw": merged_headers,
        "merged_header": _is_merged_header(preview_rows, data_start_row, merged_headers),
        "compound_headers": compound_headers,
        "filename_hint": filename_hint,
        "amount_unit": _extract_amount_unit(preview_rows, data_start_row),
    }

    fd.sheets.append(
        SheetDetection(
            file_name=filename,
            sheet_name=sheet_name,
            row_count_estimate=row_count_estimate,
            header_row_index=header_row_index,
            data_start_row=data_start_row,
            table_type="unknown",
            table_type_confidence=0,
            confidence_level="manual_required",
            adapter_id=None,
            column_mappings=[],
            has_aux_dimension=False,
            aux_dimension_columns=[],
            preview_rows=preview_rows,
            detection_evidence=detection_evidence,
            warnings=[],
        )
    )
    return fd


def _detect_xlsx_from_path(
    path: str,
    filename: str,
    *,
    file_size: int,
) -> FileDetection:
    """从文件路径探测 xlsx。

    用 openpyxl read_only=True 真流式读前 20 行 XML，内存占用与 sheet 数据量无关
    （~10-50MB 级别，取决于 shared_strings 表大小）。
    历史：曾试 calamine 加速 detect，实测 calamine 必须全量解码 sheet（YG2101 序时账
    650k 行 17.81s），不适合 "只读前 20 行" 场景，已移除。
    """
    fd = FileDetection(file_name=filename, file_size_bytes=file_size, file_type="xlsx")

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"缺少 openpyxl 依赖：{exc}",
                file=filename,
            )
        )
        return fd

    # 先尝试 read_only=True
    fd = _detect_xlsx_from_path_with_mode(openpyxl, path, filename, fd, read_only=True)

    # 检测行宽异常（与 _detect_xlsx 相同的回退逻辑）
    needs_fallback = False
    for sheet_det in fd.sheets:
        if sheet_det.preview_rows and len(sheet_det.preview_rows) >= 3:
            max_width = max(len(row) for row in sheet_det.preview_rows)
            if max_width <= 2:
                needs_fallback = True
                logger.info(
                    "read_only mode returned narrow rows for %s, falling back",
                    filename,
                )
                break

    if needs_fallback:
        fd.sheets = []
        fd.errors = []
        fd = _detect_xlsx_from_path_with_mode(openpyxl, path, filename, fd, read_only=False)

    return fd


def _detect_xlsx_from_path_with_mode(
    openpyxl: Any,
    path: str,
    filename: str,
    fd: FileDetection,
    *,
    read_only: bool,
) -> FileDetection:
    """用指定模式从文件路径打开 xlsx 并解析。"""
    try:
        wb = openpyxl.load_workbook(
            path,  # 直接传路径，openpyxl 内部流式读取
            read_only=read_only,
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001
        logger.exception(
            "openpyxl.load_workbook(path=%s, read_only=%s) failed", path, read_only
        )
        fd.errors.append(
            make_error(
                ErrorCode.CORRUPTED_FILE,
                message=f"无法打开 xlsx 文件 {filename}：{exc}",
                file=filename,
                suggestion="请检查文件是否损坏或被其他程序占用",
            )
        )
        return fd

    try:
        for sheet_name in wb.sheetnames:
            try:
                sheet_det = _detect_xlsx_sheet(
                    wb, sheet_name, filename, read_only=read_only
                )
                fd.sheets.append(sheet_det)
            except Exception as exc:  # noqa: BLE001
                logger.exception(
                    "failed to parse sheet %s in %s", sheet_name, filename
                )
                fd.errors.append(
                    make_error(
                        ErrorCode.CORRUPTED_FILE,
                        message=f"解析 sheet {sheet_name!r} 失败（文件 {filename}）：{exc}",
                        file=filename,
                        sheet=sheet_name,
                    )
                )
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass

    return fd
