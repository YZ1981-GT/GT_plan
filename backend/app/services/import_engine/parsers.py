# -*- coding: utf-8 -*-
"""数据导入解析器工厂 - BaseParser + GenericParser + 财务软件解析器

Validates: Requirements 4.1, 4.2
"""

import csv
import io
from abc import ABC, abstractmethod
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from fastapi import HTTPException


# ---------------------------------------------------------------------------
# Column name mappings for each data type (Chinese → English)
# ---------------------------------------------------------------------------

_BALANCE_COLUMN_MAP = {
    # Chinese
    "科目编码": "account_code",
    "科目代码": "account_code",
    "科目编号": "account_code",
    "编码": "account_code",
    "编号": "account_code",
    "会计科目编码": "account_code",
    "会计科目代码": "account_code",
    "科目": "account_code",
    "科目名称": "account_name",
    "名称": "account_name",
    "会计科目名称": "account_name",
    "科目全称": "account_name",
    "借贷方向": "direction",
    "方向": "direction",
    "余额方向": "direction",
    "科目方向": "direction",
    "科目级次": "level",
    "级次": "level",
    "科目类别": "category",
    "类别": "category",
    "科目类型": "category",
    "上级编码": "parent_code",
    "上级科目": "parent_code",
    "上级科目编码": "parent_code",
    "父科目编码": "parent_code",
    "期初余额": "opening_balance",
    "期初金额": "opening_balance",
    "期初": "opening_balance",
    "年初余额": "opening_balance",
    "年初金额": "opening_balance",
    "期初数": "opening_balance",
    "年初数": "opening_balance",
    "期初数量": "opening_qty",
    "期初外币": "opening_fc",
    "本期发生额": "debit_amount",
    "本年累计": "year_debit_amount",
    "借方发生额": "debit_amount",
    "借方金额": "debit_amount",
    "本期借方": "debit_amount",
    "本期借方发生额": "debit_amount",
    "借方累计": "debit_amount",
    "借方": "debit_amount",
    "贷方发生额": "credit_amount",
    "贷方金额": "credit_amount",
    "本期贷方": "credit_amount",
    "本期贷方发生额": "credit_amount",
    "贷方累计": "credit_amount",
    "贷方": "credit_amount",
    "期末余额": "closing_balance",
    "期末金额": "closing_balance",
    "期末": "closing_balance",
    "期末数": "closing_balance",
    "币种": "currency_code",
    "币种名称": "currency_code",
    "单位编码": "company_code",
    "公司编码": "company_code",
    "组织编码": "company_code",
    "期初方向": "opening_direction",
    "期末方向": "closing_direction",
    # English
    "account_code": "account_code",
    "account_name": "account_name",
    "direction": "direction",
    "level": "level",
    "category": "category",
    "parent_code": "parent_code",
    "opening_balance": "opening_balance",
    "opening_qty": "opening_qty",
    "opening_fc": "opening_fc",
    "debit_amount": "debit_amount",
    "credit_amount": "credit_amount",
    "closing_balance": "closing_balance",
    "currency_code": "currency_code",
    "company_code": "company_code",
}

_LEDGER_COLUMN_MAP = {
    "科目编码": "account_code",
    "科目代码": "account_code",
    "科目编号": "account_code",
    "编码": "account_code",
    "编号": "account_code",
    "会计科目编码": "account_code",
    "科目": "account_code",
    "科目名称": "account_name",
    "名称": "account_name",
    "会计科目名称": "account_name",
    "凭证日期": "voucher_date",
    "日期": "voucher_date",
    "记账日期": "voucher_date",
    "制单日期": "voucher_date",
    "业务日期": "voucher_date",
    "会计月份": "accounting_period",
    "会计期间": "accounting_period",
    "月份": "accounting_period",
    "凭证月份": "accounting_period",
    "期间": "accounting_period",
    "凭证类型": "voucher_type",
    "凭证字": "voucher_type",
    "凭证号": "voucher_no",
    "凭证编号": "voucher_no",
    "凭证字号": "voucher_no",
    "凭证号数": "voucher_no",
    "记账凭证号": "voucher_no",
    "分录序号": "entry_seq",
    "序号": "entry_seq",
    "分录行号": "entry_seq",
    "分录号": "entry_seq",
    "借方金额": "debit_amount",
    "借方发生额": "debit_amount",
    "借方": "debit_amount",
    "本期借方": "debit_amount",
    "本期借方发生额": "debit_amount",
    "贷方金额": "credit_amount",
    "贷方发生额": "credit_amount",
    "贷方": "credit_amount",
    "本期贷方": "credit_amount",
    "本期贷方发生额": "credit_amount",
    "借方数量": "debit_qty",
    "贷方数量": "credit_qty",
    "借方外币发生额": "debit_fc",
    "借方外币": "debit_fc",
    "贷方外币发生额": "credit_fc",
    "贷方外币": "credit_fc",
    "对方科目": "counterpart_account",
    "对方科目编码": "counterpart_account",
    "对方科目名称": "counterpart_account",
    "摘要": "summary",
    "备注": "summary",
    "制单人": "preparer",
    "编制人": "preparer",
    "录入人": "preparer",
    "操作员": "preparer",
    "填制人": "preparer",
    "币种": "currency_code",
    "币种名称": "currency_code",
    "单位编码": "company_code",
    "公司编码": "company_code",
    "审核人": "reviewer",
    "记账人": "bookkeeper",
    "辅助账核算项目": "aux_info",
    "凭证张数": "voucher_count",
    "计量单位": "unit",
    # English
    "account_code": "account_code",
    "account_name": "account_name",
    "voucher_date": "voucher_date",
    "accounting_period": "accounting_period",
    "voucher_type": "voucher_type",
    "voucher_no": "voucher_no",
    "entry_seq": "entry_seq",
    "debit_amount": "debit_amount",
    "credit_amount": "credit_amount",
    "debit_qty": "debit_qty",
    "credit_qty": "credit_qty",
    "debit_fc": "debit_fc",
    "credit_fc": "credit_fc",
    "counterpart_account": "counterpart_account",
    "summary": "summary",
    "preparer": "preparer",
    "currency_code": "currency_code",
    "company_code": "company_code",
    "reviewer": "reviewer",
    "bookkeeper": "bookkeeper",
}

_AUX_BALANCE_COLUMN_MAP = {
    "科目编码": "account_code",
    "科目代码": "account_code",
    "科目编号": "account_code",
    "编码": "account_code",
    "科目名称": "account_name",
    "名称": "account_name",
    "辅助类型": "aux_type",
    "核算类型": "aux_type",
    "核算维度": "aux_type",
    "核算项目类型编号": "aux_type",
    "辅助类型名称": "aux_type_name",
    "核算项目类型名称": "aux_type_name",
    "辅助编码": "aux_code",
    "核算编码": "aux_code",
    "核算项目编号": "aux_code",
    "核算项目编码": "aux_code",
    "辅助名称": "aux_name",
    "核算名称": "aux_name",
    "核算项目名称": "aux_name",
    "项目": "aux_name",
    "借贷方向": "direction",
    "方向": "direction",
    "期初余额": "opening_balance",
    "期初金额": "opening_balance",
    "期初": "opening_balance",
    "期初数量": "opening_qty",
    "期初外币": "opening_fc",
    "借方发生额": "debit_amount",
    "借方金额": "debit_amount",
    "贷方发生额": "credit_amount",
    "贷方金额": "credit_amount",
    "期末余额": "closing_balance",
    "期末金额": "closing_balance",
    "期末": "closing_balance",
    "币种": "currency_code",
    "单位编码": "company_code",
    "公司编码": "company_code",
    # English
    "account_code": "account_code",
    "account_name": "account_name",
    "aux_type": "aux_type",
    "aux_type_name": "aux_type_name",
    "aux_code": "aux_code",
    "aux_name": "aux_name",
    "direction": "direction",
    "opening_balance": "opening_balance",
    "opening_qty": "opening_qty",
    "opening_fc": "opening_fc",
    "debit_amount": "debit_amount",
    "credit_amount": "credit_amount",
    "closing_balance": "closing_balance",
    "currency_code": "currency_code",
    "company_code": "company_code",
    # 新增：丰桔出行等企业导出格式
    "期初方向": "opening_direction",
    "期末方向": "closing_direction",
    "借方累计": "debit_amount",
    "贷方累计": "credit_amount",
    "项目": "aux_name",
    "opening_direction": "opening_direction",
    "closing_direction": "closing_direction",
}

_AUX_LEDGER_COLUMN_MAP = {
    "科目编码": "account_code",
    "科目代码": "account_code",
    "科目编号": "account_code",
    "编码": "account_code",
    "科目名称": "account_name",
    "名称": "account_name",
    "辅助类型": "aux_type",
    "核算类型": "aux_type",
    "核算维度": "aux_type",
    "核算项目类型编号": "aux_type",
    "辅助类型名称": "aux_type_name",
    "核算项目类型名称": "aux_type_name",
    "辅助编码": "aux_code",
    "核算编码": "aux_code",
    "核算项目编号": "aux_code",
    "核算项目编码": "aux_code",
    "辅助名称": "aux_name",
    "核算名称": "aux_name",
    "核算项目名称": "aux_name",
    "项目": "aux_name",
    "凭证日期": "voucher_date",
    "日期": "voucher_date",
    "记账日期": "voucher_date",
    "会计月份": "accounting_period",
    "会计期间": "accounting_period",
    "凭证月份": "accounting_period",
    "期间": "accounting_period",
    "凭证类型": "voucher_type",
    "凭证字": "voucher_type",
    "凭证号": "voucher_no",
    "凭证编号": "voucher_no",
    "凭证字号": "voucher_no",
    "借方金额": "debit_amount",
    "借方发生额": "debit_amount",
    "贷方金额": "credit_amount",
    "贷方发生额": "credit_amount",
    "借方": "debit_amount",
    "贷方": "credit_amount",
    "借方数量": "debit_qty",
    "贷方数量": "credit_qty",
    "借方外币发生额": "debit_fc",
    "借方外币": "debit_fc",
    "贷方外币发生额": "credit_fc",
    "贷方外币": "credit_fc",
    "摘要": "summary",
    "制单人": "preparer",
    "编制人": "preparer",
    "填制人": "preparer",
    "币种": "currency_code",
    "币种名称": "currency_code",
    "单位编码": "company_code",
    "公司编码": "company_code",
    "审核人": "reviewer",
    "记账人": "bookkeeper",
    "分录号": "entry_seq",
    "单价": "unit_price",
    "汇率": "exchange_rate",
    "计量单位": "unit",
    # English
    "account_code": "account_code",
    "account_name": "account_name",
    "aux_type": "aux_type",
    "aux_type_name": "aux_type_name",
    "aux_code": "aux_code",
    "aux_name": "aux_name",
    "voucher_date": "voucher_date",
    "accounting_period": "accounting_period",
    "voucher_type": "voucher_type",
    "voucher_no": "voucher_no",
    "debit_amount": "debit_amount",
    "credit_amount": "credit_amount",
    "debit_qty": "debit_qty",
    "credit_qty": "credit_qty",
    "debit_fc": "debit_fc",
    "credit_fc": "credit_fc",
    "summary": "summary",
    "preparer": "preparer",
    "currency_code": "currency_code",
    "company_code": "company_code",
    "reviewer": "reviewer",
    "bookkeeper": "bookkeeper",
    "unit_price": "unit_price",
    "exchange_rate": "exchange_rate",
}

_COLUMN_MAPS = {
    "tb_balance": _BALANCE_COLUMN_MAP,
    "tb_ledger": _LEDGER_COLUMN_MAP,
    "tb_aux_balance": _AUX_BALANCE_COLUMN_MAP,
    "tb_aux_ledger": _AUX_LEDGER_COLUMN_MAP,
}

# Data type aliases
_DATA_TYPE_ALIASES = {
    "balance": "tb_balance",
    "ledger": "tb_ledger",
    "aux_balance": "tb_aux_balance",
    "aux_ledger": "tb_aux_ledger",
    "tb_balance": "tb_balance",
    "tb_ledger": "tb_ledger",
    "tb_aux_balance": "tb_aux_balance",
    "tb_aux_ledger": "tb_aux_ledger",
}


def normalize_data_type(data_type: str) -> str:
    """Normalize data_type to canonical form (tb_balance, tb_ledger, etc.)."""
    normalized = _DATA_TYPE_ALIASES.get(data_type)
    if not normalized:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的数据类型: {data_type}，支持: balance/ledger/aux_balance/aux_ledger",
        )
    return normalized


# ---------------------------------------------------------------------------
# Value parsing helpers
# ---------------------------------------------------------------------------


def _infer_level_from_code(code: str) -> int:
    """从科目编码推断级次（兜底逻辑）。
    点号分隔：按点号数量+1
    纯数字：4位=1级，5-6位=2级，7-8位=3级
    """
    if '.' in code:
        return code.count('.') + 1
    if len(code) <= 4:
        return 1
    if len(code) <= 6:
        return 2
    return 3


def _parse_decimal(value: str | None) -> Decimal | None:
    """Parse a string to Decimal, return None for empty/invalid."""
    if value is None:
        return None
    value = str(value).strip().replace(",", "").replace("，", "")
    if not value or value == "None" or value == "":
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(value: str | None) -> date | None:
    """Parse a string to date, supporting multiple formats."""
    if value is None:
        return None
    # openpyxl returns datetime objects directly
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    value = str(value).strip()
    if not value or value == "None":
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_period(value: str | None) -> int | None:
    """解析期间字符串为会计月份。如 '2025年1期' → 1, '2025年10期' → 10"""
    if value is None:
        return None
    import re
    m = re.search(r'(\d+)\s*期', str(value))
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# BaseParser abstract class
# ---------------------------------------------------------------------------


class BaseParser(ABC):
    """解析器基类 - 所有财务数据解析器的抽象接口。"""

    @abstractmethod
    def parse(self, content: bytes, data_type: str) -> list[dict]:
        """Parse file content into a list of normalized dicts.

        Args:
            content: Raw file bytes (Excel or CSV).
            data_type: One of tb_balance, tb_ledger, tb_aux_balance, tb_aux_ledger.

        Returns:
            List of dicts with normalized English keys and parsed values.
        """
        ...


# ---------------------------------------------------------------------------
# GenericParser - standard template parser
# ---------------------------------------------------------------------------


class GenericParser(BaseParser):
    """通用标准模板解析器 - 解析 Excel/CSV 四表数据。

    支持：
    - 大文件（百万行）：openpyxl read_only 模式流式读取
    - 多 sheet：智能识别"序时账1/序时账2"等命名模式并合并
    - 自动跳过空 sheet 和非数据 sheet（如"说明"、"目录"）
    - parse_streaming(): 流式解析，每次 yield 一批行，避免一次性加载到内存

    Validates: Requirements 4.1, 4.2
    """

    # Sheet 名称匹配模式（用于智能识别数据类型）
    # 注意：匹配时先检查更具体的类型（辅助类），再检查通用类型（序时账/余额表）
    _SHEET_PATTERNS: dict[str, list[str]] = {
        "tb_balance": ["余额", "balance", "科目余额", "试算"],
        "tb_ledger": ["序时", "ledger", "凭证", "总账", "日记账"],
        "tb_aux_balance": ["辅助余额", "核算余额", "aux_bal", "辅助科目余额", "月余额", "辅助账月余额"],
        "tb_aux_ledger": ["辅助明细", "核算明细", "aux_led", "辅助序时", "辅助账明细"],
    }

    # 匹配优先级：更具体的模式优先（辅助余额 > 辅助明细 > 序时账 > 余额表）
    _MATCH_PRIORITY = ["tb_aux_balance", "tb_aux_ledger", "tb_ledger", "tb_balance"]

    # 应跳过的 sheet 名称关键词
    _SKIP_SHEET_KEYWORDS = ["说明", "目录", "封面", "模板", "template", "readme"]

    def parse_streaming(
        self, content: bytes | BinaryIO, data_type: str, chunk_size: int = 1000,
    ):
        """流式解析 Excel 文件 — 每次 yield 一批已解析的行。

        使用 openpyxl read_only=True 模式，避免一次性加载整个文件到内存。
        适用于大文件（26万+行），峰值内存显著低于全量加载。

        CSV 文件不走此方法（CSV 本身是流式的，用 parse() 即可）。

        Args:
            content: Excel 文件字节内容
            data_type: 数据类型 (tb_balance/tb_ledger/tb_aux_balance/tb_aux_ledger)
            chunk_size: 每批行数，默认 1000

        Yields:
            list[dict]: 每批已解析、已规范化的行字典列表
        """
        data_type = normalize_data_type(data_type)
        column_map = _COLUMN_MAPS[data_type]

        workbook_input: bytes | BinaryIO
        if hasattr(content, "seek"):
            content.seek(0)
            workbook_input = content
        else:
            workbook_input = io.BytesIO(content)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                workbook_input, read_only=True, data_only=True,
            )
        except Exception:
            # 非 Excel 文件，使用 CSV 分块流式解析，避免全量加载到内存。
            for rows in self._parse_csv_streaming(content, chunk_size):
                normalized = self._normalize_columns(rows, column_map)
                parsed = self._parse_values(normalized, data_type)
                if parsed:
                    yield parsed
            return

        try:
            for ws in wb.worksheets:
                if data_type and not self._should_include_sheet(ws.title, data_type):
                    continue
                if ws.max_row is None or ws.max_row < 2:
                    continue

                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header = next(rows_iter)
                except StopIteration:
                    continue

                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header)]
                if all(h.startswith("col_") for h in headers):
                    continue

                chunk: list[dict] = []
                for row in rows_iter:
                    if all(cell is None for cell in row):
                        continue
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = str(cell) if cell is not None else ""
                    chunk.append(row_dict)

                    if len(chunk) >= chunk_size:
                        normalized = self._normalize_columns(chunk, column_map)
                        parsed = self._parse_values(normalized, data_type)
                        if parsed:
                            yield parsed
                        chunk = []

                # Yield remaining rows in this sheet
                if chunk:
                    normalized = self._normalize_columns(chunk, column_map)
                    parsed = self._parse_values(normalized, data_type)
                    if parsed:
                        yield parsed
        finally:
            wb.close()

    def parse(self, content: bytes, data_type: str) -> list[dict]:
        """Parse Excel or CSV content for the given data type."""
        data_type = normalize_data_type(data_type)
        column_map = _COLUMN_MAPS[data_type]

        # Try Excel first, then CSV
        rows = self._try_parse_excel(content, data_type)
        if rows is None:
            rows = self._parse_csv(content)

        if not rows:
            return []

        # Normalize column names
        normalized = self._normalize_columns(rows, column_map)

        # Parse values based on data type
        return self._parse_values(normalized, data_type)

    def _should_include_sheet(self, sheet_name: str, data_type: str) -> bool:
        """判断 sheet 是否应该被包含在导入中。

        策略（按优先级）：
        1. 跳过"说明"、"目录"等非数据 sheet
        2. 按优先级匹配 sheet 名称到数据类型（辅助类优先于通用类）
        3. 如果匹配到目标类型 -> 包含
        4. 如果匹配到其他类型 -> 排除
        5. 未匹配任何类型 -> 包含（兜底，让列名映射来过滤）
        """
        name_lower = sheet_name.lower().strip()

        # 1. 跳过非数据 sheet（但如果名称同时含数据关键词则不跳过）
        all_data_kw = []
        for patterns in self._SHEET_PATTERNS.values():
            all_data_kw.extend(patterns)
        is_data_sheet = any(p in name_lower for p in all_data_kw)

        if not is_data_sheet:
            for kw in self._SKIP_SHEET_KEYWORDS:
                if kw in name_lower:
                    return False

        # 2. 按优先级匹配（辅助类优先，避免"辅助明细"被"明细"误匹配）
        matched_type = None
        for dt in self._MATCH_PRIORITY:
            patterns = self._SHEET_PATTERNS.get(dt, [])
            if any(p in name_lower for p in patterns):
                matched_type = dt
                break

        if matched_type is not None:
            return matched_type == data_type

        # 3. 未匹配 -> 包含（可能是 Sheet1 这种通用名称）
        return True


    def _try_parse_excel(self, content: bytes, data_type: str = "") -> list[dict] | None:
        """Parse matching sheets from Excel file.

        智能识别 sheet：
        - "序时账"、"序时账1"、"序时账2"、"明细账" → tb_ledger
        - "余额表"、"科目余额" → tb_balance
        - 自动跳过"说明"、"目录"等非数据 sheet

        Uses read_only mode for large files (百万行级别).
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception:
            return None

        all_rows: list[dict] = []
        sheets_processed = 0

        for ws in wb.worksheets:
            # 智能判断是否应包含此 sheet
            if data_type and not self._should_include_sheet(ws.title, data_type):
                continue

            if ws.max_row is None or ws.max_row < 2:
                continue

            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue

            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header)]

            # 跳过全空表头
            if all(h.startswith("col_") for h in headers):
                continue

            for row in rows_iter:
                if all(cell is None for cell in row):
                    continue
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = str(cell) if cell is not None else ""
                all_rows.append(row_dict)

            sheets_processed += 1

        wb.close()
        return all_rows if all_rows else None

    def _parse_csv(self, content: bytes) -> list[dict]:
        """Parse CSV with encoding detection."""
        for encoding in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            raise HTTPException(
                status_code=400,
                detail="无法识别文件编码，请使用 UTF-8 或 GBK 编码",
            )

        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    def _parse_csv_streaming(self, content: bytes | BinaryIO, chunk_size: int):
        """Parse CSV in chunks to keep memory usage bounded for large files."""
        if hasattr(content, "read"):
            binary = content
            encodings = ("utf-8-sig", "gbk", "utf-8")
            for encoding in encodings:
                text_stream = None
                try:
                    if hasattr(binary, "seek"):
                        binary.seek(0)
                    text_stream = io.TextIOWrapper(binary, encoding=encoding, newline="")
                    reader = csv.DictReader(text_stream)
                    chunk: list[dict] = []
                    for row in reader:
                        chunk.append(row)
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                    if chunk:
                        yield chunk
                    return
                except UnicodeDecodeError:
                    continue
                finally:
                    if text_stream is not None:
                        try:
                            text_stream.detach()
                        except Exception:
                            pass

            raise HTTPException(
                status_code=400,
                detail="无法识别文件编码，请使用 UTF-8 或 GBK 编码",
            )
        else:
            raw = content

        for encoding in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            raise HTTPException(
                status_code=400,
                detail="无法识别文件编码，请使用 UTF-8 或 GBK 编码",
            )

        reader = csv.DictReader(io.StringIO(text))
        chunk: list[dict] = []
        for row in reader:
            chunk.append(row)
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        if chunk:
            yield chunk

    def _normalize_columns(
        self, rows: list[dict], column_map: dict[str, str]
    ) -> list[dict]:
        """Normalize column names using the provided mapping.

        支持模糊匹配：去掉方括号、圆括号、空格等特殊字符后再匹配。
        """
        import re
        normalized = []
        for row in rows:
            new_row = {}
            for key, value in row.items():
                if key is None:
                    continue
                key_str = str(key).strip()
                # 精确匹配
                mapped = column_map.get(key_str)
                if mapped is None:
                    # 清洗后匹配（去掉方括号等）
                    cleaned = re.sub(r'[\[\]【】\(\)（）\*\#\s]', '', key_str)
                    mapped = column_map.get(cleaned, key_str)
                new_row[mapped] = value
            normalized.append(new_row)
        return normalized

    def _parse_values(self, rows: list[dict], data_type: str) -> list[dict]:
        """Parse and convert values based on data type."""
        result = []
        for row in rows:
            parsed = self._parse_single_row(row, data_type)
            if parsed:
                result.append(parsed)
        return result

    def _parse_single_row(self, row: dict, data_type: str) -> dict | None:
        """Parse a single row based on data type."""
        account_code = str(row.get("account_code", "")).strip()
        if not account_code:
            return None

        base = {
            "account_code": account_code,
            "account_name": str(row.get("account_name", "")).strip() or None,
            "company_code": str(row.get("company_code", "")).strip() or "default",
            "currency_code": str(row.get("currency_code", "")).strip() or "CNY",
        }

        if data_type == "tb_balance":
            # 解析科目级次（必需字段）
            level_str = str(row.get("level", "")).strip()
            level = int(level_str) if level_str.isdigit() else None
            if level is None:
                # 尝试从编码推断级次（兜底）
                level = _infer_level_from_code(account_code)
            base.update({
                "level": level,
                "opening_balance": _parse_decimal(row.get("opening_balance")),
                "debit_amount": _parse_decimal(row.get("debit_amount")),
                "credit_amount": _parse_decimal(row.get("credit_amount")),
                "closing_balance": _parse_decimal(row.get("closing_balance")),
            })
        elif data_type == "tb_ledger":
            voucher_date = _parse_date(row.get("voucher_date"))
            voucher_no = str(row.get("voucher_no", "")).strip()
            if not voucher_date or not voucher_no:
                return None
            # 解析会计期间（支持 "2025年1期" 格式）
            ap_raw = row.get("accounting_period")
            accounting_period = _parse_period(ap_raw) if isinstance(ap_raw, str) and '期' in str(ap_raw) else None
            if accounting_period is None and ap_raw is not None:
                try:
                    accounting_period = int(ap_raw)
                except (ValueError, TypeError):
                    pass
            base.update({
                "voucher_date": voucher_date,
                "voucher_no": voucher_no,
                "accounting_period": accounting_period,
                "voucher_type": str(row.get("voucher_type", "")).strip() or None,
                "debit_amount": _parse_decimal(row.get("debit_amount")),
                "credit_amount": _parse_decimal(row.get("credit_amount")),
                "counterpart_account": str(row.get("counterpart_account", "")).strip() or None,
                "summary": str(row.get("summary", "")).strip() or None,
                "preparer": str(row.get("preparer", "")).strip() or None,
            })
        elif data_type == "tb_aux_balance":
            aux_type = str(row.get("aux_type", "")).strip()
            if not aux_type:
                return None
            base.update({
                "aux_type": aux_type,
                "aux_code": str(row.get("aux_code", "")).strip() or None,
                "aux_name": str(row.get("aux_name", "")).strip() or None,
                "opening_balance": _parse_decimal(row.get("opening_balance")),
                "debit_amount": _parse_decimal(row.get("debit_amount")),
                "credit_amount": _parse_decimal(row.get("credit_amount")),
                "closing_balance": _parse_decimal(row.get("closing_balance")),
            })
        elif data_type == "tb_aux_ledger":
            # 解析会计期间
            ap_raw = row.get("accounting_period")
            accounting_period = _parse_period(ap_raw) if isinstance(ap_raw, str) and '期' in str(ap_raw) else None
            if accounting_period is None and ap_raw is not None:
                try:
                    accounting_period = int(ap_raw)
                except (ValueError, TypeError):
                    pass
            base.update({
                "aux_type": str(row.get("aux_type", "")).strip() or None,
                "aux_code": str(row.get("aux_code", "")).strip() or None,
                "aux_name": str(row.get("aux_name", "")).strip() or None,
                "voucher_date": _parse_date(row.get("voucher_date")),
                "voucher_no": str(row.get("voucher_no", "")).strip() or None,
                "accounting_period": accounting_period,
                "voucher_type": str(row.get("voucher_type", "")).strip() or None,
                "debit_amount": _parse_decimal(row.get("debit_amount")),
                "credit_amount": _parse_decimal(row.get("credit_amount")),
                "summary": str(row.get("summary", "")).strip() or None,
                "preparer": str(row.get("preparer", "")).strip() or None,
            })

        return base


# ---------------------------------------------------------------------------
# Stub parsers for specific financial software
# ---------------------------------------------------------------------------


class YonyouParser(BaseParser):
    """用友U8/T+解析器 - 当前委托给 GenericParser。

    Validates: Requirements 4.1
    """

    def __init__(self) -> None:
        self._generic = GenericParser()

    def parse(self, content: bytes, data_type: str) -> list[dict]:
        return self._generic.parse(content, data_type)


class KingdeeParser(BaseParser):
    """金蝶K3/KIS解析器 - 当前委托给 GenericParser。

    Validates: Requirements 4.1
    """

    def __init__(self) -> None:
        self._generic = GenericParser()

    def parse(self, content: bytes, data_type: str) -> list[dict]:
        return self._generic.parse(content, data_type)


class SAPParser(BaseParser):
    """SAP解析器 - 当前委托给 GenericParser。

    Validates: Requirements 4.1
    """

    def __init__(self) -> None:
        self._generic = GenericParser()

    def parse(self, content: bytes, data_type: str) -> list[dict]:
        return self._generic.parse(content, data_type)


# ---------------------------------------------------------------------------
# ParserFactory
# ---------------------------------------------------------------------------


class ParserFactory:
    """解析器工厂 - 根据数据源类型返回对应解析器。

    Validates: Requirements 4.1
    """

    _parsers: dict[str, type[BaseParser]] = {
        "generic": GenericParser,
        "yonyou": YonyouParser,
        "kingdee": KingdeeParser,
        "sap": SAPParser,
    }

    @classmethod
    def get_parser(cls, source_type: str) -> BaseParser:
        """Get parser instance for the given source type."""
        parser_cls = cls._parsers.get(source_type)
        if not parser_cls:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的数据源类型: {source_type}，支持: {', '.join(cls._parsers.keys())}",
            )
        return parser_cls()
