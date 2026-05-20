"""底稿模板初始化服务

生成底稿时从 wp_templates/ 目录复制模板 xlsx 文件到项目底稿存储目录。
支持 prefill 预填充叠加。

流程：
1. 根据 wp_code 查找模板文件
2. 复制到项目底稿存储目录（backend/{project_id}/{wp_id}.xlsx）
3. 可选：执行 prefill 写入 =TB/=WP 计算值
4. 返回底稿文件路径供 Univer 前端加载
"""
from __future__ import annotations

import json
import logging
import re
import shutil
from pathlib import Path
from typing import Optional
from uuid import UUID

logger = logging.getLogger(__name__)

BACKEND_DIR = Path(__file__).resolve().parent.parent.parent
TEMPLATES_DIR = BACKEND_DIR / "wp_templates"
INDEX_FILE = TEMPLATES_DIR / "_index.json"
STORAGE_DIR = BACKEND_DIR / "wp_storage"

# 缓存模板索引
_index_cache: list[dict] | None = None


# ---------------------------------------------------------------------------
# F2 / F3 多文件 sheet 合并去重工具（spec workpaper-d-sales-cycle, ADR D2 + D3）
# ---------------------------------------------------------------------------


def _normalize_sheet_name(name: str) -> str:
    """归一化 sheet 名（spec D2 ADR）。

    规则（按顺序）:
      1. 中文圆括号 `（…）` → 英文 `(…)`，并 strip 首尾空白
      2. 含 `GT_Custom` → 归一为 `"GT_Custom"`（多文件 GT 内部 sheet 视为同名）
      3. 含 `底稿目录` → 归一为 `"底稿目录"`（多文件底稿目录视为同名）
      4. 其他：剔除全部空白字符（含中间空格、tab、全角空格）后返回

    幂等：normalize(normalize(x)) == normalize(x)（PBT P2 验证）
    """
    if name is None:
        return ""
    n = str(name).replace("（", "(").replace("）", ")").strip()
    if "GT_Custom" in n:
        return "GT_Custom"
    if "底稿目录" in n:
        return "底稿目录"
    # 剔除内部所有空白（中英文、全角、tab、换行）以避免空格差异导致漏去重
    return re.sub(r"\s+", "", n)


def _should_skip_historical_sheet(name: str) -> bool:
    """判断是否为历史遗留 sheet（spec D3 ADR + F-F2 ADR-F3 + J-F1 — 应在归一化前过滤）。

    匹配规则（任一命中即跳过）:
      - 含 "修订前"（如 "主营业务收入审计程序表 D4A（修订前）"、"G1A-修订前"）
      - 含 "（原）" 或 "(原)"（如 "D7A（原）"、"D8A(原)"）
      - 含 G+数字 编号且含 "删除" 或 "移至"（F 循环新模式：F2-38/F2-47/F2-52
        含 "存货计价测试程序G2-8-删除"、"G2-8-4-移至分析类" 等历史遗留 sheet）
      - 以 "-删除" 结尾（J 循环通用模式：如 "股份支付检查表J1-10-删除"、
        "IPO企业股权激励工具关注的审计重点-删除"、"首发业务解答二-删除"）
      - 含 "（示例）"/"(示例)"，或以 "示例"/"示例）"/"示例)" 结尾（F 循环示例 sheet：
        "函证差异检查表（示例）"、"合同履约成本测试（示例）"、"访谈记录与核对示例"）
    """
    if name is None:
        return False
    s = str(name)
    if "修订前" in s or "（原）" in s or "(原)" in s:
        return True
    # F-F2 ADR-F3: G+数字编号 + 删除/移至（注意 G 不限于开头位置）
    if re.search(r"G\d+", s) and ("删除" in s or "移至" in s):
        return True
    # J-F1: 以 "-删除" 结尾的通用历史遗留模式（J 循环 5 个 sheet）
    if s.endswith("-删除"):
        return True
    # F-F2 ADR-F3: 示例 sheet（括号包裹或末尾）
    if "（示例）" in s or "(示例)" in s:
        return True
    if s.endswith("示例") or s.endswith("示例）") or s.endswith("示例)"):
        return True
    return False


# ---------------------------------------------------------------------------
# F4 scenario 文件级裁剪规则（spec workpaper-d-sales-cycle ADR D4 / 任务 2.1）
#
# 5 档 scenario：normal 排除 IPO/上市/新三板/重组/舞弊应对 关键字命中的文件；
# 其他档（ipo/listed/transfer/restructure/fraud_response）加载全部 17 文件。
# 字典 + 帮助函数定义在此（实际文件加载位置），从 chain_orchestrator 重新导出
# 以满足 spec 任务 2.1 文本"在 chain_orchestrator.py 加 SCENARIO_TO_FILE_FILTER"
# 的模块位置约定（避开循环导入，与 _normalize_sheet_name 复用同一模式）。
# ---------------------------------------------------------------------------

SCENARIO_TO_FILE_FILTER: dict[str, dict[str, list[str]]] = {
    "normal": {
        "exclude_patterns": [
            "IPO",
            "上市",
            "新三板",
            "重组",
            "舞弊应对",
            "舞弊",
            "反舞弊",
        ],
    },
    "ipo": {},            # 加载全部
    "listed": {},         # 加载全部
    "transfer": {},       # 新三板挂牌：加载全部
    "restructure": {},    # 加载全部
    "fraud_response": {}, # 加载全部
}


def _filter_files_by_scenario(file_paths: list[Path], scenario: str) -> list[Path]:
    """按 scenario 过滤模板文件（spec F4 ADR D4 / 任务 2.1）。

    规则：当 scenario='normal' 时，排除文件名含任一 exclude_patterns 关键字
    的文件；其他 scenario 不过滤（加载全部）。未知 scenario 退化为 normal
    以保守为主（避免误加载 IPO 应对类文件给普通项目）。

    Args:
        file_paths: 模板文件路径列表
        scenario: 项目场景（normal/ipo/listed/transfer/restructure/fraud_response）

    Returns:
        过滤后的文件路径列表（保持原顺序）
    """
    rules = SCENARIO_TO_FILE_FILTER.get(scenario)
    if rules is None:
        # 未知 scenario 退化为 normal（保守过滤）
        rules = SCENARIO_TO_FILE_FILTER.get("normal", {})

    exclude_patterns = rules.get("exclude_patterns") or []
    if not exclude_patterns:
        return list(file_paths)

    return [
        p for p in file_paths
        if not any(kw in p.name for kw in exclude_patterns)
    ]


# ---------------------------------------------------------------------------
# H-F2 计量模式 sheet 显隐控制（spec workpaper-h-fixed-assets-cycle ADR-H2）
#
# H3 投资性房地产 + H7 生产性生物资产支持「成本模式」或「公允价值模式」。
# 独立于 SCENARIO_TO_FILE_FILTER（两者并行运行）。
# 前端 useHFixedAssetSheetGroups 读取 project.measurement_model 后
# 对命中 hide_patterns 的 sheet 设 visible=false。
# ---------------------------------------------------------------------------

MEASUREMENT_MODEL_FILTER: dict[str, dict[str, list[str]]] = {
    "cost": {
        "hide_patterns": ["（公允价值模式）", "(公允价值模式)"],
    },
    "fair_value": {
        "hide_patterns": ["（成本模式）", "(成本模式)"],
    },
}


# ---------------------------------------------------------------------------
# F4 / 任务 2.2 — B51-5 高风险触发 D4-22~D4-32 IPO 应对类底稿强制加载
# spec workpaper-d-sales-cycle ADR D4
# ---------------------------------------------------------------------------

# D4-22 至 D4-32 IPO 应对子表（含 D4-22A 总控台）
D4_IPO_CODES: list[str] = [
    "D4-22",
    "D4-22A",
    "D4-23",
    "D4-24",
    "D4-25",
    "D4-26",
    "D4-27",
    "D4-28",
    "D4-29",
    "D4-30",
    "D4-31",
    "D4-32",
]

# F2-61 至 F2-72 IPO 应对子表（存货及跌价准备 IPO/上市/新三板/重组/舞弊应对）
F2_IPO_CODES: list[str] = [
    "F2-61",
    "F2-62",
    "F2-63",
    "F2-64",
    "F2-65",
    "F2-66",
    "F2-67",
    "F2-68",
    "F2-69",
    "F2-70",
    "F2-71",
    "F2-72",
]

# IPO 配置注册表：prefix → (code 列表, audit_cycle, default_name_prefix)
_IPO_CONFIG: dict[str, dict] = {
    "D4": {
        "codes": D4_IPO_CODES,
        "audit_cycle": "D",
        "name_prefix": "营业收入 IPO 应对底稿",
    },
    "F2": {
        "codes": F2_IPO_CODES,
        "audit_cycle": "F",
        "name_prefix": "存货及跌价准备 IPO 应对底稿",
    },
    "H1": {
        "codes": [],
        "audit_cycle": "H",
        "name_prefix": "固定资产评估增值 IPO 应对底稿",
    },
    "J1": {
        "codes": [],
        "audit_cycle": "J",
        "name_prefix": "职工薪酬 IPO 应对底稿",
    },
    "K8": {
        "codes": [],
        "audit_cycle": "K",
        "name_prefix": "销售费用 IPO 应对底稿",
    },
    "L1": {
        "codes": [],
        "audit_cycle": "L",
        "name_prefix": "筹资循环 IPO 应对底稿",
    },
    "M2": {
        "codes": [],
        "audit_cycle": "M",
        "name_prefix": "权益循环 IPO 应对底稿",
    },
    "N2": {
        "codes": [],
        "audit_cycle": "N",
        "name_prefix": "应交税费 IPO 应对底稿",
    },
}


async def _ensure_ipo_loaded(
    db,  # AsyncSession (避免顶层依赖触发循环导入)
    project_id: UUID,
    year: int,
    wp_code_prefix: str = "D4",
) -> dict:
    """通用 IPO 应对底稿加载函数（D4 / F2 等多前缀参数化）。

    spec workpaper-f-purchase-inventory F-F14 / ADR-F4。

    支持的 prefix（见 ``_IPO_CONFIG``）:
      - "D4": D 销售循环 D4-22~D4-32（B51-5 高风险触发）
      - "F2": F 采购存货循环 F2-61~F2-72（B51-4 高风险触发）
      - "H1": H 固定资产循环（占位，codes=[]，暂无事件触发）

    实现策略（最小破坏式）:
      1. 对每个 IPO wp_code：
         - 查 ``wp_index`` 是否已存在 → 已存在则计入 ``skipped_existing``
         - 否则创建 ``WpIndex`` + ``WorkingPaper`` 记录，并强制以
           ``scenario='ipo'`` 调 ``init_workpaper_from_template`` 复制模板
           （绕过 ``scenario='normal'`` 的 IPO 文件过滤规则）
      2. 任一 wp_code 创建失败时回滚到该 wp_code 子事务，记录到 errors，继续下一个

    幂等：已存在的 wp_code 不会重复创建。

    Args:
        db: 异步 SQLAlchemy 会话（``AsyncSession``）
        project_id: 项目 UUID
        year: 财政年度（保留参数，未来按年度区分时使用）
        wp_code_prefix: IPO 配置 key，必须在 ``_IPO_CONFIG`` 中

    Returns:
        dict 含 ``added_codes`` / ``skipped_existing`` / ``errors`` 三个列表，
        以及 ``prefix`` 标识本次加载的循环。
    """
    import sqlalchemy as sa
    from app.models.workpaper_models import (
        WorkingPaper,
        WpIndex,
        WpSourceType,
        WpStatus,
    )

    cfg = _IPO_CONFIG.get(wp_code_prefix.upper())
    if not cfg:
        return {
            "prefix": wp_code_prefix,
            "added_codes": [],
            "skipped_existing": [],
            "errors": [{"code": "*", "error": f"unsupported prefix: {wp_code_prefix}"}],
        }

    codes: list[str] = cfg["codes"]
    audit_cycle: str = cfg["audit_cycle"]
    name_prefix: str = cfg["name_prefix"]

    result: dict[str, list | str] = {
        "prefix": wp_code_prefix,
        "added_codes": [],
        "skipped_existing": [],
        "errors": [],
    }

    # year 参数当前未做按年度筛选（wp_index 不带 year 字段），保留供未来扩展
    _ = year

    for code in codes:
        try:
            # 幂等检查：wp_index 已存在则跳过
            existing = await db.execute(
                sa.select(WpIndex.id).where(
                    WpIndex.project_id == project_id,
                    WpIndex.wp_code == code,
                )
            )
            if existing.scalar_one_or_none():
                result["skipped_existing"].append(code)  # type: ignore[union-attr]
                continue

            # 创建 WpIndex
            wp_index = WpIndex(
                project_id=project_id,
                wp_code=code,
                wp_name=f"{name_prefix} {code}",
                audit_cycle=audit_cycle,
                status=WpStatus.not_started,
            )
            db.add(wp_index)
            await db.flush()

            # 创建 WorkingPaper（file_path 占位，复制模板成功后会被覆盖）
            wp = WorkingPaper(
                wp_index_id=wp_index.id,
                project_id=project_id,
                source_type=WpSourceType.template,
                file_path=f"storage/projects/{project_id}/workpapers/{code}.xlsx",
                parsed_data={},
            )
            db.add(wp)
            await db.flush()

            # 强制 scenario='ipo' 让模板复制绕过 normal 过滤规则
            try:
                actual_path = init_workpaper_from_template(
                    project_id=project_id,
                    wp_id=wp.id,
                    wp_code=code,
                    scenario="ipo",
                    has_foreign_currency=False,
                )
                if actual_path:
                    wp.file_path = str(actual_path)
            except Exception as ie:
                # 模板复制失败不影响数据库记录（占位 file_path 已写入）
                logger.warning(
                    "_ensure_ipo_loaded[%s]: 模板复制失败 code=%s err=%s",
                    wp_code_prefix, code, ie,
                )

            result["added_codes"].append(code)  # type: ignore[union-attr]
        except Exception as e:
            result["errors"].append({"code": code, "error": str(e)})  # type: ignore[union-attr]
            try:
                await db.rollback()
            except Exception:
                pass

    return result


async def _ensure_d4_ipo_loaded(
    db,  # AsyncSession (避免顶层依赖触发循环导入)
    project_id: UUID,
    year: int,
) -> dict:
    """B51-5 高风险触发：确保 D4-22~D4-32 IPO 应对类底稿已加载到当前项目。

    spec workpaper-d-sales-cycle 任务 2.2 / ADR D4。

    向后兼容包装：转发到通用 ``_ensure_ipo_loaded(prefix='D4')``。
    保留独立函数名以避免破坏 D spec 既有调用点；
    返回值剥离 ``prefix`` 字段以保持原 schema（仅 ``added_codes`` /
    ``skipped_existing`` / ``errors`` 三键）。
    """
    raw = await _ensure_ipo_loaded(db, project_id, year, wp_code_prefix="D4")
    return {
        "added_codes": raw.get("added_codes", []),
        "skipped_existing": raw.get("skipped_existing", []),
        "errors": raw.get("errors", []),
    }


def _merge_sheets_dedup(
    target_path: Path,
    other_files: list[Path],
) -> dict:
    """合并多个 xlsx 文件的 sheet 到 target workbook，按"先过滤历史 → 再归一化去重"流程。

    spec workpaper-d-sales-cycle F2 + F3 主入口（ADR D2 + D3 + D14 双总控台过滤）。

    流程（每条 source sheet）:
      1. 若 `_should_skip_historical_sheet(name)` → 跳过（历史遗留过滤）
      2. 计算 `_normalize_sheet_name(name)`，若已存在归一化名 → 跳过（去重）
      3. 否则按原始 sheet_name（截断到 31 字符以满足 Excel 限制）创建新 sheet 并复制值

    Args:
        target_path: 主 workbook 路径（已存在）
        other_files: 待合并的 xlsx 路径列表

    Returns:
        {"merged": int, "skipped_dup": int, "skipped_historical": int} 统计信息
    """
    from openpyxl import load_workbook

    stats = {"merged": 0, "skipped_dup": 0, "skipped_historical": 0}

    target_wb = load_workbook(str(target_path))
    # 用归一化名集合去重（含 target 自身已有 sheet）
    existing_normalized: set[str] = {
        _normalize_sheet_name(s) for s in target_wb.sheetnames
    }

    try:
        for other_file in other_files:
            try:
                src_wb = load_workbook(str(other_file), read_only=True, data_only=True)
                try:
                    for sheet_name in src_wb.sheetnames:
                        # 1. 历史遗留 sheet 过滤（修订前 / （原））
                        if _should_skip_historical_sheet(sheet_name):
                            stats["skipped_historical"] += 1
                            logger.debug(
                                "merge_sheets_dedup: 跳过历史遗留 sheet '%s' from %s",
                                sheet_name, other_file.name,
                            )
                            continue

                        # 2. 归一化去重（覆盖 GT_Custom / 底稿目录 / 中英文圆括号变体）
                        normalized = _normalize_sheet_name(sheet_name)
                        if normalized in existing_normalized:
                            stats["skipped_dup"] += 1
                            logger.debug(
                                "merge_sheets_dedup: 跳过重复 sheet '%s' (normalized='%s') from %s",
                                sheet_name, normalized, other_file.name,
                            )
                            continue

                        # 3. 创建新 sheet（Excel sheet 名最长 31 字符）
                        new_ws = target_wb.create_sheet(title=sheet_name[:31])
                        existing_normalized.add(normalized)
                        src_ws = src_wb[sheet_name]
                        for row in src_ws.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    new_ws.cell(
                                        row=cell.row,
                                        column=cell.column,
                                        value=cell.value,
                                    )
                        stats["merged"] += 1
                finally:
                    src_wb.close()
            except Exception as e:
                logger.debug(
                    "merge_sheets_dedup: 合并失败 %s (%s)",
                    other_file.name, e,
                )

        target_wb.save(str(target_path))
    finally:
        target_wb.close()

    return stats


def _load_index() -> list[dict]:
    """加载模板索引（带缓存）"""
    global _index_cache
    if _index_cache is not None:
        return _index_cache
    if not INDEX_FILE.exists():
        logger.warning("模板索引文件不存在: %s", INDEX_FILE)
        return []
    with open(INDEX_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    _index_cache = data.get("files", [])
    return _index_cache


def find_template_file(wp_code: str) -> Path | None:
    """根据 wp_code 查找主模板文件

    优先匹配：含"审定表"或"常规程序"的文件 > 文件名最短的。
    对于多文件底稿（如 D2 有 D2-1至D2-4），返回审定表文件。
    对于子表（如 D2-2/E1-3），如果独立文件不存在，回退到主表文件
    （主表 xlsx 通常包含所有子表 sheet，如 "D2-1至D2-4 应收账款-审定表明细表"）。
    """
    index = _load_index()
    # 精确匹配
    candidates = [
        e for e in index
        if e["wp_code"] == wp_code and e["format"] in ("xlsx", "xlsm")
    ]
    if candidates:
        # 优先选择含"审定表"或"常规程序"的文件
        for c in candidates:
            if "审定表" in c["filename"] or "常规程序" in c["filename"]:
                full_path = TEMPLATES_DIR / c["relative_path"]
                if full_path.exists():
                    return full_path
        # 其次选择文件名最短的
        candidates.sort(key=lambda e: len(e["filename"]))
        rel_path = candidates[0]["relative_path"]
        full_path = TEMPLATES_DIR / rel_path
        if full_path.exists():
            return full_path

    # 模糊匹配：文件名以 wp_code 开头
    prefix = wp_code[0]
    template_subdir = TEMPLATES_DIR / prefix
    if template_subdir.exists():
        # 优先含"审定表"
        for f in sorted(template_subdir.iterdir()):
            if f.name.startswith(wp_code) and f.suffix.lower() in (".xlsx", ".xlsm"):
                if "审定表" in f.name or "常规程序" in f.name:
                    return f
        # 其次最短文件名
        for f in sorted(template_subdir.iterdir(), key=lambda x: len(x.name)):
            if f.name.startswith(wp_code) and f.suffix.lower() in (".xlsx", ".xlsm"):
                return f

        # 子表回退：如 D2-2 找不到，尝试包含范围式命名的文件（D2-1至D2-4）
        if "-" in wp_code:
            primary = wp_code.split("-")[0]
            sub_num = wp_code.split("-")[1]
            for f in sorted(template_subdir.iterdir()):
                if f.suffix.lower() not in (".xlsx", ".xlsm"):
                    continue
                # 匹配模式: 文件名含 "{primary}-N至{primary}-M" 且 sub_num 在范围内
                if f.name.startswith(primary + "-") and "至" in f.name:
                    # 简单包含检查
                    return f
            # 终极回退：用主表
            for f in sorted(template_subdir.iterdir()):
                if f.name.startswith(primary + " ") and f.suffix.lower() in (".xlsx", ".xlsm"):
                    return f

    return None


def find_all_template_files(wp_code: str) -> list[Path]:
    """查找 wp_code 对应的所有模板文件（多文件底稿）"""
    index = _load_index()
    candidates = [
        e for e in index
        if e["wp_code"] == wp_code and e["format"] in ("xlsx", "xlsm", "docx")
    ]
    results = []
    for c in candidates:
        full_path = TEMPLATES_DIR / c["relative_path"]
        if full_path.exists():
            results.append(full_path)
    return results


def find_template_file_any(wp_code: str) -> Path | None:
    """Find template file of any format (xlsx/xlsm/docx/doc)

    P2-2: 扩展模板查找支持 docx/doc 格式。
    优先返回 xlsx/xlsm，其次 docx/doc。
    """
    # 先尝试 xlsx/xlsm
    result = find_template_file(wp_code)
    if result:
        return result
    # 再尝试 docx/doc
    index = _load_index()
    candidates = [
        e for e in index
        if e["wp_code"] == wp_code and e["format"] in ("docx", "doc")
    ]
    if candidates:
        rel_path = candidates[0]["relative_path"]
        full_path = TEMPLATES_DIR / rel_path
        if full_path.exists():
            return full_path
    return None


def get_workpaper_storage_path(project_id: UUID, wp_id: UUID) -> Path:
    """获取底稿文件存储路径"""
    storage = STORAGE_DIR / str(project_id)
    storage.mkdir(parents=True, exist_ok=True)
    return storage / f"{wp_id}.xlsx"


def init_workpaper_from_template(
    project_id: UUID,
    wp_id: UUID,
    wp_code: str,
    merge_all_files: bool = True,
    scenario: str = "normal",
    has_foreign_currency: bool = False,
) -> Optional[Path]:
    """从模板初始化底稿文件

    复制模板 xlsx 到项目底稿存储目录。
    P1-2: 多文件底稿支持——如果 wp_code 对应多个 xlsx 文件，合并所有 sheet 到一个 workbook。
    P2-1: xlsm 文件只复制不 prefill（避免 openpyxl 损坏 VBA 宏）
    P2-2: docx/doc 文件直接复制（不做 xlsx 转换）

    E1 Sprint 2 Task 2.2: scenario 文件级裁剪（F1.2）
    - normal: 仅加载 F1+F2+F3
    - 非 normal (ipo/listed/transfer/restructure/fraud_response): 加载 F1+F2+F3+F4

    Args:
        project_id: 项目 ID
        wp_id: 底稿 ID
        wp_code: 底稿编码（如 D2）
        merge_all_files: 是否合并多文件底稿的所有 sheet
        scenario: 项目场景（normal/ipo/listed/transfer/restructure/fraud_response）
        has_foreign_currency: 是否有外币业务（保留参数，未来用于 sheet 级裁剪）

    Returns:
        底稿文件路径，如果模板不存在返回 None
    """
    # P2-2: 尝试查找任意格式模板（xlsx/xlsm/docx/doc）
    template_path = find_template_file_any(wp_code)
    if not template_path:
        logger.warning("模板文件不存在: wp_code=%s", wp_code)
        return None

    # P2-2: docx/doc 模板直接复制，不做 xlsx 转换
    if template_path.suffix.lower() in ('.docx', '.doc'):
        target_path = get_workpaper_storage_path(project_id, wp_id).with_suffix(template_path.suffix)
        shutil.copy2(template_path, target_path)
        logger.info(
            "底稿从 Word 模板初始化: wp_code=%s, template=%s, target=%s",
            wp_code, template_path.name, target_path,
        )
        return target_path

    # P2-1: xlsm 文件只复制不 prefill（openpyxl 会损坏 VBA 宏）
    if template_path.suffix.lower() == '.xlsm':
        target_path = get_workpaper_storage_path(project_id, wp_id).with_suffix('.xlsm')
        shutil.copy2(template_path, target_path)
        logger.info(
            "Skipping prefill for xlsm template: %s (VBA macros preserved)",
            wp_code,
        )
        return target_path

    # xlsx: 复制主文件
    target_path = get_workpaper_storage_path(project_id, wp_id)
    shutil.copy2(template_path, target_path)

    # P1-2: 多文件底稿合并——将其他文件的 sheet 追加到主 workbook
    if merge_all_files:
        all_files = find_all_template_files(wp_code)
        other_files = [f for f in all_files if f != template_path and f.suffix.lower() in ('.xlsx',)]
        # spec workpaper-d-sales-cycle 任务 2.1（F4 ADR D4）：
        # 通用 scenario 文件级裁剪——替换 E1 spec 的硬编码"E1 + normal"过滤。
        # 现行规则对所有 wp_code 一致：scenario='normal' 时排除 IPO/上市/新三板/
        # 重组/舞弊应对 关键字命中的文件；其他 scenario 加载全部。
        other_files = _filter_files_by_scenario(other_files, scenario)
        if other_files:
            try:
                _merge_sheets_from_other_files(target_path, other_files)
                logger.info(
                    "多文件底稿合并: wp_code=%s, scenario=%s, merged %d additional files",
                    wp_code, scenario, len(other_files),
                )
            except Exception as e:
                logger.warning("多文件合并失败（非阻断）: wp_code=%s, error=%s", wp_code, e)

    logger.info(
        "底稿从模板初始化: wp_code=%s, template=%s, target=%s",
        wp_code, template_path.name, target_path,
    )
    return target_path


def _merge_sheets_from_other_files(target_path: Path, other_files: list[Path]) -> None:
    """将其他 xlsx 文件的 sheet 追加到目标 workbook（P1-2 多文件底稿支持）。

    spec workpaper-d-sales-cycle F2/F3 重构：委派给 ``_merge_sheets_dedup``，
    新增中英文圆括号归一化 + 修订前/（原）历史 sheet 过滤。
    返回值仍保持 None（向后兼容现有 callers）；统计信息通过 logger.info 输出。
    """
    stats = _merge_sheets_dedup(target_path, other_files)
    logger.info(
        "merge_sheets_dedup stats target=%s merged=%d skipped_dup=%d skipped_historical=%d",
        target_path.name,
        stats["merged"],
        stats["skipped_dup"],
        stats["skipped_historical"],
    )


def prefill_workpaper_xlsx(
    target_path: Path,
    wp_code: str,
    tb_data: dict[str, dict[str, float]],
    adj_data: dict[str, dict[str, float]] | None = None,
    user_formulas: dict[str, dict] | None = None,
) -> int:
    """在已复制的底稿 xlsx 中写入预填充数据

    根据 prefill_formula_mapping.json 的配置，用 openpyxl 打开 xlsx，
    找到对应 sheet 和语义行，写入试算表/调整分录的值。

    E1 spec Sprint 1 Task 1.20 执行优先级:
        user_formulas(覆盖)> prefill_formula_mapping(预设)> 模板内置公式
    具体实现: 当 cell_key (sheet!cell_ref) 在 user_formulas 中时,跳过预设写入,
    交由 xlsx 自身的公式引擎在前端展示阶段计算用户公式。

    Args:
        target_path: 底稿 xlsx 文件路径
        wp_code: 底稿编码
        tb_data: 试算表数据 {科目编码: {列名: 值}}
        adj_data: 调整分录数据 {科目编码: {aje_net: 值, rje_net: 值}}
        user_formulas: 用户自定义公式 {sheet!cell_ref: {formula, formula_type, ...}}

    Returns:
        写入的单元格数量
    """
    from openpyxl import load_workbook

    # 加载公式映射
    mapping_file = BACKEND_DIR / "data" / "prefill_formula_mapping.json"
    if not mapping_file.exists():
        logger.warning("prefill_formula_mapping.json 不存在")
        return 0

    with open(mapping_file, "r", encoding="utf-8") as f:
        mapping_data = json.load(f)

    # 找到当前 wp_code 的映射
    wp_mapping = next(
        (m for m in mapping_data.get("mappings", []) if m["wp_code"] == wp_code),
        None,
    )
    if not wp_mapping:
        logger.debug("wp_code=%s 无公式映射配置", wp_code)
        return 0

    try:
        wb = load_workbook(str(target_path))
    except Exception as e:
        logger.error("无法打开底稿 xlsx: %s (%s)", target_path, e)
        return 0

    sheet_name = wp_mapping.get("sheet", "")
    ws = None
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif wb.sheetnames:
        # 尝试找包含"审定表"的 sheet
        for sn in wb.sheetnames:
            if "审定表" in sn:
                ws = wb[sn]
                break
        if not ws:
            ws = wb.active

    if not ws:
        wb.close()
        return 0

    filled_count = 0
    account_codes = wp_mapping.get("account_codes", [])
    primary_code = account_codes[0] if account_codes else ""

    # E1 spec Task 1.20: user_formulas 优先级 — 收集本 sheet 内被用户覆盖的 cell_ref
    user_formulas = user_formulas or {}
    user_overridden_cells: set[str] = set()
    if user_formulas:
        sheet_prefix = f"{ws.title}!"
        for key in user_formulas:
            if key.startswith(sheet_prefix):
                cell_ref_only = key[len(sheet_prefix):]
                user_overridden_cells.add(cell_ref_only)
        if user_overridden_cells:
            # 在 sheet 上写入用户公式(让 xlsx 自带公式引擎计算)
            for key, ufmt in user_formulas.items():
                if not key.startswith(sheet_prefix):
                    continue
                cell_ref_only = key[len(sheet_prefix):]
                if not _is_cell_coordinate(cell_ref_only):
                    continue
                user_formula_text = ufmt.get("formula") if isinstance(ufmt, dict) else str(ufmt)
                if not user_formula_text:
                    continue
                try:
                    target = ws[cell_ref_only]
                    target.value = user_formula_text
                    _mark_user_formula_cell(ws, target)
                    filled_count += 1
                except Exception as e:
                    logger.debug("user_formula 写入失败 %s: %s", key, e)

    for cell_def in wp_mapping.get("cells", []):
        formula_type = cell_def.get("formula_type", "")
        cell_ref = cell_def.get("cell_ref", "")
        # E1 spec Task 1.20: 用户自定义已覆盖此 cell,跳过预设写入
        if cell_ref in user_overridden_cells:
            continue
        # 占位 cell_ref(以下划线开头)是"汇总驱动 cell",不写入 Univer
        if cell_ref.startswith("_"):
            continue
        value = None

        if formula_type == "TB" and primary_code:
            # =TB('code','列名') → 从 tb_data 取值
            col_name = _extract_tb_column(cell_def.get("formula", ""))
            value = tb_data.get(primary_code, {}).get(col_name)
        elif formula_type == "TB_SUM" and account_codes:
            # =TB_SUM('start~end','列名') → 汇总多科目
            col_name = _extract_tb_column(cell_def.get("formula", ""))
            value = sum(
                tb_data.get(code, {}).get(col_name, 0)
                for code in account_codes
            )
        elif formula_type == "ADJ" and primary_code and adj_data:
            # =ADJ('code','type') → 从 adj_data 取值
            adj_type = _extract_adj_type(cell_def.get("formula", ""))
            value = adj_data.get(primary_code, {}).get(adj_type, 0)

        if value is not None:
            # P0-1 双策略：先尝试坐标写入，再降级语义匹配
            written = False

    # 策略 1：如果 cell_ref 是实际坐标（如 "E8"），直接写入
            if _is_cell_coordinate(cell_ref):
                try:
                    target_cell = ws[cell_ref]
                    # E1 spec Task 1.1: 不覆盖已有公式 cell
                    if _is_formula_cell(target_cell):
                        logger.debug(
                            "prefill 跳过公式 cell: wp_code=%s sheet=%s ref=%s formula=%s",
                            wp_code, ws.title, cell_ref, target_cell.value,
                        )
                    else:
                        ws[cell_ref] = value
                        _mark_prefilled_cell(ws, ws[cell_ref])
                        filled_count += 1
                    written = True
                except Exception:
                    pass

            # 策略 2：双维度定位（列头+数据行）— 致同模板标准结构
            if not written:
                col_idx = _find_column_by_keyword(ws, cell_ref)
                if col_idx:
                    # 找到列后，写入第一个数据行（列头下方第一个空行或合计行上方）
                    data_row = _find_first_data_row(ws, col_idx)
                    if data_row:
                        cell_obj = ws.cell(row=data_row, column=col_idx)
                        # E1 spec Task 1.1: 不覆盖已有公式 cell
                        if _is_formula_cell(cell_obj):
                            logger.debug(
                                "prefill 跳过公式 cell: wp_code=%s sheet=%s row=%d col=%d formula=%s",
                                wp_code, ws.title, data_row, col_idx, cell_obj.value,
                            )
                        else:
                            cell_obj.value = value
                            _mark_prefilled_cell(ws, cell_obj)
                            filled_count += 1
                        written = True

            # 策略 3：降级到语义行匹配
            if not written:
                row_num = _find_semantic_row(ws, cell_ref, wp_code, cell_ref)
                if row_num:
                    data_col = _find_data_column(ws, row_num)
                    cell_obj = ws.cell(row=row_num, column=data_col)
                    # E1 spec Task 1.1: 不覆盖已有公式 cell
                    if _is_formula_cell(cell_obj):
                        logger.debug(
                            "prefill 跳过公式 cell: wp_code=%s sheet=%s row=%d col=%d formula=%s",
                            wp_code, ws.title, row_num, data_col, cell_obj.value,
                        )
                    else:
                        cell_obj.value = value
                        _mark_prefilled_cell(ws, cell_obj)
                        filled_count += 1

    try:
        wb.save(str(target_path))
    except Exception as e:
        logger.error("保存底稿 xlsx 失败: %s (%s)", target_path, e)

    wb.close()
    logger.info("prefill 写入: wp_code=%s, filled=%d cells", wp_code, filled_count)
    return filled_count


def _extract_tb_column(formula: str) -> str:
    """从 =TB('1122','期末余额') 提取列名"""
    import re
    m = re.search(r"'([^']+)'\s*\)$", formula)
    return m.group(1) if m else "期末余额"


def _extract_adj_type(formula: str) -> str:
    """从 =ADJ('1122','aje_net') 提取类型"""
    import re
    m = re.search(r"'(aje_net|rje_net)'", formula)
    return m.group(1) if m else "aje_net"


def _is_cell_coordinate(ref: str) -> bool:
    """判断 cell_ref 是否为实际单元格坐标（如 E8, AA12）"""
    import re
    return bool(re.match(r'^[A-Z]{1,3}\d{1,5}$', ref, re.IGNORECASE))


def _is_formula_cell(cell) -> bool:
    """E1 spec Task 1.1: 判断单元格是否含 Excel 公式

    用于 prefill 写入前检查 — 凡 cell.value 以 "=" 开头的均视为公式 cell,
    prefill 必须跳过避免覆盖（如 E1-2!B22 = SUM(B15:B21) 合计公式）。
    """
    val = getattr(cell, "value", None)
    if val is None:
        return False
    if not isinstance(val, str):
        return False
    return val.lstrip().startswith("=")


def _find_column_by_keyword(ws, keyword: str) -> int | None:
    """双维度策略：在前 10 行中查找包含关键词的列号

    致同模板标准结构：第 5-6 行是列头（期初数/未审数/账项调整/重分类调整/审定数）。
    找到关键词所在列后，数据写入该列的数据行。
    """
    col_keywords = {
        "期初余额": ["期初", "年初", "期初数", "年初数"],
        "未审数": ["未审", "未审数", "期末数"],
        "AJE调整": ["AJE", "账项调整", "审计调整"],
        "RJE调整": ["RJE", "重分类调整", "重分类"],
        "上年审定数": ["上年", "审定数"],
        # 子科目分项：在 A 列找科目名称行
        "库存现金_期初": ["库存现金", "现金"],
        "银行存款_期初": ["银行存款"],
        "其他货币资金_期初": ["其他货币资金"],
        "原材料_期初": ["原材料"],
        "在产品_期初": ["在产品"],
        "库存商品_期初": ["库存商品"],
        "工程物资_期初": ["工程物资"],
        "委托加工物资_期初": ["委托加工"],
        "存货跌价准备_期初": ["跌价准备"],
        "其他业务成本_期初": ["其他业务成本"],
        "累计折旧_期初": ["累计折旧"],
        "累计摊销_期初": ["累计摊销"],
        "折旧": ["折旧"],
        "薪酬": ["薪酬", "工资"],
        "摊销": ["摊销"],
    }
    # 也处理 _未审数 后缀（与 _期初 共享科目关键词）
    base_keyword = keyword.replace("_未审数", "_期初").replace("_期初", "_期初")
    terms = col_keywords.get(keyword) or col_keywords.get(base_keyword, [keyword])

    for row in ws.iter_rows(min_row=1, max_row=10, max_col=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(t in cell.value for t in terms):
                    return cell.column
    return None


def _find_first_data_row(ws, col_idx: int) -> int | None:
    """找到指定列的第一个数据行（列头下方，跳过标题行）

    从第 7 行开始找（致同模板前 6 行通常是标题+列头），
    返回第一个空单元格或数字单元格的行号。
    """
    for row_num in range(7, 50):
        cell = ws.cell(row=row_num, column=col_idx)
        # 空单元格或已有数字（可覆盖）
        if cell.value is None or isinstance(cell.value, (int, float)):
            return row_num
        # 如果是公式（以=开头），也可以覆盖
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return row_num
    return None


def _mark_prefilled_cell(ws, cell) -> None:
    """P2-1: 为预填充的单元格设置浅蓝色背景标记

    用户可通过背景色识别哪些单元格是系统自动填入的。
    同时在单元格 comment 中记录来源信息。
    """
    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment

    # 浅蓝色背景（与 AI 内容标记一致）
    prefill_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    try:
        cell.fill = prefill_fill
        # 添加批注说明来源
        if not cell.comment:
            cell.comment = Comment("系统预填充：数据来自试算表", "系统")
    except Exception:
        pass  # 某些只读单元格无法设置样式


def _mark_user_formula_cell(ws, cell) -> None:
    """E1 spec Task 1.20: 为用户自定义公式 cell 设置浅绿色背景标记

    用户公式与系统预设公式视觉区分:
    - 系统预设(_mark_prefilled_cell): 浅蓝 #E8F4FD
    - 用户自定义(_mark_user_formula_cell): 浅绿 #E6F4EA
    """
    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment

    user_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    try:
        cell.fill = user_fill
        if not cell.comment:
            cell.comment = Comment("用户自定义公式", "用户")
    except Exception:
        pass


def _find_semantic_row(ws, keyword: str, wp_code: str = "", cell_ref: str = "") -> int | None:
    """在 sheet 中查找包含关键词的位置

    致同模板有两种布局：
    1. 行标签模式：A 列有"期初余额"等标签，数据在同行的 E/F 列
    2. 列头模式：第 5-6 行有"期初数"/"未审数"等列头，数据在该列的数据行

    本函数优先在 A-D 列的行标签中找，其次在前 10 行的列头中找。
    """
    search_terms = {
        "期初余额": ["期初", "年初", "上年末", "期初余额", "年初数", "上年期末", "期初数", "期初金额"],
        "未审数": ["未审", "期末余额", "未审数", "账面余额", "账面数", "未审定", "期末数", "未审金额"],
        "AJE调整": ["AJE", "审计调整", "调整分录", "审计调整数", "账项调整", "审计调整额", "调整数"],
        "RJE调整": ["RJE", "重分类", "重分类调整", "重分类数", "重分类金额", "重分类"],
        "上年审定数": ["上年", "上期", "上年审定", "上年审定数", "上期审定", "上年数", "上年末", "审定数"],
        # 子科目分项（E1/F2 等多科目底稿）
        "库存现金_期初": ["库存现金", "现金"],
        "库存现金_未审数": ["库存现金", "现金"],
        "银行存款_期初": ["银行存款", "银行"],
        "银行存款_未审数": ["银行存款", "银行"],
        "其他货币资金_期初": ["其他货币资金", "其他货币"],
        "其他货币资金_未审数": ["其他货币资金", "其他货币"],
        "原材料_期初": ["原材料"],
        "原材料_未审数": ["原材料"],
        "在产品_期初": ["在产品", "在制品"],
        "在产品_未审数": ["在产品", "在制品"],
        "库存商品_期初": ["库存商品", "产成品"],
        "库存商品_未审数": ["库存商品", "产成品"],
        "工程物资_期初": ["工程物资"],
        "工程物资_未审数": ["工程物资"],
        "委托加工物资_期初": ["委托加工", "委外加工"],
        "委托加工物资_未审数": ["委托加工", "委外加工"],
        "存货跌价准备_期初": ["跌价准备", "存货跌价"],
        "存货跌价准备_未审数": ["跌价准备", "存货跌价"],
        "其他业务成本_期初": ["其他业务成本", "其他成本"],
        "其他业务成本_未审数": ["其他业务成本", "其他成本"],
        "累计折旧_期初": ["累计折旧"],
        "累计折旧_未审数": ["累计折旧"],
        "累计摊销_期初": ["累计摊销"],
        "累计摊销_未审数": ["累计摊销"],
        "折旧": ["折旧", "折旧费", "折旧摊销"],
        "薪酬": ["薪酬", "工资", "职工薪酬", "人工"],
        "摊销": ["摊销", "摊销费", "无形资产摊销"],
    }
    terms = search_terms.get(keyword, [keyword])

    # 策略 1：在 A-D 列的行标签中找（标准行标签模式）
    for row in ws.iter_rows(min_row=1, max_row=80, max_col=4):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(t in cell.value for t in terms):
                    return cell.row

    # 策略 2：在前 10 行的所有列中找列头（致同列头模式）
    # 如果找到列头，返回该列头所在行+1（数据起始行）
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(t in cell.value for t in terms):
                    # 找到列头，返回下一行（数据行）或当前行
                    return cell.row

    logger.warning(
        "prefill 语义行未匹配: wp_code=%s, keyword=%s, cell_ref=%s (请人工复查模板结构)",
        wp_code, keyword, cell_ref,
    )
    return None


def _find_data_column(ws, row_num: int) -> int:
    """找到数据列（第一个数字列或 E 列）"""
    for col in range(3, 20):  # C 列开始找
        cell = ws.cell(row=row_num, column=col)
        if cell.value is None or isinstance(cell.value, (int, float)):
            return col
    return 5  # 默认 E 列


def get_workpaper_file(project_id: UUID, wp_id: UUID) -> Optional[Path]:
    """获取已存在的底稿文件路径（如果存在）"""
    path = get_workpaper_storage_path(project_id, wp_id)
    return path if path.exists() else None


def list_available_templates() -> list[dict]:
    """列出所有可用模板（供前端选择）"""
    index = _load_index()
    # 按 wp_code 去重，只返回主文件
    seen = set()
    result = []
    for e in sorted(index, key=lambda x: (x["wp_code"], len(x["filename"]))):
        code = e["wp_code"]
        if code in seen or code == "_ref":
            continue
        seen.add(code)
        result.append({
            "wp_code": code,
            "filename": e["filename"],
            "format": e["format"],
            "size_kb": e["size_kb"],
        })
    return result
