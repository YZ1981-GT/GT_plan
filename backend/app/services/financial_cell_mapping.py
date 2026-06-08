"""财务报表 cell_mapping 加载与双轨填充（内联 {{row:}} 优先，JSON 回退）."""

from __future__ import annotations

import json
import re
from decimal import Decimal
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import coordinate_to_tuple

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
MAPPING_PATH = DATA_DIR / "audit_report_templates" / "cell_mapping.json"

INLINE_ROW_RE = re.compile(r"^\{\{row:([^:]+):(current|prior)\}\}$")
HEADER_TOKEN_RE = re.compile(r"\{\{([a-zA-Z0-9_]+)\}\}")


@lru_cache(maxsize=1)
def load_cell_mapping_root() -> dict[str, Any]:
    if not MAPPING_PATH.is_file():
        return {}
    return json.loads(MAPPING_PATH.read_text(encoding="utf-8"))


def get_variant_mapping(variant_key: str) -> dict[str, Any] | None:
    variants = load_cell_mapping_root().get("variants", {})
    block = variants.get(variant_key)
    return dict(block) if isinstance(block, dict) else None


def build_amount_lookup(report_data: dict[str, list[dict]]) -> dict[str, dict[str, Any]]:
    """row_code → {current_period_amount, prior_period_amount, is_total_row, ...}."""
    lookup: dict[str, dict[str, Any]] = {}
    for rows in report_data.values():
        for row in rows:
            code = row.get("row_code")
            if code:
                lookup[code] = row
    return lookup


def scan_worksheet_row_placeholders(ws) -> dict[str, dict[str, str]]:
    """扫描 sheet 内联 ``{{row:CODE:current|prior}}`` → {code: {current: C6, prior: D6}}."""
    found: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if not isinstance(val, str):
                continue
            m = INLINE_ROW_RE.match(val.strip())
            if not m:
                continue
            code, period = m.group(1), m.group(2)
            found.setdefault(code, {})[period] = cell.coordinate
    return found


def sheet_has_row_placeholders(ws) -> bool:
    return bool(scan_worksheet_row_placeholders(ws))


def row_mappings_for_sheet(
    variant_key: str,
    report_type: str,
    ws,
) -> dict[str, dict[str, str]]:
    """合并 cell_mapping.json 与内联扫描（内联覆盖 JSON 同字段）."""
    merged: dict[str, dict[str, str]] = {}
    variant = get_variant_mapping(variant_key)
    if variant:
        for code, entry in variant.get("rows", {}).items():
            if entry.get("sheet") != report_type:
                continue
            coords: dict[str, str] = {}
            if entry.get("current"):
                coords["current"] = entry["current"]
            if entry.get("prior"):
                coords["prior"] = entry["prior"]
            if coords:
                merged[code] = coords
    inline = scan_worksheet_row_placeholders(ws)
    for code, coords in inline.items():
        merged[code] = {**merged.get(code, {}), **coords}
    return merged


def replace_header_placeholders(ws, values: dict[str, str]) -> int:
    """替换 sheet 中含 ``{{key}}`` 的单元格文本；返回替换单元格数."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if not isinstance(val, str) or "{{" not in val:
                continue
            new_text = val
            for key, repl in values.items():
                new_text = new_text.replace(f"{{{{{key}}}}}", repl or "")
            if new_text != val:
                cell.value = new_text
                count += 1
    return count


def _is_formula_cell(ws, coord: str) -> bool:
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        return True
    val = cell.value
    return isinstance(val, str) and val.startswith("=")


def fill_row_amounts(
    ws,
    mappings: dict[str, dict[str, str]],
    amounts_by_code: dict[str, dict[str, Any]],
    *,
    include_prior_year: bool = True,
    apply_amount_format,
) -> int:
    """按映射写入金额；跳过公式格。返回写入格数."""
    written = 0
    for code, coords in mappings.items():
        row_data = amounts_by_code.get(code, {})
        is_total = bool(row_data.get("is_total_row"))
        for period, coord in coords.items():
            if period == "prior" and not include_prior_year:
                continue
            if _is_formula_cell(ws, coord):
                continue
            field = (
                "current_period_amount"
                if period == "current"
                else "prior_period_amount"
            )
            amount = row_data.get(field)
            if amount is None:
                continue
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                continue
            cell.value = float(amount)
            if apply_amount_format:
                apply_amount_format(cell, is_total)
            written += 1
    return written


def build_header_values(
    *,
    company_name: str,
    year: int,
    report_type: str,
    mode: str,
) -> dict[str, str]:
    period_end = f"{year}年12月31日"
    if report_type == "balance_sheet":
        period_label = period_end
    else:
        period_label = f"{year}年度"
    mode_label = "未审数" if mode == "unadjusted" else "审定数"
    return {
        "company_full_name": company_name,
        "period_end_date": period_end,
        "audit_year": str(year),
        "currency_unit": "元",
        "report_title": "资产负债表" if report_type == "balance_sheet" else "",
        "period_label": period_label,
        "mode_label": mode_label,
    }
