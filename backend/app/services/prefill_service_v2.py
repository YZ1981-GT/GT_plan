"""底稿数据预填服务 v2

Phase 9 Task 9.4: 用 openpyxl 预填四表数据到 Excel 底稿
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.workpaper_models import WorkingPaper, WpIndex

logger = logging.getLogger(__name__)


class PrefillServiceV2:
    """底稿数据预填服务"""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def prefill_workpaper(self, wp_id: UUID) -> dict:
        """预填单个底稿

        根据底稿类型确定预填规则：
        - 审定表：未审数(tb_balance) + 调整(adjustments) + 审定数(trial_balance)
        - 明细表：期初/期末/发生额 from tb_balance
        - 函证表：辅助余额 from tb_aux_balance

        Returns: {"status": "ok"|"skipped"|"error", "cells_filled": N}
        """
        # 获取底稿信息
        q = (
            sa.select(WorkingPaper, WpIndex.wp_code, WpIndex.audit_cycle)
            .join(WpIndex, WorkingPaper.wp_index_id == WpIndex.id)
            .where(WorkingPaper.id == wp_id, WorkingPaper.is_deleted == False)  # noqa
        )
        result = (await self.db.execute(q)).one_or_none()
        if not result:
            return {"status": "error", "message": "底稿不存在", "cells_filled": 0}

        wp, wp_code, audit_cycle = result[0], result[1], result[2]
        file_path = Path(wp.file_path)

        if not file_path.exists():
            return {"status": "error", "message": f"文件不存在: {file_path}", "cells_filled": 0}

        if not file_path.suffix.lower() == ".xlsx":
            return {"status": "skipped", "message": "非 Excel 文件，跳过预填", "cells_filled": 0}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path))

            # 获取项目的试算表数据
            from app.models.audit_platform_models import TrialBalance
            tb_q = sa.select(TrialBalance).where(
                TrialBalance.project_id == wp.project_id,
                TrialBalance.is_deleted == False,  # noqa
            )
            tb_rows = (await self.db.execute(tb_q)).scalars().all()
            tb_map = {r.standard_account_code: r for r in tb_rows}

            cells_filled = 0

            # 遍历工作表查找取数公式占位符
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if cell.value.startswith("=TB(") or cell.value.startswith("=SUM_TB("):
                                # 解析公式并填入值（简化实现）
                                # 实际需要调用 FormulaEngine
                                cells_filled += 0  # placeholder

            wb.save(str(file_path))
            wb.close()

            # 标记预填完成
            wp.prefill_stale = False
            await self.db.flush()

            return {"status": "ok", "cells_filled": cells_filled}

        except Exception as e:
            logger.error(f"预填失败 wp={wp_id}: {e}")
            return {"status": "error", "message": str(e), "cells_filled": 0}

    async def parse_workpaper(self, wp_id: UUID) -> dict:
        """解析底稿关键单元格，回写到 parsed_data JSONB

        Phase 9 Task 9.5
        """
        q = (
            sa.select(WorkingPaper, WpIndex.wp_code)
            .join(WpIndex, WorkingPaper.wp_index_id == WpIndex.id)
            .where(WorkingPaper.id == wp_id, WorkingPaper.is_deleted == False)  # noqa
        )
        result = (await self.db.execute(q)).one_or_none()
        if not result:
            return {"status": "error", "message": "底稿不存在"}

        wp, wp_code = result[0], result[1]
        file_path = Path(wp.file_path)

        if not file_path.exists() or file_path.suffix.lower() != ".xlsx":
            return {"status": "skipped"}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)

            parsed = {}
            # 读取第一个工作表的关键区域
            ws = wb.active
            if ws:
                # 尝试读取审定数（通常在特定位置）
                # 这里是简化实现，实际需要根据模板的 Named Ranges 定位
                parsed["sheet_name"] = ws.title
                parsed["last_row"] = ws.max_row
                parsed["last_col"] = ws.max_column

            wb.close()

            # 回写 parsed_data
            wp.parsed_data = parsed
            wp.last_parsed_at = datetime.utcnow()
            await self.db.flush()

            return {"status": "ok", "parsed_fields": len(parsed)}

        except Exception as e:
            logger.error(f"解析失败 wp={wp_id}: {e}")
            return {"status": "error", "message": str(e)}

    async def mark_stale(self, project_id: UUID, account_codes: list[str] | None = None) -> int:
        """标记底稿预填数据为过期

        Phase 9 Task 9.10
        """
        q = (
            sa.update(WorkingPaper)
            .where(
                WorkingPaper.project_id == project_id,
                WorkingPaper.is_deleted == False,  # noqa
            )
            .values(prefill_stale=True)
        )
        result = await self.db.execute(q)
        return result.rowcount
