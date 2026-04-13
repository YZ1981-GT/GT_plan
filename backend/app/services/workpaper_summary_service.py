"""底稿跨企业汇总服务

根据选定的科目前缀和企业代码，从 trial_balance 表查询数据，
按科目行×企业列生成汇总矩阵，支持合计行和导出 Excel。
"""

from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_platform_models import TrialBalance
from app.models.core import Project


class WorkpaperSummaryService:
    """跨企业底稿汇总"""

    async def get_child_companies(
        self,
        db: AsyncSession,
        parent_project_id: UUID,
    ) -> list[dict]:
        """获取合并项目下的子公司列表（含母公司自身）"""
        # 查询 parent_project_id 指向当前项目的所有子项目
        stmt = (
            select(Project)
            .where(
                Project.parent_project_id == parent_project_id,
                Project.is_deleted == False,  # noqa: E712
            )
            .order_by(Project.company_code)
        )
        result = await db.execute(stmt)
        children = result.scalars().all()

        companies: list[dict] = []
        for p in children:
            companies.append({
                "project_id": str(p.id),
                "company_code": p.company_code or str(p.id)[:8],
                "company_name": p.client_name or p.name,
            })
        return companies

    async def generate_summary(
        self,
        db: AsyncSession,
        parent_project_id: UUID,
        year: int,
        account_codes: list[str],
        company_codes: list[str],
    ) -> dict:
        """生成跨企业汇总数据。

        Returns:
            {
                "companies": ["001", "002"],
                "company_names": {"001": "母公司", "002": "子公司A"},
                "rows": [
                    {
                        "account_code": "1122",
                        "account_name": "应收账款",
                        "values": {"001": 5000000, "002": 2000000},
                        "total": 7000000
                    }, ...
                ],
                "grand_total": {"001": ..., "002": ..., "total": ...}
            }
        """
        # 1) 获取企业名称映射
        company_names = await self._get_company_names(db, parent_project_id, company_codes)

        # 2) 获取子项目 ID 列表（按 company_code 过滤）
        project_ids = await self._get_project_ids(db, parent_project_id, company_codes)
        if not project_ids:
            return {
                "companies": company_codes,
                "company_names": company_names,
                "rows": [],
                "grand_total": {c: 0 for c in company_codes} | {"total": 0},
            }

        # 3) 查询 trial_balance
        conditions = [
            TrialBalance.project_id.in_(project_ids),
            TrialBalance.year == year,
            TrialBalance.company_code.in_(company_codes),
            TrialBalance.is_deleted == False,  # noqa: E712
        ]
        # 按科目前缀过滤（LIKE 'prefix%'）
        if account_codes:
            from sqlalchemy import or_
            prefix_filters = [
                TrialBalance.standard_account_code.like(f"{code}%")
                for code in account_codes
            ]
            conditions.append(or_(*prefix_filters))

        stmt = (
            select(TrialBalance)
            .where(*conditions)
            .order_by(TrialBalance.standard_account_code, TrialBalance.company_code)
        )
        result = await db.execute(stmt)
        rows_raw = result.scalars().all()

        # 4) 按科目分组，按企业透视
        account_map: dict[str, dict] = {}
        for row in rows_raw:
            code = row.standard_account_code
            if code not in account_map:
                account_map[code] = {
                    "account_code": code,
                    "account_name": row.account_name or "",
                    "values": {},
                }
            amount = float(row.audited_amount or row.unadjusted_amount or 0)
            cc = row.company_code
            account_map[code]["values"][cc] = (
                account_map[code]["values"].get(cc, 0) + amount
            )

        # 5) 构建结果行 + 行合计
        rows: list[dict] = []
        grand_total: dict[str, float] = {c: 0.0 for c in company_codes}
        grand_total["total"] = 0.0

        for code in sorted(account_map.keys()):
            entry = account_map[code]
            row_total = 0.0
            for cc in company_codes:
                val = entry["values"].get(cc, 0)
                row_total += val
                grand_total[cc] = grand_total.get(cc, 0) + val
            entry["total"] = row_total
            grand_total["total"] += row_total
            rows.append(entry)

        return {
            "companies": company_codes,
            "company_names": company_names,
            "rows": rows,
            "grand_total": grand_total,
        }

    async def export_excel(
        self,
        db: AsyncSession,
        parent_project_id: UUID,
        year: int,
        account_codes: list[str],
        company_codes: list[str],
    ) -> BytesIO:
        """生成 Excel 文件并返回 BytesIO 流"""
        summary = await self.generate_summary(
            db, parent_project_id, year, account_codes, company_codes,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "底稿跨企业汇总"

        # 样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4B2D77", end_color="4B2D77", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        amount_align = Alignment(horizontal="right")
        total_font = Font(bold=True, size=11)
        thin_border = Border(
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 表头
        headers = ["科目编码", "科目名称"]
        company_names = summary["company_names"]
        for cc in summary["companies"]:
            headers.append(company_names.get(cc, cc))
        headers.append("合计")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, row_data in enumerate(summary["rows"], 2):
            ws.cell(row=row_idx, column=1, value=row_data["account_code"])
            ws.cell(row=row_idx, column=2, value=row_data["account_name"])
            for col_offset, cc in enumerate(summary["companies"]):
                val = row_data["values"].get(cc, 0)
                cell = ws.cell(row=row_idx, column=3 + col_offset, value=val)
                cell.alignment = amount_align
                cell.number_format = '#,##0.00'
            total_cell = ws.cell(
                row=row_idx,
                column=3 + len(summary["companies"]),
                value=row_data["total"],
            )
            total_cell.alignment = amount_align
            total_cell.number_format = '#,##0.00'

        # 合计行
        gt = summary["grand_total"]
        total_row = len(summary["rows"]) + 2
        ws.cell(row=total_row, column=1, value="合计").font = total_font
        ws.cell(row=total_row, column=2, value="")
        for col_offset, cc in enumerate(summary["companies"]):
            cell = ws.cell(row=total_row, column=3 + col_offset, value=gt.get(cc, 0))
            cell.font = total_font
            cell.alignment = amount_align
            cell.number_format = '#,##0.00'
            cell.border = thin_border
        grand_cell = ws.cell(
            row=total_row,
            column=3 + len(summary["companies"]),
            value=gt.get("total", 0),
        )
        grand_cell.font = total_font
        grand_cell.alignment = amount_align
        grand_cell.number_format = '#,##0.00'
        grand_cell.border = thin_border

        # 列宽
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20
        for i in range(len(summary["companies"]) + 1):
            col_letter = chr(ord("C") + i)
            ws.column_dimensions[col_letter].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── 内部辅助 ──

    async def _get_company_names(
        self, db: AsyncSession, parent_project_id: UUID, company_codes: list[str],
    ) -> dict[str, str]:
        """从子项目获取 company_code → company_name 映射"""
        stmt = (
            select(Project)
            .where(
                Project.parent_project_id == parent_project_id,
                Project.company_code.in_(company_codes),
                Project.is_deleted == False,  # noqa: E712
            )
        )
        result = await db.execute(stmt)
        projects = result.scalars().all()
        return {
            (p.company_code or ""): (p.client_name or p.name)
            for p in projects
        }

    async def _get_project_ids(
        self, db: AsyncSession, parent_project_id: UUID, company_codes: list[str],
    ) -> list[UUID]:
        """获取匹配 company_code 的子项目 ID 列表"""
        stmt = (
            select(Project.id)
            .where(
                Project.parent_project_id == parent_project_id,
                Project.company_code.in_(company_codes),
                Project.is_deleted == False,  # noqa: E712
            )
        )
        result = await db.execute(stmt)
        return [row[0] for row in result.all()]
