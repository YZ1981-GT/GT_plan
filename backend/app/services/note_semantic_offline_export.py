"""附注语义离线导出扩展 — Task 10

在现有 note_offline_export_service.py 基础上扩展：
1. 新增 00_填报说明 sheet（语义版说明，含六色规范图例）
2. 新增 01_章节清单 sheet（含语义 section_id 和 variant 信息）
3. 定义语义颜色规范（六色体系，替代旧版四色体系）

本模块不修改现有 note_offline_export_service.py，作为独立扩展模块。
Requirements: 7.1, 7.2
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    "COLOR_SPEC",
    "SEMANTIC_FILLS",
    "HIDDEN_SEMANTIC_COLUMNS",
    "generate_instruction_sheet_data",
    "generate_section_list_sheet_data",
    "build_semantic_instruction_sheet",
    "build_semantic_section_list_sheet",
    "build_policy_clauses_sheet",
    "build_account_disclosure_sheet",
    "build_related_party_sheet",
    "build_validation_results_sheet",
    "create_semantic_workbook",
]

# ---------------------------------------------------------------------------
# 六色语义规范 (Task 10.4)
# ---------------------------------------------------------------------------

COLOR_SPEC: dict[str, str] = {
    "editable": "#FFFFFF",          # 白色底 — 可填
    "locked": "#E0E0E0",            # 灰色底 — 锁定
    "workpaper_source": "#D4EDDA",  # 浅绿底 — 来源底稿
    "needs_review": "#FFF3CD",      # 浅黄底 — 需复核
    "validation_failed": "#F8D7DA", # 浅红底 — 校验失败
    "prior_reference": "#D1ECF1",   # 浅蓝底 — 上年/模板参考
}

# openpyxl PatternFill 对象（hex 颜色不含 '#'）
SEMANTIC_FILLS: dict[str, PatternFill] = {
    key: PatternFill(
        start_color="FF" + color.lstrip("#"),
        end_color="FF" + color.lstrip("#"),
        fill_type="solid",
    )
    for key, color in COLOR_SPEC.items()
}

# ---------------------------------------------------------------------------
# 隐藏语义列定义 (Task 15.5)
# ---------------------------------------------------------------------------

HIDDEN_SEMANTIC_COLUMNS: list[str] = [
    "section_id",
    "table_id",
    "row_id",
    "col_id",
    "binding_id",
    "cell_mode",
]

# ---------------------------------------------------------------------------
# Fonts / Styles
# ---------------------------------------------------------------------------

_FONT_TITLE = Font(bold=True, size=14)
_FONT_HEADER = Font(bold=True, size=12)
_FONT_NORMAL = Font(size=10)
_FONT_COLOR_SAMPLE = Font(size=10, bold=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# 纯数据生成函数 (Task 10.2)
# ---------------------------------------------------------------------------


def generate_instruction_sheet_data(
    *,
    project_name: str = "",
    year: str | int = "",
    exporter_name: str = "",
    section_count: int = 0,
) -> list[list[str]]:
    """生成 00_填报说明 sheet 纯文本行数据。

    每行是 [文本内容]，用于测试或非 openpyxl 场景。
    """
    export_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    rows: list[list[str]] = [
        ["═" * 50],
        ["  附注语义离线编辑包 — 填报说明"],
        ["═" * 50],
        [""],
        ["【1. 文件用途】"],
        [f"  本文件为「{project_name} - {year} 年报附注」语义离线协作编辑包，"],
        [f"  共 {section_count} 章节，由 {exporter_name} 于 {export_time} 生成。"],
        [""],
        ["【2. 单元格颜色含义（六色规范）】"],
        ["  白色底 = 可填（editable）：您可直接输入数据"],
        ["  灰色底 = 锁定（locked）：系统保护，禁止修改"],
        ["  浅绿底 = 来源底稿（workpaper_source）：数据来自底稿自动取数"],
        ["  浅黄底 = 需复核（needs_review）：待上级复核确认"],
        ["  浅红底 = 校验失败（validation_failed）：金额校验未通过"],
        ["  浅蓝底 = 上年/模板参考（prior_reference）：参考数据，可按需修改"],
        [""],
        ["【3. 结构保护规则】"],
        ["  ⚠ 不要修改隐藏的语义列（section_id/table_id/row_id/col_id）"],
        ["  ⚠ 不要删除或重命名 _meta sheet"],
        ["  ⚠ 标题行、合计行、分组行为结构锁定行，禁止修改"],
        ["  ⚠ 公式列（灰底）由系统计算，修改将标记为公式覆盖冲突"],
        [""],
        ["【4. 动态行操作】"],
        ["  在带 ★ 标识的动态行区域："],
        ["    - 复制现有行并在标记行之间粘贴"],
        ["    - 不要在 [DYNAMIC_END] 标记下方插入数据"],
        [""],
        ["【5. 完成后回传】"],
        ["  填写完成后将本文件导入系统，系统将自动识别修改内容。"],
        ["  若存在冲突（结构/锁定/公式），系统会提示处理方式。"],
        [""],
        ["【6. 版本信息】"],
        ["  本工作包为语义版（semantic workbook），包含 _meta 隐藏 sheet。"],
        ["  系统通过 _meta sheet 区分新旧版本，请勿删除。"],
        ["═" * 50],
    ]
    return rows


# ---------------------------------------------------------------------------
# 纯数据生成函数 (Task 10.3)
# ---------------------------------------------------------------------------


def generate_section_list_sheet_data(
    sections: list[dict[str, Any]],
) -> list[list[str]]:
    """生成 01_章节清单 sheet 数据。

    返回二维列表：第一行是表头，后续行为各章节信息。
    列: [序号, 章节标题, section_id, variant, scope, 完成度(%)]
    """
    header = ["序号", "章节标题", "section_id", "variant", "scope", "完成度(%)"]
    rows: list[list[str]] = [header]

    for idx, section in enumerate(sections, start=1):
        section_id = section.get("section_id", "")
        title = section.get("section_title", section.get("title", ""))
        semantic = section.get("_semantic", {})
        variant = semantic.get("variant", "")
        scope = semantic.get("scope", "")
        completeness = _calc_section_completeness(section)

        rows.append([
            str(idx),
            title,
            section_id,
            variant,
            scope,
            str(completeness),
        ])

    return rows


def _calc_section_completeness(section: dict[str, Any]) -> int:
    """计算章节完成度百分比。"""
    table_data = section.get("table_data", {})
    all_rows = table_data.get("rows", [])
    if not all_rows:
        return 0

    total = 0
    filled = 0
    for row in all_rows:
        cells = row.get("cells", row.get("values", []))
        for cell_val in cells:
            total += 1
            if cell_val is not None and cell_val != "" and cell_val != "-":
                filled += 1

    return int((filled / total * 100) if total > 0 else 0)


# ---------------------------------------------------------------------------
# openpyxl Sheet 构建函数 (Task 10.1, 10.2, 10.3)
# ---------------------------------------------------------------------------


def build_semantic_instruction_sheet(
    ws: Worksheet,
    *,
    project_name: str = "",
    year: str | int = "",
    exporter_name: str = "",
    section_count: int = 0,
) -> None:
    """构建 00_填报说明 worksheet（含颜色示例色块）。"""
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 20

    data = generate_instruction_sheet_data(
        project_name=project_name,
        year=year,
        exporter_name=exporter_name,
        section_count=section_count,
    )

    for row_idx, row_data in enumerate(data, start=1):
        text = row_data[0] if row_data else ""
        cell = ws.cell(row=row_idx, column=1, value=text)

        # Style selection
        if text.startswith("═"):
            cell.font = _FONT_TITLE
        elif text.startswith("【"):
            cell.font = _FONT_HEADER
        else:
            cell.font = _FONT_NORMAL
        cell.alignment = _ALIGN_LEFT

    # 在颜色说明区域旁(B列)添加色块示例
    color_start_row = 10  # 对应"白色底 = 可填"行
    color_items = [
        ("editable", "可填"),
        ("locked", "锁定"),
        ("workpaper_source", "来源底稿"),
        ("needs_review", "需复核"),
        ("validation_failed", "校验失败"),
        ("prior_reference", "上年/模板参考"),
    ]
    for offset, (key, label) in enumerate(color_items):
        cell = ws.cell(row=color_start_row + offset, column=2, value=f"  {label}  ")
        cell.fill = SEMANTIC_FILLS[key]
        cell.font = _FONT_COLOR_SAMPLE
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN


def build_semantic_section_list_sheet(
    ws: Worksheet,
    sections: list[dict[str, Any]],
) -> None:
    """构建 01_章节清单 worksheet。"""
    data = generate_section_list_sheet_data(sections)

    col_widths = [8, 40, 30, 20, 15, 12]
    headers_row = data[0] if data else []

    # Header row
    for col_idx, header in enumerate(headers_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
        if col_idx <= len(col_widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # Hide section_id, variant, scope columns (C, D, E)
    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["E"].hidden = True

    # Data rows
    for row_idx, row_data in enumerate(data[1:], start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER_THIN
            cell.font = _FONT_NORMAL


# ---------------------------------------------------------------------------
# Task 15.1: 政策条款 sheet
# ---------------------------------------------------------------------------


def _write_hidden_semantic_cols(
    ws: Worksheet,
    row_idx: int,
    start_col: int,
    values: dict[str, str],
) -> None:
    """向指定行写入隐藏语义列值。"""
    for offset, col_name in enumerate(HIDDEN_SEMANTIC_COLUMNS):
        cell = ws.cell(row=row_idx, column=start_col + offset, value=values.get(col_name, ""))
        cell.font = _FONT_NORMAL


def _setup_hidden_semantic_cols(ws: Worksheet, start_col: int) -> None:
    """设置隐藏语义列表头并隐藏列。"""
    for offset, col_name in enumerate(HIDDEN_SEMANTIC_COLUMNS):
        col_letter = get_column_letter(start_col + offset)
        ws.cell(row=1, column=start_col + offset, value=col_name)
        ws.column_dimensions[col_letter].hidden = True


def build_policy_clauses_sheet(ws: Worksheet, clauses: list[dict[str, Any]]) -> None:
    """构建政策条款 sheet (Task 15.1)。

    列: 序号 | 条款标题 | 层级 | 本年内容 | 模板内容 | 上年内容 | 差异状态 | 确认状态
         + 隐藏语义列

    Args:
        ws: 目标 worksheet
        clauses: 条款列表, 每个条款包含 clause_id/title/level/current_text 等
    """
    visible_headers = [
        "序号", "条款标题", "层级", "本年内容", "模板内容", "上年内容", "差异状态", "确认状态",
    ]
    visible_col_count = len(visible_headers)
    col_widths = [6, 30, 6, 50, 50, 50, 12, 12]

    # Write visible headers
    for col_idx, header in enumerate(visible_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
        if col_idx <= len(col_widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # Setup hidden semantic columns after visible columns
    hidden_start = visible_col_count + 1
    _setup_hidden_semantic_cols(ws, hidden_start)

    # Data rows
    for row_idx, clause in enumerate(clauses, start=2):
        clause_id = clause.get("clause_id", "")
        section_id = clause.get("section_id", "")
        title = clause.get("title", "")
        level = clause.get("level", 1)
        current_text = clause.get("current_text", "")
        template_text = clause.get("template_text", "")
        prior_text = clause.get("prior_year_text", "")
        diff_status = clause.get("diff_status", "")
        confirm_status = clause.get("confirm_status", "")

        values = [row_idx - 1, title, level, current_text, template_text, prior_text, diff_status, confirm_status]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
            # Apply color based on diff_status
            if col_idx == 7:  # 差异状态列
                if diff_status == "changed":
                    cell.fill = SEMANTIC_FILLS["needs_review"]
                elif diff_status == "unchanged":
                    cell.fill = SEMANTIC_FILLS["editable"]

        # Write hidden semantic columns
        _write_hidden_semantic_cols(ws, row_idx, hidden_start, {
            "section_id": section_id,
            "table_id": "",
            "row_id": clause_id,
            "col_id": "",
            "binding_id": "",
            "cell_mode": "locked",
        })


# ---------------------------------------------------------------------------
# Task 15.2: 科目披露 sheet
# ---------------------------------------------------------------------------


def build_account_disclosure_sheet(ws: Worksheet, tables: list[dict[str, Any]]) -> None:
    """构建科目披露 sheet (Task 15.2)。

    每个表包含: section_id, table_id, table_name, columns, rows。
    多表纵向排列，表与表之间空一行。

    Args:
        ws: 目标 worksheet
        tables: 科目披露表列表
    """
    current_row = 1

    # 设置隐藏列（固定从 H 列开始，即第 8 列起）
    hidden_start = 8
    _setup_hidden_semantic_cols(ws, hidden_start)

    for table in tables:
        section_id = table.get("section_id", "")
        table_id = table.get("table_id", "")
        table_name = table.get("name", table.get("table_name", ""))
        columns = table.get("columns", [])
        rows = table.get("rows", [])

        # Table title row
        cell = ws.cell(row=current_row, column=1, value=table_name)
        cell.font = _FONT_HEADER
        cell.fill = SEMANTIC_FILLS["locked"]
        _write_hidden_semantic_cols(ws, current_row, hidden_start, {
            "section_id": section_id,
            "table_id": table_id,
            "row_id": "",
            "col_id": "",
            "binding_id": "",
            "cell_mode": "locked",
        })
        current_row += 1

        # Column headers
        for col_idx, col_def in enumerate(columns, start=1):
            col_label = col_def.get("label", col_def) if isinstance(col_def, dict) else str(col_def)
            cell = ws.cell(row=current_row, column=col_idx, value=col_label)
            cell.font = _FONT_HEADER
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_THIN
        _write_hidden_semantic_cols(ws, current_row, hidden_start, {
            "section_id": section_id,
            "table_id": table_id,
            "row_id": "__header__",
            "col_id": "",
            "binding_id": "",
            "cell_mode": "locked",
        })
        current_row += 1

        # Data rows
        for row_data in rows:
            row_id = row_data.get("row_id", "")
            row_type = row_data.get("row_type", "data")
            cell_values = row_data.get("values", row_data.get("cells", []))
            cell_modes = row_data.get("_cell_modes", {})
            cell_meta = row_data.get("_cell_meta", {})

            for col_idx, val in enumerate(cell_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = _FONT_NORMAL
                cell.border = _BORDER_THIN

                # Apply cell mode colors
                mode = cell_modes.get(str(col_idx - 1), "editable")
                if mode in SEMANTIC_FILLS:
                    cell.fill = SEMANTIC_FILLS[mode]

                # Right-align numeric values
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Hidden semantic columns
            binding_id = ""
            if cell_meta:
                first_meta = next(iter(cell_meta.values()), {})
                binding_id = first_meta.get("binding_id", "") if isinstance(first_meta, dict) else ""

            _write_hidden_semantic_cols(ws, current_row, hidden_start, {
                "section_id": section_id,
                "table_id": table_id,
                "row_id": row_id,
                "col_id": "",
                "binding_id": binding_id,
                "cell_mode": "locked" if row_type in ("total", "subtotal", "table_title") else "editable",
            })
            current_row += 1

        # Blank row between tables
        current_row += 1

    # Set column widths for visible area
    for i in range(1, hidden_start):
        ws.column_dimensions[get_column_letter(i)].width = 18


# ---------------------------------------------------------------------------
# Task 15.3: 关联方 sheet
# ---------------------------------------------------------------------------


def build_related_party_sheet(
    ws: Worksheet,
    parties: list[dict[str, Any]],
    transactions: list[dict[str, Any]],
    balances: list[dict[str, Any]],
) -> None:
    """构建关联方 sheet (Task 15.3)。

    三个子区域纵向排列：
    1. 关联方清单（主体、关系类型）
    2. 关联方交易（主体、交易类型、本期发生额、上期发生额）
    3. 关联方余额（主体、项目、期末余额、期初余额）

    Args:
        ws: 目标 worksheet
        parties: 关联方主体列表
        transactions: 关联方交易列表
        balances: 关联方余额列表
    """
    hidden_start = 8
    _setup_hidden_semantic_cols(ws, hidden_start)
    current_row = 1

    # --- Section 1: 关联方清单 ---
    cell = ws.cell(row=current_row, column=1, value="一、关联方清单")
    cell.font = _FONT_HEADER
    _write_hidden_semantic_cols(ws, current_row, hidden_start, {
        "section_id": "related_party",
        "table_id": "party_list",
        "row_id": "",
        "col_id": "",
        "binding_id": "",
        "cell_mode": "locked",
    })
    current_row += 1

    party_headers = ["序号", "关联方名称", "关系类型", "注册资本", "持股比例", "关系说明"]
    for col_idx, h in enumerate(party_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
    current_row += 1

    for idx, party in enumerate(parties, start=1):
        row_values = [
            idx,
            party.get("name", ""),
            party.get("relationship_type", ""),
            party.get("registered_capital", ""),
            party.get("shareholding_ratio", ""),
            party.get("relationship_desc", ""),
        ]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
        _write_hidden_semantic_cols(ws, current_row, hidden_start, {
            "section_id": "related_party",
            "table_id": "party_list",
            "row_id": party.get("party_id", f"party_{idx}"),
            "col_id": "",
            "binding_id": "",
            "cell_mode": "editable",
        })
        current_row += 1

    current_row += 1  # Blank row

    # --- Section 2: 关联方交易 ---
    cell = ws.cell(row=current_row, column=1, value="二、关联方交易")
    cell.font = _FONT_HEADER
    _write_hidden_semantic_cols(ws, current_row, hidden_start, {
        "section_id": "related_party",
        "table_id": "transactions",
        "row_id": "",
        "col_id": "",
        "binding_id": "",
        "cell_mode": "locked",
    })
    current_row += 1

    tx_headers = ["序号", "关联方", "交易类型", "交易内容", "本期发生额", "上期发生额"]
    for col_idx, h in enumerate(tx_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
    current_row += 1

    for idx, tx in enumerate(transactions, start=1):
        row_values = [
            idx,
            tx.get("party_name", ""),
            tx.get("transaction_type", ""),
            tx.get("description", ""),
            tx.get("current_amount", ""),
            tx.get("prior_amount", ""),
        ]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
        _write_hidden_semantic_cols(ws, current_row, hidden_start, {
            "section_id": "related_party",
            "table_id": "transactions",
            "row_id": tx.get("transaction_id", f"tx_{idx}"),
            "col_id": "",
            "binding_id": tx.get("binding_id", ""),
            "cell_mode": "editable",
        })
        current_row += 1

    current_row += 1  # Blank row

    # --- Section 3: 关联方余额 ---
    cell = ws.cell(row=current_row, column=1, value="三、关联方余额")
    cell.font = _FONT_HEADER
    _write_hidden_semantic_cols(ws, current_row, hidden_start, {
        "section_id": "related_party",
        "table_id": "balances",
        "row_id": "",
        "col_id": "",
        "binding_id": "",
        "cell_mode": "locked",
    })
    current_row += 1

    bal_headers = ["序号", "关联方", "项目", "期末余额", "期初余额", "坏账准备"]
    for col_idx, h in enumerate(bal_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
    current_row += 1

    for idx, bal in enumerate(balances, start=1):
        row_values = [
            idx,
            bal.get("party_name", ""),
            bal.get("item", ""),
            bal.get("closing_balance", ""),
            bal.get("opening_balance", ""),
            bal.get("bad_debt_provision", ""),
        ]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
        _write_hidden_semantic_cols(ws, current_row, hidden_start, {
            "section_id": "related_party",
            "table_id": "balances",
            "row_id": bal.get("balance_id", f"bal_{idx}"),
            "col_id": "",
            "binding_id": bal.get("binding_id", ""),
            "cell_mode": "editable",
        })
        current_row += 1

    # Set column widths
    widths = [6, 20, 15, 20, 15, 15, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Task 15.4: 99_校验结果 sheet
# ---------------------------------------------------------------------------


def build_validation_results_sheet(ws: Worksheet, checklist_items: list[dict[str, Any]]) -> None:
    """构建 99_校验结果 sheet (Task 15.4)。

    展示质量清单校验结果，含 level/category/message/section_id 等。

    Args:
        ws: 目标 worksheet
        checklist_items: 质量清单条目列表
    """
    visible_headers = ["序号", "级别", "类别", "章节", "表格", "行", "列", "消息", "证据"]
    visible_col_count = len(visible_headers)
    col_widths = [6, 10, 12, 20, 20, 15, 15, 50, 30]

    # Write visible headers
    for col_idx, header in enumerate(visible_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
        if col_idx <= len(col_widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # Setup hidden semantic columns
    hidden_start = visible_col_count + 1
    _setup_hidden_semantic_cols(ws, hidden_start)

    # Data rows
    for row_idx, item in enumerate(checklist_items, start=2):
        level = item.get("level", "")
        category = item.get("category", "")
        section_id = item.get("section_id", "")
        table_id = item.get("table_id", "")
        row_id = item.get("row_id", "")
        col_id = item.get("col_id", "")
        message = item.get("message", "")
        evidence = item.get("evidence", "")
        if isinstance(evidence, dict):
            evidence = "; ".join(f"{k}={v}" for k, v in evidence.items())

        row_values = [row_idx - 1, level, category, section_id, table_id, row_id, col_id, message, evidence]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN

            # Color by level
            if col_idx == 2:  # 级别列
                if level == "blocking":
                    cell.fill = SEMANTIC_FILLS["validation_failed"]
                elif level == "warning":
                    cell.fill = SEMANTIC_FILLS["needs_review"]

        # Write hidden semantic columns
        _write_hidden_semantic_cols(ws, row_idx, hidden_start, {
            "section_id": section_id,
            "table_id": table_id,
            "row_id": row_id,
            "col_id": col_id,
            "binding_id": "",
            "cell_mode": "locked",
        })


# ---------------------------------------------------------------------------
# 完整 Workbook 创建
# ---------------------------------------------------------------------------


def create_semantic_workbook(
    sections: list[dict[str, Any]],
    *,
    project_name: str = "",
    year: str | int = "",
    exporter_name: str = "",
    policy_clauses: list[dict[str, Any]] | None = None,
    disclosure_tables: list[dict[str, Any]] | None = None,
    related_parties: dict[str, Any] | None = None,
    checklist_items: list[dict[str, Any]] | None = None,
) -> Workbook:
    """创建包含语义说明页和章节清单的 Workbook。

    可选参数传入时会追加对应 sheet：
    - policy_clauses: 政策条款数据 → 「政策条款」sheet
    - disclosure_tables: 科目披露表数据 → 「科目披露」sheet
    - related_parties: 关联方数据 (parties/transactions/balances) → 「关联方」sheet
    - checklist_items: 校验结果 → 「99_校验结果」sheet

    返回 openpyxl Workbook 对象，调用者可继续追加章节 sheet。
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 00_填报说明
    ws_instructions = wb.create_sheet("00_填报说明")
    build_semantic_instruction_sheet(
        ws_instructions,
        project_name=project_name,
        year=year,
        exporter_name=exporter_name,
        section_count=len(sections),
    )

    # 01_章节清单
    ws_section_list = wb.create_sheet("01_章节清单")
    build_semantic_section_list_sheet(ws_section_list, sections)

    # 政策条款 (Task 15.1)
    if policy_clauses is not None:
        ws_policy = wb.create_sheet("政策条款")
        build_policy_clauses_sheet(ws_policy, policy_clauses)

    # 科目披露 (Task 15.2)
    if disclosure_tables is not None:
        ws_disclosure = wb.create_sheet("科目披露")
        build_account_disclosure_sheet(ws_disclosure, disclosure_tables)

    # 关联方 (Task 15.3)
    if related_parties is not None:
        ws_rp = wb.create_sheet("关联方")
        build_related_party_sheet(
            ws_rp,
            parties=related_parties.get("parties", []),
            transactions=related_parties.get("transactions", []),
            balances=related_parties.get("balances", []),
        )

    # 99_校验结果 (Task 15.4)
    if checklist_items is not None:
        ws_validation = wb.create_sheet("99_校验结果")
        build_validation_results_sheet(ws_validation, checklist_items)

    return wb
