"""表格类底稿（univer 类：F-审定表/F-明细表/G-测算 等）网格提取服务

从底稿 xlsx 模板读取单元格 + 合并区域 + 列宽 + **真实样式**（填充色/加粗/对齐/
字号/数字格式），构建只读 HTML 网格数据，供 GtGridSheet.vue 还原 Excel 外观。

动机（2026-06-02 修复）：混合底稿（含 HTML sheet + univer sheet，如 D1 同时有
A 程序表/B 目录/F 审定表）整本路由到 GtWpRenderer，但其 univer 分支只是一个
**死占位**「数据尚未导入」，模板里的审定表/明细表结构完全不显示。本服务让 univer
sheet 在 HTML 渲染器内也能显示模板网格内容（只读，且按 Excel 模板样式还原）。

样式还原（2026-06-02 用户要求按 Excel 模板样式）：提取真实 fill（rgb 直取 /
theme+tint 经 _resolve_theme_color 解析）+ bold + align + font_size + number_format，
不再用文字正则启发式（合计/小计等）。

纯函数无 DB 副作用，便于单测。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 默认 Office 主题配色（clrScheme XML 顺序：dk1 lt1 dk2 lt2 accent1..6 hlink folHlink）
_DEFAULT_SCHEME = [
    "000000", "FFFFFF", "1F497D", "EEECE1",
    "4F81BD", "C0504D", "9BBB59", "8064A2",
    "4BACC6", "F79646", "0000FF", "800080",
]

# openpyxl cell.fill.fgColor.theme 索引 → clrScheme XML 索引映射
# Excel 渲染层 theme：0=lt1 1=dk1 2=lt2 3=dk2 4=accent1…（lt1/dk1 相对 clrScheme 互换）
_THEME_TO_SCHEME = {0: 1, 1: 0, 2: 3, 3: 2, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8, 9: 9}


def _parse_scheme_from_workbook(wb: Any) -> list[str]:
    """从工作簿 theme XML 解析 clrScheme 12 色（失败回退默认 Office 配色）。"""
    import re

    try:
        raw = wb.loaded_theme
        if not raw:
            return _DEFAULT_SCHEME
        xml = raw.decode("utf-8") if isinstance(raw, bytes) else raw
        m = re.search(r"<a:clrScheme.*?</a:clrScheme>", xml, re.S)
        if not m:
            return _DEFAULT_SCHEME
        colors = re.findall(
            r"<a:(?:srgbClr|sysClr)[^>]*?(?:val|lastClr)=\"([0-9A-Fa-f]{6})\"",
            m.group(0),
        )
        return colors if len(colors) >= 12 else _DEFAULT_SCHEME
    except Exception:
        return _DEFAULT_SCHEME


def _apply_tint(hex_rgb: str, tint: float) -> str:
    """对 6 位 hex 应用 Excel tint（>0 变亮，<0 变暗），返回 6 位 hex（不含 #）。"""
    try:
        r = int(hex_rgb[0:2], 16)
        g = int(hex_rgb[2:4], 16)
        b = int(hex_rgb[4:6], 16)
    except (ValueError, IndexError):
        return hex_rgb

    def _adj(c: int) -> int:
        if tint < 0:
            v = c * (1.0 + tint)
        else:
            v = c * (1.0 - tint) + 255.0 * tint
        return max(0, min(255, int(round(v))))

    return f"{_adj(r):02X}{_adj(g):02X}{_adj(b):02X}"


def _resolve_fill(cell: Any, scheme: list[str]) -> str | None:
    """解析单元格填充色 → '#RRGGBB'（无填充 / 解析失败 → None）。"""
    fill = cell.fill
    if not fill or not getattr(fill, "patternType", None):
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    try:
        if fg.type == "rgb" and fg.rgb and isinstance(fg.rgb, str):
            # openpyxl rgb 形如 'FFE4DFEC'（前两位 alpha）
            rgb = fg.rgb[-6:]
            if rgb == "000000" and fill.patternType == "solid" and fg.rgb == "00000000":
                return None
            return f"#{rgb}"
        if fg.type == "theme":
            scheme_idx = _THEME_TO_SCHEME.get(int(fg.theme), int(fg.theme))
            if 0 <= scheme_idx < len(scheme):
                base = scheme[scheme_idx]
                return f"#{_apply_tint(base, float(fg.tint or 0.0))}"
    except Exception:
        return None
    return None


def _resolve_font_color(cell: Any) -> str | None:
    """解析字体颜色 → '#RRGGBB'（仅 rgb 类型，其余返 None 用默认）。"""
    try:
        fc = cell.font.color
        if fc and fc.type == "rgb" and fc.rgb and isinstance(fc.rgb, str):
            rgb = fc.rgb[-6:]
            if rgb in ("000000", "FFFFFF"):
                # 黑白用默认（避免暗色主题下不可见）
                return None
            return f"#{rgb}"
    except Exception:
        pass
    return None


def _cell_value(value: Any) -> Any:
    """规整单元格值：公式串(=开头)按公式处理由上层决定，这里只规整非公式标量。"""
    if value is None:
        return None
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _is_accounting_format(numfmt: str | None) -> bool:
    """判断是否会计/数字格式（空值应显示「-」占位）。"""
    if not numfmt:
        return False
    return "#,##0" in numfmt or "0.00" in numfmt


def _col_letter(c: int) -> str:
    """列号(1-based) → 字母（1→A, 27→AA）。"""
    s = ""
    while c > 0:
        c, rem = divmod(c - 1, 26)
        s = chr(65 + rem) + s
    return s


def extract_grid_from_sheet(ws: Any, scheme: list[str] | None = None, *, max_scan_rows: int = 200) -> dict:
    """从 openpyxl worksheet 提取网格数据 + 样式。

    规则（2026-06-02 用户要求）：
    - 跳过标题行（表头区：致同会计师事务所/表名/索引号/页次，由统一 preparation header 处理）
    - 裁剪空列（仅输出有实际内容的列范围）
    - 裁剪尾部空行
    - 不输出填充色（fill）——用户明确不需要背景色

    Returns:
        {
          "cells": { "A1": {v, r, c, style:{bold,align,font_size,font_color,numeric}}, ... },
          "merged_cells": [ {s:{r,c}, e:{r,c}}, ... ],
          "col_widths": { "A": 19.7, ... },
          "max_row": int, "max_col": int,
        }
        r/c 1-based，但已偏移（跳过标题行后重编号）。仅输出有值或有格式的单元格。
    """
    if scheme is None:
        scheme = _DEFAULT_SCHEME

    raw_max_row = min(ws.max_row or 0, max_scan_rows)
    raw_max_col = ws.max_column or 0
    if raw_max_row == 0 or raw_max_col == 0:
        return {"cells": {}, "merged_cells": [], "col_widths": {}, "max_row": 0, "max_col": 0}

    # ─── 1. 找到数据表起始行（跳过标题区）────────────────────────────────
    # 标题区 = 表头前几行（致同/表名/索引号/页次），数据表从第一个含"项目"/"期初"/"序号" 的行开始
    data_start_row = 1
    for r in range(1, min(raw_max_row, 15) + 1):
        for c in range(1, min(raw_max_col, 5) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() in ("项目", "序号"):
                data_start_row = r
                break
        if data_start_row > 1:
            break

    # ─── 2. 扫描有内容的列范围 ────────────────────────────────────────────
    cols_with_content: set[int] = set()
    for r in range(data_start_row, raw_max_row + 1):
        for c in range(1, raw_max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "" and not (isinstance(v, str) and v.startswith("=")):
                cols_with_content.add(c)
            elif isinstance(v, str) and v.startswith("="):
                # 公式列也算有内容（即使值为 0/None）
                cols_with_content.add(c)

    if not cols_with_content:
        return {"cells": {}, "merged_cells": [], "col_widths": {}, "max_row": 0, "max_col": 0}

    effective_max_col = max(cols_with_content)

    # ─── 3. 扫描有内容的最后行（裁剪尾部空行）────────────────────────────
    effective_max_row = data_start_row
    for r in range(raw_max_row, data_start_row - 1, -1):
        has_content = False
        for c in range(1, effective_max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "":
                has_content = True
                break
        if has_content:
            effective_max_row = r
            break

    # ─── 4. 检测列语义（公式标注用）────────────────────────────────────────
    # 扫描表头行（data_start_row）和子表头行（data_start_row+1）确定各列的业务含义
    # 审定表布局：项目 | 未审数 | 账项调整 | 重分类调整 | 审定数 | 未审数 | ... | 变动额 | 变动率 | 原因分析
    _COL_SEMANTICS: dict[str, str] = {
        "未审数": "tb_fetch",       # TB() 从试算表取数
        "账项调整": "adj_sum",      # 从调整分录汇总
        "重分类调整": "reclass_sum", # 从重分类调整汇总
        "审定数": "computed_sum",   # =未审+调整+重分类
        "变动额": "computed_diff",  # =期末审定-期初审定
        "变动率": "computed_rate",  # =变动额/期初审定
        "原因分析": "user_input",   # 用户手填
    }
    col_formula_map: dict[int, str] = {}  # col_number → formula_hint
    # 扫描 sub-header row (data_start_row + 1)
    sub_hdr_row = data_start_row + 1 if data_start_row + 1 <= effective_max_row else data_start_row
    for c in range(1, effective_max_col + 1):
        # 先查 sub-header，再查 header
        for scan_row in (sub_hdr_row, data_start_row):
            txt = ws.cell(row=scan_row, column=c).value
            if isinstance(txt, str):
                txt_clean = txt.strip().replace("\n", "")
                for keyword, hint in _COL_SEMANTICS.items():
                    if keyword in txt_clean:
                        col_formula_map[c] = hint
                        break
            if c in col_formula_map:
                break
    # 项目列（col 1）是 label，不标注
    col_formula_map.pop(1, None)

    # ─── 5. 提取单元格（重编号：数据表首行 → row 1）──────────────────────
    row_offset = data_start_row - 1
    # 确定数据行起始（跳过表头行和子表头行）
    # 表头行 = data_start_row, 子表头行 = data_start_row+1
    # 数据行 = 从第一个非表头行开始（通常 data_start_row+2 或更早）
    header_rows_count = 1
    if sub_hdr_row > data_start_row:
        # 检查子表头行是否真有内容（确认是第二级表头还是直接数据）
        sub_has_content = False
        for c in range(2, effective_max_col + 1):
            sv = ws.cell(row=sub_hdr_row, column=c).value
            if isinstance(sv, str) and sv.strip() and not sv.startswith("="):
                sub_has_content = True
                break
        if sub_has_content:
            header_rows_count = 2

    cells: dict[str, dict] = {}
    for r in range(data_start_row, effective_max_row + 1):
        for c in range(1, effective_max_col + 1):
            cell = ws.cell(row=r, column=c)
            raw = cell.value
            v = _cell_value(raw)
            is_formula = isinstance(v, str) and v.startswith("=")
            display_v = None if is_formula else v
            has_content = display_v is not None and display_v != ""

            if not has_content:
                # Include numeric-format cells even if empty (display "-" in frontend)
                if not _is_accounting_format(cell.number_format):
                    continue

            # 构建样式（不输出 fill —— 用户明确不需要背景色）
            style: dict[str, Any] = {}
            try:
                if cell.font and cell.font.bold:
                    style["bold"] = True
                if cell.font and cell.font.sz:
                    style["font_size"] = float(cell.font.sz)
            except Exception:
                pass
            fcolor = _resolve_font_color(cell)
            if fcolor:
                style["font_color"] = fcolor
            try:
                if cell.alignment and cell.alignment.horizontal:
                    style["align"] = cell.alignment.horizontal
            except Exception:
                pass
            if _is_accounting_format(cell.number_format):
                style["numeric"] = True

            # 公式标注：数据行（非表头行）的公式列加 formula_hint
            is_data_row = (r - data_start_row) >= header_rows_count
            if is_data_row and c in col_formula_map:
                style["formula_hint"] = col_formula_map[c]

            new_r = r - row_offset
            new_coord = f"{_col_letter(c)}{new_r}"
            cells[new_coord] = {
                "v": display_v if has_content else "",
                "r": new_r,
                "c": c,
                "style": style,
            }

    # ─── 5. 合并区域（仅保留在有效范围内的，并偏移行号）─────────────────
    merged: list[dict] = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row < data_start_row:
            continue  # 标题区合并跳过
        if mr.min_row > effective_max_row:
            continue
        if mr.min_col > effective_max_col:
            continue
        merged.append({
            "s": {"r": mr.min_row - row_offset, "c": mr.min_col},
            "e": {"r": min(mr.max_row, effective_max_row) - row_offset, "c": min(mr.max_col, effective_max_col)},
        })

    # ─── 6. 列宽（仅有效列）─────────────────────────────────────────────
    col_widths: dict[str, float] = {}
    for c in range(1, effective_max_col + 1):
        letter = _col_letter(c)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width:
            col_widths[letter] = round(float(dim.width), 2)

    output_max_row = effective_max_row - row_offset
    return {
        "cells": cells,
        "merged_cells": merged,
        "col_widths": col_widths,
        "max_row": output_max_row,
        "max_col": effective_max_col,
        "column_meta": {_col_letter(c): hint for c, hint in col_formula_map.items()},
        "header_rows": header_rows_count,
    }


def extract_grid(file_path: str | Path, sheet_name: str) -> dict:
    """读取 xlsx 文件指定 sheet，提取只读网格数据 + 样式（纯函数，无 DB）。

    文件不存在 / 空 / sheet 缺失 / 解析失败 → 返回空网格（降级，不抛异常）。
    """
    import openpyxl

    empty = {"cells": {}, "merged_cells": [], "col_widths": {}, "max_row": 0, "max_col": 0}

    fp = Path(file_path)
    if not fp.exists() or fp.stat().st_size == 0:
        return empty

    try:
        wb = openpyxl.load_workbook(str(fp), read_only=False, data_only=True)
    except Exception as e:
        logger.warning("extract_grid: 加载 xlsx 失败 %s: %s", fp, e)
        return empty

    try:
        if sheet_name not in wb.sheetnames:
            return empty
        scheme = _parse_scheme_from_workbook(wb)
        ws = wb[sheet_name]
        return extract_grid_from_sheet(ws, scheme)
    except Exception as e:
        logger.warning("extract_grid: 解析 sheet 失败 %s/%s: %s", fp, sheet_name, e)
        return empty
    finally:
        wb.close()
