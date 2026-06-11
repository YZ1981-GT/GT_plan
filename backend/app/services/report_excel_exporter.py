"""报表 Excel 导出引擎 — 基于致同模板填充

基于 `审计报告模板/{版本}/{范围}/` 下的 xlsx 模板文件，
使用 openpyxl 复制模板后填入数据（保留原有格式/边框/字体/列宽）。

生成 4 个 Sheet：资产负债表、利润表、现金流量表、所有者权益变动表。

Requirements: 3.1-3.15
"""
from __future__ import annotations

import copy
import json
import logging
import os
import re
from decimal import Decimal
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.core import Project
from app.models.report_models import FinancialReport, FinancialReportType

logger = logging.getLogger(__name__)

# Report type → Chinese sheet name（manifest sheet_aliases 优先，此为回退子串匹配）
REPORT_TYPE_SHEET_NAMES: dict[str, str] = {
    "balance_sheet": "资产负债表",
    "income_statement": "利润表",
    "cash_flow_statement": "现金流量表",
    "equity_statement": "所有者权益变动表",
    "asset_impairment": "资产减值损失明细表",
    "impairment_provision": "减值准备表",
}

# Amount number format
AMOUNT_FORMAT = '#,##0.00'
NEGATIVE_FORMAT = '#,##0.00;[Red](#,##0.00)'

# Font for amounts
AMOUNT_FONT = Font(name="Arial Narrow", size=10)
BOLD_AMOUNT_FONT = Font(name="Arial Narrow", size=10, bold=True)

# Indentation: 2 Chinese character widths per level ≈ 4 regular chars
INDENT_CHARS_PER_LEVEL = 4

# Inline placeholder patterns (design §6 双轨填充)
#   {{row:BS-002:current}} / {{row:BS-002:prior}}
#   {{row:BS-002:current:parent}} / {{row:BS-002:prior:parent}}（母公司个别列）
_ROW_PLACEHOLDER_RE = re.compile(
    r"\{\{row:([^:}]+):(current|prior)(?::(parent))?\}\}"
)
#   {{note_ref:BS-002}}
_NOTE_REF_RE = re.compile(r"\{\{note_ref:([^}]+)\}\}")
#   {{imp:IMP-001:opening_balance}} — 资产减值表二维占位
_IMP_PLACEHOLDER_RE = re.compile(
    r"\{\{imp:([^:}]+):([^}]+)\}\}"
)
#   header placeholders embedded in text: {{company_full_name}} 等
_HEADER_PLACEHOLDER_RE = re.compile(r"\{\{([a-zA-Z_][a-zA-Z0-9_]*)\}\}")

# Header placeholder keys handled by header replacement track
_HEADER_KEYS = frozenset(
    {"company_full_name", "period_end_date", "audit_year", "currency_unit"}
)


@lru_cache(maxsize=1)
def _load_cell_mapping() -> dict[str, Any]:
    """加载 ``cell_mapping.json``（resolve 相对模板 data 目录）.

    结构::

        {"version", "variants": {variant_key: {"headers": {...}, "rows": {...}}}}
    """
    from app.services.template_manifest_loader import resolve_template_base_dir

    path = resolve_template_base_dir() / "cell_mapping.json"
    if not path.is_file():
        logger.info("cell_mapping.json not found at %s", path)
        return {}
    try:
        with path.open(encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:  # pragma: no cover - defensive
        logger.warning("Failed to load cell_mapping.json: %s", e)
        return {}


@lru_cache(maxsize=1)
def _load_note_ref_mapping() -> dict[str, dict[str, dict]]:
    """加载 ``report_row_note_mapping.json``（row_code → section_code 映射）.

    返回结构: {variant_key: {row_code: {"section_code": "八、1", ...}}}
    """
    from app.services.template_manifest_loader import resolve_template_base_dir

    path = resolve_template_base_dir() / "report_row_note_mapping.json"
    if not path.is_file():
        logger.info("report_row_note_mapping.json not found at %s", path)
        return {}
    try:
        with path.open(encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:  # pragma: no cover - defensive
        logger.warning("Failed to load report_row_note_mapping.json: %s", e)
        return {}


class ReportExcelExporter:
    """报表 Excel 导出引擎

    Design decision D2: 基于模板复制填充（非从零生成）。
    致同模板格式复杂（合并单元格/条件格式/打印设置），从零生成无法完美复现。
    如果模板文件不存在，则从零生成（fallback）。
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export(
        self,
        project_id: UUID,
        year: int,
        mode: str = "audited",
        report_types: list[str] | None = None,
        include_prior_year: bool = True,
    ) -> BytesIO:
        """导出报表 Excel

        Args:
            project_id: 项目 ID
            year: 年度
            mode: "audited" 或 "unadjusted"
            report_types: 指定导出哪些报表（默认全部 4 张）
            include_prior_year: 是否包含上年对比列

        Returns:
            BytesIO containing the xlsx file
        """
        # 1. Load project info
        project = await self._get_project(project_id)
        company_name = project.name if project else "未知公司"

        # 2. Determine template key
        template_type = getattr(project, "template_type", None) or "soe"
        report_scope = getattr(project, "report_scope", None) or "standalone"
        template_key = f"{template_type}_{report_scope}"

        # 3. Determine which report types to export
        if report_types:
            types_to_export = report_types
        else:
            types_to_export = list(REPORT_TYPE_SHEET_NAMES.keys())

        # 4. Load report data from DB
        report_data = await self._load_report_data(project_id, year, types_to_export)

        # 4b. Load parent (母公司个别) report data for consolidated :parent columns.
        #     母分汇总已作为现成 FinancialReport 存储——读「上级代码」匹配的 standalone
        #     项目的报表行，按 row_code 建索引；缺失则留空（不崩）。
        parent_row_index = await self._load_parent_row_index(
            project, year, types_to_export
        )

        # 5. Try to load template, fallback to programmatic generation
        wb = self._load_template(template_key)
        if wb is None:
            # Fallback: generate from scratch
            wb = self._create_workbook_from_scratch(
                types_to_export, report_data, company_name, year,
                include_prior_year, mode,
            )
        else:
            # Fill template with data
            wb = self._fill_template(
                wb,
                types_to_export,
                report_data,
                company_name,
                year,
                include_prior_year,
                mode,
                template_key=template_key,
                parent_row_index=parent_row_index,
            )

        # 6. Write to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # -----------------------------------------------------------------------
    # Private helpers
    # -----------------------------------------------------------------------

    async def _get_project(self, project_id: UUID) -> Any:
        """Load project from DB."""
        result = await self.db.execute(
            select(Project).where(Project.id == project_id)
        )
        return result.scalar_one_or_none()

    async def _load_report_data(
        self, project_id: UUID, year: int, report_types: list[str]
    ) -> dict[str, list[dict]]:
        """Load generated report rows from financial_report table."""
        data: dict[str, list[dict]] = {}
        for rt in report_types:
            try:
                report_type_enum = FinancialReportType(rt)
            except ValueError:
                continue

            result = await self.db.execute(
                select(FinancialReport)
                .where(
                    FinancialReport.project_id == project_id,
                    FinancialReport.year == year,
                    FinancialReport.report_type == report_type_enum,
                    FinancialReport.is_deleted == False,  # noqa: E712
                )
                .order_by(FinancialReport.row_code)
            )
            rows = result.scalars().all()
            data[rt] = [
                {
                    "row_code": r.row_code,
                    "row_name": r.row_name or "",
                    "current_period_amount": r.current_period_amount,
                    "prior_period_amount": r.prior_period_amount,
                    "indent_level": r.indent_level,
                    "is_total_row": r.is_total_row,
                    "formula_used": r.formula_used,
                    "source_accounts": r.source_accounts,
                }
                for r in rows
            ]
        return data

    async def _load_parent_row_index(
        self, project: Any, year: int, report_types: list[str]
    ) -> dict[str, dict]:
        """加载母公司个别报表行索引（``:parent`` 占位/``current_parent`` 坐标用）.

        母公司个别数（公司列）取自「上级代码」（``Project.parent_company_code``）
        匹配的 standalone 项目（``Project.company_code == parent_company_code``，
        未软删）的已审定 ``FinancialReport`` 行，按 ``row_code`` 建索引。

        母分汇总已作为现成 FinancialReport 存储——直接读匹配项目的报表，
        无需实时聚合。找不到母公司项目则返回空 dict（导出时该列留空，不崩）。
        """
        parent_code = getattr(project, "parent_company_code", None)
        if not parent_code:
            return {}
        try:
            result = await self.db.execute(
                select(Project).where(
                    Project.company_code == parent_code,
                    Project.is_deleted == False,  # noqa: E712
                )
            )
            parent_project = result.scalars().first()
        except Exception as e:  # pragma: no cover - defensive
            logger.warning("Parent project lookup failed for code %s: %s", parent_code, e)
            return {}
        if parent_project is None:
            logger.info(
                "No parent project found for parent_company_code=%s; "
                "公司(母公司个别) columns will be left blank",
                parent_code,
            )
            return {}

        parent_data = await self._load_report_data(
            parent_project.id, year, report_types
        )
        index: dict[str, dict] = {}
        for _rt, rows in parent_data.items():
            for r in rows:
                code = r.get("row_code")
                if code:
                    index[code] = r
        return index

    def _load_template(self, template_key: str) -> Workbook | None:
        """从 ``TemplateManifestLoader`` 加载 xlsx 模板；缺失时返回 None."""
        from app.services.template_manifest_loader import get_template_manifest_loader

        parts = template_key.split("_", 1)
        if len(parts) != 2:
            logger.warning("Invalid template_key %s", template_key)
            return None
        template_type, report_scope = parts
        loader = get_template_manifest_loader()
        try:
            entry = loader.resolve_financial_statements(template_type, report_scope)
        except KeyError:
            logger.info("No manifest entry for template_key %s", template_key)
            return None
        if not entry.exists:
            logger.info("Template file missing: %s", entry.abs_path)
            return None
        try:
            wb = load_workbook(str(entry.abs_path))
            logger.info("Loaded template from %s", entry.abs_path)
            return wb
        except Exception as e:
            logger.warning("Failed to load template %s: %s", entry.abs_path, e)
            return None

    def _resolve_sheet(
        self, wb: Workbook, report_type: str, template_key: str
    ):
        """按 manifest sheet_aliases 或中文子串匹配定位 worksheet（返回第一个）."""
        sheets = self._resolve_sheets(wb, report_type, template_key)
        return sheets[0] if sheets else None

    def _resolve_sheets(
        self, wb: Workbook, report_type: str, template_key: str
    ) -> list:
        """按 manifest sheet_aliases 或中文子串匹配定位所有关联 worksheet.

        资产负债表有主表+续表（负债权益），alias 为列表时遍历全部匹配。
        """
        from app.services.template_manifest_loader import get_template_manifest_loader

        aliases = get_template_manifest_loader().get_sheet_aliases(template_key)
        alias = aliases.get(report_type)
        matched: list = []
        if alias:
            names = alias if isinstance(alias, list) else [alias]
            for name in names:
                if name in wb.sheetnames:
                    matched.append(wb[name])
        if matched:
            return matched
        # Fallback: 中文子串匹配
        sheet_hint = REPORT_TYPE_SHEET_NAMES.get(report_type)
        if sheet_hint:
            for name in wb.sheetnames:
                if sheet_hint in name:
                    matched.append(wb[name])
        return matched

    def _fill_template(
        self,
        wb: Workbook,
        types_to_export: list[str],
        report_data: dict[str, list[dict]],
        company_name: str,
        year: int,
        include_prior_year: bool,
        mode: str,
        *,
        template_key: str = "soe_standalone",
        parent_row_index: dict[str, dict] | None = None,
    ) -> Workbook:
        """按 design §6「双轨填充」填充模板工作簿，保留模板格式.

        填充顺序（每个 sheet）：
          1. 扫描内联 ``{{row:CODE:current/prior}}`` / ``{{note_ref:CODE}}`` 占位符
             → 按 ``row_code`` 从 ``report_data`` 取值回填该单元格（替换占位文本）。
          2. 表头 ``{{company_full_name}}`` 等占位符替换（合并格写左上角）。
          3. 若 sheet 无内联 ``{{row:}}`` 占位符 → 回退 ``cell_mapping.json`` 坐标。
          4. 全程跳过公式格（``data_type=='f'`` 或值以 ``=`` 开头）。
          5. ``fill_empty_as``：无数据时 ``blank`` 留空 / ``zero`` 填 0。
        """
        header_values = self._build_header_values(company_name, year, mode)
        parent_index = parent_row_index or {}

        # Build {row_code: row} index across all report types for O(1) lookup
        row_index: dict[str, dict] = {}
        for _rt, _rows in report_data.items():
            for r in _rows:
                code = r.get("row_code")
                if code:
                    row_index[code] = r

        # Build note_ref index: {row_code: "八、1"} for this template variant
        note_ref_index = self._build_note_ref_index(template_key)

        fill_empty_map = self._build_fill_empty_map(template_key)

        for rt in types_to_export:
            rows = report_data.get(rt, [])
            sheets = self._resolve_sheets(wb, rt, template_key)

            if not sheets:
                # Sheet not in template → create from scratch (only if we have data)
                if not rows:
                    continue
                fallback_title = REPORT_TYPE_SHEET_NAMES.get(rt, rt)
                ws = wb.create_sheet(title=fallback_title)
                self._write_sheet_from_scratch(
                    ws, rows, company_name, year, include_prior_year, mode, rt
                )
                continue

            # 遍历所有匹配 sheet（如资产负债表主表+续表负债权益）
            any_found_inline = False
            for ws in sheets:
                # Track 1+2: inline placeholders + header replacement
                found_inline = self._fill_by_placeholders(
                    ws, row_index, header_values, include_prior_year, fill_empty_map,
                    parent_index, note_ref_index,
                )
                if found_inline:
                    any_found_inline = True

            # Track 3: fallback to cell_mapping coordinates when no inline placeholders
            # (仅当所有 sheet 都无内联占位符时才回退 cell_mapping)
            if not any_found_inline:
                for ws in sheets:
                    self._fill_by_cell_mapping(
                        ws,
                        rt,
                        template_key,
                        row_index,
                        header_values,
                        include_prior_year,
                        fill_empty_map,
                        parent_index,
                    )

        # Remove sheets not in export list
        sheets_to_keep = set()
        for rt in types_to_export:
            resolved = self._resolve_sheets(wb, rt, template_key)
            if resolved:
                for ws in resolved:
                    sheets_to_keep.add(ws.title)
            else:
                # Fallback: 中文子串匹配
                sn = REPORT_TYPE_SHEET_NAMES.get(rt, "")
                if sn:
                    for name in wb.sheetnames:
                        if sn in name:
                            sheets_to_keep.add(name)

        # Don't remove sheets if we couldn't match any (safety)
        if sheets_to_keep:
            for name in list(wb.sheetnames):
                if name not in sheets_to_keep:
                    del wb[name]

        return wb

    # -----------------------------------------------------------------------
    # Placeholder / cell_mapping driven fill (design §6 双轨填充)
    # -----------------------------------------------------------------------

    def _build_header_values(
        self, company_name: str, year: int, mode: str
    ) -> dict[str, str]:
        """构建表头占位符 → 值映射."""
        return {
            "company_full_name": company_name,
            "period_end_date": f"{year}年12月31日",
            "audit_year": str(year),
            "currency_unit": "人民币元",
        }

    def _build_fill_empty_map(self, template_key: str) -> dict[str, str]:
        """从 cell_mapping 读取每个 row_code 的 ``fill_empty_as``（默认 blank）."""
        mapping = _load_cell_mapping()
        variant = mapping.get("variants", {}).get(template_key, {})
        result: dict[str, str] = {}
        for code, info in variant.get("rows", {}).items():
            if isinstance(info, dict):
                result[code] = info.get("fill_empty_as", "blank")
        return result

    @staticmethod
    def _build_note_ref_index(template_key: str) -> dict[str, str]:
        """构建 {row_code: section_code} 映射，用于 ``{{note_ref:CODE}}`` 填充.

        从 ``report_row_note_mapping.json`` 读取对应 variant 的映射。
        返回如 ``{"BS-002": "八、1", "BS-005": "八、2"}``。
        """
        mapping = _load_note_ref_mapping()
        variant = mapping.get(template_key, {})
        if not isinstance(variant, dict):
            return {}
        return {
            code: info["section_code"]
            for code, info in variant.items()
            if isinstance(info, dict) and "section_code" in info
        }

    @staticmethod
    def _is_formula_cell(cell) -> bool:
        """判断单元格是否为公式格（不可覆盖）."""
        if getattr(cell, "data_type", None) == "f":
            return True
        val = cell.value
        return isinstance(val, str) and val.startswith("=")

    def _resolve_amount(self, row: dict | None, period: str) -> float | None:
        """取 current/prior 金额（float 或 None）."""
        if not row:
            return None
        key = "current_period_amount" if period == "current" else "prior_period_amount"
        amount = row.get(key)
        if amount is None:
            return None
        try:
            return float(amount)
        except (TypeError, ValueError):
            return None

    def _resolve_fill_value(
        self,
        row: dict | None,
        period: str,
        row_code: str,
        fill_empty_map: dict[str, str],
    ):
        """解析回填值，应用 ``fill_empty_as``.

        返回 ``(should_write, value)``；should_write=False 表示留空（blank）。
        """
        amount = self._resolve_amount(row, period)
        if amount is not None:
            return True, amount
        # No data → honor fill_empty_as
        mode = fill_empty_map.get(row_code, "blank")
        if mode == "zero":
            return True, 0
        return False, None  # blank → clear placeholder, leave cell empty

    def _resolve_imp_value(self, row: dict | None, col_key: str) -> float | None:
        """从减值表行的 source_accounts JSONB 取多列数据.

        减值表每行在 FinancialReport.source_accounts 存储各列金额：
        {"opening_balance": 1000, "provision": 200, "reversal": 50, ...}

        若 source_accounts 无该列，尝试从 current_period_amount 回退
        （col_key 为 closing_balance 时取 current_period_amount）。
        """
        if not row:
            return None
        # 优先从 source_accounts JSONB 取分列数据
        source = row.get("source_accounts")
        if isinstance(source, dict):
            val = source.get(col_key)
            if val is not None:
                try:
                    return float(val)
                except (TypeError, ValueError):
                    pass
        # 回退：closing_balance → current_period_amount
        if col_key == "closing_balance":
            amt = row.get("current_period_amount")
            if amt is not None:
                try:
                    return float(amt)
                except (TypeError, ValueError):
                    pass
        # opening_balance → prior_period_amount
        if col_key == "opening_balance":
            amt = row.get("prior_period_amount")
            if amt is not None:
                try:
                    return float(amt)
                except (TypeError, ValueError):
                    pass
        return None

    def _fill_by_placeholders(
        self,
        ws,
        row_index: dict[str, dict],
        header_values: dict[str, str],
        include_prior_year: bool,
        fill_empty_map: dict[str, str],
        parent_row_index: dict[str, dict] | None = None,
        note_ref_index: dict[str, str] | None = None,
    ) -> bool:
        """扫描 sheet 内联占位符并回填；返回是否发现 ``{{row:}}``/``{{note_ref:}}``/``{{imp:}}`` 占位符.

        - ``{{row:CODE:current/prior}}`` → 按 row_code 取值写入该格（替换占位文本）。
        - ``{{row:CODE:current/prior:parent}}`` → 取母公司个别（公司列）值写入。
        - ``{{note_ref:CODE}}`` → 填入附注章节编号（如 "八、1"）。
        - ``{{imp:CODE:col_key}}`` → 资产减值表多列取数（从 source_accounts JSONB）。
        - 表头 ``{{key}}`` → 子串替换（合并格非锚点成员 value 为 None 自动跳过）。
        """
        parent_index = parent_row_index or {}
        note_refs = note_ref_index or {}
        found_inline = False
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or "{{" not in val:
                    continue
                if self._is_formula_cell(cell):
                    continue

                stripped = val.strip()

                # --- Track 1a: row placeholder (whole-cell) ---
                m_row = _ROW_PLACEHOLDER_RE.fullmatch(stripped)
                if m_row:
                    found_inline = True
                    code, period, is_parent = (
                        m_row.group(1),
                        m_row.group(2),
                        m_row.group(3),
                    )
                    if period == "prior" and not include_prior_year:
                        cell.value = None
                        continue
                    source = parent_index if is_parent else row_index
                    should_write, value = self._resolve_fill_value(
                        source.get(code), period, code, fill_empty_map
                    )
                    cell.value = value if should_write else None
                    continue

                # --- Track 1b: note_ref placeholder ---
                m_note = _NOTE_REF_RE.fullmatch(stripped)
                if m_note:
                    found_inline = True
                    note_code = m_note.group(1)
                    # Fill with section_code display value (e.g. "八、1")
                    display = note_refs.get(note_code)
                    cell.value = display  # None if no mapping → cell cleared
                    continue

                # --- Track 1c: imp placeholder (资产减值表多列) ---
                m_imp = _IMP_PLACEHOLDER_RE.fullmatch(stripped)
                if m_imp:
                    found_inline = True
                    imp_code, col_key = m_imp.group(1), m_imp.group(2)
                    imp_row = row_index.get(imp_code)
                    imp_value = self._resolve_imp_value(imp_row, col_key)
                    cell.value = imp_value
                    continue

                # --- Track 2: header placeholder(s) embedded in text ---
                if _HEADER_PLACEHOLDER_RE.search(val):
                    new_val = self._replace_header_placeholders(val, header_values)
                    if new_val != val:
                        cell.value = new_val

        return found_inline

    @staticmethod
    def _replace_header_placeholders(text: str, header_values: dict[str, str]) -> str:
        """替换文本中的表头占位符 ``{{key}}``；未知 key 原样保留."""

        def _sub(match: re.Match) -> str:
            key = match.group(1)
            if key in header_values:
                return str(header_values[key])
            return match.group(0)

        return _HEADER_PLACEHOLDER_RE.sub(_sub, text)

    def _fill_by_cell_mapping(
        self,
        ws,
        report_type: str,
        template_key: str,
        row_index: dict[str, dict],
        header_values: dict[str, str],
        include_prior_year: bool,
        fill_empty_map: dict[str, str],
        parent_row_index: dict[str, dict] | None = None,
    ) -> None:
        """回退：按 ``cell_mapping.json`` 坐标填充（无内联占位符的 sheet）."""
        parent_index = parent_row_index or {}
        mapping = _load_cell_mapping()
        variant = mapping.get("variants", {}).get(template_key, {})
        if not variant:
            return

        # Headers block: {report_type: {field: cellref}}
        headers = variant.get("headers", {}).get(report_type, {})
        for field, coord in headers.items():
            if field in header_values:
                self._write_header_cell(ws, coord, header_values[field])

        # Rows block: {code: {sheet, current, prior, current_parent,
        #                      prior_parent, fill_empty_as, ...}}
        for code, info in variant.get("rows", {}).items():
            if not isinstance(info, dict):
                continue
            if info.get("sheet") != report_type:
                continue
            row = row_index.get(code)
            parent_row = parent_index.get(code)

            cur_ref = info.get("current")
            if cur_ref:
                should_write, value = self._resolve_fill_value(
                    row, "current", code, fill_empty_map
                )
                self._write_mapped_cell(ws, cur_ref, value if should_write else None)

            cur_parent_ref = info.get("current_parent")
            if cur_parent_ref:
                should_write, value = self._resolve_fill_value(
                    parent_row, "current", code, fill_empty_map
                )
                self._write_mapped_cell(
                    ws, cur_parent_ref, value if should_write else None
                )

            if include_prior_year:
                prior_ref = info.get("prior")
                if prior_ref:
                    should_write, value = self._resolve_fill_value(
                        row, "prior", code, fill_empty_map
                    )
                    self._write_mapped_cell(
                        ws, prior_ref, value if should_write else None
                    )

                prior_parent_ref = info.get("prior_parent")
                if prior_parent_ref:
                    should_write, value = self._resolve_fill_value(
                        parent_row, "prior", code, fill_empty_map
                    )
                    self._write_mapped_cell(
                        ws, prior_parent_ref, value if should_write else None
                    )

    def _write_mapped_cell(self, ws, coord: str, value) -> None:
        """按坐标写入数据格，跳过公式格与合并格非锚点成员（仅写左上角）."""
        cell = self._anchor_cell(ws, coord)
        if cell is None:
            return
        if self._is_formula_cell(cell):
            return
        cell.value = value

    def _write_header_cell(self, ws, coord: str, value) -> None:
        """写入表头格（合并格写左上角锚点，跳过公式格）."""
        cell = self._anchor_cell(ws, coord)
        if cell is None:
            return
        if self._is_formula_cell(cell):
            return
        cell.value = value

    @staticmethod
    def _anchor_cell(ws, coord: str):
        """返回坐标对应的可写单元格；若是合并格非锚点成员则返回左上角锚点."""
        try:
            cell = ws[coord]
        except (ValueError, KeyError):
            return None
        if isinstance(cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if coord in rng:
                    return ws.cell(rng.min_row, rng.min_col)
            return None
        return cell

    def _create_workbook_from_scratch(
        self,
        types_to_export: list[str],
        report_data: dict[str, list[dict]],
        company_name: str,
        year: int,
        include_prior_year: bool,
        mode: str,
    ) -> Workbook:
        """Generate workbook from scratch when template is not available."""
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for rt in types_to_export:
            sheet_name = REPORT_TYPE_SHEET_NAMES.get(rt)
            if not sheet_name:
                continue

            ws = wb.create_sheet(title=sheet_name)
            rows = report_data.get(rt, [])
            self._write_sheet_from_scratch(
                ws, rows, company_name, year, include_prior_year, mode, rt
            )

        return wb

    def _write_sheet_from_scratch(
        self,
        ws,
        rows: list[dict],
        company_name: str,
        year: int,
        include_prior_year: bool,
        mode: str,
        report_type: str,
    ) -> None:
        """Write a complete sheet from scratch with proper formatting."""
        # 资产负债表使用左右两栏结构（左资产/右负债+权益）
        if report_type == "balance_sheet":
            self._write_balance_sheet_two_column(ws, rows, company_name, year, mode)
            return

        # --- Header rows ---
        # Row 1: Company name (centered, bold)
        ws.cell(1, 1).value = company_name
        ws.cell(1, 1).font = Font(name="宋体", size=14, bold=True)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 if include_prior_year else 2)

        # Row 2: Report period (centered)
        if report_type == "balance_sheet":
            period_text = f"{year}年12月31日"
        else:
            period_text = f"{year}年度"
        ws.cell(2, 1).value = period_text
        ws.cell(2, 1).font = Font(name="宋体", size=11)
        ws.cell(2, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 if include_prior_year else 2)

        # Row 3: Amount unit (right-aligned)
        mode_label = "未审数" if mode == "unadjusted" else "审定数"
        ws.cell(3, 1).value = f"单位：人民币元（{mode_label}）"
        ws.cell(3, 1).font = Font(name="宋体", size=9)
        ws.cell(3, 1).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3 if include_prior_year else 2)

        # Row 4: Column headers
        header_row = 4
        ws.cell(header_row, 1).value = "项　　目"
        ws.cell(header_row, 1).font = Font(bold=True)
        ws.cell(header_row, 1).alignment = Alignment(horizontal="center")

        ws.cell(header_row, 2).value = "期末余额" if report_type == "balance_sheet" else "本期金额"
        ws.cell(header_row, 2).font = Font(bold=True)
        ws.cell(header_row, 2).alignment = Alignment(horizontal="center")

        if include_prior_year:
            ws.cell(header_row, 3).value = "期初余额" if report_type == "balance_sheet" else "上期金额"
            ws.cell(header_row, 3).font = Font(bold=True)
            ws.cell(header_row, 3).alignment = Alignment(horizontal="center")

        # --- Data rows ---
        data_start = 5
        total_row_indices: list[int] = []  # Track total rows for SUM formulas
        child_start_map: dict[int, int] = {}  # total_row_excel_idx → first_child_excel_idx

        # Track children for SUM formula generation
        current_children_start = data_start

        for i, row in enumerate(rows):
            r = data_start + i

            # Column A: row name with indentation
            indent_level = row.get("indent_level", 0)
            indent_str = "　" * indent_level  # Full-width space
            ws.cell(r, 1).value = indent_str + row.get("row_name", "")

            # Apply indent via alignment
            ws.cell(r, 1).alignment = Alignment(
                indent=indent_level * 2,  # 2 units per level
                vertical="center",
            )

            is_total = row.get("is_total_row", False)

            # Column B: current period amount
            amount = row.get("current_period_amount")
            if is_total and i > 0:
                # Use SUM formula for total rows
                child_start = current_children_start
                child_end = r - 1
                if child_end >= child_start:
                    col_letter = get_column_letter(2)
                    ws.cell(r, 2).value = f"=SUM({col_letter}{child_start}:{col_letter}{child_end})"
                elif amount is not None:
                    ws.cell(r, 2).value = float(amount)
                # Reset children start for next group
                current_children_start = r + 1
            elif amount is not None:
                ws.cell(r, 2).value = float(amount)

            self._apply_amount_format(ws.cell(r, 2), is_total)

            # Column C: prior period amount
            if include_prior_year:
                prior = row.get("prior_period_amount")
                if is_total and i > 0:
                    child_start = child_start_map.get(r, data_start)
                    child_end = r - 1
                    if child_end >= data_start:
                        col_letter = get_column_letter(3)
                        ws.cell(r, 3).value = f"=SUM({col_letter}{child_start}:{col_letter}{child_end})"
                    elif prior is not None:
                        ws.cell(r, 3).value = float(prior)
                elif prior is not None:
                    ws.cell(r, 3).value = float(prior)

                self._apply_amount_format(ws.cell(r, 3), is_total)

            # Total row styling
            if is_total:
                self._apply_total_row_style(ws, r, 3 if include_prior_year else 2)
                total_row_indices.append(r)

            # Row name font
            if is_total:
                ws.cell(r, 1).font = Font(bold=True)

        # --- Column widths ---
        ws.column_dimensions["A"].width = 40  # Project name column
        ws.column_dimensions["B"].width = 18  # Amount column
        if include_prior_year:
            ws.column_dimensions["C"].width = 18

        # --- Print settings ---
        ws.print_area = f"A1:{get_column_letter(3 if include_prior_year else 2)}{data_start + len(rows) - 1}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    def _apply_amount_format(self, cell, is_total: bool = False) -> None:
        """Apply amount formatting to a cell."""
        cell.number_format = NEGATIVE_FORMAT
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if is_total:
            cell.font = BOLD_AMOUNT_FONT
        else:
            cell.font = AMOUNT_FONT

    @staticmethod
    def _safe_set_value(ws, row: int, col: int, value) -> None:
        """Safely set cell value, skipping MergedCell objects."""
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            return
        cell.value = value

    def _apply_total_row_style(self, ws, row_idx: int, max_col: int) -> None:
        """Apply total row styling: bold font + top border."""
        thin_border_top = Border(
            top=Side(style="thin"),
        )
        for col in range(1, max_col + 1):
            cell = ws.cell(row_idx, col)
            # Merge existing border with top border
            existing = cell.border
            cell.border = Border(
                top=Side(style="thin"),
                bottom=existing.bottom if existing else None,
                left=existing.left if existing else None,
                right=existing.right if existing else None,
            )
            if col == 1:
                cell.font = Font(bold=True)

    def _write_balance_sheet_two_column(
        self,
        ws,
        rows: list[dict],
        company_name: str,
        year: int,
        mode: str,
    ) -> None:
        """资产负债表左右两栏结构：左侧=资产，右侧=负债+权益。

        布局（参照致同国企版模板）：
        - Row 1: 公司名称（跨全列居中）
        - Row 2: 期间（yyyy年12月31日）
        - Row 3: 单位/审计模式
        - Row 4: 列标题行
          左栏：A=资产项目 | B=期末余额 | C=期初余额
          右栏：E=负债和所有者权益项目 | F=期末余额 | G=期初余额
        - Row 5+: 数据行（左右并排）
        """
        total_cols = 7  # A~G

        # --- Header ---
        ws.cell(1, 1).value = company_name
        ws.cell(1, 1).font = Font(name="宋体", size=14, bold=True)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        ws.cell(2, 1).value = f"资产负债表"
        ws.cell(2, 1).font = Font(name="宋体", size=12, bold=True)
        ws.cell(2, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

        ws.cell(3, 1).value = f"编制日期：{year}年12月31日"
        ws.cell(3, 1).font = Font(name="宋体", size=9)
        ws.cell(3, 1).alignment = Alignment(horizontal="left")
        mode_label = "未审数" if mode == "unadjusted" else "审定数"
        ws.cell(3, total_cols).value = f"单位：人民币元（{mode_label}）"
        ws.cell(3, total_cols).font = Font(name="宋体", size=9)
        ws.cell(3, total_cols).alignment = Alignment(horizontal="right")

        # --- Column headers (Row 4) ---
        header_row = 4
        headers_left = ["资　　产", "期末余额", "期初余额"]
        headers_right = ["负债和所有者权益", "期末余额", "期初余额"]

        for i, h in enumerate(headers_left):
            c = ws.cell(header_row, i + 1)
            c.value = h
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Column D = separator (空列)
        ws.column_dimensions["D"].width = 2

        for i, h in enumerate(headers_right):
            c = ws.cell(header_row, i + 5)  # E, F, G
            c.value = h
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        # --- Separate rows into assets vs liabilities+equity ---
        # Split by row_code prefix: BS-001~BS-039 = 资产, BS-040+ = 负债+权益
        # Or by section: rows before "资产合计" are assets, after are liabilities+equity
        asset_rows: list[dict] = []
        liability_rows: list[dict] = []

        is_liability_section = False
        for row in rows:
            row_name = row.get("row_name", "")
            row_code = row.get("row_code", "")
            # 资产合计行是资产最后一行
            if "资产合计" in row_name or "资产总计" in row_name:
                asset_rows.append(row)
                is_liability_section = True
                continue
            if is_liability_section:
                liability_rows.append(row)
            else:
                asset_rows.append(row)

        # 如果没有通过行名分区成功（所有行都在一侧），则按 row_code 数字分
        if not liability_rows and rows:
            asset_rows = []
            liability_rows = []
            for row in rows:
                code = row.get("row_code", "")
                # BS-039 以下为资产，BS-040 及以上为负债+权益
                try:
                    num = int(code.split("-")[-1]) if "-" in code else 999
                except (ValueError, IndexError):
                    num = 999
                if num <= 39:
                    asset_rows.append(row)
                else:
                    liability_rows.append(row)

        # --- Write data rows side by side ---
        data_start = 5
        max_rows = max(len(asset_rows), len(liability_rows))

        for i in range(max_rows):
            r = data_start + i

            # 左栏：资产
            if i < len(asset_rows):
                arow = asset_rows[i]
                indent_level = arow.get("indent_level", 0)
                indent_str = "　" * indent_level
                is_total = arow.get("is_total_row", False)

                ws.cell(r, 1).value = indent_str + arow.get("row_name", "")
                ws.cell(r, 1).alignment = Alignment(indent=indent_level * 2, vertical="center")
                if is_total:
                    ws.cell(r, 1).font = Font(bold=True)

                amount = arow.get("current_period_amount")
                if amount is not None:
                    ws.cell(r, 2).value = float(amount)
                self._apply_amount_format(ws.cell(r, 2), is_total)

                prior = arow.get("prior_period_amount")
                if prior is not None:
                    ws.cell(r, 3).value = float(prior)
                self._apply_amount_format(ws.cell(r, 3), is_total)

                if is_total:
                    self._apply_total_row_style(ws, r, 3)

            # 右栏：负债+权益
            if i < len(liability_rows):
                lrow = liability_rows[i]
                indent_level = lrow.get("indent_level", 0)
                indent_str = "　" * indent_level
                is_total = lrow.get("is_total_row", False)

                ws.cell(r, 5).value = indent_str + lrow.get("row_name", "")
                ws.cell(r, 5).alignment = Alignment(indent=indent_level * 2, vertical="center")
                if is_total:
                    ws.cell(r, 5).font = Font(bold=True)

                amount = lrow.get("current_period_amount")
                if amount is not None:
                    ws.cell(r, 6).value = float(amount)
                self._apply_amount_format(ws.cell(r, 6), is_total)

                prior = lrow.get("prior_period_amount")
                if prior is not None:
                    ws.cell(r, 7).value = float(prior)
                self._apply_amount_format(ws.cell(r, 7), is_total)

                if is_total:
                    for col in range(5, 8):
                        cell = ws.cell(r, col)
                        existing = cell.border
                        cell.border = Border(
                            top=Side(style="thin"),
                            bottom=existing.bottom if existing else None,
                        )
                    ws.cell(r, 5).font = Font(bold=True)

        # --- Column widths ---
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["E"].width = 32
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 16

        # --- Print settings ---
        last_data_row = data_start + max_rows - 1
        ws.print_area = f"A1:G{last_data_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
