"""批量建项服务 — 模板生成、批量导入、数据导出。

Feature: project-creation-enhancement, Task 8.1
"""

from io import BytesIO
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.core import Project


class BatchImportFailure(BaseModel):
    row_number: int
    errors: list[str]


class BatchImportResult(BaseModel):
    success_count: int
    fail_count: int
    failures: list[BatchImportFailure]


# 数据表列标题（与模板一致）
_TEMPLATE_COLUMNS = [
    "客户名称",
    "企业代码(USCC)",
    "项目简称",
    "审计年度",
    "项目类型",
    "会计准则",
    "报表类型",
]

# 项目类型映射：中文 → 内部值
_PROJECT_TYPE_MAP = {
    "年报审计": "annual",
    "专项审计": "special",
    "IPO审计": "ipo",
    "内控审计": "internal_control",
    "验资": "capital_verification",
    "税审": "tax_audit",
}

# 会计准则映射：中文 → 内部值
_ACCOUNTING_STANDARD_MAP = {
    "企业会计准则": "enterprise",
    "小企业会计准则": "small_enterprise",
    "金融企业会计准则": "financial",
    "政府会计准则": "government",
}

# 报表类型映射：中文 → 内部值
_REPORT_SCOPE_MAP = {
    "单户": "standalone",
    "合并": "consolidated",
}

# 最大导入行数
_MAX_IMPORT_ROWS = 500

# 默认报表类型
DEFAULT_REPORT_SCOPE = "standalone"


async def generate_template() -> BytesIO:
    """生成建项模板 Excel（数据表 + 说明事项 sheet）。

    数据表列：客户名称, 企业代码(USCC), 项目简称, 审计年度, 项目类型, 会计准则, 报表类型
    说明事项 sheet：各字段填写规则说明
    """
    wb = Workbook()

    # 数据表
    ws_data = wb.active
    ws_data.title = "数据"
    ws_data.append(_TEMPLATE_COLUMNS)

    # 说明事项 sheet
    ws_notes = wb.create_sheet("说明事项")
    instructions = [
        ["字段", "填写规则"],
        ["客户名称", "必填，被审计单位全称"],
        ["企业代码(USCC)", "必填，18位统一社会信用代码（不含I、O、Z、S、V）"],
        ["项目简称", "必填，用于审计报告等文档引用"],
        ["审计年度", "必填，4位数字年份（如 2025）"],
        ["项目类型", "必填，可选值：年报审计、专项审计、IPO审计、内控审计、验资、税审"],
        ["会计准则", "必填，可选值：企业会计准则、小企业会计准则、金融企业会计准则、政府会计准则"],
        ["报表类型", "选填，可选值：单户、合并（默认单户）"],
    ]
    for row in instructions:
        ws_notes.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def parse_and_import(file_bytes: bytes, db: AsyncSession) -> BatchImportResult:
    """解析上传文件，逐行执行校验（与单项目相同：USCC格式+short_name非空+唯一性）并创建项目。

    返回成功数+失败明细。
    """
    from app.services.uscc_validator import validate_uscc
    from app.services.uniqueness_checker import check_uniqueness

    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="文件格式不正确，请使用标准建项模板")

    # 定位数据 sheet
    if "数据" in wb.sheetnames:
        ws = wb["数据"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(rows) > _MAX_IMPORT_ROWS:
        raise HTTPException(status_code=413, detail=f"文件超过最大行数限制（{_MAX_IMPORT_ROWS} 行）")

    success_count = 0
    failures: list[BatchImportFailure] = []

    for idx, row in enumerate(rows, start=2):
        # 跳过全空行
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_errors: list[str] = []

        # 提取字段（容错：列数可能不足）
        def _cell(col_idx: int) -> str:
            if col_idx < len(row) and row[col_idx] is not None:
                return str(row[col_idx]).strip()
            return ""

        client_name = _cell(0)
        company_code = _cell(1)
        short_name = _cell(2)
        audit_year_str = _cell(3)
        project_type_cn = _cell(4)
        accounting_standard_cn = _cell(5)
        report_scope_cn = _cell(6)

        # 校验必填字段
        if not client_name:
            row_errors.append("客户名称为必填项")
        if not short_name:
            row_errors.append("项目简称为必填项")
        if not company_code:
            row_errors.append("企业代码为必填项")
        else:
            uscc_valid, uscc_error = validate_uscc(company_code)
            if not uscc_valid:
                row_errors.append(uscc_error or "企业代码格式错误")

        # 审计年度
        audit_year: int | None = None
        if not audit_year_str:
            row_errors.append("审计年度为必填项")
        else:
            try:
                audit_year = int(float(audit_year_str))
                if audit_year < 2000 or audit_year > 2100:
                    row_errors.append("审计年度须在 2000-2100 之间")
            except (ValueError, TypeError):
                row_errors.append("审计年度必须为数字")

        # 项目类型
        project_type = _PROJECT_TYPE_MAP.get(project_type_cn)
        if not project_type_cn:
            row_errors.append("项目类型为必填项")
        elif project_type is None:
            row_errors.append(f"项目类型无效：{project_type_cn}")

        # 会计准则
        accounting_standard = _ACCOUNTING_STANDARD_MAP.get(accounting_standard_cn)
        if not accounting_standard_cn:
            row_errors.append("会计准则为必填项")
        elif accounting_standard is None:
            row_errors.append(f"会计准则无效：{accounting_standard_cn}")

        # 报表类型（选填，默认单户）
        report_scope = DEFAULT_REPORT_SCOPE
        if report_scope_cn:
            mapped = _REPORT_SCOPE_MAP.get(report_scope_cn)
            if mapped is None:
                row_errors.append(f"报表类型无效：{report_scope_cn}")
            else:
                report_scope = mapped

        # 唯一性校验（仅在前面基础校验通过后执行）
        if not row_errors and company_code and audit_year:
            is_unique, uniqueness_error = await check_uniqueness(
                company_code, audit_year, report_scope, db
            )
            if not is_unique:
                row_errors.append(uniqueness_error or "唯一性校验失败")

        if row_errors:
            failures.append(BatchImportFailure(row_number=idx, errors=row_errors))
            continue

        # 创建项目（auto_commit=False，由 batch service 统一 commit）
        try:
            from app.models.audit_platform_schemas import BasicInfoSchema
            from app.services.project_wizard_service import create_project

            data = BasicInfoSchema(
                client_name=client_name,
                audit_year=audit_year,
                project_type=project_type,
                accounting_standard=accounting_standard,
                company_code=company_code,
                short_name=short_name,
                report_scope=report_scope,
            )
            await create_project(data, db, auto_commit=False)
            success_count += 1
        except HTTPException as e:
            failures.append(BatchImportFailure(row_number=idx, errors=[e.detail]))
        except Exception as e:
            # IntegrityError 等 DB 级约束违反（并发唯一性冲突）
            err_msg = str(e)
            if "uq_project_company_year_scope" in err_msg or "UNIQUE" in err_msg.upper():
                failures.append(BatchImportFailure(
                    row_number=idx, errors=["已存在该单位该年度的项目（并发冲突）"]
                ))
            else:
                failures.append(BatchImportFailure(row_number=idx, errors=[err_msg]))
            # 事务可能被污染，尝试 rollback 恢复
            try:
                await db.rollback()
            except Exception:
                pass

    # 全部行处理完毕后统一 commit
    if success_count > 0:
        try:
            await db.commit()
        except Exception as e:
            # commit 阶段 IntegrityError（并发冲突）→ 全部回滚
            await db.rollback()
            err_msg = str(e)
            if "uq_project_company_year_scope" in err_msg or "UNIQUE" in err_msg.upper():
                return BatchImportResult(
                    success_count=0,
                    fail_count=success_count + len(failures),
                    failures=failures + [BatchImportFailure(
                        row_number=0, errors=["批量提交时检测到唯一性冲突，全部回滚"]
                    )],
                )
            raise

    return BatchImportResult(
        success_count=success_count,
        fail_count=len(failures),
        failures=failures,
    )


# 反向映射（内部值 → 中文）
_PROJECT_TYPE_REVERSE = {v: k for k, v in _PROJECT_TYPE_MAP.items()}
_ACCOUNTING_STANDARD_REVERSE = {v: k for k, v in _ACCOUNTING_STANDARD_MAP.items()}
_REPORT_SCOPE_REVERSE = {v: k for k, v in _REPORT_SCOPE_MAP.items()}


async def export_projects(project_ids: list[UUID], db: AsyncSession) -> BytesIO:
    """导出选中项目为 Excel（字段结构与模板数据表一致）。"""
    result = await db.execute(
        select(Project).where(
            Project.id.in_(project_ids),
            Project.is_deleted == False,  # noqa: E712
        )
    )
    projects = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(_TEMPLATE_COLUMNS)

    for p in projects:
        ws.append([
            p.client_name,
            p.company_code or "",
            p.short_name or "",
            p.audit_year or "",
            _PROJECT_TYPE_REVERSE.get(p.project_type.value if p.project_type else "", ""),
            _ACCOUNTING_STANDARD_REVERSE.get(
                _get_accounting_standard(p), ""
            ),
            _REPORT_SCOPE_REVERSE.get(p.report_scope or DEFAULT_REPORT_SCOPE, "单户"),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _get_accounting_standard(project: Project) -> str:
    """从 wizard_state 提取 accounting_standard 值。

    create_project() 将 BasicInfoSchema.model_dump() 写入
    wizard_state.steps.basic_info.data，因此字符串值一定在此路径下。
    """
    ws = project.wizard_state or {}
    steps = ws.get("steps", {})
    # 兼容两种 key 格式：直接 "basic_info" 或嵌套对象
    basic_info_step = steps.get("basic_info", {})
    data = basic_info_step.get("data", {})
    return data.get("accounting_standard", "")
