"""审计说明智能生成服务 — 人机协同模式

Phase 12 P1-1: AI草拟 → 人工编辑 → 确认写回 → 刷新parsed_data

核心流程：
1. generate_draft: 数据采集 → Prompt构建 → LLM调用 → 返回草稿（不覆盖工作簿）
2. confirm_draft: 人工确认 → 写回工作簿 → 刷新parsed_data → 更新状态
3. refine_draft: 基于用户修改优化草稿（不绕过confirm_draft）
"""

from __future__ import annotations

import hashlib
import logging
import uuid
from datetime import datetime, timezone
from typing import AsyncGenerator
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.phase12_models import WpAiGeneration
from app.models.workpaper_models import WorkingPaper

logger = logging.getLogger(__name__)

PROMPT_VERSION = "wp_expl_v1"


class WpExplanationService:
    """审计说明智能生成"""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ------------------------------------------------------------------
    # generate_draft
    # ------------------------------------------------------------------
    async def generate_draft(
        self, project_id: UUID, year: int, wp_id: UUID, user_id: UUID | None = None
    ) -> dict:
        """生成审计说明草稿（不直接覆盖工作簿）。"""
        wp = (await self.db.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )).scalar_one_or_none()
        if not wp:
            return {"error": "底稿不存在"}

        # 1. 数据采集
        context_data = await self._collect_data(project_id, year, wp)

        # 2. 构建 Prompt
        input_text = self._build_prompt(context_data)
        input_hash = hashlib.sha256(input_text.encode()).hexdigest()[:16]

        # 2.5 RAG: 加载上年底稿作为参照
        from app.services.reference_doc_service import ReferenceDocService
        context_docs = await ReferenceDocService.load_context(
            self.db, project_id, year,
            source_type="prior_year_workpaper",
            wp_code=wp.wp_code if hasattr(wp, 'wp_code') else None,
        )

        # 2.6 加载操作手册上下文（按循环自动匹配）
        try:
            from app.services.wp_manual_service import get_context_for_llm
            # 从底稿编号推断循环前缀
            wp_code = getattr(wp, 'wp_code', '') or ''
            cycle = wp_code[0].upper() if wp_code else ''
            if cycle and cycle.isalpha():
                manual_context = get_context_for_llm(cycle, wp_code, max_total_chars=6000)
                if manual_context and not manual_context.startswith("（"):
                    if not context_docs:
                        context_docs = []
                    context_docs.append({
                        "source": f"操作手册（{cycle}循环）",
                        "content": manual_context,
                    })
        except Exception as _manual_err:
            logger.debug("load manual context failed (non-blocking): %s", _manual_err)

        # 3. 调用 LLM
        from app.services.llm_client import chat_completion
        from app.core.config import settings

        model_name = settings.DEFAULT_CHAT_MODEL
        messages = [
            {"role": "system", "content": self._system_prompt()},
            {"role": "user", "content": input_text},
        ]

        try:
            draft_text = await chat_completion(
                messages, model=model_name, temperature=0.3, max_tokens=2000,
                context_documents=context_docs if context_docs else None,
            )
        except Exception as e:
            logger.error("LLM调用失败: %s", e)
            return {"error": f"AI服务暂不可用: {e}"}

        # 4. 写入 wp_ai_generations 留痕
        generation_id = uuid.uuid4()
        gen_record = WpAiGeneration(
            id=generation_id,
            wp_id=wp_id,
            prompt_version=PROMPT_VERSION,
            model=model_name,
            input_hash=input_hash,
            output_text=draft_text,
            output_structured=self._parse_structured(draft_text),
            status="drafted",
            created_by=user_id,
        )
        self.db.add(gen_record)

        # 更新底稿 explanation_status
        wp.explanation_status = "ai_drafted"
        await self.db.flush()

        logger.info("generate_draft: wp=%s gen=%s prompt=%s", wp_id, generation_id, PROMPT_VERSION)

        return {
            "generation_id": str(generation_id),
            "prompt_version": PROMPT_VERSION,
            "draft_text": draft_text,
            "structured": gen_record.output_structured,
            "data_sources": list(context_data.keys()),
            "confidence": self._assess_confidence(context_data),
            "suggestions": self._generate_suggestions(context_data),
        }

    # ------------------------------------------------------------------
    # confirm_draft
    # ------------------------------------------------------------------
    async def confirm_draft(
        self, wp_id: UUID, generation_id: UUID, final_text: str, user_id: UUID
    ) -> dict:
        """人工确认：写回工作簿 → 刷新parsed_data → 更新状态。"""
        wp = (await self.db.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )).scalar_one_or_none()
        if not wp:
            return {"error": "底稿不存在"}

        gen = (await self.db.execute(
            sa.select(WpAiGeneration).where(WpAiGeneration.id == generation_id)
        )).scalar_one_or_none()
        if gen:
            gen.status = "confirmed"
            gen.confirmed_by = user_id
            gen.confirmed_at = datetime.utcnow()

        # 1. 写回底稿工作簿
        write_ok = await self._write_back_to_workbook(wp, final_text)

        if not write_ok:
            wp.explanation_status = "sync_failed"
            await self.db.flush()
            return {"error": "写回工作簿失败", "explanation_status": "sync_failed"}

        # 2. 刷新 parsed_data
        pd = wp.parsed_data or {}
        pd["audit_explanation"] = final_text
        pd["ai_content"] = pd.get("ai_content", {})
        pd["ai_content"]["latest_generation_id"] = str(generation_id)
        pd["ai_content"]["latest_status"] = "confirmed"
        pd["ai_content"]["last_confirmed_at"] = datetime.utcnow().isoformat()
        wp.parsed_data = pd

        # 3. 更新状态
        wp.explanation_status = "synced"
        wp.last_parsed_sync_at = datetime.utcnow()
        await self.db.flush()

        logger.info("confirm_draft: wp=%s gen=%s synced", wp_id, generation_id)
        return {
            "explanation_status": "synced",
            "last_parsed_sync_at": wp.last_parsed_sync_at.isoformat() if wp.last_parsed_sync_at else None,
        }

    # ------------------------------------------------------------------
    # refine_draft
    # ------------------------------------------------------------------
    async def refine_draft(
        self, wp_id: UUID, generation_id: UUID, user_edits: str, feedback: str | None = None
    ) -> dict:
        """基于用户修改优化草稿（不绕过confirm_draft直接生效）。"""
        gen = (await self.db.execute(
            sa.select(WpAiGeneration).where(WpAiGeneration.id == generation_id)
        )).scalar_one_or_none()
        original = gen.output_text if gen else ""

        from app.services.llm_client import chat_completion
        messages = [
            {"role": "system", "content": "你是审计说明优化助手。根据用户修改和反馈，优化审计说明草稿。保持专业、简洁。"},
            {"role": "user", "content": f"原始草稿：\n{original}\n\n用户修改：\n{user_edits}\n\n反馈：{feedback or '无'}"},
        ]
        try:
            refined = await chat_completion(messages, temperature=0.3, max_tokens=2000)
        except Exception as e:
            return {"error": f"AI优化失败: {e}"}

        return {
            "generation_id": str(generation_id),
            "refined_text": refined,
            "prompt_version": PROMPT_VERSION,
        }

    # ------------------------------------------------------------------
    # 私有方法
    # ------------------------------------------------------------------
    async def _collect_data(self, project_id: UUID, year: int, wp) -> dict:
        """采集审计说明所需数据（token预算≤6000）。"""
        from app.models.audit_platform_models import TrialBalance, Adjustment

        data = {}
        pd = wp.parsed_data or {}

        # 试算表数据
        tb_q = sa.select(TrialBalance).where(
            TrialBalance.project_id == project_id,
            TrialBalance.year == year,
            TrialBalance.is_deleted == False,
        )
        tbs = (await self.db.execute(tb_q)).scalars().all()
        if tbs:
            data["trial_balance"] = [
                {"code": t.standard_account_code, "name": t.standard_account_name,
                 "audited": float(t.audited_debit or 0) - float(t.audited_credit or 0),
                 "unadjusted": float(t.unadjusted_debit or 0) - float(t.unadjusted_credit or 0)}
                for t in tbs[:20]
            ]

        # 调整分录
        adj_q = sa.select(Adjustment).where(
            Adjustment.project_id == project_id,
            Adjustment.year == year,
            Adjustment.is_deleted == False,
        ).limit(5)
        adjs = (await self.db.execute(adj_q)).scalars().all()
        if adjs:
            data["adjustments"] = [
                {"code": a.entry_number, "amount": float(a.debit_amount or 0), "summary": a.summary or ""}
                for a in adjs
            ]

        # 底稿已有数据
        data["wp_parsed"] = {
            "audited_amount": pd.get("audited_amount"),
            "unadjusted_amount": pd.get("unadjusted_amount"),
            "conclusion": pd.get("conclusion"),
        }

        return data

    def _system_prompt(self) -> str:
        return (
            "你是一名资深审计师，负责撰写审计底稿的审计说明。\n"
            "请根据提供的试算表数据、调整分录和底稿信息，生成结构化的审计说明。\n\n"
            "格式要求：\n"
            "1. 审计目标：一句话说明本科目审计目标\n"
            "2. 执行程序：列出已执行的主要审计程序（2-4条）\n"
            "3. 审计发现：描述关键发现（金额变动、异常项等）\n"
            "4. 审计结论：给出明确结论\n\n"
            "要求：专业简洁，200-500字，金额保留2位小数，不虚构未提供的数据。"
        )

    def _build_prompt(self, context_data: dict) -> str:
        parts = []
        if "trial_balance" in context_data:
            tb_lines = []
            for t in context_data["trial_balance"][:10]:
                change = t["audited"] - t["unadjusted"]
                rate = round(change / t["unadjusted"] * 100, 1) if t["unadjusted"] else 0
                tb_lines.append(f"  {t['code']} {t['name']}: 审定{t['audited']:,.2f} 未审{t['unadjusted']:,.2f} 变动{rate}%")
            parts.append("试算表数据：\n" + "\n".join(tb_lines))

        if "adjustments" in context_data:
            adj_lines = [f"  {a['code']}: {a['amount']:,.2f} {a['summary']}" for a in context_data["adjustments"]]
            parts.append("调整分录：\n" + "\n".join(adj_lines))

        wp = context_data.get("wp_parsed", {})
        if wp.get("audited_amount") is not None:
            parts.append(f"底稿审定数: {wp['audited_amount']:,.2f}")

        return "\n\n".join(parts) if parts else "暂无数据，请生成通用审计说明模板。"

    def _parse_structured(self, text: str) -> dict | None:
        """尝试从生成文本中提取结构化字段。"""
        result = {}
        for label, key in [("审计目标", "objective"), ("执行程序", "procedures"),
                           ("审计发现", "findings"), ("审计结论", "conclusion")]:
            idx = text.find(label)
            if idx >= 0:
                end = len(text)
                for next_label, _ in [("执行程序", ""), ("审计发现", ""), ("审计结论", "")]:
                    ni = text.find(next_label, idx + len(label))
                    if ni > idx and ni < end:
                        end = ni
                result[key] = text[idx + len(label):end].strip().lstrip("：:").strip()
        return result if result else None

    def _assess_confidence(self, context_data: dict) -> str:
        score = 0
        if context_data.get("trial_balance"):
            score += 2
        if context_data.get("adjustments"):
            score += 1
        if context_data.get("wp_parsed", {}).get("audited_amount") is not None:
            score += 1
        return "high" if score >= 3 else "medium" if score >= 1 else "low"

    def _generate_suggestions(self, context_data: dict) -> list[str]:
        suggestions = []
        if not context_data.get("adjustments"):
            suggestions.append("建议补充调整分录信息")
        wp = context_data.get("wp_parsed", {})
        if wp.get("audited_amount") is None:
            suggestions.append("底稿审定数为空，建议先完成预填充")
        return suggestions

    async def _write_back_to_workbook(self, wp, text: str) -> bool:
        """将审计说明写回底稿工作簿。"""
        from pathlib import Path
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl未安装，无法写回")
            return False

        fp = Path(wp.file_path) if wp.file_path else None
        if not fp or not fp.exists():
            logger.warning("底稿文件不存在: %s", fp)
            return False

        try:
            wb = openpyxl.load_workbook(str(fp))
            ws = wb.active
            # 搜索"审计说明"关键词位置，写入右侧或下方
            target_cell = None
            for row in ws.iter_rows(max_row=100, max_col=15):
                for cell in row:
                    if cell.value and any(kw in str(cell.value) for kw in ["审计说明", "审计结论", "执行情况"]):
                        # 写入右侧单元格
                        target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        break
                if target_cell:
                    break

            if target_cell:
                target_cell.value = text
            else:
                # 兜底：写入A行最后+2行
                max_row = ws.max_row or 1
                ws.cell(row=max_row + 2, column=1, value="审计说明")
                ws.cell(row=max_row + 2, column=2, value=text)

            wb.save(str(fp))
            wb.close()
            return True
        except Exception as e:
            logger.error("写回工作簿失败: %s", e)
            return False
