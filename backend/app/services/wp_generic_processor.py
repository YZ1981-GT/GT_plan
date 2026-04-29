"""底稿通用读取处理引擎 — 用规则驱动替代逐科目写脚本

核心思路：
1. 任何底稿Excel都是"多Sheet+表头区域+数据区域"的结构
2. 通用规则自动识别：表头行位置、数据起始行、列含义、合计行
3. 按 wp_parse_rules.json 配置驱动提取，无需为每个科目写代码
4. 提取结果统一为 structure.json 格式，接入四式联动

═══ 通用处理流程 ═══
1. 打开Excel → 遍历所有Sheet
2. 每个Sheet：检测表头行（关键词扫描前10行）→ 确定数据区域
3. 按列名映射提取数据（审定数/未审数/期初/期末/增减/合计）
4. 识别合计行（"合计"/"小计"/"总计"关键词）
5. 输出结构化数据（可直接用于附注填充/报表校验/四式联动）

═══ 规则配置格式（wp_parse_rules.json） ═══
{
    "wp_code": "D2",
    "sheets": [
        {
            "name_pattern": "审定表|D2-1",
            "type": "summary",  // summary=审定表 / detail=明细表 / analysis=分析表
            "header_keywords": ["项目", "期末", "期初"],
            "data_start_offset": 1,  // 表头下方第N行开始是数据
            "columns": {
                "label": {"keywords": ["项目", "名称", "科目"]},
                "closing": {"keywords": ["期末", "审定", "余额"]},
                "opening": {"keywords": ["期初", "年初"]},
                "adjustment": {"keywords": ["调整", "AJE", "RJE"]},
            }
        }
    ]
}
"""

from __future__ import annotations

import json
import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Any

_logger = logging.getLogger(__name__)


# ═══ 通用Sheet解析 ═══

def parse_workpaper_generic(file_path: str, wp_code: str | None = None) -> dict[str, Any]:
    """通用底稿解析 — 自动识别所有Sheet的结构并提取数据

    Args:
        file_path: Excel文件路径
        wp_code: 底稿编号（用于加载特定规则，None则纯通用）

    Returns:
        {
            "wp_code": "D2",
            "file": "D2.xlsx",
            "sheets": [
                {
                    "name": "审定表D2-1",
                    "type": "summary",
                    "header_row": 5,
                    "headers": ["项目", "期末余额", "期初余额"],
                    "data_rows": [...],
                    "totals": {...},
                    "metadata": {...}
                }
            ]
        }
    """
    import openpyxl

    fp = Path(file_path)
    if not fp.exists():
        return {"error": f"文件不存在: {file_path}"}

    # 加载规则（如果有）
    rules = _load_rules_for_wp(wp_code) if wp_code else None

    try:
        wb = openpyxl.load_workbook(str(fp), data_only=True, read_only=True)
    except Exception as e:
        return {"error": f"打开文件失败: {e}"}

    result = {
        "wp_code": wp_code or fp.stem,
        "file": fp.name,
        "sheets": [],
    }

    for ws in wb.worksheets:
        sheet_data = _parse_sheet_generic(ws, rules)
        if sheet_data:
            result["sheets"].append(sheet_data)

    wb.close()
    return result


def _parse_sheet_generic(ws, rules: dict | None) -> dict | None:
    """通用Sheet解析"""
    sheet_name = ws.title

    # 如果有规则，匹配Sheet
    sheet_rule = None
    if rules:
        for sr in rules.get("sheets", []):
            pattern = sr.get("name_pattern", "")
            if pattern and re.search(pattern, sheet_name):
                sheet_rule = sr
                break

    # 读取前15行用于表头检测
    rows_data = []
    for i, row in enumerate(ws.iter_rows(max_row=min(500, ws.max_row or 100), values_only=True)):
        rows_data.append(list(row))
        if i >= 499:
            break

    if not rows_data:
        return None

    # 检测表头行
    header_row_idx = _detect_header_row(rows_data, sheet_rule)
    if header_row_idx is None:
        return None

    headers = [str(c).strip() if c else "" for c in rows_data[header_row_idx]]
    # 去掉全空的尾部列
    while headers and not headers[-1]:
        headers.pop()

    if len(headers) < 2:
        return None

    # 确定数据起始行
    data_start = header_row_idx + 1 + (sheet_rule.get("data_start_offset", 0) if sheet_rule else 0)

    # 列含义映射
    col_mapping = _map_columns(headers, sheet_rule)

    # 提取数据行
    data_rows = []
    totals = {}
    for i in range(data_start, len(rows_data)):
        row = rows_data[i]
        if not row or all(c is None for c in row):
            continue

        # 提取行数据
        row_data = _extract_row_data(row, headers, col_mapping)
        if not row_data:
            continue

        # 检测是否为合计行
        label = row_data.get("label", "")
        if _is_total_row(label):
            row_data["is_total"] = True
            totals[label] = row_data
        else:
            row_data["is_total"] = False

        data_rows.append(row_data)

    # 确定Sheet类型
    sheet_type = _detect_sheet_type(sheet_name, headers, sheet_rule)

    return {
        "name": sheet_name,
        "type": sheet_type,
        "header_row": header_row_idx,
        "headers": headers,
        "col_mapping": col_mapping,
        "data_rows": data_rows,
        "row_count": len(data_rows),
        "totals": totals,
    }


# ═══ 表头检测 ═══

_HEADER_KEYWORDS = {"项目", "名称", "科目", "期末", "期初", "余额", "借方", "贷方",
                    "审定", "未审", "调整", "合计", "金额", "本期", "上期", "增加", "减少"}

# 高权重关键词（出现即强烈暗示是表头行）
_HIGH_WEIGHT_KEYWORDS = {"期末余额", "期初余额", "审定数", "未审数", "借方金额", "贷方金额",
                         "本期增加", "本期减少", "累计折旧", "账面价值", "坏账准备"}


def _detect_header_row(rows: list[list], rule: dict | None) -> int | None:
    """检测表头行位置

    规则：
    1. 如果有规则指定 header_keywords，用规则匹配
    2. 否则扫描前15行，用加权评分找最佳表头行
       - 高权重关键词（如"期末余额"）得3分
       - 普通关键词（如"项目"）得1分
       - 跳过只有1-2个非空单元格的行（标题行/说明行）
    """
    rule_keywords = set(rule.get("header_keywords", [])) if rule else set()
    target_keywords = rule_keywords or _HEADER_KEYWORDS

    best_row = None
    best_score = 0

    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        score = 0
        non_empty = 0
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            non_empty += 1

            # 高权重匹配（完整词组）
            for hw in _HIGH_WEIGHT_KEYWORDS:
                if hw in s:
                    score += 3
                    break
            else:
                # 普通关键词匹配
                for kw in target_keywords:
                    if kw in s:
                        score += 1
                        break

        # 表头行应该有3+个非空单元格（排除标题行/说明行只有1-2个）
        if non_empty >= 3 and score >= 2 and score > best_score:
            best_score = score
            best_row = i

    return best_row


# ═══ 列含义映射 ═══

_COLUMN_KEYWORDS = {
    "label": ["项目", "名称", "科目", "内容", "类别", "摘要"],
    "closing_balance": ["期末余额", "期末数", "期末审定", "审定数", "期末"],
    "opening_balance": ["期初余额", "期初数", "年初余额", "期初", "年初"],
    "unadjusted": ["未审数", "未审", "账面"],
    "adjustment": ["调整数", "调整", "AJE", "RJE"],
    "reclassification": ["重分类", "重分"],
    "increase": ["本期增加", "增加", "本期转入"],
    "decrease": ["本期减少", "减少", "本期转出", "本期处置"],
    "provision": ["本期计提", "计提", "本期摊销", "摊销"],
    "debit": ["借方", "借方金额", "借方发生"],
    "credit": ["贷方", "贷方金额", "贷方发生"],
}


def _map_columns(headers: list[str], rule: dict | None) -> dict[str, int]:
    """将列标题映射为标准含义

    Returns: {"label": 0, "closing_balance": 1, "opening_balance": 2, ...}
    """
    # 优先用规则中的列定义
    rule_columns = rule.get("columns", {}) if rule else {}

    mapping: dict[str, int] = {}

    for col_idx, header in enumerate(headers):
        if not header:
            continue
        h = header.strip()

        # 规则匹配
        matched = False
        for field_name, field_def in rule_columns.items():
            keywords = field_def.get("keywords", []) if isinstance(field_def, dict) else []
            for kw in keywords:
                if kw in h:
                    mapping[field_name] = col_idx
                    matched = True
                    break
            if matched:
                break

        if matched:
            continue

        # 通用关键词匹配
        for field_name, keywords in _COLUMN_KEYWORDS.items():
            for kw in keywords:
                if kw in h:
                    if field_name not in mapping:  # 不覆盖已匹配的
                        mapping[field_name] = col_idx
                    break

    # 第一列默认为 label（如果未匹配）
    if "label" not in mapping and headers:
        mapping["label"] = 0

    return mapping


# ═══ 数据提取 ═══

def _extract_row_data(row: list, headers: list[str], col_mapping: dict[str, int]) -> dict | None:
    """从一行数据中按列映射提取结构化数据"""
    if not row:
        return None

    data: dict[str, Any] = {}

    for field_name, col_idx in col_mapping.items():
        if col_idx < len(row):
            val = row[col_idx]
            if field_name == "label":
                data["label"] = str(val).strip() if val else ""
            else:
                # 数值字段
                data[field_name] = _safe_number(val)

    # 跳过完全空的行（label为空且无数值）
    if not data.get("label") and all(v is None for k, v in data.items() if k != "label"):
        return None

    # 补充原始行数据（供穿透查看）
    data["_raw"] = [_safe_number(c) if isinstance(c, (int, float)) else (str(c).strip() if c else "") for c in row[:len(headers)]]

    return data


def _safe_number(val) -> float | None:
    """安全转换为数字"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None  # 0视为空
    if isinstance(val, str):
        s = val.strip().replace(",", "").replace("，", "")
        if not s or s == "-" or s == "—":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


# ═══ 辅助函数 ═══

_TOTAL_KEYWORDS = {"合计", "小计", "总计", "合 计", "总 计", "Total", "TOTAL", "Subtotal"}


def _is_total_row(label: str) -> bool:
    """判断是否为合计行"""
    if not label:
        return False
    for kw in _TOTAL_KEYWORDS:
        if kw in label:
            return True
    return False


def _detect_sheet_type(name: str, headers: list[str], rule: dict | None) -> str:
    """检测Sheet类型"""
    if rule and rule.get("type"):
        return rule["type"]

    name_lower = name.lower()
    if "审定" in name or "汇总" in name:
        return "summary"
    if "明细" in name or "清单" in name:
        return "detail"
    if "分析" in name or "测算" in name or "复核" in name:
        return "analysis"
    if "程序" in name:
        return "procedure"
    if "调整" in name:
        return "adjustment"
    if "披露" in name or "附注" in name:
        return "disclosure"

    # 从列名推断
    header_text = " ".join(headers)
    if "增加" in header_text and "减少" in header_text:
        return "movement"  # 变动表
    if "账龄" in header_text or "1年以内" in header_text:
        return "aging"  # 账龄表

    return "other"


# ═══ 规则加载 ═══

_RULES_CACHE: dict[str, dict] = {}


def _load_rules_for_wp(wp_code: str) -> dict | None:
    """加载底稿解析规则

    规则文件格式：[{wp_code: "E1", name: "...", sheets: [...]}]
    匹配逻辑：wp_code 精确匹配或前缀匹配（如 "E1" 匹配 "E1-1"）
    """
    if wp_code in _RULES_CACHE:
        return _RULES_CACHE[wp_code]

    def _load_file(path: Path):
        if not path.exists():
            return
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            # 支持新格式（数组）和旧格式（{rules: {...}}）
            if isinstance(raw, list):
                rules_list = raw
            elif isinstance(raw, dict) and "rules" in raw:
                # 旧格式兼容：转为新格式
                rules_list = [{"wp_code": k, **v} for k, v in raw["rules"].items()]
            else:
                rules_list = []

            for rule in rules_list:
                code = rule.get("wp_code", "")
                _RULES_CACHE[code] = rule
        except Exception:
            pass

    # 加载两个规则文件
    data_dir = Path(__file__).parent.parent.parent / "data"
    _load_file(data_dir / "wp_parse_rules.json")
    _load_file(data_dir / "wp_parse_rules_extended.json")

    # 精确匹配
    if wp_code in _RULES_CACHE:
        return _RULES_CACHE[wp_code]

    # 前缀匹配（如 "E1-1" 匹配规则 "E1"）
    prefix = wp_code.split("-")[0] if "-" in wp_code else wp_code
    if prefix in _RULES_CACHE:
        return _RULES_CACHE[prefix]

    return None


# ═══ 批量处理 ═══

def parse_all_project_workpapers(project_dir: str) -> list[dict]:
    """批量解析项目下所有底稿文件

    扫描 storage/projects/{id}/workpapers/ 下所有 .xlsx 文件
    """
    results = []
    wp_dir = Path(project_dir)
    if not wp_dir.exists():
        return results

    for xlsx_file in wp_dir.rglob("*.xlsx"):
        wp_code = xlsx_file.stem
        result = parse_workpaper_generic(str(xlsx_file), wp_code)
        if "error" not in result:
            results.append(result)

    return results
