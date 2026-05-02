"""将 xlsx 文件转换为 Univer IWorkbookData 格式

用 openpyxl 解析 Excel 文件，输出 Univer 前端可直接加载的 JSON 数据，
包含所有 Sheet、单元格数据、公式、基本样式（粗体、对齐、边框、背景色）。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment, Border

logger = logging.getLogger(__name__)


def xlsx_to_univer_data(file_path: str, max_rows: int = 5000) -> dict[str, Any]:
    """将 xlsx 文件转换为 Univer IWorkbookData 格式"""
    path = Path(file_path)
    if not path.exists():
        return _empty_workbook("Sheet1")

    try:
        wb = load_workbook(str(path), data_only=False)
    except Exception as e:
        logger.warning("无法解析 xlsx: %s", e)
        return _empty_workbook("Sheet1")

    sheets: dict[str, Any] = {}
    sheet_order: list[str] = []

    for idx, ws_name in enumerate(wb.sheetnames):
        ws = wb[ws_name]
        sheet_id = f"sheet{idx}"
        sheet_order.append(sheet_id)

        cell_data: dict[int, dict[int, Any]] = {}
        col_widths: dict[int, float] = {}
        row_heights: dict[int, float] = {}
        merges: list[dict] = []

        # 列宽
        for col_letter, dim in (ws.column_dimensions or {}).items():
            if dim.width and dim.width > 0:
                from openpyxl.utils import column_index_from_string
                ci = column_index_from_string(col_letter) - 1
                col_widths[ci] = dim.width * 7  # Excel 列宽单位转像素（近似）

        # 行高
        for row_num, dim in (ws.row_dimensions or {}).items():
            if dim.height and dim.height > 0:
                row_heights[row_num - 1] = dim.height

        # 合并单元格
        for merge_range in ws.merged_cells.ranges:
            merges.append({
                "startRow": merge_range.min_row - 1,
                "endRow": merge_range.max_row - 1,
                "startColumn": merge_range.min_col - 1,
                "endColumn": merge_range.max_col - 1,
            })

        # 单元格数据
        row_count = 0
        for row in ws.iter_rows(max_row=max_rows):
            r = row[0].row - 1
            row_count = r + 1
            for cell in row:
                if cell.value is None and not cell.font.bold and not cell.fill.fgColor:
                    continue
                c = cell.column - 1
                if r not in cell_data:
                    cell_data[r] = {}
                cell_data[r][c] = _convert_cell(cell)

        actual_rows = max(row_count, ws.max_row or 1)
        actual_cols = ws.max_column or 10

        sheet_data: dict[str, Any] = {
            "id": sheet_id,
            "name": ws_name,
            "rowCount": min(actual_rows + 20, max_rows + 20),
            "columnCount": max(actual_cols + 5, 20),
            "cellData": cell_data,
        }

        # 列宽
        if col_widths:
            sheet_data["columnData"] = {
                str(ci): {"w": w} for ci, w in col_widths.items()
            }

        # 行高
        if row_heights:
            sheet_data["rowData"] = {
                str(ri): {"h": h} for ri, h in row_heights.items()
            }

        # 合并单元格
        if merges:
            sheet_data["mergeData"] = merges

        # 冻结首行
        if ws.freeze_panes:
            freeze_row = ws.freeze_panes.row - 1 if ws.freeze_panes.row else 0
            freeze_col = ws.freeze_panes.column - 1 if ws.freeze_panes.column else 0
            if freeze_row > 0 or freeze_col > 0:
                sheet_data["freeze"] = {
                    "startRow": freeze_row,
                    "startColumn": freeze_col,
                    "xSplit": freeze_col,
                    "ySplit": freeze_row,
                }

        sheets[sheet_id] = sheet_data

    wb.close()

    return {
        "id": path.stem,
        "name": path.stem,
        "sheetOrder": sheet_order,
        "sheets": sheets,
    }


def _convert_cell(cell: Cell) -> dict[str, Any]:
    """将 openpyxl Cell 转换为 Univer 单元格格式"""
    result: dict[str, Any] = {}

    # 值
    val = cell.value
    if val is not None:
        if isinstance(val, (int, float)):
            result["v"] = val
            result["t"] = 2  # CellValueType.NUMBER
        elif isinstance(val, bool):
            result["v"] = 1 if val else 0
            result["t"] = 4  # CellValueType.BOOLEAN
        else:
            result["v"] = str(val)
            result["t"] = 1  # CellValueType.STRING

    # 公式（openpyxl 用 cell.value 存公式字符串，以 = 开头）
    if isinstance(cell.value, str) and cell.value.startswith("="):
        result["f"] = cell.value  # 保留完整公式
        result["v"] = ""  # 公式单元格的显示值由 Univer 计算

    # 样式
    style = _convert_style(cell)
    if style:
        result["s"] = style

    return result


def _convert_style(cell: Cell) -> dict[str, Any] | None:
    """将 openpyxl 样式转换为 Univer 样式格式"""
    s: dict[str, Any] = {}

    font = cell.font
    if font:
        if font.bold:
            s["bl"] = 1
        if font.italic:
            s["it"] = 1
        if font.underline and font.underline != "none":
            s["ul"] = {"s": 1}
        if font.size and font.size != 11:
            s["fs"] = font.size
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            rgb = str(font.color.rgb)
            if len(rgb) == 8:
                rgb = rgb[2:]  # 去掉 alpha
            if rgb != "000000":
                s["cl"] = {"rgb": f"#{rgb}"}
        if font.name and font.name not in ("Calibri", "等线", "宋体"):
            s["ff"] = font.name

    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb:
        rgb = str(fill.fgColor.rgb)
        if len(rgb) == 8:
            rgb = rgb[2:]
        if rgb and rgb != "000000" and rgb != "FFFFFF":
            s["bg"] = {"rgb": f"#{rgb}"}

    alignment = cell.alignment
    if alignment:
        if alignment.horizontal:
            ha_map = {"left": 0, "center": 1, "right": 2}
            if alignment.horizontal in ha_map:
                s["ht"] = ha_map[alignment.horizontal]
        if alignment.vertical:
            va_map = {"top": 0, "center": 1, "bottom": 2}
            if alignment.vertical in va_map:
                s["vt"] = va_map[alignment.vertical]
        if alignment.wrap_text:
            s["tb"] = 3  # WrapStrategy.WRAP

    border = cell.border
    if border:
        bd: dict[str, Any] = {}
        for side_name, side_key in [("left", "l"), ("right", "r"), ("top", "t"), ("bottom", "b")]:
            side = getattr(border, side_name, None)
            if side and side.style and side.style != "none":
                style_map = {"thin": 1, "medium": 2, "thick": 3, "dashed": 5, "dotted": 8}
                bd[side_key] = {
                    "s": style_map.get(side.style, 1),
                    "cl": {"rgb": "#D9D9D9"},
                }
        if bd:
            s["bd"] = bd

    return s if s else None


def _empty_workbook(name: str = "Sheet1") -> dict[str, Any]:
    """创建空白工作簿"""
    return {
        "id": "empty",
        "name": name,
        "sheetOrder": ["sheet0"],
        "sheets": {
            "sheet0": {
                "id": "sheet0",
                "name": name,
                "rowCount": 100,
                "columnCount": 20,
                "cellData": {},
            },
        },
    }
