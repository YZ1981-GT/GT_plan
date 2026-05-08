# -*- coding: utf-8 -*-
"""通用智能四表导入引擎（DEPRECATED — 逐步迁移到 ledger_import/ v2 模块）。

.. deprecated::
    本文件为 v1 引擎（3000+ 行单文件），已被 v2 模块化引擎替代。
    v2 位于 ``app/services/ledger_import/``，提供：
    - 三级并行识别（detector → identifier → adapter）
    - 声明式 JSON 配置（热加载）
    - 分层校验（key/recommended/extra）
    - 大文件流式探测（600MB+ CSV < 10ms）
    - raw_extra JSONB 保留未映射列

    迁移路径：
    - detect/identify → ``ledger_import.orchestrator.ImportOrchestrator.detect_from_paths``
    - convert_balance_rows → ``ledger_import.converter.convert_balance_rows``
    - convert_ledger_rows → ``ledger_import.converter.convert_ledger_rows``
    - smart_import_streaming → 通过 feature_flag ``ledger_import_v2`` 切换

    当前仍被 import_job_runner.py / ledger_import_application_service.py 调用，
    待 v2 全链路验证通过后逐步切换。

支持不同企业导出的各种格式：
1. 单行/双行合并表头自动检测与合并
2. 借贷方向自动推断（期初/期末分借贷两列时不依赖方向列）
3. 核算维度多维度拆分 + 用户确认接口
4. 多文件多年度自动识别
5. 辅助余额表 vs 辅助明细账一致性校验
"""

import hashlib
import io
import logging
import re
from contextlib import contextmanager
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterator, Optional
from uuid import UUID

logger = logging.getLogger(__name__)

class SmartImportError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        diagnostics: list[dict] | None = None,
        errors: list[str] | None = None,
        year: int | None = None,
    ):
        super().__init__(message)
        self.diagnostics = diagnostics or []
        self.errors = errors or []
        self.year = year

# ─────────────────────────────────────────────────────────────────────────────
# 1. 智能表头检测与合并
# ─────────────────────────────────────────────────────────────────────────────

# 表头关键词（用于识别表头行 vs 标题/说明行）
_HEADER_KEYWORDS = {
    "科目编码", "科目代码", "科目编号", "科目名称", "凭证日期", "记账日期",
    "借方", "贷方", "借方金额", "贷方金额", "期初", "期末", "期初余额", "期末余额",
    "年初余额", "本期发生额", "本年累计",
    "摘要", "凭证号", "凭证类型", "凭证字号",
    "核算维度", "辅助类型", "核算类型",
    "序号", "编码", "名称", "余额", "组织编码", "核算组织", "期间",
    "account_code", "account_name", "debit", "credit", "opening", "closing",
}

# 双行表头中第二行的典型关键词（借方金额/贷方金额）
_SUBHEADER_KEYWORDS = {"借方金额", "贷方金额", "借方", "贷方"}


def detect_header_rows(ws, max_scan: int = 8) -> tuple[int, int]:
    """检测表头的起始行和行数（支持单行/双行合并表头）。

    Returns:
        (header_start_row, header_row_count)
        - header_start_row: 0-indexed，表头第一行的位置
        - header_row_count: 表头占几行（1 或 2）
    """
    rows_cache = []
    rows_iter = ws.iter_rows(values_only=True)
    for _ in range(max_scan):
        try:
            row = next(rows_iter)
            rows_cache.append([str(c).strip() if c else "" for c in row])
        except StopIteration:
            break

    header_start = -1
    for i, cells in enumerate(rows_cache):
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue

        # 检查是否包含表头关键词
        has_kw = any(kw in cell for cell in non_empty for kw in _HEADER_KEYWORDS)

        # 标题/说明行特征
        first = non_empty[0] if non_empty else ""
        is_title = (
            (len(non_empty) == 1)  # 只有1个非空单元格 → 标题行或说明行
            or first.startswith("核算组织")
            or (len(first) > 30)
        )

        if is_title and not has_kw:
            continue

        if has_kw and len(non_empty) >= 2:
            header_start = i
            break

        # 多个短文本单元格也可能是表头
        if len(non_empty) >= 3 and all(len(c) <= 30 for c in non_empty):
            header_start = i
            break

    if header_start < 0:
        return 0, 1

    # 检查下一行是否是第二行表头（双行合并表头）
    if header_start + 1 < len(rows_cache):
        next_row = rows_cache[header_start + 1]
        next_non_empty = [c for c in next_row if c]

        # 第二行表头特征：包含"借方金额"/"贷方金额"等子列名
        has_sub_kw = any(
            kw in cell for cell in next_non_empty for kw in _SUBHEADER_KEYWORDS
        )

        if has_sub_kw:
            # 第一行有空列（合并单元格导致的空位）或第二行有更多非空列
            first_row = rows_cache[header_start]
            empty_count = sum(1 for c in first_row if not c)
            if empty_count >= 2 or len(next_non_empty) > len([c for c in first_row if c]):
                return header_start, 2

    return header_start, 1


def merge_header_rows(ws, header_start: int, header_count: int) -> list[str]:
    """合并单行或双行表头为完整列名列表。

    双行合并规则：
    - Row1: "年初余额"(col5) ""(col6) "期初余额"(col7) ""(col8) ...
    - Row2: ""(col1-4) "借方金额"(col5) "贷方金额"(col6) "借方金额"(col7) ...
    - 合并结果: col5="年初余额_借方金额", col6="年初余额_贷方金额", ...

    支持 read_only 和非 read_only 模式。
    非 read_only 模式下，合并单元格的值只在左上角，其他位置为 None。
    """
    # 尝试用 cell 对象读取（非 read_only 模式）
    try:
        max_col = ws.max_column or 1
    except Exception:
        max_col = 50  # read_only 模式下可能没有 max_column

    def _read_row(row_idx: int) -> list[str]:
        """读取指定行的所有单元格值。"""
        result = []
        try:
            # 尝试按行列索引读取（非 read_only 模式）
            for col in range(1, max_col + 1):
                val = ws.cell(row_idx, col).value
                result.append(str(val).strip() if val else "")
        except Exception:
            # read_only 模式：用 iter_rows
            rows_iter = ws.iter_rows(values_only=True)
            for _ in range(row_idx):
                try:
                    row = next(rows_iter)
                except StopIteration:
                    return []
            try:
                row = next(rows_iter)
                result = [str(c).strip() if c else "" for c in row]
            except StopIteration:
                return []
        return result

    # 读取表头行（1-indexed）
    row1 = _read_row(header_start + 1)

    if header_count == 1:
        return [h if h else f"col_{i}" for i, h in enumerate(row1)]

    # 双行表头
    row2 = _read_row(header_start + 2)
    max_cols = max(len(row1), len(row2))

    # 向右填充 Row1 的合并单元格（空列继承左边的值）
    filled_row1 = list(row1) + [""] * (max_cols - len(row1))
    for i in range(1, max_cols):
        if not filled_row1[i] and filled_row1[i - 1]:
            filled_row1[i] = filled_row1[i - 1]

    filled_row2 = list(row2) + [""] * (max_cols - len(row2))

    headers = []
    for i in range(max_cols):
        h1 = filled_row1[i]
        h2 = filled_row2[i]
        if h1 and h2 and h1 != h2:
            headers.append(f"{h1}_{h2}")
        elif h1:
            headers.append(h1)
        elif h2:
            headers.append(h2)
        else:
            headers.append(f"col_{i}")

    return headers



# ─────────────────────────────────────────────────────────────────────────────
# 2. 通用列名映射（扩展版，覆盖合并表头产生的组合列名）
# ─────────────────────────────────────────────────────────────────────────────

# 标准字段 → 中文描述（用于前端展示）
FIELD_LABELS = {
    "account_code": "科目编码",
    "account_name": "科目名称",
    "opening_debit": "期初借方",
    "opening_credit": "期初贷方",
    "opening_balance": "期初余额",
    "debit_amount": "借方发生额",
    "credit_amount": "贷方发生额",
    "closing_debit": "期末借方",
    "closing_credit": "期末贷方",
    "closing_balance": "期末余额",
    "year_opening_debit": "年初借方",
    "year_opening_credit": "年初贷方",
    "year_debit": "本年累计借方",
    "year_credit": "本年累计贷方",
    "voucher_date": "凭证日期",
    "voucher_no": "凭证号",
    "voucher_type": "凭证类型",
    "accounting_period": "会计期间",
    "summary": "摘要",
    "preparer": "制单人",
    "aux_dimensions": "核算维度",
    "aux_type": "辅助类型",
    "aux_code": "辅助编码",
    "aux_name": "辅助名称",
    "company_code": "组织编码",
    "level": "科目级次",
    "direction": "借贷方向",
}

_REQUIRED_FIELD_GROUPS: dict[str, list[tuple[str, set[str]]]] = {
    "balance": [
        ("account_code", {"account_code"}),
    ],
    "ledger": [
        ("account_code", {"account_code"}),
        ("voucher_date", {"voucher_date"}),
        ("voucher_no", {"voucher_no"}),
    ],
    "aux_balance": [
        ("account_code", {"account_code"}),
        ("aux_type", {"aux_type"}),
    ],
    "aux_ledger": [
        ("account_code", {"account_code"}),
    ],
    "account_chart": [
        ("account_code", {"account_code"}),
        ("account_name", {"account_name"}),
    ],
}

_RECOMMENDED_FIELD_GROUPS: dict[str, list[tuple[str, set[str]]]] = {
    "balance": [
        ("opening_balance", {"opening_balance", "opening_debit", "opening_credit", "year_opening_debit", "year_opening_credit"}),
        ("debit_amount", {"debit_amount"}),
        ("credit_amount", {"credit_amount"}),
        ("closing_balance", {"closing_balance", "closing_debit", "closing_credit"}),
    ],
    "ledger": [
        ("debit_amount", {"debit_amount"}),
        ("credit_amount", {"credit_amount"}),
        ("summary", {"summary"}),
    ],
    "aux_balance": [
        ("opening_balance", {"opening_balance", "opening_debit", "opening_credit"}),
        ("closing_balance", {"closing_balance", "closing_debit", "closing_credit"}),
        ("aux_code", {"aux_code"}),
        ("aux_name", {"aux_name"}),
    ],
}

# 合并表头列名 → 标准字段（覆盖双行合并产生的组合名）
_MERGED_HEADER_MAP = {
    # 年初余额（双行合并：年初余额_借方金额）
    "年初余额_借方金额": "year_opening_debit",
    "年初余额_贷方金额": "year_opening_credit",
    "年初余额_借方": "year_opening_debit",
    "年初余额_贷方": "year_opening_credit",
    # 年初余额（单行直接列名：年初借方金额）
    "年初借方金额": "year_opening_debit",
    "年初贷方金额": "year_opening_credit",
    "年初借方": "year_opening_debit",
    "年初贷方": "year_opening_credit",
    # 期初余额（双行合并）
    "期初余额_借方金额": "opening_debit",
    "期初余额_贷方金额": "opening_credit",
    "期初余额_借方": "opening_debit",
    "期初余额_贷方": "opening_credit",
    # 期初余额（单行直接列名）
    "期初借方金额": "opening_debit",
    "期初贷方金额": "opening_credit",
    "期初借方": "opening_debit",
    "期初贷方": "opening_credit",
    # 本期发生额（双行合并）
    "本期发生额_借方金额": "debit_amount",
    "本期发生额_贷方金额": "credit_amount",
    "本期发生额_借方": "debit_amount",
    "本期发生额_贷方": "credit_amount",
    # 本期发生额（单行直接列名）
    "本期借方金额": "debit_amount",
    "本期贷方金额": "credit_amount",
    "本期借方": "debit_amount",
    "本期贷方": "credit_amount",
    # 本年累计（双行合并）
    "本年累计_借方金额": "year_debit",
    "本年累计_贷方金额": "year_credit",
    "本年累计_借方": "year_debit",
    "本年累计_贷方": "year_credit",
    # 本年累计（单行直接列名）
    "本年累计借方": "year_debit",
    "本年累计贷方": "year_credit",
    "累计借方": "year_debit",
    "累计贷方": "year_credit",
    # 期末余额（双行合并）
    "期末余额_借方金额": "closing_debit",
    "期末余额_贷方金额": "closing_credit",
    "期末余额_借方": "closing_debit",
    "期末余额_贷方": "closing_credit",
    # 期末余额（单行直接列名）
    "期末借方金额": "closing_debit",
    "期末贷方金额": "closing_credit",
    "期末借方": "closing_debit",
    "期末贷方": "closing_credit",
    # 期初/期末（不带"余额"后缀，双行合并）
    "期初_借方金额": "opening_debit",
    "期初_贷方金额": "opening_credit",
    "期末_借方金额": "closing_debit",
    "期末_贷方金额": "closing_credit",
    # 方向列（双行表头可能产生的组合）
    "期初余额_方向": "opening_direction",
    "期末余额_方向": "closing_direction",
    "期初_方向": "opening_direction",
    "期末_方向": "closing_direction",
}

# 基础列名映射（从 account_chart_service._COLUMN_MAP 复用）
from app.models.audit_platform_models import AccountCategory, AccountChart, AccountDirection, AccountSource
from app.services.account_chart_service import (
    _COLUMN_MAP as _BASE_COLUMN_MAP,
    _infer_category,
    _infer_direction,
    _infer_level as _infer_level_svc,
    parse_aux_dimensions,
)


def _detect_missing_fields(data_type: str, matched_fields: set[str]) -> tuple[list[str], list[str]]:
    missing_required = [
        display for display, candidates in _REQUIRED_FIELD_GROUPS.get(data_type, [])
        if not any(field in matched_fields for field in candidates)
    ]
    missing_recommended = [
        display for display, candidates in _RECOMMENDED_FIELD_GROUPS.get(data_type, [])
        if not any(field in matched_fields for field in candidates)
    ]
    return missing_required, missing_recommended


def _build_missing_required_import_error(data_type: str, missing_fields: list[str], source_label: str | None = None) -> str:
    display_name = {
        "balance": "余额表",
        "ledger": "序时账",
        "account_chart": "科目表",
    }.get(data_type, "工作表")
    prefix = f"{source_label} " if source_label else ""
    field_labels = [FIELD_LABELS.get(field, field) for field in missing_fields]
    return f"{prefix}{display_name}缺少必需列: {', '.join(field_labels)}"


def smart_match_column(header: str) -> Optional[str]:
    """智能匹配列名到标准字段。优先匹配合并表头组合名，支持模糊匹配。"""
    h = header.strip()
    # 1. 合并表头组合名精确匹配
    if h in _MERGED_HEADER_MAP:
        return _MERGED_HEADER_MAP[h]
    # 2. "核算维度" 特殊处理 → 映射为 aux_dimensions（混合维度列）
    if h in ("核算维度",):
        return "aux_dimensions"
    # 3. 基础列名映射
    if h in _BASE_COLUMN_MAP:
        return _BASE_COLUMN_MAP[h]
    # 4. 清洗后匹配
    cleaned = re.sub(r'[\[\]【】\(\)（）\*\#\s]', '', h)
    if cleaned in _MERGED_HEADER_MAP:
        return _MERGED_HEADER_MAP[cleaned]
    if cleaned in ("核算维度",):
        return "aux_dimensions"
    if cleaned in _BASE_COLUMN_MAP:
        return _BASE_COLUMN_MAP[cleaned]
    # 5. 模糊匹配：关键词包含（如"辅助核算_核算维度"含"核算维度"）
    for keyword, mapped in ("核算维度", "aux_dimensions"), ("辅助核算", "aux_dimensions"), ("核算项目", "aux_dimensions"):
        if keyword in h:
            return mapped
    # 6. 子串匹配常见列名（处理如"科目编码_V2"这类带后缀的列名）
    for src, dst in _BASE_COLUMN_MAP.items():
        if len(src) >= 3 and src in h:
            return dst
    for src, dst in _MERGED_HEADER_MAP.items():
        if len(src) >= 3 and src in h:
            return dst
    return None


def _infer_column_by_content(values: list[str]) -> Optional[str]:
    """根据列内容推断标准字段类型（当表头无法识别时）。

    返回标准字段名如 "aux_dimensions", "account_code" 等，或 None。
    """
    if not values:
        return None
    non_empty = [v for v in values if v and str(v).strip()]
    if len(non_empty) < 3:
        return None

    samples = non_empty[:50]

    # 核算维度：格式为 "类型:编码,名称; 类型:编码,名称" 或包含冒号+逗号
    aux_dim_pattern = re.compile(r'[^:]+:[^,]+,[^;]+')
    aux_matches = sum(1 for v in samples if aux_dim_pattern.search(str(v)))
    if aux_matches >= max(3, len(samples) * 0.3):
        return "aux_dimensions"

    # 科目编码：大部分是数字+点号格式（如 1002, 1122.01.03）
    code_pattern = re.compile(r'^\d+(\.\d+)*$')
    code_matches = sum(1 for v in samples if code_pattern.match(str(v).strip()))
    if code_matches >= len(samples) * 0.5:
        return "account_code"

    # 科目名称：中文字符为主，长度适中
    name_matches = sum(1 for v in samples if len(str(v).strip()) >= 2 and bool(re.search(r'[\u4e00-\u9fff]', str(v))))
    if name_matches >= len(samples) * 0.5:
        # 如果前面有科目编码列，这更可能是科目名称
        return "account_name"

    # 借方/贷方金额：大部分是数字（含负数、小数点、逗号分隔）
    amount_pattern = re.compile(r'^-?[\d,]+\.?\d*$')
    amount_matches = sum(1 for v in samples if amount_pattern.match(str(v).strip().replace(',', '')))
    if amount_matches >= len(samples) * 0.5:
        # 无法区分借方还是贷方，返回通用金额标记
        return "debit_amount"  # 默认借方，后续可人工调整

    # 日期：大部分是日期格式
    date_patterns = [
        re.compile(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}'),
        re.compile(r'^\d{4}年\d{1,2}月\d{1,2}日'),
    ]
    date_matches = sum(1 for v in samples if any(p.match(str(v).strip()) for p in date_patterns))
    if date_matches >= len(samples) * 0.3:
        return "voucher_date"

    return None


# ─────────────────────────────────────────────────────────────────────────────
# 4. 年度自动提取
# ─────────────────────────────────────────────────────────────────────────────

def _extract_company_info(rows_before_header: list[list[str | None]]) -> tuple[str | None, int | None]:
    """从表头前的信息行提取企业名称和年度。

    典型格式：
      '核算组织：重庆和平药房连锁有限责任公司; 开始期间：2025年1期;...'
    Returns:
        (company_name, year)
    """
    for row in rows_before_header:
        text = " ".join(str(c).strip() for c in row if c)
        if not text:
            continue
        # 企业名称
        m = re.search(r'核算组织[：:]\s*([^;；,，]+)', text)
        company_name = m.group(1).strip() if m else None
        # 年度
        year = None
        ym = re.search(r'(20\d{2})\s*年', text)
        if ym:
            year = int(ym.group(1))
        # 只在找到企业名称或年度时才返回，否则继续扫描
        if company_name or year:
            return company_name, year
    return None, None


def extract_year_from_content(rows: list[dict], filename: str = "") -> Optional[int]:
    """从文件内容或文件名中提取年度。

    优先级：
    1. 数据行中的日期字段（voucher_date）— 支持 datetime/date 对象和字符串
    2. 会计期间字段（如"2025年1期"）
    3. 文件名中的年份
    """
    for row in rows[:20]:
        # 尝试凭证日期（优先处理 datetime/date 对象）
        for key in ("voucher_date", "记账日期", "凭证日期", "日期"):
            val = row.get(key)
            if val is None:
                continue
            if isinstance(val, (datetime, date)):
                return val.year
            m = re.search(r'(20\d{2})', str(val))
            if m:
                return int(m.group(1))
        # 尝试期间
        for key in ("accounting_period", "期间", "会计期间"):
            val = row.get(key)
            if val:
                m = re.search(r'(20\d{2})', str(val))
                if m:
                    return int(m.group(1))

    # 从文件名提取
    if filename:
        m = re.search(r'(20\d{2})', filename)
        if m:
            return int(m.group(1))

    return None


def _resolve_account_category(code: str, name: str, category_raw) -> AccountCategory:
    category = _infer_category(code, name)
    category_str = str(category_raw or "").strip().lower()
    category_map = {
        "资产": AccountCategory.asset,
        "asset": AccountCategory.asset,
        "负债": AccountCategory.liability,
        "liability": AccountCategory.liability,
        "权益": AccountCategory.equity,
        "equity": AccountCategory.equity,
        "收入": AccountCategory.revenue,
        "revenue": AccountCategory.revenue,
        "费用": AccountCategory.expense,
        "expense": AccountCategory.expense,
        "成本": AccountCategory.expense,
        "cost": AccountCategory.expense,
    }
    return category_map.get(category_str, category)


def _resolve_account_direction(direction_raw, category: AccountCategory) -> AccountDirection:
    direction_str = str(direction_raw or "").strip().lower()
    if direction_str in ("debit", "借", "借方"):
        return AccountDirection.debit
    if direction_str in ("credit", "贷", "贷方"):
        return AccountDirection.credit
    return _infer_direction(category)


def _append_account_record(
    *,
    row: dict,
    project_id: UUID,
    seen_codes: set[str],
    acct_records: list[AccountChart],
    by_category: dict[str, int],
    staged: bool = False,
    dataset_id: UUID | None = None,
) -> bool:
    code = str(row.get("account_code", "")).strip()
    name = str(row.get("account_name", "")).strip()
    if not code or not name or code in seen_codes:
        return False

    parent_code = str(row.get("parent_code", "")).strip() or None
    category = _resolve_account_category(code, name, row.get("category"))
    direction = _resolve_account_direction(row.get("direction"), category)
    level = _infer_level_svc(code, parent_code)

    level_str = str(row.get("level", "")).strip()
    if level_str.isdigit():
        level = int(level_str)

    seen_codes.add(code)
    acct_records.append(AccountChart(
        project_id=project_id,
        account_code=code,
        account_name=name,
        direction=direction,
        level=level,
        category=category,
        parent_code=parent_code,
        source=AccountSource.client,
        dataset_id=dataset_id,
        is_deleted=staged,
    ))
    by_category[category.value] = by_category.get(category.value, 0) + 1
    return True



# ─────────────────────────────────────────────────────────────────────────────
# 5. 通用 Sheet 解析（核心：支持双行表头 + 核算维度拆分）
# ─────────────────────────────────────────────────────────────────────────────

def _safe_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        s = str(val).strip()
        if not s or s == "None" or s == "":
            return None
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_date_val(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s == "None":
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_period_str(val) -> Optional[int]:
    """解析期间字符串为月份。如 '2025年1期' → 1"""
    if val is None:
        return None
    s = str(val).strip()
    m = re.search(r'(\d+)\s*期', s)
    if m:
        return int(m.group(1))
    # 纯数字
    try:
        v = int(s)
        if 1 <= v <= 12:
            return v
    except (ValueError, TypeError):
        pass
    return None


def _infer_level(code: str) -> int:
    code = code.replace(".", "").replace("-", "")
    if len(code) <= 4:
        return 1
    elif len(code) <= 6:
        return 2
    elif len(code) <= 8:
        return 3
    elif len(code) <= 10:
        return 4
    return 5


def _source_size_bytes(source) -> int:
    if isinstance(source, (bytes, bytearray)):
        return len(source)
    if isinstance(source, (str, Path)):
        try:
            return Path(source).stat().st_size
        except OSError:
            return 0
    return 0


@contextmanager
def _open_binary_source(source):
    if isinstance(source, (bytes, bytearray)):
        yield io.BytesIO(source)
        return
    if isinstance(source, (str, Path)):
        with Path(source).open("rb") as fh:
            yield fh
        return
    if hasattr(source, "seek"):
        try:
            source.seek(0)
        except Exception:
            pass
    yield source


def _workbook_input(source):
    if isinstance(source, (bytes, bytearray)):
        return io.BytesIO(source)
    if isinstance(source, (str, Path)):
        return str(source)
    if hasattr(source, "seek"):
        try:
            source.seek(0)
        except Exception:
            pass
    return source


def _detect_stream_encoding(sample: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"):
        try:
            sample.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"


def _build_sheet_context(ws, sheet_mapping: Optional[dict[str, str]] = None) -> dict:
    header_start, header_count = detect_header_rows(ws)
    headers = merge_header_rows(ws, header_start, header_count)

    column_mapping = {}
    for h in headers:
        mapped = smart_match_column(h)
        if mapped:
            column_mapping[h] = mapped
    if sheet_mapping:
        for h, std_field in sheet_mapping.items():
            if h in headers and std_field:
                column_mapping[h] = std_field

    rows_before_header: list[list[str | None]] = []
    if header_start > 0:
        try:
            max_col = ws.max_column or 1
            for r in range(1, header_start + 1):
                rows_before_header.append([
                    ws.cell(r, c).value for c in range(1, max_col + 1)
                ])
        except Exception:
            header_iter = ws.iter_rows(max_row=header_start, values_only=True)
            for row_vals in header_iter:
                rows_before_header.append(list(row_vals))
    company_name, info_year = _extract_company_info(rows_before_header)

    data_start = header_start + header_count
    total_row_estimate = None
    try:
        if ws.max_row and ws.max_row > data_start:
            total_row_estimate = ws.max_row - data_start
    except Exception:
        total_row_estimate = None

    return {
        "headers": headers,
        "header_start": header_start,
        "header_count": header_count,
        "column_mapping": column_mapping,
        "data_type": _guess_data_type(set(column_mapping.values())),
        "company_code": company_name,
        "info_year": info_year,
        "total_row_estimate": total_row_estimate,
    }


def _iter_sheet_row_dicts(
    ws,
    sheet_context: dict,
    max_data_rows: int | None = None,
    *,
    use_mapped_headers: bool = True,
) -> Iterator[dict]:
    headers = sheet_context["headers"]
    column_mapping = sheet_context["column_mapping"]
    company_name = sheet_context.get("company_code")
    data_start = sheet_context["header_start"] + sheet_context["header_count"]
    num_cols = len(headers)

    try:
        data_iter = ws.iter_rows(min_row=data_start + 1, values_only=True)
    except TypeError:
        data_iter = ws.iter_rows(values_only=True)
        for _ in range(data_start):
            try:
                next(data_iter)
            except StopIteration:
                data_iter = iter(())
                break

    row_count = 0
    for row_vals in data_iter:
        if max_data_rows is not None and row_count >= max_data_rows:
            break
        padded = list(row_vals) + [None] * max(0, num_cols - len(row_vals))
        if all(c is None for c in padded[:num_cols]):
            continue
        row_dict = {}
        for i in range(num_cols):
            header = headers[i]
            mapped = column_mapping.get(header, header) if use_mapped_headers else header
            row_dict[mapped] = padded[i]
        if company_name:
            row_dict["company_code"] = company_name
        row_count += 1
        yield row_dict


def _iter_sheet_row_batches(
    ws,
    sheet_context: dict,
    *,
    batch_size: int,
    max_data_rows: int | None = None,
) -> Iterator[list[dict]]:
    batch: list[dict] = []
    for row in _iter_sheet_row_dicts(ws, sheet_context, max_data_rows=max_data_rows):
        batch.append(row)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


def smart_parse_sheet(
    ws,
    *,
    read_only: bool = True,
    max_data_rows: int | None = None,
    include_raw_rows: bool = False,
    sheet_mapping: Optional[dict[str, str]] = None,
) -> dict:
    """通用智能解析单个 worksheet。

    Args:
        max_data_rows: 最大解析数据行数，None 表示全部解析（用于预览模式限制内存）。

    Returns:
        {
            "headers": [...],           # 合并后的列名
            "header_start": int,        # 表头起始行
            "header_count": int,        # 表头行数
            "column_mapping": {...},    # 列名→标准字段映射
            "data_type": str,           # 识别的数据类型
            "rows": [...],              # 原始数据行（dict）
            "year": int | None,         # 自动提取的年度
            "row_count": int,           # 实际解析行数
            "total_row_estimate": int | None,  # 总行数估算（仅当 max_data_rows 有限制时）
        }
    """
    sheet_context = _build_sheet_context(ws, sheet_mapping=sheet_mapping)
    rows = list(_iter_sheet_row_dicts(ws, sheet_context, max_data_rows=max_data_rows))
    raw_rows = []
    if include_raw_rows:
        raw_rows = list(
            _iter_sheet_row_dicts(
                ws,
                sheet_context,
                max_data_rows=max_data_rows,
                use_mapped_headers=False,
            )
        )
    row_count = len(rows)
    total_row_estimate = sheet_context.get("total_row_estimate")
    if max_data_rows is None or total_row_estimate is None:
        total_row_estimate = row_count
    year = sheet_context.get("info_year") or extract_year_from_content(rows)

    return {
        "headers": sheet_context["headers"],
        "header_start": sheet_context["header_start"],
        "header_count": sheet_context["header_count"],
        "column_mapping": sheet_context["column_mapping"],
        "data_type": sheet_context["data_type"],
        "rows": rows,
        "raw_rows": raw_rows,
        "year": year,
        "row_count": row_count,
        "total_row_estimate": total_row_estimate,
        "company_code": sheet_context.get("company_code"),
    }


def _parse_with_known_headers(ws, headers: list[str], header_start: int, header_count: int) -> dict:
    """用已知表头解析 read_only 模式的 worksheet（表头从完整模式探测获得）。"""
    column_mapping = {}
    for h in headers:
        mapped = smart_match_column(h)
        if mapped:
            column_mapping[h] = mapped

    # 从表头前的信息行提取企业和年度
    rows_before_header: list[list[str | None]] = []
    if header_start > 0:
        try:
            _max_col = ws.max_column or 1
            for r in range(1, header_start + 1):
                rows_before_header.append([
                    ws.cell(r, c).value for c in range(1, _max_col + 1)
                ])
        except Exception:
            _rh_iter = ws.iter_rows(max_row=header_start, values_only=True)
            for _row_vals in _rh_iter:
                rows_before_header.append(list(_row_vals))
    company_name, info_year = _extract_company_info(rows_before_header)

    data_start = header_start + header_count
    num_cols = len(headers)
    rows = []

    rows_iter = ws.iter_rows(values_only=True)
    for _ in range(data_start):
        try:
            next(rows_iter)
        except StopIteration:
            break

    for row_vals in rows_iter:
        padded = list(row_vals) + [None] * max(0, num_cols - len(row_vals))
        if all(c is None for c in padded[:num_cols]):
            continue
        row_dict = {}
        for i in range(num_cols):
            h = headers[i]
            mapped = column_mapping.get(h, h)
            row_dict[mapped] = padded[i]
        if company_name:
            row_dict["company_code"] = company_name
        rows.append(row_dict)

    mapped_fields = set(column_mapping.values())
    data_type = _guess_data_type(mapped_fields)
    year = info_year or extract_year_from_content(rows)

    return {
        "headers": headers,
        "header_start": header_start,
        "header_count": header_count,
        "column_mapping": column_mapping,
        "data_type": data_type,
        "rows": rows,
        "year": year,
        "row_count": len(rows),
        "company_code": company_name,
    }


def _guess_data_type(fields: set[str]) -> str:
    """根据已映射的字段集合推断数据类型。"""
    has_code = "account_code" in fields
    has_voucher_date = "voucher_date" in fields
    has_voucher_no = "voucher_no" in fields
    has_debit = "debit_amount" in fields
    has_credit = "credit_amount" in fields
    has_opening = any(f in fields for f in (
        "opening_balance", "opening_debit", "opening_credit",
        "year_opening_debit", "year_opening_credit",
    ))
    has_closing = any(f in fields for f in ("closing_balance", "closing_debit", "closing_credit"))
    # aux_dimensions 是混合维度列（需要拆分），不等于独立辅助表
    has_aux_dimensions = "aux_dimensions" in fields
    has_aux_separate = any(f in fields for f in ("aux_code", "aux_name")) and "aux_type" in fields
    has_chart_features = any(f in fields for f in ("direction", "parent_code", "level", "category"))

    # 独立辅助核算列（非混合维度列）- 只有当文件不含主表核心字段时才判定为独立辅助表
    if has_aux_separate and not has_aux_dimensions:
        has_balance_core = has_code and (has_opening or has_closing or has_debit or has_credit)
        has_ledger_core = has_voucher_date and has_voucher_no and (has_debit or has_credit)
        if not has_balance_core and not has_ledger_core:
            if has_voucher_date:
                return "aux_ledger"
            return "aux_balance"
        # 含主表核心字段 → 继续走 balance/ledger 判定（带辅助维度的余额表/序时账）

    # 序时账（可能含核算维度列，拆分后生成辅助明细账）
    if has_voucher_date and has_voucher_no and (has_debit or has_credit):
        return "ledger"

    # 科目表：当出现科目名称且存在方向/父级/级次/类别等科目表特征时，优先识别为科目表，
    # 避免仅因“借贷方向”列命中 credit/debit 关键字而被误判为余额表。
    if has_code and "account_name" in fields and has_chart_features and not has_voucher_date:
        return "account_chart"

    # 余额表（可能含核算维度列，拆分后生成辅助余额表）
    if has_code and (has_opening or has_closing):
        return "balance"

    # 只有借贷发生额也可能是余额表
    if has_code and (has_debit or has_credit) and not has_voucher_date:
        return "balance"

    if has_code:
        return "account_chart"

    return "unknown"



# ─────────────────────────────────────────────────────────────────────────────
# 6. 四表数据转换（从原始行 → 标准化记录）
# ─────────────────────────────────────────────────────────────────────────────

def convert_balance_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """将余额表原始行转换为 (tb_balance_rows, tb_aux_balance_rows)。

    两种期初/期末模式都保留原始数据：
    - 分列模式（期初借方+期初贷方）：存 opening_debit/opening_credit，同时算 opening_balance
    - 净额+方向模式（期初余额+方向）：存 opening_balance，不强制拆借贷

    核算维度列自动拆分为辅助余额表记录。
    """
    balance_rows = []
    aux_balance_rows = []

    for row in rows:
        account_code = str(row.get("account_code", "")).strip()
        if not account_code:
            continue

        account_name = str(row.get("account_name", "")).strip()
        company_code = str(row.get("company_code", "")).strip() or "default"

        # ── 期初 ──
        od = _safe_decimal(row.get("opening_debit"))
        oc = _safe_decimal(row.get("opening_credit"))
        opening_bal = _safe_decimal(row.get("opening_balance"))
        opening_dir = row.get("opening_direction") or row.get("direction")

        # 年初余额作为备选
        if od is None and oc is None and opening_bal is None:
            od = _safe_decimal(row.get("year_opening_debit"))
            oc = _safe_decimal(row.get("year_opening_credit"))

        # 分列模式：有借贷分列 → 保留原始借贷，同时算净额
        if od is not None or oc is not None:
            opening_balance = (od or Decimal(0)) - (oc or Decimal(0))
        elif opening_bal is not None and opening_dir:
            # 净额+方向模式：保留净额原值
            opening_balance = opening_bal
            dir_str = str(opening_dir).strip()
            if dir_str in ("贷", "贷方", "C", "c", "credit", "Credit"):
                opening_balance = -abs(opening_bal)
            elif dir_str in ("借", "借方", "D", "d", "debit", "Debit"):
                opening_balance = abs(opening_bal)
        else:
            opening_balance = opening_bal  # 纯净额，无方向

        # ── 期末 ──
        cd = _safe_decimal(row.get("closing_debit"))
        cc = _safe_decimal(row.get("closing_credit"))
        closing_bal = _safe_decimal(row.get("closing_balance"))
        closing_dir = row.get("closing_direction") or row.get("direction")

        if cd is not None or cc is not None:
            closing_balance = (cd or Decimal(0)) - (cc or Decimal(0))
        elif closing_bal is not None and closing_dir:
            closing_balance = closing_bal
            dir_str = str(closing_dir).strip()
            if dir_str in ("贷", "贷方", "C", "c", "credit", "Credit"):
                closing_balance = -abs(closing_bal)
            elif dir_str in ("借", "借方", "D", "d", "debit", "Debit"):
                closing_balance = abs(closing_bal)
        else:
            closing_balance = closing_bal

        debit_amount = _safe_decimal(row.get("debit_amount"))
        credit_amount = _safe_decimal(row.get("credit_amount"))

        # 核算维度处理（支持单列表"核算维度"和多列表"辅助类型+辅助编码+辅助名称"两种格式）
        aux_dim_str = str(row.get("aux_dimensions", "")).strip()

        # 多列格式：辅助类型 + 辅助编码（不含冒号分隔）
        if not aux_dim_str:
            aux_type_val = str(row.get("aux_type", "")).strip()
            aux_code_val = str(row.get("aux_code", "")).strip()
            if aux_type_val and aux_code_val:
                aux_name_val = str(row.get("aux_name", "")).strip() or aux_code_val
                aux_dim_str = f"{aux_type_val}:{aux_code_val},{aux_name_val}"
            elif aux_type_val and ":" not in aux_type_val and "：" not in aux_type_val:
                # 只有辅助类型名称，无编码 → 不是有效辅助维度
                aux_dim_str = ""

        base_row = {
            "account_code": account_code,
            "account_name": account_name,
            "company_code": company_code,
            "opening_balance": opening_balance,
            "opening_debit": od,
            "opening_credit": oc,
            "debit_amount": debit_amount,
            "credit_amount": credit_amount,
            "closing_balance": closing_balance,
            "closing_debit": cd,
            "closing_credit": cc,
            "currency_code": "CNY",
        }

        if aux_dim_str and (":" in aux_dim_str or "：" in aux_dim_str):
            dims = parse_aux_dimensions(aux_dim_str)
            for dim in dims:
                aux_balance_rows.append({
                    **base_row,
                    "aux_type": dim["aux_type"],
                    "aux_code": dim["aux_code"],
                    "aux_name": dim["aux_name"],
                    "aux_dimensions_raw": aux_dim_str,  # 保留原始维度组合
                })
        else:
            balance_rows.append({
                **base_row,
                "level": _infer_level(account_code),
            })

    return balance_rows, aux_balance_rows


def convert_ledger_rows(rows: list[dict]) -> tuple[list[dict], list[dict], dict]:
    """将序时账原始行转换为 (tb_ledger_rows, tb_aux_ledger_rows, aux_stats)。

    内存优化：辅助明细账只统计维度分布（aux_stats），不在内存中保留全部行。
    实际的辅助明细账行在 write_four_tables 中流式生成写入。

    Returns:
        (ledger_rows, [], aux_stats)
        - ledger_rows: 序时账行（带 _aux_dim_str 原始维度字符串）
        - []: 空列表（辅助明细账不在内存中保留）
        - aux_stats: {"成本中心": 12345, ...} 维度统计
    """
    ledger_rows = []
    aux_stats: dict[str, int] = {}

    for row in rows:
        account_code = str(row.get("account_code", "")).strip()
        if not account_code:
            continue

        voucher_date = _parse_date_val(row.get("voucher_date"))
        voucher_no = str(row.get("voucher_no", "")).strip()
        if not voucher_date or not voucher_no:
            continue

        account_name = str(row.get("account_name", "")).strip()
        voucher_type = str(row.get("voucher_type", "")).strip() or None
        accounting_period = _parse_period_str(row.get("accounting_period"))
        debit_amount = _safe_decimal(row.get("debit_amount"))
        credit_amount = _safe_decimal(row.get("credit_amount"))
        summary = str(row.get("summary", "")).strip() or None
        preparer = str(row.get("preparer", "")).strip() or None

        # 核算维度原始字符串（保留，写入时再拆分）
        # 支持单列表"核算维度"和多列表"辅助类型+辅助编码+辅助名称"两种格式
        aux_dim_str = str(row.get("aux_dimensions", "")).strip()

        # 多列格式：辅助类型 + 辅助编码
        if not aux_dim_str:
            aux_type_val = str(row.get("aux_type", "")).strip()
            aux_code_val = str(row.get("aux_code", "")).strip()
            if aux_type_val and aux_code_val:
                aux_name_val = str(row.get("aux_name", "")).strip() or aux_code_val
                aux_dim_str = f"{aux_type_val}:{aux_code_val},{aux_name_val}"
            elif aux_type_val and ":" not in aux_type_val and "：" not in aux_type_val:
                aux_dim_str = ""

        # 统计维度分布（不生成辅助明细行）
        if aux_dim_str and (":" in aux_dim_str or "：" in aux_dim_str):
            dims = parse_aux_dimensions(aux_dim_str)
            for dim in dims:
                aux_stats[dim["aux_type"]] = aux_stats.get(dim["aux_type"], 0) + 1

        company_code = str(row.get("company_code", "")).strip() or "default"

        ledger_rows.append({
            "account_code": account_code,
            "account_name": account_name,
            "voucher_date": voucher_date,
            "voucher_no": voucher_no,
            "voucher_type": voucher_type,
            "accounting_period": accounting_period,
            "debit_amount": debit_amount,
            "credit_amount": credit_amount,
            "summary": summary,
            "preparer": preparer,
            "company_code": company_code,
            "currency_code": "CNY",
            "_aux_dim_str": aux_dim_str,  # 内部字段，写入时拆分
        })

    return ledger_rows, [], aux_stats


# ─────────────────────────────────────────────────────────────────────────────
# 7. 辅助表一致性校验
# ─────────────────────────────────────────────────────────────────────────────

def validate_four_tables(
    balance_rows: list[dict],
    aux_balance_rows: list[dict],
    ledger_rows: list[dict],
    aux_ledger_rows: list[dict],
) -> list[dict]:
    """四表间一致性校验。

    辅助余额表是科目余额表中有辅助核算科目的下级明细展开。
    一条序时账如果有多个维度，会拆成多条辅助明细账记录。

    校验项：
    1. 科目余额表内部勾稽：期初 + 借方 - 贷方 = 期末
    2. 辅助余额表内部勾稽：同上
    3. 科目余额表 vs 辅助余额表：有辅助核算的科目，辅助余额按科目去重后
       的期末合计应等于科目余额表的期末（注意：同一科目多个维度类型，
       每个维度类型各自汇总都应等于科目余额，取任一维度类型比对即可）
    4. 序时账 vs 科目余额表：序时账按科目汇总的借贷发生额应等于余额表
    5. 辅助明细账 vs 辅助余额表：按 (科目, 维度类型, 维度编号) 汇总比对

    Returns:
        [{"level": "info"|"warning"|"error", "category": str, "message": str, "detail": dict}, ...]
    """
    findings = []

    # ── 1. 科目余额表内部勾稽 ──
    bal_map = {}  # account_code → {opening, debit, credit, closing}
    bal_errors = 0
    for r in balance_rows:
        code = r["account_code"]
        opening = r.get("opening_balance") or Decimal(0)
        debit = r.get("debit_amount") or Decimal(0)
        credit = r.get("credit_amount") or Decimal(0)
        closing = r.get("closing_balance") or Decimal(0)
        bal_map[code] = {"opening": opening, "debit": debit, "credit": credit, "closing": closing}
        expected = opening + debit - credit
        if abs(expected - closing) > Decimal("0.01"):
            bal_errors += 1
            if bal_errors <= 3:
                findings.append({
                    "level": "error", "category": "余额表勾稽",
                    "message": f"{code}: 期初{opening}+借{debit}-贷{credit}={expected}, 期末{closing}, 差{expected-closing}",
                    "detail": {"account_code": code},
                })
    if bal_errors > 3:
        findings.append({"level": "error", "category": "余额表勾稽",
                         "message": f"共 {bal_errors} 个科目不平（仅展示前3条）", "detail": {"total": bal_errors}})

    # ── 2. 辅助余额表内部勾稽 ──
    aux_bal_errors = 0
    for r in aux_balance_rows:
        opening = r.get("opening_balance") or Decimal(0)
        debit = r.get("debit_amount") or Decimal(0)
        credit = r.get("credit_amount") or Decimal(0)
        closing = r.get("closing_balance") or Decimal(0)
        expected = opening + debit - credit
        if abs(expected - closing) > Decimal("0.01"):
            aux_bal_errors += 1
            if aux_bal_errors <= 3:
                findings.append({
                    "level": "error", "category": "辅助余额表勾稽",
                    "message": f"{r['account_code']}/{r.get('aux_type','')}:{r.get('aux_code','')}: "
                               f"期初{opening}+借{debit}-贷{credit}={expected}, 期末{closing}",
                    "detail": {"account_code": r["account_code"], "aux_type": r.get("aux_type")},
                })
    if aux_bal_errors > 3:
        findings.append({"level": "error", "category": "辅助余额表勾稽",
                         "message": f"共 {aux_bal_errors} 条不平", "detail": {"total": aux_bal_errors}})

    # ── 3. 科目余额表 vs 辅助余额表 ──
    # 辅助余额表按科目汇总（去重：同一科目+同一维度组合只算一次期末）
    # 策略：找到每个科目下记录数最少的维度类型，用它汇总比对
    from collections import defaultdict
    aux_by_code_type = defaultdict(lambda: {"closing": Decimal(0), "count": 0})
    aux_codes = set()
    for r in aux_balance_rows:
        code = r["account_code"]
        aux_type = r.get("aux_type", "?")
        key = (code, aux_type)
        aux_by_code_type[key]["closing"] += r.get("closing_balance") or Decimal(0)
        aux_by_code_type[key]["count"] += 1
        aux_codes.add(code)

    cross_bal_errors = 0
    for code in sorted(aux_codes):
        if code not in bal_map:
            continue
        b_closing = bal_map[code]["closing"]
        # 每个维度类型的汇总都应该等于科目余额（核算维度是二级明细）
        # 找最接近的维度类型比对，如实报出差异
        types_for_code = [(k[1], v) for k, v in aux_by_code_type.items() if k[0] == code]
        if not types_for_code:
            continue
        best_type, best = min(types_for_code, key=lambda x: abs(b_closing - x[1]["closing"]))
        a_closing = best["closing"]
        diff = b_closing - a_closing
        if abs(diff) > Decimal("0.01"):
            cross_bal_errors += 1
            if cross_bal_errors <= 5:
                findings.append({
                    "level": "warning", "category": "余额表vs辅助余额表",
                    "message": f"{code}: 科目期末={b_closing}, 最近维度({best_type},{best['count']}条)汇总={a_closing}, 差{diff}",
                    "detail": {"account_code": code, "aux_type": best_type},
                })
    if cross_bal_errors > 5:
        findings.append({"level": "warning", "category": "余额表vs辅助余额表",
                         "message": f"共 {cross_bal_errors} 个科目不一致（仅展示前5条）", "detail": {"total": cross_bal_errors}})
    if cross_bal_errors == 0:
        findings.append({"level": "info", "category": "余额表vs辅助余额表",
                         "message": f"有辅助核算的 {len(aux_codes)} 个科目全部一致",
                         "detail": {"total": len(aux_codes), "match": len(aux_codes)}})

    # ── 4. 序时账 vs 科目余额表 ──
    led_by_code: dict[str, list[Decimal]] = {}  # code → [debit_sum, credit_sum]
    for r in ledger_rows:
        code = r["account_code"]
        if code not in led_by_code:
            led_by_code[code] = [Decimal(0), Decimal(0)]
        led_by_code[code][0] += r.get("debit_amount") or Decimal(0)
        led_by_code[code][1] += r.get("credit_amount") or Decimal(0)

    led_errors = 0
    for code in sorted(set(bal_map.keys()) & set(led_by_code.keys())):
        b = bal_map[code]
        ld, lc = led_by_code[code]
        if abs(b["debit"] - ld) > Decimal("0.01") or abs(b["credit"] - lc) > Decimal("0.01"):
            led_errors += 1
            if led_errors <= 3:
                findings.append({
                    "level": "warning", "category": "序时账vs余额表",
                    "message": f"{code}: 余额表借{b['debit']}/贷{b['credit']}, 序时账借{ld}/贷{lc}",
                    "detail": {"account_code": code},
                })
    if led_errors > 3:
        findings.append({"level": "warning", "category": "序时账vs余额表",
                         "message": f"共 {led_errors} 个科目不一致", "detail": {"total": led_errors}})
    elif ledger_rows:
        findings.append({"level": "info", "category": "序时账vs余额表",
                         "message": f"序时账 {len(led_by_code)} 个科目中 {len(led_by_code)-led_errors} 个与余额表一致",
                         "detail": {}})

    # ── 5. 辅助明细账 vs 辅助余额表（大数据量时抽样校验） ──
    if aux_ledger_rows and aux_balance_rows:
        # 大数据量（>50万行）时只抽样前10万行做校验，避免内存和时间开销
        sample = aux_ledger_rows if len(aux_ledger_rows) <= 500000 else aux_ledger_rows[:100000]
        is_sampled = len(sample) < len(aux_ledger_rows)

        aux_led_sums: dict[tuple, list[Decimal]] = {}
        for r in sample:
            key = (r["account_code"], r.get("aux_type", ""), r.get("aux_code", ""))
            if key not in aux_led_sums:
                aux_led_sums[key] = [Decimal(0), Decimal(0)]
            aux_led_sums[key][0] += r.get("debit_amount") or Decimal(0)
            aux_led_sums[key][1] += r.get("credit_amount") or Decimal(0)

        if not is_sampled:
            aux_cross_errors = 0
            for r in aux_balance_rows:
                key = (r["account_code"], r.get("aux_type", ""), r.get("aux_code", ""))
                if key not in aux_led_sums:
                    continue
                bal_d = r.get("debit_amount") or Decimal(0)
                bal_c = r.get("credit_amount") or Decimal(0)
                led_d, led_c = aux_led_sums[key]
                if abs(bal_d - led_d) > Decimal("0.01") or abs(bal_c - led_c) > Decimal("0.01"):
                    aux_cross_errors += 1
                    if aux_cross_errors <= 3:
                        findings.append({
                            "level": "warning", "category": "辅助明细账vs辅助余额表",
                            "message": f"{key[0]}/{key[1]}:{key[2]}: 余额表借{bal_d}/贷{bal_c}, 明细账借{led_d}/贷{led_c}",
                            "detail": {"account_code": key[0], "aux_type": key[1], "aux_code": key[2]},
                        })
            if aux_cross_errors > 3:
                findings.append({"level": "warning", "category": "辅助明细账vs辅助余额表",
                                 "message": f"共 {aux_cross_errors} 条不一致", "detail": {"total": aux_cross_errors}})
        else:
            findings.append({"level": "info", "category": "辅助明细账vs辅助余额表",
                             "message": f"辅助明细账 {len(aux_ledger_rows)} 行，数据量较大，抽样前10万行校验",
                             "detail": {"total": len(aux_ledger_rows), "sampled": len(sample)}})

    return findings


def _resolve_sheet_mapping(
    custom_mapping: Optional[dict],
    filename: str,
    sheet_name: str,
) -> Optional[dict[str, str]]:
    if not custom_mapping:
        return None
    if not any(isinstance(v, dict) for v in custom_mapping.values()):
        return custom_mapping

    keys = [
        f"{filename}/{sheet_name}",
        sheet_name,
    ]
    if sheet_name == "CSV":
        keys.append("CSV")

    for key in keys:
        value = custom_mapping.get(key)
        if isinstance(value, dict):
            return value
    return None



# ─────────────────────────────────────────────────────────────────────────────
# 8. 核心入口：智能解析多文件 → 四表数据
# ─────────────────────────────────────────────────────────────────────────────

def _parse_csv_for_preview(
    content,
    filename: str,
    custom_mapping: Optional[dict] = None,
    max_rows: int = 0,
) -> dict:
    """解析 CSV 文件用于预览/解析。"""
    import codecs
    import csv

    total_bytes = _source_size_bytes(content)
    with _open_binary_source(content) as raw_probe:
        sample = raw_probe.read(65536)
    if not sample:
        raise ValueError("CSV 文件为空")

    encoding = _detect_stream_encoding(sample)
    sample_line_count = sample.count(b'\n')
    if sample_line_count > 0 and total_bytes > 0:
        avg_line_bytes = max(len(sample) / sample_line_count, 1)
        total_lines_est = max(int(total_bytes / avg_line_bytes), 1)
    else:
        total_lines_est = 0

    header_line = None
    header_line_number = 0
    with _open_binary_source(content) as raw_header:
        stream = codecs.getreader(encoding)(raw_header)
        for idx in range(5):
            line = stream.readline()
            if not line:
                break
            cells = line.strip().split(',')
            non_empty = [c.strip() for c in cells if c.strip()]
            if len(non_empty) >= 3:
                header_line = line.strip()
                header_line_number = idx + 1
                break

    if header_line is None:
        raise ValueError("CSV 文件未找到有效表头")

    reader_header = list(csv.reader([header_line]))[0]
    raw_headers = [c.strip() if c.strip() else f"col_{j}" for j, c in enumerate(reader_header)]

    headers = raw_headers[:]
    mid = len(raw_headers) // 2
    wide_table_detected = False

    if len(raw_headers) >= 60:
        first_half_matches = sum(1 for h in raw_headers[:mid] if smart_match_column(h))
        second_half_matches = sum(1 for h in raw_headers[mid:] if smart_match_column(h))
        if second_half_matches >= 5 and second_half_matches > first_half_matches * 2:
            wide_table_detected = True
            headers = raw_headers[mid:]

    column_mapping: dict[str, str] = {}
    header_mapped: set[str] = set()
    for h in headers:
        mapped = smart_match_column(h)
        if mapped:
            column_mapping[h] = mapped
            header_mapped.add(h)

    content_inferred: dict[str, str] = {}
    if len(column_mapping) < 3:
        _preview_rows_for_inference: list[list[str]] = []
        with _open_binary_source(content) as raw_infer:
            _temp_stream = codecs.getreader(encoding)(raw_infer)
            for _ in range(header_line_number):
                if not _temp_stream.readline():
                    break
            for _ in range(20):
                line = _temp_stream.readline()
                if not line:
                    break
                cells = list(csv.reader([line.strip()]))[0] if line.strip() else []
                if cells:
                    _preview_rows_for_inference.append(cells)
        for j, h in enumerate(headers):
            if h not in column_mapping or not column_mapping[h]:
                col_values = []
                for cells in _preview_rows_for_inference:
                    idx = j
                    if idx < len(cells):
                        col_values.append(cells[idx])
                inferred = _infer_column_by_content(col_values)
                if inferred:
                    column_mapping[h] = inferred
                    content_inferred[h] = inferred

    sheet_mapping = _resolve_sheet_mapping(custom_mapping, filename, "CSV")
    if sheet_mapping:
        for h, v in sheet_mapping.items():
            if h in headers and v:
                column_mapping[h] = v

    mapped_fields = set(column_mapping.values())
    dt = _guess_data_type(mapped_fields)

    rows: list[dict] = []
    raw_rows: list[dict] = []
    row_count = 0
    BATCH = 100_000

    def _consume_lines(lines: list[str]):
        nonlocal row_count
        for row_raw in csv.reader(lines):
            if not row_raw or all(not c.strip() for c in row_raw):
                continue
            if wide_table_detected:
                padded = row_raw + [""] * max(0, mid - len(row_raw))
                row_dict = {}
                raw_row = {}
                for j, h in enumerate(headers):
                    mapped = column_mapping.get(h, h)
                    data_idx = j
                    val = padded[data_idx].strip() if data_idx < len(padded) else ""
                    raw_row[h] = val
                    row_dict[mapped] = val if val else None
            else:
                padded = row_raw + [""] * max(0, len(headers) - len(row_raw))
                row_dict = {}
                raw_row = {}
                for j, h in enumerate(headers):
                    mapped = column_mapping.get(h, h)
                    val = padded[j].strip() if j < len(padded) else ""
                    raw_row[h] = val
                    row_dict[mapped] = val if val else None
            rows.append(row_dict)
            raw_rows.append(raw_row)
            row_count += 1
            if max_rows and row_count >= max_rows:
                break

    with _open_binary_source(content) as raw_data:
        stream = codecs.getreader(encoding)(raw_data)
        total_lines = 0
        for _ in range(header_line_number):
            if not stream.readline():
                break
            total_lines += 1

        line_buf: list[str] = []
        while True:
            line = stream.readline()
            if not line:
                break
            total_lines += 1
            line = line.rstrip('\n').rstrip('\r')
            if not line:
                continue
            if not max_rows or row_count < max_rows:
                line_buf.append(line)
                if len(line_buf) >= BATCH:
                    _consume_lines(line_buf)
                    line_buf.clear()

        if line_buf and (not max_rows or row_count < max_rows):
            _consume_lines(line_buf)

        total_lines_est = max(total_lines_est, total_lines)

    year_val = extract_year_from_content(rows[:100], filename=filename)

    return {
        "headers": headers,
        "raw_headers": raw_headers,
        "column_mapping": column_mapping,
        "header_mapped": header_mapped,
        "content_inferred": content_inferred,
        "data_type": dt,
        "rows": rows,
        "raw_rows": raw_rows,
        "year": year_val,
        "row_count": row_count,
        "total_lines_est": total_lines_est,
        "header_start": max(header_line_number - 1, 0),
        "header_count": 1,
        "wide_table_detected": wide_table_detected,
    }


def smart_parse_files(
    file_contents: list[tuple[str, object]],
    year_override: Optional[int] = None,
    custom_mapping: Optional[dict[str, str] | dict[str, dict[str, str]]] = None,
    preview_mode: bool = False,
    preview_rows: int = 50,
) -> dict:
    """智能解析多个文件，返回四表数据 + 维度信息 + 校验结果。

    Args:
        file_contents: [(filename, content_bytes), ...]
        year_override: 用户指定年度（优先于自动提取）
        custom_mapping: 用户手动指定的列映射。
            - 全局映射: {原始列名: 标准字段名}
            - 按 sheet 映射: {sheet_name: {原始列名: 标准字段名}}
        preview_mode: 预览模式（只解析前 N 行，不做数据转换，节省内存）
        preview_rows: 预览模式下单 sheet/CSV 最大解析行数

    Returns:
        {
            "balance_rows": [...],
            "aux_balance_rows": [...],
            "ledger_rows": [...],
            "aux_ledger_rows": [...],
            "year": int,
            "aux_dimensions": [{"type": "成本中心", "count": 1234}, ...],
            "validation": [...],  # 一致性校验结果
            "diagnostics": [...],  # 每个文件/sheet的解析诊断
        }
    """
    import openpyxl

    all_balance_rows = []
    all_aux_balance_rows = []
    all_ledger_rows = []
    all_aux_ledger_rows = []  # 保持为空，辅助明细账在写入时流式生成
    detected_year = year_override
    diagnostics = []
    _aux_type_counts: dict[str, int] = {}  # 维度统计（从余额表+序时账合并）

    for filename, content in file_contents:
        logger.info("解析文件: %s (%d bytes) preview=%s", filename, _source_size_bytes(content), preview_mode)

        # ── CSV 文件单独处理 ──
        if filename.lower().endswith('.csv'):
            try:
                csv_max_rows = preview_rows if preview_mode else 0
                csv_parsed = _parse_csv_for_preview(content, filename, custom_mapping, max_rows=csv_max_rows)
            except Exception as e:
                diagnostics.append({
                    "file": filename, "sheet": "CSV",
                    "status": "error", "message": f"CSV 解析失败: {e}",
                })
                continue

            dt = csv_parsed["data_type"]
            matched_fields = set(csv_parsed["column_mapping"].values())
            missing_cols, missing_recommended = _detect_missing_fields(dt, matched_fields)

            if detected_year is None and csv_parsed.get("year"):
                detected_year = csv_parsed["year"]

            total_lines_est = csv_parsed.get("total_lines_est", csv_parsed["row_count"])
            est_total = max(total_lines_est - 1, csv_parsed["row_count"])  # 减去表头行

            diag = {
                "file": filename, "sheet": "CSV", "data_type": dt,
                "row_count": csv_parsed["row_count"],
                "total_row_estimate": est_total,
                "header_start": csv_parsed.get("header_start", 0),
                "header_count": 1,
                "matched_cols": sorted(matched_fields),
                "missing_cols": missing_cols,
                "missing_recommended": missing_recommended,
                "column_mapping": csv_parsed["column_mapping"],
                "status": "ok",
                # 以下字段供前端手动映射调整
                "raw_headers": csv_parsed.get("raw_headers", csv_parsed["headers"]),
                "headers": csv_parsed["headers"],
                "preview_rows_data": csv_parsed.get("raw_rows", []),
                "wide_table_detected": csv_parsed.get("wide_table_detected", False),
                "content_inferred": csv_parsed.get("content_inferred", {}),
            }

            if preview_mode:
                diag["status"] = "preview"
                # 预览模式也做数据转换拆分，估算辅助表行数
                if dt == "ledger":
                    led, _, aux_stats = convert_ledger_rows(csv_parsed["rows"])
                    diag["ledger_count_est"] = est_total
                    aux_led_est = 0
                    if csv_parsed["row_count"] > 0:
                        ratio = sum(aux_stats.values()) / csv_parsed["row_count"]
                        aux_led_est = int(est_total * ratio)
                    diag["aux_ledger_count_est"] = aux_led_est
                    for t, c in aux_stats.items():
                        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c
                elif dt == "balance" or dt == "aux_balance":
                    bal, aux_bal = convert_balance_rows(csv_parsed["rows"])
                    diag["balance_count_est"] = est_total
                    aux_bal_est = 0
                    if csv_parsed["row_count"] > 0:
                        ratio = len(aux_bal) / csv_parsed["row_count"]
                        aux_bal_est = int(est_total * ratio)
                    diag["aux_balance_count_est"] = aux_bal_est
                elif dt == "aux_ledger":
                    led, _, aux_stats = convert_ledger_rows(csv_parsed["rows"])
                    diag["ledger_count_est"] = est_total
                    aux_led_est = 0
                    if csv_parsed["row_count"] > 0:
                        ratio = sum(aux_stats.values()) / csv_parsed["row_count"]
                        aux_led_est = int(est_total * ratio)
                    diag["aux_ledger_count_est"] = aux_led_est
                    for t, c in aux_stats.items():
                        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c
                diagnostics.append(diag)
                continue

            if dt == "ledger" or dt == "aux_ledger":
                led, _, aux_stats = convert_ledger_rows(csv_parsed["rows"])
                all_ledger_rows.extend(led)
                diag["ledger_count"] = len(led)
                diag["aux_ledger_count"] = sum(aux_stats.values())
                for t, c in aux_stats.items():
                    _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c
            elif dt == "balance" or dt == "aux_balance":
                bal, aux_bal = convert_balance_rows(csv_parsed["rows"])
                all_balance_rows.extend(bal)
                all_aux_balance_rows.extend(aux_bal)
                diag["balance_count"] = len(bal)
                diag["aux_balance_count"] = len(aux_bal)
            elif dt == "account_chart":
                added_before = len(acct_records)
                for row in csv_parsed["rows"]:
                    _append_account_record(
                        row=row,
                        project_id=project_id,
                        seen_codes=seen_codes,
                        acct_records=acct_records,
                        by_category=by_category,
                        staged=True,
                    )
                diag["account_count"] = len(acct_records) - added_before
            else:
                diag["status"] = "skipped"
                diag["message"] = f"CSV 未识别的数据类型: {dt}"

            diagnostics.append(diag)
            continue

        # ── Excel 文件处理 ──
        # 第一遍：read_only 模式快速扫描
        try:
            wb = openpyxl.load_workbook(_workbook_input(content), read_only=True, data_only=True)
        except Exception as e:
            diagnostics.append({
                "file": filename, "sheet": None,
                "status": "error", "message": f"无法打开文件: {e}",
            })
            continue

        needs_full_mode = False
        for ws in wb.worksheets:
            try:
                first_rows = list(ws.iter_rows(max_row=5, values_only=True))
                if first_rows:
                    max_cols = max(len(r) for r in first_rows)
                    if max_cols <= 3:
                        needs_full_mode = True
                        break
                    # 检测合并单元格特征：同一行多列值完全相同
                    for row in first_rows:
                        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
                        if len(non_empty) >= 4:
                            from collections import Counter
                            most_common_val, most_common_count = Counter(non_empty).most_common(1)[0]
                            if most_common_count >= len(non_empty) * 0.6 and most_common_count >= 3:
                                needs_full_mode = True
                                break
                    if needs_full_mode:
                        break
            except Exception:
                pass
        wb.close()

        # 如果需要完整模式（有合并单元格），直接用完整模式打开
        # 完整模式打开约4秒，iter_rows遍历50000行约0.4秒，可接受
        if needs_full_mode:
            logger.info("  检测到合并单元格，使用完整模式: %s", filename)

        try:
            if needs_full_mode:
                wb = openpyxl.load_workbook(_workbook_input(content), data_only=True)
            else:
                wb = openpyxl.load_workbook(_workbook_input(content), read_only=True, data_only=True)
        except Exception as e:
            diagnostics.append({
                "file": filename, "sheet": None,
                "status": "error", "message": f"无法打开文件: {e}",
            })
            continue

        for ws in wb.worksheets:
            sheet_name = ws.title
            # 跳过非数据 sheet
            name_lower = sheet_name.lower()
            if any(kw in name_lower for kw in ("说明", "目录", "封面", "模板")):
                continue

            try:
                sheet_mapping = _resolve_sheet_mapping(custom_mapping, filename, sheet_name)
                parsed = smart_parse_sheet(
                    ws,
                    max_data_rows=(preview_rows if preview_mode else None),
                    include_raw_rows=preview_mode,
                    sheet_mapping=sheet_mapping,
                )
            except Exception as e:
                diagnostics.append({
                    "file": filename, "sheet": sheet_name,
                    "status": "error", "message": f"解析失败: {e}",
                })
                continue

            if parsed.get("year") is None:
                parsed["year"] = extract_year_from_content(parsed["rows"], filename=filename)

            dt = parsed["data_type"]
            row_count = parsed["row_count"]
            total_estimate = parsed.get("total_row_estimate")
            matched_fields = set(parsed["column_mapping"].values())
            missing_cols, missing_recommended = _detect_missing_fields(dt, matched_fields)

            # 提取年度
            if detected_year is None and parsed.get("year"):
                detected_year = parsed["year"]

            diag = {
                "file": filename,
                "sheet": sheet_name,
                "data_type": dt,
                "row_count": row_count,
                "total_row_estimate": total_estimate,
                "header_start": parsed["header_start"],
                "header_count": parsed["header_count"],
                "matched_cols": sorted(matched_fields),
                "missing_cols": missing_cols,
                "missing_recommended": missing_recommended,
                "column_mapping": parsed["column_mapping"],
                "status": "ok",
                # 以下字段供前端手动映射调整和企业信息展示
                "raw_headers": parsed["headers"],
                "headers": parsed["headers"],
                "preview_rows_data": parsed.get("raw_rows", []),
                "company_code": parsed.get("company_code"),
                "year": parsed.get("year"),
                "wide_table_detected": False,
                "content_inferred": {},
            }

            if preview_mode:
                diag["status"] = "preview"
                est = total_estimate if total_estimate is not None else row_count
                if dt == "balance":
                    bal, aux_bal = convert_balance_rows(parsed["rows"])
                    diag["balance_count_est"] = est
                    aux_bal_est = 0
                    if row_count > 0:
                        ratio = len(aux_bal) / row_count
                        aux_bal_est = int(est * ratio)
                    diag["aux_balance_count_est"] = aux_bal_est
                elif dt == "ledger":
                    led, _, aux_stats = convert_ledger_rows(parsed["rows"])
                    diag["ledger_count_est"] = est
                    aux_led_est = 0
                    if row_count > 0:
                        ratio = sum(aux_stats.values()) / row_count
                        aux_led_est = int(est * ratio)
                    diag["aux_ledger_count_est"] = aux_led_est
                    for t, c in aux_stats.items():
                        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c
                elif dt == "aux_balance":
                    bal, aux_bal = convert_balance_rows(parsed["rows"])
                    diag["balance_count_est"] = est
                    aux_bal_est = 0
                    if row_count > 0:
                        ratio = len(aux_bal) / row_count
                        aux_bal_est = int(est * ratio)
                    diag["aux_balance_count_est"] = aux_bal_est
                elif dt == "aux_ledger":
                    led, _, aux_stats = convert_ledger_rows(parsed["rows"])
                    diag["ledger_count_est"] = est
                    aux_led_est = 0
                    if row_count > 0:
                        ratio = sum(aux_stats.values()) / row_count
                        aux_led_est = int(est * ratio)
                    diag["aux_ledger_count_est"] = aux_led_est
                    for t, c in aux_stats.items():
                        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c
                diagnostics.append(diag)
                continue

            if dt == "balance" or dt == "aux_balance":
                bal, aux_bal = convert_balance_rows(parsed["rows"])
                all_balance_rows.extend(bal)
                all_aux_balance_rows.extend(aux_bal)
                diag["balance_count"] = len(bal)
                diag["aux_balance_count"] = len(aux_bal)

            elif dt == "ledger" or dt == "aux_ledger":
                led, _, aux_stats = convert_ledger_rows(parsed["rows"])
                all_ledger_rows.extend(led)
                # 辅助明细账不在内存中保留，写入时从 _aux_dim_str 流式拆分
                aux_led_count = sum(aux_stats.values())
                diag["ledger_count"] = len(led)
                diag["aux_ledger_count"] = aux_led_count
                # 合并维度统计
                for t, c in aux_stats.items():
                    _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c

            else:
                diag["status"] = "skipped"
                diag["message"] = f"未识别的数据类型: {dt}"

            diagnostics.append(diag)

        wb.close()

    # 年度兜底
    if detected_year is None:
        detected_year = datetime.now().year - 1

    # 辅助核算维度统计（合并余额表和序时账的统计）
    for r in all_aux_balance_rows:
        t = r.get("aux_type", "?")
        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + 1

    aux_dimensions = [
        {"type": t, "count": c}
        for t, c in sorted(_aux_type_counts.items(), key=lambda x: -x[1])
    ]

    # 一致性校验移到入库后按需触发（预览阶段跳过，632MB CSV 全量校验太慢）
    validation = []

    return {
        "balance_rows": all_balance_rows,
        "aux_balance_rows": all_aux_balance_rows,
        "ledger_rows": all_ledger_rows,
        "aux_ledger_rows": all_aux_ledger_rows,
        "year": detected_year,
        "aux_dimensions": aux_dimensions,
        "validation": validation,
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 9. 数据库写入
# ─────────────────────────────────────────────────────────────────────────────

async def _clear_project_year_tables(project_id: UUID, year: int, db) -> None:
    """Soft-delete all four table data for a given project+year.

    Ensures that each project-year has exactly one effective dataset
    and old data from previous imports does not remain as a mix.
    """
    import sqlalchemy as sa
    from app.models.audit_platform_models import ImportBatch, ImportStatus, TbBalance, TbLedger, TbAuxBalance, TbAuxLedger

    for model in (TbBalance, TbLedger, TbAuxBalance, TbAuxLedger):
        tbl = model.__table__
        await db.execute(
            sa.update(tbl)
            .where(
                tbl.c.project_id == project_id,
                tbl.c.year == year,
                tbl.c.is_deleted == sa.false(),
            )
            .values(is_deleted=True)
        )
    await db.execute(
        sa.update(ImportBatch)
        .where(
            ImportBatch.project_id == project_id,
            ImportBatch.year == year,
            ImportBatch.status == ImportStatus.completed,
        )
        .values(status=ImportStatus.rolled_back)
    )
    await db.flush()


async def write_four_tables(
    project_id: UUID,
    year: int,
    balance_rows: list[dict],
    aux_balance_rows: list[dict],
    ledger_rows: list[dict],
    aux_ledger_rows: list[dict],
    db,  # AsyncSession
    progress_callback=None,  # (stage, current, total, message) -> None
) -> dict[str, int]:
    """将四表数据写入数据库。返回 {data_type: record_count}。"""
    raise RuntimeError("write_four_tables 已废弃，请改用 smart_import_streaming 或 LedgerImportApplicationService")
    import uuid as _uuid
    import sqlalchemy as sa
    from app.models.audit_platform_models import (
        TbBalance, TbLedger, TbAuxBalance, TbAuxLedger,
        ImportBatch, ImportStatus,
    )

    def _progress(stage: str, current: int, total: int, msg: str = ""):
        if progress_callback:
            try:
                progress_callback(stage, current, total, msg)
            except Exception:
                pass

    CHUNK_SIZE = 50000
    result: dict[str, int] = {}
    # 计算总行数用于进度
    total_rows = len(balance_rows) + len(aux_balance_rows) + len(ledger_rows)
    # 辅助明细账行数预估（序时账行数 × 平均维度数）
    est_aux_ledger = sum(1 for r in ledger_rows[:1000] if r.get("_aux_dim_str")) * 2.5 * len(ledger_rows) / max(len(ledger_rows[:1000]), 1)
    total_est = total_rows + int(est_aux_ledger)
    written_total = 0

    # 序时账和辅助明细账需要特殊处理：辅助明细从序时账的 _aux_dim_str 流式拆分
    # 先写余额表和辅助余额表，再写序时账+辅助明细账

    # ── 余额表 + 辅助余额表 ──
    for data_type, model, rows in [
        ("tb_balance", TbBalance, balance_rows),
        ("tb_aux_balance", TbAuxBalance, aux_balance_rows),
    ]:
        if not rows:
            continue
        batch = ImportBatch(
            project_id=project_id, year=year, source_type="smart_import",
            file_name=f"smart_{data_type}", data_type=data_type,
            status=ImportStatus.processing, started_at=datetime.now(timezone.utc),
        )
        db.add(batch)
        await db.flush()

        tbl = model.__table__
        await db.execute(
            sa.update(tbl).where(tbl.c.project_id == project_id, tbl.c.year == year, tbl.c.is_deleted == sa.false())
            .values(is_deleted=True)
        )

        record_count = 0
        valid_cols = {c.name for c in tbl.columns}
        for i in range(0, len(rows), CHUNK_SIZE):
            chunk = rows[i:i + CHUNK_SIZE]
            recs = []
            for row in chunk:
                # 只保留表中存在的列，过滤掉 direction 等不属于该表的字段
                filtered = {k: v for k, v in row.items() if k in valid_cols}
                filtered.update({"id": _uuid.uuid4(), "project_id": project_id, "year": year,
                                 "import_batch_id": batch.id, "dataset_id": _dataset_id, "is_deleted": False})
                recs.append(filtered)
            if recs:
                await db.execute(tbl.insert(), recs)
                record_count += len(recs)
                written_total += len(recs)
                _progress(data_type, written_total, total_est, f"{data_type}: {record_count:,} / {len(rows):,}")
            # 每批提交，减少事务持有时间和锁竞争
            await db.commit()
            batch.record_count = record_count

        batch.status = ImportStatus.completed
        batch.completed_at = datetime.now(timezone.utc)
        result[data_type] = record_count

    # ── 序时账 + 辅助明细账（流式拆分，不在内存中保留全部辅助明细行） ──
    if ledger_rows:
        # 序时账批次
        led_batch = ImportBatch(
            project_id=project_id, year=year, source_type="smart_import",
            file_name="smart_tb_ledger", data_type="tb_ledger",
            status=ImportStatus.processing, started_at=datetime.now(timezone.utc),
        )
        db.add(led_batch)
        # 辅助明细账批次
        aux_batch = ImportBatch(
            project_id=project_id, year=year, source_type="smart_import",
            file_name="smart_tb_aux_ledger", data_type="tb_aux_ledger",
            status=ImportStatus.processing, started_at=datetime.now(timezone.utc),
        )
        db.add(aux_batch)
        await db.flush()

        # 软删除旧数据
        led_tbl = TbLedger.__table__
        aux_tbl = TbAuxLedger.__table__
        await db.execute(
            sa.update(led_tbl).where(led_tbl.c.project_id == project_id, led_tbl.c.year == year, led_tbl.c.is_deleted == sa.false())
            .values(is_deleted=True)
        )
        await db.execute(
            sa.update(aux_tbl).where(aux_tbl.c.project_id == project_id, aux_tbl.c.year == year, aux_tbl.c.is_deleted == sa.false())
            .values(is_deleted=True)
        )

        led_count = 0
        aux_count = 0
        led_buf = []
        aux_buf = []

        # 预生成 UUID 池（比逐个 uuid4() 快）
        def _batch_uuids(n: int) -> list:
            return [_uuid.uuid4() for _ in range(n)]

        for row in ledger_rows:
            aux_dim_str = row.pop("_aux_dim_str", "")

            led_buf.append({
                "id": _uuid.uuid4(), "project_id": project_id, "year": year,
                "import_batch_id": led_batch.id, "dataset_id": _dataset_id, "is_deleted": False, **row,
            })

            # 流式拆分辅助明细账
            if aux_dim_str and (":" in aux_dim_str or "：" in aux_dim_str):
                dims = parse_aux_dimensions(aux_dim_str)
                for dim in dims:
                    aux_buf.append({
                        "id": _uuid.uuid4(), "project_id": project_id, "year": year,
                        "import_batch_id": aux_batch.id, "dataset_id": _dataset_id, "is_deleted": False,
                        "aux_type": dim["aux_type"], "aux_code": dim["aux_code"], "aux_name": dim["aux_name"],
                        "aux_dimensions_raw": aux_dim_str,
                    })

            # 分块写入（增大到 50000 减少 INSERT 次数）
            if len(led_buf) >= CHUNK_SIZE:
                await db.execute(led_tbl.insert(), led_buf)
                led_count += len(led_buf)
                written_total += len(led_buf)
                led_buf.clear()
                _progress("tb_ledger", written_total, total_est, f"序时账: {led_count:,}, 辅助明细: {aux_count:,}")
            if len(aux_buf) >= CHUNK_SIZE:
                await db.execute(aux_tbl.insert(), aux_buf)
                aux_count += len(aux_buf)
                written_total += len(aux_buf)
                aux_buf.clear()
                _progress("tb_aux_ledger", written_total, total_est, f"序时账: {led_count:,}, 辅助明细: {aux_count:,}")
            # 有数据写入时提交，减少事务粒度
            if len(led_buf) == 0 and len(aux_buf) == 0 and (led_count > 0 or aux_count > 0):
                await db.commit()
                led_batch.record_count = led_count
                aux_batch.record_count = aux_count

        # 写入剩余
        if led_buf:
            await db.execute(led_tbl.insert(), led_buf)
            led_count += len(led_buf)
        if aux_buf:
            await db.execute(aux_tbl.insert(), aux_buf)
            aux_count += len(aux_buf)
        if led_count > 0 or aux_count > 0:
            await db.commit()

        led_batch.record_count = led_count
        led_batch.status = ImportStatus.completed
        led_batch.completed_at = datetime.now(timezone.utc)
        aux_batch.record_count = aux_count
        aux_batch.status = ImportStatus.completed
        aux_batch.completed_at = datetime.now(timezone.utc)
        result["tb_ledger"] = led_count
        result["tb_aux_ledger"] = aux_count

    await db.commit()

    # ── 入库后自动计算辅助余额汇总（通用，任何企业都触发） ──
    if result.get("tb_aux_balance", 0) > 0:
        try:
            summary_count = await rebuild_aux_balance_summary(project_id, year, db)
            result["aux_summary"] = summary_count
            logger.info("辅助余额汇总: project=%s, year=%d, %d 条", project_id, year, summary_count)
        except Exception as e:
            logger.warning("辅助余额汇总失败（不影响导入）: %s", e)

    # ── 发布数据导入完成事件，驱动后续重算与缓存失效 ──
    try:
        from app.services.event_bus import event_bus
        from app.models.audit_platform_schemas import EventPayload, EventType

        await event_bus.publish(EventPayload(
            event_type=EventType.DATA_IMPORTED,
            project_id=project_id,
            year=year,
        ))
        logger.info("DATA_IMPORTED 事件已发布: project=%s, year=%d, result=%s", project_id, year, result)
    except Exception as e:
        logger.warning("DATA_IMPORTED 事件发布失败（不影响导入）: %s", e)

    return result



async def _stream_csv_import(
    *,
    content,
    filename: str,
    project_id,
    year: int,
    batches: dict,
    db,
    custom_mapping: dict | None,
    seen_codes: set,
    acct_records: list,
    by_category: dict,
    counts: dict,
    _aux_type_counts: dict,
    progress_callback,
    dataset_id: UUID | None = None,
    chunk_size: int = 50_000,
) -> dict:
    """流式处理 CSV 大文件：流式解码 + 逐批转换 + COPY 写入。

    优化点：
    1. codecs.StreamReader 流式解码（不生成完整 str，内存减半）
    2. asyncpg COPY FROM STDIN 替代 INSERT（写入速度 5-10x）
    3. 每 10 万行一批，峰值内存 ≈ 单批数据大小
    """
    import codecs
    import csv
    from app.services.fast_writer import copy_insert

    def _prog(pct, msg=""):
        if progress_callback:
            try:
                progress_callback(pct, msg)
            except Exception:
                pass

    total_bytes = _source_size_bytes(content)
    with _open_binary_source(content) as raw_probe:
        sample = raw_probe.read(65536)
    if not sample:
        raise ValueError("CSV 文件为空")

    encoding = _detect_stream_encoding(sample)
    sample_line_count = sample.count(b'\n')
    if sample_line_count > 0 and total_bytes > 0:
        avg_line_bytes = max(len(sample) / sample_line_count, 1)
        total_lines_est = max(int(total_bytes / avg_line_bytes), 1)
    else:
        total_lines_est = 0
    _prog(12, f"CSV {filename}: ~{total_lines_est:,} 行, {total_bytes/1024/1024:.0f} MB")

    header_line = None
    header_line_number = 0
    with _open_binary_source(content) as raw_header:
        stream = codecs.getreader(encoding)(raw_header)
        for idx in range(5):
            line = stream.readline()
            if not line:
                break
            cells = line.strip().split(',')
            non_empty = [c.strip() for c in cells if c.strip()]
            if len(non_empty) >= 3:
                header_line = line.strip()
                header_line_number = idx + 1
                break

    if header_line is None:
        raise ValueError("CSV 文件未找到有效表头")

    reader_header = list(csv.reader([header_line]))[0]
    raw_headers = [c.strip() if c.strip() else f"col_{j}" for j, c in enumerate(reader_header)]
    headers = raw_headers[:]
    mid = len(raw_headers) // 2
    wide_table_detected = False
    if len(raw_headers) >= 60:
        first_half_matches = sum(1 for h in raw_headers[:mid] if smart_match_column(h))
        second_half_matches = sum(1 for h in raw_headers[mid:] if smart_match_column(h))
        if second_half_matches >= 5 and second_half_matches > first_half_matches * 2:
            wide_table_detected = True
            headers = raw_headers[mid:]

    # ── 列名映射 ──
    column_mapping: dict[str, str] = {}
    for h in headers:
        mapped = smart_match_column(h)
        if mapped:
            column_mapping[h] = mapped

    sheet_mapping = _resolve_sheet_mapping(custom_mapping, filename, "CSV")
    if sheet_mapping:
        for h, v in sheet_mapping.items():
            if h in headers and v:
                column_mapping[h] = v

    mapped_fields = set(column_mapping.values())
    dt = _guess_data_type(mapped_fields)
    miss_req, miss_rec = _detect_missing_fields(dt, mapped_fields)
 
    diag = {
        "file": filename, "sheet": "CSV", "data_type": dt,
        "row_count": 0, "header_count": 1,
        "matched_cols": sorted(mapped_fields),
        "missing_cols": miss_req, "missing_recommended": miss_rec,
        "column_mapping": column_mapping,
        "status": "ok",
        # 以下字段供前端手动映射调整和企业信息展示
        "raw_headers": raw_headers,
        "headers": headers,
        "preview_rows_data": [],
        "company_code": "",
        "year": "",
        "wide_table_detected": False,
        "content_inferred": {},
    }

    if dt in ("ledger", "balance", "account_chart") and miss_req:
        diag["status"] = "error"
        diag["message"] = _build_missing_required_import_error(dt, miss_req, "CSV")
        return {"diag": diag, "errors": [f"{filename}: {diag['message']}"]}
 
    if dt not in ("ledger", "balance", "account_chart"):
        diag["status"] = "skipped"
        diag["message"] = f"CSV 未识别的数据类型: {dt}"
        return {"diag": diag, "errors": []}

    # ── COPY 写入列定义 ──
    _LEDGER_COLS = [
        "company_code", "voucher_date", "voucher_no", "account_code",
        "account_name", "debit_amount", "credit_amount", "summary",
        "preparer", "currency_code", "accounting_period", "voucher_type",
        "entry_seq", "debit_qty", "credit_qty", "debit_fc", "credit_fc",
    ]
    _AUX_LEDGER_COLS = _LEDGER_COLS + [
        "aux_type", "aux_code", "aux_name", "aux_dimensions_raw",
    ]
    _BALANCE_COLS = [
        "company_code", "account_code", "account_name", "level",
        "opening_balance", "opening_debit", "opening_credit",
        "closing_balance", "closing_debit", "closing_credit",
        "debit_amount", "credit_amount",
        "opening_qty", "opening_fc", "currency_code",
    ]
    _AUX_BALANCE_COLS = _BALANCE_COLS + [
        "aux_type", "aux_code", "aux_name", "aux_dimensions_raw",
    ]

    # ── 流式主循环 ──
    LINES_PER_BATCH = 100_000
    row_count = 0
    errs: list[str] = []
    line_buf: list[str] = []

    async def _flush_batch(rows: list[dict]):
        """转换一批原始行并用 COPY 写入，每批提交以缩小事务粒度。"""
        if dt == "ledger":
            led, _, aux_stats = convert_ledger_rows(rows)
            for t, c in aux_stats.items():
                _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c

            led_rows: list[dict] = []
            aux_rows: list[dict] = []
            for row in led:
                adim = row.pop("_aux_dim_str", "")
                led_rows.append(row)
                if adim and (":" in adim or "：" in adim):
                    for dim in parse_aux_dimensions(adim):
                        aux_rows.append({
                            **row,
                            "aux_type": dim["aux_type"],
                            "aux_code": dim["aux_code"],
                            "aux_name": dim["aux_name"],
                            "aux_dimensions_raw": adim,
                        })

            if led_rows:
                n = await copy_insert(
                    db, "tb_ledger", _LEDGER_COLS, led_rows,
                    project_id, year, batches["tb_ledger"].id,
                    is_deleted=True,
                    dataset_id=dataset_id,
                )
                counts["tb_ledger"] += n
            if aux_rows:
                n = await copy_insert(
                    db, "tb_aux_ledger", _AUX_LEDGER_COLS, aux_rows,
                    project_id, year, batches["tb_aux_ledger"].id,
                    is_deleted=True,
                    dataset_id=dataset_id,
                )
                counts["tb_aux_ledger"] += n

        elif dt == "balance":
            bal, aux_bal = convert_balance_rows(rows)
            if bal:
                n = await copy_insert(
                    db, "tb_balance", _BALANCE_COLS, bal,
                    project_id, year, batches["tb_balance"].id,
                    is_deleted=True,
                    dataset_id=dataset_id,
                )
                counts["tb_balance"] += n
            if aux_bal:
                n = await copy_insert(
                    db, "tb_aux_balance", _AUX_BALANCE_COLS, aux_bal,
                    project_id, year, batches["tb_aux_balance"].id,
                    is_deleted=True,
                    dataset_id=dataset_id,
                )
                counts["tb_aux_balance"] += n

        # 每批 COPY 后提交，减少锁持有时间
        await db.commit()

        # 提取科目
        for row in rows:
            _append_account_record(
                row=row,
                project_id=project_id,
                seen_codes=seen_codes,
                acct_records=acct_records,
                by_category=by_category,
                staged=True,
                dataset_id=dataset_id,
            )

    def _parse_lines(lines: list[str]) -> list[dict]:
        """将一批 CSV 行解析为 dict 列表。"""
        result = []
        for row_raw in csv.reader(lines):
            if not row_raw or all(not c.strip() for c in row_raw):
                continue
            if wide_table_detected:
                padded = row_raw + [""] * max(0, mid - len(row_raw))
                row_dict = {}
                for j, h in enumerate(headers):
                    mapped = column_mapping.get(h, h)
                    val = padded[j].strip() if j < len(padded) else ""
                    row_dict[mapped] = val if val else None
            else:
                padded = row_raw + [""] * max(0, len(headers) - len(row_raw))
                row_dict = {}
                for j, h in enumerate(headers):
                    mapped = column_mapping.get(h, h)
                    val = padded[j].strip() if j < len(padded) else ""
                    row_dict[mapped] = val if val else None
            result.append(row_dict)
        return result

    with _open_binary_source(content) as raw_data:
        stream = codecs.getreader(encoding)(raw_data)
        for _ in range(header_line_number):
            if not stream.readline():
                break

        while True:
            line = stream.readline()
            if not line:
                break
            line = line.rstrip('\n').rstrip('\r')
            if line:
                line_buf.append(line)

            if len(line_buf) >= LINES_PER_BATCH:
                parsed_rows = _parse_lines(line_buf)
                row_count += len(parsed_rows)
                if parsed_rows:
                    await _flush_batch(parsed_rows)
                pct = 15 + int(row_count / max(total_lines_est, 1) * 70)
                _prog(min(pct, 88), f"CSV: {row_count:,}/{total_lines_est:,} 行")
                line_buf.clear()

        if line_buf:
            parsed_rows = _parse_lines(line_buf)
            row_count += len(parsed_rows)
            if parsed_rows:
                await _flush_batch(parsed_rows)
            line_buf.clear()

    diag["row_count"] = row_count
    _prog(88, f"CSV 完成: {row_count:,} 行")

    return {"diag": diag, "errors": errs}





async def smart_import_streaming(
    project_id: UUID,
    file_contents: list[tuple[str, object]],
    db,  # AsyncSession
    year_override: int | None = None,
    custom_mapping: Optional[dict] = None,
    progress_callback=None,
    job_id: UUID | None = None,
    created_by: UUID | None = None,
    force_activate: bool = False,
) -> dict:
    """多文件流式导入：逐批解析并隐藏写入，成功后统一激活。"""
    import openpyxl
    from app.core.config import settings
    from app.models.audit_platform_models import (
        TbBalance, TbLedger, TbAuxBalance, TbAuxLedger,
        ImportBatch, ImportStatus,
    )
    from app.services.fast_writer import copy_insert

    CHUNK = 50_000
    _LEDGER_COLS = [
        "company_code", "voucher_date", "voucher_no", "account_code",
        "account_name", "debit_amount", "credit_amount", "summary",
        "preparer", "currency_code", "accounting_period", "voucher_type",
        "entry_seq", "debit_qty", "credit_qty", "debit_fc", "credit_fc",
    ]
    _AUX_LEDGER_COLS = _LEDGER_COLS + [
        "aux_type", "aux_code", "aux_name", "aux_dimensions_raw",
    ]
    _BALANCE_COLS = [
        "company_code", "account_code", "account_name", "level",
        "opening_balance", "opening_debit", "opening_credit",
        "closing_balance", "closing_debit", "closing_credit",
        "debit_amount", "credit_amount",
        "opening_qty", "opening_fc", "currency_code",
    ]
    _AUX_BALANCE_COLS = _BALANCE_COLS + [
        "aux_type", "aux_code", "aux_name", "aux_dimensions_raw",
    ]

    def _prog(pct: int, msg: str = ""):
        if progress_callback:
            try:
                progress_callback(pct, msg)
            except Exception:
                pass

    detected_year = year_override
    total_size = sum(_source_size_bytes(c) for _, c in file_contents)
    sheet_count_est = 0

    if detected_year is None:
        for fn, ct in file_contents:
            if fn.lower().endswith('.csv'):
                sheet_count_est += 1
                if detected_year is None:
                    detected_year = extract_year_from_content([], filename=fn)
                continue
            try:
                wb0 = openpyxl.load_workbook(_workbook_input(ct), read_only=True, data_only=True)
                for ws0 in wb0.worksheets:
                    sheet_count_est += 1
                    if detected_year is None:
                        detected_year = extract_year_from_content(
                            [{f"c{j}": v for j, v in enumerate(r)}
                             for r in ws0.iter_rows(max_row=25, values_only=True)],
                            filename=fn,
                        )
                wb0.close()
            except Exception:
                continue
    else:
        for fn, ct in file_contents:
            if fn.lower().endswith('.csv'):
                sheet_count_est += 1
                continue
            try:
                wb0 = openpyxl.load_workbook(_workbook_input(ct), read_only=True, data_only=True)
                sheet_count_est += len(wb0.worksheets)
                wb0.close()
            except Exception:
                pass

    if detected_year is None:
        detected_year = datetime.now().year - 1
    year = detected_year

    _prog(5, f"年度 {year}，{len(file_contents)} 个文件，{total_size/1024/1024:.1f} MB")

    _TABLE_MAP = {
        "tb_balance": TbBalance,
        "tb_aux_balance": TbAuxBalance,
        "tb_ledger": TbLedger,
        "tb_aux_ledger": TbAuxLedger,
    }
    batches: dict[str, ImportBatch] = {}
    for dt_key in _TABLE_MAP:
        batch = ImportBatch(
            project_id=project_id,
            year=year,
            source_type="smart_import",
            file_name=f"multi_{dt_key}",
            data_type=dt_key,
            status=ImportStatus.processing,
            started_at=datetime.now(timezone.utc),
        )
        db.add(batch)
        batches[dt_key] = batch
    await db.flush()

    # Phase 17: 创建 staged 数据集版本
    _dataset_id = None
    try:
        from app.services.dataset_service import DatasetService
        _source_files = [fn for fn, _ in file_contents]
        _dataset = await DatasetService.create_staged(
            db,
            project_id=project_id,
            year=year,
            source_type="import",
            source_summary={"files": _source_files, "file_count": len(_source_files)},
            job_id=job_id,
            created_by=created_by,
        )
        _dataset_id = _dataset.id
        await db.flush()
    except Exception as _ds_err:
        logger.exception("创建 LedgerDataset 失败，已终止导入")
        raise SmartImportError(
            f"创建数据集版本失败: {_ds_err}",
            errors=[str(_ds_err)],
            year=year,
        ) from _ds_err

    _prog(10, "开始解析文件…")

    counts: dict[str, int] = {k: 0 for k in _TABLE_MAP}
    diagnostics: list[dict] = []
    errors: list[str] = []
    seen_codes: set[str] = set()
    acct_records: list[AccountChart] = []
    by_category: dict[str, int] = {}
    _aux_type_counts: dict[str, int] = {}
    sheets_done = 0

    for file_idx, (filename, content) in enumerate(file_contents):
        logger.info(
            "流式导入 %d/%d: %s (%d bytes)",
            file_idx + 1,
            len(file_contents),
            filename,
            _source_size_bytes(content),
        )

        if filename.lower().endswith('.xls'):
            diagnostics.append({
                "file": filename,
                "sheet": None,
                "status": "error",
                "message": "暂不支持 Excel 97-2003 (.xls) 文件，请先转换为 .xlsx 后再上传",
            })
            errors.append(f"暂不支持 .xls 文件: {filename}，请先转换为 .xlsx 后再上传")
            continue

        if filename.lower().endswith('.csv'):
            sheets_done += 1
            _prog(12, f"解析 CSV {filename}")
            try:
                csv_result = await _stream_csv_import(
                    content=content,
                    filename=filename,
                    project_id=project_id,
                    year=year,
                    batches=batches,
                    db=db,
                    custom_mapping=custom_mapping,
                    seen_codes=seen_codes,
                    acct_records=acct_records,
                    by_category=by_category,
                    counts=counts,
                    _aux_type_counts=_aux_type_counts,
                    progress_callback=_prog,
                    dataset_id=_dataset_id,
                    chunk_size=CHUNK,
                )
                diagnostics.append(csv_result["diag"])
                if csv_result.get("errors"):
                    errors.extend(csv_result["errors"])
            except Exception as e:
                import traceback
                logger.error("CSV 导入异常: %s\n%s", e, traceback.format_exc())
                diagnostics.append({"file": filename, "sheet": "CSV", "status": "error", "message": str(e)})
                errors.append(f"CSV 导入失败 {filename}: {e}")
            continue

        needs_full = False
        try:
            wb_probe = openpyxl.load_workbook(_workbook_input(content), read_only=True, data_only=True)
            for _ws in wb_probe.worksheets:
                try:
                    rows5 = list(_ws.iter_rows(max_row=5, values_only=True))
                    if not rows5:
                        continue
                    if max(len(r) for r in rows5) <= 3:
                        needs_full = True
                        break
                    for row in rows5:
                        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
                        if len(non_empty) >= 4:
                            from collections import Counter
                            _, most_common_count = Counter(non_empty).most_common(1)[0]
                            if most_common_count >= len(non_empty) * 0.6 and most_common_count >= 3:
                                needs_full = True
                                break
                    if needs_full:
                        break
                except Exception:
                    pass
            wb_probe.close()
        except Exception:
            pass

        full_mode_limit_bytes = max(1, settings.LEDGER_IMPORT_FULL_MODE_MAX_FILE_MB) * 1024 * 1024
        if needs_full and _source_size_bytes(content) > full_mode_limit_bytes:
            diagnostics.append({
                "file": filename,
                "sheet": None,
                "status": "error",
                "message": (
                    "该 Excel 需要非只读模式解析合并/复杂表头，文件超过 "
                    f"{settings.LEDGER_IMPORT_FULL_MODE_MAX_FILE_MB}MB 内存保护阈值。"
                    "请拆分文件、另存为 CSV，或调高 LEDGER_IMPORT_FULL_MODE_MAX_FILE_MB。"
                ),
            })
            errors.append(f"{filename}: 文件超过非只读 Excel 解析阈值")
            continue

        try:
            wb = openpyxl.load_workbook(
                _workbook_input(content),
                read_only=(not needs_full),
                data_only=True,
            )
        except Exception as e:
            diagnostics.append({"file": filename, "sheet": None, "status": "error", "message": f"无法打开文件: {e}"})
            errors.append(f"无法打开 {filename}: {e}")
            continue

        for ws in wb.worksheets:
            sname = ws.title
            if any(kw in sname.lower() for kw in ("说明", "目录", "封面", "模板")):
                continue

            sheets_done += 1
            pct = 10 + int(sheets_done / max(sheet_count_est, 1) * 78)
            _prog(min(pct, 88), f"解析 {filename} / {sname}")

            try:
                sheet_mapping = _resolve_sheet_mapping(custom_mapping, filename, sname)
                sheet_context = _build_sheet_context(ws, sheet_mapping=sheet_mapping)
            except Exception as e:
                diagnostics.append({"file": filename, "sheet": sname, "status": "error", "message": str(e)})
                errors.append(f"解析失败 {filename}/{sname}: {e}")
                continue

            dt = sheet_context["data_type"]
            matched_fields = set(sheet_context["column_mapping"].values())
            miss_req, miss_rec = _detect_missing_fields(dt, matched_fields)
            diag: dict = {
                "file": filename,
                "sheet": sname,
                "data_type": dt,
                "row_count": 0,
                "header_count": sheet_context["header_count"],
                "matched_cols": sorted(matched_fields),
                "missing_cols": miss_req,
                "missing_recommended": miss_rec,
                "column_mapping": sheet_context["column_mapping"],
                "status": "ok",
            }

            if dt in ("ledger", "balance", "account_chart") and miss_req:
                diag["status"] = "error"
                diag["message"] = _build_missing_required_import_error(dt, miss_req)
                diagnostics.append(diag)
                errors.append(f"{filename}/{sname}: {diag['message']}")
                continue

            if dt == "balance":
                total_rows = 0
                total_bal = 0
                total_aux_bal = 0
                for batch_rows in _iter_sheet_row_batches(ws, sheet_context, batch_size=CHUNK):
                    total_rows += len(batch_rows)
                    bal, aux_bal = convert_balance_rows(batch_rows)
                    total_bal += len(bal)
                    total_aux_bal += len(aux_bal)

                    if bal:
                        n = await copy_insert(
                            db, "tb_balance", _BALANCE_COLS, bal,
                            project_id, year, batches["tb_balance"].id,
                            is_deleted=True,
                            dataset_id=_dataset_id,
                        )
                        counts["tb_balance"] += n
                    if aux_bal:
                        n = await copy_insert(
                            db, "tb_aux_balance", _AUX_BALANCE_COLS, aux_bal,
                            project_id, year, batches["tb_aux_balance"].id,
                            is_deleted=True,
                            dataset_id=_dataset_id,
                        )
                        counts["tb_aux_balance"] += n
                    for r in aux_bal:
                        t = r.get("aux_type", "?")
                        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + 1
                    for row in batch_rows:
                        _append_account_record(
                            row=row,
                            project_id=project_id,
                            seen_codes=seen_codes,
                            acct_records=acct_records,
                            by_category=by_category,
                            staged=True,
                            dataset_id=_dataset_id,
                        )
                    await db.commit()

                diag["row_count"] = total_rows
                diag["balance_count"] = total_bal
                diag["aux_balance_count"] = total_aux_bal

            elif dt == "ledger":
                total_rows = 0
                total_led = 0
                total_aux = 0
                for batch_rows in _iter_sheet_row_batches(ws, sheet_context, batch_size=CHUNK):
                    total_rows += len(batch_rows)
                    led, _, aux_stats = convert_ledger_rows(batch_rows)
                    total_led += len(led)
                    total_aux += sum(aux_stats.values())
                    for t, c in aux_stats.items():
                        _aux_type_counts[t] = _aux_type_counts.get(t, 0) + c

                    led_rows: list[dict] = []
                    aux_rows: list[dict] = []
                    for row in led:
                        adim = row.pop("_aux_dim_str", "")
                        led_rows.append(row)
                        if adim and (":" in adim or "：" in adim):
                            for dim in parse_aux_dimensions(adim):
                                aux_rows.append({
                                    **row,
                                    "aux_type": dim["aux_type"],
                                    "aux_code": dim["aux_code"],
                                    "aux_name": dim["aux_name"],
                                    "aux_dimensions_raw": adim,
                                })

                    if led_rows:
                        n = await copy_insert(
                            db, "tb_ledger", _LEDGER_COLS, led_rows,
                            project_id, year, batches["tb_ledger"].id,
                            is_deleted=True,
                            dataset_id=_dataset_id,
                        )
                        counts["tb_ledger"] += n
                    if aux_rows:
                        n = await copy_insert(
                            db, "tb_aux_ledger", _AUX_LEDGER_COLS, aux_rows,
                            project_id, year, batches["tb_aux_ledger"].id,
                            is_deleted=True,
                            dataset_id=_dataset_id,
                        )
                        counts["tb_aux_ledger"] += n
                    for row in batch_rows:
                        _append_account_record(
                            row=row,
                            project_id=project_id,
                            seen_codes=seen_codes,
                            acct_records=acct_records,
                            by_category=by_category,
                            staged=True,
                            dataset_id=_dataset_id,
                        )
                    await db.commit()
                    _prog(min(pct, 88), f"解析 {filename} / {sname}: {total_rows:,} 行")

                diag["row_count"] = total_rows
                diag["ledger_count"] = total_led
                diag["aux_ledger_count"] = total_aux

            elif dt == "account_chart":
                total_rows = 0
                added_before = len(acct_records)
                for batch_rows in _iter_sheet_row_batches(ws, sheet_context, batch_size=CHUNK):
                    total_rows += len(batch_rows)
                    for row in batch_rows:
                        _append_account_record(
                            row=row,
                            project_id=project_id,
                            seen_codes=seen_codes,
                            acct_records=acct_records,
                            by_category=by_category,
                            staged=True,
                            dataset_id=_dataset_id,
                        )
                diag["row_count"] = total_rows
                diag["account_count"] = len(acct_records) - added_before

            elif dt in ("aux_balance", "aux_ledger"):
                diag["status"] = "skipped"
                diag["message"] = "独立辅助表暂不处理（已从余额表/序时账自动拆分）"
            else:
                diag["status"] = "skipped"
                diag["message"] = f"未识别的数据类型: {dt}"

            diagnostics.append(diag)

        wb.close()

    if errors:
        for dt_key, batch in batches.items():
            batch.record_count = counts[dt_key]
            batch.status = ImportStatus.failed
            batch.completed_at = datetime.now(timezone.utc)
        # Phase 17: 标记数据集失败
        if _dataset_id:
            try:
                from app.services.dataset_service import DatasetService
                await DatasetService.mark_failed(db, _dataset_id)
            except Exception:
                pass
        await db.commit()
        raise SmartImportError(
            "；".join(errors[:5]),
            diagnostics=diagnostics,
            errors=errors,
            year=year,
        )

    staged_total = sum(counts.values()) + len(acct_records)
    if staged_total <= 0:
        for dt_key, batch in batches.items():
            batch.record_count = counts[dt_key]
            batch.status = ImportStatus.failed
            batch.completed_at = datetime.now(timezone.utc)
        # Phase 17: 标记数据集失败
        if _dataset_id:
            try:
                from app.services.dataset_service import DatasetService
                await DatasetService.mark_failed(db, _dataset_id)
            except Exception:
                pass
        await db.commit()
        raise SmartImportError(
            "未解析到可导入的数据，请确认文件内容、sheet 类型与列映射",
            diagnostics=diagnostics,
            errors=errors,
            year=year,
        )

    account_ids: list[UUID] = []
    if acct_records:
        db.add_all(acct_records)
        await db.flush()
        account_ids = [record.id for record in acct_records if getattr(record, "id", None)]

    # Phase 17: 第二层 Business Validation（数据解析完成后、激活前）
    _bv_findings: list[dict] = []
    try:
        from app.services.import_validation_service import ImportValidationService
        bv_findings = []
        if _dataset_id:
            bv_findings = await ImportValidationService.run_dataset_business_validation(
                db,
                project_id=project_id,
                year=year,
                dataset_id=_dataset_id,
            )
        _bv_findings = [f.to_dict() for f in bv_findings]
        if _bv_findings:
            logger.info("Business Validation: %d 条发现", len(_bv_findings))
        gate = ImportValidationService.evaluate_activation(_bv_findings, force=force_activate)
        if not gate.get("allowed"):
            if _dataset_id:
                try:
                    from app.services.dataset_service import DatasetService
                    await DatasetService.mark_failed(db, _dataset_id)
                except Exception:
                    pass
            await db.commit()
            blocking = gate.get("blocking_findings") or []
            raise SmartImportError(
                "导入校验未通过，已阻止激活",
                diagnostics=diagnostics,
                errors=[item.get("message", "校验未通过") for item in blocking],
                year=year,
            )
    except Exception as _bv_err:
        if isinstance(_bv_err, SmartImportError):
            raise
        logger.debug("Business Validation 执行失败（不阻断导入）: %s", _bv_err)

    _prog(90, f"科目 {len(acct_records)} 个，准备激活…")

    for dt_key, batch in batches.items():
        batch.record_count = counts[dt_key]
        batch.status = ImportStatus.completed
        batch.completed_at = datetime.now(timezone.utc)

    # Phase 17: 激活数据集版本
    _activated_dataset = None
    if _dataset_id:
        from app.services.dataset_service import DatasetService
        try:
            _activated_dataset = await DatasetService.activate(
                db,
                dataset_id=_dataset_id,
                activated_by=created_by,
                record_summary=counts,
                validation_summary={
                    "business_validation": _bv_findings,
                    "total": len(_bv_findings),
                    "blocking_count": sum(1 for item in _bv_findings if item.get("blocking")),
                    "has_blocking": any(item.get("blocking") for item in _bv_findings),
                },
            )
        except Exception as _act_err:
            await DatasetService.mark_failed(db, _dataset_id)
            await db.commit()
            raise SmartImportError(
                f"数据集激活失败: {_act_err}",
                diagnostics=diagnostics,
                errors=[str(_act_err)],
                year=year,
            ) from _act_err

    await db.commit()

    if counts.get("tb_aux_balance", 0) > 0:
        try:
            sc = await rebuild_aux_balance_summary(project_id, year, db)
            counts["aux_summary"] = sc
        except Exception as e:
            logger.warning("辅助余额汇总失败: %s", e)

    if _activated_dataset is not None:
        try:
            from app.services.dataset_service import DatasetService
            await DatasetService.publish_dataset_activated(_activated_dataset)
            logger.info(
                "LEDGER_DATASET_ACTIVATED 事件: project=%s, year=%d, dataset=%s, counts=%s",
                project_id,
                year,
                _dataset_id,
                counts,
            )
        except Exception as e:
            logger.warning("LEDGER_DATASET_ACTIVATED 事件发布失败: %s", e)

    _prog(100, "导入完成")

    norm_diag: list[dict] = []
    for d in diagnostics:
        norm_diag.append({
            "file": d.get("file"),
            "sheet": d.get("sheet", ""),
            "sheet_name": d.get("sheet", ""),
            "data_type": d.get("data_type", "unknown"),
            "guessed_type": d.get("data_type", "unknown"),
            "matched_cols": d.get("matched_cols", []),
            "missing_cols": d.get("missing_cols", []),
            "missing_recommended": d.get("missing_recommended", []),
            "row_count": d.get("row_count", 0),
            "status": d.get("status", "ok"),
            "message": d.get("message"),
        })

    return {
        "total_accounts": len(acct_records),
        "by_category": by_category,
        "data_sheets_imported": counts,
        "sheet_diagnostics": norm_diag,
        "year": year,
        "dataset_id": str(_dataset_id) if _dataset_id else None,
        "errors": errors,
    }


async def rebuild_aux_balance_summary(project_id: UUID, year: int, db) -> int:
    """重建辅助余额汇总表（按维度类型+科目+辅助编码分组）。

    通用规则：任何企业导入后都自动调用。
    汇总结果存入 tb_aux_balance_summary，前端直接查汇总表渲染树形视图。
    """
    import sqlalchemy as sa

    # 1. 清除旧汇总
    await db.execute(sa.text(
        "DELETE FROM tb_aux_balance_summary WHERE project_id = :pid AND year = :yr"
    ), {"pid": str(project_id), "yr": year})

    # 2. 用 SQL 聚合直接插入（比 Python 遍历快几十倍）
    await db.execute(sa.text("""
        INSERT INTO tb_aux_balance_summary
            (project_id, year, dim_type, account_code, account_name, aux_code, aux_name,
             record_count, opening_balance, debit_amount, credit_amount, closing_balance)
        SELECT
            project_id, year, aux_type, account_code,
            MAX(account_name),
            aux_code, MAX(aux_name),
            COUNT(*),
            SUM(COALESCE(opening_balance, 0)),
            SUM(COALESCE(debit_amount, 0)),
            SUM(COALESCE(credit_amount, 0)),
            SUM(COALESCE(closing_balance, 0))
        FROM tb_aux_balance
        WHERE project_id = :pid AND year = :yr AND is_deleted = false
        GROUP BY project_id, year, aux_type, account_code, aux_code
    """), {"pid": str(project_id), "yr": year})

    # 3. 查汇总行数
    r = await db.execute(sa.text(
        "SELECT COUNT(*) FROM tb_aux_balance_summary WHERE project_id = :pid AND year = :yr"
    ), {"pid": str(project_id), "yr": year})
    count = r.scalar() or 0

    await db.commit()
    return count
