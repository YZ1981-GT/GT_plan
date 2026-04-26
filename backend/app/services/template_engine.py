"""底稿模板引擎 — 模板上传/版本管理/模板集/项目底稿生成

Validates: Requirements 1.1-1.8, 6.2, 6.3
"""

from __future__ import annotations

import logging
import uuid
from typing import Any
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.workpaper_models import (
    WpIndex,
    WpSourceType,
    WpStatus,
    WpTemplate,
    WpTemplateMeta,
    WpTemplateSet,
    WpTemplateStatus,
    WorkingPaper,
    RegionType,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 6 built-in seed template sets
# ---------------------------------------------------------------------------

BUILTIN_TEMPLATE_SETS: list[dict[str, Any]] = [
    {
        "set_name": "标准年审",
        "template_codes": ["B1-1", "C1-1", "D1-1", "E1-1", "F1-1", "G1-1"],
        "applicable_audit_type": "annual",
        "applicable_standard": "CAS",
        "description": "适用于一般企业年度审计的标准底稿模板集",
    },
    {
        "set_name": "精简版",
        "template_codes": ["B1-1", "E1-1", "F1-1"],
        "applicable_audit_type": "annual",
        "applicable_standard": "CAS",
        "description": "适用于小型企业年度审计的精简底稿模板集",
    },
    {
        "set_name": "上市公司",
        "template_codes": ["B1-1", "C1-1", "D1-1", "E1-1", "F1-1", "G1-1", "H1-1"],
        "applicable_audit_type": "annual",
        "applicable_standard": "CAS",
        "description": "适用于上市公司年度审计的底稿模板集",
    },
    {
        "set_name": "IPO",
        "template_codes": ["B1-1", "C1-1", "D1-1", "E1-1", "F1-1", "G1-1", "H1-1", "S1-1"],
        "applicable_audit_type": "ipo",
        "applicable_standard": "CAS",
        "description": "适用于IPO审计的底稿模板集",
    },
    {
        "set_name": "国企附注",
        "template_codes": ["N1-1", "N2-1", "N3-1"],
        "applicable_audit_type": "annual",
        "applicable_standard": "CAS_SOE",
        "description": "适用于国有企业附注编制的底稿模板集",
    },
    {
        "set_name": "上市附注",
        "template_codes": ["N1-1", "N2-1", "N3-1", "N4-1"],
        "applicable_audit_type": "annual",
        "applicable_standard": "CAS_LISTED",
        "description": "适用于上市公司附注编制的底稿模板集",
    },
]


class TemplateEngine:
    """底稿模板引擎

    Validates: Requirements 1.1-1.8
    """


    # ------------------------------------------------------------------
    # 6.1  upload_template
    # ------------------------------------------------------------------

    async def upload_template(
        self,
        db: AsyncSession,
        template_code: str,
        template_name: str,
        audit_cycle: str | None = None,
        applicable_standard: str | None = None,
        description: str | None = None,
        created_by: UUID | None = None,
        named_ranges: list[dict] | None = None,
    ) -> WpTemplate:
        """Upload a template: save metadata to DB, parse Named Ranges.

        In MVP mode (no actual .xlsx file), we accept metadata only and
        optionally a list of named_ranges dicts to populate wp_template_meta.

        Validates: Requirements 1.1, 1.2, 1.5
        """
        file_path = f"templates/{template_code}/1.0/{template_code}.xlsx"

        template = WpTemplate(
            template_code=template_code,
            template_name=template_name,
            audit_cycle=audit_cycle,
            applicable_standard=applicable_standard,
            version_major=1,
            version_minor=0,
            status=WpTemplateStatus.draft,
            file_path=file_path,
            description=description,
            created_by=created_by,
        )
        db.add(template)
        await db.flush()

        # Parse Named Ranges → wp_template_meta
        if named_ranges:
            for nr in named_ranges:
                meta = WpTemplateMeta(
                    template_id=template.id,
                    range_name=nr.get("range_name", ""),
                    region_type=nr.get("region_type", RegionType.manual),
                    description=nr.get("description"),
                )
                db.add(meta)

        await db.flush()
        return template

    # ------------------------------------------------------------------
    # 6.2  create_version
    # ------------------------------------------------------------------

    async def create_version(
        self,
        db: AsyncSession,
        template_code: str,
        change_type: str = "minor",
    ) -> WpTemplate:
        """Create a new version of an existing template.

        change_type="major" → major+1, minor=0
        change_type="minor" → minor+1

        Validates: Requirements 1.3
        """
        # Find latest version of this template_code
        result = await db.execute(
            sa.select(WpTemplate)
            .where(
                WpTemplate.template_code == template_code,
                WpTemplate.is_deleted == sa.false(),
            )
            .order_by(
                WpTemplate.version_major.desc(),
                WpTemplate.version_minor.desc(),
            )
            .limit(1)
        )
        latest = result.scalar_one_or_none()
        if latest is None:
            raise ValueError(f"模板 {template_code} 不存在")

        if change_type == "major":
            new_major = latest.version_major + 1
            new_minor = 0
        else:
            new_major = latest.version_major
            new_minor = latest.version_minor + 1

        version_str = f"{new_major}.{new_minor}"
        file_path = f"templates/{template_code}/{version_str}/{template_code}.xlsx"

        new_template = WpTemplate(
            template_code=template_code,
            template_name=latest.template_name,
            audit_cycle=latest.audit_cycle,
            applicable_standard=latest.applicable_standard,
            version_major=new_major,
            version_minor=new_minor,
            status=WpTemplateStatus.draft,
            file_path=file_path,
            description=latest.description,
            created_by=latest.created_by,
        )
        db.add(new_template)
        await db.flush()
        return new_template

    # ------------------------------------------------------------------
    # 6.3  delete_template
    # ------------------------------------------------------------------

    async def delete_template(
        self,
        db: AsyncSession,
        template_id: UUID,
    ) -> None:
        """Soft-delete a template after checking no project references it.

        Validates: Requirements 1.4
        """
        result = await db.execute(
            sa.select(WpTemplate).where(WpTemplate.id == template_id)
        )
        template = result.scalar_one_or_none()
        if template is None:
            raise ValueError("模板不存在")

        # Check if any working_paper references this template's code
        ref_result = await db.execute(
            sa.select(sa.func.count()).select_from(WorkingPaper).where(
                WorkingPaper.is_deleted == sa.false(),
                WorkingPaper.file_path.contains(template.template_code),
            )
        )
        ref_count = ref_result.scalar() or 0
        if ref_count > 0:
            raise ValueError(
                f"模板 {template.template_code} 已被 {ref_count} 个项目底稿引用，无法删除"
            )

        template.soft_delete()
        await db.flush()

    # ------------------------------------------------------------------
    # 6.4  Template set management
    # ------------------------------------------------------------------

    async def get_template(
        self,
        db: AsyncSession,
        template_code: str,
        version: str | None = None,
    ) -> WpTemplate | None:
        """Get a template by code. Default: latest version.

        Validates: Requirements 1.1
        """
        query = sa.select(WpTemplate).where(
            WpTemplate.template_code == template_code,
            WpTemplate.is_deleted == sa.false(),
        )
        if version and "." in version:
            major, minor = version.split(".", 1)
            query = query.where(
                WpTemplate.version_major == int(major),
                WpTemplate.version_minor == int(minor),
            )
        else:
            query = query.order_by(
                WpTemplate.version_major.desc(),
                WpTemplate.version_minor.desc(),
            )
        query = query.limit(1)
        result = await db.execute(query)
        return result.scalar_one_or_none()

    async def list_templates(
        self,
        db: AsyncSession,
        audit_cycle: str | None = None,
        applicable_standard: str | None = None,
    ) -> list[WpTemplate]:
        """List templates with optional filters."""
        query = sa.select(WpTemplate).where(
            WpTemplate.is_deleted == sa.false()
        )
        if audit_cycle:
            query = query.where(WpTemplate.audit_cycle == audit_cycle)
        if applicable_standard:
            query = query.where(WpTemplate.applicable_standard == applicable_standard)
        query = query.order_by(WpTemplate.template_code, WpTemplate.version_major.desc())
        result = await db.execute(query)
        return list(result.scalars().all())

    async def get_template_sets(self, db: AsyncSession) -> list[WpTemplateSet]:
        """Get all template sets.

        Validates: Requirements 1.6
        """
        result = await db.execute(
            sa.select(WpTemplateSet)
            .where(WpTemplateSet.is_deleted == sa.false())
            .order_by(WpTemplateSet.set_name)
        )
        return list(result.scalars().all())

    async def get_template_set(
        self, db: AsyncSession, set_id: UUID
    ) -> WpTemplateSet | None:
        """Get a single template set by ID."""
        result = await db.execute(
            sa.select(WpTemplateSet).where(
                WpTemplateSet.id == set_id,
                WpTemplateSet.is_deleted == sa.false(),
            )
        )
        return result.scalar_one_or_none()

    async def create_template_set(
        self,
        db: AsyncSession,
        set_name: str,
        template_codes: list[str] | None = None,
        applicable_audit_type: str | None = None,
        applicable_standard: str | None = None,
        description: str | None = None,
    ) -> WpTemplateSet:
        """Create a new template set."""
        ts = WpTemplateSet(
            set_name=set_name,
            template_codes=template_codes or [],
            applicable_audit_type=applicable_audit_type,
            applicable_standard=applicable_standard,
            description=description,
        )
        db.add(ts)
        await db.flush()
        return ts

    async def update_template_set(
        self,
        db: AsyncSession,
        set_id: UUID,
        set_name: str | None = None,
        template_codes: list[str] | None = None,
        applicable_audit_type: str | None = None,
        applicable_standard: str | None = None,
        description: str | None = None,
    ) -> WpTemplateSet:
        """Update an existing template set."""
        result = await db.execute(
            sa.select(WpTemplateSet).where(WpTemplateSet.id == set_id)
        )
        ts = result.scalar_one_or_none()
        if ts is None:
            raise ValueError("模板集不存在")
        if set_name is not None:
            ts.set_name = set_name
        if template_codes is not None:
            ts.template_codes = template_codes
        if applicable_audit_type is not None:
            ts.applicable_audit_type = applicable_audit_type
        if applicable_standard is not None:
            ts.applicable_standard = applicable_standard
        if description is not None:
            ts.description = description
        await db.flush()
        return ts

    async def seed_builtin_template_sets(self, db: AsyncSession) -> list[WpTemplateSet]:
        """Seed the 6 built-in template sets (idempotent).

        Validates: Requirements 1.7
        """
        created = []
        for data in BUILTIN_TEMPLATE_SETS:
            existing = await db.execute(
                sa.select(WpTemplateSet).where(
                    WpTemplateSet.set_name == data["set_name"]
                )
            )
            if existing.scalar_one_or_none() is not None:
                continue
            ts = WpTemplateSet(
                set_name=data["set_name"],
                template_codes=data["template_codes"],
                applicable_audit_type=data["applicable_audit_type"],
                applicable_standard=data["applicable_standard"],
                description=data["description"],
            )
            db.add(ts)
            created.append(ts)
        await db.flush()
        return created

    # ------------------------------------------------------------------
    # 6.5  generate_project_workpapers
    # ------------------------------------------------------------------

    async def generate_project_workpapers(
        self,
        db: AsyncSession,
        project_id: UUID,
        template_set_id: UUID,
        year: int = 2025,
        created_by: UUID | None = None,
    ) -> list[WorkingPaper]:
        """Generate project workpapers from a template set.

        1. Iterate template codes in the set
        2. For each code, find the latest published (or draft) template
        3. Create wp_index record
        4. Create working_paper record with file_path in project dir
        5. Copy template .xlsx to project workpaper directory

        Validates: Requirements 1.8, 6.2, 6.3
        """
        import json
        import shutil
        from pathlib import Path

        ts_result = await db.execute(
            sa.select(WpTemplateSet).where(WpTemplateSet.id == template_set_id)
        )
        template_set = ts_result.scalar_one_or_none()
        if template_set is None:
            raise ValueError("模板集不存在")

        template_codes = template_set.template_codes or []
        workpapers: list[WorkingPaper] = []

        # 加载程序裁剪结果（跳过被裁剪的底稿）
        from app.models.procedure_models import ProcedureInstance
        trimmed_q = sa.select(ProcedureInstance.wp_code, ProcedureInstance.status).where(
            ProcedureInstance.project_id == project_id,
            ProcedureInstance.is_deleted == sa.false(),
            ProcedureInstance.status.in_(["skip", "not_applicable"]),
        )
        trimmed_codes = set()
        for wp_code, status in (await db.execute(trimmed_q)).all():
            if wp_code:
                trimmed_codes.add(wp_code)

        # 加载模板库索引（用于查找模板文件路径）
        lib_path = Path(__file__).parent.parent.parent / "data" / "gt_template_library.json"
        template_lib: dict[str, dict] = {}
        if lib_path.exists():
            try:
                with open(lib_path, "r", encoding="utf-8-sig") as f:
                    lib_data = json.load(f)
                for item in lib_data:
                    template_lib[item.get("wp_code", "")] = item
            except Exception:
                pass

        # 项目底稿目录（按审计循环分子目录）
        project_wp_dir = Path("storage") / "projects" / str(project_id) / "workpapers"
        project_wp_dir.mkdir(parents=True, exist_ok=True)

        for code in template_codes:
            # 跳过被裁剪的底稿
            if code in trimmed_codes:
                logger.info("skip trimmed workpaper: %s", code)
                continue

            # Find latest template for this code
            tpl_result = await db.execute(
                sa.select(WpTemplate)
                .where(
                    WpTemplate.template_code == code,
                    WpTemplate.is_deleted == sa.false(),
                )
                .order_by(
                    WpTemplate.version_major.desc(),
                    WpTemplate.version_minor.desc(),
                )
                .limit(1)
            )
            tpl = tpl_result.scalar_one_or_none()

            # Determine wp_name from template or fallback
            lib_entry = template_lib.get(code, {})
            wp_name = tpl.template_name if tpl else lib_entry.get("wp_name", f"底稿{code}")
            audit_cycle = (tpl.audit_cycle if tpl else lib_entry.get("cycle_prefix")) or None

            # Create wp_index
            wp_index = WpIndex(
                project_id=project_id,
                wp_code=code,
                wp_name=wp_name,
                audit_cycle=audit_cycle,
                status=WpStatus.not_started,
            )
            db.add(wp_index)
            await db.flush()

            # 目标文件路径（按循环分子目录）
            cycle_dir = project_wp_dir / (audit_cycle or "OTHER")
            cycle_dir.mkdir(parents=True, exist_ok=True)
            dest_file = cycle_dir / f"{code}.xlsx"
            file_path = str(dest_file)

            # 尝试复制模板文件
            copied = False
            # 优先从模板库索引的 file_path 复制
            src_path_str = lib_entry.get("file_path", "")
            if src_path_str:
                src = Path(src_path_str)
                if src.exists() and src.is_file():
                    shutil.copy2(src, dest_file)
                    copied = True
                else:
                    logger.debug("template source not found: %s (code=%s)", src_path_str, code)

            # 其次从 WpTemplate.file_path 复制
            if not copied and tpl and tpl.file_path:
                tpl_src = Path(tpl.file_path)
                if tpl_src.exists() and tpl_src.is_file():
                    shutil.copy2(tpl_src, dest_file)
                    copied = True
                else:
                    logger.debug("WpTemplate file not found: %s (code=%s)", tpl.file_path, code)

            # 如果都没有源文件，创建空白 xlsx（含标准表头）
            if not copied:
                logger.info("no template source for %s, creating blank workpaper with header", code)
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = code
                    ws["A1"] = f"底稿编号: {code}"
                    ws["A2"] = f"底稿名称: {wp_name}"
                    ws["A3"] = f"审计年度: {year}"
                    wb.save(str(dest_file))
                    wb.close()
                except Exception:
                    # openpyxl 不可用时写空文件
                    dest_file.write_bytes(b"")

            # Create working_paper
            wp = WorkingPaper(
                project_id=project_id,
                wp_index_id=wp_index.id,
                file_path=file_path,
                source_type=WpSourceType.template,
                file_version=1,
                created_by=created_by,
            )
            db.add(wp)

            # 填充底稿表头（编制单位/审计期间/索引号/交叉索引等）
            try:
                from app.services.wp_header_service import fill_workpaper_header
                await fill_workpaper_header(
                    db=db, project_id=project_id, wp_id=wp.id,
                    file_path=file_path, wp_code=code, wp_name=wp_name,
                    cycle=audit_cycle,
                )
            except Exception as _e:
                logger.warning("fill header failed for %s: %s", code, _e)

            workpapers.append(wp)

        await db.flush()
        logger.info("generate_project_workpapers: project=%s codes=%d files=%d", project_id, len(template_codes), len(workpapers))
        return workpapers
