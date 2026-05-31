"""合并报表服务 (ConsolReportService)

核心功能：
- generate_consol_reports — 复用 Phase 1 Report_Engine，数据源切换为 consol_trial.consol_amount
- generate_consol_workpaper — 生成合并底稿.xlsx
- verify_balance — 合并资产负债表平衡校验
- generate_consol_notes — 生成合并附注.xlsx

Validates: Requirements (Phase 2 合并报表)
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.consolidation_models import (
    Company,
    ConsolScope,
    ConsolTrial,
    EliminationEntry,
    GoodwillCalc,
    MinorityInterest,
    AccountCategory,
)
from app.models.audit_platform_models import TrialBalance
from app.models.report_models import FinancialReport
from app.models.report_models import FinancialReportType
from app.models.consolidation_schemas import (
    ConsolTrialRow,
    BalanceCheckResult,
    ConsolWorkpaperResult,
    ConsolReportRow,
    ConsolDisclosureSection,
)
from app.services.goodwill_service import get_goodwill_list
from app.services.minority_interest_service import get_mi_list

logger = logging.getLogger(__name__)


# ============================================================================
# 辅助函数
# ============================================================================

def _decimal(v) -> Decimal:
    """安全转换为 Decimal"""
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _safe_float(v) -> float:
    """安全转换为 float"""
    try:
        return float(_decimal(v))
    except (InvalidOperation, ValueError, TypeError):
        return 0.0


# ============================================================================
# 合并报表服务
# ============================================================================

class ConsolReportService:
    """合并报表服务

    复用 Phase 1 Report_Engine 公式解析逻辑，
    数据源从 trial_balance 切换为 consol_trial。
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    # ------------------------------------------------------------------
    # 核心方法：生成合并报表
    # ------------------------------------------------------------------

    async def generate_consol_reports(
        self,
        project_id: UUID,
        year: int,
        applicable_standard: str = "enterprise",
    ) -> dict[str, list[dict]]:
        """
        生成合并报表（资产负债表、利润表）。

        A1/A2：复用 report_engine.evaluate_formula 公式解析逻辑，
        经注入 ConsolTrialResolver 把取数源切换为 consol_trial.consol_amount。
        单体/合并公式语义完全一致（仅取数源不同，关联属性 Q1）。

        Returns:
            dict: {report_type: [row_dicts]}
        """
        from app.models.report_models import ReportConfig
        from app.services.report_engine import evaluate_formula, ReportFormulaParser
        from app.services.amount_resolver import ConsolTrialResolver

        # 加载报表配置
        result = await self.db.execute(
            sa.select(ReportConfig)
            .where(
                ReportConfig.applicable_standard == applicable_standard,
                ReportConfig.is_deleted.is_(False),
            )
            .order_by(ReportConfig.report_type, ReportConfig.row_number)
        )
        configs = list(result.scalars().all())

        # 按报表类型分组
        configs_by_type: dict[FinancialReportType, list[ReportConfig]] = {}
        for cfg in configs:
            configs_by_type.setdefault(cfg.report_type, []).append(cfg)

        results: dict[str, list[dict]] = {}
        global_row_cache: dict[str, Decimal] = {}
        now = datetime.now(timezone.utc)

        # 注入合并数据源（current = 本年，prior = 上年）
        resolver_current = ConsolTrialResolver(self.db, project_id, year)
        resolver_prior = ConsolTrialResolver(self.db, project_id, year - 1)
        # 纯函数提取器（extract_account_codes 不触 DB）
        extractor = ReportFormulaParser(self.db, project_id, year)

        # 处理顺序（支持跨报表 ROW() 引用）
        type_order = [
            FinancialReportType.balance_sheet,
            FinancialReportType.income_statement,
        ]

        for report_type in type_order:
            config_rows = configs_by_type.get(report_type, [])
            if not config_rows:
                continue

            report_rows = []
            for config in sorted(config_rows, key=lambda r: r.row_number):
                # 执行公式（合并数据源，复用统一引擎）
                current_amount = await evaluate_formula(
                    config.formula, resolver=resolver_current, row_cache=global_row_cache,
                )
                prior_amount = await evaluate_formula(
                    config.formula, resolver=resolver_prior, row_cache={},
                )

                # 更新全局行缓存
                global_row_cache[config.row_code] = current_amount

                # 提取源科目
                source_accounts = extractor.extract_account_codes(config.formula)

                # 写入/更新 financial_report 表
                existing_result = await self.db.execute(
                    sa.select(FinancialReport).where(
                        FinancialReport.project_id == project_id,
                        FinancialReport.year == year,
                        FinancialReport.report_type == report_type,
                        FinancialReport.row_code == config.row_code,
                        FinancialReport.is_deleted.is_(False),
                    )
                )
                existing = existing_result.scalar_one_or_none()

                if existing:
                    existing.row_name = config.row_name
                    existing.current_period_amount = current_amount
                    existing.prior_period_amount = prior_amount
                    existing.formula_used = config.formula
                    existing.source_accounts = source_accounts if source_accounts else None
                    existing.generated_at = now
                else:
                    fr = FinancialReport(
                        project_id=project_id,
                        year=year,
                        report_type=report_type,
                        row_code=config.row_code,
                        row_name=config.row_name,
                        current_period_amount=current_amount,
                        prior_period_amount=prior_amount,
                        formula_used=config.formula,
                        source_accounts=source_accounts if source_accounts else None,
                        generated_at=now,
                    )
                    self.db.add(fr)

                report_rows.append({
                    "row_code": config.row_code,
                    "row_name": config.row_name,
                    "current_period_amount": str(current_amount),
                    "prior_period_amount": str(prior_amount),
                    "formula_used": config.formula,
                    "source_accounts": source_accounts,
                })

            results[report_type.value] = report_rows

        await self.db.flush()
        return results


    # ------------------------------------------------------------------
    # 合并资产负债表平衡校验
    # ------------------------------------------------------------------

    async def verify_balance(
        self,
        project_id: UUID,
        year: int,
    ) -> BalanceCheckResult:
        """
        合并资产负债表平衡校验。

        校验规则：
        - 资产总计 = 负债总计 + 所有者权益总计
        - 包含商誉和少数股东权益的权益合计

        Returns:
            BalanceCheckResult: 校验结果
        """
        # 从 financial_report 表获取合并报表数据
        bs_result = await self.db.execute(
            sa.select(FinancialReport)
            .where(
                FinancialReport.project_id == project_id,
                FinancialReport.year == year,
                FinancialReport.report_type == FinancialReportType.balance_sheet,
                FinancialReport.is_deleted.is_(False),
            )
            .order_by(FinancialReport.row_code)
        )
        bs_rows = list(bs_result.scalars().all())

        if not bs_rows:
            # 如果报表不存在，尝试从合并试算表计算
            return await self._verify_balance_from_consol_trial(project_id, year)

        # 解析资产负债表行
        total_assets = Decimal("0")
        total_liabilities = Decimal("0")
        total_equity = Decimal("0")
        goodwill_amount = Decimal("0")
        minority_interest_amount = Decimal("0")

        for row in bs_rows:
            amount = row.current_period_amount or Decimal("0")
            row_code = row.row_code or ""

            # 识别行次代码（需要与 report_config 中定义的 row_code 一致）
            if "BS-0" in row_code or "资产总计" in (row.row_name or ""):
                if "非流动资产" not in (row.row_name or "") and "流动资产" not in (row.row_name or ""):
                    # 资产总计行
                    total_assets = amount
            elif "BS-1" in row_code or "负债合计" in (row.row_name or ""):
                total_liabilities = amount
            elif "BS-2" in row_code or "所有者权益" in (row.row_name or ""):
                total_equity = amount
            elif "商誉" in (row.row_name or ""):
                goodwill_amount = amount
            elif "少数股东权益" in (row.row_name or ""):
                minority_interest_amount = amount

        # 计算权益合计（含少数股东权益）
        equity_with_mi = total_equity + minority_interest_amount

        # 校验
        difference = total_assets - (total_liabilities + equity_with_mi)
        is_balanced = abs(difference) < Decimal("1")  # 允许 1 元以内误差

        issues = []
        if not is_balanced:
            issues.append(
                f"资产负债表不平衡：资产总计 {total_assets} ≠ "
                f"负债 {total_liabilities} + 权益 {equity_with_mi}，"
                f"差额 {difference}"
            )

        if total_assets == Decimal("0"):
            issues.append("资产总计为 0，请先生成合并报表")

        return BalanceCheckResult(
            is_balanced=is_balanced,
            total_assets=total_assets,
            total_liabilities=total_liabilities,
            total_equity=equity_with_mi,
            minority_interest=minority_interest_amount,
            goodwill=goodwill_amount,
            difference=difference,
            issues=issues,
        )

    async def _verify_balance_from_consol_trial(
        self,
        project_id: UUID,
        year: int,
    ) -> BalanceCheckResult:
        """从合并试算表计算平衡校验"""
        consol_result = await self.db.execute(
            sa.select(ConsolTrial).where(
                ConsolTrial.project_id == project_id,
                ConsolTrial.year == year,
                ConsolTrial.is_deleted.is_(False),
            )
        )
        consol_rows = list(consol_result.scalars().all())

        if not consol_rows:
            return BalanceCheckResult(
                is_balanced=False,
                issues=["合并试算表为空，请先生成合并试算表"],
            )

        total_assets = Decimal("0")
        total_liabilities = Decimal("0")
        total_equity = Decimal("0")

        # 商誉科目代码（常见代码，需与 account_chart 对应）
        goodwill_code_patterns = ["商誉", "1604", "1n"]

        for row in consol_rows:
            amount = row.consol_amount or Decimal("0")
            category = row.account_category

            if category == AccountCategory.asset:
                total_assets += amount
            elif category == AccountCategory.liability:
                total_liabilities += amount
            elif category == AccountCategory.equity:
                total_equity += amount

        # 获取少数股东权益
        mi_records = await get_mi_list(self.db, project_id, year)
        mi_total = sum(
            (r.minority_equity or Decimal("0")) for r in mi_records
        )

        equity_with_mi = total_equity + mi_total

        # 获取商誉
        gw_records = await get_goodwill_list(self.db, project_id, year)
        gw_total = sum(
            (r.carrying_amount or Decimal("0")) for r in gw_records
        )

        # 资产包含商誉
        assets_with_goodwill = total_assets + gw_total

        # 校验
        difference = assets_with_goodwill - (total_liabilities + equity_with_mi)
        is_balanced = abs(difference) < Decimal("1")

        issues = []
        if not is_balanced:
            issues.append(
                f"资产负债表不平衡（含商誉 {gw_total}）："
                f"资产 {assets_with_goodwill} ≠ 负债 {total_liabilities} + 权益 {equity_with_mi}，"
                f"差额 {difference}"
            )

        return BalanceCheckResult(
            is_balanced=is_balanced,
            total_assets=assets_with_goodwill,
            total_liabilities=total_liabilities,
            total_equity=equity_with_mi,
            minority_interest=mi_total,
            goodwill=gw_total,
            difference=difference,
            issues=issues,
        )

    # ------------------------------------------------------------------
    # 生成合并底稿
    # ------------------------------------------------------------------

    async def generate_consol_workpaper(
        self,
        project_id: UUID,
        year: int,
    ) -> ConsolWorkpaperResult:
        """
        生成合并底稿.xlsx。

        包含 4 个 Sheet：
        1. 各公司审定数并列（列=公司，行=科目）
        2. 抵消分录汇总（按类型分组）
        3. 合并试算表（汇总数/调整/抵消/合并数）
        4. 勾稽校验（资产=负债+权益、借贷平衡等）

        Returns:
            ConsolWorkpaperResult: 包含文件路径或 BytesIO
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Sheet 1: 各公司审定数并列
        await self._create_company_trial_sheet(wb, project_id, year)

        # Sheet 2: 抵消分录汇总
        await self._create_elimination_sheet(wb, project_id, year)

        # Sheet 3: 合并试算表
        await self._create_consol_trial_sheet(wb, project_id, year)

        # Sheet 4: 勾稽校验
        await self._create_verification_sheet(wb, project_id, year)

        # 保存到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return ConsolWorkpaperResult(
            file_data=output.getvalue(),
            file_name=f"合并底稿_{year}.xlsx",
            sheet_count=4,
            message="合并底稿生成成功",
        )

    async def _create_company_trial_sheet(
        self,
        wb: "openpyxl.Workbook",
        project_id: UUID,
        year: int,
    ) -> None:
        """Sheet 1: 各公司审定数并列"""
        from openpyxl.utils import get_column_letter

        ws = wb.active
        ws.title = "各公司审定数"

        # 获取合并范围内的公司
        companies_result = await self.db.execute(
            sa.select(Company)
            .join(ConsolScope, sa.and_(
                ConsolScope.company_code == Company.company_code,
                ConsolScope.project_id == project_id,
                ConsolScope.year == year,
                ConsolScope.is_included.is_(True),
            ))
            .where(
                Company.project_id == project_id,
                Company.is_active.is_(True),
                Company.is_deleted.is_(False),
            )
        )
        companies = list(companies_result.scalars().all())

        # 获取所有试算表科目
        codes_result = await self.db.execute(
            sa.select(TrialBalance.standard_account_code)
            .where(
                TrialBalance.project_id == project_id,
                TrialBalance.year == year,
                TrialBalance.is_deleted.is_(False),
            )
            .distinct()
            .order_by(TrialBalance.standard_account_code)
        )
        all_codes = list(codes_result.all())

        # 批量预加载 (company_code, account_code) → audited_amount，避免 N×M 查询
        company_codes = [c.company_code for c in companies]
        tb_result = await self.db.execute(
            sa.select(
                TrialBalance.company_code,
                TrialBalance.standard_account_code,
                TrialBalance.audited_amount,
            ).where(
                TrialBalance.project_id == project_id,
                TrialBalance.year == year,
                TrialBalance.company_code.in_(company_codes) if company_codes else sa.false(),
                TrialBalance.is_deleted.is_(False),
            )
        )
        tb_map: dict[tuple[str, str], Decimal] = {}
        for cc, code, amount in tb_result.all():
            tb_map[(cc, code)] = amount if amount is not None else Decimal("0")

        # 表头：公司名称
        ws.cell(row=1, column=1, value="科目代码")
        ws.cell(row=1, column=2, value="科目名称")
        for i, company in enumerate(companies, start=3):
            ws.cell(row=1, column=i, value=company.company_name)

        # 填充数据
        for row_idx, (code,) in enumerate(all_codes, start=2):
            ws.cell(row=row_idx, column=1, value=code)

            # 获取该科目在每个公司的审定数
            for col_idx, company in enumerate(companies, start=3):
                amount = tb_map.get((company.company_code, code), Decimal("0"))
                ws.cell(row=row_idx, column=col_idx, value=float(amount))

            # 汇总数列
            total_col = len(companies) + 3
            if row_idx == 2:
                ws.cell(row=1, column=total_col, value="合计（审定数）")

            # 计算合计
            formula = f"=SUM({get_column_letter(3)}{row_idx}:{get_column_letter(len(companies) + 2)}{row_idx})"
            ws.cell(row=row_idx, column=total_col, value=formula)

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        for i in range(3, len(companies) + 4):
            ws.column_dimensions[get_column_letter(i)].width = 15

    async def _create_elimination_sheet(
        self,
        wb: "openpyxl.Workbook",
        project_id: UUID,
        year: int,
    ) -> None:
        """Sheet 2: 抵消分录汇总"""
        ws = wb.create_sheet(title="抵消分录汇总")

        # 表头
        headers = ["分录号", "类型", "描述", "科目代码", "科目名称", "借方金额", "贷方金额", "状态", "是否连续编制"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # 获取抵消分录
        entries_result = await self.db.execute(
            sa.select(EliminationEntry)
            .where(
                EliminationEntry.project_id == project_id,
                EliminationEntry.year == year,
                EliminationEntry.is_deleted.is_(False),
            )
            .order_by(EliminationEntry.entry_no, EliminationEntry.id)
        )
        entries = list(entries_result.scalars().all())

        for row_idx, entry in enumerate(entries, start=2):
            ws.cell(row=row_idx, column=1, value=entry.entry_no or "")
            ws.cell(row=row_idx, column=2, value=entry.entry_type.value if entry.entry_type else "")
            ws.cell(row=row_idx, column=3, value=entry.description or "")
            ws.cell(row=row_idx, column=4, value=entry.account_code or "")
            ws.cell(row=row_idx, column=5, value=entry.account_name or "")
            ws.cell(row=row_idx, column=6, value=float(entry.debit_amount or Decimal("0")))
            ws.cell(row=row_idx, column=7, value=float(entry.credit_amount or Decimal("0")))
            ws.cell(row=row_idx, column=8, value=entry.review_status.value if entry.review_status else "")
            ws.cell(row=row_idx, column=9, value="是" if entry.is_continuous else "否")

        # 汇总行
        last_row = len(entries) + 2
        ws.cell(row=last_row, column=1, value="合计")
        ws.cell(row=last_row, column=6, value=f"=SUM(F2:F{last_row - 1})")
        ws.cell(row=last_row, column=7, value=f"=SUM(G2:G{last_row - 1})")

        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15

    async def _create_consol_trial_sheet(
        self,
        wb: "openpyxl.Workbook",
        project_id: UUID,
        year: int,
    ) -> None:
        """Sheet 3: 合并试算表"""
        ws = wb.create_sheet(title="合并试算表")

        # 表头
        headers = ["科目代码", "科目名称", "类别", "个别报表合计", "合并调整", "合并抵消", "合并数"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # 获取合并试算表数据
        consol_result = await self.db.execute(
            sa.select(ConsolTrial)
            .where(
                ConsolTrial.project_id == project_id,
                ConsolTrial.year == year,
                ConsolTrial.is_deleted.is_(False),
            )
            .order_by(ConsolTrial.standard_account_code)
        )
        consol_rows = list(consol_result.scalars().all())

        for row_idx, ct_row in enumerate(consol_rows, start=2):
            ws.cell(row=row_idx, column=1, value=ct_row.standard_account_code)
            ws.cell(row=row_idx, column=2, value=ct_row.account_name or "")
            ws.cell(row=row_idx, column=3, value=ct_row.account_category.value if ct_row.account_category else "")
            ws.cell(row=row_idx, column=4, value=float(ct_row.individual_sum or Decimal("0")))
            ws.cell(row=row_idx, column=5, value=float(ct_row.consol_adjustment or Decimal("0")))
            ws.cell(row=row_idx, column=6, value=float(ct_row.consol_elimination or Decimal("0")))
            ws.cell(row=row_idx, column=7, value=float(ct_row.consol_amount or Decimal("0")))

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18

    async def _create_verification_sheet(
        self,
        wb: "openpyxl.Workbook",
        project_id: UUID,
        year: int,
    ) -> None:
        """Sheet 4: 勾稽校验"""
        from openpyxl.styles import Font

        ws = wb.create_sheet(title="勾稽校验")

        # 获取平衡校验结果
        balance_result = await self.verify_balance(project_id, year)

        # 校验项目
        ws.cell(row=1, column=1, value="合并报表勾稽校验")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        check_items = [
            ("1. 资产负债表平衡校验", ""),
            ("  资产总计", float(balance_result.total_assets)),
            ("  负债总计", float(balance_result.total_liabilities)),
            ("  所有者权益合计", float(balance_result.total_equity)),
            ("  其中：少数股东权益", float(balance_result.minority_interest or Decimal("0"))),
            ("  其中：商誉", float(balance_result.goodwill or Decimal("0"))),
            ("  差额（资产-负债-权益）", float(balance_result.difference)),
            ("  校验结果", "通过 ✓" if balance_result.is_balanced else "不通过 ✗"),
        ]

        for row_idx, (label, value) in enumerate(check_items, start=3):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        # 借贷平衡校验（抵消分录）
        row_idx = len(check_items) + 5
        ws.cell(row=row_idx, column=1, value="2. 抵消分录借贷平衡校验")
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

        entries_result = await self.db.execute(
            sa.select(EliminationEntry).where(
                EliminationEntry.project_id == project_id,
                EliminationEntry.year == year,
                EliminationEntry.is_deleted.is_(False),
            )
        )
        entries = list(entries_result.scalars().all())

        total_debit = sum(e.debit_amount or Decimal("0") for e in entries)
        total_credit = sum(e.credit_amount or Decimal("0") for e in entries)
        debit_credit_diff = total_debit - total_credit

        check_items2 = [
            ("  抵消分录借方合计", float(total_debit)),
            ("  抵消分录贷方合计", float(total_credit)),
            ("  差额", float(debit_credit_diff)),
            ("  校验结果", "通过 ✓" if abs(debit_credit_diff) < Decimal("1") else "不通过 ✗"),
        ]

        for i, (label, value) in enumerate(check_items2, start=1):
            ws.cell(row=row_idx + i, column=1, value=label)
            ws.cell(row=row_idx + i, column=2, value=value)

        # 问题列表
        if balance_result.issues:
            issue_row = row_idx + len(check_items2) + 2
            ws.cell(row=issue_row, column=1, value="发现问题：")
            ws.cell(row=issue_row, column=1).font = Font(bold=True, color="FF0000")
            for i, issue in enumerate(balance_result.issues, start=1):
                ws.cell(row=issue_row + i, column=1, value=f"  • {issue}")

        # 设置列宽
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20

    # ------------------------------------------------------------------
    # 生成合并附注
    # ------------------------------------------------------------------

    async def generate_consol_notes(
        self,
        project_id: UUID,
        year: int,
    ) -> list[ConsolDisclosureSection]:
        """
        生成合并附注。

        包含合并特有附注章节：
        1. 合并范围说明
        2. 重要子公司信息表
        3. 合并范围变动说明
        4. 商誉披露
        5. 少数股东权益披露
        6. 内部交易抵消说明
        7. 外币折算披露

        Returns:
            list[ConsolDisclosureSection]: 合并附注章节列表
        """
        sections: list[ConsolDisclosureSection] = []

        # 1. 合并范围说明
        scope_section = await self._generate_scope_section(project_id, year)
        sections.append(scope_section)

        # 2. 重要子公司信息表
        subsidiary_section = await self._generate_subsidiary_section(project_id, year)
        sections.append(subsidiary_section)

        # 3. 商誉披露
        goodwill_section = await self._generate_goodwill_section(project_id, year)
        sections.append(goodwill_section)

        # 4. 少数股东权益披露
        mi_section = await self._generate_minority_interest_section(project_id, year)
        sections.append(mi_section)

        # 5. 内部交易抵消说明
        trade_section = await self._generate_internal_trade_section(project_id, year)
        sections.append(trade_section)

        return sections

    async def _generate_scope_section(
        self,
        project_id: UUID,
        year: int,
    ) -> ConsolDisclosureSection:
        """生成合并范围说明"""
        # 获取合并范围内的公司
        included_result = await self.db.execute(
            sa.select(Company, ConsolScope)
            .join(ConsolScope, sa.and_(
                ConsolScope.company_code == Company.company_code,
                ConsolScope.project_id == project_id,
                ConsolScope.year == year,
            ))
            .where(
                Company.project_id == project_id,
                ConsolScope.is_included.is_(True),
                Company.is_deleted.is_(False),
            )
        )
        included = included_result.all()

        excluded_result = await self.db.execute(
            sa.select(Company, ConsolScope)
            .join(ConsolScope, sa.and_(
                ConsolScope.company_code == Company.company_code,
                ConsolScope.project_id == project_id,
                ConsolScope.year == year,
            ))
            .where(
                Company.project_id == project_id,
                ConsolScope.is_included.is_(False),
                Company.is_deleted.is_(False),
            )
        )
        excluded = excluded_result.all()

        rows = []
        for company, scope in included:
            rows.append({
                "公司名称": company.company_name,
                "公司代码": company.company_code,
                "持股比例": f"{company.shareholding}%" if company.shareholding else "N/A",
                "合并方法": company.consol_method.value if company.consol_method else "全额合并",
                "是否纳入": "是",
                "纳入原因": scope.inclusion_reason.value if scope.inclusion_reason else "子公司",
            })

        for company, scope in excluded:
            rows.append({
                "公司名称": company.company_name,
                "公司代码": company.company_code,
                "持股比例": f"{company.shareholding}%" if company.shareholding else "N/A",
                "合并方法": company.consol_method.value if company.consol_method else "-",
                "是否纳入": "否",
                "排除原因": scope.exclusion_reason or "-",
            })

        return ConsolDisclosureSection(
            section_code="consol_scope",
            section_title="合并范围",
            content_type="table",
            rows=rows,
            summary=f"本期纳入合并范围的子公司共 {len(included)} 家，未纳入合并范围的共 {len(excluded)} 家",
        )

    async def _generate_subsidiary_section(
        self,
        project_id: UUID,
        year: int,
    ) -> ConsolDisclosureSection:
        """生成重要子公司信息表"""
        # 获取重要子公司（持股 50% 以上或全额合并）
        subsidiaries_result = await self.db.execute(
            sa.select(Company)
            .join(ConsolScope, sa.and_(
                ConsolScope.company_code == Company.company_code,
                ConsolScope.project_id == project_id,
                ConsolScope.year == year,
                ConsolScope.is_included.is_(True),
            ))
            .where(
                Company.project_id == project_id,
                Company.is_deleted.is_(False),
                Company.is_active.is_(True),
            )
        )
        subsidiaries = list(subsidiaries_result.scalars().all())

        rows = []
        for sub in subsidiaries:
            rows.append({
                "公司名称": sub.company_name,
                "注册地": sub.functional_currency,  # 简化，实际应从公司详情取
                "业务性质": sub.consol_method.value if sub.consol_method else "其他",
                "注册资本": "-",
                "母公司持股比例": f"{sub.shareholding}%" if sub.shareholding else "N/A",
                "合并方法": "全额合并" if sub.consol_method and sub.consol_method.value == "full" else "权益法",
            })

        return ConsolDisclosureSection(
            section_code="important_subsidiaries",
            section_title="重要子公司情况",
            content_type="table",
            rows=rows,
            summary=f"共有 {len(subsidiaries)} 家重要子公司",
        )

    async def _generate_goodwill_section(
        self,
        project_id: UUID,
        year: int,
    ) -> ConsolDisclosureSection:
        """生成商誉披露"""
        goodwill_records = await get_goodwill_list(self.db, project_id, year)

        rows = []
        total_goodwill = Decimal("0")
        for gw in goodwill_records:
            carrying = gw.carrying_amount or Decimal("0")
            total_goodwill += carrying
            rows.append({
                "被投资单位": gw.subsidiary_company_code,
                "初始确认金额": float(gw.goodwill_amount or Decimal("0")),
                "本期增加": 0.0,
                "本期减少": 0.0,
                "累计减值": float(gw.accumulated_impairment or Decimal("0")),
                "期末账面价值": float(carrying),
                "负商誉标识": "是" if gw.is_negative_goodwill else "否",
            })

        return ConsolDisclosureSection(
            section_code="goodwill",
            section_title="商誉",
            content_type="table",
            rows=rows,
            summary=f"期末商誉账面价值合计 {total_goodwill} 元",
        )

    async def _generate_minority_interest_section(
        self,
        project_id: UUID,
        year: int,
    ) -> ConsolDisclosureSection:
        """生成少数股东权益披露"""
        mi_records = await get_mi_list(self.db, project_id, year)

        rows = []
        total_mi = Decimal("0")
        total_mi_profit = Decimal("0")
        for mi in mi_records:
            equity = mi.minority_equity or Decimal("0")
            profit = mi.minority_profit or Decimal("0")
            total_mi += equity
            total_mi_profit += profit
            rows.append({
                "子公司": mi.subsidiary_company_code,
                "少数股东持股比例": f"{(mi.minority_share_ratio or Decimal('0')):.2f}%"
                    if mi.minority_share_ratio is not None else "N/A",
                "期末少数股东权益": float(equity),
                "本期少数股东损益": float(profit),
                "超额亏损标识": "是" if mi.is_excess_loss else "否",
                "超额亏损金额": float(mi.excess_loss_amount or Decimal("0")),
            })

        return ConsolDisclosureSection(
            section_code="minority_interest",
            section_title="少数股东权益及少数股东损益",
            content_type="table",
            rows=rows,
            summary=f"期末少数股东权益合计 {total_mi} 元，本期少数股东损益合计 {total_mi_profit} 元",
        )

    async def _generate_internal_trade_section(
        self,
        project_id: UUID,
        year: int,
    ) -> ConsolDisclosureSection:
        """生成内部交易抵消说明"""
        # 获取内部交易汇总
        from app.models.consolidation_models import InternalTrade

        trades_result = await self.db.execute(
            sa.select(InternalTrade).where(
                InternalTrade.project_id == project_id,
                InternalTrade.year == year,
                InternalTrade.is_deleted.is_(False),
            )
        )
        trades = list(trades_result.scalars().all())

        total_trade = sum(t.trade_amount or Decimal("0") for t in trades)
        total_unrealized = sum(t.unrealized_profit or Decimal("0") for t in trades)

        rows = []
        for trade in trades:
            rows.append({
                "卖方": trade.seller_company_code,
                "买方": trade.buyer_company_code,
                "交易类型": trade.trade_type.value if trade.trade_type else "其他",
                "交易金额": float(trade.trade_amount or Decimal("0")),
                "未实现利润": float(trade.unrealized_profit or Decimal("0")),
                "期末存货中未实现比例": f"{float(trade.inventory_remaining_ratio or Decimal('0')) * 100:.2f}%"
                    if trade.inventory_remaining_ratio else "0%",
                "描述": trade.description or "",
            })

        return ConsolDisclosureSection(
            section_code="internal_trade_elimination",
            section_title="内部交易抵消",
            content_type="table",
            rows=rows,
            summary=f"本期内部交易合计 {total_trade} 元，未实现利润 {total_unrealized} 元",
        )


# ============================================================================
# 便捷函数（async 风格，供路由层 await 调用）
# A3：统一 async，消除在 AsyncSession 上调同步 Session API 的 MissingGreenlet 风险。
# 保留 *_sync 名称以兼容既有 import，但实现为 async（路由层需 await）。
# ============================================================================

async def generate_consol_reports_sync(
    db: AsyncSession,
    project_id: UUID,
    year: int,
    applicable_standard: str = "enterprise",
) -> dict[str, list[dict]]:
    """生成合并报表（async）"""
    service = ConsolReportService(db)
    return await service.generate_consol_reports(project_id, year, applicable_standard)


async def verify_balance_sync(
    db: AsyncSession,
    project_id: UUID,
    year: int,
) -> BalanceCheckResult:
    """资产负债表平衡校验（async）"""
    service = ConsolReportService(db)
    return await service.verify_balance(project_id, year)


async def generate_consol_workpaper_sync(
    db: AsyncSession,
    project_id: UUID,
    year: int,
) -> ConsolWorkpaperResult:
    """生成合并底稿（async）"""
    service = ConsolReportService(db)
    return await service.generate_consol_workpaper(project_id, year)


async def generate_consol_notes_sync(
    db: AsyncSession,
    project_id: UUID,
    year: int,
) -> list[ConsolDisclosureSection]:
    """生成合并附注（async）"""
    service = ConsolReportService(db)
    return await service.generate_consol_notes(project_id, year)
