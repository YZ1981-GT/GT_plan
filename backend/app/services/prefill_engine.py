"""底稿预填充引擎 — 真正打开 .xlsx 扫描公式并写入计算结果

替代 prefill_service.py 的 stub 实现。

流程：
1. openpyxl 打开底稿 .xlsx（保留公式和格式）
2. 扫描所有单元格，识别 =TB()/=SUM_TB()/=AUX()/=PREV()/=WP() 公式
3. 批量调用 FormulaEngine 执行公式
4. 将结果写入单元格值（保留公式文本到 comment）
5. 保存文件
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.workpaper_models import WorkingPaper

_logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 公式正则 & FormulaCell（兼容旧 prefill_service 接口）
# ---------------------------------------------------------------------------

_FORMULA_RE = re.compile(
    r'=(TB|WP|AUX|PREV|SUM_TB|LEDGER_DETAIL|COUNT_LEDGER|LEDGER|ADJ|NOTE)\s*\(([^)]*)\)',
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# 扩展公式正则（D4：6 种新公式类型）
# ---------------------------------------------------------------------------

_WP_FORMULA_RE = re.compile(r"WP\('([^']+)',\s*'([^']+)',\s*'([^']+)'\)")
_LEDGER_FORMULA_RE = re.compile(r"LEDGER\('([^']+)',\s*'([^']+)',\s*'([^']+)'\)")
_AUX_FORMULA_RE = re.compile(r"AUX\('([^']+)',\s*'([^']+)',\s*'([^']+)',\s*'([^']+)'\)")
_PREV_FORMULA_RE = re.compile(r"PREV\('([^']+)',\s*'([^']+)',\s*'([^']+)'\)")
_ADJ_FORMULA_RE = re.compile(r"ADJ\('([^']+)',\s*'([^']+)'\)")
_NOTE_FORMULA_RE = re.compile(r"NOTE\('([^']+)',\s*'([^']+)',\s*'([^']+)'\)")


class FormulaCell:
    """Represents a formula cell found during scanning."""

    def __init__(self, sheet: str, cell_ref: str, formula_type: str, raw_args: str):
        self.sheet = sheet
        self.cell_ref = cell_ref
        self.formula_type = formula_type.upper()
        self.raw_args = raw_args

    def to_dict(self) -> dict:
        return {"sheet": self.sheet, "cell_ref": self.cell_ref, "formula_type": self.formula_type, "raw_args": self.raw_args}


def _sync_prefill_to_snapshot(snap: dict, structure_data: dict, worksheets) -> None:
    """把 prefill 写入 structure_data 的 value 同步回 univer_snapshot 的 cellData

    structure_data.rows 是跨 sheet 扁平化的，需要按 sheet 拆分后写回 snap.sheets[name].cellData
    """
    sheets_meta = structure_data.get("sheets", [])
    rows_arr = structure_data.get("rows", [])
    order = snap.get("sheet_order_names") or list(snap.get("sheets", {}).keys())
    sheets_map = snap.get("sheets") or {}

    row_offset = 0
    for sheet_idx, ws in enumerate(worksheets):
        sheet_name = order[sheet_idx] if sheet_idx < len(order) else ws.title
        sheet_obj = sheets_map.get(sheet_name)
        if not isinstance(sheet_obj, dict):
            row_offset += ws.max_row or 0
            continue
        cell_data = sheet_obj.get("cellData")
        if not isinstance(cell_data, dict):
            cell_data = {}
            sheet_obj["cellData"] = cell_data

        sheet_rows = ws.max_row or 0
        for r in range(sheet_rows):
            flat_idx = row_offset + r
            if flat_idx >= len(rows_arr):
                break
            row_entry = rows_arr[flat_idx]
            cells = row_entry.get("cells", [])
            for c, cell_info in enumerate(cells):
                v = cell_info.get("value")
                if v is not None and v != "":
                    r_key = str(r)
                    c_key = str(c)
                    if r_key not in cell_data:
                        cell_data[r_key] = {}
                    if c_key not in cell_data[r_key]:
                        cell_data[r_key][c_key] = {}
                    cell_data[r_key][c_key]["v"] = v
        row_offset += sheet_rows


def _parse_args(raw: str) -> list[str]:
    """解析公式参数，处理引号和逗号"""
    args = []
    current = []
    in_quote = False
    for ch in raw:
        if ch == '"' or ch == "'":
            in_quote = not in_quote
        elif ch == ',' and not in_quote:
            args.append(''.join(current).strip().strip('"').strip("'"))
            current = []
            continue
        current.append(ch)
    if current:
        args.append(''.join(current).strip().strip('"').strip("'"))
    return args


# ---------------------------------------------------------------------------
# 6 种新公式解析器（Task 3.1）
# ---------------------------------------------------------------------------


async def _resolve_wp_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=WP('wp_code', 'sheet', 'cell') → 从其他底稿 parsed_data 取值"""
    if len(args) < 3:
        return None
    wp_code, sheet_name, cell_ref = args[0], args[1], args[2]
    from app.models.wp_optimization_models import WpTemplateMetadata
    # 通过 wp_code 找到对应底稿
    result = await db.execute(
        sa.select(WorkingPaper).where(
            WorkingPaper.project_id == project_id,
            WorkingPaper.is_deleted == False,  # noqa: E712
        )
    )
    workpapers = result.scalars().all()
    # 匹配 wp_code（从 wp_index 或 template_metadata）
    for wp in workpapers:
        if wp.parsed_data and wp.parsed_data.get("wp_code") == wp_code:
            cell_data = wp.parsed_data.get("cells", {}).get(f"{sheet_name}!{cell_ref}")
            if cell_data is not None:
                try:
                    return Decimal(str(cell_data))
                except Exception:
                    return None
    return None


async def _resolve_ledger_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=LEDGER('code', 'direction', 'period') → 从 tb_ledger 表取发生额

    支持 H 循环场景：=LEDGER('1602','credit','1月') / '1-6月' / '全年'
    """
    if len(args) < 3:
        return None
    account_code, direction, period = args[0], args[1], args[2]
    from app.models.audit_platform_models import TbLedger
    from app.services.dataset_query import get_active_filter

    active_filter = await get_active_filter(db, TbLedger, project_id, year)
    col = TbLedger.debit_amount if direction.lower() in ("debit", "借") else TbLedger.credit_amount
    q = sa.select(sa.func.coalesce(sa.func.sum(col), 0)).where(
        active_filter,
        TbLedger.account_code == account_code,
    )
    # 期间过滤 — 使用 _parse_period_range 将 "1月"/"1-6月"/"全年" 转为整数范围
    start_p, end_p = _parse_period_range(period)
    if start_p is not None and end_p is not None and (start_p, end_p) != (1, 12):
        q = q.where(TbLedger.accounting_period.between(start_p, end_p))
    result = await db.execute(q)
    val = result.scalar()
    return Decimal(str(val)) if val is not None else Decimal("0")


async def _resolve_aux_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=AUX('code', 'aux_type', 'aux_code', 'column') → 从 tb_aux_balance 取值"""
    if len(args) < 4:
        return None
    account_code, aux_type, aux_code, column = args[0], args[1], args[2], args[3]
    from app.models.dataset_models import TbAuxBalance
    from app.services.dataset_query import get_active_filter

    active_filter = await get_active_filter(db, TbAuxBalance, project_id, year)
    # 映射列名到 ORM 字段
    col_map = {
        "期末余额": TbAuxBalance.closing_balance,
        "期初余额": TbAuxBalance.opening_balance,
        "借方发生额": TbAuxBalance.debit_amount,
        "贷方发生额": TbAuxBalance.credit_amount,
    }
    col_attr = col_map.get(column, TbAuxBalance.closing_balance)
    q = sa.select(sa.func.coalesce(sa.func.sum(col_attr), 0)).where(
        active_filter,
        TbAuxBalance.account_code == account_code,
        TbAuxBalance.aux_type == aux_type,
        TbAuxBalance.aux_code == aux_code,
    )
    result = await db.execute(q)
    val = result.scalar()
    return Decimal(str(val)) if val is not None else Decimal("0")


async def _resolve_prev_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=PREV('wp_code', 'sheet', 'cell') → 从上年底稿取值"""
    if len(args) < 3:
        return None
    wp_code, sheet_name, cell_ref = args[0], args[1], args[2]
    # 查上年底稿（同项目 year-1）
    result = await db.execute(
        sa.select(WorkingPaper).where(
            WorkingPaper.project_id == project_id,
            WorkingPaper.is_deleted == False,  # noqa: E712
        )
    )
    workpapers = result.scalars().all()
    for wp in workpapers:
        if wp.parsed_data and wp.parsed_data.get("wp_code") == wp_code:
            # 尝试从 parsed_data.cells 取值
            cell_data = wp.parsed_data.get("cells", {}).get(f"{sheet_name}!{cell_ref}")
            if cell_data is not None:
                try:
                    return Decimal(str(cell_data))
                except Exception:
                    return None
    return None


async def _resolve_adj_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=ADJ('code', 'type') → 从 adjustments 表取调整金额"""
    if len(args) < 2:
        return None
    account_code, adj_type = args[0], args[1]
    from app.models.phase10_models import Adjustment, AdjustmentEntry

    # adj_type: AJE / RJE
    q = sa.select(
        sa.func.coalesce(
            sa.func.sum(AdjustmentEntry.debit_amount - AdjustmentEntry.credit_amount), 0
        )
    ).join(Adjustment, AdjustmentEntry.adjustment_id == Adjustment.id).where(
        Adjustment.project_id == project_id,
        Adjustment.year == year,
        Adjustment.is_deleted == False,  # noqa: E712
        AdjustmentEntry.standard_account_code == account_code,
    )
    if adj_type.upper() in ("AJE", "审计调整"):
        q = q.where(Adjustment.adjustment_type == "aje")
    elif adj_type.upper() in ("RJE", "重分类"):
        q = q.where(Adjustment.adjustment_type == "rje")
    result = await db.execute(q)
    val = result.scalar()
    return Decimal(str(val)) if val is not None else Decimal("0")


async def _resolve_note_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=NOTE('section', 'row', 'col') → 从 disclosure_notes 取值"""
    if len(args) < 3:
        return None
    section, row_key, col_key = args[0], args[1], args[2]
    from app.models.report_models import DisclosureNote

    result = await db.execute(
        sa.select(DisclosureNote).where(
            DisclosureNote.project_id == project_id,
            DisclosureNote.year == year,
            DisclosureNote.note_section == section,
        )
    )
    note = result.scalar_one_or_none()
    if note and note.table_data:
        # table_data 是 JSONB，按 row/col 索引取值
        rows = note.table_data.get("rows", [])
        for r in rows:
            if str(r.get("key", "")) == row_key or str(r.get("label", "")) == row_key:
                val = r.get("values", {}).get(col_key)
                if val is not None:
                    try:
                        return Decimal(str(val))
                    except Exception:
                        return None
    return None


# ---------------------------------------------------------------------------
# E1 spec Sprint 1 Task 1.4 + 1.5: LEDGER_DETAIL & COUNT_LEDGER
# ---------------------------------------------------------------------------


def _parse_period_range(period: str) -> tuple[int | None, int | None]:
    """解析期间字符串为 (start_period, end_period)。

    支持：
    - "全年"   → (1, 12)
    - "1月"    → (1, 1)
    - "1-3月"  → (1, 3)
    - "3"      → (3, 3)
    - 纯月份数字 1-12
    """
    if not period or period in ("全年", "all", "*"):
        return 1, 12
    s = str(period).strip().replace("月", "")
    if "-" in s or "~" in s:
        sep = "-" if "-" in s else "~"
        parts = [p.strip() for p in s.split(sep)]
        try:
            return int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            return None, None
    try:
        m = int(s)
        return m, m
    except ValueError:
        return None, None


async def _resolve_ledger_detail_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> list[dict[str, Any]]:
    """=LEDGER_DETAIL('科目','日期范围','金额条件') → 返回符合条件的 tb_ledger 明细行列表

    Args (3 个):
      0: account_code   科目编码（精确匹配，可加 % 模糊）
      1: period_range   "全年" / "1月" / "1-3月" / "12月"
      2: amount_filter  金额条件，如 ">=100000" / "<0" / "*"（不过滤）

    Returns:
      [{voucher_date, voucher_no, summary, debit_amount, credit_amount, counterpart_account}, ...]
      按 voucher_date 升序，最多返回 200 行（截断保护）
    """
    if len(args) < 1:
        return []
    account_code = args[0]
    period = args[1] if len(args) > 1 else "全年"
    amount_filter = args[2] if len(args) > 2 else "*"

    from app.models.audit_platform_models import TbLedger
    from app.services.dataset_query import get_active_filter

    active_filter = await get_active_filter(db, TbLedger, project_id, year)
    q = sa.select(
        TbLedger.voucher_date,
        TbLedger.voucher_no,
        TbLedger.summary,
        TbLedger.debit_amount,
        TbLedger.credit_amount,
        TbLedger.counterpart_account,
        TbLedger.account_code,
        TbLedger.account_name,
    ).where(
        active_filter,
    )

    # 科目过滤 — 支持精确 / LIKE 前缀
    if "%" in account_code:
        q = q.where(TbLedger.account_code.like(account_code))
    else:
        q = q.where(TbLedger.account_code == account_code)

    # 期间过滤
    start_p, end_p = _parse_period_range(period)
    if start_p is not None and end_p is not None and (start_p, end_p) != (1, 12):
        q = q.where(
            TbLedger.accounting_period.between(start_p, end_p)
        )

    # 金额条件过滤（>= / > / <= / < / ==）
    af = (amount_filter or "").strip()
    if af and af != "*":
        import re as _re
        m = _re.match(r"^(>=|<=|>|<|==|=)\s*(-?\d+(?:\.\d+)?)\s*$", af)
        if m:
            op_map = {
                ">=": lambda x: (TbLedger.debit_amount >= x) | (TbLedger.credit_amount >= x),
                "<=": lambda x: (TbLedger.debit_amount <= x) | (TbLedger.credit_amount <= x),
                ">": lambda x: (TbLedger.debit_amount > x) | (TbLedger.credit_amount > x),
                "<": lambda x: (TbLedger.debit_amount < x) | (TbLedger.credit_amount < x),
                "==": lambda x: (TbLedger.debit_amount == x) | (TbLedger.credit_amount == x),
                "=": lambda x: (TbLedger.debit_amount == x) | (TbLedger.credit_amount == x),
            }
            try:
                threshold = Decimal(m.group(2))
                q = q.where(op_map[m.group(1)](threshold))
            except Exception:
                pass

    q = q.order_by(TbLedger.voucher_date.asc(), TbLedger.voucher_no.asc()).limit(200)
    result = await db.execute(q)
    rows = result.all()

    out: list[dict[str, Any]] = []
    for r in rows:
        out.append({
            "voucher_date": r.voucher_date.isoformat() if r.voucher_date else None,
            "voucher_no": r.voucher_no,
            "summary": r.summary or "",
            "debit_amount": str(Decimal(str(r.debit_amount))) if r.debit_amount is not None else "0",
            "credit_amount": str(Decimal(str(r.credit_amount))) if r.credit_amount is not None else "0",
            "counterpart_account": r.counterpart_account or "",
            "account_code": r.account_code,
            "account_name": r.account_name or "",
        })
    return out


async def _resolve_count_ledger_formula(
    db: AsyncSession, project_id: UUID, year: int, args: list[str]
) -> Decimal | None:
    """=COUNT_LEDGER('科目','日期范围') → 返回符合条件的 tb_ledger 笔数

    Args (1-2 个):
      0: account_code   科目编码
      1: period_range   "全年" / "1月" / "1-3月" / "12月"
    """
    if len(args) < 1:
        return Decimal("0")
    account_code = args[0]
    period = args[1] if len(args) > 1 else "全年"

    from app.models.audit_platform_models import TbLedger
    from app.services.dataset_query import get_active_filter

    active_filter = await get_active_filter(db, TbLedger, project_id, year)
    q = sa.select(sa.func.count()).select_from(TbLedger).where(active_filter)
    if "%" in account_code:
        q = q.where(TbLedger.account_code.like(account_code))
    else:
        q = q.where(TbLedger.account_code == account_code)

    start_p, end_p = _parse_period_range(period)
    if start_p is not None and end_p is not None and (start_p, end_p) != (1, 12):
        q = q.where(
            TbLedger.accounting_period.between(start_p, end_p)
        )

    result = await db.execute(q)
    val = result.scalar()
    return Decimal(str(val or 0))


# 公式类型 → 解析器映射
_FORMULA_RESOLVERS = {
    "WP": _resolve_wp_formula,
    "LEDGER": _resolve_ledger_formula,
    "AUX": _resolve_aux_formula,
    "PREV": _resolve_prev_formula,
    "ADJ": _resolve_adj_formula,
    "NOTE": _resolve_note_formula,
    # E1 spec Sprint 1 Task 1.4 + 1.5
    "LEDGER_DETAIL": _resolve_ledger_detail_formula,
    "COUNT_LEDGER": _resolve_count_ledger_formula,
    "TB_AUX": None,  # handled separately below
}


async def _resolve_tb_aux(
    db: AsyncSession, project_id: UUID, year: int,
    account_code: str, aux_type: str, column: str,
) -> list[dict[str, Any]]:
    """=TB_AUX('account_code', 'aux_type', 'column') → 从 tb_aux_balance 表查询辅助余额明细

    返回 [{aux_code, aux_name, value}, ...] 列表。
    """
    from app.models.dataset_models import TbAuxBalance
    from app.services.dataset_query import get_active_filter

    active_filter = await get_active_filter(db, TbAuxBalance, project_id, year)

    # 映射列名到 ORM 字段
    col_map = {
        "期末余额": TbAuxBalance.closing_balance,
        "期初余额": TbAuxBalance.opening_balance,
        "借方发生额": TbAuxBalance.debit_amount,
        "贷方发生额": TbAuxBalance.credit_amount,
    }
    col_attr = col_map.get(column, TbAuxBalance.closing_balance)

    q = sa.select(
        TbAuxBalance.aux_code,
        TbAuxBalance.aux_name,
        col_attr.label("value"),
    ).where(
        active_filter,
        TbAuxBalance.account_code == account_code,
        TbAuxBalance.aux_type == aux_type,
    ).order_by(TbAuxBalance.aux_code)

    result = await db.execute(q)
    rows = result.all()

    return [
        {
            "aux_code": r.aux_code or "",
            "aux_name": r.aux_name or "",
            "value": float(r.value) if r.value is not None else 0.0,
        }
        for r in rows
    ]


async def resolve_extended_formula(
    db: AsyncSession, project_id: UUID, year: int,
    formula_type: str, raw_args: str,
) -> Decimal | None:
    """统一入口：解析并执行扩展公式"""
    ft = formula_type.upper()

    # TB_AUX 返回列表而非单值，调用方需特殊处理
    if ft == "TB_AUX":
        args = _parse_args(raw_args)
        if len(args) >= 3:
            results = await _resolve_tb_aux(db, project_id, year, args[0], args[1], args[2])
            # 返回合计值作为 Decimal（调用方如需明细列表应直接调 _resolve_tb_aux）
            total = sum(r["value"] for r in results)
            return Decimal(str(total))
        return None

    # E1 spec Task 1.4: LEDGER_DETAIL 返回列表，统一入口返回笔数（合计意义不大）
    # 调用方如需明细应直接调 _resolve_ledger_detail_formula
    if ft == "LEDGER_DETAIL":
        args = _parse_args(raw_args)
        rows = await _resolve_ledger_detail_formula(db, project_id, year, args)
        return Decimal(str(len(rows)))

    resolver = _FORMULA_RESOLVERS.get(ft)
    if resolver is None:
        return None
    args = _parse_args(raw_args)
    return await resolver(db, project_id, year, args)


async def prefill_workpaper_real(
    db: AsyncSession,
    project_id: UUID,
    year: int,
    wp_id: UUID,
) -> dict[str, Any]:
    """
    真正的预填充：打开 .xlsx → 扫描公式 → 执行 → 写回
    """
    try:
        import openpyxl
    except ImportError:
        return {"wp_id": str(wp_id), "status": "error", "message": "openpyxl 未安装"}

    # 获取底稿文件路径
    wp = (await db.execute(
        sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
    )).scalar_one_or_none()
    if not wp or not wp.file_path:
        return {"wp_id": str(wp_id), "status": "error", "message": "底稿文件不存在"}

    fp = Path(wp.file_path)
    if not fp.exists():
        return {"wp_id": str(wp_id), "status": "error", "message": f"文件不存在: {fp}"}

    # 打开 Excel（保留格式）
    try:
        wb = openpyxl.load_workbook(str(fp), data_only=False)
    except Exception as e:
        return {"wp_id": str(wp_id), "status": "error", "message": f"打开文件失败: {e}"}

    # 扫描所有公式
    formulas_found = []
    for sheet_idx, ws in enumerate(wb.worksheets):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    match = _FORMULA_RE.search(cell.value)
                    if match:
                        formulas_found.append({
                            "sheet": ws.title,
                            "sheet_idx": sheet_idx,
                            "cell": cell.coordinate,
                            "row": cell.row - 1,       # 0-based row index
                            "col": cell.column - 1,    # 0-based column index
                            "formula_type": match.group(1).upper(),
                            "raw_args": match.group(2).strip(),
                            "original": cell.value,
                            "cell_obj": cell,
                        })

    if not formulas_found:
        wb.close()
        return {
            "wp_id": str(wp_id), "status": "ok",
            "formulas_found": 0, "formulas_filled": 0,
            "message": "未发现取数公式",
        }

    # 批量执行公式（需求 46：只更新 parsed_data['univer_snapshot'] 的 v 字段，保留 xlsx 公式）
    from app.services.formula_engine import FormulaEngine
    import json as _json
    engine = FormulaEngine()

    # 读取 parsed_data['univer_snapshot']（Req 6 单源化：从 JSONB 读取，不再读 structure.json）
    structure_data: dict | None = None
    if wp.parsed_data and isinstance(wp.parsed_data, dict):
        univer_snap = wp.parsed_data.get("univer_snapshot")
        if isinstance(univer_snap, dict) and univer_snap.get("sheets"):
            # 从 slim snapshot 重建 structure 格式供 prefill 写值
            from app.routers.workpaper_html_preview import _structure_from_univer_snapshot
            structure_data = _structure_from_univer_snapshot(univer_snap)
    if structure_data is None:
        _logger.warning("prefill_real: univer_snapshot 不存在 wp=%s，预填充将跳过值写入", wp_id)

    filled = 0
    errors = []

    for f in formulas_found:
        args = _parse_args(f["raw_args"])
        ft = f["formula_type"]

        params: dict[str, Any] = {}
        if ft == "TB" and len(args) >= 2:
            params = {"account_code": args[0], "column_name": args[1]}
        elif ft == "SUM_TB" and len(args) >= 2:
            params = {"account_range": args[0], "column_name": args[1]}
        elif ft == "AUX" and len(args) >= 4:
            params = {"account_code": args[0], "aux_dimension": args[1], "dimension_value": args[2], "column_name": args[3]}
        elif ft == "WP" and len(args) >= 2:
            params = {"wp_code": args[0], "cell_ref": args[1]}
        elif ft == "PREV" and len(args) >= 2:
            params = {"formula_type": args[0], "account_code": args[1] if len(args) > 1 else "", "column_name": args[2] if len(args) > 2 else ""}
        else:
            errors.append({"cell": f["cell"], "error": f"参数不足: {f['original']}"})
            continue

        try:
            result = await engine.execute(
                db=db,
                project_id=project_id,
                year=year if ft != "PREV" else year - 1,
                formula_type=ft if ft != "PREV" else params.get("formula_type", "TB"),
                params=params,
            )

            if hasattr(result, 'message'):
                # FormulaError
                errors.append({"cell": f["cell"], "error": result.message})
                continue

            value = float(result) if isinstance(result, Decimal) else result

            # 需求 46.1/46.2：只更新 parsed_data['univer_snapshot'] 的 v 字段，保留 xlsx 中的公式
            # 不再写 cell_obj.value（避免覆盖公式），不再将公式移入 comment
            if structure_data is not None:
                # structure 的 rows 是跨 sheet 扁平化的，需累加前序 sheet 的行数
                row_offset = 0
                sheets_meta = structure_data.get("sheets", [])
                snapshot_sheets = wb.worksheets
                for prior_idx in range(f["sheet_idx"]):
                    if prior_idx < len(snapshot_sheets):
                        row_offset += snapshot_sheets[prior_idx].max_row or 0

                target_row_idx = row_offset + f["row"]
                rows_arr = structure_data.get("rows", [])
                if 0 <= target_row_idx < len(rows_arr):
                    cells_arr = rows_arr[target_row_idx].get("cells", [])
                    # 补齐列（若之前是稀疏写入）
                    while len(cells_arr) <= f["col"]:
                        cells_arr.append({"value": "", "formula": None})
                    # 保留 formula 字段不变，仅更新 value
                    cells_arr[f["col"]]["value"] = value
                    if cells_arr[f["col"]].get("formula") is None:
                        cells_arr[f["col"]]["formula"] = f["original"]
                    rows_arr[target_row_idx]["cells"] = cells_arr
                    filled += 1
                else:
                    errors.append({"cell": f["cell"], "error": f"univer_snapshot 行越界: {target_row_idx}"})
            else:
                # 无 univer_snapshot：跳过本单元格（避免破坏公式），但不算错误
                # 上面已记录 warning 日志
                pass

        except Exception as e:
            errors.append({"cell": f["cell"], "error": str(e)})

    # 保存 prefill 结果到 parsed_data['univer_snapshot']（Req 6 单源化：不再写 structure.json 文件）
    if structure_data is not None and filled > 0:
        try:
            from sqlalchemy.orm.attributes import flag_modified
            # 把 structure 格式的 prefill 结果回写到 univer_snapshot JSONB
            existing_pd = dict(wp.parsed_data) if isinstance(wp.parsed_data, dict) else {}
            snap = existing_pd.get("univer_snapshot")
            if isinstance(snap, dict) and snap.get("sheets"):
                # 把 structure_data 的 value 更新回 slim snapshot 的 cellData
                _sync_prefill_to_snapshot(snap, structure_data, wb.worksheets)
                existing_pd["univer_snapshot"] = snap
                wp.parsed_data = existing_pd
                flag_modified(wp, "parsed_data")
                await db.flush()
        except Exception as e:
            wb.close()
            return {"wp_id": str(wp_id), "status": "error", "message": f"parsed_data 保存失败: {e}"}

    wb.close()

    # ── proposal-remaining-18 task 2.3 / L-3：写入 TB 快照供下次比对 ──
    # 仅在 filled > 0 时写入，避免无公式底稿误覆盖既有快照。
    # 从 formulas_found 收集 TB / AUX 类首参 account_code（WP/PREV/NOTE 不进快照）。
    if filled > 0:
        try:
            from app.models.audit_platform_models import TrialBalance
            snapshot_codes: set[str] = set()
            for f in formulas_found:
                args = _parse_args(f["raw_args"])
                ft = f["formula_type"]
                if ft in ("TB", "AUX") and args:
                    snapshot_codes.add(args[0])
            if snapshot_codes:
                rows = (await db.execute(
                    sa.select(
                        TrialBalance.standard_account_code,
                        TrialBalance.audited_amount,
                    ).where(
                        TrialBalance.project_id == project_id,
                        TrialBalance.year == year,
                        TrialBalance.is_deleted == sa.false(),
                        TrialBalance.standard_account_code.in_(snapshot_codes),
                    )
                )).all()
                snapshot = {
                    row.standard_account_code: float(row.audited_amount)
                    if row.audited_amount is not None else 0.0
                    for row in rows
                }
                wp.prefill_tb_snapshot = snapshot
                from sqlalchemy.orm.attributes import flag_modified
                flag_modified(wp, "prefill_tb_snapshot")
                _logger.info(
                    "prefill_real: wp=%s snapshot recorded for %d accounts",
                    wp_id, len(snapshot),
                )
        except Exception as e:
            # 快照写入失败不阻断 prefill 主流程（仅警告）
            _logger.warning("prefill_real: snapshot write failed wp=%s err=%s", wp_id, e)

    _logger.info("prefill_real: wp=%s found=%d filled=%d errors=%d", wp_id, len(formulas_found), filled, len(errors))

    return {
        "wp_id": str(wp_id),
        "status": "ok",
        "formulas_found": len(formulas_found),
        "formulas_filled": filled,
        "errors": errors[:10],
        "message": f"预填充完成：{filled}/{len(formulas_found)} 个公式已计算",
    }


async def parse_workpaper_real(
    db: AsyncSession,
    project_id: UUID,
    wp_id: UUID,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    真正的解析回写：打开 .xlsx → 提取关键数据 → 写入 parsed_data

    参数：
      dry_run=True  — 仅返回解析预览，不写入 parsed_data（用于两步确认流程的步骤1）
      dry_run=False — 解析并写入 parsed_data（默认，用于步骤2确认后）

    提取内容：
    1. 审定数（审定表中的审定数合计）
    2. 未审数
    3. AJE/RJE 调整金额
    4. 结论文本（搜索"审计结论"/"结论"附近的单元格）
    5. 交叉引用（=WP() 公式）
    """
    try:
        import openpyxl
    except ImportError:
        return {"wp_id": str(wp_id), "status": "error", "message": "openpyxl 未安装"}

    wp = (await db.execute(
        sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
    )).scalar_one_or_none()
    if not wp or not wp.file_path:
        return {"wp_id": str(wp_id), "status": "error", "message": "底稿文件不存在"}

    fp = Path(wp.file_path)
    if not fp.exists():
        return {"wp_id": str(wp_id), "status": "error", "message": f"文件不存在: {fp}"}

    try:
        wb = openpyxl.load_workbook(str(fp), data_only=True, read_only=False)
    except Exception as e:
        return {"wp_id": str(wp_id), "status": "error", "message": f"打开文件失败: {e}"}

    parsed = {
        "audited_amount": None,
        "unadjusted_amount": None,
        "aje_adjustment": 0,
        "rje_adjustment": 0,
        "conclusion": None,
        "conclusion_text": None,
        "audit_explanation": None,
        "cross_refs": [],
        "ai_content": [],
        "extracted_at": None,
    }

    # 关键词搜索
    _AUDITED_KW = ["审定数", "审定金额", "审定余额", "审计后金额", "Audited"]
    _UNADJUSTED_KW = ["未审数", "未审金额", "账面数", "账面金额", "Unadjusted"]
    _AJE_KW = ["AJE", "审计调整", "调整分录"]
    _RJE_KW = ["RJE", "重分类"]
    _CONCLUSION_KW = ["审计结论", "结论", "审计意见", "Conclusion"]
    _EXPLANATION_KW = ["审计说明", "审计结论", "执行情况", "审计程序执行", "审计发现", "Explanation"]
    _WP_REF_RE = re.compile(r'=WP\s*\(\s*["\']([^"\']+)["\']', re.IGNORECASE)

    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=200, max_col=20):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                s = str(val).strip()

                # 检查是否是关键词标签（左侧标签，右侧是数值）
                for kw_list, key in [
                    (_AUDITED_KW, "audited_amount"),
                    (_UNADJUSTED_KW, "unadjusted_amount"),
                    (_AJE_KW, "aje_adjustment"),
                    (_RJE_KW, "rje_adjustment"),
                ]:
                    if any(kw in s for kw in kw_list):
                        # 取右侧单元格的值
                        try:
                            right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if right_cell.value is not None:
                                num = float(right_cell.value)
                                if parsed[key] is None or key in ("aje_adjustment", "rje_adjustment"):
                                    parsed[key] = num
                        except (ValueError, TypeError):
                            pass

                # 结论文本（标题行含关键词，下方或右侧是结论内容）
                for kw in _CONCLUSION_KW:
                    if kw in s:
                        # 情况1：关键词和结论在同一单元格（长文本）
                        if len(s) > len(kw) + 10:
                            parsed["conclusion"] = s
                            parsed["conclusion_text"] = s
                            break
                        # 情况2：关键词是标题，结论在右侧单元格
                        try:
                            right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if right_cell.value and len(str(right_cell.value).strip()) > 5:
                                parsed["conclusion"] = str(right_cell.value).strip()
                                parsed["conclusion_text"] = parsed["conclusion"]
                                break
                        except Exception:
                            pass
                        # 情况3：关键词是标题，结论在下方单元格
                        try:
                            below_cell = ws.cell(row=cell.row + 1, column=cell.column)
                            if below_cell.value and len(str(below_cell.value).strip()) > 5:
                                parsed["conclusion"] = str(below_cell.value).strip()
                                parsed["conclusion_text"] = parsed["conclusion"]
                                break
                        except Exception:
                            pass

                # 审计说明提取
                if parsed["audit_explanation"] is None:
                    for ekw in _EXPLANATION_KW:
                        if ekw in s:
                            parts = []
                            if len(s) > len(ekw) + 20:
                                parts.append(s)
                            for co in range(1, 6):
                                try:
                                    rc = ws.cell(row=cell.row, column=cell.column + co)
                                    if rc.value and str(rc.value).strip():
                                        parts.append(str(rc.value).strip())
                                    else:
                                        break
                                except Exception:
                                    break
                            if not parts:
                                for ro in range(1, 10):
                                    try:
                                        bc = ws.cell(row=cell.row + ro, column=cell.column)
                                        if bc.value and str(bc.value).strip():
                                            parts.append(str(bc.value).strip())
                                        else:
                                            break
                                    except Exception:
                                        break
                            if parts:
                                parsed["audit_explanation"] = "\n".join(parts)
                            break

                # 交叉引用
                if isinstance(val, str):
                    for m in _WP_REF_RE.finditer(val):
                        parsed["cross_refs"].append(m.group(1))

    wb.close()

    # 写入 parsed_data（dry_run=True 时跳过写入，仅返回预览）
    from datetime import datetime, timezone
    parsed["extracted_at"] = datetime.now(timezone.utc).isoformat()
    if not dry_run:
        wp_write = (await db.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )).scalar_one_or_none()
        if wp_write:
            wp_write.parsed_data = parsed
            wp_write.last_parsed_at = datetime.now(timezone.utc)
            await db.flush()

    _logger.info(
        "parse_real: wp=%s audited=%s unadj=%s conclusion=%s refs=%d dry_run=%s",
        wp_id, parsed["audited_amount"], parsed["unadjusted_amount"],
        "yes" if parsed["conclusion"] else "no", len(parsed["cross_refs"]), dry_run,
    )

    return {
        "wp_id": str(wp_id),
        "status": "ok",
        "dry_run": dry_run,
        "audited_amount": parsed["audited_amount"],
        "unadjusted_amount": parsed["unadjusted_amount"],
        "has_conclusion": parsed["conclusion"] is not None,
        "cross_ref_count": len(parsed["cross_refs"]),
        "parsed_data": parsed,
        "message": "解析预览（未写入）" if dry_run else "解析完成",
    }


# ---------------------------------------------------------------------------
# mark_stale — 标记底稿预填数据为过期（从 prefill_service_v2 迁移）
# ---------------------------------------------------------------------------

async def mark_stale(
    db: AsyncSession,
    project_id: UUID,
    account_codes: list[str] | None = None,
) -> int:
    """标记底稿预填数据为过期"""
    q = (
        sa.update(WorkingPaper)
        .where(
            WorkingPaper.project_id == project_id,
            WorkingPaper.is_deleted == False,  # noqa: E712
        )
        .values(prefill_stale=True)
    )
    result = await db.execute(q)
    return result.rowcount


# ---------------------------------------------------------------------------
# scan_formulas — 从文本列表中扫描公式（兼容旧 PrefillService 接口）
# ---------------------------------------------------------------------------

def scan_formulas(cell_texts: list[dict]) -> list[FormulaCell]:
    """Scan a list of cell text entries for formula patterns."""
    results: list[FormulaCell] = []
    for entry in cell_texts:
        text = entry.get("text", "")
        sheet = entry.get("sheet", "Sheet1")
        cell_ref = entry.get("cell_ref", "")
        for match in _FORMULA_RE.finditer(text):
            results.append(FormulaCell(
                sheet=sheet, cell_ref=cell_ref,
                formula_type=match.group(1).upper(),
                raw_args=match.group(2).strip(),
            ))
    return results


# ---------------------------------------------------------------------------
# detect_conflicts — 版本冲突检测（从 prefill_service 迁移）
# ---------------------------------------------------------------------------

async def detect_conflicts(
    db: AsyncSession,
    project_id: UUID,
    wp_id: UUID,
    uploaded_version: int,
) -> dict[str, Any]:
    """Detect conflicts between uploaded file and server version."""
    result = await db.execute(
        sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
    )
    wp = result.scalar_one_or_none()
    if wp is None:
        return {"has_conflict": False, "error": "底稿不存在"}
    current_version = wp.file_version
    has_conflict = uploaded_version < current_version
    return {
        "has_conflict": has_conflict,
        "uploaded_version": uploaded_version,
        "server_version": current_version,
        "conflicts": [] if not has_conflict else [
            {"message": f"版本冲突: 上传版本 {uploaded_version} < 服务器版本 {current_version}"}
        ],
    }


# ---------------------------------------------------------------------------
# 兼容类（供旧测试代码导入，代理到模块级函数）
# ---------------------------------------------------------------------------

class PrefillService:
    """兼容旧接口"""
    def _scan_formulas(self, cell_texts: list[dict]) -> list[FormulaCell]:
        return scan_formulas(cell_texts)

    async def prefill_workpaper(self, db: AsyncSession, project_id: UUID, year: int, wp_id: UUID) -> dict:
        return await prefill_workpaper_real(db, project_id, year, wp_id)

    async def batch_prefill(self, db: AsyncSession, project_id: UUID, year: int, wp_ids: list[UUID]) -> dict:
        return await batch_prefill_ordered(db, project_id, year, wp_ids)

    async def _get_cached_prefill(self, wp_id: UUID) -> dict | None:
        return None

    async def _set_cached_prefill(self, wp_id: UUID, data: dict) -> None:
        pass


# ---------------------------------------------------------------------------
# 批量预填充（按依赖顺序处理多底稿）— Task 3.8
# ---------------------------------------------------------------------------


async def batch_prefill_ordered(
    db: AsyncSession,
    project_id: UUID,
    year: int,
    wp_ids: list[UUID],
) -> dict[str, Any]:
    """按依赖顺序批量预填充多底稿

    流程：
    1. 收集所有底稿的公式
    2. 构建依赖图
    3. 拓扑排序确定执行顺序
    4. 按顺序逐个执行预填充
    """
    from app.services.wp_formula_dependency import (
        build_dependency_graph,
        topological_sort,
        detect_cycles,
    )

    # 获取底稿信息
    result = await db.execute(
        sa.select(WorkingPaper).where(
            WorkingPaper.id.in_(wp_ids),
            WorkingPaper.project_id == project_id,
            WorkingPaper.is_deleted == False,  # noqa: E712
        )
    )
    workpapers = result.scalars().all()
    wp_map: dict[str, WorkingPaper] = {}
    for wp in workpapers:
        wp_code = (wp.parsed_data or {}).get("wp_code", str(wp.id))
        wp_map[wp_code] = wp

    # 收集公式构建依赖图
    all_formulas: list[dict] = []
    for wp_code, wp in wp_map.items():
        if wp.parsed_data and wp.parsed_data.get("cell_provenance"):
            for cell_ref, prov in wp.parsed_data["cell_provenance"].items():
                all_formulas.append({
                    "wp_code": wp_code,
                    "sheet": cell_ref.split("!")[0] if "!" in cell_ref else "",
                    "cell_ref": cell_ref.split("!")[-1] if "!" in cell_ref else cell_ref,
                    "formula_type": prov.get("formula_type", ""),
                    "raw_args": prov.get("raw_args", ""),
                })

    # 构建依赖图
    graph = build_dependency_graph(all_formulas)

    # 检测循环
    cycles = detect_cycles(graph)
    if cycles:
        return {
            "total": len(wp_ids),
            "success_count": 0,
            "error_count": len(wp_ids),
            "errors": {"circular_reference": f"检测到循环引用: {cycles[0]}"},
            "results": {},
        }

    # 拓扑排序
    try:
        order = topological_sort(graph)
    except ValueError as e:
        return {
            "total": len(wp_ids),
            "success_count": 0,
            "error_count": len(wp_ids),
            "errors": {"topological_sort": str(e)},
            "results": {},
        }

    # 按拓扑顺序执行（只处理请求的 wp_ids）
    requested_codes = set(wp_map.keys())
    ordered_codes = [c for c in order if c in requested_codes]
    # 补充不在依赖图中的底稿（无公式依赖的）
    remaining = requested_codes - set(ordered_codes)
    execution_order = list(remaining) + ordered_codes

    success: dict[str, Any] = {}
    errors: dict[str, str] = {}

    for wp_code in execution_order:
        wp = wp_map.get(wp_code)
        if wp is None:
            continue
        try:
            r = await prefill_workpaper_real(db, project_id, year, wp.id)
            success[str(wp.id)] = r
        except Exception as e:
            errors[str(wp.id)] = str(e)

    return {
        "total": len(wp_ids),
        "success_count": len(success),
        "error_count": len(errors),
        "results": success,
        "errors": errors,
        "execution_order": execution_order,
    }


class ParseService:
    """兼容旧接口"""
    async def parse_workpaper(self, db: AsyncSession, project_id: UUID, wp_id: UUID) -> dict:
        return await parse_workpaper_real(db, project_id, wp_id)

    async def detect_conflicts(self, db: AsyncSession, project_id: UUID, wp_id: UUID, uploaded_version: int) -> dict:
        return await detect_conflicts(db, project_id, wp_id, uploaded_version)


# ---------------------------------------------------------------------------
# Round 4 需求 7: 预填充 provenance 回写
# ---------------------------------------------------------------------------

PREFILL_SERVICE_VERSION = "prefill_v1.2"

# 公式类型 → provenance source 映射
_FORMULA_TYPE_TO_SOURCE: dict[str, str] = {
    "TB": "trial_balance",
    "SUM_TB": "trial_balance",
    "AUX": "aux_balance",
    "PREV": "prior_year",
    "WP": "workpaper_ref",
    "LEDGER": "ledger",
    "ADJ": "adjustment",
    "NOTE": "disclosure_note",
}


def _build_source_ref(
    formula_type: str,
    raw_args: str,
    params: dict[str, str],
) -> str | None:
    """根据公式类型和参数构建 source_ref 字符串

    Returns:
        source_ref 字符串，或 None（参数不足/未知类型时）
    """
    ft = formula_type.upper()

    if ft in ("TB", "SUM_TB"):
        account = params.get("account_code") or params.get("account_range", "")
        column = params.get("column_name", "")
        if not account:
            return None
        return f"{account}:{column}"

    elif ft == "AUX":
        account = params.get("account_code", "")
        dimension = params.get("aux_dimension", "")
        value = params.get("dimension_value", "")
        column = params.get("column_name", "")
        if not account:
            return None
        return f"{account}:{dimension}:{value}:{column}"

    elif ft == "PREV":
        # =PREV('wp_code', 'sheet', 'cell')
        wp_code = params.get("wp_code", "")
        sheet = params.get("sheet", "")
        cell = params.get("cell_ref", "")
        if wp_code:
            return f"prev:{wp_code}!{sheet}!{cell}"
        # 兼容旧格式
        inner_type = params.get("formula_type", "")
        account = params.get("account_code", "")
        if not inner_type:
            return None
        return f"{inner_type}:{account}"

    elif ft == "WP":
        wp_code = params.get("wp_code", "")
        sheet = params.get("sheet", "")
        cell_ref = params.get("cell_ref", "")
        if not wp_code:
            return None
        return f"{wp_code}!{sheet}!{cell_ref}"

    elif ft == "LEDGER":
        account = params.get("account_code", "")
        direction = params.get("direction", "")
        period = params.get("period", "")
        if not account:
            return None
        return f"ledger:{account}:{direction}:{period}"

    elif ft == "ADJ":
        account = params.get("account_code", "")
        adj_type = params.get("adj_type", "")
        if not account:
            return None
        return f"adj:{account}:{adj_type}"

    elif ft == "NOTE":
        section = params.get("section", "")
        row = params.get("row", "")
        col = params.get("col", "")
        if not section:
            return None
        return f"note:{section}:{row}:{col}"

    return None


def _write_cell_provenance(wp, provenance: dict[str, dict]) -> None:
    """将 provenance 数据写入 wp.parsed_data.cell_provenance

    Supersede 策略：
    - 重填时覆盖旧值
    - 保留最多 1 次历史（_prev 字段）
    - 相同 filled_at 时不创建 _prev（幂等重入）
    """
    if wp.parsed_data is None:
        wp.parsed_data = {}

    if "cell_provenance" not in wp.parsed_data:
        wp.parsed_data["cell_provenance"] = {}

    cp = wp.parsed_data["cell_provenance"]

    for cell_ref, new_entry in provenance.items():
        existing = cp.get(cell_ref)

        if existing is None:
            # 首次写入
            cp[cell_ref] = new_entry
        else:
            # Supersede: 检查是否幂等重入（相同 filled_at）
            if existing.get("filled_at") == new_entry.get("filled_at"):
                # 幂等，不创建 _prev
                cp[cell_ref] = new_entry
            else:
                # 保留上一次为 _prev（丢弃更早的历史）
                prev_snapshot = {
                    k: v for k, v in existing.items() if k != "_prev"
                }
                new_entry["_prev"] = prev_snapshot
                cp[cell_ref] = new_entry


# ---------------------------------------------------------------------------
# E1 spec Sprint 1 Task 1.17/1.18: 表头自动填充
# ---------------------------------------------------------------------------


async def fill_header_cells(
    db: AsyncSession,
    wp_id: UUID,
    project: Any | None = None,
) -> dict[str, Any]:
    """按 wp_template_metadata.header_cells 配置填充底稿 R3/R4 表头

    锚定 requirements F4.2:
    - R3 左:被审计单位名称(Project.company_name)
    - R3 中:审计期间(Project.year)
    - R3 右:索引号(wp_code + sheet 后缀)
    - R4 左:编制人/日期(从 WorkingPaper.created_by 取 + first_open_at)
    - R4 中:复核人/日期(L1 复核通过后回填)
    - R4 右:页次(自动计算)

    支持 sheet_overrides:不同 sheet 表头位置可能不同(E1A vs E1-2 等)。
    严格遵守 Task 1.1: 不覆盖已有公式 cell(如 A3 含 =底稿目录!A2 引用)。

    Returns:
        {wp_id, status, cells_filled, sheets_processed, errors}
    """
    try:
        import openpyxl
    except ImportError:
        return {"wp_id": str(wp_id), "status": "error", "message": "openpyxl 未安装"}

    from datetime import datetime, timezone

    # 1. 加载底稿 + wp_index + 元数据
    from app.models.audit_platform_models import WpIndex
    from app.models.wp_optimization_models import WpTemplateMetadata
    from app.models.core import Project

    wp_q = sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
    wp = (await db.execute(wp_q)).scalar_one_or_none()
    if not wp or not wp.file_path:
        return {"wp_id": str(wp_id), "status": "skip", "message": "底稿无文件路径"}

    fp = Path(wp.file_path)
    if not fp.exists():
        return {"wp_id": str(wp_id), "status": "skip", "message": f"文件不存在: {fp}"}

    # 仅处理 xlsx/xlsm
    if fp.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"wp_id": str(wp_id), "status": "skip", "message": f"非 xlsx 类型: {fp.suffix}"}

    # 2. 取 wp_code(从 wp_index)
    wp_code = None
    if wp.wp_index_id:
        idx_q = sa.select(WpIndex.wp_code, WpIndex.wp_name).where(WpIndex.id == wp.wp_index_id)
        idx_row = (await db.execute(idx_q)).first()
        if idx_row:
            wp_code = idx_row[0]

    if not wp_code:
        return {"wp_id": str(wp_id), "status": "skip", "message": "wp_code 未知"}

    # 3. 取项目元数据(Project)
    company_name = ""
    audit_year = ""
    if project is not None:
        company_name = getattr(project, "company_name", "") or getattr(project, "name", "") or ""
        audit_year = str(getattr(project, "year", "") or "")
    else:
        proj_q = sa.select(Project.company_name, Project.year).where(Project.id == wp.project_id)
        proj_row = (await db.execute(proj_q)).first()
        if proj_row:
            company_name = proj_row[0] or ""
            audit_year = str(proj_row[1] or "")

    # 4. 取 wp_template_metadata.header_cells
    meta_q = sa.select(WpTemplateMetadata).where(
        WpTemplateMetadata.wp_code == wp_code
    ).limit(1)
    meta = (await db.execute(meta_q)).scalar_one_or_none()
    if meta is None or not meta.header_cells:
        return {
            "wp_id": str(wp_id),
            "status": "skip",
            "message": f"wp_code={wp_code} 无 header_cells 配置",
        }

    header_cfg = meta.header_cells or {}
    default_cfg = header_cfg.get("default", {})
    sheet_overrides = header_cfg.get("sheet_overrides", {})

    # 5. 取编制人/复核人(从 WorkingPaper / 用户表)
    preparer_str = ""
    reviewer_str = ""
    try:
        if wp.created_by:
            from app.models.core import User
            usr = (await db.execute(
                sa.select(User.full_name, User.username).where(User.id == wp.created_by)
            )).first()
            if usr:
                preparer_name = usr[0] or usr[1] or ""
                preparer_date = (wp.created_at or datetime.now(timezone.utc)).strftime("%Y-%m-%d")
                preparer_str = f"{preparer_name} {preparer_date}"
    except Exception as e:
        _logger.debug("fill_header_cells get preparer fallback: %s", e)

    # 复核人 / 页次 — 当前阶段留空,L1 复核通过后由 review 服务回填
    reviewer_str = ""
    page_no_str = ""

    # 6. 打开 xlsx,逐 sheet 写入
    try:
        wb = openpyxl.load_workbook(str(fp), data_only=False)
    except Exception as e:
        return {"wp_id": str(wp_id), "status": "error", "message": f"打开文件失败: {e}"}

    cells_filled = 0
    sheets_processed = 0
    errors: list[str] = []

    def _is_formula(cell) -> bool:
        v = cell.value
        return isinstance(v, str) and v.lstrip().startswith("=")

    def _apply_sheet(ws, cfg: dict) -> int:
        """对一个 sheet 应用一份 header_cells 配置;返回 filled 数量"""
        n = 0
        # 字段映射 — 配置 key → 值的来源
        index_no = f"{wp_code}-{ws.title.split(wp_code)[-1] if wp_code in ws.title else ws.title}"
        value_map = {
            "company_name": company_name,
            "audit_period": audit_year,
            "audit_period_subtitle": audit_year,
            "index_no": index_no,
            "preparer": preparer_str,
            "reviewer": reviewer_str,
            "page_no": page_no_str,
            "preparer_name_with_date": preparer_str,
            "reviewer_name_with_date": reviewer_str,
        }
        for key, item in cfg.items():
            if not isinstance(item, dict):
                continue
            cell_ref = item.get("cell")
            if not cell_ref:
                continue

            # 优先级: source 引用(底稿目录!A2) → field 字段 → key 直接查表
            new_value: str | None = None
            src = item.get("source")
            field_key = item.get("field")
            if src:
                # 引用其他 sheet 的 cell - 让 Excel 公式自然生效,我们不覆盖
                # 检查目标 cell 是否已是公式,是则跳过(模板自带 =底稿目录!A2)
                try:
                    target_cell = ws[cell_ref]
                except Exception:
                    continue
                if _is_formula(target_cell):
                    continue
                # 如果不是公式,则写入 source 引用作为 Excel 公式
                new_value = f"={src}"
            elif field_key:
                new_value = str(value_map.get(field_key, "") or "")
            else:
                # 直接用 key(向后兼容)
                new_value = str(value_map.get(key, "") or "")

            if not new_value:
                continue

            try:
                target_cell = ws[cell_ref]
            except Exception:
                continue

            # Task 1.1: 不覆盖已有公式
            if _is_formula(target_cell):
                continue
            # 已有非空值且非默认占位时跳过(避免覆盖用户填写)
            if target_cell.value and not isinstance(target_cell.value, (int, float)):
                existing = str(target_cell.value).strip()
                # 占位符(空白/单一日期占位)允许覆盖
                if existing and existing != "  " and existing != "—":
                    continue

            target_cell.value = new_value
            n += 1
        return n

    for ws in wb.worksheets:
        cfg = sheet_overrides.get(ws.title) or default_cfg
        if not cfg:
            continue
        try:
            sheets_processed += 1
            cells_filled += _apply_sheet(ws, cfg)
        except Exception as e:
            errors.append(f"{ws.title}: {e}")

    try:
        wb.save(str(fp))
    except Exception as e:
        wb.close()
        return {"wp_id": str(wp_id), "status": "error", "message": f"保存失败: {e}"}

    wb.close()

    _logger.info(
        "fill_header_cells: wp=%s code=%s sheets=%d filled=%d errors=%d",
        wp_id, wp_code, sheets_processed, cells_filled, len(errors),
    )
    return {
        "wp_id": str(wp_id),
        "wp_code": wp_code,
        "status": "ok",
        "cells_filled": cells_filled,
        "sheets_processed": sheets_processed,
        "errors": errors[:5],
    }
