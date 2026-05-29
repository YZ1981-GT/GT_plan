"""Sprint 4 Task 15 — 底稿离线导出服务 (US-14).

主要 API:
- export_workpaper_template(sheets_data, ...) → (xlsx_bytes, sha256_hash)

功能：
- 15.1: 导出填写模板 xlsx
- 15.2: 注意事项 sheet 生成（7 节内容 + 项目元数据填充）
- 15.3: cell 颜色标记（黄=可填 / 灰=公式 / 红=禁改 / 绿=必填）+ sheet 保护
- 15.4: _meta_ sheet 生成（hidden + base64+gzip + sha256 checksum）
- 15.5: 可选 AES 加密（Fernet）

纯函数设计，DB 操作通过外部传入数据。
"""
from __future__ import annotations

import base64
import gzip
import hashlib
import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["WpOfflineExportService", "export_workpaper_template"]

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# 4-color semantic fills (15.3)
FILL_EDITABLE = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # 黄=可填
FILL_FORMULA = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")  # 灰=公式
FILL_LOCKED = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")  # 红=禁改
FILL_REQUIRED = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")  # 绿=必填

FONT_HEADER = Font(bold=True, size=12)
FONT_TITLE = Font(bold=True, size=14)
FONT_NORMAL = Font(size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MAX_SHEET_NAME = 31
FORMAT_VERSION = "1.0"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _truncate_sheet_name(name: str) -> str:
    """Truncate sheet name to Excel 31-char limit, replace invalid chars."""
    for ch in r"[]:*?/\\":
        name = name.replace(ch, "_")
    if len(name) > MAX_SHEET_NAME:
        name = name[: MAX_SHEET_NAME - 2] + ".."
    return name


def _classify_cell(cell_meta: dict[str, Any]) -> str:
    """Classify cell type for coloring.

    Returns: 'editable' | 'formula' | 'locked' | 'required'
    """
    source = cell_meta.get("source", "manual")
    is_required = cell_meta.get("is_required", False)
    mode = cell_meta.get("mode", "manual")

    if mode == "formula" or source == "formula":
        return "formula"
    if source in ("wp_data", "trial_balance", "cross_ref", "auto_fill"):
        return "locked"
    if is_required:
        return "required"
    return "editable"


def _get_fill_for_type(cell_type: str) -> PatternFill:
    """Get fill color for cell type."""
    return {
        "editable": FILL_EDITABLE,
        "formula": FILL_FORMULA,
        "locked": FILL_LOCKED,
        "required": FILL_REQUIRED,
    }.get(cell_type, FILL_EDITABLE)


def _compute_sha256(data: bytes) -> str:
    """Compute SHA-256 hex digest."""
    return hashlib.sha256(data).hexdigest()


def _compress_meta(meta_dict: dict[str, Any]) -> str:
    """Compress metadata dict to base64+gzip string (15.4)."""
    json_bytes = json.dumps(meta_dict, ensure_ascii=False, default=str).encode("utf-8")
    compressed = gzip.compress(json_bytes)
    return base64.b64encode(compressed).decode("ascii")


def _decompress_meta(encoded: str) -> dict[str, Any]:
    """Decompress base64+gzip string back to dict."""
    compressed = base64.b64decode(encoded.encode("ascii"))
    json_bytes = gzip.decompress(compressed)
    return json.loads(json_bytes.decode("utf-8"))


def _encrypt_bytes(data: bytes, password: str) -> bytes:
    """AES encrypt bytes using Fernet (15.5)."""
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

    salt = b"wp_offline_export_salt_v1"
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100_000)
    key = base64.urlsafe_b64encode(kdf.derive(password.encode("utf-8")))
    f = Fernet(key)
    return f.encrypt(data)


def _decrypt_bytes(data: bytes, password: str) -> bytes:
    """AES decrypt bytes using Fernet."""
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

    salt = b"wp_offline_export_salt_v1"
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100_000)
    key = base64.urlsafe_b64encode(kdf.derive(password.encode("utf-8")))
    f = Fernet(key)
    return f.decrypt(data)


# ---------------------------------------------------------------------------
# 15.2: 注意事项 Sheet 模板（7 节内容 + 项目元数据）
# ---------------------------------------------------------------------------


def _build_instructions_sheet(ws: Worksheet, context: dict[str, Any]) -> None:
    """Fill the 注意事项 sheet with 7-section usage guide (15.2)."""
    project_name = context.get("project_name", "{项目名称}")
    year = context.get("year", "{年度}")
    exporter_name = context.get("exporter_name", "{导出人}")
    export_time = context.get("export_time", _now_str())
    sheet_count = context.get("sheet_count", 0)
    wp_code = context.get("wp_code", "{底稿编码}")
    deadline = context.get("deadline", "")
    contact_name = context.get("contact_name", "")
    contact_email = context.get("contact_email", "")
    contact_phone = context.get("contact_phone", "")

    ws.column_dimensions["A"].width = 80

    lines = [
        ("═" * 50, FONT_TITLE),
        ("  底稿离线填写模板  —  使用说明", FONT_TITLE),
        ("═" * 50, FONT_TITLE),
        ("", None),
        ("【1. 文件用途】", FONT_HEADER),
        (f"  本文件为「{project_name} - {year} 年度」底稿 {wp_code} 离线填写模板，", FONT_NORMAL),
        (f"  共 {sheet_count} 个工作表，由 {exporter_name} 于 {export_time} 生成。", FONT_NORMAL),
        ("", None),
        ("【2. 单元格颜色含义】", FONT_HEADER),
        ("  ⬛ 黄底 = 您可填写的数据（如金额 / 客户名称 / 描述）", FONT_NORMAL),
        ("  ⬛ 灰底 = 系统自动计算（公式驱动，请勿修改）", FONT_NORMAL),
        ("  ⬛ 红底 = 锁定单元格（如来自试算表 / 交叉引用，禁止改动）", FONT_NORMAL),
        ("  ⬛ 绿底 = 必填项（导入时若为空将告警）", FONT_NORMAL),
        ("", None),
        ("【3. 填写规则】", FONT_HEADER),
        ("  - 仅修改黄底和绿底单元格", FONT_NORMAL),
        ("  - 数字请输入纯数字（不要加千分位逗号）", FONT_NORMAL),
        ("  - 日期格式：YYYY-MM-DD", FONT_NORMAL),
        ("  - 百分比请输入小数（如 0.15 表示 15%）", FONT_NORMAL),
        ("", None),
        ("【4. 公式区域】", FONT_HEADER),
        ("  灰底单元格的公式见单元格批注（鼠标悬停查看）。", FONT_NORMAL),
        ("  如需修改公式逻辑，请在对应黄底原始数据修改后由系统重算。", FONT_NORMAL),
        ("", None),
        ("【5. 文件名 / sheet 名规则】", FONT_HEADER),
        ("  ⚠ 不要重命名文件 / 不要修改 sheet 名（含隐藏的 sheet ID 用于回传匹配）", FONT_NORMAL),
        ("  ⚠ 不要删除 _meta_ 隐藏 sheet（删了无法导回）", FONT_NORMAL),
        ("", None),
        ("【6. 完成后回传】", FONT_HEADER),
        (f"  填写完成后，将本文件发回联系人（见下方），", FONT_NORMAL),
        ("  在系统内点「底稿 → 📥 导入填写结果」即可合并您的填报内容。", FONT_NORMAL),
        (f"  截止日期：{deadline or '无限制'}", FONT_NORMAL),
        ("", None),
        ("【7. 注意事项】", FONT_HEADER),
        ("  - 不要修改 sheet 顺序（导入按 sheet 名匹配，顺序无关，但避免误删）", FONT_NORMAL),
        ("  - 不要新增 sheet（新 sheet 不会被导入）", FONT_NORMAL),
        ("  - 多人协作时合并冲突由项目负责人在系统内决策", FONT_NORMAL),
        ("  - 如遇问题请联系下方联系人", FONT_NORMAL),
        ("", None),
        ("【联系人】", FONT_HEADER),
        (f"  姓名: {contact_name}  邮箱: {contact_email}  电话: {contact_phone}", FONT_NORMAL),
        ("═" * 50, FONT_TITLE),
    ]

    for row_idx, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = ALIGN_LEFT


# ---------------------------------------------------------------------------
# 15.3: Content Sheet Builder (4-color + protection)
# ---------------------------------------------------------------------------


def _build_content_sheet(
    ws: Worksheet,
    sheet_data: dict[str, Any],
) -> None:
    """Build a single content sheet with data + 4-color marking + protection (15.3)."""
    sheet_name = sheet_data.get("sheet_name", "")
    headers = sheet_data.get("headers", [])
    rows = sheet_data.get("rows", [])
    cell_meta = sheet_data.get("cell_meta", {})
    formulas = sheet_data.get("formulas", {})

    # Row 1: sheet identifier (hidden for import matching)
    ws.cell(row=1, column=1, value=f"sheet_id:{sheet_name}")
    ws.row_dimensions[1].hidden = True

    # Row 2: headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(header)) * 2)

    # Data rows (starting row 3)
    for row_idx, row_data in enumerate(rows, start=3):
        cells = row_data.get("cells", [])
        for col_idx, cell_val in enumerate(cells, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.border = BORDER_THIN

            # Cell key for metadata lookup
            cell_key = f"{row_idx - 3}:{col_idx - 1}"

            # Determine cell type and apply fill (15.3)
            meta = cell_meta.get(cell_key, {})
            cell_type = _classify_cell(meta)
            cell.fill = _get_fill_for_type(cell_type)

            # Protection: lock formula/locked cells
            if cell_type in ("formula", "locked"):
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

            # Comments for formulas
            if cell_key in formulas:
                formula_info = formulas[cell_key]
                expr = formula_info if isinstance(formula_info, str) else formula_info.get("expression", "")
                if expr:
                    cell.comment = Comment(f"公式: {expr}", "系统")

    # Enable sheet protection (allow editing unlocked cells)
    ws.protection.sheet = True
    ws.protection.enable()


# ---------------------------------------------------------------------------
# 15.4: _meta_ Sheet
# ---------------------------------------------------------------------------


def _build_meta_sheet(ws: Worksheet, sheets_data: list[dict[str, Any]], context: dict[str, Any]) -> None:
    """Build hidden _meta_ sheet with compressed metadata (15.4)."""
    meta_payload: dict[str, Any] = {
        "wp_id": context.get("wp_id", ""),
        "wp_code": context.get("wp_code", ""),
        "project_id": context.get("project_id", ""),
        "year": context.get("year", ""),
        "export_time": context.get("export_time", _now_str()),
        "sheets": {},
    }

    for sheet in sheets_data:
        sheet_name = sheet.get("sheet_name", "")
        meta_payload["sheets"][sheet_name] = {
            "cell_meta": sheet.get("cell_meta", {}),
            "formulas": sheet.get("formulas", {}),
            "row_meta": sheet.get("row_meta", []),
            "headers": sheet.get("headers", []),
            "row_count": len(sheet.get("rows", [])),
        }

    # Compress and store
    compressed = _compress_meta(meta_payload)

    # A1: meta_data
    ws.cell(row=1, column=1, value="meta_data")
    ws.cell(row=1, column=2, value=compressed)

    # A2: sheet_names list
    sheet_names = [s.get("sheet_name", "") for s in sheets_data]
    ws.cell(row=2, column=1, value="sheet_names")
    ws.cell(row=2, column=2, value=json.dumps(sheet_names, ensure_ascii=False))

    # A3: binding hash (sha256 of compressed meta)
    binding_hash = _compute_sha256(compressed.encode("utf-8"))
    ws.cell(row=3, column=1, value="binding_hash")
    ws.cell(row=3, column=2, value=binding_hash)

    # A4: export timestamp
    ws.cell(row=4, column=1, value="export_time")
    ws.cell(row=4, column=2, value=context.get("export_time", _now_str()))

    # A5: format version
    ws.cell(row=5, column=1, value="format_version")
    ws.cell(row=5, column=2, value=FORMAT_VERSION)

    # A6: wp_id
    ws.cell(row=6, column=1, value="wp_id")
    ws.cell(row=6, column=2, value=str(context.get("wp_id", "")))


# ---------------------------------------------------------------------------
# Core Export Function (15.1 orchestration)
# ---------------------------------------------------------------------------


def export_workpaper_template(
    sheets_data: list[dict[str, Any]],
    *,
    wp_id: str = "",
    wp_code: str = "",
    project_id: str = "",
    year: int | str = "",
    project_name: str = "",
    exporter_name: str = "",
    password: str | None = None,
    deadline: str = "",
    contact_name: str = "",
    contact_email: str = "",
    contact_phone: str = "",
) -> tuple[bytes, str]:
    """Export workpaper sheets to xlsx template for offline filling.

    Args:
        sheets_data: List of sheet dicts, each with:
            - sheet_name: str
            - headers: list[str]
            - rows: list[dict] with 'cells' key
            - cell_meta: dict[str, dict] keyed by "row:col"
            - formulas: dict[str, Any] keyed by "row:col"
        wp_id: Workpaper UUID.
        wp_code: Workpaper code (e.g. A1-11).
        project_id: Project UUID.
        year: Fiscal year.
        project_name: Project display name.
        exporter_name: Name of the person exporting.
        password: Optional AES encryption password.
        deadline: Deadline date string.
        contact_name: Contact person name.
        contact_email: Contact email.
        contact_phone: Contact phone.

    Returns:
        (xlsx_bytes, sha256_hash) — if password provided, bytes are encrypted.
    """
    export_time = _now_str()

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # 15.2: 注意事项 sheet (first)
    ws_instructions = wb.create_sheet("注意事项")
    context = {
        "project_name": project_name,
        "year": year,
        "exporter_name": exporter_name,
        "export_time": export_time,
        "sheet_count": len(sheets_data),
        "wp_code": wp_code,
        "deadline": deadline,
        "contact_name": contact_name,
        "contact_email": contact_email,
        "contact_phone": contact_phone,
    }
    _build_instructions_sheet(ws_instructions, context)

    # 15.3: Content sheets
    for sheet in sheets_data:
        sheet_name = sheet.get("sheet_name", "Sheet")
        display_name = _truncate_sheet_name(sheet_name)

        # Ensure unique sheet name
        existing_names = [ws.title for ws in wb.worksheets]
        if display_name in existing_names:
            display_name = _truncate_sheet_name(f"{sheet_name[:25]}_{len(existing_names)}")

        ws_content = wb.create_sheet(display_name)
        _build_content_sheet(ws_content, sheet)

    # 15.4: _meta_ sheet (last, hidden)
    ws_meta = wb.create_sheet("_meta_")
    meta_context = {
        "wp_id": wp_id,
        "wp_code": wp_code,
        "project_id": project_id,
        "year": year,
        "export_time": export_time,
    }
    _build_meta_sheet(ws_meta, sheets_data, meta_context)
    ws_meta.sheet_state = "hidden"

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()

    # Compute hash
    file_hash = _compute_sha256(xlsx_bytes)

    # 15.5: Optional AES encryption
    if password:
        xlsx_bytes = _encrypt_bytes(xlsx_bytes, password)

    return xlsx_bytes, file_hash


# ---------------------------------------------------------------------------
# Async Service Class
# ---------------------------------------------------------------------------


class WpOfflineExportService:
    """底稿离线导出服务 (Sprint 4 Task 15).

    Async wrapper that loads workpaper data from DB then delegates to pure export function.
    """

    def __init__(self, db: Any = None):
        self.db = db

    async def export_template(
        self,
        wp_id: UUID,
        sheet_names: list[str] | None = None,
        password: str | None = None,
        exporter_name: str = "",
        deadline: str = "",
        contact_name: str = "",
        contact_email: str = "",
        contact_phone: str = "",
    ) -> tuple[bytes, str]:
        """Export workpaper template from DB to xlsx.

        Returns:
            (xlsx_bytes, sha256_hash)
        """
        wp_data = await self._load_workpaper(wp_id)
        if not wp_data:
            return b"", ""

        sheets_data = self._extract_sheets(wp_data, sheet_names)
        project_name = await self._get_project_name(wp_data.get("project_id"))

        return export_workpaper_template(
            sheets_data,
            wp_id=str(wp_id),
            wp_code=wp_data.get("wp_code", ""),
            project_id=str(wp_data.get("project_id", "")),
            year=wp_data.get("year", ""),
            project_name=project_name,
            exporter_name=exporter_name,
            password=password,
            deadline=deadline,
            contact_name=contact_name,
            contact_email=contact_email,
            contact_phone=contact_phone,
        )

    async def _load_workpaper(self, wp_id: UUID) -> dict[str, Any]:
        """Load workpaper data from DB."""
        if self.db is None:
            return {}
        try:
            from sqlalchemy import select as sa_select
            from app.models.models import ProjectWorkpaper

            result = await self.db.execute(
                sa_select(ProjectWorkpaper).where(ProjectWorkpaper.id == wp_id)
            )
            wp = result.scalar_one_or_none()
            if not wp:
                return {}
            return {
                "wp_id": str(wp.id),
                "wp_code": wp.wp_code or "",
                "project_id": wp.project_id,
                "year": getattr(wp, "year", ""),
                "parsed_data": wp.parsed_data or {},
            }
        except Exception:
            return {}

    def _extract_sheets(
        self, wp_data: dict[str, Any], sheet_names: list[str] | None = None
    ) -> list[dict[str, Any]]:
        """Extract sheet data from workpaper parsed_data."""
        parsed = wp_data.get("parsed_data", {})
        sheets = parsed.get("sheets", {})

        result = []
        for name, sheet_content in sheets.items():
            if sheet_names and name not in sheet_names:
                continue
            result.append({
                "sheet_name": name,
                "headers": sheet_content.get("headers", []),
                "rows": sheet_content.get("rows", []),
                "cell_meta": sheet_content.get("cell_meta", {}),
                "formulas": sheet_content.get("formulas", {}),
                "row_meta": sheet_content.get("row_meta", []),
            })

        return result

    async def _get_project_name(self, project_id: Any) -> str:
        """Get project display name."""
        if self.db is None or project_id is None:
            return ""
        try:
            from sqlalchemy import select as sa_select
            from app.models.models import Project

            result = await self.db.execute(
                sa_select(Project.name).where(Project.id == project_id)
            )
            return result.scalar_one_or_none() or ""
        except Exception:
            return ""
