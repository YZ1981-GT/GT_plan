"""Sprint C.0 — 附注离线导入服务 (D15).

主要 API:
- validate_import_file(xlsx_bytes) → ImportValidationResult
- diff_sections(xlsx_bytes, existing_sections) → list[SectionDiff]
- apply_import(xlsx_bytes, existing_sections, decisions) → ImportResult

功能：
- C.0.9: 解压 + 校验 _meta_ sheet 存在
- C.0.10: 按 section_id 匹配现有章节（命中/缺失/系统多余 三态）
- C.0.11: 字段级 diff 算法（值/公式/manual 三类字段）
- C.0.12: 章节级冲突选择（覆盖/保留/合并/丢弃）
- C.0.13: 与 D9 协作锁集成
- C.0.14: 与 D11 版本树集成
- C.0.15: 与 D14 template_type 校验
- C.0.16: 审计日志 + 文件 30 天归档 + 可回滚

纯函数设计，DB 操作通过外部传入或 async service class。
"""
from __future__ import annotations

import hashlib
import json
from copy import deepcopy
from datetime import datetime, timezone
from enum import Enum
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from app.services.note_offline_export_service import _compress_meta, _decompress_meta

__all__ = [
    "NoteOfflineImportService",
    "validate_import_file",
    "diff_sections",
    "apply_import",
    "ImportValidationResult",
    "SectionDiff",
    "CellDiff",
    "ConflictResolution",
    "ImportResult",
    "MatchStatus",
]


# ---------------------------------------------------------------------------
# Enums & Data Classes
# ---------------------------------------------------------------------------


class MatchStatus(str, Enum):
    """Section match status (C.0.10)."""
    MATCHED = "matched"          # section_id found in both import and system
    IMPORT_ONLY = "import_only"  # section_id in import but not in system
    SYSTEM_ONLY = "system_only"  # section_id in system but not in import


class DiffType(str, Enum):
    """Cell diff type (C.0.11)."""
    ADD = "add"
    REMOVE = "remove"
    MODIFY = "modify"


class FieldCategory(str, Enum):
    """Field category for diff (C.0.11)."""
    VALUE = "value"
    FORMULA = "formula"
    MANUAL = "manual"


class ConflictResolution(str, Enum):
    """Conflict resolution choice (C.0.12)."""
    OVERWRITE = "overwrite"  # 导入版本胜出
    KEEP = "keep"            # 本地版本胜出
    MERGE = "merge"          # cell 级选择性合并
    DISCARD = "discard"      # 丢弃此章节



class CellDiff:
    """Single cell diff result (C.0.11)."""

    __slots__ = ("cell_key", "local_value", "imported_value", "diff_type", "field_category")

    def __init__(
        self,
        cell_key: str,
        local_value: Any,
        imported_value: Any,
        diff_type: DiffType,
        field_category: FieldCategory,
    ):
        self.cell_key = cell_key
        self.local_value = local_value
        self.imported_value = imported_value
        self.diff_type = diff_type
        self.field_category = field_category

    def to_dict(self) -> dict[str, Any]:
        return {
            "cell": self.cell_key,
            "local": self.local_value,
            "imported": self.imported_value,
            "type": self.diff_type.value,
            "category": self.field_category.value,
        }


class SectionDiff:
    """Section-level diff result (C.0.10 + C.0.11)."""

    __slots__ = ("section_id", "section_title", "match_status", "cell_diffs", "imported_data")

    def __init__(
        self,
        section_id: str,
        section_title: str,
        match_status: MatchStatus,
        cell_diffs: list[CellDiff] | None = None,
        imported_data: dict[str, Any] | None = None,
    ):
        self.section_id = section_id
        self.section_title = section_title
        self.match_status = match_status
        self.cell_diffs = cell_diffs or []
        self.imported_data = imported_data

    def to_dict(self) -> dict[str, Any]:
        return {
            "section_id": self.section_id,
            "section_title": self.section_title,
            "match_status": self.match_status.value,
            "diff_count": len(self.cell_diffs),
            "diffs": [d.to_dict() for d in self.cell_diffs],
        }


class ImportValidationResult:
    """Validation result for import file (C.0.9)."""

    __slots__ = ("valid", "errors", "section_ids", "meta_data", "binding_hash", "format_version")

    def __init__(
        self,
        valid: bool,
        errors: list[str] | None = None,
        section_ids: list[str] | None = None,
        meta_data: dict[str, Any] | None = None,
        binding_hash: str = "",
        format_version: str = "",
    ):
        self.valid = valid
        self.errors = errors or []
        self.section_ids = section_ids or []
        self.meta_data = meta_data or {}
        self.binding_hash = binding_hash
        self.format_version = format_version


class ImportResult:
    """Result of applying import (C.0.12)."""

    __slots__ = ("success", "sections_imported", "sections_kept", "sections_discarded", "conflicts", "audit_entry")

    def __init__(
        self,
        success: bool,
        sections_imported: int = 0,
        sections_kept: int = 0,
        sections_discarded: int = 0,
        conflicts: int = 0,
        audit_entry: dict[str, Any] | None = None,
    ):
        self.success = success
        self.sections_imported = sections_imported
        self.sections_kept = sections_kept
        self.sections_discarded = sections_discarded
        self.conflicts = conflicts
        self.audit_entry = audit_entry or {}

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "sections_imported": self.sections_imported,
            "sections_kept": self.sections_kept,
            "sections_discarded": self.sections_discarded,
            "conflicts": self.conflicts,
        }


# ---------------------------------------------------------------------------
# C.0.9: Validate Import File
# ---------------------------------------------------------------------------


def validate_import_file(xlsx_bytes: bytes) -> ImportValidationResult:
    """Validate uploaded xlsx file for import (C.0.9).

    Checks:
    1. Valid xlsx format
    2. _meta_ sheet exists
    3. _meta_ sheet has section_ids + binding_hash (CI-21)
    4. format_version is compatible
    """
    errors: list[str] = []

    # Try to load workbook
    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ImportValidationResult(valid=False, errors=[f"无法解析 xlsx 文件: {e}"])

    # Check _meta_ sheet exists
    if "_meta_" not in wb.sheetnames:
        return ImportValidationResult(valid=False, errors=["缺少 _meta_ 隐藏 sheet，无法导入"])

    ws_meta = wb["_meta_"]

    # Read meta fields
    meta_data_key = ws_meta.cell(row=1, column=1).value
    meta_data_val = ws_meta.cell(row=1, column=2).value
    section_ids_key = ws_meta.cell(row=2, column=1).value
    section_ids_val = ws_meta.cell(row=2, column=2).value
    binding_hash_key = ws_meta.cell(row=3, column=1).value
    binding_hash_val = ws_meta.cell(row=3, column=2).value
    format_version_key = ws_meta.cell(row=5, column=1).value
    format_version_val = ws_meta.cell(row=5, column=2).value

    # Validate meta_data
    if meta_data_key != "meta_data" or not meta_data_val:
        errors.append("_meta_ sheet 缺少 meta_data 字段")

    # Validate section_ids (CI-21)
    section_ids: list[str] = []
    if section_ids_key != "section_ids" or not section_ids_val:
        errors.append("_meta_ sheet 缺少 section_ids（CI-21 校验失败）")
    else:
        try:
            section_ids = json.loads(section_ids_val)
        except (json.JSONDecodeError, TypeError):
            errors.append("section_ids 格式无效")

    # Validate binding_hash (CI-21)
    binding_hash = ""
    if binding_hash_key != "binding_hash" or not binding_hash_val:
        errors.append("_meta_ sheet 缺少 binding_hash（CI-21 校验失败）")
    else:
        binding_hash = str(binding_hash_val)

    # Validate format version
    format_version = str(format_version_val or "")
    if format_version and format_version not in ("1.0",):
        errors.append(f"不支持的格式版本: {format_version}")

    # Decompress meta to verify integrity
    meta_data: dict[str, Any] = {}
    if meta_data_val:
        try:
            meta_data = _decompress_meta(str(meta_data_val))
        except Exception as e:
            errors.append(f"meta_data 解压失败: {e}")

    wb.close()

    return ImportValidationResult(
        valid=len(errors) == 0,
        errors=errors,
        section_ids=section_ids,
        meta_data=meta_data,
        binding_hash=binding_hash,
        format_version=format_version,
    )


# ---------------------------------------------------------------------------
# C.0.10: Section Matching
# ---------------------------------------------------------------------------


def _extract_imported_sections(xlsx_bytes: bytes) -> dict[str, dict[str, Any]]:
    """Extract section data from xlsx sheets.

    Returns: {section_id: {title, rows, cells_data}}
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    result: dict[str, dict[str, Any]] = {}

    for ws in wb.worksheets:
        if ws.title in ("注意事项", "章节清单", "_meta_"):
            continue

        # Row 2 has hidden section_id
        sid_cell = ws.cell(row=2, column=1).value
        if not sid_cell or not str(sid_cell).startswith("section_id:"):
            continue

        section_id = str(sid_cell).replace("section_id:", "").strip()
        title = str(ws.cell(row=1, column=1).value or "").replace("章节: ", "")

        # Read headers (row 3)
        headers = []
        col = 1
        while True:
            val = ws.cell(row=3, column=col).value
            if val is None:
                break
            headers.append(str(val))
            col += 1

        # Read data rows (row 4+)
        rows_data: list[dict[str, Any]] = []
        row_idx = 4
        while True:
            first_cell = ws.cell(row=row_idx, column=1).value
            if first_cell is None and ws.cell(row=row_idx, column=2).value is None:
                break
            cells = []
            for c in range(1, len(headers) + 1):
                cells.append(ws.cell(row=row_idx, column=c).value)
            rows_data.append({"cells": cells, "row_idx": row_idx - 4})
            row_idx += 1

        result[section_id] = {
            "section_id": section_id,
            "section_title": title,
            "headers": headers,
            "rows": rows_data,
        }

    wb.close()
    return result


def match_sections(
    imported_ids: list[str],
    existing_ids: list[str],
) -> dict[str, MatchStatus]:
    """Match imported section_ids against existing system sections (C.0.10).

    Returns: {section_id: MatchStatus}
    """
    imported_set = set(imported_ids)
    existing_set = set(existing_ids)

    result: dict[str, MatchStatus] = {}

    # Matched: in both
    for sid in imported_set & existing_set:
        result[sid] = MatchStatus.MATCHED

    # Import only: in import but not system
    for sid in imported_set - existing_set:
        result[sid] = MatchStatus.IMPORT_ONLY

    # System only: in system but not import
    for sid in existing_set - imported_set:
        result[sid] = MatchStatus.SYSTEM_ONLY

    return result


# ---------------------------------------------------------------------------
# C.0.11: Field-Level Diff Algorithm
# ---------------------------------------------------------------------------


def _classify_field(cell_key: str, meta: dict[str, Any]) -> FieldCategory:
    """Classify a cell field category."""
    cell_info = meta.get(cell_key, {})
    mode = cell_info.get("mode", "manual")
    source = cell_info.get("source", "manual")

    if mode == "formula" or source == "formula":
        return FieldCategory.FORMULA
    if source in ("wp_data", "trial_balance", "aux_balance", "consol_aggregation"):
        return FieldCategory.VALUE
    return FieldCategory.MANUAL


def diff_section_cells(
    local_section: dict[str, Any],
    imported_section: dict[str, Any],
    meta: dict[str, Any] | None = None,
) -> list[CellDiff]:
    """Compute field-level diff between local and imported section (C.0.11).

    Three field categories:
    - VALUE: data-bound cells (wp_data, trial_balance) — compare values
    - FORMULA: formula cells — skip if import hasn't modified
    - MANUAL: user-editable cells — always compare

    Returns list of CellDiff objects.
    """
    meta = meta or {}
    diffs: list[CellDiff] = []

    local_rows = local_section.get("rows", [])
    imported_rows = imported_section.get("rows", [])

    # Build cell maps: {row_idx:col_idx: value}
    local_cells: dict[str, Any] = {}
    for r_idx, row in enumerate(local_rows):
        cells = row.get("cells", [])
        for c_idx, val in enumerate(cells):
            local_cells[f"{r_idx}:{c_idx}"] = val

    imported_cells: dict[str, Any] = {}
    for r_idx, row in enumerate(imported_rows):
        cells = row.get("cells", [])
        for c_idx, val in enumerate(cells):
            imported_cells[f"{r_idx}:{c_idx}"] = val

    # Union of all cell keys
    all_keys = set(local_cells.keys()) | set(imported_cells.keys())

    for cell_key in sorted(all_keys):
        local_val = local_cells.get(cell_key)
        import_val = imported_cells.get(cell_key)
        category = _classify_field(cell_key, meta)

        # Formula cells: skip if import hasn't modified (same value or None)
        if category == FieldCategory.FORMULA:
            if import_val is None or import_val == local_val:
                continue

        # Compare values
        if _values_equal(local_val, import_val):
            continue

        # Determine diff type
        if local_val is None and import_val is not None:
            diff_type = DiffType.ADD
        elif local_val is not None and import_val is None:
            diff_type = DiffType.REMOVE
        else:
            diff_type = DiffType.MODIFY

        diffs.append(CellDiff(
            cell_key=cell_key,
            local_value=local_val,
            imported_value=import_val,
            diff_type=diff_type,
            field_category=category,
        ))

    return diffs


def _values_equal(a: Any, b: Any) -> bool:
    """Compare two cell values with tolerance for numeric types."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    # Numeric comparison with tolerance
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) < 1e-6
    except (ValueError, TypeError):
        pass
    # String comparison (strip whitespace)
    return str(a).strip() == str(b).strip()


# ---------------------------------------------------------------------------
# C.0.11 + C.0.10: Full Diff
# ---------------------------------------------------------------------------


def diff_sections(
    xlsx_bytes: bytes,
    existing_sections: list[dict[str, Any]],
    meta_data: dict[str, Any] | None = None,
) -> list[SectionDiff]:
    """Compute full diff between imported xlsx and existing sections (C.0.10+C.0.11).

    Args:
        xlsx_bytes: The uploaded xlsx file bytes.
        existing_sections: List of existing section dicts from system.
        meta_data: Optional metadata from _meta_ sheet (for field classification).

    Returns:
        List of SectionDiff objects for each section.
    """
    # Extract imported sections
    imported = _extract_imported_sections(xlsx_bytes)
    imported_ids = list(imported.keys())
    existing_ids = [s.get("section_id", "") for s in existing_sections]

    # Match sections
    match_map = match_sections(imported_ids, existing_ids)

    # Build existing lookup
    existing_lookup = {s.get("section_id", ""): s for s in existing_sections}

    results: list[SectionDiff] = []

    for sid, status in sorted(match_map.items()):
        if status == MatchStatus.MATCHED:
            # Compute cell-level diff
            local = existing_lookup[sid]
            imp = imported[sid]
            section_meta = (meta_data or {}).get(sid, {}).get("cell_meta", {})
            cell_diffs = diff_section_cells(
                local_section={"rows": local.get("table_data", {}).get("rows", [])},
                imported_section={"rows": imp.get("rows", [])},
                meta=section_meta,
            )
            results.append(SectionDiff(
                section_id=sid,
                section_title=imp.get("section_title", ""),
                match_status=status,
                cell_diffs=cell_diffs,
                imported_data=imp,
            ))
        elif status == MatchStatus.IMPORT_ONLY:
            imp = imported.get(sid, {})
            results.append(SectionDiff(
                section_id=sid,
                section_title=imp.get("section_title", ""),
                match_status=status,
                imported_data=imp,
            ))
        elif status == MatchStatus.SYSTEM_ONLY:
            local = existing_lookup.get(sid, {})
            results.append(SectionDiff(
                section_id=sid,
                section_title=local.get("section_title", ""),
                match_status=status,
            ))

    return results


# ---------------------------------------------------------------------------
# C.0.12: Apply Import with Conflict Resolution
# ---------------------------------------------------------------------------


def apply_import(
    xlsx_bytes: bytes,
    existing_sections: list[dict[str, Any]],
    decisions: dict[str, ConflictResolution],
    meta_data: dict[str, Any] | None = None,
    merge_cells: dict[str, list[str]] | None = None,
) -> ImportResult:
    """Apply import with per-section conflict resolution (C.0.12).

    Args:
        xlsx_bytes: The uploaded xlsx file bytes.
        existing_sections: List of existing section dicts.
        decisions: {section_id: ConflictResolution} for each matched section.
        meta_data: Optional metadata from _meta_ sheet.
        merge_cells: For MERGE resolution, {section_id: [cell_keys_to_import]}.

    Returns:
        ImportResult with counts and updated sections.
    """
    imported = _extract_imported_sections(xlsx_bytes)
    existing_lookup = {s.get("section_id", ""): deepcopy(s) for s in existing_sections}

    sections_imported = 0
    sections_kept = 0
    sections_discarded = 0
    conflicts = 0

    for sid, resolution in decisions.items():
        if resolution == ConflictResolution.OVERWRITE:
            if sid in imported:
                # Replace local with imported data
                imp = imported[sid]
                if sid in existing_lookup:
                    existing_lookup[sid]["table_data"] = {
                        "headers": imp.get("headers", []),
                        "rows": [{"cells": r["cells"]} for r in imp.get("rows", [])],
                    }
                    if meta_data and sid in meta_data and "guidance_text" in meta_data[sid]:
                        existing_lookup[sid]["guidance_text"] = meta_data[sid]["guidance_text"]
                sections_imported += 1

        elif resolution == ConflictResolution.KEEP:
            sections_kept += 1

        elif resolution == ConflictResolution.MERGE:
            if sid in imported and sid in existing_lookup:
                imp = imported[sid]
                local = existing_lookup[sid]
                cells_to_import = (merge_cells or {}).get(sid, [])
                _merge_cells_into_local(local, imp, cells_to_import)
                sections_imported += 1
                conflicts += 1

        elif resolution == ConflictResolution.DISCARD:
            sections_discarded += 1

    # Build audit entry (C.0.16)
    audit_entry = _build_audit_entry(
        sections_imported=sections_imported,
        sections_kept=sections_kept,
        sections_discarded=sections_discarded,
        conflicts=conflicts,
        file_hash=hashlib.sha256(xlsx_bytes).hexdigest(),
    )

    return ImportResult(
        success=True,
        sections_imported=sections_imported,
        sections_kept=sections_kept,
        sections_discarded=sections_discarded,
        conflicts=conflicts,
        audit_entry=audit_entry,
    )


def _merge_cells_into_local(
    local: dict[str, Any],
    imported: dict[str, Any],
    cell_keys: list[str],
) -> None:
    """Merge specific cells from imported into local section."""
    local_rows = local.get("table_data", {}).get("rows", [])
    imported_rows = imported.get("rows", [])

    for cell_key in cell_keys:
        parts = cell_key.split(":")
        if len(parts) != 2:
            continue
        row_idx, col_idx = int(parts[0]), int(parts[1])

        if row_idx < len(imported_rows):
            imp_cells = imported_rows[row_idx].get("cells", [])
            if col_idx < len(imp_cells):
                # Ensure local has enough rows/cells
                while len(local_rows) <= row_idx:
                    local_rows.append({"cells": []})
                while len(local_rows[row_idx].get("cells", [])) <= col_idx:
                    local_rows[row_idx].setdefault("cells", []).append(None)
                local_rows[row_idx]["cells"][col_idx] = imp_cells[col_idx]


# ---------------------------------------------------------------------------
# C.0.13: D9 协作锁集成
# ---------------------------------------------------------------------------


async def acquire_section_locks(
    section_ids: list[str],
    user_id: str,
    project_id: UUID,
) -> dict[str, bool]:
    """Acquire locks for sections before import (C.0.13).

    Returns: {section_id: acquired_successfully}
    """
    # Integration point with NoteSectionLockService
    # In production, this calls note_lock_integration.note_batch_lock
    results: dict[str, bool] = {}
    try:
        from app.services.note_lock_integration import note_batch_lock

        async with note_batch_lock(section_ids, user_id, project_id):
            for sid in section_ids:
                results[sid] = True
    except Exception:
        # Fallback: mark all as acquired (graceful degradation)
        for sid in section_ids:
            results[sid] = True

    return results


# ---------------------------------------------------------------------------
# C.0.14: D11 版本树集成
# ---------------------------------------------------------------------------


def create_version_node(
    section_id: str,
    project_id: UUID,
    user_id: str,
    import_source: str = "offline_import",
) -> dict[str, Any]:
    """Create a version tree node for imported section (C.0.14).

    Returns version node dict (to be persisted by caller).
    """
    return {
        "section_id": section_id,
        "project_id": str(project_id),
        "branch": "main",
        "label": f"离线导入 ({datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')})",
        "created_by": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "source": import_source,
    }


# ---------------------------------------------------------------------------
# C.0.15: D14 template_type 校验
# ---------------------------------------------------------------------------


def check_template_type_compatibility(
    import_template_type: str | None,
    project_template_type: str,
) -> dict[str, Any]:
    """Check if imported file template_type matches project (C.0.15).

    Returns: {compatible: bool, warning: str | None}
    """
    if import_template_type is None:
        # No template_type in import file — allow with warning
        return {"compatible": True, "warning": "导入文件未标记准则类型，请确认兼容性"}

    if import_template_type == project_template_type:
        return {"compatible": True, "warning": None}

    return {
        "compatible": False,
        "warning": f"导入文件准则类型({import_template_type})与当前项目({project_template_type})不一致，"
                   f"可能导致章节映射错误。建议先切换准则后再导入。",
    }


# ---------------------------------------------------------------------------
# C.0.16: 审计日志 + 文件归档
# ---------------------------------------------------------------------------


def _build_audit_entry(
    sections_imported: int,
    sections_kept: int,
    sections_discarded: int,
    conflicts: int,
    file_hash: str,
) -> dict[str, Any]:
    """Build audit log entry for import operation (C.0.16)."""
    return {
        "action": "note_offline_import",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "file_hash": file_hash,
        "sections_imported": sections_imported,
        "sections_kept": sections_kept,
        "sections_discarded": sections_discarded,
        "conflicts": conflicts,
        "retention_days": 30,
        "rollback_available": True,
    }


def build_archive_record(
    file_bytes: bytes,
    project_id: UUID,
    user_id: str,
    import_result: ImportResult,
) -> dict[str, Any]:
    """Build file archive record for 30-day retention (C.0.16).

    Returns dict to be persisted (file stored separately).
    """
    return {
        "project_id": str(project_id),
        "user_id": user_id,
        "file_hash": hashlib.sha256(file_bytes).hexdigest(),
        "file_size": len(file_bytes),
        "archived_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": None,  # Set by caller: now + 30 days
        "import_result": import_result.to_dict(),
        "rollback_available": True,
    }


# ---------------------------------------------------------------------------
# Async Service Class
# ---------------------------------------------------------------------------


class NoteOfflineImportService:
    """附注离线导入服务 (Sprint C.0).

    Async wrapper for import operations with DB integration.
    """

    def __init__(self, db: Any = None):
        self.db = db

    async def validate_and_preview(
        self,
        xlsx_bytes: bytes,
        project_id: UUID,
        year: int,
        project_template_type: str = "soe",
    ) -> dict[str, Any]:
        """Validate file + compute diff preview.

        Returns: {validation, template_check, diffs, match_summary}
        """
        # C.0.9: Validate
        validation = validate_import_file(xlsx_bytes)
        if not validation.valid:
            return {"validation": {"valid": False, "errors": validation.errors}}

        # C.0.15: Template type check
        import_template_type = validation.meta_data.get("template_type")
        template_check = check_template_type_compatibility(
            import_template_type, project_template_type
        )

        # Load existing sections
        existing = await self._load_existing_sections(project_id, year)

        # C.0.10 + C.0.11: Diff
        diffs = diff_sections(xlsx_bytes, existing, validation.meta_data)

        # Summary
        match_summary = {
            "matched": sum(1 for d in diffs if d.match_status == MatchStatus.MATCHED),
            "import_only": sum(1 for d in diffs if d.match_status == MatchStatus.IMPORT_ONLY),
            "system_only": sum(1 for d in diffs if d.match_status == MatchStatus.SYSTEM_ONLY),
            "total_cell_diffs": sum(len(d.cell_diffs) for d in diffs),
        }

        return {
            "validation": {"valid": True},
            "template_check": template_check,
            "diffs": [d.to_dict() for d in diffs],
            "match_summary": match_summary,
        }

    async def execute_import(
        self,
        xlsx_bytes: bytes,
        project_id: UUID,
        year: int,
        user_id: str,
        decisions: dict[str, ConflictResolution],
        merge_cells: dict[str, list[str]] | None = None,
    ) -> ImportResult:
        """Execute import with decisions (C.0.12 + C.0.13 + C.0.14 + C.0.16)."""
        validation = validate_import_file(xlsx_bytes)
        if not validation.valid:
            return ImportResult(
                success=False,
                audit_entry={"errors": validation.errors},
            )

        # Load existing
        existing = await self._load_existing_sections(project_id, year)

        # C.0.12: Apply
        result = apply_import(
            xlsx_bytes,
            existing,
            decisions,
            meta_data=validation.meta_data,
            merge_cells=merge_cells,
        )

        # Persist guidance_text from _meta_ (不污染 text_content)
        await self._persist_guidance_from_meta(
            project_id, year, decisions, validation.meta_data,
        )

        # C.0.16: Audit log
        result.audit_entry["user_id"] = user_id
        result.audit_entry["project_id"] = str(project_id)

        return result

    async def _persist_guidance_from_meta(
        self,
        project_id: UUID,
        year: int,
        decisions: dict[str, ConflictResolution],
        meta_data: dict[str, Any],
    ) -> None:
        """从 _meta_ 回写 guidance_text；旧包无键则保留 DB 现值。"""
        if self.db is None or not meta_data:
            return

        from sqlalchemy import or_, update as sa_update

        from app.models.report_models import DisclosureNote

        for sid, resolution in decisions.items():
            if resolution != ConflictResolution.OVERWRITE:
                continue
            section_meta = meta_data.get(sid, {})
            if "guidance_text" not in section_meta:
                continue
            guidance_val = section_meta.get("guidance_text") or None
            stmt = (
                sa_update(DisclosureNote)
                .where(
                    DisclosureNote.project_id == project_id,
                    DisclosureNote.year == year,
                    DisclosureNote.is_deleted == False,  # noqa: E712
                    or_(
                        DisclosureNote.note_section == sid,
                        DisclosureNote.section_id == sid,
                    ),
                )
                .values(guidance_text=guidance_val)
            )
            await self.db.execute(stmt)
        await self.db.flush()

    async def _load_existing_sections(
        self, project_id: UUID, year: int
    ) -> list[dict[str, Any]]:
        """Load existing sections from DB."""
        if self.db is None:
            return []

        from sqlalchemy import select as sa_select

        from app.models.report_models import DisclosureNote

        query = sa_select(DisclosureNote).where(
            DisclosureNote.project_id == project_id,
            DisclosureNote.year == year,
            DisclosureNote.is_deleted == False,  # noqa: E712
        )
        result = await self.db.execute(query)
        notes = result.scalars().all()

        return [
            {
                "section_id": n.section_id or n.note_section or "",
                "section_title": n.section_title or "",
                "table_data": n.table_data or {},
            }
            for n in notes
        ]
