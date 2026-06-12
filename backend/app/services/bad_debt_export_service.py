"""坏账准备明细表 D2-3 致同模板导出引擎（BadDebtExportService）

对应 design.md「Components and Interfaces」导出能力与 requirements Req 7
「致同 D2-3 模板对齐与导出」。

致同模板 14 列结构（A~N）
------------------------------------------------------------------
A 项目名 / B 期初未审数 / C 期初账项调整 / D 重分类调整(期初) / E 期初审定数 /
F 本期计提 / G 其他增加 / H 本期转回 / I 核销 / J 其他减少 /
K 期末未审数 / L 期末账项调整 / M 重分类调整(期末) / N 期末审定数

行布局
------------------------------------------------------------------
R1-R9   表头元信息区（事务所名称 / 被审计单位 / 审计期间 等，占位或传参）
R10-R11 列标题区（两行表头：分组标题 + 列名）
R12+    单项评估计提父行 → 其子行 → 信用风险组合计提父行 → 其子行 → … → 合计行

口径铁律
------------------------------------------------------------------
- 数据来自 NestedTableService.get_tree（父行 amounts 已是子行汇总，子行 amounts
  为各自值，summary 为合计行）。
- Child_Row 空金额列输出空单元格（None，非 0）。
- "其中"子行 A 列前加两个空格缩进（如 "  甲公司"）以区分层级。
- 父行 A 列 = row_label，合计行 A 列 = "合计"。

Requirements: 7.1, 7.2, 7.3, 7.4, 7.5
"""

from __future__ import annotations

import io
import uuid
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.bad_debt_schemas import RowAmounts
from app.services.bad_debt_auto_sum import AutoSumEngine
from app.services.bad_debt_nested_table_service import NestedTableService

# 13 金额列 amount_b ~ amount_n（A 项目名为标签列不在此列表）
_AMOUNT_COLUMNS = AutoSumEngine.AMOUNT_COLUMNS  # ["amount_b", ..., "amount_n"]

# 子行缩进前缀（A 列前加两个空格）
_CHILD_INDENT = "  "

# 表头元信息起始行 / 列标题行 / 数据起始行（1-based）
_META_START_ROW = 1            # R1-R9 元信息
_HEADER_GROUP_ROW = 10         # R10 分组标题
_HEADER_COL_ROW = 11           # R11 列名
_DATA_START_ROW = 12           # R12 起数据区

# 14 列分组标题（R10）：A 列 + 期初系列(B-E) + 本期增加(F-G) + 本期减少(H-J) +
# 期末系列(K-N)。用于合并单元格的分组表头。
_COLUMN_TITLES = [
    "项目",            # A
    "期初未审数",      # B
    "期初账项调整",    # C
    "重分类调整(期初)",  # D
    "期初审定数",      # E
    "本期计提",        # F
    "其他增加",        # G
    "本期转回",        # H
    "核销",            # I
    "其他减少",        # J
    "期末未审数",      # K
    "期末账项调整",    # L
    "重分类调整(期末)",  # M
    "期末审定数",      # N
]


class BadDebtExportMeta(BaseModel):
    """导出表头元信息（R1-R9 区域）。

    全部可选，未提供时输出空表头占位，不影响 14 列数据结构。
    """

    firm_name: str | None = None          # 事务所名称
    entity_name: str | None = None        # 被审计单位
    audit_period: str | None = None       # 审计期间
    sheet_title: str | None = None        # 表标题（默认"坏账准备明细表"）
    currency_unit: str | None = None      # 金额单位（如"元"）


class BadDebtExportService:
    """坏账准备明细表 D2-3 致同 14 列模板 xlsx 导出。"""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ─── 公开导出接口 ────────────────────────────────────────────────────────

    async def export_workbook(
        self, wp_index_id: uuid.UUID, meta: BadDebtExportMeta | None = None
    ) -> Workbook:
        """构建并返回 openpyxl Workbook（不落盘）。

        Requirements: 7.1, 7.2, 7.3, 7.4, 7.5
        """
        tree = await NestedTableService(self.db).get_tree(wp_index_id)
        meta = meta or BadDebtExportMeta()

        wb = Workbook()
        ws = wb.active
        ws.title = "D2-3"

        self._write_meta(ws, meta)
        self._write_header(ws)
        next_row = self._write_data_rows(ws, tree)
        self._write_summary_row(ws, tree, next_row)
        self._apply_column_widths(ws)

        return wb

    async def export_bytes(
        self, wp_index_id: uuid.UUID, meta: BadDebtExportMeta | None = None
    ) -> io.BytesIO:
        """导出为 BytesIO（供 HTTP 下载 / 流式响应）。"""
        wb = await self.export_workbook(wp_index_id, meta)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ─── 表头元信息区 R1-R9 ──────────────────────────────────────────────────

    def _write_meta(self, ws: Worksheet, meta: BadDebtExportMeta) -> None:
        """写入 R1-R9 元信息区（Req 7.4）。

        固定行位：
        R1 表标题 / R3 事务所名称 / R4 被审计单位 / R5 审计期间 / R7 金额单位
        其余行留空作为占位，保持模板上方留白结构。
        """
        title = meta.sheet_title or "坏账准备明细表"
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        if meta.firm_name is not None:
            ws.cell(row=3, column=1, value=f"事务所：{meta.firm_name}")
        if meta.entity_name is not None:
            ws.cell(row=4, column=1, value=f"被审计单位：{meta.entity_name}")
        if meta.audit_period is not None:
            ws.cell(row=5, column=1, value=f"审计期间：{meta.audit_period}")
        if meta.currency_unit is not None:
            # 金额单位通常右上角；此处放 R7 N 列附近作占位
            ws.cell(row=7, column=14, value=f"金额单位：{meta.currency_unit}")

    # ─── 列标题区 R10-R11 ─────────────────────────────────────────────────────

    def _write_header(self, ws: Worksheet) -> None:
        """写入 R10 分组标题 + R11 列名（Req 7.1）。

        R10：A 列「项目」+ 分组合并标题（期初系列 / 本期增加 / 本期减少 / 期末系列）。
        R11：14 列逐列列名（与 _COLUMN_TITLES 顺序一致）。
        """
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # R10 分组标题：A 项目 + 四个分组
        ws.cell(row=_HEADER_GROUP_ROW, column=1, value="项目")
        groups = [
            (2, 5, "期初"),       # B-E
            (6, 7, "本期增加"),   # F-G
            (8, 10, "本期减少"),  # H-J
            (11, 14, "期末"),     # K-N
        ]
        for start_col, end_col, label in groups:
            ws.cell(row=_HEADER_GROUP_ROW, column=start_col, value=label)
            if end_col > start_col:
                ws.merge_cells(
                    start_row=_HEADER_GROUP_ROW,
                    start_column=start_col,
                    end_row=_HEADER_GROUP_ROW,
                    end_column=end_col,
                )

        # R11 列名（14 列）
        for idx, title in enumerate(_COLUMN_TITLES, start=1):
            cell = ws.cell(row=_HEADER_COL_ROW, column=idx, value=title)
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

        # R10 分组标题样式
        for col in range(1, 15):
            cell = ws.cell(row=_HEADER_GROUP_ROW, column=col)
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

    # ─── 数据区 R12+ ─────────────────────────────────────────────────────────

    def _write_data_rows(self, ws: Worksheet, tree) -> int:
        """按模板行顺序写父行 + 子行，返回下一个可写行号（用于合计行）。

        行顺序：父行(加粗) → 其子行(缩进"  XXX") → 下一父行 → … （Req 7.2, 7.5）
        子行空金额列输出空单元格（None，非 0）（Req 7.3）。
        """
        row_idx = _DATA_START_ROW
        for parent in tree.parents:
            # 父行：A 列 = row_label，加粗；13 金额列为子行汇总（已在 get_tree 算好）
            self._write_row(
                ws, row_idx, label=parent.row_label, amounts=parent.amounts, bold=True
            )
            row_idx += 1
            # 子行：A 列缩进两空格，空金额列留空
            for child in parent.children:
                self._write_row(
                    ws,
                    row_idx,
                    label=f"{_CHILD_INDENT}{child.row_label}",
                    amounts=child.amounts,
                    bold=False,
                )
                row_idx += 1
        return row_idx

    def _write_summary_row(self, ws: Worksheet, tree, row_idx: int) -> None:
        """写合计行（Req 7.2）：A 列 = "合计"，13 金额列 = Summary 汇总值。"""
        self._write_row(
            ws, row_idx, label="合计", amounts=tree.summary.amounts, bold=True
        )

    def _write_row(
        self,
        ws: Worksheet,
        row_idx: int,
        *,
        label: str,
        amounts: RowAmounts,
        bold: bool,
    ) -> None:
        """写单行：A 列标签 + B~N 13 金额列。

        金额列为 None 时输出空单元格（不写 0）（Req 7.3）。
        """
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        if bold:
            label_cell.font = Font(bold=True)

        # B~N 列：column 2..14 对应 amount_b..amount_n
        for offset, col_name in enumerate(_AMOUNT_COLUMNS):
            col_idx = offset + 2  # B 列起
            value = getattr(amounts, col_name)
            if value is None:
                continue  # 空值不写（保持空单元格，Req 7.3）
            cell = ws.cell(row=row_idx, column=col_idx, value=_to_number(value))
            cell.number_format = "#,##0.00"
            if bold:
                cell.font = Font(bold=True)

    # ─── 列宽 ────────────────────────────────────────────────────────────────

    def _apply_column_widths(self, ws: Worksheet) -> None:
        """A 列宽（项目名），B~N 数值列统一宽度。"""
        ws.column_dimensions["A"].width = 28
        for col_letter in "BCDEFGHIJKLMN":
            ws.column_dimensions[col_letter].width = 14


def _to_number(value: Decimal) -> float | Decimal:
    """金额写入 openpyxl：保留 Decimal 以避免 float 漂移（openpyxl 支持 Decimal）。"""
    return value


__all__ = ["BadDebtExportService", "BadDebtExportMeta"]
