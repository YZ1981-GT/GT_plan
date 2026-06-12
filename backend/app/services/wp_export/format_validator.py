"""FormatValidator - 导入文件格式校验器

验证导入文件的结构和数据完整性：
- MIME 类型一致性（扩展名与实际内容）
- Sheet 页签结构匹配 render_schema
- 必填单元格检查
- 数值型字段类型校验

返回结构化 ValidationReport (passed/warnings/errors)。

Requirements: 5.1, 5.2, 5.3, 5.4, 5.5, 5.6
"""

from __future__ import annotations

import io
import zipfile
from typing import Any

from openpyxl import load_workbook

from app.schemas.wp_export_schemas import (
    ValidationItem,
    ValidationLevel,
    ValidationReport,
)


class FormatValidator:
    """格式校验器

    校验导入文件的 MIME 类型、sheet 结构、必填单元格、数值类型。
    """

    # 扩展名 → 预期 MIME 类型
    MIME_MAP: dict[str, str] = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }

    # ZIP magic bytes: PK\x03\x04
    _ZIP_MAGIC = b"PK\x03\x04"

    def validate(
        self,
        file_content: bytes,
        filename: str,
        render_schema: dict | None = None,
    ) -> ValidationReport:
        """完整校验流程，返回结构化 ValidationReport。

        Args:
            file_content: 文件二进制内容
            filename: 文件名（含扩展名）
            render_schema: 底稿渲染 schema（含 sheets 定义）

        Returns:
            ValidationReport: overall 为最严重级别，items 按级别分类
        """
        items: list[ValidationItem] = []

        # 1. MIME 类型校验
        items.extend(self._check_mime_type(file_content, filename))

        # 2~4 需要 render_schema 才能进行结构/必填/数值校验
        if render_schema is not None:
            # 仅 xlsx 才做 sheet 结构、必填、数值校验
            ext = self._get_extension(filename)
            if ext == ".xlsx":
                items.extend(
                    self._check_sheet_structure(file_content, render_schema)
                )
                items.extend(
                    self._check_required_cells(file_content, render_schema)
                )
                items.extend(
                    self._check_numeric_types(file_content, render_schema)
                )

        return self._build_report(items)

    # ─── 私有校验方法 ─────────────────────────────────────────────────────

    def _check_mime_type(
        self, content: bytes, filename: str
    ) -> list[ValidationItem]:
        """检查文件扩展名与实际内容类型一致性。

        xlsx/docx 均为 ZIP 格式，通过 [Content_Types].xml 区分。
        """
        items: list[ValidationItem] = []
        ext = self._get_extension(filename)

        if ext not in self.MIME_MAP:
            items.append(
                ValidationItem(
                    level=ValidationLevel.ERROR,
                    location="file",
                    message=f"不支持的文件扩展名: {ext}",
                    field="extension",
                )
            )
            return items

        # 检查 ZIP magic bytes
        if not content.startswith(self._ZIP_MAGIC):
            items.append(
                ValidationItem(
                    level=ValidationLevel.ERROR,
                    location="file",
                    message=f"文件内容不是有效的 ZIP 格式，与扩展名 {ext} 不匹配",
                    field="mime_type",
                )
            )
            return items

        # 检查 ZIP 内 [Content_Types].xml 确认真实类型
        try:
            with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
                if "[Content_Types].xml" not in zf.namelist():
                    items.append(
                        ValidationItem(
                            level=ValidationLevel.ERROR,
                            location="file",
                            message="ZIP 缺少 [Content_Types].xml，非有效 Office 文件",
                            field="mime_type",
                        )
                    )
                    return items

                content_types = zf.read("[Content_Types].xml").decode("utf-8")

                if ext == ".xlsx":
                    if "spreadsheetml" not in content_types:
                        items.append(
                            ValidationItem(
                                level=ValidationLevel.ERROR,
                                location="file",
                                message="文件扩展名为 .xlsx 但内容不是 spreadsheetml 格式",
                                field="mime_type",
                            )
                        )
                elif ext == ".docx":
                    if "wordprocessingml" not in content_types:
                        items.append(
                            ValidationItem(
                                level=ValidationLevel.ERROR,
                                location="file",
                                message="文件扩展名为 .docx 但内容不是 wordprocessingml 格式",
                                field="mime_type",
                            )
                        )
        except zipfile.BadZipFile:
            items.append(
                ValidationItem(
                    level=ValidationLevel.ERROR,
                    location="file",
                    message="文件虽以 ZIP 头开始但不是有效 ZIP 归档",
                    field="mime_type",
                )
            )

        return items

    def _check_sheet_structure(
        self, file_content: bytes, render_schema: dict
    ) -> list[ValidationItem]:
        """检查 sheet 页签名称与 render_schema 中定义的 sheets 匹配。

        render_schema.sheets 的 key 为期望的 sheet 名称。
        """
        items: list[ValidationItem] = []
        expected_sheets = set(render_schema.get("sheets", {}).keys())

        if not expected_sheets:
            return items

        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            actual_sheets = set(wb.sheetnames)
            wb.close()
        except Exception as e:
            items.append(
                ValidationItem(
                    level=ValidationLevel.ERROR,
                    location="file",
                    message=f"无法解析 xlsx 文件: {str(e)}",
                    field="sheet_structure",
                )
            )
            return items

        # 找出缺失的 sheet
        missing = expected_sheets - actual_sheets
        for sheet_name in sorted(missing):
            items.append(
                ValidationItem(
                    level=ValidationLevel.ERROR,
                    location=f"Sheet:{sheet_name}",
                    message=f"缺少期望的 sheet 页签: {sheet_name}",
                    field="sheet_structure",
                )
            )

        return items

    def _check_required_cells(
        self, file_content: bytes, render_schema: dict
    ) -> list[ValidationItem]:
        """检查 render_schema 中标记为 required 的字段是否有值。

        schema 结构示例:
        {
          "sheets": {
            "Sheet1": {
              "dynamic_table": {
                "start_row": 3,
                "columns": {
                  "A": {"field": "code", "required": true},
                  "B": {"field": "name", "required": false}
                }
              }
            }
          }
        }
        """
        items: list[ValidationItem] = []
        sheets_schema = render_schema.get("sheets", {})

        if not sheets_schema:
            return items

        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception:
            return items

        try:
            for sheet_name, sheet_schema in sheets_schema.items():
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                table = sheet_schema.get("dynamic_table", {})
                columns = table.get("columns", {})
                start_row = table.get("start_row", 1)

                # Find required columns
                required_cols: list[tuple[str, str]] = []
                for col_letter, col_def in columns.items():
                    if isinstance(col_def, dict) and col_def.get("required", False):
                        field = col_def.get("field", col_letter)
                        required_cols.append((col_letter, field))

                if not required_cols:
                    continue

                # Check each data row for required values
                max_row = ws.max_row or start_row
                for row_idx in range(start_row, max_row + 1):
                    # Check if row has any data at all (skip completely empty rows)
                    row_has_data = False
                    for col_letter, _ in columns.items():
                        cell = ws[f"{col_letter}{row_idx}"]
                        if cell.value is not None and cell.value != "":
                            row_has_data = True
                            break

                    if not row_has_data:
                        continue

                    # Check required cells in data rows
                    for col_letter, field in required_cols:
                        cell = ws[f"{col_letter}{row_idx}"]
                        if cell.value is None or (
                            isinstance(cell.value, str) and cell.value.strip() == ""
                        ):
                            items.append(
                                ValidationItem(
                                    level=ValidationLevel.ERROR,
                                    location=f"{sheet_name}!{col_letter}{row_idx}",
                                    message=f"必填字段 '{field}' 为空",
                                    field=field,
                                )
                            )
        finally:
            wb.close()

        return items

    def _check_numeric_types(
        self, file_content: bytes, render_schema: dict
    ) -> list[ValidationItem]:
        """检查 render_schema 中标记为 number 类型的列内容是否为数值。

        非数值内容填入金额单元格标记为 warning 级。
        """
        items: list[ValidationItem] = []
        sheets_schema = render_schema.get("sheets", {})

        if not sheets_schema:
            return items

        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception:
            return items

        try:
            for sheet_name, sheet_schema in sheets_schema.items():
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                table = sheet_schema.get("dynamic_table", {})
                columns = table.get("columns", {})
                start_row = table.get("start_row", 1)

                # Find numeric columns
                numeric_cols: list[tuple[str, str]] = []
                for col_letter, col_def in columns.items():
                    if isinstance(col_def, dict) and col_def.get("type") == "number":
                        field = col_def.get("field", col_letter)
                        numeric_cols.append((col_letter, field))

                if not numeric_cols:
                    continue

                max_row = ws.max_row or start_row
                for row_idx in range(start_row, max_row + 1):
                    for col_letter, field in numeric_cols:
                        cell = ws[f"{col_letter}{row_idx}"]
                        value = cell.value
                        # Skip None/empty (not a type error)
                        if value is None or (isinstance(value, str) and value.strip() == ""):
                            continue
                        # Check if value is numeric
                        if not isinstance(value, (int, float)):
                            # Try to parse string as number
                            if isinstance(value, str):
                                try:
                                    float(value)
                                except (ValueError, TypeError):
                                    items.append(
                                        ValidationItem(
                                            level=ValidationLevel.WARNING,
                                            location=f"{sheet_name}!{col_letter}{row_idx}",
                                            message=f"数值列 '{field}' 包含非数值内容: {repr(value)}",
                                            field=field,
                                        )
                                    )
                            else:
                                items.append(
                                    ValidationItem(
                                        level=ValidationLevel.WARNING,
                                        location=f"{sheet_name}!{col_letter}{row_idx}",
                                        message=f"数值列 '{field}' 包含非数值类型: {type(value).__name__}",
                                        field=field,
                                    )
                                )
        finally:
            wb.close()

        return items

    # ─── 工具方法 ─────────────────────────────────────────────────────────

    @staticmethod
    def _get_extension(filename: str) -> str:
        """提取文件扩展名（小写）"""
        dot_idx = filename.rfind(".")
        if dot_idx == -1:
            return ""
        return filename[dot_idx:].lower()

    @staticmethod
    def _build_report(items: list[ValidationItem]) -> ValidationReport:
        """构建 ValidationReport，overall 取最严重级别。"""
        error_count = sum(1 for i in items if i.level == ValidationLevel.ERROR)
        warning_count = sum(1 for i in items if i.level == ValidationLevel.WARNING)
        passed_count = sum(1 for i in items if i.level == ValidationLevel.PASSED)

        if error_count > 0:
            overall = ValidationLevel.ERROR
        elif warning_count > 0:
            overall = ValidationLevel.WARNING
        else:
            overall = ValidationLevel.PASSED

        return ValidationReport(
            overall=overall,
            items=items,
            passed_count=passed_count,
            warning_count=warning_count,
            error_count=error_count,
        )
