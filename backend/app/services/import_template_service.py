"""统一 Excel 导入模板服务

为系统中所有业务数据提供标准化的导入模板生成、格式校验、数据解析能力。
支持的导入类型：
  - adjustments: 调整分录
  - report: 报表数据（单张/全套）
  - disclosure_note: 附注数据（单个/全套）
  - workpaper: 底稿数据
  - formula: 公式（报表/附注）
  - staff: 人员库
"""

from __future__ import annotations

import io
import logging
from enum import Enum
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ImportType(str, Enum):
    adjustments = "adjustments"
    report = "report"
    disclosure_note = "disclosure_note"
    workpaper = "workpaper"
    formula = "formula"
    staff = "staff"
    trial_balance = "trial_balance"


# ── 模板列定义 ──────────────────────────────────────────────

# 每种导入类型的列定义：(列名, 是否必填, 数据类型说明, 示例值)
TEMPLATE_COLUMNS: dict[ImportType, list[tuple[str, bool, str, str]]] = {
    ImportType.adjustments: [
        # 新模板格式 (export-template 端点产出): 每行 1 个科目, 借贷分行
        # 9 列联动结构: 编号/类型/摘要/二级编码/二级名称(主输入)/一级编码/一级名称/报表项目/借方/贷方
        # parse_import_data 也兼容旧格式 (科目名称/科目编码 7 列模板 + 分录编号/借方科目代码 等)
        # 示例值要与 export-template 模板里的示例完全一致 (用于 _is_example_row 跳过)
        ("编号", False, "文本", "AJE-001"),
        ("类型", True, "AJE/RJE", "AJE"),
        ("摘要", True, "文本", "补提应收账款减值"),
        ("二级科目编码", False, "标准科目代码(自动填充)", ""),
        ("二级科目名称", True, "文本(下拉选择)", "应收账款"),
        ("一级科目编码", False, "标准科目代码(自动联动)", ""),
        ("一级科目名称", False, "文本(自动联动)", ""),
        ("报表项目", False, "文本(自动联动)", ""),
        ("借方金额", False, "数值", ""),
        ("贷方金额", False, "数值", "100000"),
    ],
    ImportType.report: [
        ("行次", True, "整数", "1"),
        ("项目", True, "文本", "货币资金"),
        ("本期金额", False, "数值", "1000000.00"),
        ("上期金额", False, "数值", "900000.00"),
        ("备注", False, "文本", ""),
    ],
    ImportType.disclosure_note: [
        ("章节编号", True, "文本", "F01"),
        ("章节标题", True, "文本", "货币资金"),
        ("表格编号", False, "文本", "T01"),
        ("行名称", True, "文本", "库存现金"),
        ("期末余额", False, "数值", "50000.00"),
        ("期初余额", False, "数值", "45000.00"),
        ("本期增加", False, "数值", "5000.00"),
        ("本期减少", False, "数值", "0.00"),
        ("备注", False, "文本", ""),
    ],
    ImportType.workpaper: [
        ("底稿编码", True, "文本", "D2-1"),
        ("底稿名称", True, "文本", "应收账款明细表"),
        ("审计循环", False, "文本", "D"),
        ("行序号", True, "整数", "1"),
        ("行名称", True, "文本", "客户A"),
        ("未审数", False, "数值", "100000.00"),
        ("调整数", False, "数值", "0.00"),
        ("审定数", False, "数值", "100000.00"),
        ("说明", False, "文本", "已函证确认"),
    ],
    ImportType.formula: [
        ("公式编号", True, "文本", "F-BS-001"),
        ("适用模块", True, "report/note", "report"),
        ("报表类型", False, "文本", "balance_sheet"),
        ("行次/章节", True, "文本", "1"),
        ("列标识", False, "文本", "本期金额"),
        ("公式表达式", True, "文本", "TB('1001','期末')"),
        ("说明", False, "文本", "货币资金=库存现金+银行存款"),
    ],
    ImportType.staff: [
        ("姓名", True, "文本", "张三"),
        ("工号", False, "文本", "SJ2-001"),
        ("部门", False, "文本", "审计二部"),
        ("职级", False, "文本", "高级审计员"),
        ("所属合伙人", False, "文本", "李四"),
        ("专业领域", False, "文本", "金融审计"),
        ("联系电话", False, "文本", "13800138000"),
        ("邮箱", False, "文本", "zhangsan@example.com"),
    ],
    ImportType.trial_balance: [
        ("科目代码", True, "文本", "1001"),
        ("科目名称", True, "文本", "库存现金"),
        ("方向", False, "借/贷", "借"),
        ("科目类别", False, "资产/负债/权益/收入/费用", "资产"),
        ("期初余额", False, "数值", "50000.00"),
        ("本期借方", False, "数值", "10000.00"),
        ("本期贷方", False, "数值", "5000.00"),
        ("期末余额", False, "数值", "55000.00"),
        ("未审数", False, "数值", "55000.00"),
        ("调整数", False, "数值", "0.00"),
        ("审定数", False, "数值", "55000.00"),
    ],
}

# 导入类型的中文名称
IMPORT_TYPE_LABELS: dict[ImportType, str] = {
    ImportType.adjustments: "调整分录",
    ImportType.report: "报表数据",
    ImportType.disclosure_note: "附注数据",
    ImportType.workpaper: "底稿数据",
    ImportType.formula: "公式",
    ImportType.staff: "人员库",
    ImportType.trial_balance: "试算表",
}


# ── 样式常量 ──────────────────────────────────────────────

_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4B2D77", end_color="4B2D77", fill_type="solid")
_REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_EXAMPLE_FONT = Font(name="微软雅黑", size=10, color="999999", italic=True)
_NORMAL_FONT = Font(name="微软雅黑", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ── 模板生成 ──────────────────────────────────────────────


def generate_template(
    import_type: ImportType,
    *,
    extra_context: dict[str, Any] | None = None,
) -> bytes:
    """生成标准导入模板 Excel 文件。

    返回 xlsx 文件的 bytes，可直接作为 StreamingResponse 返回。
    extra_context 可传入额外参数（如报表类型、附注章节等）用于定制模板。
    """
    columns = TEMPLATE_COLUMNS.get(import_type)
    if not columns:
        raise ValueError(f"不支持的导入类型: {import_type}")

    label = IMPORT_TYPE_LABELS.get(import_type, str(import_type.value))
    wb = Workbook()

    # ── Sheet 1: 数据填写 ──
    ws = wb.active
    ws.title = "数据"

    # 写表头
    for col_idx, (col_name, required, dtype, _example) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = f"{'* ' if required else ''}{col_name}"
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) * 2.5, 15)

    # 写示例行（第 2 行，灰色斜体）
    for col_idx, (_col_name, _required, _dtype, example) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = example
        cell.font = _EXAMPLE_FONT
        cell.border = _THIN_BORDER

    # 写第二个示例行（第 3 行，空白供用户填写）
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.border = _THIN_BORDER

    # 高亮必填列
    for col_idx, (_col_name, required, _dtype, _example) in enumerate(columns, 1):
        if required:
            for row in range(2, 4):
                ws.cell(row=row, column=col_idx).fill = _REQUIRED_FILL

    # ── Sheet 2: 填写说明 ──
    ws_help = wb.create_sheet("填写说明")
    ws_help.column_dimensions["A"].width = 20
    ws_help.column_dimensions["B"].width = 12
    ws_help.column_dimensions["C"].width = 20
    ws_help.column_dimensions["D"].width = 40

    help_headers = ["列名", "是否必填", "数据类型", "说明"]
    for col_idx, h in enumerate(help_headers, 1):
        cell = ws_help.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (col_name, required, dtype, example) in enumerate(columns, 2):
        ws_help.cell(row=row_idx, column=1, value=col_name).font = _NORMAL_FONT
        ws_help.cell(row=row_idx, column=2, value="必填" if required else "选填").font = _NORMAL_FONT
        ws_help.cell(row=row_idx, column=3, value=dtype).font = _NORMAL_FONT
        ws_help.cell(row=row_idx, column=4, value=f"示例: {example}").font = _NORMAL_FONT
        if required:
            for c in range(1, 5):
                ws_help.cell(row=row_idx, column=c).fill = _REQUIRED_FILL

    # 底部注意事项
    note_row = len(columns) + 3
    notes = [
        f"【{label}导入模板】",
        "1. 请在「数据」Sheet 中填写数据，第 2 行为示例（导入时会自动跳过）",
        "2. 带 * 号的列为必填项，黄色底色标记",
        "3. 请勿修改表头列名，否则系统无法识别",
        "4. 数值列请填写纯数字，不要包含千分位逗号或货币符号",
        "5. 日期格式统一为 YYYY-MM-DD（如 2025-12-31）",
    ]
    for i, note in enumerate(notes):
        cell = ws_help.cell(row=note_row + i, column=1)
        cell.value = note
        cell.font = Font(name="微软雅黑", size=10, color="666666")
        ws_help.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 格式校验 ──────────────────────────────────────────────


class ValidationError:
    """单条校验错误"""
    def __init__(self, row: int | None, column: str, message: str, severity: str = "error"):
        self.row = row
        self.column = column
        self.message = message
        self.severity = severity  # error / warning

    def to_dict(self) -> dict:
        return {"row": self.row, "column": self.column, "message": self.message, "severity": self.severity}


class ImportValidationResult:
    """校验结果"""
    def __init__(self):
        self.errors: list[ValidationError] = []
        self.warnings: list[ValidationError] = []
        self.row_count: int = 0
        self.preview_rows: list[dict[str, Any]] = []

    @property
    def valid(self) -> bool:
        return len(self.errors) == 0

    def to_dict(self) -> dict:
        return {
            "valid": self.valid,
            "row_count": self.row_count,
            "errors": [e.to_dict() for e in self.errors],
            "warnings": [w.to_dict() for w in self.warnings],
            "preview_rows": self.preview_rows[:10],  # 最多预览 10 行
        }


# 数值类型关键词集合
_NUMERIC_TYPES = {"数值", "整数"}


def _is_example_row(row_values: list[str], columns: list[tuple]) -> bool:
    """判断是否为模板示例行（宽松匹配：超过一半的值和示例一致即视为示例行）

    注: 仅在前 6 行使用 (覆盖 adjustments 模板的 5 行示例区).
    特殊处理: 编号 == AJE-001 / RJE-001 视为示例行 (导出模板预填的固定示例编号).
    """
    if not row_values:
        return False

    # 特殊路径: 第一列是编号,等于 AJE-001 / RJE-001 直接视为示例
    if row_values and str(row_values[0]).strip() in ("AJE-001", "RJE-001"):
        return True

    example_values = [col[3] for col in columns]

    def _norm(v) -> str:
        """规范化值: 去前后空白 + 数字 100000 == 字符串 100000"""
        s = str(v).strip() if v is not None else ""
        try:
            f = float(s.replace(",", ""))
            if f == int(f):
                return str(int(f))
            return str(f)
        except (ValueError, TypeError):
            return s

    matchable_pairs = [(a, b) for a, b in zip(row_values, example_values) if str(b).strip()]
    if not matchable_pairs:
        return False
    match_count = sum(1 for a, b in matchable_pairs if _norm(a) == _norm(b))
    return match_count >= max(2, len(matchable_pairs) * 0.6)


def _is_numeric_column(dtype: str) -> bool:
    """判断列的数据类型是否为数值型"""
    return any(t in dtype for t in _NUMERIC_TYPES)


def _try_parse_numeric(value: Any) -> tuple[bool, float | None]:
    """尝试将值解析为数值，返回 (是否成功, 数值)"""
    if value is None or str(value).strip() == "":
        return True, None  # 空值不算数值错误
    try:
        # 去除千分位逗号和货币符号
        cleaned = str(value).replace(",", "").replace("，", "").replace("¥", "").replace("$", "").strip()
        return True, float(cleaned)
    except (ValueError, TypeError):
        return False, None


def validate_import_file(
    import_type: ImportType,
    file_bytes: bytes,
    filename: str = "",
) -> ImportValidationResult:
    """校验上传的 Excel 文件是否符合模板格式。

    返回校验结果，包含错误列表和预览数据。
    """
    result = ImportValidationResult()
    columns = TEMPLATE_COLUMNS.get(import_type)
    if not columns:
        result.errors.append(ValidationError(None, "", f"不支持的导入类型: {import_type}"))
        return result

    expected_names = [col[0] for col in columns]
    required_names = {col[0] for col in columns if col[1]}
    # 列名→数据类型映射
    col_dtype_map = {col[0]: col[2] for col in columns}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        result.errors.append(ValidationError(None, "", f"无法解析文件: {e}"))
        return result

    # ─── 自动选 sheet (与 parse_import_data 同款策略) ─────
    # adjustments 类型: 找模板 sheet, 跳过 关注事项/标准科目库 等说明 sheet
    ws = None
    if import_type == ImportType.adjustments:
        skip_names = {"关注事项", "标准科目库", "说明", "Sheet1"}
        for sn in wb.sheetnames:
            if sn in skip_names:
                continue
            if "模板" in sn or "AJE" in sn.upper() or "RJE" in sn.upper():
                ws = wb[sn]
                break
    if ws is None:
        ws = wb.active
    if ws is None:
        result.errors.append(ValidationError(None, "", "文件中没有工作表"))
        return result

    # 读取表头（第 1 行）
    header_row = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
        val = str(cell.value or "").strip().lstrip("* ").strip()
        header_row.append(val)

    if not header_row:
        result.errors.append(ValidationError(1, "", "第一行为空，请确认表头"))
        return result

    # 检查必填列是否存在
    header_set = set(header_row)
    missing = required_names - header_set
    if missing:
        result.errors.append(ValidationError(
            1, "", f"缺少必填列: {', '.join(sorted(missing))}"
        ))

    # 检查是否有未知列（警告）
    known_set = set(expected_names)
    unknown = header_set - known_set - {""}
    if unknown:
        result.warnings.append(ValidationError(
            1, "", f"包含未知列（将被忽略）: {', '.join(sorted(unknown))}"
        ))

    # 建立列名→索引映射
    col_map: dict[str, int] = {}
    for idx, name in enumerate(header_row):
        if name in known_set:
            col_map[name] = idx

    # 逐行校验数据（从第 2 行开始）
    data_rows: list[dict[str, Any]] = []
    error_count_limit = 50  # 最多报 50 个错误，避免大文件刷屏

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过示例行 (扩展到前 6 行覆盖 adjustments 模板的 5 行示例区)
        if row_idx <= 6:
            row_values = [str(v or "").strip() for v in row[:len(columns)]]
            if _is_example_row(row_values, columns):
                continue

        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_data: dict[str, Any] = {}
        for col_name, col_idx in col_map.items():
            if col_idx < len(row):
                row_data[col_name] = row[col_idx]

        # 校验必填字段
        if len(result.errors) < error_count_limit:
            for req_name in required_names:
                if req_name in col_map:
                    val = row_data.get(req_name)
                    if val is None or str(val).strip() == "":
                        result.errors.append(ValidationError(
                            row_idx, req_name, f"第 {row_idx} 行「{req_name}」不能为空"
                        ))

        # 校验数值类型
        if len(result.errors) < error_count_limit:
            for col_name, val in row_data.items():
                dtype = col_dtype_map.get(col_name, "")
                if _is_numeric_column(dtype) and val is not None and str(val).strip() != "":
                    ok, _ = _try_parse_numeric(val)
                    if not ok:
                        result.errors.append(ValidationError(
                            row_idx, col_name, f"第 {row_idx} 行「{col_name}」应为数值，实际值: {val}"
                        ))

        data_rows.append(row_data)

    result.row_count = len(data_rows)
    result.preview_rows = [{k: _serialize_cell(v) for k, v in r.items()} for r in data_rows[:10]]

    if result.row_count == 0 and not result.errors:
        result.warnings.append(ValidationError(None, "", "文件中没有有效数据行"))

    if len(result.errors) >= error_count_limit:
        result.warnings.append(ValidationError(None, "", f"错误过多，仅显示前 {error_count_limit} 条"))

    wb.close()
    return result


def parse_import_data(
    import_type: ImportType,
    file_bytes: bytes,
) -> list[dict[str, Any]]:
    """解析已校验通过的 Excel 文件，返回结构化数据列表。

    特殊处理 (adjustments):
    - 优先读取所有匹配 AJE/RJE 模板的 sheet (跳过 关注事项 / 标准科目库 等说明型 sheet)
    - 列名兼容新旧两套格式:
      * 新模板: 编号 / 类型 / 摘要 / 科目名称 / 科目编码 / 借方金额 / 贷方金额
      * 旧模板: 分录编号 / 调整类型 / 借方科目代码 / 借方金额 / 贷方科目代码 / 贷方金额 / 摘要
    """
    columns = TEMPLATE_COLUMNS.get(import_type, [])
    expected_names = [col[0] for col in columns]
    known_set = set(expected_names)

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # ─── 自动选 sheet ──────────────────────────────────────
    # adjustments: 找名称含 模板 且不是 关注事项/标准科目库 的 sheet (支持多 sheet 合并)
    target_sheets: list = []
    if import_type == ImportType.adjustments:
        skip_names = {"关注事项", "标准科目库", "说明", "Sheet1"}
        for sn in wb.sheetnames:
            if sn in skip_names:
                continue
            # 含"模板"关键词或不在跳过名单的都试着读
            if "模板" in sn or "AJE" in sn.upper() or "RJE" in sn.upper():
                target_sheets.append(wb[sn])
        # 兜底: 没找到任何模板 sheet 就用 active
        if not target_sheets:
            target_sheets = [wb.active] if wb.active is not None else []
    else:
        # 其他类型仍用 active
        target_sheets = [wb.active] if wb.active is not None else []

    # ─── 兼容旧字段名: 把旧字段名映射到新字段 ──────────────
    # 旧 → 新 别名表 (parse 阶段统一规范化, 后续 _import_adjustments 只看新字段)
    legacy_alias_map = {
        "分录编号": "编号",
        "调整类型": "类型",
        # 7 列模板兼容 (科目名称/科目编码 → 二级科目名称/二级科目编码)
        "科目名称": "二级科目名称",
        "科目编码": "二级科目编码",
    }

    # ─── 解析数据行 ──────────────────────────────────────
    rows: list[dict[str, Any]] = []
    for ws in target_sheets:
        if ws is None:
            continue

        # 读取表头
        header_row = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
            val = str(cell.value or "").strip().lstrip("* ").strip()
            header_row.append(val)

        col_map: dict[str, int] = {}
        for idx, name in enumerate(header_row):
            # 别名规范化
            normalized = legacy_alias_map.get(name, name)
            if normalized in known_set or name in (
                "借方科目代码", "借方科目名称", "贷方科目代码", "贷方科目名称",
                "借方金额", "贷方金额",  # 可能直接含,也属于已知
            ):
                # 旧模板专用字段保留原名供 _import_adjustments 兜底处理
                col_map[normalized if normalized in known_set else name] = idx

        if not col_map:
            continue  # 这个 sheet 没识别到任何列, 跳过

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过示例行 (扩展到前 6 行覆盖 adjustments 模板的 5 行示例区)
            if row_idx <= 6:
                row_values = [str(v or "").strip() for v in row[:len(columns)]]
                if _is_example_row(row_values, columns):
                    continue

            # 跳过全空行
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            row_data: dict[str, Any] = {}
            for col_name, col_idx in col_map.items():
                if col_idx < len(row):
                    row_data[col_name] = row[col_idx]

            rows.append(row_data)

    wb.close()
    return rows


def _serialize_cell(value: Any) -> Any:
    """将单元格值序列化为 JSON 安全类型"""
    if value is None:
        return None
    from datetime import date, datetime
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value
