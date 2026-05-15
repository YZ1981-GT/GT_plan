"""归档时自动生成 evidence_index.xlsx（证据索引表）

Sprint 6 Task 6.8:
  归档时调用 generate_evidence_index(db, project_id) 生成 Excel 文件，
  列：底稿编码 / 单元格 / 附件名 / 页码 / 证据类型 / 结论
"""

from __future__ import annotations

import io
import logging
from uuid import UUID

import sqlalchemy as sa
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.wp_optimization_models import EvidenceLink

logger = logging.getLogger(__name__)


async def generate_evidence_index(db: AsyncSession, project_id: UUID) -> bytes:
    """生成项目级证据索引 Excel 文件

    Returns:
        bytes: xlsx 文件内容
    """
    # 查询该项目所有底稿的证据链接（通过 working_paper.project_id 关联）
    q = sa.text("""
        SELECT
            wp.wp_code,
            el.sheet_name,
            el.cell_ref,
            a.file_name,
            el.page_ref,
            el.evidence_type,
            el.check_conclusion,
            el.created_at
        FROM evidence_links el
        JOIN working_paper wp ON wp.id = el.wp_id
        LEFT JOIN attachments a ON a.id = el.attachment_id
        WHERE wp.project_id = :pid
        ORDER BY wp.wp_code, el.sheet_name, el.cell_ref
    """)
    rows = (await db.execute(q, {"pid": str(project_id)})).fetchall()

    # 生成 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "证据索引"

    # 表头
    headers = ["底稿编码", "Sheet", "单元格", "附件名", "页码", "证据类型", "结论", "创建时间"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F0EDF5", end_color="F0EDF5", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row[0])  # wp_code
        ws.cell(row=row_idx, column=2, value=row[1])  # sheet_name
        ws.cell(row=row_idx, column=3, value=row[2])  # cell_ref
        ws.cell(row=row_idx, column=4, value=row[3])  # file_name
        ws.cell(row=row_idx, column=5, value=row[4])  # page_ref
        ws.cell(row=row_idx, column=6, value=row[5])  # evidence_type
        ws.cell(row=row_idx, column=7, value=row[6])  # check_conclusion
        ws.cell(row=row_idx, column=8, value=str(row[7]) if row[7] else "")  # created_at

    # 列宽
    col_widths = [14, 12, 8, 30, 10, 12, 30, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 输出 bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
