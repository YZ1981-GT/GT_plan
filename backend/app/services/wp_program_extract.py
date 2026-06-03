"""A 类程序表（a-program-console）程序行提取服务

从底稿 xlsx 模板的「审计程序表」sheet 解析出 ProgramRow 列表，结构对齐
GtAProgramConsole.vue 消费的 html_data.programs 形态。

动机（2026-06-02 修复）：render-config 对 a-program-console **从未**生成
programs（不像 b-index 有 _generate_b_index_data），导致程序表中控台永远显示
"暂无审计程序" —— 模板里 18 条审计程序内容完全没体现出来，无法裁剪/新增。

设计要点（与模板布局解耦，避免硬编码行号）：
- 表头行 = 列 A（或任一列）值为「序号」的行；
- 认定子表头行 = 表头行 + 1（存在/完整性/权利和义务/准确性/列报）；
- 程序描述列 = 表头含「审计程序」的列；分类列 = 含「分类」的列；
  底稿索引列 = 含「索引」的列；
- 5 项认定列 = 分类列与索引列之间的列，按子表头关键词映射认定键；
- 数据行 = 序号列为整数的行。

纯函数无 DB 副作用，便于单测。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 认定子表头关键词 → ProgramRow.assertions 键
_ASSERTION_KEYWORDS: list[tuple[str, str]] = [
    ("存在", "existence"),
    ("完整", "completeness"),
    ("权利", "rights"),
    ("准确", "accuracy"),
    ("计价", "accuracy"),
    ("列报", "presentation"),
]

# 5 项认定列（分类列与索引列之间无子表头时的定位兜底顺序）
_ASSERTION_ORDER = ["existence", "completeness", "rights", "accuracy", "presentation"]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_int_like(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value == int(value)
    if isinstance(value, str):
        return value.strip().isdigit()
    return False


def _map_assertion_key(header_text: str) -> str | None:
    for kw, key in _ASSERTION_KEYWORDS:
        if kw in header_text:
            return key
    return None


def extract_program_rows_from_sheet(ws: Any) -> list[dict]:
    """从 openpyxl worksheet 提取程序行列表。

    Returns:
        list[dict]: 每个元素结构对齐 GtAProgramConsole 的 ProgramRow：
        {
            "id": "row-{n}",
            "program_no": int,
            "program_desc": str,
            "program_category": str,
            "assertions": {existence/completeness/rights/accuracy/presentation: bool},
            "linked_workpapers": str,
            "status": "pending",
        }
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    # ─── 1. 定位表头行（含「序号」的行）+ 序号列 ────────────────────────
    header_row = None
    no_col = 1
    for r in range(1, min(max_row, 60) + 1):
        for c in range(1, max_col + 1):
            if _cell_text(ws.cell(row=r, column=c).value) == "序号":
                header_row = r
                no_col = c
                break
        if header_row:
            break

    if header_row is None:
        return []

    # ─── 2. 定位描述列 / 分类列 / 索引列 ──────────────────────────────
    desc_col = None
    category_col = None
    idx_col = None
    for c in range(1, max_col + 1):
        txt = _cell_text(ws.cell(row=header_row, column=c).value)
        if not txt:
            continue
        if desc_col is None and "审计程序" in txt:
            desc_col = c
        if category_col is None and "分类" in txt:
            category_col = c
        if idx_col is None and "索引" in txt:
            idx_col = c

    # 兜底：描述列默认序号列右一列；分类列默认描述列右一列
    if desc_col is None:
        desc_col = no_col + 1
    if category_col is None:
        category_col = desc_col + 1

    # ─── 3. 认定列映射（分类列与索引列之间）────────────────────────────
    sub_header_row = header_row + 1
    assertion_end = (idx_col - 1) if idx_col and idx_col > category_col else min(category_col + 5, max_col)
    assertion_cols: list[tuple[int, str]] = []
    fallback_idx = 0
    for c in range(category_col + 1, assertion_end + 1):
        sub_txt = _cell_text(ws.cell(row=sub_header_row, column=c).value)
        key = _map_assertion_key(sub_txt)
        if key is None and fallback_idx < len(_ASSERTION_ORDER):
            # 子表头无文字时按固定顺序兜底
            key = _ASSERTION_ORDER[fallback_idx]
        if key is not None:
            assertion_cols.append((c, key))
        fallback_idx += 1

    # ─── 4. 提取数据行（序号列为整数）────────────────────────────────
    programs: list[dict] = []
    # 数据起始行：跳过认定子表头行（若其序号列非整数）
    start_row = header_row + 1
    if not _is_int_like(ws.cell(row=start_row, column=no_col).value):
        start_row = header_row + 2

    for r in range(start_row, max_row + 1):
        no_val = ws.cell(row=r, column=no_col).value
        if not _is_int_like(no_val):
            continue

        program_no = int(float(no_val)) if not isinstance(no_val, str) else int(no_val.strip())
        desc = _cell_text(ws.cell(row=r, column=desc_col).value)
        category = _cell_text(ws.cell(row=r, column=category_col).value)

        assertions: dict[str, bool] = {}
        for col, key in assertion_cols:
            mark = _cell_text(ws.cell(row=r, column=col).value)
            if mark:
                assertions[key] = True

        linked = ""
        if idx_col:
            raw = _cell_text(ws.cell(row=r, column=idx_col).value)
            # 模板里索引常含换行，统一转 / 分隔
            linked = raw.replace("\n", "/").replace("//", "/").strip("/ ")

        programs.append({
            "id": f"row-{program_no}",
            "program_no": program_no,
            "program_desc": desc,
            "program_category": category,
            "assertions": assertions,
            "linked_workpapers": linked,
            "status": "pending",
        })

    return programs


def extract_program_rows(file_path: str | Path, sheet_name: str) -> list[dict]:
    """读取 xlsx 文件指定 sheet，提取程序行列表（纯函数，无 DB）。

    文件不存在 / 空 / sheet 缺失 / 解析失败 → 返回 []（降级，不抛异常）。
    """
    import openpyxl

    fp = Path(file_path)
    if not fp.exists() or fp.stat().st_size == 0:
        return []

    try:
        wb = openpyxl.load_workbook(str(fp), read_only=False, data_only=True)
    except Exception as e:
        logger.warning("extract_program_rows: 加载 xlsx 失败 %s: %s", fp, e)
        return []

    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return extract_program_rows_from_sheet(ws)
    except Exception as e:
        logger.warning("extract_program_rows: 解析 sheet 失败 %s/%s: %s", fp, sheet_name, e)
        return []
    finally:
        wb.close()
