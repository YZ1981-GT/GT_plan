"""审定表（audit-sheet）行项目提取服务

从底稿 xlsx 模板的「审定表」sheet（class_code=`F-审定表`，如 应收票据「审定表D1-1」）
解析出结构化的行项目列表 AuditSheetRow，供前端 GtAuditSheet.vue 渲染可编辑表格。

动机（spec `audit-sheet-editable`）：审定表从只读 GtGridSheet（HTML 死网格）升级为
结构化可编辑表，行项目需从模板动态解析——不同科目（应收票据/固定资产/无形资产）
自动展示对应的行结构（分节 / 明细 / 小计 / 合计），并支持层级缩进。

解析策略（与模板布局解耦，避免硬编码行号）：
- 定位表头行 = 项目列（通常 A 列）值为「项目」/「科目」的行；
- 跳过子表头行（未审数/账项调整/审定数 等——项目列为空）；
- 数据行 = 表头行之后、项目列非空的行，直至首个空行（截断 审计说明/结论 等说明区）；
- 缩进：优先取前导空格 / 全角「　」推断，无前导空格时按「分节内的明细行」结构兜底缩进 1 级；
- 分节行：项目名以中文数字序号「一、二、三…」开头 → isSection=True, bold=True；
- 合计行：项目名含「合计」/「小计」 → isComputed=True, bold=True。

输出对齐 design 的 AuditSheetRow 后端结构（TB 取数列不在此处填，见组④ Task 13）。

纯函数无 DB 副作用，便于单测。文件不存在 / 解析失败 → 返回 []（降级，不抛异常）。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 表头行项目列关键词（命中其一即认定为表头行）
# 覆盖审定表（"项目"/"科目"）+ 明细表（"票据种类"/"种类"/"类别"/"名称"/"客户名称"）
_HEADER_KEYWORDS = ("项目", "科目", "票据种类", "种类", "类别", "名称", "客户名称")

# 分节行：以中文数字序号开头（一、 二、 三．等），允许序号后接顿号/点/句点
_SECTION_PATTERN = re.compile(r"^[一二三四五六七八九十百零]+\s*[、.．)）]")

# 合计行关键词
_TOTAL_KEYWORDS = ("合计", "小计")

# 「审计说明」区标题关键词（数据表体之后）
_NOTES_HEADING_KEYWORDS = ("审计说明",)

# 「审计结论」区标题关键词
_CONCLUSION_HEADING_KEYWORDS = ("审计结论", "总体结论")

# 全角空格 / 半角空格 / 制表符（用于前导缩进推断）
_INDENT_CHARS = (" ", "\u3000", "\t")


def _cell_text(value: Any) -> str:
    """单元格值 → 去首尾空白的字符串（None → ''）。"""
    if value is None:
        return ""
    return str(value).strip()


def _is_section(item: str) -> bool:
    """是否分节标题行（中文数字序号开头）。"""
    return bool(_SECTION_PATTERN.match(item.strip()))


def _is_total(item: str) -> bool:
    """是否合计/小计行。排除"按组合计提"/"按单项计提"等误匹配。"""
    stripped = item.strip()
    # 排除已知误匹配模式
    if "计提" in stripped:
        return False
    return any(kw in stripped for kw in _TOTAL_KEYWORDS)


def _leading_indent(raw_item: str) -> int:
    """从前导空格 / 全角空格推断缩进层级。

    规则：全角「　」每个计 1 级；半角空格每 2 个计 1 级；制表符每个计 1 级。
    """
    full = 0
    half = 0
    tab = 0
    for ch in raw_item:
        if ch == "\u3000":
            full += 1
        elif ch == " ":
            half += 1
        elif ch == "\t":
            tab += 1
        else:
            break
    return full + tab + (half // 2)


def _locate_header(ws: Any, max_row: int, max_col: int) -> tuple[int | None, int]:
    """定位表头行（项目列值为「项目」/「科目」）+ 项目列号。

    Returns (header_row, item_col)；未找到表头则 header_row=None。
    """
    for r in range(1, min(max_row, 60) + 1):
        for c in range(1, min(max_col, 5) + 1):
            if _cell_text(ws.cell(row=r, column=c).value) in _HEADER_KEYWORDS:
                return r, c
    return None, 1


def extract_audit_rows_from_sheet(ws: Any) -> list[dict]:
    """从 openpyxl worksheet 提取审定表行项目列表。

    Returns:
        list[dict]: 每个元素对齐 design 的 AuditSheetRow 后端结构：
        {
            "id": "row-{n}",
            "item": str,                 # 项目名（去前导空格）
            "indent": int,               # 缩进层级（0/1/2…）
            "bold": bool,                # 是否粗体（分节/合计强制 True）
            "isSection": bool,           # 是否分节标题行（一、二、三）
            "isComputed": bool,          # 是否合计/小计行
            "account_code": None,        # 科目编码（TB 取数映射，组④ Task 13 填）
            "adj_amount": None,          # 账项调整（用户编辑，初始 None）
            "reclass_amount": None,      # 重分类调整（用户编辑，初始 None）
            "reason": "",                # 原因分析（用户编辑，初始空）
        }
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    # ─── 1. 定位表头行（项目列值为「项目」/「科目」）+ 项目列 ────────────
    header_row, item_col = _locate_header(ws, max_row, max_col)

    if header_row is None:
        return []

    # ─── 2. 定位数据起始行（跳过项目列为空的子表头行）────────────────────
    data_start = header_row + 1
    while data_start <= max_row and not _cell_text(ws.cell(row=data_start, column=item_col).value):
        data_start += 1

    # ─── 3. 逐行提取（项目列非空，遇首个空行截断说明/结论区）──────────────
    rows: list[dict] = []
    section_active = False
    n = 0
    for r in range(data_start, max_row + 1):
        cell = ws.cell(row=r, column=item_col)
        raw = "" if cell.value is None else str(cell.value)
        item = raw.strip()

        if not item:
            if rows:
                break  # 已收集到行 → 首个空行视为表体结束（截断说明区）
            continue

        # 遇到审计说明/审计结论标题 → 表体结束（明细表无空行分隔直接接说明区）
        if any(kw in item for kw in _NOTES_HEADING_KEYWORDS + _CONCLUSION_HEADING_KEYWORDS):
            break

        is_section = _is_section(item)
        is_total = _is_total(item)

        # 缩进：优先前导空格推断；无前导缩进时，分节内明细行结构兜底缩进 1 级
        indent = _leading_indent(raw)
        if indent == 0 and not is_section and not is_total and section_active:
            indent = 1
        if is_section:
            section_active = True

        try:
            font_bold = bool(cell.font and cell.font.bold)
        except Exception:
            font_bold = False

        n += 1
        rows.append({
            "id": f"row-{n}",
            "item": item,
            "indent": indent,
            "bold": font_bold or is_section or is_total,
            "isSection": is_section,
            "isComputed": is_total,
            "account_code": None,
            "adj_amount": None,
            "reclass_amount": None,
            "reason": "",
        })

    return rows


def extract_audit_sheet_columns(ws: Any) -> list[dict] | None:
    """从 worksheet 提取数据表的列定义（表头行的所有列标题）。

    当列数 > 标准审定表的列数（期初/期末/调整/审定/变动/变动率/原因 共 9 列）时，
    返回完整列定义供前端动态渲染。列数 ≤ 标准审定表时返回 None（用默认布局）。

    多行表头处理：如果主表头行（如"项目 | 期初余额 | 本期增加..."）只有分组标题，
    但下一行有细分列标题（如"期初未审数 | 账项调整 | 重分类调整..."），取细分行。

    Returns:
        list[dict]: [{"key": "col_2", "label": "期初未审数", "col_idx": 2}, ...]
        或 None（标准审定表布局，走默认）
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return None

    header_row, item_col = _locate_header(ws, max_row, max_col)
    if header_row is None:
        return None

    # 提取表头行所有非空列标题（跳过项目列本身）
    columns: list[dict] = []
    for c in range(1, max_col + 1):
        if c == item_col:
            continue
        label = _cell_text(ws.cell(row=header_row, column=c).value)
        if not label:
            continue
        columns.append({
            "key": f"col_{c}",
            "label": label,
            "col_idx": c,
        })

    # 多行表头：如果主表头列数少（≤8），检查下一行是否有更多细分列标题
    if len(columns) <= 8 and header_row + 1 <= max_row:
        sub_columns: list[dict] = []
        for c in range(1, max_col + 1):
            if c == item_col:
                continue
            label = _cell_text(ws.cell(row=header_row + 1, column=c).value)
            if not label:
                continue
            sub_columns.append({
                "key": f"col_{c}",
                "label": label,
                "col_idx": c,
            })
        # 子表头更详细（列数更多）→ 用子表头
        if len(sub_columns) > len(columns):
            columns = sub_columns

    # 标准审定表 ≤ 8 数据列（不含项目列），此时走默认布局
    if len(columns) <= 8:
        return None

    return columns


def extract_audit_rows_with_values(ws: Any) -> tuple[list[dict], list[dict] | None]:
    """从 worksheet 提取审定表/明细表的行项目 + 列定义 + 初始值。

    对于标准审定表（≤8 数据列），返回 (rows, None) — 走默认布局。
    对于多列明细表（如 D1-2 有 10 数据列），返回 (rows, column_defs)，
    其中 rows 的每行额外包含 col_{n} 字段携带模板初始值。

    Returns:
        (rows, column_defs): rows 为行列表，column_defs 为列定义或 None
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return [], None

    header_row, item_col = _locate_header(ws, max_row, max_col)
    if header_row is None:
        return [], None

    # 列定义
    col_defs = extract_audit_sheet_columns(ws)

    # 数据起始行（多列明细表：不跳过空行，因为可能是动态行占位）
    data_start = header_row + 1
    if not col_defs:
        # 标准审定表：跳过项目列为空的子表头行
        while data_start <= max_row and not _cell_text(ws.cell(row=data_start, column=item_col).value):
            data_start += 1

    # 逐行提取
    rows: list[dict] = []
    section_active = False
    n = 0
    consecutive_empty = 0
    for r in range(data_start, max_row + 1):
        cell = ws.cell(row=r, column=item_col)
        raw = "" if cell.value is None else str(cell.value)
        item = raw.strip()

        if not item:
            if col_defs:
                # 多列明细表：空行可能是动态行占位，检查该行其他列是否有值
                has_other_data = any(
                    ws.cell(row=r, column=cd["col_idx"]).value is not None
                    for cd in col_defs
                )
                if has_other_data:
                    # 有数据的占位行 → 保留为可编辑空行（最多预留 6 行）
                    blank_count = sum(1 for row in rows if row.get("isCustom") and not row.get("item"))
                    if blank_count >= 6:
                        continue  # 已有 6 行空白占位，跳过多余的
                    n += 1
                    row_data: dict = {
                        "id": f"row-{n}", "item": "", "indent": 0, "bold": False,
                        "isSection": False, "isComputed": False, "isCustom": True,
                        "account_code": None, "adj_amount": None,
                        "reclass_amount": None, "reason": "",
                    }
                    for cd in col_defs:
                        val = ws.cell(row=r, column=cd["col_idx"]).value
                        if val is not None:
                            try:
                                row_data[cd["key"]] = float(val) if isinstance(val, (int, float)) else val
                            except (TypeError, ValueError):
                                row_data[cd["key"]] = val
                    rows.append(row_data)
                    consecutive_empty = 0
                    continue
                # 完全空行
                consecutive_empty += 1
                if consecutive_empty >= 3 and rows:
                    break  # 连续3行全空 → 表体结束
                continue
            else:
                if rows:
                    break  # 标准审定表：首个空行截断
                continue

        consecutive_empty = 0

        # 遇到审计说明/审计结论标题 → 表体结束
        if any(kw in item for kw in _NOTES_HEADING_KEYWORDS + _CONCLUSION_HEADING_KEYWORDS):
            break

        is_section = _is_section(item)
        is_total = _is_total(item)

        indent = _leading_indent(raw)
        if indent == 0 and not is_section and not is_total and section_active:
            indent = 1
        if is_section:
            section_active = True

        try:
            font_bold = bool(cell.font and cell.font.bold)
        except Exception:
            font_bold = False

        n += 1
        row: dict = {
            "id": f"row-{n}",
            "item": item,
            "indent": indent,
            "bold": font_bold or is_section or is_total,
            "isSection": is_section,
            "isComputed": is_total,
            "account_code": None,
            "adj_amount": None,
            "reclass_amount": None,
            "reason": "",
        }

        # 多列明细表：提取每个数据列的初始值
        if col_defs:
            for cd in col_defs:
                val = ws.cell(row=r, column=cd["col_idx"]).value
                if val is not None:
                    try:
                        row[cd["key"]] = float(val) if isinstance(val, (int, float)) else val
                    except (TypeError, ValueError):
                        row[cd["key"]] = val

        rows.append(row)

    return rows, col_defs


def extract_audit_sections_from_sheet(ws: Any) -> dict:
    """从 worksheet 提取「审计说明」与「审计结论」区文本（数据表体之后的说明段落）。

    审定表标准版式（致同/通用）数据网格之后有两块说明区：
      - 「1.审计说明」：记录重要审计发现/判断，主要对变动大（如变动率>30%）的科目
        说明原因，含质押/贴现等特殊事项说明 —— 多行子项；
      - 「2.审计结论」：给出明确审计结论（是否认可被审计单位列报金额）。

    解析策略（布局解耦，不硬编码行号）：
      - 定位「审计说明」标题行 → 其后到「审计结论」标题行之间的非空文本行 = notes；
      - 「审计结论」标题行之后的非空文本行 = conclusion；
      - 标题行自身（如「1.审计说明」「2.审计结论」）不计入内容。

    Returns:
        {
            "notes": str,        # 审计说明正文（多行以 \\n 连接，无则空串）
            "conclusion": str,   # 审计结论正文（无则空串）
            "notes_label": str,  # 「审计说明」标题原文（用于前端区块标题，默认「审计说明」）
            "conclusion_label": str,  # 「审计结论」标题原文（默认「审计结论」）
        }
    """
    empty = {
        "notes": "",
        "conclusion": "",
        "notes_label": "审计说明",
        "conclusion_label": "审计结论",
    }
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return empty

    def _is_heading(text: str, keywords: tuple[str, ...]) -> bool:
        # 命中关键词且文本较短（标题行，非正文）——避免正文里偶含「审计结论」误判
        return any(kw in text for kw in keywords) and len(text) <= 20

    # 行文本 = 该行所有非空单元格拼接（跨列标题/正文都收）
    def _row_text(r: int) -> str:
        parts: list[str] = []
        for c in range(1, max_col + 1):
            t = _cell_text(ws.cell(row=r, column=c).value)
            if t:
                parts.append(t)
        return "  ".join(parts).strip()

    notes_row: int | None = None
    conclusion_row: int | None = None
    notes_label = "审计说明"
    conclusion_label = "审计结论"
    for r in range(1, max_row + 1):
        first = _cell_text(ws.cell(row=r, column=1).value)
        if not first:
            continue
        if notes_row is None and _is_heading(first, _NOTES_HEADING_KEYWORDS):
            notes_row = r
            notes_label = re.sub(r"^[0-9.、\s]+", "", first) or "审计说明"
        elif _is_heading(first, _CONCLUSION_HEADING_KEYWORDS):
            conclusion_row = r
            conclusion_label = re.sub(r"^[0-9.、\s]+", "", first) or "审计结论"

    if notes_row is None and conclusion_row is None:
        return empty

    notes_lines: list[str] = []
    conclusion_lines: list[str] = []
    if notes_row is not None:
        end = conclusion_row if (conclusion_row and conclusion_row > notes_row) else (max_row + 1)
        for r in range(notes_row + 1, end):
            txt = _row_text(r)
            if txt:
                notes_lines.append(txt)
    if conclusion_row is not None:
        for r in range(conclusion_row + 1, max_row + 1):
            txt = _row_text(r)
            if txt:
                conclusion_lines.append(txt)

    return {
        "notes": "\n".join(notes_lines),
        "conclusion": "\n".join(conclusion_lines),
        "notes_label": notes_label,
        "conclusion_label": conclusion_label,
    }


def _locate_header_grid(grid: list[list[Any]], max_row: int, max_col: int) -> tuple[int | None, int]:
    """定位表头行（项目列值为 _HEADER_KEYWORDS）— grid 版本（1-based 坐标）。"""
    for r in range(1, min(max_row, 60) + 1):
        for c in range(1, min(max_col, 5) + 1):
            r_idx, c_idx = r - 1, c - 1
            if r_idx >= len(grid):
                continue
            row_data = grid[r_idx] or []
            val = row_data[c_idx] if c_idx < len(row_data) else None
            if _cell_text(val) in _HEADER_KEYWORDS:
                return r, c
    return None, 1


def extract_audit_rows_from_grid(grid: list[list[Any]]) -> list[dict]:
    """从二维值数组提取审定表行项目列表（calamine / openpyxl 共用）。

    与 extract_audit_rows_from_sheet 逻辑一致，但不依赖 openpyxl worksheet。
    注意：font.bold 信息丢失（grid 无样式），改用 is_section/is_total 推断 bold。
    """
    if not grid:
        return []

    max_row = len(grid)
    max_col = max((len(r or []) for r in grid), default=0)
    if max_row == 0 or max_col == 0:
        return []

    def _grid_val(r: int, c: int) -> Any:
        """1-based row/col access."""
        ri, ci = r - 1, c - 1
        if ri >= len(grid):
            return None
        row_data = grid[ri] or []
        return row_data[ci] if ci < len(row_data) else None

    header_row, item_col = _locate_header_grid(grid, max_row, max_col)
    if header_row is None:
        return []

    data_start = header_row + 1
    while data_start <= max_row and not _cell_text(_grid_val(data_start, item_col)):
        data_start += 1

    rows: list[dict] = []
    section_active = False
    n = 0
    for r in range(data_start, max_row + 1):
        raw_val = _grid_val(r, item_col)
        raw = "" if raw_val is None else str(raw_val)
        item = raw.strip()

        if not item:
            if rows:
                break
            continue

        if any(kw in item for kw in _NOTES_HEADING_KEYWORDS + _CONCLUSION_HEADING_KEYWORDS):
            break

        is_section = _is_section(item)
        is_total = _is_total(item)

        indent = _leading_indent(raw)
        if indent == 0 and not is_section and not is_total and section_active:
            indent = 1
        if is_section:
            section_active = True

        n += 1
        rows.append({
            "id": f"row-{n}",
            "item": item,
            "indent": indent,
            "bold": is_section or is_total,
            "isSection": is_section,
            "isComputed": is_total,
            "account_code": None,
            "adj_amount": None,
            "reclass_amount": None,
            "reason": "",
        })

    return rows


def extract_audit_sections_from_grid(grid: list[list[Any]]) -> dict:
    """从二维值数组提取「审计说明 / 审计结论」区文本（grid 版本）。"""
    empty = {
        "notes": "",
        "conclusion": "",
        "notes_label": "审计说明",
        "conclusion_label": "审计结论",
    }
    if not grid:
        return empty

    max_row = len(grid)
    max_col = max((len(r or []) for r in grid), default=0)
    if max_row == 0 or max_col == 0:
        return empty

    def _grid_val(r: int, c: int) -> Any:
        ri, ci = r - 1, c - 1
        if ri >= len(grid):
            return None
        row_data = grid[ri] or []
        return row_data[ci] if ci < len(row_data) else None

    def _is_heading(text: str, keywords: tuple[str, ...]) -> bool:
        return any(kw in text for kw in keywords) and len(text) <= 20

    def _row_text(r: int) -> str:
        parts: list[str] = []
        for c in range(1, max_col + 1):
            t = _cell_text(_grid_val(r, c))
            if t:
                parts.append(t)
        return "  ".join(parts).strip()

    notes_row: int | None = None
    conclusion_row: int | None = None
    notes_label = "审计说明"
    conclusion_label = "审计结论"
    for r in range(1, max_row + 1):
        first = _cell_text(_grid_val(r, 1))
        if not first:
            continue
        if notes_row is None and _is_heading(first, _NOTES_HEADING_KEYWORDS):
            notes_row = r
            notes_label = re.sub(r"^[0-9.、\s]+", "", first) or "审计说明"
        elif _is_heading(first, _CONCLUSION_HEADING_KEYWORDS):
            conclusion_row = r
            conclusion_label = re.sub(r"^[0-9.、\s]+", "", first) or "审计结论"

    if notes_row is None and conclusion_row is None:
        return empty

    notes_lines: list[str] = []
    conclusion_lines: list[str] = []
    if notes_row is not None:
        end = conclusion_row if (conclusion_row and conclusion_row > notes_row) else (max_row + 1)
        for r in range(notes_row + 1, end):
            txt = _row_text(r)
            if txt:
                notes_lines.append(txt)
    if conclusion_row is not None:
        for r in range(conclusion_row + 1, max_row + 1):
            txt = _row_text(r)
            if txt:
                conclusion_lines.append(txt)

    return {
        "notes": "\n".join(notes_lines),
        "conclusion": "\n".join(conclusion_lines),
        "notes_label": notes_label,
        "conclusion_label": conclusion_label,
    }


def extract_audit_rows(file_path: str | Path, sheet_name: str) -> list[dict]:
    """读取 xlsx 文件指定 sheet，提取审定表行项目列表（纯函数，无 DB）。

    文件不存在 / 空 / sheet 缺失 / 解析失败 → 返回 []（降级，不抛异常）。
    使用 read_sheet_values 统一适配器（calamine/openpyxl 自动切换）。
    """
    from app.services.xlsx_read_adapter import read_sheet_values, list_sheet_names

    fp = Path(file_path)
    if not fp.exists() or fp.stat().st_size == 0:
        return []

    try:
        names = list_sheet_names(fp)
        if sheet_name not in names:
            return []
        grid = read_sheet_values(fp, sheet_name)
        return extract_audit_rows_from_grid(grid)
    except Exception as e:
        logger.warning("extract_audit_rows: 解析 sheet 失败 %s/%s: %s", fp, sheet_name, e)
        return []


def extract_audit_sections(file_path: str | Path, sheet_name: str) -> dict:
    """读取 xlsx 指定 sheet，提取「审计说明 / 审计结论」区文本（纯函数，无 DB）。

    文件不存在 / 空 / sheet 缺失 / 解析失败 → 返回空区块（降级，不抛异常）。
    使用 read_sheet_values 统一适配器（calamine/openpyxl 自动切换）。
    """
    from app.services.xlsx_read_adapter import read_sheet_values, list_sheet_names

    empty = {
        "notes": "",
        "conclusion": "",
        "notes_label": "审计说明",
        "conclusion_label": "审计结论",
    }
    fp = Path(file_path)
    if not fp.exists() or fp.stat().st_size == 0:
        return empty

    try:
        names = list_sheet_names(fp)
        if sheet_name not in names:
            return empty
        grid = read_sheet_values(fp, sheet_name)
        return extract_audit_sections_from_grid(grid)
    except Exception as e:
        logger.warning("extract_audit_sections: 解析 sheet 失败 %s/%s: %s", fp, sheet_name, e)
        return empty


def extract_audit_rows_with_values_from_file(
    file_path: str | Path, sheet_name: str
) -> tuple[list[dict], list[dict] | None]:
    """读取 xlsx 指定 sheet，提取行项目 + 列定义 + 初始值（文件级包装）。

    标准审定表（≤8 数据列）→ (rows, None)；多列明细表 → (rows, column_defs)。
    文件不存在 / 空 / sheet 缺失 / 解析失败 → ([], None)（降级，不抛异常）。

    注意：此函数仍使用 openpyxl（需要 ws.cell 访问模式用于 extract_audit_rows_with_values
    的多列值提取）。calamine 切换仅应用于纯行提取的 extract_audit_rows / extract_audit_sections。
    """
    import openpyxl

    fp = Path(file_path)
    if not fp.exists() or fp.stat().st_size == 0:
        return [], None

    try:
        wb = openpyxl.load_workbook(str(fp), read_only=False, data_only=True)
    except Exception as e:
        logger.warning("extract_audit_rows_with_values: 加载 xlsx 失败 %s: %s", fp, e)
        return [], None

    try:
        if sheet_name not in wb.sheetnames:
            return [], None
        ws = wb[sheet_name]
        return extract_audit_rows_with_values(ws)
    except Exception as e:
        logger.warning(
            "extract_audit_rows_with_values: 解析 sheet 失败 %s/%s: %s",
            fp, sheet_name, e,
        )
        return [], None
    finally:
        wb.close()
