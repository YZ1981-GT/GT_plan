"""统一 xlsx 只读取值入口（calamine / openpyxl 可切换）。

xlsx-read-acceleration spec：底稿模板纯取值提取点共用此模块。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from app.core.config import settings

logger = logging.getLogger(__name__)

_calamine_available_cache: bool | None = None


def _calamine_available() -> bool:
    global _calamine_available_cache
    if _calamine_available_cache is not None:
        return _calamine_available_cache
    try:
        import python_calamine  # noqa: F401
    except ImportError:
        _calamine_available_cache = False
    else:
        _calamine_available_cache = True
    return _calamine_available_cache


def _normalize_cell(value: Any) -> Any:
    if value == "":
        return None
    return value


def _normalize_rows(rows: list[list[Any]]) -> list[list[Any]]:
    return [[_normalize_cell(c) for c in (row or [])] for row in rows]


def _read_openpyxl(path: str | Path, sheet_name: str | None) -> list[list[Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
        else:
            ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_calamine(path: str | Path, sheet_name: str | None) -> list[list[Any]]:
    from app.services.ledger_import.parsers.excel_parser_calamine import _load_sheet
    from python_calamine import CalamineWorkbook

    path_str = str(path)
    if sheet_name is None:
        wb = CalamineWorkbook.from_path(path_str)
        sheet_name = wb.sheet_names[0] if wb.sheet_names else ""
        if not sheet_name:
            return []
    return _load_sheet(path_str, sheet_name)


def read_sheet_values(
    path: str | Path,
    sheet_name: str | None = None,
    *,
    prefer_calamine: bool = True,
) -> list[list[Any]]:
    """读取 sheet 二维值数组（不含样式/公式）。"""
    fp = Path(path)
    if not fp.exists() or fp.stat().st_size == 0:
        return []

    use_calamine = (
        prefer_calamine
        and settings.XLSX_READ_USE_CALAMINE
        and _calamine_available()
    )

    try:
        if use_calamine:
            rows = _read_calamine(fp, sheet_name)
        else:
            rows = _read_openpyxl(fp, sheet_name)
    except Exception as exc:
        if use_calamine:
            logger.warning(
                "read_sheet_values calamine 失败，降级 openpyxl: %s/%s: %s",
                fp,
                sheet_name,
                exc,
            )
            rows = _read_openpyxl(fp, sheet_name)
        else:
            logger.warning(
                "read_sheet_values openpyxl 失败: %s/%s: %s",
                fp,
                sheet_name,
                exc,
            )
            return []

    return _normalize_rows(rows)


def list_sheet_names(path: str | Path) -> list[str]:
    fp = Path(path)
    if not fp.exists():
        return []
    if settings.XLSX_READ_USE_CALAMINE and _calamine_available():
        try:
            from python_calamine import CalamineWorkbook

            return list(CalamineWorkbook.from_path(str(fp)).sheet_names)
        except Exception as exc:
            logger.warning("list_sheet_names calamine 失败，降级 openpyxl: %s", exc)
    import openpyxl

    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
