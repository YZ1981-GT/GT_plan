"""底稿 xlsx 导出服务 — 致同模板填值还原

按 design §6.2 实现 4 路径写入策略：
  1. fixed_cells: 写 schema 默认值（项目元数据 + 索引号）
  2. dynamic_table: html_data[sheet][rows] 按 columns 映射写入
  3. formulas: 跳过（保留原公式 → 用户打开 xlsx 自动重算）
  4. static_text + merged_cells: openpyxl 加载时已含，跳过

Requirements: 4.3.1.a-g（方案 C 还原 7 项约束）
"""

from __future__ import annotations

import asyncio
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# 模板根目录
_TEMPLATES_ROOT = Path(__file__).resolve().parent.parent.parent / "wp_templates"

# 并发信号量：限制同时进行的 openpyxl 操作数（design §9.6）
_EXPORT_SEMAPHORE = asyncio.Semaphore(10)


class TemplateNotFoundError(Exception):
    """模板 xlsx 文件不存在"""
    pass


class ExportValidationError(Exception):
    """导出时必填字段缺失"""

    def __init__(self, missing_fields: list[str]):
        self.missing_fields = missing_fields
        super().__init__(f"Missing required fields: {missing_fields}")


def _resolve_template_path(schema: dict) -> Path:
    """从 schema 解析模板文件路径。

    schema['template_path'] 格式：
      'backend/wp_templates/D/D2-1至D2-4  应收账款- 审定表明细表（Leap-常规程序）.xlsx'

    Returns:
        模板文件的绝对路径

    Raises:
        TemplateNotFoundError: 模板文件不存在
    """
    template_path_str = schema.get("template_path", "")
    if not template_path_str:
        raise TemplateNotFoundError(
            f"Schema missing 'template_path' for wp_code={schema.get('wp_code', '?')}"
        )

    # template_path 可能是相对于项目根的路径
    # 尝试从项目根解析
    project_root = Path(__file__).resolve().parent.parent.parent.parent
    candidate = project_root / template_path_str
    if candidate.is_file():
        return candidate

    # 也尝试直接从 _TEMPLATES_ROOT 解析（去掉 'backend/wp_templates/' 前缀）
    stripped = template_path_str.replace("backend/wp_templates/", "")
    candidate2 = _TEMPLATES_ROOT / stripped
    if candidate2.is_file():
        return candidate2

    raise TemplateNotFoundError(
        f"Template not found: '{template_path_str}' "
        f"(searched: {candidate}, {candidate2})"
    )


def _render_template_var(
    value_template: str,
    project_meta: dict[str, Any],
) -> str:
    """替换占位符变量 ${var_name} 为实际值。

    支持的变量：
      - ${entity_name}: project_meta['entity_name'] 或 project_meta['client_name']
      - ${period_end}: project_meta['period_end']
      - ${index_no}: project_meta['index_no']
      - ${page_no}: project_meta['page_no']
      - 其他 ${xxx}: 从 project_meta[xxx] 取值

    如果 value_template 不含 ${...}，原样返回。
    """
    if "${" not in value_template:
        return value_template

    def _replace(match: re.Match) -> str:
        var_name = match.group(1)
        # 别名映射
        if var_name == "entity_name":
            return str(
                project_meta.get("entity_name")
                or project_meta.get("client_name", "")
            )
        return str(project_meta.get(var_name, ""))

    return re.sub(r"\$\{(\w+)\}", _replace, value_template)


def _get_nested(data: dict, field_path: str) -> Any:
    """按点分路径从 dict 取值。

    例如 'assertion.existence' → data['assertion']['existence']
    """
    parts = field_path.split(".")
    current: Any = data
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    return current


def _col_letter_to_index(col_letter: str) -> int:
    """列字母转 1-indexed 数字（A=1, B=2, ...）"""
    return column_index_from_string(col_letter)


def _write_dynamic_table(
    ws: Worksheet,
    table_schema: dict,
    rows: list[dict],
) -> None:
    """写入动态表格区域。

    按 schema 中 columns 定义，将 rows 数据逐行写入 worksheet。
    公式列跳过（保留原 formulas）。
    """
    start_row = table_schema.get("start_row", 1)
    columns = table_schema.get("columns", {})

    for i, row_data in enumerate(rows):
        excel_row = start_row + i
        for col_letter, col_def in columns.items():
            # col_def 可能是 dict 或 string
            if isinstance(col_def, str):
                col_def = {"field": col_def, "type": "text"}

            cell_ref = f"{col_letter}{excel_row}"
            cell = ws[cell_ref]

            # 路径 3: 公式列跳过（保留原公式）
            if (
                cell.value is not None
                and isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                continue

            field_path = col_def.get("field", "")
            if not field_path:
                continue

            value = _get_nested(row_data, field_path)

            # 类型转换
            render_type = col_def.get("render", "")
            col_type = col_def.get("type", "text")

            if render_type == "checkmark":
                cell.value = "√" if value else ""
            elif col_type == "number":
                if value is not None and value != "":
                    try:
                        cell.value = float(value)
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = None
            elif col_type == "boolean":
                cell.value = "√" if value else ""
            else:
                cell.value = value if value is not None else ""


def _sync_export_workpaper_xlsx(
    schema: dict,
    html_data: dict[str, Any],
    project_meta: dict[str, Any],
) -> BytesIO:
    """同步导出函数（在 run_in_executor 中调用）。

    加载致同模板 xlsx → 按 schema 4 路径写入 → 返回 BytesIO。

    Args:
        schema: wp_render_schema YAML 解析后的 dict
        html_data: working_paper.parsed_data['html_data'] 中对应数据
        project_meta: 项目元数据（entity_name, period_end, index_no, page_no 等）

    Returns:
        BytesIO 包含导出的 xlsx 数据

    Raises:
        TemplateNotFoundError: 模板文件不存在
    """
    # 1. 解析模板路径并加载
    template_path = _resolve_template_path(schema)
    logger.debug("Loading template: %s", template_path)

    # 加载模板（NOT read_only — 需要写入；data_only=False — 保留公式）
    wb = openpyxl.load_workbook(
        str(template_path),
        data_only=False,
        keep_vba=False,
    )

    # 2. 遍历 schema 中定义的 sheets
    sheets_schema = schema.get("sheets", {})
    for sheet_name, sheet_schema in sheets_schema.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(
                "Sheet '%s' not found in template, skipping", sheet_name
            )
            continue

        ws = wb[sheet_name]
        sheet_data = html_data.get(sheet_name, {})

        # 路径 1: fixed_cells — 写入 schema 默认值（项目元数据 + 索引号）
        fixed_cells = sheet_schema.get("fixed_cells", {})
        for cell_ref, value_template in fixed_cells.items():
            rendered_value = _render_template_var(
                str(value_template), project_meta
            )
            ws[cell_ref] = rendered_value

        # 路径 2: dynamic_table — 写入用户数据行
        if "dynamic_table" in sheet_schema:
            dynamic_rows = sheet_data.get("rows", [])
            # 也支持 sheet_data 直接作为 rows list 的情况
            if isinstance(sheet_data, list):
                dynamic_rows = sheet_data
            _write_dynamic_table(
                ws,
                sheet_schema["dynamic_table"],
                dynamic_rows,
            )

        # 路径 3: formulas — 跳过（openpyxl 加载时已保留原公式）
        # 路径 4: static_text + merged_cells — 跳过（openpyxl 加载时已含）

        # 兜底：确保 worksheet 维度正确
        try:
            ws.calculate_dimension(force=True)
        except Exception:
            pass  # 某些 sheet 可能无法计算维度，忽略

    # 3. 保存到 BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return buf


async def export_workpaper_xlsx(
    wp_code: str,
    html_data: dict[str, Any],
    schema: dict,
    project_meta: dict[str, Any],
) -> BytesIO:
    """异步导出底稿 xlsx（致同模板填值还原）。

    使用 asyncio.run_in_executor 包装同步 openpyxl 操作，
    配合 Semaphore(10) 限制并发数（design §9.6，6000 并发安全）。

    Args:
        wp_code: 底稿编码（如 'D2A'）
        html_data: 用户填写的 HTML 数据（按 sheet_name 分组）
        schema: wp_render_schema YAML 解析后的 dict
        project_meta: 项目元数据 dict，包含：
            - entity_name: 被审计单位名称
            - period_end: 审计截止日
            - index_no: 索引号
            - page_no: 页码（默认 '1/1'）

    Returns:
        BytesIO 包含导出的 xlsx 数据

    Raises:
        TemplateNotFoundError: 模板 xlsx 文件不存在
        ExportValidationError: 必填字段缺失
    """
    # 验证必填字段
    missing = []
    if not project_meta.get("entity_name") and not project_meta.get("client_name"):
        missing.append("entity_name")
    if not project_meta.get("period_end"):
        missing.append("period_end")
    if missing:
        raise ExportValidationError(missing)

    # 使用信号量限制并发 openpyxl 操作
    async with _EXPORT_SEMAPHORE:
        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(
            None,
            _sync_export_workpaper_xlsx,
            schema,
            html_data,
            project_meta,
        )

    return result
