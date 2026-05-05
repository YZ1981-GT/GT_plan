"""增强审计日志服务 — 真实落库 + 哈希链 + 异步队列缓冲

Refinement Round 1 — 需求 9：审计日志真实落库 + 不可篡改。

架构：
- log_action 推入 Redis List（或降级 asyncio.Queue），立即返回（P95 < 50ms）
- audit_log_writer_worker 批量消费队列 → 计算哈希链 → 批量 INSERT
- 内存 _recent_actions 保留最近 1000 条作为查询加速缓存
- entry_hash = sha256(ts|user_id|action_type|object_id|payload_json|prev_hash)
- prev_hash 按 project_id 分链（跨项目不链）
"""

from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import logging
import time
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any
from uuid import UUID

logger = logging.getLogger(__name__)

# Redis queue key for audit log entries
AUDIT_LOG_QUEUE_KEY = "audit:log:queue"
AUDIT_LOG_RETRY_QUEUE_KEY = "audit:log:retry"

# Anomaly detection thresholds
ANOMALY_THRESHOLDS = {
    "bulk_download": {"count": 10, "window_seconds": 300},
    "off_hours_operation": {"start_hour": 22, "end_hour": 6},
    "rapid_export": {"count": 5, "window_seconds": 60},
}


def _compute_entry_hash(
    ts: str,
    user_id: str,
    action_type: str,
    object_id: str | None,
    payload_json: str,
    prev_hash: str,
) -> str:
    """计算审计日志条目哈希。

    entry_hash = sha256(ts|user_id|action_type|object_id|payload_json|prev_hash)
    """
    raw = f"{ts}|{user_id}|{action_type}|{object_id or ''}|{payload_json}|{prev_hash}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


class AuditLoggerEnhanced:
    """增强审计日志服务 — 异步队列 + 批量落库 + 哈希链。"""

    # 创世哈希（每个 project_id 链的第一条 prev_hash）
    GENESIS_HASH = "0" * 64

    def __init__(self):
        self._recent_actions: list[dict] = []
        self._anomaly_alerts: list[dict] = []
        self._fallback_queue: asyncio.Queue | None = None
        self._redis_available: bool | None = None  # None = 未检测

    async def _get_redis(self):
        """尝试获取 Redis 客户端，不可用返回 None。"""
        try:
            from app.core.redis import redis_client
            await asyncio.wait_for(redis_client.ping(), timeout=0.3)
            self._redis_available = True
            return redis_client
        except Exception:
            if self._redis_available is not False:
                logger.warning("[audit_logger] Redis 不可用，降级到 asyncio.Queue")
            self._redis_available = False
            return None

    def _get_fallback_queue(self) -> asyncio.Queue:
        """获取进程内降级队列（Redis 不可用时使用）。"""
        if self._fallback_queue is None:
            self._fallback_queue = asyncio.Queue(maxsize=50000)
        return self._fallback_queue

    async def log_action(
        self,
        user_id: UUID | str,
        action: str,
        object_type: str,
        object_id: UUID | str | None = None,
        project_id: UUID | str | None = None,
        details: dict | None = None,
        ip_address: str | None = None,
        request_id: str | None = None,
        session_id: str | None = None,
        ua: str | None = None,
    ) -> dict:
        """记录操作日志 — 推入队列立即返回（P95 < 50ms）。

        日志条目推入 Redis List 或降级 asyncio.Queue，
        由 audit_log_writer_worker 批量消费落库。
        """
        now = datetime.now(timezone.utc)
        entry = {
            "user_id": str(user_id) if user_id else None,
            "action_type": action,
            "object_type": object_type,
            "object_id": str(object_id) if object_id else None,
            "project_id": str(project_id) if project_id else None,
            "payload": details or {},
            "ip": ip_address,
            "ua": ua,
            "trace_id": request_id,
            "session_id": session_id,
            "ts": now.isoformat(),
            "timestamp": time.time(),  # 兼容旧查询接口
        }

        # 内存缓存（查询加速，最近 1000 条）
        self._recent_actions.append(entry)
        if len(self._recent_actions) > 1000:
            self._recent_actions = self._recent_actions[-1000:]

        # 推入队列（非阻塞）
        try:
            redis = await self._get_redis()
            if redis:
                await redis.lpush(
                    AUDIT_LOG_QUEUE_KEY,
                    json.dumps(entry, ensure_ascii=False, default=str),
                )
            else:
                # 降级到进程内队列
                queue = self._get_fallback_queue()
                try:
                    queue.put_nowait(entry)
                except asyncio.QueueFull:
                    logger.error("[audit_logger] 降级队列已满，日志丢失: %s", entry.get("action_type"))
        except Exception as e:
            logger.error("[audit_logger] 推入队列失败: %s", e)

        # 异常检测（保留原有逻辑）
        self._check_anomalies(entry)

        return entry

    def _check_anomalies(self, entry: dict):
        """检测异常操作。"""
        user_id = entry.get("user_id", "")
        action = entry.get("action_type", "")
        now = entry.get("timestamp", time.time())

        if action in ("download", "export", "batch_download"):
            recent = [
                a for a in self._recent_actions
                if a.get("user_id") == user_id
                and a.get("action_type") in ("download", "export", "batch_download")
                and now - a.get("timestamp", 0) < ANOMALY_THRESHOLDS["bulk_download"]["window_seconds"]
            ]
            if len(recent) >= ANOMALY_THRESHOLDS["bulk_download"]["count"]:
                self._add_anomaly_alert(
                    "bulk_download",
                    f"用户 {user_id} 在5分钟内下载/导出 {len(recent)} 次",
                    entry,
                )

        hour = datetime.now(timezone.utc).hour
        threshold = ANOMALY_THRESHOLDS["off_hours_operation"]
        if hour >= threshold["start_hour"] or hour < threshold["end_hour"]:
            if action in ("delete", "export", "batch_download", "modify"):
                self._add_anomaly_alert(
                    "off_hours",
                    f"用户 {user_id} 在非工作时间执行 {action} 操作",
                    entry,
                )

    def _add_anomaly_alert(self, alert_type: str, message: str, entry: dict):
        self._anomaly_alerts.append({
            "type": alert_type,
            "message": message,
            "user_id": entry.get("user_id", ""),
            "action": entry.get("action_type", ""),
            "timestamp": time.time(),
        })
        if len(self._anomaly_alerts) > 500:
            self._anomaly_alerts = self._anomaly_alerts[-500:]
        logger.warning("ANOMALY ALERT [%s]: %s", alert_type, message)

    # ---- Query & Analysis ----

    def query_logs(
        self,
        user_id: str | None = None,
        action: str | None = None,
        object_type: str | None = None,
        project_id: str | None = None,
        limit: int = 100,
    ) -> list[dict]:
        """查询内存缓存（非权威来源，仅用于实时监控面板；合规审计查询请使用
        GET /api/audit-logs/verify-chain 或直接查 audit_log_entries 表）。

        R1 Bug Fix 10: 明确语义——此方法仅返回进程内最近 1000 条缓存记录，
        不保证完整性和持久性。正式审计取证必须走数据库查询。
        """
        results = self._recent_actions
        if user_id:
            results = [r for r in results if r.get("user_id") == user_id]
        if action:
            results = [r for r in results if r.get("action_type") == action]
        if object_type:
            results = [r for r in results if r.get("object_type") == object_type]
        if project_id:
            results = [r for r in results if r.get("project_id") == project_id]
        return results[-limit:]

    def get_anomaly_alerts(self, limit: int = 50) -> list[dict]:
        """获取异常告警。"""
        return self._anomaly_alerts[-limit:]

    # ---- Export ----

    def export_csv(self, logs: list[dict] | None = None) -> bytes:
        """导出为 CSV。"""
        data = logs or self._recent_actions
        output = io.StringIO()
        fields = ["ts", "user_id", "action_type", "object_type", "object_id", "project_id", "ip"]
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for entry in data:
            writer.writerow(entry)
        return output.getvalue().encode("utf-8-sig")

    def export_excel(self, logs: list[dict] | None = None) -> bytes:
        """导出为 Excel（openpyxl）。"""
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "审计日志"

            headers = ["ID", "事件类型", "操作者", "角色", "对象类型", "对象ID", "操作", "时间"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = openpyxl.styles.Font(bold=True, name="仿宋_GB2312")

            data = logs if logs is not None else []
            for row_idx, log in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=log.get("id", ""))
                ws.cell(row=row_idx, column=2, value=log.get("event_type", ""))
                ws.cell(row=row_idx, column=3, value=log.get("actor_id", ""))
                ws.cell(row=row_idx, column=4, value=log.get("actor_role", ""))
                ws.cell(row=row_idx, column=5, value=log.get("resource_type", log.get("object_type", "")))
                ws.cell(row=row_idx, column=6, value=log.get("resource_id", log.get("object_id", "")))
                ws.cell(row=row_idx, column=7, value=log.get("action_type", ""))
                ws.cell(row=row_idx, column=8, value=log.get("ts", ""))

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ImportError:
            return self.export_csv(logs)

    # ---- Retention ----

    def cleanup_old_logs(self, max_age_days: int = 365) -> int:
        """清理超过保留期的内存缓存日志。"""
        cutoff = time.time() - (max_age_days * 86400)
        before = len(self._recent_actions)
        self._recent_actions = [a for a in self._recent_actions if a.get("timestamp", 0) > cutoff]
        return before - len(self._recent_actions)


# Global instance
audit_logger = AuditLoggerEnhanced()
