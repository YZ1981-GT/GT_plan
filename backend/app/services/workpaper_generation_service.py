"""自定义底稿 working_paper 幂等生成（custom-workpaper-formula-binding 任务 6）。"""

from __future__ import annotations

import logging
import uuid
from pathlib import Path

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.workpaper_models import (
    WorkingPaper,
    WpIndex,
    WpSourceType,
)

logger = logging.getLogger(__name__)


class WorkpaperGenerationService:
    """指派 / 手动入口幂等创建 working_paper。"""

    async def ensure_working_paper(
        self,
        db: AsyncSession,
        project_id: uuid.UUID,
        wp_index_id: uuid.UUID,
        *,
        source_type: WpSourceType = WpSourceType.manual,
        created_by: uuid.UUID | None = None,
    ) -> WorkingPaper:
        """先查后建；已存在则直接返回（uq_working_paper_project_index）。"""
        existing = (
            await db.execute(
                sa.select(WorkingPaper).where(
                    WorkingPaper.project_id == project_id,
                    WorkingPaper.wp_index_id == wp_index_id,
                    WorkingPaper.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one_or_none()
        if existing is not None:
            return existing

        wp_index = (
            await db.execute(
                sa.select(WpIndex).where(
                    WpIndex.id == wp_index_id,
                    WpIndex.project_id == project_id,
                    WpIndex.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one_or_none()
        if wp_index is None:
            raise ValueError("底稿索引不存在")

        cycle = wp_index.audit_cycle or "A"
        dest_file = (
            Path("storage")
            / "projects"
            / str(project_id)
            / "workpapers"
            / cycle
            / f"{wp_index.wp_code}.xlsx"
        )
        dest_file.parent.mkdir(parents=True, exist_ok=True)
        if not dest_file.exists():
            try:
                import openpyxl

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = wp_index.wp_code[:31] or "Sheet1"
                ws["A1"] = f"底稿编号: {wp_index.wp_code}"
                ws["A2"] = f"底稿名称: {wp_index.wp_name or wp_index.wp_code}"
                wb.save(str(dest_file))
                wb.close()
            except Exception:
                dest_file.write_bytes(b"")

        wp = WorkingPaper(
            project_id=project_id,
            wp_index_id=wp_index_id,
            file_path=str(dest_file).replace("\\", "/"),
            source_type=source_type,
            file_version=1,
            parsed_data=None,
            created_by=created_by,
        )
        db.add(wp)
        await db.flush()
        logger.info(
            "ensure_working_paper 新建 wp_index=%s wp_code=%s",
            wp_index_id,
            wp_index.wp_code,
        )
        return wp


workpaper_generation_service = WorkpaperGenerationService()
