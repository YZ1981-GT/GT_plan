"""Sprint C.0 — 附注离线导出服务 (D15).

主要 API:
- export_sections(project_id, year, section_ids, ...) → (xlsx_bytes, sha256_hash)

功能：
- C.0.1: 按 section_id list 选择章节子集
- C.0.2: xlsx 包结构（注意事项 + 章节清单 + N 章节 + _meta_）
- C.0.3: 4 色语义渲染 + DataValidation 锁定
- C.0.4: 单元格批注（公式说明 + 数据源）
- C.0.5: _meta_ sheet（base64+gzip 压缩 binding/formula/row_meta）
- C.0.6: 注意事项 sheet 模板
- C.0.7: 章节清单 TOC
- C.0.8: 可选 AES 加密 + 文件 hash

纯函数设计，DB 操作通过外部传入数据。
"""
from __future__ import annotations

import base64
import gzip
import hashlib
import json
from copy import deepcopy
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["NoteOfflineExportService", "export_sections_to_xlsx"]


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# 4-color semantic fills (C.0.3)
FILL_EDITABLE = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # 黄=可填
FILL_FORMULA = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")  # 灰=公式
FILL_LOCKED = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")  # 红=锁定
FILL_REQUIRED = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")  # 绿=必填


FONT_HEADER = Font(bold=True, size=12)
FONT_TITLE = Font(bold=True, size=14)
FONT_NORMAL = Font(size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Max Excel sheet name length
MAX_SHEET_NAME = 31


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _truncate_sheet_name(name: str) -> str:
    """Truncate sheet name to Excel 31-char limit, replace invalid chars."""
    # Replace invalid Excel sheet name characters
    for ch in r"[]:*?/\\":
        name = name.replace(ch, "_")
    if len(name) > MAX_SHEET_NAME:
        name = name[: MAX_SHEET_NAME - 2] + ".."
    return name


def _classify_cell(cell_meta: dict[str, Any]) -> str:
    """Classify cell type for coloring.

    Returns: 'editable' | 'formula' | 'locked' | 'required'
    """
    source = cell_meta.get("source", "manual")
    is_required = cell_meta.get("is_required", False)
    mode = cell_meta.get("mode", "manual")

    if mode == "formula" or source == "formula":
        return "formula"
    if source in ("wp_data", "trial_balance", "aux_balance", "aux_ledger_aging", "consol_aggregation"):
        return "locked"
    if is_required:
        return "required"
    return "editable"


def _get_fill_for_type(cell_type: str) -> PatternFill:
    """Get fill color for cell type."""
    return {
        "editable": FILL_EDITABLE,
        "formula": FILL_FORMULA,
        "locked": FILL_LOCKED,
        "required": FILL_REQUIRED,
    }.get(cell_type, FILL_EDITABLE)


def _compute_sha256(data: bytes) -> str:
    """Compute SHA-256 hex digest."""
    return hashlib.sha256(data).hexdigest()


def _compress_meta(meta_dict: dict[str, Any]) -> str:
    """Compress metadata dict to base64+gzip string (C.0.5)."""
    json_bytes = json.dumps(meta_dict, ensure_ascii=False, default=str).encode("utf-8")
    compressed = gzip.compress(json_bytes)
    return base64.b64encode(compressed).decode("ascii")


def _decompress_meta(encoded: str) -> dict[str, Any]:
    """Decompress base64+gzip string back to dict."""
    compressed = base64.b64decode(encoded.encode("ascii"))
    json_bytes = gzip.decompress(compressed)
    return json.loads(json_bytes.decode("utf-8"))


def _encrypt_bytes(data: bytes, password: str) -> bytes:
    """AES encrypt bytes using Fernet (C.0.8)."""
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

    # Derive key from password
    salt = b"note_offline_export_salt_v1"  # Fixed salt for deterministic key derivation
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100_000)
    key = base64.urlsafe_b64encode(kdf.derive(password.encode("utf-8")))
    f = Fernet(key)
    return f.encrypt(data)


def _decrypt_bytes(data: bytes, password: str) -> bytes:
    """AES decrypt bytes using Fernet."""
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

    salt = b"note_offline_export_salt_v1"
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100_000)
    key = base64.urlsafe_b64encode(kdf.derive(password.encode("utf-8")))
    f = Fernet(key)
    return f.decrypt(data)


# ---------------------------------------------------------------------------
# C.0.6: 注意事项 Sheet 模板
# ---------------------------------------------------------------------------


def _build_instructions_sheet(ws: Worksheet, context: dict[str, Any]) -> None:
    """Fill the 注意事项 sheet with 7-section usage guide (C.0.6)."""
    project_name = context.get("project_name", "{项目名称}")
    year = context.get("year", "{年度}")
    exporter_name = context.get("exporter_name", "{导出人}")
    export_time = context.get("export_time", _now_str())
    section_count = context.get("section_count", 0)
    partner_name = context.get("partner_name", "{partner_name}")
    partner_email = context.get("partner_email", "{partner_email}")
    partner_phone = context.get("partner_phone", "{partner_phone}")

    ws.column_dimensions["A"].width = 80

    lines = [
        ("═" * 50, FONT_TITLE),
        ("  附注离线编辑包  —  使用说明", FONT_TITLE),
        ("═" * 50, FONT_TITLE),
        ("", None),
        ("【1. 文件用途】", FONT_HEADER),
        (f"  本文件为「{project_name} - {year} 年报附注」离线协作编辑包，", FONT_NORMAL),
        (f"  共 {section_count} 章节，由 {exporter_name} 于 {export_time} 生成。", FONT_NORMAL),
        ("", None),
        ("【2. 单元格颜色含义】", FONT_HEADER),
        ("  ⬛ 黄底 = 您可填写的数据（如金额 / 客户名称 / 描述）", FONT_NORMAL),
        ("  ⬛ 灰底 = 系统自动计算（公式驱动，请勿修改）", FONT_NORMAL),
        ("  ⬛ 红底 = 锁定单元格（如来自试算表 / 底稿，禁止改动）", FONT_NORMAL),
        ("  ⬛ 绿底 = 必填项（导入时若为空将告警）", FONT_NORMAL),
        ("", None),
        ("【3. 动态行操作】", FONT_HEADER),
        ('  在带 ★ 标识的"动态行"区域：', FONT_NORMAL),
        ("    - 复制现有行 → 在两个标记行之间粘贴 → 填写新数据", FONT_NORMAL),
        ("    - 不要在 [DYNAMIC_END:xxx] 标记下方插入数据", FONT_NORMAL),
        ("    - 删除行：选中整行 → 右键删除（不要清空内容留空行）", FONT_NORMAL),
        ("", None),
        ("【4. 公式区域】", FONT_HEADER),
        ("  灰底单元格的公式见单元格批注（鼠标悬停查看）。", FONT_NORMAL),
        ("  如需修改公式逻辑，请在对应黄底原始数据修改后由系统重算。", FONT_NORMAL),
        ("", None),
        ("【5. 文件名 / sheet 名规则】", FONT_HEADER),
        ("  ⚠ 不要重命名文件 / 不要修改 sheet 名（含隐藏的章节 ID 用于回传匹配）", FONT_NORMAL),
        ("  ⚠ 不要删除 _meta_ 隐藏 sheet（删了无法导回）", FONT_NORMAL),
        ("", None),
        ("【6. 完成后回传】", FONT_HEADER),
        (f"  填写完成后，将本文件发回 {partner_email}，", FONT_NORMAL),
        ("  partner 在系统内点「附注 → 一键导入」即可合并您的填报内容。", FONT_NORMAL),
        ("", None),
        ("【7. 注意事项】", FONT_HEADER),
        ("  - 不要修改 sheet 顺序（导入按 section_id 匹配，顺序无关，但避免误删）", FONT_NORMAL),
        ("  - 不要新增 sheet（新 sheet 不会被导入）", FONT_NORMAL),
        ("  - 文字段落（如会计政策）可改但建议保留模板措辞", FONT_NORMAL),
        ("  - 多人协作时合并冲突由 partner 在系统内决策", FONT_NORMAL),
        ("", None),
        ("【联系人】", FONT_HEADER),
        (f"  partner: {partner_name}  邮箱: {partner_email}  电话: {partner_phone}", FONT_NORMAL),
        ("═" * 50, FONT_TITLE),
    ]

    for row_idx, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = ALIGN_LEFT


# ---------------------------------------------------------------------------
# C.0.7: 章节清单 TOC Sheet
# ---------------------------------------------------------------------------


def _build_toc_sheet(ws: Worksheet, sections: list[dict[str, Any]]) -> None:
    """Build TOC sheet with section list (C.0.7)."""
    headers = ["序号", "章节标题", "完成度(%)", "必填项数", "section_id"]
    col_widths = [8, 40, 12, 12, 30]

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Hide section_id column (column E)
    ws.column_dimensions["E"].hidden = True

    # Data rows
    for row_idx, section in enumerate(sections, start=2):
        section_id = section.get("section_id", "")
        title = section.get("section_title", "")
        completeness = _calc_completeness(section)
        required_count = _count_required_cells(section)

        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = BORDER_THIN
        ws.cell(row=row_idx, column=2, value=title).border = BORDER_THIN
        ws.cell(row=row_idx, column=3, value=completeness).border = BORDER_THIN
        ws.cell(row=row_idx, column=4, value=required_count).border = BORDER_THIN
        ws.cell(row=row_idx, column=5, value=section_id).border = BORDER_THIN


def _calc_completeness(section: dict[str, Any]) -> int:
    """Calculate section completeness percentage.

    真实结构：rows[].values（非 cells），多表则遍历 _tables。
    """
    table_data = section.get("table_data", {}) or {}
    tables = table_data.get("_tables")
    if not isinstance(tables, list) or not tables:
        tables = [table_data]

    total_cells = 0
    filled_cells = 0
    for tbl in tables:
        if not isinstance(tbl, dict):
            continue
        for row in tbl.get("rows", []) or []:
            for cell_val in row.get("values", []) or []:
                total_cells += 1
                if cell_val is not None and cell_val != "" and cell_val != "-":
                    filled_cells += 1

    return int((filled_cells / total_cells * 100) if total_cells > 0 else 0)


def _count_required_cells(section: dict[str, Any]) -> int:
    """Count required cells in section."""
    cell_meta = section.get("_cell_meta", {})
    return sum(1 for m in cell_meta.values() if m.get("is_required", False))


# ---------------------------------------------------------------------------
# C.0.2/C.0.3/C.0.4: Section Sheet Builder
# ---------------------------------------------------------------------------


def _build_section_sheet(
    ws: Worksheet,
    section: dict[str, Any],
    include_formulas: bool = True,
    include_provenance: bool = True,
) -> None:
    """Build a single section sheet with data + 4-color + comments (C.0.2/3/4).

    真实 ``table_data`` 结构（与 DisclosureNote 一致）：
        {
          "name": "货币资金",
          "headers": ["期末余额", "期初余额"],   # 可缺省（默认两列）
          "rows": [
            {"label": "库存现金", "values": [0.0, 0.0], "is_total": false,
             "row_type": "data",
             "_cell_meta": {"0": {...}, "1": {...}}, "_cell_modes": {"0":"auto"}}
          ],
          "_tables": [ ...多表... ]   # 多表章节
        }

    注意：单元格值字段是 **``values``**（非 ``cells``），且行级 ``_cell_meta`` /
    ``_cell_modes`` 按 **列索引** 键（"0"/"1"...）内嵌在每行；首列是 ``label``，
    其后才是 ``values`` 各列。早期版本误读 ``cells`` / section 级 ``_cell_meta``
    导致导出表格全空（2026-06-13 修）。
    """
    section_id = section.get("section_id", "")
    title = section.get("section_title", "")
    table_data = section.get("table_data", {}) or {}

    # Row 1: section title (metadata)
    ws.cell(row=1, column=1, value=f"章节: {title}")
    ws.cell(row=1, column=1).font = FONT_TITLE

    # Row 2: section_id (hidden row for import matching)
    ws.cell(row=2, column=1, value=f"section_id:{section_id}")
    ws.row_dimensions[2].hidden = True

    # 多表章节：_tables 优先；否则单表用顶层 table_data
    tables = table_data.get("_tables")
    if not isinstance(tables, list) or not tables:
        tables = [table_data]

    cur_row = 3
    for t_idx, tbl in enumerate(tables):
        if not isinstance(tbl, dict):
            continue
        cur_row = _render_section_table(
            ws, tbl, start_row=cur_row,
            include_formulas=include_formulas,
            include_provenance=include_provenance,
        )
        cur_row += 1  # 表间空一行

    # Enable sheet protection (allow editing unlocked cells)
    ws.protection.sheet = True
    ws.protection.enable()


def _render_section_table(
    ws: Worksheet,
    tbl: dict[str, Any],
    *,
    start_row: int,
    include_formulas: bool,
    include_provenance: bool,
) -> int:
    """渲染单张附注表到 worksheet，返回下一可用行号。

    列布局：第 1 列 = 行标签（label），其后 = ``values`` 各列。
    表头 = ["项目"] + headers（headers 缺省时按 values 列数生成"列1/列2"）。
    """
    name = tbl.get("name") or ""
    headers = tbl.get("headers") or []
    rows = tbl.get("rows") or []

    # 推断数据列数（取各行 values 最大长度）
    n_value_cols = 0
    for r in rows:
        vals = r.get("values")
        if isinstance(vals, list):
            n_value_cols = max(n_value_cols, len(vals))
    if not n_value_cols and headers:
        n_value_cols = max(0, len(headers) - 1)

    # headers 约定：第 1 个元素是 label 列表头（如"项目"），其余对应 values 各列。
    # 容错：若 headers 长度==n_value_cols（无 label 头），则 label 头用默认"项目"。
    if len(headers) == n_value_cols:
        label_header = "项目"
        value_headers = list(headers)
    else:
        label_header = str(headers[0]).replace("<br/>", "").replace("<br>", "") if headers else "项目"
        value_headers = list(headers[1:]) if len(headers) > 1 else []

    row_cursor = start_row

    # 表名行（可选）
    if name:
        cell = ws.cell(row=row_cursor, column=1, value=name)
        cell.font = FONT_HEADER
        row_cursor += 1

    # 表头行：第 1 列 label 头 + value 列头
    header_cell = ws.cell(row=row_cursor, column=1, value=label_header)
    header_cell.font = FONT_HEADER
    header_cell.alignment = ALIGN_CENTER
    header_cell.border = BORDER_THIN
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 30)
    for c in range(n_value_cols):
        raw = value_headers[c] if c < len(value_headers) else f"列{c + 1}"
        htext = str(raw).replace("<br/>", "").replace("<br>", "")
        cell = ws.cell(row=row_cursor, column=c + 2, value=htext)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        col_letter = get_column_letter(c + 2)
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 0, 16
        )
    row_cursor += 1

    # 数据行
    for row_data in rows:
        if not isinstance(row_data, dict):
            continue
        label = row_data.get("label", "")
        values = row_data.get("values") or []
        is_total = row_data.get("is_total", False)
        row_type = row_data.get("row_type", "data")
        cell_meta = row_data.get("_cell_meta") or {}
        cell_modes = row_data.get("_cell_modes") or {}
        is_dynamic = isinstance(row_type, str) and row_type.startswith("dynamic_")

        # 第 1 列：label（动态行加 ★）
        label_text = f"★ {label}" if is_dynamic else label
        lcell = ws.cell(row=row_cursor, column=1, value=label_text)
        lcell.border = BORDER_THIN
        if is_total:
            lcell.font = FONT_HEADER

        # 数据列：values
        for c in range(n_value_cols):
            val = values[c] if c < len(values) else None
            cell = ws.cell(row=row_cursor, column=c + 2, value=val)
            cell.border = BORDER_THIN
            if is_total:
                cell.font = FONT_HEADER

            col_meta = cell_meta.get(str(c), {}) if isinstance(cell_meta, dict) else {}
            mode = cell_modes.get(str(c)) if isinstance(cell_modes, dict) else None
            # 分类着色：mode=auto / 有 binding_id → 锁定（来自系统取数）
            has_binding = bool(col_meta.get("binding_id"))
            if has_binding or mode in ("auto", "formula"):
                cell.fill = _get_fill_for_type("locked" if has_binding else "formula")
                cell.protection = Protection(locked=True)
            else:
                cell.fill = _get_fill_for_type("editable")
                cell.protection = Protection(locked=False)

            # 注：不再写 openpyxl Comment（legacy VML 批注 + sheet protection 组合
            # 在 WPS 下会触发"无法打开指定的文件"。绑定/公式溯源信息已通过 4 色语义
            # + 隐藏 _meta_ sheet 完整承载，批注为冗余提示，移除以保 WPS 兼容）。

        row_cursor += 1

    return row_cursor


# ---------------------------------------------------------------------------
# C.0.5: _meta_ Sheet
# ---------------------------------------------------------------------------


def _build_meta_sheet(ws: Worksheet, sections: list[dict[str, Any]]) -> None:
    """Build hidden _meta_ sheet with compressed metadata (C.0.5)."""
    # Collect all metadata
    meta_payload: dict[str, Any] = {}
    for section in sections:
        sid = section.get("section_id", "")
        meta_payload[sid] = {
            "bindings": section.get("_bindings", {}),
            "formulas": section.get("_formulas", {}),
            "row_meta": section.get("_row_meta", []),
            "cell_modes": section.get("_cell_modes", {}),
            "cell_meta": section.get("_cell_meta", {}),
            "dynamic_regions": section.get("_dynamic_regions", []),
        }

    # Compress and store
    compressed = _compress_meta(meta_payload)

    # A1: key label
    ws.cell(row=1, column=1, value="meta_data")
    ws.cell(row=1, column=2, value=compressed)

    # A2: section_id list (for CI-21 validation)
    section_ids = [s.get("section_id", "") for s in sections]
    ws.cell(row=2, column=1, value="section_ids")
    ws.cell(row=2, column=2, value=json.dumps(section_ids, ensure_ascii=False))

    # A3: binding hash (for CI-21 validation)
    binding_hash = _compute_sha256(compressed.encode("utf-8"))
    ws.cell(row=3, column=1, value="binding_hash")
    ws.cell(row=3, column=2, value=binding_hash)

    # A4: export timestamp
    ws.cell(row=4, column=1, value="export_time")
    ws.cell(row=4, column=2, value=_now_str())

    # A5: version
    ws.cell(row=5, column=1, value="format_version")
    ws.cell(row=5, column=2, value="1.0")


# ---------------------------------------------------------------------------
# Navigation Hyperlinks (TOC ↔ Section sheets)
# ---------------------------------------------------------------------------

FONT_LINK = Font(color="0000FF", underline="single", size=10)
FONT_BACK_LINK = Font(color="0000FF", underline="single", size=9)


def _add_navigation_links(
    ws_toc: Worksheet,
    section_sheet_names: list[str],
    sections: list[dict[str, Any]],
) -> None:
    """Add bidirectional hyperlinks between TOC and section sheets.

    - TOC: title column cells link to corresponding section sheet A1
    - Each section sheet: row 1 right side gets a "← 返回目录" link back to TOC
    """
    from openpyxl.utils import quote_sheetname

    # TOC → section sheets (title is column B, starting row 2)
    for idx, sheet_name in enumerate(section_sheet_names):
        toc_row = idx + 2  # row 1 is header
        cell = ws_toc.cell(row=toc_row, column=2)
        # Internal hyperlink format: #'Sheet Name'!A1
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = FONT_LINK

    # Section sheets → back to TOC
    for sheet_name in section_sheet_names:
        # Find the worksheet by name
        ws = ws_toc.parent[sheet_name]
        # Add "← 返回目录" in column D row 1 (after the title)
        back_cell = ws.cell(row=1, column=4, value="← 返回目录")
        back_cell.hyperlink = "#'章节清单'!A1"
        back_cell.font = FONT_BACK_LINK


# ---------------------------------------------------------------------------
# Core Export Function (C.0.1 + C.0.2 orchestration)
# ---------------------------------------------------------------------------


def export_sections_to_xlsx(
    sections: list[dict[str, Any]],
    *,
    section_ids: list[str] | None = None,
    include_formulas: bool = True,
    include_provenance: bool = True,
    password: str | None = None,
    exporter_name: str = "",
    partner_info: dict[str, str] | None = None,
    project_name: str = "",
    year: int | str = "",
) -> tuple[bytes, str]:
    """Export disclosure note sections to xlsx bytes.

    Args:
        sections: List of section dicts (from DB or service layer).
        section_ids: Optional filter — only export these section_ids.
        include_formulas: Include formula expressions in comments.
        include_provenance: Include data source info in comments.
        password: Optional AES encryption password.
        exporter_name: Name of the person exporting.
        partner_info: Dict with partner_name, partner_email, partner_phone.
        project_name: Project display name.
        year: Fiscal year.

    Returns:
        (xlsx_bytes, sha256_hash) — if password provided, bytes are encrypted.
    """
    # C.0.1: Filter sections by section_id list
    if section_ids is not None:
        id_set = set(section_ids)
        filtered = [s for s in sections if s.get("section_id", "") in id_set]
    else:
        filtered = list(sections)

    if not filtered:
        filtered = []

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # C.0.6: 注意事项 sheet (first)
    ws_instructions = wb.create_sheet("注意事项")
    context = {
        "project_name": project_name,
        "year": year,
        "exporter_name": exporter_name,
        "export_time": _now_str(),
        "section_count": len(filtered),
        "partner_name": (partner_info or {}).get("partner_name", ""),
        "partner_email": (partner_info or {}).get("partner_email", ""),
        "partner_phone": (partner_info or {}).get("partner_phone", ""),
    }
    _build_instructions_sheet(ws_instructions, context)

    # C.0.7: 章节清单 sheet (second)
    ws_toc = wb.create_sheet("章节清单")
    _build_toc_sheet(ws_toc, filtered)

    # C.0.2: Section sheets — track sheet_name mapping for hyperlinks
    section_sheet_names: list[str] = []
    for idx, section in enumerate(filtered, start=1):
        section_id = section.get("section_id", "unknown")
        title = section.get("section_title", section_id)
        # 使用序号+标题作为 sheet 名，更直观
        numbered_title = f"{idx:02d}_{title}"
        sheet_name = _truncate_sheet_name(numbered_title)

        # Ensure unique sheet name
        existing_names = [ws.title for ws in wb.worksheets]
        if sheet_name in existing_names:
            sheet_name = _truncate_sheet_name(f"{title[:25]}_{section_id[:5]}")

        section_sheet_names.append(sheet_name)
        ws_section = wb.create_sheet(sheet_name)
        _build_section_sheet(ws_section, section, include_formulas, include_provenance)

    # Add hyperlinks: TOC → section sheets, section sheets → TOC
    _add_navigation_links(ws_toc, section_sheet_names, filtered)

    # C.0.5: _meta_ sheet (last, hidden)
    ws_meta = wb.create_sheet("_meta_")
    _build_meta_sheet(ws_meta, filtered)
    ws_meta.sheet_state = "hidden"

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()

    # C.0.8: Compute hash
    file_hash = _compute_sha256(xlsx_bytes)

    # C.0.8: Optional AES encryption
    if password:
        xlsx_bytes = _encrypt_bytes(xlsx_bytes, password)

    return xlsx_bytes, file_hash


# ---------------------------------------------------------------------------
# Async Service Class (wraps pure function with DB access)
# ---------------------------------------------------------------------------


class NoteOfflineExportService:
    """附注离线导出服务 (Sprint C.0).

    Async wrapper that loads sections from DB then delegates to pure export function.
    """

    def __init__(self, db: Any = None):
        """Initialize with optional DB session."""
        self.db = db

    async def export_sections(
        self,
        project_id: UUID,
        year: int,
        section_ids: list[str] | None = None,
        include_formulas: bool = True,
        include_provenance: bool = True,
        password: str | None = None,
        exporter_name: str = "",
        partner_info: dict[str, str] | None = None,
    ) -> tuple[bytes, str]:
        """Export sections from DB to xlsx.

        Returns:
            (xlsx_bytes, sha256_hash)
        """
        # Load sections from DB
        sections = await self._load_sections(project_id, year, section_ids)
        project_name = await self._get_project_name(project_id)

        return export_sections_to_xlsx(
            sections,
            section_ids=section_ids,
            include_formulas=include_formulas,
            include_provenance=include_provenance,
            password=password,
            exporter_name=exporter_name,
            partner_info=partner_info,
            project_name=project_name,
            year=year,
        )

    async def _load_sections(
        self,
        project_id: UUID,
        year: int,
        section_ids: list[str] | None = None,
    ) -> list[dict[str, Any]]:
        """Load disclosure note sections from DB."""
        if self.db is None:
            return []

        from sqlalchemy import or_, select as sa_select

        from app.models.report_models import DisclosureNote

        query = sa_select(DisclosureNote).where(
            DisclosureNote.project_id == project_id,
            DisclosureNote.year == year,
            DisclosureNote.is_deleted == False,  # noqa: E712
        ).order_by(DisclosureNote.sort_order.asc().nulls_last(), DisclosureNote.note_section.asc())
        if section_ids:
            # 前端自定义勾选传的是 note_section（如"八、1"），且 DB section_id 列多为空，
            # 故按 note_section 过滤（兼容极少数 section_id 有值的情况）。
            query = query.where(
                or_(
                    DisclosureNote.note_section.in_(section_ids),
                    DisclosureNote.section_id.in_(section_ids),
                )
            )

        result = await self.db.execute(query)
        notes = result.scalars().all()

        sections = []
        for note in notes:
            section_dict = {
                "section_id": note.section_id or note.note_section or "",
                "section_title": note.section_title or "",
                "table_data": note.table_data or {},
                "_formulas": (note.table_data or {}).get("_formulas", {}),
                "_cell_provenance": (note.table_data or {}).get("_cell_provenance", {}),
                "_cell_meta": (note.table_data or {}).get("_cell_meta", {}),
                "_cell_modes": (note.table_data or {}).get("_cell_modes", {}),
                "_bindings": (note.table_data or {}).get("_bindings", {}),
                "_row_meta": (note.table_data or {}).get("_row_meta", []),
                "_dynamic_regions": (note.table_data or {}).get("_dynamic_regions", []),
            }
            sections.append(section_dict)

        return sections

    async def _get_project_name(self, project_id: UUID) -> str:
        """Get project display name."""
        if self.db is None:
            return ""
        try:
            from sqlalchemy import select as sa_select

            from app.models.models import Project

            result = await self.db.execute(
                sa_select(Project.name).where(Project.id == project_id)
            )
            row = result.scalar_one_or_none()
            return row or ""
        except Exception:
            return ""
