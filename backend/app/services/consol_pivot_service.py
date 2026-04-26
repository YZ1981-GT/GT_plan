"""自定义透视查询 + Excel 导出 + 模板 CRUD

execute_query: 行/列维度 + 值字段 + 筛选
_pivot_account_by_company: 行=科目，列=企业
_pivot_company_by_account: 行=企业，列=科目
_transpose: 行列互换
export_excel: openpyxl 导出
save_template / list_templates: 查询模板 CRUD
"""

from __future__ import annotations

import io
import uuid
from collections import defaultdict
from decimal import Decimal
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.consolidation_models import ConsolQueryTemplate, ConsolWorksheet
from app.services.consol_tree_service import TreeNode, build_tree, get_descendants


ZERO = Decimal("0")

# Valid value fields on ConsolWorksheet
VALID_VALUE_FIELDS = {
    "consolidated_amount", "children_amount_sum", "net_difference",
    "adjustment_debit", "adjustment_credit",
    "elimination_debit", "elimination_credit",
}


async def execute_query(
    db: AsyncSession,
    project_id: UUID,
    year: int,
    row_dimension: str = "account",
    col_dimension: str = "company",
    value_field: str = "consolidated_amount",
    filters: dict | None = None,
    transpose: bool = False,
    node_company_code: str | None = None,
    aggregation_mode: str = "self",
) -> dict:
    """执行透视查询。

    Returns {headers: [...], rows: [[...]], totals: [...]}
    """
    if value_field not in VALID_VALUE_FIELDS:
        value_field = "consolidated_amount"

    # Determine which company codes to include
    company_codes = await _resolve_company_codes(
        db, project_id, node_company_code, aggregation_mode
    )

    # Load worksheet data
    query = sa.select(ConsolWorksheet).where(
        ConsolWorksheet.project_id == project_id,
        ConsolWorksheet.year == year,
        ConsolWorksheet.is_deleted == sa.false(),
    )
    if company_codes is not None:
        query = query.where(ConsolWorksheet.node_company_code.in_(company_codes))

    # Apply filters
    if filters:
        if "account_codes" in filters and filters["account_codes"]:
            query = query.where(ConsolWorksheet.account_code.in_(filters["account_codes"]))
        if "company_codes" in filters and filters["company_codes"]:
            query = query.where(ConsolWorksheet.node_company_code.in_(filters["company_codes"]))

    result = await db.execute(query)
    rows = result.scalars().all()

    if row_dimension == "account" and col_dimension == "company":
        pivot_result = _pivot_account_by_company(rows, value_field)
    elif row_dimension == "company" and col_dimension == "account":
        pivot_result = _pivot_company_by_account(rows, value_field)
    else:
        pivot_result = _pivot_account_by_company(rows, value_field)

    if transpose:
        pivot_result = _transpose(pivot_result)

    return pivot_result


async def _resolve_company_codes(
    db: AsyncSession,
    project_id: UUID,
    node_company_code: str | None,
    aggregation_mode: str,
) -> list[str] | None:
    """根据汇总模式确定要包含的企业编码列表"""
    if not node_company_code:
        return None  # All codes

    tree = await build_tree(db, project_id)
    if not tree:
        return [node_company_code]

    from app.services.consol_tree_service import find_node as _find
    target = _find(tree, node_company_code)
    if not target:
        return [node_company_code]

    if aggregation_mode == "self":
        return [target.company_code]
    elif aggregation_mode == "children":
        return [target.company_code] + [c.company_code for c in target.children]
    elif aggregation_mode == "descendants":
        all_nodes = [target] + get_descendants(target)
        return [n.company_code for n in all_nodes]
    return [node_company_code]


def _pivot_account_by_company(
    rows: list[ConsolWorksheet], value_field: str
) -> dict:
    """行=科目，列=企业，含合计列"""
    # Collect unique dimensions
    accounts: list[str] = sorted(set(r.account_code for r in rows))
    companies: list[str] = sorted(set(r.node_company_code for r in rows))

    # Build data matrix
    data_map: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: ZERO))
    for r in rows:
        val = getattr(r, value_field, ZERO) or ZERO
        data_map[r.account_code][r.node_company_code] = val

    headers = ["科目编码"] + companies + ["合计"]
    result_rows = []
    col_totals = defaultdict(lambda: ZERO)

    for acct in accounts:
        row_data = [acct]
        row_total = ZERO
        for comp in companies:
            val = data_map[acct][comp]
            row_data.append(str(val))
            row_total += val
            col_totals[comp] += val
        row_data.append(str(row_total))
        result_rows.append(row_data)

    # Totals row
    totals = ["合计"]
    grand_total = ZERO
    for comp in companies:
        totals.append(str(col_totals[comp]))
        grand_total += col_totals[comp]
    totals.append(str(grand_total))

    return {"headers": headers, "rows": result_rows, "totals": totals}


def _pivot_company_by_account(
    rows: list[ConsolWorksheet], value_field: str
) -> dict:
    """行=企业，列=科目，含合计列"""
    accounts: list[str] = sorted(set(r.account_code for r in rows))
    companies: list[str] = sorted(set(r.node_company_code for r in rows))

    data_map: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: ZERO))
    for r in rows:
        val = getattr(r, value_field, ZERO) or ZERO
        data_map[r.node_company_code][r.account_code] = val

    headers = ["企业编码"] + accounts + ["合计"]
    result_rows = []
    col_totals = defaultdict(lambda: ZERO)

    for comp in companies:
        row_data = [comp]
        row_total = ZERO
        for acct in accounts:
            val = data_map[comp][acct]
            row_data.append(str(val))
            row_total += val
            col_totals[acct] += val
        row_data.append(str(row_total))
        result_rows.append(row_data)

    totals = ["合计"]
    grand_total = ZERO
    for acct in accounts:
        totals.append(str(col_totals[acct]))
        grand_total += col_totals[acct]
    totals.append(str(grand_total))

    return {"headers": headers, "rows": result_rows, "totals": totals}


def _transpose(pivot_result: dict) -> dict:
    """行列互换"""
    headers = pivot_result["headers"]
    rows = pivot_result["rows"]
    totals = pivot_result.get("totals", [])

    # Build full matrix including totals
    all_rows = rows + ([totals] if totals else [])
    if not all_rows:
        return pivot_result

    # New headers = first column of each row
    new_headers = [headers[0]] + [r[0] for r in all_rows]

    # New rows = each original column becomes a row
    new_rows = []
    for col_idx in range(1, len(headers)):
        new_row = [headers[col_idx]]
        for r in all_rows:
            new_row.append(r[col_idx] if col_idx < len(r) else "0")
        new_rows.append(new_row)

    return {"headers": new_headers, "rows": new_rows, "totals": []}


async def export_excel(
    db: AsyncSession,
    project_id: UUID,
    year: int,
    row_dimension: str = "account",
    col_dimension: str = "company",
    value_field: str = "consolidated_amount",
    filters: dict | None = None,
    transpose: bool = False,
    node_company_code: str | None = None,
    aggregation_mode: str = "self",
) -> bytes:
    """导出透视查询结果为 Excel 字节流"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")

    pivot_data = await execute_query(
        db, project_id, year, row_dimension, col_dimension,
        value_field, filters, transpose, node_company_code, aggregation_mode,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "合并透视表"

    # Fonts
    header_font = Font(name="仿宋_GB2312", bold=True, size=11)
    data_font = Font(name="Arial Narrow", size=10)
    total_font = Font(name="Arial Narrow", bold=True, size=10)
    thin_border = Border(
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(pivot_data["headers"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(pivot_data["rows"], 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="right")
                # Try to convert to number for Excel
                try:
                    cell.value = float(val)
                    cell.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    pass

    # Write totals row
    if pivot_data.get("totals"):
        total_row = len(pivot_data["rows"]) + 2
        for col_idx, val in enumerate(pivot_data["totals"], 1):
            cell = ws.cell(row=total_row, column=col_idx, value=val)
            cell.font = total_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="right")
                try:
                    cell.value = float(val)
                    cell.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    pass

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Template CRUD
# ---------------------------------------------------------------------------


async def save_template(
    db: AsyncSession,
    project_id: UUID,
    name: str,
    row_dimension: str,
    col_dimension: str,
    value_field: str,
    filters: dict | None = None,
    transpose: bool = False,
    aggregation_mode: str = "self",
) -> dict:
    """保存查询模板"""
    tpl = ConsolQueryTemplate(
        id=uuid.uuid4(),
        project_id=project_id,
        name=name,
        row_dimension=row_dimension,
        col_dimension=col_dimension,
        value_field=value_field,
        filters=filters or {},
        transpose=transpose,
        aggregation_mode=aggregation_mode,
    )
    db.add(tpl)
    await db.commit()
    return _tpl_to_dict(tpl)


async def list_templates(
    db: AsyncSession,
    project_id: UUID,
) -> list[dict]:
    """列出项目的查询模板"""
    result = await db.execute(
        sa.select(ConsolQueryTemplate).where(
            ConsolQueryTemplate.project_id == project_id,
            ConsolQueryTemplate.is_deleted == sa.false(),
        ).order_by(ConsolQueryTemplate.created_at.desc())
    )
    templates = result.scalars().all()
    return [_tpl_to_dict(t) for t in templates]


def _tpl_to_dict(tpl: ConsolQueryTemplate) -> dict:
    return {
        "id": str(tpl.id),
        "name": tpl.name,
        "row_dimension": tpl.row_dimension,
        "col_dimension": tpl.col_dimension,
        "value_field": tpl.value_field,
        "filters": tpl.filters,
        "transpose": tpl.transpose,
        "aggregation_mode": tpl.aggregation_mode,
    }
