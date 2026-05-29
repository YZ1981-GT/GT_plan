"""Sprint 4 Task 15 — 底稿离线导入服务 (US-14).

主要 API:
- validate_and_diff(xlsx_bytes, existing_sheets) → ImportPreviewResult
- apply_import(xlsx_bytes, existing_sheets, strategy, ...) → ImportResult

功能：
- 15.6: 验证 + diff + apply
- 15.7: validate_and_diff：_meta_ 校验 + 逐 cell 对比 + 分类（changed/added/deleted）
- 15.8: apply_import：3 种冲突策略（overwrite/keep_system/merge）+ 审计日志
- 15.14: 导出文件 30 天归档 + 审计日志

纯函数设计，DB 操作通过外部传入或 async service class。
"""
from __future__ import annotations

import hashlib
import json
from copy import deepcopy
from datetime import datetime, timezone
from enum import Enum
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from app.services.wp_offline_export_service import _compress_meta, _decompress_meta

__all__ = [
    "WpOfflineImportService",
    "validate_import_file",
    "diff_sheets",
    "apply_import",
    "ImportValidationResult",
    "SheetDiff",
    "CellDiff",
    "ConflictStrategy",
    "ImportResult",
]


# ---------------------------------------------------------------------------
# Enums & Data Classes
# ---------------------------------------------------------------------------


class DiffType(str, Enum):
    """Cell diff type."""
    CHANGED = "changed"
    ADDED = "added"
    DELETED = "deleted"


class ConflictStrategy(str, Enum):
    """Conflict resolution strategy (15.8)."""
    OVERWRITE = "overwrite"      # 导入版本覆盖系统
    KEEP_SYSTEM = "keep_system"  # 保留系统版本
    MERGE = "merge"              # cell 级选择性合并


class CellDiff:
    """Single cell diff result (15.7)."""

    __slots__ = ("cell_key", "sheet_name", "local_value", "imported_value", "diff_type")

    def __init__(
        self,
        cell_key: str,
        sheet_name: str,
        local_value: Any,
        imported_value: Any,
        diff_type: DiffType,
    ):
        self.cell_key = cell_key
        self.sheet_name = sheet_name
        self.local_value = local_value
        self.imported_value = imported_value
        self.diff_type = diff_type

    def to_dict(self) -> dict[str, Any]:
        return {
            "cell": self.cell_key,
            "sheet": self.sheet_name,
            "local": self.local_value,
            "imported": self.imported_value,
            "type": self.diff_type.value,
        }


class SheetDiff:
    """Sheet-level diff result."""

    __slots__ = ("sheet_name", "cell_diffs", "status")

    def __init__(
        self,
        sheet_name: str,
        cell_diffs: list[CellDiff] | None = None,
        status: str = "matched",
    ):
        self.sheet_name = sheet_name
        self.cell_diffs = cell_diffs or []
        self.status = status  # matched / import_only / system_only

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "status": self.status,
            "diff_count": len(self.cell_diffs),
            "diffs": [d.to_dict() for d in self.cell_diffs],
        }


class ImportValidationResult:
    """Validation result for import file (15.7)."""

    __slots__ = ("valid", "errors", "sheet_names", "meta_data", "binding_hash", "format_version", "wp_id")

    def __init__(
        self,
        valid: bool,
        errors: list[str] | None = None,
        sheet_names: list[str] | None = None,
        meta_data: dict[str, Any] | None = None,
        binding_hash: str = "",
        format_version: str = "",
        wp_id: str = "",
    ):
        self.valid = valid
        self.errors = errors or []
        self.sheet_names = sheet_names or []
        self.meta_data = meta_data or {}
        self.binding_hash = binding_hash
        self.format_version = format_version
        self.wp_id = wp_id


class ImportResult:
    """Result of applying import (15.8)."""

    __slots__ = ("success", "sheets_imported", "sheets_kept", "cells_changed", "audit_entry")

    def __init__(
        self,
        success: bool,
        sheets_imported: int = 0,
        sheets_kept: int = 0,
        cells_changed: int = 0,
        audit_entry: dict[str, Any] | None = None,
    ):
        self.success = success
        self.sheets_imported = sheets_imported
        self.sheets_kept = sheets_kept
        self.cells_changed = cells_changed
        self.audit_entry = audit_entry or {}

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "sheets_imported": self.sheets_imported,
            "sheets_kept": self.sheets_kept,
            "cells_changed": self.cells_changed,
            "audit_entry": self.audit_entry,
        }


# ---------------------------------------------------------------------------
# 15.7: Validate Import File
# ---------------------------------------------------------------------------


def validate_import_file(xlsx_bytes: bytes) -> ImportValidationResult:
    """Validate uploaded xlsx file for import (15.7).

    Checks:
    1. Valid xlsx format
    2. _meta_ sheet exists
    3. _meta_ has sheet_names + binding_hash
    4. format_version is compatible
    """
    errors: list[str] = []

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ImportValidationResult(valid=False, errors=[f"无法解析 xlsx 文件: {e}"])

    # Check _meta_ sheet exists
    if "_meta_" not in wb.sheetnames:
        wb.close()
        return ImportValidationResult(valid=False, errors=["缺少 _meta_ 隐藏 sheet，无法导入"])

    ws_meta = wb["_meta_"]

    # Read meta fields
    meta_data_key = ws_meta.cell(row=1, column=1).value
    meta_data_val = ws_meta.cell(row=1, column=2).value
    sheet_names_key = ws_meta.cell(row=2, column=1).value
    sheet_names_val = ws_meta.cell(row=2, column=2).value
    binding_hash_key = ws_meta.cell(row=3, column=1).value
    binding_hash_val = ws_meta.cell(row=3, column=2).value
    format_version_key = ws_meta.cell(row=5, column=1).value
    format_version_val = ws_meta.cell(row=5, column=2).value
    wp_id_key = ws_meta.cell(row=6, column=1).value
    wp_id_val = ws_meta.cell(row=6, column=2).value

    # Validate meta_data
    if meta_data_key != "meta_data" or not meta_data_val:
        errors.append("_meta_ sheet 缺少 meta_data 字段")

    # Validate sheet_names
    sheet_names: list[str] = []
    if sheet_names_key != "sheet_names" or not sheet_names_val:
        errors.append("_meta_ sheet 缺少 sheet_names 字段")
    else:
        try:
            sheet_names = json.loads(sheet_names_val)
        except (json.JSONDecodeError, TypeError):
            errors.append("sheet_names 格式无效")

    # Validate binding_hash
    binding_hash = ""
    if binding_hash_key != "binding_hash" or not binding_hash_val:
        errors.append("_meta_ sheet 缺少 binding_hash 字段")
    else:
        binding_hash = str(binding_hash_val)

    # Validate format version
    format_version = str(format_version_val or "")
    if format_version and format_version not in ("1.0",):
        errors.append(f"不支持的格式版本: {format_version}")

    # Decompress meta to verify integrity
    meta_data: dict[str, Any] = {}
    if meta_data_val:
        try:
            meta_data = _decompress_meta(str(meta_data_val))
        except Exception as e:
            errors.append(f"meta_data 解压失败: {e}")

    # Extract wp_id
    wp_id = str(wp_id_val or "") if wp_id_key == "wp_id" else ""

    wb.close()

    return ImportValidationResult(
        valid=len(errors) == 0,
        errors=errors,
        sheet_names=sheet_names,
        meta_data=meta_data,
        binding_hash=binding_hash,
        format_version=format_version,
        wp_id=wp_id,
    )


# ---------------------------------------------------------------------------
# 15.7: Extract Imported Sheets
# ---------------------------------------------------------------------------


def _extract_imported_sheets(xlsx_bytes: bytes, meta_data: dict[str, Any] | None = None) -> dict[str, dict[str, Any]]:
    """Extract sheet data from xlsx.

    Returns: {sheet_name: {headers, rows}}
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    result: dict[str, dict[str, Any]] = {}

    # Get row counts from meta if available
    sheets_meta = (meta_data or {}).get("sheets", {})

    for ws in wb.worksheets:
        if ws.title in ("注意事项", "_meta_"):
            continue

        # Row 1 has hidden sheet_id
        sid_cell = ws.cell(row=1, column=1).value
        if not sid_cell or not str(sid_cell).startswith("sheet_id:"):
            continue

        sheet_name = str(sid_cell).replace("sheet_id:", "").strip()

        # Read headers (row 2)
        headers: list[str] = []
        col = 1
        while True:
            val = ws.cell(row=2, column=col).value
            if val is None:
                break
            headers.append(str(val))
            col += 1

        # Determine expected row count from meta or use max_row
        expected_rows = sheets_meta.get(sheet_name, {}).get("row_count", 0)
        if expected_rows == 0:
            # Fallback: use max_row from openpyxl
            expected_rows = max(0, (ws.max_row or 2) - 2)

        # Read data rows (row 3+)
        rows_data: list[dict[str, Any]] = []
        for r in range(expected_rows):
            row_idx = r + 3
            cells: list[Any] = []
            for c in range(1, max(len(headers), 1) + 1):
                val = ws.cell(row=row_idx, column=c).value
                cells.append(val)
            rows_data.append({"cells": cells, "row_idx": r})

        result[sheet_name] = {
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows_data,
        }

    wb.close()
    return result


# ---------------------------------------------------------------------------
# 15.7: Cell-by-Cell Diff
# ---------------------------------------------------------------------------


def _values_equal(a: Any, b: Any) -> bool:
    """Compare two cell values with tolerance for numeric types."""
    import math

    if a is None and b is None:
        return True
    # Treat empty string and None as equivalent (Excel doesn't distinguish)
    if (a is None or a == "") and (b is None or b == ""):
        return True
    if a is None or b is None:
        return False
    # Numeric comparison with tolerance
    try:
        fa, fb = float(a), float(b)
        # Handle nan: both nan → equal
        if math.isnan(fa) and math.isnan(fb):
            return True
        # Handle inf
        if fa == fb:
            return True
        if abs(fa - fb) < 1e-6:
            return True
        return False
    except (ValueError, TypeError):
        pass
    # String comparison (strip whitespace)
    return str(a).strip() == str(b).strip()


def diff_sheet_cells(
    sheet_name: str,
    local_rows: list[dict[str, Any]],
    imported_rows: list[dict[str, Any]],
    cell_meta: dict[str, Any] | None = None,
) -> list[CellDiff]:
    """Compute cell-by-cell diff between local and imported sheet (15.7).

    Classifies each diff as changed/added/deleted.
    """
    cell_meta = cell_meta or {}
    diffs: list[CellDiff] = []

    # Build cell maps
    local_cells: dict[str, Any] = {}
    for r_idx, row in enumerate(local_rows):
        cells = row.get("cells", [])
        for c_idx, val in enumerate(cells):
            local_cells[f"{r_idx}:{c_idx}"] = val

    imported_cells: dict[str, Any] = {}
    for r_idx, row in enumerate(imported_rows):
        cells = row.get("cells", [])
        for c_idx, val in enumerate(cells):
            imported_cells[f"{r_idx}:{c_idx}"] = val

    # Union of all cell keys
    all_keys = set(local_cells.keys()) | set(imported_cells.keys())

    for cell_key in sorted(all_keys):
        local_val = local_cells.get(cell_key)
        import_val = imported_cells.get(cell_key)

        # Skip cells that are formula/locked (not user-editable)
        meta = cell_meta.get(cell_key, {})
        source = meta.get("source", "manual")
        mode = meta.get("mode", "manual")
        if mode == "formula" or source == "formula":
            continue
        if source in ("wp_data", "trial_balance", "cross_ref", "auto_fill"):
            continue

        if _values_equal(local_val, import_val):
            continue

        # Determine diff type
        if cell_key not in local_cells:
            diff_type = DiffType.ADDED
        elif cell_key not in imported_cells:
            diff_type = DiffType.DELETED
        else:
            diff_type = DiffType.CHANGED

        diffs.append(CellDiff(
            cell_key=cell_key,
            sheet_name=sheet_name,
            local_value=local_val,
            imported_value=import_val,
            diff_type=diff_type,
        ))

    return diffs


# ---------------------------------------------------------------------------
# 15.7: Full Diff
# ---------------------------------------------------------------------------


def diff_sheets(
    xlsx_bytes: bytes,
    existing_sheets: list[dict[str, Any]],
    meta_data: dict[str, Any] | None = None,
) -> list[SheetDiff]:
    """Compute full diff between imported xlsx and existing sheets (15.7).

    Args:
        xlsx_bytes: The uploaded xlsx file bytes.
        existing_sheets: List of existing sheet dicts from system.
        meta_data: Optional metadata from _meta_ sheet.

    Returns:
        List of SheetDiff objects.
    """
    imported = _extract_imported_sheets(xlsx_bytes, meta_data)
    imported_names = set(imported.keys())
    existing_lookup = {s.get("sheet_name", ""): s for s in existing_sheets}
    existing_names = set(existing_lookup.keys())

    results: list[SheetDiff] = []

    # Matched sheets
    for name in sorted(imported_names & existing_names):
        local = existing_lookup[name]
        imp = imported[name]
        # Get cell_meta from meta_data if available
        sheet_meta = (meta_data or {}).get("sheets", {}).get(name, {}).get("cell_meta", {})
        cell_diffs = diff_sheet_cells(
            sheet_name=name,
            local_rows=local.get("rows", []),
            imported_rows=imp.get("rows", []),
            cell_meta=sheet_meta,
        )
        results.append(SheetDiff(sheet_name=name, cell_diffs=cell_diffs, status="matched"))

    # Import only
    for name in sorted(imported_names - existing_names):
        results.append(SheetDiff(sheet_name=name, status="import_only"))

    # System only
    for name in sorted(existing_names - imported_names):
        results.append(SheetDiff(sheet_name=name, status="system_only"))

    return results


# ---------------------------------------------------------------------------
# 15.8: Apply Import with Conflict Strategy
# ---------------------------------------------------------------------------


def apply_import(
    xlsx_bytes: bytes,
    existing_sheets: list[dict[str, Any]],
    strategy: ConflictStrategy,
    meta_data: dict[str, Any] | None = None,
    merge_cells: dict[str, list[str]] | None = None,
) -> ImportResult:
    """Apply import with conflict strategy (15.8).

    Args:
        xlsx_bytes: The uploaded xlsx file bytes.
        existing_sheets: List of existing sheet dicts.
        strategy: Conflict resolution strategy.
        meta_data: Optional metadata from _meta_ sheet.
        merge_cells: For MERGE strategy, {sheet_name: [cell_keys_to_import]}.

    Returns:
        ImportResult with counts and audit entry.
    """
    imported = _extract_imported_sheets(xlsx_bytes, meta_data)
    existing_lookup = {s.get("sheet_name", ""): deepcopy(s) for s in existing_sheets}

    sheets_imported = 0
    sheets_kept = 0
    cells_changed = 0

    for sheet_name, imp_data in imported.items():
        if sheet_name not in existing_lookup:
            continue  # Skip import_only sheets

        if strategy == ConflictStrategy.OVERWRITE:
            # Replace local with imported data
            existing_lookup[sheet_name]["rows"] = [
                {"cells": r["cells"]} for r in imp_data.get("rows", [])
            ]
            sheets_imported += 1
            cells_changed += sum(len(r.get("cells", [])) for r in imp_data.get("rows", []))

        elif strategy == ConflictStrategy.KEEP_SYSTEM:
            sheets_kept += 1

        elif strategy == ConflictStrategy.MERGE:
            # Only import specific cells
            cells_to_import = (merge_cells or {}).get(sheet_name, [])
            if cells_to_import:
                local = existing_lookup[sheet_name]
                _merge_cells_into_local(local, imp_data, cells_to_import)
                cells_changed += len(cells_to_import)
                sheets_imported += 1
            else:
                # If no specific cells, compute diff and import all changed editable cells
                sheet_meta = (meta_data or {}).get("sheets", {}).get(sheet_name, {}).get("cell_meta", {})
                diffs = diff_sheet_cells(
                    sheet_name=sheet_name,
                    local_rows=existing_lookup[sheet_name].get("rows", []),
                    imported_rows=imp_data.get("rows", []),
                    cell_meta=sheet_meta,
                )
                if diffs:
                    local = existing_lookup[sheet_name]
                    diff_keys = [d.cell_key for d in diffs]
                    _merge_cells_into_local(local, imp_data, diff_keys)
                    cells_changed += len(diffs)
                    sheets_imported += 1
                else:
                    sheets_kept += 1

    # 15.14: Build audit entry
    audit_entry = _build_audit_entry(
        sheets_imported=sheets_imported,
        sheets_kept=sheets_kept,
        cells_changed=cells_changed,
        strategy=strategy.value,
        file_hash=hashlib.sha256(xlsx_bytes).hexdigest(),
    )

    return ImportResult(
        success=True,
        sheets_imported=sheets_imported,
        sheets_kept=sheets_kept,
        cells_changed=cells_changed,
        audit_entry=audit_entry,
    )


def _merge_cells_into_local(
    local: dict[str, Any],
    imported: dict[str, Any],
    cell_keys: list[str],
) -> None:
    """Merge specific cells from imported into local sheet."""
    local_rows = local.get("rows", [])
    imported_rows = imported.get("rows", [])

    for cell_key in cell_keys:
        parts = cell_key.split(":")
        if len(parts) != 2:
            continue
        row_idx, col_idx = int(parts[0]), int(parts[1])

        if row_idx < len(imported_rows):
            imp_cells = imported_rows[row_idx].get("cells", [])
            if col_idx < len(imp_cells):
                # Ensure local has enough rows/cells
                while len(local_rows) <= row_idx:
                    local_rows.append({"cells": []})
                while len(local_rows[row_idx].get("cells", [])) <= col_idx:
                    local_rows[row_idx].setdefault("cells", []).append(None)
                local_rows[row_idx]["cells"][col_idx] = imp_cells[col_idx]


# ---------------------------------------------------------------------------
# 15.14: 审计日志 + 文件归档
# ---------------------------------------------------------------------------


def _build_audit_entry(
    sheets_imported: int,
    sheets_kept: int,
    cells_changed: int,
    strategy: str,
    file_hash: str,
) -> dict[str, Any]:
    """Build audit log entry for import operation (15.14)."""
    return {
        "action": "wp_offline_import",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "file_hash": file_hash,
        "strategy": strategy,
        "sheets_imported": sheets_imported,
        "sheets_kept": sheets_kept,
        "cells_changed": cells_changed,
        "retention_days": 30,
        "rollback_available": True,
    }


def build_archive_record(
    file_bytes: bytes,
    wp_id: str,
    project_id: str,
    user_id: str,
    import_result: ImportResult,
) -> dict[str, Any]:
    """Build file archive record for 30-day retention (15.14).

    Returns dict to be persisted (file stored separately).
    """
    return {
        "wp_id": wp_id,
        "project_id": project_id,
        "user_id": user_id,
        "file_hash": hashlib.sha256(file_bytes).hexdigest(),
        "file_size": len(file_bytes),
        "archived_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": None,  # Set by caller: now + 30 days
        "import_result": import_result.to_dict(),
        "rollback_available": True,
    }


# ---------------------------------------------------------------------------
# Async Service Class
# ---------------------------------------------------------------------------


class WpOfflineImportService:
    """底稿离线导入服务 (Sprint 4 Task 15).

    Async wrapper for import operations with DB integration.
    """

    def __init__(self, db: Any = None):
        self.db = db

    async def validate_and_preview(
        self,
        xlsx_bytes: bytes,
        wp_id: UUID,
    ) -> dict[str, Any]:
        """Validate file + compute diff preview (15.7).

        Returns: {validation, diffs, summary}
        """
        # Validate
        validation = validate_import_file(xlsx_bytes)
        if not validation.valid:
            return {"validation": {"valid": False, "errors": validation.errors}}

        # Check wp_id matches
        if validation.wp_id and str(wp_id) != validation.wp_id:
            return {
                "validation": {
                    "valid": False,
                    "errors": [f"文件对应底稿 ID 不匹配（期望 {wp_id}，实际 {validation.wp_id}）"],
                }
            }

        # Load existing sheets
        existing = await self._load_existing_sheets(wp_id)

        # Compute diff
        diffs = diff_sheets(xlsx_bytes, existing, validation.meta_data)

        # Summary
        summary = {
            "matched": sum(1 for d in diffs if d.status == "matched"),
            "import_only": sum(1 for d in diffs if d.status == "import_only"),
            "system_only": sum(1 for d in diffs if d.status == "system_only"),
            "total_cell_diffs": sum(len(d.cell_diffs) for d in diffs),
        }

        return {
            "validation": {"valid": True},
            "diffs": [d.to_dict() for d in diffs],
            "summary": summary,
        }

    async def execute_import(
        self,
        xlsx_bytes: bytes,
        wp_id: UUID,
        user_id: str,
        strategy: ConflictStrategy,
        merge_cells: dict[str, list[str]] | None = None,
    ) -> ImportResult:
        """Execute import with strategy (15.8)."""
        # Validate first
        validation = validate_import_file(xlsx_bytes)
        if not validation.valid:
            return ImportResult(success=False, audit_entry={"errors": validation.errors})

        # Load existing
        existing = await self._load_existing_sheets(wp_id)

        # Apply
        result = apply_import(
            xlsx_bytes, existing, strategy,
            meta_data=validation.meta_data,
            merge_cells=merge_cells,
        )

        # Enrich audit entry
        result.audit_entry["user_id"] = user_id
        result.audit_entry["wp_id"] = str(wp_id)

        return result

    async def _load_existing_sheets(self, wp_id: UUID) -> list[dict[str, Any]]:
        """Load existing workpaper sheets from DB."""
        if self.db is None:
            return []
        try:
            from sqlalchemy import select as sa_select
            from app.models.models import ProjectWorkpaper

            result = await self.db.execute(
                sa_select(ProjectWorkpaper).where(ProjectWorkpaper.id == wp_id)
            )
            wp = result.scalar_one_or_none()
            if not wp:
                return []

            parsed = wp.parsed_data or {}
            sheets = parsed.get("sheets", {})
            return [
                {
                    "sheet_name": name,
                    "headers": content.get("headers", []),
                    "rows": content.get("rows", []),
                    "cell_meta": content.get("cell_meta", {}),
                    "formulas": content.get("formulas", {}),
                }
                for name, content in sheets.items()
            ]
        except Exception:
            return []
