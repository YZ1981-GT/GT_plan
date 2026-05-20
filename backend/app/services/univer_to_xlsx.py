"""将 Univer IWorkbookData 格式回写为 xlsx 文件

接收前端 Univer 编辑器的 snapshot 数据，转换为 openpyxl Workbook 并保存。
保留单元格值、公式、基本样式（粗体、字体颜色、背景色、对齐、边框）。
"""

from __future__ import annotations

import logging
from datetime import timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def univer_data_to_xlsx(data: dict[str, Any], output_path: str) -> dict:
    """将 Univer IWorkbookData 写入 xlsx 文件。

    返回 {"sheets": int, "rows": int, "cells": int}
    """
    wb = Workbook()
    # 删除默认 sheet
    if wb.active:
        wb.remove(wb.active)

    sheets = data.get("sheets", {})
    sheet_order = data.get("sheetOrder", list(sheets.keys()))
    # Univer 全局 styles 表（cell.s 可能是 styleId 字符串引用）
    global_styles = data.get("styles", {}) or {}
    total_rows = 0
    total_cells = 0

    for sheet_id in sheet_order:
        sheet = sheets.get(sheet_id)
        if not sheet:
            continue

        ws = wb.create_sheet(title=sheet.get("name", sheet_id)[:31])
        cell_data = sheet.get("cellData", {})

        # 列宽
        col_data = sheet.get("columnData", {})
        for ci_str, col_info in col_data.items():
            ci = int(ci_str) + 1
            w = col_info.get("w")
            if w and w > 0:
                ws.column_dimensions[get_column_letter(ci)].width = w / 7  # 像素→Excel 列宽

        # 行高
        row_data_meta = sheet.get("rowData", {})
        for ri_str, row_info in row_data_meta.items():
            ri = int(ri_str) + 1
            h = row_info.get("h")
            if h and h > 0:
                ws.row_dimensions[ri].height = h

        # 合并单元格
        for merge in sheet.get("mergeData", []):
            sr = merge.get("startRow", 0) + 1
            er = merge.get("endRow", 0) + 1
            sc = merge.get("startColumn", 0) + 1
            ec = merge.get("endColumn", 0) + 1
            if sr <= er and sc <= ec:
                ws.merge_cells(
                    start_row=sr, start_column=sc,
                    end_row=er, end_column=ec,
                )

        # 冻结窗格
        freeze = sheet.get("freeze")
        if freeze:
            fr = freeze.get("ySplit", 0)
            fc = freeze.get("xSplit", 0)
            if fr > 0 or fc > 0:
                ws.freeze_panes = ws.cell(row=fr + 1, column=fc + 1)

        # 写入单元格
        for r_str, row_cells in cell_data.items():
            r = int(r_str)
            total_rows += 1
            for c_str, cell_info in row_cells.items():
                c = int(c_str)
                total_cells += 1
                cell = ws.cell(row=r + 1, column=c + 1)

                # 公式优先
                formula = cell_info.get("f")
                if formula:
                    cell.value = formula if formula.startswith("=") else f"={formula}"
                else:
                    val = cell_info.get("v")
                    cell_type = cell_info.get("t")
                    if val is not None:
                        if cell_type == 2:  # number
                            try:
                                cell.value = float(val)
                            except (ValueError, TypeError):
                                cell.value = val
                        elif cell_type == 4:  # boolean
                            cell.value = bool(val)
                        else:
                            cell.value = val

                # 样式
                style = cell_info.get("s")
                if style:
                    # Univer style 可能是 styleId 字符串引用，需从全局 styles 表查
                    if isinstance(style, str):
                        style = global_styles.get(style)
                    if isinstance(style, dict):
                        _apply_style(cell, style)

    # 确保至少有一个 sheet
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return {"sheets": len(sheet_order), "rows": total_rows, "cells": total_cells}


def _normalize_color_argb(rgb_value: Any) -> str | None:
    """将 Univer 颜色转为 openpyxl 期望的 aRGB hex（8 字符）。

    输入示例：
    - "#FF5149" → "FFFF5149"  (加 alpha=FF)
    - "FF5149" → "FFFF5149"
    - "#A1B2C3D4" → "A1B2C3D4"  (已是 aRGB)
    - "rgb(255,0,0)" → "FFFF0000"
    - 无效或空 → None
    """
    if not rgb_value:
        return None
    s = str(rgb_value).strip()
    if not s:
        return None
    # rgb(...) 形式
    if s.lower().startswith("rgb("):
        try:
            inner = s[4:-1]
            parts = [int(x.strip()) for x in inner.split(",")[:3]]
            if len(parts) >= 3:
                return "FF" + "".join(f"{c:02X}" for c in parts[:3])
        except Exception:
            return None
    # 去掉 # 前缀
    if s.startswith("#"):
        s = s[1:]
    # 仅保留 hex 字符
    s = "".join(c for c in s if c in "0123456789abcdefABCDEF")
    if len(s) == 8:
        return s.upper()
    if len(s) == 6:
        return ("FF" + s).upper()
    if len(s) == 3:
        # 短 hex #RGB → #RRGGBB → FFRRGGBB
        full = "".join(c * 2 for c in s)
        return ("FF" + full).upper()
    return None


def _apply_style(cell: Any, style: dict) -> None:
    """将 Univer 样式应用到 openpyxl Cell"""
    font_kwargs: dict[str, Any] = {}

    if style.get("bl"):
        font_kwargs["bold"] = True
    if style.get("it"):
        font_kwargs["italic"] = True
    if style.get("fs"):
        font_kwargs["size"] = style["fs"]
    if style.get("ff"):
        font_kwargs["name"] = style["ff"]

    cl = style.get("cl")
    if cl and isinstance(cl, dict):
        argb = _normalize_color_argb(cl.get("rgb"))
        if argb:
            font_kwargs["color"] = argb

    ul = style.get("ul")
    if ul and isinstance(ul, dict) and ul.get("s"):
        font_kwargs["underline"] = "single"

    if font_kwargs:
        try:
            cell.font = Font(**font_kwargs)
        except Exception:
            # 颜色异常等场景降级（不应阻断整体保存）
            font_kwargs.pop("color", None)
            try:
                cell.font = Font(**font_kwargs)
            except Exception:
                pass

    # 背景色
    bg = style.get("bg")
    if bg and isinstance(bg, dict):
        argb = _normalize_color_argb(bg.get("rgb"))
        if argb and argb[2:] not in ("000000", "FFFFFF"):
            try:
                cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
            except Exception:
                pass

    # 对齐
    align_kwargs: dict[str, Any] = {}
    ht = style.get("ht")
    if ht is not None:
        ha_map = {0: "left", 1: "center", 2: "right"}
        align_kwargs["horizontal"] = ha_map.get(ht, "left")
    vt = style.get("vt")
    if vt is not None:
        va_map = {0: "top", 1: "center", 2: "bottom"}
        align_kwargs["vertical"] = va_map.get(vt, "center")
    tb = style.get("tb")
    if tb == 3:
        align_kwargs["wrap_text"] = True
    if align_kwargs:
        cell.alignment = Alignment(**align_kwargs)

    # 边框
    bd = style.get("bd")
    if bd and isinstance(bd, dict):
        border_kwargs: dict[str, Any] = {}
        style_map = {1: "thin", 2: "medium", 3: "thick", 5: "dashed", 8: "dotted"}
        default_color = "FFD9D9D9"  # aRGB 8 字符
        for side_key, side_name in [("l", "left"), ("r", "right"), ("t", "top"), ("b", "bottom")]:
            side_info = bd.get(side_key)
            if side_info:
                bs = style_map.get(side_info.get("s", 1), "thin")
                argb = _normalize_color_argb((side_info.get("cl") or {}).get("rgb")) or default_color
                try:
                    border_kwargs[side_name] = Side(style=bs, color=argb)
                except Exception:
                    border_kwargs[side_name] = Side(style=bs, color=default_color)
        if border_kwargs:
            try:
                cell.border = Border(**border_kwargs)
            except Exception:
                pass


def univer_snapshot_to_structure(snapshot: dict[str, Any]) -> dict[str, Any]:
    """将 Univer snapshot 转换为 structure.json 格式（供三式联动使用）"""
    sheets_data = snapshot.get("sheets", {})
    sheet_order = snapshot.get("sheetOrder", list(sheets_data.keys()))
    all_rows: list[dict] = []
    sheet_names: list[str] = []

    for sheet_id in sheet_order:
        sheet = sheets_data.get(sheet_id)
        if not sheet:
            continue
        sheet_names.append(sheet.get("name", sheet_id))

        cell_data = sheet.get("cellData", {})
        if not cell_data:
            continue

        max_row = max((int(k) for k in cell_data.keys()), default=-1)
        for r in range(max_row + 1):
            row_cells = cell_data.get(str(r), cell_data.get(r, {}))
            cells: list[dict] = []
            if row_cells:
                max_col = max((int(k) for k in row_cells.keys()), default=-1)
                for c in range(max_col + 1):
                    cell = row_cells.get(str(c), row_cells.get(c))
                    if cell:
                        # cell.s 可能是 styleId 字符串引用，安全提取 bold
                        s = cell.get("s")
                        bold = False
                        if isinstance(s, dict):
                            bold = bool(s.get("bl"))
                        cells.append({
                            "value": cell.get("v", ""),
                            "formula": cell.get("f"),
                            "bold": bold,
                        })
                    else:
                        cells.append({"value": "", "formula": None})
            all_rows.append({"cells": cells})

    return {
        "rows": all_rows,
        "sheets": [{"name": n} for n in sheet_names],
        "sheet_names": sheet_names,
        "metadata": {
            "editor": "univer",
            "saved_at": __import__("datetime").datetime.now(timezone.utc).isoformat(),
        },
    }
