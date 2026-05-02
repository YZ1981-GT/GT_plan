"""
底稿精细化规则引擎

加载 wp_fine_rules/{code}.json 精细化规则，执行：
1. 精确提取审定表关键行数据（按固定行号+列号）
2. 执行交叉引用校验（审定表↔明细表↔余额调节表↔试算表↔报表）
3. 执行审计检查（余额核对/完整性/变动分析）
4. 生成结构化提取结果（供 parsed_data 和 AI 使用）
"""
import json
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional
from uuid import UUID

logger = logging.getLogger(__name__)

_FINE_RULES_DIR = Path(__file__).parent.parent.parent / "data" / "wp_fine_rules"


def load_fine_rule(wp_code: str) -> Optional[dict]:
    """加载精细化规则文件"""
    # 精确匹配
    code = wp_code.upper().split("-")[0] if "-" in wp_code else wp_code.upper()
    for fp in _FINE_RULES_DIR.glob("*.json"):
        try:
            with open(fp, "r", encoding="utf-8-sig") as f:
                rule = json.load(f)
            if rule.get("wp_code", "").upper() == code:
                return rule
        except Exception:
            continue
    return None


def list_fine_rules() -> list[dict]:
    """列出所有精细化规则"""
    rules = []
    if not _FINE_RULES_DIR.exists():
        return rules
    for fp in sorted(_FINE_RULES_DIR.glob("*.json")):
        try:
            with open(fp, "r", encoding="utf-8-sig") as f:
                rule = json.load(f)
            rules.append({
                "wp_code": rule.get("wp_code"),
                "name": rule.get("name"),
                "sheets": len(rule.get("sheet_rules", rule.get("sheets", []))),
                "checks": len(rule.get("audit_checks", [])),
                "cross_refs": len(rule.get("cross_references", [])),
                "file": fp.name,
            })
        except Exception:
            continue
    return rules


def extract_with_fine_rule(
    file_path: str,
    wp_code: str,
    project_id: str = "",
    year: int = 0,
) -> dict[str, Any]:
    """使用精细化规则从底稿Excel提取结构化数据

    返回：
    {
        "wp_code": "E1",
        "sheets": {
            "E1-1": {
                "rows": {"cash": {"closing_audited": 1000, ...}, "total": {...}},
                "raw_data": [...],
            },
            "E1-2": {...},
        },
        "summary": {"total_closing_audited": 1000, ...},
        "checks": [{"code": "E1-CHK-01", "passed": true, ...}],
        "cross_refs": [{"from": "...", "to": "...", "matched": true}],
    }
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}

    rule = load_fine_rule(wp_code)
    if not rule:
        return {"error": f"No fine rule for {wp_code}"}

    fp = Path(file_path)
    if not fp.exists():
        return {"error": f"File not found: {file_path}"}

    # 收集所有需要打开的文件（主文件 + source_file 指定的其他文件）
    files_to_open: dict[str, Path] = {"_main": fp}
    base_dir = fp.parent
    for sr in rule.get("sheet_rules", []):
        sf = sr.get("source_file", "")
        if sf and sf not in files_to_open:
            sf_path = base_dir / sf
            if not sf_path.exists():
                # 尝试在模板目录中查找
                sf_path = base_dir.parent / sf
            if sf_path.exists():
                files_to_open[sf] = sf_path

    # 打开所有工作簿
    workbooks: dict[str, Any] = {}
    try:
        for key, path in files_to_open.items():
            try:
                workbooks[key] = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            except Exception as e:
                logger.warning("Cannot open %s: %s", path, e)
    except Exception as e:
        return {"error": f"Cannot open files: {e}"}

    result: dict[str, Any] = {
        "wp_code": rule["wp_code"],
        "name": rule.get("name", rule.get("wp_name", rule["wp_code"])),
        "sheets": {},
        "summary": {},
        "checks": [],
        "cross_refs": [],
    }

    # 遍历规则中的每个Sheet
    for sheet_rule in rule.get("sheet_rules", rule.get("sheets", [])):
        code = sheet_rule.get("code", sheet_rule.get("exact_name", ""))
        # 优先用 exact_name 精确匹配
        exact_name = sheet_rule.get("exact_name", "")
        pattern = exact_name or sheet_rule.get("sheet_pattern", sheet_rule.get("name_pattern", ""))
        sheet_type = sheet_rule.get("type", "")

        # 确定在哪个工作簿中查找
        source_file = sheet_rule.get("source_file", "")
        wb = workbooks.get(source_file) or workbooks.get("_main")
        if not wb:
            result["sheets"][code] = {"found": False, "reason": "workbook not available"}
            continue

        # 匹配Sheet
        ws = _find_sheet(wb, pattern)
        if not ws:
            result["sheets"][code] = {"found": False}
            continue

        sheet_data: dict[str, Any] = {"found": True, "type": sheet_type, "rows": {}}

        # 按类型提取 — 兼容新旧两种列定义格式
        # 旧格式: sheet_rule["layout"]["columns"]
        # 新格式: sheet_rule["columns"]
        effective_rule = sheet_rule
        if "layout" in sheet_rule and "columns" not in sheet_rule:
            # 旧格式转换：将 layout.columns 提升到 sheet_rule 级别供提取函数使用
            effective_rule = {**sheet_rule, "columns": sheet_rule["layout"]["columns"]}

        if sheet_type == "summary" and "key_rows" in effective_rule:
            sheet_data["rows"] = _extract_summary_rows(ws, effective_rule)
        elif sheet_type == "detail":
            sheet_data["detail_rows"] = _extract_detail_rows(ws, effective_rule)
        elif sheet_type == "adjustment":
            sheet_data["adjustments"] = _extract_adjustments(ws, effective_rule)

        result["sheets"][code] = sheet_data

    # 关闭所有工作簿
    for wb in workbooks.values():
        try:
            wb.close()
        except Exception:
            pass

    # 构建摘要
    summary_code = f"{rule['wp_code']}-1"
    summary_sheet = result["sheets"].get(summary_code, {})
    summary_rows = summary_sheet.get("rows", {})
    total_row = summary_rows.get("total", {})
    result["summary"] = {
        "closing_audited": total_row.get("closing_audited") or total_row.get("closing_balance"),
        "opening_audited": total_row.get("opening_audited"),
        "change_amount": total_row.get("change_amount"),
    }

    # 执行审计检查（实际校验）
    result["checks"] = _run_audit_checks(rule, result["sheets"], result["summary"])

    logger.info("extract_with_fine_rule: %s → %d sheets extracted, %d checks",
                wp_code, sum(1 for s in result["sheets"].values() if s.get("found")),
                len(result["checks"]))
    return result


def _find_sheet(wb, pattern: str):
    """根据精确名称或模式匹配Sheet"""
    import re
    # 优先精确匹配
    for ws in wb.worksheets:
        if ws.title == pattern:
            return ws
    # 降级模式匹配
    parts = pattern.split("|")
    for ws in wb.worksheets:
        title = ws.title
        for p in parts:
            if p in title or re.search(p, title, re.IGNORECASE):
                return ws
    return None


def _safe_num(val) -> Optional[float]:
    """安全转换为数字"""
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, Decimal):
            return float(val)
        s = str(val).strip().replace(",", "")
        if not s or s in ("None", "-", "—", ""):
            return None
        return float(s)
    except (ValueError, InvalidOperation):
        return None


def _extract_summary_rows(ws, sheet_rule: dict) -> dict[str, dict]:
    """从审定表提取关键行数据

    兼容两种 key_rows 格式：
    - 旧格式: {"cash": {"row": 7, "label": "库存现金"}}
    - 新格式: {"section1_total": 13} 或 {"data_items": [8, 9, 10]}
    """
    key_rows = sheet_rule.get("key_rows", {})
    columns = sheet_rule.get("layout", {}).get("columns", sheet_rule.get("columns", {}))
    result = {}

    for key, row_def in key_rows.items():
        # 兼容新格式：值为整数（单行号）
        if isinstance(row_def, int):
            row_num = row_def
            row_data: dict[str, Any] = {
                "label": str(ws.cell(row_num, 1).value or ""),
                "row": row_num,
                "is_total": "total" in key.lower(),
            }
            for col_key, col_def in columns.items():
                if col_key == "label":
                    continue
                col_num = col_def.get("col", 0) if isinstance(col_def, dict) else 0
                if col_num > 0:
                    row_data[col_key] = _safe_num(ws.cell(row_num, col_num).value)
            result[key] = row_data
            continue

        # 兼容新格式：值为数组（多行号）— 逐行提取
        if isinstance(row_def, list):
            for i, rn in enumerate(row_def):
                if not isinstance(rn, int):
                    continue
                item_key = f"{key}_{i}"
                item_data: dict[str, Any] = {
                    "label": str(ws.cell(rn, 1).value or ""),
                    "row": rn,
                }
                for col_key, col_def in columns.items():
                    if col_key == "label":
                        continue
                    col_num = col_def.get("col", 0) if isinstance(col_def, dict) else 0
                    if col_num > 0:
                        item_data[col_key] = _safe_num(ws.cell(rn, col_num).value)
                result[item_key] = item_data
            continue

        # 旧格式：值为字典
        if not isinstance(row_def, dict):
            continue
        row_num = row_def.get("row", 0)
        if row_num <= 0:
            continue

        row_data = {
            "label": row_def.get("label", ""),
            "row": row_num,
            "is_total": row_def.get("is_total", False),
            "account_code": row_def.get("account_code", ""),
        }

        # 按列定义提取值
        for col_key, col_def in columns.items():
            if col_key == "label":
                continue
            col_num = col_def.get("col", 0) if isinstance(col_def, dict) else 0
            if col_num > 0:
                row_data[col_key] = _safe_num(ws.cell(row_num, col_num).value)

        result[key] = row_data

    return result


def _extract_detail_rows(ws, sheet_rule: dict) -> list[dict]:
    """从明细表提取数据行"""
    data_start = sheet_rule.get("data_start_row", 1)
    columns = sheet_rule.get("layout", {}).get("columns", sheet_rule.get("columns", {}))
    max_row = ws.max_row or 100
    rows = []

    for r in range(data_start, min(max_row + 1, data_start + 200)):
        row_data = {}
        has_data = False
        for col_key, col_def in columns.items():
            col_num = col_def.get("col", 0) if isinstance(col_def, dict) else 0
            if col_num > 0:
                val = ws.cell(r, col_num).value
                if val is not None:
                    has_data = True
                row_data[col_key] = val
        if has_data:
            row_data["_row"] = r
            rows.append(row_data)

    return rows


def _extract_adjustments(ws, sheet_rule: dict) -> list[dict]:
    """从调整分录表提取"""
    data_start = sheet_rule.get("data_start_row", 1)
    columns = sheet_rule.get("layout", {}).get("columns", sheet_rule.get("columns", {}))
    max_row = ws.max_row or 50
    items = []

    for r in range(data_start, min(max_row + 1, data_start + 100)):
        row_data = {}
        has_data = False
        for col_key, col_def in columns.items():
            col_num = col_def.get("col", 0)
            if col_num > 0:
                val = ws.cell(r, col_num).value
                if val is not None:
                    has_data = True
                row_data[col_key] = val
        if has_data:
            items.append(row_data)

    return items


# ═══════════════════════════════════════════
# 审计检查执行引擎
# ═══════════════════════════════════════════

def _run_audit_checks(rule: dict, sheets: dict, summary: dict) -> list[dict]:
    """执行精细化规则中定义的审计检查

    根据 check.type 分派到不同的校验逻辑：
    - balance: 余额核对（审定表合计 vs 试算表/报表）
    - cross_ref: 交叉引用（审定表行 vs 明细表合计）
    - movement: 变动校验（期初+变动=期末）
    - completeness: 完整性（账户数核对）
    - reconciliation: 余额调节（账面±未达=对账单±未达）
    - confirmation: 函证核对
    - analysis: 分析程序
    - cutoff: 截止测试
    """
    results = []
    wp_code = rule.get("wp_code", "")
    summary_code = f"{wp_code}-1"
    summary_sheet = sheets.get(summary_code, {})
    summary_rows = summary_sheet.get("rows", {})

    for check in rule.get("audit_checks", []):
        check_code = check.get("id", check.get("code", ""))
        check_type = check.get("type", "")
        severity = check.get("severity", "info")
        description = check.get("description", "")

        passed = None
        actual = None
        expected = None
        diff = None
        message = ""

        try:
            if check_type == "balance":
                passed, actual, expected, diff, message = _check_balance(
                    check_code, summary_rows, sheets, rule
                )
            elif check_type == "cross_ref":
                passed, actual, expected, diff, message = _check_cross_ref(
                    check_code, summary_rows, sheets, rule
                )
            elif check_type == "movement":
                passed, message = _check_movement(summary_rows)
                actual = expected = diff = None
            elif check_type == "formula":
                passed, actual, expected, diff, message = _check_formula(
                    check_code, summary_rows, sheets, rule
                )
            elif check_type == "completeness":
                passed, message = _check_completeness(check_code, sheets, rule)
            elif check_type == "aging":
                passed, actual, expected, diff, message = _check_aging(
                    check_code, summary_rows, sheets, rule
                )
            elif check_type == "confirmation":
                passed, message = _check_confirmation(check_code, sheets, rule)
            elif check_type in ("check", "analysis", "cutoff"):
                # 这些是定性检查，需要人工判断或LLM辅助
                # 从对应Sheet检查是否有数据填写
                passed, message = _check_sheet_filled(check_code, sheets, rule)
            elif check_type == "reconciliation":
                passed, message = _check_reconciliation(check_code, sheets, rule)
            else:
                message = f"未知检查类型: {check_type}"
        except Exception as e:
            message = f"检查执行异常: {e}"

        results.append({
            "code": check_code,
            "type": check_type,
            "severity": severity,
            "description": description,
            "passed": passed,
            "actual": actual,
            "expected": expected,
            "diff": diff,
            "message": message,
        })

    return results


def _check_balance(check_code: str, summary_rows: dict, sheets: dict, rule: dict):
    """余额核对检查"""
    total = summary_rows.get("total", {})
    total_closing = total.get("closing_audited") or total.get("closing_balance")
    tb_row = summary_rows.get("tb_balance", {})
    tb_closing = tb_row.get("closing_audited") or tb_row.get("closing_balance")

    if "CHK-01" in check_code:
        # 审定表合计 vs 试算平衡表数
        if total_closing is not None and tb_closing is not None:
            diff = abs(total_closing - tb_closing)
            passed = diff < 0.01
            return passed, total_closing, tb_closing, round(diff, 2), \
                "通过" if passed else f"差异 {diff:.2f}"
        return None, total_closing, tb_closing, None, "数据不完整，无法校验"

    if "CHK-02" in check_code:
        # 审定表合计 vs 报表（需要外部数据，标记待验证）
        if total_closing is not None:
            return None, total_closing, None, None, "需要报表数据验证（REPORT('BS-002','期末')）"
        return None, None, None, None, "审定表合计为空"

    return None, None, None, None, "未匹配的余额检查"


def _check_cross_ref(check_code: str, summary_rows: dict, sheets: dict, rule: dict):
    """交叉引用检查"""
    wp_code = rule.get("wp_code", "")

    if "CHK-03" in check_code:
        # 库存现金审定数 vs 现金明细表合计
        cash_row = summary_rows.get("cash", {})
        cash_audited = cash_row.get("closing_audited") or cash_row.get("closing_balance")
        detail_sheet = sheets.get(f"{wp_code}-2", {})
        detail_rows = detail_sheet.get("detail_rows", [])
        if detail_rows:
            # 合计明细表的审定数列
            detail_total = sum(
                _safe_num(r.get("closing_audited") or r.get("closing_rmb")) or 0
                for r in detail_rows
            )
            if cash_audited is not None:
                diff = abs(cash_audited - detail_total)
                passed = diff < 0.01
                return passed, cash_audited, detail_total, round(diff, 2), \
                    "通过" if passed else f"差异 {diff:.2f}"
        return None, cash_audited, None, None, "现金明细表无数据"

    if "CHK-04" in check_code:
        # 银行存款审定数 vs 银行明细表合计
        bank_row = summary_rows.get("bank_deposit", {})
        bank_audited = bank_row.get("closing_audited") or bank_row.get("closing_balance")
        # 银行明细可能在 E1-3-rmb 或 E1-3
        for detail_code in [f"{wp_code}-3-rmb", f"{wp_code}-3"]:
            detail_sheet = sheets.get(detail_code, {})
            detail_rows = detail_sheet.get("detail_rows", [])
            if detail_rows:
                detail_total = sum(
                    _safe_num(r.get("closing_audited") or r.get("closing_balance")) or 0
                    for r in detail_rows
                )
                if bank_audited is not None:
                    diff = abs(bank_audited - detail_total)
                    passed = diff < 0.01
                    return passed, bank_audited, detail_total, round(diff, 2), \
                        "通过" if passed else f"差异 {diff:.2f}"
        return None, bank_audited, None, None, "银行明细表无数据"

    return None, None, None, None, "未匹配的交叉引用检查"


def _check_movement(summary_rows: dict):
    """变动校验：各行期初+变动=期末"""
    issues = []
    for key, row in summary_rows.items():
        if row.get("is_total") or key in ("tb_balance", "diff", "overseas"):
            continue
        opening = row.get("opening_audited")
        closing = row.get("closing_audited") or row.get("closing_balance")
        change = row.get("change_amount")
        if opening is not None and closing is not None and change is not None:
            expected = opening + change
            if abs(expected - closing) > 0.01:
                issues.append(f"{row.get('label', key)}: 期初{opening}+变动{change}={expected} ≠ 期末{closing}")

    if not issues:
        return True, "各行变动校验通过"
    return False, f"{len(issues)}行变动不一致: " + "; ".join(issues[:3])


def _check_formula(check_code: str, summary_rows: dict, sheets: dict, rule: dict):
    """公式校验：净值 = 原值 - 坏账准备（逐行）"""
    wp_code = rule.get("wp_code", "")
    summary_code = f"{wp_code}-1"
    sheet = sheets.get(summary_code, {})
    rows = sheet.get("rows", {})

    # 查找三段小计行
    gross_key = None
    bad_debt_key = None
    net_key = None
    for key, row in rows.items():
        label = (row.get("label") or "").lower()
        if row.get("is_total"):
            if "原值" in label or "gross" in key:
                gross_key = key
            elif "坏账" in label or "bad_debt" in key:
                bad_debt_key = key
            elif "净值" in label or "net" in key or "合计" in label:
                net_key = key

    if not (gross_key and bad_debt_key and net_key):
        return None, None, None, None, "未找到原值/坏账/净值小计行"

    gross = rows[gross_key].get("closing_audited") or rows[gross_key].get("closing_balance") or 0
    bad_debt = rows[bad_debt_key].get("closing_audited") or rows[bad_debt_key].get("closing_balance") or 0
    net = rows[net_key].get("closing_audited") or rows[net_key].get("closing_balance") or 0

    expected_net = gross - bad_debt
    diff = abs(net - expected_net)
    passed = diff < 0.01
    return passed, net, expected_net, round(diff, 2), \
        "通过" if passed else f"净值{net} ≠ 原值{gross}-坏账{bad_debt}={expected_net}"


def _check_completeness(check_code: str, sheets: dict, rule: dict):
    """完整性检查：检查关键Sheet是否有数据"""
    wp_code = rule.get("wp_code", "")

    # 检查明细表是否有数据行
    for code_suffix in ["-2", "-3"]:
        detail_code = f"{wp_code}{code_suffix}"
        detail_sheet = sheets.get(detail_code, {})
        detail_rows = detail_sheet.get("detail_rows", [])
        if detail_rows and len(detail_rows) > 0:
            return True, f"明细表有 {len(detail_rows)} 条数据"

    # 检查是否找到了Sheet但无数据
    found_sheets = sum(1 for s in sheets.values() if s.get("found"))
    total_sheets = len(sheets)
    if found_sheets == 0:
        return None, "未找到任何Sheet数据"
    return None, f"已找到 {found_sheets}/{total_sheets} 个Sheet，明细表暂无数据"


def _check_aging(check_code: str, summary_rows: dict, sheets: dict, rule: dict):
    """账龄相关检查"""
    # CHK-12: 账龄分段合计 = 原值合计
    if "CHK-12" in check_code:
        # 需要从明细表的账龄列汇总，当前精细化提取未细化到账龄列
        return None, None, None, None, "需要明细表账龄列数据验证"

    # CHK-13: 计提比例与政策一致
    if "CHK-13" in check_code:
        return None, None, None, None, "需要与会计政策比对验证"

    # CHK-14: 迁徙率合理性
    if "CHK-14" in check_code:
        return None, None, None, None, "需要上年账龄数据计算迁徙率"

    return None, None, None, None, "未匹配的账龄检查"


def _check_confirmation(check_code: str, sheets: dict, rule: dict):
    """函证检查：函证结果汇总表是否有数据"""
    wp_code = rule.get("wp_code", "")

    # 查找函证结果汇总Sheet
    for code, sheet in sheets.items():
        if "D0-1" in code or "E0-1" in code or "函证结果" in code:
            if sheet.get("found") and sheet.get("detail_rows"):
                count = len(sheet["detail_rows"])
                return True, f"函证结果汇总有 {count} 条记录"
            elif sheet.get("found"):
                return None, "函证结果汇总Sheet已找到但暂无数据"

    return None, "未找到函证结果汇总Sheet"


def _check_sheet_filled(check_code: str, sheets: dict, rule: dict):
    """定性检查：检查对应Sheet是否已填写（有数据行）"""
    # 从check_code推断对应的Sheet
    # 如 D1-CHK-08 → 检查D1-8（贴现/背书）
    # 如 D2-CHK-11 → 检查D2-10（ECL测试）
    wp_code = rule.get("wp_code", "")

    # 遍历所有已找到的Sheet，检查是否有数据
    filled_count = 0
    total_found = 0
    for code, sheet in sheets.items():
        if not sheet.get("found"):
            continue
        total_found += 1
        has_data = bool(
            sheet.get("detail_rows") or
            sheet.get("adjustments") or
            (sheet.get("rows") and any(
                v for k, v in sheet["rows"].items()
                if not k.endswith("_title") and v.get("closing_audited") or v.get("closing_balance")
            ))
        )
        if has_data:
            filled_count += 1

    if total_found == 0:
        return None, "未找到相关Sheet"
    rate = round(filled_count / total_found * 100)
    return True if rate > 50 else None, f"{filled_count}/{total_found} 个Sheet已填写（{rate}%）"


def _check_reconciliation(check_code: str, sheets: dict, rule: dict):
    """余额调节检查"""
    wp_code = rule.get("wp_code", "")

    # 查找余额调节表Sheet
    for code, sheet in sheets.items():
        if "调节" in code or "reconciliation" in sheet.get("type", ""):
            if sheet.get("found"):
                return None, "余额调节表已找到，需人工确认调节项"

    return None, "未找到余额调节表"
