"""科目表管理服务 — 标准科目加载 + 客户科目导入

Validates: Requirements 2.1, 2.2, 2.3, 2.4, 2.5
"""

import csv
import io
import json
import logging
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException, UploadFile
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_platform_models import (
    AccountCategory,
    AccountChart,
    AccountDirection,
    AccountSource,
)
from app.models.audit_platform_schemas import (
    AccountChartResponse,
    AccountImportResult,
    AccountTreeNode,
)

logger = logging.getLogger(__name__)

# Seed data directory
_DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"

# Accounting standard → JSON file mapping
_STANDARD_FILES = {
    "enterprise": "standard_account_chart.json",
}

# Account code prefix → category mapping
_CODE_CATEGORY_MAP: dict[str, AccountCategory] = {
    "1": AccountCategory.asset,
    "2": AccountCategory.liability,
    "3": AccountCategory.equity,
    "4": AccountCategory.equity,    # 部分企业4xxx=权益类（实收资本/资本公积）
    "5": AccountCategory.revenue,   # revenue + expense (5001-5111=revenue, 5401+=expense)
    "6": AccountCategory.revenue,   # 6xxx 损益类（6001-6201=revenue, 6401+=expense）
}

# More precise: 5001-5399 = revenue, 5400+ = expense
_EXPENSE_CODES_START = "54"

# 6xxx 损益类细分
_6XXX_EXPENSE_START = "64"

# 按科目名称关键词推断分类（优先于编码推断）
_NAME_CATEGORY_KEYWORDS: dict[str, AccountCategory] = {
    # 权益类
    "实收资本": AccountCategory.equity,
    "股本": AccountCategory.equity,
    "资本公积": AccountCategory.equity,
    "盈余公积": AccountCategory.equity,
    "利润分配": AccountCategory.equity,
    "未分配利润": AccountCategory.equity,
    "本年利润": AccountCategory.equity,
    "其他综合收益": AccountCategory.equity,
    "其他权益工具": AccountCategory.equity,
    "库存股": AccountCategory.equity,
    # 收入类
    "主营业务收入": AccountCategory.revenue,
    "其他业务收入": AccountCategory.revenue,
    "投资收益": AccountCategory.revenue,
    "公允价值变动损益": AccountCategory.revenue,
    "资产处置收益": AccountCategory.revenue,
    "其他收益": AccountCategory.revenue,
    "汇兑损益": AccountCategory.revenue,
    "营业外收入": AccountCategory.revenue,
    # 费用/成本类
    "主营业务成本": AccountCategory.expense,
    "其他业务成本": AccountCategory.expense,
    "营业税金及附加": AccountCategory.expense,
    "税金及附加": AccountCategory.expense,
    "销售费用": AccountCategory.expense,
    "管理费用": AccountCategory.expense,
    "财务费用": AccountCategory.expense,
    "研发费用": AccountCategory.expense,
    "资产减值损失": AccountCategory.expense,
    "信用减值损失": AccountCategory.expense,
    "营业外支出": AccountCategory.expense,
    "所得税费用": AccountCategory.expense,
    "以前年度损益调整": AccountCategory.expense,
    "生产成本": AccountCategory.expense,
    "制造费用": AccountCategory.expense,
    "研发支出": AccountCategory.expense,
    "工程施工": AccountCategory.expense,
}


def _infer_category(code: str, name: str = "") -> AccountCategory:
    """Infer account category from code prefix and account name.

    优先用科目名称关键词推断（更准确），编码首位作为兜底。
    """
    # 1. 优先按名称关键词匹配
    if name:
        for kw, cat in _NAME_CATEGORY_KEYWORDS.items():
            if kw in name:
                return cat

    if not code:
        return AccountCategory.asset

    first = code[0]

    # 2. 处理 5xxx 损益类细分
    if first == "5":
        if code[:2] >= _EXPENSE_CODES_START:
            return AccountCategory.expense
        return AccountCategory.revenue

    # 3. 处理 6xxx 损益类细分
    if first == "6":
        if code[:2] >= _6XXX_EXPENSE_START:
            return AccountCategory.expense
        return AccountCategory.revenue

    # 4. 处理 4xxx — 需要区分成本类（生产成本4001）和权益类（实收资本4001）
    # 标准体系中 4xxx 是成本类，但部分企业用 4xxx 表示权益类
    # 如果名称没匹配到，按编码兜底
    if first == "4":
        # 标准成本类编码
        if code[:4] in ("4001", "4101", "4301", "4401", "4403"):
            return AccountCategory.expense
        return AccountCategory.equity

    return _CODE_CATEGORY_MAP.get(first, AccountCategory.asset)


def _infer_direction(category: AccountCategory) -> AccountDirection:
    """Infer default direction from category."""
    if category in (AccountDirection.credit,):
        return AccountDirection.credit
    if category in (AccountCategory.liability, AccountCategory.equity, AccountCategory.revenue):
        return AccountDirection.credit
    return AccountDirection.debit


def _infer_level(code: str, parent_code: str | None) -> int:
    """Infer account level from code length or parent."""
    if parent_code:
        return 2
    if len(code) <= 4:
        return 1
    return 2


# ---------------------------------------------------------------------------
# load_standard_template
# ---------------------------------------------------------------------------


async def load_standard_template(
    project_id: UUID,
    accounting_standard: str,
    db: AsyncSession,
) -> list[AccountChartResponse]:
    """Load standard account chart template into account_chart table.

    Reads the JSON seed data file and bulk inserts into account_chart
    with source=standard.

    Validates: Requirements 2.1, 2.2
    """
    file_name = _STANDARD_FILES.get(accounting_standard)
    if not file_name:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的会计准则: {accounting_standard}，当前仅支持: {', '.join(_STANDARD_FILES.keys())}",
        )

    json_path = _DATA_DIR / file_name
    if not json_path.exists():
        raise HTTPException(
            status_code=500,
            detail=f"标准科目表数据文件不存在: {file_name}",
        )

    with open(json_path, encoding="utf-8-sig") as f:
        data = json.load(f)

    accounts_data = data.get("accounts", [])
    if not accounts_data:
        raise HTTPException(status_code=500, detail="标准科目表数据为空")

    # Check existing standard accounts — support incremental update
    existing_result = await db.execute(
        select(AccountChart.account_code).where(
            AccountChart.project_id == project_id,
            AccountChart.source == AccountSource.standard,
            AccountChart.is_deleted == False,  # noqa: E712
        )
    )
    existing_codes = {row[0] for row in existing_result.all()}

    # Filter out already-loaded codes (incremental: only insert new ones)
    new_items = [
        item for item in accounts_data
        if item["code"] not in existing_codes
    ]

    # Bulk insert new accounts only
    if new_items:
        records: list[AccountChart] = []
        for item in new_items:
            record = AccountChart(
                project_id=project_id,
                account_code=item["code"],
                account_name=item["name"],
                direction=AccountDirection(item["direction"]),
                level=item["level"],
                category=AccountCategory(item["category"]),
                parent_code=item.get("parent_code"),
                source=AccountSource.standard,
            )
            records.append(record)
        db.add_all(records)
        await db.flush()
        await db.commit()
        logger.info("增量加载标准科目: project=%s, 新增 %d 个", project_id, len(records))

    # Return all standard accounts (existing + new)
    all_result = await db.execute(
        select(AccountChart).where(
            AccountChart.project_id == project_id,
            AccountChart.source == AccountSource.standard,
            AccountChart.is_deleted == False,  # noqa: E712
        ).order_by(AccountChart.account_code)
    )
    all_records = all_result.scalars().all()

    return [
        AccountChartResponse(
            id=r.id,
            project_id=r.project_id,
            account_code=r.account_code,
            account_name=r.account_name,
            direction=r.direction,
            level=r.level,
            category=r.category,
            parent_code=r.parent_code,
            source=r.source,
            created_at=r.created_at,
        )
        for r in all_records
    ]


# ---------------------------------------------------------------------------
# import_client_chart
# ---------------------------------------------------------------------------


async def import_client_chart(
    project_id: UUID,
    file: UploadFile,
    db: AsyncSession,
    column_mapping: dict[str, str | None] | None = None,
) -> AccountImportResult:
    """Import client account chart from Excel/CSV file.

    Parses account_code, account_name, direction, parent_code.
    Validates required columns (account_code + account_name).
    Writes to account_chart with source=client.

    Args:
        column_mapping: Optional user-confirmed mapping {original_header: standard_field}.
                        If provided, overrides the default _COLUMN_MAP normalization.

    Validates: Requirements 2.3, 2.4, 2.5
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")

    filename_lower = file.filename.lower()
    content = await file.read()

    rows: list[dict] = []

    if filename_lower.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename_lower.endswith((".xlsx", ".xls")):
        rows = _parse_excel(content)  # skip_rows=None → 自动检测表头行
    else:
        raise HTTPException(
            status_code=400,
            detail="不支持的文件格式，请上传 Excel (.xlsx/.xls) 或 CSV (.csv) 文件",
        )

    if not rows:
        raise HTTPException(status_code=400, detail="文件中未解析到有效数据")

    # Apply user-provided column mapping or default normalization
    if column_mapping:
        normalized_rows = _apply_custom_mapping(rows, column_mapping)
    else:
        # Validate required columns (legacy path)
        first_row = rows[0]
        missing_cols = []
        if "account_code" not in first_row and "科目编码" not in first_row:
            missing_cols.append("account_code/科目编码")
        if "account_name" not in first_row and "科目名称" not in first_row:
            missing_cols.append("account_name/科目名称")

        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"缺少必填列: {', '.join(missing_cols)}",
            )
        normalized_rows = _normalize_columns(rows)

    # Validate required fields after normalization
    if normalized_rows:
        sample = normalized_rows[0]
        missing = []
        if "account_code" not in sample:
            missing.append("account_code/科目编码")
        if "account_name" not in sample:
            missing.append("account_name/科目名称")
        if missing and column_mapping:
            raise HTTPException(
                status_code=400,
                detail=f"列映射中缺少必填字段: {', '.join(missing)}",
            )

    # Build records
    records: list[AccountChart] = []
    errors: list[str] = []
    by_category: dict[str, int] = {}

    for i, row in enumerate(normalized_rows, start=2):  # row 2+ (header is row 1)
        code = str(row.get("account_code", "")).strip()
        name = str(row.get("account_name", "")).strip()

        if not code or not name:
            errors.append(f"第{i}行: 科目编码或名称为空，已跳过")
            continue

        parent_code = str(row.get("parent_code", "")).strip() or None
        direction_str = str(row.get("direction", "")).strip().lower()

        # Parse direction
        if direction_str in ("debit", "借", "借方"):
            direction = AccountDirection.debit
        elif direction_str in ("credit", "贷", "贷方"):
            direction = AccountDirection.credit
        else:
            # Infer from code + name
            category = _infer_category(code, name)
            direction = _infer_direction(category)

        category = _infer_category(code, name)
        level = _infer_level(code, parent_code)

        # 优先使用用户映射的 level 和 category（如果有）
        level_str = str(row.get("level", "")).strip()
        if level_str and level_str.isdigit():
            level = int(level_str)

        category_str = str(row.get("category", "")).strip().lower()
        if category_str:
            _cat_map = {
                "资产": AccountCategory.asset, "asset": AccountCategory.asset,
                "负债": AccountCategory.liability, "liability": AccountCategory.liability,
                "权益": AccountCategory.equity, "equity": AccountCategory.equity,
                "收入": AccountCategory.revenue, "revenue": AccountCategory.revenue,
                "费用": AccountCategory.expense, "expense": AccountCategory.expense,
                "成本": AccountCategory.expense, "cost": AccountCategory.expense,
            }
            if category_str in _cat_map:
                category = _cat_map[category_str]

        record = AccountChart(
            project_id=project_id,
            account_code=code,
            account_name=name,
            direction=direction,
            level=level,
            category=category,
            parent_code=parent_code,
            source=AccountSource.client,
        )
        records.append(record)

        cat_name = category.value
        by_category[cat_name] = by_category.get(cat_name, 0) + 1

    if not records:
        raise HTTPException(status_code=400, detail="未解析到有效科目数据")

    # 去重：同一 account_code 只保留第一条（科目表可能有多行同编码的辅助核算维度）
    seen_codes: set[str] = set()
    unique_records: list[AccountChart] = []
    for r in records:
        if r.account_code not in seen_codes:
            seen_codes.add(r.account_code)
            unique_records.append(r)
    records = unique_records

    # 软删除旧的客户科目（覆盖导入）
    from sqlalchemy import update as sa_update
    existing_result = await db.execute(
        select(func.count()).select_from(AccountChart).where(
            AccountChart.project_id == project_id,
            AccountChart.source == AccountSource.client,
            AccountChart.is_deleted == False,  # noqa: E712
        )
    )
    existing_count = existing_result.scalar() or 0
    if existing_count > 0:
        await db.execute(
            sa_update(AccountChart)
            .where(
                AccountChart.project_id == project_id,
                AccountChart.source == AccountSource.client,
                AccountChart.is_deleted == False,  # noqa: E712
            )
            .values(is_deleted=True)
        )
        await db.flush()

    db.add_all(records)
    await db.commit()

    # ── 自动导入其他 sheet 的四表数据 ──
    other_sheets_result = {}
    sheet_diagnostics = []
    if filename_lower.endswith((".xlsx", ".xls")):
        try:
            print(f"[AUTO_IMPORT] 开始自动导入四表数据: project={project_id}, content_size={len(content)}")
            other_sheets_result, sheet_diagnostics = await _auto_import_data_sheets(
                project_id, content, year=None, db=db,
            )
            print(f"[AUTO_IMPORT] 四表导入结果: {other_sheets_result}")
            print(f"[AUTO_IMPORT] 诊断: {sheet_diagnostics}")
        except Exception as e:
            import traceback
            print(f"[AUTO_IMPORT] 自动导入四表数据失败: {e}")
            traceback.print_exc()
            errors.append(f"四表数据自动导入失败: {str(e)}")

    return AccountImportResult(
        total_imported=len(records),
        by_category=by_category,
        errors=errors,
        data_sheets_imported=other_sheets_result,
        sheet_diagnostics=sheet_diagnostics,
    )


# ---------------------------------------------------------------------------
# _auto_import_data_sheets — 自动识别并导入四表数据
# ---------------------------------------------------------------------------

# sheet 类型 → import_service data_type 映射
_SHEET_TYPE_TO_DATA_TYPE = {
    "balance": "tb_balance",
    "ledger": "tb_ledger",
    "aux_balance": "tb_aux_balance",
    "aux_ledger": "tb_aux_ledger",
}


async def _auto_import_data_sheets(
    project_id: UUID,
    content: bytes,
    year: int | None,
    db: AsyncSession,
) -> tuple[dict[str, int], list[dict]]:
    """解析 Excel 所有 sheet，自动识别余额表/序时账/辅助账并导入。

    Returns: (imported_counts, diagnostics)
      - imported_counts: {data_type: record_count}
      - diagnostics: [{sheet_name, guessed_type, matched_cols, missing_cols, row_count}]
    """
    from app.services.import_engine.parsers import GenericParser

    # 硬性必需列（缺少会导致解析器丢弃整行，无法入库）
    _REQUIRED_COLS: dict[str, list[str]] = {
        "balance": ["account_code"],
        "ledger": ["account_code", "voucher_date", "voucher_no"],
        "aux_balance": ["account_code", "aux_type"],
        "aux_ledger": ["account_code"],
    }
    # 建议列（缺少不阻止导入，但数据不完整，查账功能受限）
    _RECOMMENDED_COLS: dict[str, list[str]] = {
        "balance": ["opening_balance", "debit_amount", "credit_amount", "closing_balance"],
        "ledger": ["debit_amount", "credit_amount", "summary"],
        "aux_balance": ["opening_balance", "closing_balance", "aux_code", "aux_name"],
        "aux_ledger": ["aux_type", "voucher_date", "debit_amount", "credit_amount"],
    }

    # 如果没有指定年度，从项目的 wizard_state 中获取
    if year is None:
        from sqlalchemy import select as sa_select
        from app.models.core import Project
        proj_result = await db.execute(
            sa_select(Project.wizard_state).where(Project.id == project_id)
        )
        ws = proj_result.scalar_one_or_none() or {}
        basic = ws.get("steps", {}).get("basic_info", {}).get("data", {})
        year = basic.get("audit_year") or basic.get("year") or (
            __import__("datetime").datetime.now().year - 1
        )

    # 1. 解析所有 sheet，识别每个 sheet 的数据类型，生成诊断信息
    all_sheets = _parse_excel_multi_sheet(content)
    print(f"[AUTO_IMPORT] 解析到 {len(all_sheets)} 个 sheet: {list(all_sheets.keys())}")
    types_to_import: set[str] = set()
    diagnostics: list[dict] = []

    for sheet_name, rows in all_sheets.items():
        if not rows:
            diagnostics.append({
                "sheet_name": sheet_name, "guessed_type": "empty",
                "matched_cols": [], "missing_cols": [], "row_count": 0,
            })
            continue

        first_row = rows[0]
        normalized_cols = set()
        for col_name in first_row.keys():
            clean = _clean_header(col_name)
            mapped = _COLUMN_MAP.get(clean)
            if mapped:
                normalized_cols.add(mapped)

        guessed = _guess_file_type(normalized_cols)
        data_type = _SHEET_TYPE_TO_DATA_TYPE.get(guessed)

        # 计算缺失的必需列和建议列
        required = _REQUIRED_COLS.get(guessed, [])
        missing = [c for c in required if c not in normalized_cols]
        recommended = _RECOMMENDED_COLS.get(guessed, [])
        missing_recommended = [c for c in recommended if c not in normalized_cols]

        diag = {
            "sheet_name": sheet_name,
            "guessed_type": guessed,
            "matched_cols": sorted(normalized_cols),
            "missing_cols": missing,
            "missing_recommended": missing_recommended,
            "row_count": len(rows),
        }
        diagnostics.append(diag)

        logger.info("  sheet '%s': %d rows, guessed=%s, data_type=%s, missing=%s",
                     sheet_name, len(rows), guessed, data_type, missing)
        if data_type and not missing:
            types_to_import.add(data_type)
        elif data_type and missing:
            logger.warning("  sheet '%s' 缺少必需列 %s，跳过导入", sheet_name, missing)

    print(f"[AUTO_IMPORT] 识别到的数据类型: {types_to_import}")
    if not types_to_import:
        return {}, diagnostics

    # 2. 对每种识别出的 data_type，用 GenericParser 解析并导入
    from app.services.import_service import (
        _build_record_dict,
        _TABLE_MAP,
        _soft_delete_existing,
        _backfill_account_names,
        CHUNK_SIZE,
    )
    from app.models.audit_platform_models import ImportBatch, ImportStatus
    from datetime import datetime

    parser = GenericParser()
    result: dict[str, int] = {}

    for data_type in types_to_import:
        try:
            print(f"[AUTO_IMPORT] 解析 {data_type}...")
            parsed_data = parser.parse(content, data_type)
            print(f"[AUTO_IMPORT] {data_type}: parsed {len(parsed_data) if parsed_data else 0} rows")
            if not parsed_data:
                continue

            batch = ImportBatch(
                project_id=project_id,
                year=year,
                source_type="generic",
                file_name=f"auto_{data_type}",
                data_type=data_type,
                status=ImportStatus.processing,
                started_at=datetime.utcnow(),
            )
            db.add(batch)
            await db.flush()

            # 覆盖旧数据
            await _soft_delete_existing(project_id, year, data_type, db)

            # 批量写入
            table_model = _TABLE_MAP[data_type]
            record_count = 0
            for i in range(0, len(parsed_data), CHUNK_SIZE):
                chunk = parsed_data[i:i + CHUNK_SIZE]
                recs = [
                    r for row in chunk
                    if (r := _build_record_dict(data_type, row, project_id, year, batch.id))
                ]
                if recs:
                    await db.execute(table_model.__table__.insert(), recs)
                    record_count += len(recs)

            batch.record_count = record_count
            batch.status = ImportStatus.completed
            batch.completed_at = datetime.utcnow()
            await db.commit()

            await _backfill_account_names(project_id, batch.id, data_type, db)

            if record_count > 0:
                result[data_type] = record_count
                logger.info("自动导入 %s: project=%s, records=%d", data_type, project_id, record_count)

        except Exception as e:
            print(f"[AUTO_IMPORT] 自动导入 {data_type} 失败: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            try:
                await db.rollback()
            except Exception:
                pass
            continue

    return result, diagnostics


# ---------------------------------------------------------------------------
# get_standard_chart / get_client_chart
# ---------------------------------------------------------------------------


async def get_standard_chart(
    project_id: UUID,
    db: AsyncSession,
) -> list[AccountChartResponse]:
    """Get standard account chart for a project."""
    result = await db.execute(
        select(AccountChart).where(
            AccountChart.project_id == project_id,
            AccountChart.source == AccountSource.standard,
            AccountChart.is_deleted == False,  # noqa: E712
        ).order_by(AccountChart.account_code)
    )
    accounts = result.scalars().all()
    return [
        AccountChartResponse.model_validate(a)
        for a in accounts
    ]


async def get_client_chart_tree(
    project_id: UUID,
    db: AsyncSession,
) -> dict[str, list[AccountTreeNode]]:
    """Get client account chart as tree structure grouped by category.

    Validates: Requirements 2.6
    """
    result = await db.execute(
        select(AccountChart).where(
            AccountChart.project_id == project_id,
            AccountChart.source == AccountSource.client,
            AccountChart.is_deleted == False,  # noqa: E712
        ).order_by(AccountChart.account_code)
    )
    accounts = result.scalars().all()

    # Group by category
    by_category: dict[str, list[AccountChart]] = {}
    for a in accounts:
        cat = a.category.value
        by_category.setdefault(cat, []).append(a)

    # Build tree per category
    tree: dict[str, list[AccountTreeNode]] = {}
    for cat, accts in by_category.items():
        tree[cat] = _build_tree(accts)

    return tree


def _build_tree(accounts: list[AccountChart]) -> list[AccountTreeNode]:
    """Build tree from flat account list using parent_code."""
    nodes: dict[str, AccountTreeNode] = {}
    roots: list[AccountTreeNode] = []

    # Create all nodes
    for a in accounts:
        node = AccountTreeNode(
            account_code=a.account_code,
            account_name=a.account_name,
            direction=a.direction,
            level=a.level,
            category=a.category,
            parent_code=a.parent_code,
            children=[],
        )
        nodes[a.account_code] = node

    # Link children to parents
    for a in accounts:
        node = nodes[a.account_code]
        if a.parent_code and a.parent_code in nodes:
            nodes[a.parent_code].children.append(node)
        else:
            roots.append(node)

    return roots


# ---------------------------------------------------------------------------
# File parsing helpers
# ---------------------------------------------------------------------------

# Column name mapping: Chinese → English (extended for all file types)
_COLUMN_MAP = {
    # 科目信息
    "科目编码": "account_code",
    "科目代码": "account_code",
    "科目编号": "account_code",
    "编码": "account_code",
    "编号": "account_code",
    "会计科目编码": "account_code",
    "会计科目代码": "account_code",
    "科目": "account_code",
    "科目名称": "account_name",
    "名称": "account_name",
    "会计科目名称": "account_name",
    "科目全称": "account_name",
    "借贷方向": "direction",
    "方向": "direction",
    "余额方向": "direction",
    "借贷": "direction",
    "父科目编码": "parent_code",
    "上级科目编码": "parent_code",
    "上级编码": "parent_code",
    "上级科目": "parent_code",
    "parent_code": "parent_code",
    "account_code": "account_code",
    "account_name": "account_name",
    "direction": "direction",
    # 凭证信息
    "凭证日期": "voucher_date",
    "日期": "voucher_date",
    "记账日期": "voucher_date",
    "制单日期": "voucher_date",
    "业务日期": "voucher_date",
    "voucher_date": "voucher_date",
    "凭证号": "voucher_no",
    "凭证编号": "voucher_no",
    "凭证字号": "voucher_no",
    "凭证号数": "voucher_no",
    "记账凭证号": "voucher_no",
    "voucher_no": "voucher_no",
    "摘要": "summary",
    "备注": "summary",
    "说明": "summary",
    "summary": "summary",
    "对方科目": "counterpart_account",
    "对方科目编码": "counterpart_account",
    "对方科目名称": "counterpart_account",
    "counterpart_account": "counterpart_account",
    "制单人": "preparer",
    "编制人": "preparer",
    "录入人": "preparer",
    "操作员": "preparer",
    "preparer": "preparer",
    # 金额信息
    "借方金额": "debit_amount",
    "借方发生额": "debit_amount",
    "借方": "debit_amount",
    "本期借方": "debit_amount",
    "本期借方发生额": "debit_amount",
    "debit": "debit_amount",
    "debit_amount": "debit_amount",
    "贷方金额": "credit_amount",
    "贷方发生额": "credit_amount",
    "贷方": "credit_amount",
    "本期贷方": "credit_amount",
    "本期贷方发生额": "credit_amount",
    "credit": "credit_amount",
    "credit_amount": "credit_amount",
    "期初余额": "opening_balance",
    "年初余额": "opening_balance",
    "期初金额": "opening_balance",
    "年初金额": "opening_balance",
    "期初数": "opening_balance",
    "年初数": "opening_balance",
    "opening": "opening_balance",
    "opening_balance": "opening_balance",
    "期末余额": "closing_balance",
    "期末金额": "closing_balance",
    "期末数": "closing_balance",
    "closing": "closing_balance",
    "closing_balance": "closing_balance",
    # 辅助核算
    "辅助类型": "aux_type",
    "核算类型": "aux_type",
    "核算维度": "aux_type",
    "aux_type": "aux_type",
    "辅助编码": "aux_code",
    "核算编码": "aux_code",
    "核算项目编码": "aux_code",
    "aux_code": "aux_code",
    "辅助名称": "aux_name",
    "核算名称": "aux_name",
    "核算项目名称": "aux_name",
    "aux_name": "aux_name",
    # 新增字段（基于实际序时账 Excel 表头）
    "会计月份": "accounting_period",
    "会计期间": "accounting_period",
    "月份": "accounting_period",
    "accounting_period": "accounting_period",
    "凭证类型": "voucher_type",
    "凭证字": "voucher_type",
    "voucher_type": "voucher_type",
    "分录序号": "entry_seq",
    "序号": "entry_seq",
    "entry_seq": "entry_seq",
    "借方数量": "debit_qty",
    "debit_qty": "debit_qty",
    "贷方数量": "credit_qty",
    "credit_qty": "credit_qty",
    "借方外币发生额": "debit_fc",
    "借方外币": "debit_fc",
    "debit_fc": "debit_fc",
    "贷方外币发生额": "credit_fc",
    "贷方外币": "credit_fc",
    "credit_fc": "credit_fc",
    "核算项目类型编号": "aux_type",
    "核算项目类型名称": "aux_type_name",
    "辅助类型名称": "aux_type_name",
    "aux_type_name": "aux_type_name",
    "核算项目编号": "aux_code",
    "科目级次": "level",
    "级次": "level",
    "level": "level",
    "科目类别": "category",
    "类别": "category",
    "category": "category",
    "期初数量": "opening_qty",
    "opening_qty": "opening_qty",
    "期初外币": "opening_fc",
    "opening_fc": "opening_fc",
    # 新增：丰桔出行等企业导出格式
    "科目方向": "direction",
    "科目类型": "category",
    "末级科目": "is_leaf",
    "凭证月份": "accounting_period",
    "分录行号": "entry_seq",
    "分录号": "entry_seq",
    "币种名称": "currency_code",
    "币种": "currency_code",
    "填制人": "preparer",
    "审核人": "reviewer",
    "记账人": "bookkeeper",
    "期初方向": "opening_direction",
    "期末方向": "closing_direction",
    "借方累计": "debit_amount",
    "贷方累计": "credit_amount",
    "辅助账核算项目": "aux_info",
    "凭证张数": "voucher_count",
    "单价": "unit_price",
    "汇率": "exchange_rate",
    "计量单位": "unit",
    "项目": "aux_name",
}


# ---------------------------------------------------------------------------
# preview_file — file preview + auto column mapping
# ---------------------------------------------------------------------------


import re as _re

def _clean_header(h: str) -> str:
    """清洗列名：去掉方括号、圆括号、空格、星号等特殊字符。"""
    s = str(h).strip()
    s = _re.sub(r'[\[\]【】\(\)（）\*\#\s]', '', s)
    return s


def _match_column(header: str) -> str | None:
    """模糊匹配列名到标准字段。先精确匹配原始列名，再清洗后匹配。"""
    h = str(header).strip()
    # 1. 精确匹配
    if h in _COLUMN_MAP:
        return _COLUMN_MAP[h]
    # 2. 清洗后匹配
    cleaned = _clean_header(h)
    if cleaned in _COLUMN_MAP:
        return _COLUMN_MAP[cleaned]
    # 3. 去掉常见前缀后匹配（如"GT_"）
    for prefix in ("GT_", "gt_"):
        if cleaned.startswith(prefix):
            without_prefix = cleaned[len(prefix):]
            if without_prefix in _COLUMN_MAP:
                return _COLUMN_MAP[without_prefix]
    return None


def _guess_file_type(mapped_cols: set[str]) -> str:
    """Guess file type based on which standard columns are present."""
    has_account_code = "account_code" in mapped_cols
    has_account_name = "account_name" in mapped_cols
    has_voucher_date = "voucher_date" in mapped_cols
    has_voucher_no = "voucher_no" in mapped_cols
    has_debit = "debit_amount" in mapped_cols
    has_credit = "credit_amount" in mapped_cols
    has_opening = "opening_balance" in mapped_cols
    has_closing = "closing_balance" in mapped_cols
    has_aux_type = "aux_type" in mapped_cols or "aux_type_name" in mapped_cols
    has_aux_code = "aux_code" in mapped_cols
    has_aux_name = "aux_name" in mapped_cols

    # 辅助核算类（优先判断，因为也可能有凭证日期）
    if has_aux_type or has_aux_code or has_aux_name:
        if has_voucher_date and has_voucher_no:
            return "aux_ledger"  # 辅助明细账（有凭证信息）
        return "aux_balance"  # 辅助余额表（无凭证信息）

    # 序时账/凭证表
    if has_voucher_date and has_voucher_no and (has_debit or has_credit):
        return "ledger"

    # 余额表
    if has_account_code and (has_opening or has_closing):
        return "balance"

    # 科目表（只有编码+名称，无金额）
    if has_account_code and has_account_name and not has_debit and not has_credit:
        return "account_chart"

    return "unknown"


async def preview_file(file: UploadFile, skip_rows: int | None = None) -> dict:
    """Parse file, return first 20 rows per sheet + auto-matched column mapping.

    Args:
        skip_rows: 跳过前N行。None=自动检测表头行位置（推荐）

    Returns:
        {
            "sheets": [
                {
                    "sheet_name": "Sheet1",
                    "headers": ["col1", "col2", ...],
                    "rows": [...],  # first 20 rows (after skipping)
                    "total_rows": 150,
                    "column_mapping": {"col1": "account_code", ...},
                    "file_type_guess": "ledger"
                },
                ...
            ],
            "active_sheet": 0
        }
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")

    filename_lower = file.filename.lower()
    content = await file.read()

    if filename_lower.endswith(".csv"):
        rows = _parse_csv(content, skip_rows=skip_rows)
        if not rows:
            raise HTTPException(status_code=400, detail="文件中未解析到有效数据")
        headers = list(rows[0].keys())
        column_mapping = {h: _match_column(h) for h in headers}
        mapped_cols = {v for v in column_mapping.values() if v}
        return {
            "sheets": [{
                "sheet_name": "CSV",
                "headers": headers,
                "rows": rows[:20],
                "total_rows": len(rows),
                "column_mapping": column_mapping,
                "file_type_guess": _guess_file_type(mapped_cols),
            }],
            "active_sheet": 0,
        }
    elif filename_lower.endswith((".xlsx", ".xls")):
        sheets_data = _parse_excel_multi_sheet(content, skip_rows=skip_rows)
        if not sheets_data:
            raise HTTPException(status_code=400, detail="文件中未解析到有效数据")
        sheets = []
        for sheet_name, rows in sheets_data.items():
            if not rows:
                continue
            headers = list(rows[0].keys())
            column_mapping = {h: _match_column(h) for h in headers}
            mapped_cols = {v for v in column_mapping.values() if v}
            sheets.append({
                "sheet_name": sheet_name,
                "headers": headers,
                "rows": rows[:20],
                "total_rows": len(rows),
                "column_mapping": column_mapping,
                "file_type_guess": _guess_file_type(mapped_cols),
            })
        if not sheets:
            raise HTTPException(status_code=400, detail="所有工作表均无有效数据")
        return {"sheets": sheets, "active_sheet": 0}
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式")


def _normalize_columns(rows: list[dict]) -> list[dict]:
    """Normalize column names from Chinese to English."""
    normalized = []
    for row in rows:
        new_row = {}
        for key, value in row.items():
            if key is None:
                continue
            key_str = str(key).strip()
            mapped = _match_column(key_str) or key_str
            new_row[mapped] = value
        normalized.append(new_row)
    return normalized


def _apply_custom_mapping(rows: list[dict], mapping: dict[str, str | None]) -> list[dict]:
    """Apply user-confirmed column mapping to rows."""
    normalized = []
    for row in rows:
        new_row = {}
        for key, value in row.items():
            if key is None:
                continue
            key_str = str(key).strip()
            target = mapping.get(key_str)
            if target:  # skip None / "(忽略)"
                new_row[target] = value
        normalized.append(new_row)
    return normalized


def _parse_csv(content: bytes, skip_rows: int = 0) -> list[dict]:
    """Parse CSV file content into list of dicts."""
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        raise HTTPException(status_code=400, detail="无法识别文件编码，请使用 UTF-8 或 GBK 编码")

    lines = text.strip().splitlines()
    if skip_rows > 0 and len(lines) > skip_rows:
        lines = lines[skip_rows:]
    reader = csv.DictReader(io.StringIO("\n".join(lines)))
    return list(reader)


def _parse_excel(content: bytes, skip_rows: int | None = None) -> list[dict]:
    """Parse Excel active sheet into list of dicts.

    skip_rows=None → 自动检测表头行位置。
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl 库")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel 文件中没有工作表")

    result = _parse_sheet(ws, skip_rows=skip_rows)
    wb.close()
    return result


def _parse_excel_multi_sheet(content: bytes, skip_rows: int | None = None) -> dict[str, list[dict]]:
    """Parse ALL sheets in an Excel file, each returning list of dicts.

    skip_rows=None → 每个 sheet 自动检测表头行位置。
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl 库")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        rows = _parse_sheet(ws, skip_rows=skip_rows)
        sheets[ws_name] = rows
    wb.close()
    return sheets


def _detect_header_row(ws, max_scan: int = 5) -> int:
    """自动检测表头行位置（0-indexed skip_rows）。

    策略：扫描前 max_scan 行，找到第一行"看起来像表头"的行。
    表头特征：多个非空单元格，且内容是短文本（非长段落说明）。
    说明行特征：第一个单元格很长（>30字符），或以数字序号开头（如"1."、"2."）。
    """
    rows_iter = ws.iter_rows(values_only=True)
    for i in range(max_scan):
        try:
            row = next(rows_iter)
        except StopIteration:
            return 0

        cells = [str(c).strip() if c else "" for c in row]
        non_empty = [c for c in cells if c]

        if not non_empty:
            continue

        first_cell = non_empty[0]

        # 说明行特征：以数字序号开头（"1."、"2."），或第一个单元格超长
        is_instruction = (
            len(first_cell) > 30
            or (len(first_cell) > 2 and first_cell[0].isdigit() and first_cell[1] == '.')
        )

        if is_instruction:
            continue

        # 表头特征：至少3个非空单元格，且都是短文本
        if len(non_empty) >= 3 and all(len(c) <= 30 for c in non_empty):
            return i

        # 只有1-2个非空单元格但内容短，也可能是表头（如只有2列的简单表）
        if len(non_empty) >= 1 and all(len(c) <= 20 for c in non_empty):
            return i

    return 0


def _parse_sheet(ws, skip_rows: int | None = None) -> list[dict]:
    """Parse a single worksheet into list of dicts.

    skip_rows=None 时自动检测表头行位置。
    """
    if skip_rows is None:
        skip_rows = _detect_header_row(ws)

    rows_iter = ws.iter_rows(values_only=True)

    # Skip rows
    for _ in range(skip_rows):
        try:
            next(rows_iter)
        except StopIteration:
            return []

    # Header row (the row after skipped rows)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header)]

    result = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(cell) if cell is not None else ""
        result.append(row_dict)
    return result
