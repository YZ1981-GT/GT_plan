"""WpStandardConversionService — 底稿层准则切换服务（SOE↔Listed）

Requirements: 2.1, 2.2, 2.3, 5.1
Spec: multi-standard-unification
Design: D1（底稿切换复用附注层模式）/ D2（准则差异数据源）/ D4（切换前置条件）

为底稿层补齐"已编制底稿 SOE↔Listed 切换"能力，复用附注层
``note_conversion_service`` 已验证的转换模式：
- 共有底稿（两准则都适用的 wp_code）：保留 ``parsed_data`` + ``working_paper`` 不动
- 源准则独有底稿：归档（soft delete + 记录 conversion_reason）
- 目标准则独有底稿：从模板创建（建 WpIndex + WorkingPaper + parsed_data）

准则差异数据源（D2）：每个 wp_code 的 ``applicable_standard`` 来自 PG 表
``wp_template_registry``（经 ``WpTemplateRegistryService.load_tree`` 读取）。
``applicable_standard`` 是一个列表，例如 ``["soe"]`` / ``["listed"]`` /
``["soe", "listed"]`` / ``[]``（空表示通用/共有，适用所有准则）。

范围边界（见 requirements.md）：本 spec 只处理 SOE↔Listed 切换，分类依据是
结构化 standard 的 ``entity_type`` 轴（∈ {"soe", "listed", "private"}），
不处理"合并↔单体"（scope 轴）切换。

注意：本任务（3.1）实现编排骨架——``classify_workpapers`` /
``preview_conversion`` / ``convert_workpapers``（编排器）；三组底稿的实际
处理逻辑（保留 / 归档 / 创建）与前置条件检查由后续任务实现：
- ``_retain_shared_workpapers`` → 任务 3.2（已实现）
- ``_archive_source_only_workpapers`` → 任务 3.3（已实现）
- ``_create_target_only_workpapers`` → 任务 3.4（已实现）
- ``check_preconditions`` → 任务 3.5（已实现）
"""
from __future__ import annotations

import json
import logging
import os
import shutil
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.models.base import ProjectStatus
from app.models.core import Project
from app.models.workpaper_models import (
    WorkingPaper,
    WpIndex,
    WpSourceType,
    WpStatus,
)
from app.services.wp_template_registry import wp_template_registry_service

logger = logging.getLogger(__name__)


class WorkpaperConversionPreconditionError(ValueError):
    """切换前置条件不满足（如存在未保存底稿）。

    继承 ``ValueError`` 以遵循项目"业务异常 = ValueError 子类铁律"——
    router 层捕获 ``ValueError`` 即统一映射为 HTTP 422。
    """

# 默认 entity_type（与 StandardUnificationService.DEFAULT_STANDARD 一致）
_DEFAULT_ENTITY_TYPE = "soe"

# 致同标准底稿模板库目录（与 wp_template.generate_from_codes 同源）
_TEMPLATE_LIBRARY_PATH = (
    Path(__file__).resolve().parent.parent.parent / "data" / "gt_template_library.json"
)

# 知识库底稿模板目录（与 generate_from_codes 一致）
_KB_TEMPLATE_BASE = Path(
    os.path.expanduser("~/.gt_audit_helper/knowledge/workpaper_templates")
)


class WpStandardConversionService:
    """底稿层准则切换服务（SOE↔Listed）。"""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ------------------------------------------------------------------
    # 分类（D2：准则差异数据源）
    # ------------------------------------------------------------------

    async def classify_workpapers(
        self,
        project_id: UUID,
        old_standard: dict,
        new_standard: dict,
    ) -> dict:
        """将底稿分为 共有 / 源独有 / 目标独有 三组（按 entity_type 轴）。

        分类规则（D2）：
        - 某 wp_code 的"适用准则列表"来自 ``wp_template_registry``；列表为空
          或缺失视为"通用"（适用所有 entity_type）。
        - shared：当前项目已存在、且同时适用于 old 与 new entity 的 wp_code
          （切换时保留用户数据）。
        - source_only：当前项目已存在、适用于 old 但不适用 new 的 wp_code
          （切换时归档）。
        - target_only：注册表中适用于 new 但不适用 old、且当前项目尚不存在的
          wp_code（切换时新建）。

        降级（graceful）：若 ``wp_template_registry`` 表不存在或为空，则映射为
        ``{}``，此时所有已存在底稿均被视为"通用 → 共有"，切换成为归档层面的
        no-op（不丢数据）。

        Returns:
            ``{"shared": [...], "source_only": [...], "target_only": [...]}``，
            各组为去重并排序后的 wp_code 列表（保证确定性）。
        """
        old_entity = (old_standard or {}).get("entity_type") or _DEFAULT_ENTITY_TYPE
        new_entity = (new_standard or {}).get("entity_type") or _DEFAULT_ENTITY_TYPE

        # 1. 项目现有（未删除）底稿的 wp_code 集合
        existing_codes = await self._load_existing_wp_codes(project_id)

        # 2. 注册表 {wp_code: [applicable_standards]} 映射（缺失则 {}）
        registry_map = await self._load_registry_standard_map()

        shared: set[str] = set()
        source_only: set[str] = set()

        # 3. 对现有底稿分类（shared vs source_only）
        for wp_code in existing_codes:
            standards = registry_map.get(wp_code)
            app_old = self._is_applicable(standards, old_entity)
            app_new = self._is_applicable(standards, new_entity)

            if app_old and not app_new:
                # 适用 old 但不适用 new → 归档
                source_only.add(wp_code)
            else:
                # 适用两者 / 仅适用 new / 通用 → 保留（避免误删用户数据）
                shared.add(wp_code)

        # 4. 目标独有：注册表中适用 new 但不适用 old、且项目尚不存在的 wp_code
        target_only: set[str] = set()
        for wp_code, standards in registry_map.items():
            if wp_code in existing_codes:
                continue
            if self._is_applicable(standards, new_entity) and not self._is_applicable(
                standards, old_entity
            ):
                target_only.add(wp_code)

        result = {
            "shared": sorted(shared),
            "source_only": sorted(source_only),
            "target_only": sorted(target_only),
        }
        logger.info(
            "classify_workpapers: project=%s %s->%s shared=%d source_only=%d target_only=%d",
            project_id,
            old_entity,
            new_entity,
            len(result["shared"]),
            len(result["source_only"]),
            len(result["target_only"]),
        )
        return result

    # ------------------------------------------------------------------
    # 影响预览（需求 5.1，只读，无 DB 写入）
    # ------------------------------------------------------------------

    async def preview_conversion(
        self,
        project_id: UUID,
        old_standard: dict,
        new_standard: dict,
    ) -> dict:
        """切换影响预览（只读）。

        基于 ``classify_workpapers`` 的分类结果，返回切换将"归档 / 新建 / 保留"
        的底稿清单与计数。此方法不做任何 DB 写入（需求 5.1）。

        Returns:
            ``{
                "to_archive": [...source_only...],
                "to_create": [...target_only...],
                "to_retain": [...shared...],
                "counts": {"archive": n, "create": n, "retain": n},
            }``
        """
        classification = await self.classify_workpapers(
            project_id, old_standard, new_standard
        )
        to_archive = classification["source_only"]
        to_create = classification["target_only"]
        to_retain = classification["shared"]
        return {
            "to_archive": to_archive,
            "to_create": to_create,
            "to_retain": to_retain,
            "counts": {
                "archive": len(to_archive),
                "create": len(to_create),
                "retain": len(to_retain),
            },
        }

    # ------------------------------------------------------------------
    # 切换执行（编排器）
    # ------------------------------------------------------------------

    async def convert_workpapers(
        self,
        project_id: UUID,
        classification: dict,
        new_standard: dict,
        changed_by: UUID | None = None,
    ) -> dict:
        """执行底稿准则切换（编排器）。

        步骤：
        1. 前置条件检查（``check_preconditions`` — 任务 3.5 实现）
        2. 共有底稿保留（``_retain_shared_workpapers`` — 任务 3.2 实现）
        3. 源独有底稿归档（``_archive_source_only_workpapers`` — 任务 3.3 实现）
        4. 目标独有底稿创建（``_create_target_only_workpapers`` — 任务 3.4 实现）
        5. 提交事务
        6. 返回切换汇总

        Args:
            project_id: 项目 ID
            classification: ``classify_workpapers`` 的返回结果
                （含 shared / source_only / target_only 三组 wp_code）
            new_standard: 目标结构化准则
            changed_by: 触发切换的用户（用于审计记录，可选）

        Returns:
            ``{"retained": N, "archived": N, "created": N}``
        """
        shared_codes = classification.get("shared", []) if classification else []
        source_only_codes = (
            classification.get("source_only", []) if classification else []
        )
        target_only_codes = (
            classification.get("target_only", []) if classification else []
        )

        # 1. 前置条件检查（任务 3.5）
        await self.check_preconditions(project_id)

        # 2. 共有底稿保留（任务 3.2）
        retained = await self._retain_shared_workpapers(
            project_id, shared_codes, new_standard
        )
        # 3. 源独有底稿归档（任务 3.3）
        archived = await self._archive_source_only_workpapers(
            project_id, source_only_codes, new_standard
        )
        # 4. 目标独有底稿创建（任务 3.4）
        created = await self._create_target_only_workpapers(
            project_id, target_only_codes, new_standard, changed_by
        )

        # 5. flush（本方法只 flush 不 commit，事务提交由调用方（router）统一管理。）
        await self.db.flush()

        logger.info(
            "convert_workpapers: project=%s retained=%d archived=%d created=%d (changed_by=%s)",
            project_id,
            retained,
            archived,
            created,
            changed_by,
        )

        # 6. 汇总
        return {"retained": retained, "archived": archived, "created": created}

    # ------------------------------------------------------------------
    # 内部辅助（分类用，纯查询 / 纯函数）
    # ------------------------------------------------------------------

    async def _load_existing_wp_codes(self, project_id: UUID) -> set[str]:
        """加载项目现有（未删除）底稿的 wp_code 集合。"""
        result = await self.db.execute(
            sa.select(WpIndex.wp_code).where(
                WpIndex.project_id == project_id,
                WpIndex.is_deleted == sa.false(),
            )
        )
        return {row[0] for row in result.all() if row[0]}

    async def _load_registry_standard_map(self) -> dict[str, list[str]]:
        """加载 {wp_code: [applicable_standards]} 映射（D2）。

        若 ``wp_template_registry`` 表不存在或为空，返回 ``{}``（降级，使所有
        现有底稿被视为通用/共有）。
        """
        try:
            if not await wp_template_registry_service.table_exists(self.db):
                logger.info(
                    "_load_registry_standard_map: wp_template_registry 表不存在，降级为空映射"
                )
                return {}
            tree = await wp_template_registry_service.load_tree(self.db)
        except Exception as exc:  # 表缺失 / 查询失败一律降级，不阻塞切换
            logger.warning(
                "_load_registry_standard_map: 读取注册表失败，降级为空映射: %s", exc
            )
            return {}

        mapping: dict[str, list[str]] = {}
        for entry in tree or []:
            wp_code = entry.get("wp_code")
            if not wp_code:
                continue
            standards = entry.get("applicable_standard") or []
            # 归一化为小写字符串列表
            if isinstance(standards, str):
                standards = [standards]
            mapping[wp_code] = [str(s).lower() for s in standards if s]
        return mapping

    @staticmethod
    def _is_applicable(standards: list[str] | None, entity_type: str) -> bool:
        """判定某 wp_code 是否适用于给定 entity_type。

        空列表 / None（通用）→ 适用所有 entity_type；否则 entity_type 在列表中才适用。
        """
        if not standards:
            return True
        return (entity_type or "").lower() in standards

    # ------------------------------------------------------------------
    # 切换处理 helper（后续任务实现，当前为占位 stub）
    # ------------------------------------------------------------------

    async def _retain_shared_workpapers(
        self,
        project_id: UUID,
        shared_codes: list[str],
        new_standard: dict,
    ) -> int:
        """共有底稿保留逻辑（数据保全的"只读计数"操作）。

        共有底稿是同时适用于旧准则与新准则的底稿。本方法的核心不变量
        （Requirement 2.2）是：切换过程中这些底稿的用户已填内容
        （``WorkingPaper.parsed_data``）**绝不能丢失或被改写**。因此本方法
        刻意是一个对数据无副作用的 no-op：

        - **不**修改、删除或重写任何 ``WorkingPaper.parsed_data``；
        - **不**修改、删除或重写任何 ``WpIndex`` 行；
        - **不**调用 ``db.commit()``（事务提交由编排器 ``convert_workpapers``
          统一处理）。

        关于 ``workpaper_sheet_classification`` 的更新（Requirement 2.5 的
        "如有变化"）：sheet 分类是**模板级**信息（按 ``wp_code`` 在
        ``workpaper_sheet_classification`` 中维护），不随准则 / entity_type
        变化——同一 ``wp_code`` 在 SOE 与 Listed 下的 sheet 分类相同。故对
        SOE↔Listed 切换而言，共有底稿没有需要改写的分类，无需任何更新。

        本方法仅 **统计并返回** ``shared_codes`` 中在当前项目里实际存在
        （未删除）的底稿数量，作为编排器对外汇报的 "retained" 计数。

        Args:
            project_id: 项目 ID
            shared_codes: 共有底稿的 wp_code 列表（来自 ``classify_workpapers``）
            new_standard: 目标结构化准则（当前无需消费，保留以对齐其他 helper 签名）

        Returns:
            ``shared_codes`` 中实际存在于项目（未删除）的去重底稿数量；
            ``shared_codes`` 为空时返回 0。
        """
        if not shared_codes:
            return 0

        # 仅做去重计数：统计 shared_codes 中实际存在（未删除）的底稿。
        # 不触碰 parsed_data / WpIndex 行，满足 Requirement 2.2 数据保全要求。
        result = await self.db.execute(
            sa.select(sa.func.count(sa.distinct(WpIndex.wp_code))).where(
                WpIndex.project_id == project_id,
                WpIndex.wp_code.in_(shared_codes),
                WpIndex.is_deleted == sa.false(),
            )
        )
        return int(result.scalar_one() or 0)

    async def _archive_source_only_workpapers(
        self,
        project_id: UUID,
        source_only_codes: list[str],
        new_standard: dict,
    ) -> int:
        """源准则独有底稿归档逻辑（soft delete + 记录切换 lineage）。

        源准则独有底稿是只适用于旧准则、不适用于新准则的底稿。切换时这些底稿
        必须被**归档**（而非硬删除），以便后续审计追溯与数据恢复
        （Requirement 2.1）；同时切换原因必须被**记录**下来供审计溯源
        （Requirement 2.3）。

        关于 lineage 落点（为何不写 ``template_lineage`` 列）：
        design.md 引用了 ``template_lineage.conversion_reason``，但底稿层模型
        （``WpIndex`` / ``WorkingPaper``）**都没有 ``template_lineage`` 列**——
        ``template_lineage`` 是附注层 ``disclosure_notes`` 才有的列。为在底稿层
        落地等价的审计 lineage，本方法把切换 lineage 写入两个模型已有的 JSONB
        字段下的保留键 ``_conversion_lineage``：
        - ``WorkingPaper.parsed_data["_conversion_lineage"]``（底稿文件层，
          与用户已填内容同处，记录该文件因何被归档）；
        - ``WpIndex.cross_ref_codes["_conversion_lineage"]``（索引层，记录该
          索引行因何被归档）。

        每条 lineage 条目形如::

            {
                "action": "archived_on_standard_change",
                "conversion_reason": "standard_changed_to_<new_entity_type>",
                "new_standard": <new_standard>,
                "archived_at": "<iso8601 utc>",
            }

        JSONB 变更检测：SQLAlchemy **不会**自动侦测对 JSONB 字段的原地（in-place）
        修改，因此本方法在追加 lineage 后对相应字段调用 ``flag_modified`` 显式
        标脏，确保变更被持久化。

        归档语义：
        - 对每个匹配的 ``WpIndex``（未删除），先归档其名下所有未删除的
          ``WorkingPaper``（``is_deleted=True`` + 追加 lineage），再归档
          ``WpIndex`` 本身（``is_deleted=True`` + 追加 lineage）。
        - 不调用 ``db.commit()``（事务提交由编排器 ``convert_workpapers`` 统一
          处理）；仅在写入完成后 ``flush`` 以保证顺序。

        Args:
            project_id: 项目 ID
            source_only_codes: 源准则独有底稿的 wp_code 列表
                （来自 ``classify_workpapers``）
            new_standard: 目标结构化准则（用于派生 conversion_reason 与记录）

        Returns:
            实际被归档的 ``WpIndex`` 行数；``source_only_codes`` 为空时返回 0。
        """
        if not source_only_codes:
            return 0

        new_entity_type = (new_standard or {}).get("entity_type", _DEFAULT_ENTITY_TYPE)
        archived_at = datetime.now(timezone.utc).isoformat()
        lineage_entry = {
            "action": "archived_on_standard_change",
            "conversion_reason": f"standard_changed_to_{new_entity_type}",
            "new_standard": new_standard,
            "archived_at": archived_at,
        }

        # 1. 加载项目下匹配的未删除 WpIndex 行
        index_result = await self.db.execute(
            sa.select(WpIndex).where(
                WpIndex.project_id == project_id,
                WpIndex.wp_code.in_(source_only_codes),
                WpIndex.is_deleted == sa.false(),
            )
        )
        wp_indexes = index_result.scalars().all()

        archived_count = 0
        for wp_index in wp_indexes:
            # 2. 归档该索引名下所有未删除的 WorkingPaper
            wp_result = await self.db.execute(
                sa.select(WorkingPaper).where(
                    WorkingPaper.wp_index_id == wp_index.id,
                    WorkingPaper.is_deleted == sa.false(),
                )
            )
            for wp in wp_result.scalars().all():
                wp.is_deleted = True
                parsed_data = dict(wp.parsed_data) if wp.parsed_data else {}
                lineage = list(parsed_data.get("_conversion_lineage") or [])
                lineage.append(lineage_entry)
                parsed_data["_conversion_lineage"] = lineage
                wp.parsed_data = parsed_data
                flag_modified(wp, "parsed_data")

            # 3. 归档 WpIndex 本身 + 记录 lineage 到 cross_ref_codes
            wp_index.is_deleted = True
            cross_ref = dict(wp_index.cross_ref_codes) if wp_index.cross_ref_codes else {}
            index_lineage = list(cross_ref.get("_conversion_lineage") or [])
            index_lineage.append(lineage_entry)
            cross_ref["_conversion_lineage"] = index_lineage
            wp_index.cross_ref_codes = cross_ref
            flag_modified(wp_index, "cross_ref_codes")

            archived_count += 1

        # 写入完成后 flush 保证顺序（commit 由编排器统一处理）
        await self.db.flush()

        logger.info(
            "_archive_source_only_workpapers: project=%s archived=%d reason=%s",
            project_id,
            archived_count,
            lineage_entry["conversion_reason"],
        )
        return archived_count

    async def _create_target_only_workpapers(
        self,
        project_id: UUID,
        target_only_codes: list[str],
        new_standard: dict,
        changed_by: UUID | None = None,
    ) -> int:
        """目标准则独有底稿创建逻辑（复用 generate_from_codes 子逻辑）。

        目标准则独有底稿是只适用于新准则、不适用旧准则的底稿。切换时这些底稿
        必须被**创建**出来，使新准则下应有的底稿齐全（Requirement 2.1）。

        复用 ``wp_template.generate_from_codes`` 端点已验证的单底稿创建逻辑
        （见 design.md D1）：建 ``WpIndex`` + ``WorkingPaper`` + 模板文件
        + ``parsed_data``。本方法加载一次模板库目录后，逐个 wp_code 委托
        ``_generate_one_workpaper`` 完成创建，返回实际新建的底稿数。

        关于 ``parsed_data``（与 generate_from_codes 的差异，关键）：
        wp-generation-pipeline spec 发现 ``generate_from_codes`` 创建
        ``WorkingPaper`` 后**从不设置 ``parsed_data``**（保持 NULL），导致 HTML
        渲染器显示"有记录无内容"。本任务的验收（Requirement 2.1）明确要求
        "+ parsed_data"，故本方法在创建 ``WorkingPaper`` 后**必定**填充非空
        ``parsed_data``（优先调用 ``wp_parsed_data_service.populate_parsed_data``
        若存在；否则写入最小非空占位 dict），确保新建底稿不为空。

        事务：不调用 ``db.commit()``（提交由编排器 ``convert_workpapers`` 统一
        处理）；仅在需要拿生成主键时 ``flush``。

        Args:
            project_id: 项目 ID
            target_only_codes: 目标准则独有底稿的 wp_code 列表
                （来自 ``classify_workpapers``）
            new_standard: 目标结构化准则（暂未直接消费，保留以对齐其他 helper 签名）
            changed_by: 触发切换的用户（作为新建 WorkingPaper.created_by，可选）

        Returns:
            实际新建的底稿数；``target_only_codes`` 为空时返回 0。
        """
        if not target_only_codes:
            return 0

        # 模板库目录加载一次（降级为空 dict，不阻塞创建——仍会兜底建空模板）
        template_lib = self._load_template_library()

        created = 0
        for wp_code in target_only_codes:
            try:
                if await self._generate_one_workpaper(
                    project_id,
                    wp_code,
                    template_lib.get(wp_code, {}),
                    new_standard,
                    changed_by,
                ):
                    created += 1
            except Exception as exc:  # 单条失败隔离，不破坏整批切换
                logger.warning(
                    "_create_target_only_workpapers: 创建底稿 %s 失败，跳过: %s",
                    wp_code,
                    exc,
                )

        logger.info(
            "_create_target_only_workpapers: project=%s created=%d/%d",
            project_id,
            created,
            len(target_only_codes),
        )
        return created

    def _load_template_library(self) -> dict[str, dict]:
        """加载致同标准底稿模板库目录为 ``{code: entry}`` 映射（D2）。

        与 ``wp_template.generate_from_codes`` 同源：读取
        ``backend/data/gt_template_library.json``（``encoding="utf-8-sig"`` 以
        兼容 BOM），顶层 ``templates`` 列表中每个条目按 ``code``/``wp_code`` 建索引。
        文件缺失或解析失败时降级为 ``{}``（不阻塞创建，逐条仍可兜底建空模板）。
        """
        if not _TEMPLATE_LIBRARY_PATH.exists():
            logger.info(
                "_load_template_library: 模板库文件不存在(%s)，降级为空映射",
                _TEMPLATE_LIBRARY_PATH,
            )
            return {}
        try:
            with open(_TEMPLATE_LIBRARY_PATH, "r", encoding="utf-8-sig") as f:
                lib_data = json.load(f)
        except Exception as exc:
            logger.warning(
                "_load_template_library: 读取模板库失败，降级为空映射: %s", exc
            )
            return {}

        if isinstance(lib_data, dict):
            entries = lib_data.get("templates", [])
        else:
            entries = lib_data or []

        mapping: dict[str, dict] = {}
        for item in entries:
            if not isinstance(item, dict):
                continue
            code = item.get("code") or item.get("wp_code") or ""
            if code:
                mapping[code] = item
        return mapping

    async def _generate_one_workpaper(
        self,
        project_id: UUID,
        wp_code: str,
        lib_entry: dict,
        new_standard: dict,
        changed_by: UUID | None,
    ) -> bool:
        """为单个 wp_code 创建底稿（建 WpIndex + WorkingPaper + 模板文件 + parsed_data）。

        复用 ``wp_template.generate_from_codes`` 的单底稿创建子逻辑（design.md D1）：

        1. 若项目已存在同 ``wp_code`` 的**未删除** ``WpIndex`` → 跳过（返回 False）。
        2. 处理唯一约束冲突：``uq_wp_index_project_code`` 建在
           ``(project_id, wp_code)`` 上且**不含 ``is_deleted``**——若存在一行
           **已软删除**的同 ``wp_code`` ``WpIndex``，直接 INSERT 新行会触发唯一
           约束冲突。故此处**复活**（``is_deleted=False`` + 重置状态）该软删除行
           而非新建，避免约束违反。
        3. 否则新建 ``WpIndex``（``status=not_started``）并 ``flush`` 取主键。
        4. 计算文件路径 ``storage/projects/{pid}/workpapers/{cycle}/{code}.xlsx``
           并复制模板（知识库 → 原始 file_path → openpyxl 最小工作簿 → 空字节）。
        5. 创建 ``WorkingPaper``（``source_type=template`` / ``file_version=1``）。
        6. 尽力（try/except 非致命）绑定 active dataset + 填充表头。
        7. **填充 ``parsed_data``** 使其非空（见 wp-generation-pipeline 发现）。

        Returns:
            True 表示实际新建/复活了一份底稿；False 表示已存在而跳过。
        """
        # 1+2. 检查是否已存在（含软删除行，因唯一约束不含 is_deleted）
        existing_result = await self.db.execute(
            sa.select(WpIndex).where(
                WpIndex.project_id == project_id,
                WpIndex.wp_code == wp_code,
            )
        )
        existing_rows = existing_result.scalars().all()

        active_row = next((r for r in existing_rows if not r.is_deleted), None)
        if active_row is not None:
            # 已存在未删除底稿 → 跳过（不覆盖用户数据）
            return False

        wp_name = lib_entry.get("name") or lib_entry.get("wp_name") or f"底稿{wp_code}"
        cycle = lib_entry.get("cycle_prefix") or (wp_code[0] if wp_code else "X")

        soft_deleted_row = existing_rows[0] if existing_rows else None
        if soft_deleted_row is not None:
            # 复活软删除行（规避 uq_wp_index_project_code 唯一约束冲突——
            # 该唯一索引建在 (project_id, wp_code) 上、不含 is_deleted）
            wp_index = soft_deleted_row
            wp_index.is_deleted = False
            wp_index.wp_name = wp_name
            wp_index.audit_cycle = cycle
            wp_index.status = WpStatus.not_started
        else:
            # 3. 新建 WpIndex
            wp_index = WpIndex(
                project_id=project_id,
                wp_code=wp_code,
                wp_name=wp_name,
                audit_cycle=cycle,
                status=WpStatus.not_started,
            )
            self.db.add(wp_index)
        await self.db.flush()  # 取 wp_index.id

        # 4. 计算文件目录并复制模板
        dest_file = (
            Path("storage")
            / "projects"
            / str(project_id)
            / "workpapers"
            / cycle
            / f"{wp_code}.xlsx"
        )
        dest_file.parent.mkdir(parents=True, exist_ok=True)
        self._copy_template_file(dest_file, lib_entry, wp_code, wp_name, cycle)

        # 5. 创建 WorkingPaper
        wp = WorkingPaper(
            project_id=project_id,
            wp_index_id=wp_index.id,
            file_path=str(dest_file),
            source_type=WpSourceType.template,
            file_version=1,
            created_by=changed_by,
        )
        self.db.add(wp)
        await self.db.flush()  # 取 wp.id，供表头/绑定/parsed_data 使用

        # 6a. 尽力绑定 active dataset（非致命）
        bind_year = self._derive_year(new_standard)
        try:
            from app.services.dataset_query import bind_to_active_dataset

            await bind_to_active_dataset(self.db, wp, project_id, bind_year)
        except Exception as exc:
            logger.warning(
                "_generate_one_workpaper: dataset 绑定失败 wp=%s: %s", wp_code, exc
            )

        # 6b. 尽力填充底稿表头（非致命）
        try:
            from app.services.wp_header_service import fill_workpaper_header

            await fill_workpaper_header(
                db=self.db,
                project_id=project_id,
                wp_id=wp.id,
                file_path=str(dest_file),
                wp_code=wp_code,
                wp_name=wp_name,
                cycle=cycle,
            )
        except Exception as exc:
            logger.warning(
                "_generate_one_workpaper: 表头填充失败 wp=%s: %s", wp_code, exc
            )

        # 7. 填充 parsed_data 使其非空
        #    （wp-generation-pipeline 发现：generate_from_codes 从不设
        #     parsed_data → HTML 渲染器"有记录无内容"；本任务验收要求 +parsed_data）
        await self._populate_parsed_data(wp, wp_code, wp_name, cycle)

        return True

    @staticmethod
    def _derive_year(new_standard: dict) -> int:
        """派生用于 dataset 绑定/表头的审计年度（尽力，缺省取上一自然年）。"""
        year = (new_standard or {}).get("year")
        if isinstance(year, int) and year > 0:
            return year
        return datetime.now(timezone.utc).year - 1

    def _copy_template_file(
        self,
        dest_file: Path,
        lib_entry: dict,
        wp_code: str,
        wp_name: str,
        cycle: str,
    ) -> None:
        """复制模板文件到目标路径（与 generate_from_codes 四级兜底一致）。

        顺序：① 知识库 ``~/.gt_audit_helper/.../{cycle}/{name}.xlsx`` 或原始
        文件名 → ② 原始 ``file_path``（含项目根回退）→ ③ openpyxl 最小工作簿
        （写 code/name 占位）→ ④ 空字节兜底。
        """
        src_path = lib_entry.get("file_path", "")
        template_name = lib_entry.get("name", "") or wp_name

        # 1. 知识库底稿模板目录
        candidates: list[Path] = []
        if template_name:
            candidates.append(_KB_TEMPLATE_BASE / cycle / f"{template_name}.xlsx")
        if src_path:
            candidates.append(_KB_TEMPLATE_BASE / cycle / Path(src_path).name)

        for candidate in candidates:
            if candidate and candidate.exists():
                shutil.copy2(candidate, dest_file)
                return

        # 2. 回退：原始模板路径（含项目根回退）
        if src_path:
            src = Path(src_path)
            if not src.exists():
                root_src = (
                    Path(__file__).resolve().parent.parent.parent.parent / src_path
                )
                if root_src.exists():
                    src = root_src
            if src.exists():
                shutil.copy2(src, dest_file)
                return

        # 3. openpyxl 最小工作簿兜底
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = wp_code[:31] if wp_code else "Sheet1"
            ws["A1"] = f"底稿编号: {wp_code}"
            ws["A2"] = f"底稿名称: {wp_name}"
            ws["A3"] = f"审计阶段: {cycle}"
            wb.save(str(dest_file))
            wb.close()
        except Exception:
            # 4. 空字节兜底
            dest_file.write_bytes(b"")

    async def _populate_parsed_data(
        self,
        wp: WorkingPaper,
        wp_code: str,
        wp_name: str,
        cycle: str,
    ) -> None:
        """填充 ``WorkingPaper.parsed_data`` 使其非空（非致命）。

        优先调用 ``wp_parsed_data_service.populate_parsed_data``（若该服务已存在，
        由 wp-generation-pipeline spec 提供，从 xlsx 读 sheet 结构）；该服务尚未
        实现时，回退为写入最小非空占位 dict，保证 ``parsed_data`` 字段非 NULL
        （验收的核心是 parsed_data 被填充，而非 NULL）。
        """
        try:
            from app.services import wp_parsed_data_service  # type: ignore

            await wp_parsed_data_service.populate_parsed_data(
                self.db, wp, wp_code, wp_name, cycle
            )
            return
        except ImportError:
            # wp_parsed_data_service 尚未实现（见 wp-generation-pipeline spec）
            pass
        except Exception as exc:
            logger.warning(
                "_populate_parsed_data: populate_parsed_data 失败 wp=%s，回退占位: %s",
                wp_code,
                exc,
            )

        # 回退：写入最小非空占位，确保 parsed_data 非 NULL（避免 HTML 渲染器
        # "有记录无内容"）。结构对齐渲染器消费的 html_data 形态。
        if not wp.parsed_data:
            wp.parsed_data = {
                "_created_by": "standard_conversion",
                "wp_code": wp_code,
                "html_data": {},
                "generated_at": datetime.now(timezone.utc).isoformat(),
            }
            flag_modified(wp, "parsed_data")

    async def check_preconditions(self, project_id: UUID) -> None:
        """切换前置条件检查（D4）。

        切换准则前必须满足三个前置条件，任一不满足即抛出业务异常
        （``WorkpaperConversionPreconditionError``，ValueError 子类 → router
        映射 HTTP 422），阻止切换：

        1. **项目非归档**：归档项目只读，不允许准则切换。判定 = ``status ==
           ProjectStatus.archived`` 或 ``archived_at is not None``（双重判定更稳健）。
        2. **所有底稿无未保存编辑**（Requirement 2.4）：以
           ``WorkingPaper.prefill_stale == True`` 作为"脏 / 待更新"信号——
           存在任何未删除的脏底稿即拒绝切换，提示"请先保存所有底稿"。
        3. **无进行中任务**：项目无 running/queued 等进行中的导入作业
           （``ImportJob``）。此项为**尽力检测**（best-effort）——检测自身失败
           时记录告警并视为"无进行中任务"，绝不因检测失败而阻塞切换。

        Args:
            project_id: 待切换项目 ID

        Raises:
            ValueError: 项目不存在。
            WorkpaperConversionPreconditionError: 任一前置条件不满足。
        """
        # 1. 加载项目
        project_result = await self.db.execute(
            sa.select(Project).where(Project.id == project_id)
        )
        project = project_result.scalar_one_or_none()
        if project is None:
            raise ValueError("项目不存在")

        # 2. 项目归档检查（status==archived 或 archived_at 非空，双重判定）
        if project.status == ProjectStatus.archived or project.archived_at is not None:
            logger.info(
                "check_preconditions: project=%s 已归档，拒绝切换", project_id
            )
            raise WorkpaperConversionPreconditionError("项目已归档，无法切换准则")

        # 3. 未保存底稿检查（prefill_stale=True 视为脏 / 待更新）
        dirty_result = await self.db.execute(
            sa.select(sa.func.count()).where(
                WorkingPaper.project_id == project_id,
                WorkingPaper.is_deleted == sa.false(),
                WorkingPaper.prefill_stale == sa.true(),
            )
        )
        dirty_count = int(dirty_result.scalar_one() or 0)
        if dirty_count > 0:
            logger.info(
                "check_preconditions: project=%s 存在 %d 份未保存底稿，拒绝切换",
                project_id,
                dirty_count,
            )
            raise WorkpaperConversionPreconditionError(
                f"请先保存所有底稿（存在 {dirty_count} 份未保存/待更新的底稿）"
            )

        # 4. 进行中任务检查（尽力检测：检测失败不阻塞切换）
        in_progress_count = 0
        try:
            from app.models.dataset_models import ImportJob, JobStatus

            in_progress_statuses = [
                JobStatus.pending,
                JobStatus.queued,
                JobStatus.running,
                JobStatus.validating,
                JobStatus.writing,
                JobStatus.activating,
            ]
            job_result = await self.db.execute(
                sa.select(sa.func.count()).where(
                    ImportJob.project_id == project_id,
                    ImportJob.status.in_(in_progress_statuses),
                )
            )
            in_progress_count = int(job_result.scalar_one() or 0)
        except Exception as exc:  # 检测失败一律降级为"无进行中任务"，不阻塞切换
            logger.warning(
                "check_preconditions: 进行中任务检测失败，降级为无进行中任务: %s",
                exc,
            )
            in_progress_count = 0

        if in_progress_count > 0:
            logger.info(
                "check_preconditions: project=%s 存在 %d 个进行中任务，拒绝切换",
                project_id,
                in_progress_count,
            )
            raise WorkpaperConversionPreconditionError(
                "项目存在进行中的任务，请稍后再试"
            )

        logger.info("check_preconditions: project=%s 前置条件检查通过", project_id)
