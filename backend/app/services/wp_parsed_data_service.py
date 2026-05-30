"""底稿 parsed_data 填充服务

从底稿的 xlsx 模板文件读取 sheet 结构，写入 WorkingPaper.parsed_data，
结构对齐 HTML 渲染器消费的 parsed_data['html_data'][sheet_name] 形态。

Feature: wp-generation-pipeline
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

logger = logging.getLogger(__name__)


def _read_xlsx_structure(file_path: str | Path) -> dict[str, dict[str, Any]]:
    """纯函数：openpyxl 读 xlsx 各 sheet → 构建 {sheet_name: {cells, columns}}

    无 DB、无副作用，便于测试。

    Returns:
        dict: {sheet_name: {"cells": {"A1": {"v": value}, ...}, "columns": [...]}}
    """
    import openpyxl

    file_path = Path(file_path)
    if not file_path.exists() or file_path.stat().st_size == 0:
        return {}

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to load workbook %s: %s", file_path, e)
        return {}

    structure: dict[str, dict[str, Any]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells: dict[str, dict[str, Any]] = {}
        columns: list[str] = []

        # 读取所有有值的单元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    coord = cell.coordinate  # e.g. "A1"
                    cells[coord] = {"v": cell.value}
                    # 收集列字母
                    col_letter = cell.column_letter
                    if col_letter not in columns:
                        columns.append(col_letter)

        structure[sheet_name] = {
            "cells": cells,
            "columns": sorted(columns),
        }

    wb.close()
    return structure


async def populate_parsed_data(
    db: AsyncSession,
    wp: Any,
    wp_code: str,
    wp_name: str,
    cycle: str,
) -> None:
    """读取 wp.file_path 的 xlsx，构建 parsed_data.html_data。

    结构（对齐 GtWpRenderer / wp_html_save 的 html_data 形态）：
    parsed_data = {
        "html_data": {
            "<sheet_name>": {
                "cells": { "A1": {"v": ...}, ... },
                "columns": [...],
            },
            ...
        },
        "wp_code": wp_code,
        "generated_at": iso8601,
    }
    """
    structure = _read_xlsx_structure(wp.file_path)

    wp.parsed_data = {
        "html_data": structure,
        "wp_code": wp_code,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    flag_modified(wp, "parsed_data")
