"""穿透查询 API 路由

高性能四表联查：余额→序时账→凭证、余额→辅助余额→辅助明细

- GET /api/projects/{id}/ledger/penetrate          — 统一穿透查询
- GET /api/projects/{id}/ledger/balance             — 科目余额
- GET /api/projects/{id}/ledger/entries/{code}      — 序时账明细
- GET /api/projects/{id}/ledger/voucher/{no}        — 凭证分录
- GET /api/projects/{id}/ledger/aux-balance/{code}  — 辅助余额
- GET /api/projects/{id}/ledger/aux-entries/{code}  — 辅助明细
- DELETE /api/projects/{id}/ledger/cache             — 清除缓存

Validates: Requirements 15.1-15.4
"""

from __future__ import annotations

import io
import urllib.parse
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, Query, UploadFile, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import require_project_access
from app.models.core import User
from app.core.redis import get_redis
from app.services.dataset_query import get_active_filter
from app.services.ledger_penetration_service import LedgerPenetrationService

router = APIRouter(prefix="/api/projects/{project_id}/ledger", tags=["ledger-penetration"])


def _svc(db: AsyncSession, redis) -> LedgerPenetrationService:
    return LedgerPenetrationService(db, redis)


@router.get("/penetrate")
async def penetrate(
    project_id: UUID,
    year: int = Query(...),
    account_code: str | None = None,
    drill_level: str = "all",
    date_from: str | None = None,
    date_to: str | None = None,
    page: int = 1,
    page_size: int = 100,
    db: AsyncSession = Depends(get_db),
    redis=Depends(get_redis),
):
    """统一穿透查询（带缓存）"""
    svc = _svc(db, redis)
    return await svc.penetrate_cached(
        project_id, year, account_code, drill_level,
        date_from, date_to, page, page_size,
    )


@router.get("/balance")
async def get_balance(
    project_id: UUID,
    year: int = Query(...),
    account_code: str | None = None,
    db: AsyncSession = Depends(get_db),
    redis=Depends(get_redis),
    current_user: User = Depends(require_project_access("readonly")),
):
    """科目余额汇总（Redis 缓存 5 分钟）"""
    import json
    cache_key = f"ledger:balance:{project_id}:{year}:{account_code or 'all'}"
    if redis:
        try:
            cached = await redis.get(cache_key)
            if cached:
                return json.loads(cached)
        except Exception:
            pass
    svc = _svc(db, None)
    result = await svc.get_balance_summary(project_id, year, account_code)
    if redis:
        try:
            await redis.set(cache_key, json.dumps(result, default=str), ex=300)
        except Exception:
            pass
    return result


@router.get("/opening-balance/{account_code}")
async def get_opening_balance(
    project_id: UUID,
    account_code: str,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """获取科目期初余额（用于序时账 running_balance 计算）"""
    svc = _svc(db, None)
    opening = await svc.get_account_opening_balance(project_id, year, account_code)
    return {"opening_balance": float(opening), "account_code": account_code}


@router.get("/entries/{account_code}")
async def get_ledger_entries(
    project_id: UUID,
    account_code: str,
    year: int = Query(...),
    date_from: str | None = None,
    date_to: str | None = None,
    cursor: str | None = Query(None, description="游标分页: date|id 格式"),
    limit: int = Query(100, ge=1, le=1000),
    page: int = 1,
    page_size: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """序时账明细（按科目穿透）— 支持游标分页和传统分页

    当提供 cursor 参数时使用游标分页（推荐大数据量场景），
    否则使用传统 OFFSET 分页。
    """
    svc = _svc(db, None)
    # 优先使用游标分页（首次请求不传 cursor 也走游标分页，用 limit 控制条数）
    if cursor is not None or limit != 100:
        return await svc.get_ledger_entries_cursor(
            project_id, year, account_code,
            cursor=cursor, limit=limit,
            date_from=date_from, date_to=date_to,
        )
    return await svc.get_ledger_entries(
        project_id, year, account_code, date_from, date_to, page, page_size,
    )


@router.get("/voucher/{voucher_no}")
async def get_voucher_entries(
    project_id: UUID,
    voucher_no: str,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """凭证分录明细（按凭证号穿透）"""
    svc = _svc(db, None)
    return await svc.get_voucher_entries(project_id, year, voucher_no)


@router.get("/aux-balance-summary")
async def get_aux_balance_summary(
    project_id: UUID,
    year: int = Query(...),
    dim_type: Optional[str] = Query(None, description="维度类型"),
    db: AsyncSession = Depends(get_db),
    redis=Depends(get_redis),
    current_user: User = Depends(require_project_access("readonly")),
):
    """辅助余额汇总（预计算，按维度+科目+辅助编码分组，Redis 缓存 5 分钟）。

    前端树形视图用这个接口，不再加载12万行原始数据。
    返回：维度类型列表 + 汇总行数据
    """
    import json
    import sqlalchemy as sa

    cache_key = f"ledger:aux_summary:{project_id}:{year}:{dim_type or 'all'}"
    if redis:
        try:
            cached = await redis.get(cache_key)
            if cached:
                return json.loads(cached)
        except Exception:
            pass

    # 维度类型列表（含各类型的记录数）
    r = await db.execute(sa.text("""
        SELECT dim_type, SUM(record_count) as total_records, COUNT(*) as group_count
        FROM tb_aux_balance_summary
        WHERE project_id = :pid AND year = :yr
        GROUP BY dim_type ORDER BY total_records DESC
    """), {"pid": str(project_id), "yr": year})
    dim_types = [{"type": row[0], "total_records": int(row[1]), "group_count": int(row[2])} for row in r.fetchall()]

    # 如果只请求维度类型列表（不需要行数据）
    if dim_type == "__types_only__":
        return {"dim_types": dim_types, "rows": [], "total": 0}

    # 汇总数据（按维度类型筛选）
    params: dict = {"pid": str(project_id), "yr": year}
    sql = """
        SELECT dim_type, account_code, account_name, aux_code, aux_name,
               record_count, opening_balance, debit_amount, credit_amount, closing_balance
        FROM tb_aux_balance_summary
        WHERE project_id = :pid AND year = :yr
    """
    if dim_type:
        sql += " AND dim_type = :dt"
        params["dt"] = dim_type
    sql += " ORDER BY account_code, aux_code"

    r = await db.execute(sa.text(sql), params)
    rows = [
        {
            "dim_type": row[0], "account_code": row[1], "account_name": row[2],
            "aux_code": row[3], "aux_name": row[4], "record_count": row[5],
            "opening_balance": float(row[6]) if row[6] else 0,
            "debit_amount": float(row[7]) if row[7] else 0,
            "credit_amount": float(row[8]) if row[8] else 0,
            "closing_balance": float(row[9]) if row[9] else 0,
        }
        for row in r.fetchall()
    ]

    result = {"dim_types": dim_types, "rows": rows, "total": len(rows)}
    if redis:
        try:
            import json
            await redis.set(cache_key, json.dumps(result, default=str), ex=300)
        except Exception:
            pass
    return result


@router.get("/aux-balance-paged")
async def get_aux_balance_paged(
    project_id: UUID,
    year: int = Query(...),
    dim_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    filter: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """辅助余额表分页查询（后端筛选+分页，前端不加载全量数据）"""
    import sqlalchemy as sa
    from app.models.audit_platform_models import TbAuxBalance
    tbl = TbAuxBalance.__table__

    where = [await get_active_filter(db, tbl, project_id, year)]
    if dim_type and dim_type != '全部':
        where.append(tbl.c.aux_type == dim_type)
    if search:
        kw = f"%{search}%"
        where.append(sa.or_(
            tbl.c.account_code.ilike(kw), tbl.c.account_name.ilike(kw),
            tbl.c.aux_name.ilike(kw), tbl.c.aux_code.ilike(kw),
        ))
    if filter == "closing":
        where.append(tbl.c.closing_balance != 0)
    elif filter == "opening":
        where.append(tbl.c.opening_balance != 0)
    elif filter == "changed":
        where.append(sa.or_(tbl.c.debit_amount != 0, tbl.c.credit_amount != 0))

    # 总数
    count_r = await db.execute(sa.select(sa.func.count()).select_from(
        sa.select(tbl.c.id).where(*where).subquery()
    ))
    total = count_r.scalar() or 0

    # 分页数据
    stmt = (
        sa.select(
            tbl.c.account_code, tbl.c.account_name, tbl.c.aux_type,
            tbl.c.aux_code, tbl.c.aux_name, tbl.c.opening_balance,
            tbl.c.debit_amount, tbl.c.credit_amount, tbl.c.closing_balance,
            tbl.c.aux_dimensions_raw,
        )
        .where(*where)
        .order_by(tbl.c.account_code, tbl.c.aux_type, tbl.c.aux_code)
        .offset((page - 1) * page_size).limit(page_size)
    )
    result = await db.execute(stmt)
    rows = [dict(r._mapping) for r in result.fetchall()]

    return {"rows": rows, "total": total, "page": page, "page_size": page_size}


@router.get("/aux-balance-detail")
async def get_aux_balance_detail(
    project_id: UUID,
    year: int = Query(...),
    account_code: str = Query(...),
    dim_type: str = Query(...),
    aux_code: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """辅助余额明细查询（树形展开时按需加载）"""
    import sqlalchemy as sa
    from app.models.audit_platform_models import TbAuxBalance
    tbl = TbAuxBalance.__table__

    where = [
        await get_active_filter(db, tbl, project_id, year),
        tbl.c.account_code == account_code,
        tbl.c.aux_type == dim_type,
    ]
    if aux_code:
        where.append(tbl.c.aux_code == aux_code)

    stmt = sa.select(
        tbl.c.account_code, tbl.c.account_name, tbl.c.aux_type,
        tbl.c.aux_code, tbl.c.aux_name, tbl.c.opening_balance,
        tbl.c.debit_amount, tbl.c.credit_amount, tbl.c.closing_balance,
        tbl.c.aux_dimensions_raw,
    ).where(*where).order_by(tbl.c.aux_code)

    result = await db.execute(stmt)
    return [dict(r._mapping) for r in result.fetchall()]


@router.get("/aux-balance-all")
async def get_all_aux_balance(
    project_id: UUID,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """全量辅助余额（所有科目的辅助核算维度）"""
    svc = _svc(db, None)
    return await svc.get_all_aux_balance(project_id, year)


@router.get("/aux-balance/{account_code}")
async def get_aux_balance(
    project_id: UUID,
    account_code: str,
    year: int = Query(...),
    aux_type: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """辅助余额（按科目穿透到辅助维度）"""
    svc = _svc(db, None)
    return await svc.get_aux_balance(project_id, year, account_code, aux_type)


@router.get("/aux-entries/{account_code}")
async def get_aux_ledger_entries(
    project_id: UUID,
    account_code: str,
    year: int = Query(...),
    aux_type: str | None = None,
    aux_code: str | None = None,
    cursor: str | None = Query(None, description="游标分页: date|id 格式"),
    limit: int = Query(100, ge=1, le=1000),
    page: int = 1,
    page_size: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """辅助明细账（按辅助维度穿透）— 支持游标分页和传统分页"""
    svc = _svc(db, None)
    if cursor is not None:
        return await svc.get_aux_ledger_entries_cursor(
            project_id, year, account_code,
            cursor=cursor, limit=limit,
            aux_type=aux_type, aux_code=aux_code,
        )
    return await svc.get_aux_ledger_entries(
        project_id, year, account_code, aux_type, aux_code, page, page_size,
    )


@router.get("/aux/by-triplet")
async def get_aux_by_triplet(
    project_id: UUID,
    year: int = Query(...),
    account_code: str = Query(..., description="科目编码"),
    aux_type: str = Query(..., description="辅助维度类型，如 客户/成本中心/金融机构"),
    aux_code: str | None = Query(None, description="辅助编码（缺省则返回该维度类型下所有）"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """S6-8 三元组精确穿透（account_code + aux_type + aux_code）。

    同时返回辅助余额 + 辅助明细账，解决"税率"等维度类型跨科目重名问题。
    前端点击科目下的某条"客户:041108,重庆医药..."时调用此端点。
    """
    svc = _svc(db, None)
    return await svc.get_aux_by_triplet(
        project_id, year, account_code, aux_type, aux_code,
        page=page, page_size=page_size,
    )


@router.delete("/cache")
async def clear_cache(
    project_id: UUID,
    year: int = Query(...),
    redis=Depends(get_redis),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("edit")),
):
    """清除穿透查询缓存"""
    svc = _svc(db, redis)
    count = await svc.invalidate_cache(project_id, year)
    return {"cleared": count, "message": f"已清除 {count} 条缓存"}


@router.post("/upload")
async def upload_data(
    project_id: UUID,
    year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("edit")),
):
    """上传四表数据文件（支持历史年度）."""
    raise HTTPException(
        status_code=410,
        detail="旧 /ledger/upload 导入入口已废弃，请改用 /ledger/smart-preview 与 /ledger/smart-import",
    )


@router.post("/upload-multi")
async def upload_multi_files(
    project_id: UUID,
    year: int = Query(...),
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("edit")),
):
    """上传多个四表数据文件（支持多个序时账文件合并导入）."""
    raise HTTPException(
        status_code=410,
        detail="旧 /ledger/upload-multi 导入入口已废弃，请改用 /ledger/smart-preview 与 /ledger/smart-import",
    )


# ─────────────────────────────────────────────────────────────────────────────
# 智能导入（通用引擎，支持双行表头 + 核算维度拆分 + 多文件多年度）
# ─────────────────────────────────────────────────────────────────────────────


@router.post("/smart-preview")
async def smart_preview(
    project_id: UUID,
    files: list[UploadFile] | None = File(None),
    upload_token: Optional[str] = Query(None, description="上传产物令牌，可复用预览上传的文件"),
    year: Optional[int] = Query(None, description="指定年度（不指定则自动提取）"),
    preview_rows: int = Query(50, description="预览模式最大解析行数（默认50，限制内存占用）"),
    current_user: User = Depends(require_project_access("readonly")),
):
    """智能预览：轻量解析前 N 行，返回识别结果 + 列映射 + 总行估算。

    不写入数据库，不转换数据，只解析表头 + 前 N 行做数据类型识别，
    供用户确认后再调用 smart-import 后台异步写入。
    """
    from app.services.ledger_import_application_service import LedgerImportApplicationService

    try:
        return await LedgerImportApplicationService.preview(
            project_id=project_id,
            user_id=str(current_user.id),
            files=files,
            upload_token=upload_token,
            year=year,
            preview_rows=preview_rows,
        )
    except Exception as e:
        import traceback, logging
        logging.getLogger(__name__).error(f"smart-preview error: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"文件解析失败: {str(e)[:200]}")


@router.post("/smart-import")
async def smart_import(
    project_id: UUID,
    files: list[UploadFile] | None = File(None),
    upload_token: Optional[str] = Query(None, description="预览阶段上传产物令牌"),
    year: Optional[int] = Query(None, description="指定年度（不指定则自动提取）"),
    custom_mapping: Optional[str] = Form(None, description="自定义列映射JSON"),
    custom_mapping_query: Optional[str] = Query(None, description="兼容旧版 query custom_mapping"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("edit")),
):
    """智能导入（异步后台）：解析多个文件并写入数据库。

    立即返回 accepted，前端通过 /api/data-lifecycle/import-queue/{project_id} 轮询进度。
    适合大文件（>10MB 或 >100万行）避免 HTTP 超时。
    """
    from app.services.ledger_import_application_service import LedgerImportApplicationService

    return await LedgerImportApplicationService.submit_import_job(
        project_id=project_id,
        user_id=str(current_user.id),
        db=db,
        files=files,
        upload_token=upload_token,
        year=year,
        custom_mapping=custom_mapping or custom_mapping_query,
        payload_style="ledger",
    )


@router.get("/years")
async def get_available_years(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """获取该项目有数据的年度列表"""
    import sqlalchemy as sa
    from app.models.dataset_models import DatasetStatus, LedgerDataset
    from app.models.audit_platform_models import TbBalance

    dataset_result = await db.execute(
        sa.select(sa.distinct(LedgerDataset.year))
        .where(
            LedgerDataset.project_id == project_id,
            LedgerDataset.status == DatasetStatus.active,
        )
        .order_by(LedgerDataset.year.desc())
    )
    dataset_years = [row[0] for row in dataset_result.fetchall()]
    if dataset_years:
        return {"years": dataset_years}

    tbl = TbBalance.__table__
    result = await db.execute(
        sa.select(sa.distinct(tbl.c.year))
        .where(tbl.c.project_id == project_id, tbl.c.is_deleted == sa.false())
        .order_by(tbl.c.year.desc())
    )
    years = [row[0] for row in result.fetchall()]
    return {"years": years}


@router.get("/stats")
async def get_data_stats(
    project_id: UUID,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """获取四表数据统计（行数、最后导入时间）"""
    import sqlalchemy as sa
    from app.models.audit_platform_models import TbBalance, TbAuxBalance, TbLedger, TbAuxLedger, ImportBatch

    stats = {}
    for name, model in [("balance", TbBalance), ("aux_balance", TbAuxBalance),
                         ("ledger", TbLedger), ("aux_ledger", TbAuxLedger)]:
        tbl = model.__table__
        active_filter = await get_active_filter(db, tbl, project_id, year)
        r = await db.execute(
            sa.select(sa.func.count()).where(active_filter)
        )
        stats[name] = r.scalar() or 0

    # 最后导入时间
    r = await db.execute(
        sa.select(ImportBatch.completed_at)
        .where(ImportBatch.project_id == project_id, ImportBatch.year == year)
        .order_by(ImportBatch.completed_at.desc())
        .limit(1)
    )
    last_import = r.scalar()

    return {
        "year": year,
        "counts": stats,
        "total": sum(stats.values()),
        "last_import": last_import.isoformat() if last_import else None,
    }


@router.get("/export-ledger/{account_code}")
async def export_ledger_excel(
    project_id: UUID,
    account_code: str,
    year: int = Query(...),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """导出序时账为 Excel（含期初余额行+月小计+累计余额）"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.models.audit_platform_models import TbLedger, TbBalance
    import sqlalchemy as sa
    from decimal import Decimal

    # 获取期初余额
    tbl_b = TbBalance.__table__
    if account_code.endswith('*'):
        prefix = account_code[:-1]
        code_filter_b = tbl_b.c.account_code.like(prefix + '%')
    else:
        code_filter_b = (tbl_b.c.account_code == account_code)
    ob_r = await db.execute(
        sa.select(sa.func.coalesce(sa.func.sum(tbl_b.c.opening_balance), 0))
        .where(await get_active_filter(db, tbl_b, project_id, year), code_filter_b)
    )
    opening = float(ob_r.scalar() or 0)

    # 获取序时账数据
    tbl = TbLedger.__table__
    if account_code.endswith('*'):
        code_filter = tbl.c.account_code.like(prefix + '%')
    else:
        code_filter = (tbl.c.account_code == account_code)
    where = [await get_active_filter(db, tbl, project_id, year), code_filter]
    if date_from:
        where.append(tbl.c.voucher_date >= date_from)
    if date_to:
        where.append(tbl.c.voucher_date <= date_to)
    stmt = (
        sa.select(tbl.c.voucher_date, tbl.c.voucher_no, tbl.c.summary,
                   tbl.c.debit_amount, tbl.c.credit_amount, tbl.c.counterpart_account)
        .where(*where).order_by(tbl.c.voucher_date, tbl.c.voucher_no)
    )
    result = await db.execute(stmt)
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    acct_label = account_code.replace('*', '')
    ws.title = f"序时账_{acct_label}"

    headers = ["日期", "凭证号", "摘要", "借方", "贷方", "余额", "对方科目"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0ECF7", end_color="F0ECF7", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 期初余额行
    opening_font = Font(bold=True, italic=True, color="4B2D77")
    ws.cell(2, 1, "")
    ws.cell(2, 2, "")
    ws.cell(2, 3, "期初余额")
    ws.cell(2, 6, opening)
    for c in range(1, 8):
        ws.cell(2, c).font = opening_font

    subtotal_font = Font(bold=True)
    subtotal_fill = PatternFill(start_color="FEF6E6", end_color="FEF6E6", fill_type="solid")

    balance = opening
    month_debit = 0.0
    month_credit = 0.0
    last_month = ""
    excel_row = 3

    for row in rows:
        vd = row[0]
        vd_str = vd.isoformat() if hasattr(vd, 'isoformat') else str(vd or '')
        month = vd_str[:7]
        d = float(row[3] or 0)
        c = float(row[4] or 0)
        balance += d - c
        month_debit += d
        month_credit += c

        if not last_month:
            last_month = month

        # 月份变化时插入上月小计
        if month != last_month and last_month:
            ws.cell(excel_row, 3, f"{last_month} 本月合计")
            ws.cell(excel_row, 4, month_debit - d)
            ws.cell(excel_row, 5, month_credit - c)
            ws.cell(excel_row, 6, balance - d + c)
            for col in range(1, 8):
                ws.cell(excel_row, col).font = subtotal_font
                ws.cell(excel_row, col).fill = subtotal_fill
            excel_row += 1
            month_debit = d
            month_credit = c
            last_month = month

        ws.cell(excel_row, 1, vd_str)
        ws.cell(excel_row, 2, row[1])
        ws.cell(excel_row, 3, row[2])
        ws.cell(excel_row, 4, d if d else None)
        ws.cell(excel_row, 5, c if c else None)
        ws.cell(excel_row, 6, balance)
        ws.cell(excel_row, 7, row[5])
        excel_row += 1

    # 最后一个月的小计
    if rows:
        ws.cell(excel_row, 3, f"{last_month} 本月合计")
        ws.cell(excel_row, 4, month_debit)
        ws.cell(excel_row, 5, month_credit)
        ws.cell(excel_row, 6, balance)
        for col in range(1, 8):
            ws.cell(excel_row, col).font = subtotal_font
            ws.cell(excel_row, col).fill = subtotal_fill

    widths = [12, 12, 30, 16, 16, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for col in [4, 5, 6]:
        for r in range(2, excel_row + 1):
            ws.cell(r, col).number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"序时账_{acct_label}_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"},
    )


@router.get("/export-balance")
async def export_balance_excel(
    project_id: UUID,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """导出科目余额表为 Excel"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from app.models.audit_platform_models import TbBalance
    import sqlalchemy as sa

    tbl = TbBalance.__table__
    stmt = (
        sa.select(
            tbl.c.account_code, tbl.c.account_name, tbl.c.level,
            tbl.c.opening_balance, tbl.c.debit_amount,
            tbl.c.credit_amount, tbl.c.closing_balance,
        )
        .where(await get_active_filter(db, tbl, project_id, year))
        .order_by(tbl.c.account_code)
    )
    result = await db.execute(stmt)
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "科目余额表"

    headers = ["科目编号", "科目名称", "级次", "期初余额", "借方发生额", "贷方发生额", "期末余额"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0ECF7", end_color="F0ECF7", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    level1_font = Font(bold=True)
    level1_fill = PatternFill(start_color="F8F5FC", end_color="F8F5FC", fill_type="solid")

    for idx, row in enumerate(rows, 2):
        level = row[2] or 1
        ws.cell(idx, 1, row[0])
        ws.cell(idx, 2, row[1])
        ws.cell(idx, 3, level)
        ws.cell(idx, 4, float(row[3]) if row[3] else None)
        ws.cell(idx, 5, float(row[4]) if row[4] else None)
        ws.cell(idx, 6, float(row[5]) if row[5] else None)
        ws.cell(idx, 7, float(row[6]) if row[6] else None)
        # 一级科目加粗+浅紫背景
        if level == 1:
            for c in range(1, 8):
                ws.cell(idx, c).font = level1_font
                ws.cell(idx, c).fill = level1_fill

    widths = [16, 24, 6, 16, 16, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for col in [4, 5, 6, 7]:
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row_idx, col).number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"科目余额表_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"},
    )


@router.get("/export-aux-balance")
async def export_aux_balance_excel(
    project_id: UUID,
    year: int = Query(...),
    dim_type: Optional[str] = Query(None, description="维度类型筛选"),
    search: Optional[str] = Query(None, description="搜索关键词"),
    filter: Optional[str] = Query(None, description="筛选条件: closing/opening/changed"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """导出辅助余额表为 Excel（支持当前视图条件）"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.models.audit_platform_models import TbAuxBalance
    import sqlalchemy as sa

    tbl = TbAuxBalance.__table__
    stmt = (
        sa.select(
            tbl.c.account_code, tbl.c.account_name,
            tbl.c.aux_type, tbl.c.aux_code, tbl.c.aux_name,
            tbl.c.opening_balance, tbl.c.debit_amount,
            tbl.c.credit_amount, tbl.c.closing_balance,
            tbl.c.aux_dimensions_raw,
        )
        .where(await get_active_filter(db, tbl, project_id, year))
        .order_by(tbl.c.account_code, tbl.c.aux_type, tbl.c.aux_code)
    )
    if dim_type:
        stmt = stmt.where(tbl.c.aux_type == dim_type)
    if search:
        kw = f"%{search}%"
        stmt = stmt.where(sa.or_(
            tbl.c.account_code.ilike(kw), tbl.c.account_name.ilike(kw),
            tbl.c.aux_name.ilike(kw), tbl.c.aux_code.ilike(kw),
        ))
    if filter == "closing":
        stmt = stmt.where(tbl.c.closing_balance != 0)
    elif filter == "opening":
        stmt = stmt.where(tbl.c.opening_balance != 0)
    elif filter == "changed":
        stmt = stmt.where(sa.or_(tbl.c.debit_amount != 0, tbl.c.credit_amount != 0))

    result = await db.execute(stmt)
    rows = result.fetchall()

    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "辅助余额表"

    headers = ["科目编号", "科目名称", "辅助类型", "辅助编码", "辅助名称", "关联维度", "期初余额", "借方发生额", "贷方发生额", "期末余额"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0ECF7", end_color="F0ECF7", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 按 account_code + aux_code 分组，插入小计行
    from collections import OrderedDict
    from decimal import Decimal

    subtotal_font = Font(bold=True)
    subtotal_fill = PatternFill(start_color="FEF6E6", end_color="FEF6E6", fill_type="solid")

    groups = OrderedDict()  # key -> list of row tuples
    for row in rows:
        key = f"{row[0]}|{row[3]}"  # account_code|aux_code
        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    excel_row = 2
    for key, group_rows in groups.items():
        # 多条时先写小计行
        if len(group_rows) > 1:
            s_open = sum(float(r[5] or 0) for r in group_rows)
            s_debit = sum(float(r[6] or 0) for r in group_rows)
            s_credit = sum(float(r[7] or 0) for r in group_rows)
            s_close = sum(float(r[8] or 0) for r in group_rows)
            first = group_rows[0]
            ws.cell(excel_row, 1, first[0])
            ws.cell(excel_row, 2, first[1])
            ws.cell(excel_row, 3, first[2])
            ws.cell(excel_row, 4, first[3])
            ws.cell(excel_row, 5, f"{first[4]} 小计({len(group_rows)}条)")
            ws.cell(excel_row, 6, "")
            ws.cell(excel_row, 7, s_open)
            ws.cell(excel_row, 8, s_debit)
            ws.cell(excel_row, 9, s_credit)
            ws.cell(excel_row, 10, s_close)
            for col in range(1, 11):
                ws.cell(excel_row, col).font = subtotal_font
                ws.cell(excel_row, col).fill = subtotal_fill
            excel_row += 1

        # 写明细行
        for row in group_rows:
            ws.cell(excel_row, 1, row[0])
            ws.cell(excel_row, 2, row[1])
            ws.cell(excel_row, 3, row[2])
            ws.cell(excel_row, 4, row[3])
            ws.cell(excel_row, 5, row[4])
            raw = row[9] or ""
            if raw and dim_type:
                parts = [p.strip() for p in raw.split(";") if p.strip() and not p.strip().startswith(dim_type + ":")]
                ws.cell(excel_row, 6, "; ".join(parts))
            else:
                ws.cell(excel_row, 6, raw)
            ws.cell(excel_row, 7, float(row[5]) if row[5] else None)
            ws.cell(excel_row, 8, float(row[6]) if row[6] else None)
            ws.cell(excel_row, 9, float(row[7]) if row[7] else None)
            ws.cell(excel_row, 10, float(row[8]) if row[8] else None)
            excel_row += 1

    total_rows = excel_row - 1

    # 设置列宽
    widths = [14, 18, 10, 14, 24, 40, 16, 16, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # 数字格式
    for col in [7, 8, 9, 10]:
        for row_idx in range(2, total_rows + 1):
            ws.cell(row_idx, col).number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"辅助余额表_{dim_type or '全部'}_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"},
    )


# ── 入库后数据一致性校验（按需触发） ──

@router.get("/validate")
async def validate_data(
    project_id: UUID,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """四表数据一致性校验（入库后按需触发）。

    从数据库查询做校验，比预览阶段更准确（有完整的辅助明细账数据）。
    校验项：
    1. 科目余额表内部勾稽（期初+借-贷=期末）
    2. 辅助余额表内部勾稽
    3. 科目余额表 vs 辅助余额表（每个维度类型汇总应等于科目余额）
    4. 序时账 vs 科目余额表（按科目汇总借贷发生额）
    """
    import sqlalchemy as sa
    from decimal import Decimal
    from app.models.audit_platform_models import TbAuxBalance, TbBalance, TbLedger

    findings = []

    # 1. 科目余额表内部勾稽
    bal_tbl = TbBalance.__table__
    bal_rows = await db.execute(
        sa.select(
            bal_tbl.c.account_code,
            bal_tbl.c.opening_balance,
            bal_tbl.c.debit_amount,
            bal_tbl.c.credit_amount,
            bal_tbl.c.closing_balance,
        ).where(
            await get_active_filter(db, bal_tbl, project_id, year),
            sa.or_(
                bal_tbl.c.opening_balance.isnot(None),
                bal_tbl.c.debit_amount.isnot(None),
                bal_tbl.c.credit_amount.isnot(None),
                bal_tbl.c.closing_balance.isnot(None),
            ),
        )
    )
    bal_data = bal_rows.fetchall()

    bal_map = {}
    bal_errors = 0
    for r in bal_data:
        code, opening, debit, credit, closing = r
        opening = opening or Decimal(0)
        debit = debit or Decimal(0)
        credit = credit or Decimal(0)
        closing = closing or Decimal(0)
        bal_map[code] = {"opening": opening, "debit": debit, "credit": credit, "closing": closing}
        expected = opening + debit - credit
        if abs(expected - closing) > Decimal("0.01"):
            bal_errors += 1
            if bal_errors <= 5:
                findings.append({
                    "level": "error", "category": "余额表勾稽",
                    "message": f"{code}: 期初{opening}+借{debit}-贷{credit}={expected}, 期末{closing}, 差{expected-closing}",
                })
    if bal_errors > 5:
        findings.append({"level": "error", "category": "余额表勾稽",
                         "message": f"共 {bal_errors} 个科目不平"})
    elif bal_errors == 0:
        findings.append({"level": "info", "category": "余额表勾稽",
                         "message": f"{len(bal_data)} 个科目全部勾稽通过"})

    # 2. 科目余额表 vs 辅助余额表
    aux_tbl = TbAuxBalance.__table__
    aux_agg = await db.execute(
        sa.select(
            aux_tbl.c.account_code,
            aux_tbl.c.aux_type,
            sa.func.count().label("cnt"),
            sa.func.sum(sa.func.coalesce(aux_tbl.c.closing_balance, 0)).label("total_closing"),
        )
        .where(await get_active_filter(db, aux_tbl, project_id, year))
        .group_by(aux_tbl.c.account_code, aux_tbl.c.aux_type)
    )
    aux_agg_data = aux_agg.fetchall()

    from collections import defaultdict
    aux_by_code = defaultdict(list)
    for code, aux_type, cnt, total_closing in aux_agg_data:
        aux_by_code[code].append((aux_type, cnt, total_closing or Decimal(0)))

    cross_errors = 0
    for code in sorted(aux_by_code.keys()):
        if code not in bal_map:
            continue
        b_closing = bal_map[code]["closing"]
        types = aux_by_code[code]
        best_type, best_cnt, best_closing = min(types, key=lambda x: abs(b_closing - x[2]))
        diff = b_closing - best_closing
        if abs(diff) > Decimal("0.01"):
            cross_errors += 1
            if cross_errors <= 5:
                findings.append({
                    "level": "warning", "category": "余额表vs辅助余额表",
                    "message": f"{code}: 科目期末={b_closing}, 维度({best_type},{best_cnt}条)汇总={best_closing}, 差{diff}",
                })
    if cross_errors > 5:
        findings.append({"level": "warning", "category": "余额表vs辅助余额表",
                         "message": f"共 {cross_errors} 个科目不一致"})
    elif cross_errors == 0 and aux_by_code:
        findings.append({"level": "info", "category": "余额表vs辅助余额表",
                         "message": f"{len(aux_by_code)} 个有辅助核算的科目全部一致"})

    # 3. 序时账 vs 科目余额表
    led_tbl = TbLedger.__table__
    led_agg = await db.execute(
        sa.select(
            led_tbl.c.account_code,
            sa.func.sum(sa.func.coalesce(led_tbl.c.debit_amount, 0)).label("total_debit"),
            sa.func.sum(sa.func.coalesce(led_tbl.c.credit_amount, 0)).label("total_credit"),
        )
        .where(await get_active_filter(db, led_tbl, project_id, year))
        .group_by(led_tbl.c.account_code)
    )
    led_agg_data = led_agg.fetchall()

    led_errors = 0
    for code, led_debit, led_credit in led_agg_data:
        if code not in bal_map:
            continue
        b = bal_map[code]
        led_debit = led_debit or Decimal(0)
        led_credit = led_credit or Decimal(0)
        if abs(b["debit"] - led_debit) > Decimal("0.01") or abs(b["credit"] - led_credit) > Decimal("0.01"):
            led_errors += 1
            if led_errors <= 5:
                findings.append({
                    "level": "warning", "category": "序时账vs余额表",
                    "message": f"{code}: 余额表借{b['debit']}/贷{b['credit']}, 序时账借{led_debit}/贷{led_credit}",
                })
    if led_errors > 5:
        findings.append({"level": "warning", "category": "序时账vs余额表",
                         "message": f"共 {led_errors} 个科目不一致"})
    elif led_errors == 0 and led_agg_data:
        findings.append({"level": "info", "category": "序时账vs余额表",
                         "message": f"{len(led_agg_data)} 个科目全部一致"})

    return {
        "year": year,
        "findings": findings,
        "summary": {
            "balance_count": len(bal_data),
            "aux_account_count": len(aux_by_code),
            "ledger_account_count": len(led_agg_data),
            "errors": sum(1 for f in findings if f["level"] == "error"),
            "warnings": sum(1 for f in findings if f["level"] == "warning"),
        },
    }


@router.get("/balance-tree")
async def get_balance_tree(
    project_id: UUID,
    year: int = Query(...),
    company_code: str | None = Query(None, description="可选公司代码，默认合并全部"),
    page: int = Query(1, ge=1, description="页码（1-based）"),
    page_size: int = Query(
        100, ge=1, le=200,
        description="每页科目数，用户自定义，最多 200",
    ),
    keyword: str | None = Query(
        None,
        description="按 account_code / account_name 模糊过滤（大小写不敏感）",
    ),
    only_with_children: bool = Query(
        False, description="仅返回含辅助维度的科目（前端维度过滤）",
    ),
    only_with_activity: bool = Query(
        False,
        description=(
            "仅返回有金额活动的科目：对资产/负债/权益类要求 opening/closing/debit/credit "
            "任一非零；对损益类（5/6 开头）只要求 debit/credit 任一非零"
            "（损益类期末结转后 opening/closing 天然为 NULL，不应参与判定）"
        ),
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """科目余额树形视图（Layer 2 v2 / Sprint 8 P3 分页）。

    返回主表 + 辅助余额嵌套结构，前端 el-table :tree-props 直接渲染：

        [
          {account_code, account_name, level, opening_balance, ...,
           aggregated_from_aux: bool, aux_row_count: int,
           children: [
             {aux_type, aux_code, aux_name, aux_dimensions_raw,
              opening_balance, closing_balance, ...},
             ...
           ]},
          ...
        ]

    children 按 (aux_type, aux_code) 排序；父节点是 tb_balance 行，
    子节点来自 tb_aux_balance 同 account_code 聚合。
    """
    import sqlalchemy as sa
    from app.models.audit_platform_models import TbBalance, TbAuxBalance

    bal_tbl = TbBalance.__table__
    aux_tbl = TbAuxBalance.__table__

    bal_where = [await get_active_filter(db, bal_tbl, project_id, year)]
    aux_where = [await get_active_filter(db, aux_tbl, project_id, year)]
    if company_code:
        bal_where.append(bal_tbl.c.company_code == company_code)
        aux_where.append(aux_tbl.c.company_code == company_code)

    # keyword 模糊过滤（作用于主表 account_code / account_name）
    if keyword:
        kw = f"%{keyword.strip()}%"
        bal_where.append(
            sa.or_(
                bal_tbl.c.account_code.ilike(kw),
                bal_tbl.c.account_name.ilike(kw),
            )
        )

    # only_with_children=True 时，先取有辅助的 account_code 集合，加到 where 过滤主表
    if only_with_children:
        codes_stmt = sa.select(aux_tbl.c.account_code).where(*aux_where).distinct()
        codes_result = await db.execute(codes_stmt)
        codes_with_aux = [r[0] for r in codes_result.fetchall()]
        if codes_with_aux:
            bal_where.append(bal_tbl.c.account_code.in_(codes_with_aux))
        else:
            # 没有任何含辅助的科目 → 返回空页
            return {
                "year": year,
                "company_code": company_code,
                "tree": [],
                "pagination": {
                    "page": page, "page_size": page_size,
                    "total": 0, "total_pages": 0,
                },
                "summary": {
                    "account_count": 0,
                    "aggregated_count": 0,
                    "with_children_count": 0,
                    "aux_total_rows": 0,
                    "mismatches": [],
                },
            }

    # only_with_activity：按科目类型差异化判断
    # - 资产/负债/权益类（1/2/3/4 开头）：opening/closing/debit/credit 任一非零
    # - 损益类（5/6 开头）：debit/credit 任一非零（opening/closing 天然 NULL 不参与判定）
    # 实现：用 OR 拼两种场景而非 CASE，让查询优化器能走索引
    # 用 substr 而非 left（SQLite 测试兼容，PG 也支持）
    if only_with_activity:
        def _nonzero(col):
            return sa.and_(col.is_not(None), col != 0)

        first_char = sa.func.substr(bal_tbl.c.account_code, 1, 1)
        # 场景 1：损益类（5/6）只看 debit/credit
        loss_gain_active = sa.and_(
            first_char.in_(("5", "6")),
            sa.or_(
                _nonzero(bal_tbl.c.debit_amount),
                _nonzero(bal_tbl.c.credit_amount),
            ),
        )
        # 场景 2：非损益类 + 任一金额字段非零
        other_active = sa.and_(
            sa.not_(first_char.in_(("5", "6"))),
            sa.or_(
                _nonzero(bal_tbl.c.debit_amount),
                _nonzero(bal_tbl.c.credit_amount),
                _nonzero(bal_tbl.c.opening_balance),
                _nonzero(bal_tbl.c.closing_balance),
            ),
        )
        bal_where.append(sa.or_(loss_gain_active, other_active))

    # 分页前先查总数
    total_stmt = sa.select(sa.func.count()).select_from(bal_tbl).where(*bal_where)
    total = (await db.execute(total_stmt)).scalar() or 0
    total_pages = (total + page_size - 1) // page_size if page_size else 1

    # 主表（带分页）
    offset = (page - 1) * page_size
    bal_stmt = (
        sa.select(
            bal_tbl.c.account_code, bal_tbl.c.account_name, bal_tbl.c.level,
            bal_tbl.c.company_code,
            bal_tbl.c.opening_balance, bal_tbl.c.opening_debit, bal_tbl.c.opening_credit,
            bal_tbl.c.debit_amount, bal_tbl.c.credit_amount,
            bal_tbl.c.closing_balance, bal_tbl.c.closing_debit, bal_tbl.c.closing_credit,
            bal_tbl.c.currency_code, bal_tbl.c.raw_extra,
        )
        .where(*bal_where)
        .order_by(bal_tbl.c.account_code)
        .limit(page_size)
        .offset(offset)
    )
    bal_result = await db.execute(bal_stmt)
    bal_rows_list = bal_result.fetchall()
    # 本页主表的 account_code 集合（用于收窄 aux 查询，避免拉全量）
    page_account_codes = [r[0] for r in bal_rows_list]
    if page_account_codes:
        aux_where.append(aux_tbl.c.account_code.in_(page_account_codes))
    else:
        # 本页无主表数据 → 返回空
        return {
            "year": year,
            "company_code": company_code,
            "tree": [],
            "pagination": {
                "page": page, "page_size": page_size,
                "total": total, "total_pages": total_pages,
            },
            "summary": {
                "account_count": 0,
                "aggregated_count": 0,
                "with_children_count": 0,
                "aux_total_rows": 0,
                "mismatches": [],
            },
        }

    # 辅助余额（按 (company_code, account_code) 分组到父节点，多公司合并场景下正确隔离）
    aux_stmt = sa.select(
        aux_tbl.c.company_code,
        aux_tbl.c.account_code, aux_tbl.c.aux_type, aux_tbl.c.aux_code,
        aux_tbl.c.aux_name, aux_tbl.c.aux_dimensions_raw,
        aux_tbl.c.opening_balance, aux_tbl.c.opening_debit, aux_tbl.c.opening_credit,
        aux_tbl.c.debit_amount, aux_tbl.c.credit_amount,
        aux_tbl.c.closing_balance, aux_tbl.c.closing_debit, aux_tbl.c.closing_credit,
        aux_tbl.c.currency_code,
    ).where(*aux_where).order_by(
        aux_tbl.c.company_code,
        aux_tbl.c.account_code, aux_tbl.c.aux_type, aux_tbl.c.aux_code,
    )
    aux_result = await db.execute(aux_stmt)

    # aux rows 按 (company_code, account_code, aux_type) 三层聚合
    # 真实账务场景：一行多维度在 tb_aux_balance 里会冗余存 N 条（N=维度个数），
    # 按"单一维度类型"聚合才等于主表金额（任一维度完备覆盖全部金额）。
    # 多公司合并场景下，相同 account_code 在不同 company_code 下独立，不能合并。
    from collections import defaultdict
    aux_by_comp_code_type: dict[
        tuple[str, str], dict[str, list[dict]]
    ] = defaultdict(lambda: defaultdict(list))
    for r in aux_result.fetchall():
        aux_by_comp_code_type[(r[0], r[1])][r[2] or ""].append({
            "aux_type": r[2],
            "aux_code": r[3],
            "aux_name": r[4],
            "aux_dimensions_raw": r[5],
            "opening_balance": float(r[6]) if r[6] is not None else None,
            "opening_debit": float(r[7]) if r[7] is not None else None,
            "opening_credit": float(r[8]) if r[8] is not None else None,
            "debit_amount": float(r[9]) if r[9] is not None else None,
            "credit_amount": float(r[10]) if r[10] is not None else None,
            "closing_balance": float(r[11]) if r[11] is not None else None,
            "closing_debit": float(r[12]) if r[12] is not None else None,
            "closing_credit": float(r[13]) if r[13] is not None else None,
            "currency_code": r[14],
        })

    def _build_dimension_nodes(
        company_code: str, account_code: str, account_name: str | None,
    ) -> tuple[list[dict], int, dict[str, float]]:
        """构造 aux_type 分组节点 + 返回每个维度类型的 closing 求和。"""
        dim_nodes: list[dict] = []
        total_aux_rows = 0
        type_sums: dict[str, float] = {}
        for aux_type, rows in aux_by_comp_code_type.get((company_code, account_code), {}).items():
            total_aux_rows += len(rows)
            dim_sum = sum((r["closing_balance"] or 0) for r in rows)
            type_sums[aux_type] = dim_sum
            dim_nodes.append({
                "_is_dimension_group": True,
                "aux_type": aux_type,
                "account_code": account_code,
                "account_name": account_name,
                "closing_balance": dim_sum,
                "opening_balance": sum(
                    (r["opening_balance"] or 0) for r in rows
                ),
                "debit_amount": sum(
                    (r["debit_amount"] or 0) for r in rows
                ),
                "credit_amount": sum(
                    (r["credit_amount"] or 0) for r in rows
                ),
                "record_count": len(rows),
                "has_children": True,
                "children": rows,
            })
        return dim_nodes, total_aux_rows, type_sums

    tree: list[dict] = []
    for r in bal_rows_list:
        raw_extra = r[13] or {}
        aggregated = bool(raw_extra.get("_aggregated_from_aux"))
        aux_row_count = int(raw_extra.get("_aux_row_count") or 0)
        # 传 (company_code=r[3], account_code=r[0], account_name=r[1])，
        # 多公司合并场景下 aux 分组隔离
        dim_nodes, total_aux_rows, _type_sums = _build_dimension_nodes(r[3], r[0], r[1])
        tree.append({
            "account_code": r[0],
            "account_name": r[1],
            "level": r[2],
            "company_code": r[3],
            "opening_balance": float(r[4]) if r[4] is not None else None,
            "opening_debit": float(r[5]) if r[5] is not None else None,
            "opening_credit": float(r[6]) if r[6] is not None else None,
            "debit_amount": float(r[7]) if r[7] is not None else None,
            "credit_amount": float(r[8]) if r[8] is not None else None,
            "closing_balance": float(r[9]) if r[9] is not None else None,
            "closing_debit": float(r[10]) if r[10] is not None else None,
            "closing_credit": float(r[11]) if r[11] is not None else None,
            "currency_code": r[12],
            "aggregated_from_aux": aggregated,
            "aux_row_count": aux_row_count,
            "aux_types": list(_type_sums.keys()),
            "aux_rows_total": total_aux_rows,
            "has_children": len(dim_nodes) > 0,
            "children": dim_nodes,
        })

    # 辅助一致性检查：任一维度类型求和 ≠ 主表即视为 mismatch
    # （按单一 aux_type 聚合应等于主表，这是正确的辅助核算语义）
    # 多公司合并场景下带 company_code 标识，避免用户看两家同 account_code 时混淆
    mismatches: list[dict] = []
    for node in tree:
        if not node["children"]:
            continue
        parent_val = node["closing_balance"] or 0
        for dim_node in node["children"]:
            dim_sum = dim_node["closing_balance"] or 0
            diff = abs(dim_sum - parent_val)
            if diff > 1.0:
                mismatches.append({
                    "company_code": node["company_code"],
                    "account_code": node["account_code"],
                    "aux_type": dim_node["aux_type"],
                    "parent_closing": parent_val,
                    "dim_sum": dim_sum,
                    "record_count": dim_node["record_count"],
                    "diff": diff,
                })

    return {
        "year": year,
        "company_code": company_code,
        "tree": tree,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
        },
        "summary": {
            "account_count": len(tree),
            "aggregated_count": sum(1 for n in tree if n["aggregated_from_aux"]),
            "with_children_count": sum(1 for n in tree if n["has_children"]),
            "aux_total_rows": sum(n["aux_rows_total"] for n in tree),
            "mismatches": mismatches,
        },
    }
