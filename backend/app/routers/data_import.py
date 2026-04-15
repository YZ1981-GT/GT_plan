"""数据导入 API 路由

Validates: Requirements 4.3, 4.4, 4.23
"""

from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.config import settings
from app.deps import get_current_user
from app.models.audit_platform_schemas import (
    ImportBatchResponse,
    ImportProgress,
)
from app.models.core import User
from app.services import import_service

router = APIRouter(prefix="/api/projects/{project_id}/import", tags=["import"])


@router.post("", response_model=ImportBatchResponse)
async def upload_and_import(
    project_id: UUID,
    file: UploadFile = File(...),
    source_type: str = Form(default="generic"),
    data_type: str = Form(...),
    year: int = Form(...),
    on_duplicate: str = Form(default="skip"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportBatchResponse:
    """Upload file and start import (multipart form).

    on_duplicate: skip(跳过重复) / overwrite(覆盖旧数据) / error(报错中止)

    Validates: Requirements 4.3
    """
    # 文件大小校验
    if file.size and file.size > settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=413,
            detail=f"文件大小超过限制（最大 {settings.MAX_UPLOAD_SIZE_MB}MB）",
        )

    batch = await import_service.start_import(
        project_id=project_id,
        file=file,
        source_type=source_type,
        data_type=data_type,
        year=year,
        db=db,
        on_duplicate=on_duplicate,
    )
    return ImportBatchResponse.model_validate(batch)


@router.get("/{batch_id}/progress", response_model=ImportProgress)
async def get_import_progress(
    project_id: UUID,
    batch_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportProgress:
    """Get import progress for a batch.

    Validates: Requirements 4.4
    """
    return await import_service.get_import_progress(batch_id, db)


@router.get("/batches", response_model=list[ImportBatchResponse])
async def list_import_batches(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[ImportBatchResponse]:
    """List all import batches for a project."""
    return await import_service.get_import_batches(project_id, db)


@router.get("/check-duplicates")
async def check_duplicates(
    project_id: UUID,
    data_type: str,
    year: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """导入前检查是否存在重复数据。

    前端在用户点"确认导入"前调用此接口，如果有重复则弹窗让用户选择覆盖或跳过。
    """
    return await import_service.check_duplicates(project_id, year, data_type, db)


@router.post("/{batch_id}/rollback", response_model=ImportBatchResponse)
async def rollback_import(
    project_id: UUID,
    batch_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportBatchResponse:
    """Rollback an import batch.

    Validates: Requirements 4.23
    """
    batch = await import_service.rollback_import(batch_id, db)
    return ImportBatchResponse.model_validate(batch)


@router.post("/preview")
async def preview_file(
    project_id: UUID,
    file: UploadFile = File(...),
    skip_rows: int = Form(default=0),
    current_user: User = Depends(get_current_user),
) -> dict:
    """预览上传文件：返回 sheet 列表、每个 sheet 的表头和前 20 行数据。

    用于大文件导入前让用户确认数据类型和列映射。
    """
    if file.size and file.size > settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=413,
            detail=f"文件大小超过限制（最大 {settings.MAX_UPLOAD_SIZE_MB}MB）",
        )

    content = await file.read()
    sheets = []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            __import__("io").BytesIO(content), read_only=True, data_only=True
        )

        for ws in wb.worksheets:
            if ws.max_row is None or ws.max_row < 1:
                continue

            rows_iter = ws.iter_rows(values_only=True)

            # 跳过指定行数
            for _ in range(skip_rows):
                try:
                    next(rows_iter)
                except StopIteration:
                    break

            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue

            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header_row)]

            # 读取前 20 行预览
            preview_rows = []
            for j, row in enumerate(rows_iter):
                if j >= 20:
                    break
                if all(cell is None for cell in row):
                    continue
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = str(cell) if cell is not None else ""
                preview_rows.append(row_data)

            # 猜测数据类型
            guessed_type = _guess_data_type(headers)

            sheets.append({
                "sheet_name": ws.title,
                "total_rows": (ws.max_row or 1) - 1 - skip_rows,
                "headers": headers,
                "preview_rows": preview_rows,
                "guessed_data_type": guessed_type,
            })

        wb.close()
    except Exception:
        # 尝试 CSV
        for encoding in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            return {"sheets": [], "error": "无法识别文件格式"}

        import csv
        reader = csv.reader(__import__("io").StringIO(text))
        all_rows = list(reader)
        if len(all_rows) > skip_rows:
            headers = [str(h).strip() for h in all_rows[skip_rows]]
            preview = [
                dict(zip(headers, row))
                for row in all_rows[skip_rows + 1 : skip_rows + 21]
            ]
            sheets.append({
                "sheet_name": "CSV",
                "total_rows": len(all_rows) - 1 - skip_rows,
                "headers": headers,
                "preview_rows": preview,
                "guessed_data_type": _guess_data_type(headers),
            })

    return {
        "file_name": file.filename,
        "file_size": file.size,
        "sheets": sheets,
    }


def _guess_data_type(headers: list[str]) -> str | None:
    """根据列名猜测数据类型。"""
    header_set = {h.lower().strip() for h in headers}

    # 序时账特征：有凭证日期+凭证号
    ledger_kw = {"凭证日期", "日期", "凭证号", "凭证编号", "voucher_date", "voucher_no"}
    if header_set & ledger_kw:
        # 区分辅助明细账和序时账
        aux_kw = {"辅助类型", "核算类型", "辅助编码", "aux_type", "aux_code"}
        if header_set & aux_kw:
            return "tb_aux_ledger"
        return "tb_ledger"

    # 余额表特征：有期初+期末
    balance_kw = {"期初余额", "期末余额", "期初", "期末", "opening_balance", "closing_balance"}
    if header_set & balance_kw:
        aux_kw = {"辅助类型", "核算类型", "辅助编码", "aux_type", "aux_code"}
        if header_set & aux_kw:
            return "tb_aux_balance"
        return "tb_balance"

    return None
