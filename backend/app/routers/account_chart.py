"""科目表管理 API 路由

Validates: Requirements 2.2, 2.5, 2.6
"""

from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.audit_platform_schemas import (
    AccountChartResponse,
    AccountImportResult,
    AccountTreeNode,
)
from app.models.core import User
from app.services import account_chart_service

router = APIRouter(prefix="/api/projects", tags=["account-chart"])


def _build_sheet_diagnostics(diagnostics: list[dict] | None) -> list[dict]:
    """规范化智能导入诊断结构，供前端结果页稳定展示。"""
    normalized: list[dict] = []
    for sheet in diagnostics or []:
        column_mapping = sheet.get("column_mapping") or {}
        matched_cols = sheet.get("matched_cols")
        if not isinstance(matched_cols, list):
            matched_cols = sorted(set(column_mapping.values())) if isinstance(column_mapping, dict) else []
        normalized.append({
            "sheet_name": sheet.get("sheet", ""),
            "guessed_type": sheet.get("data_type", "unknown"),
            "matched_cols": matched_cols,
            "missing_cols": sheet.get("missing_cols") or [],
            "missing_recommended": sheet.get("missing_recommended") or [],
            "row_count": sheet.get("row_count", 0),
        })
    return normalized


@router.get(
    "/{project_id}/account-chart/standard",
    response_model=list[AccountChartResponse],
)
async def get_standard_chart(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[AccountChartResponse]:
    """获取项目的标准科目表。

    如果尚未加载，则根据项目会计准则自动加载标准科目模板。

    Validates: Requirements 2.2
    """
    # Always try incremental load (adds missing standard accounts)
    from app.services.project_wizard_service import get_wizard_state
    state = await get_wizard_state(project_id, db)
    accounting_standard = "enterprise"  # default
    basic_info = state.steps.get("basic_info")
    if basic_info and basic_info.data:
        accounting_standard = basic_info.data.get("accounting_standard", "enterprise")

    return await account_chart_service.load_standard_template(
        project_id, accounting_standard, db
    )


@router.post("/{project_id}/account-chart/preview")
async def preview_file(
    project_id: UUID,
    file: UploadFile = File(...),
    skip_rows: int | None = Query(default=None, description="跳过前N行，None=自动检测表头"),
    current_user: User = Depends(get_current_user),
):
    """Preview uploaded file: return first 20 rows per sheet + auto-matched column mapping.

    大文件优化：使用 smart_import_engine 的智能表头检测（支持双行合并表头），
    只读取前 20 行数据用于预览，不加载全部数据。
    """
    import io
    import openpyxl
    from app.services.smart_import_engine import (
        detect_header_rows, merge_header_rows, smart_match_column, _guess_data_type,
    )
    from app.services.account_chart_service import _COLUMN_MAP, _match_column, _guess_file_type

    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")

    content = await file.read()
    filename_lower = file.filename.lower()

    if filename_lower.endswith(".csv"):
        # CSV 用原有逻辑（小文件）—— 需要 seek 回起点因为 content 已经 read 过
        await file.seek(0)
        return await account_chart_service.preview_file(file, skip_rows=skip_rows)

    # Excel：小文件用完整模式（正确处理合并单元格），大文件用 read_only（快速）
    file_size_mb = len(content) / 1024 / 1024
    use_full_mode = file_size_mb < 10  # 10MB 以下用完整模式

    try:
        if use_full_mode:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法打开文件: {e}")

    sheets = []
    for ws in wb.worksheets:
        name_lower = ws.title.lower()
        if any(kw in name_lower for kw in ("说明", "目录", "封面")):
            continue

        max_col = ws.max_column or 50  # read_only 模式下可能没有 max_column

        # 第一步：读前 10 行原始数据（供用户确认表头位置）
        raw_rows = []
        if use_full_mode:
            for row_idx in range(1, min(11, (ws.max_row or 1) + 1)):
                row_vals = []
                for col in range(1, max_col + 1):
                    v = ws.cell(row_idx, col).value
                    row_vals.append(str(v).strip() if v is not None else "")
                raw_rows.append(row_vals)
        else:
            # read_only 模式用 iter_rows
            for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                raw_rows.append([str(c).strip() if c else "" for c in row])

        # 第二步：智能检测表头位置
        hs, hc = detect_header_rows(ws)
        headers = merge_header_rows(ws, hs, hc)

        # 第三步：列名映射
        column_mapping = {}
        for h in headers:
            mapped = smart_match_column(h)
            if not mapped:
                mapped = _match_column(h)
            if mapped:
                column_mapping[h] = mapped

        # 第四步：读前 20 行数据（表头之后）
        data_start = hs + hc
        num_cols = len(headers)
        rows = []
        if use_full_mode:
            for row_idx in range(data_start + 1, min(data_start + 21, (ws.max_row or 0) + 1)):
                row_dict = {}
                all_none = True
                for col in range(1, num_cols + 1):
                    v = ws.cell(row_idx, col).value
                    h = headers[col - 1] if col - 1 < len(headers) else f"col_{col}"
                    row_dict[h] = str(v).strip() if v is not None else ""
                    if v is not None:
                        all_none = False
                if not all_none:
                    rows.append(row_dict)
        else:
            # read_only 模式
            count = 0
            for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
                if i < data_start:
                    continue
                if count >= 20:
                    break
                padded = list(row_vals) + [None] * max(0, num_cols - len(row_vals))
                if all(c is None for c in padded[:num_cols]):
                    continue
                row_dict = {}
                for j in range(num_cols):
                    h = headers[j]
                    row_dict[h] = str(padded[j]).strip() if padded[j] is not None else ""
                rows.append(row_dict)
                count += 1

        total_rows = max(0, (ws.max_row or 0) - data_start) if use_full_mode else 0

        mapped_cols = set(column_mapping.values())
        file_type = _guess_data_type(mapped_cols) if mapped_cols else "unknown"

        sheets.append({
            "sheet_name": ws.title,
            "headers": headers,
            "rows": rows,
            "total_rows": total_rows,
            "column_mapping": column_mapping,
            "file_type_guess": file_type,
            "header_count": hc,
            "header_start": hs,
            "raw_first_rows": raw_rows,  # 前 10 行原始数据
        })

    wb.close()

    if not sheets:
        raise HTTPException(status_code=400, detail="文件中未解析到有效数据")

    return {"sheets": sheets, "active_sheet": 0}


@router.post(
    "/{project_id}/account-chart/import",
    response_model=AccountImportResult,
)
async def import_client_chart(
    project_id: UUID,
    files: list[UploadFile] = File(default=[]),
    file: UploadFile | None = File(default=None),
    column_mapping: str | None = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> AccountImportResult:
    """导入科目表 + 四表数据（支持多文件一次上传）。

    使用 smart_import_streaming 统一处理：
    - 多文件安全：只 soft-delete 一次，不会后续文件覆盖前面的
    - 内存可控：逐 sheet 解析并写入
    - 自动提取科目表、四表数据、诊断信息
    """
    import json
    from app.services.smart_import_engine import smart_import_streaming
    from app.services.import_queue_service import ImportQueueService

    # 兼容单文件 / 多文件两种前端传参方式
    all_files: list[UploadFile] = list(files) if files else []
    if file is not None:
        all_files.append(file)
    if not all_files:
        raise HTTPException(status_code=400, detail="未提供文件")

    parsed_mapping = None
    if column_mapping:
        try:
            parsed_mapping = json.loads(column_mapping)
        except (json.JSONDecodeError, TypeError):
            pass

    # 读取所有文件内容
    file_contents: list[tuple[str, bytes]] = []
    for f in all_files:
        ct = await f.read()
        file_contents.append((f.filename or "upload.xlsx", ct))

    def _on_progress(pct: int, msg: str):
        ImportQueueService.update_progress(project_id, min(pct, 99), msg)

    result = await smart_import_streaming(
        project_id=project_id,
        file_contents=file_contents,
        db=db,
        custom_mapping=parsed_mapping,
        progress_callback=_on_progress,
    )

    return AccountImportResult(
        total_imported=result["total_accounts"],
        by_category=result["by_category"],
        errors=result["errors"],
        data_sheets_imported=result["data_sheets_imported"],
        sheet_diagnostics=result["sheet_diagnostics"],
        year=result["year"],
    )


@router.post("/{project_id}/account-chart/import-async")
async def import_async(
    project_id: UUID,
    files: list[UploadFile] = File(default=[]),
    file: UploadFile | None = File(default=None),
    column_mapping: str | None = Form(None),
    current_user: User = Depends(get_current_user),
):
    """异步导入（多文件支持）：立即返回，后台处理，前端轮询进度。

    适合大文件（>10MB）或多文件场景，避免 HTTP 超时。
    """
    import asyncio
    import json
    from app.services.import_queue_service import ImportQueueService

    # 获取导入锁
    ok, msg = ImportQueueService.acquire_lock(project_id, str(current_user.id))
    if not ok:
        raise HTTPException(status_code=409, detail=msg)

    # 兼容单文件 / 多文件
    all_files: list[UploadFile] = list(files) if files else []
    if file is not None:
        all_files.append(file)
    if not all_files:
        ImportQueueService.release_lock(project_id)
        raise HTTPException(status_code=400, detail="未提供文件")

    parsed_mapping = None
    if column_mapping:
        try:
            parsed_mapping = json.loads(column_mapping)
        except (json.JSONDecodeError, TypeError):
            pass

    # 预读所有文件到内存（multipart 上传完成后释放连接）
    file_contents: list[tuple[str, bytes]] = []
    for f in all_files:
        ct = await f.read()
        file_contents.append((f.filename or "upload.xlsx", ct))

    # 后台任务
    async def _do_import():
        from app.core.database import async_session
        from app.services.smart_import_engine import smart_import_streaming

        try:
            ImportQueueService.update_progress(
                project_id, 2,
                f"开始导入 {len(file_contents)} 个文件…",
            )

            def _on_progress(pct: int, msg: str):
                ImportQueueService.update_progress(project_id, pct, msg)

            async with async_session() as db:
                result = await smart_import_streaming(
                    project_id=project_id,
                    file_contents=file_contents,
                    db=db,
                    custom_mapping=parsed_mapping,
                    progress_callback=_on_progress,
                )

                result_payload = {
                    "total_imported": result["total_accounts"],
                    "by_category": result["by_category"],
                    "errors": result["errors"],
                    "data_sheets_imported": result["data_sheets_imported"],
                    "sheet_diagnostics": result["sheet_diagnostics"],
                    "year": result["year"],
                }
                ImportQueueService.update_progress(
                    project_id, 100,
                    f"导入完成: {result['data_sheets_imported']}",
                    result=result_payload,
                )

        except Exception as e:
            import traceback
            import logging
            logging.getLogger(__name__).error(
                "异步导入失败: %s\n%s", e, traceback.format_exc(),
            )
            ImportQueueService.update_progress(
                project_id, -1, f"导入失败: {e}",
                result={
                    "total_imported": 0,
                    "by_category": {},
                    "errors": [f"导入失败: {e}"],
                    "data_sheets_imported": {},
                    "sheet_diagnostics": [],
                    "year": None,
                },
            )
        finally:
            await asyncio.sleep(3)
            ImportQueueService.release_lock(project_id)

    asyncio.create_task(_do_import())

    return {
        "status": "accepted",
        "message": f"导入任务已提交（{len(file_contents)} 个文件），请轮询进度",
        "project_id": str(project_id),
    }


@router.get(
    "/{project_id}/account-chart/client",
    response_model=dict[str, list[AccountTreeNode]],
)
async def get_client_chart(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, list[AccountTreeNode]]:
    """获取客户科目表（树形结构按类别分组）。

    Validates: Requirements 2.6
    """
    return await account_chart_service.get_client_chart_tree(project_id, db)


# ── 列映射保存/加载（持久化到 wizard_state） ──

from pydantic import BaseModel as _BaseModel


class SaveColumnMappingRequest(_BaseModel):
    """保存列映射"""
    file_type: str  # account_chart / ledger / balance / aux_balance
    sheet_name: str | None = None
    mapping: dict[str, str]  # {原列名: 标准字段名}


class ReferenceMatchRequest(_BaseModel):
    """参照匹配请求"""
    source_project_id: str
    source_year: int | None = None


@router.post("/{project_id}/column-mappings")
async def save_column_mapping(
    project_id: UUID,
    body: SaveColumnMappingRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """保存用户确认的列映射（持久化到项目 wizard_state）"""
    from sqlalchemy import select
    from app.models.core import Project

    result = await db.execute(
        select(Project).where(Project.id == project_id, Project.is_deleted == False)  # noqa: E712
    )
    project = result.scalar_one_or_none()
    if not project:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="项目不存在")

    # 保存到 wizard_state.column_mappings
    state = project.wizard_state or {}
    if "column_mappings" not in state:
        state["column_mappings"] = {}

    key = f"{body.file_type}"
    if body.sheet_name:
        key = f"{body.file_type}:{body.sheet_name}"
    state["column_mappings"][key] = body.mapping
    project.wizard_state = state

    await db.commit()
    return {"saved": True, "key": key, "field_count": len(body.mapping)}


@router.get("/{project_id}/column-mappings")
async def get_column_mappings(
    project_id: UUID,
    file_type: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取已保存的列映射"""
    from sqlalchemy import select
    from app.models.core import Project

    result = await db.execute(
        select(Project).where(Project.id == project_id, Project.is_deleted == False)  # noqa: E712
    )
    project = result.scalar_one_or_none()
    if not project:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="项目不存在")

    state = project.wizard_state or {}
    mappings = state.get("column_mappings", {})

    if file_type:
        # 过滤指定文件类型
        filtered = {k: v for k, v in mappings.items() if k.startswith(file_type)}
        return filtered
    return mappings


@router.get("/{project_id}/column-mappings/reference-projects")
async def get_reference_projects(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取可参照的项目列表（有已保存列映射的其他项目）"""
    from sqlalchemy import select, func, cast, String
    from app.models.core import Project

    # 查找所有有 column_mappings 的项目（排除当前项目）
    result = await db.execute(
        select(Project.id, Project.name, Project.client_name, Project.wizard_state)
        .where(
            Project.id != project_id,
            Project.is_deleted == False,  # noqa: E712
            Project.wizard_state.isnot(None),
        )
        .order_by(Project.created_at.desc())
        .limit(20)
    )
    projects = []
    for row in result.all():
        ws = row.wizard_state or {}
        mappings = ws.get("column_mappings", {})
        if mappings:
            projects.append({
                "id": str(row.id),
                "name": row.name,
                "client_name": row.client_name,
                "mapping_count": len(mappings),
                "file_types": list(set(k.split(":")[0] for k in mappings.keys())),
            })
    return projects


@router.post("/{project_id}/column-mappings/reference-copy")
async def reference_copy_mappings(
    project_id: UUID,
    body: ReferenceMatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从其他项目参照复制列映射"""
    from sqlalchemy import select
    from app.models.core import Project

    # 获取源项目的映射
    source = await db.execute(
        select(Project).where(Project.id == UUID(body.source_project_id))
    )
    source_project = source.scalar_one_or_none()
    if not source_project or not source_project.wizard_state:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="源项目不存在或无映射数据")

    source_mappings = (source_project.wizard_state or {}).get("column_mappings", {})
    if not source_mappings:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="源项目无已保存的列映射")

    # 复制到目标项目
    target = await db.execute(
        select(Project).where(Project.id == project_id, Project.is_deleted == False)  # noqa: E712
    )
    target_project = target.scalar_one_or_none()
    if not target_project:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="目标项目不存在")

    state = target_project.wizard_state or {}
    if "column_mappings" not in state:
        state["column_mappings"] = {}
    state["column_mappings"].update(source_mappings)
    target_project.wizard_state = state

    await db.commit()
    return {
        "copied": True,
        "source_project": body.source_project_id,
        "mapping_count": len(source_mappings),
    }


class AccountUpdateItem(_BaseModel):
    """单个科目更新"""
    account_code: str
    account_name: str | None = None
    direction: str | None = None


class BatchUpdateRequest(_BaseModel):
    """批量更新科目"""
    updates: list[AccountUpdateItem]


@router.put("/{project_id}/account-chart/batch-update")
async def batch_update_accounts(
    project_id: UUID,
    body: BatchUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """批量更新客户科目（名称、借贷方向）"""
    from app.models.audit_platform_models import AccountChart, AccountSource, AccountDirection

    updated = 0
    for item in body.updates:
        result = await db.execute(
            select(AccountChart).where(
                AccountChart.project_id == project_id,
                AccountChart.account_code == item.account_code,
                AccountChart.source == AccountSource.client,
                AccountChart.is_deleted == False,  # noqa: E712
            )
        )
        row = result.scalar_one_or_none()
        if not row:
            continue

        if item.account_name is not None:
            row.account_name = item.account_name
        if item.direction is not None:
            try:
                row.direction = AccountDirection(item.direction)
            except ValueError:
                pass
        updated += 1

    await db.commit()
    return {"updated": updated, "total": len(body.updates)}
