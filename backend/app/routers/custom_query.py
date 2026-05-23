"""自定义查询 API — 支持多维度跨模块数据查询

支持查询维度：
  - report: 报表数据（report_config）
  - trial_balance: 试算表数据（trial_balance_entries）
  - disclosure: 附注数据（consol_note_data）
  - adjustment: 调整分录（adjustments）
  - worksheet: 工作底稿数据（consol_worksheet_data）
  - workpaper: 底稿列表（working_paper）— Sprint 6 新增
  - account_balance: 科目余额（tb_balance）— Sprint 6 新增
  - ledger_entries: 序时账（tb_ledger）— Sprint 6 新增
  - report_lines: 报表行次（report_config / financial_report）— Sprint 6 新增
  - workhours: 工时记录（work_hours）— Sprint 6 新增

支持过滤：
  - project_id: 项目
  - year: 年度
  - company_code: 单位
  - report_type: 报表类型
  - account_name: 科目名
  - section_id: 附注章节

API:
  POST   /api/custom-query/execute              — 执行查询
  GET    /api/custom-query/indicators           — 获取可查询指标库（树形）
  GET    /api/custom-query/templates            — 列出查询模板（私有 + 全局）
  POST   /api/custom-query/templates            — 创建查询模板
  GET    /api/custom-query/templates/{id}       — 获取模板详情
  PUT    /api/custom-query/templates/{id}       — 更新模板
  DELETE /api/custom-query/templates/{id}       — 删除模板（仅创建者或 admin）
"""

import asyncio
import json
import logging
import uuid
from datetime import datetime
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from pydantic import BaseModel, Field
from sqlalchemy import or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.core import Project, User
from app.models.custom_query_models import CustomQueryTemplate

router = APIRouter(prefix="/api/custom-query", tags=["custom-query"])

logger = logging.getLogger(__name__)


class QueryRequest(BaseModel):
    project_id: str
    year: int
    source: str  # report | trial_balance | disclosure | adjustment | worksheet | workpaper | account_balance | ledger_entries | report_lines | workhours
    filters: dict = {}  # report_type, account_name, section_id, company_code, etc.
    columns: list[str] = []  # 要查询的列（空=全部）
    limit: int = 500
    offset: int = 0


@router.get("/indicators")
async def get_indicators(
    project_id: str | None = Query(default=None, description="项目 ID — 传入后报表/附注树会按项目模板类型动态生成"),
    response: Response = None,  # type: ignore[assignment]
    db: AsyncSession = Depends(get_db),
):
    """获取可查询指标库（树形结构）

    响应头：
      - X-Indicators-Schema-Version：树结构版本号，前端用此值做缓存键，schema 升版自动失效
      - Cache-Control：private, max-age=0, must-revalidate（每次都校验，但允许 ETag 复用）

    传入 project_id 时：
      - 报表大类下展示项目模板类型（国企/上市）对应的具体报表
      - 附注大类下展示项目实际配置的章节（按 parent_section 分组成子大类）
    未传 project_id 时：报表/附注用通用降级树（仅显示报表类型，无国企/上市区分）
    """
    template_type = await _resolve_project_template_type(db, project_id)
    standard_label = _STANDARD_LABEL.get(template_type, "通用")

    # ─── 报表树：项目模板类型决定大类标签 + 报表清单 ───────────────
    # 6 张报表：BS / IS / CFS / CFS-补充资料 / 权益变动 / 减值准备表
    # 减值准备表 (impairment_provision) 仅 soe 模板有数据，listed 自动隐藏避免误点
    report_children = [
        {"key": "report_balance_sheet", "label": f"资产负债表（{standard_label}）", "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "standard": template_type},
        {"key": "report_income_statement", "label": f"利润表（{standard_label}）", "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "standard": template_type},
        {"key": "report_cash_flow_statement", "label": f"现金流量表（{standard_label}）", "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "standard": template_type},
        {"key": "report_cash_flow_supplement", "label": f"现金流量表补充资料（{standard_label}）", "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "standard": template_type},
        {"key": "report_equity_statement", "label": f"所有者权益变动表（{standard_label}）", "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "standard": template_type},
    ]
    if template_type == "soe":
        report_children.append(
            {"key": "report_impairment_provision", "label": f"减值准备表（{standard_label}）", "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "standard": template_type}
        )

    # ─── 附注树：按 parent_section 分组成大类→明细两层 ─────────────
    disclosure_children = await _build_disclosure_tree(template_type)

    # ─── 合并范围单位树（仅合并项目可见）────────────────────────
    consol_units_node = await _build_consol_units_tree(db, project_id)

    base_tree: list[dict] = []
    base_tree.append({
        "key": "report",
        "label": f"📊 报表（{standard_label}）" if project_id else "📊 报表",
        "icon": "📊",
        "children": report_children,
    })
    base_tree.append({
        "key": "trial_balance", "label": "📋 试算表", "icon": "📋",
        "children": [
            {"key": "tb_detail", "label": "科目明细", "columns": ["account_code", "account_name", "opening_balance", "closing_balance", "debit_amount", "credit_amount"]},
            {"key": "tb_summary", "label": "试算平衡表", "columns": ["row_code", "row_name", "unadjusted", "aje_dr", "aje_cr", "rcl_dr", "rcl_cr", "audited"]},
        ],
    })
    base_tree.append({
        "key": "disclosure",
        "label": f"📝 附注（{standard_label}）" if project_id else "📝 附注",
        "icon": "📝",
        "children": disclosure_children,
    })
    # 合并范围节点：仅合并项目（report_scope=consolidated 且至少 1 家纳入单位）展示
    if consol_units_node:
        base_tree.append(consol_units_node)
    base_tree.append({
        "key": "adjustment", "label": "📐 调整分录", "icon": "📐",
        "children": [
            {"key": "adj_aje", "label": "审计调整分录(AJE)", "columns": ["entry_number", "account_name", "debit_amount", "credit_amount", "description"]},
            {"key": "adj_rcl", "label": "重分类调整(RCL)", "columns": ["entry_number", "account_name", "debit_amount", "credit_amount", "description"]},
        ],
    })
    base_tree.append({
        "key": "worksheet", "label": "📑 工作底稿", "icon": "📑",
        "children": [
            {"key": "ws_info", "label": "基本信息表", "columns": ["company_name", "company_code", "holding_type", "non_common_ratio"]},
            {"key": "ws_elimination", "label": "抵消分录", "columns": ["direction", "subject", "amount", "desc"]},
            {"key": "ws_consol_tb", "label": "合并试算平衡表", "columns": ["row_code", "row_name", "summary", "equity_dr", "equity_cr", "audited"]},
        ],
    })
    base_tree.append({
        "key": "workpaper", "label": "📄 底稿列表", "icon": "📄",
        "children": await _build_workpaper_tree(db, project_id),
    })
    base_tree.append({
        "key": "account_balance", "label": "💰 科目余额", "icon": "💰",
        "children": [
            {"key": "account_balance", "label": "科目余额表", "columns": ["account_code", "account_name", "opening_balance", "closing_balance", "debit_amount", "credit_amount"]},
        ],
    })
    base_tree.append({
        "key": "ledger_entries", "label": "📜 序时账", "icon": "📜",
        "children": [
            {"key": "ledger_entries", "label": "序时账明细", "columns": ["voucher_date", "voucher_no", "account_code", "account_name", "debit_amount", "credit_amount", "summary"]},
        ],
    })
    base_tree.append({
        "key": "report_lines", "label": "📈 报表行次", "icon": "📈",
        "children": [
            {"key": "report_lines", "label": "报表行次配置", "columns": ["row_code", "row_name", "report_type", "applicable_standard", "indent_level", "is_total_row", "formula"]},
        ],
    })
    base_tree.append({
        "key": "workhours", "label": "⏱️ 工时记录", "icon": "⏱️",
        "children": [
            {"key": "workhours", "label": "工时记录", "columns": ["work_date", "hours", "description", "status", "staff_id"]},
        ],
    })
    if response is not None:
        response.headers["X-Indicators-Schema-Version"] = str(_INDICATORS_SCHEMA_VERSION)
        response.headers["Cache-Control"] = "private, max-age=0, must-revalidate"
    return base_tree


# ─── 辅助函数：项目模板类型解析 + 附注树构建 ───────────────────────────────

_STANDARD_LABEL = {"soe": "国企", "listed": "上市"}
# indicators 树结构 schema 版本：树字段/叶子 key 形态有变更时升 1，前端用此值自动失效旧缓存
# 历史演进：v1 扁平 → v2 项目级 → v3 6 报表 → v4 合并模块 → v5 底稿 3 层 → v6 sheet 全集 → v7 disabled → v8 sheet 辅助灰度 + ancestorKeys
_INDICATORS_SCHEMA_VERSION = 8
_DISCLOSURE_CACHE: dict[str, list[dict]] = {}
_WP_MAPPING_CACHE: list[dict] | None = None

# 14 审计循环代号 → 中文名（与 memory.md 沉淀对齐：A=报表/调整 / B=控制了解 / C=控制测试 / D=销售收入 / E=货币资金 / F=采购存货 / G=投资 / H=固定资产+在建工程+使用权资产+租赁负债 / I=无形资产+商誉+开发支出 / J=职工薪酬+股份支付 / K=管理 / L=筹资 / M=股东权益 / N=税费 / S=专项程序）
_CYCLE_NAMES: dict[str, str] = {
    "A": "报表/调整",
    "B": "控制了解",
    "C": "控制测试",
    "D": "销售收入",
    "E": "货币资金",
    "F": "采购存货",
    "G": "投资",
    "H": "固定资产+在建+使用权",
    "I": "无形资产+商誉",
    "J": "职工薪酬+股份支付",
    "K": "管理",
    "L": "筹资",
    "M": "股东权益",
    "N": "税费",
    "S": "专项程序",
}


async def _resolve_project_template_type(db: AsyncSession, project_id: str | None) -> str:
    """从项目读 template_type，未配置或未传 project_id 则降级为 'soe'"""
    if not project_id:
        return "soe"
    try:
        proj_uuid = uuid.UUID(project_id)
    except (TypeError, ValueError):
        return "soe"
    proj = await db.get(Project, proj_uuid)
    if proj and proj.template_type in _STANDARD_LABEL:
        return proj.template_type
    return "soe"


async def _resolve_project_report_scope(db: AsyncSession, project_id: str | None) -> str:
    """从项目读 report_scope（standalone / consolidated），默认 standalone"""
    if not project_id:
        return "standalone"
    try:
        proj_uuid = uuid.UUID(project_id)
    except (TypeError, ValueError):
        return "standalone"
    proj = await db.get(Project, proj_uuid)
    if proj and proj.report_scope in ("standalone", "consolidated"):
        return proj.report_scope
    return "standalone"


async def _build_consol_units_tree(db: AsyncSession, project_id: str | None) -> dict | None:
    """合并项目下生成「合并范围」顶层节点，children 是每家纳入合并的单位。

    判定 = report_scope=consolidated **或** consol_scope 表已有数据（向导未走完时兜底）
    每家单位下挂 4 个明细：科目余额 / 序时账 / 试算表（个体） / 调整分录（按 company_code）
    返回 None 表示当前项目非合并 / 无单位 / 无 project_id（不渲染该节点）
    """
    if not project_id:
        return None
    try:
        proj_uuid = uuid.UUID(project_id)
    except (TypeError, ValueError):
        return None
    # 取最新年度的纳入合并单位（按 ownership_ratio 排序）
    sql = text("""
        SELECT DISTINCT ON (company_code) company_code, company_name, company_type, ownership_ratio
        FROM consol_scope
        WHERE project_id = :pid AND is_included = true AND is_deleted = false
        ORDER BY company_code, year DESC
    """)
    try:
        rows = (await db.execute(sql, {"pid": str(proj_uuid)})).fetchall()
    except Exception:
        return None
    if not rows:
        # 兜底：若 consol_scope 无数据但 project.report_scope=consolidated，仍提示「待录入」节点
        scope = await _resolve_project_report_scope(db, project_id)
        if scope == "consolidated":
            return {
                "key": "consol_units",
                "label": "🏢 合并范围（待录入）",
                "icon": "🏢",
                "children": [
                    {"key": "consol_units_empty", "label": "⚠ 尚未配置纳入合并的单位"},
                ],
            }
        return None
    units: list[dict] = []
    for r in rows:
        cc = r[0]
        cname = r[1] or cc
        ctype = r[2] or ""
        ratio = float(r[3]) if r[3] is not None else None
        ratio_label = f" {ratio:.2f}%" if ratio is not None else ""
        type_label = {"parent": "母", "subsidiary": "子", "associate": "联", "joint_venture": "合营"}.get(str(ctype), "")
        unit_label = f"[{type_label}] {cname}{ratio_label}".strip()
        units.append({
            "key": f"consol_unit_group_{cc}",
            "label": unit_label,
            "company_code": cc,
            "children": [
                {"key": f"consol_unit:{cc}:account_balance", "label": "科目余额", "columns": ["account_code", "account_name", "opening_balance", "closing_balance", "debit_amount", "credit_amount"], "company_code": cc},
                {"key": f"consol_unit:{cc}:ledger_entries", "label": "序时账明细", "columns": ["voucher_date", "voucher_no", "account_code", "account_name", "debit_amount", "credit_amount", "summary"], "company_code": cc},
                {"key": f"consol_unit:{cc}:tb_detail", "label": "试算表（个体）", "columns": ["account_code", "account_name", "opening_balance", "closing_balance", "debit_amount", "credit_amount"], "company_code": cc},
                {"key": f"consol_unit:{cc}:adjustment", "label": "调整分录", "columns": ["entry_number", "account_name", "debit_amount", "credit_amount", "description"], "company_code": cc},
            ],
        })
    return {
        "key": "consol_units",
        "label": f"🏢 合并范围（{len(units)}家）",
        "icon": "🏢",
        "children": units,
    }


async def _build_disclosure_tree(template_type: str) -> list[dict]:
    """读 consol_note_sections_{template_type}.json，按 parent_section 分组成两层树。

    返回结构：
      [
        { key: 'note_group_货币资金', label: '货币资金', children: [
            { key: 'disclosure_note:五-1-1', label: '五-1-1 货币资金', section_id: '五-1-1' },
            { key: 'disclosure_note:五-1-2', label: '五-1-2 受限制货币资金明细', section_id: '五-1-2' },
        ] },
        ...
      ]
    """
    sections = _load_disclosure_sections(template_type)
    if not sections:
        # 降级：未读到 JSON 时返回单一通用入口
        return [{"key": "disclosure_note", "label": "附注章节数据（通用）", "columns": ["section_id", "headers", "rows"]}]

    # 按 parent_section 分组（按 parent_seq 排序）
    groups: dict[str, dict] = {}
    for sec in sections:
        parent = sec.get("parent_section") or sec.get("title") or "其他"
        parent_seq = sec.get("parent_seq", 999)
        if parent not in groups:
            groups[parent] = {
                "key": f"note_group_{parent}",
                "label": parent,
                "_seq": parent_seq,
                "children": [],
            }
        section_id = sec.get("section_id", "")
        title = sec.get("title", "")
        groups[parent]["children"].append({
            "key": f"disclosure_note:{section_id}",
            "label": f"{section_id} {title}".strip(),
            "section_id": section_id,
            "columns": ["section_id", "headers", "rows"],
            "_seq": sec.get("seq", 999),
        })

    # 按 parent_seq 排序大类，按 seq 排序明细
    sorted_groups = sorted(groups.values(), key=lambda g: g["_seq"])
    for g in sorted_groups:
        g["children"].sort(key=lambda c: c["_seq"])
        for c in g["children"]:
            c.pop("_seq", None)
        g.pop("_seq", None)
    return sorted_groups


def _load_disclosure_sections(template_type: str) -> list[dict]:
    """读 backend/data/consol_note_sections_{template_type}.json（带模块级缓存）"""
    if template_type in _DISCLOSURE_CACHE:
        return _DISCLOSURE_CACHE[template_type]
    data_path = Path(__file__).resolve().parent.parent.parent / "data" / f"consol_note_sections_{template_type}.json"
    if not data_path.exists():
        _DISCLOSURE_CACHE[template_type] = []
        return []
    try:
        with data_path.open("r", encoding="utf-8") as f:
            sections = json.load(f)
        _DISCLOSURE_CACHE[template_type] = sections if isinstance(sections, list) else []
    except Exception:
        _DISCLOSURE_CACHE[template_type] = []
    return _DISCLOSURE_CACHE[template_type]


def _load_wp_mapping() -> list[dict]:
    """读 backend/data/wp_account_mapping.json（带模块级缓存）"""
    global _WP_MAPPING_CACHE
    if _WP_MAPPING_CACHE is not None:
        return _WP_MAPPING_CACHE
    data_path = Path(__file__).resolve().parent.parent.parent / "data" / "wp_account_mapping.json"
    if not data_path.exists():
        _WP_MAPPING_CACHE = []
        return []
    try:
        with data_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            _WP_MAPPING_CACHE = data.get("mappings", []) or []
        elif isinstance(data, list):
            _WP_MAPPING_CACHE = data
        else:
            _WP_MAPPING_CACHE = []
    except Exception:
        _WP_MAPPING_CACHE = []
    return _WP_MAPPING_CACHE


_STEP_SHEET_CACHE: dict[str, dict] | None = None
# wp_code → 真实模板 xlsx 的 sheet 列表（启动后惰性扫描，缺失则降级为 step_sheet_mapping）
_REAL_SHEETS_CACHE: dict[str, list[str]] = {}

# sheet 灰度判定：辅助 sheet 关键词（与 chain_orchestrator scenarioFilter 对齐）
# 业务规则：模板里这些 sheet 是给审计师参考的辅助资料，不参与正式数据填报，查询无意义
_AUX_SHEET_KEYWORDS = ("(修订前)", "（修订前）", "(示例)", "（示例）", "(提示)", "（提示）", "-修订前", "GT_Custom")


def _is_aux_sheet(sheet_name: str) -> bool:
    """判断 sheet 名是否为辅助 sheet（应灰显）"""
    if not sheet_name:
        return False
    return any(kw in sheet_name for kw in _AUX_SHEET_KEYWORDS)


def _load_step_sheet_mapping() -> dict[str, dict]:
    """读 backend/data/step_sheet_mapping.json — 全集 wp_code → {wp_name, available_sheets[]}

    覆盖率 100% (179 底稿 / 1040 sheet)，是底稿全 sheet 列表的权威源。
    首次加载时与 wp_account_mapping.json 主底稿做漂移检测，差异写 logger.warning（提醒维护对齐）。
    """
    global _STEP_SHEET_CACHE
    if _STEP_SHEET_CACHE is not None:
        return _STEP_SHEET_CACHE
    data_path = Path(__file__).resolve().parent.parent.parent / "data" / "step_sheet_mapping.json"
    if not data_path.exists():
        _STEP_SHEET_CACHE = {}
        return {}
    try:
        with data_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        _STEP_SHEET_CACHE = data.get("mappings", {}) if isinstance(data, dict) else {}
    except Exception:
        _STEP_SHEET_CACHE = {}

    # 双源漂移检测（首次加载时执行一次）
    try:
        wp_codes_step = {c for c in _STEP_SHEET_CACHE.keys() if "-" not in c}
        wp_codes_acc = {m.get("wp_code") for m in _load_wp_mapping() if m.get("wp_code") and "-" not in m.get("wp_code", "")}
        only_in_step = wp_codes_step - wp_codes_acc
        only_in_acc = wp_codes_acc - wp_codes_step
        if only_in_step or only_in_acc:
            logger.warning(
                "wp mapping double-source drift: step_sheet_mapping has %d codes not in wp_account_mapping (%s...), "
                "wp_account_mapping has %d codes not in step_sheet_mapping (%s...). "
                "Maintain both files in sync to avoid tree inconsistency.",
                len(only_in_step), sorted(list(only_in_step))[:5],
                len(only_in_acc), sorted(list(only_in_acc))[:5],
            )
    except Exception as e:
        logger.debug("drift detection skipped: %s", e)
    return _STEP_SHEET_CACHE


async def _build_workpaper_tree(db: AsyncSession, project_id: str | None) -> list[dict]:
    """底稿树：循环 → 主底稿（科目名）→ sheet 程序明细。

    - 树永远基于全集（wp_account_mapping + step_sheet_mapping）
    - 传入 project_id 时，从 wp_index 取该项目实际有的 wp_code 集合（裁剪后），
      不在集合里的节点标 disabled=true 灰显不可选；
      项目实际有但全集没的（罕见）补充到全集末尾
    """
    mapping = _load_wp_mapping()
    sheet_map = _load_step_sheet_mapping()
    map_by_code: dict[str, dict] = {m.get("wp_code", ""): m for m in mapping if m.get("wp_code")}

    # 项目级实际有的 wp_code 集合（用于判定 disabled）
    available_codes: set[str] | None = None
    if project_id:
        try:
            proj_uuid = uuid.UUID(project_id)
            sql = text("""
                SELECT wp_code FROM wp_index
                WHERE project_id = :pid AND is_deleted = false
            """)
            rows = (await db.execute(sql, {"pid": str(proj_uuid)})).fetchall()
            available_codes = {r[0] for r in rows if r[0]}
        except (TypeError, ValueError):
            available_codes = None

    # 主底稿全集：mapping JSON 主底稿 + step_sheet_mapping 主底稿（去重）
    primaries: list[dict] = []
    seen: set[str] = set()
    for m in mapping:
        code = m.get("wp_code", "")
        if not code or "-" in code or code in seen:
            continue
        seen.add(code)
        primaries.append({
            "wp_code": code,
            "wp_name": m.get("wp_name", ""),
            "cycle": m.get("cycle") or (code[:1] if code else ""),
            "account_name": m.get("account_name") or "",
        })
    for code, info in sheet_map.items():
        if "-" in code or code in seen:
            continue
        seen.add(code)
        primaries.append({
            "wp_code": code,
            "wp_name": info.get("wp_name", ""),
            "cycle": code[:1] if code else "",
            "account_name": "",
        })

    # 按 cycle 分组
    cycle_groups: dict[str, list[dict]] = {}
    for p in primaries:
        cy = (p["cycle"] or "").upper() or "其他"
        cycle_groups.setdefault(cy, []).append(p)

    def _cycle_sort_key(cy: str) -> tuple:
        order = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "S"]
        return (order.index(cy) if cy in order else 99, cy)

    def _is_disabled(code: str) -> bool:
        """传 project_id 时，wp_code 不在 available_codes 即灰显"""
        if available_codes is None:
            return False
        return code not in available_codes

    result: list[dict] = []
    for cy in sorted(cycle_groups.keys(), key=_cycle_sort_key):
        primary_list = sorted(cycle_groups[cy], key=lambda x: x["wp_code"])
        primary_nodes: list[dict] = []
        cycle_active_count = 0
        for p in primary_list:
            code = p["wp_code"]
            acc = p.get("account_name") or ""
            wpname = p.get("wp_name") or ""
            primary_label = f"{code} {acc or wpname}".strip()
            primary_disabled = _is_disabled(code)
            if not primary_disabled:
                cycle_active_count += 1
            sheets = sheet_map.get(code, {}).get("available_sheets") or []
            children_nodes: list[dict] = []
            children_nodes.append({
                "key": f"workpaper:{code}",
                "label": f"{code}（汇总）",
                "wp_code": code,
                "disabled": primary_disabled,
                "columns": ["wp_code", "wp_name", "audit_cycle", "status", "review_status"],
            })
            for s_name in sheets:
                # sheet 灰度规则:
                # 1) 主底稿在项目内被裁剪 → 该 sheet 也灰
                # 2) sheet 名带 "(修订前)" / "(示例)" / "(提示)" / "GT_Custom" 等模板辅助标识 → 灰显（与 chain_orchestrator scenarioFilter 对齐）
                sheet_aux = _is_aux_sheet(s_name)
                children_nodes.append({
                    "key": f"workpaper:{code}|{s_name}",
                    "label": s_name,
                    "wp_code": code,
                    "sheet_name": s_name,
                    "disabled": primary_disabled or sheet_aux,
                    "disabled_reason": "辅助 sheet" if sheet_aux and not primary_disabled else None,
                    "columns": ["wp_code", "wp_name", "audit_cycle", "status", "review_status"],
                })
            primary_nodes.append({
                "key": f"wp_primary_{code}",
                "label": f"{primary_label}（{len(sheets)}）",
                "disabled": primary_disabled,
                "children": children_nodes,
            })
        # 循环节点 label 显示 "活跃数/总数"（项目模式）或 "总数"（全集模式）
        if available_codes is not None:
            cycle_label = f"{cy} {_CYCLE_NAMES.get(cy, '')}".strip() + f"（{cycle_active_count}/{len(primary_list)}）"
        else:
            cycle_label = f"{cy} {_CYCLE_NAMES.get(cy, '')}".strip() + f"（{len(primary_list)}）"
        # 整个循环全部灰则循环本身也 disabled
        cycle_all_disabled = available_codes is not None and cycle_active_count == 0
        result.append({
            "key": f"wp_cycle_{cy}",
            "label": cycle_label,
            "disabled": cycle_all_disabled,
            "children": primary_nodes,
        })
    return result


def _sync_parse_sheet_preview(content: bytes, sheet_name: str | None) -> dict:
    """同步函数：openpyxl 解析单 sheet 返回 cellData（供 run_in_executor 调用避免阻塞 event loop）"""
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    target = sheet_name if sheet_name and sheet_name in wb.sheetnames else (wb.sheetnames[0] if wb.sheetnames else None)
    if not target:
        wb.close()
        return {"error": "工作簿无 sheet"}
    ws = wb[target]
    rows = min(ws.max_row or 0, 100)
    cols = min(ws.max_column or 0, 30)
    cells: dict[str, dict] = {}
    for row_iter in ws.iter_rows(min_row=1, max_row=rows, max_col=cols):
        for cell in row_iter:
            if cell.value is None:
                continue
            key = f"{cell.row - 1},{cell.column - 1}"
            v = cell.value
            cells[key] = {"v": v if isinstance(v, (int, float)) else str(v)}
    merges = []
    for mr in (ws.merged_cells.ranges if hasattr(ws, "merged_cells") else []):
        merges.append({"r1": mr.min_row - 1, "c1": mr.min_col - 1, "r2": mr.max_row - 1, "c2": mr.max_col - 1})
    all_sheets = list(wb.sheetnames)
    wb.close()
    return {
        "sheet_name": target,
        "all_sheets": all_sheets,
        "rows": rows,
        "cols": cols,
        "cells": cells,
        "merges": merges,
    }


def _sync_extract_cell_range(content: bytes, sheet_name: str | None, r1: int, c1: int, r2: int, c2: int, max_cells: int = 500) -> dict:
    """同步函数：按 cell_range 提取单元格值清单（供 run_in_executor 调用）"""
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    target = sheet_name if sheet_name and sheet_name in wb.sheetnames else (wb.sheetnames[0] if wb.sheetnames else None)
    if not target:
        wb.close()
        return {"error": "工作簿无 sheet"}
    ws = wb[target]
    out_rows: list[dict] = []
    idx = 0
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            idx += 1
            if idx > max_cells:
                break
            cell = ws.cell(row=r, column=c)
            col_letters = ""
            cc = c
            while cc > 0:
                cc, rem = divmod(cc - 1, 26)
                col_letters = chr(65 + rem) + col_letters
            cell_ref = f"{col_letters}{r}"
            v = cell.value
            if v is None:
                val_str: str | int | float = ""
            elif isinstance(v, (int, float)):
                val_str = v
            else:
                val_str = str(v)
            out_rows.append({"cell_ref": cell_ref, "value": val_str})
        if idx >= max_cells:
            break
    wb.close()
    return {"rows": out_rows, "sheet_name": target}


@router.get("/wp-id-by-code")
async def wp_id_by_code(
    project_id: str = Query(..., description="项目 ID"),
    wp_code: str = Query(..., description="底稿编码（如 D2）"),
    db: AsyncSession = Depends(get_db),
):
    """wp_code → working_paper.id 映射（供前端穿透跳转 WorkpaperEditor 用）"""
    try:
        proj_uuid = uuid.UUID(project_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="invalid project_id")
    row = (await db.execute(text("""
        SELECT w.id FROM working_paper w
        JOIN wp_index i ON w.wp_index_id = i.id
        WHERE i.project_id = :pid AND i.wp_code = :wpc AND w.is_deleted = false
        LIMIT 1
    """), {"pid": str(proj_uuid), "wpc": wp_code})).first()
    if not row:
        raise HTTPException(status_code=404, detail=f"底稿 {wp_code} 在当前项目不存在")
    return {"wp_id": str(row[0]), "wp_code": wp_code, "project_id": project_id}


@router.get("/wp-sheet-preview")
async def wp_sheet_preview(
    project_id: str = Query(..., description="项目 ID"),
    wp_code: str = Query(..., description="底稿编码（如 D2）"),
    sheet_name: str | None = Query(None, description="目标 sheet 名称（缺省返回首个 sheet）"),
    db: AsyncSession = Depends(get_db),
):
    """高级查询 sheet 选区器专用端点：返回模板 cellData 用于网格预览。

    流程：wp_code → wp_index.id → working_paper.id → storage xlsx → openpyxl 解析
    返回单 sheet 的 {rows, cols, cells: {[r,c]: {v, f}}, merges}，体积比 Univer 完整 JSON 小
    """
    try:
        proj_uuid = uuid.UUID(project_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="invalid project_id")

    # 1. 找 working_paper.id
    row = (await db.execute(text("""
        SELECT w.id FROM working_paper w
        JOIN wp_index i ON w.wp_index_id = i.id
        WHERE i.project_id = :pid AND i.wp_code = :wpc AND w.is_deleted = false
        LIMIT 1
    """), {"pid": str(proj_uuid), "wpc": wp_code})).first()

    if not row:
        return {"available": False, "reason": "底稿在当前项目中不存在", "wp_code": wp_code}

    wp_id = row[0]

    # 2. 找 storage 文件
    from app.services.wp_template_init_service import get_workpaper_file, init_workpaper_from_template
    storage_path = get_workpaper_file(proj_uuid, wp_id)
    if not storage_path:
        # 从模板初始化
        init_result = init_workpaper_from_template(proj_uuid, wp_id, wp_code)
        if not init_result:
            return {"available": False, "reason": f"模板文件不存在: {wp_code}", "wp_code": wp_code}
        storage_path = init_result

    if storage_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"available": False, "reason": "非 xlsx 类底稿", "wp_code": wp_code}

    # 3. openpyxl 解析（异步化：放线程池避免阻塞 event loop）
    try:
        content = storage_path.read_bytes()
        loop = asyncio.get_running_loop()
        parsed = await loop.run_in_executor(None, _sync_parse_sheet_preview, content, sheet_name)
        if "error" in parsed:
            return {"available": False, "reason": parsed["error"], "wp_code": wp_code}
        return {
            "available": True,
            "wp_code": wp_code,
            "wp_id": str(wp_id),
            **parsed,
        }
    except Exception as e:
        logger.warning("wp-sheet-preview parse failed for %s: %s", wp_code, e)
        return {"available": False, "reason": f"解析失败: {e}", "wp_code": wp_code}


@router.post("/execute")
async def execute_query(
    body: QueryRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """执行自定义查询（关键路径自动写 audit_log）"""
    source = body.source
    pid = body.project_id
    year = body.year
    filters = body.filters
    limit = min(body.limit, 2000)

    result: dict = {"rows": [], "columns": [], "total": 0}
    try:
        # disclosure_note:section_id 形式（树叶子点击）→ 走 disclosure 并把 section_id 注入 filters
        if source.startswith("disclosure_note:"):
            sid = source.split(":", 1)[1]
            new_filters = {**filters, "section_id": sid}
            result = await _query_disclosure(db, pid, year, new_filters, limit)
        # consol_unit:{company_code}:{kind} 形式（合并单位树叶子点击）
        elif source.startswith("consol_unit:"):
            parts = source.split(":", 2)
            if len(parts) == 3:
                _, cc, kind = parts
                new_filters = {**filters, "company_code": cc}
                if kind == "account_balance":
                    result = await _query_account_balance(db, pid, year, new_filters, limit)
                elif kind == "ledger_entries":
                    result = await _query_ledger_entries(db, pid, year, new_filters, limit)
                elif kind == "tb_detail":
                    result = await _query_trial_balance(db, pid, year, new_filters, limit)
                elif kind == "adjustment":
                    result = await _query_adjustments(db, pid, year, {**filters, "adjustment_type": "AJE"}, limit)
                else:
                    result = {"rows": [], "columns": [], "total": 0, "error": f"未知合并单位查询: {source}"}
            else:
                result = {"rows": [], "columns": [], "total": 0, "error": f"未知合并单位查询: {source}"}
        # workpaper:{wp_code} 或 workpaper:{wp_code}|{sheet_name} 形式（底稿树叶子点击）
        elif source.startswith("workpaper:"):
            tail = source.split(":", 1)[1]
            if "|" in tail:
                wp_code, sheet_name = tail.split("|", 1)
                result = await _query_workpaper(db, pid, year, {**filters, "wp_code": wp_code, "sheet_name": sheet_name}, limit)
            else:
                result = await _query_workpaper(db, pid, year, {**filters, "wp_code": tail}, limit)
        elif source == 'report' or source.startswith('report_'):
            result = await _query_report(db, pid, year, filters, limit)
        elif source == 'trial_balance' or source == 'tb_detail':
            result = await _query_trial_balance(db, pid, year, filters, limit)
        elif source == 'tb_summary':
            result = await _query_tb_summary(db, pid, year, filters, limit)
        elif source == 'disclosure' or source == 'disclosure_note':
            result = await _query_disclosure(db, pid, year, filters, limit)
        elif source.startswith('adj_') or source == 'adjustment':
            result = await _query_adjustments(db, pid, year, filters, limit)
        elif source.startswith('ws_') or source == 'worksheet':
            result = await _query_worksheet(db, pid, year, filters, limit)
        elif source == 'workpaper':
            result = await _query_workpaper(db, pid, year, filters, limit)
        elif source == 'account_balance':
            result = await _query_account_balance(db, pid, year, filters, limit)
        elif source == 'ledger_entries':
            result = await _query_ledger_entries(db, pid, year, filters, limit)
        elif source == 'report_lines':
            result = await _query_report_lines(db, pid, year, filters, limit)
        elif source == 'workhours':
            result = await _query_workhours(db, pid, year, filters, limit)
        else:
            result = {"rows": [], "columns": [], "total": 0, "error": f"未知数据源: {source}"}
    except Exception as e:
        # 失败时回滚事务（防止 asyncpg session 污染影响后续请求）
        try:
            await db.rollback()
        except Exception:
            pass
        result = {"rows": [], "columns": [], "total": 0, "error": str(e)}

    # 审计日志（关键路径：底稿/合并/附注/调整 写日志，普通试算/报表/工时跳过避免噪声）
    _SENSITIVE_PREFIXES = ("workpaper:", "consol_unit:", "disclosure_note:", "adj_", "disclosure")
    if source.startswith(_SENSITIVE_PREFIXES) or source in ("workpaper", "adjustment", "disclosure"):
        try:
            from app.services.audit_logger_enhanced import audit_logger
            await audit_logger.log_action(
                user_id=current_user.id,
                action="custom_query.execute",
                object_type="custom_query",
                object_id=None,
                project_id=pid,
                details={
                    "source": source,
                    "year": year,
                    "filters": {k: v for k, v in filters.items() if not isinstance(v, (dict, list)) or len(str(v)) < 200},
                    "row_count": result.get("total", 0),
                    "has_error": "error" in result,
                },
            )
        except Exception as audit_e:
            logger.warning("audit_log enqueue failed for custom_query: %s", audit_e)

    return result


async def _query_report(db, pid, year, filters, limit):
    report_type = filters.get("report_type", "balance_sheet")
    # standard 优先级：filters.standard > 项目 (template_type + report_scope) 推断 > 默认 soe_standalone
    standard = filters.get("standard")
    if not standard:
        template_type = await _resolve_project_template_type(db, pid)
        scope = await _resolve_project_report_scope(db, pid)
        # report_config 取值：soe_standalone / listed_standalone / listed_consolidated
        # soe 暂无 consolidated 数据，强制走 standalone
        if template_type == "soe":
            standard = "soe_standalone"
        else:
            standard = f"listed_{scope or 'standalone'}"
    # 优先查项目级数据，降级查全局模板
    query = "SELECT row_code, row_name, current_period_amount, prior_period_amount, indent_level, is_total_row FROM report_config WHERE report_type = :rt AND applicable_standard = :std AND is_deleted = false"
    params: dict = {"rt": report_type, "std": standard, "lim": limit}
    if pid:
        query += " AND (project_id = :pid OR project_id IS NULL)"
        params["pid"] = pid
    query += " ORDER BY row_number LIMIT :lim"
    result = await db.execute(text(query), params)
    rows = [{"row_code": r[0], "row_name": r[1], "current_period_amount": float(r[2]) if r[2] else None, "prior_period_amount": float(r[3]) if r[3] else None, "indent": r[4], "is_total": r[5]} for r in result.fetchall()]
    return {"rows": rows, "columns": ["row_code", "row_name", "current_period_amount", "prior_period_amount"], "total": len(rows), "applicable_standard": standard}


async def _query_trial_balance(db, pid, year, filters, limit):
    query = "SELECT account_code, account_name, opening_balance, closing_balance, debit_amount, credit_amount FROM trial_balance_entries WHERE project_id = :pid AND year = :y"
    params: dict = {"pid": pid, "y": year, "lim": limit}
    if filters.get("account_name"):
        query += " AND account_name LIKE :an"
        params["an"] = f"%{filters['account_name']}%"
    if filters.get("company_code"):
        query += " AND company_code = :cc"
        params["cc"] = filters["company_code"]
    query += " ORDER BY account_code LIMIT :lim"
    result = await db.execute(text(query), params)
    rows = [{"account_code": r[0], "account_name": r[1], "opening_balance": float(r[2]) if r[2] else None, "closing_balance": float(r[3]) if r[3] else None, "debit_amount": float(r[4]) if r[4] else None, "credit_amount": float(r[5]) if r[5] else None} for r in result.fetchall()]
    return {"rows": rows, "columns": ["account_code", "account_name", "opening_balance", "closing_balance", "debit_amount", "credit_amount"], "total": len(rows)}


async def _query_tb_summary(db, pid, year, filters, limit):
    sheet_type = filters.get("report_type", "balance_sheet")
    result = await db.execute(
        text("SELECT data FROM consol_worksheet_data WHERE project_id = :pid AND year = :y AND sheet_key = :sk"),
        {"pid": pid, "y": year, "sk": f"tb_summary_{sheet_type}"},
    )
    row = result.fetchone()
    if row and isinstance(row[0], dict):
        rows = row[0].get("rows", [])[:limit]
        return {"rows": rows, "columns": ["row_code", "row_name", "unadjusted", "aje_dr", "aje_cr", "rcl_dr", "rcl_cr", "audited"], "total": len(rows)}
    return {"rows": [], "columns": [], "total": 0}


async def _query_disclosure(db, pid, year, filters, limit):
    section_id = filters.get("section_id", "")
    if section_id:
        result = await db.execute(
            text("SELECT section_id, data FROM consol_note_data WHERE project_id = :pid AND year = :y AND section_id = :sid"),
            {"pid": pid, "y": year, "sid": section_id},
        )
    else:
        result = await db.execute(
            text("SELECT section_id, data FROM consol_note_data WHERE project_id = :pid AND year = :y LIMIT :lim"),
            {"pid": pid, "y": year, "lim": limit},
        )
    # 将附注数据展平为表格行（每个章节的每行数据变成一条记录）
    flat_rows = []
    all_headers: list[str] = []
    for r in result.fetchall():
        data = r[1] if isinstance(r[1], dict) else {}
        headers = data.get("headers", [])
        rows = data.get("rows", [])
        if headers and not all_headers:
            all_headers = ["section_id"] + headers
        for row_data in rows[:100]:  # 每章节最多100行
            obj: dict = {"section_id": r[0]}
            for hi, h in enumerate(headers):
                obj[h] = row_data[hi] if hi < len(row_data) else ''
            flat_rows.append(obj)
    columns = all_headers if all_headers else ["section_id"]
    return {"rows": flat_rows[:limit], "columns": columns, "total": len(flat_rows)}


async def _query_adjustments(db, pid, year, filters, limit):
    adj_type = filters.get("adjustment_type", "AJE")
    result = await db.execute(
        text("SELECT entry_number, account_name, debit_amount, credit_amount, description, status FROM adjustments WHERE project_id = :pid AND year = :y AND adjustment_type = :at AND is_deleted = false ORDER BY entry_number LIMIT :lim"),
        {"pid": pid, "y": year, "at": adj_type, "lim": limit},
    )
    rows = [{"entry_number": r[0], "account_name": r[1], "debit_amount": float(r[2]) if r[2] else None, "credit_amount": float(r[3]) if r[3] else None, "description": r[4], "status": r[5]} for r in result.fetchall()]
    return {"rows": rows, "columns": ["entry_number", "account_name", "debit_amount", "credit_amount", "description", "status"], "total": len(rows)}


async def _query_worksheet(db, pid, year, filters, limit):
    sheet_key = filters.get("sheet_key", "info")
    result = await db.execute(
        text("SELECT data, updated_at FROM consol_worksheet_data WHERE project_id = :pid AND year = :y AND sheet_key = :sk"),
        {"pid": pid, "y": year, "sk": sheet_key},
    )
    row = result.fetchone()
    if row and isinstance(row[0], dict):
        data = row[0]
        rows = data.get("rows", [])[:limit]
        columns = list(rows[0].keys()) if rows else []
        return {"rows": rows, "columns": columns, "total": len(rows), "updated_at": str(row[1]) if row[1] else None}
    return {"rows": [], "columns": [], "total": 0}



# ──────────────────────────────────────────────────────────────────────────
# Sprint 6 Task 6.4 — 新增 5 个数据源查询函数
# ──────────────────────────────────────────────────────────────────────────


async def _query_workpaper(db, pid, year, filters, limit):
    """底稿列表（working_paper + wp_index）— working_paper 无 year 字段，year 仅用于过滤兼容

    特殊路径：当 filters 含 wp_code + cell_range 时，
    走「单元格区域提取」分支：解析模板 xlsx 的指定 sheet，按 cell_range 提取真实单元格值返回（用于高级查询 sheet 选区器）
    """
    # 单元格区域提取分支（高级查询 sheet picker 专用）
    if filters.get("wp_code") and filters.get("cell_range"):
        return await _query_workpaper_cell_range(db, pid, filters)

    sql = """
        SELECT wi.wp_code, wi.wp_name, wi.audit_cycle, wp.status, wp.review_status,
               wp.assigned_to, wp.created_at, wp.updated_at
        FROM working_paper wp
        LEFT JOIN wp_index wi ON wi.id = wp.wp_index_id
        WHERE wp.project_id = :pid AND wp.is_deleted = false
    """
    params: dict = {"pid": pid, "lim": limit}
    if filters.get("status"):
        sql += " AND wp.status = :st"
        params["st"] = filters["status"]
    if filters.get("review_status"):
        sql += " AND wp.review_status = :rs"
        params["rs"] = filters["review_status"]
    if filters.get("cycle"):
        sql += " AND wi.audit_cycle = :cy"
        params["cy"] = filters["cycle"]
    if filters.get("wp_code"):
        # 支持精确单底稿查询（树叶子点击）+ 同 cycle 下子表前缀模糊（wp_code='E1' 时也带出 E1-1/E1-2）
        sql += " AND (wi.wp_code = :wpc OR wi.wp_code LIKE :wpc_like)"
        params["wpc"] = filters["wp_code"]
        params["wpc_like"] = f"{filters['wp_code']}-%"
    sql += " ORDER BY wi.wp_code LIMIT :lim"
    result = await db.execute(text(sql), params)
    sheet_echo = filters.get("sheet_name")
    rows = []
    for r in result.fetchall():
        row = {
            "wp_code": r[0],
            "wp_name": r[1],
            "audit_cycle": r[2],
            "status": r[3],
            "review_status": r[4],
            "assigned_to": str(r[5]) if r[5] else None,
            "created_at": str(r[6]) if r[6] else None,
            "updated_at": str(r[7]) if r[7] else None,
        }
        if sheet_echo:
            row["sheet_name"] = sheet_echo
        rows.append(row)
    columns = ["wp_code", "wp_name", "audit_cycle", "status", "review_status", "assigned_to"]
    if sheet_echo:
        columns.append("sheet_name")
    return {
        "rows": rows,
        "columns": columns,
        "total": len(rows),
    }


async def _query_workpaper_cell_range(db, pid: str, filters: dict):
    """按 cell_range 从模板 xlsx 提取真实单元格值

    支持语法：
      - 'A1:C3' 矩形区域
      - 'B5' 单 cell
      - 'A1:A10,C1:C5' 多区域逗号分隔
      - 'A:A' 整列（限 100 行）
      - 'A:C' 整列范围
    """
    wp_code = filters["wp_code"]
    sheet_name = filters.get("sheet_name")
    cell_range = filters["cell_range"]

    try:
        proj_uuid = uuid.UUID(pid)
    except (TypeError, ValueError):
        return {"rows": [], "columns": [], "total": 0, "error": "invalid project_id"}

    # 1. wp_code → wp_id
    row = (await db.execute(text("""
        SELECT w.id FROM working_paper w
        JOIN wp_index i ON w.wp_index_id = i.id
        WHERE i.project_id = :pid AND i.wp_code = :wpc AND w.is_deleted = false
        LIMIT 1
    """), {"pid": str(proj_uuid), "wpc": wp_code})).first()
    if not row:
        return {"rows": [], "columns": [], "total": 0, "error": f"底稿 {wp_code} 在当前项目不存在"}
    wp_id = row[0]

    # 2. storage 文件
    from app.services.wp_template_init_service import get_workpaper_file, init_workpaper_from_template
    storage_path = get_workpaper_file(proj_uuid, wp_id)
    if not storage_path:
        init_result = init_workpaper_from_template(proj_uuid, wp_id, wp_code)
        if not init_result:
            return {"rows": [], "columns": [], "total": 0, "error": f"模板文件不存在: {wp_code}"}
        storage_path = init_result

    if storage_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"rows": [], "columns": [], "total": 0, "error": "非 xlsx 类底稿"}

    # 3. 解析 cell_range 多区域语法
    ranges = _parse_cell_ranges(cell_range)
    if not ranges:
        return {"rows": [], "columns": [], "total": 0, "error": f"无效 cell_range: {cell_range}"}

    # 4. 异步化读取 + 多区域累计
    try:
        content = storage_path.read_bytes()
        loop = asyncio.get_running_loop()
        all_rows: list[dict] = []
        used_sheet = ""
        idx = 0
        MAX_TOTAL = 500
        for (r1, c1, r2, c2) in ranges:
            remaining = MAX_TOTAL - idx
            if remaining <= 0:
                break
            parsed = await loop.run_in_executor(
                None, _sync_extract_cell_range, content, sheet_name, r1, c1, r2, c2, remaining
            )
            if "error" in parsed:
                return {"rows": [], "columns": [], "total": 0, "error": parsed["error"]}
            used_sheet = parsed.get("sheet_name", used_sheet)
            for cell_row in parsed["rows"]:
                idx += 1
                all_rows.append({
                    "index": idx,
                    "wp_code": wp_code,
                    "sheet_name": used_sheet,
                    "cell_ref": cell_row["cell_ref"],
                    "value": cell_row["value"],
                })
                if idx >= MAX_TOTAL:
                    break
        return {
            "rows": all_rows,
            "columns": ["index", "wp_code", "sheet_name", "cell_ref", "value"],
            "total": len(all_rows),
        }
    except Exception as e:
        logger.warning("cell_range extract failed for %s/%s: %s", wp_code, cell_range, e)
        return {"rows": [], "columns": [], "total": 0, "error": f"解析失败: {e}"}


def _parse_cell_ranges(spec: str) -> list[tuple[int, int, int, int]]:
    """解析 cell_range 字符串为 (r1, c1, r2, c2) 元组列表（1-indexed）

    支持：
      - 'A1:C3'    → [(1, 1, 3, 3)]
      - 'B5'       → [(5, 2, 5, 2)]
      - 'A1:A10,C1:C5' → [(1, 1, 10, 1), (1, 3, 5, 3)]
      - 'A:A'      → [(1, 1, 100, 1)]  整列限 100 行
      - 'A:C'      → [(1, 1, 100, 3)]
    """
    import re as _re
    INTEGER_COL_LIMIT = 100  # 整列默认行数上限

    def _col_to_idx(s: str) -> int:
        n = 0
        for ch in s:
            n = n * 26 + (ord(ch) - 64)
        return n

    out: list[tuple[int, int, int, int]] = []
    for part in spec.split(","):
        part = part.strip().upper()
        if not part:
            continue
        # 单 cell 'B5'
        m = _re.match(r"^([A-Z]+)(\d+)$", part)
        if m:
            col = _col_to_idx(m.group(1)); r = int(m.group(2))
            out.append((r, col, r, col))
            continue
        # 整列 'A:A' / 'A:C'
        m = _re.match(r"^([A-Z]+):([A-Z]+)$", part)
        if m:
            c1 = _col_to_idx(m.group(1)); c2 = _col_to_idx(m.group(2))
            out.append((1, min(c1, c2), INTEGER_COL_LIMIT, max(c1, c2)))
            continue
        # 矩形 'A1:C3'
        m = _re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", part)
        if m:
            c1 = _col_to_idx(m.group(1)); r1 = int(m.group(2))
            c2 = _col_to_idx(m.group(3)); r2 = int(m.group(4))
            out.append((min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)))
            continue
    return out


async def _query_account_balance(db, pid, year, filters, limit):
    """科目余额表（tb_balance）— B' 视图架构按 is_deleted=false 过滤"""
    sql = """
        SELECT account_code, account_name, opening_balance, closing_balance,
               debit_amount, credit_amount, currency_code
        FROM tb_balance
        WHERE project_id = :pid AND year = :y AND is_deleted = false
    """
    params: dict = {"pid": pid, "y": year, "lim": limit}
    if filters.get("account_code"):
        sql += " AND account_code LIKE :ac"
        params["ac"] = f"{filters['account_code']}%"
    if filters.get("account_name"):
        sql += " AND account_name LIKE :an"
        params["an"] = f"%{filters['account_name']}%"
    if filters.get("company_code"):
        sql += " AND company_code = :cc"
        params["cc"] = filters["company_code"]
    sql += " ORDER BY account_code LIMIT :lim"
    result = await db.execute(text(sql), params)
    rows = [
        {
            "account_code": r[0],
            "account_name": r[1],
            "opening_balance": float(r[2]) if r[2] is not None else None,
            "closing_balance": float(r[3]) if r[3] is not None else None,
            "debit_amount": float(r[4]) if r[4] is not None else None,
            "credit_amount": float(r[5]) if r[5] is not None else None,
            "currency_code": r[6],
        }
        for r in result.fetchall()
    ]
    return {
        "rows": rows,
        "columns": ["account_code", "account_name", "opening_balance", "closing_balance", "debit_amount", "credit_amount", "currency_code"],
        "total": len(rows),
    }


async def _query_ledger_entries(db, pid, year, filters, limit):
    """序时账（tb_ledger）"""
    sql = """
        SELECT voucher_date, voucher_no, account_code, account_name,
               debit_amount, credit_amount, summary
        FROM tb_ledger
        WHERE project_id = :pid AND year = :y AND is_deleted = false
    """
    params: dict = {"pid": pid, "y": year, "lim": limit}
    if filters.get("account_code"):
        sql += " AND account_code LIKE :ac"
        params["ac"] = f"{filters['account_code']}%"
    if filters.get("voucher_no"):
        sql += " AND voucher_no LIKE :vn"
        params["vn"] = f"%{filters['voucher_no']}%"
    if filters.get("summary"):
        sql += " AND summary LIKE :sm"
        params["sm"] = f"%{filters['summary']}%"
    if filters.get("date_from"):
        sql += " AND voucher_date >= :df"
        params["df"] = filters["date_from"]
    if filters.get("date_to"):
        sql += " AND voucher_date <= :dt"
        params["dt"] = filters["date_to"]
    sql += " ORDER BY voucher_date DESC, voucher_no LIMIT :lim"
    result = await db.execute(text(sql), params)
    rows = [
        {
            "voucher_date": str(r[0]) if r[0] else None,
            "voucher_no": r[1],
            "account_code": r[2],
            "account_name": r[3],
            "debit_amount": float(r[4]) if r[4] is not None else None,
            "credit_amount": float(r[5]) if r[5] is not None else None,
            "summary": r[6],
        }
        for r in result.fetchall()
    ]
    return {
        "rows": rows,
        "columns": ["voucher_date", "voucher_no", "account_code", "account_name", "debit_amount", "credit_amount", "summary"],
        "total": len(rows),
    }


async def _query_report_lines(db, pid, year, filters, limit):
    """报表行次配置（report_config）"""
    sql = """
        SELECT row_code, row_name, report_type, applicable_standard,
               indent_level, is_total_row, formula, sort_order
        FROM report_config
        WHERE is_deleted = false
    """
    params: dict = {"lim": limit}
    if filters.get("report_type"):
        sql += " AND report_type = :rt"
        params["rt"] = filters["report_type"]
    if filters.get("applicable_standard"):
        sql += " AND applicable_standard = :std"
        params["std"] = filters["applicable_standard"]
    if filters.get("has_formula") is True:
        sql += " AND formula IS NOT NULL AND formula != ''"
    elif filters.get("has_formula") is False:
        sql += " AND (formula IS NULL OR formula = '')"
    if filters.get("row_name"):
        sql += " AND row_name LIKE :rn"
        params["rn"] = f"%{filters['row_name']}%"
    sql += " ORDER BY applicable_standard, report_type, sort_order LIMIT :lim"
    result = await db.execute(text(sql), params)
    rows = [
        {
            "row_code": r[0],
            "row_name": r[1],
            "report_type": r[2],
            "applicable_standard": r[3],
            "indent_level": r[4],
            "is_total_row": r[5],
            "formula": r[6],
            "sort_order": r[7],
        }
        for r in result.fetchall()
    ]
    return {
        "rows": rows,
        "columns": ["row_code", "row_name", "report_type", "applicable_standard", "indent_level", "is_total_row", "formula", "sort_order"],
        "total": len(rows),
    }


async def _query_workhours(db, pid, year, filters, limit):
    """工时记录（work_hours）"""
    sql = """
        SELECT work_date, hours, description, status, staff_id, project_id, created_at
        FROM work_hours
        WHERE is_deleted = false
    """
    params: dict = {"lim": limit}
    # 可选项目过滤（pid 可能为空字符串）
    if pid:
        sql += " AND project_id = :pid"
        params["pid"] = pid
    if year:
        sql += " AND EXTRACT(YEAR FROM work_date) = :y"
        params["y"] = year
    if filters.get("staff_id"):
        sql += " AND staff_id = :sid"
        params["sid"] = filters["staff_id"]
    if filters.get("status"):
        sql += " AND status = :st"
        params["st"] = filters["status"]
    if filters.get("date_from"):
        sql += " AND work_date >= :df"
        params["df"] = filters["date_from"]
    if filters.get("date_to"):
        sql += " AND work_date <= :dt"
        params["dt"] = filters["date_to"]
    sql += " ORDER BY work_date DESC LIMIT :lim"
    result = await db.execute(text(sql), params)
    rows = [
        {
            "work_date": str(r[0]) if r[0] else None,
            "hours": float(r[1]) if r[1] is not None else 0,
            "description": r[2],
            "status": r[3],
            "staff_id": str(r[4]) if r[4] else None,
            "project_id": str(r[5]) if r[5] else None,
            "created_at": str(r[6]) if r[6] else None,
        }
        for r in result.fetchall()
    ]
    return {
        "rows": rows,
        "columns": ["work_date", "hours", "description", "status", "staff_id"],
        "total": len(rows),
    }


# ──────────────────────────────────────────────────────────────────────────
# Sprint 6 Task 6.6 — 自定义查询模板 CRUD 端点
# ──────────────────────────────────────────────────────────────────────────


class TemplateCreateRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    description: str | None = None
    data_source: str = Field(..., min_length=1, max_length=50)
    config: dict
    scope: Literal["private", "global"] = "private"


class TemplateUpdateRequest(BaseModel):
    name: str | None = Field(None, min_length=1, max_length=200)
    description: str | None = None
    data_source: str | None = None
    config: dict | None = None
    scope: Literal["private", "global"] | None = None


class TemplateResponse(BaseModel):
    id: str
    name: str
    description: str | None
    data_source: str
    config: dict
    scope: str
    created_by: str
    is_owner: bool
    created_at: str
    updated_at: str


def _serialize_template(tpl: CustomQueryTemplate, current_user_id: uuid.UUID) -> dict:
    return {
        "id": str(tpl.id),
        "name": tpl.name,
        "description": tpl.description,
        "data_source": tpl.data_source,
        "config": tpl.config or {},
        "scope": tpl.scope,
        "created_by": str(tpl.created_by),
        "is_owner": tpl.created_by == current_user_id,
        "created_at": tpl.created_at.isoformat() if tpl.created_at else None,
        "updated_at": tpl.updated_at.isoformat() if tpl.updated_at else None,
    }


@router.get("/templates")
async def list_templates(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """列出可见模板：本人创建的所有 + 全局共享。

    按 updated_at 倒序排列（最近编辑置顶）。
    """
    stmt = (
        select(CustomQueryTemplate)
        .where(
            or_(
                CustomQueryTemplate.created_by == current_user.id,
                CustomQueryTemplate.scope == "global",
            )
        )
        .order_by(CustomQueryTemplate.updated_at.desc())
    )
    result = await db.execute(stmt)
    templates = result.scalars().all()
    return {
        "templates": [_serialize_template(t, current_user.id) for t in templates],
        "total": len(templates),
    }


@router.post("/templates", status_code=201)
async def create_template(
    body: TemplateCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建查询模板。"""
    tpl = CustomQueryTemplate(
        id=uuid.uuid4(),
        name=body.name,
        description=body.description,
        data_source=body.data_source,
        config=body.config,
        scope=body.scope,
        created_by=current_user.id,
    )
    db.add(tpl)
    await db.commit()
    await db.refresh(tpl)
    return _serialize_template(tpl, current_user.id)


@router.get("/templates/{template_id}")
async def get_template(
    template_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取模板详情。"""
    try:
        tpl_uuid = uuid.UUID(template_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail={"error_code": "INVALID_TEMPLATE_ID"})
    tpl = await db.get(CustomQueryTemplate, tpl_uuid)
    if not tpl:
        raise HTTPException(status_code=404, detail={"error_code": "TEMPLATE_NOT_FOUND"})
    # 权限：本人 or 全局可见
    if tpl.scope != "global" and tpl.created_by != current_user.id:
        raise HTTPException(status_code=403, detail={"error_code": "TEMPLATE_NOT_VISIBLE"})
    return _serialize_template(tpl, current_user.id)


@router.put("/templates/{template_id}")
async def update_template(
    template_id: str,
    body: TemplateUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新模板（仅创建者）。"""
    try:
        tpl_uuid = uuid.UUID(template_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail={"error_code": "INVALID_TEMPLATE_ID"})
    tpl = await db.get(CustomQueryTemplate, tpl_uuid)
    if not tpl:
        raise HTTPException(status_code=404, detail={"error_code": "TEMPLATE_NOT_FOUND"})
    user_role = getattr(current_user, "role", "")
    if tpl.created_by != current_user.id and user_role != "admin":
        raise HTTPException(status_code=403, detail={"error_code": "ONLY_OWNER_CAN_UPDATE"})

    if body.name is not None:
        tpl.name = body.name
    if body.description is not None:
        tpl.description = body.description
    if body.data_source is not None:
        tpl.data_source = body.data_source
    if body.config is not None:
        tpl.config = body.config
    if body.scope is not None:
        tpl.scope = body.scope
    tpl.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(tpl)
    return _serialize_template(tpl, current_user.id)


@router.delete("/templates/{template_id}", status_code=204)
async def delete_template(
    template_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除模板（仅创建者或 admin）。"""
    try:
        tpl_uuid = uuid.UUID(template_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail={"error_code": "INVALID_TEMPLATE_ID"})
    tpl = await db.get(CustomQueryTemplate, tpl_uuid)
    if not tpl:
        raise HTTPException(status_code=404, detail={"error_code": "TEMPLATE_NOT_FOUND"})
    user_role = getattr(current_user, "role", "")
    if tpl.created_by != current_user.id and user_role != "admin":
        raise HTTPException(status_code=403, detail={"error_code": "ONLY_OWNER_OR_ADMIN_CAN_DELETE"})
    await db.delete(tpl)
    await db.commit()
    return None
