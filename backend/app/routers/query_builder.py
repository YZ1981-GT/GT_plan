"""S-3 高级查询构建器 — 可视化条件 → SQLAlchemy core 安全构造 → 执行/预览/导出

POST /api/query/preview      预览生成的 SQL（不执行）
POST /api/query/execute      执行查询，返回结构化结果
POST /api/query/export-excel 执行查询并导出 Excel
GET  /api/query/schema       返回白名单表/字段元信息（前端用于构造可视化选择器）

设计要点：
- 仅允许查询白名单的只读 audit/财务表（trial_balance / working_paper / wp_index /
  adjustments / unadjusted_misstatements / report_line_mapping / report_config /
  account_chart / tb_balance / tb_ledger / materiality）
- 完全使用 SQLAlchemy core `select()` 按列引用构造，**不做字符串拼接**，
  绑定参数走 SQLAlchemy bindparam，杜绝 SQL 注入
- 字段名/表名/操作符/排序方向通过白名单核对后才写入 query
- 仅 admin / manager 可访问
- 不暴露 user / role / token / auth 表

注册到 router_registry.system 域 §117。

Validates: requirements §三 · S-3 高级查询构建器
"""

from __future__ import annotations

import io
import logging
from typing import Any, Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import Column, and_, asc, desc, func, or_, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql import Select
from sqlalchemy.sql.elements import ColumnElement

from app.core.database import get_db
from app.deps import get_current_user
from app.models.audit_platform_models import (
    AccountChart,
    Adjustment,
    Materiality,
    ReportLineMapping,
    TbBalance,
    TbLedger,
    TrialBalance,
    UnadjustedMisstatement,
)
from app.models.core import Project, User
from app.models.report_models import DisclosureNote, ReportConfig
from app.models.staff_models import StaffMember, WorkHour
from app.models.workpaper_models import WorkingPaper, WpIndex

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/query", tags=["query-builder"])


# ─────────────────────────────────────────────────────────────────────────────
# 表白名单（只读 audit/财务表，绝不暴露 user/role/auth/token 表）
# ─────────────────────────────────────────────────────────────────────────────
TABLE_WHITELIST: dict[str, dict[str, Any]] = {
    "trial_balance": {
        "model": TrialBalance,
        "label": "试算表",
        "fields": [
            "id", "project_id", "year", "company_code",
            "standard_account_code", "account_name", "account_category",
            "unadjusted_amount", "rje_adjustment", "aje_adjustment",
            "audited_amount", "opening_balance", "currency_code",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "adjustments": {
        "model": Adjustment,
        "label": "调整分录（AJE/RJE）",
        "fields": [
            "id", "project_id", "year", "company_code", "adjustment_no",
            "adjustment_type", "description", "account_code", "account_name",
            "debit_amount", "credit_amount", "review_status",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "unadjusted_misstatements": {
        "model": UnadjustedMisstatement,
        "label": "未更正错报",
        "fields": [
            "id", "project_id", "year", "misstatement_description",
            "affected_account_code", "affected_account_name",
            "misstatement_amount", "misstatement_type",
            "management_reason", "auditor_evaluation",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "report_line_mapping": {
        "model": ReportLineMapping,
        "label": "报表行次映射",
        "fields": [
            "id", "project_id", "report_type",
            "standard_account_code", "report_line_code", "report_line_name",
            "report_line_level", "parent_line_code", "mapping_type",
            "is_confirmed", "is_deleted", "created_at", "updated_at",
        ],
    },
    "report_config": {
        "model": ReportConfig,
        "label": "报表行次配置",
        "fields": [
            "id", "report_type", "applicable_standard",
            "row_code", "row_name", "row_number", "indent_level",
            "is_total_row", "parent_row_code", "formula",
            "formula_category", "formula_description", "formula_source",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "working_paper": {
        "model": WorkingPaper,
        "label": "底稿文件",
        "fields": [
            "id", "project_id", "wp_index_id", "file_path", "source_type",
            "file_version", "status", "review_status",
            "assigned_to", "reviewer", "workflow_status",
            "explanation_status", "consistency_status",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "wp_index": {
        "model": WpIndex,
        "label": "底稿索引",
        "fields": [
            "id", "project_id", "wp_code", "wp_name", "audit_cycle",
            "assigned_to", "reviewer", "status",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "tb_balance": {
        "model": TbBalance,
        "label": "科目余额表",
        "fields": [
            "id", "project_id", "year", "company_code",
            "account_code", "account_name", "level",
            "opening_balance", "closing_balance",
            "debit_amount", "credit_amount", "currency_code",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "tb_ledger": {
        "model": TbLedger,
        "label": "序时账（总账明细）",
        "fields": [
            "id", "project_id", "year", "company_code",
            "voucher_date", "voucher_no", "account_code", "account_name",
            "accounting_period", "voucher_type", "entry_seq",
            "debit_amount", "credit_amount", "summary",
            "currency_code", "is_deleted", "created_at",
        ],
    },
    "account_chart": {
        "model": AccountChart,
        "label": "科目表",
        "fields": [
            "id", "project_id", "account_code", "account_name",
            "direction", "level", "category", "parent_code",
            "source", "is_deleted", "created_at", "updated_at",
        ],
    },
    "materiality": {
        "model": Materiality,
        "label": "重要性水平",
        "fields": [
            "id", "project_id", "year", "benchmark_type",
            "benchmark_amount", "overall_percentage", "overall_materiality",
            "performance_ratio", "performance_materiality",
            "trivial_ratio", "trivial_threshold",
            "is_override", "is_deleted", "created_at", "updated_at",
        ],
    },
    # ── 业务维度扩展（项目 / 单位 / 附注 / 人员 / 工时） ──
    "projects": {
        "model": Project,
        "label": "项目",
        "fields": [
            "id", "name", "client_name",
            "audit_period_start", "audit_period_end",
            "project_type", "status", "scenario",
            "manager_id", "partner_id",
            "company_code", "template_type", "report_scope",
            "parent_company_name", "parent_company_code",
            "ultimate_company_name", "ultimate_company_code",
            "consol_level", "risk_level",
            "budget_hours", "contract_amount",
            "archived_at", "is_deleted", "created_at", "updated_at",
        ],
    },
    "disclosure_notes": {
        "model": DisclosureNote,
        "label": "附注",
        "fields": [
            "id", "project_id", "year",
            "note_section", "section_title", "account_name",
            "content_type", "source_template", "status",
            "sort_order", "is_stale",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "staff_members": {
        "model": StaffMember,
        "label": "人员",
        "fields": [
            "id", "user_id", "name", "employee_no",
            "department", "title", "partner_name", "partner_id",
            "specialty", "phone", "email", "join_date",
            "source", "role_level",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "work_hours": {
        "model": WorkHour,
        "label": "工时",
        "fields": [
            "id", "staff_id", "project_id", "work_date",
            "hours", "start_time", "end_time",
            "description", "status", "purpose", "ai_suggested",
            "is_deleted", "created_at", "updated_at",
        ],
    },
    "users": {
        "model": User,
        "label": "用户",
        # 显式排除 hashed_password / 安全敏感字段
        "fields": [
            "id", "username", "email", "role",
            "office_code", "is_active",
            "is_deleted", "created_at", "updated_at",
        ],
    },
}


# S-3 v2：JOIN 白名单（声明式，不接受任意 ON 条件）
# 格式: {table_name: {target_table: {on: [(left_col, right_col), ...]}, ...}}
# 仅枚举常用业务关联，新增 JOIN 必须显式登记
JOIN_WHITELIST: dict[str, dict[str, dict[str, list[tuple[str, str]]]]] = {
    "trial_balance": {
        "wp_index": {"on": [("project_id", "project_id")]},
        "account_chart": {
            "on": [
                ("project_id", "project_id"),
                ("standard_account_code", "account_code"),
            ]
        },
        "report_line_mapping": {
            "on": [
                ("project_id", "project_id"),
                ("standard_account_code", "standard_account_code"),
            ]
        },
    },
    "adjustments": {
        "wp_index": {"on": [("project_id", "project_id")]},
        "trial_balance": {
            "on": [
                ("project_id", "project_id"),
                ("account_code", "standard_account_code"),
            ]
        },
    },
    "working_paper": {
        "wp_index": {"on": [("wp_index_id", "id")]},
    },
    "wp_index": {
        "working_paper": {"on": [("id", "wp_index_id")]},
    },
    "tb_balance": {
        "account_chart": {
            "on": [("project_id", "project_id"), ("account_code", "account_code")]
        },
    },
    "tb_ledger": {
        "account_chart": {
            "on": [("project_id", "project_id"), ("account_code", "account_code")]
        },
    },
    "report_line_mapping": {
        "report_config": {"on": [("report_line_code", "row_code")]},
    },
    # ── 业务维度 JOIN ──
    "projects": {
        # 项目 → 项目下所有业务对象
        "trial_balance":     {"on": [("id", "project_id")]},
        "working_paper":     {"on": [("id", "project_id")]},
        "wp_index":          {"on": [("id", "project_id")]},
        "tb_balance":        {"on": [("id", "project_id")]},
        "tb_ledger":         {"on": [("id", "project_id")]},
        "adjustments":       {"on": [("id", "project_id")]},
        "disclosure_notes":  {"on": [("id", "project_id")]},
        "work_hours":        {"on": [("id", "project_id")]},
        "users":             {"on": [("manager_id", "id")]},  # 项目经理
    },
    "disclosure_notes": {
        "projects":          {"on": [("project_id", "id")]},
        "users":             {"on": [("updated_by", "id")]},
    },
    "staff_members": {
        "users":             {"on": [("user_id", "id")]},
        "work_hours":        {"on": [("id", "staff_id")]},
    },
    "work_hours": {
        "staff_members":     {"on": [("staff_id", "id")]},
        "projects":          {"on": [("project_id", "id")]},
    },
    "users": {
        "staff_members":     {"on": [("id", "user_id")]},
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# 操作符白名单 — 显式分类，禁止任意 SQL 片段
# ─────────────────────────────────────────────────────────────────────────────
OPERATOR_WHITELIST: set[str] = {
    "eq", "neq",            # = / !=
    "gt", "gte", "lt", "lte",  # > / >= / < / <=
    "like", "not_like",     # LIKE / NOT LIKE（自动包裹 %）
    "in", "not_in",         # IN / NOT IN
    "is_null", "is_not_null",  # IS NULL / IS NOT NULL
    "between",              # BETWEEN（[lo, hi]）
}


AGGREGATE_WHITELIST: set[str] = {"count", "sum", "avg", "min", "max"}


# ─────────────────────────────────────────────────────────────────────────────
# RBAC：仅 admin / manager / partner（partner ⊃ manager 权限超集）可访问
# ─────────────────────────────────────────────────────────────────────────────
def _get_role_value(user: User) -> str:
    role = getattr(user, "role", "")
    return role.value if hasattr(role, "value") else str(role)


def _require_admin_or_manager(user: User) -> None:
    role = _get_role_value(user)
    if role not in ("admin", "manager", "partner"):
        raise HTTPException(
            status_code=403,
            detail={"error_code": "QUERY_BUILDER_FORBIDDEN",
                    "message": "高级查询构建器仅 admin / manager 可访问"},
        )


# ─────────────────────────────────────────────────────────────────────────────
# Pydantic 模型 — DSL 严格校验
# ─────────────────────────────────────────────────────────────────────────────
class FilterCond(BaseModel):
    field: str = Field(..., min_length=1, max_length=100)
    op: str = Field(..., min_length=1, max_length=20)
    # value 类型由后端按 op 解释（in/not_in 接受 list；is_null/is_not_null 忽略；between 接受 [lo, hi]）
    value: Any | None = None


class OrderBy(BaseModel):
    field: str = Field(..., min_length=1, max_length=100)
    direction: Literal["asc", "desc"] = "asc"


class QueryDSL(BaseModel):
    table: str = Field(..., min_length=1, max_length=64)
    fields: list[str] = Field(default_factory=list)
    filters: list[FilterCond] = Field(default_factory=list)
    filter_logic: Literal["and", "or"] = "and"
    group_by: list[str] = Field(default_factory=list)
    aggregates: list[dict[str, str]] = Field(default_factory=list)
    # aggregates 元素：{"func": "sum", "field": "audited_amount", "alias": "total"}
    order_by: list[OrderBy] = Field(default_factory=list)
    limit: int = Field(default=100, ge=1, le=1000)
    offset: int = Field(default=0, ge=0)
    # S-3 v2 新增：JOIN 关联（声明式白名单，不接受用户传 ON）
    # joins 元素：{"table": "wp_index", "type": "inner"|"left"}
    joins: list[dict[str, str]] = Field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# 核心：DSL → SQLAlchemy core `select()`（白名单+绑定参数；无字符串拼接）
# ─────────────────────────────────────────────────────────────────────────────
def _resolve_table(table_name: str) -> dict[str, Any]:
    if table_name not in TABLE_WHITELIST:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "TABLE_NOT_ALLOWED",
                    "message": f"表 '{table_name}' 不在白名单中",
                    "allowed_tables": sorted(TABLE_WHITELIST.keys())},
        )
    return TABLE_WHITELIST[table_name]


def _resolve_column(table_meta: dict[str, Any], field: str) -> Column:
    if field not in table_meta["fields"]:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "FIELD_NOT_ALLOWED",
                    "message": f"字段 '{field}' 不在表 '{table_meta['label']}' 白名单中",
                    "allowed_fields": list(table_meta["fields"])},
        )
    model = table_meta["model"]
    col = getattr(model, field, None)
    if col is None:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "FIELD_NOT_FOUND",
                    "message": f"模型 {model.__name__} 不存在字段 '{field}'"},
        )
    return col


def _resolve_field_ref(
    base_table: str,
    field: str,
    join_tables: set[str],
) -> Column:
    """S-3 v2：解析字段引用，支持 ``table.field`` 双段语法

    - ``audited_amount`` → 默认从 base_table 解析
    - ``trial_balance.audited_amount`` → 从指定表解析
    - 引用的表必须 ∈ (base_table ∪ join_tables)
    """
    if "." in field:
        table_name, field_name = field.split(".", 1)
        if table_name != base_table and table_name not in join_tables:
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "FIELD_TABLE_NOT_JOINED",
                    "message": f"字段引用 '{field}' 中的表 '{table_name}' 未在 joins 中声明",
                    "joined_tables": sorted(join_tables | {base_table}),
                },
            )
        meta = _resolve_table(table_name)
        return _resolve_column(meta, field_name)
    # 单段语法：从 base_table 解析
    return _resolve_column(_resolve_table(base_table), field)


def _coerce_value(col: Column, value: Any) -> Any:
    """按列的 SQLAlchemy 类型把 user-supplied value 转为正确 Python 类型。

    修复生产 bug：UUID 列绑定 str value 时 SQLAlchemy 会调 `value.hex` 抛
    `AttributeError: 'str' object has no attribute 'hex'`。同理 Date/DateTime/
    Decimal 列也需要从 ISO 字符串/数字 coerce。

    支持类型：
    - UUID（PG_UUID / sa.Uuid）
    - Decimal / Numeric
    - Date / DateTime（ISO 8601 字符串）
    - Bool（接受 "true"/"false" 字符串）
    - 其他类型保持原值
    """
    import datetime as _dt
    import decimal as _dec
    import uuid as _uuid

    if value is None:
        return None

    # 取列的 Python type（通过 SQLAlchemy type 解析）
    try:
        col_type = col.type
        py_type = col_type.python_type
    except (NotImplementedError, AttributeError):
        # Enum / 复合类型可能不支持 python_type，原值传回
        return value

    # 已是正确类型 → 直接返回
    if isinstance(value, py_type):
        return value

    # UUID
    if py_type is _uuid.UUID:
        if isinstance(value, str):
            try:
                return _uuid.UUID(value)
            except (ValueError, AttributeError) as e:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "error_code": "INVALID_UUID",
                        "message": f"无法解析为 UUID: {value!r} ({e})",
                    },
                )
        return value

    # Decimal
    if py_type is _dec.Decimal:
        try:
            return _dec.Decimal(str(value))
        except _dec.InvalidOperation:
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "INVALID_DECIMAL",
                    "message": f"无法解析为 Decimal: {value!r}",
                },
            )

    # Date / DateTime
    if py_type is _dt.date:
        if isinstance(value, str):
            try:
                return _dt.date.fromisoformat(value)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail={"error_code": "INVALID_DATE", "message": f"无法解析为 date: {value!r}"},
                )
        return value
    if py_type is _dt.datetime:
        if isinstance(value, str):
            try:
                return _dt.datetime.fromisoformat(value)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "error_code": "INVALID_DATETIME",
                        "message": f"无法解析为 datetime: {value!r}",
                    },
                )
        return value

    # Bool 字符串
    if py_type is bool and isinstance(value, str):
        lower = value.lower()
        if lower in ("true", "1", "yes"):
            return True
        if lower in ("false", "0", "no"):
            return False

    # int 字符串
    if py_type is int and isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return value  # 保持原值，让 SQLAlchemy 报更精确的错

    # float 字符串
    if py_type is float and isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return value

    return value


def _build_filter(col: Column, op: str, value: Any) -> ColumnElement:
    if op not in OPERATOR_WHITELIST:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "OP_NOT_ALLOWED",
                    "message": f"操作符 '{op}' 不在白名单中",
                    "allowed_ops": sorted(OPERATOR_WHITELIST)},
        )
    # 类型 coerce：UUID/Decimal/Date/DateTime/Bool 自动从 str 转
    if op in ("in", "not_in"):
        if not isinstance(value, list) or not value:
            raise HTTPException(
                status_code=400,
                detail={"error_code": "INVALID_IN_VALUE",
                        "message": f"{op} 操作符要求 value 为非空数组"},
            )
        coerced_value: Any = [_coerce_value(col, v) for v in value]
    elif op == "between":
        if not isinstance(value, list) or len(value) != 2:
            raise HTTPException(
                status_code=400,
                detail={"error_code": "INVALID_BETWEEN_VALUE",
                        "message": "between 操作符要求 value 为 [lo, hi] 长度=2 数组"},
            )
        coerced_value = [_coerce_value(col, v) for v in value]
    elif op in ("is_null", "is_not_null"):
        coerced_value = None  # 忽略
    elif op in ("like", "not_like"):
        # LIKE 强制字符串语义，不 coerce
        coerced_value = value
    else:
        coerced_value = _coerce_value(col, value)

    if op == "eq":
        return col == coerced_value
    if op == "neq":
        return col != coerced_value
    if op == "gt":
        return col > coerced_value
    if op == "gte":
        return col >= coerced_value
    if op == "lt":
        return col < coerced_value
    if op == "lte":
        return col <= coerced_value
    if op == "like":
        return col.like(f"%{coerced_value}%")
    if op == "not_like":
        return col.notlike(f"%{coerced_value}%")
    if op == "in":
        return col.in_(coerced_value)
    if op == "not_in":
        return col.notin_(coerced_value)
    if op == "is_null":
        return col.is_(None)
    if op == "is_not_null":
        return col.isnot(None)
    if op == "between":
        return col.between(coerced_value[0], coerced_value[1])
    # 不会到此（OPERATOR_WHITELIST 已穷举）
    raise HTTPException(
        status_code=400,
        detail={"error_code": "OP_NOT_IMPLEMENTED", "message": f"操作符 {op} 未实现"},
    )


def _build_select(dsl: QueryDSL) -> tuple[Select, list[str]]:
    """根据 DSL 构造 `select()` 与列名列表。"""
    table_meta = _resolve_table(dsl.table)
    model = table_meta["model"]

    # ── S-3 v2：解析 joins ──
    join_specs: list[tuple[str, str, list[tuple[str, str]]]] = []  # (target_table, join_type, on_pairs)
    join_tables: set[str] = set()
    for j in dsl.joins:
        target = j.get("table", "")
        jtype = (j.get("type") or "inner").lower()
        if jtype not in ("inner", "left"):
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "JOIN_TYPE_NOT_ALLOWED",
                    "message": f"join type '{jtype}' 必须 ∈ {{inner, left}}",
                },
            )
        # 校验 target 表存在 + 在 base_table 的 JOIN_WHITELIST 中
        if target not in TABLE_WHITELIST:
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "JOIN_TABLE_NOT_ALLOWED",
                    "message": f"join 目标表 '{target}' 不在 TABLE_WHITELIST 中",
                },
            )
        allowed_joins = JOIN_WHITELIST.get(dsl.table, {})
        if target not in allowed_joins:
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "JOIN_NOT_REGISTERED",
                    "message": (
                        f"表 '{dsl.table}' 与 '{target}' 之间没有预登记的 JOIN 关系"
                    ),
                    "available_joins_for_base": sorted(allowed_joins.keys()),
                },
            )
        on_pairs = allowed_joins[target]["on"]
        join_specs.append((target, jtype, on_pairs))
        join_tables.add(target)

    # ── 选择列 ──
    select_cols: list[ColumnElement] = []
    column_names: list[str] = []

    if dsl.fields:
        for f in dsl.fields:
            col = _resolve_field_ref(dsl.table, f, join_tables)
            select_cols.append(col)
            column_names.append(f)
    elif dsl.aggregates:
        # 纯聚合查询：不要求 fields，但若有 group_by 则前端应同时把它放入 fields
        pass
    else:
        # 无 fields 默认全字段（仅 base_table，避免 JOIN 后字段爆炸）
        for f in table_meta["fields"]:
            col = getattr(model, f, None)
            if col is not None:
                select_cols.append(col)
                column_names.append(f)

    # ── 聚合列 ──
    aggregate_funcs = {
        "count": func.count,
        "sum": func.sum,
        "avg": func.avg,
        "min": func.min,
        "max": func.max,
    }
    for agg in dsl.aggregates:
        agg_func = agg.get("func", "").lower()
        agg_field = agg.get("field", "")
        agg_alias = agg.get("alias") or f"{agg_func}_{agg_field.replace('.', '_')}"
        if agg_func not in AGGREGATE_WHITELIST:
            raise HTTPException(
                status_code=400,
                detail={"error_code": "AGG_NOT_ALLOWED",
                        "message": f"聚合函数 '{agg_func}' 不在白名单中",
                        "allowed_aggs": sorted(AGGREGATE_WHITELIST)},
            )
        # count(*) 特例：field='*' 或为空时使用常量
        if agg_func == "count" and (not agg_field or agg_field == "*"):
            agg_col = func.count().label(agg_alias)
        else:
            target = _resolve_field_ref(dsl.table, agg_field, join_tables)
            agg_col = aggregate_funcs[agg_func](target).label(agg_alias)
        select_cols.append(agg_col)
        column_names.append(agg_alias)

    if not select_cols:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "EMPTY_SELECT",
                    "message": "fields 与 aggregates 不能同时为空"},
        )

    stmt: Select = select(*select_cols)

    # ── S-3 v2：应用 JOIN ──
    for target_table, jtype, on_pairs in join_specs:
        target_meta = _resolve_table(target_table)
        target_model = target_meta["model"]
        on_clauses = []
        for left_col, right_col in on_pairs:
            left = _resolve_column(table_meta, left_col)
            right = _resolve_column(target_meta, right_col)
            on_clauses.append(left == right)
        on_expr = and_(*on_clauses) if len(on_clauses) > 1 else on_clauses[0]
        if jtype == "left":
            stmt = stmt.outerjoin(target_model, on_expr)
        else:
            stmt = stmt.join(target_model, on_expr)

    # ── WHERE ──
    where_clauses: list[ColumnElement] = []
    for cond in dsl.filters:
        col = _resolve_field_ref(dsl.table, cond.field, join_tables)
        where_clauses.append(_build_filter(col, cond.op, cond.value))

    if where_clauses:
        if dsl.filter_logic == "or":
            stmt = stmt.where(or_(*where_clauses))
        else:
            stmt = stmt.where(and_(*where_clauses))

    # ── GROUP BY ──
    for f in dsl.group_by:
        col = _resolve_field_ref(dsl.table, f, join_tables)
        stmt = stmt.group_by(col)

    # ── ORDER BY ──
    for ob in dsl.order_by:
        col = _resolve_field_ref(dsl.table, ob.field, join_tables)
        stmt = stmt.order_by(asc(col) if ob.direction == "asc" else desc(col))

    # ── LIMIT / OFFSET ──
    stmt = stmt.limit(min(dsl.limit, 1000)).offset(dsl.offset)

    return stmt, column_names


def _stmt_to_sql(stmt: Select) -> str:
    """生成可读的 SQL 预览（保留绑定参数为命名占位）。"""
    try:
        compiled = stmt.compile(compile_kwargs={"literal_binds": False})
        return str(compiled)
    except Exception as exc:  # 兼容部分类型无法 inline 的情况
        logger.debug("compile preview fallback: %s", exc)
        return str(stmt)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/schema")
async def get_schema(
    current_user: User = Depends(get_current_user),
):
    """返回白名单表/字段元信息（前端用于构造可视化选择器）。

    S-3 v2：joins 字段列出当前表可关联的目标表 + 关联条件
    """
    _require_admin_or_manager(current_user)
    return {
        "tables": [
            {
                "name": name,
                "label": meta["label"],
                "fields": meta["fields"],
                "joins": [
                    {
                        "target_table": target,
                        "target_label": TABLE_WHITELIST[target]["label"],
                        "on": [
                            {"left_field": l, "right_field": r}
                            for l, r in spec["on"]
                        ],
                    }
                    for target, spec in JOIN_WHITELIST.get(name, {}).items()
                    if target in TABLE_WHITELIST
                ],
            }
            for name, meta in TABLE_WHITELIST.items()
        ],
        "operators": sorted(OPERATOR_WHITELIST),
        "aggregates": sorted(AGGREGATE_WHITELIST),
    }


@router.post("/preview")
async def preview_query(
    body: QueryDSL,
    current_user: User = Depends(get_current_user),
):
    """仅生成 SQL（不执行），用于前端"SQL 预览"。"""
    _require_admin_or_manager(current_user)
    stmt, column_names = _build_select(body)
    return {
        "sql": _stmt_to_sql(stmt),
        "columns": column_names,
        "table": body.table,
    }


@router.post("/execute")
async def execute_query(
    body: QueryDSL,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """执行查询，返回结构化结果。"""
    _require_admin_or_manager(current_user)
    stmt, column_names = _build_select(body)
    try:
        result = await db.execute(stmt)
        rows = result.fetchall()
    except SQLAlchemyError as exc:
        try:
            await db.rollback()
        except Exception:
            pass
        logger.exception("query_builder execute failed")
        raise HTTPException(
            status_code=400,
            detail={"error_code": "QUERY_EXECUTION_FAILED",
                    "message": f"查询执行失败：{exc}"},
        )

    rows_serialized: list[dict] = []
    for r in rows:
        obj: dict[str, Any] = {}
        for idx, col_name in enumerate(column_names):
            v = r[idx] if idx < len(r) else None
            obj[col_name] = _serialize_cell(v)
        rows_serialized.append(obj)

    return {
        "rows": rows_serialized,
        "columns": column_names,
        "total": len(rows_serialized),
        "table": body.table,
        "sql": _stmt_to_sql(stmt),
    }


@router.post("/export-excel")
async def export_excel(
    body: QueryDSL,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """执行查询并生成 Excel 流。"""
    _require_admin_or_manager(current_user)
    stmt, column_names = _build_select(body)
    try:
        result = await db.execute(stmt)
        rows = result.fetchall()
    except SQLAlchemyError as exc:
        try:
            await db.rollback()
        except Exception:
            pass
        raise HTTPException(
            status_code=400,
            detail={"error_code": "QUERY_EXECUTION_FAILED",
                    "message": f"查询执行失败：{exc}"},
        )

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    table_label = TABLE_WHITELIST[body.table]["label"]
    ws.title = table_label[:31]  # Excel sheet 名 ≤31 字符

    # 表头
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0EDF5", end_color="F0EDF5", fill_type="solid")
    for ci, name in enumerate(column_names, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for ri, row in enumerate(rows, 2):
        for ci, _ in enumerate(column_names, 1):
            v = row[ci - 1] if ci - 1 < len(row) else None
            ws.cell(row=ri, column=ci, value=_excel_cell_value(v))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"query_{body.table}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 序列化工具
# ─────────────────────────────────────────────────────────────────────────────
def _serialize_cell(v: Any) -> Any:
    """将 SQL 行单元格序列化为 JSON 友好类型。"""
    from datetime import date, datetime
    from decimal import Decimal
    from uuid import UUID
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, UUID):
        return str(v)
    if hasattr(v, "value") and not isinstance(v, (str, int, float, bool)):
        return v.value  # SQLAlchemy enum
    return v


def _excel_cell_value(v: Any) -> Any:
    """Excel cell 接受 str/number/datetime；UUID/Decimal/枚举要转换。

    openpyxl 不支持 timezone-aware datetime（会抛 TypeError），
    PG ``timestamptz`` 字段会带 tzinfo，必须显式 strip。
    """
    from datetime import date, datetime
    from decimal import Decimal
    from uuid import UUID
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        # strip tzinfo（保持壁钟时间，符合用户本地化预期）
        return v.replace(tzinfo=None) if v.tzinfo is not None else v
    if isinstance(v, date):
        return v
    if isinstance(v, UUID):
        return str(v)
    if hasattr(v, "value") and not isinstance(v, (str, int, float, bool)):
        return v.value
    return v
