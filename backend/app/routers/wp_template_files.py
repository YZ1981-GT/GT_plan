"""底稿模板文件路由

提供底稿 xlsx 文件的获取和初始化端点。
WorkpaperEditor 前端通过这些端点加载底稿 xlsx 文件供 Univer 渲染。
"""
from __future__ import annotations

import logging
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from starlette.requests import Request
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.services.wp_template_init_service import (
    find_template_file,
    get_workpaper_file,
    get_workpaper_storage_path,
    init_workpaper_from_template,
    list_available_templates,
    prefill_workpaper_xlsx,
)

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/projects/{project_id}/workpapers/{wp_id}/template-file",
    tags=["workpaper-template-files"],
)


@router.get("")
async def get_workpaper_xlsx(
    project_id: str,
    wp_id: str,
    db: AsyncSession = Depends(get_db),
):
    """获取底稿 xlsx 文件（供 Univer importXLSX 加载）

    如果底稿文件已存在，直接返回。
    如果不存在，从模板初始化后自动 prefill 试算表数据再返回。
    """
    pid = UUID(project_id)
    wid = UUID(wp_id)

    # 检查是否已有底稿文件
    existing = get_workpaper_file(pid, wid)
    if existing:
        return FileResponse(
            str(existing),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=existing.name,
        )

    # 从数据库获取 wp_code（支持多种关联方式）
    row = (await db.execute(text("""
        SELECT i.wp_code
        FROM working_paper w
        LEFT JOIN wp_index i ON w.wp_index_id = i.id
        WHERE w.id = :wid
    """), {"wid": wp_id})).first()

    if not row or not row[0]:
        raise HTTPException(status_code=404, detail="底稿不存在或无编码，无法加载模板")
    else:
        wp_code = row[0]

    # 从模板初始化
    result = init_workpaper_from_template(pid, wid, wp_code)
    if not result:
        raise HTTPException(
            status_code=404,
            detail=f"模板文件不存在: {wp_code}（系统将使用空白 Univer 编辑器）",
        )

    # P1-1: Auto-prefill on first initialization
    # Only prefill xlsx files (not xlsm/docx)
    if result.suffix.lower() == '.xlsx':
        tb_data = await _get_tb_data_for_prefill(db, pid)
        if tb_data:
            try:
                filled = prefill_workpaper_xlsx(result, wp_code, tb_data)
                if filled > 0:
                    logger.info("Auto-prefill on init: wp_code=%s, filled=%d cells", wp_code, filled)
            except Exception as e:
                logger.warning("Auto-prefill failed (non-blocking): wp_code=%s, error=%s", wp_code, e)

    return FileResponse(
        str(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=result.name,
    )


@router.post("/init")
async def init_from_template(
    project_id: str,
    wp_id: str,
    db: AsyncSession = Depends(get_db),
):
    """手动触发从模板初始化底稿文件（覆盖已有文件）"""
    pid = UUID(project_id)
    wid = UUID(wp_id)

    row = (await db.execute(text("""
        SELECT i.wp_code FROM working_paper w
        JOIN wp_index i ON w.wp_index_id = i.id
        WHERE w.id = :wid
    """), {"wid": wp_id})).first()

    if not row or not row[0]:
        raise HTTPException(status_code=404, detail="底稿不存在或无编码")

    result = init_workpaper_from_template(pid, wid, row[0])
    if not result:
        raise HTTPException(status_code=404, detail=f"模板文件不存在: {row[0]}")

    return {
        "message": f"底稿已从模板初始化: {row[0]}",
        "file_path": str(result.name),
        "size_kb": round(result.stat().st_size / 1024, 1),
    }


@router.get("/available-templates", include_in_schema=True)
async def get_available_templates():
    """列出所有可用模板（供前端展示）"""
    templates = list_available_templates()
    return {"items": templates, "total": len(templates)}


@router.post("/upload-xlsx")
async def upload_xlsx_file(
    project_id: str,
    wp_id: str,
    request: Request,
):
    """接收 Univer 导出的 xlsx blob 并覆盖保存

    P2-2: 加入文件级冲突检测——如果服务端文件的修改时间比客户端打开时更新，
    说明有其他人在此期间保存过，返回 409 提示冲突。
    前端通过 header X-File-Opened-At 传递打开时间戳。
    """
    pid = UUID(project_id)
    wid = UUID(wp_id)
    storage_path = get_workpaper_storage_path(pid, wid)

    # P2-2: 文件级冲突检测
    opened_at = request.headers.get("X-File-Opened-At")
    if opened_at and storage_path.exists():
        try:
            file_mtime = storage_path.stat().st_mtime
            client_opened = float(opened_at)
            if file_mtime > client_opened:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "error_code": "XLSX_FILE_CONFLICT",
                        "message": "底稿文件已被其他用户修改，请刷新后重试",
                        "server_mtime": file_mtime,
                        "client_opened_at": client_opened,
                    },
                )
        except (ValueError, TypeError):
            pass  # 无法解析时间戳，跳过冲突检测

    form = await request.form()
    file = form.get("file")
    if file:
        content = await file.read()
        storage_path.parent.mkdir(parents=True, exist_ok=True)
        with open(storage_path, "wb") as f:
            f.write(content)
        logger.info("xlsx saved: %s (%d bytes)", storage_path.name, len(content))
        return {"message": "xlsx 已保存", "size_kb": round(len(content) / 1024, 1)}

    raise HTTPException(status_code=400, detail="未收到文件")


@router.post("/to-json")
async def convert_xlsx_to_json(
    project_id: str,
    wp_id: str,
    request: Request,
):
    """将 xlsx 转换为 Univer workbook JSON（含格式信息）

    增强版：保留合并单元格/列宽/行高/字体/边框/底纹/公式，恢复 70-80% 原始格式。
    使用 read_only=False 模式读取完整样式信息。
    """
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="未收到文件")

    content = await file.read()
    if len(content) < 100:
        raise HTTPException(status_code=400, detail="文件内容过小")

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = load_workbook(BytesIO(content), read_only=False, data_only=False)
        sheets: dict = {}
        sheet_order: list[str] = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheet_id = f"sheet{idx}"
            sheet_order.append(sheet_id)
            cell_data: dict = {}
            merge_data: list = []
            column_data: dict = {}
            row_data_map: dict = {}

            # 合并单元格
            for merged_range in ws.merged_cells.ranges:
                merge_data.append({
                    "startRow": merged_range.min_row - 1,
                    "endRow": merged_range.max_row - 1,
                    "startColumn": merged_range.min_col - 1,
                    "endColumn": merged_range.max_col - 1,
                })

            # 列宽
            for col_letter, dim in ws.column_dimensions.items():
                if dim.width and dim.width != 8.43:  # 非默认宽度
                    col_idx = ord(col_letter.upper()) - 65 if len(col_letter) == 1 else (ord(col_letter[0].upper()) - 64) * 26 + ord(col_letter[1].upper()) - 65
                    column_data[col_idx] = {"w": int(dim.width * 7.5)}  # Excel 宽度单位转像素

            # 行高
            for row_num, dim in ws.row_dimensions.items():
                if dim.height and dim.height != 15:  # 非默认高度
                    row_data_map[row_num - 1] = {"h": int(dim.height * 1.33)}  # pt 转像素

            # 单元格数据 + 样式 + 公式
            row_count = 0
            col_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    row_idx = cell.row - 1
                    col_idx = cell.column - 1
                    row_count = max(row_count, row_idx + 1)
                    col_count = max(col_count, col_idx + 1)

                    if cell.value is None and not _has_style(cell):
                        continue

                    cell_obj: dict = {}

                    # 值
                    if cell.value is not None:
                        val = cell.value
                        if isinstance(val, str) and val.startswith("="):
                            # 公式
                            cell_obj["f"] = val
                        elif isinstance(val, (int, float)):
                            cell_obj["v"] = val
                        else:
                            cell_obj["v"] = str(val)

                    # 样式
                    style = _extract_cell_style(cell)
                    if style:
                        cell_obj["s"] = style

                    if cell_obj:
                        if row_idx not in cell_data:
                            cell_data[row_idx] = {}
                        cell_data[row_idx][col_idx] = cell_obj

            sheet_obj: dict = {
                "id": sheet_id,
                "name": sheet_name,
                "rowCount": max(row_count, 100),
                "columnCount": max(col_count, 26),
                "cellData": cell_data,
            }
            if merge_data:
                sheet_obj["mergeData"] = merge_data
            if column_data:
                sheet_obj["columnData"] = column_data
            if row_data_map:
                sheet_obj["rowData"] = row_data_map

            # 冻结窗格
            if ws.freeze_panes:
                freeze_cell = str(ws.freeze_panes)
                # 解析 "B3" → row=2, col=1
                import re as _re
                m = _re.match(r"([A-Z]+)(\d+)", freeze_cell)
                if m:
                    freeze_col = 0
                    for ch in m.group(1):
                        freeze_col = freeze_col * 26 + (ord(ch) - 64)
                    freeze_col -= 1
                    freeze_row = int(m.group(2)) - 1
                    if freeze_row > 0 or freeze_col > 0:
                        sheet_obj["freeze"] = {
                            "startRow": freeze_row,
                            "startColumn": freeze_col,
                            "xSplit": freeze_col,
                            "ySplit": freeze_row,
                        }

            # 默认列宽（如果 sheet 有设置）
            if ws.sheet_format and ws.sheet_format.defaultColWidth:
                sheet_obj["defaultColumnWidth"] = int(ws.sheet_format.defaultColWidth * 7.5)
            if ws.sheet_format and ws.sheet_format.defaultRowHeight:
                sheet_obj["defaultRowHeight"] = int(ws.sheet_format.defaultRowHeight * 1.33)

            # 条件格式
            cf_rules = _extract_conditional_formatting(ws)
            if cf_rules:
                sheet_obj["conditionalFormattingRules"] = cf_rules

            # 数据验证/下拉列表
            dv_rules = _extract_data_validations(ws)
            if dv_rules:
                sheet_obj["dataValidations"] = dv_rules

            # 图片
            images = _extract_images(ws)
            if images:
                sheet_obj["drawings"] = images

            sheets[sheet_id] = sheet_obj

        wb.close()

        workbook_json = {
            "id": f"wp-{wp_id[:8]}",
            "name": f"Workpaper {wp_id[:8]}",
            "sheetOrder": sheet_order,
            "sheets": sheets,
        }
        return JSONResponse(content=workbook_json)

    except Exception as e:
        logger.error("xlsx-to-json conversion failed: %s", e)
        raise HTTPException(status_code=500, detail=f"xlsx 转换失败: {str(e)}")


@router.get("/xlsx-to-json")
async def convert_xlsx_storage_to_json(
    project_id: str,
    wp_id: str,
    db: AsyncSession = Depends(get_db),
):
    """直接读 storage 中的 xlsx 文件转 Univer JSON。

    PoC D2 修复：前端原本走 POST /to-json + FormData 上传，但同 wp_id 文件已在
    storage 里，没必要再传一次。前端调用 GET /xlsx-to-json 即可拿到 Univer
    workbook JSON（含完整 cellData/合并/列宽/样式/公式/图片/数据验证/条件格式）。

    流程：
    1. 找 storage/{project_id}/{wp_id}.xlsx；不存在则 init_workpaper_from_template
    2. openpyxl 读取 + _has_style + _extract_cell_style + _extract_conditional_formatting + _extract_data_validations + _extract_images
    3. 返回 Univer 的 IWorkbookData JSON snapshot
    """
    from io import BytesIO

    pid = UUID(project_id)
    wid = UUID(wp_id)

    # 1. 找 storage 文件，不存在则从模板初始化
    storage_path = get_workpaper_file(pid, wid)
    if not storage_path:
        # 查 wp_code
        row = (await db.execute(text("""
            SELECT i.wp_code FROM working_paper w
            LEFT JOIN wp_index i ON w.wp_index_id = i.id
            WHERE w.id = :wid
        """), {"wid": wp_id})).first()
        if not row or not row[0]:
            raise HTTPException(status_code=404, detail="底稿不存在或无编码")
        result = init_workpaper_from_template(pid, wid, row[0])
        if not result:
            raise HTTPException(status_code=404, detail=f"模板文件不存在: {row[0]}")
        storage_path = result

    if storage_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise HTTPException(
            status_code=400,
            detail=f"非 xlsx 类底稿: {storage_path.name}（docx 走 /docx-to-json）",
        )

    # 2. 读取文件并复用 POST /to-json 的转换逻辑
    try:
        content = storage_path.read_bytes()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), read_only=False, data_only=False)
        sheets: dict = {}
        sheet_order: list[str] = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheet_id = f"sheet{idx}"
            sheet_order.append(sheet_id)
            cell_data: dict = {}
            merge_data: list = []
            column_data: dict = {}
            row_data_map: dict = {}

            for merged_range in ws.merged_cells.ranges:
                merge_data.append({
                    "startRow": merged_range.min_row - 1,
                    "endRow": merged_range.max_row - 1,
                    "startColumn": merged_range.min_col - 1,
                    "endColumn": merged_range.max_col - 1,
                })

            for col_letter, dim in ws.column_dimensions.items():
                if dim.width and dim.width != 8.43:
                    col_idx = ord(col_letter.upper()) - 65 if len(col_letter) == 1 else (ord(col_letter[0].upper()) - 64) * 26 + ord(col_letter[1].upper()) - 65
                    column_data[col_idx] = {"w": int(dim.width * 7.5)}

            for row_num, dim in ws.row_dimensions.items():
                if dim.height and dim.height != 15:
                    row_data_map[row_num - 1] = {"h": int(dim.height * 1.33)}

            row_count = 0
            col_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    row_idx = cell.row - 1
                    col_idx = cell.column - 1
                    row_count = max(row_count, row_idx + 1)
                    col_count = max(col_count, col_idx + 1)

                    if cell.value is None and not _has_style(cell):
                        continue

                    cell_obj: dict = {}
                    if cell.value is not None:
                        val = cell.value
                        if isinstance(val, str) and val.startswith("="):
                            cell_obj["f"] = val
                        elif isinstance(val, (int, float)):
                            cell_obj["v"] = val
                        else:
                            cell_obj["v"] = str(val)
                    style = _extract_cell_style(cell)
                    if style:
                        cell_obj["s"] = style
                    if cell_obj:
                        if row_idx not in cell_data:
                            cell_data[row_idx] = {}
                        cell_data[row_idx][col_idx] = cell_obj

            sheet_obj: dict = {
                "id": sheet_id,
                "name": sheet_name,
                "rowCount": max(row_count, 100),
                "columnCount": max(col_count, 26),
                "cellData": cell_data,
            }
            if merge_data:
                sheet_obj["mergeData"] = merge_data
            if column_data:
                sheet_obj["columnData"] = column_data
            if row_data_map:
                sheet_obj["rowData"] = row_data_map

            if ws.freeze_panes:
                freeze_cell = str(ws.freeze_panes)
                import re as _re
                m = _re.match(r"([A-Z]+)(\d+)", freeze_cell)
                if m:
                    freeze_col = 0
                    for ch in m.group(1):
                        freeze_col = freeze_col * 26 + (ord(ch) - 64)
                    freeze_col -= 1
                    freeze_row = int(m.group(2)) - 1
                    if freeze_row > 0 or freeze_col > 0:
                        sheet_obj["freeze"] = {
                            "startRow": freeze_row,
                            "startColumn": freeze_col,
                            "xSplit": freeze_col,
                            "ySplit": freeze_row,
                        }

            cf_rules = _extract_conditional_formatting(ws)
            if cf_rules:
                sheet_obj["conditionalFormattingRules"] = cf_rules

            dv_rules = _extract_data_validations(ws)
            if dv_rules:
                sheet_obj["dataValidations"] = dv_rules

            images = _extract_images(ws)
            if images:
                sheet_obj["drawings"] = images

            sheets[sheet_id] = sheet_obj

        wb.close()

        # ── Foundation Task 1.2: 注入 prefill_source 元数据到 cellData ──
        # 读取 prefill_formula_mapping.json，按 wp_code 匹配当前底稿
        try:
            _wp_code_row = (await db.execute(text("""
                SELECT i.wp_code FROM working_paper w
                LEFT JOIN wp_index i ON w.wp_index_id = i.id
                WHERE w.id = :wid
            """), {"wid": wp_id})).first()
            _current_wp_code = _wp_code_row[0] if _wp_code_row else None

            if _current_wp_code:
                import json as _json
                _mapping_path = Path(__file__).resolve().parent.parent.parent / "data" / "prefill_formula_mapping.json"
                if _mapping_path.exists():
                    with open(_mapping_path, "r", encoding="utf-8") as _mf:
                        _mapping_data = _json.load(_mf)

                    # 颜色映射：formula_type → 背景色
                    _SOURCE_COLOR_MAP = {
                        "TB": "#E3F2FD",
                        "TB_SUM": "#E3F2FD",
                        "TB_AUX": "#E3F2FD",
                        "AJE": "#E8F5E9",
                        "ADJ": "#E8F5E9",
                        "PREV": "#F3E5F5",
                        "WP": "#E0F7FA",
                    }

                    # 找到当前 wp_code 的所有 mapping 条目
                    for _mapping in _mapping_data.get("mappings", []):
                        if _mapping.get("wp_code") != _current_wp_code:
                            continue
                        _target_sheet = _mapping.get("sheet", "")
                        # 找到对应的 sheet_id
                        _target_sheet_id = None
                        for _sid, _sobj in sheets.items():
                            if _sobj.get("name") == _target_sheet:
                                _target_sheet_id = _sid
                                break
                        if not _target_sheet_id:
                            continue

                        _sheet_cell_data = sheets[_target_sheet_id].get("cellData", {})
                        for _cell_mapping in _mapping.get("cells", []):
                            _cell_ref = _cell_mapping.get("cell_ref", "")
                            _formula = _cell_mapping.get("formula", "")
                            _formula_type = _cell_mapping.get("formula_type", "")

                            # 跳过语义名称（非坐标），只处理有 Excel 坐标的 cell_ref
                            # 当前 mapping 使用语义名称（如"期初余额"），不是坐标
                            # 按设计文档：对有坐标的注入，语义名称跳过（future sprint）
                            # 但我们仍然注入 custom 元数据到 sheet 级别供前端 composable 使用
                            _bg_color = _SOURCE_COLOR_MAP.get(_formula_type, "#E3F2FD")

                            # 将 prefill 元数据存到 sheet 级别的 custom 字段
                            if "custom" not in sheets[_target_sheet_id]:
                                sheets[_target_sheet_id]["custom"] = {}
                            if "prefill_mappings" not in sheets[_target_sheet_id]["custom"]:
                                sheets[_target_sheet_id]["custom"]["prefill_mappings"] = []
                            sheets[_target_sheet_id]["custom"]["prefill_mappings"].append({
                                "cell_ref": _cell_ref,
                                "formula": _formula,
                                "formula_type": _formula_type,
                                "bg_color": _bg_color,
                            })
        except Exception as _prefill_err:
            logger.warning("prefill metadata injection failed (non-blocking): %s", _prefill_err)

        return JSONResponse(content={
            "id": f"wp-{wp_id[:8]}",
            "name": storage_path.stem,
            "sheetOrder": sheet_order,
            "sheets": sheets,
        })

    except Exception as e:
        logger.error("xlsx storage-to-json conversion failed: %s", e)
        raise HTTPException(status_code=500, detail=f"xlsx 转换失败: {str(e)}")


@router.get("/docx-to-json")
async def convert_docx_to_univer_doc(
    project_id: str,
    wp_id: str,
    db: AsyncSession = Depends(get_db),
):
    """将底稿 docx 模板转换为 Univer Doc IDocumentData JSON snapshot。

    R10 复盘补丁 — wp_templates 109 个 docx 类底稿前端编辑支持。

    流程：
    1. 查 wp_index.wp_code 获取模板编码
    2. find_template_file(wp_code) 找到 .docx 文件
    3. python-docx 解析 → docx_to_univer_doc_service 转 IDocumentData JSON
    4. 失败时由前端 fallback 到 mammoth → HTML → TipTap

    不同于 xlsx 链路：xlsx 走 importXLSX API，docx 必须后端先转 JSON 再前端 createUnit。
    """
    from app.services.docx_to_univer_doc_service import docx_path_to_univer_doc

    pid = UUID(project_id)
    wid = UUID(wp_id)

    # 查 wp_code
    row = (await db.execute(text("""
        SELECT i.wp_code
        FROM working_paper w
        LEFT JOIN wp_index i ON w.wp_index_id = i.id
        WHERE w.id = :wid
    """), {"wid": wp_id})).first()

    if not row or not row[0]:
        raise HTTPException(status_code=404, detail="底稿不存在或无编码")

    wp_code = row[0]

    # 找模板文件（必须是 docx/doc）
    template_path = find_template_file(wp_code)
    if not template_path:
        raise HTTPException(status_code=404, detail=f"模板文件不存在: {wp_code}")

    suffix = template_path.suffix.lower()
    if suffix not in (".docx", ".doc"):
        raise HTTPException(
            status_code=400,
            detail=f"模板不是 Word 文档: {template_path.name}（component_type 应为 univer 走 xlsx 链路）",
        )

    # .doc 旧格式 python-docx 不支持，前端必须 fallback
    if suffix == ".doc":
        raise HTTPException(
            status_code=415,
            detail=f"旧版 .doc 格式不支持直接转换，请前端走 mammoth fallback: {template_path.name}",
        )

    try:
        snapshot = docx_path_to_univer_doc(template_path, doc_id=f"wp-{wid.hex[:12]}")
        return JSONResponse(
            content={
                "wp_code": wp_code,
                "filename": template_path.name,
                "snapshot": snapshot,
            }
        )
    except Exception as e:
        logger.error("docx-to-univer-doc conversion failed: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"docx 转换失败（前端应 fallback mammoth）: {type(e).__name__}: {e}",
        )


def _has_style(cell) -> bool:
    """检查单元格是否有非默认样式"""
    try:
        if cell.font and (cell.font.bold or cell.font.italic or (cell.font.size and cell.font.size != 11)):
            return True
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and str(cell.fill.fgColor.rgb) not in ("00000000", "0"):
            return True
        if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
            return True
        if cell.number_format and cell.number_format != "General":
            return True
        if cell.alignment and (cell.alignment.wrap_text or cell.alignment.vertical != "bottom"):
            return True
    except Exception:
        pass
    return False


def _extract_cell_style(cell) -> dict | None:
    """从 openpyxl cell 提取样式转为 Univer 格式

    支持：字体/底纹/边框/对齐/数字格式/文本换行/下划线/删除线
    """
    style: dict = {}

    try:
        # 字体
        font = cell.font
        if font:
            if font.name and font.name not in ("等线", "Calibri", "宋体"):
                style["ff"] = font.name
            if font.size and font.size != 11:
                style["fs"] = int(font.size)
            if font.bold:
                style["bl"] = 1
            if font.italic:
                style["it"] = 1
            if font.underline and font.underline != "none":
                style["ul"] = 1
            if font.strikethrough:
                style["st"] = 1
            if font.color and font.color.rgb and str(font.color.rgb) not in ("00000000", "0", "FF000000"):
                rgb = str(font.color.rgb)[-6:]
                if rgb != "000000":
                    style["cl"] = {"rgb": f"#{rgb}"}

        # 底纹/填充
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if len(rgb) >= 6 and rgb[-6:] != "000000" and rgb not in ("00000000", "0"):
                style["bg"] = {"rgb": f"#{rgb[-6:]}"}

        # 边框（含线型粗细）
        border = cell.border
        if border:
            bd: dict = {}
            border_style_map = {"thin": 1, "medium": 2, "thick": 3, "double": 4, "dotted": 5, "dashed": 6}
            for side_name, side_key in [("left", "l"), ("right", "r"), ("top", "t"), ("bottom", "b")]:
                side = getattr(border, side_name, None)
                if side and side.style:
                    s_val = border_style_map.get(side.style, 1)
                    side_color = "#000000"
                    if side.color and side.color.rgb:
                        c = str(side.color.rgb)
                        if len(c) >= 6:
                            side_color = f"#{c[-6:]}"
                    bd[side_key] = {"s": s_val, "cl": {"rgb": side_color}}
            if bd:
                style["bd"] = bd

        # 对齐
        align = cell.alignment
        if align:
            # 水平对齐
            if align.horizontal:
                h_map = {"left": 0, "center": 1, "right": 2, "justify": 3, "fill": 4}
                if align.horizontal in h_map:
                    style["ht"] = h_map[align.horizontal]
            # 垂直对齐
            if align.vertical and align.vertical != "bottom":
                v_map = {"top": 0, "center": 1, "bottom": 2}
                if align.vertical in v_map:
                    style["vt"] = v_map[align.vertical]
            # 文本换行
            if align.wrap_text:
                style["tb"] = 1  # textBreak: wrap
            # 文本旋转
            if align.text_rotation and align.text_rotation != 0:
                style["tr"] = align.text_rotation

        # 数字格式
        nf = cell.number_format
        if nf and nf != "General":
            # Univer 数字格式映射
            nf_map = {
                "0": "0",
                "0.00": "0.00",
                "#,##0": "#,##0",
                "#,##0.00": "#,##0.00",
                "0%": "0%",
                "0.00%": "0.00%",
                "yyyy-mm-dd": "yyyy-mm-dd",
                "yyyy/m/d": "yyyy/m/d",
            }
            style["n"] = {"pattern": nf_map.get(nf, nf)}

    except Exception:
        pass  # 样式提取失败不阻断转换

    return style if style else None


async def _get_tb_data_for_prefill(
    db: AsyncSession,
    project_id: UUID,
) -> dict[str, dict[str, float]]:
    """P1-1: Query trial_balance data for prefill.

    Returns: {standard_account_code: {列名: 值}} dict suitable for prefill_workpaper_xlsx.
    """
    try:
        result = await db.execute(text("""
            SELECT
                tb.account_code,
                tb.opening_balance,
                tb.closing_balance,
                tb.debit_amount,
                tb.credit_amount
            FROM trial_balance tb
            WHERE tb.project_id = :pid
            LIMIT 5000
        """), {"pid": str(project_id)})
        rows = result.fetchall()
    except Exception:
        # trial_balance table may not exist or have different schema
        return {}

    if not rows:
        return {}

    tb_data: dict[str, dict[str, float]] = {}
    for row in rows:
        code = row[0]
        if not code:
            continue
        tb_data[code] = {
            "期初余额": float(row[1] or 0),
            "期末余额": float(row[2] or 0),
            "未审数": float(row[2] or 0),  # 未审数 = 期末余额
            "借方发生额": float(row[3] or 0),
            "贷方发生额": float(row[4] or 0),
        }
    return tb_data


def _extract_conditional_formatting(ws) -> list[dict]:
    """从 openpyxl worksheet 提取条件格式规则转为 Univer 格式

    支持的条件格式类型：
    - cellIs（单元格值比较：大于/小于/等于/介于）
    - colorScale（色阶）
    - dataBar（数据条）
    """
    rules = []
    try:
        for cf in ws.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                rule_obj: dict = {
                    "ranges": _parse_range_string(range_str),
                    "type": rule.type or "cellIs",
                }

                # cellIs 类型（最常见：值大于/小于某阈值时变色）
                if rule.type == "cellIs" or rule.type == "expression":
                    rule_obj["operator"] = rule.operator or "greaterThan"
                    if rule.formula:
                        rule_obj["formula"] = [str(f) for f in rule.formula]
                    # 提取格式（字体颜色/底纹）
                    if rule.dxf:
                        fmt: dict = {}
                        if rule.dxf.font and rule.dxf.font.color and rule.dxf.font.color.rgb:
                            rgb = str(rule.dxf.font.color.rgb)[-6:]
                            fmt["cl"] = {"rgb": f"#{rgb}"}
                        if rule.dxf.fill and rule.dxf.fill.fgColor and rule.dxf.fill.fgColor.rgb:
                            rgb = str(rule.dxf.fill.fgColor.rgb)[-6:]
                            fmt["bg"] = {"rgb": f"#{rgb}"}
                        if fmt:
                            rule_obj["style"] = fmt

                # colorScale 色阶
                elif rule.type == "colorScale" and rule.colorScale:
                    cs = rule.colorScale
                    rule_obj["colorScale"] = {
                        "colors": [
                            f"#{str(c.rgb)[-6:]}" if c and c.rgb else "#FFFFFF"
                            for c in (cs.color or [])
                        ],
                    }

                # dataBar 数据条
                elif rule.type == "dataBar" and rule.dataBar:
                    db = rule.dataBar
                    rule_obj["dataBar"] = {
                        "color": f"#{str(db.color.rgb)[-6:]}" if db.color and db.color.rgb else "#4472C4",
                    }

                rules.append(rule_obj)
    except Exception:
        pass  # 条件格式提取失败不阻断

    return rules


def _extract_data_validations(ws) -> list[dict]:
    """从 openpyxl worksheet 提取数据验证/下拉列表转为 Univer 格式

    支持的验证类型：
    - list（下拉列表）
    - whole/decimal（整数/小数范围）
    - date（日期范围）
    - textLength（文本长度）
    """
    validations = []
    try:
        for dv in ws.data_validations.dataValidation:
            dv_obj: dict = {
                "ranges": _parse_range_string(str(dv.sqref)),
                "type": dv.type or "list",
            }

            # 下拉列表
            if dv.type == "list":
                formula = str(dv.formula1) if dv.formula1 else ""
                if formula.startswith('"') and formula.endswith('"'):
                    # 内联列表："是,否,不适用"
                    options = formula.strip('"').split(",")
                    dv_obj["options"] = options
                else:
                    # 引用范围
                    dv_obj["formula"] = formula
                dv_obj["showDropDown"] = True

            # 数值范围
            elif dv.type in ("whole", "decimal"):
                dv_obj["operator"] = dv.operator or "between"
                if dv.formula1:
                    dv_obj["min"] = str(dv.formula1)
                if dv.formula2:
                    dv_obj["max"] = str(dv.formula2)

            # 日期范围
            elif dv.type == "date":
                dv_obj["operator"] = dv.operator or "between"
                if dv.formula1:
                    dv_obj["min"] = str(dv.formula1)
                if dv.formula2:
                    dv_obj["max"] = str(dv.formula2)

            # 文本长度
            elif dv.type == "textLength":
                dv_obj["operator"] = dv.operator or "lessThanOrEqual"
                if dv.formula1:
                    dv_obj["maxLength"] = str(dv.formula1)

            # 通用属性
            if dv.error:
                dv_obj["errorMessage"] = dv.error
            if dv.errorTitle:
                dv_obj["errorTitle"] = dv.errorTitle
            if dv.prompt:
                dv_obj["inputMessage"] = dv.prompt
            if dv.promptTitle:
                dv_obj["inputTitle"] = dv.promptTitle
            dv_obj["allowBlank"] = bool(dv.allow_blank)

            validations.append(dv_obj)
    except Exception:
        pass  # 数据验证提取失败不阻断

    return validations


def _parse_range_string(range_str: str) -> list[dict]:
    """将 Excel 范围字符串（如 'A1:D10' 或 'B2:B100 D2:D100'）转为 Univer ranges 数组"""
    import re as _re
    ranges = []
    # 分割多个范围（空格或逗号分隔）
    parts = _re.split(r"[,\s]+", range_str.strip())
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 解析 A1:D10 格式
        m = _re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", part, _re.IGNORECASE)
        if m:
            start_col = _col_letter_to_idx(m.group(1))
            start_row = int(m.group(2)) - 1
            end_col = _col_letter_to_idx(m.group(3))
            end_row = int(m.group(4)) - 1
            ranges.append({
                "startRow": start_row,
                "endRow": end_row,
                "startColumn": start_col,
                "endColumn": end_col,
            })
        else:
            # 单个单元格 A1
            m2 = _re.match(r"([A-Z]+)(\d+)", part, _re.IGNORECASE)
            if m2:
                col = _col_letter_to_idx(m2.group(1))
                row = int(m2.group(2)) - 1
                ranges.append({
                    "startRow": row, "endRow": row,
                    "startColumn": col, "endColumn": col,
                })
    return ranges


def _col_letter_to_idx(letters: str) -> int:
    """将列字母转为 0-based 索引（A=0, B=1, ..., AA=26）"""
    result = 0
    for ch in letters.upper():
        result = result * 26 + (ord(ch) - 64)
    return result - 1


def _extract_images(ws) -> list[dict]:
    """从 openpyxl worksheet 提取嵌入图片转为 Univer drawing 格式

    提取图片的 base64 数据、位置（锚点单元格）、尺寸。
    Univer 通过 UniverSheetsDrawingPreset 渲染图片。
    """
    import base64
    drawings = []
    try:
        for img in ws._images:
            drawing: dict = {
                "type": "image",
                "width": img.width or 200,
                "height": img.height or 150,
            }

            # 位置（锚点）
            if hasattr(img, 'anchor') and img.anchor:
                anchor = img.anchor
                if hasattr(anchor, '_from'):
                    drawing["anchorRow"] = anchor._from.row
                    drawing["anchorCol"] = anchor._from.col
                elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                    drawing["anchorRow"] = anchor.row
                    drawing["anchorCol"] = anchor.col

            # 图片数据（base64）
            if hasattr(img, '_data') and img._data:
                img_data = img._data()
                if img_data:
                    b64 = base64.b64encode(img_data).decode('ascii')
                    # 推断 MIME 类型
                    mime = "image/png"
                    if hasattr(img, 'format') and img.format:
                        mime = f"image/{img.format.lower()}"
                    drawing["src"] = f"data:{mime};base64,{b64}"
            elif hasattr(img, 'ref') and img.ref:
                # 引用路径（相对于 xlsx 包内）
                drawing["ref"] = str(img.ref)

            drawings.append(drawing)
    except Exception:
        pass  # 图片提取失败不阻断

    return drawings
