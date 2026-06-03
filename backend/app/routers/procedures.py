"""审计程序裁剪与委派 API 路由

Phase 9 Task 9.12
"""

from __future__ import annotations

from uuid import UUID

from fastapi import APIRouter, Depends, Query
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.services.procedure_service import ProcedureService

router = APIRouter(prefix="/api/projects", tags=["procedures"])


class TrimItem(BaseModel):
    id: str
    status: str  # execute / skip / not_applicable
    skip_reason: str | None = None


class TrimRequest(BaseModel):
    items: list[TrimItem]


class CustomProcedureRequest(BaseModel):
    procedure_name: str
    procedure_code: str | None = None
    sort_order: int | None = None


class AssignRequest(BaseModel):
    assignments: list[dict]  # [{procedure_id, staff_id}]


class BatchApplyRequest(BaseModel):
    target_project_ids: list[str]


@router.get("/{project_id}/procedures/{cycle}")
async def get_procedures(
    project_id: UUID, cycle: str,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    return await svc.get_procedures(project_id, cycle)


@router.post("/{project_id}/procedures/{cycle}/init")
async def init_procedures(
    project_id: UUID, cycle: str,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    result = await svc.init_from_templates(project_id, cycle)
    await db.commit()
    return {"count": len(result), "procedures": result}


@router.put("/{project_id}/procedures/{cycle}/trim")
async def save_trim(
    project_id: UUID, cycle: str, data: TrimRequest,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    count = await svc.save_trim(project_id, cycle, [i.model_dump() for i in data.items])

    # ── Phase 15: 发布裁剪事件到事件总线 ──
    try:
        from app.services.task_event_bus import task_event_bus
        from app.services.trace_event_service import generate_trace_id
        for item in data.items:
            if item.status in ("skip", "not_applicable"):
                event_type = "trim_applied"
            else:
                event_type = "trim_rollback"
            await task_event_bus.publish(
                db=db,
                project_id=project_id,
                event_type=event_type,
                task_node_id=None,
                payload={
                    "procedure_id": item.id,
                    "cycle": cycle,
                    "status": item.status,
                    "skip_reason": item.skip_reason,
                    "ref_id": item.id,
                    "version": "1",
                },
                trace_id=generate_trace_id(),
            )
    except Exception as _evt_err:
        import logging
        logging.getLogger(__name__).warning(f"[EVENT_BUS] trim event publish failed: {_evt_err}")

    # ── Phase 14: trace 留痕 ──
    try:
        from app.services.trace_event_service import trace_event_service, generate_trace_id as _gen_tid
        await trace_event_service.write(
            db=db,
            project_id=project_id,
            event_type="trim_applied",
            object_type="procedure",
            object_id=project_id,
            actor_id=user.id,
            action=f"save_trim:{cycle}:{count}_items",
            trace_id=_gen_tid(),
        )
    except Exception:
        pass

    await db.commit()
    return {"updated": count}


@router.post("/{project_id}/procedures/{cycle}/custom")
async def add_custom(
    project_id: UUID, cycle: str, data: CustomProcedureRequest,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    result = await svc.add_custom(project_id, cycle, data.model_dump())
    await db.commit()
    return result


@router.post("/{project_id}/procedures/{cycle}/custom-with-template")
async def add_custom_with_template(
    project_id: UUID,
    cycle: str,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
    procedure_name: str = Query(...),
    procedure_code: str | None = Query(None),
):
    """新增自定义程序 + 上传底稿模板文件（multipart）.

    文件通过后续的 upload 步骤上传，此接口先创建程序记录和 wp_index 占位。
    """
    from app.models.workpaper_models import WpIndex, WpStatus

    svc = ProcedureService(db)

    # 生成 wp_code（如 CUSTOM-D-001）
    import sqlalchemy as sa
    from app.models.procedure_models import ProcedureInstance
    existing_count = (await db.execute(
        sa.select(sa.func.count()).select_from(ProcedureInstance).where(
            ProcedureInstance.project_id == project_id,
            ProcedureInstance.audit_cycle == cycle,
            ProcedureInstance.is_custom == True,  # noqa: E712
            ProcedureInstance.is_deleted == False,  # noqa: E712
        )
    )).scalar() or 0
    wp_code = procedure_code or f"{cycle}-C{existing_count + 1:02d}"

    # 创建 ProcedureInstance
    result = await svc.add_custom(project_id, cycle, {
        "procedure_name": procedure_name,
        "procedure_code": wp_code,
        "wp_code": wp_code,
    })

    # 创建 wp_index 占位记录（后续上传文件时关联）
    wp_idx = WpIndex(
        project_id=project_id,
        wp_code=wp_code,
        wp_name=procedure_name,
        audit_cycle=cycle,
        status=WpStatus.not_started,
    )
    db.add(wp_idx)
    await db.flush()

    await db.commit()
    return {
        **result,
        "wp_index_id": str(wp_idx.id),
        "wp_code": wp_code,
        "message": f"已创建自定义程序 {wp_code}，可在底稿列表中上传模板文件",
    }


@router.get("/{project_id}/procedures/{cycle}/blank-template")
async def download_blank_template(
    project_id: UUID,
    cycle: str,
    procedure_name: str = Query("自定义底稿"),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """生成并下载空白底稿模板 Excel."""
    from io import BytesIO
    from fastapi import HTTPException
    from fastapi.responses import StreamingResponse
    import sqlalchemy as sa
    import logging

    log = logging.getLogger(__name__)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl 未安装")

    try:
        from app.models.audit_platform_models import TrialBalance, AccountChart, AccountSource
    except ImportError:
        AccountSource = None  # type: ignore

    CYCLE_NAMES = {
        'A': '报表与调整', 'B': '风险评估', 'C': '控制测试', 'D': '销售收入',
        'E': '货币资金', 'F': '采购存货', 'G': '投资', 'H': '固定资产',
        'I': '无形资产', 'J': '职工薪酬', 'K': '管理费用', 'L': '筹资',
        'M': '股东权益', 'N': '税费', 'S': '专项程序',
    }
    cycle_name = CYCLE_NAMES.get(cycle.upper(), cycle)

    # 获取项目年度（防御性）
    from datetime import datetime
    year = datetime.now().year - 1
    entity_name = ""
    period_end = ""
    preparer_name = ""
    reviewer_name = ""
    try:
        from app.models.core import Project
        proj = (await db.execute(sa.select(Project).where(Project.id == project_id))).scalar_one_or_none()
        if proj:
            entity_name = getattr(proj, "name", "") or ""
            pe = getattr(proj, "audit_period_end", None)
            if pe:
                period_end = str(pe)[:10]
                year = int(str(pe)[:4])
            ws = getattr(proj, 'wizard_state', None)
            if isinstance(ws, dict):
                y = ws.get("steps", {}).get("basic_info", {}).get("data", {}).get("audit_year")
                if y:
                    year = int(y)
    except Exception as e:
        log.warning("blank-template: 获取项目信息失败: %s", e)

    # 获取编制人/复核人（项目派单人员）
    try:
        staff_q = sa.text("""
            SELECT pa.role, s.name FROM project_assignments pa
            JOIN staff_members s ON s.id = pa.staff_id
            WHERE pa.project_id = :pid AND pa.role IN ('preparer','auditor','reviewer','manager')
        """)
        staff_rows = (await db.execute(staff_q, {"pid": str(project_id)})).all()
        for role, name in staff_rows:
            if role in ("preparer", "auditor") and not preparer_name:
                preparer_name = name or ""
            elif role in ("reviewer", "manager") and not reviewer_name:
                reviewer_name = name or ""
    except Exception as e:
        log.warning("blank-template: 获取编制/复核人失败: %s", e)

    # 获取试算表余额
    tb_rows = []
    try:
        tb_result = await db.execute(
            sa.select(
                TrialBalance.standard_account_code,
                TrialBalance.unadjusted_amount,
                TrialBalance.audited_amount,
                TrialBalance.aje_adjustment,
                TrialBalance.rje_adjustment,
            ).where(
                TrialBalance.project_id == project_id,
                TrialBalance.year == year,
                TrialBalance.is_deleted == False,  # noqa: E712
            )
        )
        tb_rows = list(tb_result.all())
    except Exception as e:
        log.warning("blank-template: 获取试算表失败: %s", e)

    # 获取科目表
    all_accounts = []
    try:
        chart_result = await db.execute(
            sa.select(AccountChart.account_code, AccountChart.account_name, AccountChart.source).where(
                AccountChart.project_id == project_id,
                AccountChart.is_deleted == False,  # noqa: E712
            ).order_by(AccountChart.account_code)
        )
        all_accounts = list(chart_result.all())
    except Exception as e:
        log.warning("blank-template: 获取科目表失败: %s", e)

    # 生成 Excel
    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
    )

    # ─── Sheet 1: 编制要求 ───
    ws1 = wb.active
    ws1.title = "编制要求"
    requirements = [
        (f"📋 {procedure_name} — 编制要求", "header"),
        ("", ""),
        (f"审计循环：{cycle} {cycle_name}", ""),
        (f"审计年度：{year}", ""),
        ("", ""),
        ("一、底稿编制目标", "section"),
        ("记录审计程序的执行过程、获取的审计证据和形成的审计结论。", ""),
        ("", ""),
        ("二、编制要求", "section"),
        ("1. 「数据表」顶部为编制信息表头（被审计单位/编制人/复核人/截止日/索引号），已自动配齐，编制/复核人及日期请签署", ""),
        ("2. 在表头下方的明细区填写审计程序执行结果", ""),
        ("3. 科目余额已从试算表预填（未审数/审定数），请勿修改", ""),
        ("4. 需要填写的列：审计程序描述、执行结果、审计结论、发现问题", ""),
        ("5. 如有调整建议，请在「调整建议」列填写金额和方向", ""),
        ("", ""),
        ("三、数据表字段说明", "section"),
        ("• 科目编码：标准科目编码（自动填充）", ""),
        ("• 科目名称：科目中文名称（自动填充）", ""),
        ("• 未审余额：试算表未审数（自动填充，勿改）", ""),
        ("• 审定余额：试算表审定数（自动填充，勿改）", ""),
        ("• 审计程序：描述执行的具体审计程序（必填）", ""),
        ("• 执行结果：程序执行的具体结果和发现（必填）", ""),
        ("• 审计结论：对该科目的审计结论（必填）", ""),
        ("• 调整建议：如需调整，填写建议调整金额（选填）", ""),
        ("• 备注：其他需要说明的事项（选填）", ""),
        ("", ""),
        ("四、注意事项", "section"),
        ("• 完成编辑后通过底稿列表的「上传」功能上传本文件", ""),
        ("• 系统会自动识别数据表中的内容并入库", ""),
        ("• 请勿修改表头结构和 sheet 名称", ""),
    ]
    title_font = Font(bold=True, size=14, color="6750A4")
    section_font = Font(bold=True, size=11, color="6750A4")
    for i, (text, style) in enumerate(requirements, 1):
        c = ws1.cell(row=i, column=1, value=text)
        if style == "header":
            c.font = title_font
        elif style == "section":
            c.font = section_font
        else:
            c.font = Font(size=10, color="333333")
    ws1.column_dimensions["A"].width = 60

    # ─── Sheet 2: 数据表 ───
    ws2 = wb.create_sheet("数据表")
    header_fill = PatternFill(start_color="F4F0FA", end_color="F4F0FA", fill_type="solid")
    label_font = Font(bold=True, size=10, color="4B2D77")

    # ── 编制信息表头（自动配齐，从项目信息填充）──
    # 索引号 = procedure_code（自定义程序编码）；空值留「—」供编制人手填
    prep_info = [
        ("被审计单位", entity_name or "—", "截止日", period_end or "—"),
        ("编制人", preparer_name or "—", "编制日期", "—"),
        ("复核人", reviewer_name or "—", "复核日期", "—"),
        ("索引号", "—", "审计循环", f"{cycle} {cycle_name}"),
    ]
    for r, (l1, v1, l2, v2) in enumerate(prep_info, 1):
        c1 = ws2.cell(row=r, column=1, value=l1); c1.font = label_font; c1.fill = header_fill; c1.border = thin_border
        c2 = ws2.cell(row=r, column=2, value=v1); c2.border = thin_border
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c3 = ws2.cell(row=r, column=5, value=l2); c3.font = label_font; c3.fill = header_fill; c3.border = thin_border
        c4 = ws2.cell(row=r, column=6, value=v2); c4.border = thin_border
        ws2.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)

    # ── 审计程序明细表头（编制信息下方）──
    data_header_row = len(prep_info) + 2  # 空 1 行
    headers = ["科目编码", "科目名称", "未审余额", "审定余额", "审计程序", "执行结果", "审计结论", "调整建议", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=data_header_row, column=col, value=h)
        c.font = Font(bold=True, size=11)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # 预填试算表余额
    row_idx = data_header_row + 1
    for code, unadj, audited, aje, rje in tb_rows:
        if not code:
            continue
        # 查找科目名称
        acct_name = ""
        for ac, an, _ in all_accounts:
            if ac == code:
                acct_name = an or ""
                break
        ws2.cell(row=row_idx, column=1, value=code).border = thin_border
        ws2.cell(row=row_idx, column=2, value=acct_name).border = thin_border
        c3 = ws2.cell(row=row_idx, column=3, value=float(unadj or 0))
        c3.border = thin_border
        c3.number_format = '#,##0.00'
        c4 = ws2.cell(row=row_idx, column=4, value=float(audited or 0))
        c4.border = thin_border
        c4.number_format = '#,##0.00'
        for col in range(5, 10):
            ws2.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    # 如果没有试算表数据，留 20 行空行
    if row_idx == data_header_row + 1:
        for r in range(data_header_row + 1, data_header_row + 21):
            for col in range(1, 10):
                ws2.cell(row=r, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 30
    ws2.column_dimensions["F"].width = 30
    ws2.column_dimensions["G"].width = 20
    ws2.column_dimensions["H"].width = 14
    ws2.column_dimensions["I"].width = 20
    # 冻结编制信息表头 + 明细表头（数据从 data_header_row+1 起滚动）
    ws2.freeze_panes = f"A{data_header_row + 1}"

    # ─── Sheet 3: 科目参考 ───
    ws3 = wb.create_sheet("科目参考")
    ref_headers = ["科目编码", "科目名称", "来源"]
    for col, h in enumerate(ref_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill(start_color="6750A4", end_color="6750A4", fill_type="solid")
        c.border = thin_border
    for ri, (code, name, source) in enumerate(all_accounts, 2):
        ws3.cell(row=ri, column=1, value=code).border = thin_border
        ws3.cell(row=ri, column=2, value=name).border = thin_border
        ws3.cell(row=ri, column=3, value="客户" if (AccountSource and source == AccountSource.client) else ("客户" if str(source) == "client" else "标准")).border = thin_border
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 10
    ws3.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"blank_template_{cycle}_{procedure_name}.xlsx"
    # 中文文件名需 RFC5987 编码（HTTP 头按 latin-1，直接放中文会 UnicodeEncodeError）
    from urllib.parse import quote
    ascii_name = filename.encode("ascii", "ignore").decode() or "blank_template.xlsx"
    utf8_name = quote(filename, safe="")
    disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


@router.put("/{project_id}/procedures/assign")
async def assign_procedures(
    project_id: UUID, data: AssignRequest,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    count = await svc.assign_procedures(project_id, data.assignments)
    await db.commit()
    return {"assigned": count}


@router.get("/{project_id}/procedures/{cycle}/trim-scheme")
async def get_trim_scheme(
    project_id: UUID, cycle: str,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    return await svc.get_trim_scheme(project_id, cycle) or {}


@router.post("/{project_id}/procedures/{cycle}/apply-scheme")
async def apply_scheme(
    project_id: UUID, cycle: str,
    source_project_id: UUID = Query(...),
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    count = await svc.apply_scheme(project_id, cycle, source_project_id)
    await db.commit()
    return {"applied": count}


@router.post("/{project_id}/procedures/{cycle}/batch-apply")
async def batch_apply(
    project_id: UUID, cycle: str, data: BatchApplyRequest,
    db: AsyncSession = Depends(get_db), user=Depends(get_current_user),
):
    svc = ProcedureService(db)
    result = await svc.batch_apply(project_id, cycle, [UUID(t) for t in data.target_project_ids])
    await db.commit()
    return result


class ExecutionStatusUpdate(BaseModel):
    execution_status: str  # not_started / in_progress / completed


@router.put("/{project_id}/procedures/instance/{procedure_id}/execution")
async def update_execution_status(
    project_id: UUID,
    procedure_id: UUID,
    data: ExecutionStatusUpdate,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """更新程序执行状态（审计助理标记进度）"""
    from app.models.procedure_models import ProcedureInstance
    import sqlalchemy as sa

    result = await db.execute(
        sa.select(ProcedureInstance).where(
            ProcedureInstance.id == procedure_id,
            ProcedureInstance.project_id == project_id,
        )
    )
    proc = result.scalar_one_or_none()
    if not proc:
        from fastapi import HTTPException
        raise HTTPException(404, "程序不存在")

    proc.execution_status = data.execution_status
    await db.flush()
    await db.commit()
    return {"id": str(proc.id), "execution_status": proc.execution_status}
