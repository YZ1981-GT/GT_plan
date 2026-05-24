"""审计调整分录 API

覆盖：
- GET  列表（支持 type/status 筛选）
- POST 创建
- PUT  修改
- DELETE 软删除
- POST review 变更状态
- GET  summary 汇总统计
- GET  account-dropdown 科目下拉
- GET  wp-summary/{wp_code} 底稿审定表数据
"""

from __future__ import annotations

import logging
from uuid import UUID

import sqlalchemy as sa
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.bulk_operations import BulkRequest, bulk_execute
from app.core.database import get_db
from app.core.field_selection import parse_fields, BLOCKED_FIELDS
from app.core.pagination import PaginationParams
from app.deps import get_current_user, check_consol_lock, require_project_access, get_user_scope_cycles
from app.models.core import User
from app.models.audit_platform_models import (
    Adjustment,
    AdjustmentType,
    ReviewStatus,
    UnadjustedMisstatement,
)
from app.models.audit_platform_schemas import (
    AccountOption,
    AdjustmentCreate,
    AdjustmentSummary,
    AdjustmentUpdate,
    ReviewStatusChange,
    WPAdjustmentSummary,
)
from app.services.adjustment_service import AdjustmentService
from app.services.adjustment_impact_service import preview_impact as preview_impact_service
from app.services.mapping_service import get_codes_by_cycles
from app.services.misstatement_service import UnadjustedMisstatementService

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/projects/{project_id}/adjustments",
    tags=["adjustments"],
)


@router.get("")
async def list_adjustments(
    project_id: UUID,
    year: int = Query(...),
    adjustment_type: AdjustmentType | None = Query(None),
    review_status: ReviewStatus | None = Query(None),
    fields: str | None = Query(None, description="逗号分隔的字段名，如 id,adjustment_no,description"),
    pagination: PaginationParams = Depends(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """分录列表（支持 type/status 筛选，需项目成员权限）"""
    svc = AdjustmentService(db)
    result = await svc.list_entries(
        project_id, year,
        adjustment_type=adjustment_type,
        review_status=review_status,
        page=pagination.page, page_size=pagination.page_size,
    )

    # scope_cycles 过滤：非 admin/partner 用户只能看到被分配循环对应的科目
    try:
        scope_cycles = await get_user_scope_cycles(current_user, project_id, db)
        if scope_cycles is not None:
            allowed_codes = await get_codes_by_cycles(project_id, scope_cycles)
            if isinstance(result, dict) and "items" in result:
                result["items"] = [
                    e for e in result["items"]
                    if any(
                        li.get("standard_account_code") in allowed_codes
                        for li in (e.get("line_items") or [])
                    )
                ]
                result["total"] = len(result["items"])
    except Exception:
        pass  # scope filtering failure should not block the response

    # 字段选择：过滤返回字段（仅过滤 items 内的字段，保留分页元数据）
    requested_fields = parse_fields(fields)
    if requested_fields is not None and isinstance(result, dict) and "items" in result:
        allowed = requested_fields - BLOCKED_FIELDS
        allowed.add("id")
        result["items"] = [
            {k: v for k, v in item.items() if k in allowed}
            for item in result["items"]
        ]

    return result


@router.post("")
async def create_adjustment(
    project_id: UUID,
    data: AdjustmentCreate,
    batch_mode: bool = Query(False, description="批量模式：暂不触发重算事件"),
    db: AsyncSession = Depends(get_db),
    user=Depends(require_project_access("edit")),
    _lock_check=Depends(check_consol_lock),
):
    """创建调整分录（合并锁定期间禁止，需编辑权限）"""
    svc = AdjustmentService(db)
    try:
        result = await svc.create_entry(project_id, data, user.id, batch_mode=batch_mode)
        await db.commit()
        return result.model_dump()
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/batch-commit")
async def batch_commit(
    project_id: UUID,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user=Depends(require_project_access("edit")),
):
    """批量提交：统一触发一次重算事件（配合 batch_mode=true 使用）"""
    svc = AdjustmentService(db)
    result = await svc.batch_commit(project_id, year)
    return result


# ---------------------------------------------------------------------------
# L-2 task 2.1: 调整分录影响预览（不写 DB）
# ---------------------------------------------------------------------------


class PreviewLineItem(BaseModel):
    """preview-impact 单条 line_item（兼容多种字段名）"""
    model_config = {"extra": "ignore"}

    account_code: str | None = Field(default=None, description="标准科目编码")
    standard_account_code: str | None = Field(default=None, description="兼容字段：等同 account_code")
    debit: float | None = Field(default=0)
    credit: float | None = Field(default=0)
    debit_amount: float | None = Field(default=None)
    credit_amount: float | None = Field(default=None)


class AdjustmentPreviewRequest(BaseModel):
    line_items: list[PreviewLineItem] = Field(min_length=1)
    year: int | None = None


@router.post("/preview-impact")
async def preview_adjustment_impact(
    project_id: UUID,
    body: AdjustmentPreviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """模拟调整分录影响（**不写 DB**）

    返回受影响的报表行（按 report_type/row_code 聚合 delta）+ 受影响底稿 wp_code 列表。
    用于前端 AdjustmentImpactPreview 弹窗的实时预览（debounce 500ms 触发）。

    Validates: proposal-remaining-18 §二 L-2，design.md ADR-2
    """
    raw_items: list[dict] = []
    for it in body.line_items:
        d = it.model_dump(exclude_none=False)
        if d.get("debit_amount") is not None:
            d["debit"] = d["debit_amount"]
        if d.get("credit_amount") is not None:
            d["credit"] = d["credit_amount"]
        raw_items.append(d)

    result = await preview_impact_service(
        db=db,
        project_id=project_id,
        line_items=raw_items,
        year=body.year,
    )
    serialized_rows = [
        {
            "report_type": r["report_type"],
            "row_code": r["row_code"],
            "row_name": r["row_name"],
            "field": r["field"],
            "delta": str(r["delta"]),
        }
        for r in result["affected_report_rows"]
    ]
    return {
        "affected_report_rows": serialized_rows,
        "affected_workpapers": result["affected_workpapers"],
        "unmapped_accounts": result["unmapped_accounts"],
    }


@router.put("/{entry_group_id}")
async def update_adjustment(
    project_id: UUID,
    entry_group_id: UUID,
    data: AdjustmentUpdate,
    db: AsyncSession = Depends(get_db),
    user=Depends(require_project_access("edit")),
    _lock_check=Depends(check_consol_lock),
):
    """修改调整分录（合并锁定期间禁止，需编辑权限）"""
    svc = AdjustmentService(db)
    try:
        result = await svc.update_entry(project_id, entry_group_id, data, user.id)
        await db.commit()
        return result.model_dump()
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{entry_group_id}")
async def delete_adjustment(
    project_id: UUID,
    entry_group_id: UUID,
    db: AsyncSession = Depends(get_db),
    _lock_check=Depends(check_consol_lock),
    current_user: User = Depends(require_project_access("edit")),
):
    """软删除调整分录（合并锁定期间禁止，需编辑权限）"""
    svc = AdjustmentService(db)
    try:
        await svc.delete_entry(project_id, entry_group_id)
        await db.commit()
        return {"message": "删除成功"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/batch-delete")
async def batch_delete_adjustments(
    project_id: UUID,
    body: BulkRequest,
    db: AsyncSession = Depends(get_db),
    _lock_check=Depends(check_consol_lock),
    current_user: User = Depends(require_project_access("edit")),
):
    """批量软删除调整分录（合并锁定期间禁止，需编辑权限）"""
    svc = AdjustmentService(db)

    async def _delete_one(_db, _row):
        # 复用 service 层的删除逻辑（含状态校验+事件发布）
        await svc.delete_entry(project_id, _row.id)

    from app.models.audit_platform_models import Adjustment
    result = await bulk_execute(db, Adjustment, body.ids, _delete_one)
    await db.commit()
    return result


@router.post("/{entry_group_id}/review")
async def review_adjustment(
    project_id: UUID,
    entry_group_id: UUID,
    change: ReviewStatusChange,
    db: AsyncSession = Depends(get_db),
    user=Depends(require_project_access("review")),
):
    """变更复核状态（需复核权限）"""
    svc = AdjustmentService(db)
    try:
        await svc.change_review_status(
            project_id, entry_group_id, change, user.id
        )
        await db.commit()
        return {"message": "状态变更成功"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ---------------------------------------------------------------------------
# R1 需求 3 — AJE 一键转错报
# ---------------------------------------------------------------------------


class ConvertToMisstatementRequest(BaseModel):
    """将 rejected AJE 组转为未更正错报的请求体。"""

    force: bool = Field(
        default=False,
        description="当该 AJE 组已存在对应错报时是否强制再建一条（默认否，返回 409）",
    )


@router.post("/{entry_group_id}/convert-to-misstatement")
async def convert_adjustment_to_misstatement(
    project_id: UUID,
    entry_group_id: UUID,
    body: ConvertToMisstatementRequest | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("edit")),
):
    """将被驳回的 AJE 组一键转为《未更正错报汇总表》的一条记录。

    对齐 R1 需求 3 验收 7 与 gate 规则 `R1-AJE-UNCONVERTED` 的 suggested_action。
    封装 `UnadjustedMisstatementService.create_from_rejected_aje`，并补：
      - 组存在性 / 项目归属校验（400）
      - 幂等：已转过则返回 409 `ALREADY_CONVERTED`（可通过 force=true 强制再建）
      - year 从 group 内实际分录推断（不信任请求）
      - 审计日志 `adjustment.converted_to_misstatement`（失败仅 warning）
    """
    force = bool(body.force) if body else False

    # 1) 校验 group 存在且属于该项目
    group_rows = (
        await db.execute(
            sa.select(Adjustment.id, Adjustment.year, Adjustment.adjustment_type)
            .where(
                Adjustment.project_id == project_id,
                Adjustment.entry_group_id == entry_group_id,
                Adjustment.is_deleted.is_(False),
            )
        )
    ).all()
    if not group_rows:
        raise HTTPException(
            status_code=400,
            detail="AJE 组不存在或不属于该项目",
        )

    adj_ids = [r.id for r in group_rows]
    # year 取该组所有行 year 的一致值；如不一致打 warning 并取第一行
    years = {r.year for r in group_rows}
    if len(years) > 1:
        logger.warning(
            "[convert_to_misstatement] group=%s has inconsistent years %s, using first row",
            entry_group_id, sorted(years),
        )
    year = group_rows[0].year

    # 2) 幂等检查：查该组内任一 adjustment_id 是否已被错报引用
    existing_ms_row = (
        await db.execute(
            sa.select(UnadjustedMisstatement.id)
            .where(
                UnadjustedMisstatement.project_id == project_id,
                UnadjustedMisstatement.source_adjustment_id.in_(adj_ids),
                UnadjustedMisstatement.is_deleted.is_(False),
            )
            .limit(1)
        )
    ).first()

    if existing_ms_row and not force:
        raise HTTPException(
            status_code=409,
            detail={
                "error_code": "ALREADY_CONVERTED",
                "message": "该 AJE 组已转为未更正错报，如需再次转换请传入 force=true",
                "existing_misstatement_id": str(existing_ms_row.id),
            },
        )

    # 3) 调 service 落库
    svc = UnadjustedMisstatementService(db)
    try:
        result = await svc.create_from_rejected_aje(
            project_id=project_id,
            entry_group_id=entry_group_id,
            year=year,
            created_by=current_user.id,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 4) 审计日志（失败仅 warning 不阻断）
    try:
        from app.services.audit_logger_enhanced import audit_logger

        await audit_logger.log_action(
            user_id=current_user.id,
            action="adjustment.converted_to_misstatement",
            object_type="adjustment_group",
            object_id=entry_group_id,
            project_id=project_id,
            details={
                "misstatement_id": str(result.id),
                "source_adjustment_id": str(result.source_adjustment_id)
                if result.source_adjustment_id
                else None,
                "adjustment_count": len(adj_ids),
                "net_amount": str(result.misstatement_amount),
                "misstatement_type": result.misstatement_type.value
                if hasattr(result.misstatement_type, "value")
                else str(result.misstatement_type),
                "force": force,
                "year": year,
            },
        )
    except Exception as log_err:  # noqa: BLE001
        logger.warning(
            "[convert_to_misstatement] audit log failed (non-blocking): %s",
            log_err,
        )

    await db.commit()

    return {
        "misstatement_id": str(result.id),
        "source_entry_group_id": str(entry_group_id),
        "source_adjustment_id": str(result.source_adjustment_id)
        if result.source_adjustment_id
        else None,
        "net_amount": str(result.misstatement_amount),
        "misstatement_type": result.misstatement_type.value
        if hasattr(result.misstatement_type, "value")
        else str(result.misstatement_type),
        "year": year,
        "adjustment_count": len(adj_ids),
        "created_at": result.created_at.isoformat() if result.created_at else None,
    }


@router.get("/summary")
async def get_summary(
    project_id: UUID,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """汇总统计"""
    svc = AdjustmentService(db)
    result = await svc.get_summary(project_id, year)
    return result.model_dump()


@router.get("/account-dropdown")
async def get_account_dropdown(
    project_id: UUID,
    report_line_code: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """科目下拉选项"""
    svc = AdjustmentService(db)
    options = await svc.get_account_dropdown(project_id, report_line_code)
    return [o.model_dump() for o in options]


@router.get("/wp-summary/{wp_code}")
async def get_wp_summary(
    project_id: UUID,
    wp_code: str,
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """底稿审定表数据"""
    svc = AdjustmentService(db)
    result = await svc.get_wp_adjustment_summary(project_id, year, wp_code)
    return result.model_dump()


@router.get("/export-summary")
async def export_adjustment_summary(
    project_id: UUID,
    year: int = Query(...),
    format: str = Query("excel", description="导出格式: excel 或 json"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """导出审计调整汇总表（AJE + RJE）"""
    import logging
    _logger = logging.getLogger(__name__)
    _logger.info(
        "adjustment_export: user=%s project=%s year=%s",
        str(current_user.id), str(project_id), year,
    )

    svc = AdjustmentService(db)
    aje_result = await svc.list_entries(project_id, year, adjustment_type=AdjustmentType.aje, page_size=500)
    rje_result = await svc.list_entries(project_id, year, adjustment_type=AdjustmentType.rje, page_size=500)
    # list_entries 返回 dict {"items": [...], "total": ...}
    aje_list = aje_result.get("items", []) if isinstance(aje_result, dict) else aje_result
    rje_list = rje_result.get("items", []) if isinstance(rje_result, dict) else rje_result

    if format == "json":
        return {
            "aje": [_adj_to_dict(a) for a in aje_list],
            "rje": [_adj_to_dict(a) for a in rje_list],
            "aje_count": len(aje_list),
            "rje_count": len(rje_list),
        }

    # Excel 导出
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl 未安装")

    wb = openpyxl.Workbook()

    # AJE sheet
    ws_aje = wb.active
    ws_aje.title = "AJE审计调整"
    _write_adj_sheet(ws_aje, aje_list, "AJE")

    # RJE sheet
    ws_rje = wb.create_sheet("RJE重分类")
    _write_adj_sheet(ws_rje, rje_list, "RJE")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=adjustment_summary_{year}.xlsx"},
    )


@router.get("/export-template")
async def export_adjustment_template(
    project_id: UUID,
    year: int = Query(...),
    template_type: str | None = Query(
        None,
        description="模板类型: soe=国企版 / listed=上市版,不传则从项目派生",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_access("readonly")),
):
    """导出调整分录空白模板 Excel (按项目实际数据生成,支持国企/上市两版).

    包含 4 sheet:
    - 关注事项 (使用说明 / 字段含义 / 校验规则)
    - AJE 模板 (国企版或上市版,根据 template_type)
    - RJE 模板 (同上)
    - 项目科目库 (本项目实际使用的二级明细 + 一级标准 + 报表行联动数据)

    数据维度 (4 列联动):
    - 二级科目编码 / 二级科目名称 (用户主输入,从下拉选择本项目实际科目)
    - 一级科目编码 / 报表项目 (公式自动联动填充)
    """
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(500, "openpyxl 未安装")

    from app.models.audit_platform_models import (
        AccountChart, AccountSource, AccountMapping, ReportLineMapping,
    )
    from app.models.core import Project
    from sqlalchemy import select as sa_select

    # ─── Step 1: 派生 template_type ─────────────────────────────────
    proj_result = await db.execute(sa_select(Project).where(Project.id == project_id))
    proj = proj_result.scalar_one_or_none()
    if not proj:
        raise HTTPException(404, "项目不存在")

    if not template_type:
        template_type = (proj.template_type or "soe").lower()
    template_type = template_type.lower()
    if template_type not in ("soe", "listed"):
        template_type = "soe"
    template_label = "国企版" if template_type == "soe" else "上市版"

    # ─── Step 2: 自愈增量补齐标准科目 ─────────────────────────────
    try:
        from app.services import account_chart_service
        await account_chart_service.load_standard_template(project_id, "enterprise", db)
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"export-template 自愈加载标准科目失败,继续: {e}")

    # ─── Step 3: 装载本项目科目库(client + standard) ─────────────────
    # 客户科目 (source=client) 是本项目实际使用的二级明细 (如 1122.001 应收账款-客户A)
    # 标准科目 (source=standard) 作为兜底 (客户没拆二级时用一级)
    chart_result = await db.execute(
        sa_select(
            AccountChart.account_code,
            AccountChart.account_name,
            AccountChart.source,
        ).where(
            AccountChart.project_id == project_id,
            AccountChart.is_deleted == False,  # noqa: E712
        ).order_by(AccountChart.account_code)
    )
    all_accounts = list(chart_result.all())

    # ─── Step 4: 装载 ReportLineMapping (按当前项目维度,已含国企/上市差异) ───
    rlm_result = await db.execute(
        sa_select(
            ReportLineMapping.standard_account_code,
            ReportLineMapping.report_line_code,
            ReportLineMapping.report_line_name,
        ).where(
            ReportLineMapping.project_id == project_id,
            ReportLineMapping.is_deleted == False,  # noqa: E712
        )
    )
    code_to_report: dict[str, tuple[str, str]] = {
        row[0]: (row[1] or "", row[2] or "") for row in rlm_result.all()
    }

    # ─── Step 5: 装载 AccountMapping (客户码→标准码,用于一级反查) ─────
    mapping_result = await db.execute(
        sa_select(
            AccountMapping.original_account_code,
            AccountMapping.standard_account_code,
        ).where(
            AccountMapping.project_id == project_id,
            AccountMapping.is_deleted == False,  # noqa: E712
        )
    )
    client_to_std: dict[str, str] = {
        row[0]: row[1] for row in mapping_result.all() if row[1]
    }

    def _normalize_to_level1(code: str) -> str:
        """归一化到一级编码: 6401.01/6401-01 → 6401, 1231-01 保留"""
        if not code:
            return ""
        if code.startswith("1231-"):
            return code
        if code in client_to_std:
            return client_to_std[code]
        clean = code.strip()
        for sep in (".", "-", "/", "_", "\\", " "):
            clean = clean.split(sep)[0]
        return clean[:4] if len(clean) >= 4 else clean

    # ─── Step 5b: 装载试算表 (P1-2: 当前余额 + 已有调整数) ──────────
    # 试算表按 standard_account_code 聚合,用于显示参考余额
    from app.models.audit_platform_models import TrialBalance
    tb_result = await db.execute(
        sa_select(
            TrialBalance.standard_account_code,
            TrialBalance.audited_amount,
            TrialBalance.unadjusted_amount,
            TrialBalance.aje_adjustment,
            TrialBalance.rje_adjustment,
        ).where(
            TrialBalance.project_id == project_id,
            TrialBalance.year == year,
            TrialBalance.is_deleted == False,  # noqa: E712
        )
    )
    # 一级编码 → (审定数, 未审数, AJE 调整数, RJE 调整数)
    tb_data: dict[str, tuple] = {}
    tb_codes_with_balance: set[str] = set()  # 试算表中有余额的一级码
    for row in tb_result.all():
        code = row[0]
        tb_data[code] = (
            row[1] or 0,  # audited_amount (审定数 = 当前余额)
            row[2] or 0,  # unadjusted_amount (未审数)
            row[3] or 0,  # aje_adjustment (已有 AJE 调整)
            row[4] or 0,  # rje_adjustment (已有 RJE 调整)
        )
        # 仅余额非零的科目才视为"有余额"
        audited = float(row[1] or 0)
        unadjusted = float(row[2] or 0)
        if abs(audited) > 0.01 or abs(unadjusted) > 0.01:
            tb_codes_with_balance.add(code)

    # ─── Step 6: 拼接科目库 7 列宽表 ─────────────────────────────
    # 7 列: A=二级编码 / B=二级名称 / C=一级编码 / D=一级名称 / E=报表项目 / F=当前余额 / G=已有调整
    # 排序原则:
    # 1. 优先 client 二级 (本项目实际明细)
    # 2. standard 父级在已被 client 子明细覆盖时不再展示 (P0-3 去重)
    # 3. 试算表有余额的科目排前 (P0-1 优先级提示)
    # 未映射科目在 E 列(报表项目)标 ⚠ 未映射 (P0-2)

    seen_codes: set[str] = set()
    # 计算被 client 已覆盖的 standard 一级码 (P0-3 去重)
    client_level1_set: set[str] = set()
    for code, name, source in all_accounts:
        if source == AccountSource.client and code:
            level1 = _normalize_to_level1(code)
            client_level1_set.add(level1)

    # 7 元组: (二级码, 二级名, 一级码, 一级名, 报表行, 余额, 已有调整, has_balance_flag)
    sub_accounts: list[tuple[str, str, str, str, str, float, float, bool]] = []

    # 一级名称查询
    level1_name_map: dict[str, str] = {}
    for code, name, source in all_accounts:
        if source != AccountSource.standard:
            continue
        if "-" not in code and len(code) == 4 and code.isdigit():
            level1_name_map[code] = name or ""
    for code, name, source in all_accounts:
        if source == AccountSource.standard and code.startswith("1231-"):
            level1_name_map[code] = name or ""

    def _build_entry(code: str, name: str) -> tuple[str, str, str, str, str, float, float, bool]:
        """组装一行 sub_accounts 元组"""
        level1 = _normalize_to_level1(code)
        level1_name = level1_name_map.get(level1, "")
        if not level1_name and code == level1:
            level1_name = name or ""
        rl = code_to_report.get(level1, ("", ""))
        report_line = rl[1] if rl[1] else "⚠ 未映射"  # P0-2
        # 余额: 优先按当前码查,再按一级码查
        tb_row = tb_data.get(code) or tb_data.get(level1) or (0, 0, 0, 0)
        audited = float(tb_row[0])
        adj_total = float(tb_row[2]) + float(tb_row[3])  # AJE + RJE
        has_balance = code in tb_codes_with_balance or level1 in tb_codes_with_balance
        return (code, name or code, level1, level1_name, report_line, audited, adj_total, has_balance)

    # 先放 client 科目 (本项目实际明细)
    for code, name, source in all_accounts:
        if source != AccountSource.client or not code or code in seen_codes:
            continue
        seen_codes.add(code)
        sub_accounts.append(_build_entry(code, name))

    # 再放 standard 科目 (P0-3: 已被 client 子明细覆盖的一级科目跳过)
    for code, name, source in all_accounts:
        if source != AccountSource.standard or not code or code in seen_codes:
            continue
        # P0-3: 如果该 standard 一级码已被 client 子明细覆盖,跳过避免重复
        if code in client_level1_set and "-" not in code:
            # 但保留 1231-01~05 这种二级标准 (因为客户子明细可能不全)
            continue
        seen_codes.add(code)
        sub_accounts.append(_build_entry(code, name))

    # P0-1: 试算表有余额的排前面 (按 has_balance 降序稳定排序)
    sub_accounts.sort(key=lambda x: (not x[7], x[0]))  # has_balance=True 在前,然后按编码字典序

    wb = openpyxl.Workbook()

    # ─── Sheet 1: 关注事项 ───────────────────────────────
    ws_notes = wb.active
    ws_notes.title = "关注事项"
    notes = [
        (f"📋 调整分录导入模板使用说明 - {template_label}", "header"),
        ("", ""),
        ("一、字段说明 (9 列结构)", "section"),
        ("• A 编号 (adjustment_no)", "可留空,系统自动生成 (AJE-001 / RJE-001)"),
        ("• B 类型 (type)", "必填,AJE=审计调整 / RJE=重分类调整"),
        ("• C 摘要 (description)", "必填,简要说明调整原因 (≤ 200 字)"),
        ("• D 二级科目编码", "自动填充 (公式),根据 E 列名称反查"),
        ("• E 二级科目名称 ⭐主输入", "必填,从下拉选择本项目实际科目 (含客户子明细)"),
        ("• F 一级科目编码", "自动填充 (公式),归一到一级 (4 位编码或 1231-0x)"),
        ("• G 一级科目名称", "自动填充 (公式),即报表项目对应的标准科目"),
        ("• H 报表项目", f"自动填充 (公式),依据【映射规则】关联到 {template_label} 的报表行"),
        ("• I 借方金额", "正数,与贷方互斥 (同一行只填一个)"),
        ("• J 贷方金额", "正数,与借方互斥 (同一行只填一个)"),
        ("", ""),
        ("二、4 列联动机制", "section"),
        ("第一步: 在 E 列下拉选择 二级科目名称 (本项目实际明细科目)", ""),
        ("第二步: D 列 二级编码 自动填 (INDEX/MATCH 反查)", ""),
        ("第三步: F 列 一级编码 自动归一 (取前 4 位 / 1231 系列保留二级)", ""),
        ("第四步: G/H 列 一级名称 + 报表项目 自动填 (依据本项目映射规则)", ""),
        ("", ""),
        ("三、校验规则", "section"),
        ("✅ 借贷必须平衡", "同一编号下所有行的借方合计 = 贷方合计 (容差 0.01)"),
        ("✅ 科目必须在项目库", "下拉只显示本项目已加载的标准科目 + 客户实际科目"),
        ("✅ 一级映射必须存在", "F/G/H 列联动失败 = 该科目还未做映射规则,请先到 [试算表 → 映射规则] 完善"),
        ("✅ 金额方向", "借方/贷方至少填一个,且只能填一个 (互斥)"),
        ("", ""),
        ("四、AJE / RJE 区别", "section"),
        ("AJE 审计调整", "影响审定数 (P/L 损益类调整),会更新报表数据"),
        ("RJE 重分类", "不影响损益,仅在 BS 内重分类 (如长投↔可供出售)"),
        ("", ""),
        ("五、示例", "section"),
        ("分录 AJE-001 (借应收账款 100, 贷信用减值损失 100)", "→ 两行,同 adjustment_no=AJE-001"),
        ("第 1 行: 编号=AJE-001 类型=AJE 摘要=补提坏账 二级名称=应收账款 借方=100", ""),
        ("第 2 行: 编号=AJE-001 类型=AJE 摘要=补提坏账 二级名称=信用减值损失 贷方=100", ""),
        ("", ""),
        ("⚠️ 重要提示", "section"),
        ("• 修改本模板表结构会导致导入失败", ""),
        ("• 完成填写后通过页面 'Excel 导入' 按钮上传", ""),
        ("• 联动失败 (F/G/H 显示 #N/A) → 该科目缺映射,请先在 [试算表 → 映射规则] 配置", ""),
        ("", ""),
        ("六、项目科目库 (最右侧 sheet)", "section"),
        (f"参见本工作簿最右侧的【项目科目库】sheet,共 {len(sub_accounts)} 个科目", ""),
        ("7 列结构: 二级编码 / 二级名称 / 一级编码 / 一级名称 / 报表项目 / 当前余额 / 已有调整数", ""),
        ("• 浅绿底 = 试算表有余额的科目 (按余额优先排序在前)", ""),
        ("• 浅灰底 = 试算表无余额的科目 (兜底,可用于新增预提)", ""),
        ("• 浅橙底报表项目列 = ⚠ 未映射,请先到 [试算表 → 映射规则] 完善", ""),
        ("• 当前余额 / 已有调整数 = 一级聚合值,客户子明细共享同一聚合 (仅供参考,导入时不参与校验)", ""),
        ("该 sheet 是 E 列下拉数据源 + D/F/G/H 列联动公式取值源,请勿删除", ""),
    ]

    title_font = Font(bold=True, size=14, color="6750A4")
    section_font = Font(bold=True, size=11, color="6750A4")
    field_font = Font(bold=True, size=10)
    note_font = Font(size=10, color="606060")

    for i, (col1, col2) in enumerate(notes, 1):
        c1 = ws_notes.cell(row=i, column=1, value=col1)
        c2 = ws_notes.cell(row=i, column=2, value=col2 if col2 not in ("header", "section") else "")
        if col2 == "header":
            c1.font = title_font
            ws_notes.row_dimensions[i].height = 28
        elif col2 == "section":
            c1.font = section_font
        elif col1.startswith(("•", "✅", "⚠")):
            c1.font = field_font
            c2.font = note_font
        elif col1.startswith("第"):
            c1.font = note_font
        else:
            c1.font = note_font
            c2.font = note_font

    ws_notes.column_dimensions["A"].width = 50
    ws_notes.column_dimensions["B"].width = 60

    # ─── Sheet 2 & 3: AJE / RJE 模板 (9 列结构,4 列联动) ─────────
    # 列结构: A编号 / B类型 / C摘要 / D二级编码(联动) / E二级名称(下拉⭐) /
    #         F一级编码(联动) / G一级名称(联动) / H报表项目(联动) / I借方 / J贷方
    # 设计原则: 用户在 E 列选名称, D/F/G/H 全部 INDEX/MATCH 自动联动
    headers = [
        "编号", "类型", "摘要",
        "二级科目编码", "二级科目名称",
        "一级科目编码", "一级科目名称", "报表项目",
        "借方金额", "贷方金额",
    ]
    header_fill = PatternFill(start_color="F4F0FA", end_color="F4F0FA", fill_type="solid")
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # 找一个真实存在的二级名称做示例(优先客户科目)
    example_name1 = "应收账款"
    example_name2 = "信用减值损失"
    example_name3 = "其他权益工具投资"
    example_name4 = "长期股权投资"
    sub_names_set = {sa[1] for sa in sub_accounts}
    if example_name1 not in sub_names_set:
        for sa in sub_accounts:
            if sa[2].startswith("11"):  # 一级码 1xxx
                example_name1 = sa[1]
                break

    for adj_type, sheet_name, examples in [
        ("AJE", "AJE模板", [
            # 仅填 编号/类型/摘要/二级名称(E)/借贷,其他列由公式自动算
            ("AJE-001", "AJE", "补提应收账款减值", "", example_name1, "", "", "", 0, 100000.00),
            ("AJE-001", "AJE", "补提应收账款减值", "", example_name2, "", "", "", 100000.00, 0),
            ("", "", "", "", "", "", "", "", 0, 0),
            ("", "", "", "", "", "", "", "", 0, 0),
            ("", "", "", "", "", "", "", "", 0, 0),
        ]),
        ("RJE", "RJE模板", [
            ("RJE-001", "RJE", "长投重分类为其他权益工具", "", example_name3, "", "", "", 50000.00, 0),
            ("RJE-001", "RJE", "长投重分类为其他权益工具", "", example_name4, "", "", "", 0, 50000.00),
            ("", "", "", "", "", "", "", "", 0, 0),
            ("", "", "", "", "", "", "", "", 0, 0),
            ("", "", "", "", "", "", "", "", 0, 0),
        ]),
    ]:
        ws = wb.create_sheet(sheet_name)
        # 表头
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        # 示例数据
        example_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        formula_fill = PatternFill(start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")
        for ri, row_data in enumerate(examples, 2):
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val if val != 0 else None)
                c.border = thin_border
                if ri <= 3:
                    c.fill = example_fill

        # 列宽
        ws.column_dimensions["A"].width = 14   # 编号
        ws.column_dimensions["B"].width = 8    # 类型
        ws.column_dimensions["C"].width = 28   # 摘要
        ws.column_dimensions["D"].width = 14   # 二级编码(联动)
        ws.column_dimensions["E"].width = 26   # 二级名称(下拉⭐)
        ws.column_dimensions["F"].width = 14   # 一级编码(联动)
        ws.column_dimensions["G"].width = 22   # 一级名称(联动)
        ws.column_dimensions["H"].width = 22   # 报表项目(联动)
        ws.column_dimensions["I"].width = 14   # 借方
        ws.column_dimensions["J"].width = 14   # 贷方
        ws.freeze_panes = "A2"

        # ─── 4 列联动公式 (第 7 行起,跳过 5 行示例 + 表头) ─────
        # 项目科目库 sheet 7 列: A=二级编码 / B=二级名称 / C=一级编码 / D=一级名称 / E=报表项目 / F=余额 / G=调整
        if sub_accounts:
            n = len(sub_accounts)
            for ri in range(7, 201):
                # D 列: 用 E 名称反查 二级编码 (项目科目库 A 列)
                d_cell = ws.cell(row=ri, column=4,
                    value=f'=IFERROR(INDEX(项目科目库!$A$2:$A${n + 1},MATCH(E{ri},项目科目库!$B$2:$B${n + 1},0)),"")',
                )
                # F 列: 用 E 名称反查 一级编码 (项目科目库 C 列)
                f_cell = ws.cell(row=ri, column=6,
                    value=f'=IFERROR(INDEX(项目科目库!$C$2:$C${n + 1},MATCH(E{ri},项目科目库!$B$2:$B${n + 1},0)),"")',
                )
                # G 列: 用 E 名称反查 一级名称 (项目科目库 D 列)
                g_cell = ws.cell(row=ri, column=7,
                    value=f'=IFERROR(INDEX(项目科目库!$D$2:$D${n + 1},MATCH(E{ri},项目科目库!$B$2:$B${n + 1},0)),"")',
                )
                # H 列: 用 E 名称反查 报表项目 (项目科目库 E 列)
                h_cell = ws.cell(row=ri, column=8,
                    value=f'=IFERROR(INDEX(项目科目库!$E$2:$E${n + 1},MATCH(E{ri},项目科目库!$B$2:$B${n + 1},0)),"")',
                )
                for c in (d_cell, f_cell, g_cell, h_cell):
                    c.border = thin_border
                    c.fill = formula_fill
                    # 公式列填浅黄底色 + 斜体提示,但**不开 sheet 保护**:
                    # 用户可手动覆盖输入(场景=没有 Excel 公式的轻量编辑器/复制粘贴会丢公式),
                    # 导入端 _resolve_code 三级容错(一级码/二级码/二级名称)兜底匹配。

    # ─── Sheet 4: 项目科目库 (7 列: A 二级编码/B 二级名称/C 一级编码/D 一级名称/E 报表项目/F 当前余额/G 已有调整) ───
    ws_lib = wb.create_sheet("项目科目库")
    lib_headers = [
        "二级科目编码", "二级科目名称",
        "一级科目编码", "一级科目名称",
        "报表项目",
        "当前余额", "已有调整数",
    ]
    for col, h in enumerate(lib_headers, 1):
        cell = ws_lib.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="6750A4", end_color="6750A4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 数据行 (sub_accounts 是 8 元组: 二级码/二级名/一级码/一级名/报表行/余额/调整/has_balance)
    no_balance_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # 灰底:无余额
    has_balance_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 浅绿:有余额
    unmapped_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # 浅橙:未映射
    for ri, entry in enumerate(sub_accounts, 2):
        sub_code, sub_name, level1_code, level1_name, report_line, balance, adj_total, has_balance = entry
        ws_lib.cell(row=ri, column=1, value=sub_code).border = thin_border
        ws_lib.cell(row=ri, column=2, value=sub_name).border = thin_border
        ws_lib.cell(row=ri, column=3, value=level1_code).border = thin_border
        ws_lib.cell(row=ri, column=4, value=level1_name).border = thin_border
        c5 = ws_lib.cell(row=ri, column=5, value=report_line)
        c5.border = thin_border
        # P0-2: 未映射科目 标橙色
        if "⚠" in (report_line or ""):
            c5.fill = unmapped_fill
        c6 = ws_lib.cell(row=ri, column=6, value=balance if balance else None)
        c6.border = thin_border
        c6.alignment = Alignment(horizontal="right")
        c6.number_format = '#,##0.00;-#,##0.00'
        c7 = ws_lib.cell(row=ri, column=7, value=adj_total if adj_total else None)
        c7.border = thin_border
        c7.alignment = Alignment(horizontal="right")
        c7.number_format = '#,##0.00;-#,##0.00'
        # P0-1: 有余额的浅绿底色,无余额浅灰底色 (整行)
        if has_balance:
            for col in range(1, 6):  # 1-5 列(科目相关)
                ws_lib.cell(row=ri, column=col).fill = has_balance_fill
        else:
            for col in range(1, 6):
                ws_lib.cell(row=ri, column=col).fill = no_balance_fill

    ws_lib.column_dimensions["A"].width = 16
    ws_lib.column_dimensions["B"].width = 32
    ws_lib.column_dimensions["C"].width = 14
    ws_lib.column_dimensions["D"].width = 22
    ws_lib.column_dimensions["E"].width = 24
    ws_lib.column_dimensions["F"].width = 16
    ws_lib.column_dimensions["G"].width = 16
    ws_lib.freeze_panes = "A2"

    # ─── 数据校验 (E 列科目名称下拉 + B 列 AJE/RJE 类型下拉) ─────
    if sub_accounts:
        n = len(sub_accounts)
        # 引用 项目科目库 B 列 (二级科目名称)
        formula_range_name = f"=项目科目库!$B$2:$B${n + 1}"
        for sheet_name in ("AJE模板", "RJE模板"):
            ws_target = wb[sheet_name]
            # E 列: 二级科目名称下拉
            dv_name = DataValidation(
                type="list",
                formula1=formula_range_name,
                allow_blank=True,
                showDropDown=False,
            )
            dv_name.error = "请从下拉列表中选择本项目实际科目名称"
            dv_name.errorTitle = "科目名称必须在项目科目库"
            dv_name.prompt = "从下拉选择二级科目名称, D/F/G/H 列将自动联动填充"
            dv_name.promptTitle = "提示"
            dv_name.add(f"E2:E200")
            ws_target.add_data_validation(dv_name)

            # B 列: AJE/RJE 类型下拉
            dv_type = DataValidation(
                type="list",
                formula1='"AJE,RJE"',
                allow_blank=False,
            )
            dv_type.add("B2:B200")
            ws_target.add_data_validation(dv_type)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=adjustment_template_{template_type}_{year}.xlsx"
        },
    )


def _adj_to_dict(adj) -> dict:
    """将分录组（dict 或 ORM 对象）转为导出用字典。"""
    if isinstance(adj, dict):
        return {
            "adjustment_no": adj.get("adjustment_no", ""),
            "description": adj.get("description", ""),
            "account_code": adj.get("account_code", ""),
            "account_name": adj.get("account_name", ""),
            "debit_amount": str(adj.get("debit_amount") or adj.get("total_debit") or 0),
            "credit_amount": str(adj.get("credit_amount") or adj.get("total_credit") or 0),
        }
    return {
        "adjustment_no": getattr(adj, "adjustment_no", ""),
        "description": getattr(adj, "description", ""),
        "account_code": getattr(adj, "account_code", ""),
        "account_name": getattr(adj, "account_name", ""),
        "debit_amount": str(getattr(adj, "debit_amount", 0) or 0),
        "credit_amount": str(getattr(adj, "credit_amount", 0) or 0),
    }


def _write_adj_sheet(ws, entries, adj_type: str):
    from openpyxl.styles import Font, Alignment, PatternFill

    headers = ["编号", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额"]
    header_fill = PatternFill(start_color="F4F0FA", end_color="F4F0FA", fill_type="solid")
    header_font = Font(bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for adj in entries:
        if isinstance(adj, dict) and "line_items" in adj:
            # 分录组模式：每个 line_item 一行
            line_items = adj.get("line_items", [])
            if not line_items:
                # 无明细行时输出汇总行
                ws.cell(row=row_idx, column=1, value=adj.get("adjustment_no", ""))
                ws.cell(row=row_idx, column=2, value=adj.get("description", ""))
                ws.cell(row=row_idx, column=3, value="")
                ws.cell(row=row_idx, column=4, value="")
                ws.cell(row=row_idx, column=5, value=float(adj.get("total_debit") or 0))
                ws.cell(row=row_idx, column=6, value=float(adj.get("total_credit") or 0))
                row_idx += 1
            else:
                for li in line_items:
                    ws.cell(row=row_idx, column=1, value=adj.get("adjustment_no", ""))
                    ws.cell(row=row_idx, column=2, value=adj.get("description", ""))
                    ws.cell(row=row_idx, column=3, value=li.get("standard_account_code", ""))
                    ws.cell(row=row_idx, column=4, value=li.get("account_name", ""))
                    ws.cell(row=row_idx, column=5, value=float(li.get("debit_amount") or 0))
                    ws.cell(row=row_idx, column=6, value=float(li.get("credit_amount") or 0))
                    row_idx += 1
        else:
            # 扁平模式兼容
            d = _adj_to_dict(adj)
            ws.cell(row=row_idx, column=1, value=d["adjustment_no"])
            ws.cell(row=row_idx, column=2, value=d["description"])
            ws.cell(row=row_idx, column=3, value=d["account_code"])
            ws.cell(row=row_idx, column=4, value=d["account_name"])
            ws.cell(row=row_idx, column=5, value=float(d["debit_amount"] or 0))
            ws.cell(row=row_idx, column=6, value=float(d["credit_amount"] or 0))
            row_idx += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16


# R10 Spec B / Sprint 3.2.3 — 调整分录组关联底稿
@router.get("/{entry_group_id}/related-workpapers")
async def get_adjustment_related_workpapers(
    project_id: UUID,
    entry_group_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """R10 Spec B / F8：根据调整分录组中所有 line_items 的科目反查关联底稿。"""
    from app.models.adjustment_models import Adjustment, AdjustmentEntry
    from app.services.workpaper_query import find_workpapers_by_account_codes

    # 找该 group_id 的所有分录行
    stmt = (
        sa.select(Adjustment)
        .where(
            Adjustment.entry_group_id == entry_group_id,
            Adjustment.project_id == project_id,
            Adjustment.is_deleted == False,  # noqa: E712
        )
    )
    adjustments = (await db.execute(stmt)).scalars().all()
    if not adjustments:
        raise HTTPException(status_code=404, detail="调整分录组不存在")

    # 收集所有 line_items 的科目编码
    codes: set[str] = set()
    for adj in adjustments:
        # AdjustmentEntry 关联的 line_items
        entries_stmt = sa.select(AdjustmentEntry).where(
            AdjustmentEntry.adjustment_id == adj.id,
            AdjustmentEntry.is_deleted == False,  # noqa: E712
        )
        try:
            entries = (await db.execute(entries_stmt)).scalars().all()
            for e in entries:
                code = getattr(e, "standard_account_code", None) or getattr(e, "account_code", None)
                if code:
                    codes.add(str(code))
        except Exception:
            pass

    workpapers = await find_workpapers_by_account_codes(db, project_id, list(codes))
    return {
        "entry_group_id": entry_group_id,
        "account_codes": list(codes),
        "workpapers": workpapers,
    }
