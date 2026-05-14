"""批量重分类导入/导出 API

GET /template — 导出重分类 Excel 模板（预填科目列表）
POST /import — 读取 Excel 按连续借贷平衡组拆分
POST /inline-submit — 多行录入一键提交（借贷平衡门控）

Validates: Requirements 16.1, 16.2, 16.3, 16.4, 16.5, 16.6, 16.7, 16.8
"""

from __future__ import annotations

import io
import logging
from decimal import Decimal
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.core import User

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/projects/{project_id}/adjustments/reclassification",
    tags=["reclassification"],
)


# ── Pydantic Models ──

class ReclassificationRow(BaseModel):
    """重分类单行"""
    account_code: str
    account_name: str = ""
    debit_amount: float = 0
    credit_amount: float = 0
    summary: str = ""


class InlineSubmitRequest(BaseModel):
    """多行录入一键提交请求"""
    year: int
    rows: list[ReclassificationRow] = Field(..., min_length=2)


class ImportResult(BaseModel):
    """导入结果"""
    groups: list[dict] = []
    unbalanced_groups: list[dict] = []
    total_rows: int = 0


# ── 工具函数 ──

def _split_balanced_groups(rows: list[dict]) -> tuple[list[list[dict]], list[list[dict]]]:
    """按连续借贷平衡组拆分行。

    逻辑：遍历行，累计借方/贷方合计。当 sum(debit)==sum(credit) 时，
    该组成为一笔独立分录。如果最后一组不平衡，标记为"待修正"。
    """
    balanced_groups: list[list[dict]] = []
    unbalanced_groups: list[list[dict]] = []

    current_group: list[dict] = []
    sum_debit = Decimal("0")
    sum_credit = Decimal("0")

    for row in rows:
        debit = Decimal(str(row.get("debit_amount", 0) or 0))
        credit = Decimal(str(row.get("credit_amount", 0) or 0))

        # 跳过空行（借贷都为 0）
        if debit == 0 and credit == 0:
            continue

        current_group.append(row)
        sum_debit += debit
        sum_credit += credit

        # 检查是否平衡
        if sum_debit == sum_credit and len(current_group) >= 2:
            balanced_groups.append(current_group)
            current_group = []
            sum_debit = Decimal("0")
            sum_credit = Decimal("0")

    # 处理剩余未平衡的组
    if current_group:
        unbalanced_groups.append(current_group)

    return balanced_groups, unbalanced_groups


# ── 端点 ──

@router.get("/template")
async def export_reclassification_template(
    project_id: UUID,
    year: int = 2025,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出重分类 Excel 模板（预填科目列表）。

    模板列：科目编码 / 科目名称 / 借方金额 / 贷方金额 / 摘要
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # 查询项目的科目列表
    account_query = text("""
        SELECT DISTINCT account_code, account_name
        FROM tb_balance
        WHERE project_id = :project_id
          AND year = :year
          AND is_deleted = false
        ORDER BY account_code
    """)
    result = await db.execute(
        account_query,
        {"project_id": str(project_id), "year": year},
    )
    accounts = result.fetchall()

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重分类调整"

    # 表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0EDF5", end_color="F0EDF5", fill_type="solid")

    # 写表头
    headers = ["科目编码", "科目名称", "借方金额", "贷方金额", "摘要"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 30

    # 预填科目列表（科目编码+名称，金额留空）
    for i, acct in enumerate(accounts, 2):
        ws.cell(row=i, column=1, value=acct[0])
        ws.cell(row=i, column=2, value=acct[1])

    # 添加科目参考 sheet
    ws_ref = wb.create_sheet("科目参考")
    ws_ref.cell(row=1, column=1, value="科目编码").font = header_font
    ws_ref.cell(row=1, column=2, value="科目名称").font = header_font
    for i, acct in enumerate(accounts, 2):
        ws_ref.cell(row=i, column=1, value=acct[0])
        ws_ref.cell(row=i, column=2, value=acct[1])

    # 输出为字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reclassification_template_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_reclassification(
    project_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导入重分类 Excel，按连续借贷平衡组拆分为多笔独立分录。"""
    import openpyxl

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式文件")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # 解析行数据
    rows: list[dict] = []
    header_row = None

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            # 跳过表头
            header_row = row
            continue
        if not row or all(cell is None for cell in row):
            continue

        account_code = str(row[0] or "").strip()
        account_name = str(row[1] or "").strip() if len(row) > 1 else ""
        debit = float(row[2] or 0) if len(row) > 2 and row[2] else 0
        credit = float(row[3] or 0) if len(row) > 3 and row[3] else 0
        summary = str(row[4] or "").strip() if len(row) > 4 else ""

        if not account_code and debit == 0 and credit == 0:
            continue

        rows.append({
            "account_code": account_code,
            "account_name": account_name,
            "debit_amount": debit,
            "credit_amount": credit,
            "summary": summary,
            "row_number": i,
        })

    wb.close()

    # 按借贷平衡组拆分
    balanced_groups, unbalanced_groups = _split_balanced_groups(rows)

    return {
        "groups": [
            {
                "group_index": idx + 1,
                "rows": group,
                "total_debit": float(sum(Decimal(str(r["debit_amount"])) for r in group)),
                "total_credit": float(sum(Decimal(str(r["credit_amount"])) for r in group)),
                "is_balanced": True,
            }
            for idx, group in enumerate(balanced_groups)
        ],
        "unbalanced_groups": [
            {
                "group_index": idx + 1,
                "rows": group,
                "total_debit": float(sum(Decimal(str(r["debit_amount"])) for r in group)),
                "total_credit": float(sum(Decimal(str(r["credit_amount"])) for r in group)),
                "is_balanced": False,
                "status": "pending_correction",
            }
            for idx, group in enumerate(unbalanced_groups)
        ],
        "total_rows": len(rows),
        "balanced_count": len(balanced_groups),
        "unbalanced_count": len(unbalanced_groups),
    }


@router.post("/inline-submit")
async def inline_submit_reclassification(
    project_id: UUID,
    body: InlineSubmitRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """多行录入一键提交（借贷平衡门控）。

    当 sum(debit_amount) ≠ sum(credit_amount) 时拒绝提交（返回 400）。
    """
    # 借贷平衡校验
    total_debit = sum(Decimal(str(r.debit_amount)) for r in body.rows)
    total_credit = sum(Decimal(str(r.credit_amount)) for r in body.rows)

    if total_debit != total_credit:
        raise HTTPException(
            status_code=400,
            detail={
                "error_code": "DEBIT_CREDIT_IMBALANCE",
                "message": f"借贷不平衡：借方合计 {float(total_debit)}，贷方合计 {float(total_credit)}，差额 {float(total_debit - total_credit)}",
                "total_debit": float(total_debit),
                "total_credit": float(total_credit),
                "difference": float(total_debit - total_credit),
            },
        )

    # 创建重分类调整分录
    from uuid import uuid4
    from datetime import datetime, timezone
    from app.services.event_bus import event_bus
    from app.models.audit_platform_schemas import EventPayload, EventType

    entry_group_id = uuid4()
    account_codes: list[str] = []

    # 获取下一个分录号
    no_query = text("""
        SELECT MAX(CAST(SUBSTRING(adjustment_no FROM '[0-9]+$') AS INTEGER))
        FROM adjustments
        WHERE project_id = :project_id AND year = :year AND is_deleted = false
    """)
    try:
        no_result = await db.execute(
            no_query, {"project_id": str(project_id), "year": body.year}
        )
        max_no = no_result.scalar() or 0
    except Exception:
        max_no = 0

    adjustment_no = f"RJE-{max_no + 1:03d}"

    for row in body.rows:
        row_id = uuid4()
        insert_query = text("""
            INSERT INTO adjustments
                (id, project_id, year, entry_group_id, adjustment_no,
                 account_code, account_name, debit_amount, credit_amount,
                 summary, adjustment_type, status, is_deleted, created_at, updated_at, version)
            VALUES
                (:id, :project_id, :year, :entry_group_id, :adjustment_no,
                 :account_code, :account_name, :debit_amount, :credit_amount,
                 :summary, :adjustment_type, :status, false, :now, :now, 1)
        """)
        now = datetime.now(timezone.utc)
        await db.execute(
            insert_query,
            {
                "id": str(row_id),
                "project_id": str(project_id),
                "year": body.year,
                "entry_group_id": str(entry_group_id),
                "adjustment_no": adjustment_no,
                "account_code": row.account_code,
                "account_name": row.account_name,
                "debit_amount": float(row.debit_amount) if row.debit_amount else None,
                "credit_amount": float(row.credit_amount) if row.credit_amount else None,
                "summary": row.summary,
                "adjustment_type": "reclassification",
                "status": "committed",
                "now": now,
            },
        )
        if row.account_code:
            account_codes.append(row.account_code)

    await db.commit()

    # 发布批量提交事件（单次级联）
    unique_codes = list(set(account_codes))
    if unique_codes:
        await event_bus.publish(EventPayload(
            event_type=EventType.ADJUSTMENT_BATCH_COMMITTED,
            project_id=project_id,
            year=body.year,
            account_codes=unique_codes,
            entry_group_id=entry_group_id,
        ))

    return {
        "message": "重分类提交成功",
        "entry_group_id": str(entry_group_id),
        "adjustment_no": adjustment_no,
        "rows_count": len(body.rows),
        "total_debit": float(total_debit),
        "total_credit": float(total_credit),
    }
