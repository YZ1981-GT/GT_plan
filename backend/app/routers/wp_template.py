"""底稿模板 API 路由

- POST   /api/templates              — 上传模板
- GET    /api/templates              — 模板列表
- GET    /api/templates/{code}       — 获取模板详情
- POST   /api/templates/{code}/versions — 创建新版本
- DELETE /api/templates/{id}         — 删除模板
- GET    /api/template-sets          — 模板集列表
- GET    /api/template-sets/{id}     — 模板集详情
- POST   /api/template-sets          — 创建模板集
- PUT    /api/template-sets/{id}     — 更新模板集
- POST   /api/template-sets/seed     — 初始化内置模板集

Validates: Requirements 1.1-1.8
"""

from __future__ import annotations

from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.deps import get_current_user
from app.models.core import User
from app.models.workpaper_models import WpIndex, WorkingPaper, WpStatus, WpSourceType
from app.models.workpaper_schemas import TemplateResponse, TemplateSetResponse
from app.services.template_engine import TemplateEngine
import sqlalchemy as sa

router = APIRouter(tags=["templates"])


# ---------------------------------------------------------------------------
# Request schemas
# ---------------------------------------------------------------------------


class TemplateUploadRequest(BaseModel):
    template_code: str
    template_name: str
    audit_cycle: str | None = None
    applicable_standard: str | None = None
    description: str | None = None
    named_ranges: list[dict] | None = None


class VersionCreateRequest(BaseModel):
    change_type: str = "minor"  # "major" or "minor"


class TemplateSetCreateRequest(BaseModel):
    set_name: str
    template_codes: list[str] | None = None
    applicable_audit_type: str | None = None
    applicable_standard: str | None = None
    description: str | None = None


class TemplateSetUpdateRequest(BaseModel):
    set_name: str | None = None
    template_codes: list[str] | None = None
    applicable_audit_type: str | None = None
    applicable_standard: str | None = None
    description: str | None = None


class GenerateWorkpapersRequest(BaseModel):
    template_set_id: UUID
    year: int = 2025


# ---------------------------------------------------------------------------
# Template endpoints
# ---------------------------------------------------------------------------


@router.post("/api/templates", response_model=TemplateResponse)
async def upload_template(
    data: TemplateUploadRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """上传模板文件（MVP: 仅保存元数据）"""
    engine = TemplateEngine()
    template = await engine.upload_template(
        db=db,
        template_code=data.template_code,
        template_name=data.template_name,
        audit_cycle=data.audit_cycle,
        applicable_standard=data.applicable_standard,
        description=data.description,
        named_ranges=data.named_ranges,
    )
    await db.commit()
    return template


@router.get("/api/templates", response_model=list[TemplateResponse])
async def list_templates(
    audit_cycle: str | None = None,
    applicable_standard: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """模板列表（支持按循环、准则筛选）"""
    engine = TemplateEngine()
    return await engine.list_templates(
        db=db,
        audit_cycle=audit_cycle,
        applicable_standard=applicable_standard,
    )


@router.get("/api/templates/{code}", response_model=TemplateResponse)
async def get_template(
    code: str,
    version: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取模板详情（默认最新版本）"""
    engine = TemplateEngine()
    tpl = await engine.get_template(db=db, template_code=code, version=version)
    if tpl is None:
        raise HTTPException(status_code=404, detail=f"模板 {code} 不存在")
    return tpl


@router.post("/api/templates/{code}/versions", response_model=TemplateResponse)
async def create_version(
    code: str,
    data: VersionCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建新版本"""
    engine = TemplateEngine()
    try:
        tpl = await engine.create_version(
            db=db,
            template_code=code,
            change_type=data.change_type,
        )
        await db.commit()
        return tpl
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/api/templates/{template_id}")
async def delete_template(
    template_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除模板（校验无引用）"""
    engine = TemplateEngine()
    try:
        await engine.delete_template(db=db, template_id=template_id)
        await db.commit()
        return {"message": "模板已删除"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ---------------------------------------------------------------------------
# Template set endpoints
# ---------------------------------------------------------------------------


@router.get("/api/template-sets", response_model=list[TemplateSetResponse])
async def list_template_sets(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """模板集列表"""
    engine = TemplateEngine()
    return await engine.get_template_sets(db=db)


@router.get("/api/template-sets/{set_id}", response_model=TemplateSetResponse)
async def get_template_set(
    set_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """模板集详情"""
    engine = TemplateEngine()
    ts = await engine.get_template_set(db=db, set_id=set_id)
    if ts is None:
        raise HTTPException(status_code=404, detail="模板集不存在")
    return ts


@router.post("/api/template-sets", response_model=TemplateSetResponse)
async def create_template_set(
    data: TemplateSetCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建模板集"""
    engine = TemplateEngine()
    ts = await engine.create_template_set(
        db=db,
        set_name=data.set_name,
        template_codes=data.template_codes,
        applicable_audit_type=data.applicable_audit_type,
        applicable_standard=data.applicable_standard,
        description=data.description,
    )
    await db.commit()
    return ts


@router.put("/api/template-sets/{set_id}", response_model=TemplateSetResponse)
async def update_template_set(
    set_id: UUID,
    data: TemplateSetUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新模板集"""
    engine = TemplateEngine()
    try:
        ts = await engine.update_template_set(
            db=db,
            set_id=set_id,
            set_name=data.set_name,
            template_codes=data.template_codes,
            applicable_audit_type=data.applicable_audit_type,
            applicable_standard=data.applicable_standard,
            description=data.description,
        )
        await db.commit()
        return ts
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/api/template-sets/seed")
async def seed_template_sets(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """初始化6个内置模板集（幂等）"""
    engine = TemplateEngine()
    created = await engine.seed_builtin_template_sets(db=db)
    await db.commit()
    return {"message": f"已创建 {len(created)} 个内置模板集", "count": len(created)}


# ---------------------------------------------------------------------------
# Generate project workpapers
# ---------------------------------------------------------------------------


@router.post("/api/projects/{project_id}/working-papers/generate")
async def generate_project_workpapers(
    project_id: UUID,
    data: GenerateWorkpapersRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从模板集生成项目底稿"""
    engine = TemplateEngine()
    try:
        workpapers = await engine.generate_project_workpapers(
            db=db,
            project_id=project_id,
            template_set_id=data.template_set_id,
            year=data.year,
        )
        await db.commit()
        return {
            "message": f"已生成 {len(workpapers)} 个底稿",
            "count": len(workpapers),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


class GenerateFromCodesRequest(BaseModel):
    wp_codes: list[str]
    year: int = 2025


@router.post("/api/projects/{project_id}/working-papers/generate-from-codes")
async def generate_from_codes(
    project_id: UUID,
    data: GenerateFromCodesRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从推荐的底稿编码列表直接生成底稿文件（不需要模板集）"""
    import json
    import shutil
    from pathlib import Path

    lib_path = Path(__file__).parent.parent.parent / "data" / "gt_template_library.json"
    template_lib: dict[str, dict] = {}
    if lib_path.exists():
        try:
            with open(lib_path, "r", encoding="utf-8-sig") as f:
                lib_data = json.load(f)
            for item in lib_data.get("templates", lib_data) if isinstance(lib_data, dict) else lib_data:
                template_lib[item.get("code", item.get("wp_code", ""))] = item
        except Exception:
            pass

    project_wp_dir = Path("storage") / "projects" / str(project_id) / "workpapers"
    created = 0
    skipped = 0

    for code in data.wp_codes:
        # 检查是否已存在
        existing = await db.execute(
            sa.select(WpIndex).where(
                WpIndex.project_id == project_id,
                WpIndex.wp_code == code,
                WpIndex.is_deleted == sa.false(),
            )
        )
        if existing.scalar_one_or_none():
            skipped += 1
            continue

        lib_entry = template_lib.get(code, {})
        wp_name = lib_entry.get("name", lib_entry.get("wp_name", f"底稿{code}"))
        cycle = lib_entry.get("cycle_prefix", code[0] if code else "X")

        # 创建 wp_index
        wp_index = WpIndex(
            project_id=project_id,
            wp_code=code,
            wp_name=wp_name,
            audit_cycle=cycle,
            status=WpStatus.not_started,
        )
        db.add(wp_index)
        await db.flush()

        # 文件目录
        cycle_dir = project_wp_dir / cycle
        cycle_dir.mkdir(parents=True, exist_ok=True)
        dest_file = cycle_dir / f"{code}.xlsx"

        # 复制模板文件（优先从知识库底稿模板目录查找）
        copied = False
        src_path = lib_entry.get("file_path", "")
        template_name = lib_entry.get("name", "") or wp_name

        # 1. 知识库底稿模板目录: ~/.gt_audit_helper/knowledge/workpaper_templates/{cycle}/
        import os
        kb_base = Path(os.path.expanduser("~/.gt_audit_helper/knowledge/workpaper_templates"))
        kb_file = kb_base / cycle / f"{template_name}.xlsx" if template_name else None
        # 也尝试用原始文件名
        if src_path:
            kb_file_by_name = kb_base / cycle / Path(src_path).name
        else:
            kb_file_by_name = None

        for candidate in [kb_file, kb_file_by_name]:
            if candidate and candidate.exists():
                shutil.copy2(candidate, dest_file)
                copied = True
                break

        # 2. 回退：从原始模板路径查找（项目根目录）
        if not copied and src_path:
            src = Path(src_path)
            if not src.exists():
                root_src = Path(__file__).resolve().parent.parent.parent.parent / src_path
                if root_src.exists():
                    src = root_src
            if src.exists():
                shutil.copy2(src, dest_file)
                copied = True

        if not copied:
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = code
                ws["A1"] = f"底稿编号: {code}"
                ws["A2"] = f"底稿名称: {wp_name}"
                ws["A3"] = f"审计年度: {data.year}"
                wb.save(str(dest_file))
                wb.close()
            except Exception:
                dest_file.write_bytes(b"")

        # 创建 working_paper
        wp = WorkingPaper(
            project_id=project_id,
            wp_index_id=wp_index.id,
            file_path=str(dest_file),
            source_type=WpSourceType.template,
            file_version=1,
            created_by=current_user.id,
        )
        db.add(wp)

        # F50 / Sprint 8.17: 底稿快照绑定 —— 绑定当前 active dataset
        try:
            from app.services.dataset_query import bind_to_active_dataset
            await bind_to_active_dataset(db, wp, project_id, data.year)
        except Exception as _bind_err:
            import logging
            logging.getLogger(__name__).warning(
                "dataset binding failed for wp %s: %s", code, _bind_err
            )

        # 填充底稿表头（编制单位/审计期间/索引号/交叉索引等）
        try:
            from app.services.wp_header_service import fill_workpaper_header
            await fill_workpaper_header(
                db=db, project_id=project_id, wp_id=wp.id,
                file_path=str(dest_file), wp_code=code, wp_name=wp_name,
                cycle=cycle,
            )
        except Exception as _e:
            import logging
            logging.getLogger(__name__).warning("fill header failed for %s: %s", code, _e)

        created += 1

    await db.flush()
    await db.commit()
    return {"created": created, "skipped": skipped, "message": f"已生成 {created} 个底稿，跳过 {skipped} 个已存在"}


class CreateCustomWorkpaperRequest(BaseModel):
    wp_code: str
    wp_name: str
    audit_cycle: str | None = None
    year: int = 2025


@router.post("/api/projects/{project_id}/working-papers/create-custom")
async def create_custom_workpaper(
    project_id: UUID,
    data: CreateCustomWorkpaperRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建自定义底稿（用户自建，非模板生成）

    自动填充致同标准表头（编制单位/审计期间/索引号/交叉索引等）。
    """
    from pathlib import Path

    # 检查编号是否已存在
    existing = await db.execute(
        sa.select(WpIndex).where(
            WpIndex.project_id == project_id,
            WpIndex.wp_code == data.wp_code,
            WpIndex.is_deleted == sa.false(),
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail=f"底稿编号 {data.wp_code} 已存在")

    cycle = data.audit_cycle or (data.wp_code[0] if data.wp_code else "X")

    # 创建 wp_index
    wp_index = WpIndex(
        project_id=project_id,
        wp_code=data.wp_code,
        wp_name=data.wp_name,
        audit_cycle=cycle,
        status=WpStatus.not_started,
    )
    db.add(wp_index)
    await db.flush()

    # 创建空白 xlsx
    project_wp_dir = Path("storage") / "projects" / str(project_id) / "workpapers"
    cycle_dir = project_wp_dir / cycle
    cycle_dir.mkdir(parents=True, exist_ok=True)
    dest_file = cycle_dir / f"{data.wp_code}.xlsx"

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data.wp_code
        wb.save(str(dest_file))
        wb.close()
    except Exception:
        dest_file.write_bytes(b"")

    # 创建 working_paper
    wp = WorkingPaper(
        project_id=project_id,
        wp_index_id=wp_index.id,
        file_path=str(dest_file),
        source_type=WpSourceType.manual,
        file_version=1,
        created_by=current_user.id,
    )
    db.add(wp)
    await db.flush()

    # F50 / Sprint 8.17: 自定义底稿同样绑定当前 active dataset
    try:
        from app.services.dataset_query import bind_to_active_dataset
        await bind_to_active_dataset(db, wp, project_id, data.year)
    except Exception as _bind_err:
        import logging
        logging.getLogger(__name__).warning(
            "dataset binding failed for custom wp %s: %s", data.wp_code, _bind_err
        )

    # 自定义底稿强制写入标准表头（is_custom=True）
    try:
        from app.services.wp_header_service import fill_workpaper_header
        await fill_workpaper_header(
            db=db, project_id=project_id, wp_id=wp.id,
            file_path=str(dest_file), wp_code=data.wp_code, wp_name=data.wp_name,
            cycle=cycle, is_custom=True,
        )
    except Exception as _e:
        import logging
        logging.getLogger(__name__).warning("fill custom header failed for %s: %s", data.wp_code, _e)

    await db.commit()
    return {
        "wp_id": str(wp.id),
        "wp_code": data.wp_code,
        "wp_name": data.wp_name,
        "file_path": str(dest_file),
        "message": "自定义底稿创建成功，表头已自动填充",
    }
