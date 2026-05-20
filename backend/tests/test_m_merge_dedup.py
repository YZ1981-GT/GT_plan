"""M 循环合并去重验证测试（spec workpaper-m-equity-cycle Task 1.1）

验证 chain_orchestrator 对 M 循环 10 文件复用 _merge_sheets_dedup 合并去重：

Sprint 0 实测基线（requirements.md §一·B / design.md ADR-M1）：
- 10 模板文件（M1 应付股利 / M2 实收资本 / M3 库存股 / M4 资本公积 /
  M5 盈余公积 / M6 未分配利润 / M7 专项储备 / M8 一般风险准备 /
  M9 其他综合收益 / M10 其他权益工具）
- 102 raw sheet
- 4 历史遗留 sheet（3"修订前" + 1"-删除"）：
  - '其他权益工具实质性程序表 Q10A (修订前)'
  - '未分配利润实质性程序表 Q6A  (修订前)'
  - '一般风险准备实质性程序表 Q8A  (修订前)'
  - '针对性测试M8-5-删除'
- 3 末尾空格 sheet：
  - '未分配利润实质性程序表 M6A '（M6 文件）
  - ' 专项储备实质性程序表 M7A '（M7 文件，首尾均有空格）
  - '一般风险准备实质性程序表 M8A '（M8 文件）
- 33 跨文件归一化重复：
  底稿目录(10×→9) / GT_Custom(10×→9) / 附注披露信息(上市公司)(9×→8) /
  附注披露信息(国有企业)(8×→7) = 33
- 合并去重后有效 sheet = 65（102 raw - 4 历史 - 33 跨文件去重 = 65）

测试目标：
1. M 循环 10 文件被正确发现
2. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
3. 102 raw sheets → 4 历史遗留 → 33 跨文件去重 → 65 有效 sheet
4. 3 末尾空格 sheet（M6A/M7A/M8A）被正确保留（不被过滤）
5. 跨文件去重（底稿目录 / GT_Custom / 附注披露变体）正确工作
6. D/F/H/I/G/J/K/L 历史遗留过滤回归无影响

_Requirements: M-F1_
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "M"

# Collect all M cycle template files
M_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_m_sheet_names() -> list[str]:
    """从 10 个 M 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in M_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程：先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_m_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestMCycleMergeDedup:
    """M-F1: 10 文件合并去重验证（Sprint 0 baseline: 10 files / 102 raw / 4 historical / 65 dedup）"""

    def test_m_cycle_has_10_template_files(self):
        """M 循环应有 10 个模板文件（M1~M10）"""
        assert len(M_TEMPLATE_FILES) == 10, (
            f"Expected 10 M cycle template files, got {len(M_TEMPLATE_FILES)}: "
            f"{[f.name for f in M_TEMPLATE_FILES]}"
        )

    def test_m_cycle_template_files_named_correctly(self):
        """M 循环 10 个模板文件命名应覆盖 M1~M10 全部编号"""
        names = [f.name for f in M_TEMPLATE_FILES]
        for i in range(1, 11):
            prefix = f"M{i}"
            assert any(n.startswith(prefix + " ") or n.startswith(prefix + ".") for n in names), (
                f"Expected file starting with '{prefix}' in {names}"
            )

    def test_m_cycle_raw_sheet_count_102(self):
        """M 循环 10 文件原始 sheet 总数 = 102（Sprint 0 实测基线）"""
        all_names = _get_all_m_sheet_names()
        assert len(all_names) == 102, (
            f"Expected 102 raw sheets, got {len(all_names)}"
        )

    def test_m_cycle_dedup_sheet_count_65(self):
        """M 循环合并去重后 sheet 数 = 65（Sprint 0 实测基线）

        102 raw - 4 historical - 33 跨文件归一化去重 = 65
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 65, (
            f"Expected 65 dedup sheets, got {dedup_count}"
        )

    def test_m_cycle_historical_sheets_count_is_4(self):
        """M 循环 102 sheet 中恰好 4 个命中历史遗留过滤（3"修订前" + 1"-删除"）"""
        all_names = _get_all_m_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 4, (
            f"Expected 4 historical sheet hits, got {len(historical_hits)}: {historical_hits}"
        )

    def test_m_cycle_historical_3_xiuding_qian(self):
        """M 循环 4 个历史遗留中 3 个含"修订前"模式"""
        all_names = _get_all_m_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        xiuding_qian = [h for h in historical_hits if "修订前" in h]
        assert len(xiuding_qian) == 3, (
            f"Expected 3 '修订前' historical sheets, got {len(xiuding_qian)}: {xiuding_qian}"
        )

    def test_m_cycle_historical_1_shanchu(self):
        """M 循环 4 个历史遗留中 1 个以"-删除"结尾"""
        all_names = _get_all_m_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        shanchu = [h for h in historical_hits if h.endswith("-删除")]
        assert len(shanchu) == 1, (
            f"Expected 1 '-删除' historical sheet, got {len(shanchu)}: {shanchu}"
        )
        assert "M8-5" in shanchu[0], (
            f"Expected '-删除' sheet to contain 'M8-5', got '{shanchu[0]}'"
        )

    def test_m_cycle_trailing_space_sheet_count_3(self):
        """M 循环恰好 3 个 sheet 含末尾空格（M6A/M7A/M8A）"""
        all_names = _get_all_m_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        assert len(trailing_space_sheets) == 3, (
            f"Expected 3 trailing-space sheets, got {len(trailing_space_sheets)}: "
            f"{[repr(s) for s in trailing_space_sheets]}"
        )

    def test_m_cycle_trailing_space_sheets_are_m6a_m7a_m8a(self):
        """末尾空格 sheet 分别对应 M6A / M7A / M8A"""
        all_names = _get_all_m_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        # 验证包含 M6A / M7A / M8A 标识
        trailing_codes = []
        for s in trailing_space_sheets:
            if "M6A" in s:
                trailing_codes.append("M6A")
            elif "M7A" in s:
                trailing_codes.append("M7A")
            elif "M8A" in s:
                trailing_codes.append("M8A")
        assert sorted(trailing_codes) == ["M6A", "M7A", "M8A"], (
            f"Expected trailing-space sheets to be M6A/M7A/M8A, got {trailing_codes} "
            f"from {[repr(s) for s in trailing_space_sheets]}"
        )

    def test_m_cycle_trailing_space_sheets_preserved_in_dedup(self):
        """3 个末尾空格 sheet 在去重后被正确保留（不被历史遗留过滤）"""
        all_names = _get_all_m_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        for sheet in trailing_space_sheets:
            assert _should_skip_historical_sheet(sheet) is False, (
                f"Trailing-space sheet should NOT be filtered as historical: {repr(sheet)}"
            )

    def test_m_cycle_trailing_space_normalization_distinct(self):
        """3 个末尾空格 sheet 归一化后各自唯一（不会互相去重）"""
        all_names = _get_all_m_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        normalized = [_normalize_sheet_name(s) for s in trailing_space_sheets]
        assert len(set(normalized)) == 3, (
            f"3 trailing-space sheets should have 3 distinct normalized names, "
            f"got {normalized}"
        )

    def test_m_cycle_cross_file_redundant_count_33(self):
        """M 循环跨文件去重应移除 33 个重复 sheet

        10 文件包含的跨文件重复 sheet（归一化后）：
        - 底稿目录(10×→9 dup)
        - GT_Custom(10×→9 dup)
        - 附注披露信息(上市公司)(9×→8 dup)
        - 附注披露信息(国有企业)(8×→7 dup)
        合计 9+9+8+7 = 33 重复需去除
        """
        all_names = _get_all_m_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)
        dups = {k: v for k, v in norm_counter.items() if v > 1}
        total_redundant = sum(v - 1 for v in dups.values())
        assert total_redundant == 33, (
            f"Expected 33 cross-file redundant sheets, got {total_redundant}: {dups}"
        )

    def test_m_cycle_cross_file_dedup_targets_identified(self):
        """验证跨文件去重目标 sheet 分布：底稿目录 10× / GT_Custom 10×"""
        all_names = _get_all_m_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 底稿目录 10 个文件各 1 份
        assert norm_counter.get("底稿目录", 0) == 10, (
            f"'底稿目录' should appear 10 times (one per file), "
            f"got {norm_counter.get('底稿目录', 0)}"
        )

        # GT_Custom 出现在 10 个文件
        assert norm_counter.get("GT_Custom", 0) == 10, (
            f"'GT_Custom' should appear 10 times, got {norm_counter.get('GT_Custom', 0)}"
        )

    def test_m_cycle_fzhu_pilu_variants(self):
        """验证附注披露变体去重分布"""
        all_names = _get_all_m_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 附注披露信息(上市公司) 出现 9 次
        shangshi_key = "附注披露信息(上市公司)"
        assert norm_counter.get(shangshi_key, 0) == 9, (
            f"'{shangshi_key}' should appear 9 times, "
            f"got {norm_counter.get(shangshi_key, 0)}"
        )

        # 附注披露信息(国有企业) 出现 8 次
        guoqi_key = "附注披露信息(国有企业)"
        assert norm_counter.get(guoqi_key, 0) == 8, (
            f"'{guoqi_key}' should appear 8 times, "
            f"got {norm_counter.get(guoqi_key, 0)}"
        )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        M 循环 chain 必须复用这条 re-export 路径（D/F/H/I/G/J/K/L spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_merge_sheets_dedup_on_m_cycle(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 M 循环 10 文件实际执行合并去重

        以 M1 作为目标 workbook，将其余 9 文件作为 other_files 合并；验证：
        - skipped_historical >= 3（其余 9 文件中的历史遗留 sheet）
        - skipped_dup >= 20（跨文件归一化重复）
        - 合并后逻辑有效 sheet 数 = 65（去重后基线）
        """
        import shutil

        # 选 M1 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in M_TEMPLATE_FILES if f.name.startswith("M1 "))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # 其余 9 个文件作为 other_files
        other_files = [f for f in M_TEMPLATE_FILES if not f.name.startswith("M1 ")]
        assert len(other_files) == 9

        stats = _merge_sheets_dedup(target, other_files)

        # 跨文件重复 sheet 应被去重
        assert stats["skipped_dup"] >= 20, (
            f"应至少跳过 20 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 历史遗留 sheet 应被过滤（M1 自身可能无历史遗留，但其余文件有）
        assert stats["skipped_historical"] >= 3, (
            f"应至少跳过 3 个历史遗留 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 物理 sheet 数
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            all_sheets = wb.sheetnames
        finally:
            wb.close()

        # 逻辑有效 sheet 数 = 65（去重后基线）
        logical_count = sum(
            1 for s in all_sheets if not _should_skip_historical_sheet(s)
        )
        assert logical_count == 65, (
            f"合并后逻辑有效 sheet 数应为 65（dedup baseline），实际 {logical_count}"
        )


class TestPriorCyclesHistoricalFilterRegression:
    """D/F/H/I/G/J/K/L 历史遗留过滤回归测试 — 确保 M spec 无副作用

    M 循环未引入新的过滤模式（仅复用已有"修订前"/"(原)"/"删除"模式），本测试保证
    D/F/H/I/G/J/K/L 既有模式不被破坏，且 M 正常 sheet 不被误过滤。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾）
        ("参考－商誉减值测试示例", True),
        # G 循环历史遗留模式
        ("投资收益实质性程序表G11A-修订前", True),
        # J 循环新增"-删除"模式
        ("股份支付检查表J1-10-删除", True),
        ("应付职工薪酬余额及期后事项检查表J1-11-删除", True),
        # L 循环历史遗留（"（示例）"模式）
        ("函证差异检查表（示例）", True),
        # M 循环 4 个历史遗留
        ("其他权益工具实质性程序表 Q10A (修订前)", True),
        ("未分配利润实质性程序表 Q6A  (修订前)", True),
        ("一般风险准备实质性程序表 Q8A  (修订前)", True),
        ("针对性测试M8-5-删除", True),
        # M 循环正常 sheet 不应被过滤
        ("审定表M2-1", False),
        ("明细表（上市公司）M2-2", False),
        ("明细表（非上市公司）M2-2", False),
        ("明细表M4-2", False),
        ("明细表M5-2", False),
        ("明细表M6-2", False),
        ("明细表M9-2", False),
        ("明细表M10-2", False),
        ("未分配利润实质性程序表 M6A ", False),  # 末尾空格保留
        (" 专项储备实质性程序表 M7A ", False),  # 首尾空格保留
        ("一般风险准备实质性程序表 M8A ", False),  # 末尾空格保留
        ("实收资本实质性程序表M2A", False),
        ("库存股实质性程序表M3A", False),
        ("资本公积实质性程序表M4A", False),
        ("盈余公积实质性程序表M5A", False),
        ("其他综合收益实质性程序表M9A", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露信息(上市公司)", False),
        ("附注披露信息（国有企业）", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I/G/J/K/L 历史遗留过滤模式仍正确工作，M 正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
