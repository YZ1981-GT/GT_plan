"""I 循环合并去重验证测试（spec workpaper-i-intangible-assets-cycle Task 1.1）

验证 chain_orchestrator 对 I 循环 6 文件复用 _merge_sheets_dedup 合并去重：
1. I 循环 6 文件（I1~I6）合并后 sheet 数 = 67（去重后）
2. _should_skip_historical_sheet 对 I3 "参考－商誉减值测试示例" 返回 True（1 历史遗留命中）
3. 跨文件"底稿目录"/"GT_Custom"/"附注披露(上市公司)"/"附注披露(国有企业)"去重
4. I1-10/I1-11 摊销 2 版本 + I4-6/I4-7 摊销 2 版本不被误去重（含括号修饰词）
5. D/F/H 历史遗留过滤回归（现有模式仍正确工作）
6. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "I"

# Collect all I cycle template files (I1~I6)
I_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_i_sheet_names() -> list[str]:
    """从 6 个 I 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in I_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程: 先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_i_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestICycleMergeDedup:
    """I-F1: 6 文件合并去重验证（Sprint 0 baseline: 6 files / 86 raw / 1 historical / 67 dedup）"""

    def test_i_cycle_has_6_template_files(self):
        """I 循环应有 6 个模板文件（I1~I6）"""
        assert len(I_TEMPLATE_FILES) == 6, (
            f"Expected 6 I cycle template files, got {len(I_TEMPLATE_FILES)}: "
            f"{[f.name for f in I_TEMPLATE_FILES]}"
        )

    def test_i_cycle_template_files_named_i1_to_i6(self):
        """I 循环 6 个模板文件命名应为 I1~I6 开头"""
        names = [f.name for f in I_TEMPLATE_FILES]
        for prefix in ["I1", "I2", "I3", "I4", "I5", "I6"]:
            assert any(n.startswith(prefix) for n in names), (
                f"Expected file starting with '{prefix}' in {names}"
            )

    def test_i_cycle_raw_sheet_count_is_86(self):
        """I 循环 6 文件原始 sheet 总数 = 86（Sprint 0 实测基线）"""
        all_names = _get_all_i_sheet_names()
        assert len(all_names) == 86, (
            f"Expected 86 raw sheets, got {len(all_names)}"
        )

    def test_i_cycle_dedup_sheet_count_is_67(self):
        """I 循环合并去重后 sheet 数 = 67（Sprint 0 实测基线）

        86 raw - 1 historical (I3 参考－商誉减值测试示例) - 18 跨文件归一化去重 = 67
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 67, (
            f"Expected 67 dedup sheets, got {dedup_count}"
        )

    def test_i_cycle_historical_sheets_count_is_1(self):
        """I 循环 86 sheet 中恰好 1 个命中历史遗留过滤（I3 参考－商誉减值测试示例）"""
        all_names = _get_all_i_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 1, (
            f"Expected 1 historical sheet hit, got {len(historical_hits)}: {historical_hits}"
        )

    def test_i3_reference_goodwill_impairment_example_is_filtered(self):
        """I3 "参考－商誉减值测试示例" 被现行 regex（"示例"末尾模式）正确过滤

        ADR-I 关键发现：I 循环唯一历史遗留 sheet 已被 D/F spec 现有 regex 覆盖，0 代码改动。
        """
        sheet_name = "参考－商誉减值测试示例"
        assert _should_skip_historical_sheet(sheet_name) is True, (
            f"'{sheet_name}' 应被 _should_skip_historical_sheet 过滤（'示例'末尾模式）"
        )

    def test_cross_file_dedup_removes_duplicates(self):
        """跨文件"底稿目录"/"GT_Custom"/"附注披露"被正确去重

        Sprint 0 实测：底稿目录×6 + GT_Custom×6 + 附注披露(上市)×5 + 附注披露(国企)×5 = 22 raw
        归一化后各仅保留 1 份 → 18 重复需去除（22 - 4 = 18）
        """
        all_names = _get_all_i_sheet_names()

        dedup_targets = {
            "底稿目录": 0,
            "GT_Custom": 0,
            "附注披露(上市公司)": 0,
            "附注披露(国有企业)": 0,
        }

        for name in all_names:
            normalized = _normalize_sheet_name(name)
            if normalized == "底稿目录":
                dedup_targets["底稿目录"] += 1
            elif normalized == "GT_Custom":
                dedup_targets["GT_Custom"] += 1
            elif normalized == "附注披露(上市公司)":
                dedup_targets["附注披露(上市公司)"] += 1
            elif normalized == "附注披露(国有企业)":
                dedup_targets["附注披露(国有企业)"] += 1

        # 每种跨文件重复 sheet 应出现多次（>1 表示有重复需去重）
        for key, count in dedup_targets.items():
            assert count > 1, (
                f"'{key}' 应在多个文件中出现（跨文件去重目标），实际出现 {count} 次"
            )

        # 跨文件归一化总冗余 = sum - 4（每种各保留 1 份）
        total_redundant = sum(dedup_targets.values()) - len(dedup_targets)
        assert total_redundant == 18, (
            f"跨文件归一化冗余预期 18 sheet，实际 {total_redundant}: {dedup_targets}"
        )

    def test_i1_amortization_two_versions_not_misdeduped(self):
        """I1 摊销测算 2 版本（I1-10 不含减值 / I1-11 含减值）不被误去重

        ADR-I1 关键约束：I1-10 / I1-11 各有独立 wp_code 且 sheet 名后缀含括号修饰词，
        归一化后 normalized key 不同，应全部保留。
        """
        all_names = _get_all_i_sheet_names()

        # I1-10 / I1-11 摊销测算 2 版本
        i1_10_sheets = [n for n in all_names if "I1-10" in n]
        i1_11_sheets = [n for n in all_names if "I1-11" in n]

        assert len(i1_10_sheets) >= 1, (
            f"I1-10 摊销测算（不含减值）应至少 1 个 sheet，实际 {i1_10_sheets}"
        )
        assert len(i1_11_sheets) >= 1, (
            f"I1-11 摊销测算（含减值）应至少 1 个 sheet，实际 {i1_11_sheets}"
        )

        # 归一化后不应互相覆盖
        i1_10_normalized = {_normalize_sheet_name(n) for n in i1_10_sheets}
        i1_11_normalized = {_normalize_sheet_name(n) for n in i1_11_sheets}
        assert i1_10_normalized.isdisjoint(i1_11_normalized), (
            f"I1-10 / I1-11 归一化后应不重叠，但有交集: "
            f"I1-10={i1_10_normalized}, I1-11={i1_11_normalized}"
        )

    def test_i4_amortization_two_methods_not_misdeduped(self):
        """I4 长期待摊费用摊销 2 方法（I4-6 直线法 / I4-7 工作量法）不被误去重"""
        all_names = _get_all_i_sheet_names()

        i4_6_sheets = [n for n in all_names if "I4-6" in n]
        i4_7_sheets = [n for n in all_names if "I4-7" in n]

        assert len(i4_6_sheets) >= 1, (
            f"I4-6 摊销测算（直线法）应至少 1 个 sheet，实际 {i4_6_sheets}"
        )
        assert len(i4_7_sheets) >= 1, (
            f"I4-7 摊销测算（工作量法）应至少 1 个 sheet，实际 {i4_7_sheets}"
        )

        i4_6_normalized = {_normalize_sheet_name(n) for n in i4_6_sheets}
        i4_7_normalized = {_normalize_sheet_name(n) for n in i4_7_sheets}
        assert i4_6_normalized.isdisjoint(i4_7_normalized), (
            f"I4-6 / I4-7 归一化后应不重叠，但有交集: "
            f"I4-6={i4_6_normalized}, I4-7={i4_7_normalized}"
        )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        spec 任务文本声明 chain_orchestrator 是 `_merge_sheets_dedup` 等的访问入口；
        I 循环 chain 必须复用这条 0 代码改动的 re-export 路径（D/F/H spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup（D/F/H spec 已注册路径）"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象（确保 0 代码改动复用）
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_find_all_template_files_returns_6_for_i_cycle_codes(self):
        """find_all_template_files 对 I1~I6 任一 wp_code 应能定位到 I 循环模板（多文件支持）"""
        # 实测主入口：任一 I 循环 wp_code 调用 find_all_template_files 应返回 ≥ 1 文件
        for wp_code in ["I1", "I2", "I3", "I4", "I5", "I6"]:
            files = find_all_template_files(wp_code)
            assert len(files) >= 1, (
                f"find_all_template_files('{wp_code}') 应至少返回 1 个文件，实际 {files}"
            )

    def test_merge_sheets_dedup_on_i_cycle_filters_and_dedups(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 I 循环 6 文件实际执行合并去重

        以 I1 为目标，将 I2~I6 作为 other_files 合并；验证：
        - merged 数量 = (raw - historical) - duplicate_dropped = 67 - I1 自己的 sheet 数
        - skipped_historical 包含 1 (I3 参考－商誉减值测试示例)
        - skipped_dup 体现跨文件归一化重复（≥ 18）
        """
        import shutil

        # 复制 I1 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in I_TEMPLATE_FILES if f.name.startswith("I1"))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # I2~I6 作为 other_files
        other_files = [f for f in I_TEMPLATE_FILES if not f.name.startswith("I1")]
        assert len(other_files) == 5

        stats = _merge_sheets_dedup(target, other_files)

        # I3 历史遗留 sheet 应被跳过
        assert stats["skipped_historical"] >= 1, (
            f"应至少跳过 1 个历史遗留 sheet（I3 参考－商誉减值测试示例），"
            f"实际 stats={stats}"
        )

        # 跨文件重复 sheet 应被去重
        assert stats["skipped_dup"] >= 18, (
            f"应至少跳过 18 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 总 sheet 数 = 67（去重总数）
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            final_count = len(wb.sheetnames)
        finally:
            wb.close()
        assert final_count == 67, (
            f"合并去重后目标 workbook sheet 数应为 67，实际 {final_count}"
        )


class TestDFHHistoricalFilterRegression:
    """D/F/H 历史遗留过滤回归测试 — 确保 I spec 不影响现有模式

    I-F1 ADR 关键约束：_should_skip_historical_sheet 现行 regex 已覆盖 I3 唯一历史遗留 sheet，
    无需扩展函数（0 代码改动）。本测试保证 D/F/H 既有模式不被本 spec 引入的回归破坏。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾，I-F1 验证项）
        ("参考－商誉减值测试示例", True),
        # H 循环正常 sheet 不应被过滤
        ("审定表H1-1", False),
        ("折旧测算表（不含减值）-直线法H1-12", False),
        ("审定表（成本模式）H3-1", False),
        # I 循环正常 sheet 不应被过滤（含 "减值" 但不是 "示例" 后缀）
        ("摊销测算表（不含减值）I1-10（剩余年限法）", False),
        ("摊销测算表（含减值）I1-11", False),
        ("减值准备测试表I1-12", False),
        ("商誉减值测试I3-6", False),
        ("研发项目资本化时点判断I2-6", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露(上市公司)", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I 历史遗留过滤模式仍正确工作，正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
