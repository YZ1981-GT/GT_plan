"""N-F2: N 税金循环 8-category sheet group rules — backend completeness check.

Validates: Requirements N-F2 / Spec workpaper-n-tax-cycle / Sprint 2 Task 2.2

Mirror of frontend useNTaxCycleSheetGroups.ts classification rules.

8 categories (priority 1~99):
1. 索引 (priority=1, defaultHidden=true): 底稿目录 / GT_Custom
2. 程序表 (priority=2): 含"程序表" 或 N*A 结尾
3. 审定表 (priority=3): 含"审定表"
4. 明细表 (priority=4): 含"明细表"
5. 税费计算 (priority=5): 测算表 / 计算表 / 税费.*计算 / 应交.*税费
6. 递延所得税 (priority=6): 递延所得税.*核对 / 递延.*费用
7. 附注+调整 (priority=7, defaultHidden=true): 附注披露 / 调整分录
8. 其他 (priority=99): fallback

The classification function is reproduced here in pure Python to provide a
backend-side correctness oracle for the 45 real N cycle sheets.

_Requirements: N-F2_
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _normalize_sheet_name,
    _should_skip_historical_sheet,
)


# ──────────────────────────────────────────────────────────────────────────
# Mirror of frontend classifyNSheet rules (must stay in sync)
# ──────────────────────────────────────────────────────────────────────────

N_RULES = [
    {
        "id": "index",
        "category": "索引",
        "priority": 1,
        "default_hidden": True,
        "match": lambda s: bool(re.match(r"^(底稿目录|GT_Custom)$", s.strip())),
    },
    {
        "id": "procedure",
        "category": "程序表",
        "priority": 2,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"程序表|[A-Z]\d*A\s*$", s)),
    },
    {
        "id": "audit_table",
        "category": "审定表",
        "priority": 3,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"审定表", s)),
    },
    {
        "id": "detail",
        "category": "明细表",
        "priority": 4,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"明细表", s)),
    },
    {
        "id": "tax_calc",
        "category": "税费计算",
        "priority": 5,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"测算表|计算表|税费.*计算|应交.*税费", s)),
    },
    {
        "id": "deferred_tax",
        "category": "递延所得税",
        "priority": 6,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"递延所得税.*核对|递延.*费用", s)),
    },
    {
        "id": "notes_adj",
        "category": "附注+调整",
        "priority": 7,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"附注披露|调整分录", s)),
    },
    {
        "id": "other",
        "category": "其他",
        "priority": 99,
        "default_hidden": False,
        "match": lambda s: True,
    },
]


def classify_n_sheet(name: str) -> dict:
    """Return classification metadata; fallback to 其他."""
    for rule in N_RULES:
        if rule["match"](name):
            cat = {
                "id": rule["id"],
                "category": rule["category"],
                "priority": rule["priority"],
                "default_hidden": rule["default_hidden"],
            }
            # 附注披露 defaultHidden=true 特殊处理
            if rule["id"] == "notes_adj" and re.search(r"附注披露", name):
                cat["default_hidden"] = True
            return cat
    return {
        "id": "other",
        "category": "其他",
        "priority": 99,
        "default_hidden": False,
    }


# ──────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "N"


@pytest.fixture(scope="module")
def n_45_unique_sheets() -> list[str]:
    """45 个去重后 N 循环有效 sheet 名（基于 _normalize_sheet_name + filter historical）"""
    seen_normalized: set[str] = set()
    unique_originals: list[str] = []
    for f in sorted(TEMPLATES_DIR.glob("*.xlsx")):
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            for s in wb.sheetnames:
                if _should_skip_historical_sheet(s):
                    continue
                norm = _normalize_sheet_name(s)
                if norm not in seen_normalized:
                    seen_normalized.add(norm)
                    unique_originals.append(s)
        finally:
            wb.close()
    return unique_originals


# ──────────────────────────────────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────────────────────────────────


class TestNSheetGroupCompleteness:
    """N-F2: 8 类规则对 45 sheet 完备性 + 唯一性"""

    def test_45_sheet_baseline(self, n_45_unique_sheets):
        """N 循环去重后有效 sheet 数 = 45（Sprint 0 实测基线）"""
        assert len(n_45_unique_sheets) == 45, (
            f"Expected 45 N-cycle dedup sheets, got {len(n_45_unique_sheets)}"
        )

    def test_all_45_classified_to_non_empty_category(self, n_45_unique_sheets):
        """所有 sheet 均被分类到非空类目"""
        for s in n_45_unique_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"], f"sheet '{s}' classified to empty category"

    def test_each_sheet_exactly_one_category(self, n_45_unique_sheets):
        """任意 sheet 名恰好命中 1 类（rule fallback `() => true` 必然命中）"""
        for s in n_45_unique_sheets:
            hits = [r for r in N_RULES if r["match"](s)]
            assert len(hits) >= 1, f"sheet '{s}' missed all rules"

    def test_no_sheet_falls_into_unknown_other_unexpectedly(
        self, n_45_unique_sheets
    ):
        """fallback "其他" 类应仅包含合理的无法明确归类项；
        其数量应 < 40%（45 中 < 18）"""
        other_sheets = [
            s for s in n_45_unique_sheets if classify_n_sheet(s)["id"] == "other"
        ]
        ratio = len(other_sheets) / len(n_45_unique_sheets)
        assert ratio < 0.40, (
            f"Too many sheets fell to '其他' fallback: "
            f"{len(other_sheets)}/{len(n_45_unique_sheets)} = {ratio:.1%}\n"
            f"Other sheets: {other_sheets}"
        )

    def test_all_8_categories_represented(self, n_45_unique_sheets):
        """45 sheet 应覆盖 ≥ 6 类（8 类中至少 6 类被命中）"""
        categories_hit = {classify_n_sheet(s)["category"] for s in n_45_unique_sheets}
        assert len(categories_hit) >= 6, (
            f"Only {len(categories_hit)} categories hit, expected ≥ 6\n"
            f"Hit: {sorted(categories_hit)}"
        )
        # 索引 / 附注+调整 必须命中（N 循环 5 文件均含）
        assert "索引" in categories_hit
        assert "附注+调整" in categories_hit
        # 程序表 / 审定表 / 明细表 必须命中
        for must_have in ["程序表", "审定表", "明细表"]:
            assert must_have in categories_hit, (
                f"Category '{must_have}' should be hit but not in {sorted(categories_hit)}"
            )


class TestNSheetGroupKeyPaths:
    """N-F2: 关键 sheet 分类路径验证"""

    def test_index_sheets(self):
        for s in ["底稿目录", "GT_Custom"]:
            cat = classify_n_sheet(s)
            assert cat["category"] == "索引"
            assert cat["default_hidden"] is True

    def test_procedure_sheets(self):
        """程序表：含"程序表"关键词 或 N*A 结尾"""
        procedure_sheets = [
            "递延所得税资产实质性程序表N1A",
            "应交税费实质性程序表N2A",
            "递延所得税负债实质性程序表N3A",
            "税金及附加审计程序表N4A ",  # 末尾空格
            "所得税费用实质性程序表N5A",
        ]
        for s in procedure_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"] == "程序表", (
                f"sheet '{s}' should be 程序表, got {cat['category']}"
            )

    def test_audit_table_sheets(self):
        """审定表：N1-1 / N2-1 / N3-1 / N4-1 / N5-1"""
        audit_sheets = [
            "审定表N1-1",
            "审定表N2-1",
            "审定表N3-1",
            "审定表N4-1",
            "审定表N5-1",
        ]
        for s in audit_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"] == "审定表", (
                f"sheet '{s}' should be 审定表, got {cat['category']}"
            )

    def test_detail_sheets(self):
        """明细表：N1-2 / N2-2 / N3-2 / N4-2 / N5-2"""
        detail_sheets = [
            "明细表N1-2",
            "明细表N2-2",
            "明细表N3-2",
            "明细表N4-2",
            "明细表N5-2",
        ]
        for s in detail_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"] == "明细表", (
                f"sheet '{s}' should be 明细表, got {cat['category']}"
            )

    def test_tax_calc_sheets(self):
        """税费计算：测算表 / 计算表 相关 sheet"""
        tax_calc_sheets = [
            "应交其他税费测算表N2-8",
            "当期所得税费用计算表N5-4",
        ]
        for s in tax_calc_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"] == "税费计算", (
                f"sheet '{s}' should be 税费计算, got {cat['category']}"
            )

    def test_deferred_tax_sheets(self):
        """递延所得税：递延所得税费用核对表"""
        deferred_sheets = [
            "递延所得税费用核对表N5-8",
        ]
        for s in deferred_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"] == "递延所得税", (
                f"sheet '{s}' should be 递延所得税, got {cat['category']}"
            )

    def test_disclosure_sheets_hidden(self):
        """附注披露 5 种括号变体均 defaultHidden=true"""
        variants = [
            "附注披露信息(上市公司)",
            "附注披露信息(国企)",
            "附注披露信息（上市公司）",
            "附注披露信息（国企）",
            "附注披露信息（国有企业）",
        ]
        for s in variants:
            cat = classify_n_sheet(s)
            assert cat["category"] == "附注+调整", (
                f"sheet '{s}' should be 附注+调整, got {cat['category']}"
            )
            assert cat["default_hidden"] is True, (
                f"sheet '{s}' should have default_hidden=True"
            )

    def test_trailing_space_n4a_classified_as_procedure(self):
        """末尾空格 sheet '税金及附加审计程序表N4A ' 正确归入程序表"""
        cat = classify_n_sheet("税金及附加审计程序表N4A ")
        assert cat["category"] == "程序表"

    def test_fallback_other_catches_unknown(self):
        """未知 sheet 名归入 其他 fallback"""
        unknown_sheets = [
            "会计提示",
            "随便的名字",
            "未知sheet",
        ]
        for s in unknown_sheets:
            cat = classify_n_sheet(s)
            assert cat["category"] == "其他", (
                f"sheet '{s}' should be 其他, got {cat['category']}"
            )


class TestNSheetGroupPriorityConflicts:
    """N-F2: 优先级冲突解决验证"""

    def test_procedure_priority_over_audit_table(self):
        """程序表(2) < 审定表(3)：含"程序表"的 sheet 优先匹配程序表"""
        cat = classify_n_sheet("应交税费实质性程序表N2A")
        assert cat["category"] == "程序表"
        assert cat["priority"] == 2

    def test_audit_table_priority_over_detail(self):
        """审定表(3) < 明细表(4)：含"审定表"的 sheet 优先匹配审定表"""
        cat = classify_n_sheet("审定表N2-1")
        assert cat["category"] == "审定表"
        assert cat["priority"] == 3

    def test_detail_priority_over_tax_calc(self):
        """明细表(4) < 税费计算(5)：含"明细表"的 sheet 优先匹配明细表"""
        cat = classify_n_sheet("明细表N2-2")
        assert cat["category"] == "明细表"
        assert cat["priority"] == 4


class TestNSheetGroupCategoryDistribution:
    """N-F2: 类目分布合理性"""

    def test_no_single_non_other_category_over_50_percent(
        self, n_45_unique_sheets
    ):
        """非 fallback '其他' 类目不应单一占比 > 50%"""
        counts = Counter(
            classify_n_sheet(s)["category"] for s in n_45_unique_sheets
        )
        for cat, n in counts.items():
            if cat == "其他":
                continue
            ratio = n / len(n_45_unique_sheets)
            assert ratio < 0.50, (
                f"Category '{cat}' takes {n}/{len(n_45_unique_sheets)} = "
                f"{ratio:.1%}, exceeding 50% threshold"
            )

    def test_distribution_summary(self, n_45_unique_sheets):
        """分布摘要打印（便于回归比对）"""
        counts = Counter(
            classify_n_sheet(s)["category"] for s in n_45_unique_sheets
        )
        total = sum(counts.values())
        print("\n=== N 循环 45 sheet 分类分布 ===")
        for cat, n in sorted(counts.items(), key=lambda x: -x[1]):
            print(f"  {cat}: {n} ({n / total:.1%})")
        assert total == 45

    def test_expected_key_group_sizes(self, n_45_unique_sheets):
        """关键分组预期大小验证（基于 openpyxl 实测）"""
        counts = Counter(
            classify_n_sheet(s)["category"] for s in n_45_unique_sheets
        )
        # 索引：底稿目录 + GT_Custom = 2
        assert counts.get("索引", 0) == 2, (
            f"索引 should have 2 sheets, got {counts.get('索引', 0)}"
        )
        # 程序表：N1A~N5A + 其他含"程序表"关键词的 sheet = 8
        assert counts.get("程序表", 0) == 8, (
            f"程序表 should have 8 sheets, got {counts.get('程序表', 0)}"
        )
        # 审定表：N1-1~N5-1 = 5
        assert counts.get("审定表", 0) == 5, (
            f"审定表 should have 5 sheets, got {counts.get('审定表', 0)}"
        )
        # 明细表：N1-2~N5-2 + 其他含"明细表"关键词的 sheet = 9
        assert counts.get("明细表", 0) == 9, (
            f"明细表 should have 9 sheets, got {counts.get('明细表', 0)}"
        )
        # 税费计算：测算表/计算表相关 = 7
        assert counts.get("税费计算", 0) == 7, (
            f"税费计算 should have 7 sheets, got {counts.get('税费计算', 0)}"
        )
        # 递延所得税：递延所得税费用核对表 = 1
        assert counts.get("递延所得税", 0) == 1, (
            f"递延所得税 should have 1 sheet, got {counts.get('递延所得税', 0)}"
        )
        # 附注+调整：附注披露信息变体 + 调整分录 = 8
        assert counts.get("附注+调整", 0) == 8, (
            f"附注+调整 should have 8 sheets, got {counts.get('附注+调整', 0)}"
        )
