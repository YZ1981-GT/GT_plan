"""审定表（audit-sheet）行项目提取服务单测

守护 render-config audit-sheet 从模板自动解析行项目结构（分节/明细/合计/缩进），
为 GtAuditSheet.vue 可编辑表格提供动态行结构（spec `audit-sheet-editable` Task 3）。

测试覆盖：
- 正常审定表模板解析 → 行数>0 + 含分节(isSection)/合计(isComputed)/缩进(indent>=1)
- 真实模板 D1 应收票据「审定表D1-1」解析（文件缺失则 skip，不阻塞 CI）
- 文件不存在 → 返回 []
- sheet 不存在 → 返回 []
- 空模板（无 项目/科目 表头）→ 返回 []
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.services.wp_audit_sheet_extract import (
    extract_audit_rows,
    extract_audit_rows_from_sheet,
    extract_audit_sections,
)

# 真实模板（cwd=backend 运行；以测试文件定位 backend 根，路径稳定）
_BACKEND_ROOT = Path(__file__).resolve().parents[1]
_REAL_TEMPLATE = _BACKEND_ROOT / "wp_templates" / "D" / "D1 应收票据.xlsx"
_REAL_SHEET = "审定表D1-1"


def _build_audit_sheet(tmp_path, sheet_name="审定表D1-1"):
    """构建一个贴合致同审定表布局的最小 xlsx。

    布局（与 design 的 AuditSheetRow 行结构一致）：
    - 标题区（致同/表名）→ 应被跳过
    - 表头行：A 列「项目」+ 数值列标题
    - 子表头行：项目列为空（未审数/审定数）→ 应被跳过
    - 数据行：分节(一、二、三) + 缩进明细 + 小计/合计
    """
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 标题区（应被跳过）
    ws["A1"] = "致同会计师事务所"
    ws["A2"] = "应收票据审定表"
    # 表头行 5：项目列 + 数值列标题
    ws["A5"] = "项目"
    ws["B5"] = "期初数"
    ws["E5"] = "期末数"
    # 子表头行 6：项目列为空（应被跳过）
    ws["B6"] = "未审数"
    ws["E6"] = "审定数"
    # 数据行
    ws["A7"] = "一、应收票据原值"          # 分节行
    ws["A8"] = "  银行承兑汇票"            # 缩进明细（半角空格×2 → indent 1）
    ws["A9"] = "  商业承兑汇票"
    ws["A10"] = "小计"                     # 小计 → isComputed
    ws["A11"] = "二、应收票据坏账准备"      # 分节行
    ws["A12"] = "  银行承兑汇票"
    ws["A13"] = "小计"
    ws["A14"] = "三、应收票据净值"          # 分节行
    ws["A15"] = "合计"                     # 合计 → isComputed
    # 表体后说明区（应被首个空行截断：row16 空，row17 起为说明/结论区）
    ws["A17"] = "1.审计说明"
    ws["A18"] = "（1）期末净值较期初变动："
    ws["A19"] = "主要原因（比例超过30%的）："
    ws["A20"] = "2.审计结论"
    ws["A21"] = "经审计，认可被审计单位列报金额。"

    # 分节行/合计行加粗
    for coord in ("A7", "A11", "A14", "A10", "A13", "A15"):
        ws[coord].font = Font(bold=True, size=10)

    fp = tmp_path / "audit.xlsx"
    wb.save(str(fp))
    wb.close()
    return fp, sheet_name


# ─────────────────────────────────────────────────────────────────────────
# 正常解析（合成模板，确定性、自包含）
# ─────────────────────────────────────────────────────────────────────────


def test_extract_basic_rows(tmp_path):
    """正常审定表解析：行数>0 + 含分节/合计/缩进。"""
    fp, sn = _build_audit_sheet(tmp_path)
    rows = extract_audit_rows(fp, sn)

    # 行数 > 0
    assert len(rows) > 0
    # 至少一个分节行（一、二、三）
    sections = [r for r in rows if r["isSection"]]
    assert len(sections) >= 1
    # 分节行必为粗体
    assert all(r["bold"] for r in sections)
    # 至少一个合计/小计行
    computed = [r for r in rows if r["isComputed"]]
    assert len(computed) >= 1
    # 至少一个缩进 >= 1 的明细行
    indented = [r for r in rows if r["indent"] >= 1]
    assert len(indented) >= 1


def test_section_total_flags(tmp_path):
    """分节/合计/明细标记正确，且标题区被跳过、说明区被截断。"""
    fp, sn = _build_audit_sheet(tmp_path)
    rows = extract_audit_rows(fp, sn)
    items = [r["item"] for r in rows]

    # 标题行不应出现
    assert "致同会计师事务所" not in items
    assert "应收票据审定表" not in items
    # 子表头行（项目列空）被跳过 → 首行应是第一个分节
    assert rows[0]["item"] == "一、应收票据原值"
    assert rows[0]["isSection"] is True
    assert rows[0]["isComputed"] is False
    # 说明区被空行截断（说明/结论标题及正文不进入行项目）
    assert "1.审计说明" not in items
    assert "2.审计结论" not in items
    assert "经审计，认可被审计单位列报金额。" not in items

    # 分节行
    sec = next(r for r in rows if r["item"] == "二、应收票据坏账准备")
    assert sec["isSection"] is True
    assert sec["bold"] is True
    # 合计行
    total = next(r for r in rows if r["item"] == "合计")
    assert total["isComputed"] is True
    assert total["bold"] is True
    assert total["isSection"] is False
    # 缩进明细行
    detail = next(r for r in rows if r["item"] == "银行承兑汇票")
    assert detail["indent"] >= 1
    assert detail["isSection"] is False
    assert detail["isComputed"] is False


def test_row_schema_fields(tmp_path):
    """每行输出对齐 design 的 AuditSheetRow 后端结构。"""
    fp, sn = _build_audit_sheet(tmp_path)
    rows = extract_audit_rows(fp, sn)
    assert rows  # 非空
    r = rows[0]
    for key in (
        "id", "item", "indent", "bold", "isSection", "isComputed",
        "account_code", "adj_amount", "reclass_amount", "reason",
    ):
        assert key in r, f"缺少字段 {key}"
    # 用户编辑列初始值
    assert r["account_code"] is None
    assert r["adj_amount"] is None
    assert r["reclass_amount"] is None
    assert r["reason"] == ""
    # id 形如 row-{n}，按顺序递增
    ids = [row["id"] for row in rows]
    assert ids[0] == "row-1"
    assert ids == [f"row-{i + 1}" for i in range(len(rows))]


def test_from_sheet_helper(tmp_path):
    """worksheet 级 helper 与文件级一致。"""
    fp, sn = _build_audit_sheet(tmp_path)
    wb = openpyxl.load_workbook(str(fp), data_only=True)
    try:
        ws = wb[sn]
        rows = extract_audit_rows_from_sheet(ws)
    finally:
        wb.close()
    assert len(rows) > 0
    assert any(r["isSection"] for r in rows)
    assert any(r["isComputed"] for r in rows)


# ─────────────────────────────────────────────────────────────────────────
# 真实模板（缺失则 skip，CI 无模板不阻塞）
# ─────────────────────────────────────────────────────────────────────────


@pytest.mark.skipif(
    not _REAL_TEMPLATE.exists(),
    reason=f"真实模板缺失：{_REAL_TEMPLATE}",
)
def test_real_template_d1():
    """真实 D1 应收票据「审定表D1-1」解析：行数>0 + 含分节/合计/缩进。"""
    rows = extract_audit_rows(_REAL_TEMPLATE, _REAL_SHEET)
    assert len(rows) > 0
    # 含分节行
    assert any(r["isSection"] for r in rows)
    # 含合计/小计行
    assert any(r["isComputed"] for r in rows)
    # 含缩进明细行
    assert any(r["indent"] >= 1 for r in rows)
    # 所有行均有完整字段结构
    for r in rows:
        assert set(r.keys()) >= {
            "id", "item", "indent", "bold", "isSection", "isComputed",
            "account_code", "adj_amount", "reclass_amount", "reason",
        }


# ─────────────────────────────────────────────────────────────────────────
# 审计说明 / 审计结论区提取（extract_audit_sections）
# ─────────────────────────────────────────────────────────────────────────


def test_extract_sections_basic(tmp_path):
    """合成模板：审计说明正文 + 审计结论正文均被提取，标题保留。"""
    fp, sn = _build_audit_sheet(tmp_path)
    secs = extract_audit_sections(fp, sn)
    assert secs["notes_label"] == "审计说明"
    assert secs["conclusion_label"] == "审计结论"
    # 说明区正文（标题「1.审计说明」自身不计入正文）
    assert "1.审计说明" not in secs["notes"]
    assert "期末净值较期初变动" in secs["notes"]
    assert "比例超过30%" in secs["notes"]
    # 结论区正文
    assert "认可被审计单位列报金额" in secs["conclusion"]
    assert "2.审计结论" not in secs["conclusion"]


def test_extract_sections_schema(tmp_path):
    """返回结构含 4 个键，均为字符串。"""
    fp, sn = _build_audit_sheet(tmp_path)
    secs = extract_audit_sections(fp, sn)
    for key in ("notes", "conclusion", "notes_label", "conclusion_label"):
        assert key in secs
        assert isinstance(secs[key], str)


def test_extract_sections_missing_file(tmp_path):
    """文件不存在 → 返回空区块（默认标题，正文为空）。"""
    secs = extract_audit_sections(tmp_path / "nope.xlsx", "X")
    assert secs == {
        "notes": "", "conclusion": "",
        "notes_label": "审计说明", "conclusion_label": "审计结论",
    }


def test_extract_sections_no_section_headings(tmp_path):
    """模板无说明/结论标题 → 正文空，标题回退默认。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审定表X"
    ws["A5"] = "项目"
    ws["A6"] = "银行存款"
    fp = tmp_path / "nosec.xlsx"
    wb.save(str(fp))
    wb.close()
    secs = extract_audit_sections(fp, "审定表X")
    assert secs["notes"] == ""
    assert secs["conclusion"] == ""


@pytest.mark.skipif(
    not _REAL_TEMPLATE.exists(),
    reason=f"真实模板缺失：{_REAL_TEMPLATE}",
)
def test_extract_sections_real_template_d1():
    """真实 D1：审计说明区含变动/质押贴现说明，标题为审计说明/审计结论。"""
    secs = extract_audit_sections(_REAL_TEMPLATE, _REAL_SHEET)
    assert secs["notes_label"] == "审计说明"
    assert secs["conclusion_label"] == "审计结论"
    # 说明区含期末净值变动 + 质押贴现说明（模板默认文本）
    assert "净值" in secs["notes"]
    assert "质押" in secs["notes"] or "贴现" in secs["notes"]


# ─────────────────────────────────────────────────────────────────────────
# 降级：返回 []
# ─────────────────────────────────────────────────────────────────────────


def test_missing_file_returns_empty(tmp_path):
    """文件不存在 → 返回 []。"""
    assert extract_audit_rows(tmp_path / "nope.xlsx", "X") == []


def test_missing_sheet_returns_empty(tmp_path):
    """sheet 名不存在 → 返回 []。"""
    fp, _ = _build_audit_sheet(tmp_path)
    assert extract_audit_rows(fp, "不存在的Sheet") == []


def test_empty_template_returns_empty(tmp_path):
    """空模板（无 项目/科目 表头）→ 返回 []。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空表"
    ws["A1"] = "无表头内容"
    fp = tmp_path / "empty.xlsx"
    wb.save(str(fp))
    wb.close()
    assert extract_audit_rows(fp, "空表") == []


def test_zero_byte_file_returns_empty(tmp_path):
    """空字节文件 → 返回 []（不抛异常）。"""
    fp = tmp_path / "zero.xlsx"
    fp.write_bytes(b"")
    assert extract_audit_rows(fp, "X") == []
