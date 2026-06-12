"""Integration tests for ReportExcelExporter cell_mapping / inline-placeholder fill.

Task 20 (audit-report-template-integration): verifies the double-track fill
(design §6) against the real POC template ``soe_standalone.xlsx``:

- Inline ``{{row:BS-xxx:current/prior}}`` placeholders are replaced by audited
  values at their actual coordinates (C6/D6 …), NOT by row order.
- Header placeholders (``{{company_full_name}}`` …) are substituted in-text.
- Formula/total rows (``=SUM`` at C31/D31/C57/D57) are NOT overwritten.
- ``fill_empty_as`` honored for row_codes with no data.

If the POC xlsx is absent, a synthetic workbook with the same inline +
SUM pattern is constructed and the same behavior is asserted.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
import pytest_asyncio
from openpyxl import Workbook, load_workbook
from sqlalchemy import MetaData
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.core import Project
from app.models.report_models import FinancialReport, FinancialReportType

SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON
if not hasattr(SQLiteTypeCompiler, "visit_ARRAY"):
    SQLiteTypeCompiler.visit_ARRAY = lambda self, type_, **kw: "TEXT"
from app.services.report_excel_exporter import ReportExcelExporter
from app.services.template_manifest_loader import resolve_template_base_dir

_POC_PATH = (
    resolve_template_base_dir()
    / "financial_statements"
    / "soe_standalone.xlsx"
)
_BS_SHEET = "1,2-资产负债表(企财01表）"

_engine = create_async_engine("sqlite+aiosqlite:///:memory:")
_TEST_TABLES = [Project.__table__, FinancialReport.__table__]


@pytest_asyncio.fixture
async def db_session():
    async with _engine.begin() as conn:
        await conn.run_sync(
            lambda sync_conn: MetaData().create_all(sync_conn, tables=_TEST_TABLES)
        )
    factory = async_sessionmaker(_engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as session:
        yield session
    async with _engine.begin() as conn:
        await conn.run_sync(
            lambda sync_conn: MetaData().drop_all(sync_conn, tables=_TEST_TABLES)
        )


@pytest_asyncio.fixture
async def soe_project(db_session: AsyncSession):
    project = Project(
        id=uuid.uuid4(),
        name="致同测试国企有限公司",
        client_name="致同测试国企有限公司",
        template_type="soe",
        report_scope="standalone",
    )
    db_session.add(project)
    await db_session.flush()
    return project


@pytest_asyncio.fixture
async def bs_rows(db_session: AsyncSession, soe_project: Project):
    """BS-002 货币资金 has values; BS-003 left empty (fill_empty_as=blank)."""
    now = datetime.now(timezone.utc)
    rows = [
        FinancialReport(
            id=uuid.uuid4(),
            project_id=soe_project.id,
            year=2024,
            report_type=FinancialReportType.balance_sheet,
            row_code="BS-002",
            row_name="货币资金",
            current_period_amount=Decimal("1234567.89"),
            prior_period_amount=Decimal("987654.32"),
            indent_level=1,
            is_total_row=False,
            generated_at=now,
        ),
        FinancialReport(
            id=uuid.uuid4(),
            project_id=soe_project.id,
            year=2024,
            report_type=FinancialReportType.balance_sheet,
            row_code="BS-008",
            row_name="应收账款",
            current_period_amount=Decimal("-50000.00"),
            prior_period_amount=Decimal("60000.00"),
            indent_level=1,
            is_total_row=False,
            generated_at=now,
        ),
    ]
    for r in rows:
        db_session.add(r)
    await db_session.flush()
    return rows


@pytest.mark.skipif(not _POC_PATH.is_file(), reason="POC template absent in checkout")
@pytest.mark.asyncio
async def test_poc_template_inline_fill(db_session, soe_project, bs_rows):
    """Real POC template: inline placeholders filled at C6/D6, SUM preserved."""
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=soe_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
        include_prior_year=True,
    )
    wb = load_workbook(output)
    assert _BS_SHEET in wb.sheetnames
    ws = wb[_BS_SHEET]

    # BS-002 (row 6): audited values replace the {{row:...}} placeholders
    assert ws["C6"].value == pytest.approx(1234567.89)
    assert ws["D6"].value == pytest.approx(987654.32)

    # BS-008 (row 12): negative current value
    assert ws["C12"].value == pytest.approx(-50000.00)
    assert ws["D12"].value == pytest.approx(60000.00)

    # BS-003 (row 7): no data + fill_empty_as=blank → cleared, NOT a placeholder
    assert ws["C7"].value in (None, "")
    assert "{{" not in str(ws["C7"].value or "")

    # Formula/total rows must NOT be overwritten
    for coord in ("C31", "D31", "C56", "D56", "C57", "D57"):
        val = ws[coord].value
        assert isinstance(val, str) and val.startswith("="), (
            f"{coord} formula was overwritten: {val!r}"
        )

    # Header placeholder replaced (A3='编制单位：{{company_full_name}}')
    assert "致同测试国企有限公司" in str(ws["A3"].value)
    assert "{{" not in str(ws["A3"].value)
    # A2 period_end_date
    assert "2024年12月31日" in str(ws["A2"].value)


@pytest.mark.asyncio
async def test_synthetic_inline_fill(db_session, soe_project, bs_rows, tmp_path, monkeypatch):
    """Synthetic template with same inline + SUM pattern (path-independent guard)."""
    # Build a synthetic template mirroring the POC layout
    wb = Workbook()
    ws = wb.active
    ws.title = _BS_SHEET
    ws["A2"] = "{{period_end_date}}"
    ws["A3"] = "编制单位：{{company_full_name}}"
    ws["A6"] = "货币资金"
    ws["C6"] = "{{row:BS-002:current}}"
    ws["D6"] = "{{row:BS-002:prior}}"
    ws["A7"] = "应收账款"
    ws["C7"] = "{{row:BS-008:current}}"
    ws["D7"] = "{{row:BS-008:prior}}"
    ws["A8"] = "空行"
    ws["C8"] = "{{row:BS-003:current}}"  # no data → blank
    ws["A9"] = "合计"
    ws["C9"] = "=SUM(C6:C8)"
    ws["D9"] = "=SUM(D6:D8)"
    synth = tmp_path / "synthetic.xlsx"
    wb.save(synth)

    # Force exporter to load the synthetic template
    def _fake_load(self, template_key):
        return load_workbook(str(synth))

    monkeypatch.setattr(ReportExcelExporter, "_load_template", _fake_load)

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=soe_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
        include_prior_year=True,
    )
    out_wb = load_workbook(output)
    out = out_wb[_BS_SHEET]

    assert out["C6"].value == pytest.approx(1234567.89)
    assert out["D6"].value == pytest.approx(987654.32)
    assert out["C7"].value == pytest.approx(-50000.00)
    # SUM formula preserved
    assert str(out["C9"].value).startswith("=SUM")
    assert str(out["D9"].value).startswith("=SUM")
    # Header substituted
    assert "致同测试国企有限公司" in str(out["A3"].value)
    assert "2024年12月31日" in str(out["A2"].value)


# ─────────────────────────────────────────────────────────────────────────────
# 权益变动表二维矩阵占位符 {{eq:CODE:year_key:col_key}} 填充
# ─────────────────────────────────────────────────────────────────────────────

_EQ_SHEET = "5-所有者权益变动表（企财04表-合并）"


@pytest_asyncio.fixture
async def eq_rows(db_session: AsyncSession, soe_project: Project):
    """EQ-001 上年年末余额：含 eq_matrix 二维数据（本年/上年 × 多列）。

    EQ-002 仅扁平 source_accounts（过渡兼容路径）；EQ-003 无 source_accounts。
    """
    now = datetime.now(timezone.utc)
    rows = [
        FinancialReport(
            id=uuid.uuid4(),
            project_id=soe_project.id,
            year=2024,
            report_type=FinancialReportType.equity_statement,
            row_code="EQ-001",
            row_name="一、上年年末余额",
            current_period_amount=Decimal("0"),
            prior_period_amount=Decimal("0"),
            source_accounts={
                "eq_matrix": {
                    "current_year": {
                        "share_capital": 1000000.00,
                        "capital_reserve": 250000.50,
                        "retained_earnings": -30000.00,
                    },
                    "prior_year": {
                        "share_capital": 900000.00,
                        "capital_reserve": 200000.00,
                    },
                }
            },
            indent_level=0,
            is_total_row=False,
            generated_at=now,
        ),
        FinancialReport(
            id=uuid.uuid4(),
            project_id=soe_project.id,
            year=2024,
            report_type=FinancialReportType.equity_statement,
            row_code="EQ-002",
            row_name="加：会计政策变更",
            source_accounts={"share_capital": 12345.67},  # 扁平过渡结构
            indent_level=1,
            is_total_row=False,
            generated_at=now,
        ),
        FinancialReport(
            id=uuid.uuid4(),
            project_id=soe_project.id,
            year=2024,
            report_type=FinancialReportType.equity_statement,
            row_code="EQ-003",
            row_name="前期差错更正",
            source_accounts=None,
            indent_level=1,
            is_total_row=False,
            generated_at=now,
        ),
    ]
    for r in rows:
        db_session.add(r)
    await db_session.flush()
    return rows


@pytest.mark.asyncio
async def test_eq_matrix_inline_fill(db_session, soe_project, eq_rows, tmp_path, monkeypatch):
    """权益变动表 {{eq:}} 占位被矩阵值替换；N 列 SUM 公式保留；无数据清空。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _EQ_SHEET
    ws["A3"] = "编制单位：{{company_full_name}}"
    # EQ-001 行（含矩阵数据）
    ws["C9"] = "{{eq:EQ-001:current_year:share_capital}}"
    ws["G9"] = "{{eq:EQ-001:current_year:capital_reserve}}"
    ws["M9"] = "{{eq:EQ-001:current_year:retained_earnings}}"
    # 上年列
    ws["P9"] = "{{eq:EQ-001:prior_year:share_capital}}"
    # 缺列（current_year 无 special_reserve）→ 清空
    ws["J9"] = "{{eq:EQ-001:current_year:special_reserve}}"
    # N 列合计公式（不可覆盖）
    ws["N9"] = "=SUM(C9:G9)+SUM(I9:M9)-H9"
    # EQ-002 扁平过渡结构
    ws["C10"] = "{{eq:EQ-002:current_year:share_capital}}"
    # EQ-003 无 source_accounts → 清空
    ws["C11"] = "{{eq:EQ-003:current_year:share_capital}}"
    synth = tmp_path / "synthetic_eq.xlsx"
    wb.save(synth)

    monkeypatch.setattr(
        ReportExcelExporter, "_load_template",
        lambda self, key: load_workbook(str(synth)),
    )

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=soe_project.id,
        year=2024,
        mode="audited",
        report_types=["equity_statement"],
        include_prior_year=True,
    )
    out = load_workbook(output)[_EQ_SHEET]

    # 矩阵单元取数正确
    assert out["C9"].value == pytest.approx(1000000.00)
    assert out["G9"].value == pytest.approx(250000.50)
    assert out["M9"].value == pytest.approx(-30000.00)
    # 上年列
    assert out["P9"].value == pytest.approx(900000.00)
    # 缺列 → 清空，且不残留占位符
    assert out["J9"].value in (None, "")
    assert "{{" not in str(out["J9"].value or "")
    # N 列合计公式保留
    assert str(out["N9"].value).startswith("=SUM")
    # 扁平过渡结构 current_year 命中
    assert out["C10"].value == pytest.approx(12345.67)
    # 无 source_accounts → 清空
    assert out["C11"].value in (None, "")
    assert "{{" not in str(out["C11"].value or "")
    # 全表无残留 eq 占位符
    for row in out.iter_rows():
        for cell in row:
            assert "{{eq:" not in str(cell.value or ""), f"残留: {cell.coordinate}={cell.value!r}"


@pytest.mark.asyncio
async def test_eq_prior_year_excluded_when_disabled(
    db_session, soe_project, eq_rows, tmp_path, monkeypatch
):
    """include_prior_year=False 时 prior_year 占位被清空（不写值、不残留）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _EQ_SHEET
    ws["C9"] = "{{eq:EQ-001:current_year:share_capital}}"
    ws["P9"] = "{{eq:EQ-001:prior_year:share_capital}}"
    synth = tmp_path / "synthetic_eq2.xlsx"
    wb.save(synth)

    monkeypatch.setattr(
        ReportExcelExporter, "_load_template",
        lambda self, key: load_workbook(str(synth)),
    )

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=soe_project.id,
        year=2024,
        mode="audited",
        report_types=["equity_statement"],
        include_prior_year=False,
    )
    out = load_workbook(output)[_EQ_SHEET]

    assert out["C9"].value == pytest.approx(1000000.00)
    # prior_year 被禁用 → 清空
    assert out["P9"].value in (None, "")
    assert "{{" not in str(out["P9"].value or "")
