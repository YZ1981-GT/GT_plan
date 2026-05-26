"""Tests for wp_xlsx_export_service — 致同模板填值导出

验证 4 路径写入策略：
  1. fixed_cells 写入
  2. dynamic_table 写入
  3. formulas 保留（跳过）
  4. static_text + merged_cells 保留（跳过）

Requirements: 4.3.1.a-g（方案 C 还原 7 项约束）
"""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from app.services.wp_xlsx_export_service import (
    ExportValidationError,
    TemplateNotFoundError,
    _get_nested,
    _render_template_var,
    _sync_export_workpaper_xlsx,
    _write_dynamic_table,
    export_workpaper_xlsx,
)


# ─── Unit tests for helper functions ─────────────────────────────────────


class TestRenderTemplateVar:
    """Test _render_template_var placeholder substitution."""

    def test_no_placeholder_returns_original(self):
        assert _render_template_var("致同会计师事务所", {}) == "致同会计师事务所"

    def test_entity_name_from_entity_name(self):
        meta = {"entity_name": "测试公司"}
        assert _render_template_var("${entity_name}", meta) == "测试公司"

    def test_entity_name_fallback_to_client_name(self):
        meta = {"client_name": "客户公司"}
        assert _render_template_var("${entity_name}", meta) == "客户公司"

    def test_period_end(self):
        meta = {"period_end": "2025-12-31"}
        assert _render_template_var("${period_end}", meta) == "2025-12-31"

    def test_index_no(self):
        meta = {"index_no": "D2A-001"}
        assert _render_template_var("${index_no}", meta) == "D2A-001"

    def test_page_no(self):
        meta = {"page_no": "1/1"}
        assert _render_template_var("${page_no}", meta) == "1/1"

    def test_missing_var_returns_empty(self):
        assert _render_template_var("${nonexistent}", {}) == ""

    def test_mixed_text_and_placeholder(self):
        meta = {"entity_name": "ABC公司"}
        result = _render_template_var("被审计单位：${entity_name}", meta)
        assert result == "被审计单位：ABC公司"

    def test_multiple_placeholders(self):
        meta = {"entity_name": "ABC", "period_end": "2025"}
        result = _render_template_var("${entity_name} - ${period_end}", meta)
        assert result == "ABC - 2025"


class TestGetNested:
    """Test _get_nested dot-path accessor."""

    def test_simple_key(self):
        assert _get_nested({"name": "test"}, "name") == "test"

    def test_nested_key(self):
        data = {"assertion": {"existence": True}}
        assert _get_nested(data, "assertion.existence") is True

    def test_missing_key_returns_none(self):
        assert _get_nested({"a": 1}, "b") is None

    def test_missing_nested_key_returns_none(self):
        assert _get_nested({"a": {"b": 1}}, "a.c") is None

    def test_non_dict_intermediate_returns_none(self):
        assert _get_nested({"a": "string"}, "a.b") is None

    def test_deeply_nested(self):
        data = {"a": {"b": {"c": {"d": 42}}}}
        assert _get_nested(data, "a.b.c.d") == 42


class TestWriteDynamicTable:
    """Test _write_dynamic_table writes rows correctly."""

    def test_writes_text_values(self):
        """Text fields are written to cells."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        table_schema = {
            "start_row": 2,
            "columns": {
                "A": {"field": "name", "type": "text"},
                "B": {"field": "desc", "type": "text"},
            },
        }
        rows = [
            {"name": "程序1", "desc": "描述1"},
            {"name": "程序2", "desc": "描述2"},
        ]

        _write_dynamic_table(ws, table_schema, rows)

        assert ws["A2"].value == "程序1"
        assert ws["B2"].value == "描述1"
        assert ws["A3"].value == "程序2"
        assert ws["B3"].value == "描述2"

    def test_writes_checkmark_for_boolean(self):
        """Boolean fields with render='checkmark' write √ or empty."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        table_schema = {
            "start_row": 1,
            "columns": {
                "A": {"field": "checked", "type": "boolean", "render": "checkmark"},
            },
        }
        rows = [{"checked": True}, {"checked": False}, {"checked": None}]

        _write_dynamic_table(ws, table_schema, rows)

        assert ws["A1"].value == "√"
        assert ws["A2"].value == ""
        assert ws["A3"].value == ""

    def test_writes_number_values(self):
        """Number fields are converted to float."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        table_schema = {
            "start_row": 1,
            "columns": {
                "A": {"field": "seq", "type": "number"},
            },
        }
        rows = [{"seq": 1}, {"seq": "2.5"}, {"seq": None}]

        _write_dynamic_table(ws, table_schema, rows)

        assert ws["A1"].value == 1.0
        assert ws["A2"].value == 2.5
        assert ws["A3"].value is None

    def test_preserves_formula_cells(self):
        """Cells with existing formulas are not overwritten."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(B1:C1)"  # Pre-existing formula

        table_schema = {
            "start_row": 1,
            "columns": {
                "A": {"field": "value", "type": "text"},
            },
        }
        rows = [{"value": "should not overwrite"}]

        _write_dynamic_table(ws, table_schema, rows)

        # Formula should be preserved
        assert ws["A1"].value == "=SUM(B1:C1)"

    def test_nested_field_path(self):
        """Dot-separated field paths access nested data."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        table_schema = {
            "start_row": 1,
            "columns": {
                "A": {"field": "assertion.existence", "type": "boolean", "render": "checkmark"},
            },
        }
        rows = [{"assertion": {"existence": True}}]

        _write_dynamic_table(ws, table_schema, rows)

        assert ws["A1"].value == "√"

    def test_string_col_def_treated_as_text(self):
        """String column definitions are treated as text type."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        table_schema = {
            "start_row": 1,
            "columns": {
                "A": "name",  # Simple string format
            },
        }
        rows = [{"name": "test value"}]

        _write_dynamic_table(ws, table_schema, rows)

        assert ws["A1"].value == "test value"


class TestSyncExportWorkpaperXlsx:
    """Test _sync_export_workpaper_xlsx with a real openpyxl workbook."""

    def test_template_not_found_raises(self):
        """Missing template raises TemplateNotFoundError."""
        schema = {
            "wp_code": "NONEXIST",
            "template_path": "backend/wp_templates/X/nonexistent.xlsx",
            "sheets": {},
        }
        with pytest.raises(TemplateNotFoundError):
            _sync_export_workpaper_xlsx(schema, {}, {})

    def test_export_with_temp_template(self, tmp_path):
        """Full export flow with a temporary template file."""
        import openpyxl

        # Create a temporary template
        template_file = tmp_path / "test_template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header"
        ws["B1"] = "=A1"  # Formula to preserve
        ws.merge_cells("C1:D1")  # Merged cells to preserve
        wb.save(str(template_file))

        schema = {
            "wp_code": "TEST",
            "template_path": str(template_file),
            "sheets": {
                "TestSheet": {
                    "fixed_cells": {
                        "A1": "${entity_name}",
                    },
                    "dynamic_table": {
                        "start_row": 3,
                        "columns": {
                            "A": {"field": "name", "type": "text"},
                        },
                    },
                },
            },
        }
        html_data = {
            "TestSheet": {
                "rows": [{"name": "Row 1"}, {"name": "Row 2"}],
            },
        }
        project_meta = {"entity_name": "测试公司", "period_end": "2025-12-31"}

        result = _sync_export_workpaper_xlsx(schema, html_data, project_meta)

        # Verify result is BytesIO
        assert isinstance(result, BytesIO)

        # Reload and verify
        result.seek(0)
        wb2 = openpyxl.load_workbook(result, data_only=False)
        ws2 = wb2["TestSheet"]

        # fixed_cells: entity_name written
        assert ws2["A1"].value == "测试公司"

        # formula preserved
        assert ws2["B1"].value == "=A1"

        # merged_cells preserved
        assert len(ws2.merged_cells.ranges) > 0

        # dynamic_table rows written
        assert ws2["A3"].value == "Row 1"
        assert ws2["A4"].value == "Row 2"

        wb2.close()


class TestExportWorkpaperXlsx:
    """Test the async export_workpaper_xlsx function."""

    def test_missing_entity_name_raises_validation_error(self):
        """Missing entity_name raises ExportValidationError."""
        with pytest.raises(ExportValidationError) as exc_info:
            asyncio.run(
                export_workpaper_xlsx(
                    wp_code="D2A",
                    html_data={},
                    schema={"template_path": "x.xlsx", "sheets": {}},
                    project_meta={"period_end": "2025-12-31"},
                )
            )
        assert "entity_name" in exc_info.value.missing_fields

    def test_missing_period_end_raises_validation_error(self):
        """Missing period_end raises ExportValidationError."""
        with pytest.raises(ExportValidationError) as exc_info:
            asyncio.run(
                export_workpaper_xlsx(
                    wp_code="D2A",
                    html_data={},
                    schema={"template_path": "x.xlsx", "sheets": {}},
                    project_meta={"entity_name": "Test"},
                )
            )
        assert "period_end" in exc_info.value.missing_fields

    def test_successful_export_with_temp_template(self, tmp_path):
        """Async export succeeds with valid inputs."""
        import openpyxl

        # Create a temporary template
        template_file = tmp_path / "async_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "placeholder"
        wb.save(str(template_file))

        schema = {
            "wp_code": "TEST",
            "template_path": str(template_file),
            "sheets": {
                "Sheet1": {
                    "fixed_cells": {"A1": "${entity_name}"},
                },
            },
        }
        html_data = {}
        project_meta = {"entity_name": "异步测试公司", "period_end": "2025-12-31"}

        result = asyncio.run(
            export_workpaper_xlsx("TEST", html_data, schema, project_meta)
        )

        assert isinstance(result, BytesIO)
        result.seek(0)
        wb2 = openpyxl.load_workbook(result)
        assert wb2["Sheet1"]["A1"].value == "异步测试公司"
        wb2.close()

    def test_template_not_found_propagates(self):
        """TemplateNotFoundError propagates through async wrapper."""
        schema = {
            "wp_code": "MISSING",
            "template_path": "nonexistent/path.xlsx",
            "sheets": {},
        }
        project_meta = {"entity_name": "Test", "period_end": "2025-12-31"}

        with pytest.raises(TemplateNotFoundError):
            asyncio.run(
                export_workpaper_xlsx("MISSING", {}, schema, project_meta)
            )
