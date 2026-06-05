"""批量建项服务 Property-Based Tests + 单元测试。

Feature: project-creation-enhancement
Properties 10-11 + unit tests
"""

import uuid

import pytest
import pytest_asyncio
import hypothesis.strategies as st
from hypothesis import given, settings, HealthCheck
from openpyxl import Workbook
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.base import Base
from app.services.uscc_validator import USCC_CHARSET, _CHAR_TO_VALUE, _WEIGHTS
from app.services.batch_project_service import (
    BatchImportResult,
    generate_template,
    parse_and_import,
    export_projects,
    _TEMPLATE_COLUMNS,
)


# ---------------------------------------------------------------------------
# 辅助：构造合法 USCC
# ---------------------------------------------------------------------------

def _compute_check_digit(prefix: str) -> str:
    """根据 17 位前缀计算第 18 位校验码字符。"""
    total = 0
    for i in range(17):
        total += _CHAR_TO_VALUE[prefix[i]] * _WEIGHTS[i]
    remainder = total % 31
    check_digit = 31 - remainder
    if check_digit == 31:
        check_digit = 0
    return USCC_CHARSET[check_digit]


def make_valid_uscc(prefix_17: str) -> str:
    """从 17 位前缀构造合法 18 位 USCC。"""
    return prefix_17 + _compute_check_digit(prefix_17)


FIXED_USCC_PREFIX = "91110000710931130"
FIXED_USCC = make_valid_uscc(FIXED_USCC_PREFIX)


# ---------------------------------------------------------------------------
# DB Fixtures（in-memory SQLite）
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture
async def engine():
    """创建测试用 in-memory SQLite 引擎。"""
    from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
    SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON
    if not hasattr(SQLiteTypeCompiler, "visit_ARRAY"):
        SQLiteTypeCompiler.visit_ARRAY = lambda self, type_, **kw: "TEXT"

    eng = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)
    async with eng.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield eng
    await eng.dispose()


@pytest_asyncio.fixture
async def db_session(engine):
    """创建独立的 DB session。"""
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as session:
        yield session


# ---------------------------------------------------------------------------
# 辅助：创建包含数据行的 Excel 文件
# ---------------------------------------------------------------------------

def _make_import_excel(rows: list[list]) -> bytes:
    """构造包含给定数据行的 Excel 文件 bytes。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(_TEMPLATE_COLUMNS)
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Property 10: 批量导入结果计数一致性
# **Validates: Requirements 5.6**
#
# For any batch import file with N rows, success_count + fail_count == N,
# and len(failures) == fail_count.
# ---------------------------------------------------------------------------

# Strategy: generate N rows with mix of valid/invalid data
_uscc_prefix_st = st.text(alphabet=USCC_CHARSET, min_size=17, max_size=17)


@given(
    n_valid=st.integers(min_value=0, max_value=3),
    n_invalid=st.integers(min_value=0, max_value=3),
    prefix=_uscc_prefix_st,
)
@settings(max_examples=5, suppress_health_check=[HealthCheck.function_scoped_fixture])
@pytest.mark.asyncio
async def test_batch_import_count_consistency(n_valid: int, n_invalid: int, prefix: str, engine):
    """Property 10: success_count + fail_count == N, len(failures) == fail_count."""
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    rows: list[list] = []

    # Generate valid rows with unique USCCs
    for i in range(n_valid):
        # Ensure unique USCC per row by modifying prefix suffix
        p = prefix[:15] + USCC_CHARSET[i % len(USCC_CHARSET)] + USCC_CHARSET[(i + 5) % len(USCC_CHARSET)]
        uscc = make_valid_uscc(p)
        rows.append([
            f"客户{i}",
            uscc,
            f"简称{i}",
            2090 + i,  # unique year to avoid collision
            "年报审计",
            "企业会计准则",
            "单户",
        ])

    # Generate invalid rows (empty company_code)
    for i in range(n_invalid):
        rows.append([
            f"无效客户{i}",
            "",  # empty USCC → fail
            f"无效简称{i}",
            2025,
            "年报审计",
            "企业会计准则",
            "单户",
        ])

    if not rows:
        return  # N=0, nothing to test

    file_bytes = _make_import_excel(rows)

    async with factory() as db:
        result = await parse_and_import(file_bytes, db)

    total_rows = n_valid + n_invalid
    assert result.success_count + result.fail_count == total_rows
    assert len(result.failures) == result.fail_count


# ---------------------------------------------------------------------------
# Property 11: 批量导出/导入回环解析
# **Validates: Requirements 5.8, 5.9**
#
# For any set of existing projects, exporting them to Excel and then parsing
# that Excel through the batch import parser SHALL successfully parse every row.
# ---------------------------------------------------------------------------

@given(prefix=_uscc_prefix_st)
@settings(max_examples=5, suppress_health_check=[HealthCheck.function_scoped_fixture])
@pytest.mark.asyncio
async def test_export_import_round_trip(prefix: str, engine):
    """Property 11: Export → re-parse → all rows parseable (field structure compatible)."""
    from app.models.audit_platform_schemas import BasicInfoSchema
    from app.services.project_wizard_service import create_project

    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with factory() as db:
        # Create a project
        uscc = make_valid_uscc(prefix)
        data = BasicInfoSchema(
            client_name="回环测试客户",
            audit_year=2060,
            project_type="annual",
            accounting_standard="enterprise",
            company_code=uscc,
            short_name="回环简称",
            report_scope="standalone",
        )
        project = await create_project(data, db)

        # Export
        export_output = await export_projects([project.id], db)
        export_bytes = export_output.getvalue()

    # 验证导出 Excel 列值正确（包括 accounting_standard 中文反映射）
    from openpyxl import load_workbook as _load_wb
    _wb = _load_wb(BytesIO(export_bytes), read_only=True)
    _ws = _wb["数据"]
    _exported_row = list(_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert _exported_row[0] == "回环测试客户"  # 客户名称
    assert _exported_row[2] == "回环简称"  # 项目简称
    assert _exported_row[5] == "企业会计准则"  # accounting_standard 正确反映射
    assert _exported_row[6] == "单户"  # report_scope 正确反映射

    # Parse the exported file in a fresh session (to avoid uniqueness collision,
    # we test that parsing succeeds structurally by checking no format errors)
    async with factory() as db2:
        # The exported file will fail uniqueness (same project exists),
        # but the parsing itself should work — we verify the errors are
        # ONLY uniqueness-related, not format/structure errors.
        result = await parse_and_import(export_bytes, db2)

        # Every row should be parseable: either success (if unique) or
        # uniqueness error (not a parsing/format error)
        for failure in result.failures:
            for err in failure.errors:
                # Structure/format errors would mention things like:
                # "为必填项", "格式错误", "无效"
                # Uniqueness errors mention "已存在"
                assert "已存在" in err, (
                    f"Round-trip produced non-uniqueness error: {err}"
                )


# ---------------------------------------------------------------------------
# Unit Test: 空文件
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_empty_file_returns_zero(db_session):
    """Empty Excel (no data rows) → success=0, fail=0."""
    file_bytes = _make_import_excel([])
    result = await parse_and_import(file_bytes, db_session)
    assert result.success_count == 0
    assert result.fail_count == 0
    assert result.failures == []


# ---------------------------------------------------------------------------
# Unit Test: 全部无效行
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_all_invalid_rows(db_session):
    """All rows invalid → fail_count == row count, success_count == 0.
    Note: fully empty rows are skipped (not counted as failures).
    """
    rows = [
        ["有客户名", "INVALID_USCC", "", "abc", "无效类型", "无效准则", "无效范围"],
        ["客户", "12345", "简称", "2025", "年报审计", "企业会计准则", "单户"],  # USCC too short
        ["客户B", "", "简称B", "2025", "年报审计", "企业会计准则", "单户"],  # empty USCC
    ]
    file_bytes = _make_import_excel(rows)
    result = await parse_and_import(file_bytes, db_session)
    assert result.success_count == 0
    assert result.fail_count == 3
    assert len(result.failures) == 3


# ---------------------------------------------------------------------------
# Unit Test: 混合有效/无效行
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_mixed_valid_invalid(db_session):
    """Mix of valid and invalid rows → correct counts."""
    uscc1 = make_valid_uscc("91320000715837650")
    rows = [
        # Valid row
        ["好客户", uscc1, "好简称", 2025, "年报审计", "企业会计准则", "单户"],
        # Invalid row (no short_name)
        ["另一客户", FIXED_USCC, "", 2025, "年报审计", "企业会计准则", ""],
    ]
    file_bytes = _make_import_excel(rows)
    result = await parse_and_import(file_bytes, db_session)
    assert result.success_count == 1
    assert result.fail_count == 1
    assert result.failures[0].row_number == 3  # row 3 (header is row 1, data starts row 2)
    assert "项目简称为必填项" in result.failures[0].errors


# ---------------------------------------------------------------------------
# Unit Test: generate_template 结构验证
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_generate_template_structure():
    """Template has correct sheets and column headers."""
    from openpyxl import load_workbook

    output = await generate_template()
    wb = load_workbook(output)

    assert "数据" in wb.sheetnames
    assert "说明事项" in wb.sheetnames

    ws_data = wb["数据"]
    headers = [cell.value for cell in ws_data[1]]
    assert headers == _TEMPLATE_COLUMNS
