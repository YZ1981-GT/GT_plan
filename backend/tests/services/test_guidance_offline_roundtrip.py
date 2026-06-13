"""离线导出 _meta_ 含 guidance_text — note-guidance-text-separation Task 8.3 (xlsx part)."""

from __future__ import annotations

from app.services.note_offline_export_service import export_sections_to_xlsx, _decompress_meta


def test_offline_export_meta_includes_guidance_text():
    sections = [{
        "section_id": "八、1",
        "section_title": "货币资金",
        "guidance_text": "（注：应披露受限货币资金。）",
        "table_data": {"headers": ["项目"], "rows": []},
        "_bindings": {},
        "_formulas": {},
        "_row_meta": [],
        "_cell_modes": {},
        "_cell_meta": {},
        "_dynamic_regions": [],
    }]
    xlsx_bytes, _ = export_sections_to_xlsx(
        sections,
        include_formulas=False,
        include_provenance=False,
        exporter_name="test",
        project_name="test",
        year=2025,
    )
    assert len(xlsx_bytes) > 100

    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(xlsx_bytes))
    assert "_meta_" in wb.sheetnames
    ws = wb["_meta_"]
    compressed = ws.cell(row=1, column=2).value
    meta = _decompress_meta(str(compressed))
    assert meta["八、1"]["guidance_text"] == "（注：应披露受限货币资金。）"


def test_offline_import_meta_roundtrip_from_export():
    """Property 13: 导出 _meta_ 含 guidance，validate 可解出。"""
    from app.services.note_offline_import_service import validate_import_file

    guidance = "（注：应披露受限货币资金。）"
    sections = [{
        "section_id": "八、1",
        "section_title": "货币资金",
        "guidance_text": guidance,
        "table_data": {"headers": ["项目"], "rows": []},
        "_bindings": {}, "_formulas": {}, "_row_meta": [],
        "_cell_modes": {}, "_cell_meta": {}, "_dynamic_regions": [],
    }]
    xlsx_bytes, _ = export_sections_to_xlsx(
        sections, include_formulas=False, include_provenance=False,
        exporter_name="test", project_name="test", year=2025,
    )
    validation = validate_import_file(xlsx_bytes)
    assert validation.valid
    assert validation.meta_data["八、1"]["guidance_text"] == guidance
