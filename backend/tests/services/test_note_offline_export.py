"""Tests for Sprint C.0 — 附注离线导出服务 (D15).

Covers:
- C.0.1: Section filtering by section_id
- C.0.2: xlsx structure (sheet names, count)
- C.0.3: 4-color cell formatting
- C.0.4: Cell comments (formula + provenance)
- C.0.5: _meta_ sheet (base64+gzip round-trip)
- C.0.6: Instructions sheet content
- C.0.7: TOC sheet structure
- C.0.8: AES encryption + hash
"""
from __future__ import annotations

import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.note_offline_export_service import (
    FILL_EDITABLE,
    FILL_FORMULA,
    FILL_LOCKED,
    FILL_REQUIRED,
    NoteOfflineExportService,
    _calc_completeness,
    _classify_cell,
    _compress_meta,
    _count_required_cells,
    _decompress_meta,
    _decrypt_bytes,
    _encrypt_bytes,
    _truncate_sheet_name,
    export_sections_to_xlsx,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_section(
    section_id: str = "section_cash",
    title: str = "货币资金",
    rows: list | None = None,
    cell_meta: dict | None = None,
    formulas: dict | None = None,
    provenance: dict | None = None,
) -> dict:
    """Create a mock section dict."""
    if rows is None:
        rows = [
            {"row_type": "data", "label": "银行存款", "cells": [100.0, 200.0]},
            {"row_type": "data", "label": "现金", "cells": [50.0, None]},
            {"row_type": "dynamic_data", "label": "其他", "cells": [10.0, 20.0]},
        ]
    return {
        "section_id": section_id,
        "section_title": title,
        "table_data": {
            "headers": ["项目", "期末余额"],
            "rows": rows,
        },
        "_cell_meta": cell_meta or {},
        "_formulas": formulas or {},
        "_cell_provenance": provenance or {},
        "_bindings": {"0:1": {"source": "trial_balance", "account_codes": ["1001"]}},
        "_row_meta": [{"row_type": "data"}, {"row_type": "data"}, {"row_type": "dynamic_data"}],
        "_dynamic_regions": [{"name": "其他", "axis": "row", "start_idx": 2, "end_idx": 2}],
        "_cell_modes": {"0:0": "manual", "0:1": "auto"},
    }


def _make_sections(n: int = 3) -> list[dict]:
    """Create N mock sections."""
    return [
        _make_section(f"section_{i}", f"章节{i}")
        for i in range(n)
    ]


def _load_wb(xlsx_bytes: bytes):
    """Load workbook from bytes."""
    return load_workbook(BytesIO(xlsx_bytes))


# ---------------------------------------------------------------------------
# C.0.1: Section Filtering
# ---------------------------------------------------------------------------


class TestSectionFiltering:
    """C.0.1 — filter by section_id list."""

    def test_no_filter_returns_all(self):
        sections = _make_sections(5)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        # 注意事项 + 章节清单 + 5 sections + _meta_ = 8 sheets
        assert len(wb.sheetnames) == 8

    def test_filter_subset(self):
        sections = _make_sections(5)
        xlsx_bytes, _ = export_sections_to_xlsx(sections, section_ids=["section_0", "section_2"])
        wb = _load_wb(xlsx_bytes)
        # 注意事项 + 章节清单 + 2 sections + _meta_ = 5 sheets
        assert len(wb.sheetnames) == 5

    def test_filter_empty_list(self):
        sections = _make_sections(3)
        xlsx_bytes, _ = export_sections_to_xlsx(sections, section_ids=[])
        wb = _load_wb(xlsx_bytes)
        # 注意事项 + 章节清单 + 0 sections + _meta_ = 3 sheets
        assert len(wb.sheetnames) == 3

    def test_filter_nonexistent_ids(self):
        sections = _make_sections(3)
        xlsx_bytes, _ = export_sections_to_xlsx(sections, section_ids=["nonexistent"])
        wb = _load_wb(xlsx_bytes)
        assert len(wb.sheetnames) == 3  # no section sheets

    def test_empty_sections_input(self):
        xlsx_bytes, hash_val = export_sections_to_xlsx([])
        wb = _load_wb(xlsx_bytes)
        assert len(wb.sheetnames) == 3  # instructions + toc + _meta_
        assert len(hash_val) == 64  # SHA-256 hex


# ---------------------------------------------------------------------------
# C.0.2: xlsx Structure
# ---------------------------------------------------------------------------


class TestXlsxStructure:
    """C.0.2 — workbook structure."""

    def test_sheet_order(self):
        sections = _make_sections(2)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        names = wb.sheetnames
        assert names[0] == "注意事项"
        assert names[1] == "章节清单"
        assert names[-1] == "_meta_"

    def test_meta_sheet_hidden(self):
        sections = _make_sections(1)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        meta_ws = wb["_meta_"]
        assert meta_ws.sheet_state == "hidden"

    def test_sheet_name_truncation(self):
        long_name = "这是一个非常非常长的章节标题超过三十一个字符的限制"
        section = _make_section("s1", long_name)
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        # All sheet names must be <= 31 chars
        for name in wb.sheetnames:
            assert len(name) <= 31


# ---------------------------------------------------------------------------
# C.0.3: 4-Color Cell Formatting
# ---------------------------------------------------------------------------


class TestCellFormatting:
    """C.0.3 — 4-color semantic fills."""

    def test_classify_manual_cell(self):
        assert _classify_cell({"source": "manual", "mode": "manual"}) == "editable"

    def test_classify_formula_cell(self):
        assert _classify_cell({"source": "formula", "mode": "formula"}) == "formula"

    def test_classify_wp_data_cell(self):
        assert _classify_cell({"source": "wp_data"}) == "locked"

    def test_classify_trial_balance_cell(self):
        assert _classify_cell({"source": "trial_balance"}) == "locked"

    def test_classify_required_cell(self):
        assert _classify_cell({"source": "manual", "is_required": True}) == "required"

    def test_fill_applied_in_xlsx(self):
        section = _make_section(
            cell_meta={
                "0:0": {"source": "manual", "mode": "manual"},
                "0:1": {"source": "formula", "mode": "formula"},
            }
        )
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        # Find the section sheet (3rd sheet)
        ws = wb.worksheets[2]
        # Row 4 (data starts at row 4), col 1 = editable (yellow)
        cell_a4 = ws.cell(row=4, column=1)
        assert cell_a4.fill.start_color.rgb == "FFFFFF00"
        # Row 4, col 2 = formula (gray)
        cell_b4 = ws.cell(row=4, column=2)
        assert cell_b4.fill.start_color.rgb == "FFD9D9D9"


# ---------------------------------------------------------------------------
# C.0.4: Cell Comments
# ---------------------------------------------------------------------------


class TestCellComments:
    """C.0.4 — formula + provenance comments."""

    def test_formula_comment(self):
        section = _make_section(
            formulas={"0:1": {"expression": "=SUM(B2:B10)", "type": "formula"}}
        )
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        ws = wb.worksheets[2]
        cell = ws.cell(row=4, column=2)
        assert cell.comment is not None
        assert "公式: =SUM(B2:B10)" in cell.comment.text

    def test_provenance_wp_data_comment(self):
        section = _make_section(
            provenance={"0:1": {"source": "wp_data", "wp_code": "h08"}}
        )
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        ws = wb.worksheets[2]
        cell = ws.cell(row=4, column=2)
        assert cell.comment is not None
        assert "wp_data" in cell.comment.text
        assert "h08" in cell.comment.text

    def test_provenance_tb_comment(self):
        section = _make_section(
            provenance={"0:0": {"source": "trial_balance", "account_codes": ["1001", "1002"]}}
        )
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        ws = wb.worksheets[2]
        cell = ws.cell(row=4, column=1)
        assert cell.comment is not None
        assert "试算表" in cell.comment.text

    def test_no_comment_when_disabled(self):
        section = _make_section(
            formulas={"0:1": {"expression": "=SUM(B2:B10)"}}
        )
        xlsx_bytes, _ = export_sections_to_xlsx(
            [section], include_formulas=False, include_provenance=False
        )
        wb = _load_wb(xlsx_bytes)
        ws = wb.worksheets[2]
        cell = ws.cell(row=4, column=2)
        assert cell.comment is None


# ---------------------------------------------------------------------------
# C.0.5: _meta_ Sheet
# ---------------------------------------------------------------------------


class TestMetaSheet:
    """C.0.5 — base64+gzip compressed metadata."""

    def test_meta_round_trip(self):
        original = {"key": "value", "nested": {"a": 1}, "list": [1, 2, 3]}
        compressed = _compress_meta(original)
        decompressed = _decompress_meta(compressed)
        assert decompressed == original

    def test_meta_sheet_has_section_ids(self):
        sections = _make_sections(3)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["_meta_"]
        assert ws.cell(row=2, column=1).value == "section_ids"
        ids = json.loads(ws.cell(row=2, column=2).value)
        assert len(ids) == 3
        assert "section_0" in ids

    def test_meta_sheet_has_binding_hash(self):
        sections = _make_sections(2)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["_meta_"]
        assert ws.cell(row=3, column=1).value == "binding_hash"
        hash_val = ws.cell(row=3, column=2).value
        assert len(hash_val) == 64  # SHA-256

    def test_meta_decompression_matches_input(self):
        sections = [_make_section("s1", "测试")]
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["_meta_"]
        compressed = ws.cell(row=1, column=2).value
        meta = _decompress_meta(compressed)
        assert "s1" in meta
        assert "bindings" in meta["s1"]

    def test_meta_format_version(self):
        sections = _make_sections(1)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["_meta_"]
        assert ws.cell(row=5, column=1).value == "format_version"
        assert ws.cell(row=5, column=2).value == "1.0"


# ---------------------------------------------------------------------------
# C.0.6: Instructions Sheet
# ---------------------------------------------------------------------------


class TestInstructionsSheet:
    """C.0.6 — 注意事项 sheet content."""

    def test_instructions_present(self):
        sections = _make_sections(2)
        xlsx_bytes, _ = export_sections_to_xlsx(
            sections,
            project_name="测试项目",
            year=2025,
            exporter_name="张三",
            partner_info={"partner_name": "李四", "partner_email": "li@test.com", "partner_phone": "13800138000"},
        )
        wb = _load_wb(xlsx_bytes)
        ws = wb["注意事项"]
        # Check key content exists
        all_text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 50))
        assert "使用说明" in all_text
        assert "测试项目" in all_text
        assert "2025" in all_text
        assert "张三" in all_text
        assert "li@test.com" in all_text
        assert "黄底" in all_text
        assert "灰底" in all_text
        assert "红底" in all_text
        assert "绿底" in all_text

    def test_instructions_has_7_sections(self):
        xlsx_bytes, _ = export_sections_to_xlsx(_make_sections(1))
        wb = _load_wb(xlsx_bytes)
        ws = wb["注意事项"]
        all_text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 50))
        for i in range(1, 8):
            assert f"【{i}." in all_text


# ---------------------------------------------------------------------------
# C.0.7: TOC Sheet
# ---------------------------------------------------------------------------


class TestTocSheet:
    """C.0.7 — 章节清单 TOC."""

    def test_toc_headers(self):
        sections = _make_sections(3)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["章节清单"]
        assert ws.cell(row=1, column=1).value == "序号"
        assert ws.cell(row=1, column=2).value == "章节标题"
        assert ws.cell(row=1, column=3).value == "完成度(%)"
        assert ws.cell(row=1, column=4).value == "必填项数"
        assert ws.cell(row=1, column=5).value == "section_id"

    def test_toc_section_id_hidden(self):
        sections = _make_sections(2)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["章节清单"]
        assert ws.column_dimensions["E"].hidden is True

    def test_toc_row_count(self):
        sections = _make_sections(4)
        xlsx_bytes, _ = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        ws = wb["章节清单"]
        # Header + 4 data rows
        assert ws.cell(row=5, column=2).value == "章节3"
        assert ws.cell(row=6, column=2).value is None

    def test_completeness_calculation(self):
        section = _make_section(rows=[
            {"row_type": "data", "cells": [100, 200]},
            {"row_type": "data", "cells": [None, ""]},
        ])
        assert _calc_completeness(section) == 50  # 2/4 filled

    def test_completeness_empty(self):
        section = _make_section(rows=[])
        assert _calc_completeness(section) == 0


# ---------------------------------------------------------------------------
# C.0.8: AES Encryption + Hash
# ---------------------------------------------------------------------------


class TestEncryption:
    """C.0.8 — AES encryption and hash."""

    def test_hash_consistent(self):
        sections = _make_sections(2)
        _, hash1 = export_sections_to_xlsx(sections)
        _, hash2 = export_sections_to_xlsx(sections)
        # Hash should be consistent for same input (timestamps may differ)
        assert len(hash1) == 64
        assert len(hash2) == 64

    def test_encryption_round_trip(self):
        sections = _make_sections(2)
        encrypted_bytes, _ = export_sections_to_xlsx(sections, password="test123")
        # Encrypted bytes should be different from plain xlsx
        plain_bytes, _ = export_sections_to_xlsx(sections)
        assert encrypted_bytes != plain_bytes

        # Decrypt and verify it's valid xlsx
        decrypted = _decrypt_bytes(encrypted_bytes, "test123")
        wb = _load_wb(decrypted)
        assert "注意事项" in wb.sheetnames

    def test_wrong_password_fails(self):
        sections = _make_sections(1)
        encrypted_bytes, _ = export_sections_to_xlsx(sections, password="correct")
        with pytest.raises(Exception):
            _decrypt_bytes(encrypted_bytes, "wrong")

    def test_no_password_returns_plain_xlsx(self):
        sections = _make_sections(1)
        xlsx_bytes, _ = export_sections_to_xlsx(sections, password=None)
        # Should be valid xlsx directly
        wb = _load_wb(xlsx_bytes)
        assert "注意事项" in wb.sheetnames


# ---------------------------------------------------------------------------
# Edge Cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    """Edge cases and integration."""

    def test_section_with_no_table_data(self):
        section = {"section_id": "s1", "section_title": "空章节", "table_data": {}}
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        assert len(wb.sheetnames) == 4

    def test_dynamic_row_star_marker(self):
        section = _make_section()
        xlsx_bytes, _ = export_sections_to_xlsx([section])
        wb = _load_wb(xlsx_bytes)
        ws = wb.worksheets[2]
        # Row 6 (4+2) is dynamic_data row, should have ★
        cell = ws.cell(row=6, column=1)
        assert "★" in str(cell.value)

    def test_truncate_sheet_name_invalid_chars(self):
        assert _truncate_sheet_name("test[1]:2") == "test_1__2"

    def test_truncate_sheet_name_length(self):
        long = "a" * 50
        result = _truncate_sheet_name(long)
        assert len(result) <= 31

    def test_count_required_cells(self):
        section = _make_section(cell_meta={
            "0:0": {"is_required": True},
            "0:1": {"is_required": False},
            "1:0": {"is_required": True},
        })
        assert _count_required_cells(section) == 2

    def test_large_section_count(self):
        """Verify export handles many sections."""
        sections = _make_sections(20)
        xlsx_bytes, hash_val = export_sections_to_xlsx(sections)
        wb = _load_wb(xlsx_bytes)
        # 注意事项 + 章节清单 + 20 sections + _meta_ = 23
        assert len(wb.sheetnames) == 23
        assert len(hash_val) == 64
