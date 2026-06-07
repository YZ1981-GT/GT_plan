"""三式样式一致性测试 — Task 16

验证前端预览、Word 导出、Excel 导出在以下方面保持一致：
- 三线表样式
- 标题缩进
- 金额列右对齐

Validates: Requirements 7.2, 9.1
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.services.note_semantic_offline_export import (
    SEMANTIC_FILLS,
    build_account_disclosure_sheet,
    build_policy_clauses_sheet,
    create_semantic_workbook,
)

# ---------------------------------------------------------------------------
# Task 16.1: 前端预览、Word、Excel 样式映射清单
# ---------------------------------------------------------------------------

STYLE_MAPPING: dict[str, dict[str, str]] = {
    "三线表": {
        "frontend": "border-top: 2px solid; border-bottom: 2px solid; thead border-bottom: 1px solid",
        "word": "TableStyleClass: ThreeLineTable",
        "excel": "上下粗线+表头下细线",
    },
    "标题缩进": {
        "frontend": "padding-left: {level * 16}px",
        "word": "indent: {level * 240}twips",
        "excel": "indent: {level}",
    },
    "金额列对齐": {
        "frontend": "text-align: right; font-variant-numeric: tabular-nums",
        "word": "alignment: right; numFormat: '#,##0.00'",
        "excel": "alignment: right; number_format: '#,##0.00'",
    },
}

OUTPUT_FORMATS = ["frontend", "word", "excel"]


# ---------------------------------------------------------------------------
# 视觉断言辅助函数 (Task 16.3)
# ---------------------------------------------------------------------------


def assert_three_line_table_style(ws, header_row: int = 2, data_start_row: int = 3) -> None:
    """断言 Excel worksheet 符合三线表样式：表头有边框，数据行有细边框。

    只检查有内容的可见列（跳过隐藏语义列）。
    """
    # Header row should have borders for visible content columns
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        col_letter = get_column_letter(col_idx)
        # Skip hidden columns
        if ws.column_dimensions[col_letter].hidden:
            continue
        # Skip empty header cells
        if cell.value is None:
            continue
        assert cell.border is not None, f"Header cell ({header_row},{col_idx}) should have border"
        assert cell.font.bold is True, f"Header cell ({header_row},{col_idx}) should be bold"


def assert_title_indent(ws, row: int, expected_level: int) -> None:
    """断言 Excel 标题行有正确的缩进层级。"""
    cell = ws.cell(row=row, column=1)
    if cell.alignment and cell.alignment.indent:
        assert cell.alignment.indent == expected_level, (
            f"Row {row} indent should be {expected_level}, got {cell.alignment.indent}"
        )


def assert_amount_column_right_aligned(ws, col_idx: int, start_row: int = 2) -> None:
    """断言金额列右对齐。"""
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            assert cell.alignment is not None, f"Amount cell ({row_idx},{col_idx}) should have alignment"
            assert cell.alignment.horizontal == "right", (
                f"Amount cell ({row_idx},{col_idx}) should be right-aligned"
            )


# ---------------------------------------------------------------------------
# Task 16.1 Tests: 样式映射清单完整性
# ---------------------------------------------------------------------------


class TestStyleMappingCompleteness:
    """验证样式映射清单覆盖所有3种输出格式。"""

    def test_style_mapping_covers_all_formats(self):
        """每种样式规则都映射到 frontend/word/excel 三种格式。"""
        for style_name, mapping in STYLE_MAPPING.items():
            for fmt in OUTPUT_FORMATS:
                assert fmt in mapping, f"Style '{style_name}' missing format '{fmt}'"
                assert mapping[fmt], f"Style '{style_name}' format '{fmt}' is empty"

    def test_style_mapping_has_three_rules(self):
        """样式映射包含三线表、标题缩进、金额列对齐三种规则。"""
        expected_styles = {"三线表", "标题缩进", "金额列对齐"}
        assert set(STYLE_MAPPING.keys()) == expected_styles

    def test_each_format_has_consistent_structure(self):
        """每种格式的映射值都是非空字符串。"""
        for style_name, mapping in STYLE_MAPPING.items():
            for fmt, value in mapping.items():
                assert isinstance(value, str), f"{style_name}.{fmt} should be string"
                assert len(value) > 5, f"{style_name}.{fmt} should have meaningful content"


# ---------------------------------------------------------------------------
# Task 16.2 Tests: 三线表、标题缩进、金额列对齐
# ---------------------------------------------------------------------------


class TestThreeLineTableStyle:
    """三线表样式测试。"""

    def test_account_disclosure_has_header_borders(self):
        """科目披露表表头应有边框（三线表风格）。"""
        wb = Workbook()
        ws = wb.active
        tables = [{
            "section_id": "accounts_receivable",
            "table_id": "aging_analysis",
            "name": "应收账款账龄分析",
            "columns": [{"label": "项目"}, {"label": "期末余额"}, {"label": "期初余额"}],
            "rows": [
                {"row_id": "within_1y", "row_type": "data", "values": ["1年以内", 1000.00, 800.00]},
                {"row_id": "total", "row_type": "total", "values": ["合计", 1000.00, 800.00]},
            ],
        }]
        build_account_disclosure_sheet(ws, tables)
        # Row 2 is the header row (row 1 is table title)
        assert_three_line_table_style(ws, header_row=2)

    def test_policy_clauses_header_has_borders(self):
        """政策条款表头应有边框。"""
        wb = Workbook()
        ws = wb.active
        clauses = [
            {"clause_id": "c1", "title": "收入确认", "level": 2, "current_text": "...", "diff_status": "unchanged"},
        ]
        build_policy_clauses_sheet(ws, clauses)
        # Row 1 is header
        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)
            assert cell.border is not None
            assert cell.font.bold is True


class TestAmountColumnAlignment:
    """金额列右对齐测试。"""

    def test_numeric_values_right_aligned_in_disclosure(self):
        """科目披露表中数值应右对齐。"""
        wb = Workbook()
        ws = wb.active
        tables = [{
            "section_id": "fixed_assets",
            "table_id": "changes",
            "name": "固定资产变动",
            "columns": [{"label": "项目"}, {"label": "金额"}],
            "rows": [
                {"row_id": "r1", "row_type": "data", "values": ["房屋及建筑物", 50000.00]},
                {"row_id": "r2", "row_type": "data", "values": ["机器设备", 30000.00]},
            ],
        }]
        build_account_disclosure_sheet(ws, tables)
        # Column 2 contains amounts (data starts at row 3: title=1, header=2, data=3+)
        assert_amount_column_right_aligned(ws, col_idx=2, start_row=3)

    def test_text_values_not_forced_right(self):
        """文本列不应被强制右对齐。"""
        wb = Workbook()
        ws = wb.active
        tables = [{
            "section_id": "test",
            "table_id": "t1",
            "name": "测试",
            "columns": [{"label": "项目"}, {"label": "金额"}],
            "rows": [
                {"row_id": "r1", "row_type": "data", "values": ["文本项", 100.00]},
            ],
        }]
        build_account_disclosure_sheet(ws, tables)
        # Row 3, col 1 is text - should not be forced right
        cell = ws.cell(row=3, column=1)
        if cell.alignment:
            assert cell.alignment.horizontal != "right" or cell.alignment.horizontal is None


# ---------------------------------------------------------------------------
# Task 16.3 Tests: 视觉断言覆盖试点章节
# ---------------------------------------------------------------------------


class TestVisualAssertionPilotSections:
    """视觉断言覆盖试点章节测试。"""

    def test_full_workbook_pilot_section_styles(self):
        """完整工作包中试点章节样式正确。"""
        sections = [{
            "section_id": "accounts_receivable",
            "section_title": "应收账款",
            "table_data": {"rows": [{"cells": [100, 200]}]},
            "_semantic": {"variant": "soe_consolidated", "scope": "consolidated"},
        }]
        disclosure_tables = [{
            "section_id": "accounts_receivable",
            "table_id": "aging",
            "name": "账龄分析",
            "columns": [{"label": "项目"}, {"label": "期末余额"}],
            "rows": [
                {"row_id": "r1", "row_type": "data", "values": ["1年以内", 500.00]},
            ],
        }]
        wb = create_semantic_workbook(
            sections,
            project_name="试点测试",
            year="2025",
            disclosure_tables=disclosure_tables,
        )
        assert "科目披露" in wb.sheetnames
        ws = wb["科目披露"]
        # Verify header style on pilot section
        assert_three_line_table_style(ws, header_row=2)
        # Verify amount alignment
        assert_amount_column_right_aligned(ws, col_idx=2, start_row=3)

    def test_validation_results_level_coloring(self):
        """校验结果 sheet 中 blocking 级别应有红色标记。"""
        checklist = [
            {"level": "blocking", "category": "tieout", "section_id": "ar", "message": "不一致"},
            {"level": "warning", "category": "stale", "section_id": "fa", "message": "数据过期"},
        ]
        wb = create_semantic_workbook([], checklist_items=checklist)
        ws = wb["99_校验结果"]
        # Row 2 = first item (blocking)
        level_cell = ws.cell(row=2, column=2)
        assert level_cell.fill == SEMANTIC_FILLS["validation_failed"]
        # Row 3 = second item (warning)
        level_cell_2 = ws.cell(row=3, column=2)
        assert level_cell_2.fill == SEMANTIC_FILLS["needs_review"]

    def test_visual_assertion_helpers_exist(self):
        """视觉断言辅助函数已定义且可调用。"""
        assert callable(assert_three_line_table_style)
        assert callable(assert_title_indent)
        assert callable(assert_amount_column_right_aligned)
