"""Tests for 附注语义离线导出扩展 — Task 10.

Covers:
- 10.2: 00_填报说明 sheet 数据生成
- 10.3: 01_章节清单 sheet 数据生成
- 10.4: 六色颜色规范完整性
- 10.6: 导出 workbook 包含说明页和章节清单

Requirements: 7.1, 7.2
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.note_semantic_offline_export import (
    COLOR_SPEC,
    SEMANTIC_FILLS,
    build_semantic_instruction_sheet,
    build_semantic_section_list_sheet,
    create_semantic_workbook,
    generate_instruction_sheet_data,
    generate_section_list_sheet_data,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_sections(n: int = 3) -> list[dict]:
    """Create N mock sections with semantic info."""
    sections = []
    for i in range(n):
        sections.append({
            "section_id": f"section_{i}",
            "section_title": f"章节{i}",
            "table_data": {
                "headers": ["项目", "金额"],
                "rows": [
                    {"row_type": "data", "label": "行1", "cells": [100.0, 200.0]},
                    {"row_type": "data", "label": "行2", "cells": [None, 50.0]},
                ],
            },
            "_semantic": {
                "variant": "soe_consolidated" if i % 2 == 0 else "listed_standalone",
                "scope": "consolidated" if i % 2 == 0 else "standalone",
            },
        })
    return sections


# ---------------------------------------------------------------------------
# Task 10.4: 颜色规范完整性
# ---------------------------------------------------------------------------


class TestColorSpec:
    """测试六色规范定义完整性。"""

    def test_color_spec_has_six_entries(self):
        assert len(COLOR_SPEC) == 6

    def test_color_spec_keys(self):
        expected_keys = {
            "editable",
            "locked",
            "workpaper_source",
            "needs_review",
            "validation_failed",
            "prior_reference",
        }
        assert set(COLOR_SPEC.keys()) == expected_keys

    def test_color_spec_values_are_hex(self):
        for key, color in COLOR_SPEC.items():
            assert color.startswith("#"), f"{key} color should start with #"
            assert len(color) == 7, f"{key} color should be #RRGGBB format"

    def test_semantic_fills_match_color_spec(self):
        """SEMANTIC_FILLS 和 COLOR_SPEC 一一对应。"""
        assert set(SEMANTIC_FILLS.keys()) == set(COLOR_SPEC.keys())

    def test_semantic_fills_are_pattern_fill(self):
        from openpyxl.styles import PatternFill

        for key, fill in SEMANTIC_FILLS.items():
            assert isinstance(fill, PatternFill), f"{key} should be PatternFill"
            assert fill.fill_type == "solid"


# ---------------------------------------------------------------------------
# Task 10.2: 00_填报说明
# ---------------------------------------------------------------------------


class TestInstructionSheet:
    """测试填报说明 sheet 数据生成。"""

    def test_generate_instruction_data_returns_list(self):
        data = generate_instruction_sheet_data(
            project_name="测试项目",
            year="2025",
            exporter_name="张三",
            section_count=5,
        )
        assert isinstance(data, list)
        assert len(data) > 0

    def test_instruction_data_contains_project_info(self):
        data = generate_instruction_sheet_data(
            project_name="致同审计",
            year="2025",
            exporter_name="李四",
            section_count=3,
        )
        flat_text = "\n".join(row[0] for row in data)
        assert "致同审计" in flat_text
        assert "2025" in flat_text
        assert "李四" in flat_text
        assert "3" in flat_text

    def test_instruction_data_contains_color_descriptions(self):
        data = generate_instruction_sheet_data()
        flat_text = "\n".join(row[0] for row in data)
        assert "可填" in flat_text
        assert "锁定" in flat_text
        assert "来源底稿" in flat_text
        assert "需复核" in flat_text
        assert "校验失败" in flat_text
        assert "上年/模板参考" in flat_text

    def test_instruction_data_contains_structure_warnings(self):
        data = generate_instruction_sheet_data()
        flat_text = "\n".join(row[0] for row in data)
        assert "语义列" in flat_text
        assert "_meta" in flat_text

    def test_build_instruction_sheet_writes_cells(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        build_semantic_instruction_sheet(
            ws,
            project_name="项目A",
            year="2025",
            section_count=2,
        )
        # 第一行应有标题
        assert ws.cell(row=1, column=1).value is not None
        # 颜色样本列
        assert ws.cell(row=10, column=2).value is not None


# ---------------------------------------------------------------------------
# Task 10.3: 01_章节清单
# ---------------------------------------------------------------------------


class TestSectionListSheet:
    """测试章节清单 sheet 数据生成。"""

    def test_generate_section_list_has_header(self):
        sections = _make_sections(2)
        data = generate_section_list_sheet_data(sections)
        assert data[0] == ["序号", "章节标题", "section_id", "variant", "scope", "完成度(%)"]

    def test_generate_section_list_row_count(self):
        sections = _make_sections(3)
        data = generate_section_list_sheet_data(sections)
        # 1 header + 3 data rows
        assert len(data) == 4

    def test_section_list_contains_section_ids(self):
        sections = _make_sections(2)
        data = generate_section_list_sheet_data(sections)
        section_ids = [row[2] for row in data[1:]]
        assert "section_0" in section_ids
        assert "section_1" in section_ids

    def test_section_list_contains_variants(self):
        sections = _make_sections(2)
        data = generate_section_list_sheet_data(sections)
        variants = [row[3] for row in data[1:]]
        assert "soe_consolidated" in variants
        assert "listed_standalone" in variants

    def test_section_list_completeness_calculation(self):
        sections = [{
            "section_id": "test",
            "section_title": "测试",
            "table_data": {
                "rows": [
                    {"cells": [100, 200]},  # all filled
                    {"cells": [None, ""]},   # none filled
                ],
            },
            "_semantic": {},
        }]
        data = generate_section_list_sheet_data(sections)
        # 2 filled / 4 total = 50%
        assert data[1][5] == "50"

    def test_build_section_list_sheet_hides_columns(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        sections = _make_sections(2)
        build_semantic_section_list_sheet(ws, sections)
        # C, D, E columns should be hidden
        assert ws.column_dimensions["C"].hidden is True
        assert ws.column_dimensions["D"].hidden is True
        assert ws.column_dimensions["E"].hidden is True


# ---------------------------------------------------------------------------
# Task 10.6: 导出 workbook 包含说明页和章节清单
# ---------------------------------------------------------------------------


class TestCreateSemanticWorkbook:
    """测试导出 workbook 结构完整性。"""

    def test_workbook_contains_instruction_sheet(self):
        sections = _make_sections(2)
        wb = create_semantic_workbook(sections, project_name="项目X", year="2025")
        assert "00_填报说明" in wb.sheetnames

    def test_workbook_contains_section_list_sheet(self):
        sections = _make_sections(2)
        wb = create_semantic_workbook(sections, project_name="项目X", year="2025")
        assert "01_章节清单" in wb.sheetnames

    def test_workbook_sheet_order(self):
        sections = _make_sections(2)
        wb = create_semantic_workbook(sections, project_name="项目X", year="2025")
        assert wb.sheetnames[0] == "00_填报说明"
        assert wb.sheetnames[1] == "01_章节清单"

    def test_workbook_can_be_saved_and_loaded(self):
        """验证生成的 workbook 可以保存并重新加载。"""
        sections = _make_sections(3)
        wb = create_semantic_workbook(
            sections,
            project_name="致同测试",
            year="2025",
            exporter_name="王五",
        )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        loaded_wb = load_workbook(buf)
        assert "00_填报说明" in loaded_wb.sheetnames
        assert "01_章节清单" in loaded_wb.sheetnames

    def test_workbook_instruction_sheet_has_content(self):
        sections = _make_sections(1)
        wb = create_semantic_workbook(sections, project_name="项目Y")
        ws = wb["00_填报说明"]
        # First row should have content
        assert ws.cell(row=1, column=1).value is not None

    def test_workbook_section_list_has_data_rows(self):
        sections = _make_sections(3)
        wb = create_semantic_workbook(sections)
        ws = wb["01_章节清单"]
        # Header + 3 data rows
        assert ws.cell(row=1, column=1).value == "序号"
        assert ws.cell(row=2, column=1).value == "1"
        assert ws.cell(row=4, column=1).value == "3"

    def test_empty_sections_produces_valid_workbook(self):
        wb = create_semantic_workbook([], project_name="空项目")
        assert "00_填报说明" in wb.sheetnames
        assert "01_章节清单" in wb.sheetnames
