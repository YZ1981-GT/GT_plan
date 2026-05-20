"""F2 11 文件合并去重单测（spec workpaper-f-purchase-inventory F-F1）

Validates: Requirements F-F1.1, F-F1.2, F-F1.3, F-F1.4

F2 真实模板有 10 个 F2-* xlsx 文件（normal scenario 后 9 个），合并去重后预期：
- merged: 新增 sheet
- skipped_dup: 重复 sheet（底稿目录/GT_Custom/修订说明）
- skipped_historical: 历史遗留 sheet（含 G2-*-删除/移至/示例）

实测基线（Sprint 0 后）：normal scenario 合并 9 文件 → ≤ 67 sheet
"""
from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _filter_files_by_scenario,
    _merge_sheets_dedup,
    find_all_template_files,
)


@pytest.fixture
def f2_xlsx_files():
    """获取 F2 全部 xlsx 模板文件"""
    files = find_all_template_files("F2")
    xlsx = [f for f in files if f.suffix.lower() == ".xlsx"]
    if not xlsx:
        pytest.skip("F2 模板文件未找到")
    return xlsx


@pytest.fixture
def merge_target_in_tmp(tmp_path, f2_xlsx_files):
    """复制第一个 F2 文件到 tmp 作为合并目标"""
    target = tmp_path / "F2_merged.xlsx"
    shutil.copy2(f2_xlsx_files[0], target)
    return target, f2_xlsx_files[1:]


class TestF2MergeDedupNormalScenario:
    """F-F1: scenario=normal 时 F2 合并去重"""

    def test_normal_scenario_excludes_ipo_file(self, f2_xlsx_files):
        """F-F1: scenario=normal 应排除 F2-61~F2-72 IPO 文件"""
        filtered = _filter_files_by_scenario(f2_xlsx_files, "normal")
        # 至少排除 1 个 IPO 文件
        assert len(filtered) < len(f2_xlsx_files)
        for f in filtered:
            assert "IPO" not in f.name and "舞弊应对" not in f.name

    def test_ipo_scenario_loads_all_f2_files(self, f2_xlsx_files):
        """F-F1: scenario=ipo 应加载全部 F2 文件"""
        filtered = _filter_files_by_scenario(f2_xlsx_files, "ipo")
        assert len(filtered) == len(f2_xlsx_files)


class TestF2MergeDedupExecution:
    """F-F1: 执行合并去重并验证结果"""

    def test_merge_dedup_normal_scenario(self, tmp_path, f2_xlsx_files):
        """合并 normal scenario 下的 F2 文件，验证去重统计"""
        normal_files = _filter_files_by_scenario(f2_xlsx_files, "normal")
        if len(normal_files) < 2:
            pytest.skip("F2 文件不足以测试合并")

        target = tmp_path / "F2_merged.xlsx"
        shutil.copy2(normal_files[0], target)
        others = normal_files[1:]
        stats = _merge_sheets_dedup(target, others)

        # 应有 merged + skipped_dup + skipped_historical 统计
        assert "merged" in stats
        assert "skipped_dup" in stats
        assert "skipped_historical" in stats
        assert stats["merged"] >= 0
        # F-F1.4: 同名 sheet 应被去重
        assert stats["skipped_dup"] > 0, "应有重复 sheet 被去重（底稿目录/GT_Custom）"
        # F-F2: 历史遗留 sheet 应被过滤
        assert stats["skipped_historical"] > 0, (
            "应有历史遗留 sheet 被过滤（G2-*-删除/移至 + 示例）"
        )

    def test_merge_result_no_duplicate_sheets(self, tmp_path, f2_xlsx_files):
        """合并后 workbook 不应有重复 sheet 名"""
        normal_files = _filter_files_by_scenario(f2_xlsx_files, "normal")
        if len(normal_files) < 2:
            pytest.skip("F2 文件不足以测试合并")

        target = tmp_path / "F2_merged.xlsx"
        shutil.copy2(normal_files[0], target)
        _merge_sheets_dedup(target, normal_files[1:])

        wb = load_workbook(str(target), read_only=True)
        try:
            sheet_names = wb.sheetnames
            assert len(sheet_names) == len(set(sheet_names)), (
                f"合并后存在重复 sheet 名: {sheet_names}"
            )
        finally:
            wb.close()

    def test_merge_result_no_historical_sheets(self, tmp_path, f2_xlsx_files):
        """合并后 workbook 不应包含历史遗留 sheet"""
        normal_files = _filter_files_by_scenario(f2_xlsx_files, "normal")
        if len(normal_files) < 2:
            pytest.skip("F2 文件不足以测试合并")

        target = tmp_path / "F2_merged.xlsx"
        shutil.copy2(normal_files[0], target)
        _merge_sheets_dedup(target, normal_files[1:])

        wb = load_workbook(str(target), read_only=True)
        try:
            for s in wb.sheetnames:
                # 不应有 G2-*-删除/移至 / 示例 / 修订前
                assert not (
                    ("G" in s and ("删除" in s or "移至" in s))
                    or "（示例）" in s
                    or "(示例)" in s
                    or s.endswith("示例")
                    or "修订前" in s
                ), f"合并后仍含历史遗留 sheet: {s}"
        finally:
            wb.close()

    def test_merge_result_sheet_count_reasonable(self, tmp_path, f2_xlsx_files):
        """合并后 sheet 数应在合理范围（去重后 < 原始总数）"""
        normal_files = _filter_files_by_scenario(f2_xlsx_files, "normal")
        if len(normal_files) < 2:
            pytest.skip("F2 文件不足以测试合并")

        # 计算原始 sheet 总数
        raw_total = 0
        for f in normal_files:
            wb = load_workbook(str(f), read_only=True)
            raw_total += len(wb.sheetnames)
            wb.close()

        target = tmp_path / "F2_merged.xlsx"
        shutil.copy2(normal_files[0], target)
        _merge_sheets_dedup(target, normal_files[1:])

        wb = load_workbook(str(target), read_only=True)
        try:
            final = len(wb.sheetnames)
        finally:
            wb.close()

        # 合并后应少于原始总数（有去重发生）
        assert final < raw_total, (
            f"合并后 sheet 数 {final} 不少于原始 {raw_total}，去重未生效"
        )
        # F-F1.2: ≤ N_f2_dedup_sheets（实测预估 ~75）
        assert final <= 75, f"合并后 sheet 数 {final} 超过预估 75"
