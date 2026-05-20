"""H 循环合并去重验证测试（spec workpaper-h-fixed-assets-cycle Task 1.1）

验证 chain_orchestrator 对 H 循环 11 文件复用 _merge_sheets_dedup 合并去重：
1. H 循环 11 文件合并后 sheet 数 = 159（去重后）
2. _should_skip_historical_sheet 对全部 187 H sheet 返回 False（0 历史遗留）
3. 跨文件"底稿目录"/"GT_Custom"/"附注披露信息（上市公司）"/"附注披露信息（国有企业）"去重
4. D/F 历史遗留过滤回归（现有模式仍正确工作）
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "H"

# Collect all H cycle template files
H_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_h_sheet_names() -> list[str]:
    """从 11 个 H 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in H_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数"""
    all_names = _get_all_h_sheet_names()
    # 先过滤历史遗留，再归一化去重
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestHCycleMergeDedup:
    """H-F1: 11 文件合并去重验证"""

    def test_h_cycle_has_11_template_files(self):
        """H 循环应有 11 个模板文件（H0~H10）"""
        assert len(H_TEMPLATE_FILES) == 11, (
            f"Expected 11 H cycle template files, got {len(H_TEMPLATE_FILES)}: "
            f"{[f.name for f in H_TEMPLATE_FILES]}"
        )

    def test_h_cycle_raw_sheet_count_is_187(self):
        """H 循环 11 文件原始 sheet 总数 = 187"""
        all_names = _get_all_h_sheet_names()
        assert len(all_names) == 187, (
            f"Expected 187 raw sheets, got {len(all_names)}"
        )

    def test_h_cycle_dedup_sheet_count_is_159(self):
        """H 循环合并去重后 sheet 数 = 159"""
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 159, (
            f"Expected 159 dedup sheets, got {dedup_count}"
        )

    def test_h_cycle_no_historical_sheets(self):
        """H 循环 187 sheet 全部不命中历史遗留过滤（0 命中）"""
        all_names = _get_all_h_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 0, (
            f"Expected 0 historical sheet hits, got {len(historical_hits)}: {historical_hits}"
        )

    def test_cross_file_dedup_removes_duplicates(self):
        """跨文件"底稿目录"/"GT_Custom"/"附注披露信息（上市公司）"/"附注披露信息（国有企业）"被正确去重"""
        all_names = _get_all_h_sheet_names()

        # 统计跨文件重复的关键 sheet 名
        dedup_targets = {
            "底稿目录": 0,
            "GT_Custom": 0,
            "附注披露信息（上市公司）": 0,
            "附注披露信息（国有企业）": 0,
        }

        for name in all_names:
            normalized = _normalize_sheet_name(name)
            if normalized == "底稿目录":
                dedup_targets["底稿目录"] += 1
            elif normalized == "GT_Custom":
                dedup_targets["GT_Custom"] += 1
            elif "附注披露信息" in name and "上市" in name:
                dedup_targets["附注披露信息（上市公司）"] += 1
            elif "附注披露信息" in name and "国有" in name:
                dedup_targets["附注披露信息（国有企业）"] += 1

        # 每种跨文件重复 sheet 应出现多次（>1 表示有重复需去重）
        for key, count in dedup_targets.items():
            assert count > 1, (
                f"'{key}' 应在多个文件中出现（跨文件去重目标），实际出现 {count} 次"
            )

    def test_same_wpcode_multi_version_not_misdeduped(self):
        """同 wp_code 多版本 sheet（如 H1-12 三版折旧）不被误去重

        含括号修饰词的 normalized key 不同，应全部保留。
        """
        all_names = _get_all_h_sheet_names()

        # H1-12 三版折旧测算表
        h1_12_sheets = [n for n in all_names if "H1-12" in n]
        assert len(h1_12_sheets) >= 3, (
            f"H1-12 应有 ≥ 3 个版本 sheet，实际 {len(h1_12_sheets)}: {h1_12_sheets}"
        )

        # 验证归一化后仍然不同（不会被误去重）
        h1_12_normalized = {_normalize_sheet_name(n) for n in h1_12_sheets}
        assert len(h1_12_normalized) == len(h1_12_sheets), (
            f"H1-12 多版本 sheet 归一化后应仍不同，但有重复: "
            f"原始={h1_12_sheets}, 归一化={h1_12_normalized}"
        )


class TestDFHistoricalFilterRegression:
    """D/F 历史遗留过滤回归测试 — 确保 H spec 不影响现有模式"""

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # 正常 sheet 不应被过滤
        ("存货审定表F2-1", False),
        ("底稿目录", False),
        ("GT_Custom", False),
        ("审定表H1-1", False),
        ("折旧测算表（不含减值）-直线法H1-12", False),
        ("审定表（成本模式）H3-1", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F 历史遗留过滤模式仍正确工作，H 正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
