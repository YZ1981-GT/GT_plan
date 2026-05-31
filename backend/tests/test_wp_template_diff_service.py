"""模板版本 diff 引擎单元测试

构造两版本模板 xlsx → 断言 diff 正确。

Spec: wp-template-migration
Requirements: 1.1
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from app.services.wp_template_diff_service import (
    ColumnDiff,
    TemplateDiff,
    _compute_column_diff,
    _detect_renamed_columns,
    _detect_renamed_sheets,
    _read_template_structure,
    generate_template_diff,
)


def _create_xlsx(sheets: dict[str, list[str]], path: Path) -> Path:
    """辅助：创建测试用 xlsx 文件

    Args:
        sheets: {sheet_name: [column_headers]}
        path: 输出路径
    """
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    default_ws = wb.active
    if default_ws:
        wb.remove(default_ws)

    for sheet_name, headers in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

    wb.save(str(path))
    return path


class TestReadTemplateStructure:
    """测试 _read_template_structure"""

    def test_reads_sheet_headers(self, tmp_path: Path):
        xlsx_path = tmp_path / "test.xlsx"
        _create_xlsx(
            {"Sheet1": ["科目", "期末余额", "期初余额"], "Sheet2": ["项目", "金额"]},
            xlsx_path,
        )
        result = _read_template_structure(xlsx_path)
        assert result == {
            "Sheet1": ["科目", "期末余额", "期初余额"],
            "Sheet2": ["项目", "金额"],
        }

    def test_empty_file_returns_empty(self, tmp_path: Path):
        xlsx_path = tmp_path / "empty.xlsx"
        xlsx_path.write_bytes(b"")
        result = _read_template_structure(xlsx_path)
        assert result == {}

    def test_nonexistent_file_returns_empty(self, tmp_path: Path):
        result = _read_template_structure(tmp_path / "nonexistent.xlsx")
        assert result == {}


class TestDetectRenamedSheets:
    """测试 sheet 改名检测"""

    def test_detects_rename_by_column_similarity(self):
        old_structure = {"旧表": ["科目", "期末", "期初", "差异"]}
        new_structure = {"新表": ["科目", "期末", "期初", "差异"]}
        result = _detect_renamed_sheets(
            {"旧表"}, {"新表"}, old_structure, new_structure
        )
        assert result == [("旧表", "新表")]

    def test_no_rename_when_columns_differ(self):
        old_structure = {"旧表": ["A", "B", "C"]}
        new_structure = {"新表": ["X", "Y", "Z"]}
        result = _detect_renamed_sheets(
            {"旧表"}, {"新表"}, old_structure, new_structure
        )
        assert result == []

    def test_no_rename_when_no_removed_or_added(self):
        result = _detect_renamed_sheets(
            {"Sheet1"}, {"Sheet1"}, {"Sheet1": ["A"]}, {"Sheet1": ["A"]}
        )
        assert result == []


class TestDetectRenamedColumns:
    """测试列改名检测"""

    def test_detects_positional_rename(self):
        old_cols = ["科目", "期末余额", "期初余额"]
        new_cols = ["科目", "本期余额", "期初余额"]
        result = _detect_renamed_columns(old_cols, new_cols)
        assert result == [("期末余额", "本期余额")]

    def test_no_rename_when_same(self):
        cols = ["A", "B", "C"]
        result = _detect_renamed_columns(cols, cols)
        assert result == []


class TestComputeColumnDiff:
    """测试列级 diff 计算"""

    def test_added_columns(self):
        result = _compute_column_diff("Sheet1", ["A", "B"], ["A", "B", "C"])
        assert result is not None
        assert result.added == ["C"]
        assert result.removed == []

    def test_removed_columns(self):
        result = _compute_column_diff("Sheet1", ["A", "B", "C"], ["A", "B"])
        assert result is not None
        assert result.removed == ["C"]
        assert result.added == []

    def test_no_change_returns_none(self):
        result = _compute_column_diff("Sheet1", ["A", "B"], ["A", "B"])
        assert result is None


class TestGenerateTemplateDiff:
    """测试完整 diff 生成"""

    def test_added_sheet(self, tmp_path: Path):
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"Sheet1": ["A", "B"]}, old_path)
        _create_xlsx({"Sheet1": ["A", "B"], "Sheet2": ["X", "Y"]}, new_path)

        diff = generate_template_diff(old_path, new_path)
        assert "Sheet2" in diff.added_sheets
        assert diff.removed_sheets == []

    def test_removed_sheet(self, tmp_path: Path):
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"Sheet1": ["A"], "Sheet2": ["B"]}, old_path)
        _create_xlsx({"Sheet1": ["A"]}, new_path)

        diff = generate_template_diff(old_path, new_path)
        assert "Sheet2" in diff.removed_sheets
        assert diff.added_sheets == []

    def test_renamed_sheet(self, tmp_path: Path):
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"旧名称": ["科目", "期末", "期初", "差异"]}, old_path)
        _create_xlsx({"新名称": ["科目", "期末", "期初", "差异"]}, new_path)

        diff = generate_template_diff(old_path, new_path)
        assert ("旧名称", "新名称") in diff.renamed_sheets
        assert diff.added_sheets == []
        assert diff.removed_sheets == []

    def test_column_added_in_existing_sheet(self, tmp_path: Path):
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"Sheet1": ["科目", "期末"]}, old_path)
        _create_xlsx({"Sheet1": ["科目", "期末", "备注"]}, new_path)

        diff = generate_template_diff(old_path, new_path)
        assert len(diff.column_diffs) == 1
        assert diff.column_diffs[0].sheet_name == "Sheet1"
        assert "备注" in diff.column_diffs[0].added

    def test_no_changes(self, tmp_path: Path):
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"Sheet1": ["A", "B"]}, old_path)
        _create_xlsx({"Sheet1": ["A", "B"]}, new_path)

        diff = generate_template_diff(old_path, new_path)
        assert not diff.has_changes

    def test_summary(self, tmp_path: Path):
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"Sheet1": ["A"], "Sheet2": ["B"]}, old_path)
        _create_xlsx({"Sheet1": ["A", "C"], "Sheet3": ["D"]}, new_path)

        diff = generate_template_diff(old_path, new_path)
        s = diff.summary()
        assert s["has_changes"] is True
        assert s["added_sheets"] >= 1
