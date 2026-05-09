"""生成最小合成账套样本（balance + ledger 2 sheets）供 CI smoke 使用。

产出：minimal_balance_ledger.xlsx，含 10 行余额 + 10 行序时账，
结构完全对齐真实导出格式（标题行 + 合并表头 + 数据）。

用法：
    from backend.tests.fixtures.ledger_samples.minimal_balance_ledger import (
        create_minimal_sample,
    )
    path = create_minimal_sample()  # 返回 Path
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl


def create_minimal_sample(output_path: str | Path | None = None) -> Path:
    """创建最小合成样本 xlsx。

    Args:
        output_path: 输出路径，None 则写到本 fixture 目录下 minimal_balance_ledger.xlsx

    Returns:
        生成文件的 Path
    """
    if output_path is None:
        output_path = Path(__file__).parent / "minimal_balance_ledger.xlsx"
    output_path = Path(output_path)

    wb = openpyxl.Workbook()

    # ── Sheet 1: 科目余额表（合并表头 + 核算维度） ──
    ws1 = wb.active
    ws1.title = "科目余额表"
    ws1["A1"] = "科目余额表"
    ws1.merge_cells("A1:H1")
    ws1["A2"] = "核算组织：测试公司; 开始期间：2025年1期; 结束期间：2025年12期"
    ws1.merge_cells("A2:H2")
    # 合并表头：第 3 行主标题，第 4 行子列
    headers_main = ["科目编码", "科目名称", "核算维度", "组织编码",
                    "期初余额", "期初余额", "期末余额", "期末余额"]
    headers_sub = ["", "", "", "", "借方金额", "贷方金额", "借方金额", "贷方金额"]
    for i, h in enumerate(headers_main, 1):
        ws1.cell(row=3, column=i, value=h)
    for i, h in enumerate(headers_sub, 1):
        ws1.cell(row=4, column=i, value=h)

    # 数据行
    balance_data = [
        ("1001", "库存现金", "", "COMP01", 100.00, None, 120.00, None),
        ("1002", "银行存款", "", "COMP01", 50000.00, None, 48000.00, None),
        ("1002", "银行存款", "金融机构:A001,工商银行", "COMP01", 30000.00, None, 28000.00, None),
        ("1002", "银行存款", "金融机构:A002,建设银行", "COMP01", 20000.00, None, 20000.00, None),
        ("1122", "应收账款", "客户:C001 甲公司", "COMP01", 5000.00, None, 4500.00, None),
        ("1122", "应收账款", "客户:C002 乙公司", "COMP01", 3000.00, None, 3500.00, None),
        ("2202", "应付账款", "供应商:V001 某供应商", "COMP01", None, 2000.00, None, 2500.00),
        ("4103", "本年利润", "", "COMP01", 0.00, None, 10000.00, None),
    ]
    for i, row in enumerate(balance_data, start=5):
        for j, val in enumerate(row, start=1):
            ws1.cell(row=i, column=j, value=val)

    # ── Sheet 2: 序时账（简单表头 + 核算维度列） ──
    ws2 = wb.create_sheet("序时账")
    headers = [
        "记账日期", "凭证号", "凭证类型", "摘要", "科目编码", "科目名称",
        "借方", "贷方", "核算维度", "制单人",
    ]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=1, column=i, value=h)

    ledger_data = [
        (date(2025, 1, 15), "记-001", "记", "收到货款", "1002", "银行存款",
         5000.00, None, "金融机构:A001,工商银行", "张三"),
        (date(2025, 1, 15), "记-001", "记", "收到货款", "1122", "应收账款",
         None, 5000.00, "客户:C001 甲公司", "张三"),
        (date(2025, 2, 10), "记-002", "记", "支付采购款", "2202", "应付账款",
         3000.00, None, "供应商:V001 某供应商", "李四"),
        (date(2025, 2, 10), "记-002", "记", "支付采购款", "1002", "银行存款",
         None, 3000.00, "金融机构:A002,建设银行", "李四"),
        (date(2025, 3, 5), "记-003", "记", "结转成本", "5401", "主营业务成本",
         2000.00, None, "成本中心:CC01 生产部", "王五"),
        (date(2025, 12, 31), "记-100", "ZZ", "结转本年利润", "4103", "本年利润",
         10000.00, None, "", "张三"),
    ]
    for i, row in enumerate(ledger_data, start=2):
        for j, val in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=val)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # CLI: python minimal_balance_ledger.py [output_path]
    import sys
    path = create_minimal_sample(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"Created: {path}")
