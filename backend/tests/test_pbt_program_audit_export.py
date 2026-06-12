"""Property 23 & 25: Program Sheet Export Columns + Audit Sheet Export Completeness (PBT)

Property 23: 程序表导出包含 procedure_code/description/status/conclusion/executor 五列
Property 25: 审定表导出含5列+末尾汇总行

**Validates: Requirements 8.1, 9.1, 9.4**

Testing framework: hypothesis
"""

from __future__ import annotations

from io import BytesIO

from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook

from app.services.wp_export.export_engine import WpExportEngine

# ─── Hypothesis Strategies ────────────────────────────────────────────────────


# Alphabet safe for openpyxl: printable characters only (no control chars)
_SAFE_ALPHABET = st.characters(
    whitelist_categories=("L", "N", "P", "S", "Z"),
    blacklist_characters="\x00\x01\x02\x03\x04\x05\x06\x07\x08\x09\x0a\x0b\x0c\x0d"
    "\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1a\x1b\x1c\x1d\x1e\x1f",
)


@st.composite
def st_program_procedures(draw: st.DrawFn) -> list[dict]:
    """Generate a list of program procedure dicts for export testing."""
    num_procedures = draw(st.integers(min_value=1, max_value=5))
    procedures = []
    for i in range(num_procedures):
        proc = {
            "procedure_code": draw(
                st.text(alphabet=_SAFE_ALPHABET, min_size=1, max_size=10)
            ),
            "description": draw(
                st.text(alphabet=_SAFE_ALPHABET, min_size=1, max_size=30)
            ),
            "execution_status": draw(
                st.sampled_from(["未执行", "已执行", "不适用", "进行中", ""])
            ),
            "execution_conclusion": draw(
                st.text(alphabet=_SAFE_ALPHABET, min_size=0, max_size=30)
            ),
            "executor": draw(
                st.text(alphabet=_SAFE_ALPHABET, min_size=0, max_size=10)
            ),
        }
        procedures.append(proc)
    return procedures


@st.composite
def st_audit_accounts(draw: st.DrawFn) -> list[dict]:
    """Generate a list of audit account dicts for export testing."""
    num_accounts = draw(st.integers(min_value=1, max_value=5))
    accounts = []
    for i in range(num_accounts):
        unadj = draw(
            st.floats(min_value=-1e8, max_value=1e8, allow_nan=False, allow_infinity=False)
        )
        adj = draw(
            st.floats(min_value=-1e6, max_value=1e6, allow_nan=False, allow_infinity=False)
        )
        audited = unadj + adj

        account = {
            "account_code": draw(
                st.text(
                    alphabet=st.characters(whitelist_categories=("N",)),
                    min_size=4,
                    max_size=6,
                )
            ),
            "account_name": draw(
                st.text(alphabet=_SAFE_ALPHABET, min_size=1, max_size=15)
            ),
            "unadjusted_amount": unadj,
            "adjustment_amount": adj,
            "audited_amount": audited,
            "adjustment_source": draw(
                st.one_of(
                    st.just(""),
                    st.text(alphabet=_SAFE_ALPHABET, min_size=1, max_size=10),
                )
            ),
        }
        accounts.append(account)
    return accounts


# ─── Property 23: Program Sheet Export Contains Required Columns ───────────────


class TestProgramSheetExportColumns:
    """Property 23: 程序表导出包含 procedure_code/description/status/conclusion/executor 五列

    **Validates: Requirements 8.1**
    """

    # Expected column header keywords (mapped from field → header text)
    REQUIRED_FIELDS = [
        "程序编号",
        "程序描述",
        "执行状态",
        "执行结论",
        "执行人",
    ]

    @given(procedures=st_program_procedures())
    @settings(max_examples=5, deadline=None)
    def test_program_sheet_contains_five_required_columns(
        self, procedures: list[dict]
    ) -> None:
        """Program sheet export contains all 5 required columns in headers.

        **Validates: Requirements 8.1**
        """
        buf: BytesIO = WpExportEngine.build_program_sheet_bytes(procedures)

        wb = load_workbook(buf)
        ws = wb.active

        assert ws is not None, "Workbook has no active sheet"
        assert ws.title == "程序表", f"Sheet title should be '程序表', got '{ws.title}'"

        # Read header row (row 1)
        headers = [ws.cell(row=1, column=col).value for col in range(1, 6)]
        headers_text = " ".join(str(h) for h in headers if h)

        for required_field in self.REQUIRED_FIELDS:
            assert required_field in headers_text, (
                f"Required column '{required_field}' not found in headers: {headers}"
            )

        # Verify exactly 5 columns have non-empty headers
        non_empty_headers = [h for h in headers if h]
        assert len(non_empty_headers) == 5, (
            f"Expected 5 column headers, got {len(non_empty_headers)}: {non_empty_headers}"
        )

        wb.close()

    @given(procedures=st_program_procedures())
    @settings(max_examples=5, deadline=None)
    def test_program_sheet_readonly_columns_marked(
        self, procedures: list[dict]
    ) -> None:
        """Program sheet marks readonly columns (code, description) with [只读] prefix.

        **Validates: Requirements 8.2**
        """
        buf: BytesIO = WpExportEngine.build_program_sheet_bytes(procedures)

        wb = load_workbook(buf)
        ws = wb.active

        # Columns 1, 2 should have [只读] prefix
        header1 = ws.cell(row=1, column=1).value
        header2 = ws.cell(row=1, column=2).value
        assert header1 and "[只读]" in header1, (
            f"Column 1 header should contain '[只读]', got: {header1}"
        )
        assert header2 and "[只读]" in header2, (
            f"Column 2 header should contain '[只读]', got: {header2}"
        )

        # Columns 3, 4, 5 should have [可编辑] prefix
        header3 = ws.cell(row=1, column=3).value
        header4 = ws.cell(row=1, column=4).value
        header5 = ws.cell(row=1, column=5).value
        assert header3 and "[可编辑]" in header3, (
            f"Column 3 header should contain '[可编辑]', got: {header3}"
        )
        assert header4 and "[可编辑]" in header4, (
            f"Column 4 header should contain '[可编辑]', got: {header4}"
        )
        assert header5 and "[可编辑]" in header5, (
            f"Column 5 header should contain '[可编辑]', got: {header5}"
        )

        wb.close()

    @given(procedures=st_program_procedures())
    @settings(max_examples=5, deadline=None)
    def test_program_sheet_data_row_count(
        self, procedures: list[dict]
    ) -> None:
        """Program sheet contains exactly len(procedures) data rows.

        **Validates: Requirements 8.1**
        """
        buf: BytesIO = WpExportEngine.build_program_sheet_bytes(procedures)

        wb = load_workbook(buf)
        ws = wb.active

        # Data starts at row 2, max_row should be len(procedures) + 1 (header row)
        expected_max_row = len(procedures) + 1
        assert ws.max_row == expected_max_row, (
            f"Expected {expected_max_row} rows (1 header + {len(procedures)} data), "
            f"got {ws.max_row}"
        )

        wb.close()


# ─── Property 25: Audit Sheet Export Completeness ─────────────────────────────


class TestAuditSheetExportCompleteness:
    """Property 25: 审定表导出含5列+末尾汇总行

    **Validates: Requirements 9.1, 9.4**
    """

    REQUIRED_COLUMNS = [
        "科目编码",
        "科目名称",
        "未审数",
        "调整数",
        "审定数",
    ]

    @given(accounts=st_audit_accounts())
    @settings(max_examples=5, deadline=None)
    def test_audit_sheet_contains_five_columns(
        self, accounts: list[dict]
    ) -> None:
        """Audit sheet export contains all 5 required columns.

        **Validates: Requirements 9.1**
        """
        buf: BytesIO = WpExportEngine.build_audit_sheet_bytes(accounts)

        wb = load_workbook(buf)
        ws = wb.active

        assert ws is not None, "Workbook has no active sheet"
        assert ws.title == "审定表", f"Sheet title should be '审定表', got '{ws.title}'"

        # Read header row (row 1)
        headers = [ws.cell(row=1, column=col).value for col in range(1, 6)]

        for required_col in self.REQUIRED_COLUMNS:
            assert required_col in headers, (
                f"Required column '{required_col}' not found in headers: {headers}"
            )

        # Verify exactly 5 columns
        non_empty_headers = [h for h in headers if h]
        assert len(non_empty_headers) == 5, (
            f"Expected 5 column headers, got {len(non_empty_headers)}: {non_empty_headers}"
        )

        wb.close()

    @given(accounts=st_audit_accounts())
    @settings(max_examples=5, deadline=None)
    def test_audit_sheet_has_summary_row(
        self, accounts: list[dict]
    ) -> None:
        """Audit sheet has a summary row at the end with totals.

        **Validates: Requirements 9.4**
        """
        buf: BytesIO = WpExportEngine.build_audit_sheet_bytes(accounts)

        wb = load_workbook(buf)
        ws = wb.active

        # Summary row should be at row = len(accounts) + 2 (header + data + summary)
        expected_summary_row = len(accounts) + 2
        summary_label = ws.cell(row=expected_summary_row, column=1).value

        assert summary_label == "合计", (
            f"Expected summary row label '合计' at row {expected_summary_row}, "
            f"got: {summary_label}"
        )

        # Verify summary values are numeric (totals)
        total_unadj = ws.cell(row=expected_summary_row, column=3).value
        total_adj = ws.cell(row=expected_summary_row, column=4).value
        total_aud = ws.cell(row=expected_summary_row, column=5).value

        assert total_unadj is not None, "Summary unadjusted_amount should not be None"
        assert total_adj is not None, "Summary adjustment_amount should not be None"
        assert total_aud is not None, "Summary audited_amount should not be None"

        # Verify totals are numeric
        assert isinstance(total_unadj, (int, float)), (
            f"Summary unadjusted should be numeric, got {type(total_unadj)}"
        )
        assert isinstance(total_adj, (int, float)), (
            f"Summary adjustment should be numeric, got {type(total_adj)}"
        )
        assert isinstance(total_aud, (int, float)), (
            f"Summary audited should be numeric, got {type(total_aud)}"
        )

        wb.close()

    @given(accounts=st_audit_accounts())
    @settings(max_examples=5, deadline=None)
    def test_audit_sheet_summary_row_is_last(
        self, accounts: list[dict]
    ) -> None:
        """Summary row is the last row in the sheet (after all data rows).

        **Validates: Requirements 9.1, 9.4**
        """
        buf: BytesIO = WpExportEngine.build_audit_sheet_bytes(accounts)

        wb = load_workbook(buf)
        ws = wb.active

        # Total rows = 1 (header) + len(accounts) (data) + 1 (summary)
        expected_total_rows = len(accounts) + 2
        assert ws.max_row == expected_total_rows, (
            f"Expected {expected_total_rows} total rows, got {ws.max_row}"
        )

        # Last row should be the summary
        last_row_label = ws.cell(row=ws.max_row, column=1).value
        assert last_row_label == "合计", (
            f"Last row should be summary '合计', got: {last_row_label}"
        )

        wb.close()
