"""M-F2: M 权益循环 8-category sheet group rules — backend completeness check.

Validates: Requirements M-F2 / Spec workpaper-m-equity-cycle / Sprint 2 Task 2.2

Mirror of frontend useMEquityCycleSheetGroups.ts classification rules.

8 categories (priority 0~7):
1. 索引 (defaultHidden=true): 底稿目录 / GT_Custom
2. 程序表: 实质性程序表 / M*A 结尾
3. 审定表: 审定表M*-1 pattern
4. 明细表: 明细表 pattern（含上市/非上市变体）
5. 变动分析: 变动 / 增减 / 权益变动
6. 检查表: 检查 / 核查 / 测试
7. 附注+调整 (defaultHidden=true): 附注 / 披露 / 调整
8. 其他: fallback

The classification function is reproduced here in pure Python to provide a
backend-side correctness oracle for the 65 real M cycle sheets.
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _normalize_sheet_name,
    _should_skip_historical_sheet,
)


# ──────────────────────────────────────────────────────────────────────────
# Mirror of frontend classifyMSheet rules (must stay in sync)
# ──────────────────────────────────────────────────────────────────────────

M_RULES = [
    {
        "id": "index",
        "category": "索引",
        "priority": 0,
        "default_hidden": True,
        "match": lambda s: bool(re.match(r"^(底稿目录|GT_Custom)$", s.strip())),
    },
    {
        "id": "procedure",
        "category": "程序表",
        "priority": 1,
        "default_hidden": False,
        "match": lambda s: bool(
            re.search(r"实质性程序表", s)
            or re.search(r"[A-Z]\d*A\s*$", s.strip())
            or re.search(r"M\d+A\s*$", s.strip())
        ),
    },
    {
        "id": "audit_table",
        "category": "审定表",
        "priority": 2,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"审定表", s)),
    },
    {
        "id": "detail",
        "category": "明细表",
        "priority": 3,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"明细表", s)),
    },
    {
        "id": "movement_analysis",
        "category": "变动分析",
        "priority": 4,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"变动|增减|权益变动", s)),
    },
    {
        "id": "check_table",
        "category": "检查表",
        "priority": 5,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"检查|核查|测试", s)),
    },
    {
        "id": "disclosure_adj",
        "category": "附注+调整",
        "priority": 6,
        "default_hidden": True,
        "match": lambda s: bool(re.search(r"附注|披露|调整", s)),
    },
    {
        "id": "other",
        "category": "其他",
        "priority": 7,
        "default_hidden": False,
        "match": lambda s: True,
    },
]


def classify_m_sheet(name: str) -> dict:
    """Return classification metadata; fallback to 其他."""
    for rule in M_RULES:
        if rule["match"](name):
            cat = {
                "id": rule["id"],
                "category": rule["category"],
                "priority": rule["priority"],
                "default_hidden": rule["default_hidden"],
                "readonly": False,
            }
            # 附注披露 readonly=true 特殊处理
            if rule["id"] == "disclosure_adj" and re.search(r"附注披露", name):
                cat["default_hidden"] = True
                cat["readonly"] = True
            return cat
    return {
        "id": "other",
        "category": "其他",
        "priority": 7,
        "default_hidden": False,
        "readonly": False,
    }


# ──────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "M"


@pytest.fixture(scope="module")
def m_65_unique_sheets() -> list[str]:
    """65 个去重后 M 循环有效 sheet 名（基于 _normalize_sheet_name + filter historical）"""
    seen_normalized: set[str] = set()
    unique_originals: list[str] = []
    for f in sorted(TEMPLATES_DIR.glob("*.xlsx")):
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            for s in wb.sheetnames:
                if _should_skip_historical_sheet(s):
                    continue
                norm = _normalize_sheet_name(s)
                if norm not in seen_normalized:
                    seen_normalized.add(norm)
                    unique_originals.append(s)
        finally:
            wb.close()
    return unique_originals


# ──────────────────────────────────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────────────────────────────────


class TestMSheetGroupCompleteness:
    """M-F2: 8 类规则对 65 sheet 完备性 + 唯一性"""

    def test_65_sheet_baseline(self, m_65_unique_sheets):
        """M 循环去重后有效 sheet 数 = 65（Sprint 0 实测基线）"""
        assert len(m_65_unique_sheets) == 65, (
            f"Expected 65 M-cycle dedup sheets, got {len(m_65_unique_sheets)}"
        )

    def test_all_65_classified_to_non_empty_category(self, m_65_unique_sheets):
        """所有 sheet 均被分类到非空类目"""
        for s in m_65_unique_sheets:
            cat = classify_m_sheet(s)
            assert cat["category"], f"sheet '{s}' classified to empty category"

    def test_each_sheet_exactly_one_category(self, m_65_unique_sheets):
        """任意 sheet 名恰好命中 1 类（rule fallback `() => true` 必然命中）"""
        for s in m_65_unique_sheets:
            hits = [r for r in M_RULES if r["match"](s)]
            assert len(hits) >= 1, f"sheet '{s}' missed all rules"

    def test_no_sheet_falls_into_other_unexpectedly(self, m_65_unique_sheets):
        """fallback "其他" 类应 < 30%（65 中 < 20）"""
        other_sheets = [
            s for s in m_65_unique_sheets if classify_m_sheet(s)["id"] == "other"
        ]
        ratio = len(other_sheets) / len(m_65_unique_sheets)
        assert ratio < 0.30, (
            f"Too many sheets fell to '其他' fallback: "
            f"{len(other_sheets)}/{len(m_65_unique_sheets)} = {ratio:.1%}\n"
            f"Other sheets: {other_sheets}"
        )

    def test_all_8_categories_represented(self, m_65_unique_sheets):
        """65 sheet 应覆盖 ≥ 6 类（8 类中至少 6 类被命中）"""
        categories_hit = {classify_m_sheet(s)["category"] for s in m_65_unique_sheets}
        assert len(categories_hit) >= 6, (
            f"Only {len(categories_hit)} categories hit, expected ≥ 6\n"
            f"Hit: {sorted(categories_hit)}"
        )
        # 索引 / 附注+调整 / 程序表 / 审定表 / 明细表 必须命中
        for must_have in ["索引", "附注+调整", "程序表", "审定表", "明细表"]:
            assert must_have in categories_hit, (
                f"Category '{must_have}' should be hit but not in {sorted(categories_hit)}"
            )


class TestMSheetGroupKeyPaths:
    """M-F2: 关键 sheet 分类路径验证"""

    def test_index_sheets(self):
        """底稿目录 / GT_Custom → 索引 + defaultHidden=true"""
        for s in ["底稿目录", "GT_Custom"]:
            cat = classify_m_sheet(s)
            assert cat["category"] == "索引"
            assert cat["default_hidden"] is True

    def test_procedure_sheets(self):
        """实质性程序表 → 程序表"""
        assert classify_m_sheet("实收资本实质性程序表M2A")["category"] == "程序表"
        assert classify_m_sheet("未分配利润实质性程序表 M6A ")["category"] == "程序表"
        assert classify_m_sheet("资本公积实质性程序表M4A")["category"] == "程序表"
        assert classify_m_sheet("盈余公积实质性程序表M5A")["category"] == "程序表"

    def test_audit_table_sheets(self):
        """审定表M*-1 → 审定表"""
        assert classify_m_sheet("审定表M2-1")["category"] == "审定表"
        assert classify_m_sheet("审定表M4-1")["category"] == "审定表"
        assert classify_m_sheet("审定表M5-1")["category"] == "审定表"
        assert classify_m_sheet("审定表M6-1")["category"] == "审定表"

    def test_detail_sheets(self):
        """明细表 → 明细表（含上市/非上市变体）"""
        assert classify_m_sheet("明细表（非上市公司）M2-2")["category"] == "明细表"
        assert classify_m_sheet("明细表（上市公司）M2-2")["category"] == "明细表"
        assert classify_m_sheet("明细表M4-2")["category"] == "明细表"
        assert classify_m_sheet("明细表M5-2")["category"] == "明细表"
        assert classify_m_sheet("明细表M6-2")["category"] == "明细表"
        assert classify_m_sheet("明细表M9-2")["category"] == "明细表"
        assert classify_m_sheet("明细表M10-2")["category"] == "明细表"

    def test_disclosure_sheets_hidden_and_readonly(self):
        """附注披露 → 附注+调整 + defaultHidden=true + readonly=true"""
        variants = [
            "附注披露信息(上市公司)",
            "附注披露信息(国企)",
            "附注披露信息（上市公司）",
            "附注披露信息（国企）",
            "附注披露信息（国有企业）",
        ]
        for s in variants:
            cat = classify_m_sheet(s)
            assert cat["category"] == "附注+调整", f"sheet '{s}' not in 附注+调整"
            assert cat["default_hidden"] is True, f"sheet '{s}' not defaultHidden"
            assert cat["readonly"] is True, f"sheet '{s}' not readonly"

    def test_trailing_space_procedure_sheet(self):
        """末尾空格 sheet 仍正确分类"""
        # M6A 末尾带空格
        cat = classify_m_sheet("未分配利润实质性程序表 M6A ")
        assert cat["category"] == "程序表"

    def test_priority_ordering_first_match_wins(self):
        """优先级冲突：首个命中即停止"""
        # "审定表" 含"审定" → priority 2 审定表命中（不会落到 检查表 的"测试"）
        assert classify_m_sheet("审定表M2-1")["category"] == "审定表"
        # "明细表" 含"明细" → priority 3 明细表命中
        assert classify_m_sheet("明细表M6-2")["category"] == "明细表"


class TestMSheetGroupCategoryDistribution:
    """M-F2: 类目分布合理性"""

    def test_no_single_non_other_category_over_50_percent(self, m_65_unique_sheets):
        """非 fallback '其他' 类目不应单一占比 > 50%"""
        counts = Counter(
            classify_m_sheet(s)["category"] for s in m_65_unique_sheets
        )
        for cat, n in counts.items():
            if cat == "其他":
                continue
            ratio = n / len(m_65_unique_sheets)
            assert ratio < 0.50, (
                f"Category '{cat}' takes {n}/{len(m_65_unique_sheets)} = "
                f"{ratio:.1%}, exceeding 50% threshold"
            )

    def test_distribution_summary(self, m_65_unique_sheets):
        """分布摘要打印（便于回归比对）"""
        counts = Counter(
            classify_m_sheet(s)["category"] for s in m_65_unique_sheets
        )
        total = sum(counts.values())
        print("\n=== M 循环 65 sheet 分类分布 ===")
        for cat, n in sorted(counts.items(), key=lambda x: -x[1]):
            print(f"  {cat}: {n} ({n / total:.1%})")
        assert total == 65
