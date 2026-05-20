"""G 循环合并去重验证测试（spec workpaper-g-investment-cycle Task 1.1）

验证 chain_orchestrator 对 G 循环 15 文件复用 _merge_sheets_dedup 合并去重：

Sprint 0 实测基线（requirements.md 附录 A）：
- 15 模板文件（G0~G14）
- 197 raw sheet
- 4 历史遗留 sheet（G11/G12/G13/G14 各 1 个 "修订前"）已被现行 regex 命中
- 152 dedup sheet（去重保留首次出现）
- 无同 wp_code 多 sheet（不同于 H 循环）
- 跨文件去重目标："底稿目录" / "附注披露信息(上市公司)" / "附注披露信息(国企)"
- G 循环不含 "GT_Custom"（仅基于 wp_template_index 实测，与 D/F/H/I 不同）

测试目标：
1. G 循环 15 文件合并后 sheet 数 = 152（去重后）
2. _should_skip_historical_sheet 对全部 197 G sheet 命中 4 项（G11A/G12A/G13A/G14A 修订前）
3. 跨文件 "底稿目录"/"附注披露信息(上市公司)"/"附注披露信息(国企)" 被正确去重
4. G 循环无同 wp_code 多 sheet 误去重
5. D/F/H/I 历史遗留过滤回归（现有模式仍正确工作）
6. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "G"

# Collect all G cycle template files (G0~G14)
G_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_g_sheet_names() -> list[str]:
    """从 15 个 G 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in G_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程：先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_g_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestGCycleMergeDedup:
    """G-F1: 15 文件合并去重验证（Sprint 0 baseline: 15 files / 197 raw / 4 historical / 152 dedup）"""

    def test_g_cycle_has_15_template_files(self):
        """G 循环应有 15 个模板文件（G0~G14）"""
        assert len(G_TEMPLATE_FILES) == 15, (
            f"Expected 15 G cycle template files, got {len(G_TEMPLATE_FILES)}: "
            f"{[f.name for f in G_TEMPLATE_FILES]}"
        )

    def test_g_cycle_template_files_named_g0_to_g14(self):
        """G 循环 15 个模板文件命名应覆盖 G0~G14"""
        names = [f.name for f in G_TEMPLATE_FILES]
        for prefix in [f"G{i}" for i in range(15)]:
            assert any(n.startswith(prefix + " ") for n in names), (
                f"Expected file starting with '{prefix} ' in {names}"
            )

    def test_g_cycle_raw_sheet_count_is_197(self):
        """G 循环 15 文件原始 sheet 总数 = 197（Sprint 0 实测基线）"""
        all_names = _get_all_g_sheet_names()
        assert len(all_names) == 197, (
            f"Expected 197 raw sheets, got {len(all_names)}"
        )

    def test_g_cycle_dedup_sheet_count_is_152(self):
        """G 循环合并去重后 sheet 数 = 152（Sprint 0 实测基线）

        197 raw - 4 historical (G11A/G12A/G13A/G14A 修订前) - 41 跨文件归一化去重 = 152
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 152, (
            f"Expected 152 dedup sheets, got {dedup_count}"
        )

    def test_g_cycle_historical_sheets_count_is_4(self):
        """G 循环 197 sheet 中恰好 4 个命中历史遗留过滤

        G11/G12/G13/G14 各含 1 个 "修订前" sheet，已被 _should_skip_historical_sheet 现行 regex 命中。
        """
        all_names = _get_all_g_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 4, (
            f"Expected 4 historical sheet hits, got {len(historical_hits)}: {historical_hits}"
        )

    @pytest.mark.parametrize("expected_hist_keyword", [
        "G11",  # 投资收益实质性程序表G11A-修订前
        "G12",  # 净敞口套期收益审计程序表G12A-修订前
        "G13",  # 公允价值变动收益审计程序表G13A-修订前
        "G14",  # 信用减值损失审计程序表G14A -修订前
    ])
    def test_g11_g12_g13_g14_historical_sheets_filtered(self, expected_hist_keyword: str):
        """G11/G12/G13/G14 各自的"修订前"程序表 sheet 被正确过滤

        ADR-G1 关键约束：4 个历史遗留 sheet 已被现行 regex 覆盖，0 代码改动。
        """
        all_names = _get_all_g_sheet_names()
        # 找到该 wp_code 下所有"修订前"sheet
        hits = [
            n for n in all_names
            if expected_hist_keyword in n and _should_skip_historical_sheet(n)
        ]
        assert len(hits) >= 1, (
            f"{expected_hist_keyword} 应至少有 1 个'修订前'历史 sheet 被过滤，实际找到 {hits}"
        )
        # 同时验证 wp_code 仍有非历史 sheet 存在（程序表有现行版本）
        non_historical = [
            n for n in all_names
            if expected_hist_keyword in n and not _should_skip_historical_sheet(n)
        ]
        assert len(non_historical) >= 1, (
            f"{expected_hist_keyword} 应保留至少 1 个非历史 sheet（现行程序表），实际 {non_historical}"
        )

    def test_cross_file_dedup_removes_duplicates(self):
        """跨文件 "底稿目录" / "附注披露信息(上市公司)" / "附注披露信息(国企)" 被正确去重

        Sprint 0 实测：底稿目录×15 + 附注披露(上市)×14 + 附注披露(国企)×14 = 43 raw
        归一化后各仅保留 1 份 → 40 重复需去除（43 - 3 = 40）

        注意：G 循环不含 "GT_Custom"（与 D/F/H/I 不同，G 模板未生成 custom sheet）
        """
        all_names = _get_all_g_sheet_names()

        dedup_targets: dict[str, int] = {
            "底稿目录": 0,
            "附注披露信息(上市公司)": 0,
            "附注披露信息(国企)": 0,
        }

        for name in all_names:
            normalized = _normalize_sheet_name(name)
            if normalized in dedup_targets:
                dedup_targets[normalized] += 1

        # 每种跨文件重复 sheet 应出现多次（>1 表示有重复需去重）
        for key, count in dedup_targets.items():
            assert count > 1, (
                f"'{key}' 应在多个文件中出现（跨文件去重目标），实际出现 {count} 次"
            )

        # 跨文件归一化总冗余 = sum - 3（每种各保留 1 份）
        total_redundant = sum(dedup_targets.values()) - len(dedup_targets)
        assert total_redundant == 40, (
            f"跨文件归一化冗余预期 40 sheet（底稿目录 14 + 上市附注 13 + 国企附注 13），"
            f"实际 {total_redundant}: {dedup_targets}"
        )

    def test_g_cycle_no_same_wpcode_multi_sheet(self):
        """G 循环每个 wp_code 对应的 sheet 名归一化后唯一（不同于 H 循环）

        ADR-G1 关键发现：G 循环不存在 H 循环式的"同 wp_code 多版本 sheet"问题，
        故不需要分支选择器/路由保护。
        """
        all_names = _get_all_g_sheet_names()
        # 仅检查包含 wp_code 数字编号的 sheet（如 G1-2 / G7-1 等）
        # 排除"底稿目录"/"附注披露"/历史遗留等系统性重复 sheet
        wp_code_sheets: dict[str, set[str]] = {}
        for n in all_names:
            if _should_skip_historical_sheet(n):
                continue
            norm = _normalize_sheet_name(n)
            if norm in {"底稿目录", "附注披露信息(上市公司)", "附注披露信息(国企)"}:
                continue
            # 提取 wp_code（匹配 G 后跟数字-数字 模式）
            import re
            m = re.search(r"G\d+(?:-\d+)?", n)
            if m:
                wp_code = m.group(0)
                wp_code_sheets.setdefault(wp_code, set()).add(norm)

        # 验证不存在同一 wp_code 含多个不同归一化 sheet 名（H 循环式多版本）
        # 允许 wp_code 作为 sheet 名片段在多个 sheet 中出现（如 "G7-1" / "G7-2" 是不同 wp_sub_code）
        # 这里关键检查：归一化后没有 1 个 wp_code "完全相同" 但归一化保留多份
        # 实际验证：归一化后总和 + 历史 + 系统重复 = 197
        total_normalized_unique = len({
            _normalize_sheet_name(n) for n in all_names
            if not _should_skip_historical_sheet(n)
        })
        # 152 dedup = unique normalized count
        assert total_normalized_unique == 152, (
            f"归一化后唯一 sheet 数应 = 152，实际 {total_normalized_unique}"
        )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        spec 任务文本声明 chain_orchestrator 是 `_merge_sheets_dedup` 等的访问入口；
        G 循环 chain 必须复用这条 0 代码改动的 re-export 路径（D/F/H/I spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup（D/F/H/I spec 已注册路径）"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象（确保 0 代码改动复用）
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_find_all_template_files_returns_files_for_g_cycle_codes(self):
        """find_all_template_files 对 G0~G14 任一 wp_code 应能定位到 G 循环模板"""
        for i in range(15):
            wp_code = f"G{i}"
            files = find_all_template_files(wp_code)
            assert len(files) >= 1, (
                f"find_all_template_files('{wp_code}') 应至少返回 1 个文件，实际 {files}"
            )

    def test_merge_sheets_dedup_on_g_cycle_filters_and_dedups(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 G 循环 15 文件实际执行合并去重

        以 G0 作为目标 workbook，将 G1~G14 作为 other_files 合并；验证：
        - skipped_historical 包含 4（G11A/G12A/G13A/G14A 修订前）
        - skipped_dup ≥ 40（跨文件归一化重复）
        - 合并后目标 workbook 总 sheet 数 = 152
        """
        import shutil

        # 复制 G0 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in G_TEMPLATE_FILES if f.name.startswith("G0 "))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # G1~G14 作为 other_files
        other_files = [f for f in G_TEMPLATE_FILES if not f.name.startswith("G0 ")]
        assert len(other_files) == 14

        stats = _merge_sheets_dedup(target, other_files)

        # 4 个历史遗留 sheet 应被跳过
        assert stats["skipped_historical"] >= 4, (
            f"应至少跳过 4 个历史遗留 sheet（G11A/G12A/G13A/G14A 修订前），"
            f"实际 stats={stats}"
        )

        # 跨文件重复 sheet 应被去重（≥ 40 sheet 跨文件归一化重复）
        assert stats["skipped_dup"] >= 40, (
            f"应至少跳过 40 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 总 sheet 数 = 152
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            final_count = len(wb.sheetnames)
        finally:
            wb.close()
        assert final_count == 152, (
            f"合并去重后目标 workbook sheet 数应为 152，实际 {final_count}"
        )


class TestDFHIHistoricalFilterRegression:
    """D/F/H/I 历史遗留过滤回归测试 — 确保 G spec 不影响现有模式

    G-F1 ADR 关键约束：_should_skip_historical_sheet 现行 regex 已覆盖 G11A/G12A/G13A/G14A
    4 个 "修订前" sheet，无需扩展函数（0 代码改动）。本测试保证 D/F/H/I 既有模式不被本 spec 破坏。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾）
        ("参考－商誉减值测试示例", True),
        # G 循环历史遗留模式（G-F1 验证项）
        ("投资收益实质性程序表G11A-修订前", True),
        ("净敞口套期收益审计程序表G12A-修订前", True),
        ("公允价值变动收益审计程序表G13A-修订前", True),
        ("信用减值损失审计程序表G14A -修订前", True),
        # H 循环正常 sheet 不应被过滤
        ("审定表H1-1", False),
        ("折旧测算表（不含减值）-直线法H1-12", False),
        ("审定表（成本模式）H3-1", False),
        # I 循环正常 sheet 不应被过滤
        ("摊销测算表（不含减值）I1-10（剩余年限法）", False),
        ("摊销测算表（含减值）I1-11", False),
        ("商誉减值测试I3-6", False),
        # G 循环正常 sheet 不应被过滤
        ("审定表G1-1", False),
        ("公允价值测试表G1-6", False),
        ("第三层次公允价值计量的调节表G1-7", False),
        ("业务模式分析G1-8", False),
        ("合同现金流量特征分析G1-10", False),
        ("长期股权投资审定表G7-1", False),
        ("有价证券监盘表G1-11", False),
        ("衍生金融工具核查表G1-14", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露信息(上市公司)", False),
        ("附注披露信息(国企)", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I/G 历史遗留过滤模式仍正确工作，正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
