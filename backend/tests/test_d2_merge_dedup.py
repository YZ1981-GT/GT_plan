"""D2 三文件合并 sheet 去重单测（spec workpaper-d-sales-cycle 任务 1.6）。

Validates: Requirements F2 + F3（D 销售循环 spec — D2 多文件合并去重 / D4 修订前过滤）
ADR: D2（sheet 名归一化）+ D3（修订前/(原) 历史 sheet 过滤）

覆盖 4 类去重规则：
  1. 中英文圆括号归一化：`审定表（D2-1）` ↔ `审定表(D2-1)` 视为同名
  2. GT_Custom 归一化：含 `GT_Custom` 字串一律归一为 `"GT_Custom"`
  3. 底稿目录归一化：含 `底稿目录` 字串一律归一为 `"底稿目录"`
  4. 修订前/(原) 历史 sheet 过滤：含 `修订前` / `（原）` / `(原)` 的 sheet 不参与合并

测试既覆盖纯函数级（_normalize_sheet_name / _should_skip_historical_sheet）也
覆盖端到端（_merge_sheets_dedup 用真实 openpyxl xlsx 文件）。
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
)


# ---------------------------------------------------------------------------
# 辅助：openpyxl 工具
# ---------------------------------------------------------------------------


def _create_xlsx(path: Path, sheet_names: list[str]) -> Path:
    """在 ``path`` 创建一个 xlsx，按顺序包含 ``sheet_names`` 列出的 sheet。

    每个 sheet 写入一个能识别的 marker 单元格 A1，方便断言 sheet 来源未串号。
    """
    wb = Workbook()
    # openpyxl 默认带一个 "Sheet"，先 remove
    default = wb.active
    wb.remove(default)
    for sn in sheet_names:
        # Excel sheet 名上限 31 字符；调用方传入超长名时由 _merge_sheets_dedup 截断
        ws = wb.create_sheet(title=sn[:31])
        ws["A1"] = f"marker:{sn}"
    wb.save(str(path))
    return path


def _read_sheetnames(path: Path) -> list[str]:
    wb = load_workbook(str(path), read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# TestNormalizeSheetName: 4 类归一化规则 + 幂等
# ---------------------------------------------------------------------------


class TestNormalizeSheetName:
    """_normalize_sheet_name 纯函数测试。"""

    def test_chinese_paren_normalized_to_english(self) -> None:
        """规则 1：中文圆括号 `（…）` 应归一为英文 `(…)`，与英文括号同名视为重复。"""
        cn = _normalize_sheet_name("应收账款审定表（D2-1）")
        en = _normalize_sheet_name("应收账款审定表(D2-1)")
        assert cn == en, (
            f"中英文圆括号未归一为同名：cn={cn!r}, en={en!r}"
        )
        # 进一步：归一化结果应仅含英文括号
        assert "（" not in cn and "）" not in cn
        assert "(" in cn and ")" in cn

    def test_gt_custom_collapsed_to_canonical(self) -> None:
        """规则 2：所有含 `GT_Custom` 的变体归一为字面量 `"GT_Custom"`。"""
        variants = [
            "GT_Custom",
            "GT_Custom_xxx",
            "GT_Custom 自定义",
            "GT_Custom_2",
            "  GT_Custom_company  ",
            "GT_Custom（合并）",
        ]
        normalized = {_normalize_sheet_name(v) for v in variants}
        assert normalized == {"GT_Custom"}, (
            f"GT_Custom 变体未全部归一为 'GT_Custom'，实际 {normalized}"
        )

    def test_index_sheet_collapsed_to_canonical(self) -> None:
        """规则 3：所有含 `底稿目录` 的变体归一为字面量 `"底稿目录"`。"""
        variants = [
            "底稿目录",
            "底稿目录(2)",
            "底稿目录（2）",
            "  底稿目录  ",
            "底稿目录_old",
        ]
        normalized = {_normalize_sheet_name(v) for v in variants}
        assert normalized == {"底稿目录"}, (
            f"底稿目录 变体未全部归一为 '底稿目录'，实际 {normalized}"
        )

    def test_idempotent(self) -> None:
        """幂等：normalize(normalize(x)) == normalize(x)（PBT P2 设计）。"""
        samples = [
            "应收账款审定表（D2-1）",
            "应收账款审定表(D2-1)",
            "GT_Custom_xxx",
            "底稿目录(2)",
            "明细表D2-2",
            "  前后空白  ",
            "中间 空格 表",  # 空白会被剔除
            "",
        ]
        for s in samples:
            once = _normalize_sheet_name(s)
            twice = _normalize_sheet_name(once)
            assert once == twice, (
                f"normalize 非幂等：input={s!r}, once={once!r}, twice={twice!r}"
            )

    def test_none_input_is_safe(self) -> None:
        """边界：None 输入返回空串，不抛异常。"""
        assert _normalize_sheet_name(None) == ""  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# TestShouldSkipHistoricalSheet: 修订前 / (原) 历史 sheet 检测
# ---------------------------------------------------------------------------


class TestShouldSkipHistoricalSheet:
    """_should_skip_historical_sheet 纯函数测试。"""

    @pytest.mark.parametrize(
        "name,expected",
        [
            # 应跳过：含 "修订前"
            ("主营业务收入审计程序表 D4A（修订前）", True),
            ("D4A（修订前）", True),
            ("收入审计修订前版本", True),
            # 应跳过：含 "（原）" / "(原)"
            ("D7A（原）", True),
            ("D8A(原)", True),
            ("应收账款审定表（原）", True),
            ("D9审定表(原)", True),
            # 不应跳过：正常 sheet
            ("底稿目录", False),
            ("应收账款审定表(D2-1)", False),
            ("应收账款审定表（D2-1）", False),  # 中文括号但内容非"原"
            ("GT_Custom", False),
            ("明细表D2-2", False),
            ("D4A", False),
            # 边界
            ("", False),
        ],
    )
    def test_historical_sheet_detection(self, name: str, expected: bool) -> None:
        assert _should_skip_historical_sheet(name) is expected, (
            f"sheet_name={name!r} 期望 skip={expected}，实际 {not expected}"
        )

    def test_none_input_is_safe(self) -> None:
        """边界：None 输入不抛异常，返回 False。"""
        assert _should_skip_historical_sheet(None) is False  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# TestMergeSheetsDedupE2E: 用真实 openpyxl xlsx 验证端到端合并
# ---------------------------------------------------------------------------


class TestMergeSheetsDedupE2E:
    """_merge_sheets_dedup 端到端测试，使用临时 xlsx 文件。"""

    def test_dedup_chinese_paren_variants(self, tmp_path: Path) -> None:
        """中英文圆括号变体应去重为同一 sheet。"""
        target = _create_xlsx(
            tmp_path / "target.xlsx",
            ["应收账款审定表(D2-1)"],
        )
        other = _create_xlsx(
            tmp_path / "other.xlsx",
            ["应收账款审定表（D2-1）"],  # 中文括号变体
        )

        stats = _merge_sheets_dedup(target, [other])

        assert stats == {"merged": 0, "skipped_dup": 1, "skipped_historical": 0}, (
            f"中英文括号去重 stats 不符：{stats}"
        )
        # target 仍只有原 sheet，未新增中文括号版本
        assert _read_sheetnames(target) == ["应收账款审定表(D2-1)"]

    def test_dedup_gt_custom_variants(self, tmp_path: Path) -> None:
        """target 已有 GT_Custom_xxx，other 的 GT_Custom_yyy 应被去重。"""
        target = _create_xlsx(
            tmp_path / "target.xlsx",
            ["GT_Custom_company"],
        )
        other = _create_xlsx(
            tmp_path / "other.xlsx",
            ["GT_Custom_other_company"],
        )

        stats = _merge_sheets_dedup(target, [other])

        assert stats == {"merged": 0, "skipped_dup": 1, "skipped_historical": 0}, (
            f"GT_Custom 去重 stats 不符：{stats}"
        )
        # target 仅保留原 GT_Custom_company（先到先得）
        assert _read_sheetnames(target) == ["GT_Custom_company"]

    def test_dedup_index_variants(self, tmp_path: Path) -> None:
        """target 已有 `底稿目录`，other 的 `底稿目录(2)` 应被去重。"""
        target = _create_xlsx(
            tmp_path / "target.xlsx",
            ["底稿目录"],
        )
        other = _create_xlsx(
            tmp_path / "other.xlsx",
            ["底稿目录(2)"],
        )

        stats = _merge_sheets_dedup(target, [other])

        assert stats == {"merged": 0, "skipped_dup": 1, "skipped_historical": 0}, (
            f"底稿目录 去重 stats 不符：{stats}"
        )
        assert _read_sheetnames(target) == ["底稿目录"]

    def test_filter_historical_sheets(self, tmp_path: Path) -> None:
        """修订前 / (原) 历史 sheet 应被 skip，不参与合并。"""
        target = _create_xlsx(
            tmp_path / "target.xlsx",
            ["底稿目录"],
        )
        other = _create_xlsx(
            tmp_path / "other.xlsx",
            [
                "主营业务收入审计程序表D4A（修订前）",
                "应收账款审定表D7A（原）",
                "存货审定表D8A(原)",
            ],
        )

        stats = _merge_sheets_dedup(target, [other])

        assert stats == {
            "merged": 0,
            "skipped_dup": 0,
            "skipped_historical": 3,
        }, f"历史 sheet 过滤 stats 不符：{stats}"
        # target 不应包含任何"修订前 / 原"版本
        names = _read_sheetnames(target)
        assert names == ["底稿目录"]
        for n in names:
            assert "修订前" not in n
            assert "（原）" not in n and "(原)" not in n

    def test_merge_dedup_returns_correct_stats(self, tmp_path: Path) -> None:
        """端到端综合场景：target + 2 个 other 文件，覆盖 4 类去重规则。

        target.xlsx     : ["底稿目录", "审定表(D2-1)", "GT_Custom_target"]
        other1.xlsx     : ["底稿目录(2)", "审定表（D2-1）", "GT_Custom_other"]
                          → 三者均与 target 归一化重复 → skipped_dup=3
        other2.xlsx     : ["明细表D2-2", "审计程序表D4A（修订前）"]
                          → 明细表D2-2 是新 sheet → merged=1
                          → 修订前 → skipped_historical=1

        预期 stats = {merged: 1, skipped_dup: 3, skipped_historical: 1}
        """
        target = _create_xlsx(
            tmp_path / "target.xlsx",
            ["底稿目录", "审定表(D2-1)", "GT_Custom_target"],
        )
        other1 = _create_xlsx(
            tmp_path / "other1.xlsx",
            ["底稿目录(2)", "审定表（D2-1）", "GT_Custom_other"],
        )
        other2 = _create_xlsx(
            tmp_path / "other2.xlsx",
            ["明细表D2-2", "审计程序表D4A（修订前）"],
        )

        stats = _merge_sheets_dedup(target, [other1, other2])

        assert stats == {
            "merged": 1,
            "skipped_dup": 3,
            "skipped_historical": 1,
        }, f"综合场景 stats 不符：{stats}"

        names = _read_sheetnames(target)
        # target 应在原 3 个 sheet 末尾追加 明细表D2-2
        assert names == [
            "底稿目录",
            "审定表(D2-1)",
            "GT_Custom_target",
            "明细表D2-2",
        ], f"target sheetnames 不符：{names}"

        # 反向断言：不应出现修订前 / 原 / 中文括号变体 / 底稿目录(2) / 多余的 GT_Custom 副本
        for n in names:
            assert "修订前" not in n
            assert "（原）" not in n and "(原)" not in n
        assert "底稿目录(2)" not in names
        assert "审定表（D2-1）" not in names
        assert "GT_Custom_other" not in names

    def test_new_sheet_content_copied(self, tmp_path: Path) -> None:
        """合并新 sheet 时单元格内容应被复制（验证 _create_sheet + iter_rows 流程）。"""
        target = _create_xlsx(tmp_path / "target.xlsx", ["底稿目录"])
        other = _create_xlsx(tmp_path / "other.xlsx", ["明细表D2-2"])

        stats = _merge_sheets_dedup(target, [other])
        assert stats == {"merged": 1, "skipped_dup": 0, "skipped_historical": 0}

        # 重新打开 target，新 sheet A1 应为 source 写入的 marker
        wb = load_workbook(str(target), read_only=True)
        try:
            assert "明细表D2-2" in wb.sheetnames
            assert wb["明细表D2-2"]["A1"].value == "marker:明细表D2-2"
        finally:
            wb.close()

    def test_empty_other_files_no_op(self, tmp_path: Path) -> None:
        """边界：other_files 为空列表不应改变 target，stats 全 0。"""
        target = _create_xlsx(tmp_path / "target.xlsx", ["底稿目录", "审定表(D2-1)"])
        before = _read_sheetnames(target)

        stats = _merge_sheets_dedup(target, [])

        assert stats == {"merged": 0, "skipped_dup": 0, "skipped_historical": 0}
        assert _read_sheetnames(target) == before
