"""P2-9: 交付成果审定财务报表导出 — 权益变动表 {{eq:}} 回填集成测试.

模拟交付中心 render_financial_reports 走 ReportExcelExporter mode=audited，
验证 enrich_equity_statement_rows 内存回填后占位符有值。
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
import pytest_asyncio
from openpyxl import Workbook, load_workbook
from sqlalchemy import MetaData, select
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.core import Project
from app.models.report_models import FinancialReport, FinancialReportType
from app.services.report_excel_exporter import ReportExcelExporter

SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON
if not hasattr(SQLiteTypeCompiler, "visit_ARRAY"):
    SQLiteTypeCompiler.visit_ARRAY = lambda self, type_, **kw: "TEXT"

_engine = create_async_engine("sqlite+aiosqlite:///:memory:")
_EQ_SHEET = "5-所有者权益变动表（企财04表-合并）"

_TEST_TABLES = [Project.__table__, FinancialReport.__table__]


@pytest_asyncio.fixture
async def db_session() -> AsyncSession:
    async with _engine.begin() as conn:
        await conn.run_sync(
            lambda sync_conn: MetaData().create_all(sync_conn, tables=_TEST_TABLES)
        )
    factory = async_sessionmaker(_engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as session:
        yield session
    async with _engine.begin() as conn:
        await conn.run_sync(
            lambda sync_conn: MetaData().drop_all(sync_conn, tables=_TEST_TABLES)
        )


def _bs(pid, code, name, current, prior):
    now = datetime.now(timezone.utc)
    return FinancialReport(
        id=uuid.uuid4(),
        project_id=pid,
        year=2024,
        report_type=FinancialReportType.balance_sheet,
        row_code=code,
        row_name=name,
        current_period_amount=Decimal(str(current)),
        prior_period_amount=Decimal(str(prior)),
        generated_at=now,
    )


def _eq(pid, code, name):
    now = datetime.now(timezone.utc)
    return FinancialReport(
        id=uuid.uuid4(),
        project_id=pid,
        year=2024,
        report_type=FinancialReportType.equity_statement,
        row_code=code,
        row_name=name,
        generated_at=now,
    )


@pytest_asyncio.fixture
async def audited_equity_seed(db_session: AsyncSession):
    project = Project(
        id=uuid.uuid4(),
        name="交付权益导出测试",
        client_name="交付权益导出测试",
        template_type="soe",
        report_scope="standalone",
    )
    db_session.add(project)
    rows = [
        _bs(project.id, "BS-101", "所有者权益：", 0, 0),
        _bs(project.id, "BS-102", "实收资本", 5000000, 4000000),
        _bs(project.id, "BS-113", "资本公积", 800000, 700000),
        _bs(project.id, "BS-125", "未分配利润", 2000000, 1500000),
        _bs(project.id, "BS-128", "所有者权益合计", 7800000, 6200000),
        _eq(project.id, "EQ-001", "一、上年年末余额"),
    ]
    for r in rows:
        db_session.add(r)
    await db_session.flush()
    return project


@pytest.mark.asyncio
async def test_deliverable_audited_equity_eq_placeholders_filled(
    db_session: AsyncSession, audited_equity_seed, tmp_path, monkeypatch,
):
    """审定导出（交付路径）：EQ-001 矩阵由 BS prior enrich 后写入 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _EQ_SHEET
    ws["C9"] = "{{eq:EQ-001:current_year:share_capital}}"
    ws["G9"] = "{{eq:EQ-001:current_year:capital_reserve}}"
    ws["M9"] = "{{eq:EQ-001:current_year:retained_earnings}}"
    synth = tmp_path / "synthetic_deliverable_eq.xlsx"
    wb.save(synth)

    monkeypatch.setattr(
        ReportExcelExporter, "_load_template",
        lambda self, key: load_workbook(str(synth)),
    )

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=audited_equity_seed.id,
        year=2024,
        mode="audited",
        report_types=["equity_statement"],
        include_prior_year=True,
    )
    out = load_workbook(output)[_EQ_SHEET]

    assert out["C9"].value == pytest.approx(4000000.0)
    assert out["G9"].value == pytest.approx(700000.0)
    assert out["M9"].value == pytest.approx(1500000.0)

    res = await db_session.execute(
        select(FinancialReport).where(
            FinancialReport.project_id == audited_equity_seed.id,
            FinancialReport.row_code == "EQ-001",
        )
    )
    eq1 = res.scalar_one()
    # 导出 enrich 为内存路径，不写库
    assert eq1.source_accounts is None

    for row in out.iter_rows():
        for cell in row:
            assert "{{eq:" not in str(cell.value or "")
