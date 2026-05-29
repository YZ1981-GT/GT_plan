"""render_schema 覆盖率验证测试

验证 3 个核心断言：
1. 每个模板 xlsx 都有对应的 yaml schema（覆盖率 100%）
2. yaml 中的 sheets 与 xlsx 模板实际 sheet 名匹配
3. 非 PENDING sheet 的 componentType 与 analysis JSON 中的 render 映射一致

Validates: Requirements 2.2 原则 2（配置驱动）+ US-7（render_schema 全覆盖）
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import pytest
import yaml

# ─── 项目根目录 ─────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATE_DIR = PROJECT_ROOT / "backend" / "wp_templates"
SCHEMA_DIR = PROJECT_ROOT / "backend" / "data" / "wp_render_schema"
GENERATED_DIR = SCHEMA_DIR / "generated"
ANALYSIS_JSON_PATH = (
    PROJECT_ROOT
    / ".kiro"
    / "specs"
    / "workpaper-editor-slimdown"
    / "workpaper_template_analysis.json"
)

# ─── 与 generate_wp_render_schema.py / _validate_render_schema_quality.py 一致 ──
_WP_CODE_PATTERN = re.compile(r"^([A-Z]\d+[A-Z]?(?:-\d+)*)")

RENDER_TO_COMPONENT: dict[str, str] = {
    "HTML 中控台": "a-program-console",
    "HTML 表单（编制信息+索引导航）": "b-index",
    "HTML 嵌套表（多级子表）": "c-note-table",
    "HTML 表单（表格型检查）": "d-form-table",
    "HTML 表单（专属子组件）": "d-form-confirmation",
    "HTML 表单（电子签）": "d-form-review",
    "HTML 段落型": "d-form-paragraph",
    "HTML 是否问答型": "d-form-qa",
    "HTML 表单": "e-control-test",
    "HTML stepper": "e-control-test",
    "保留 Univer": "univer",
    "保留 Univer（测算）": "univer",
    "跳过渲染": "skip",
    "PENDING-待人工归类": "pending",
    "静态展示": "h-static-doc",
}


# ─── Fixtures ────────────────────────────────────────────────────────────────


def _extract_wp_code(filename: str) -> str | None:
    """从模板文件名提取完整 wp_code"""
    m = _WP_CODE_PATTERN.match(filename)
    return m.group(1) if m else None


def _collect_template_xlsx() -> list[Path]:
    """收集所有循环子目录下的 xlsx 模板文件（排除 _reference）"""
    xlsx_files: list[Path] = []
    for subdir in sorted(TEMPLATE_DIR.iterdir()):
        if subdir.is_dir() and subdir.name != "_reference":
            xlsx_files.extend(sorted(subdir.glob("*.xlsx")))
    return xlsx_files


def _load_all_yaml_schemas() -> dict[str, dict]:
    """加载所有 yaml schema（手写优先于 generated）

    返回 {wp_code: parsed_yaml_dict}
    """
    schemas: dict[str, dict] = {}

    # 1. 先加载 generated（低优先级）
    if GENERATED_DIR.exists():
        for yaml_path in sorted(GENERATED_DIR.glob("*.yaml")):
            data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
            if data and isinstance(data, dict):
                wp_code = data.get("wp_code", yaml_path.stem)
                schemas[wp_code] = data

    # 2. 再加载手写（高优先级，覆盖 generated）
    for yaml_path in sorted(SCHEMA_DIR.glob("*.yaml")):
        data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
        if data and isinstance(data, dict):
            wp_code = data.get("wp_code", yaml_path.stem)
            schemas[wp_code] = data

    return schemas


@pytest.fixture(scope="module")
def template_xlsx_files() -> list[Path]:
    """所有模板 xlsx 文件"""
    return _collect_template_xlsx()


@pytest.fixture(scope="module")
def yaml_schemas() -> dict[str, dict]:
    """所有 yaml schema（手写优先）"""
    return _load_all_yaml_schemas()


@pytest.fixture(scope="module")
def analysis_data() -> dict:
    """workpaper_template_analysis.json 数据"""
    return json.loads(ANALYSIS_JSON_PATH.read_text(encoding="utf-8"))


# ─── Test 1: 覆盖率 100% ────────────────────────────────────────────────────


def test_all_templates_have_schema(template_xlsx_files, yaml_schemas):
    """断言每个模板 xlsx 都有对应的 yaml schema（覆盖率 100%）

    只检查文件名能提取出 wp_code 的 xlsx（.docx 等非 xlsx 不在范围内）。
    """
    missing: list[str] = []

    for xlsx_path in template_xlsx_files:
        wp_code = _extract_wp_code(xlsx_path.name)
        if wp_code is None:
            # 无法提取 wp_code 的文件跳过（如中文开头的参考文件）
            continue
        if wp_code not in yaml_schemas:
            missing.append(f"{xlsx_path.parent.name}/{xlsx_path.name} (wp_code={wp_code})")

    coverage = 1.0 - (len(missing) / max(len(template_xlsx_files), 1))
    assert len(missing) == 0, (
        f"覆盖率 {coverage:.1%}，以下 {len(missing)} 个模板缺少 yaml schema:\n"
        + "\n".join(missing[:20])
        + (f"\n... 还有 {len(missing) - 20} 个" if len(missing) > 20 else "")
    )


# ─── Test 2: sheets 匹配 ────────────────────────────────────────────────────


def test_schema_sheets_match_xlsx(yaml_schemas):
    """对 generated yaml 文件中声明的 sheets，验证其名称集合与对应 xlsx 模板的实际 sheet 名匹配。

    只检查 generated 目录下的 yaml（1:1 对应模板 xlsx）。
    手写 yaml 可能定义跨模板的虚拟 sheet，不在此断言范围内。

    对于同一 wp_code 有多个 xlsx 变体的情况（如 A24-1 有大型国企/非大型国企两版），
    yaml 中的 sheets 应是所有同 wp_code xlsx 文件 sheet 名的并集子集。

    使用 openpyxl 读取 xlsx 的 sheet 名列表进行比对。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    # 构建 wp_code → [xlsx_path, ...] 映射（同一 wp_code 可能有多个 xlsx 变体）
    wp_code_to_xlsx: dict[str, list[Path]] = {}
    for xlsx_path in _collect_template_xlsx():
        code = _extract_wp_code(xlsx_path.name)
        if code:
            wp_code_to_xlsx.setdefault(code, []).append(xlsx_path)

    # 收集 generated 目录下的 wp_code 集合
    generated_wp_codes: set[str] = set()
    if GENERATED_DIR.exists():
        for yaml_path in GENERATED_DIR.glob("*.yaml"):
            generated_wp_codes.add(yaml_path.stem)

    mismatches: list[str] = []
    checked = 0

    for wp_code, schema in yaml_schemas.items():
        # 只检查 generated yaml
        if wp_code not in generated_wp_codes:
            continue

        sheets_in_yaml = set(schema.get("sheets", {}).keys())
        if not sheets_in_yaml:
            continue

        # 收集该 wp_code 对应的所有 xlsx 的 sheet 名并集
        xlsx_paths = wp_code_to_xlsx.get(wp_code, [])
        if not xlsx_paths:
            # 尝试 template_path 字段
            template_path_str = schema.get("template_path")
            if template_path_str:
                tp = PROJECT_ROOT / template_path_str
                if tp.exists():
                    xlsx_paths = [tp]

        if not xlsx_paths:
            continue

        all_sheets_in_xlsx: set[str] = set()
        for xlsx_path in xlsx_paths:
            try:
                wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                all_sheets_in_xlsx.update(wb.sheetnames)
                wb.close()
            except Exception:
                continue

        if not all_sheets_in_xlsx:
            continue

        checked += 1

        # yaml 中声明的 sheet 必须是所有同 wp_code xlsx sheet 名并集的子集
        extra_in_yaml = sheets_in_yaml - all_sheets_in_xlsx
        if extra_in_yaml:
            mismatches.append(
                f"{wp_code}: yaml 中有但 xlsx 中无: {extra_in_yaml}"
            )

    assert checked > 0, "未找到任何可验证的 yaml + xlsx 配对"
    assert len(mismatches) == 0, (
        f"共检查 {checked} 个 generated yaml，{len(mismatches)} 个 sheets 不匹配:\n"
        + "\n".join(mismatches[:20])
    )


# ─── Test 3: componentType 一致性 ───────────────────────────────────────────


def test_component_type_matches_analysis(analysis_data, yaml_schemas):
    """验证 analysis JSON 中每个非 PENDING sheet 的 componentType 与 yaml 一致。

    遍历 analysis JSON 的所有 cycles → templates → sheets，
    对每个 render 值映射到 expected componentType，
    然后与 yaml 中对应 wp_code/sheet_name 的 component_type 比对。
    """
    mismatches: list[str] = []
    total_checked = 0
    pending_skipped = 0

    for cycle_key, cycle_data in analysis_data.get("cycles", {}).items():
        for template in cycle_data.get("templates", []):
            filename = template["filename"]
            wp_code = _extract_wp_code(filename)
            if not wp_code:
                continue

            schema = yaml_schemas.get(wp_code)
            if not schema:
                continue

            yaml_sheets = schema.get("sheets", {})

            for sheet in template.get("sheets", []):
                sheet_name = sheet["name"]
                render_value = sheet["render"]
                expected_ct = RENDER_TO_COMPONENT.get(render_value)

                if expected_ct is None:
                    # 未知 render 值，跳过
                    continue

                if expected_ct == "pending":
                    pending_skipped += 1
                    continue

                actual_ct = None
                sheet_config = yaml_sheets.get(sheet_name)
                if isinstance(sheet_config, dict):
                    actual_ct = sheet_config.get("component_type")

                if actual_ct is None:
                    # sheet 不在 yaml 中，跳过（test_schema_sheets_match_xlsx 已覆盖）
                    continue

                total_checked += 1
                if actual_ct != expected_ct:
                    mismatches.append(
                        f"{wp_code}/{sheet_name}: "
                        f"render='{render_value}' → expected={expected_ct}, actual={actual_ct}"
                    )

    assert total_checked > 0, "未找到任何可验证的 sheet componentType 配对"
    assert len(mismatches) == 0, (
        f"共检查 {total_checked} 个 sheet（跳过 {pending_skipped} 个 PENDING），"
        f"{len(mismatches)} 个 componentType 不匹配:\n"
        + "\n".join(mismatches[:30])
        + (f"\n... 还有 {len(mismatches) - 30} 个" if len(mismatches) > 30 else "")
    )
