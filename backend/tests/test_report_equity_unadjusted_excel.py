"""P2-8: 未审 Excel 导出 {{eq:}} 占位符集成测试.

验证 mode=unadjusted 时，权益变动表占位从四表未审数动态计算的 eq_matrix 回填。
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
import pytest_asyncio
from openpyxl import Workbook, load_workbook
from sqlalchemy import MetaData
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.audit_platform_models import AccountCategory, TrialBalance
from app.models.core import Project, ProjectStatus, ProjectType
from app.models.report_models import FinancialReportType, ReportConfig
from app.services.report_excel_exporter import ReportExcelExporter

SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON
if not hasattr(SQLiteTypeCompiler, "visit_ARRAY"):
    SQLiteTypeCompiler.visit_ARRAY = lambda self, type_, **kw: "TEXT"

TEST_DATABASE_URL = "sqlite+aiosqlite:///:memory:"
test_engine = create_async_engine(TEST_DATABASE_URL, echo=False)

FAKE_USER_ID = uuid.uuid4()
FAKE_PROJECT_ID = uuid.uuid4()
_EQ_SHEET = "5-所有者权益变动表（企财04表-合并）"

_TEST_TABLES = [
    Project.__table__,
    TrialBalance.__table__,
    ReportConfig.__table__,
]


@pytest_asyncio.fixture
async def db_session() -> AsyncSession:
    async with test_engine.begin() as conn:
        await conn.run_sync(
            lambda sync_conn: MetaData().create_all(sync_conn, tables=_TEST_TABLES)
        )
    session_factory = async_sessionmaker(
        test_engine, class_=AsyncSession, expire_on_commit=False,
    )
    async with session_factory() as session:
        yield session
    async with test_engine.begin() as conn:
        await conn.run_sync(
            lambda sync_conn: MetaData().drop_all(sync_conn, tables=_TEST_TABLES)
        )


@pytest_asyncio.fixture
async def unadjusted_equity_seed(db_session: AsyncSession):
    project = Project(
        id=FAKE_PROJECT_ID,
        name="未审Excel权益测试",
        client_name="未审Excel权益测试",
        project_type=ProjectType.annual,
        status=ProjectStatus.planning,
        created_by=FAKE_USER_ID,
        template_type="soe",
        report_scope="standalone",
    )
    db_session.add(project)
    await db_session.flush()

    tb_rows = [
        ("4001", "实收资本", AccountCategory.equity, Decimal("5000000"), Decimal("4800000"), Decimal("4000000")),
        ("4002", "资本公积", AccountCategory.equity, Decimal("800000"), Decimal("750000"), Decimal("700000")),
        ("4104", "未分配利润", AccountCategory.equity, Decimal("2000000"), Decimal("1900000"), Decimal("1500000")),
    ]
    for year in (2023, 2024):
        for code, name, cat, audited, unadjusted, opening in tb_rows:
            yr_audited = opening if year == 2023 else audited
            yr_unadjusted = opening if year == 2023 else unadjusted
            db_session.add(
                TrialBalance(
                    project_id=FAKE_PROJECT_ID,
                    year=year,
                    company_code="default",
                    standard_account_code=code,
                    account_name=name,
                    account_category=cat,
                    audited_amount=yr_audited,
                    unadjusted_amount=yr_unadjusted,
                    opening_balance=opening,
                )
            )

    for row_code, row_name, formula, row_num, indent, is_total, rt in [
        ("BS-101", "所有者权益：", None, 101, 0, False, FinancialReportType.balance_sheet),
        ("BS-102", "实收资本", "TB('4001','期末余额')", 102, 0, False, FinancialReportType.balance_sheet),
        ("BS-113", "资本公积", "TB('4002','期末余额')", 113, 0, False, FinancialReportType.balance_sheet),
        ("BS-125", "未分配利润", "TB('4104','期末余额')", 125, 0, False, FinancialReportType.balance_sheet),
        ("BS-128", "所有者权益合计", "ROW('BS-102')+ROW('BS-113')+ROW('BS-125')", 128, 0, True, FinancialReportType.balance_sheet),
        ("EQ-001", "一、上年年末余额", None, 1, 0, False, FinancialReportType.equity_statement),
    ]:
        db_session.add(
            ReportConfig(
                applicable_standard="soe_standalone",
                report_type=rt,
                row_code=row_code,
                row_name=row_name,
                formula=formula,
                row_number=row_num,
                indent_level=indent,
                is_total_row=is_total,
            )
        )
    await db_session.flush()


@pytest.mark.asyncio
async def test_unadjusted_excel_eq_placeholder_fill(
    db_session: AsyncSession, unadjusted_equity_seed, tmp_path, monkeypatch,
):
    """未审导出：EQ-001 share_capital 来自未审 BS 上期（opening 4000000）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _EQ_SHEET
    ws["C9"] = "{{eq:EQ-001:current_year:share_capital}}"
    ws["G9"] = "{{eq:EQ-001:current_year:capital_reserve}}"
    ws["P9"] = "{{eq:EQ-001:prior_year:share_capital}}"
    synth = tmp_path / "synthetic_unadj_eq.xlsx"
    wb.save(synth)

    monkeypatch.setattr(
        ReportExcelExporter, "_load_template",
        lambda self, key: load_workbook(str(synth)),
    )

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=FAKE_PROJECT_ID,
        year=2024,
        mode="unadjusted",
        report_types=["equity_statement"],
        include_prior_year=True,
    )
    assert isinstance(output, BytesIO)
    out = load_workbook(output)[_EQ_SHEET]

    assert out["C9"].value == pytest.approx(4000000.0)
    assert out["G9"].value == pytest.approx(700000.0)
    assert out["P9"].value == pytest.approx(4000000.0)
    for row in out.iter_rows():
        for cell in row:
            assert "{{eq:" not in str(cell.value or ""), f"残留: {cell.coordinate}={cell.value!r}"
