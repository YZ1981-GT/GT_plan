"""L 循环合并去重验证测试（spec workpaper-l-debt-cycle Task 1.1）

验证 chain_orchestrator 对 L 循环 9 文件复用 _merge_sheets_dedup 合并去重：

Sprint 0 实测基线（requirements.md 附录 A / design.md ADR-L1）：
- 9 模板文件（L0 债务循环函证 / L1 短期借款 / L2 应付利息 / L3 长期借款 /
  L4 应付债券 / L5 长期应付款 / L6 专项应付款 / L7 其他非流动负债 / L8 财务费用）
- 100 raw sheet
- 1 历史遗留 sheet（'函证差异检查表（示例）' in L0，"（示例）"模式命中）
- 1 末尾空格 sheet（'应付债券实质性程序表L4A '，L4 文件）
- 20 跨文件归一化重复：
  底稿目录(9×→8) / GT_Custom(4×→3) / 附注披露信息核对(上市)(3×→2) /
  附注披露信息(上市)(3×→2) / 附注披露信息(国企)(3×→2) /
  附注披露信息核对(国企)(2×→1) / 附注披露(国企)信息(2×→1) /
  附注披露(上市)信息(2×→1) = 20
- 合并去重后有效 sheet = 79（100 raw - 1 历史 - 20 跨文件去重 = 79）

测试目标：
1. L 循环 9 文件被 find_all_template_files 正确发现
2. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
3. 100 raw sheets → 1 历史遗留 → 20 跨文件去重 → 79 有效 sheet
4. 1 末尾空格 sheet（'应付债券实质性程序表L4A '）被正确保留
5. 跨文件去重（底稿目录 / GT_Custom / 附注披露变体）正确工作
6. D/F/H/I/G/J/K 历史遗留过滤回归无影响

_Requirements: L-F1_
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "L"

# Collect all L cycle template files
L_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_l_sheet_names() -> list[str]:
    """从 9 个 L 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in L_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程：先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_l_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestLCycleMergeDedup:
    """L-F1: 9 文件合并去重验证（Sprint 0 baseline: 9 files / 100 raw / 1 historical / 79 dedup）"""

    def test_l_cycle_has_9_template_files(self):
        """L 循环应有 9 个模板文件（L0~L8）"""
        assert len(L_TEMPLATE_FILES) == 9, (
            f"Expected 9 L cycle template files, got {len(L_TEMPLATE_FILES)}: "
            f"{[f.name for f in L_TEMPLATE_FILES]}"
        )

    def test_l_cycle_template_files_named_correctly(self):
        """L 循环 9 个模板文件命名应覆盖 L0~L8 全部编号"""
        names = [f.name for f in L_TEMPLATE_FILES]
        for prefix in [f"L{i}" for i in range(9)]:
            assert any(n.startswith(prefix + " ") for n in names), (
                f"Expected file starting with '{prefix} ' in {names}"
            )

    def test_l_cycle_raw_sheet_count_100(self):
        """L 循环 9 文件原始 sheet 总数 = 100（Sprint 0 实测基线）"""
        all_names = _get_all_l_sheet_names()
        assert len(all_names) == 100, (
            f"Expected 100 raw sheets, got {len(all_names)}"
        )

    def test_l_cycle_dedup_sheet_count_79(self):
        """L 循环合并去重后 sheet 数 = 79（Sprint 0 实测基线）

        100 raw - 1 historical（函证差异检查表（示例）） - 20 跨文件归一化去重 = 79
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 79, (
            f"Expected 79 dedup sheets, got {dedup_count}"
        )

    def test_l_cycle_historical_sheets_count_is_1(self):
        """L 循环 100 sheet 中恰好 1 个命中历史遗留过滤（'函证差异检查表（示例）'）"""
        all_names = _get_all_l_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 1, (
            f"Expected 1 historical sheet hit, got {len(historical_hits)}: {historical_hits}"
        )

    def test_l_cycle_historical_sheet_is_hanzheng_example(self):
        """L 循环唯一的历史遗留 sheet 是 '函证差异检查表（示例）'（L0 文件）"""
        all_names = _get_all_l_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 1
        assert "函证差异检查表" in historical_hits[0], (
            f"Expected '函证差异检查表（示例）', got '{historical_hits[0]}'"
        )
        assert "示例" in historical_hits[0], (
            f"Expected '（示例）' pattern in '{historical_hits[0]}'"
        )

    def test_l_cycle_trailing_space_sheet_count_1(self):
        """L 循环恰好 1 个 sheet 末尾带空格（'应付债券实质性程序表L4A '）"""
        all_names = _get_all_l_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        assert len(trailing_space_sheets) == 1, (
            f"Expected 1 trailing-space sheet, got {len(trailing_space_sheets)}: "
            f"{trailing_space_sheets}"
        )

    def test_l_cycle_trailing_space_sheet_is_l4a(self):
        """末尾空格 sheet 是 '应付债券实质性程序表L4A '（L4 文件）"""
        all_names = _get_all_l_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        assert len(trailing_space_sheets) == 1
        assert "L4A" in trailing_space_sheets[0], (
            f"Expected trailing-space sheet to contain 'L4A', got '{trailing_space_sheets[0]}'"
        )
        assert trailing_space_sheets[0].endswith(" "), (
            f"Sheet name should end with space: repr='{trailing_space_sheets[0]!r}'"
        )

    def test_l_cycle_trailing_space_sheet_preserved_in_dedup(self):
        """末尾空格 sheet '应付债券实质性程序表L4A ' 在去重后被正确保留（不被过滤）"""
        trailing_sheet = "应付债券实质性程序表L4A "
        # 不应被历史遗留过滤
        assert _should_skip_historical_sheet(trailing_sheet) is False, (
            f"Trailing-space sheet should NOT be filtered as historical"
        )

    def test_l_cycle_cross_file_redundant_count_20(self):
        """L 循环跨文件去重应移除 20 个重复 sheet

        9 文件包含的跨文件重复 sheet（归一化后）：
        - 底稿目录(9×→8 dup)
        - GT_Custom(4×→3 dup)
        - 附注披露信息核对(上市)(3×→2 dup)
        - 附注披露信息(上市)(3×→2 dup)
        - 附注披露信息(国企)(3×→2 dup)
        - 附注披露信息核对(国企)(2×→1 dup)
        - 附注披露(国企)信息(2×→1 dup)
        - 附注披露(上市)信息(2×→1 dup)
        合计 8+3+2+2+2+1+1+1 = 20 重复需去除
        """
        all_names = _get_all_l_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)
        dups = {k: v for k, v in norm_counter.items() if v > 1}
        total_redundant = sum(v - 1 for v in dups.values())
        assert total_redundant == 20, (
            f"Expected 20 cross-file redundant sheets, got {total_redundant}: {dups}"
        )

    def test_l_cycle_cross_file_dedup_targets_identified(self):
        """验证跨文件去重目标 sheet 分布：底稿目录 9× / GT_Custom 4×"""
        all_names = _get_all_l_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 底稿目录 9 个文件各 1 份
        assert norm_counter.get("底稿目录", 0) == 9, (
            f"'底稿目录' should appear 9 times (one per file), "
            f"got {norm_counter.get('底稿目录', 0)}"
        )

        # GT_Custom 出现在 4 个文件
        assert norm_counter.get("GT_Custom", 0) == 4, (
            f"'GT_Custom' should appear 4 times, got {norm_counter.get('GT_Custom', 0)}"
        )

    def test_l_cycle_no_problematic_wp_code_multi_sheet(self):
        """L 循环同 wp_code 多 sheet 仅限 L4-7/L4-8 债券两种付息方式（不影响去重）

        L4-7 有 2 个 sheet（到期一次还本付息 / 分期付息到期一次还本）
        L4-8 有 2 个 sheet（同上两种付息方式）
        这是合法的业务多版本（类似 H 循环成本/公允模式），归一化后 key 不同不会被误去重。
        除此之外无其他同 wp_code 多 sheet 情况。
        """
        all_names = _get_all_l_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]

        # 提取含 L+数字 的 sheet 名中的 wp_code 部分
        import re
        wp_code_pattern = re.compile(r"L\d+[A-Za-z]?-\d+")
        wp_codes_found: dict[str, list[str]] = {}
        for name in non_hist:
            matches = wp_code_pattern.findall(name)
            for m in matches:
                wp_codes_found.setdefault(m, []).append(name)

        # 检查同一 wp_code 对应多个不同 sheet 名（归一化后）
        multi_sheet_codes = {}
        for code, sheets in wp_codes_found.items():
            unique_normalized = set(_normalize_sheet_name(s) for s in sheets)
            if len(unique_normalized) > 1:
                multi_sheet_codes[code] = list(unique_normalized)

        # 仅 L4-7 和 L4-8 有合法的两种付息方式变体
        assert set(multi_sheet_codes.keys()) == {"L4-7", "L4-8"}, (
            f"Only L4-7 and L4-8 should have multi-sheet variants (bond payment modes), "
            f"but found: {multi_sheet_codes}"
        )
        # 每个都恰好 2 个变体
        for code in ["L4-7", "L4-8"]:
            assert len(multi_sheet_codes[code]) == 2, (
                f"{code} should have exactly 2 variants, got {multi_sheet_codes[code]}"
            )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        L 循环 chain 必须复用这条 re-export 路径（D/F/H/I/G/J/K spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_find_all_template_files_returns_l_cycle_files(self):
        """find_all_template_files 对 L0~L8 各 wp_code 应能定位到 L 循环模板"""
        for i in range(9):
            wp_code = f"L{i}"
            files = find_all_template_files(wp_code)
            assert len(files) >= 1, (
                f"find_all_template_files('{wp_code}') should return >= 1 file, got {files}"
            )

    def test_find_all_template_files_l_cycle_total_9(self):
        """find_all_template_files 对 L 循环 9 个 wp_code 合计返回 9 个文件"""
        all_files = []
        for i in range(9):
            wp_code = f"L{i}"
            all_files.extend(find_all_template_files(wp_code))
        assert len(all_files) == 9, (
            f"L cycle should have 9 template files total, got {len(all_files)}: "
            f"{[f.name for f in all_files]}"
        )

    def test_merge_sheets_dedup_on_l_cycle(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 L 循环 9 文件实际执行合并去重

        以 L0 作为目标 workbook，将其余 8 文件作为 other_files 合并；验证：
        - skipped_historical >= 0（L0 自身含 '函证差异检查表（示例）' 但 target 内不删除）
        - skipped_dup >= 10（跨文件归一化重复）
        - 合并后逻辑有效 sheet 数 = 79（去重后基线）
        """
        import shutil

        # 选 L0 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in L_TEMPLATE_FILES if f.name.startswith("L0 "))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # 其余 8 个文件作为 other_files
        other_files = [f for f in L_TEMPLATE_FILES if not f.name.startswith("L0 ")]
        assert len(other_files) == 8

        stats = _merge_sheets_dedup(target, other_files)

        # 跨文件重复 sheet 应被去重
        assert stats["skipped_dup"] >= 10, (
            f"应至少跳过 10 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 物理 sheet 数
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            all_sheets = wb.sheetnames
        finally:
            wb.close()

        # 逻辑有效 sheet 数 = 79（去重后基线）
        logical_count = sum(
            1 for s in all_sheets if not _should_skip_historical_sheet(s)
        )
        assert logical_count == 79, (
            f"合并后逻辑有效 sheet 数应为 79（dedup baseline），实际 {logical_count}"
        )


class TestPriorCyclesHistoricalFilterRegression:
    """D/F/H/I/G/J/K 历史遗留过滤回归测试 — 确保 L spec 无副作用

    L 循环未引入新的过滤模式（仅复用已有"（示例）"模式），本测试保证
    D/F/H/I/G/J/K 既有模式不被破坏，且 L 正常 sheet 不被误过滤。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾）
        ("参考－商誉减值测试示例", True),
        # G 循环历史遗留模式
        ("投资收益实质性程序表G11A-修订前", True),
        # J 循环新增"-删除"模式
        ("股份支付检查表J1-10-删除", True),
        ("应付职工薪酬余额及期后事项检查表J1-11-删除", True),
        # L 循环唯一历史遗留（"（示例）"模式）
        ("函证差异检查表（示例）", True),
        # L 循环正常 sheet 不应被过滤
        ("审定表L1-1", False),
        ("明细表L1-2", False),
        ("利息测算表L1-5", False),
        ("审定表L3-1", False),
        ("明细表L3-2", False),
        ("利息测算表L3-5", False),
        ("应付债券实质性程序表L4A ", False),  # 末尾空格保留
        ("审定表L5-1", False),
        ("明细表L5-2", False),
        ("审定表L6-1", False),
        ("明细表L6-2", False),
        ("审定表L8-1", False),
        ("明细表L8-2", False),
        ("短期借款实质性程序表L1A", False),
        ("长期借款实质性程序表L3A", False),
        ("财务费用实质性程序表L8A", False),
        # K 循环正常 sheet 不应被过滤
        ("审定表K8-1", False),
        ("审定表K9-1", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露信息(上市公司)", False),
        ("附注披露信息(国企)", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I/G/J/K 历史遗留过滤模式仍正确工作，L 正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
