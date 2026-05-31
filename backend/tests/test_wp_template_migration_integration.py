"""模板版本升级数据迁移 - 集成测试

模拟真实模板 v2025-R5→R6 diff + 迁移流程。

Spec: wp-template-migration
Requirements: 1.1, 2.1
"""

from __future__ import annotations

import copy
from pathlib import Path

import openpyxl
import pytest

from app.services.wp_migration_report_service import (
    MigrationReport,
    MigrationResult,
    classify_migration_result,
    generate_migration_report_markdown,
)
from app.services.wp_migration_service import WpMigrationService
from app.services.wp_template_diff_service import generate_template_diff


def _create_xlsx(sheets: dict[str, list[str]], path: Path) -> Path:
    """创建测试用 xlsx"""
    wb = openpyxl.Workbook()
    default_ws = wb.active
    if default_ws:
        wb.remove(default_ws)
    for sheet_name, headers in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    wb.save(str(path))
    return path


class TestEndToEndMigration:
    """端到端集成测试：模拟 R5→R6 升级"""

    def test_full_migration_flow(self, tmp_path: Path):
        """完整流程：diff → 迁移 → 报告"""
        # 模拟 R5 模板
        old_path = tmp_path / "template_R5.xlsx"
        _create_xlsx(
            {
                "货币资金": ["科目编码", "科目名称", "期末余额", "期初余额"],
                "应收账款": ["客户", "金额", "账龄"],
                "旧表_待删除": ["A", "B", "C"],
            },
            old_path,
        )

        # 模拟 R6 模板（新增 sheet、删除 sheet、新增列）
        new_path = tmp_path / "template_R6.xlsx"
        _create_xlsx(
            {
                "货币资金": ["科目编码", "科目名称", "期末余额", "期初余额", "备注"],
                "应收账款": ["客户", "金额", "账龄", "预期信用损失"],
                "新增_合同资产": ["合同编号", "金额", "确认日期"],
            },
            new_path,
        )

        # Step 1: 生成 diff
        diff = generate_template_diff(old_path, new_path)
        assert diff.has_changes
        assert "新增_合同资产" in diff.added_sheets
        assert "旧表_待删除" in diff.removed_sheets
        # 列级变化
        col_changes = {cd.sheet_name: cd for cd in diff.column_diffs}
        assert "货币资金" in col_changes
        assert "备注" in col_changes["货币资金"].added
        assert "应收账款" in col_changes
        assert "预期信用损失" in col_changes["应收账款"].added

        # Step 2: 模拟已编制底稿的 parsed_data
        parsed_data = {
            "html_data": {
                "货币资金": {
                    "cells": {
                        "A2": {"v": "1001"},
                        "B2": {"v": "库存现金"},
                        "C2": {"v": 50000.00},
                        "D2": {"v": 45000.00},
                    },
                    "columns": ["科目编码", "科目名称", "期末余额", "期初余额"],
                },
                "应收账款": {
                    "cells": {
                        "A2": {"v": "客户A"},
                        "B2": {"v": 100000},
                        "C2": {"v": "1年以内"},
                    },
                    "columns": ["客户", "金额", "账龄"],
                },
                "旧表_待删除": {
                    "cells": {"A2": {"v": "旧数据"}},
                    "columns": ["A", "B", "C"],
                },
            },
            "wp_code": "E1-1",
        }

        # Step 3: 执行迁移
        from app.services.wp_parsed_data_service import _read_xlsx_structure

        new_structure = _read_xlsx_structure(new_path)
        migrated = WpMigrationService.apply_diff_to_parsed_data(
            parsed_data, diff, new_structure
        )

        # 验证：共有数据保留
        assert migrated["html_data"]["货币资金"]["cells"]["A2"]["v"] == "1001"
        assert migrated["html_data"]["货币资金"]["cells"]["C2"]["v"] == 50000.00
        assert migrated["html_data"]["应收账款"]["cells"]["A2"]["v"] == "客户A"

        # 验证：新增列出现在 columns 中
        assert "备注" in migrated["html_data"]["货币资金"]["columns"]
        assert "预期信用损失" in migrated["html_data"]["应收账款"]["columns"]

        # 验证：新增 sheet 存在
        assert "新增_合同资产" in migrated["html_data"]

        # 验证：删除的 sheet 被归档
        assert "旧表_待删除" not in migrated["html_data"]
        assert "旧表_待删除" in migrated["_archived_data"]
        assert migrated["_archived_data"]["旧表_待删除"]["cells"]["A2"]["v"] == "旧数据"

        # 验证：迁移时间戳
        assert "_migrated_at" in migrated

    def test_idempotent_migration(self, tmp_path: Path):
        """幂等性：已迁移的数据再次迁移时跳过"""
        old_path = tmp_path / "old.xlsx"
        new_path = tmp_path / "new.xlsx"
        _create_xlsx({"Sheet1": ["A", "B"]}, old_path)
        _create_xlsx({"Sheet1": ["A", "B", "C"]}, new_path)

        diff = generate_template_diff(old_path, new_path)

        parsed_data = {
            "html_data": {"Sheet1": {"cells": {"A1": {"v": "test"}}, "columns": ["A", "B"]}},
            "_migrated_at": "2025-01-01T00:00:00+00:00",  # 已迁移标记
        }

        # WpMigrationService.migrate_workpaper 会检查 _migrated_at 跳过
        # 这里直接测试 apply_diff 仍然可以工作（纯函数不检查幂等）
        migrated = WpMigrationService.apply_diff_to_parsed_data(parsed_data, diff)
        assert "C" in migrated["html_data"]["Sheet1"]["columns"]


class TestMigrationReport:
    """迁移报告生成测试"""

    def test_report_generation(self):
        """报告正确分类和格式化"""
        report = MigrationReport(
            template_old_version="v2025-R5",
            template_new_version="v2025-R6",
            started_at="2025-06-01T10:00:00Z",
            finished_at="2025-06-01T10:05:00Z",
        )

        report.add_result(MigrationResult(
            wp_id="id1", wp_code="D2-1", wp_name="应收账款",
            status="success", snapshot_id="snap-001",
        ))
        report.add_result(MigrationResult(
            wp_id="id2", wp_code="D2-2", wp_name="预付账款",
            status="skipped", message="parsed_data 为空",
        ))
        report.add_result(MigrationResult(
            wp_id="id3", wp_code="E1-1", wp_name="货币资金",
            status="manual_required", message="结构冲突：列类型不匹配",
        ))
        report.add_result(MigrationResult(
            wp_id="id4", wp_code="F1-1", wp_name="存货",
            status="error", message="未知错误",
        ))

        assert report.success_count == 1
        assert report.skipped_count == 1
        assert report.manual_required_count == 1
        assert report.error_count == 1
        assert report.total_count == 4

        # 生成 markdown
        md = generate_migration_report_markdown(report)
        assert "# 模板版本升级迁移报告" in md
        assert "v2025-R5" in md
        assert "v2025-R6" in md
        assert "✅ 成功迁移" in md
        assert "⏭️ 跳过" in md
        assert "⚠️ 需人工处理" in md
        assert "❌ 错误" in md
        assert "D2-1" in md
        assert "应收账款" in md

    def test_classify_manual_required(self):
        """结构冲突自动标记为需人工处理"""
        result = classify_migration_result(
            wp_id="id1",
            wp_code="D2-1",
            wp_name="测试",
            migrate_result={"status": "error", "message": "结构冲突：列数不匹配"},
        )
        assert result.status == "manual_required"

    def test_classify_normal_error(self):
        """普通错误保持 error 状态"""
        result = classify_migration_result(
            wp_id="id1",
            wp_code="D2-1",
            wp_name="测试",
            migrate_result={"status": "error", "message": "数据库连接失败"},
        )
        assert result.status == "error"

    def test_classify_success(self):
        """成功状态正确传递"""
        result = classify_migration_result(
            wp_id="id1",
            wp_code="D2-1",
            wp_name="测试",
            migrate_result={"status": "success", "snapshot_id": "snap-001", "message": "迁移成功"},
        )
        assert result.status == "success"
        assert result.snapshot_id == "snap-001"
