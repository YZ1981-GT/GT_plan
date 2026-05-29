"""Sprint 4 Task 15.15 — 底稿离线导出→导入 round-trip PBT.

**Validates: Requirements US-14**

Uses hypothesis to generate random workpaper fill data and verifies
that export → import round-trip produces correct diffs.

Property: For any set of editable cells with random values,
exporting a workpaper template and then importing it back should
correctly identify all changed cells in the diff.
"""
from __future__ import annotations

import string

from hypothesis import given, settings, HealthCheck
from hypothesis import strategies as st

from app.services.wp_offline_export_service import (
    export_workpaper_template,
    _decompress_meta,
)
from app.services.wp_offline_import_service import (
    validate_import_file,
    diff_sheets,
    apply_import,
    ConflictStrategy,
    DiffType,
)


# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------

# Generate a valid sheet name (no special chars, 1-20 chars)
sheet_name_st = st.text(
    alphabet=string.ascii_letters + string.digits + "_",
    min_size=1,
    max_size=20,
)

# Generate cell values (numbers, strings, None)
cell_value_st = st.one_of(
    st.none(),
    st.integers(min_value=-1_000_000, max_value=1_000_000),
    st.floats(min_value=-1e6, max_value=1e6, allow_nan=False, allow_infinity=False),
    st.text(alphabet=string.ascii_letters + string.digits + " ", min_size=0, max_size=30),
)

# Generate a row (list of cell values, 1-8 columns)
row_st = st.lists(cell_value_st, min_size=1, max_size=8)

# Generate headers matching row width
def headers_for_width(width: int):
    return st.lists(
        st.text(alphabet=string.ascii_letters, min_size=1, max_size=10),
        min_size=width,
        max_size=width,
    )


# Generate a single sheet with consistent dimensions
@st.composite
def sheet_st(draw):
    name = draw(sheet_name_st)
    num_cols = draw(st.integers(min_value=1, max_value=8))
    num_rows = draw(st.integers(min_value=1, max_value=10))
    headers = draw(st.lists(
        st.text(alphabet=string.ascii_letters, min_size=1, max_size=10),
        min_size=num_cols,
        max_size=num_cols,
    ))
    rows = []
    for _ in range(num_rows):
        cells = draw(st.lists(cell_value_st, min_size=num_cols, max_size=num_cols))
        rows.append({"cells": cells})

    # All cells are editable (manual source)
    cell_meta = {}
    for r_idx in range(num_rows):
        for c_idx in range(num_cols):
            cell_meta[f"{r_idx}:{c_idx}"] = {"source": "manual", "mode": "manual"}

    return {
        "sheet_name": name,
        "headers": headers,
        "rows": rows,
        "cell_meta": cell_meta,
        "formulas": {},
        "row_meta": [],
    }


# Generate multiple sheets with unique names
@st.composite
def sheets_data_st(draw):
    num_sheets = draw(st.integers(min_value=1, max_value=3))
    sheets = []
    used_names = set()
    for _ in range(num_sheets):
        sheet = draw(sheet_st())
        # Ensure unique names
        while sheet["sheet_name"] in used_names:
            sheet["sheet_name"] = sheet["sheet_name"] + "_"
        used_names.add(sheet["sheet_name"])
        sheets.append(sheet)
    return sheets


# ---------------------------------------------------------------------------
# Properties
# ---------------------------------------------------------------------------


@given(sheets_data=sheets_data_st())
@settings(
    max_examples=30,
    deadline=10000,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture],
)
def test_export_import_roundtrip_validates(sheets_data: list[dict]):
    """Property: exported xlsx always passes validation on re-import."""
    xlsx_bytes, file_hash = export_workpaper_template(
        sheets_data,
        wp_id="test-wp-001",
        wp_code="A1-1",
        project_id="test-project-001",
        year=2025,
        project_name="测试项目",
        exporter_name="测试用户",
    )

    assert len(xlsx_bytes) > 0
    assert len(file_hash) == 64  # SHA-256 hex

    # Validate the exported file
    result = validate_import_file(xlsx_bytes)
    assert result.valid, f"Validation failed: {result.errors}"
    assert result.wp_id == "test-wp-001"
    assert len(result.sheet_names) == len(sheets_data)


@given(sheets_data=sheets_data_st())
@settings(
    max_examples=30,
    deadline=10000,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture],
)
def test_export_import_no_change_diff_empty(sheets_data: list[dict]):
    """Property: importing unchanged file produces zero diffs."""
    xlsx_bytes, _ = export_workpaper_template(
        sheets_data,
        wp_id="test-wp-002",
        wp_code="B1-1",
        project_id="test-project-002",
        year=2025,
    )

    # Validate to get meta_data (needed for row count)
    validation = validate_import_file(xlsx_bytes)
    assert validation.valid

    # Diff against same data should be empty
    diffs = diff_sheets(xlsx_bytes, sheets_data, validation.meta_data)
    for sheet_diff in diffs:
        if sheet_diff.status == "matched":
            assert len(sheet_diff.cell_diffs) == 0, (
                f"Sheet {sheet_diff.sheet_name} has unexpected diffs: "
                f"{[d.to_dict() for d in sheet_diff.cell_diffs]}"
            )


@given(
    sheets_data=sheets_data_st(),
    modified_value=st.one_of(
        st.integers(min_value=1, max_value=999999),
        st.text(alphabet=string.ascii_letters, min_size=1, max_size=10),
    ),
)
@settings(
    max_examples=30,
    deadline=10000,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture],
)
def test_export_modify_import_detects_change(sheets_data: list[dict], modified_value):
    """Property: modifying an editable cell is detected as 'changed' in diff."""
    from copy import deepcopy
    from io import BytesIO
    from openpyxl import load_workbook

    # Export
    xlsx_bytes, _ = export_workpaper_template(
        sheets_data,
        wp_id="test-wp-003",
        wp_code="C1-1",
        project_id="test-project-003",
        year=2025,
    )

    # Modify a cell in the exported xlsx
    wb = load_workbook(BytesIO(xlsx_bytes))
    # Find first content sheet (skip 注意事项 and _meta_)
    content_sheets = [ws for ws in wb.worksheets if ws.title not in ("注意事项", "_meta_")]
    if not content_sheets:
        return  # Skip if no content sheets

    target_ws = content_sheets[0]
    # Row 3 is first data row (row 1 = hidden sheet_id, row 2 = headers)
    target_ws.cell(row=3, column=1, value=modified_value)

    # Save modified xlsx
    buf = BytesIO()
    wb.save(buf)
    modified_bytes = buf.getvalue()

    # Diff against original data
    validation = validate_import_file(modified_bytes)
    assert validation.valid

    diffs = diff_sheets(modified_bytes, sheets_data, validation.meta_data)

    # Find the sheet that was modified
    target_sheet_name = sheets_data[0]["sheet_name"]
    target_diff = next((d for d in diffs if d.sheet_name == target_sheet_name), None)

    if target_diff and target_diff.status == "matched":
        # The original value at 0:0
        original_val = sheets_data[0]["rows"][0]["cells"][0] if sheets_data[0]["rows"] else None

        # If modified_value differs from original, we should see a diff
        from app.services.wp_offline_import_service import _values_equal
        if not _values_equal(original_val, modified_value):
            assert len(target_diff.cell_diffs) > 0, (
                f"Expected diff for cell 0:0 (original={original_val}, modified={modified_value})"
            )
            # Check the diff is of type CHANGED
            cell_0_0_diff = next(
                (d for d in target_diff.cell_diffs if d.cell_key == "0:0"), None
            )
            if cell_0_0_diff:
                assert cell_0_0_diff.diff_type == DiffType.CHANGED


@given(sheets_data=sheets_data_st())
@settings(
    max_examples=20,
    deadline=10000,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture],
)
def test_apply_overwrite_imports_all(sheets_data: list[dict]):
    """Property: overwrite strategy imports all matched sheets."""
    xlsx_bytes, _ = export_workpaper_template(
        sheets_data,
        wp_id="test-wp-004",
        wp_code="D1-1",
        project_id="test-project-004",
        year=2025,
    )

    result = apply_import(
        xlsx_bytes,
        sheets_data,
        ConflictStrategy.OVERWRITE,
    )

    assert result.success
    assert result.sheets_imported == len(sheets_data)
    assert result.audit_entry["strategy"] == "overwrite"
    assert result.audit_entry["retention_days"] == 30


@given(sheets_data=sheets_data_st())
@settings(
    max_examples=20,
    deadline=10000,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture],
)
def test_apply_keep_system_imports_none(sheets_data: list[dict]):
    """Property: keep_system strategy keeps all sheets unchanged."""
    xlsx_bytes, _ = export_workpaper_template(
        sheets_data,
        wp_id="test-wp-005",
        wp_code="E1-1",
        project_id="test-project-005",
        year=2025,
    )

    result = apply_import(
        xlsx_bytes,
        sheets_data,
        ConflictStrategy.KEEP_SYSTEM,
    )

    assert result.success
    assert result.sheets_imported == 0
    assert result.sheets_kept == len(sheets_data)


@given(sheets_data=sheets_data_st())
@settings(
    max_examples=20,
    deadline=10000,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture],
)
def test_encryption_roundtrip(sheets_data: list[dict]):
    """Property: encrypted export can be decrypted and validated."""
    from app.services.wp_offline_export_service import _decrypt_bytes

    password = "test_password_123"
    xlsx_bytes, file_hash = export_workpaper_template(
        sheets_data,
        wp_id="test-wp-006",
        wp_code="F1-1",
        project_id="test-project-006",
        year=2025,
        password=password,
    )

    # Encrypted bytes should not be valid xlsx directly
    direct_validation = validate_import_file(xlsx_bytes)
    assert not direct_validation.valid

    # Decrypt and validate
    decrypted = _decrypt_bytes(xlsx_bytes, password)
    result = validate_import_file(decrypted)
    assert result.valid, f"Decrypted validation failed: {result.errors}"
    assert result.wp_id == "test-wp-006"
