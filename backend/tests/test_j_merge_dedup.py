"""J 循环合并去重验证测试（spec workpaper-j-payroll-cycle Task 1.1）

验证 chain_orchestrator 对 J 循环 3 文件复用 _merge_sheets_dedup 合并去重：

Sprint 0 实测基线（requirements.md 附录 A）：
- 3 模板文件（J1 应付职工薪酬 / J2 长期应付职工薪酬 / J3 股份支付）
- 38 raw sheet
- 5 历史遗留 sheet（"-删除"后缀）已被 _should_skip_historical_sheet 命中
- 2 个"原版"sheet（J1A-原版 / L1A-原）保留不过滤（修改模板前旧版本）
- 4 跨文件去重（底稿目录×2 + 附注披露上市×1 + 附注披露国企×1）
- 合并后有效 sheet = 29（38 - 5 - 4 = 29）

测试目标：
1. J 循环 3 文件被 find_all_template_files 正确发现
2. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
3. 38 raw sheets → 5 个"-删除"被过滤 → 4 跨文件去重 → 29 有效 sheet
4. "J1A-原版" / "L1A-原" 不被过滤（保留）
5. 跨文件去重（底稿目录 / 附注披露）正确工作
6. D/F/H/I/G 历史遗留过滤回归无影响

_Requirements: J-F1_
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "J"

# Collect all J cycle template files
J_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_j_sheet_names() -> list[str]:
    """从 3 个 J 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in J_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程：先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_j_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestJCycleMergeDedup:
    """J-F1: 3 文件合并去重验证（Sprint 0 baseline: 3 files / 38 raw / 5 historical / 29 dedup）"""

    def test_j_cycle_has_3_template_files(self):
        """J 循环应有 3 个模板文件"""
        assert len(J_TEMPLATE_FILES) == 3, (
            f"Expected 3 J cycle template files, got {len(J_TEMPLATE_FILES)}: "
            f"{[f.name for f in J_TEMPLATE_FILES]}"
        )

    def test_j_cycle_template_files_named_correctly(self):
        """J 循环 3 个模板文件命名应覆盖 J1/J2/J3"""
        names = [f.name for f in J_TEMPLATE_FILES]
        for prefix in ["J1", "J2", "J3"]:
            assert any(n.startswith(prefix + " ") for n in names), (
                f"Expected file starting with '{prefix} ' in {names}"
            )

    def test_j_cycle_raw_sheet_count_is_38(self):
        """J 循环 3 文件原始 sheet 总数 = 38（Sprint 0 实测基线）"""
        all_names = _get_all_j_sheet_names()
        assert len(all_names) == 38, (
            f"Expected 38 raw sheets, got {len(all_names)}"
        )

    def test_j_cycle_dedup_sheet_count_is_29(self):
        """J 循环合并去重后 sheet 数 = 29（Sprint 0 实测基线）

        38 raw - 5 historical ("-删除") - 4 跨文件归一化去重 = 29
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 29, (
            f"Expected 29 dedup sheets, got {dedup_count}"
        )

    def test_j_cycle_historical_sheets_count_is_5(self):
        """J 循环 38 sheet 中恰好 5 个命中历史遗留过滤（"-删除"后缀）"""
        all_names = _get_all_j_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 5, (
            f"Expected 5 historical sheet hits, got {len(historical_hits)}: {historical_hits}"
        )

    @pytest.mark.parametrize("expected_sheet", [
        "股份支付检查表J1-10-删除",
        "应付职工薪酬余额及期后事项检查表J1-11-删除",
        "职工薪酬政策统计表J1-12-删除",
        "IPO企业股权激励工具关注的审计重点-删除",
        "首发业务解答二-删除",
    ])
    def test_specific_historical_sheet_filtered(self, expected_sheet: str):
        """5 个"-删除"sheet 各自被 _should_skip_historical_sheet 正确命中"""
        assert _should_skip_historical_sheet(expected_sheet) is True, (
            f"_should_skip_historical_sheet('{expected_sheet}') should be True"
        )

    @pytest.mark.parametrize("sheet_name", [
        "应付职工薪酬实质性程序表 J1A-原版",
        "应付职工薪酬实质性程序表 L1A-原",
    ])
    def test_yuanban_sheets_not_filtered(self, sheet_name: str):
        """'原版'/'原' sheet 不被过滤（修改模板前旧版本，保留供参考）

        ADR-J1 关键约束：J1A-原版 / L1A-原 是修改模板前的旧版本，保留不过滤。
        """
        assert _should_skip_historical_sheet(sheet_name) is False, (
            f"_should_skip_historical_sheet('{sheet_name}') should be False "
            f"(原版 sheets are kept, not filtered)"
        )

    def test_yuanban_sheets_exist_in_templates(self):
        """确认"原版"sheet 确实存在于 J 循环模板中"""
        all_names = _get_all_j_sheet_names()
        yuanban_sheets = [n for n in all_names if "原版" in n or n.endswith("原")]
        assert len(yuanban_sheets) == 2, (
            f"Expected 2 '原版/原' sheets in J templates, got {len(yuanban_sheets)}: {yuanban_sheets}"
        )

    def test_cross_file_dedup_removes_duplicates(self):
        """跨文件 "底稿目录" / "附注披露信息(上市公司)" / "附注披露信息(国有企业)" 被正确去重

        Sprint 0 实测：底稿目录×3 + 附注披露(上市)×2 + 附注披露(国企)×2 = 7 raw
        归一化后各仅保留 1 份 → 4 重复需去除（7 - 3 = 4）
        """
        all_names = _get_all_j_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]

        # Count occurrences of cross-file duplicate targets after normalization
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)
        dups = {k: v for k, v in norm_counter.items() if v > 1}

        # Should have cross-file duplicates
        assert len(dups) > 0, "Should have cross-file duplicates to dedup"

        # Total redundant = sum of (count - 1) for each duplicate
        total_redundant = sum(v - 1 for v in dups.values())
        assert total_redundant == 4, (
            f"Expected 4 cross-file redundant sheets (底稿目录×2 + 附注上市×1 + 附注国企×1), "
            f"got {total_redundant}: {dups}"
        )

    def test_cross_file_dedup_targets_identified(self):
        """验证跨文件去重目标：底稿目录 / 附注披露信息(上市公司) / 附注披露信息(国有企业)"""
        all_names = _get_all_j_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 底稿目录应出现 3 次（每个文件各 1 份）
        assert norm_counter.get("底稿目录", 0) == 3, (
            f"'底稿目录' should appear 3 times (one per file), got {norm_counter.get('底稿目录', 0)}"
        )

        # 附注披露信息应各出现 2 次
        shangshi_key = None
        guoqi_key = None
        for key in norm_counter:
            if "附注披露" in key and "上市" in key:
                shangshi_key = key
            elif "附注披露" in key and ("国有" in key or "国企" in key):
                guoqi_key = key

        assert shangshi_key is not None, "Should find '附注披露信息(上市公司)' in normalized names"
        assert norm_counter[shangshi_key] == 2, (
            f"'{shangshi_key}' should appear 2 times, got {norm_counter[shangshi_key]}"
        )

        assert guoqi_key is not None, "Should find '附注披露信息(国有企业)' in normalized names"
        assert norm_counter[guoqi_key] == 2, (
            f"'{guoqi_key}' should appear 2 times, got {norm_counter[guoqi_key]}"
        )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        J 循环 chain 必须复用这条 re-export 路径（D/F/H/I/G spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_find_all_template_files_returns_j_cycle_files(self):
        """find_all_template_files 对 J1/J2/J3 各 wp_code 应能定位到 J 循环模板"""
        for wp_code in ["J1", "J2", "J3"]:
            files = find_all_template_files(wp_code)
            assert len(files) >= 1, (
                f"find_all_template_files('{wp_code}') should return >= 1 file, got {files}"
            )

    def test_find_all_template_files_j_cycle_total_3(self):
        """find_all_template_files 对 J 循环 3 个 wp_code 合计返回 3 个文件"""
        all_files = []
        for wp_code in ["J1", "J2", "J3"]:
            all_files.extend(find_all_template_files(wp_code))
        assert len(all_files) == 3, (
            f"J cycle should have 3 template files total, got {len(all_files)}: "
            f"{[f.name for f in all_files]}"
        )

    def test_merge_sheets_dedup_on_j_cycle(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 J 循环 3 文件实际执行合并去重

        以 J1 作为目标 workbook，将 J2/J3 作为 other_files 合并；验证：
        - skipped_dup >= 4（跨文件归一化重复：底稿目录×2 + 附注×2）
        - merged = 11（J2 9 sheets + J3 6 sheets - 4 dup = 11）
        - 合并后目标 workbook 物理 sheet 数 = 34（J1 的 23 + 合并的 11）
        - 逻辑有效 sheet 数 = 29（34 物理 - 5 个"-删除"在 J1 中）

        注意：_merge_sheets_dedup 不删除 target 自身的历史遗留 sheet，
        仅阻止 other_files 中的历史遗留 sheet 被合并进来。
        5 个"-删除"sheet 全在 J1（target）中，故 skipped_historical=0。
        """
        import shutil

        # 复制 J1 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in J_TEMPLATE_FILES if f.name.startswith("J1 "))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # J2/J3 作为 other_files
        other_files = [f for f in J_TEMPLATE_FILES if not f.name.startswith("J1 ")]
        assert len(other_files) == 2

        stats = _merge_sheets_dedup(target, other_files)

        # J2/J3 不含"-删除"sheet，故 skipped_historical = 0
        assert stats["skipped_historical"] == 0, (
            f"J2/J3 不含'-删除'sheet，skipped_historical 应为 0，实际 stats={stats}"
        )

        # 跨文件重复 sheet 应被去重（底稿目录×2 + 附注×2 = 4）
        assert stats["skipped_dup"] >= 4, (
            f"应至少跳过 4 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 物理 sheet 数 = J1(23) + merged(11) = 34
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            all_sheets = wb.sheetnames
            physical_count = len(all_sheets)
        finally:
            wb.close()

        # 物理 sheet 数 = 23 + merged
        expected_physical = 23 + stats["merged"]
        assert physical_count == expected_physical, (
            f"合并后物理 sheet 数应为 {expected_physical}（J1=23 + merged={stats['merged']}），"
            f"实际 {physical_count}"
        )

        # 逻辑有效 sheet 数 = 物理 - 5 个"-删除"（在 J1 target 中）= 29
        logical_count = sum(
            1 for s in all_sheets if not _should_skip_historical_sheet(s)
        )
        assert logical_count == 29, (
            f"合并后逻辑有效 sheet 数应为 29（物理 {physical_count} - 5 删除），"
            f"实际 {logical_count}"
        )


class TestDFHIGHistoricalFilterRegression:
    """D/F/H/I/G 历史遗留过滤回归测试 — 确保 J spec 扩展不影响现有模式

    J-F1 扩展：新增 "-删除" 结尾通用模式。本测试保证 D/F/H/I/G 既有模式不被破坏，
    且正常 sheet 不被误过滤。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾）
        ("参考－商誉减值测试示例", True),
        # G 循环历史遗留模式
        ("投资收益实质性程序表G11A-修订前", True),
        ("净敞口套期收益审计程序表G12A-修订前", True),
        ("公允价值变动收益审计程序表G13A-修订前", True),
        ("信用减值损失审计程序表G14A -修订前", True),
        # J 循环新增"-删除"模式
        ("股份支付检查表J1-10-删除", True),
        ("应付职工薪酬余额及期后事项检查表J1-11-删除", True),
        ("职工薪酬政策统计表J1-12-删除", True),
        ("IPO企业股权激励工具关注的审计重点-删除", True),
        ("首发业务解答二-删除", True),
        # J 循环"原版"sheet 不应被过滤
        ("应付职工薪酬实质性程序表 J1A-原版", False),
        ("应付职工薪酬实质性程序表 L1A-原", False),
        # H 循环正常 sheet 不应被过滤
        ("审定表H1-1", False),
        ("折旧测算表（不含减值）-直线法H1-12", False),
        ("审定表（成本模式）H3-1", False),
        # I 循环正常 sheet 不应被过滤
        ("摊销测算表（不含减值）I1-10（剩余年限法）", False),
        ("摊销测算表（含减值）I1-11", False),
        # G 循环正常 sheet 不应被过滤
        ("审定表G1-1", False),
        ("长期股权投资审定表G7-1", False),
        # J 循环正常 sheet 不应被过滤
        ("审定表J1-1 ", False),
        ("明细表J1-2 ", False),
        ("月度分析表J1-4", False),
        ("计提情况检查表J1-6", False),
        ("分配情况检查表J1-7", False),
        ("明细表J2-2", False),
        ("股份支付检查表J3-2", False),
        ("IPO企业薪酬审计提示", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露信息(上市公司)", False),
        ("附注披露信息(国企)", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I/G/J 历史遗留过滤模式仍正确工作，正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
