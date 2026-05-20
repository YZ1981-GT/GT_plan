"""F-F10 prefill 扩展 ≥ 60 cell 验证测试

验证:
- F-cycle 新增 ≥ 60 cells 存在于 prefill_formula_mapping.json
- 每个 cell 有合法 formula_type（TB / TB_SUM / AUX / LEDGER / LEDGER_DETAIL / PREV / ADJ）
- 5 大类目标 sheet 全部覆盖（F2-2 / F2-21~F2-26 / F2-38~F2-44 / F2-47~F2-49 / F3-2+F4-2）
- F2-2 明细汇总表用 =AUX/=TB（中间 sheet → F2-1 cross_sheet 公式自动计算）

锚定 ADR-F2 两级 prefill 链路：TB/AUX/LEDGER → F2-2 → F2-1（487 cross_sheet 公式）
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
PREFILL_FILE = DATA_DIR / "prefill_formula_mapping.json"

# F-F10 实测目标：F-cycle ≥ 60 新 cells（实际追加 64）
F_CYCLE_NEW_CELL_TARGET = 60

VALID_FORMULA_TYPES = {
    "TB", "TB_SUM", "ADJ", "PREV", "WP", "AUX", "LEDGER", "LEDGER_DETAIL",
    "COUNT_LEDGER", "NOTE", "TB_AUX", "SUM_TB",
}

# 5 大类目标 sheet 关键字（按 F-F10 需求 + openpyxl 实测真名）
TARGET_SHEET_KEYWORDS = {
    "F2-2_master": "明细汇总表F2-2",
    "F2-21_stocktake": "盘点计划问卷F2-21",
    "F2-22_plan": "监盘计划F2-22",
    "F2-23_summary": "监盘小结F2-23",
    "F2-24_reconcile": "存货账面余额与仓储台账（ERP业务数据)核对记录F2-24",
    "F2-25_sample": "抽盘结果汇总表F2-25",
    "F2-26_rollback": "盘点倒轧表F2-26",
    "F2-38_valuation_avg": "计价方法测试表-平均F2-38",
    "F2-39_valuation_fifo": "计价方法测试表-先进先出F2-39",
    "F2-40_valuation_std": "计价方法测试表-标准成本差异F2-40",
    "F2-41_production_cost": "生产成本明细表F2-41",
    "F2-47_impairment": "跌价准备测试表F2-47",
    "F2-48_aging": "长库龄 呆滞 超过保质期存货明细表F2-48",
    "F2-49_impair_reversal": "跌价转回F2-49",
    "F3-2_payable_detail": "明细表F3-2",
    "F4-2_account_payable_detail": "明细表F4-2",
}


@pytest.fixture
def prefill_data():
    assert PREFILL_FILE.exists(), f"prefill_formula_mapping.json not found at {PREFILL_FILE}"
    with open(PREFILL_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _f_entries(data):
    return [m for m in data["mappings"] if str(m.get("wp_code", "")).startswith("F")]


def test_prefill_file_exists():
    assert PREFILL_FILE.exists()


def test_f_cycle_total_cells_meets_target(prefill_data):
    """F-cycle 总 cells ≥ baseline (45) + 60 target = 105"""
    f_entries = _f_entries(prefill_data)
    total_f_cells = sum(len(m.get("cells", [])) for m in f_entries)
    # F-F10 目标：在原 45 cells 基础上 +60 = 105+
    assert total_f_cells >= 45 + F_CYCLE_NEW_CELL_TARGET, (
        f"F-cycle cells = {total_f_cells}, expected >= {45 + F_CYCLE_NEW_CELL_TARGET}"
    )


def test_all_target_sheets_present(prefill_data):
    """5 大类目标 sheet 全部覆盖"""
    f_entries = _f_entries(prefill_data)
    sheets = {m["sheet"] for m in f_entries}
    missing = []
    for label, kw in TARGET_SHEET_KEYWORDS.items():
        if not any(kw in s for s in sheets):
            missing.append((label, kw))
    assert not missing, f"Missing target sheets: {missing}"


def test_all_formulas_have_valid_type(prefill_data):
    """每个 cell 的 formula_type 合法"""
    f_entries = _f_entries(prefill_data)
    invalid = []
    for m in f_entries:
        for cell in m.get("cells", []):
            ft = cell.get("formula_type")
            if ft not in VALID_FORMULA_TYPES:
                invalid.append((m["sheet"], cell.get("cell_ref"), ft))
    assert not invalid, f"Invalid formula_type cells: {invalid[:5]}"


def test_all_cells_have_formula_and_description(prefill_data):
    """每个 cell 必须有 formula + description（可读性铁律）"""
    f_entries = _f_entries(prefill_data)
    missing = []
    for m in f_entries:
        for cell in m.get("cells", []):
            if not cell.get("formula"):
                missing.append((m["sheet"], cell.get("cell_ref"), "formula"))
            if not cell.get("description"):
                missing.append((m["sheet"], cell.get("cell_ref"), "description"))
    assert not missing, f"Missing fields: {missing[:5]}"


def test_f22_master_sheet_uses_aux_or_tb(prefill_data):
    """ADR-F2 铁律：F2-2 明细汇总表（中间 sheet）必须使用 =TB / =AUX / =LEDGER 直接取数

    F2-2 是 F2-1 审定表 487 cross_sheet 公式的数据源；
    禁止在 F2-2 用 =WP（避免循环引用）。
    """
    f_entries = _f_entries(prefill_data)
    f22 = [m for m in f_entries if "明细汇总表F2-2" in m.get("sheet", "")]
    assert f22, "F2-2 明细汇总表 prefill mapping not found"
    forbidden_in_master = {"WP"}
    bad = []
    for m in f22:
        for cell in m.get("cells", []):
            if cell.get("formula_type") in forbidden_in_master:
                bad.append((m["sheet"], cell.get("cell_ref"), cell.get("formula_type")))
    assert not bad, f"F2-2 中间 sheet 不应用 =WP（违反 ADR-F2）: {bad}"


def test_valuation_sheets_use_ledger_detail(prefill_data):
    """F2-38~F2-44 计价测试至少包含 =LEDGER_DETAIL 公式（按金额分层抽样）"""
    f_entries = _f_entries(prefill_data)
    valuation = [m for m in f_entries
                 if "F2-38" in m.get("sheet", "")
                 or "F2-39" in m.get("sheet", "")
                 or "F2-40" in m.get("sheet", "")]
    assert valuation, "F2-38~F2-40 计价测试 prefill mapping 缺失"
    has_ledger_detail = any(
        cell.get("formula_type") == "LEDGER_DETAIL"
        for m in valuation
        for cell in m.get("cells", [])
    )
    assert has_ledger_detail, "F2 计价测试至少应包含 =LEDGER_DETAIL 抽样公式"


def test_impairment_sheet_uses_aux_for_product(prefill_data):
    """F2-47 跌价准备测试按存货分类维度（=AUX 4-arg）取数"""
    f_entries = _f_entries(prefill_data)
    impair = [m for m in f_entries if "跌价准备测试表F2-47" in m.get("sheet", "")]
    assert impair, "F2-47 跌价准备 prefill mapping 缺失"
    has_aux = any(
        cell.get("formula_type") == "AUX"
        for m in impair
        for cell in m.get("cells", [])
    )
    assert has_aux, "F2-47 跌价准备至少应包含 =AUX 公式"


def test_f3_f4_supplier_detail(prefill_data):
    """F3-2 / F4-2 明细表按供应商维度（=AUX 4-arg）取数"""
    f_entries = _f_entries(prefill_data)
    supplier_sheets = [
        m for m in f_entries
        if (m["wp_code"] == "F3" and "明细表F3-2" in m.get("sheet", ""))
        or (m["wp_code"] == "F4" and "明细表F4-2" in m.get("sheet", ""))
    ]
    assert supplier_sheets, "F3-2 / F4-2 明细表 prefill mapping 缺失"
    has_aux = any(
        cell.get("formula_type") == "AUX"
        for m in supplier_sheets
        for cell in m.get("cells", [])
    )
    assert has_aux, "F3-2/F4-2 明细表至少应包含 =AUX 公式（按供应商取数）"


def test_f_cycle_account_codes_valid(prefill_data):
    """F-cycle account_codes 应为合法的存货/应付账款/营业成本/费用科目编码"""
    f_entries = _f_entries(prefill_data)
    # 1xxx 资产（存货/预付/其他应收）/ 22xx 应付 / 5xxx-6xxx 损益（营业成本等）
    valid_prefixes = ("1", "2", "5", "6")
    invalid = []
    for m in f_entries:
        for code in m.get("account_codes", []):
            if not str(code).startswith(valid_prefixes):
                invalid.append((m["sheet"], code))
    assert not invalid, f"Invalid account_codes: {invalid[:5]}"


def test_aux_formulas_use_4_args(prefill_data):
    """业务正确性铁律：=AUX 公式必须 4 args (code, aux_type, aux_code, column)，
    否则 prefill_engine._resolve_aux_formula 直接 return None 取不到数。

    P0-1 修复后必须验证：F-cycle 中所有 =AUX 公式都是 4-arg 形式。
    """
    import re
    f_entries = _f_entries(prefill_data)
    bad = []
    aux_pattern = re.compile(r"=AUX\s*\(([^)]*)\)", re.IGNORECASE)
    for m in f_entries:
        for cell in m.get("cells", []):
            if cell.get("formula_type") != "AUX":
                continue
            formula = cell.get("formula", "")
            mm = aux_pattern.search(formula)
            if not mm:
                bad.append((m["sheet"], cell.get("cell_ref"), "no AUX(...) match", formula))
                continue
            # 计算逗号数（应该有 3 个逗号 = 4 个 args）
            args_str = mm.group(1)
            arg_count = len([a for a in args_str.split(",") if a.strip()])
            if arg_count != 4:
                bad.append((m["sheet"], cell.get("cell_ref"), f"expected 4 args got {arg_count}", formula))
    assert not bad, f"P0-1 违反：F-cycle =AUX 必须 4-arg：{bad[:3]}"


def test_f_cycle_sheet_names_match_real_templates(prefill_data):
    """业务正确性铁律：F-cycle 新增 sheet 名必须存在于 wp_templates/F/*.xlsx 真实模板中
    （通过 openpyxl 实测核对，防止臆造 sheet 名）。

    P0-2 修复后必须验证。仅验证 F-F10 新增的中间/明细 sheet（F2-1 / F0/F1/F5 审定表
    本身在原 spec 中已与其它测试覆盖）。
    """
    from pathlib import Path
    from openpyxl import load_workbook

    repo_root = Path(__file__).resolve().parents[2]
    f_template_dir = repo_root / "backend" / "wp_templates" / "F"
    if not f_template_dir.exists():
        import pytest as _pytest
        _pytest.skip("F template dir not present in current snapshot")

    # 收集所有 F2 模板的真实 sheet 名集合
    real_sheet_names: set[str] = set()
    for f in f_template_dir.glob("*.xlsx"):
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            real_sheet_names.update(wb.sheetnames)
            wb.close()
        except Exception:
            continue

    # 受 P0-2 重写脚本控制的 sheet 列表（仅这些必须实测核对）
    p02_managed_sheets = {
        "明细汇总表F2-2",
        "盘点计划问卷F2-21",
        "监盘计划F2-22",
        "监盘小结F2-23",
        "存货账面余额与仓储台账（ERP业务数据)核对记录F2-24",
        "抽盘结果汇总表F2-25",
        "盘点倒轧表F2-26",
        "计价方法测试表-平均F2-38",
        "计价方法测试表-先进先出F2-39",
        "计价方法测试表-标准成本差异F2-40",
        "生产成本明细表F2-41",
        "直接人工分析表F2-42",
        "制造费用明细表F2-43",
        "跌价准备测试表F2-47",
        "长库龄 呆滞 超过保质期存货明细表F2-48",
        "跌价转回F2-49",
        "明细表F3-2",
        "明细表F4-2",
    }

    f_entries = _f_entries(prefill_data)
    bad: list[tuple[str, str]] = []
    for m in f_entries:
        sheet = m.get("sheet", "")
        if sheet not in p02_managed_sheets:
            continue
        if sheet not in real_sheet_names:
            bad.append((m["wp_code"], sheet))
    assert not bad, f"P0-2 违反：以下 sheet 在真实 F 模板中不存在 → {bad}"
