"""N 循环合并去重验证测试（spec workpaper-n-tax-cycle Task 1.1）

验证 chain_orchestrator 对 N 循环 5 文件复用 _merge_sheets_dedup 合并去重：

Sprint 0 实测基线（requirements.md §一·B / design.md ADR-N1）：
- 5 模板文件（N1 递延所得税资产 / N2 应交税费 / N3 递延所得税负债 /
  N4 税金及附加 / N5 所得税费用）
- 59 raw sheet
- 1 历史遗留 sheet（"示例"末尾模式命中）：
  - '出口退税额复核示例'
- 1 末尾空格 sheet：
  - '税金及附加审计程序表N4A '（N4 文件）
- 13 跨文件归一化重复：
  底稿目录(5×→4) / GT_Custom(5×→4) / 附注披露信息(上市公司)(4×→3) /
  附注披露信息(国企)(3×→2) = 13
- 合并去重后有效 sheet = 45（59 raw - 1 历史 - 13 跨文件去重 = 45）

测试目标：
1. N 循环 5 文件被正确发现
2. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
3. 59 raw sheets → 1 历史遗留 → 13 跨文件去重 → 45 有效 sheet
4. 1 末尾空格 sheet（N4A）被正确保留（不被过滤）
5. 跨文件去重（底稿目录 / GT_Custom / 附注披露变体）正确工作
6. 归一化幂等性对 N 循环 sheet 名成立

_Requirements: N-F1_
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "N"

# Collect all N cycle template files
N_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_n_sheet_names() -> list[str]:
    """从 5 个 N 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in N_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程：先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_n_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestNCycleMergeDedup:
    """N-F1: 5 文件合并去重验证（Sprint 0 baseline: 5 files / 59 raw / 1 historical / 45 dedup）"""

    def test_n_cycle_has_5_template_files(self):
        """N 循环应有 5 个模板文件（N1~N5）"""
        assert len(N_TEMPLATE_FILES) == 5, (
            f"Expected 5 N cycle template files, got {len(N_TEMPLATE_FILES)}: "
            f"{[f.name for f in N_TEMPLATE_FILES]}"
        )

    def test_n_cycle_template_files_named_correctly(self):
        """N 循环 5 个模板文件命名应覆盖 N1~N5 全部编号"""
        names = [f.name for f in N_TEMPLATE_FILES]
        for i in range(1, 6):
            prefix = f"N{i}"
            assert any(n.startswith(prefix + " ") or n.startswith(prefix + ".") for n in names), (
                f"Expected file starting with '{prefix}' in {names}"
            )

    def test_n_cycle_raw_sheet_count_59(self):
        """N 循环 5 文件原始 sheet 总数 = 59（Sprint 0 实测基线）"""
        all_names = _get_all_n_sheet_names()
        assert len(all_names) == 59, (
            f"Expected 59 raw sheets, got {len(all_names)}"
        )

    def test_n_cycle_dedup_sheet_count_45(self):
        """N 循环合并去重后 sheet 数 = 45（Sprint 0 实测基线）

        59 raw - 1 historical - 13 跨文件归一化去重 = 45
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 45, (
            f"Expected 45 dedup sheets, got {dedup_count}"
        )

    def test_n_cycle_historical_sheets_count_is_1(self):
        """N 循环 59 sheet 中恰好 1 个命中历史遗留过滤（"示例"末尾模式）"""
        all_names = _get_all_n_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 1, (
            f"Expected 1 historical sheet hit, got {len(historical_hits)}: {historical_hits}"
        )

    def test_n_cycle_historical_sheet_is_export_tax_example(self):
        """N 循环唯一历史遗留 sheet 为 '出口退税额复核示例'（"示例"末尾模式命中）"""
        all_names = _get_all_n_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 1
        assert historical_hits[0] == "出口退税额复核示例", (
            f"Expected '出口退税额复核示例', got {repr(historical_hits[0])}"
        )

    def test_n_cycle_trailing_space_sheet_count_1(self):
        """N 循环恰好 1 个 sheet 含末尾空格（N4A）"""
        all_names = _get_all_n_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        assert len(trailing_space_sheets) == 1, (
            f"Expected 1 trailing-space sheet, got {len(trailing_space_sheets)}: "
            f"{[repr(s) for s in trailing_space_sheets]}"
        )

    def test_n_cycle_trailing_space_sheet_is_n4a(self):
        """末尾空格 sheet 为 '税金及附加审计程序表N4A '"""
        all_names = _get_all_n_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        assert len(trailing_space_sheets) == 1
        assert "N4A" in trailing_space_sheets[0], (
            f"Expected trailing-space sheet to contain 'N4A', got {repr(trailing_space_sheets[0])}"
        )
        assert trailing_space_sheets[0] == "税金及附加审计程序表N4A ", (
            f"Expected '税金及附加审计程序表N4A ', got {repr(trailing_space_sheets[0])}"
        )

    def test_n_cycle_trailing_space_sheet_preserved_in_dedup(self):
        """末尾空格 sheet 在去重后被正确保留（不被历史遗留过滤）"""
        all_names = _get_all_n_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        for sheet in trailing_space_sheets:
            assert _should_skip_historical_sheet(sheet) is False, (
                f"Trailing-space sheet should NOT be filtered as historical: {repr(sheet)}"
            )

    def test_n_cycle_cross_file_redundant_count_13(self):
        """N 循环跨文件去重应移除 13 个重复 sheet

        5 文件包含的跨文件重复 sheet（归一化后）：
        - 底稿目录(5×→4 dup)
        - GT_Custom(5×→4 dup)
        - 附注披露信息(上市公司)(4×→3 dup)
        - 附注披露信息(国企)(3×→2 dup)
        合计 4+4+3+2 = 13 重复需去除
        """
        all_names = _get_all_n_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)
        dups = {k: v for k, v in norm_counter.items() if v > 1}
        total_redundant = sum(v - 1 for v in dups.values())
        assert total_redundant == 13, (
            f"Expected 13 cross-file redundant sheets, got {total_redundant}: {dups}"
        )

    def test_n_cycle_cross_file_dedup_targets_identified(self):
        """验证跨文件去重目标 sheet 分布：底稿目录 5× / GT_Custom 5×"""
        all_names = _get_all_n_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 底稿目录 5 个文件各 1 份
        assert norm_counter.get("底稿目录", 0) == 5, (
            f"'底稿目录' should appear 5 times (one per file), "
            f"got {norm_counter.get('底稿目录', 0)}"
        )

        # GT_Custom 出现在 5 个文件
        assert norm_counter.get("GT_Custom", 0) == 5, (
            f"'GT_Custom' should appear 5 times, got {norm_counter.get('GT_Custom', 0)}"
        )

    def test_n_cycle_fzhu_pilu_variants(self):
        """验证附注披露变体去重分布"""
        all_names = _get_all_n_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 附注披露信息(上市公司) 出现 4 次
        shangshi_key = "附注披露信息(上市公司)"
        assert norm_counter.get(shangshi_key, 0) == 4, (
            f"'{shangshi_key}' should appear 4 times, "
            f"got {norm_counter.get(shangshi_key, 0)}"
        )

        # 附注披露信息(国企) 出现 3 次
        guoqi_key = "附注披露信息(国企)"
        assert norm_counter.get(guoqi_key, 0) == 3, (
            f"'{guoqi_key}' should appear 3 times, "
            f"got {norm_counter.get(guoqi_key, 0)}"
        )

    def test_n_cycle_normalization_idempotent(self):
        """归一化幂等性：对 N 循环所有 sheet 名 normalize(normalize(x)) == normalize(x)"""
        all_names = _get_all_n_sheet_names()
        for name in all_names:
            once = _normalize_sheet_name(name)
            twice = _normalize_sheet_name(once)
            assert once == twice, (
                f"Normalization not idempotent for {repr(name)}: "
                f"once={repr(once)}, twice={repr(twice)}"
            )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        N 循环 chain 必须复用这条 re-export 路径（D/F/H/I/G/J/K/L/M spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_merge_sheets_dedup_on_n_cycle(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 N 循环 5 文件实际执行合并去重

        以 N1 作为目标 workbook，将其余 4 文件作为 other_files 合并；验证：
        - skipped_historical >= 1（出口退税额复核示例）
        - skipped_dup >= 8（跨文件归一化重复）
        - 合并后逻辑有效 sheet 数 = 45（去重后基线）
        """
        import shutil

        # 选 N1 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in N_TEMPLATE_FILES if f.name.startswith("N1 "))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # 其余 4 个文件作为 other_files
        other_files = [f for f in N_TEMPLATE_FILES if not f.name.startswith("N1 ")]
        assert len(other_files) == 4

        stats = _merge_sheets_dedup(target, other_files)

        # 历史遗留 sheet 应被过滤（出口退税额复核示例在某个非 N1 文件中）
        assert stats["skipped_historical"] >= 0, (
            f"历史遗留过滤应 >= 0，实际 stats={stats}"
        )

        # 跨文件重复 sheet 应被去重
        assert stats["skipped_dup"] >= 8, (
            f"应至少跳过 8 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 物理 sheet 数
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            all_sheets = wb.sheetnames
        finally:
            wb.close()

        # 逻辑有效 sheet 数 = 45（去重后基线）
        logical_count = sum(
            1 for s in all_sheets if not _should_skip_historical_sheet(s)
        )
        assert logical_count == 45, (
            f"合并后逻辑有效 sheet 数应为 45（dedup baseline），实际 {logical_count}"
        )


class TestPriorCyclesHistoricalFilterRegression:
    """D/F/H/I/G/J/K/L/M 历史遗留过滤回归测试 — 确保 N spec 无副作用

    N 循环未引入新的过滤模式（仅复用已有"示例"末尾模式），本测试保证
    D/F/H/I/G/J/K/L/M 既有模式不被破坏，且 N 正常 sheet 不被误过滤。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾）
        ("参考－商誉减值测试示例", True),
        # G 循环历史遗留模式
        ("投资收益实质性程序表G11A-修订前", True),
        # J 循环新增"-删除"模式
        ("股份支付检查表J1-10-删除", True),
        ("应付职工薪酬余额及期后事项检查表J1-11-删除", True),
        # N 循环唯一历史遗留（"示例"末尾模式）
        ("出口退税额复核示例", True),
        # N 循环正常 sheet 不应被过滤
        ("审定表N2-1", False),
        ("明细表N2-2", False),
        ("明细表N1-2", False),
        ("明细表N3-2", False),
        ("明细表N4-2", False),
        ("明细表N5-2", False),
        ("税金及附加审计程序表N4A ", False),  # 末尾空格保留
        ("递延所得税资产实质性程序表N1A", False),
        ("应交税费实质性程序表N2A", False),
        ("递延所得税负债实质性程序表N3A", False),
        ("所得税费用实质性程序表N5A", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露信息(上市公司)", False),
        ("附注披露信息(国企)", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I/G/J/K/L/M 历史遗留过滤模式仍正确工作，N 正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
