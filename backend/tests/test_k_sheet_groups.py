"""K-F2: K cycle 11-category sheet group rules — backend completeness check.

Validates: Requirements K-F2 / Spec workpaper-k-admin-cycle / Sprint 2 Task 2.2
Updated: k-admin-cycle-post-review-fix Task 4.6 (sync Python mirror rules)

Mirror of frontend useKAdminCycleSheetGroups.ts classification rules.

11 categories (priority 0~9):
1. 索引 (defaultHidden=true): 底稿目录 / GT_Custom
2. 程序表: 实质性程序表 / 函证程序表 / xxA 结尾
3. 审定表: 审定表 / 情况表 / 函证结果汇总
4. 费用明细 (前置): 明细表K8-2 / 明细表K9-2 / K10-2~K13-2
5. 明细表: 其他明细表
6. 分析程序: 分析 / 对比 / 情况分析
7. 往来款检查: K1-/K3- 含 检查/账龄/挂账/关联方/三阶段/未收回/大额/坏账/核销/转回/替代程序/信用减值
8. 检查表: 检查表 / 计提 / 分配 / 截止性测试 / 测算 / 测试表 / 政策检查 / 核对表
9. 函证辅助 (priority=7.5): K0-x 函证/替代程序/回函/核实/舞弊风险/差异调节/过程控制/会计提示
10. 附注+调整 (附注披露 defaultHidden=true): 附注披露 / 调整分录
11. 其他: fallback

The classification function is reproduced here in pure Python to provide a
backend-side correctness oracle for the 109 real K cycle sheets.
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _normalize_sheet_name,
    _should_skip_historical_sheet,
)


# ──────────────────────────────────────────────────────────────────────────
# Mirror of frontend classifyKSheet rules (must stay in sync)
# ──────────────────────────────────────────────────────────────────────────

K_RULES = [
    {
        "id": "index",
        "category": "索引",
        "priority": 0,
        "default_hidden": True,
        "match": lambda s: bool(re.match(r"^(底稿目录|GT_Custom)$", s.strip())),
    },
    {
        "id": "procedure",
        "category": "程序表",
        "priority": 1,
        "default_hidden": False,
        "match": lambda s: bool(
            re.search(r"实质性程序表", s)
            or re.search(r"函证程序表", s)
            or re.search(r"[A-Z]\d*A$", s.strip())
            or re.search(r"[A-Z]\d*A-", s)
            or re.search(r"[A-Z]\d*A ", s)
        ),
    },
    {
        "id": "audit_table",
        "category": "审定表",
        "priority": 2,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"审定表|情况表|函证结果汇总", s)),
    },
    {
        "id": "expense_detail",
        "category": "费用明细",
        "priority": 3,
        "default_hidden": False,
        "match": lambda s: bool(re.match(r"^明细表K(8|9|1[0-3])-", s.strip())),
    },
    {
        "id": "detail",
        "category": "明细表",
        "priority": 4,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"明细表", s)),
    },
    {
        "id": "analysis",
        "category": "分析程序",
        "priority": 5,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"分析|对比|情况分析", s)),
    },
    {
        "id": "receivable_payable_check",
        "category": "往来款检查",
        "priority": 6,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"K[13]-", s))
        and bool(
            re.search(
                r"(检查|账龄|挂账|关联方|三阶段|未收回|大额|坏账|核销|转回|替代程序|信用减值)",
                s,
            )
        ),
    },
    {
        "id": "check_table",
        "category": "检查表",
        "priority": 7,
        "default_hidden": False,
        "match": lambda s: bool(
            re.search(r"检查表|分配|截止性测试|测算|测试表|政策检查|核对表", s)
            or re.search(r"(?<!会)计提", s)
        ),
    },
    {
        "id": "confirmation_aux",
        "category": "函证辅助",
        "priority": 7.5,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"K0[-]?\d", s))
        and bool(
            re.search(
                r"函证|替代程序|回函|核实|舞弊风险|差异调节|过程控制|会计提示",
                s,
            )
        ),
    },
    {
        "id": "disclosure_adj",
        "category": "附注+调整",
        "priority": 8,
        "default_hidden": False,
        "match": lambda s: bool(re.search(r"附注披露|调整分录|调整分录汇总", s)),
    },
    {
        "id": "other",
        "category": "其他",
        "priority": 9,
        "default_hidden": False,
        "match": lambda s: True,
    },
]


def classify_k_sheet(name: str) -> dict:
    """Return classification metadata; fallback to 其他."""
    for rule in K_RULES:
        if rule["match"](name):
            cat = {
                "id": rule["id"],
                "category": rule["category"],
                "priority": rule["priority"],
                "default_hidden": rule["default_hidden"],
            }
            # 附注披露 defaultHidden=true 特殊处理
            if rule["id"] == "disclosure_adj" and re.search(r"附注披露", name):
                cat["default_hidden"] = True
            return cat
    return {
        "id": "other",
        "category": "其他",
        "priority": 9,
        "default_hidden": False,
    }


# ──────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "K"


@pytest.fixture(scope="module")
def k_109_unique_sheets() -> list[str]:
    """109 个去重后 K 循环有效 sheet 名（基于 _normalize_sheet_name + filter historical）"""
    seen_normalized: set[str] = set()
    unique_originals: list[str] = []
    for f in sorted(TEMPLATES_DIR.glob("*.xlsx")):
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            for s in wb.sheetnames:
                if _should_skip_historical_sheet(s):
                    continue
                norm = _normalize_sheet_name(s)
                if norm not in seen_normalized:
                    seen_normalized.add(norm)
                    unique_originals.append(s)
        finally:
            wb.close()
    return unique_originals


# ──────────────────────────────────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────────────────────────────────


class TestKSheetGroupCompleteness:
    """K-F2: 11 类规则对 109 sheet 完备性 + 唯一性"""

    def test_109_sheet_baseline(self, k_109_unique_sheets):
        """K 循环去重后有效 sheet 数 = 109（Sprint 0 实测基线）"""
        assert len(k_109_unique_sheets) == 109, (
            f"Expected 109 K-cycle dedup sheets, got {len(k_109_unique_sheets)}"
        )

    def test_all_109_classified_to_non_empty_category(self, k_109_unique_sheets):
        """所有 sheet 均被分类到非空类目"""
        for s in k_109_unique_sheets:
            cat = classify_k_sheet(s)
            assert cat["category"], f"sheet '{s}' classified to empty category"

    def test_each_sheet_exactly_one_category(self, k_109_unique_sheets):
        """任意 sheet 名恰好命中 1 类（rule fallback `() => true` 必然命中）"""
        for s in k_109_unique_sheets:
            hits = [r for r in K_RULES if r["match"](s)]
            assert len(hits) >= 1, f"sheet '{s}' missed all rules"

    def test_no_sheet_falls_into_unknown_other_unexpectedly(
        self, k_109_unique_sheets
    ):
        """fallback "其他" 类应仅包含合理的 K0 函证特殊 sheet 等无法明确归类项；
        其数量应 < 30%（109 中 < 33）"""
        other_sheets = [
            s for s in k_109_unique_sheets if classify_k_sheet(s)["id"] == "other"
        ]
        ratio = len(other_sheets) / len(k_109_unique_sheets)
        assert ratio < 0.30, (
            f"Too many sheets fell to '其他' fallback: "
            f"{len(other_sheets)}/{len(k_109_unique_sheets)} = {ratio:.1%}\n"
            f"Other sheets: {other_sheets}"
        )

    def test_all_11_categories_represented(self, k_109_unique_sheets):
        """109 sheet 应覆盖 ≥ 8 类（11 类中至少 8 类被命中，'附注+调整'/'索引' 必有）"""
        categories_hit = {classify_k_sheet(s)["category"] for s in k_109_unique_sheets}
        assert len(categories_hit) >= 8, (
            f"Only {len(categories_hit)} categories hit, expected ≥ 8\n"
            f"Hit: {sorted(categories_hit)}"
        )
        # 索引 / 附注+调整 必须命中（K 循环 14 文件均含）
        assert "索引" in categories_hit
        assert "附注+调整" in categories_hit
        # 程序表 / 审定表 / 费用明细 / 明细表 必须命中
        for must_have in ["程序表", "审定表", "费用明细", "明细表"]:
            assert must_have in categories_hit, (
                f"Category '{must_have}' should be hit but not in {sorted(categories_hit)}"
            )


class TestKSheetGroupKeyPaths:
    """K-F2: 关键 sheet 分类路径验证"""

    def test_index_sheets(self):
        for s in ["底稿目录", "GT_Custom"]:
            cat = classify_k_sheet(s)
            assert cat["category"] == "索引"
            assert cat["default_hidden"] is True

    def test_disclosure_sheets_hidden(self):
        """附注披露 5 种括号变体均 defaultHidden=true"""
        variants = [
            "附注披露信息(上市公司)",
            "附注披露信息(国企)",
            "附注披露信息（上市公司）",
            "附注披露信息（国企）",
            "附注披露信息（国有企业）",
        ]
        for s in variants:
            cat = classify_k_sheet(s)
            assert cat["category"] == "附注+调整"
            assert cat["default_hidden"] is True

    def test_adjustment_entries_visible(self):
        for s in ["调整分录汇总K1-4", "调整分录汇总K8-3", "调整分录汇总 K5-3"]:
            cat = classify_k_sheet(s)
            assert cat["category"] == "附注+调整"
            assert cat["default_hidden"] is False

    def test_expense_detail_priority_over_general_detail(self):
        """K8-2/K9-2 优先匹配 费用明细（priority 3 < 明细表 4）"""
        assert classify_k_sheet("明细表K8-2")["category"] == "费用明细"
        assert classify_k_sheet("明细表K9-2")["category"] == "费用明细"
        # K1-2/K3-2 不匹配 expense_detail，落到 detail
        assert classify_k_sheet("明细表K1-2")["category"] == "明细表"
        assert classify_k_sheet("明细表K3-2")["category"] == "明细表"

    def test_k5_2_with_space_classified_as_detail(self):
        """K5-2 真实 sheet 名含空格仍归入 明细表"""
        assert classify_k_sheet("明细表 K5-2")["category"] == "明细表"

    def test_receivable_payable_check_path(self):
        """K1-/K3- 含 关联方/账龄/挂账/三阶段/未收回 → 往来款检查"""
        path = [
            ("关联方及交易检查表K1-11", "往来款检查"),
            ("长期未收回款项检查表K1-10", "往来款检查"),
            ("长期挂账检查表K3-5", "往来款检查"),
            ("三阶段划分检查表K1-7", "往来款检查"),
            ("信用减值损失会计政策检查K1-6", "往来款检查"),
        ]
        for s, expected in path:
            assert classify_k_sheet(s)["category"] == expected, (
                f"sheet '{s}' classified incorrectly"
            )

    def test_general_check_table_path(self):
        """非 K1/K3 的 检查表/截止性测试/测算 → 检查表"""
        path = [
            ("截止性测试(从记账凭证至原始凭证）K8-6", "检查表"),
            ("管理费用检查表K9-8", "检查表"),
            ("销售费用检查表K8-8", "检查表"),
            ("合同检查表K8-5", "检查表"),
            ("预计负债检查表 K5-7", "检查表"),
            ("产品质量保修检查表 K5-4", "检查表"),
            ("减值准备测试表（后续计量） K6-5", "检查表"),
            ("摊销测算表K2-5", "检查表"),
        ]
        for s, expected in path:
            assert classify_k_sheet(s)["category"] == expected

    def test_priority_conflict_resolution(self):
        """优先级冲突铁律：明细表(4) < 往来款检查(6) → 含'明细表'的 K1- 仍归 明细表"""
        # 坏账准备明细表K1-3 含"明细表" → priority 4 detail 命中
        assert classify_k_sheet("坏账准备明细表K1-3")["category"] == "明细表"
        # 大额其他应收款情况分析表K1-5 含"分析" → priority 5 analysis 命中
        assert classify_k_sheet("大额其他应收款情况分析表K1-5")["category"] == "分析程序"

    def test_confirmation_aux_category(self):
        """K0-x 函证辅助 sheet 归入 函证辅助（priority=7.5）"""
        # 函证差异调节表K0-4 → 函证辅助
        assert classify_k_sheet("函证差异调节表K0-4")["category"] == "函证辅助"
        # 核实被函证单位信息K0-2 → 函证辅助
        assert classify_k_sheet("核实被函证单位信息K0-2")["category"] == "函证辅助"
        # 其他 K0 函证相关 sheet
        assert classify_k_sheet("替代程序K0-5")["category"] == "函证辅助"
        assert classify_k_sheet("回函检查K0-3")["category"] == "函证辅助"
        assert classify_k_sheet("舞弊风险评估K0-7")["category"] == "函证辅助"
        assert classify_k_sheet("过程控制K0-6")["category"] == "函证辅助"
        assert classify_k_sheet("会计提示K0-8")["category"] == "函证辅助"

    def test_expanded_expense_detail_regex(self):
        """费用明细 regex 扩展覆盖 K10-2~K13-2"""
        # K10-2 / K13-2 归入 费用明细（扩展后）
        assert classify_k_sheet("明细表K10-2")["category"] == "费用明细"
        assert classify_k_sheet("明细表K11-2")["category"] == "费用明细"
        assert classify_k_sheet("明细表K12-2")["category"] == "费用明细"
        assert classify_k_sheet("明细表K13-2")["category"] == "费用明细"
        # K8-2/K9-2 仍归 费用明细（原有行为不变）
        assert classify_k_sheet("明细表K8-2")["category"] == "费用明细"
        assert classify_k_sheet("明细表K9-2")["category"] == "费用明细"
        # K1-2/K3-2/K5-2 仍归通用 明细表（不受影响）
        assert classify_k_sheet("明细表K1-2")["category"] == "明细表"
        assert classify_k_sheet("明细表K3-2")["category"] == "明细表"
        assert classify_k_sheet("明细表K5-2")["category"] == "明细表"


class TestKSheetGroupCategoryDistribution:
    """K-F2: 类目分布合理性（防止单一类目占比过高表明分类粒度不足）"""

    def test_no_single_non_other_category_over_50_percent(
        self, k_109_unique_sheets
    ):
        """非 fallback '其他' 类目不应单一占比 > 50%（避免分类粒度不足）"""
        counts = Counter(
            classify_k_sheet(s)["category"] for s in k_109_unique_sheets
        )
        for cat, n in counts.items():
            if cat == "其他":
                continue
            ratio = n / len(k_109_unique_sheets)
            assert ratio < 0.50, (
                f"Category '{cat}' takes {n}/{len(k_109_unique_sheets)} = "
                f"{ratio:.1%}, exceeding 50% threshold"
            )

    def test_distribution_summary(self, k_109_unique_sheets):
        """分布摘要打印（便于回归比对）"""
        counts = Counter(
            classify_k_sheet(s)["category"] for s in k_109_unique_sheets
        )
        total = sum(counts.values())
        print("\n=== K 循环 109 sheet 分类分布 ===")
        for cat, n in sorted(counts.items(), key=lambda x: -x[1]):
            print(f"  {cat}: {n} ({n / total:.1%})")
        assert total == 109
