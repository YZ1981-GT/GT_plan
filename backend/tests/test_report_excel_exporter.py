"""Tests for ReportExcelExporter

Covers:
- Export with data (multiple report types)
- Export empty report (no data rows)
- Format verification (amount format, total row style, indentation)
- Parameter handling (mode, report_types, include_prior_year)
- Endpoint parameter validation
- Filename sanitization
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

from app.models.base import Base
from app.models.core import Project
from app.models.report_models import FinancialReport, FinancialReportType
from app.services.report_excel_exporter import (
    AMOUNT_FORMAT,
    NEGATIVE_FORMAT,
    REPORT_TYPE_SHEET_NAMES,
    ReportExcelExporter,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_engine = create_async_engine("sqlite+aiosqlite:///:memory:")
_async_session = sessionmaker(_engine, class_=AsyncSession, expire_on_commit=False)


@pytest_asyncio.fixture
async def db_session():
    """Create in-memory SQLite session with required tables."""
    async with _engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    async with _async_session() as session:
        yield session
    async with _engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)


@pytest_asyncio.fixture
async def sample_project(db_session: AsyncSession):
    """Create a sample project."""
    project = Project(
        id=uuid.uuid4(),
        name="测试科技有限公司",
        client_name="测试科技有限公司",
        template_type="soe",
        report_scope="standalone",
    )
    db_session.add(project)
    await db_session.flush()
    return project


@pytest_asyncio.fixture
async def sample_report_data(db_session: AsyncSession, sample_project: Project):
    """Create sample financial report rows."""
    project_id = sample_project.id
    year = 2024
    now = datetime.now(timezone.utc)

    rows = [
        FinancialReport(
            id=uuid.uuid4(),
            project_id=project_id,
            year=year,
            report_type=FinancialReportType.balance_sheet,
            row_code="BS-001",
            row_name="流动资产：",
            current_period_amount=None,
            prior_period_amount=None,
            indent_level=0,
            is_total_row=False,
            generated_at=now,
        ),
        FinancialReport(
            id=uuid.uuid4(),
            project_id=project_id,
            year=year,
            report_type=FinancialReportType.balance_sheet,
            row_code="BS-002",
            row_name="货币资金",
            current_period_amount=Decimal("1234567.89"),
            prior_period_amount=Decimal("987654.32"),
            indent_level=1,
            is_total_row=False,
            generated_at=now,
        ),
        FinancialReport(
            id=uuid.uuid4(),
            project_id=project_id,
            year=year,
            report_type=FinancialReportType.balance_sheet,
            row_code="BS-003",
            row_name="应收账款",
            current_period_amount=Decimal("-500000.00"),
            prior_period_amount=Decimal("300000.00"),
            indent_level=1,
            is_total_row=False,
            generated_at=now,
        ),
        FinancialReport(
            id=uuid.uuid4(),
            project_id=project_id,
            year=year,
            report_type=FinancialReportType.balance_sheet,
            row_code="BS-010",
            row_name="流动资产合计",
            current_period_amount=Decimal("734567.89"),
            prior_period_amount=Decimal("1287654.32"),
            indent_level=0,
            is_total_row=True,
            generated_at=now,
        ),
        # Income statement
        FinancialReport(
            id=uuid.uuid4(),
            project_id=project_id,
            year=year,
            report_type=FinancialReportType.income_statement,
            row_code="IS-001",
            row_name="营业收入",
            current_period_amount=Decimal("5000000.00"),
            prior_period_amount=Decimal("4500000.00"),
            indent_level=0,
            is_total_row=False,
            generated_at=now,
        ),
    ]

    for r in rows:
        db_session.add(r)
    await db_session.flush()
    return rows


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_with_data(db_session: AsyncSession, sample_project, sample_report_data):
    """Test export generates valid xlsx with data."""
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        include_prior_year=True,
    )

    assert isinstance(output, BytesIO)
    assert output.getvalue()  # Non-empty

    # Load and verify
    wb = load_workbook(output)
    assert len(wb.sheetnames) >= 1

    # Find balance sheet
    bs_sheet = None
    for name in wb.sheetnames:
        if "资产负债" in name:
            bs_sheet = wb[name]
            break

    assert bs_sheet is not None
    # Verify data was written (check that some cells have values)
    has_data = False
    for row in bs_sheet.iter_rows(min_row=4, max_row=bs_sheet.max_row):
        for cell in row:
            if cell.value is not None and cell.value != "":
                has_data = True
                break
        if has_data:
            break
    assert has_data


@pytest.mark.asyncio
async def test_export_empty_report(db_session: AsyncSession, sample_project):
    """Test export with no report data generates valid xlsx (empty sheets)."""
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
    )

    assert isinstance(output, BytesIO)
    # Should still produce a valid workbook
    wb = load_workbook(output)
    assert len(wb.sheetnames) >= 1


@pytest.mark.asyncio
async def test_template_preserves_number_format(db_session: AsyncSession, sample_project, sample_report_data):
    """Template-driven fill preserves the template's own number format on filled cells.

    Task 20: writing only ``.value`` keeps the template style. BS-002 maps to C6.
    """
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
        include_prior_year=True,
    )

    wb = load_workbook(output)
    ws = None
    for name in wb.sheetnames:
        if "资产负债" in name:
            ws = wb[name]
            break

    assert ws is not None

    # BS-002 current value lands at C6 (real placeholder coordinate, not row order)
    c6 = ws["C6"]
    assert isinstance(c6.value, (int, float))
    # Template's own number format preserved (POC uses '#,##0.00_ ')
    assert c6.number_format != "General"


@pytest.mark.asyncio
async def test_formula_rows_not_overwritten(db_session: AsyncSession, sample_project, sample_report_data):
    """Total/SUM rows in the template must NOT be overwritten by the fill (Task 20.4)."""
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
    )

    wb = load_workbook(output)
    ws = None
    for name in wb.sheetnames:
        if "资产负债" in name:
            ws = wb[name]
            break

    assert ws is not None

    # POC SUM/total rows must remain formulas (not clobbered)
    for coord in ("C31", "D31", "C57", "D57"):
        val = ws[coord].value
        assert isinstance(val, str) and val.startswith("="), (
            f"{coord} formula was overwritten: {val!r}"
        )


@pytest.mark.asyncio
async def test_parameter_mode_unadjusted(db_session: AsyncSession, sample_project, sample_report_data):
    """Test export with mode=unadjusted produces valid output."""
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="unadjusted",
    )

    assert isinstance(output, BytesIO)
    wb = load_workbook(output)
    # Should produce a valid workbook regardless of mode
    assert len(wb.sheetnames) >= 1
    # Check that the workbook has content (mode affects data source, not structure)
    ws = wb[wb.sheetnames[0]]
    assert ws.max_row >= 4  # At least headers


@pytest.mark.asyncio
async def test_parameter_report_types_filter(db_session: AsyncSession, sample_project, sample_report_data):
    """Test export with specific report_types only exports those sheets."""
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        report_types=["income_statement"],
    )

    wb = load_workbook(output)
    # Should only have income statement sheet
    assert len(wb.sheetnames) == 1
    assert "利润" in wb.sheetnames[0]


@pytest.mark.asyncio
async def test_parameter_no_prior_year(db_session: AsyncSession, sample_project, sample_report_data):
    """Test export without prior year column has fewer data columns."""
    exporter = ReportExcelExporter(db_session)

    # Export WITH prior year
    output_with = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
        include_prior_year=True,
    )

    # Export WITHOUT prior year
    output_without = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
        include_prior_year=False,
    )

    wb_with = load_workbook(output_with)
    wb_without = load_workbook(output_without)

    ws_with = wb_with[wb_with.sheetnames[0]]
    ws_without = wb_without[wb_without.sheetnames[0]]

    # Template puts prior-period values in column D. Without prior year, the D
    # placeholders are cleared, so column D should have fewer numeric cells.
    col_d_with = sum(
        1 for row in ws_with.iter_rows(min_row=5, max_row=ws_with.max_row, min_col=4, max_col=4)
        for cell in row if isinstance(cell.value, (int, float))
    )
    col_d_without = sum(
        1 for row in ws_without.iter_rows(min_row=5, max_row=ws_without.max_row, min_col=4, max_col=4)
        for cell in row if isinstance(cell.value, (int, float))
    )

    # With prior year should have more numeric data in column D than without
    assert col_d_with > col_d_without


@pytest.mark.asyncio
async def test_value_placement_by_row_code(db_session: AsyncSession, sample_project, sample_report_data):
    """Values land at the template's real coordinates by row_code, not row order.

    Task 20.1: BS-002 → C6, BS-003 → C7 (POC soe_standalone layout).
    The template's own row names in column A are preserved untouched.
    """
    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=sample_project.id,
        year=2024,
        mode="audited",
        report_types=["balance_sheet"],
    )

    wb = load_workbook(output)
    ws = None
    for name in wb.sheetnames:
        if "资产负债" in name:
            ws = wb[name]
            break

    assert ws is not None

    # BS-002 货币资金 current → C6
    assert ws["C6"].value == pytest.approx(1234567.89)
    # BS-003 应收账款 current (negative) → C7
    assert ws["C7"].value == pytest.approx(-500000.00)
    # Template's own row label in column A preserved (not overwritten by exporter)
    assert "货币资金" in str(ws["A6"].value)
    assert "{{" not in str(ws["A6"].value)


def test_filename_sanitization():
    """Test filename sanitization removes special characters."""
    from app.routers.report_export import _sanitize_filename

    assert _sanitize_filename("test/file:name.xlsx") == "test_file_name.xlsx"
    assert _sanitize_filename('a*b?c"d.xlsx') == "a_b_c_d.xlsx"
    assert _sanitize_filename("normal_name.xlsx") == "normal_name.xlsx"


def test_company_short_name():
    """Test company short name extraction."""
    from app.routers.report_export import _get_company_short_name

    assert _get_company_short_name("测试科技有限公司") == "测试科技"
    assert _get_company_short_name("AB公司") == "AB公司"
    assert _get_company_short_name("短名") == "短名"
