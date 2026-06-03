"""univer 表格类底稿网格提取服务单测

守护 render-config univer sheet（审定表/明细表/测算表）从模板自动提取网格数据，
修复混合底稿里 univer sheet 只显示死占位「数据尚未导入」的问题。
"""

from __future__ import annotations

import openpyxl

from app.services.wp_grid_extract import extract_grid, extract_grid_from_sheet


def _build_grid_sheet(tmp_path, sheet_name="审定表D1-1"):
    """构建一个贴合致同审定表布局的最小 xlsx（含合并单元格 + 样式）。"""
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = "致同会计师事务所"
    ws["A2"] = "应收票据审定表"
    ws["A5"] = "项目"
    ws["B5"] = "期初数"
    ws["F5"] = "期末数"
    ws["B6"] = "未审数"
    ws["E6"] = "审定数"
    ws["A7"] = "一、应收票据原值"
    ws["A8"] = "银行承兑汇票"
    ws["A10"] = "小计"

    # 样式：表头加粗+居中+紫色填充；分节行紫填充加粗；数据列会计格式
    purple = PatternFill(fill_type="solid", fgColor="FFE4DFEC")
    for coord in ("A5", "B5", "F5", "B6", "E6"):
        ws[coord].font = Font(bold=True, size=10)
        ws[coord].alignment = Alignment(horizontal="center")
        ws[coord].fill = purple
    ws["A7"].font = Font(bold=True, size=10)
    ws["A7"].fill = purple
    ws["A7"].alignment = Alignment(horizontal="left")
    ws["A10"].font = Font(bold=True, size=10)
    # 数据单元格会计格式 + 填充（确保被提取）
    ws["B8"].number_format = '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ '
    ws["B8"].alignment = Alignment(horizontal="right")
    ws["B8"].fill = purple

    # 合并：标题横跨 A1:F1，项目列 A5:A6 纵向合并
    ws.merge_cells("A1:F1")
    ws.merge_cells("A5:A6")
    ws.merge_cells("B5:E5")
    # 列宽
    ws.column_dimensions["A"].width = 20

    fp = tmp_path / "grid.xlsx"
    wb.save(str(fp))
    wb.close()
    return fp, sheet_name


def test_extract_styles(tmp_path):
    """提取真实样式：加粗/对齐/会计格式标记（无 fill）。"""
    fp, sn = _build_grid_sheet(tmp_path)
    g = extract_grid(fp, sn)
    # 项目 (偏移后 A1) → bold + center
    a1 = g["cells"]["A1"]["style"]
    assert a1.get("bold") is True
    assert a1.get("align") == "center"
    assert a1.get("fill") is None  # 不输出 fill
    # 一、应收票据原值 (偏移后 A3) → bold + left
    a3 = g["cells"]["A3"]["style"]
    assert a3.get("bold") is True
    assert a3.get("align") == "left"
    assert a3.get("fill") is None


def test_extract_basic_grid(tmp_path):
    fp, sn = _build_grid_sheet(tmp_path)
    g = extract_grid(fp, sn)
    # 数据表从"项目"行开始（标题行被跳过）
    # 第一行应该是"项目"（偏移后 r=1）
    assert g["max_row"] >= 6
    assert g["max_col"] >= 6
    # 第一行 A1 = "项目"
    assert g["cells"]["A1"]["v"] == "项目"
    assert g["cells"]["A1"]["r"] == 1
    assert g["cells"]["A1"]["c"] == 1
    # "一、应收票据原值" 偏移到 row 3（原 row7 - offset4 = row3）
    a3 = g["cells"].get("A3")
    assert a3 is not None
    assert a3["v"] == "一、应收票据原值"
    # 标题行（致同/审定表）不应出现
    has_title = any("致同" in str(c.get("v", "")) for c in g["cells"].values())
    assert not has_title
    # 无 fill 样式
    fills = [c["style"].get("fill") for c in g["cells"].values() if c.get("style", {}).get("fill")]
    assert len(fills) == 0


def test_merged_ranges_extracted(tmp_path):
    fp, sn = _build_grid_sheet(tmp_path)
    g = extract_grid(fp, sn)
    merged = g["merged_cells"]
    # 标题区合并 (A1:F1) 被跳过，只保留数据区合并
    # A5:A6 (原 row5-6) → 偏移后 row1-2：项目列纵向合并
    a_merge = next((m for m in merged if m["s"]["c"] == 1 and m["s"]["r"] == 1), None)
    assert a_merge is not None
    assert a_merge["e"]["r"] == 2  # 2 行合并
    # B5:E5 (原 row5) → 偏移后 row1: 期初数横向合并
    b_merge = next((m for m in merged if m["s"]["c"] == 2 and m["s"]["r"] == 1), None)
    assert b_merge is not None
    assert b_merge["e"]["c"] == 5  # B~E


def test_col_widths_extracted(tmp_path):
    fp, sn = _build_grid_sheet(tmp_path)
    g = extract_grid(fp, sn)
    assert g["col_widths"].get("A") == 20.0


def test_missing_file_returns_empty(tmp_path):
    g = extract_grid(tmp_path / "nope.xlsx", "X")
    assert g["cells"] == {}
    assert g["merged_cells"] == []
    assert g["max_row"] == 0


def test_missing_sheet_returns_empty(tmp_path):
    fp, _ = _build_grid_sheet(tmp_path)
    g = extract_grid(fp, "不存在的Sheet")
    assert g["cells"] == {}
    assert g["max_row"] == 0


def test_empty_sheet_returns_empty_grid(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空表"
    fp = tmp_path / "empty.xlsx"
    wb.save(str(fp))
    wb.close()
    g = extract_grid(fp, "空表")
    assert g["cells"] == {}
