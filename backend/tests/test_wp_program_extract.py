"""A 类程序表提取服务单测

守护 render-config a-program-console 自动生成程序行的正确性。
锚定修复：程序表中控台「暂无审计程序」——模板程序内容未体现。
"""

from __future__ import annotations

import openpyxl
import pytest

from app.services.wp_program_extract import (
    extract_program_rows,
    extract_program_rows_from_sheet,
)


def _build_program_sheet(tmp_path, sheet_name="应收票据审计程序表D1A"):
    """构建一个贴合致同 A 程序表布局的最小 xlsx。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 表头区（行 1-14 杂项）
    ws["A1"] = "致同会计师事务所"
    ws["A14"] = "三、计划实施的实质性程序"
    # 表头行 15
    ws["A15"] = "序号"
    ws["B15"] = "审计程序"
    ws["C15"] = "程序分类"
    ws["D15"] = "财务报表的认定"
    ws["I15"] = "底稿索引"
    # 认定子表头行 16
    ws["D16"] = "存在"
    ws["E16"] = "完整性"
    ws["F16"] = "权利和\n义务"
    ws["G16"] = "准确性、计价和分摊"
    ws["H16"] = "列报"
    ws["J16"] = "裁剪修改栏"
    # 数据行 17-19
    ws["A17"] = 1
    ws["B17"] = "获取或编制应收票据明细表"
    ws["C17"] = "常规★"
    ws["D17"] = "√"
    ws["E17"] = "√"
    ws["G17"] = "√"
    ws["I17"] = "D1-1\n/D1-2"

    ws["A18"] = 2
    ws["B18"] = "实施函证程序"
    ws["C18"] = "备选"
    ws["F18"] = "√"
    ws["I18"] = "D0"

    ws["A19"] = 3
    ws["B19"] = "检查列报披露"
    ws["C19"] = "舞弊应对"
    ws["H19"] = "√"

    fp = tmp_path / "prog.xlsx"
    wb.save(str(fp))
    wb.close()
    return fp, sheet_name


def test_extract_basic_rows(tmp_path):
    fp, sn = _build_program_sheet(tmp_path)
    programs = extract_program_rows(fp, sn)
    assert len(programs) == 3
    p1 = programs[0]
    assert p1["program_no"] == 1
    assert p1["program_desc"] == "获取或编制应收票据明细表"
    assert p1["program_category"] == "常规★"
    assert p1["status"] == "pending"
    # 认定：D/E/G → existence/completeness/accuracy
    assert p1["assertions"].get("existence") is True
    assert p1["assertions"].get("completeness") is True
    assert p1["assertions"].get("accuracy") is True
    assert "rights" not in p1["assertions"]
    assert "presentation" not in p1["assertions"]
    # 链接底稿换行转 /
    assert p1["linked_workpapers"] == "D1-1/D1-2"


def test_extract_assertion_mapping(tmp_path):
    fp, sn = _build_program_sheet(tmp_path)
    programs = extract_program_rows(fp, sn)
    # row2: 仅 F(权利义务)
    p2 = programs[1]
    assert p2["assertions"] == {"rights": True}
    assert p2["linked_workpapers"] == "D0"
    # row3: 仅 H(列报)
    p3 = programs[2]
    assert p3["assertions"] == {"presentation": True}
    assert p3["program_category"] == "舞弊应对"


def test_missing_file_returns_empty(tmp_path):
    assert extract_program_rows(tmp_path / "nope.xlsx", "X") == []


def test_missing_sheet_returns_empty(tmp_path):
    fp, _ = _build_program_sheet(tmp_path)
    assert extract_program_rows(fp, "不存在的Sheet") == []


def test_no_header_returns_empty(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空表"
    ws["A1"] = "无序号表头"
    fp = tmp_path / "empty.xlsx"
    wb.save(str(fp))
    wb.close()
    assert extract_program_rows(fp, "空表") == []


def test_program_no_monotonic(tmp_path):
    """程序序号应按模板顺序递增（守护数据行不丢不乱序）。"""
    fp, sn = _build_program_sheet(tmp_path)
    programs = extract_program_rows(fp, sn)
    nos = [p["program_no"] for p in programs]
    assert nos == sorted(nos)
    assert nos == [1, 2, 3]


# ─── render-config 模板路径回退（修复「暂无审计程序」回归）─────────────────


@pytest.mark.asyncio
async def test_generate_a_program_data_from_template_path(tmp_path):
    """_generate_a_program_data 用解析到的模板路径能提取出程序行。

    回归守护：当 working_paper.file_path 为空时，render-config 回退到
    wp_templates/ 标准模板（find_template_file_any），A 程序表中控台不再空。
    """
    from app.routers.wp_render_config import _generate_a_program_data

    fp, sn = _build_program_sheet(tmp_path)
    result = await _generate_a_program_data(file_path=str(fp), sheet_name=sn)
    assert isinstance(result, dict)
    assert len(result["programs"]) == 3
    assert result["programs"][0]["program_no"] == 1
    assert result["trim_decisions"] == []


@pytest.mark.asyncio
async def test_generate_a_program_data_none_path_degrades():
    """file_path 为 None（模板也找不到）→ programs 空列表，不抛异常。"""
    from app.routers.wp_render_config import _generate_a_program_data

    result = await _generate_a_program_data(file_path=None, sheet_name="X")
    assert result["programs"] == []
    assert result["trim_decisions"] == []


def test_find_template_file_any_resolves_d1():
    """标准模板库能按 wp_code 解析出 D1 主模板（回退源存在性守护）。

    若该断言失败说明 wp_templates/ 缺 D1 模板或索引未含 D1 —— render-config
    的模板回退将取不到文件，A 程序表/网格会退回空态。
    """
    from app.services.wp_template_init_service import find_template_file_any

    p = find_template_file_any("D1")
    # 环境缺模板库时跳过（CI 可能不含全量 wp_templates）
    if p is None:
        pytest.skip("wp_templates D1 模板不存在（环境未铺模板库）")
    assert p.is_file()
    # 含 D1A 审计程序表 sheet
    import openpyxl

    wb = openpyxl.load_workbook(str(p), read_only=True)
    try:
        assert any("D1A" in s for s in wb.sheetnames)
    finally:
        wb.close()
