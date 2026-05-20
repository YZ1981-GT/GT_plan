"""K 循环合并去重验证测试（spec workpaper-k-admin-cycle Task 1.1）

验证 chain_orchestrator 对 K 循环 14 文件复用 _merge_sheets_dedup 合并去重：

Sprint 0 实测基线（任务执行时 openpyxl 实测复核 2026-05-19）：
- 14 模板文件（K0~K13）
- 152 raw sheet
- 0 历史遗留 sheet（K 模板干净，`_should_skip_historical_sheet` 命中数=0）
- 0 末尾空格 sheet
- 43 跨文件归一化重复（`底稿目录` x14 + `附注披露信息(上市公司)` x13 +
  `附注披露信息(国企)` x12 + `GT_Custom` x8 → 14+13+12+8 - 4 = 43）
- 合并去重后有效 sheet = 109（152 raw - 0 历史 - 43 跨文件去重 = 109）

注：requirements.md v1.0 起草时引用 N_k_dedup_sheets=114 / N_k_cross_file_dups=38，
与 Sprint 0 实测复核（109 / 43）存在偏差。本测试以实测值为准，
spec 三件套基线已同步修正（参见 requirements.md §三·B 偏差归零段落）。

测试目标：
1. K 循环 14 文件被 find_all_template_files 正确发现
2. chain_orchestrator re-export _merge_sheets_dedup / _should_skip_historical_sheet
3. 152 raw sheets → 0 历史遗留 → 43 跨文件去重 → 109 有效 sheet
4. 0 末尾空格 sheet（K 模板干净）
5. 跨文件去重（底稿目录 / 附注披露 / GT_Custom）正确工作
6. D/F/H/I/G/J 历史遗留过滤回归无影响

_Requirements: K-F1_
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _merge_sheets_dedup,
    _normalize_sheet_name,
    _should_skip_historical_sheet,
    find_all_template_files,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "K"

# Collect all K cycle template files
K_TEMPLATE_FILES = sorted(TEMPLATES_DIR.glob("*.xlsx"))


def _get_all_k_sheet_names() -> list[str]:
    """从 14 个 K 循环模板文件中提取全部 sheet 名"""
    all_names: list[str] = []
    for f in K_TEMPLATE_FILES:
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            all_names.extend(wb.sheetnames)
        finally:
            wb.close()
    return all_names


def _get_dedup_sheet_count() -> int:
    """模拟 _merge_sheets_dedup 流程计算去重后 sheet 数

    流程：先过滤历史遗留 → 再归一化去重，与 wp_template_init_service 行为一致。
    """
    all_names = _get_all_k_sheet_names()
    seen_normalized: set[str] = set()
    dedup_count = 0
    for name in all_names:
        if _should_skip_historical_sheet(name):
            continue
        normalized = _normalize_sheet_name(name)
        if normalized not in seen_normalized:
            seen_normalized.add(normalized)
            dedup_count += 1
    return dedup_count


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestKCycleMergeDedup:
    """K-F1: 14 文件合并去重验证（Sprint 0 baseline: 14 files / 152 raw / 0 historical / 109 dedup）"""

    def test_k_cycle_has_14_template_files(self):
        """K 循环应有 14 个模板文件（K0~K13）"""
        assert len(K_TEMPLATE_FILES) == 14, (
            f"Expected 14 K cycle template files, got {len(K_TEMPLATE_FILES)}: "
            f"{[f.name for f in K_TEMPLATE_FILES]}"
        )

    def test_k_cycle_template_files_named_correctly(self):
        """K 循环 14 个模板文件命名应覆盖 K0~K13 全部编号"""
        names = [f.name for f in K_TEMPLATE_FILES]
        for prefix in [f"K{i}" for i in range(14)]:
            assert any(n.startswith(prefix + " ") for n in names), (
                f"Expected file starting with '{prefix} ' in {names}"
            )

    def test_k_cycle_raw_sheet_count_152(self):
        """K 循环 14 文件原始 sheet 总数 = 152（Sprint 0 实测基线）"""
        all_names = _get_all_k_sheet_names()
        assert len(all_names) == 152, (
            f"Expected 152 raw sheets, got {len(all_names)}"
        )

    def test_k_cycle_dedup_sheet_count_109(self):
        """K 循环合并去重后 sheet 数 = 109（Sprint 0 实测基线）

        152 raw - 0 historical - 43 跨文件归一化去重 = 109
        """
        dedup_count = _get_dedup_sheet_count()
        assert dedup_count == 109, (
            f"Expected 109 dedup sheets, got {dedup_count}"
        )

    def test_k_cycle_zero_historical_skipped(self):
        """K 循环 152 sheet 全部不命中历史遗留过滤（K 模板干净，0 命中）"""
        all_names = _get_all_k_sheet_names()
        historical_hits = [
            name for name in all_names if _should_skip_historical_sheet(name)
        ]
        assert len(historical_hits) == 0, (
            f"K cycle templates should be clean (0 historical hits), "
            f"got {len(historical_hits)}: {historical_hits}"
        )

    def test_k_cycle_zero_trailing_space(self):
        """K 循环 152 sheet 全部无末尾空格（与 J 循环不同，K 模板 sheet 名干净）"""
        all_names = _get_all_k_sheet_names()
        trailing_space_sheets = [n for n in all_names if n != n.rstrip()]
        assert len(trailing_space_sheets) == 0, (
            f"K cycle templates should have 0 trailing-space sheets, "
            f"got {len(trailing_space_sheets)}: {trailing_space_sheets}"
        )

    def test_k_cycle_cross_file_redundant_count_43(self):
        """K 循环跨文件去重应移除 43 个重复 sheet

        14 文件包含的跨文件重复 sheet（归一化后）：
        - `底稿目录`：14 份（每个文件 1 份）→ 13 重复
        - `附注披露信息(上市公司)`：13 份 → 12 重复
        - `附注披露信息(国企)`：12 份 → 11 重复
        - `GT_Custom`：8 份（6 个文件无 GT_Custom）→ 7 重复
        合计 13 + 12 + 11 + 7 = 43 重复需去除
        """
        all_names = _get_all_k_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)
        dups = {k: v for k, v in norm_counter.items() if v > 1}
        total_redundant = sum(v - 1 for v in dups.values())
        assert total_redundant == 43, (
            f"Expected 43 cross-file redundant sheets, got {total_redundant}: {dups}"
        )

    def test_k_cycle_cross_file_dedup_targets_identified(self):
        """验证 4 类跨文件去重目标 sheet：底稿目录 / 附注披露(上市公司) / 附注披露(国企) / GT_Custom

        注：还存在一个边缘 case `附注披露信息(国有企业)`（仅 K0 1 份）— 与 `附注披露信息(国企)`
        归一化 key 不同（"国企" vs "国有企业"），不会被去重。属于模板命名遗留差异。
        """
        all_names = _get_all_k_sheet_names()
        non_hist = [n for n in all_names if not _should_skip_historical_sheet(n)]
        norm_counter = Counter(_normalize_sheet_name(n) for n in non_hist)

        # 底稿目录 14 个文件各 1 份
        assert norm_counter.get("底稿目录", 0) == 14, (
            f"'底稿目录' should appear 14 times (one per file), "
            f"got {norm_counter.get('底稿目录', 0)}"
        )

        # GT_Custom 出现在 8 个文件
        assert norm_counter.get("GT_Custom", 0) == 8, (
            f"'GT_Custom' should appear 8 times, got {norm_counter.get('GT_Custom', 0)}"
        )

        # 附注披露信息(上市公司) 13 份（K0~K13 14 文件中除 1 个外都有）
        assert norm_counter.get("附注披露信息(上市公司)", 0) == 13, (
            f"'附注披露信息(上市公司)' should appear 13 times, "
            f"got {norm_counter.get('附注披露信息(上市公司)', 0)}"
        )

        # 附注披露信息(国企) 12 份（精确匹配该归一化 key）
        assert norm_counter.get("附注披露信息(国企)", 0) == 12, (
            f"'附注披露信息(国企)' should appear 12 times, "
            f"got {norm_counter.get('附注披露信息(国企)', 0)}"
        )

        # 边缘 case：附注披露信息(国有企业) 1 份（与"国企"归一化 key 不同，不去重）
        assert norm_counter.get("附注披露信息(国有企业)", 0) == 1, (
            f"'附注披露信息(国有企业)' edge-case should appear 1 time, "
            f"got {norm_counter.get('附注披露信息(国有企业)', 0)}"
        )

    def test_chain_orchestrator_reexports_merge_helpers(self):
        """chain_orchestrator 模块 re-export _merge_sheets_dedup / _should_skip_historical_sheet

        K 循环 chain 必须复用这条 re-export 路径（D/F/H/I/G/J spec 已验证）。
        """
        from app.services import chain_orchestrator as co

        assert hasattr(co, "_merge_sheets_dedup"), (
            "chain_orchestrator 应 re-export _merge_sheets_dedup"
        )
        assert hasattr(co, "_normalize_sheet_name"), (
            "chain_orchestrator 应 re-export _normalize_sheet_name"
        )
        assert hasattr(co, "_should_skip_historical_sheet"), (
            "chain_orchestrator 应 re-export _should_skip_historical_sheet"
        )

        # 验证 re-export 与原始实现是同一函数对象
        from app.services import wp_template_init_service as wts
        assert co._merge_sheets_dedup is wts._merge_sheets_dedup
        assert co._normalize_sheet_name is wts._normalize_sheet_name
        assert co._should_skip_historical_sheet is wts._should_skip_historical_sheet

    def test_find_all_template_files_returns_k_cycle_files(self):
        """find_all_template_files 对 K0~K13 各 wp_code 应能定位到 K 循环模板"""
        for i in range(14):
            wp_code = f"K{i}"
            files = find_all_template_files(wp_code)
            assert len(files) >= 1, (
                f"find_all_template_files('{wp_code}') should return >= 1 file, got {files}"
            )

    def test_find_all_template_files_k_cycle_total_14(self):
        """find_all_template_files 对 K 循环 14 个 wp_code 合计返回 14 个文件"""
        all_files = []
        for i in range(14):
            wp_code = f"K{i}"
            all_files.extend(find_all_template_files(wp_code))
        # K0~K13 各 1 个文件
        assert len(all_files) == 14, (
            f"K cycle should have 14 template files total, got {len(all_files)}: "
            f"{[f.name for f in all_files]}"
        )

    def test_merge_sheets_dedup_on_k_cycle(self, tmp_path: Path):
        """端到端验证 _merge_sheets_dedup 对 K 循环 14 文件实际执行合并去重

        以 K0 作为目标 workbook（11 sheet），将其余 13 文件作为 other_files 合并；验证：
        - skipped_historical = 0（K 模板干净，无任何历史遗留 sheet）
        - skipped_dup >= 30（跨文件归一化重复，K0 自带 1 份底稿目录/附注/GT_Custom 后，
          其余 13 文件再带的同名 sheet 全部归一化重复）
        - 合并后目标 workbook 物理 sheet 数 = 11 + merged
        - 逻辑有效 sheet 数 = 109（去重后基线）
        """
        import shutil

        # 选 K0 作为目标 workbook（避免污染源模板）
        target_src = next(f for f in K_TEMPLATE_FILES if f.name.startswith("K0 "))
        target = tmp_path / target_src.name
        shutil.copy(target_src, target)

        # 其余 13 个文件作为 other_files
        other_files = [f for f in K_TEMPLATE_FILES if not f.name.startswith("K0 ")]
        assert len(other_files) == 13

        stats = _merge_sheets_dedup(target, other_files)

        # K 模板干净，0 历史遗留过滤
        assert stats["skipped_historical"] == 0, (
            f"K templates clean, skipped_historical 应为 0，实际 stats={stats}"
        )

        # 跨文件重复 sheet 应被去重（至少 30 个 — K0 含底稿目录/附注上市/附注国企/GT_Custom 4 类，
        # 其余 13 文件带的同名 sheet 全部归一化重复）
        assert stats["skipped_dup"] >= 30, (
            f"应至少跳过 30 个跨文件归一化重复 sheet，实际 stats={stats}"
        )

        # 合并后目标 workbook 物理 sheet 数 = K0(11) + merged
        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            all_sheets = wb.sheetnames
            physical_count = len(all_sheets)
        finally:
            wb.close()

        # K0 = 11 sheets
        expected_physical = 11 + stats["merged"]
        assert physical_count == expected_physical, (
            f"合并后物理 sheet 数应为 {expected_physical}（K0=11 + merged={stats['merged']}），"
            f"实际 {physical_count}"
        )

        # 逻辑有效 sheet 数 = 109（K 模板无历史遗留）
        logical_count = sum(
            1 for s in all_sheets if not _should_skip_historical_sheet(s)
        )
        assert logical_count == 109, (
            f"合并后逻辑有效 sheet 数应为 109（dedup baseline），实际 {logical_count}"
        )


class TestPriorCyclesHistoricalFilterRegression:
    """D/F/H/I/G/J 历史遗留过滤回归测试 — 确保 K spec 无副作用

    K 循环未引入新的过滤模式（K 模板干净），本测试保证 D/F/H/I/G/J 既有模式不被破坏，
    且 K 正常 sheet 不被误过滤。
    """

    @pytest.mark.parametrize("name,expected", [
        # D 循环历史遗留模式（"修订前" / "（原）" / "(原)"）
        ("主营业务收入审计程序表 D4A（修订前）", True),
        ("D7A（原）", True),
        ("D8A(原)", True),
        ("G1A-修订前", True),
        # F 循环历史遗留模式（G+数字 + 删除/移至 / 示例）
        ("存货计价测试程序G2-8-删除", True),
        ("G2-8-4-移至分析类", True),
        ("产品年度成本比较G2-9-3-删除", True),
        ("函证差异检查表（示例）", True),
        ("合同履约成本测试（示例）", True),
        ("访谈记录与核对示例", True),
        # I 循环历史遗留模式（"示例"末尾）
        ("参考－商誉减值测试示例", True),
        # G 循环历史遗留模式
        ("投资收益实质性程序表G11A-修订前", True),
        # J 循环新增"-删除"模式
        ("股份支付检查表J1-10-删除", True),
        ("应付职工薪酬余额及期后事项检查表J1-11-删除", True),
        # K 循环正常 sheet 不应被过滤
        ("审定表K8-1", False),
        ("审定表K9-1", False),
        ("审定表K1-1", False),
        ("审定表K11-1", False),
        ("明细表K8-2", False),
        ("明细表K9-2", False),
        ("明细表K1-2", False),
        ("明细表 K5-2", False),  # K5-2 sheet 名含空格
        ("明细表K3-2", False),
        ("销售费用实质性程序表K8A", False),
        ("管理费用实质性程序表K9A", False),
        ("资产减值损失实质性程序表K11A", False),
        # 通用正常 sheet 不应被过滤
        ("底稿目录", False),
        ("GT_Custom", False),
        ("附注披露信息(上市公司)", False),
        ("附注披露信息(国企)", False),
    ])
    def test_historical_filter_regression(self, name: str, expected: bool):
        """D/F/H/I/G/J 历史遗留过滤模式仍正确工作，K 正常 sheet 不被误过滤"""
        result = _should_skip_historical_sheet(name)
        assert result is expected, (
            f"_should_skip_historical_sheet('{name}') = {result}, expected {expected}"
        )
