"""L-3 一致性快照单测（spec proposal-remaining-18 task 2.3）

验证 prefill_workpaper_real 执行后将本次 prefill 涉及账户的
TB audited_amount 快照写入 working_paper.prefill_tb_snapshot JSONB。

Validates:
- prefill 执行后 prefill_tb_snapshot 字段被写入
- 快照内容形如 {account_code: amount}（仅 TB/AUX 类公式涉及的科目）
- prefill 失败（无文件/无公式）时不应覆盖既有快照
"""

from __future__ import annotations

import json
import uuid
from pathlib import Path

import pytest
import pytest_asyncio
import sqlalchemy as sa
from openpyxl import Workbook
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.audit_platform_models import AccountCategory, TrialBalance
from app.models.base import Base
from app.models.core import Project, ProjectStatus, ProjectType
from app.models.workpaper_models import (
    WorkingPaper,
    WpIndex,
    WpSourceType,
)
from app.services.prefill_engine import prefill_workpaper_real

# SQLite 不识别 PG JSONB；映射到 JSON 以便 in-memory 测试
SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON


FAKE_USER_ID = uuid.uuid4()
FAKE_PROJECT_ID = uuid.uuid4()


@pytest_asyncio.fixture
async def db_session() -> AsyncSession:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(
        engine, class_=AsyncSession, expire_on_commit=False
    )
    async with session_factory() as session:
        yield session
    await engine.dispose()


def _build_xlsx_with_tb_formula(tmp: Path, formula: str = "=TB('1001','期末余额')") -> Path:
    """构造含 TB 公式的最小 xlsx + 同名 structure.json（prefill 写值需要）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "PFTest"
    ws["A1"] = "账户"
    ws["B1"] = formula  # row=1, col=2 → 0-based row=0, col=1
    fp = tmp / "pftest.xlsx"
    wb.save(fp)
    wb.close()

    # 构造 minimal structure.json，prefill 才会增加 filled 计数
    structure = {
        "sheets": [{"name": "PFTest"}],
        "rows": [
            {
                "cells": [
                    {"value": "账户", "formula": None},
                    {"value": "", "formula": formula},
                ]
            }
        ],
    }
    structure_path = fp.with_suffix(".structure.json")
    with open(structure_path, "w", encoding="utf-8") as sf:
        json.dump(structure, sf, ensure_ascii=False)

    return fp


async def _seed_project_and_wp(
    db: AsyncSession, file_path: Path
) -> uuid.UUID:
    project = Project(
        id=FAKE_PROJECT_ID,
        name="L-3 快照测试",
        client_name="L-3 快照测试",
        project_type=ProjectType.annual,
        status=ProjectStatus.planning,
        created_by=FAKE_USER_ID,
    )
    db.add(project)
    await db.flush()

    wp_index = WpIndex(
        project_id=FAKE_PROJECT_ID,
        wp_code="PF-TEST",
        wp_name="L-3 快照测试底稿",
    )
    db.add(wp_index)
    await db.flush()

    wp = WorkingPaper(
        project_id=FAKE_PROJECT_ID,
        wp_index_id=wp_index.id,
        file_path=str(file_path),
        source_type=WpSourceType.template,
        file_version=1,
    )
    db.add(wp)
    await db.commit()
    return wp.id


def _add_tb_row(
    db: AsyncSession,
    account_code: str,
    amount: float,
    year: int = 2025,
) -> None:
    tb = TrialBalance(
        project_id=FAKE_PROJECT_ID,
        year=year,
        company_code="C001",
        standard_account_code=account_code,
        account_name=f"科目-{account_code}",
        account_category=AccountCategory.asset,
        unadjusted_amount=amount,
        audited_amount=amount,
    )
    db.add(tb)


@pytest.mark.asyncio
async def test_prefill_records_tb_snapshot_for_tb_formula(
    db_session: AsyncSession, tmp_path: Path
):
    """L-3 task 2.3: prefill 后 prefill_tb_snapshot 应记录 TB 公式涉及账户的 audited_amount"""
    fp = _build_xlsx_with_tb_formula(tmp_path, "=TB('1001','期末余额')")
    wp_id = await _seed_project_and_wp(db_session, fp)
    _add_tb_row(db_session, "1001", 12345.67, year=2025)
    await db_session.commit()

    result = await prefill_workpaper_real(
        db=db_session,
        project_id=FAKE_PROJECT_ID,
        year=2025,
        wp_id=wp_id,
    )
    await db_session.commit()

    assert result["status"] == "ok", result
    # 至少扫到 1 条公式
    assert result["formulas_found"] >= 1

    # 重新加载 wp 验证 prefill_tb_snapshot 已写入
    wp = (
        await db_session.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )
    ).scalar_one()
    assert wp.prefill_tb_snapshot is not None, "prefill_tb_snapshot 应被写入"
    assert "1001" in wp.prefill_tb_snapshot
    assert float(wp.prefill_tb_snapshot["1001"]) == pytest.approx(12345.67)


@pytest.mark.asyncio
async def test_prefill_snapshot_extracts_only_tb_aux_account_codes(
    db_session: AsyncSession, tmp_path: Path
):
    """快照只覆盖 TB / AUX 类公式（不抓 WP/PREV/NOTE 等首参为 wp_code 的公式）"""
    # AUX 4-arg + TB 2-arg 各一条；外加 WP 公式（首参是 wp_code 不应被抓）
    wb = Workbook()
    ws = wb.active
    ws.title = "PFTest"
    ws["A1"] = "h1"
    ws["A2"] = "h2"
    ws["A3"] = "h3"
    ws["B1"] = "=TB('2001','期末余额')"
    ws["B2"] = "=AUX('1122','客户','C001','期末余额')"
    ws["B3"] = "=WP('E1-1','Sheet1','B5')"
    fp = tmp_path / "pf_multi.xlsx"
    wb.save(fp)
    wb.close()

    structure = {
        "sheets": [{"name": "PFTest"}],
        "rows": [
            {"cells": [{"value": "h1", "formula": None}, {"value": "", "formula": "=TB('2001','期末余额')"}]},
            {"cells": [{"value": "h2", "formula": None}, {"value": "", "formula": "=AUX('1122','客户','C001','期末余额')"}]},
            {"cells": [{"value": "h3", "formula": None}, {"value": "", "formula": "=WP('E1-1','Sheet1','B5')"}]},
        ],
    }
    with open(fp.with_suffix(".structure.json"), "w", encoding="utf-8") as sf:
        json.dump(structure, sf, ensure_ascii=False)

    wp_id = await _seed_project_and_wp(db_session, fp)
    _add_tb_row(db_session, "2001", 9876.54, year=2025)
    _add_tb_row(db_session, "1122", 555.00, year=2025)
    await db_session.commit()

    result = await prefill_workpaper_real(
        db=db_session,
        project_id=FAKE_PROJECT_ID,
        year=2025,
        wp_id=wp_id,
    )
    await db_session.commit()

    assert result["status"] == "ok"
    wp = (
        await db_session.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )
    ).scalar_one()
    snap = wp.prefill_tb_snapshot
    assert snap is not None
    # TB('2001') 和 AUX('1122') 首参都进快照
    assert "2001" in snap
    assert "1122" in snap
    # WP('E1-1', ...) 的首参是 wp_code 不应进快照
    assert "E1-1" not in snap


@pytest.mark.asyncio
async def test_prefill_failure_keeps_existing_snapshot(
    db_session: AsyncSession, tmp_path: Path
):
    """prefill 失败（文件不存在）时不应覆盖已有快照"""
    fp = tmp_path / "nonexistent.xlsx"
    wp_id = await _seed_project_and_wp(db_session, fp)

    # 预先写入历史快照
    pre_existing = {"9999": 111.11}
    wp = (
        await db_session.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )
    ).scalar_one()
    wp.prefill_tb_snapshot = pre_existing
    await db_session.commit()

    result = await prefill_workpaper_real(
        db=db_session,
        project_id=FAKE_PROJECT_ID,
        year=2025,
        wp_id=wp_id,
    )
    await db_session.commit()

    assert result["status"] == "error"

    wp = (
        await db_session.execute(
            sa.select(WorkingPaper).where(WorkingPaper.id == wp_id)
        )
    ).scalar_one()
    # 历史快照保留
    assert wp.prefill_tb_snapshot == pre_existing
