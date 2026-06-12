"""Property 2: Metadata Embed-Extract Round-Trip (PBT)

对任意有效 MetadataBundle，embed→extract 返回相同值。

**Validates: Requirements 1.4, 3.1, 3.2, 3.3**

Testing framework: hypothesis
"""

from __future__ import annotations

import tempfile
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID

from docx import Document
from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import Workbook, load_workbook

from app.schemas.wp_export_schemas import MetadataBundle
from app.services.wp_export.metadata_codec import MetadataCodec


# ─── Hypothesis Strategy ─────────────────────────────────────────────────────


@st.composite
def st_metadata_bundle(draw: st.DrawFn) -> MetadataBundle:
    """Generate valid MetadataBundle instances.

    - wp_code: non-empty string (e.g. "D1", "E2-3")
    - project_id: UUID
    - file_version: positive int
    - export_timestamp: datetime with timezone
    - preparer: str | None
    - reviewer: str | None
    - review_status: str | None
    """
    wp_code = draw(
        st.from_regex(r"[A-Z][0-9](?:-[0-9]{1,2})?", fullmatch=True)
    )
    project_id = draw(st.uuids())
    file_version = draw(st.integers(min_value=1, max_value=9999))
    # Generate timezone-aware datetime (use UTC to avoid DST issues)
    export_timestamp = draw(
        st.datetimes(
            min_value=datetime(2020, 1, 1),
            max_value=datetime(2030, 12, 31),
            timezones=st.just(timezone.utc),
        )
    )
    # Use printable chars only (openpyxl rejects control chars in custom properties)
    _safe_chars = st.characters(whitelist_categories=("L", "N", "P", "Z"), blacklist_characters="\x00")
    preparer = draw(st.one_of(st.none(), st.text(alphabet=_safe_chars, min_size=1, max_size=20).filter(lambda s: s.strip())))
    reviewer = draw(st.one_of(st.none(), st.text(alphabet=_safe_chars, min_size=1, max_size=20).filter(lambda s: s.strip())))
    review_status = draw(
        st.one_of(
            st.none(),
            st.sampled_from(["draft", "approved", "rejected", "pending"]),
        )
    )

    return MetadataBundle(
        wp_code=wp_code,
        project_id=project_id,
        file_version=file_version,
        export_timestamp=export_timestamp,
        preparer=preparer,
        reviewer=reviewer,
        review_status=review_status,
    )


# ─── Property Tests ──────────────────────────────────────────────────────────


class TestMetadataRoundTrip:
    """Property 2: 对任意有效 MetadataBundle，embed→extract 返回相同值"""

    @given(metadata=st_metadata_bundle())
    @settings(max_examples=5)
    def test_xlsx_embed_extract_roundtrip(self, metadata: MetadataBundle) -> None:
        """xlsx: embed→save→load→extract 返回相同 MetadataBundle

        **Validates: Requirements 1.4, 3.1, 3.2, 3.3**
        """
        codec = MetadataCodec()
        wb = Workbook()
        codec.embed_xlsx(wb, metadata)

        # openpyxl custom properties require save/load to round-trip
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)

        try:
            wb.save(str(path))
            wb2 = load_workbook(str(path))
            extracted = codec.extract_xlsx(wb2)
        finally:
            path.unlink(missing_ok=True)

        assert extracted is not None, "extract_xlsx returned None"
        assert extracted.wp_code == metadata.wp_code
        assert extracted.project_id == metadata.project_id
        assert extracted.file_version == metadata.file_version
        assert extracted.export_timestamp == metadata.export_timestamp
        assert extracted.preparer == metadata.preparer
        assert extracted.reviewer == metadata.reviewer
        assert extracted.review_status == metadata.review_status

    @given(metadata=st_metadata_bundle())
    @settings(max_examples=5)
    def test_docx_embed_extract_roundtrip(self, metadata: MetadataBundle) -> None:
        """docx: embed→save→load→extract 返回相同 MetadataBundle

        **Validates: Requirements 1.4, 3.1, 3.2, 3.3**
        """
        codec = MetadataCodec()
        doc = Document()
        codec.embed_docx(doc, metadata)

        # python-docx requires save/load to round-trip
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
            path = Path(f.name)

        try:
            doc.save(str(path))
            doc2 = Document(str(path))
            extracted = codec.extract_docx(doc2)
        finally:
            path.unlink(missing_ok=True)

        assert extracted is not None, "extract_docx returned None"
        assert extracted.wp_code == metadata.wp_code
        assert extracted.project_id == metadata.project_id
        assert extracted.file_version == metadata.file_version
        assert extracted.export_timestamp == metadata.export_timestamp
        assert extracted.preparer == metadata.preparer
        assert extracted.reviewer == metadata.reviewer
        assert extracted.review_status == metadata.review_status
