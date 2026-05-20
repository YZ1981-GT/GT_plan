"""Task 2.24: reseed + prefill_engine 验证 H1 两级链路

验证 H1 两级 prefill 链路：
  Level 1: TB/LEDGER → H1-2 明细表 (prefill formulas populate H1-2 cells)
  Level 2: H1-1 审定表 has cross_sheet formulas that reference H1-2 cells → auto-calculate

类似 F2 两级链路（TB/AUX → F2-2 → F2-1），H1 的链路是：
  - H1-2 明细表从 TB 取数（=TB('1601','期初余额') 等）
  - H1-1 审定表既有自己的 TB prefill（期初/未审数），也通过 Excel 模板内
    cross_sheet 公式引用 H1-2 的数据自动计算

验证内容：
1. prefill_formula_mapping.json 中 H1-2 明细表有 ≥ 10 cells（TB 取数）
2. prefill_formula_mapping.json 中 H1-1 审定表有 ≥ 5 cells（TB/ADJ/PREV）
3. H1 模板文件中 H1-1 审定表含 cross_sheet 公式引用 H1-2（openpyxl 验证）
4. prefill_formula_mapping.json 可正确加载（reseed 模拟）
5. H1-2 明细表不使用 =WP 公式（防循环引用，与 F2-2 同铁律）
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
PREFILL_FILE = DATA_DIR / "prefill_formula_mapping.json"
TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "wp_templates" / "H"

VALID_FORMULA_TYPES = {
    "TB", "TB_SUM", "ADJ", "PREV", "WP", "AUX", "LEDGER", "LEDGER_DETAIL",
    "COUNT_LEDGER", "NOTE", "TB_AUX", "SUM_TB",
}


@pytest.fixture
def prefill_data():
    assert PREFILL_FILE.exists(), f"prefill_formula_mapping.json not found at {PREFILL_FILE}"
    with open(PREFILL_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _h1_entries(data):
    """获取所有 wp_code='H1' 的 prefill entries"""
    return [m for m in data["mappings"] if m.get("wp_code") == "H1"]


def _h1_2_entries(data):
    """获取 H1-2 明细表的 prefill entries"""
    return [m for m in data["mappings"]
            if m.get("wp_code") == "H1" and "明细表H1-2" in m.get("sheet", "")]


def _h1_1_entries(data):
    """获取 H1-1 审定表的 prefill entries"""
    return [m for m in data["mappings"]
            if m.get("wp_code") == "H1" and "审定表H1-1" in m.get("sheet", "")]


# ─── Test 1: prefill_formula_mapping.json 可正确加载（reseed 模拟）───────────


def test_prefill_file_loads_correctly():
    """模拟 reseed：验证 prefill_formula_mapping.json 可正确解析为 JSON"""
    assert PREFILL_FILE.exists(), "prefill_formula_mapping.json 不存在"
    with open(PREFILL_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    assert "mappings" in data, "JSON 缺少 mappings 字段"
    assert isinstance(data["mappings"], list), "mappings 应为 list"
    assert len(data["mappings"]) > 0, "mappings 不应为空"


def test_prefill_h1_entries_exist(prefill_data):
    """H1 wp_code 在 prefill_formula_mapping.json 中有多个 entry"""
    h1_entries = _h1_entries(prefill_data)
    # H1 应有多个 entry：审定表 + 明细表 + 折旧测算(3版) + 折旧分配 + 减值测算 + 分析程序
    assert len(h1_entries) >= 3, (
        f"H1 entries = {len(h1_entries)}, expected >= 3 "
        f"(审定表 + 明细表 + 折旧/减值)"
    )


# ─── Test 2: Level 1 — H1-2 明细表 prefill cells ≥ 10 ─────────────────────


def test_h1_2_detail_sheet_has_sufficient_cells(prefill_data):
    """Level 1: H1-2 明细表从 TB 取数，cells ≥ 10"""
    h1_2 = _h1_2_entries(prefill_data)
    assert h1_2, "H1-2 明细表 prefill mapping 不存在"
    total_cells = sum(len(m.get("cells", [])) for m in h1_2)
    assert total_cells >= 10, (
        f"H1-2 明细表 cells = {total_cells}, expected >= 10 "
        f"(固定资产原值+累计折旧+减值准备 × 期初/期末/借方/贷方)"
    )


def test_h1_2_uses_tb_formulas(prefill_data):
    """H1-2 明细表主要使用 =TB 公式从试算表取数"""
    h1_2 = _h1_2_entries(prefill_data)
    assert h1_2, "H1-2 明细表 prefill mapping 不存在"
    tb_cells = []
    for m in h1_2:
        for cell in m.get("cells", []):
            if cell.get("formula_type") == "TB":
                tb_cells.append(cell)
    # H1-2 应有多个 TB 公式（1601 期初/期末/借方/贷方 + 1602 同理 + 1603）
    assert len(tb_cells) >= 6, (
        f"H1-2 TB cells = {len(tb_cells)}, expected >= 6"
    )


def test_h1_2_no_wp_formula(prefill_data):
    """ADR-H4 铁律：H1-2 明细表（中间 sheet）禁止使用 =WP（防循环引用）

    H1-2 是 H1-1 审定表 cross_sheet 公式的数据源；
    如果 H1-2 用 =WP 引用 H1-1，会形成循环依赖。
    """
    h1_2 = _h1_2_entries(prefill_data)
    assert h1_2, "H1-2 明细表 prefill mapping 不存在"
    bad = []
    for m in h1_2:
        for cell in m.get("cells", []):
            if cell.get("formula_type") == "WP":
                bad.append((m["sheet"], cell.get("cell_ref"), cell.get("formula")))
    assert not bad, f"H1-2 中间 sheet 不应用 =WP（违反 ADR-H4）: {bad}"


def test_h1_2_account_codes_correct(prefill_data):
    """H1-2 明细表 account_codes 应包含 1601/1602/1603（固定资产/累计折旧/减值准备）"""
    h1_2 = _h1_2_entries(prefill_data)
    assert h1_2, "H1-2 明细表 prefill mapping 不存在"
    all_codes = set()
    for m in h1_2:
        all_codes.update(m.get("account_codes", []))
    # 至少包含 1601（固定资产原值）和 1602（累计折旧）
    assert "1601" in all_codes, "H1-2 应包含 1601（固定资产原值）"
    assert "1602" in all_codes, "H1-2 应包含 1602（累计折旧）"


# ─── Test 3: Level 2 — H1-1 审定表 prefill cells ≥ 5 ──────────────────────


def test_h1_1_audit_table_has_sufficient_cells(prefill_data):
    """Level 2: H1-1 审定表有自己的 prefill cells ≥ 5"""
    h1_1 = _h1_1_entries(prefill_data)
    assert h1_1, "H1-1 审定表 prefill mapping 不存在"
    total_cells = sum(len(m.get("cells", [])) for m in h1_1)
    assert total_cells >= 5, (
        f"H1-1 审定表 cells = {total_cells}, expected >= 5 "
        f"(期初/未审数/AJE/RJE/上年审定数/累计折旧)"
    )


def test_h1_1_has_tb_and_adj_formulas(prefill_data):
    """H1-1 审定表应有 TB + ADJ + PREV 公式"""
    h1_1 = _h1_1_entries(prefill_data)
    assert h1_1, "H1-1 审定表 prefill mapping 不存在"
    formula_types = set()
    for m in h1_1:
        for cell in m.get("cells", []):
            formula_types.add(cell.get("formula_type"))
    assert "TB" in formula_types, "H1-1 应有 =TB 公式（期初/未审数）"
    assert "ADJ" in formula_types, "H1-1 应有 =ADJ 公式（AJE/RJE 调整）"
    assert "PREV" in formula_types, "H1-1 应有 =PREV 公式（上年审定数）"


# ─── Test 4: H1 模板 cross_sheet 公式验证（openpyxl）─────────────────────────


def test_h1_template_cross_sheet_formulas():
    """验证 H1 模板文件中 H1-1 审定表含 cross_sheet 公式引用 H1-2

    H1-1 审定表通过 Excel 内部公式（如 ='明细表H1-2'!D35）引用 H1-2 数据，
    实现两级链路的第二级：H1-2 prefill 填充后 → H1-1 自动计算。

    实测确认：H1-1 Row 8~10+ 有大量 ='明细表H1-2'!Dxx / !Jxx / !Ixx 等公式。
    """
    from openpyxl import load_workbook

    h1_file = TEMPLATE_DIR / "H1 固定资产.xlsx"
    if not h1_file.exists():
        pytest.skip("H1 固定资产.xlsx 模板文件不存在")

    wb = load_workbook(h1_file, read_only=False, data_only=False)

    # 确认两个关键 sheet 都存在
    assert "审定表H1-1" in wb.sheetnames, (
        f"H1 模板缺少 '审定表H1-1' sheet, 实际 sheets: {wb.sheetnames[:10]}"
    )
    assert "明细表H1-2" in wb.sheetnames, (
        f"H1 模板缺少 '明细表H1-2' sheet, 实际 sheets: {wb.sheetnames[:10]}"
    )

    # 扫描 H1-1 审定表中引用 H1-2 的公式
    ws = wb["审定表H1-1"]
    cross_sheet_refs = []
    for row in ws.iter_rows(min_row=1, max_row=50, max_col=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                # Excel cross_sheet 公式格式：='明细表H1-2'!D35
                if "明细表H1-2" in val and "!" in val:
                    cross_sheet_refs.append({
                        "cell": cell.coordinate,
                        "formula": val,
                    })

    wb.close()

    # H1-1 审定表必须有 cross_sheet 公式引用 H1-2（两级链路核心）
    assert len(cross_sheet_refs) >= 3, (
        f"H1-1 审定表 cross_sheet 引用 H1-2 的公式数 = {len(cross_sheet_refs)}, "
        f"expected >= 3（实测有 10+ 个 ='明细表H1-2'!Dxx 等公式）"
    )


# ─── Test 5: 全部 H1 entries formula_type 合法 ─────────────────────────────


def test_h1_all_formulas_valid_type(prefill_data):
    """H1 所有 prefill cells 的 formula_type 合法"""
    h1_entries = _h1_entries(prefill_data)
    invalid = []
    for m in h1_entries:
        for cell in m.get("cells", []):
            ft = cell.get("formula_type")
            if ft not in VALID_FORMULA_TYPES:
                invalid.append((m["sheet"], cell.get("cell_ref"), ft))
    assert not invalid, f"Invalid formula_type: {invalid[:5]}"


def test_h1_all_cells_have_formula_and_description(prefill_data):
    """H1 所有 prefill cells 必须有 formula + description"""
    h1_entries = _h1_entries(prefill_data)
    missing = []
    for m in h1_entries:
        for cell in m.get("cells", []):
            if not cell.get("formula"):
                missing.append((m["sheet"], cell.get("cell_ref"), "formula"))
            if not cell.get("description"):
                missing.append((m["sheet"], cell.get("cell_ref"), "description"))
    assert not missing, f"Missing fields: {missing[:5]}"


# ─── Test 6: H1 两级链路完整性验证 ─────────────────────────────────────────


def test_h1_two_level_chain_integrity(prefill_data):
    """验证 H1 两级 prefill 链路完整性：

    Level 1: TB → H1-2 明细表（prefill 填充 1601/1602/1603 科目数据）
    Level 2: H1-1 审定表（prefill 填充 + Excel cross_sheet 引用 H1-2 自动计算）

    完整性条件：
    1. H1-2 有 TB 公式取 1601/1602 数据
    2. H1-1 有 TB 公式取 1601/1602 数据（直接取数）
    3. H1-2 和 H1-1 在同一 workbook（支持 cross_sheet）
    4. H1-2 不用 =WP（防循环引用）
    """
    # Level 1: H1-2 从 TB 取数
    h1_2 = _h1_2_entries(prefill_data)
    assert h1_2, "链路断裂：H1-2 明细表无 prefill entry"
    h1_2_tb_formulas = [
        cell for m in h1_2 for cell in m.get("cells", [])
        if cell.get("formula_type") == "TB"
    ]
    assert len(h1_2_tb_formulas) >= 6, (
        f"链路薄弱：H1-2 TB cells = {len(h1_2_tb_formulas)}, "
        f"需 >= 6 才能支撑 H1-1 cross_sheet 计算"
    )

    # Level 2: H1-1 有自己的 prefill
    h1_1 = _h1_1_entries(prefill_data)
    assert h1_1, "链路断裂：H1-1 审定表无 prefill entry"
    h1_1_cells = sum(len(m.get("cells", [])) for m in h1_1)
    assert h1_1_cells >= 5, (
        f"H1-1 cells = {h1_1_cells}, 审定表至少需 5 cells"
    )

    # 验证 H1-2 不用 =WP（防循环引用）
    for m in h1_2:
        for cell in m.get("cells", []):
            assert cell.get("formula_type") != "WP", (
                f"循环引用风险：H1-2 使用了 =WP 公式 "
                f"({m['sheet']}/{cell.get('cell_ref')})"
            )


# ─── Test 7: reseed 加载验证（模拟 _load_prefill_mappings）──────────────────


def test_reseed_loads_h1_entries():
    """模拟 reseed 端点的 _load_prefill_mappings 逻辑，验证 H1 entries 可正确加载"""
    # 模拟 template_library_mgmt._load_prefill_mappings
    with open(PREFILL_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    mappings = data.get("mappings", [])
    h1_entries = [m for m in mappings if m.get("wp_code") == "H1"]

    # 验证加载后的 H1 entries 结构完整
    for entry in h1_entries:
        assert "wp_code" in entry
        assert "sheet" in entry
        assert "cells" in entry
        assert isinstance(entry["cells"], list)
        for cell in entry["cells"]:
            assert "cell_ref" in cell
            assert "formula" in cell
            assert "formula_type" in cell

    # 验证 H1 总 cells 数量（降级后目标 ≥ 70 cells 中 H1 贡献最大）
    total_h1_cells = sum(len(m.get("cells", [])) for m in h1_entries)
    # H1 应有：审定表 7 + 明细表 11 + 折旧测算(3版) 10+6+5 + 折旧分配 8 + 减值 8 + 分析程序 = 55+
    assert total_h1_cells >= 40, (
        f"H1 total cells = {total_h1_cells}, expected >= 40 "
        f"(审定表+明细表+折旧3版+折旧分配+减值+分析程序)"
    )


# ─── Test 8: H1 折旧测算 3 版 sheet 都有 prefill ──────────────────────────


def test_h1_12_all_three_versions_have_prefill(prefill_data):
    """H1-12 折旧测算表 3 个版本都有 prefill entries（LEDGER 取数）"""
    h1_entries = _h1_entries(prefill_data)
    h1_12_sheets = [
        m["sheet"] for m in h1_entries
        if "H1-12" in m.get("sheet", "")
    ]
    expected_versions = [
        "折旧测算表（不含减值）-直线法H1-12",
        "折旧测算表（含减值）H1-12",
        "折旧测算表（多次减值）H1-12",
    ]
    for version in expected_versions:
        assert version in h1_12_sheets, (
            f"H1-12 缺少版本: {version}, 实际: {h1_12_sheets}"
        )


def test_h1_12_uses_ledger_formulas(prefill_data):
    """H1-12 折旧测算表使用 =LEDGER 公式按月取折旧数据"""
    h1_entries = _h1_entries(prefill_data)
    h1_12 = [m for m in h1_entries if "H1-12" in m.get("sheet", "")]
    assert h1_12, "H1-12 折旧测算表 prefill mapping 缺失"
    has_ledger = any(
        cell.get("formula_type") == "LEDGER"
        for m in h1_12
        for cell in m.get("cells", [])
    )
    assert has_ledger, "H1-12 折旧测算表至少应包含 =LEDGER 公式（按月取折旧）"
