# Feature: workpaper-bad-debt-nested-structure — Task 12.2 导出单元测试
"""BadDebtExportService 致同 14 列模板导出单元测试。

覆盖（Req 7.1/7.3/7.5）：
- 14 列顺序：列标题行(R11) A~N 与致同模板一致
- 行布局：父行 → 子行 → 合计行
- 空值非零：Child_Row 空金额列输出空单元格（None，非 0）
- 缩进格式："其中"子行 A 列前加两个空格
- 元信息保留：R1-R9 事务所/被审计单位/审计期间

DB：in-process 内存 SQLite，仅建 bad_debt_detail_rows 表。
用 openpyxl 读回断言。
"""

from __future__ import annotations

import io
import uuid
from decimal import Decimal

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.base import Base
from app.models.bad_debt_models import BadDebtDetailRow, ProvisionMethod
from app.schemas.bad_debt_schemas import (
    CreateChildRowDTO,
    CreateParentRowDTO,
    RowAmounts,
    UpdateRowDTO,
)
from app.services.bad_debt_export_service import (
    BadDebtExportMeta,
    BadDebtExportService,
)
from app.services.bad_debt_nested_table_service import NestedTableService

# 列标题行 / 数据起始行（与 service 常量一致）
_HEADER_COL_ROW = 11
_DATA_START_ROW = 12

# 期望的 14 列标题（R11，A~N）
_EXPECTED_TITLES = [
    "项目",
    "期初未审数",
    "期初账项调整",
    "重分类调整(期初)",
    "期初审定数",
    "本期计提",
    "其他增加",
    "本期转回",
    "核销",
    "其他减少",
    "期末未审数",
    "期末账项调整",
    "重分类调整(期末)",
    "期末审定数",
]


@pytest_asyncio.fixture
async def session() -> AsyncSession:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(
            Base.metadata.create_all, tables=[BadDebtDetailRow.__table__]
        )
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as s:
        yield s
    await engine.dispose()


async def _build_sample(session: AsyncSession) -> uuid.UUID:
    """构建：单项评估计提(2子行) + 信用风险组合(1子行) 的样例树。"""
    svc = NestedTableService(session)
    wp = uuid.uuid4()
    p1 = await svc.create_parent_row(
        wp,
        CreateParentRowDTO(
            provision_method=ProvisionMethod.INDIVIDUAL, row_label="按单项评估计提"
        ),
    )
    # 子行只填 E/N（K 留空 → 导出应为空单元格）
    await svc.create_child_row(
        p1.id,
        CreateChildRowDTO(
            row_label="甲公司", amount_e=Decimal("100.00"), amount_n=Decimal("80.00")
        ),
    )
    await svc.create_child_row(
        p1.id,
        CreateChildRowDTO(
            row_label="乙公司", amount_e=Decimal("50.00"), amount_n=Decimal("40.00")
        ),
    )
    p2 = await svc.create_parent_row(
        wp,
        CreateParentRowDTO(
            provision_method=ProvisionMethod.CREDIT_RISK_AGING,
            row_label="信用风险组合-账龄分析法",
        ),
    )
    await svc.create_child_row(
        p2.id,
        CreateChildRowDTO(
            row_label="组合A", amount_e=Decimal("200.00"), amount_n=Decimal("180.00")
        ),
    )
    return wp


def _read_ws(buf: io.BytesIO):
    return load_workbook(buf).active


@pytest.mark.asyncio
async def test_export_14_column_header_order(session: AsyncSession):
    """14 列顺序：R11 列标题 A~N 与致同模板一致。"""
    wp = await _build_sample(session)
    buf = await BadDebtExportService(session).export_bytes(wp)
    ws = _read_ws(buf)
    actual = [ws.cell(row=_HEADER_COL_ROW, column=c).value for c in range(1, 15)]
    assert actual == _EXPECTED_TITLES


@pytest.mark.asyncio
async def test_export_row_layout_parent_child_summary(session: AsyncSession):
    """行布局：父行 → 子行 → 父行 → 子行 → 合计行。"""
    wp = await _build_sample(session)
    buf = await BadDebtExportService(session).export_bytes(wp)
    ws = _read_ws(buf)

    labels = []
    r = _DATA_START_ROW
    while True:
        v = ws.cell(row=r, column=1).value
        if v is None:
            break
        labels.append(v)
        r += 1

    assert labels == [
        "按单项评估计提",
        "  甲公司",
        "  乙公司",
        "信用风险组合-账龄分析法",
        "  组合A",
        "合计",
    ]


@pytest.mark.asyncio
async def test_export_child_indent(session: AsyncSession):
    """缩进格式：子行 A 列前加两个空格。"""
    wp = await _build_sample(session)
    buf = await BadDebtExportService(session).export_bytes(wp)
    ws = _read_ws(buf)
    # R13 = 第一个子行
    child_label = ws.cell(row=_DATA_START_ROW + 1, column=1).value
    assert child_label.startswith("  ")
    assert child_label == "  甲公司"


@pytest.mark.asyncio
async def test_export_empty_amount_is_blank_not_zero(session: AsyncSession):
    """空值非零：子行未填的 K 列(amount_k, column 11)应为 None 而非 0。"""
    wp = await _build_sample(session)
    buf = await BadDebtExportService(session).export_bytes(wp)
    ws = _read_ws(buf)
    # R13 甲公司：K 列(column 11)未填 → None
    k_cell = ws.cell(row=_DATA_START_ROW + 1, column=11).value
    assert k_cell is None
    # E 列(column 5)已填 100
    e_cell = ws.cell(row=_DATA_START_ROW + 1, column=5).value
    assert Decimal(str(e_cell)) == Decimal("100.00")


@pytest.mark.asyncio
async def test_export_summary_row_sums(session: AsyncSession):
    """合计行金额 = 全部父行汇总：E 合计 = 100+50+200 = 350，N = 80+40+180 = 300。"""
    wp = await _build_sample(session)
    buf = await BadDebtExportService(session).export_bytes(wp)
    ws = _read_ws(buf)
    # 合计行 = R17（12父1 13甲 14乙 15父2 16组合A 17合计）
    summary_row = _DATA_START_ROW + 5
    assert ws.cell(row=summary_row, column=1).value == "合计"
    assert Decimal(str(ws.cell(row=summary_row, column=5).value)) == Decimal("350.00")
    assert Decimal(str(ws.cell(row=summary_row, column=14).value)) == Decimal("300.00")


@pytest.mark.asyncio
async def test_export_meta_preserved(session: AsyncSession):
    """元信息保留：R1-R9 事务所/被审计单位/审计期间。"""
    wp = await _build_sample(session)
    meta = BadDebtExportMeta(
        firm_name="致同会计师事务所",
        entity_name="某某有限公司",
        audit_period="2025年度",
        sheet_title="应收账款坏账准备明细表",
    )
    buf = await BadDebtExportService(session).export_bytes(wp, meta)
    ws = _read_ws(buf)
    assert ws.cell(row=1, column=1).value == "应收账款坏账准备明细表"
    assert "致同会计师事务所" in ws.cell(row=3, column=1).value
    assert "某某有限公司" in ws.cell(row=4, column=1).value
    assert "2025年度" in ws.cell(row=5, column=1).value


@pytest.mark.asyncio
async def test_export_default_title_when_no_meta(session: AsyncSession):
    """无 meta 时使用默认表标题，14 列结构仍完整。"""
    wp = await _build_sample(session)
    buf = await BadDebtExportService(session).export_bytes(wp)
    ws = _read_ws(buf)
    assert ws.cell(row=1, column=1).value == "坏账准备明细表"
    actual = [ws.cell(row=_HEADER_COL_ROW, column=c).value for c in range(1, 15)]
    assert actual == _EXPECTED_TITLES
