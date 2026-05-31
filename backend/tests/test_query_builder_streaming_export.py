"""Task 7: query_builder 大结果集流式导出 — 单元测试 + PBT

**Validates: Requirements 4.2, 4.4**

属性:
- P4 流式导出等价: 流式/分页构建 Excel 内容与全量导出一致（仅内存峰值降低）

验证:
- write_only=True 模式生成的 Excel 可被正常解析
- 分页读取（fetchmany）产出的行数据与全量一致
- 表头样式、sheet 名称保持不变
"""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from decimal import Decimal

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.database import get_db
from app.deps import get_current_user
from app.models.audit_platform_models import AccountCategory, TrialBalance
from app.models.base import Base
from app.models.core import Project, ProjectStatus, ProjectType, User, UserRole
from app.routers.query_builder import router as query_builder_router, _EXPORT_FETCH_SIZE

SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON

PROJECT_ID = uuid.uuid4()
USER_ID = uuid.uuid4()


class _FakeUser:
    def __init__(self, role: UserRole = UserRole.admin):
        self.id = USER_ID
        self.role = role
        self.email = "tester@example.com"
        self.username = "tester"
        self.is_active = True


@pytest_asyncio.fixture
async def db_session() -> AsyncSession:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as session:
        yield session
    await engine.dispose()


def _make_app(db_session: AsyncSession) -> FastAPI:
    app = FastAPI()
    app.include_router(query_builder_router)

    async def _override_db():
        yield db_session

    async def _override_user():
        return _FakeUser()

    app.dependency_overrides[get_db] = _override_db
    app.dependency_overrides[get_current_user] = _override_user
    return app


async def _seed_many_rows(db: AsyncSession, count: int) -> None:
    """Seed project + N trial_balance rows."""
    db.add(Project(
        id=PROJECT_ID,
        name="Streaming Export Test",
        client_name="SE Test",
        project_type=ProjectType.annual,
        status=ProjectStatus.planning,
        created_by=USER_ID,
        audit_period_start=date(2025, 1, 1),
        audit_period_end=date(2025, 12, 31),
    ))
    await db.flush()

    categories = list(AccountCategory)
    for i in range(count):
        db.add(TrialBalance(
            id=uuid.uuid4(),
            project_id=PROJECT_ID,
            year=2025,
            company_code="C001",
            standard_account_code=f"{1000 + i:04d}",
            account_name=f"科目_{i}",
            account_category=categories[i % len(categories)],
            audited_amount=Decimal(f"{i * 100.5:.2f}"),
        ))
    await db.flush()


# ═══════════════════════════════════════════════════════════════
# 单元测试: 流式导出基本正确性
# ═══════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_streaming_export_basic(db_session: AsyncSession):
    """流式导出返回有效 xlsx，表头+数据行完整。"""
    await _seed_many_rows(db_session, 5)
    app = _make_app(db_session)
    payload = {
        "table": "trial_balance",
        "fields": ["standard_account_code", "account_name", "audited_amount"],
        "limit": 100,
    }
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as ac:
        resp = await ac.post("/api/query/export-excel", json=payload)

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 表头
    assert rows[0] == ("standard_account_code", "account_name", "audited_amount")
    # 数据行数
    assert len(rows) == 6  # 1 header + 5 data


@pytest.mark.asyncio
async def test_streaming_export_exceeds_fetch_size(db_session: AsyncSession):
    """当行数超过 _EXPORT_FETCH_SIZE 时，分页读取仍产出全部行。"""
    # 用较小的 fetch size 来测试分页逻辑
    import app.routers.query_builder as qb_mod
    original = qb_mod._EXPORT_FETCH_SIZE
    qb_mod._EXPORT_FETCH_SIZE = 3  # 强制每批只取 3 行

    try:
        await _seed_many_rows(db_session, 10)
        app = _make_app(db_session)
        payload = {
            "table": "trial_balance",
            "fields": ["standard_account_code", "account_name"],
            "limit": 100,
        }
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as ac:
            resp = await ac.post("/api/query/export-excel", json=payload)

        assert resp.status_code == 200
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("standard_account_code", "account_name")
        assert len(rows) == 11  # 1 header + 10 data rows
    finally:
        qb_mod._EXPORT_FETCH_SIZE = original


@pytest.mark.asyncio
async def test_streaming_export_empty_result(db_session: AsyncSession):
    """空结果集仍返回有效 xlsx（仅表头）。"""
    # 只建 project 不建 trial_balance
    db_session.add(Project(
        id=PROJECT_ID,
        name="Empty Test",
        client_name="E",
        project_type=ProjectType.annual,
        status=ProjectStatus.planning,
        created_by=USER_ID,
        audit_period_start=date(2025, 1, 1),
        audit_period_end=date(2025, 12, 31),
    ))
    await db_session.flush()

    app = _make_app(db_session)
    payload = {
        "table": "trial_balance",
        "fields": ["standard_account_code"],
        "limit": 50,
    }
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as ac:
        resp = await ac.post("/api/query/export-excel", json=payload)

    assert resp.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("standard_account_code",)
    assert len(rows) == 1  # 仅表头


@pytest.mark.asyncio
async def test_streaming_export_sheet_name(db_session: AsyncSession):
    """Sheet 名称使用表的中文 label（截断 ≤31 字符）。"""
    db_session.add(Project(
        id=PROJECT_ID,
        name="Sheet Name Test",
        client_name="SN",
        project_type=ProjectType.annual,
        status=ProjectStatus.planning,
        created_by=USER_ID,
        audit_period_start=date(2025, 1, 1),
        audit_period_end=date(2025, 12, 31),
    ))
    await db_session.flush()

    app = _make_app(db_session)
    payload = {
        "table": "trial_balance",
        "fields": ["id"],
        "limit": 10,
    }
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as ac:
        resp = await ac.post("/api/query/export-excel", json=payload)

    assert resp.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # sheet 名应为 trial_balance 的中文 label
    assert ws.title == "试算表"


# ═══════════════════════════════════════════════════════════════
# PBT: P4 流式导出等价性
# ═══════════════════════════════════════════════════════════════

from hypothesis import given, settings
from hypothesis import strategies as st


@st.composite
def row_counts(draw):
    """生成合理的行数（覆盖 0、小于 fetch_size、大于 fetch_size 边界）。"""
    return draw(st.sampled_from([0, 1, 3, 5, 10, 15]))


@st.composite
def field_subsets(draw):
    """从 trial_balance 可用字段中随机选子集。"""
    all_fields = ["standard_account_code", "account_name", "audited_amount", "year"]
    chosen = draw(st.lists(
        st.sampled_from(all_fields),
        min_size=1,
        max_size=len(all_fields),
        unique=True,
    ))
    return chosen


@pytest.mark.asyncio
@settings(max_examples=10, deadline=10000)
@given(count=row_counts(), fields=field_subsets())
async def test_pbt_streaming_export_equivalence(count, fields):
    """P4 流式导出等价: 分页流式导出内容与全量一致。

    **Validates: Requirements 4.2, 4.4**

    策略: 随机行数 × 随机字段子集，强制 fetch_size=2 触发多批次，
    验证流式导出的 Excel 行数据与直接 execute 查询结果完全一致。
    """
    import app.routers.query_builder as qb_mod
    from unittest.mock import patch, AsyncMock

    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with factory() as session:
        pid = uuid.uuid4()
        uid = uuid.uuid4()
        session.add(Project(
            id=pid,
            name="PBT Test",
            client_name="PBT",
            project_type=ProjectType.annual,
            status=ProjectStatus.planning,
            created_by=uid,
            audit_period_start=date(2025, 1, 1),
            audit_period_end=date(2025, 12, 31),
        ))
        await session.flush()

        categories = list(AccountCategory)
        for i in range(count):
            session.add(TrialBalance(
                id=uuid.uuid4(),
                project_id=pid,
                year=2025,
                company_code="C001",
                standard_account_code=f"{1000 + i:04d}",
                account_name=f"科目_{i}",
                account_category=categories[i % len(categories)],
                audited_amount=Decimal(f"{i * 100.5:.2f}"),
            ))
        await session.flush()

        # 构建 app
        app = FastAPI()
        app.include_router(query_builder_router)

        async def _override_db():
            yield session

        async def _override_user():
            return _FakeUser()

        app.dependency_overrides[get_db] = _override_db
        app.dependency_overrides[get_current_user] = _override_user

        payload = {"table": "trial_balance", "fields": fields, "limit": 1000}

        # 强制小 fetch_size 触发多批次分页
        original = qb_mod._EXPORT_FETCH_SIZE
        qb_mod._EXPORT_FETCH_SIZE = 2

        # 禁用 Redis 缓存避免跨 hypothesis example 数据泄漏
        with patch("app.services.query_cache.get_cached_result", new_callable=AsyncMock, return_value=None), \
             patch("app.services.query_cache.set_cached_result", new_callable=AsyncMock):
            try:
                async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as ac:
                    # 获取 execute 结果作为基准
                    exec_resp = await ac.post("/api/query/execute", json=payload)
                    assert exec_resp.status_code == 200
                    exec_data = exec_resp.json()
                    expected_rows = exec_data["rows"]

                    # 获取流式导出
                    export_resp = await ac.post("/api/query/export-excel", json=payload)
                    assert export_resp.status_code == 200
            finally:
                qb_mod._EXPORT_FETCH_SIZE = original

        # 解析 Excel
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        ws = wb.active
        excel_rows = list(ws.iter_rows(values_only=True))

        # 表头一致
        assert excel_rows[0] == tuple(fields), f"Header mismatch: {excel_rows[0]} vs {fields}"

        # 数据行数一致
        data_rows = excel_rows[1:]
        assert len(data_rows) == len(expected_rows), (
            f"Row count mismatch: excel={len(data_rows)} vs execute={len(expected_rows)}"
        )

        # 逐行逐列值一致（考虑类型转换：execute 返回 JSON 序列化值，Excel 存原始类型）
        for ri, (excel_row, exec_row) in enumerate(zip(data_rows, expected_rows)):
            for ci, field in enumerate(fields):
                excel_val = excel_row[ci]
                exec_val = exec_row[field]
                # 数值类型统一用 float 比较（Excel int 0 vs JSON float 0.0 等价）
                if isinstance(excel_val, (int, float)) and isinstance(exec_val, (int, float)):
                    assert float(excel_val) == float(exec_val), (
                        f"Row {ri} field '{field}': excel={excel_val!r} vs execute={exec_val!r}"
                    )
                else:
                    ev_str = str(excel_val) if excel_val is not None else None
                    ex_str = str(exec_val) if exec_val is not None else None
                    assert ev_str == ex_str, (
                        f"Row {ri} field '{field}': excel={excel_val!r} vs execute={exec_val!r}"
                    )

    await engine.dispose()
