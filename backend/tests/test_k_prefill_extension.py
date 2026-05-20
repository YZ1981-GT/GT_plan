"""K-F6: K cycle prefill extension validation (≥ 40 new cells).

Validates: Requirements K-F6 / Spec workpaper-k-admin-cycle / Sprint 1 Task 1.6

Sprint 0.X aux 实测结果（design.md ADR-K3）:
- 6601 销售费用 aux_type='客户' 与费用类别维度不匹配 → K8-2 用 =LEDGER_DETAIL
- 6602 管理费用 aux_type='区域2'+'客户' 不匹配 → K9-2 用 =LEDGER_DETAIL
- 1221 其他应收款 aux_type='三方收款标识' → K1-2 用 =AUX 4-arg
- 2241 其他应付款 aux_type='代收代付类别' → K3-2 用 =AUX 4-arg
- K5-2 sheet 真名 '明细表 K5-2'（K5 与 -2 之间有空格）
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

DATA_FILE = Path(__file__).parent.parent / "data" / "prefill_formula_mapping.json"
TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "K"


@pytest.fixture(scope="module")
def prefill_data():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


@pytest.fixture(scope="module")
def k_extension_entries(prefill_data):
    """K cycle 6 new entries from task 1.6 (by wp_code + sheet matching)."""
    target_keys = {
        ("K8", "明细表K8-2"),
        ("K9", "明细表K9-2"),
        ("K1", "明细表K1-2"),
        ("K3", "明细表K3-2"),
        ("K5", "明细表 K5-2"),  # 注意空格
        ("K8", "实质性分析K8-4"),
    }
    return [
        e
        for e in prefill_data["mappings"]
        if (e["wp_code"], e["sheet"]) in target_keys
    ]


# ──────────────────────────────────────────────────────────────────────────
# Total cell count + per-target distribution
# ──────────────────────────────────────────────────────────────────────────


class TestKF6TotalCellCount:
    def test_total_new_cells_at_least_40(self, k_extension_entries):
        total = sum(len(e["cells"]) for e in k_extension_entries)
        assert total >= 40, f"Expected ≥ 40 new prefill cells, got {total}"

    def test_six_entries_present(self, k_extension_entries):
        assert len(k_extension_entries) == 6, (
            f"Expected 6 new K extension entries (K8-2/K9-2/K1-2/K3-2/K5-2/K8-4), "
            f"got {len(k_extension_entries)}"
        )


class TestKF6DistributionPerSheet:
    @pytest.fixture
    def by_sheet(self, k_extension_entries):
        return {(e["wp_code"], e["sheet"]): e for e in k_extension_entries}

    def test_k8_2_at_least_10(self, by_sheet):
        e = by_sheet[("K8", "明细表K8-2")]
        assert len(e["cells"]) >= 10, (
            f"K8-2 should have ≥ 10 cells, got {len(e['cells'])}"
        )

    def test_k9_2_at_least_10(self, by_sheet):
        e = by_sheet[("K9", "明细表K9-2")]
        assert len(e["cells"]) >= 10, (
            f"K9-2 should have ≥ 10 cells, got {len(e['cells'])}"
        )

    def test_k1_2_at_least_6(self, by_sheet):
        e = by_sheet[("K1", "明细表K1-2")]
        assert len(e["cells"]) >= 6, (
            f"K1-2 should have ≥ 6 cells, got {len(e['cells'])}"
        )

    def test_k3_2_at_least_6(self, by_sheet):
        e = by_sheet[("K3", "明细表K3-2")]
        assert len(e["cells"]) >= 6, (
            f"K3-2 should have ≥ 6 cells, got {len(e['cells'])}"
        )

    def test_k5_2_at_least_4(self, by_sheet):
        e = by_sheet[("K5", "明细表 K5-2")]
        assert len(e["cells"]) >= 4, (
            f"K5-2 should have ≥ 4 cells, got {len(e['cells'])}"
        )

    def test_k8_4_at_least_4(self, by_sheet):
        e = by_sheet[("K8", "实质性分析K8-4")]
        assert len(e["cells"]) >= 4, (
            f"K8-4 should have ≥ 4 cells, got {len(e['cells'])}"
        )


# ──────────────────────────────────────────────────────────────────────────
# Real sheet names (openpyxl-verified)
# ──────────────────────────────────────────────────────────────────────────


class TestKF6RealSheetNames:
    """Verify sheet field matches openpyxl-extracted real sheet name."""

    @pytest.fixture(scope="class")
    def real_sheets_per_template(self):
        """Map K wp_code → real sheet names list from openpyxl."""
        result: dict[str, list[str]] = {}
        for f in TEMPLATES_DIR.glob("*.xlsx"):
            # filename starts with 'K{N} '
            prefix = f.name.split(" ")[0]
            wb = load_workbook(str(f), read_only=True, data_only=True)
            try:
                result[prefix] = list(wb.sheetnames)
            finally:
                wb.close()
        return result

    def test_k8_2_sheet_real_name(self, real_sheets_per_template):
        assert "明细表K8-2" in real_sheets_per_template["K8"], (
            f"明细表K8-2 should exist in K8 template, got {real_sheets_per_template['K8']}"
        )

    def test_k9_2_sheet_real_name(self, real_sheets_per_template):
        assert "明细表K9-2" in real_sheets_per_template["K9"]

    def test_k1_2_sheet_real_name(self, real_sheets_per_template):
        assert "明细表K1-2" in real_sheets_per_template["K1"]

    def test_k3_2_sheet_real_name(self, real_sheets_per_template):
        assert "明细表K3-2" in real_sheets_per_template["K3"]

    def test_k5_2_sheet_real_name_with_space(self, real_sheets_per_template):
        """K5-2 真实 sheet 名 K5 与 -2 之间有空格 — 关键！"""
        assert "明细表 K5-2" in real_sheets_per_template["K5"], (
            f"明细表 K5-2（含空格）必须存在于 K5 模板，"
            f"got {real_sheets_per_template['K5']}"
        )
        # 反向校验：无空格版本不应存在
        assert "明细表K5-2" not in real_sheets_per_template["K5"], (
            "明细表K5-2（无空格）不应存在于 K5 模板（实际真名含空格）"
        )

    def test_k8_4_sheet_real_name(self, real_sheets_per_template):
        assert "实质性分析K8-4" in real_sheets_per_template["K8"]


# ──────────────────────────────────────────────────────────────────────────
# AUX 4-arg validation (strict)
# ──────────────────────────────────────────────────────────────────────────


class TestKF6AUX4ArgValidation:
    """All AUX formulas must be 4-arg (3 commas inside =AUX(...))."""

    def test_all_aux_formulas_are_4_arg(self, k_extension_entries):
        offending: list[str] = []
        for entry in k_extension_entries:
            for cell in entry["cells"]:
                if cell["formula_type"] != "AUX":
                    continue
                formula = cell["formula"]
                # Strip "=AUX(" and trailing ")"
                assert formula.startswith("=AUX(") and formula.endswith(")")
                inner = formula[len("=AUX("):-1]
                comma_count = inner.count(",")
                if comma_count != 3:
                    offending.append(
                        f"{entry['wp_code']}/{entry['sheet']}/{cell['cell_ref']}: "
                        f"{formula} ({comma_count} commas)"
                    )
        assert not offending, (
            f"AUX formulas must be 4-arg (3 commas). Offending:\n"
            + "\n".join(offending)
        )

    def test_k1_2_aux_uses_1221_three_party(self, k_extension_entries):
        """K1-2 aux entries must use account 1221 + aux_type 三方收款标识"""
        e = next(
            x for x in k_extension_entries
            if x["wp_code"] == "K1" and x["sheet"] == "明细表K1-2"
        )
        for cell in e["cells"]:
            assert "'1221'" in cell["formula"]
            assert "'三方收款标识'" in cell["formula"]
            assert cell["formula_type"] == "AUX"

    def test_k3_2_aux_uses_2241_dispatch_category(self, k_extension_entries):
        """K3-2 aux entries must use account 2241 + aux_type 代收代付类别"""
        e = next(
            x for x in k_extension_entries
            if x["wp_code"] == "K3" and x["sheet"] == "明细表K3-2"
        )
        for cell in e["cells"]:
            assert "'2241'" in cell["formula"]
            assert "'代收代付类别'" in cell["formula"]
            assert cell["formula_type"] == "AUX"


# ──────────────────────────────────────────────────────────────────────────
# Formula type distribution
# ──────────────────────────────────────────────────────────────────────────


class TestKF6FormulaTypeDistribution:
    def test_k8_2_uses_ledger_detail(self, k_extension_entries):
        """K8-2 should use LEDGER_DETAIL (aux_type='客户' 与费用类别不匹配)"""
        e = next(
            x for x in k_extension_entries
            if x["wp_code"] == "K8" and x["sheet"] == "明细表K8-2"
        )
        types = {c["formula_type"] for c in e["cells"]}
        assert types == {"LEDGER_DETAIL"}, (
            f"K8-2 should use only LEDGER_DETAIL, got {types}"
        )

    def test_k9_2_uses_ledger_detail(self, k_extension_entries):
        e = next(
            x for x in k_extension_entries
            if x["wp_code"] == "K9" and x["sheet"] == "明细表K9-2"
        )
        types = {c["formula_type"] for c in e["cells"]}
        assert types == {"LEDGER_DETAIL"}, (
            f"K9-2 should use only LEDGER_DETAIL, got {types}"
        )

    def test_k5_2_uses_tb(self, k_extension_entries):
        e = next(
            x for x in k_extension_entries
            if x["wp_code"] == "K5" and x["sheet"] == "明细表 K5-2"
        )
        types = {c["formula_type"] for c in e["cells"]}
        assert types == {"TB"}, f"K5-2 should use only TB, got {types}"

    def test_k8_4_uses_tb_and_prev(self, k_extension_entries):
        e = next(
            x for x in k_extension_entries
            if x["wp_code"] == "K8" and x["sheet"] == "实质性分析K8-4"
        )
        types = {c["formula_type"] for c in e["cells"]}
        assert types == {"TB", "PREV"}, (
            f"K8-4 should use TB + PREV, got {types}"
        )


# ──────────────────────────────────────────────────────────────────────────
# Schema integrity
# ──────────────────────────────────────────────────────────────────────────


class TestKF6SchemaIntegrity:
    REQUIRED_ENTRY_FIELDS = {"wp_code", "wp_name", "sheet", "account_codes", "cells"}
    REQUIRED_CELL_FIELDS = {"cell_ref", "formula", "formula_type", "description"}

    def test_entry_required_fields(self, k_extension_entries):
        for entry in k_extension_entries:
            missing = self.REQUIRED_ENTRY_FIELDS - set(entry.keys())
            assert not missing, (
                f"Entry {entry.get('wp_code')}/{entry.get('sheet')} missing fields: {missing}"
            )

    def test_cell_required_fields(self, k_extension_entries):
        for entry in k_extension_entries:
            for cell in entry["cells"]:
                missing = self.REQUIRED_CELL_FIELDS - set(cell.keys())
                assert not missing, (
                    f"Cell in {entry['wp_code']}/{entry['sheet']} "
                    f"missing fields: {missing}"
                )

    def test_no_duplicate_cell_refs_per_entry(self, k_extension_entries):
        for entry in k_extension_entries:
            refs = [c["cell_ref"] for c in entry["cells"]]
            assert len(refs) == len(set(refs)), (
                f"Duplicate cell_ref in {entry['wp_code']}/{entry['sheet']}: "
                f"{[r for r in refs if refs.count(r) > 1]}"
            )

    def test_descriptions_non_empty(self, k_extension_entries):
        for entry in k_extension_entries:
            for cell in entry["cells"]:
                assert cell["description"].strip(), (
                    f"Empty description: {entry['wp_code']}/{entry['sheet']}/{cell['cell_ref']}"
                )
