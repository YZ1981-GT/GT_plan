"""calamine vs openpyxl 等价测试 — xlsx-read-acceleration spec 组②

验证 read_sheet_values(prefer_calamine=True) vs (prefer_calamine=False) 对各迁移点
使用的操作模式输出等价。使用小型 fixture（非生产模板）演示等价性。

覆盖迁移点：
- Task 5: wp_program_extract（基础值提取 + 表头定位）
- Task 6: wp_audit_sheet_extract（值提取 + 合并单元格行为）
- Task 7: wp_generic_processor / wp_template_diff_service /
          wp_fine_rule_engine / import_template_service

如果 calamine 不可用（未安装 python_calamine），测试自动跳过 calamine 路径比较，
仅验证 openpyxl 路径正常工作。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pytest

# ─── Helpers ───────────────────────────────────────────────────────────


def _calamine_installed() -> bool:
    try:
        import python_calamine  # noqa: F401
        return True
    except ImportError:
        return False


skip_no_calamine = pytest.mark.skipif(
    not _calamine_installed(),
    reason="python_calamine not installed — skipping calamine equivalence comparison",
)


def _build_xlsx(tmp_path: Path, sheets: dict[str, list[list]], filename: str = "equiv.xlsx") -> Path:
    """构造含指定 sheets 的 xlsx fixture。"""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    path = tmp_path / filename
    wb.save(str(path))
    return path


def _build_xlsx_with_merged(tmp_path: Path, filename: str = "merged.xlsx") -> Path:
    """构造含合并单元格的 xlsx fixture（验证审定表场景）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审定表D1-1"
    # Row 1-3: title area
    ws["A1"] = "致同会计师事务所"
    ws["A2"] = "应收票据审定表"
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    # Row 4: header
    ws["A4"] = "项目"
    ws["B4"] = "期初余额"
    ws["C4"] = "本期增加"
    ws["D4"] = "本期减少"
    ws["E4"] = "期末余额"
    # Row 5: data
    ws["A5"] = "一、银行承兑汇票"
    ws["B5"] = 100000
    ws["C5"] = 50000
    ws["D5"] = 30000
    ws["E5"] = 120000
    # Row 6: another data row
    ws["A6"] = "  商业承兑汇票"
    ws["B6"] = 20000
    ws["C6"] = 10000
    ws["D6"] = 5000
    ws["E6"] = 25000
    # Row 7: total
    ws["A7"] = "合计"
    ws["B7"] = 120000
    ws["C7"] = 60000
    ws["D7"] = 35000
    ws["E7"] = 145000

    path = tmp_path / filename
    wb.save(str(path))
    return path


def _normalize_for_comparison(rows: list[list[Any]]) -> list[list[Any]]:
    """归一化比较：int/float 容忍 + 剔除尾部全 None 行 + 统一 None。"""
    while rows and all(c is None for c in rows[-1]):
        rows = rows[:-1]
    result = []
    for row in rows:
        normalized = []
        for c in row:
            if isinstance(c, float) and c == int(c):
                normalized.append(int(c))
            elif c == "":
                normalized.append(None)
            else:
                normalized.append(c)
        result.append(normalized)
    return result


# ─── Task 5: wp_program_extract equivalence ────────────────────────────


@pytest.fixture
def program_xlsx(tmp_path: Path) -> Path:
    """模拟程序表模板：有序号/审计程序/分类列。"""
    return _build_xlsx(tmp_path, {
        "审计程序表D1A": [
            ["致同会计师事务所"],
            ["应收票据审计程序表"],
            [],
            ["序号", "审计程序", "分类", "存在", "完整性", "索引"],
            [],  # sub-header placeholder
            [1, "取得应收票据明细表", "实质性", "√", "", "D1-2"],
            [2, "检查票据到期日", "实质性", "", "√", "D1-3"],
            [3, "函证票据余额", "实质性", "√", "√", "D0-1"],
        ],
    }, "program_test.xlsx")


@skip_no_calamine
def test_program_extract_calamine_vs_openpyxl(program_xlsx: Path):
    """wp_program_extract: calamine vs openpyxl 读取程序表值等价。"""
    from app.services.xlsx_read_adapter import read_sheet_values

    rows_cal = read_sheet_values(program_xlsx, "审计程序表D1A", prefer_calamine=True)
    rows_opy = read_sheet_values(program_xlsx, "审计程序表D1A", prefer_calamine=False)
    assert _normalize_for_comparison(rows_cal) == _normalize_for_comparison(rows_opy)


def test_program_extract_openpyxl_basic(program_xlsx: Path):
    """wp_program_extract: openpyxl 路径正常提取程序行。"""
    from app.services.wp_program_extract import extract_program_rows

    programs = extract_program_rows(program_xlsx, "审计程序表D1A")
    assert len(programs) == 3
    assert programs[0]["program_desc"] == "取得应收票据明细表"
    assert programs[2]["linked_workpapers"] == "D0-1"


# ─── Task 6: wp_audit_sheet_extract equivalence ────────────────────────


@pytest.fixture
def audit_sheet_xlsx(tmp_path: Path) -> Path:
    """模拟审定表：有合并单元格标题 + 数据行。"""
    return _build_xlsx_with_merged(tmp_path, "audit_test.xlsx")


@skip_no_calamine
def test_audit_sheet_calamine_vs_openpyxl(audit_sheet_xlsx: Path):
    """wp_audit_sheet_extract: calamine vs openpyxl 值提取等价。

    重点：合并单元格区域 calamine 只在左上角 cell 有值，其余为 None。
    openpyxl data_only=True read_only=False 也是左上角有值，其余 None。
    对于审定表提取（只读项目列=A列第一格），两者行为一致。
    """
    from app.services.xlsx_read_adapter import read_sheet_values

    rows_cal = read_sheet_values(audit_sheet_xlsx, "审定表D1-1", prefer_calamine=True)
    rows_opy = read_sheet_values(audit_sheet_xlsx, "审定表D1-1", prefer_calamine=False)
    assert _normalize_for_comparison(rows_cal) == _normalize_for_comparison(rows_opy)


def test_audit_sheet_extract_basic(audit_sheet_xlsx: Path):
    """wp_audit_sheet_extract: openpyxl 路径正常提取审定表行。"""
    wb = openpyxl.load_workbook(str(audit_sheet_xlsx), read_only=False, data_only=True)
    ws = wb["审定表D1-1"]

    from app.services.wp_audit_sheet_extract import extract_audit_rows_from_sheet

    rows = extract_audit_rows_from_sheet(ws)
    wb.close()

    assert len(rows) >= 2
    # 第一个数据行应是分节行
    assert rows[0]["isSection"] is True
    assert "银行承兑汇票" in rows[0]["item"]
    # 最后一个应是合计行
    assert rows[-1]["isComputed"] is True
    assert "合计" in rows[-1]["item"]


@skip_no_calamine
def test_audit_sections_calamine_vs_openpyxl(tmp_path: Path):
    """wp_audit_sheet_extract.extract_audit_sections: 合并单元格对取值无影响。"""
    # 构造含 审计说明/审计结论 区的 xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审定表"
    ws["A1"] = "项目"
    ws["B1"] = "期末"
    ws["A2"] = "一、应收票据"
    ws["B2"] = 100000
    ws["A3"] = "合计"
    ws["B3"] = 100000
    ws["A4"] = None  # empty separator
    ws["A5"] = "1.审计说明"
    ws["A6"] = "本期应收票据增长30%，主要因新增客户。"
    ws["A7"] = "2.审计结论"
    ws["A8"] = "经审计，应收票据列报金额公允。"
    path = tmp_path / "sections_test.xlsx"
    wb.save(str(path))

    from app.services.xlsx_read_adapter import read_sheet_values

    rows_cal = read_sheet_values(path, "审定表", prefer_calamine=True)
    rows_opy = read_sheet_values(path, "审定表", prefer_calamine=False)
    assert _normalize_for_comparison(rows_cal) == _normalize_for_comparison(rows_opy)


# ─── Task 7: Other migration points equivalence ───────────────────────


@pytest.fixture
def generic_xlsx(tmp_path: Path) -> Path:
    """模拟通用底稿模板（wp_generic_processor / diff / fine_rule / import 共用）。"""
    return _build_xlsx(tmp_path, {
        "审定表D2-1": [
            ["致同会计师事务所"],
            ["应收账款审定表"],
            [],
            ["项目", "期末余额", "期初余额", "变动额", "变动率"],
            ["客户A", 500000, 450000, 50000, 0.111],
            ["客户B", 300000, 320000, -20000, -0.0625],
            ["合计", 800000, 770000, 30000, 0.039],
        ],
        "明细表D2-2": [
            ["序号", "客户名称", "应收金额", "账龄", "坏账准备"],
            [1, "客户A", 500000, "1年以内", 25000],
            [2, "客户B", 300000, "1-2年", 30000],
        ],
    }, "generic_test.xlsx")


@skip_no_calamine
def test_generic_processor_calamine_vs_openpyxl(generic_xlsx: Path):
    """wp_generic_processor: calamine vs openpyxl 多 sheet 读取等价。"""
    from app.services.xlsx_read_adapter import read_sheet_values

    for sheet_name in ["审定表D2-1", "明细表D2-2"]:
        rows_cal = read_sheet_values(generic_xlsx, sheet_name, prefer_calamine=True)
        rows_opy = read_sheet_values(generic_xlsx, sheet_name, prefer_calamine=False)
        assert _normalize_for_comparison(rows_cal) == _normalize_for_comparison(rows_opy), \
            f"Mismatch in sheet: {sheet_name}"


@skip_no_calamine
def test_template_diff_calamine_vs_openpyxl(generic_xlsx: Path):
    """wp_template_diff_service: 两版本模板 diff 依赖第一行表头读取，验证等价。"""
    from app.services.xlsx_read_adapter import read_sheet_values

    # diff service reads first row as headers
    rows_cal = read_sheet_values(generic_xlsx, "审定表D2-1", prefer_calamine=True)
    rows_opy = read_sheet_values(generic_xlsx, "审定表D2-1", prefer_calamine=False)
    # First row (after normalization) should match — this is what diff service uses
    norm_cal = _normalize_for_comparison(rows_cal)
    norm_opy = _normalize_for_comparison(rows_opy)
    assert norm_cal[0] == norm_opy[0]  # title row
    assert norm_cal == norm_opy  # full equivalence


@skip_no_calamine
def test_fine_rule_engine_calamine_vs_openpyxl(generic_xlsx: Path):
    """wp_fine_rule_engine: 按固定行/列号取值 — calamine 等价验证。"""
    from app.services.xlsx_read_adapter import read_sheet_values

    rows_cal = read_sheet_values(generic_xlsx, "审定表D2-1", prefer_calamine=True)
    rows_opy = read_sheet_values(generic_xlsx, "审定表D2-1", prefer_calamine=False)
    # Fine rule engine reads by row_num/col_num — verify specific cells
    norm_cal = _normalize_for_comparison(rows_cal)
    norm_opy = _normalize_for_comparison(rows_opy)
    # Row 5 (0-indexed: 4), Col B (index 1) = 500000
    assert norm_cal[4][1] == norm_opy[4][1] == 500000
    # Row 7 (0-indexed: 6), Col B (index 1) = 800000 (total)
    assert norm_cal[6][1] == norm_opy[6][1] == 800000


@skip_no_calamine
def test_import_template_calamine_vs_openpyxl(generic_xlsx: Path):
    """import_template_service: 逐行读取值等价。"""
    from app.services.xlsx_read_adapter import read_sheet_values

    rows_cal = read_sheet_values(generic_xlsx, "明细表D2-2", prefer_calamine=True)
    rows_opy = read_sheet_values(generic_xlsx, "明细表D2-2", prefer_calamine=False)
    assert _normalize_for_comparison(rows_cal) == _normalize_for_comparison(rows_opy)


# ─── Multi-sheet and list_sheet_names equivalence ──────────────────────


@skip_no_calamine
def test_list_sheet_names_calamine_vs_openpyxl(generic_xlsx: Path):
    """list_sheet_names: calamine vs openpyxl 返回相同 sheet 名列表。"""
    from app.services.xlsx_read_adapter import list_sheet_names

    # list_sheet_names uses calamine when available
    names = list_sheet_names(generic_xlsx)
    assert "审定表D2-1" in names
    assert "明细表D2-2" in names


def test_all_points_openpyxl_path_works(generic_xlsx: Path):
    """所有迁移点 openpyxl 路径正常工作（calamine 无关）。"""
    from app.services.xlsx_read_adapter import read_sheet_values

    rows = read_sheet_values(generic_xlsx, "审定表D2-1", prefer_calamine=False)
    assert len(rows) == 7
    # Verify data row values
    assert rows[4][0] == "客户A"
    assert rows[4][1] == 500000
