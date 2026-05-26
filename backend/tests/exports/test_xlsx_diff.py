"""Property 2 PBT: 方案 C 字符级还原（导出 xlsx ≡ 致同模板手填）

**Validates: Requirements 4.3.1.a-g + 5.3**

For any (template, schema, html_data) 合法组合，wp_xlsx_export_service 的输出
xlsx 与"在原致同模板上手填同样数据"的 xlsx 在 cell value 维度上字符级 diff 为
空集（除用户填写位置标记的差异）。

三条 property（hypothesis max_examples=20，因 openpyxl I/O 慢）：

- Property 2a (确定性): 同 (template, schema, html_data) 输入 → 字节级恒等输出
  （除 xlsx zip metadata 时间戳外，cell value 集合恒等）
- Property 2b (字符级 diff): 修改 fixed_cells 后 → 仅修改的 cell 反映 html_data，
  其余 cell（含 formulas / merged_cells / static_text）值与原模板恒等
- Property 2c (空 html_data 等价模板): html_data 为 {} 时 → cell value 全集与原模板
  完全一致（除 fixed_cells 的占位符替换后留空字符串）

Spec: `.kiro/specs/workpaper-html-renderer/`
"""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from hypothesis import HealthCheck, given, settings as h_settings
from hypothesis import strategies as st
from openpyxl.workbook.workbook import Workbook

from app.services.wp_xlsx_export_service import _sync_export_workpaper_xlsx


# ─── 可重用 fixture：构建测试模板 + schema ───────────────────────────────────


# 测试模板共享布局（5×5 + 公式 + 合并单元格 + 静态文本 + dynamic_table 区域）：
#   A1:B1 (merged) = "致同模板表头"           (path 4: static_text + merged)
#   C1               = "${entity_name}"          (path 1: fixed_cells)
#   D1               = "${period_end}"           (path 1: fixed_cells)
#   A2               = "项目"  / B2 = "金额"     (path 4: static_text)
#   A3 / A4 / A5     = (空，dynamic_table 区域)  (path 2)
#   A6               = "合计"                    (path 4: static_text)
#   B6               = "=SUM(B3:B5)"             (path 3: formula 保留)


def _build_template_in_tmp(tmp_path: Path, name: str = "test_template.xlsx") -> Path:
    """创建测试用致同模板 xlsx（含公式 + 合并单元格 + 静态文本）。

    Returns:
        模板文件绝对路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # 静态文本 + 合并单元格（path 4）
    ws.merge_cells("A1:B1")
    ws["A1"] = "致同模板表头"
    ws["A2"] = "项目"
    ws["B2"] = "金额"
    ws["A6"] = "合计"

    # 占位符（path 1: fixed_cells）
    ws["C1"] = "${entity_name}"  # 模板原始占位符（被覆盖时验证）
    ws["D1"] = "${period_end}"

    # 公式（path 3: formula 保留不动）
    ws["B6"] = "=SUM(B3:B5)"

    template_file = tmp_path / name
    wb.save(str(template_file))
    wb.close()
    return template_file


def _build_schema(template_path: str) -> dict:
    """对应模板的 schema 定义（4 路径全覆盖）。"""
    return {
        "wp_code": "TEST",
        "template_path": template_path,
        "sheets": {
            "TestSheet": {
                # path 1: fixed_cells
                "fixed_cells": {
                    "C1": "${entity_name}",
                    "D1": "${period_end}",
                },
                # path 2: dynamic_table（A3:B5 区域写用户数据）
                "dynamic_table": {
                    "start_row": 3,
                    "columns": {
                        "A": {"field": "name", "type": "text"},
                        "B": {"field": "amount", "type": "number"},
                    },
                },
                # path 3 & 4 由 openpyxl 加载时自动保留，无需 schema 配置
            },
        },
    }


def _read_cell_values(buf: BytesIO) -> dict[str, object]:
    """从 BytesIO 读取所有 cell value（含公式字符串），返回 {coord: value} 字典。

    使用 data_only=False 保留公式字符串原样（=SUM(...) 而非缓存值）。
    """
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=False)
    ws = wb["TestSheet"]
    result: dict[str, object] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                result[cell.coordinate] = cell.value
    wb.close()
    return result


def _read_template_baseline(template_path: Path) -> dict[str, object]:
    """读取原始模板 cell 值（基准），用于 diff 对比。"""
    wb = openpyxl.load_workbook(str(template_path), data_only=False)
    ws = wb["TestSheet"]
    result: dict[str, object] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                result[cell.coordinate] = cell.value
    wb.close()
    return result


# ─── Hypothesis Strategies ───────────────────────────────────────────────────


# 简单 cell 值策略：纯文本 / 整数 / 浮点（不含公式开头 = 避免歧义）
st_simple_text = st.text(
    alphabet="abcABC一二三0123456789测试公司项目 ",
    min_size=0,
    max_size=20,
).filter(lambda s: not s.startswith("="))

st_simple_number = st.one_of(
    st.integers(min_value=-99999, max_value=99999),
    st.floats(
        min_value=-99999.0,
        max_value=99999.0,
        allow_nan=False,
        allow_infinity=False,
    ),
)


# 行字典策略：name 是文本，amount 是数字
st_dynamic_row = st.fixed_dictionaries({
    "name": st_simple_text,
    "amount": st_simple_number,
})


# html_data 策略：subset of {TestSheet: {rows: [...]}}（0~3 行随机）
st_html_data = st.fixed_dictionaries({
    "TestSheet": st.fixed_dictionaries({
        "rows": st.lists(st_dynamic_row, min_size=0, max_size=3),
    }),
})


# project_meta 策略：entity_name 必填 + period_end 必填
st_project_meta = st.fixed_dictionaries({
    "entity_name": st.text(
        alphabet="abcABC一二三测试公司有限责任",
        min_size=1,
        max_size=15,
    ),
    "period_end": st.sampled_from([
        "2024-12-31", "2025-06-30", "2025-12-31", "2026-03-31",
    ]),
})


# ─── Property 2a: 确定性（同输入 → 同输出 cell 集合） ─────────────────────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(html_data=st_html_data, project_meta=st_project_meta)
def test_property_2a_deterministic_export(
    tmp_path_factory: pytest.TempPathFactory,
    html_data: dict,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.a-g** — 同 (template, schema, html_data) 两次导出 cell 值集合恒等。

    导出 1 次 vs 导出 2 次：
      - 所有 cell 的 (coordinate, value) 字典严格相等
      - 字节流可能因 zip metadata 时间戳不同（不验证字节恒等）
    """
    tmp_path = tmp_path_factory.mktemp("prop2a")
    template_path = _build_template_in_tmp(tmp_path)
    schema = _build_schema(str(template_path))

    buf1 = _sync_export_workpaper_xlsx(schema, html_data, project_meta)
    buf2 = _sync_export_workpaper_xlsx(schema, html_data, project_meta)

    cells1 = _read_cell_values(buf1)
    cells2 = _read_cell_values(buf2)

    assert cells1 == cells2, (
        f"两次导出 cell 集合不一致：\n"
        f"  diff (only in 1): {set(cells1.items()) - set(cells2.items())}\n"
        f"  diff (only in 2): {set(cells2.items()) - set(cells1.items())}"
    )


# ─── Property 2b: 字符级 diff（仅修改的 cell 反映 html_data） ─────────────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(html_data=st_html_data, project_meta=st_project_meta)
def test_property_2b_char_level_diff_only_at_user_positions(
    tmp_path_factory: pytest.TempPathFactory,
    html_data: dict,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.a-g + 5.3** — 字符级 diff 仅在用户填写位置出现。

    导出 xlsx vs 原模板：
      - 公式 cell B6（path 3）: '=SUM(B3:B5)' 字符串恒等
      - 静态 cell A1/A2/B2/A6（path 4）: 值与模板恒等
      - 合并单元格集合恒等（A1:B1）
      - fixed_cells C1/D1（path 1）: 反映 project_meta 替换后的值
      - dynamic_table A3:B5（path 2）: 反映 html_data['TestSheet']['rows'] 内容
    """
    tmp_path = tmp_path_factory.mktemp("prop2b")
    template_path = _build_template_in_tmp(tmp_path)
    schema = _build_schema(str(template_path))

    template_cells = _read_template_baseline(template_path)
    out = _sync_export_workpaper_xlsx(schema, html_data, project_meta)
    out_cells = _read_cell_values(out)

    # ─── path 3: 公式 cell 字符串恒等 ─────────────────────────
    assert out_cells.get("B6") == "=SUM(B3:B5)", (
        f"公式 B6 未保留，期望 '=SUM(B3:B5)'，实际 {out_cells.get('B6')!r}"
    )
    assert out_cells.get("B6") == template_cells.get("B6"), (
        "公式 cell 与模板不恒等（违反 path 3 跳过策略）"
    )

    # ─── path 4: 静态文本 cell 与模板恒等 ──────────────────────
    static_coords = ["A1", "A2", "B2", "A6"]
    for coord in static_coords:
        assert out_cells.get(coord) == template_cells.get(coord), (
            f"静态 cell {coord} 被错误覆盖：模板={template_cells.get(coord)!r} "
            f"导出={out_cells.get(coord)!r}"
        )

    # ─── 合并单元格恒等（A1:B1）─────────────────────────────────
    out.seek(0)
    wb_out = openpyxl.load_workbook(out, data_only=False)
    out_merged = sorted(str(r) for r in wb_out["TestSheet"].merged_cells.ranges)
    wb_out.close()

    wb_tpl = openpyxl.load_workbook(str(template_path), data_only=False)
    tpl_merged = sorted(str(r) for r in wb_tpl["TestSheet"].merged_cells.ranges)
    wb_tpl.close()

    assert out_merged == tpl_merged, (
        f"merged_ranges 不恒等：\n  模板: {tpl_merged}\n  导出: {out_merged}"
    )

    # ─── path 1: fixed_cells 反映 project_meta ────────────────
    assert out_cells.get("C1") == project_meta["entity_name"], (
        f"fixed_cell C1 未反映 entity_name={project_meta['entity_name']!r}，"
        f"实际 {out_cells.get('C1')!r}"
    )
    assert out_cells.get("D1") == project_meta["period_end"], (
        f"fixed_cell D1 未反映 period_end={project_meta['period_end']!r}，"
        f"实际 {out_cells.get('D1')!r}"
    )

    # ─── path 2: dynamic_table 反映 html_data ──────────────────
    rows = html_data["TestSheet"]["rows"]
    for i, row in enumerate(rows):
        excel_row = 3 + i
        # name 列（A）
        a_coord = f"A{excel_row}"
        expected_a = row["name"] if row["name"] is not None else ""
        # 空字符串 / None 时不写入到 out_cells（_read_cell_values 跳过 None）
        if expected_a == "":
            assert out_cells.get(a_coord) in (None, ""), (
                f"dynamic_table {a_coord} 应为空，实际 {out_cells.get(a_coord)!r}"
            )
        else:
            assert out_cells.get(a_coord) == expected_a, (
                f"dynamic_table {a_coord} 未反映 name={expected_a!r}，"
                f"实际 {out_cells.get(a_coord)!r}"
            )
        # amount 列（B）
        b_coord = f"B{excel_row}"
        expected_b = float(row["amount"])
        actual_b = out_cells.get(b_coord)
        if actual_b is not None:
            assert isinstance(actual_b, (int, float))
            # 允许浮点精度误差
            assert abs(float(actual_b) - expected_b) < 1e-6, (
                f"dynamic_table {b_coord} 未反映 amount={expected_b}，"
                f"实际 {actual_b!r}"
            )


# ─── Property 2c: 空 html_data 等价于模板（除 fixed_cells 占位符替换） ────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(project_meta=st_project_meta)
def test_property_2c_empty_html_data_equals_template(
    tmp_path_factory: pytest.TempPathFactory,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.a-g** — html_data 为空时输出 ≡ 模板（除 fixed_cells 替换）。

    html_data = {'TestSheet': {'rows': []}} → 不写 dynamic_table；
    导出 cell 值集合 = 模板 cell 值集合 \\ {C1, D1} ∪ {C1=entity, D1=period}
    （即只有 fixed_cells 两个位置被替换，其他位置完全恒等）。
    """
    tmp_path = tmp_path_factory.mktemp("prop2c")
    template_path = _build_template_in_tmp(tmp_path)
    schema = _build_schema(str(template_path))

    template_cells = _read_template_baseline(template_path)
    out = _sync_export_workpaper_xlsx(
        schema,
        {"TestSheet": {"rows": []}},
        project_meta,
    )
    out_cells = _read_cell_values(out)

    # 除 C1/D1 外，所有模板 cell 在导出中恒等
    for coord, tpl_val in template_cells.items():
        if coord in ("C1", "D1"):
            continue
        assert out_cells.get(coord) == tpl_val, (
            f"空 html_data 时 cell {coord} 与模板不恒等：\n"
            f"  模板: {tpl_val!r}\n"
            f"  导出: {out_cells.get(coord)!r}"
        )

    # C1/D1 反映 project_meta
    assert out_cells.get("C1") == project_meta["entity_name"]
    assert out_cells.get("D1") == project_meta["period_end"]

    # 不应在 dynamic_table 区域 (A3:B5) 留任何残留值
    for coord in ["A3", "A4", "A5", "B3", "B4", "B5"]:
        assert out_cells.get(coord) is None, (
            f"空 html_data 不应写入 dynamic_table 区域 {coord}，"
            f"实际 {out_cells.get(coord)!r}"
        )
