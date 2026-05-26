"""Property 3 PBT: 公式与合并单元格保留不变量

**Validates: Requirements 4.3.1.b + 4.3.1.c**

For any 模板 xlsx 加载后的 (formula_cells, merged_ranges) 集合，导出 xlsx 重新
加载后的对应集合 SHALL 与之恒等（即 export 是公式与合并区的恒等映射）。

四条 property（hypothesis max_examples=20，因 openpyxl I/O 慢）：

- Property 3a (formulas 保留): 任意 html_data，导出 → 重载后 formula_cells 集合
  （cells where value starts with '='）= 模板 formula_cells 集合
- Property 3b (merged_ranges 恒等): 任意 html_data，导出 → 重载后 merged_ranges
  排序列表 = 模板 merged_ranges 排序列表
- Property 3c (formulas 不被 dynamic_table 覆盖): 即使 dynamic_table 试图写入到
  公式 cell 地址，公式仍被保留（service 路径 3 的 startswith('=') 保护）
- Property 3d (styles 保留): 至少 cell.fill.fgColor / cell.font.bold 与模板恒等
  （取样 1-2 个 cell 验证 openpyxl 加载-保存不破坏样式）

Spec: `.kiro/specs/workpaper-html-renderer/`
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from hypothesis import HealthCheck, given, settings as h_settings
from hypothesis import strategies as st
from openpyxl.styles import Font, PatternFill

from app.services.wp_xlsx_export_service import _sync_export_workpaper_xlsx


# ─── 测试模板构建 ────────────────────────────────────────────────────────────
#
# 布局（含多个公式 + 多个合并区域 + 样式 cell）：
#   A1:B1 (merged) = "标题" (font: bold)             (merged + style)
#   C1:D1 (merged) = "副标题"                        (merged)
#   A2 = "项目"                                      (style: red fill)
#   B2 = "金额"
#   A3 = "${entity_name}"                             (fixed_cells)
#   B3 = "=B6+B7"                                    (formula 1)
#   A4..A6 = (空，dynamic_table 写入区)
#   B4..B6 = (空，dynamic_table 写入区)
#   A7 = "合计"
#   B7 = "=SUM(B4:B6)"                               (formula 2)
#   D5 = "=A1"                                       (formula 3 - 跨列公式)


def _build_invariants_template(tmp_path: Path) -> Path:
    """创建含 3 个公式 + 2 个合并区域 + 2 处样式的测试模板。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InvSheet"

    # 合并区域 1
    ws.merge_cells("A1:B1")
    ws["A1"] = "标题"
    ws["A1"].font = Font(bold=True, name="宋体", size=12)

    # 合并区域 2
    ws.merge_cells("C1:D1")
    ws["C1"] = "副标题"

    # 样式 cell（红色填充）
    ws["A2"] = "项目"
    ws["A2"].fill = PatternFill(
        start_color="FFFF0000",
        end_color="FFFF0000",
        fill_type="solid",
    )
    ws["B2"] = "金额"

    # 占位符 + 公式
    ws["A3"] = "${entity_name}"
    ws["B3"] = "=B6+B7"  # formula 1（用户可能误以为这是数据区会被覆盖）

    # 静态文本 + 合计公式
    ws["A7"] = "合计"
    ws["B7"] = "=SUM(B4:B6)"  # formula 2

    # 跨列公式
    ws["D5"] = "=A1"  # formula 3

    template_file = tmp_path / "invariants_template.xlsx"
    wb.save(str(template_file))
    wb.close()
    return template_file


def _build_invariants_schema(template_path: str) -> dict:
    """对应模板的 schema（dynamic_table 故意覆盖到 B3 公式 cell 测试 path 3 保护）。"""
    return {
        "wp_code": "INV",
        "template_path": template_path,
        "sheets": {
            "InvSheet": {
                "fixed_cells": {
                    "A3": "${entity_name}",
                },
                # dynamic_table 起点 row=3，会覆盖到 B3（公式 cell）
                # path 3 保护应阻止覆盖
                "dynamic_table": {
                    "start_row": 3,
                    "columns": {
                        "A": {"field": "name", "type": "text"},
                        "B": {"field": "amount", "type": "number"},
                    },
                },
            },
        },
    }


# ─── Helpers: 提取不变量 ─────────────────────────────────────────────────────


def _extract_formula_cells(wb: openpyxl.Workbook, sheet_name: str) -> dict[str, str]:
    """提取所有公式 cell 的 {coord: formula_string} 字典。"""
    ws = wb[sheet_name]
    result: dict[str, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if (
                isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                result[cell.coordinate] = cell.value
    return result


def _extract_merged_ranges(wb: openpyxl.Workbook, sheet_name: str) -> list[str]:
    """提取所有合并区域字符串（如 'A1:B1'），排序后返回。"""
    ws = wb[sheet_name]
    return sorted(str(r) for r in ws.merged_cells.ranges)


def _extract_template_invariants(template_path: Path) -> tuple[dict[str, str], list[str]]:
    """从模板提取 (formula_cells, merged_ranges) 基准。"""
    wb = openpyxl.load_workbook(str(template_path), data_only=False)
    formulas = _extract_formula_cells(wb, "InvSheet")
    merged = _extract_merged_ranges(wb, "InvSheet")
    wb.close()
    return formulas, merged


def _extract_export_invariants(buf: BytesIO) -> tuple[dict[str, str], list[str]]:
    """从导出 BytesIO 提取 (formula_cells, merged_ranges)。"""
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=False)
    formulas = _extract_formula_cells(wb, "InvSheet")
    merged = _extract_merged_ranges(wb, "InvSheet")
    wb.close()
    return formulas, merged


# ─── Hypothesis Strategies ───────────────────────────────────────────────────


st_simple_text = st.text(
    alphabet="abcABC一二三0123456789测试 ",
    min_size=0,
    max_size=15,
).filter(lambda s: not s.startswith("="))

st_simple_number = st.one_of(
    st.integers(min_value=-9999, max_value=9999),
    st.floats(
        min_value=-9999.0,
        max_value=9999.0,
        allow_nan=False,
        allow_infinity=False,
    ),
)


st_dynamic_row = st.fixed_dictionaries({
    "name": st_simple_text,
    "amount": st_simple_number,
})


st_html_data = st.fixed_dictionaries({
    "InvSheet": st.fixed_dictionaries({
        "rows": st.lists(st_dynamic_row, min_size=0, max_size=4),
    }),
})


st_project_meta = st.fixed_dictionaries({
    "entity_name": st.text(
        alphabet="abcABC一二三测试公司有限",
        min_size=1,
        max_size=12,
    ),
    "period_end": st.sampled_from(["2024-12-31", "2025-06-30", "2025-12-31"]),
})


# ─── Property 3a: formulas 集合恒等 ───────────────────────────────────────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(html_data=st_html_data, project_meta=st_project_meta)
def test_property_3a_formulas_preserved(
    tmp_path_factory: pytest.TempPathFactory,
    html_data: dict,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.b** — 任意 html_data 导出后 formula_cells 集合恒等。

    导出 → 重载 → set of {coord: formula_string} == 模板 formula_cells set
    （包括 B3 即使被 dynamic_table 试图覆盖也应保留）。
    """
    tmp_path = tmp_path_factory.mktemp("prop3a")
    template_path = _build_invariants_template(tmp_path)
    schema = _build_invariants_schema(str(template_path))

    tpl_formulas, _ = _extract_template_invariants(template_path)
    out = _sync_export_workpaper_xlsx(schema, html_data, project_meta)
    out_formulas, _ = _extract_export_invariants(out)

    assert out_formulas == tpl_formulas, (
        f"formula_cells 集合不恒等（违反 Requirement 4.3.1.b）：\n"
        f"  模板: {tpl_formulas}\n"
        f"  导出: {out_formulas}\n"
        f"  diff (only in template): "
        f"{set(tpl_formulas.items()) - set(out_formulas.items())}\n"
        f"  diff (only in export): "
        f"{set(out_formulas.items()) - set(tpl_formulas.items())}"
    )


# ─── Property 3b: merged_ranges 集合恒等 ──────────────────────────────────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(html_data=st_html_data, project_meta=st_project_meta)
def test_property_3b_merged_ranges_preserved(
    tmp_path_factory: pytest.TempPathFactory,
    html_data: dict,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.c** — 任意 html_data 导出后 merged_ranges 排序列表恒等。

    导出 → 重载 → sorted(merged_ranges) == 模板 sorted(merged_ranges)
    """
    tmp_path = tmp_path_factory.mktemp("prop3b")
    template_path = _build_invariants_template(tmp_path)
    schema = _build_invariants_schema(str(template_path))

    _, tpl_merged = _extract_template_invariants(template_path)
    out = _sync_export_workpaper_xlsx(schema, html_data, project_meta)
    _, out_merged = _extract_export_invariants(out)

    assert out_merged == tpl_merged, (
        f"merged_ranges 集合不恒等（违反 Requirement 4.3.1.c）：\n"
        f"  模板: {tpl_merged}\n"
        f"  导出: {out_merged}"
    )


# ─── Property 3c: formulas 不被 dynamic_table 覆盖 ─────────────────────────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(html_data=st_html_data, project_meta=st_project_meta)
def test_property_3c_dynamic_table_does_not_overwrite_formulas(
    tmp_path_factory: pytest.TempPathFactory,
    html_data: dict,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.b** — dynamic_table 试图写入到公式 cell 时公式仍保留。

    模板 B3 = '=B6+B7'，schema 的 dynamic_table 起点是 row=3 列 B 写 amount，
    第一行用户数据会试图写到 B3（公式 cell）。path 3 startswith('=') 保护应生效。
    """
    tmp_path = tmp_path_factory.mktemp("prop3c")
    template_path = _build_invariants_template(tmp_path)
    schema = _build_invariants_schema(str(template_path))

    rows = html_data["InvSheet"]["rows"]
    if not rows:
        # 无数据时该 property 平凡成立
        return

    out = _sync_export_workpaper_xlsx(schema, html_data, project_meta)
    out_formulas, _ = _extract_export_invariants(out)

    # B3 必须保留为公式
    assert out_formulas.get("B3") == "=B6+B7", (
        f"dynamic_table 第一行 amount={rows[0]['amount']} 错误覆盖了 B3 公式，"
        f"实际 B3 = {out_formulas.get('B3')!r}（应保持 '=B6+B7'）"
    )

    # B7 (=SUM(B4:B6)) 也必须保留
    assert out_formulas.get("B7") == "=SUM(B4:B6)", (
        f"B7 SUM 公式被破坏，实际 {out_formulas.get('B7')!r}"
    )


# ─── Property 3d: styles 保留（fill / font 恒等抽样） ──────────────────────────


@h_settings(max_examples=20, deadline=None,
            suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture])
@given(html_data=st_html_data, project_meta=st_project_meta)
def test_property_3d_styles_preserved(
    tmp_path_factory: pytest.TempPathFactory,
    html_data: dict,
    project_meta: dict,
) -> None:
    """**Validates: Requirements 4.3.1.d** — cell.fill / cell.font 抽样保留。

    抽样 A1（font.bold=True）和 A2（fill 红色）：导出后样式属性与模板恒等。
    """
    tmp_path = tmp_path_factory.mktemp("prop3d")
    template_path = _build_invariants_template(tmp_path)
    schema = _build_invariants_schema(str(template_path))

    out = _sync_export_workpaper_xlsx(schema, html_data, project_meta)
    out.seek(0)
    wb_out = openpyxl.load_workbook(out, data_only=False)
    ws_out = wb_out["InvSheet"]

    # A1 font.bold 必须为 True
    assert ws_out["A1"].font.bold is True, (
        f"A1 font.bold 未保留：实际 {ws_out['A1'].font.bold!r}"
    )

    # A2 fill 必须是红色（FFFF0000）
    a2_fill = ws_out["A2"].fill
    a2_color = (
        a2_fill.fgColor.rgb if a2_fill.fgColor and a2_fill.fgColor.rgb else None
    )
    assert a2_color == "FFFF0000", (
        f"A2 fill 颜色未保留：实际 {a2_color!r}（期望 'FFFF0000'）"
    )

    wb_out.close()
