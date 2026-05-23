"""proposal-remaining-18 D-1 大底稿懒加载 — 后端单测

覆盖：
- GET /xlsx-to-json?sheets=active：仅返回 active sheet 完整 cellData，
  其余 sheet 仅返回 {id, name, rowCount, columnCount, cellData={}, custom._lazy=True}
- GET /sheet/{sheet_name}：按需返回单个 sheet 完整数据
- GET /sheet/{sheet_name} 404：sheet 不存在返回 404
- GET /xlsx-to-json（不传 sheets 参数）：向后兼容，返回全部 sheet 完整数据
- 首屏 JSON 体积下降基准（active vs 全量）

Validates: spec proposal-remaining-18 task 0.4
"""

from __future__ import annotations

import io
import json
import uuid
from pathlib import Path
from unittest.mock import patch

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.base import Base

SQLiteTypeCompiler.visit_JSONB = SQLiteTypeCompiler.visit_JSON

TEST_DATABASE_URL = "sqlite+aiosqlite:///:memory:"
test_engine = create_async_engine(TEST_DATABASE_URL, echo=False)


def _make_multi_sheet_xlsx(path: Path, num_sheets: int = 5, rows_per_sheet: int = 20) -> None:
    """生成包含 num_sheets 个 sheet 的 xlsx 文件，模拟大底稿。

    每个 sheet 填 rows_per_sheet 行 × 6 列数据，便于验证 cellData 体积差异。
    """
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)
    for idx in range(num_sheets):
        ws = wb.create_sheet(f"Sheet_{idx}")
        for r in range(1, rows_per_sheet + 1):
            for c in range(1, 7):
                ws.cell(row=r, column=c, value=f"S{idx}-R{r}-C{c}")
    # 第一个 sheet 设为 active
    wb.active = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()


@pytest_asyncio.fixture
async def db_session() -> AsyncSession:
    async with test_engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(test_engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as session:
        yield session


@pytest_asyncio.fixture
async def lazy_load_setup(tmp_path, db_session: AsyncSession):
    """准备：项目 + 底稿 + xlsx 文件 + 注入路径解析 patch。

    返回 (client, project_id, wp_id, xlsx_path)。
    """
    from app.core.database import get_db
    from app.deps import get_current_user, require_project_access
    from app.main import app

    project_id = uuid.uuid4()
    wp_id = uuid.uuid4()

    # 生成 5 sheet × 20 行 × 6 列 的 xlsx
    xlsx_path = tmp_path / f"{wp_id}.xlsx"
    _make_multi_sheet_xlsx(xlsx_path, num_sheets=5, rows_per_sheet=20)

    # 建最小化 wp_index + working_paper（端点查询 i.wp_code）
    await db_session.execute(text("""
        INSERT INTO wp_index (id, project_id, wp_code, wp_name, audit_cycle, status, is_deleted)
        VALUES (:id, :pid, :code, :name, 'D', 'todo', 0)
    """), {"id": str(uuid.uuid4()), "pid": str(project_id), "code": "TEST1", "name": "test"})
    wp_idx_row = (await db_session.execute(text(
        "SELECT id FROM wp_index WHERE project_id = :pid"
    ), {"pid": str(project_id)})).first()
    await db_session.execute(text("""
        INSERT INTO working_paper (
            id, project_id, wp_index_id, status, file_version,
            file_path, source_type, is_deleted
        )
        VALUES (:id, :pid, :wpi, 'draft', 1, :fp, 'manual', 0)
    """), {
        "id": str(wp_id),
        "pid": str(project_id),
        "wpi": str(wp_idx_row[0]),
        "fp": str(xlsx_path),
    })
    await db_session.commit()

    async def override_get_db():
        yield db_session

    class _FakeUser:
        id = uuid.uuid4()

        class _Role:
            value = "admin"

        role = _Role()

    async def override_user():
        return _FakeUser()

    def override_require_project_access(*_a, **_kw):
        async def _dep():
            return _FakeUser()
        return _dep

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[get_current_user] = override_user

    # patch storage path resolution → 返回我们生成的 xlsx
    with patch(
        "app.routers.wp_template_files.get_workpaper_file",
        return_value=xlsx_path,
    ):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            yield client, project_id, wp_id, xlsx_path

    app.dependency_overrides.clear()


def _unwrap(resp_json: dict) -> dict:
    """全局 response wrapper：统一拆出 data 字段（如有）"""
    if isinstance(resp_json, dict) and "code" in resp_json and "data" in resp_json:
        return resp_json["data"]
    return resp_json


# ─────────────────────────────────────────────────────────────────────────────
# 测试组 1：sheets=active 懒加载行为
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_xlsx_to_json_lazy_active_only_returns_active_full(lazy_load_setup):
    """sheets=active：仅 active sheet 有完整 cellData，其余 cellData 为空且带 _lazy 标记"""
    client, project_id, wp_id, _ = lazy_load_setup
    resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/xlsx-to-json",
        params={"sheets": "active"},
    )
    assert resp.status_code == 200, resp.text
    data = _unwrap(resp.json())
    assert "sheets" in data
    sheets = data["sheets"]
    assert len(sheets) == 5  # 元数据全部返回，确保前端可渲染左侧 sheet 导航

    # 至少一个 sheet 有完整 cellData（active sheet）
    full_sheets = [s for s in sheets.values() if s.get("cellData")]
    lazy_sheets = [s for s in sheets.values() if s.get("custom", {}).get("_lazy") is True]
    assert len(full_sheets) == 1, "懒加载下应只有 active sheet 完整加载"
    assert len(lazy_sheets) == 4, "其余 4 个 sheet 应仅返回元数据"

    # 懒加载 sheet 必须含 name + rowCount + columnCount，方便前端建立 sheet 导航
    for s in lazy_sheets:
        assert s["name"]
        assert s["rowCount"] >= 100
        assert s["columnCount"] >= 26
        assert s["cellData"] == {}


@pytest.mark.asyncio
async def test_xlsx_to_json_active_payload_smaller_than_full(lazy_load_setup):
    """首屏 JSON 体积基准：sheets=active 序列化后体积 < 全量 50%"""
    client, project_id, wp_id, _ = lazy_load_setup

    full_resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/xlsx-to-json"
    )
    active_resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/xlsx-to-json",
        params={"sheets": "active"},
    )
    assert full_resp.status_code == 200
    assert active_resp.status_code == 200

    full_size = len(json.dumps(_unwrap(full_resp.json())))
    active_size = len(json.dumps(_unwrap(active_resp.json())))
    # active 体积应明显小于 full（5 sheet 中 1 个含数据 → 期望 < 50%）
    assert active_size < full_size * 0.5, (
        f"懒加载体积未按预期下降：active={active_size}, full={full_size}, "
        f"占比={active_size / full_size:.1%}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# 测试组 2：/sheet/{sheet_name} 按需加载
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_get_sheet_by_name_returns_full_data(lazy_load_setup):
    """按 sheet 名加载单个 sheet，返回完整 cellData"""
    client, project_id, wp_id, _ = lazy_load_setup
    resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/sheet/Sheet_2"
    )
    assert resp.status_code == 200, resp.text
    data = _unwrap(resp.json())
    assert data["name"] == "Sheet_2"
    assert data["id"] == "sheet2"
    # cellData 应含 20 行数据
    assert len(data["cellData"]) >= 20
    # 验证某个具体 cell 内容
    row0 = data["cellData"]["0"] if "0" in data["cellData"] else data["cellData"][0]
    assert any("S2-R1" in str(c.get("v", "")) for c in row0.values())


@pytest.mark.asyncio
async def test_get_sheet_by_name_404_when_sheet_not_exists(lazy_load_setup):
    """sheet 名不存在时返回 404"""
    client, project_id, wp_id, _ = lazy_load_setup
    resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/sheet/NonExistent"
    )
    assert resp.status_code == 404
    body = resp.json()
    # 全局 wrapper 把 detail 改为 message 字段
    msg = body.get("detail") or body.get("message", "")
    assert "sheet 不存在" in msg


# ─────────────────────────────────────────────────────────────────────────────
# 测试组 3：向后兼容（无 sheets 参数返回全量）
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_xlsx_to_json_default_returns_full_for_all_sheets(lazy_load_setup):
    """缺省（不传 sheets）：所有 sheet 都返回完整 cellData，无 _lazy 标记"""
    client, project_id, wp_id, _ = lazy_load_setup
    resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/xlsx-to-json"
    )
    assert resp.status_code == 200
    data = _unwrap(resp.json())
    assert len(data["sheets"]) == 5
    for s in data["sheets"].values():
        # 全量模式下不应该有 _lazy 标记
        assert s.get("custom", {}).get("_lazy") is not True
        # 每个 sheet 都该有 cellData
        assert len(s["cellData"]) > 0


@pytest.mark.asyncio
async def test_xlsx_to_json_unknown_sheets_param_falls_back_to_full(lazy_load_setup):
    """未识别的 sheets 参数值（非 'active'）→ 退化为全量返回（向后兼容）"""
    client, project_id, wp_id, _ = lazy_load_setup
    resp = await client.get(
        f"/api/projects/{project_id}/workpapers/{wp_id}/template-file/xlsx-to-json",
        params={"sheets": "all"},  # 非约定值
    )
    assert resp.status_code == 200
    data = _unwrap(resp.json())
    # 所有 sheet 都该有 cellData（不进入懒加载分支）
    for s in data["sheets"].values():
        assert s.get("custom", {}).get("_lazy") is not True
