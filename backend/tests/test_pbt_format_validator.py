"""Property-Based Tests for FormatValidator (Properties 14-18)

Tests:
- Property 14: 扩展名与 MIME 不匹配时报 error 级
- Property 15: sheet 名与 render_schema 不匹配时报 error 级
- Property 16: required 字段为空时报 error 级
- Property 17: 数值列非数值内容报 warning 级
- Property 18: overall=最差级别，三类计数=len(items)正确分区

Testing framework: hypothesis
"""

from __future__ import annotations

import io
import tempfile
from pathlib import Path

from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import Workbook

from app.schemas.wp_export_schemas import (
    ValidationItem,
    ValidationLevel,
    ValidationReport,
)
from app.services.wp_export.format_validator import FormatValidator


# ─── Helpers ──────────────────────────────────────────────────────────────────


def _make_xlsx_bytes(sheet_names: list[str] | None = None, data: dict | None = None) -> bytes:
    """Create a valid xlsx file in memory and return bytes.

    Args:
        sheet_names: list of sheet names to create (default: ["Sheet"])
        data: dict mapping sheet_name -> list of (col_letter, row, value) tuples
    """
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    if sheet_names is None:
        sheet_names = ["Sheet"]

    for name in sheet_names:
        wb.create_sheet(title=name)

    if data:
        for sheet_name, cells in data.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for col_letter, row, value in cells:
                    ws[f"{col_letter}{row}"] = value

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_docx_bytes() -> bytes:
    """Create a minimal valid docx file in memory."""
    from docx import Document

    doc = Document()
    doc.add_paragraph("test")
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


# ─── Property 14: MIME Type Validation ────────────────────────────────────────


class TestProperty14MimeTypeValidation:
    """Property 14: 扩展名与 MIME 不匹配时报 error 级

    **Validates: Requirements 5.1**
    """

    @given(
        random_bytes=st.binary(min_size=10, max_size=100).filter(
            lambda b: not b.startswith(b"PK\x03\x04")
        )
    )
    @settings(max_examples=5)
    def test_non_zip_content_with_xlsx_extension_reports_error(
        self, random_bytes: bytes
    ) -> None:
        """Random non-ZIP bytes with .xlsx extension should report error.

        **Validates: Requirements 5.1**
        """
        validator = FormatValidator()
        report = validator.validate(random_bytes, "test.xlsx")

        assert report.overall == ValidationLevel.ERROR
        assert report.error_count > 0
        # At least one item about MIME mismatch
        mime_errors = [
            i for i in report.items
            if i.level == ValidationLevel.ERROR and "mime" in (i.field or "")
        ]
        assert len(mime_errors) > 0

    @given(
        random_bytes=st.binary(min_size=10, max_size=100).filter(
            lambda b: not b.startswith(b"PK\x03\x04")
        )
    )
    @settings(max_examples=5)
    def test_non_zip_content_with_docx_extension_reports_error(
        self, random_bytes: bytes
    ) -> None:
        """Random non-ZIP bytes with .docx extension should report error.

        **Validates: Requirements 5.1**
        """
        validator = FormatValidator()
        report = validator.validate(random_bytes, "report.docx")

        assert report.overall == ValidationLevel.ERROR
        assert report.error_count > 0
        mime_errors = [
            i for i in report.items
            if i.level == ValidationLevel.ERROR and "mime" in (i.field or "")
        ]
        assert len(mime_errors) > 0

    @settings(max_examples=5)
    @given(st.data())
    def test_xlsx_content_with_docx_extension_reports_error(
        self, data: st.DataObject
    ) -> None:
        """Valid xlsx content but named .docx should report error.

        **Validates: Requirements 5.1**
        """
        xlsx_bytes = _make_xlsx_bytes(["Sheet1"])
        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "wrong.docx")

        assert report.overall == ValidationLevel.ERROR
        assert report.error_count > 0

    @settings(max_examples=5)
    @given(st.data())
    def test_docx_content_with_xlsx_extension_reports_error(
        self, data: st.DataObject
    ) -> None:
        """Valid docx content but named .xlsx should report error.

        **Validates: Requirements 5.1**
        """
        docx_bytes = _make_docx_bytes()
        validator = FormatValidator()
        report = validator.validate(docx_bytes, "wrong.xlsx")

        assert report.overall == ValidationLevel.ERROR
        assert report.error_count > 0


# ─── Property 15: Sheet Structure Validation ──────────────────────────────────


class TestProperty15SheetStructureValidation:
    """Property 15: sheet 名与 render_schema 不匹配时报 error 级

    **Validates: Requirements 5.2**
    """

    @given(
        expected_sheet=st.text(
            alphabet=st.characters(whitelist_categories=("L", "N")),
            min_size=1,
            max_size=20,
        )
    )
    @settings(max_examples=5)
    def test_missing_expected_sheet_reports_error(
        self, expected_sheet: str
    ) -> None:
        """When file lacks a sheet expected by render_schema, report error.

        **Validates: Requirements 5.2**
        """
        # Create xlsx with only "OtherSheet"
        xlsx_bytes = _make_xlsx_bytes(["OtherSheet"])

        # Schema expects a different sheet name
        render_schema = {
            "sheets": {
                expected_sheet: {
                    "dynamic_table": {
                        "start_row": 2,
                        "columns": {"A": {"field": "code", "type": "text"}},
                    }
                }
            }
        }

        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "test.xlsx", render_schema)

        # If expected_sheet != "OtherSheet", should have error
        if expected_sheet != "OtherSheet":
            assert report.overall == ValidationLevel.ERROR
            assert report.error_count > 0
            structure_errors = [
                i for i in report.items
                if i.level == ValidationLevel.ERROR
                and i.field == "sheet_structure"
            ]
            assert len(structure_errors) > 0

    @given(
        sheet_names=st.lists(
            st.text(
                alphabet=st.characters(whitelist_categories=("L", "N")),
                min_size=1,
                max_size=15,
            ),
            min_size=2,
            max_size=4,
            unique=True,
        )
    )
    @settings(max_examples=5)
    def test_all_sheets_present_no_structure_error(
        self, sheet_names: list[str]
    ) -> None:
        """When file contains all expected sheets, no structure error.

        **Validates: Requirements 5.2**
        """
        xlsx_bytes = _make_xlsx_bytes(sheet_names)
        render_schema = {
            "sheets": {
                name: {"dynamic_table": {"start_row": 2, "columns": {}}}
                for name in sheet_names
            }
        }

        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "test.xlsx", render_schema)

        structure_errors = [
            i for i in report.items
            if i.field == "sheet_structure" and i.level == ValidationLevel.ERROR
        ]
        assert len(structure_errors) == 0


# ─── Property 16: Required Cell Validation ────────────────────────────────────


class TestProperty16RequiredCellValidation:
    """Property 16: required 字段为空时报 error 级

    **Validates: Requirements 5.3**
    """

    @given(
        num_rows=st.integers(min_value=1, max_value=3),
    )
    @settings(max_examples=5)
    def test_empty_required_cell_reports_error(self, num_rows: int) -> None:
        """Rows with data where required column is empty should report error.

        **Validates: Requirements 5.3**
        """
        # Create xlsx with data rows that have some content but required col empty
        sheet_name = "Data"
        data_cells: list[tuple[str, int, object]] = []
        start_row = 2

        for row_idx in range(start_row, start_row + num_rows):
            # Column B has data (non-required), Column A (required) is empty
            data_cells.append(("B", row_idx, f"value_{row_idx}"))
            # Column A explicitly left empty (not added)

        xlsx_bytes = _make_xlsx_bytes([sheet_name], {sheet_name: data_cells})

        render_schema = {
            "sheets": {
                sheet_name: {
                    "dynamic_table": {
                        "start_row": start_row,
                        "columns": {
                            "A": {"field": "code", "type": "text", "required": True},
                            "B": {"field": "name", "type": "text", "required": False},
                        },
                    }
                }
            }
        }

        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "test.xlsx", render_schema)

        assert report.overall == ValidationLevel.ERROR
        assert report.error_count >= num_rows
        # Each row missing required field -> one error per row
        required_errors = [
            i for i in report.items
            if i.level == ValidationLevel.ERROR and i.field == "code"
        ]
        assert len(required_errors) == num_rows

    @given(
        value=st.text(
            alphabet=st.characters(
                whitelist_categories=("L", "N", "P", "S"),
                blacklist_characters="\x00",
            ),
            min_size=1,
            max_size=10,
        ).filter(lambda s: s.strip() != ""),
    )
    @settings(max_examples=5)
    def test_filled_required_cell_no_error(self, value: str) -> None:
        """Required fields with non-empty values should not produce error.

        **Validates: Requirements 5.3**
        """
        sheet_name = "Data"
        start_row = 2
        data_cells = [("A", start_row, value), ("B", start_row, "info")]
        xlsx_bytes = _make_xlsx_bytes([sheet_name], {sheet_name: data_cells})

        render_schema = {
            "sheets": {
                sheet_name: {
                    "dynamic_table": {
                        "start_row": start_row,
                        "columns": {
                            "A": {"field": "code", "type": "text", "required": True},
                            "B": {"field": "name", "type": "text", "required": False},
                        },
                    }
                }
            }
        }

        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "test.xlsx", render_schema)

        required_errors = [
            i for i in report.items
            if i.level == ValidationLevel.ERROR and i.field == "code"
        ]
        assert len(required_errors) == 0


# ─── Property 17: Numeric Type Validation ─────────────────────────────────────


class TestProperty17NumericTypeValidation:
    """Property 17: 数值列非数值内容报 warning 级

    **Validates: Requirements 5.5**
    """

    @given(
        non_numeric_text=st.text(
            alphabet=st.characters(whitelist_categories=("L",)),
            min_size=1,
            max_size=10,
        )
    )
    @settings(max_examples=5)
    def test_non_numeric_in_number_column_reports_warning(
        self, non_numeric_text: str
    ) -> None:
        """Non-numeric text in a number-typed column should report warning.

        **Validates: Requirements 5.5**
        """
        sheet_name = "Data"
        start_row = 2
        # Put non-numeric text in a number column
        data_cells = [
            ("A", start_row, "code1"),
            ("B", start_row, non_numeric_text),  # This is the number column
        ]
        xlsx_bytes = _make_xlsx_bytes([sheet_name], {sheet_name: data_cells})

        render_schema = {
            "sheets": {
                sheet_name: {
                    "dynamic_table": {
                        "start_row": start_row,
                        "columns": {
                            "A": {"field": "code", "type": "text"},
                            "B": {"field": "amount", "type": "number"},
                        },
                    }
                }
            }
        }

        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "test.xlsx", render_schema)

        # Should be warning level (not error)
        warning_items = [
            i for i in report.items
            if i.level == ValidationLevel.WARNING and i.field == "amount"
        ]
        assert len(warning_items) > 0
        assert report.warning_count > 0

    @given(
        numeric_value=st.floats(
            min_value=-1e10, max_value=1e10, allow_nan=False, allow_infinity=False
        ),
    )
    @settings(max_examples=5)
    def test_numeric_in_number_column_no_warning(
        self, numeric_value: float
    ) -> None:
        """Numeric values in a number-typed column should not produce warning.

        **Validates: Requirements 5.5**
        """
        sheet_name = "Data"
        start_row = 2
        data_cells = [
            ("A", start_row, "code1"),
            ("B", start_row, numeric_value),
        ]
        xlsx_bytes = _make_xlsx_bytes([sheet_name], {sheet_name: data_cells})

        render_schema = {
            "sheets": {
                sheet_name: {
                    "dynamic_table": {
                        "start_row": start_row,
                        "columns": {
                            "A": {"field": "code", "type": "text"},
                            "B": {"field": "amount", "type": "number"},
                        },
                    }
                }
            }
        }

        validator = FormatValidator()
        report = validator.validate(xlsx_bytes, "test.xlsx", render_schema)

        # No warnings about the number column
        type_warnings = [
            i for i in report.items
            if i.level == ValidationLevel.WARNING and i.field == "amount"
        ]
        assert len(type_warnings) == 0


# ─── Property 18: Validation Report Structure ─────────────────────────────────


class TestProperty18ValidationReportStructure:
    """Property 18: overall=最差级别，三类计数=len(items)正确分区

    **Validates: Requirements 5.6**
    """

    @given(
        errors=st.lists(
            st.builds(
                ValidationItem,
                level=st.just(ValidationLevel.ERROR),
                location=st.just("test!A1"),
                message=st.text(min_size=1, max_size=20),
                field=st.just("f1"),
            ),
            min_size=0,
            max_size=3,
        ),
        warnings=st.lists(
            st.builds(
                ValidationItem,
                level=st.just(ValidationLevel.WARNING),
                location=st.just("test!B2"),
                message=st.text(min_size=1, max_size=20),
                field=st.just("f2"),
            ),
            min_size=0,
            max_size=3,
        ),
        passed_items=st.lists(
            st.builds(
                ValidationItem,
                level=st.just(ValidationLevel.PASSED),
                location=st.just("test!C3"),
                message=st.text(min_size=1, max_size=20),
                field=st.just("f3"),
            ),
            min_size=0,
            max_size=3,
        ),
    )
    @settings(max_examples=5)
    def test_report_structure_counts_and_overall(
        self,
        errors: list[ValidationItem],
        warnings: list[ValidationItem],
        passed_items: list[ValidationItem],
    ) -> None:
        """Report overall = worst level, counts match items partition.

        **Validates: Requirements 5.6**
        """
        all_items = errors + warnings + passed_items

        # Use the validator's internal _build_report to test structure
        validator = FormatValidator()
        report = validator._build_report(all_items)

        # Count invariants
        assert report.error_count == len(errors)
        assert report.warning_count == len(warnings)
        assert report.passed_count == len(passed_items)
        assert len(report.items) == len(all_items)

        # Overall level invariant
        if len(errors) > 0:
            assert report.overall == ValidationLevel.ERROR
        elif len(warnings) > 0:
            assert report.overall == ValidationLevel.WARNING
        else:
            assert report.overall == ValidationLevel.PASSED

    @given(
        num_errors=st.integers(min_value=1, max_value=5),
        num_warnings=st.integers(min_value=0, max_value=5),
    )
    @settings(max_examples=5)
    def test_overall_is_worst_level_with_mixed_items(
        self, num_errors: int, num_warnings: int
    ) -> None:
        """When both errors and warnings exist, overall is ERROR.

        **Validates: Requirements 5.6**
        """
        items: list[ValidationItem] = []
        for i in range(num_errors):
            items.append(
                ValidationItem(
                    level=ValidationLevel.ERROR,
                    location=f"S!A{i}",
                    message=f"err {i}",
                    field="x",
                )
            )
        for i in range(num_warnings):
            items.append(
                ValidationItem(
                    level=ValidationLevel.WARNING,
                    location=f"S!B{i}",
                    message=f"warn {i}",
                    field="y",
                )
            )

        validator = FormatValidator()
        report = validator._build_report(items)

        assert report.overall == ValidationLevel.ERROR
        assert report.error_count == num_errors
        assert report.warning_count == num_warnings
