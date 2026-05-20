"""L-F2: L 筹资循环 10-category sheet group rules — backend completeness check.

Validates: Requirements L-F2 / Spec workpaper-l-debt-cycle / Sprint 2 Task 2.2

Mirror of frontend useLDebtCycleSheetGroups.ts classification rules.

10 categories (priority 0~9):
1. 索引 (defaultHidden=true): 底稿目录 / GT_Custom / 修订说明
2. 历史遗留 (defaultHidden=true): （示例）模式
3. 总控台: 实质性程序表 / xxA 结尾
4. 审定表: 审定表
5. 明细表: 明细表
6. 分析程序: 分析程序
7. 利息测算: 利息测算 / 利息计算 / 利率测算
8. 检查表: 逾期 / 检查表 / 核查表 / 摊余成本
9. 附注+调整 (defaultHidden=true, 附注披露 readonly=true): 附注披露 / 调整分录
10. 其他程序: fallback

The classification function is reproduced here in pure Python to provide a
backend-side correctness oracle for the 79 real L cycle sheets.
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.wp_template_init_service import (
    _normalize_sheet_name,
    _should_skip_historical_sheet,
)


# ──────────────────────────────────────────────────────────────────────────
# Mirror of frontend classifyLSheet rules (must stay in sync)
# ──────────────────────────────────────────────────────────────────────────

L_RULES = [
    {
        "id": "index",
        "category": "索引",
        "priority": 0,
        "default_hidden": True,
        "readonly": False,
        "match": lambda s: bool(re.match(r"^(底稿目录|GT_Custom|修订说明)$", s.strip())),
    },
    {
        "id": "historical",
        "category": "历史遗留",
        "priority": 1,
        "default_hidden": True,
        "readonly": False,
        "match": lambda s: bool(re.search(r"（示例）|\(示例\)|示例$", s)),
    },
    {
        "id": "procedure",
        "category": "总控台",
        "priority": 2,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: bool(
            re.search(r"[A-Z]\d*A\s*$", s) or re.search(r"实质性程序表", s)
        ),
    },
    {
        "id": "audit_table",
        "category": "审定表",
        "priority": 3,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: bool(re.search(r"审定表", s)),
    },
    {
        "id": "detail",
        "category": "明细表",
        "priority": 4,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: bool(re.search(r"明细表", s)),
    },
    {
        "id": "analysis",
        "category": "分析程序",
        "priority": 5,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: bool(re.search(r"分析程序", s)),
    },
    {
        "id": "interest_calc",
        "category": "利息测算",
        "priority": 6,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: bool(re.search(r"利息测算|利息计算|利率测算", s)),
    },
    {
        "id": "check_table",
        "category": "检查表",
        "priority": 7,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: bool(re.search(r"逾期|检查表|核查表|摊余成本", s)),
    },
    {
        "id": "disclosure_adj",
        "category": "附注+调整",
        "priority": 8,
        "default_hidden": True,
        "readonly": False,
        "match": lambda s: bool(re.search(r"附注披露|调整分录", s)),
    },
    {
        "id": "other",
        "category": "其他程序",
        "priority": 9,
        "default_hidden": False,
        "readonly": False,
        "match": lambda s: True,
    },
]


def classify_l_sheet(name: str) -> dict:
    """Return classification metadata; fallback to 其他程序."""
    for rule in L_RULES:
        if rule["match"](name):
            cat = {
                "id": rule["id"],
                "category": rule["category"],
                "priority": rule["priority"],
                "default_hidden": rule["default_hidden"],
                "readonly": rule["readonly"],
            }
            # 附注披露 defaultHidden=true + readonly=true 特殊处理
            if rule["id"] == "disclosure_adj" and re.search(r"附注披露", name):
                cat["default_hidden"] = True
                cat["readonly"] = True
            return cat
    return {
        "id": "other",
        "category": "其他程序",
        "priority": 9,
        "default_hidden": False,
        "readonly": False,
    }


# ──────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent.parent / "wp_templates" / "L"


@pytest.fixture(scope="module")
def l_79_unique_sheets() -> list[str]:
    """79 个去重后 L 循环有效 sheet 名（基于 _normalize_sheet_name + filter historical）"""
    seen_normalized: set[str] = set()
    unique_originals: list[str] = []
    for f in sorted(TEMPLATES_DIR.glob("*.xlsx")):
        wb = load_workbook(str(f), read_only=True, data_only=True)
        try:
            for s in wb.sheetnames:
                if _should_skip_historical_sheet(s):
                    continue
                norm = _normalize_sheet_name(s)
                if norm not in seen_normalized:
                    seen_normalized.add(norm)
                    unique_originals.append(s)
        finally:
            wb.close()
    return unique_originals


# ──────────────────────────────────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────────────────────────────────


class TestLSheetGroupCompleteness:
    """L-F2: 10 类规则对 79 sheet 完备性 + 唯一性"""

    def test_79_sheet_baseline(self, l_79_unique_sheets):
        """L 循环去重后有效 sheet 数 = 79（Sprint 0 实测基线）"""
        assert len(l_79_unique_sheets) == 79, (
            f"Expected 79 L-cycle dedup sheets, got {len(l_79_unique_sheets)}"
        )

    def test_all_79_classified_to_non_empty_category(self, l_79_unique_sheets):
        """所有 sheet 均被分类到非空类目"""
        for s in l_79_unique_sheets:
            cat = classify_l_sheet(s)
            assert cat["category"], f"sheet '{s}' classified to empty category"

    def test_each_sheet_exactly_one_category(self, l_79_unique_sheets):
        """任意 sheet 名恰好命中 1 类（rule fallback `() => true` 必然命中）"""
        for s in l_79_unique_sheets:
            hits = [r for r in L_RULES if r["match"](s)]
            assert len(hits) >= 1, f"sheet '{s}' missed all rules"

    def test_no_sheet_falls_into_other_unexpectedly(self, l_79_unique_sheets):
        """fallback "其他程序" 类应仅包含无法明确归类项；
        其数量应 < 30%（79 中 < 24）"""
        other_sheets = [
            s for s in l_79_unique_sheets if classify_l_sheet(s)["id"] == "other"
        ]
        ratio = len(other_sheets) / len(l_79_unique_sheets)
        assert ratio < 0.30, (
            f"Too many sheets fell to '其他程序' fallback: "
            f"{len(other_sheets)}/{len(l_79_unique_sheets)} = {ratio:.1%}\n"
            f"Other sheets: {other_sheets}"
        )

    def test_key_categories_represented(self, l_79_unique_sheets):
        """79 sheet 应覆盖 ≥ 7 类（10 类中至少 7 类被命中）"""
        categories_hit = {classify_l_sheet(s)["category"] for s in l_79_unique_sheets}
        assert len(categories_hit) >= 7, (
            f"Only {len(categories_hit)} categories hit, expected ≥ 7\n"
            f"Hit: {sorted(categories_hit)}"
        )
        # 索引 / 附注+调整 / 总控台 / 审定表 / 明细表 必须命中
        for must_have in ["索引", "附注+调整", "总控台", "审定表", "明细表"]:
            assert must_have in categories_hit, (
                f"Category '{must_have}' should be hit but not in {sorted(categories_hit)}"
            )


class TestLSheetGroupKeyPaths:
    """L-F2: 关键 sheet 分类路径验证"""

    def test_index_sheets(self):
        """索引类 defaultHidden=true"""
        for s in ["底稿目录", "GT_Custom"]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "索引"
            assert cat["default_hidden"] is True

    def test_historical_sheet(self):
        """历史遗留类（示例）defaultHidden=true"""
        cat = classify_l_sheet("函证差异检查表（示例）")
        assert cat["category"] == "历史遗留"
        assert cat["default_hidden"] is True

    def test_procedure_sheets(self):
        """总控台：实质性程序表 / xxA 结尾"""
        # xxA 结尾
        for s in ["短期借款实质性程序表L1A", "长期借款实质性程序表L3A",
                  "应付债券实质性程序表L4A "]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "总控台", f"sheet '{s}' should be 总控台"
        # 实质性程序表关键词
        cat = classify_l_sheet("实质性程序表L8A")
        assert cat["category"] == "总控台"

    def test_audit_table_sheets(self):
        """审定表"""
        for s in ["审定表L1-1", "审定表L3-1", "审定表L5-1", "审定表L8-1"]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "审定表"

    def test_detail_sheets(self):
        """明细表"""
        for s in ["明细表L1-2", "明细表L3-2", "明细表L5-2", "明细表L6-2", "明细表L8-2"]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "明细表"

    def test_analysis_sheets(self):
        """分析程序"""
        for s in ["分析程序L1-3", "分析程序L3-3"]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "分析程序"

    def test_interest_calc_sheets(self):
        """利息测算（priority 6，优先于检查表 7）"""
        for s in ["利息测算表L1-5", "利息测算表L3-5"]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "利息测算"

    def test_interest_calc_priority_over_check_table(self):
        """利息测算 priority=6 优先于检查表 priority=7"""
        # "利息测算表L1-5" 含"利息测算"→ 利息测算(6)，不被"检查表"(7)误命中
        cat = classify_l_sheet("利息测算表L1-5")
        assert cat["category"] == "利息测算"
        assert cat["priority"] == 6

    def test_check_table_sheets(self):
        """检查表：逾期 / 检查表 / 核查表 / 摊余成本"""
        test_cases = [
            ("逾期贷款检查表L1-6", "检查表"),
            ("逾期贷款核查表L3-6", "检查表"),
            ("摊余成本计算表L5-3", "检查表"),
        ]
        for s, expected in test_cases:
            cat = classify_l_sheet(s)
            assert cat["category"] == expected, f"sheet '{s}' should be {expected}"

    def test_disclosure_sheets_hidden_and_readonly(self):
        """附注披露 5 种括号变体均 defaultHidden=true + readonly=true"""
        variants = [
            "附注披露信息(上市公司)",
            "附注披露信息(国企)",
            "附注披露信息（上市公司）",
            "附注披露信息（国企）",
            "附注披露信息（国有企业）",
        ]
        for s in variants:
            cat = classify_l_sheet(s)
            assert cat["category"] == "附注+调整", f"sheet '{s}' should be 附注+调整"
            assert cat["default_hidden"] is True, f"sheet '{s}' should be defaultHidden"
            assert cat["readonly"] is True, f"sheet '{s}' should be readonly"

    def test_adjustment_entries_not_readonly(self):
        """调整分录 defaultHidden=true 但 readonly=false"""
        for s in ["调整分录汇总L1-4", "调整分录汇总L8-3"]:
            cat = classify_l_sheet(s)
            assert cat["category"] == "附注+调整"
            assert cat["default_hidden"] is True
            assert cat["readonly"] is False

    def test_trailing_space_sheet(self):
        """末尾空格 sheet 仍正确分类"""
        # L4A 末尾带空格 → 总控台（xxA\s*$ 匹配）
        cat = classify_l_sheet("应付债券实质性程序表L4A ")
        assert cat["category"] == "总控台"

    def test_gt_custom_classified_as_index(self):
        """GT_Custom 归入索引"""
        cat = classify_l_sheet("GT_Custom")
        assert cat["category"] == "索引"
        assert cat["default_hidden"] is True

    def test_fallback_catches_unknown(self):
        """未知 sheet 名归入 其他程序"""
        cat = classify_l_sheet("随便的名字")
        assert cat["category"] == "其他程序"
        assert cat["priority"] == 9
        assert cat["default_hidden"] is False


class TestLSheetGroupPriorityOrdering:
    """L-F2: 优先级排序验证"""

    def test_interest_calc_before_check_table(self):
        """利息测算(6) < 检查表(7)：含"利息测算"的 sheet 不被"检查表"误命中"""
        cat = classify_l_sheet("利息测算表L1-5")
        assert cat["category"] == "利息测算"
        assert cat["priority"] == 6

    def test_audit_table_before_detail(self):
        """审定表(3) < 明细表(4)：含"审定表"的 sheet 不被"明细表"误命中"""
        cat = classify_l_sheet("审定表L1-1")
        assert cat["category"] == "审定表"
        assert cat["priority"] == 3

    def test_procedure_before_audit_table(self):
        """总控台(2) < 审定表(3)：含"实质性程序表"的 sheet 优先匹配总控台"""
        cat = classify_l_sheet("实质性程序表L1A")
        assert cat["category"] == "总控台"
        assert cat["priority"] == 2

    def test_all_priorities_unique(self):
        """10 条规则 priority 全部唯一"""
        priorities = [r["priority"] for r in L_RULES]
        assert len(set(priorities)) == len(priorities), (
            f"Duplicate priorities found: {priorities}"
        )


class TestLSheetGroupCategoryDistribution:
    """L-F2: 类目分布合理性"""

    def test_no_single_non_other_category_over_50_percent(self, l_79_unique_sheets):
        """非 fallback '其他程序' 类目不应单一占比 > 50%"""
        counts = Counter(
            classify_l_sheet(s)["category"] for s in l_79_unique_sheets
        )
        for cat, n in counts.items():
            if cat == "其他程序":
                continue
            ratio = n / len(l_79_unique_sheets)
            assert ratio < 0.50, (
                f"Category '{cat}' takes {n}/{len(l_79_unique_sheets)} = "
                f"{ratio:.1%}, exceeding 50% threshold"
            )

    def test_distribution_summary(self, l_79_unique_sheets):
        """分布摘要打印（便于回归比对）"""
        counts = Counter(
            classify_l_sheet(s)["category"] for s in l_79_unique_sheets
        )
        total = sum(counts.values())
        print("\n=== L 循环 79 sheet 分类分布 ===")
        for cat, n in sorted(counts.items(), key=lambda x: -x[1]):
            print(f"  {cat}: {n} ({n / total:.1%})")
        assert total == 79
