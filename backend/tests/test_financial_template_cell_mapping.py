"""Task 12 (audit-report-template-integration): cell_mapping 补全 + :parent 导出.

两部分：
1. **模板占位结构断言**（4 变体物理 xlsx）：
   - 每变体 BS 主表有 ``{{row:...:current}}``；
   - BS 续表也有占位（流动负债/所有者权益）；
   - consolidated BS 公司列有 ``:parent`` 占位；standalone 无；
   - 占位绝不落在 附注列（col2）或指引提示文本格（whole-cell only）；
   - 公式格未被占位污染。
2. **Exporter ``:parent`` 解析**（in-memory SQLite）：
   - consolidated 项目 + 「上级代码」匹配的 standalone 母公司项目（带 FinancialReport）
     → 公司列填入母公司个别值，合并列填本项目值；
   - 母公司缺失 → 公司列留空（不崩）。
"""
from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from decimal import Decimal

import openpyxl
import pytest
import pytest_asyncio
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

from app.models.base import Base
from app.models.core import Project
from app.models.report_models import FinancialReport, FinancialReportType
from app.services.report_excel_exporter import ReportExcelExporter
from app.services.template_manifest_loader import resolve_template_base_dir

_FS_DIR = resolve_template_base_dir() / "financial_statements"

_ROW_RE = re.compile(r"\{\{row:[^}]+\}\}")
_ROW_CURRENT_RE = re.compile(r"\{\{row:[^:}]+:current(?::parent)?\}\}")
_PARENT_RE = re.compile(r"\{\{row:[^:}]+:(?:current|prior):parent\}\}")
_WHOLE_RE = re.compile(r"^\{\{row:[^}]+\}\}$")

_VARIANTS = [
    "soe_standalone",
    "soe_consolidated",
    "listed_standalone",
    "listed_consolidated",
]


def _bs_sheets(wb) -> list[str]:
    return [
        n for n in wb.sheetnames
        if "资产负债表" in n and n != "GT_Custom"
    ]


def _variant_path(variant: str):
    return _FS_DIR / f"{variant}.xlsx"


# ---------------------------------------------------------------------------
# Part 1: template placeholder structure
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("variant", _VARIANTS)
def test_bs_main_has_current_placeholder(variant):
    """(a) 每变体 BS 主表至少一个 {{row:...:current}}."""
    path = _variant_path(variant)
    if not path.is_file():
        pytest.skip(f"{variant}.xlsx absent")
    wb = load_workbook(path, data_only=False)
    main_sheets = [n for n in _bs_sheets(wb) if "续" not in n]
    assert main_sheets, f"{variant}: no BS main sheet"
    ws = wb[main_sheets[0]]
    found = any(
        isinstance(c.value, str) and _ROW_CURRENT_RE.search(c.value)
        for row in ws.iter_rows() for c in row
    )
    assert found, f"{variant}: BS main sheet missing :current placeholder"


@pytest.mark.parametrize("variant", _VARIANTS)
def test_bs_continuation_has_placeholders(variant):
    """(b) BS 续表也有占位（负债/权益行）."""
    path = _variant_path(variant)
    if not path.is_file():
        pytest.skip(f"{variant}.xlsx absent")
    wb = load_workbook(path, data_only=False)
    cont = [n for n in _bs_sheets(wb) if "续" in n]
    assert cont, f"{variant}: no BS continuation sheet"
    ws = wb[cont[0]]
    count = sum(
        1 for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and _ROW_RE.search(c.value)
    )
    assert count > 0, f"{variant}: BS 续表 has no {{row:}} placeholders"


@pytest.mark.parametrize("variant", ["soe_consolidated", "listed_consolidated"])
def test_consolidated_bs_has_parent_placeholders(variant):
    """(c) consolidated BS 公司列有 :parent 占位."""
    path = _variant_path(variant)
    if not path.is_file():
        pytest.skip(f"{variant}.xlsx absent")
    wb = load_workbook(path, data_only=False)
    main = [n for n in _bs_sheets(wb) if "续" not in n][0]
    ws = wb[main]
    parent_found = any(
        isinstance(c.value, str) and _PARENT_RE.search(c.value)
        for row in ws.iter_rows() for c in row
    )
    assert parent_found, f"{variant}: BS main has no :parent placeholder"


@pytest.mark.parametrize("variant", ["soe_standalone", "listed_standalone"])
def test_standalone_bs_has_no_parent_placeholders(variant):
    """(c') standalone BS 无 :parent 占位."""
    path = _variant_path(variant)
    if not path.is_file():
        pytest.skip(f"{variant}.xlsx absent")
    wb = load_workbook(path, data_only=False)
    for name in _bs_sheets(wb):
        ws = wb[name]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and _PARENT_RE.search(c.value):
                    pytest.fail(
                        f"{variant}/{name}!{c.coordinate} unexpected :parent"
                    )


@pytest.mark.parametrize("variant", _VARIANTS)
def test_no_placeholder_in_note_or_guidance(variant):
    """(d) 占位绝不落在附注列(col2)或与指引提示文本混排的格."""
    path = _variant_path(variant)
    if not path.is_file():
        pytest.skip(f"{variant}.xlsx absent")
    wb = load_workbook(path, data_only=False)
    for name in wb.sheetnames:
        if name == "GT_Custom":
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str) or not _ROW_RE.search(c.value):
                    continue
                # col2 = 附注 → must never hold a row placeholder
                assert c.column != 2, (
                    f"{variant}/{name}!{c.coordinate} placeholder in 附注 col2"
                )
                # whole-cell only (not mixed with guidance hint text)
                assert _WHOLE_RE.fullmatch(c.value.strip()), (
                    f"{variant}/{name}!{c.coordinate} placeholder mixed "
                    f"with text: {c.value[:40]!r}"
                )


@pytest.mark.parametrize("variant", _VARIANTS)
def test_formula_cells_not_polluted(variant):
    """(e) 公式格不含占位符（=SUM 合计行未被覆盖）."""
    path = _variant_path(variant)
    if not path.is_file():
        pytest.skip(f"{variant}.xlsx absent")
    wb = load_workbook(path, data_only=False)
    for name in wb.sheetnames:
        if name == "GT_Custom":
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    assert "{{" not in v, (
                        f"{variant}/{name}!{c.coordinate} formula has placeholder"
                    )


def test_cell_mapping_consolidated_has_parent_coords():
    """cell_mapping.json consolidated 变体行含 current_parent/prior_parent 坐标."""
    import json

    path = resolve_template_base_dir() / "cell_mapping.json"
    if not path.is_file():
        pytest.skip("cell_mapping.json absent")
    data = json.loads(path.read_text(encoding="utf-8"))
    variants = data.get("variants", {})

    # consolidated: at least one row with parent coords
    for v in ("soe_consolidated", "listed_consolidated"):
        rows = variants.get(v, {}).get("rows", {})
        has_parent = any(
            isinstance(info, dict)
            and (info.get("current_parent") or info.get("prior_parent"))
            for info in rows.values()
        )
        assert has_parent, f"{v}: cell_mapping rows missing parent coords"

    # standalone: NO parent coords
    for v in ("soe_standalone", "listed_standalone"):
        rows = variants.get(v, {}).get("rows", {})
        any_parent = any(
            isinstance(info, dict)
            and (info.get("current_parent") or info.get("prior_parent"))
            for info in rows.values()
        )
        assert not any_parent, f"{v}: unexpected parent coords in standalone"


# ---------------------------------------------------------------------------
# Part 2: exporter :parent resolution (in-memory SQLite)
# ---------------------------------------------------------------------------

_engine = create_async_engine("sqlite+aiosqlite:///:memory:")
_async_session = sessionmaker(_engine, class_=AsyncSession, expire_on_commit=False)

_BS_SHEET = "1,2-资产负债表(企财01表）"


@pytest_asyncio.fixture
async def db_session():
    async with _engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    async with _async_session() as session:
        yield session
    async with _engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)


def _synthetic_consolidated_template(tmp_path):
    """合并 BS 4 列：C=合并期末 D=公司期末 E=合并期初 F=公司期初."""
    wb = Workbook()
    ws = wb.active
    ws.title = _BS_SHEET
    ws["A2"] = "{{period_end_date}}"
    ws["A3"] = "编制单位：{{company_full_name}}"
    ws["A4"] = "项目"
    ws["C4"] = "期末余额"
    ws["E4"] = "期初余额"
    ws["C5"] = "合并"
    ws["D5"] = "公司"
    ws["E5"] = "合并"
    ws["F5"] = "公司"
    ws["A6"] = "流动资产："
    ws["A7"] = "货币资金"
    ws["C7"] = "{{row:BS-002:current}}"
    ws["D7"] = "{{row:BS-002:current:parent}}"
    ws["E7"] = "{{row:BS-002:prior}}"
    ws["F7"] = "{{row:BS-002:prior:parent}}"
    ws["A8"] = "合计"
    ws["C8"] = "=SUM(C7:C7)"
    p = tmp_path / "consol.xlsx"
    wb.save(p)
    return p


async def _add_project(db, *, company_code, parent_code, scope):
    proj = Project(
        id=uuid.uuid4(),
        name=f"测试{company_code}",
        client_name=f"测试{company_code}",
        template_type="soe",
        report_scope=scope,
        company_code=company_code,
        parent_company_code=parent_code,
    )
    db.add(proj)
    await db.flush()
    return proj


async def _add_bs_row(db, project_id, *, current, prior):
    db.add(FinancialReport(
        id=uuid.uuid4(),
        project_id=project_id,
        year=2024,
        report_type=FinancialReportType.balance_sheet,
        row_code="BS-002",
        row_name="货币资金",
        current_period_amount=Decimal(str(current)),
        prior_period_amount=Decimal(str(prior)),
        indent_level=1,
        is_total_row=False,
        generated_at=datetime.now(timezone.utc),
    ))
    await db.flush()


@pytest.mark.asyncio
async def test_exporter_resolves_parent_column(db_session, tmp_path, monkeypatch):
    """公司列填母公司个别值，合并列填本项目值."""
    parent = await _add_project(
        db_session, company_code="PARENT01", parent_code=None, scope="standalone"
    )
    child = await _add_project(
        db_session, company_code="CHILD01", parent_code="PARENT01",
        scope="consolidated",
    )
    await _add_bs_row(db_session, child.id, current=1000.0, prior=900.0)
    await _add_bs_row(db_session, parent.id, current=2000.0, prior=1800.0)

    synth = _synthetic_consolidated_template(tmp_path)
    monkeypatch.setattr(
        ReportExcelExporter, "_load_template",
        lambda self, key: load_workbook(str(synth)),
    )

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=child.id, year=2024, mode="audited",
        report_types=["balance_sheet"], include_prior_year=True,
    )
    ws = load_workbook(output)[_BS_SHEET]
    # 合并列 = 本项目（child）值
    assert ws["C7"].value == pytest.approx(1000.0)
    assert ws["E7"].value == pytest.approx(900.0)
    # 公司列 = 母公司个别（parent）值
    assert ws["D7"].value == pytest.approx(2000.0)
    assert ws["F7"].value == pytest.approx(1800.0)
    # SUM 未被覆盖
    assert str(ws["C8"].value).startswith("=SUM")


@pytest.mark.asyncio
async def test_exporter_missing_parent_leaves_blank(db_session, tmp_path, monkeypatch):
    """母公司项目缺失 → 公司列留空，不崩."""
    child = await _add_project(
        db_session, company_code="CHILD02", parent_code="NONEXISTENT",
        scope="consolidated",
    )
    await _add_bs_row(db_session, child.id, current=1000.0, prior=900.0)

    synth = _synthetic_consolidated_template(tmp_path)
    monkeypatch.setattr(
        ReportExcelExporter, "_load_template",
        lambda self, key: load_workbook(str(synth)),
    )

    exporter = ReportExcelExporter(db_session)
    output = await exporter.export(
        project_id=child.id, year=2024, mode="audited",
        report_types=["balance_sheet"], include_prior_year=True,
    )
    ws = load_workbook(output)[_BS_SHEET]
    # 合并列正常
    assert ws["C7"].value == pytest.approx(1000.0)
    # 公司列留空（fill_empty_as=blank → None），无占位残留
    assert ws["D7"].value in (None, "")
    assert "{{" not in str(ws["D7"].value or "")
    assert ws["F7"].value in (None, "")
